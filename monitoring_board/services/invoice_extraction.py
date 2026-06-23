from __future__ import annotations

import csv
import hashlib
import re
import zipfile
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from monitoring_board.reporting.invoices import (
    extraction_result_from_candidates,
    is_supported_invoice_extension,
    normalize_decimal,
    normalize_nif,
)
from monitoring_board.reporting.models import InvoiceCandidate, InvoiceExtractionResult


@dataclass(frozen=True)
class InvoiceFileReadResult:
    text: str
    method: str
    warnings: tuple[str, ...] = ()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def extract_invoice_file(path: Path) -> InvoiceExtractionResult:
    if not is_supported_invoice_extension(path.name):
        return extraction_result_from_candidates(method="unsupported", candidates=(), warnings=("unsupported_invoice_format",), errors=("unsupported_invoice_format",))
    try:
        validate_invoice_file_content(path)
    except ValueError as exc:
        return extraction_result_from_candidates(method="unsupported", candidates=(), warnings=("invalid_invoice_file_content",), errors=(str(exc),))
    read = read_invoice_text(path)
    if not read.text.strip():
        return extraction_result_from_candidates(method=read.method, candidates=(), warnings=(*read.warnings, "invoice_requires_manual_review"))
    candidates = tuple(extract_candidates_from_text(read.text, source=read.method))
    warnings = [*read.warnings, *detect_ambiguous_price_warnings(read.text, candidates)]
    if not candidates:
        warnings.append("invoice_values_incomplete")
    return extraction_result_from_candidates(method=read.method, candidates=candidates, warnings=tuple(warnings))


def validate_invoice_file_content(path: Path) -> None:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        with path.open("rb") as handle:
            if handle.read(5) != b"%PDF-":
                raise ValueError("invalid_pdf_signature")
        return
    if suffix in {".xlsx", ".xlsm"}:
        if not zipfile.is_zipfile(path):
            raise ValueError("invalid_office_zip")
        with zipfile.ZipFile(path) as archive:
            names = set(archive.namelist())
            if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
                raise ValueError("invalid_office_structure")
        return
    if suffix in {".txt", ".csv"}:
        sample = path.read_bytes()[:4096]
        if b"\x00" in sample:
            raise ValueError("binary_text_invoice")
        if sample:
            control = sum(1 for byte in sample if byte < 32 and byte not in {9, 10, 13})
            high = sum(1 for byte in sample if byte >= 128)
            if control / len(sample) > 0.02 or high / len(sample) > 0.40:
                raise ValueError("binary_text_invoice")
        return
    raise ValueError("unsupported_invoice_format")


def read_invoice_text(path: Path) -> InvoiceFileReadResult:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return read_pdf_text(path)
    if suffix in {".xlsx", ".xlsm"}:
        return read_excel_text(path)
    if suffix == ".csv":
        return read_csv_text(path)
    if suffix == ".txt":
        return InvoiceFileReadResult(path.read_text(encoding="utf-8", errors="replace"), "txt")
    return InvoiceFileReadResult("", "unsupported", ("unsupported_invoice_format",))


def read_pdf_text(path: Path) -> InvoiceFileReadResult:
    try:
        from pypdf import PdfReader
    except ImportError:
        return InvoiceFileReadResult("", "pdf", ("unsupported_invoice_format",))
    reader = PdfReader(str(path))
    parts: list[str] = []
    for page_index, page in enumerate(reader.pages[:8], start=1):
        text = page.extract_text() or ""
        if text:
            parts.append(f"[page {page_index}]\n{text}")
    if not parts:
        return InvoiceFileReadResult("", "pdf", ("scanned_pdf_requires_manual_review",))
    return InvoiceFileReadResult("\n".join(parts), "pdf")


def read_excel_text(path: Path) -> InvoiceFileReadResult:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    parts: list[str] = []
    for sheet in workbook.worksheets[:4]:
        for row in sheet.iter_rows(max_row=200, max_col=20, values_only=True):
            values = [str(value).strip() for value in row if value not in (None, "")]
            if values:
                parts.append(f"{sheet.title}: " + " | ".join(values))
    workbook.close()
    return InvoiceFileReadResult("\n".join(parts), "excel")


def read_csv_text(path: Path) -> InvoiceFileReadResult:
    content = path.read_text(encoding="utf-8", errors="replace")
    sample = content[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    rows: list[str] = []
    for index, row in enumerate(csv.reader(content.splitlines(), dialect), start=1):
        if index > 300:
            break
        if any(cell.strip() for cell in row):
            rows.append(" | ".join(cell.strip() for cell in row))
    return InvoiceFileReadResult("\n".join(rows), "csv")


def extract_candidates_from_text(text: str, *, source: str) -> list[InvoiceCandidate]:
    normalized = re.sub(r"[ \t]+", " ", text.replace("\r", "\n"))
    candidates: list[InvoiceCandidate] = []
    add_regex_candidate(candidates, "invoice_number", normalized, r"(?:Fatura|Factura|Invoice)\s*(?:n[.Âºo]*|number|#)?\s*[:\-]?\s*([A-Z0-9][A-Z0-9\/\-.]{2,})", "0.86", source)
    add_nif_candidates(candidates, normalized, source)
    add_regex_candidate(candidates, "issue_date", normalized, r"(?:Data de emiss[aÃ£]o|Issue date|Data)\s*[:\-]?\s*(\d{1,2}[\/\-.]\d{1,2}[\/\-.]\d{4}|\d{4}-\d{2}-\d{2})", "0.80", source)
    period = re.search(r"(?:Per[iÃ­]odo(?: de fatura[cÃ§][aÃ£]o)?|Billing period)\s*[:\-]?\s*(\d{1,2}[\/\-.]\d{1,2}[\/\-.]\d{4}|\d{4}-\d{2}-\d{2})\s*(?:a|atÃ©|-|to)\s*(\d{1,2}[\/\-.]\d{1,2}[\/\-.]\d{4}|\d{4}-\d{2}-\d{2})", normalized, flags=re.I)
    if period:
        candidates.append(InvoiceCandidate("billing_period_start", period.group(1), Decimal("0.84"), short_evidence(period.group(0)), source))
        candidates.append(InvoiceCandidate("billing_period_end", period.group(2), Decimal("0.84"), short_evidence(period.group(0)), source))
    add_decimal_candidate(candidates, "total_amount", normalized, r"(?:Total(?: da fatura)?|Total amount)\s*[:\-]?\s*([0-9][0-9\s.,]*)\s*(?:â‚¬|EUR)", "0.76", source)
    add_decimal_candidate(candidates, "total_energy_kwh", normalized, r"(?:Energia ativa|Consumo|Total energia).*?([0-9][0-9\s.,]*)\s*kWh", "0.68", source)
    add_price_candidates(candidates, normalized, source)
    return dedupe_candidates(candidates)


def add_nif_candidates(candidates: list[InvoiceCandidate], text: str, source: str) -> None:
    nif_value = r"((?:PT\s*)?\d(?:[\s\-]?\d){8})(?![\s\-]?\d)"
    explicit_patterns = (
        ("supplier_nif", rf"(?:Fornecedor|Supplier|Emitente)[^\n]{{0,100}}(?:NIF|VAT)\s*[:\-]?\s*{nif_value}", "0.90"),
        ("customer_nif", rf"(?:Cliente|Customer|Titular)[^\n]{{0,100}}(?:NIF|VAT)\s*[:\-]?\s*{nif_value}", "0.90"),
        ("supplier_nif", rf"(?:NIF|VAT)\s+(?:Fornecedor|Supplier|Emitente)\s*[:\-]?\s*{nif_value}", "0.88"),
        ("customer_nif", rf"(?:NIF|VAT)\s+(?:Cliente|Customer|Titular)\s*[:\-]?\s*{nif_value}", "0.88"),
    )
    for field_name, pattern, confidence in explicit_patterns:
        add_nif_candidate(candidates, field_name, text, pattern, confidence, source)
    explicit_fields = {candidate.field_name for candidate in candidates if candidate.field_name in {"supplier_nif", "customer_nif"}}
    generic_nifs = re.findall(rf"(?:NIF|VAT)\s*[:\-]?\s*{nif_value}", text, flags=re.I)
    if not generic_nifs:
        return
    if "customer_nif" not in explicit_fields:
        candidates.append(InvoiceCandidate("customer_nif", normalize_nif(generic_nifs[-1]), Decimal("0.55" if len(generic_nifs) == 1 else "0.65"), short_evidence(f"NIF: {generic_nifs[-1]}"), source))
    if len(generic_nifs) > 1 and "supplier_nif" not in explicit_fields:
        candidates.append(InvoiceCandidate("supplier_nif", normalize_nif(generic_nifs[0]), Decimal("0.60"), short_evidence(f"NIF: {generic_nifs[0]}"), source))


def add_nif_candidate(candidates: list[InvoiceCandidate], field_name: str, text: str, pattern: str, confidence: str, source: str) -> None:
    match = re.search(pattern, text, flags=re.I)
    if match:
        candidates.append(InvoiceCandidate(field_name, normalize_nif(match.group(1)), Decimal(confidence), short_evidence(match.group(0)), source))


def add_price_candidates(candidates: list[InvoiceCandidate], text: str, source: str) -> None:
    labels = {
        "super_vazio_price_eur_kwh": r"super\s+vazio",
        "ponta_price_eur_kwh": r"ponta",
        "cheia_price_eur_kwh": r"cheias?|fora\s+de\s+vazio",
        "vazio_price_eur_kwh": r"(?<!super\s)vazio",
        "simple_price_eur_kwh": r"pre[cç]o\s+unit[aá]rio|preco\s+unitario|energia\s+ativa",
    }
    for field_name, label in labels.items():
        for match in re.finditer(label, text, flags=re.I):
            context = short_evidence(text[match.start() : match.start() + 140])
            value = price_from_line(context)
            if value is not None:
                candidates.append(InvoiceCandidate(field_name, str(value), Decimal("0.82"), context, source))


def price_from_line(line: str) -> Decimal | None:
    unit_pattern = r"(?:EUR\s*/\s*kWh|\u20ac\s*/\s*kWh|â‚¬\s*/\s*kWh|Eur\s+por\s+kWh|pre[cç]o\s+unit[aá]rio|preco\s+unitario)"
    number_pattern = r"[0-9]+(?:[.,][0-9]+)?"
    matches = list(re.finditer(rf"({number_pattern})\s*{unit_pattern}|{unit_pattern}\s*[:\-]?\s*({number_pattern})", line, flags=re.I))
    for match in matches:
        raw = match.group(1) or match.group(2)
        try:
            value = normalize_decimal(raw)
        except ValueError:
            continue
        if Decimal("0") <= value <= Decimal("2"):
            return value
    return None


def detect_ambiguous_price_warnings(text: str, candidates: tuple[InvoiceCandidate, ...]) -> tuple[str, ...]:
    price_fields = {candidate.field_name for candidate in candidates if candidate.field_name.endswith("_price_eur_kwh")}
    if price_fields:
        return ()
    for line in text.splitlines():
        if re.search(r"\b(?:Ponta|Cheias?|Vazio|Super vazio|Preco unitario|Preço unitário)\b", line, flags=re.I) and re.search(r"\d+[,.]\d+", line):
            return ("ambiguous_invoice_price",)
    return ()


def add_regex_candidate(candidates: list[InvoiceCandidate], field_name: str, text: str, pattern: str, confidence: str, source: str) -> None:
    match = re.search(pattern, text, flags=re.I | re.S)
    if match:
        candidates.append(InvoiceCandidate(field_name, match.group(1).strip(), Decimal(confidence), short_evidence(match.group(0)), source))


def add_decimal_candidate(candidates: list[InvoiceCandidate], field_name: str, text: str, pattern: str, confidence: str, source: str) -> None:
    match = re.search(pattern, text, flags=re.I | re.S)
    if not match:
        return
    try:
        value = normalize_decimal(match.group(1))
    except ValueError:
        return
    candidates.append(InvoiceCandidate(field_name, str(value), Decimal(confidence), short_evidence(match.group(0)), source))


def short_evidence(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()[:160]


def dedupe_candidates(candidates: list[InvoiceCandidate]) -> list[InvoiceCandidate]:
    best: dict[str, InvoiceCandidate] = {}
    for candidate in candidates:
        current = best.get(candidate.field_name)
        if current is None or candidate.confidence > current.confidence:
            best[candidate.field_name] = candidate
    return list(best.values())
