from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from monitoring_board.portfolio_reports import build_portfolio_report_rows
from monitoring_board.reporting.financial_quality import apply_production_financial_gate
from monitoring_board.reporting.periods import build_period
from monitoring_board.reporting.portfolio import (
    ENGINE_VERSION,
    METRIC_CATALOG,
    PortfolioComparisonResult,
    PortfolioReportProfile,
    PortfolioReportRequest,
    PortfolioReportResult,
    PortfolioReportRow,
    aggregate_rows,
    apply_filters,
    apply_sort,
    comparison_values,
    data_coverage,
    evaluate_thresholds,
)


def prepare_portfolio_report(
    conn,
    *,
    portfolio_id: int,
    portfolio_name: str,
    profile: PortfolioReportProfile,
    period_type: str = "monthly",
    report_month: str | None = None,
    year: int | str | None = None,
    quarter: int | str | None = None,
    semester: int | str | None = None,
    comparison: str = "",
    profile_version: int = 1,
    reference_date: date | None = None,
) -> PortfolioReportResult:
    period = build_period(period_type, report_month=report_month, year=year, month=(report_month or "")[-2:] if report_month else None, quarter=quarter, semester=semester)
    request = PortfolioReportRequest(portfolio_id=portfolio_id, period=period, profile_id=profile.id, comparison=comparison)
    reference_date = reference_date or date.today()
    rows = load_period_rows(conn, request, reference_date=reference_date)
    rows = apply_filters(rows, profile)
    rows = tuple(PortfolioReportRow(row.asset_id, row.values, row.warnings, evaluate_thresholds(row, profile.thresholds)) for row in rows)
    rows = apply_sort(rows, profile)
    summary = aggregate_rows(rows, profile.columns)
    coverage = data_coverage(rows, period.included_months)
    comparison_result: PortfolioComparisonResult | None = None
    if comparison:
        previous_period = comparison_period_args(period.start, period_type, comparison)
        previous = prepare_portfolio_report(
            conn,
            portfolio_id=portfolio_id,
            portfolio_name=portfolio_name,
            profile=profile,
            period_type=period_type,
            report_month=previous_period.get("report_month"),
            year=previous_period.get("year"),
            quarter=previous_period.get("quarter"),
            semester=previous_period.get("semester"),
            comparison="",
            profile_version=profile_version,
            reference_date=reference_date,
        )
        comparison_result = comparison_values(summary, previous.summary, comparison)
    warnings = tuple(sorted({warning for row in rows for warning in row.warnings} | set(summary.warnings) | (set(comparison_result.warnings) if comparison_result else set())))
    return PortfolioReportResult(
        portfolio_id=portfolio_id,
        portfolio_name=portfolio_name,
        profile=profile,
        profile_version=profile_version,
        period=period,
        columns=tuple(column for column in profile.columns if column.visible),
        rows=rows,
        summary=summary,
        comparison=comparison_result,
        coverage=coverage,
        warnings=warnings,
        metadata={
            "period_type": period.period_type.value,
            "period_start": period.start.isoformat(),
            "period_end": period.end.isoformat(),
            "row_count": len(rows),
        },
        engine_version=ENGINE_VERSION,
        generated_at=datetime.now(),
    )


def load_period_rows(conn, request: PortfolioReportRequest, *, reference_date: date) -> tuple[PortfolioReportRow, ...]:
    by_asset: dict[int | None, dict[str, Any]] = {}
    warnings_by_asset: dict[int | None, set[str]] = {}
    missing_months_by_asset: dict[int | None, set[str]] = {}
    for month in request.period.included_months:
        monthly_rows = build_portfolio_report_rows(
            conn,
            request.portfolio_id,
            month.strftime("%Y-%m"),
            reference_date=reference_date,
        )
        seen_assets: set[int | None] = set()
        for monthly in monthly_rows:
            asset_id = monthly.get("asset_id")
            seen_assets.add(asset_id)
            target = by_asset.setdefault(asset_id, base_values(monthly))
            warnings_by_asset.setdefault(asset_id, set()).update(monthly.get("warnings", []))
            accumulate_month(target, monthly)
            accumulate_source_coverage(target, monthly)
        for asset_id in set(by_asset) - seen_assets:
            missing_months_by_asset.setdefault(asset_id, set()).add(month.isoformat())
    rows: list[PortfolioReportRow] = []
    for asset_id, values in by_asset.items():
        values["_expected_months"] = len(request.period.included_months)
        finalize_values(values)
        missing_sources = missing_sources_for_values(values, warnings_by_asset.get(asset_id, set()))
        values["missing_sources"] = tuple(sorted(missing_sources))
        values["missing_months"] = tuple(sorted(missing_months_by_asset.get(asset_id, set())))
        values["warning_count"] = len(warnings_by_asset.get(asset_id, set()))
        values["coverage_pct"] = coverage_for_row(missing_sources)
        rows.append(PortfolioReportRow(asset_id=asset_id, values=values, warnings=tuple(sorted(warnings_by_asset.get(asset_id, set())))))
    return tuple(rows)


def base_values(monthly: dict[str, Any]) -> dict[str, Any]:
    return {
        "asset_id": monthly.get("asset_id"),
        "installation": monthly.get("installation"),
        "local_installation": monthly.get("local_installation"),
        "nif": monthly.get("nif"),
        "sub_account": monthly.get("sub_account"),
        "installed_power_kwp": monthly.get("installed_power_kwp"),
        "mapping_confidence": monthly.get("mapping_confidence"),
        "invoice_status": monthly.get("invoice_status"),
        "tariff_type": monthly.get("tariff_type"),
        "data_status": monthly.get("data_status"),
        "warning_labels": monthly.get("warning_labels", []),
        "production_quality_status": monthly.get("production_quality_status"),
        "production_source": monthly.get("production_source"),
        "production_expected_days": 0,
        "production_available_days": 0,
        "raw_daily_production_kwh": None,
        "_production_statuses": [],
        "_complete_production_months": 0,
        "_source_slots": {"production": 0, "helioscope": 0, "availability": 0, "tariff": 0, "self_use": 0, "invoice": 0, "mapping": 0},
        "_availability_weighted": Decimal("0"),
        "_availability_weight": Decimal("0"),
    }


def accumulate_month(target: dict[str, Any], monthly: dict[str, Any]) -> None:
    production_status = str(monthly.get("production_quality_status") or ("complete" if monthly.get("actual_production_kwh") is not None else "missing"))
    target.setdefault("_production_statuses", []).append(production_status)
    if production_status == "complete":
        target["_complete_production_months"] = int(target.get("_complete_production_months", 0)) + 1
    target["production_expected_days"] = int(target.get("production_expected_days") or 0) + int(monthly.get("production_expected_days") or 0)
    target["production_available_days"] = int(target.get("production_available_days") or 0) + int(monthly.get("production_available_days") or 0)
    if monthly.get("raw_daily_production_kwh") is not None:
        target["raw_daily_production_kwh"] = Decimal(str(target.get("raw_daily_production_kwh") or 0)) + Decimal(str(monthly["raw_daily_production_kwh"]))
    for key in (
        "actual_production_kwh",
        "helioscope_expected_kwh",
        "expected_production_kwh",
        "expected_consumption_kwh",
        "expected_self_use_kwh",
        "expected_export_kwh",
        "expected_grid_import_kwh",
        "adjusted_expected_kwh",
        "production_ponta_kwh",
        "production_cheia_kwh",
        "production_vazio_kwh",
        "production_super_vazio_kwh",
        "self_use_ponta_kwh",
        "self_use_cheia_kwh",
        "self_use_vazio_kwh",
        "self_use_super_vazio_kwh",
        "self_use_value_ponta_eur",
        "self_use_value_cheia_eur",
        "self_use_value_vazio_eur",
        "self_use_value_super_vazio_eur",
        "estimated_value_eur",
        "self_use_kwh",
        "self_use_simple_kwh",
        "self_use_value_simple_eur",
        "export_kwh",
        "consumption_kwh",
        "grid_import_kwh",
        "export_revenue_eur",
        "esco_payment_eur",
        "fixed_fee_eur",
        "net_benefit_eur",
    ):
        if monthly.get(key) is not None:
            target[key] = Decimal(str(target.get(key) or 0)) + Decimal(str(monthly[key]))
    for key in (
        "expected_self_consumption_rate_pct",
        "expected_self_sufficiency_rate_pct",
        "expected_specific_yield",
        "expected_production_source",
    ):
        if monthly.get(key) is not None and target.get(key) is None:
            target[key] = monthly.get(key)
    if monthly.get("availability_pct") is not None and monthly.get("installed_power_kwp"):
        target["_availability_weighted"] += Decimal(str(monthly["availability_pct"])) * Decimal(str(monthly["installed_power_kwp"]))
        target["_availability_weight"] += Decimal(str(monthly["installed_power_kwp"]))


def accumulate_source_coverage(target: dict[str, Any], monthly: dict[str, Any]) -> None:
    warnings = set(monthly.get("warnings") or ())
    slots = target.setdefault("_source_slots", {})
    source_warnings = {
        "production": {"missing_monthly_production"},
        "helioscope": {"missing_helioscope_expected"},
        "availability": {"missing_availability"},
        "tariff": {"missing_tariff", "expired_tariff", "tariff_validity_gap", "overlapping_tariffs"},
        "self_use": {"missing_hourly_self_use", "missing_self_use"},
        "invoice": {"missing_invoice", "review_required", "extraction_failed", "incompatible_invoice"},
        "mapping": {"mapping_pending", "mapping_conflict"},
    }
    for source, missing_codes in source_warnings.items():
        if source == "production":
            production_status = monthly.get("production_quality_status") or ("complete" if monthly.get("actual_production_kwh") is not None else "missing")
            if production_status == "complete":
                slots[source] = int(slots.get(source, 0)) + 1
        elif source == "self_use" and "inferred_hourly_self_use" in warnings and not warnings.intersection(missing_codes):
            slots[source] = int(slots.get(source, 0)) + 1
        elif not warnings.intersection(missing_codes):
            slots[source] = int(slots.get(source, 0)) + 1


def finalize_values(values: dict[str, Any]) -> None:
    actual = _decimal_or_none(values.get("actual_production_kwh"))
    complete_months = int(values.get("_complete_production_months") or 0)
    expected_months = int(values.get("_expected_months") or 0)
    statuses = list(values.get("_production_statuses") or [])
    values["complete_production_subtotal_kwh"] = actual
    if complete_months != expected_months:
        values["actual_production_kwh"] = None
        actual = None
    status_priority = {"complete": 0, "missing": 1, "partial": 2, "in_progress": 3, "conflict": 4}
    values["production_quality_status"] = max(statuses, key=lambda item: status_priority.get(item, 1), default="missing")
    expected_days = int(values.get("production_expected_days") or 0)
    values["production_coverage_ratio"] = (
        Decimal(int(values.get("production_available_days") or 0)) / Decimal(expected_days)
        if expected_days
        else Decimal("0")
    )
    values["production_coverage_pct"] = values["production_coverage_ratio"] * Decimal("100")
    adjusted = _decimal_or_none(values.get("adjusted_expected_kwh"))
    installed = _decimal_or_none(values.get("installed_power_kwp"))
    self_use = _decimal_or_none(values.get("self_use_kwh"))
    export = _decimal_or_none(values.get("export_kwh"))
    consumption = _decimal_or_none(values.get("consumption_kwh"))
    values["deviation_kwh"] = actual - adjusted if actual is not None and adjusted is not None else None
    values["deviation_pct"] = ((actual - adjusted) / adjusted * Decimal("100")) if actual is not None and adjusted else None
    values["specific_yield"] = actual / installed if actual is not None and installed else None
    values["self_consumption_rate_pct"] = self_use / (self_use + export) * Decimal("100") if self_use is not None and export is not None and (self_use + export) else None
    values["self_sufficiency_rate_pct"] = self_use / consumption * Decimal("100") if self_use is not None and consumption else None
    values["availability_pct"] = values["_availability_weighted"] / values["_availability_weight"] if values["_availability_weight"] else None
    apply_production_financial_gate(values)
    values.pop("_availability_weighted", None)
    values.pop("_availability_weight", None)
    values.pop("_production_statuses", None)
    values.pop("_complete_production_months", None)
    for key, definition in METRIC_CATALOG.items():
        if key in values and isinstance(values[key], Decimal):
            values[key] = values[key].quantize(Decimal("1") if definition.decimals <= 0 else Decimal("1." + ("0" * definition.decimals)))


def _decimal_or_none(value: Any) -> Decimal | None:
    if value is None:
        return None
    return Decimal(str(value))


def missing_sources_for_values(values: dict[str, Any], warnings: set[str]) -> set[str]:
    missing = set()
    if values.get("production_quality_status") != "complete" or "missing_monthly_production" in warnings:
        missing.add("production")
    if "missing_helioscope_expected" in warnings:
        missing.add("helioscope")
    if "missing_availability" in warnings:
        missing.add("availability")
    if any(warning in warnings for warning in {"missing_tariff", "expired_tariff", "tariff_validity_gap", "overlapping_tariffs"}):
        missing.add("tariff")
    if any(warning in warnings for warning in {"missing_hourly_self_use", "missing_self_use"}):
        missing.add("self_use")
    if any(warning in warnings for warning in {"missing_invoice", "review_required", "extraction_failed", "incompatible_invoice"}):
        missing.add("invoice")
    if any(warning in warnings for warning in {"mapping_pending", "mapping_conflict"}):
        missing.add("mapping")
    return missing


def coverage_for_row(missing_sources: set[str]) -> Decimal:
    total = Decimal("7")
    complete = total - Decimal(len(missing_sources))
    return (complete / total * Decimal("100")).quantize(Decimal("1.00"))


def comparison_period_args(start: date, period_type: str, comparison: str) -> dict[str, Any]:
    if comparison == "previous_year":
        previous = start.replace(year=start.year - 1)
    elif period_type == "quarterly":
        previous_month = start.month - 3
        previous = date(start.year - 1, 10, 1) if previous_month < 1 else date(start.year, previous_month, 1)
    elif period_type == "semiannual":
        previous = date(start.year - 1, 7, 1) if start.month == 1 else date(start.year, 1, 1)
    elif period_type == "annual":
        previous = date(start.year - 1, 1, 1)
    else:
        previous = date(start.year - 1, 12, 1) if start.month == 1 else date(start.year, start.month - 1, 1)
    if period_type == "monthly":
        return {"report_month": previous.strftime("%Y-%m")}
    if period_type == "quarterly":
        return {"year": previous.year, "quarter": ((previous.month - 1) // 3) + 1}
    if period_type == "semiannual":
        return {"year": previous.year, "semester": 1 if previous.month == 1 else 2}
    return {"year": previous.year}


def export_portfolio_result_workbook(result: PortfolioReportResult) -> Workbook:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumo"
    summary.append(["Portfolio", result.portfolio_name])
    summary.append(["Periodo", result.period.label])
    summary.append(["Perfil", result.profile.name])
    summary.append(["Versao do perfil", result.profile_version])
    summary.append(["Engine", result.engine_version])
    summary.append(["Cobertura global", float(result.coverage.global_pct)])
    summary.append([])
    summary.append(["Metrica", "Valor"])
    for key, value in result.summary.values.items():
        summary.append([METRIC_CATALOG[key].label if key in METRIC_CATALOG else key, format_cell(value)])
    if result.comparison:
        summary.append([])
        summary.append(["Comparacao", result.comparison.mode])
        summary.append(["Metrica", "Atual", "Anterior", "Diferenca", "Diferenca %"])
        for key, item in result.comparison.values.items():
            summary.append([
                METRIC_CATALOG[key].label if key in METRIC_CATALOG else key,
                format_cell(item.get("current")),
                format_cell(item.get("previous")),
                format_cell(item.get("delta")),
                format_cell(item.get("delta_pct")),
            ])
    summary.append([])
    summary.append(["Warnings", ", ".join(result.warnings)])
    sheet = workbook.create_sheet("Instalacoes")
    sheet.append([column.label for column in result.columns])
    for row in result.rows:
        sheet.append([format_cell(row.values.get(column.metric_key)) for column in result.columns])
    if result.summary.values:
        sheet.append([
            format_cell(result.summary.values.get(column.metric_key)) if column.metric_key in result.summary.values else ("TOTAL" if index == 0 else None)
            for index, column in enumerate(result.columns)
        ])
    sheet.freeze_panes = "B2"
    if sheet.max_row and sheet.max_column:
        sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(result.columns, start=1):
        definition = METRIC_CATALOG.get(column.metric_key)
        number_format = number_format_for(definition.value_type if definition else "", column.decimals if column.decimals is not None else (definition.decimals if definition else 2))
        if number_format:
            for cell in sheet.iter_cols(min_col=index, max_col=index, min_row=2):
                for item in cell:
                    item.number_format = number_format
    quality = workbook.create_sheet("Qualidade dos dados")
    quality.append(["Fonte", "Cobertura"])
    for source, value in result.coverage.by_source.items():
        quality.append([source, float(value)])
    quality.append([])
    quality.append(["Instalacao", "Estado producao", "Cobertura diaria %", "Subtotal diario bruto kWh", "Codigo", "Severidade", "Fonte", "Acao sugerida"])
    for row in result.rows:
        if not row.warnings:
            quality.append([
                row.values.get("installation") or row.asset_id or "-",
                row.values.get("production_quality_status") or "-",
                format_cell(row.values.get("production_coverage_pct")),
                format_cell(row.values.get("raw_daily_production_kwh")),
                "",
                "",
                "production",
                "",
            ])
        for warning in row.warnings:
            quality.append([
                row.values.get("installation") or row.asset_id or "-",
                row.values.get("production_quality_status") or "-",
                format_cell(row.values.get("production_coverage_pct")),
                format_cell(row.values.get("raw_daily_production_kwh")),
                warning,
                warning_severity(warning),
                warning_source(warning),
                warning_action(warning),
            ])
    metadata = workbook.create_sheet("Metadados")
    metadata.append(["engine_version", result.engine_version])
    metadata.append(["generated_at", result.generated_at.isoformat(timespec="seconds")])
    if result.metadata.get("snapshot_id"):
        metadata.append(["snapshot_id", result.metadata["snapshot_id"]])
    metadata.append(["profile", result.profile.name])
    metadata.append(["profile_version", result.profile_version])
    metadata.append(["period_type", result.period.period_type.value])
    metadata.append(["period_start", result.period.start.isoformat()])
    metadata.append(["period_end", result.period.end.isoformat()])
    metadata.append(["months", ", ".join(month.isoformat() for month in result.period.included_months)])
    metadata.append(["sources", ", ".join(result.coverage.by_source.keys())])
    metadata.append(["columns", ", ".join(column.metric_key for column in result.columns)])
    for sheet_obj in workbook.worksheets:
        if sheet_obj.max_row:
            for cell in sheet_obj[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
        for column in sheet_obj.columns:
            width = max(len(str(cell.value or "")) for cell in column)
            sheet_obj.column_dimensions[column[0].column_letter].width = min(max(width + 2, 12), 48)
    return workbook


def format_cell(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (list, tuple)):
        return "; ".join(str(item) for item in value)
    return value


def number_format_for(value_type: str, decimals: int) -> str:
    if value_type in {"number", "money"}:
        base = "#,##0" if decimals <= 0 else "#,##0." + ("0" * decimals)
        return base + ' "EUR"' if value_type == "money" else base
    return ""


def warning_severity(code: str) -> str:
    if "conflict" in code or "overlapping" in code:
        return "critical"
    if code.startswith("missing") or code in {"expired_tariff", "incompatible_invoice"}:
        return "missing"
    return "warning"


def warning_source(code: str) -> str:
    if "helioscope" in code:
        return "helioscope"
    if "availability" in code:
        return "availability"
    if "tariff" in code or "price" in code:
        return "tariff"
    if "invoice" in code or code in {"review_required", "extraction_failed", "incompatible_invoice"}:
        return "invoice"
    if "mapping" in code:
        return "mapping"
    if "self_use" in code:
        return "self_use"
    return "production"


def warning_action(code: str) -> str:
    if code.startswith("missing"):
        return "Importar ou confirmar a fonte em falta."
    if code == "inferred_hourly_self_use":
        return "Confirmar dados horarios quando disponiveis."
    if "conflict" in code:
        return "Resolver conflito antes de publicar."
    return "Rever configuracao ou dados de origem."
