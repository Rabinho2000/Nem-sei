from __future__ import annotations

import hashlib
import html
import io
import shutil
import re
import unicodedata
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from monitoring_board.runtime import resolve_runtime_file_path
from monitoring_board.reporting.portfolio import METRIC_CATALOG, PortfolioReportResult
from monitoring_board.reporting.templates import ReportTemplate, enabled_sections_in_order
from monitoring_board.services.portfolio_reporting import export_portfolio_result_workbook, format_cell


MAX_BATCH_ASSETS = 25
MAX_BATCH_PERIODS = 12
MAX_ZIP_FILES = 80
MAX_TOTAL_OUTPUTS = 120
MAX_RENDERED_FILE_BYTES = 25 * 1024 * 1024
RESERVED_WINDOWS_NAMES = {"CON", "PRN", "AUX", "NUL", *(f"COM{index}" for index in range(1, 10)), *(f"LPT{index}" for index in range(1, 10))}


@dataclass(frozen=True)
class RenderedFile:
    filename: str
    content: bytes
    mimetype: str
    fmt: str
    asset_id: int | None = None
    portfolio_id: int | None = None
    snapshot_id: int | None = None
    period_type: str = ""
    period_start: str = ""
    period_end: str = ""
    is_auxiliary: bool = False
    warnings: tuple[str, ...] = ()

    @property
    def sha256(self) -> str:
        return hashlib.sha256(self.content).hexdigest()

    @property
    def size_bytes(self) -> int:
        return len(self.content)


def safe_filename(value: str, *, extension: str = "") -> str:
    if any(item in (value or "") for item in ("..", "/", "\\")) or ":" in (value or ""):
        raise ValueError("unsafe_filename")
    normalized = unicodedata.normalize("NFKD", value or "relatorio")
    ascii_text = "".join(char for char in normalized if not unicodedata.combining(char))
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", ascii_text).strip("._-") or "relatorio"
    if cleaned.upper() in RESERVED_WINDOWS_NAMES:
        cleaned = f"_{cleaned}"
    cleaned = cleaned[:100]
    suffix = extension if extension.startswith(".") or not extension else f".{extension}"
    if suffix and not cleaned.lower().endswith(suffix.lower()):
        cleaned += suffix
    if any(item in cleaned for item in ("..", "/", "\\")) or ":" in cleaned:
        raise ValueError("unsafe_filename")
    return cleaned


def validate_formats(report_type: str, formats: list[str]) -> tuple[str, ...]:
    allowed = {"pdf", "excel", "zip"}
    selected = tuple(dict.fromkeys(item.strip().lower() for item in formats if item.strip()))
    if not selected:
        raise ValueError("no_output_formats")
    unknown = set(selected) - allowed
    if unknown:
        raise ValueError("unsupported_output_format")
    main = tuple(item for item in selected if item != "zip")
    if "zip" in selected and not main:
        raise ValueError("zip_requires_main_format")
    if report_type == "individual" and any(item not in {"pdf", "excel", "zip"} for item in selected):
        raise ValueError("unsupported_individual_format")
    if report_type == "portfolio" and any(item not in {"pdf", "excel", "zip"} for item in selected):
        raise ValueError("unsupported_portfolio_format")
    return selected


def render_portfolio_html(result: PortfolioReportResult, template: ReportTemplate) -> str:
    title = html.escape(expand_pattern(template.title, result=result) or f"{result.portfolio_name} - {result.period.label}")
    parts = [
        f"<article class='report-preview' style='--primary:{html.escape(template.branding.primary_color)};--secondary:{html.escape(template.branding.secondary_color)}'>",
        logo_html(template),
        f"<h1>{title}</h1>",
        f"<p>{html.escape(template.subtitle)}</p>",
        f"<p>{html.escape(template.branding.company_name)} - {html.escape(template.branding.client_name)} - {html.escape(result.period.label)}</p>",
    ]
    for section in enabled_sections_in_order(template):
        parts.append(render_portfolio_html_section(result, section.key, section.title))
    if template.branding.footer or template.branding.disclaimer:
        parts.append(f"<footer>{html.escape(template.branding.footer)} {html.escape(template.branding.contacts)} {html.escape(template.branding.disclaimer)}</footer>")
    parts.append("</article>")
    return "".join(parts)


def render_portfolio_html_section(result: PortfolioReportResult, key: str, title: str) -> str:
    if key in {"cover", "executive_summary", "kpis"}:
        body = "".join(f"<dt>{html.escape(METRIC_CATALOG[item].label if item in METRIC_CATALOG else item)}</dt><dd>{html.escape(str(format_cell(value)))}</dd>" for item, value in result.summary.values.items())
        return f"<section data-section='{html.escape(key)}'><h2>{html.escape(title)}</h2><dl>{body}</dl></section>"
    if key == "comparison":
        if not result.comparison:
            return ""
        rows = "".join(f"<tr><td>{html.escape(METRIC_CATALOG[item].label if item in METRIC_CATALOG else item)}</td><td>{html.escape(str(values.get('current')))}</td><td>{html.escape(str(values.get('previous')))}</td><td>{html.escape(str(values.get('delta')))}</td></tr>" for item, values in result.comparison.values.items())
        return f"<section data-section='comparison'><h2>{html.escape(title)}</h2><table><tbody>{rows}</tbody></table></section>"
    if key == "installations_table":
        header = "".join(f"<th>{html.escape(column.label)}</th>" for column in result.columns)
        rows = "".join("<tr>" + "".join(f"<td>{html.escape(str(format_cell(row.values.get(column.metric_key)) or '-'))}</td>" for column in result.columns) + "</tr>" for row in result.rows)
        return f"<section data-section='installations_table'><h2>{html.escape(title)}</h2><table><thead><tr>{header}</tr></thead><tbody>{rows}</tbody></table></section>"
    if key in {"availability", "quality", "warnings", "metadata", "financial", "top_performers", "underperformers"}:
        return portfolio_detail_section(result, key, title)
    return ""


def render_portfolio_pdf(result: PortfolioReportResult, template: ReportTemplate) -> RenderedFile:
    buffer = io.BytesIO()
    pagesize = landscape(A4) if template.orientation == "landscape" else A4
    top, right, bottom, left = template.margins_mm
    doc = SimpleDocTemplate(buffer, pagesize=pagesize, leftMargin=left * mm, rightMargin=right * mm, topMargin=top * mm, bottomMargin=bottom * mm)
    styles = getSampleStyleSheet()
    story: list[Any] = [
        Paragraph(expand_pattern(template.title, result=result) or f"{result.portfolio_name} - {result.period.label}", styles["Title"]),
        Paragraph(template.subtitle or template.branding.company_name, styles["Normal"]),
        Spacer(1, 12),
    ]
    append_logo(story, template)
    for section in enabled_sections_in_order(template):
        append_portfolio_pdf_section(story, styles, result, template, section.key, section.title)
    doc.build(story, onFirstPage=page_footer(template), onLaterPages=page_footer(template))
    return checked_file(
        RenderedFile(
            filename=safe_filename(expand_pattern(template.filename_pattern, result=result), extension="pdf"),
            content=buffer.getvalue(),
            mimetype="application/pdf",
            fmt="pdf",
            portfolio_id=result.portfolio_id,
            period_type=result.period.period_type.value,
            period_start=result.period.start.isoformat(),
            period_end=result.period.end.isoformat(),
            warnings=tuple(result.warnings),
        )
    )


def append_portfolio_pdf_section(story: list[Any], styles: Any, result: PortfolioReportResult, template: ReportTemplate, key: str, title: str) -> None:
    if key in {"cover", "executive_summary", "kpis"}:
        story.append(Paragraph(title, styles["Heading2"]))
        story.append(Table([["Metrica", "Valor"], *[[METRIC_CATALOG[item].label if item in METRIC_CATALOG else item, str(format_cell(value))] for item, value in result.summary.values.items()]], hAlign="LEFT", repeatRows=1))
        story.append(Spacer(1, 10))
    elif key == "comparison" and result.comparison:
        story.append(Paragraph(title, styles["Heading2"]))
        story.append(Table([["Metrica", "Atual", "Anterior", "Diferenca"], *[[METRIC_CATALOG[item].label if item in METRIC_CATALOG else item, str(values.get("current")), str(values.get("previous")), str(values.get("delta"))] for item, values in result.comparison.values.items()]], hAlign="LEFT", repeatRows=1))
        story.append(Spacer(1, 10))
    elif key == "installations_table":
        first_column = result.columns[:1]
        metric_columns = result.columns[1:]
        chunks = [metric_columns[index : index + 8] for index in range(0, len(metric_columns), 8)] or [()]
        for chunk_index, chunk in enumerate(chunks, start=1):
            columns = tuple(first_column) + tuple(chunk)
            table_rows = [[column.label for column in columns]]
            for row in result.rows:
                table_rows.append([clip(str(format_cell(row.values.get(column.metric_key)) or "-")) for column in columns])
            table = Table(table_rows, repeatRows=1, hAlign="LEFT")
            table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(template.branding.primary_color)), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey), ("FONTSIZE", (0, 0), (-1, -1), 7)]))
            story.append(Paragraph(f"{title} {chunk_index}/{len(chunks)}", styles["Heading2"]))
            story.append(table)
            story.append(Spacer(1, 8))
    elif key in {"availability", "quality", "warnings", "metadata", "financial", "top_performers", "underperformers"}:
        story.append(Paragraph(title, styles["Heading2"]))
        for row in portfolio_detail_rows(result, key):
            story.append(Paragraph(" - ".join(str(item) for item in row), styles["Normal"]))
        story.append(Spacer(1, 8))


def render_individual_pdf(report: dict[str, Any], template: ReportTemplate) -> RenderedFile:
    buffer = io.BytesIO()
    top, right, bottom, left = template.margins_mm
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=left * mm, rightMargin=right * mm, topMargin=top * mm, bottomMargin=bottom * mm)
    styles = getSampleStyleSheet()
    asset = report.get("asset") or {}
    title = expand_individual_pattern(template.title or "{asset} - {period}", report)
    story: list[Any] = [Paragraph(title, styles["Title"]), Paragraph(template.subtitle, styles["Normal"]), Spacer(1, 12)]
    append_logo(story, template)
    for section in enabled_sections_in_order(template):
        rows = individual_section_rows(report, section.key)
        if not rows:
            continue
        story.append(Paragraph(section.title, styles["Heading2"]))
        story.append(Table(rows, hAlign="LEFT", repeatRows=1))
        story.append(Spacer(1, 8))
    doc.build(story, onFirstPage=page_footer(template), onLaterPages=page_footer(template))
    return checked_file(
        RenderedFile(
            filename=safe_filename(expand_individual_pattern(template.filename_pattern or "{asset}_{period}", report), extension="pdf"),
            content=buffer.getvalue(),
            mimetype="application/pdf",
            fmt="pdf",
            asset_id=int(asset.get("id") or asset.get("asset_id") or report.get("asset_id") or 0) or None,
            period_type=str(report.get("period_type") or "monthly"),
            period_start=str(report.get("period_start") or report.get("month_start") or ""),
            period_end=str(report.get("period_end") or report.get("month_end") or ""),
            warnings=tuple(report.get("warnings") or report.get("billing_warnings") or ()),
        )
    )


def render_individual_excel(report: dict[str, Any], template: ReportTemplate) -> RenderedFile:
    workbook = Workbook()
    asset = report.get("asset") or {}
    summary = workbook.active
    summary.title = "Resumo"
    summary.append(["Empresa", template.branding.company_name])
    summary.append(["Cliente", template.branding.client_name])
    summary.append(["Instalacao", asset.get("project_name") or asset.get("name") or ""])
    summary.append(["Periodo", report.get("period_label") or report.get("report_month") or ""])
    summary.append(["Motor", report.get("engine_version") or "individual-report-v1"])
    energy = workbook.create_sheet("Energia")
    energy.append(["Metrica", "Valor"])
    for key in ("production_kwh", "self_use_kwh", "export_kwh", "consumption_kwh", "grid_import_kwh"):
        energy.append([key, report.get(key)])
    financial = workbook.create_sheet("Financeiro")
    financial.append(["Metrica", "Valor"])
    for key in ("savings_eur", "export_revenue_eur", "solcor_payment_eur", "fixed_monthly_fee_eur", "net_benefit_eur"):
        financial.append([key, report.get(key)])
    quality = workbook.create_sheet("Qualidade dos dados")
    quality.append(["Codigo"])
    for warning in list(report.get("warnings") or []) + list(report.get("billing_warnings") or []):
        quality.append([warning])
    metadata = workbook.create_sheet("Metadados")
    for key in ("period_type", "period_start", "period_end", "months_count", "tariff_type", "billing_mode", "billing_energy_base"):
        metadata.append([key, str(report.get(key) or "")])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return checked_file(
        RenderedFile(
            filename=safe_filename(expand_individual_pattern(template.filename_pattern or "{asset}_{period}", report), extension="xlsx"),
            content=buffer.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            fmt="xlsx",
            asset_id=int(asset.get("id") or asset.get("asset_id") or report.get("asset_id") or 0) or None,
            period_type=str(report.get("period_type") or "monthly"),
            period_start=str(report.get("period_start") or report.get("month_start") or ""),
            period_end=str(report.get("period_end") or report.get("month_end") or ""),
            warnings=tuple(report.get("warnings") or report.get("billing_warnings") or ()),
        )
    )


def individual_rows(report: dict[str, Any]) -> list[list[str]]:
    rows = [["Metrica", "Valor"]]
    for label, key in (
        ("Periodo", "period_label"),
        ("Producao", "production_kwh"),
        ("Autoconsumo", "self_use_kwh"),
        ("Excedente", "export_kwh"),
        ("Consumo", "consumption_kwh"),
        ("Importacao", "grid_import_kwh"),
        ("Pagamento ESCO", "solcor_payment_eur"),
        ("Mensalidade", "fixed_monthly_fee_eur"),
        ("Beneficio liquido", "net_benefit_eur"),
    ):
        rows.append([label, str(report.get(key, "dados indisponiveis"))])
    return rows


def individual_section_rows(report: dict[str, Any], key: str) -> list[list[str]]:
    asset = report.get("asset") or {}
    mapping = {
        "cover": [["Campo", "Valor"], ["Instalacao", asset.get("project_name") or "Dados indisponiveis"], ["Periodo", report.get("period_label") or "Dados indisponiveis"]],
        "identification": [["Campo", "Valor"], ["NIF", asset.get("nif") or "Dados indisponiveis"], ["Tipo", report.get("report_type") or "Dados indisponiveis"]],
        "executive_summary": individual_rows(report),
        "production": [["Metrica", "Valor"], ["Producao", report.get("production_kwh") or "Dados indisponiveis"]],
        "self_consumption": [["Metrica", "Valor"], ["Autoconsumo", report.get("self_use_kwh") or "Dados indisponiveis"], ["Taxa", report.get("autoconsumption_pct") or "Dados indisponiveis"]],
        "financial": [["Metrica", "Valor"], ["Poupanca", report.get("savings_eur") or "Dados indisponiveis"], ["Beneficio liquido", report.get("net_benefit_eur") or "Dados indisponiveis"]],
        "tariffs": [["Campo", "Valor"], ["Tarifa", report.get("tariff_type") or "Dados indisponiveis"], ["Fonte", report.get("tariff_source") or "Dados indisponiveis"]],
        "data_quality": [["Campo", "Valor"], ["Cobertura", report.get("coverage_pct") or "Dados indisponiveis"]],
        "warnings": [["Codigo"], *[[warning] for warning in list(report.get("warnings") or []) + list(report.get("billing_warnings") or [])]],
        "metadata": [["Campo", "Valor"], ["Engine", report.get("engine_version") or "individual-report-v1"], ["Periodo", report.get("period_type") or "monthly"]],
    }
    return mapping.get(key, [["Campo", "Valor"], [key, "Dados indisponiveis"]])


def portfolio_detail_section(result: PortfolioReportResult, key: str, title: str) -> str:
    rows = "".join("<tr>" + "".join(f"<td>{html.escape(str(value))}</td>" for value in row) + "</tr>" for row in portfolio_detail_rows(result, key))
    return f"<section data-section='{html.escape(key)}'><h2>{html.escape(title)}</h2><table><tbody>{rows}</tbody></table></section>"


def portfolio_detail_rows(result: PortfolioReportResult, key: str) -> list[list[Any]]:
    if key == "financial":
        keys = ("estimated_value_eur", "export_revenue_eur", "esco_payment_eur", "fixed_fee_eur", "net_benefit_eur")
        return [[METRIC_CATALOG[item].label, format_cell(result.summary.values.get(item)) or "Dados indisponiveis"] for item in keys if item in result.summary.values]
    if key == "availability":
        low = [row.values.get("installation") for row in result.rows if row.values.get("availability_pct") is not None and row.values.get("availability_pct") < 95]
        return [["Disponibilidade", format_cell(result.summary.values.get("availability_pct")) or "Dados indisponiveis"], ["Abaixo threshold", ", ".join(str(item) for item in low) or "nenhuma"]]
    if key == "top_performers":
        ranked = sorted((row for row in result.rows if row.values.get("deviation_pct") is not None), key=lambda row: row.values["deviation_pct"], reverse=True)[:5]
        return [[row.values.get("installation"), format_cell(row.values.get("deviation_pct"))] for row in ranked] or [["Dados indisponiveis"]]
    if key == "underperformers":
        ranked = sorted((row for row in result.rows if row.values.get("deviation_pct") is not None), key=lambda row: row.values["deviation_pct"])[:5]
        return [[row.values.get("installation"), format_cell(row.values.get("deviation_pct")), ", ".join(row.warnings)] for row in ranked] or [["Dados indisponiveis"]]
    if key == "quality":
        return [["Fonte", "Cobertura"], *[[source, value] for source, value in result.coverage.by_source.items()], ["Incompletas", result.coverage.incomplete_installations], ["Meses em falta", ", ".join(result.coverage.missing_months)]]
    if key == "metadata":
        return [["Engine", result.engine_version], ["Perfil", result.profile.name], ["Versao", result.profile_version], ["Periodo", result.period.label], ["Gerado", result.generated_at.isoformat(timespec="seconds")], ["Snapshot", result.metadata.get("snapshot_id") or "-"]]
    return [["Warnings", ", ".join(result.warnings) or "sem warnings"]]


def logo_html(template: ReportTemplate) -> str:
    if not template.branding.logo_path:
        return ""
    src = html.escape(template.branding.logo_path)
    return f"<img alt='Logo' src='/{src}' style='max-height:80px'>"


def append_logo(story: list[Any], template: ReportTemplate) -> None:
    if not template.branding.logo_path:
        return
    path = resolve_runtime_file_path(template.branding.logo_path)
    if path.exists() and path.is_file():
        story.append(Image(str(path), width=80, height=40, kind="proportional"))
        story.append(Spacer(1, 8))


def render_portfolio_excel(result: PortfolioReportResult, template: ReportTemplate) -> RenderedFile:
    workbook = export_portfolio_result_workbook(result)
    workbook["Resumo"].insert_rows(1)
    workbook["Resumo"]["A1"] = template.branding.company_name
    buffer = io.BytesIO()
    workbook.save(buffer)
    return checked_file(
        RenderedFile(
            filename=safe_filename(expand_pattern(template.filename_pattern, result=result), extension="xlsx"),
            content=buffer.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            fmt="xlsx",
            portfolio_id=result.portfolio_id,
            period_type=result.period.period_type.value,
            period_start=result.period.start.isoformat(),
            period_end=result.period.end.isoformat(),
            warnings=tuple(result.warnings),
        )
    )


def render_zip(files: list[RenderedFile], filename: str = "reports.zip") -> RenderedFile:
    if len(files) > MAX_ZIP_FILES:
        raise ValueError("too_many_zip_files")
    buffer = io.BytesIO()
    used: set[str] = set()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for file in files:
            name = unique_name(file.filename, used)
            archive.writestr(name, file.content)
    return checked_file(RenderedFile(filename=safe_filename(filename, extension="zip"), content=buffer.getvalue(), mimetype="application/zip", fmt="zip", is_auxiliary=True))


def store_rendered_file(base_dir: Path, run_id: int, file: RenderedFile) -> tuple[Path, str]:
    run_dir = (base_dir / str(run_id)).resolve()
    run_dir.mkdir(parents=True, exist_ok=True)
    staging_dir = run_dir / ".staging"
    staging_dir.mkdir(parents=True, exist_ok=True)
    target = unique_path(run_dir, file.filename)
    if run_dir not in target.parents:
        raise ValueError("unsafe_output_path")
    staging_target = (staging_dir / target.name).resolve()
    if staging_dir not in staging_target.parents:
        raise ValueError("unsafe_output_path")
    staging_target.write_bytes(file.content)
    if hashlib.sha256(staging_target.read_bytes()).hexdigest() != file.sha256:
        staging_target.unlink(missing_ok=True)
        raise ValueError("staging_hash_mismatch")
    shutil.move(str(staging_target), str(target))
    cleanup_empty_staging(staging_dir)
    return target, str(target)


def cleanup_empty_staging(staging_dir: Path) -> None:
    try:
        if staging_dir.exists() and not any(staging_dir.iterdir()):
            staging_dir.rmdir()
    except OSError:
        pass


def unique_path(run_dir: Path, filename: str) -> Path:
    target = (run_dir / filename).resolve()
    if run_dir not in target.parents:
        raise ValueError("unsafe_output_path")
    if not target.exists():
        return target
    stem = target.stem
    suffix = target.suffix
    for index in range(2, 1000):
        candidate = (run_dir / f"{stem}_{index}{suffix}").resolve()
        if run_dir not in candidate.parents:
            raise ValueError("unsafe_output_path")
        if not candidate.exists():
            return candidate
    raise ValueError("too_many_duplicate_filenames")


def expand_pattern(pattern: str, *, result: PortfolioReportResult) -> str:
    return (pattern or "{portfolio}_{period}").format(
        portfolio=result.portfolio_name,
        period=result.period.label.replace(" ", "-"),
        profile=result.profile.name,
        engine=result.engine_version,
    )


def expand_individual_pattern(pattern: str, report: dict[str, Any]) -> str:
    asset = report.get("asset") or {}
    return (pattern or "{asset}_{period}").format(
        asset=asset.get("project_name") or asset.get("name") or "Instalacao",
        period=str(report.get("period_label") or report.get("report_month") or "periodo").replace(" ", "-"),
    )


def unique_name(filename: str, used: set[str]) -> str:
    if any(item in filename for item in ("/", "\\", "..")):
        raise ValueError("unsafe_zip_filename")
    if filename not in used:
        used.add(filename)
        return filename
    stem, dot, suffix = filename.rpartition(".")
    stem = stem or filename
    for index in range(2, 1000):
        candidate = f"{stem}_{index}{dot}{suffix}" if dot else f"{stem}_{index}"
        if candidate not in used:
            used.add(candidate)
            return candidate
    raise ValueError("too_many_duplicate_filenames")


def page_footer(template: ReportTemplate):
    def draw(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        footer = " - ".join(item for item in (template.branding.footer, template.branding.contacts, template.branding.disclaimer) if item)
        canvas.drawString(doc.leftMargin, 12, clip(footer, 160))
        canvas.drawRightString(doc.pagesize[0] - doc.rightMargin, 12, f"Pag. {doc.page}")
        canvas.restoreState()

    return draw


def clip(value: str, limit: int = 80) -> str:
    return value if len(value) <= limit else value[: limit - 1] + "..."


def checked_file(file: RenderedFile) -> RenderedFile:
    if file.size_bytes > MAX_RENDERED_FILE_BYTES:
        raise ValueError("rendered_file_too_large")
    return file
