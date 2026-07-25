from __future__ import annotations

import argparse
import calendar
import html
import io
import json
import logging
import math
import mimetypes
import os
import re
import secrets
import sqlite3
import threading
import time
import unicodedata
import struct
from contextvars import ContextVar
from contextlib import closing, contextmanager
from dataclasses import dataclass, replace
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import requests
from apscheduler.schedulers.background import BackgroundScheduler
from flask import current_app
from flask import Flask, abort, flash, g, has_app_context, has_request_context, redirect, render_template, request, send_file, session, url_for
from werkzeug.utils import secure_filename
from monitoring_board.db import configure_database_for_runtime, create_database_backup, ensure_column, get_db, query_all, query_scalar
from monitoring_board.logging_config import configure_logging
from monitoring_board.portfolio_reports import (
    aggregate_portfolio_total,
    auto_map_portfolio_assets,
    build_portfolio_kpis,
    build_portfolio_report_rows,
    export_portfolio_report_workbook,
    filter_report_rows,
    import_helioscope_file,
    map_external_portfolio_entity,
    seed_external_portfolio_rows,
)
from monitoring_board.routes.auth import auth_bp
from monitoring_board.routes.field_routes import field_routes_bp
from monitoring_board import runtime as runtime_module
from monitoring_board.runtime import (
    BACKUP_DIR,
    BASE_DIR,
    CONTRACTS_DIR,
    DB_PATH,
    DEFAULT_EXCEL_PATH,
    LOG_DIR,
    RUNTIME_PATHS,
    UPLOAD_DIR,
    ensure_runtime_directories,
    env_flag,
    max_upload_bytes,
    path_is_within,
    resolve_runtime_file_path_within,
    store_runtime_relative_path,
)
from monitoring_board.security import app_password_configured, csrf_token, flask_secret_key
from monitoring_board.customer_reports import (
    build_customer_report_pdf,
    detect_report_type,
    prepare_customer_report,
)
from monitoring_board.financial_model_repository import (
    ensure_financial_model_schema,
    get_asset_model as get_financial_model_for_asset,
    get_model_source as get_financial_model_source,
    list_model_monthly as list_financial_model_monthly,
)
from monitoring_board.portfolio_repository import (
    add_member as portfolio_add_member,
    apply_import_run,
    archive_portfolio,
    confirm_mapping,
    copy_members,
    create_import_preview,
    create_portfolio,
    delete_alias as portfolio_delete_alias,
    delete_portfolio,
    detect_portfolio_conflicts,
    duplicate_portfolio,
    ensure_portfolio_management_schema,
    export_configuration_workbook,
    get_import_run,
    import_preview_from_json,
    get_portfolio,
    list_aliases as portfolio_list_aliases,
    list_available_assets as portfolio_list_available_assets,
    list_portfolio_members,
    list_portfolios,
    mapping_context as portfolio_mapping_context,
    move_member_up_down,
    reactivate_portfolio,
    rebuild_asset_alias_blob as portfolio_rebuild_asset_alias_blob,
    remove_members,
    reorder_members,
    suggest_mapping as portfolio_suggest_mapping,
    sync_portfolio_asset_members,
    toggle_alias,
    unmap_member,
    update_alias as portfolio_update_alias,
    update_member as portfolio_update_member,
    update_portfolio,
    upsert_alias,
)

from monitoring_board.portfolio_report_repository import (
    archive_profile as archive_portfolio_report_profile,
    duplicate_profile as duplicate_portfolio_report_profile,
    ensure_portfolio_reporting_schema,
    get_default_profile as get_default_portfolio_report_profile,
    get_profile as get_portfolio_report_profile,
    get_snapshot_result as get_portfolio_snapshot_result,
    latest_profile_version as latest_portfolio_report_profile_version,
    list_profiles as list_portfolio_report_profiles,
    list_report_history as list_portfolio_report_history,
    save_profile as save_portfolio_report_profile,
    set_default_profile as set_default_portfolio_report_profile,
    snapshot_portfolio_result,
)
from monitoring_board.report_template_repository import (
    add_generated_file,
    archive_template as archive_report_template,
    create_generation_run,
    duplicate_template as duplicate_report_template,
    ensure_report_template_schema,
    finish_generation_run,
    get_default_template,
    get_generated_file,
    get_template,
    latest_template_version,
    list_generated_files,
    list_generation_runs,
    list_templates,
    save_template,
    set_default_template,
)
from monitoring_board.reporting_storage import reconcile_generated_reports
from monitoring_board.reporting.portfolio import (
    METRIC_CATALOG,
    aggregate_rows,
    profile_from_config,
    profile_to_config,
)
from monitoring_board.reporting.templates import default_template, template_from_config, template_to_config, validate_template_scope
from monitoring_board.services.report_rendering import (
    MAX_BATCH_ASSETS,
    MAX_BATCH_PERIODS,
    MAX_TOTAL_OUTPUTS,
    render_individual_excel,
    render_individual_pdf,
    render_portfolio_excel,
    render_portfolio_html,
    render_portfolio_pdf,
    render_zip,
    store_rendered_file,
    validate_formats,
)
from monitoring_board.services.portfolio_reporting import export_portfolio_result_workbook, prepare_portfolio_report
from monitoring_board.reporting.billing import decimal_from_value
from monitoring_board.reporting.data_quality import (
    evaluate_monthly_production_quality,
    production_quality_notice,
)
from monitoring_board.reporting.invoices import normalize_date, is_supported_invoice_extension, validate_invoice_values, warnings_require_override
from monitoring_board.reporting.availability import (
    apply_inverter_edge_tolerance as reporting_apply_inverter_edge_tolerance,
    calculate_inverter_daily_availability as reporting_calculate_inverter_daily_availability,
    calculate_weighted_plant_availability as reporting_calculate_weighted_plant_availability,
    invalid_power_warnings as reporting_invalid_power_warnings,
    inverter_availability_slot as reporting_inverter_availability_slot,
    is_inverter_available as reporting_is_inverter_available,
)
from monitoring_board.reporting.models import BillingConfig, InvoiceCandidate, InvoiceStatus, ReportPeriodType, ReportType, ReportingPeriod, TariffType
from monitoring_board.reporting.periods import (
    ReportingPeriodError,
    build_period,
    month_bounds,
    monthly_period,
    normalize_report_month as reporting_normalize_report_month,
    period_from_form,
)
from monitoring_board.reporting.repositories import (
    add_tariff_period_rule,
    apply_tri_hourly_template,
    archive_invoice_document,
    billing_config_to_form_values,
    create_invoice_document,
    create_invoice_extraction_run,
    create_source_file_record,
    delete_tariff_period_rule,
    delete_asset_tariff,
    detect_tariff_validity_warnings,
    ensure_billing_config_schema,
    get_asset_billing_config,
    get_asset_billing_config_row,
    get_invoice_document,
    get_monthly_availability,
    get_tariff_config_for_date,
    get_tariff_resolution_warnings,
    list_asset_tariffs,
    list_tariffs_at,
    list_daily_production_records,
    list_invoice_extraction_runs,
    list_portfolio_invoice_documents,
    list_hourly_production_records,
    list_monthly_production_records,
    list_tariffs_intersecting_period,
    row_to_hourly_energy_record,
    save_asset_tariff,
    duplicate_asset_tariff,
    find_invoice_by_hash,
    reject_invoice_document,
    update_invoice_from_candidates,
    update_invoice_review,
    update_tariff_period_rule,
    upsert_asset_billing_config,
)
from monitoring_board.reporting.validation import (
    BillingValidationError,
    parse_billing_config_form,
    parse_billing_values_source,
    validate_report_asset_selection,
)
from monitoring_board.reporting.tariffs import result_to_legacy_dict, value_tariff_energy, with_billing_fallback
from monitoring_board.services.invoice_extraction import extract_invoice_file, sha256_file, validate_invoice_file_content
from monitoring_board.services.fusionsolar import (
    classify_fusionsolar_inverter_availability,
    describe_fusionsolar_health_state,
    map_fusionsolar_status,
)
from monitoring_board.services.financial_models import (
    FinancialModelError,
    activate_financial_model,
    archive_financial_model,
    build_asset_financial_model_context,
    cancel_financial_model_preview,
    compare_financial_models,
    confirm_financial_model_import,
    create_financial_model_preview,
    parse_model_details_json,
    resolve_financial_model_path,
    sha256_file as financial_model_sha256_file,
)
from monitoring_board.services import fusionsolar_client as fusionsolar_client_module
from monitoring_board.services.fusionsolar_client import FusionSolarClient
from monitoring_board.services.fusionsolar_errors import (
    FusionSolarApiError,
    FusionSolarRateLimitError,
    FusionSolarSessionExpiredError,
)
from monitoring_board.services.fusionsolar_models import (
    FusionSolarEndpoints,
    collect_time_noon_of_month_ms,
    collect_time_start_of_day_ms,
    normalize_kpi_rows as normalize_client_kpi_rows,
    parse_collect_date as parse_client_collect_date,
)
from monitoring_board.services import sigenergy as sigenergy_service
from monitoring_board.services.api_client_base import retry_api_call
from monitoring_board.services.api_rate_limit import (
    ApiRateLimitError,
    active_cooldown_until,
    ensure_api_call_state_schema,
    get_api_call_state,
    mark_api_cooldown,
    record_api_attempt,
    record_api_success,
    require_not_in_cooldown,
)
from monitoring_board.services.background_job_time import (
    as_background_job_utc,
    background_job_timestamp_is_due,
    background_job_timestamp_to_lisbon,
    background_job_utc_now,
    parse_background_job_timestamp,
    serialize_background_job_timestamp,
)
from monitoring_board.services.production_api_queue import (
    ApiQueuePolicy,
    ApiSlotUnavailableError,
    PRODUCTION_KPI_AREA,
    WAT_HISTORY_AREA,
    account_key as production_api_account_key,
    ensure_api_queue_state,
    ensure_api_queue_schema,
    get_account_queue_state,
    list_api_queue_states,
    record_account_407,
    record_api_407 as record_production_api_407,
    recover_expired_leases,
    release_account_lease,
    release_api_lease,
    reserve_account_lease,
    reserve_api_slot,
)
from monitoring_board.services.sampled_availability import (
    cleanup_realtime_snapshot_payloads,
    ensure_sampled_availability_schema,
    expected_devices_for_date,
    materialize_sampled_availability_day,
    record_device_configuration,
)

from monitoring_board.services.telegram_service import (
    get_telegram_config,
    is_telegram_configured,
    send_telegram_message,
    telegram_daily_summary_enabled,
    test_telegram_connection,
)
from openpyxl import load_workbook
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


PORTFOLIO_MANAGER_ERROR_MESSAGES = {
    "member_already_exists": "Esta instalacao ja pertence a este portfolio.",
    "member_not_found": "A entrada selecionada ja nao existe neste portfolio.",
    "no_members_copied": "Nenhuma instalacao pode ser copiada.",
    "no_members_moved": "Nenhuma instalacao pode ser movida.",
    "portfolio_not_found": "O portfolio selecionado nao existe.",
    "alias_conflict": "Este alias ja esta associado a outra instalacao.",
    "portfolio_has_report_history": "Este portfolio nao pode ser apagado porque tem historico de relatorios.",
    "delete_confirmation_mismatch": "Escreve o nome exato do portfolio para confirmar.",
    "asset_not_found": "A instalacao selecionada ja nao existe.",
    "duplicate_ids": "A selecao contem entradas repetidas.",
    "invalid_member_order": "A ordem enviada ja nao corresponde a este portfolio.",
    "no_ids": "Seleciona pelo menos uma instalacao.",
}

TARIFF_WARNING_MESSAGES = {
    "overlapping_tariffs": "Existem tarifas com datas de validade sobrepostas.",
    "missing_tariff_rules": "A tarifa ainda nao tem horarios configurados.",
    "incomplete_tariff_coverage": "Os horarios nao cobrem todas as horas.",
    "missing_hourly_self_use": "Nao existem dados horarios suficientes para calcular o autoconsumo.",
    "unclassified_hourly_energy": "Existem registos horarios que nao correspondem a nenhuma regra.",
    "missing_tariff": "Nao existe uma tarifa valida para este periodo.",
    "expired_tariff": "A ultima tarifa terminou antes deste periodo.",
    "tariff_validity_gap": "Existem dias sem tarifa valida neste periodo.",
    "tariff_change_within_month": "Existem varias tarifas aplicaveis dentro do mes.",
}


def _portfolio_error_message(exc: Exception) -> str:
    code = str(exc)
    return PORTFOLIO_MANAGER_ERROR_MESSAGES.get(code, code or "Nao foi possivel concluir a operacao.")


def _portfolio_manager_redirect(portfolio_id: int | None = None, **overrides: Any):
    values: dict[str, Any] = {}
    selected_id = portfolio_id if portfolio_id is not None else int(request.form.get("portfolio_id", request.args.get("portfolio_id", "0")) or 0)
    if selected_id:
        values["portfolio_id"] = selected_id
    for key in ("tab", "search", "asset_filter", "alias_search", "alias_filter"):
        value = overrides.pop(key, None)
        if value is None:
            value = request.form.get(key, request.args.get(key, ""))
        if value:
            values[key] = value
    values.update({key: value for key, value in overrides.items() if value not in (None, "")})
    return redirect(url_for("portfolio_manager", **values))


LOGGER = logging.getLogger(__name__)

build_runtime_paths = runtime_module.build_runtime_paths
load_local_env = runtime_module.load_local_env
resolve_runtime_file_path = runtime_module.resolve_runtime_file_path


INTEGRATION_PROVIDER_FUSIONSOLAR = "FusionSolar"
INTEGRATION_PROVIDER_SIGENERGY = "Sigenergy"
INTEGRATION_PROVIDER_OPTIONS = [INTEGRATION_PROVIDER_FUSIONSOLAR, INTEGRATION_PROVIDER_SIGENERGY]
API_AREA_STATE = "state"
API_AREA_PRODUCTION = "production"
API_AREA_DIAGNOSTICS = "diagnostics"
BACKGROUND_JOB_TYPES_PERFORMANCE = (
    "fusionsolar_state_sync",
    "fusionsolar_production_sync",
    "fusionsolar_production_backfill",
    "fusionsolar_inverter_availability_backfill",
    "fusionsolar_month_cycle",
    "fusionsolar_month_close",
    "fusionsolar_report_production_request",
    "fusionsolar_report_wat_request",
    "fusionsolar_realtime_materialize_cleanup",
    "sigenergy_state_sync",
    "performance_reference_recalculation",
)
FUSIONSOLAR_PRODUCTION_JOB_TYPES = (
    "fusionsolar_production_sync",
    "fusionsolar_production_backfill",
    "fusionsolar_month_cycle",
    "fusionsolar_month_close",
    "fusionsolar_report_production_request",
)
FUSIONSOLAR_WAT_JOB_TYPES = (
    "fusionsolar_inverter_availability_backfill",
    "fusionsolar_report_wat_request",
)
FUSIONSOLAR_BACKGROUND_JOB_TYPES = (
    "fusionsolar_state_sync",
    *FUSIONSOLAR_PRODUCTION_JOB_TYPES,
    *FUSIONSOLAR_WAT_JOB_TYPES,
    "fusionsolar_realtime_materialize_cleanup",
)
BACKGROUND_JOB_STALE_RUNNING_MINUTES = 30
DEFAULT_FUSIONSOLAR_SYNC_HOURS = "08:00,14:00"
DEFAULT_STATE_SYNC_INTERVAL_HOURS = 1
DEFAULT_FUSIONSOLAR_PRODUCTION_SYNC_TIME = "00:10"
DEFAULT_FUSIONSOLAR_DIAGNOSTICS_SYNC_TIME = "00:30"
DEFAULT_FUSIONSOLAR_LOGIN_ENDPOINT = "/thirdData/login"
DEFAULT_FUSIONSOLAR_STATIONS_ENDPOINT = "/thirdData/stations"
DEFAULT_FUSIONSOLAR_REALTIME_ENDPOINT = "/thirdData/getStationRealKpi"
DEFAULT_FUSIONSOLAR_DEVICES_ENDPOINT = "/thirdData/getDevList"
DEFAULT_FUSIONSOLAR_DEVICE_REALTIME_ENDPOINT = "/thirdData/getDevRealKpi"
DEFAULT_FUSIONSOLAR_DEVICE_HISTORY_ENDPOINT = "/thirdData/getDevHistoryKpi"
DEFAULT_FUSIONSOLAR_ALARMS_ENDPOINT = "/thirdData/getAlarmList"
DEFAULT_FUSIONSOLAR_DAY_KPI_ENDPOINT = "/thirdData/getKpiStationDay"
DEFAULT_FUSIONSOLAR_MONTH_KPI_ENDPOINT = "/thirdData/getKpiStationMonth"
DEFAULT_FUSIONSOLAR_ALARMS_LANGUAGE = "en_US"
DEFAULT_SIGENERGY_BASE_URL = "https://api-eu.sigencloud.com"
DEFAULT_SIGENERGY_AUTH_ENDPOINT = "/openapi/auth/login/key"
DEFAULT_SIGENERGY_SYSTEMS_ENDPOINT = "/openapi/system"
DEFAULT_SIGENERGY_ENERGY_FLOW_ENDPOINT = "/openapi/systems/{system_id}/energyFlow"
DEFAULT_SIGENERGY_ONBOARD_ENDPOINT = "/openapi/board/onboard"
DEFAULT_SIGENERGY_REGION = "eu"
DEFAULT_SIGENERGY_SYNC_HOURS = "08:00,14:00"
DEFAULT_SIGENERGY_SNAPSHOT_RETENTION_DAYS = 90
FUSIONSOLAR_PERFORMANCE_RATE_LIMIT_MINUTES = 60
FUSIONSOLAR_PERFORMANCE_KPI_DELAY_SECONDS = 65
DEFAULT_FUSIONSOLAR_PRODUCTION_KPI_DAILY_BUDGET = 20
DEFAULT_FUSIONSOLAR_PRODUCTION_KPI_DAILY_RESERVED_CALLS = 2
DEFAULT_FUSIONSOLAR_PRODUCTION_KPI_MONTH_CLOSE_RESERVED_CALLS = 2
DEFAULT_FUSIONSOLAR_PRODUCTION_KPI_MIN_INTERVAL_SECONDS = 65
DEFAULT_FUSIONSOLAR_WAT_DAILY_BUDGET = 36
DEFAULT_FUSIONSOLAR_REALTIME_SNAPSHOT_RETENTION_DAYS = 30
FUSIONSOLAR_PERFORMANCE_MAX_API_CALLS = 20
FUSIONSOLAR_PERFORMANCE_RATE_LIMIT_UNTIL: datetime | None = None
DEFAULT_FUSIONSOLAR_RATE_LIMIT_MINUTES = 60
DEFAULT_FUSIONSOLAR_SESSION_CACHE_MINUTES = 55
DEFAULT_DEVICE_COMMUNICATION_THRESHOLD_MINUTES = 15
FUSIONSOLAR_INVERTER_DEVICE_TYPE_IDS = {1, 38}
INVERTER_AVAILABILITY_SLOT_MINUTES = 15
INVERTER_AVAILABILITY_EDGE_TOLERANCE_MINUTES = 30
LOW_INVERTER_AVAILABILITY_PCT = 90.0
LISBON_TIMEZONE = ZoneInfo("Europe/Lisbon")
DEFAULT_STRING_PRESENT_VOLTAGE_THRESHOLD = 100.0
DEFAULT_STRING_AUTO_LEARN_OBSERVATIONS = 2
SIGENERGY_SYNC_LOCK = threading.Lock()

STATUS_COLORS = {
    "Erro": "danger",
    "Desconectada": "warning",
    "Resolvido": "success",
    "Operacional": "success",
    "OK": "success",
    "AtenÃ§Ã£o": "warning",
    "Alerta": "warning",
    "CrÃ­tico": "danger",
    "Sem referÃªncia": "muted",
    "Sem dados": "muted",
    "Aberto": "danger",
    "Em analise": "warning",
    "Agendado": "accent",
    "Em visita": "accent",
    "Fechado": "muted",
}

TICKET_STATUSES = ["Aberto", "Em analise", "Agendado", "Em visita", "Resolvido", "Fechado"]
TICKET_URGENCIES = ["Baixa", "Media", "Alta", "Critica"]
TICKET_MATERIAL_STATUSES = ["Nao definido", "Sem material", "Necessario", "Pronto", "Bloqueado"]
TICKET_WORK_TYPES = ["Diagnostico", "Comunicacao", "Inversor", "String", "Estrutura", "Limpeza", "Preventiva", "Outro"]
MONTH_NAMES_PT = [
    "",
    "Janeiro",
    "Fevereiro",
    "MarÃ§o",
    "Abril",
    "Maio",
    "Junho",
    "Julho",
    "Agosto",
    "Setembro",
    "Outubro",
    "Novembro",
    "Dezembro",
]
MONITORING_SOURCES = ["FusionSolar", "Sigenergy", "Manual / Outro"]
ASSET_MONITORING_STATUSES = ["active", "silenced", "maintenance", "out_of_scope", "disabled"]
OK_MONITORING_STATUSES = {"Operacional", "Resolvido", "OK"}
PROBLEM_MONITORING_STATUSES = {"Erro", "Desconectada"}
ALERT_SCOPE_OPTIONS = ["all_assets", "only_o&m", "only_active_contracts", "only_selected_assets"]
ALERT_SETTING_DEFAULTS = {
    "TELEGRAM_ALERTS_ENABLED": "true",
    "ALERT_SCOPE": "only_o&m",
    "SEND_NEW_ERROR_ALERTS": "true",
    "SEND_OFFLINE_ALERTS": "true",
    "SEND_RESOLVED_ALERTS": "true",
    "SEND_PERSISTENT_ALERTS": "true",
    "SEND_RECURRENT_ALERTS": "false",
    "DAYTIME_OFFLINE_ONLY": "true",
    "IGNORE_HISTORICAL_ALERTS": "true",
    "MINIMUM_ALERT_SEVERITY": "info",
    "NEW_ERROR_COOLDOWN_MINUTES": "0",
    "OFFLINE_COOLDOWN_MINUTES": "120",
    "RESOLVED_COOLDOWN_MINUTES": "0",
    "PERSISTENT_COOLDOWN_HOURS": "24",
    "RECURRENT_COOLDOWN_HOURS": "24",
    "ALERT_BASELINE_AT": "",
}
RENEWAL_STATUSES = ["Por contactar", "Email enviado", "Em negociacao", "Renovado", "Sem interesse"]
INTEGRATION_STATUS_COLORS = {
    "success": "success",
    "error": "danger",
    "warning": "warning",
    "pending": "accent",
}
EXPORT_DATASETS = {
    "assets": {
        "label": "Instalacoes / centrais",
        "columns": [
            ("project_name", "Central"),
            ("location", "Localizacao"),
            ("address", "Morada"),
            ("contact_phone", "Contacto"),
            ("contact_name", "Nome"),
            ("access_type", "Acesso"),
            ("coverage_type", "Tipo de cobertura"),
            ("contract_type", "Contrato"),
            ("active_contract", "O&M"),
            ("company_name", "Empresa"),
            ("contact_email", "Email"),
        ],
    },
    "monitoring": {
        "label": "Monitorizacao filtrada",
        "columns": [
            ("record_date", "Data"),
            ("imported_at", "Importado em"),
            ("project_name", "Central"),
            ("location", "Localizacao"),
            ("contract_type", "Contrato"),
            ("active_contract", "O&M"),
            ("status", "Estado"),
            ("notes", "Notas"),
            ("source", "Origem"),
        ],
    },
    "tickets": {
        "label": "Intervencoes O&M",
        "columns": [
            ("project_name", "Central"),
            ("location", "Localizacao"),
            ("contract_type", "Contrato"),
            ("active_contract", "O&M"),
            ("title", "Titulo"),
            ("status", "Estado"),
            ("urgency", "Urgencia"),
            ("installation_ref", "Referencia"),
            ("next_action", "Proxima acao"),
            ("work_type", "Tipo de trabalho"),
            ("material_status", "Material"),
            ("planned_date", "Data planeada"),
            ("due_date", "Data limite"),
            ("estimated_minutes", "Minutos previstos"),
            ("assigned_to", "Equipa"),
            ("planning_notes", "Notas planeamento"),
            ("notes", "Notas"),
            ("created_at", "Criado em"),
            ("updated_at", "Atualizado em"),
        ],
    },
    "executive_report": {
        "label": "Relatorio executivo O&M",
        "columns": [
            ("section", "Seccao"),
            ("priority", "Prioridade"),
            ("project_name", "Central"),
            ("status", "Estado"),
            ("problem_days", "Dias em problema"),
            ("recurrence_count", "Recorrencias 90d"),
            ("open_tickets", "Tickets abertos"),
            ("source", "Origem"),
            ("notes", "Notas"),
        ],
    },
    "monitoring_report": {
        "label": "Relatorio limpo de monitorizacao",
        "columns": [
            ("period", "Periodo"),
            ("project_name", "Instalacao"),
            ("location", "Localizacao"),
            ("current_status", "Estado atual"),
            ("last_record_date", "Ultima monitorizacao"),
            ("monitoring_records", "Registos no periodo"),
            ("error_records", "Erros no periodo"),
            ("distinct_errors", "Erros diferentes"),
            ("error_types", "Tipos de erro"),
            ("open_tickets", "Tickets abertos"),
            ("visits_period", "Visitas no periodo"),
            ("last_visit_date", "Ultima visita"),
            ("latest_notes", "Notas"),
        ],
    },
    "production_report": {
        "label": "Relatorio de producao mensal/anual",
        "columns": [
            ("period", "Periodo"),
            ("project_name", "Instalacao"),
            ("location", "Localizacao"),
            ("provider", "Origem API"),
            ("production_kwh", "Producao kWh"),
            ("specific_yield", "kWh/kWp"),
            ("expected_kwh", "Producao esperada kWh"),
            ("deviation_pct", "Desvio %"),
            ("performance_status", "Estado performance"),
            ("data_points", "Pontos de dados"),
            ("data_source", "Tipo de dados"),
            ("last_update", "Ultima atualizacao"),
            ("notes", "Notas"),
        ],
    },
}

GROUP_INHERITED_FIELDS = [
    "company_name",
    "location",
    "address",
    "contract_type",
    "contact_name",
    "contact_role",
    "contact_email",
    "contact_phone",
    "access_type",
    "coverage_type",
]

SCHEDULER: BackgroundScheduler | None = None
FUSIONSOLAR_SESSION_CACHE: dict[str, Any] = {}
FUSIONSOLAR_SESSION_LOCK = threading.Lock()
FUSIONSOLAR_SYNC_LOCK = threading.Lock()
PRODUCTION_KPI_CALL_CONTEXT: ContextVar[dict[str, Any] | None] = ContextVar(
    "production_kpi_call_context",
    default=None,
)


@contextmanager
def release_fusionsolar_sync_lock() -> Any:
    try:
        yield
    finally:
        FUSIONSOLAR_SYNC_LOCK.release()


def create_app() -> Flask:
    ensure_runtime_directories(RUNTIME_PATHS)
    app = Flask(
        __name__,
        template_folder=str(BASE_DIR / "templates"),
        static_folder=str(BASE_DIR / "static"),
    )
    app.config["SECRET_KEY"] = flask_secret_key()
    app.config["DATA_DIR"] = str(RUNTIME_PATHS.data_dir)
    app.config["DATABASE"] = str(DB_PATH)
    app.config["EXCEL_PATH"] = str(DEFAULT_EXCEL_PATH) if DEFAULT_EXCEL_PATH else ""
    app.config["MAX_CONTENT_LENGTH"] = max_upload_bytes()
    app.config["SESSION_COOKIE_HTTPONLY"] = True
    app.config["SESSION_COOKIE_SAMESITE"] = os.environ.get("SESSION_COOKIE_SAMESITE", "Lax").strip() or "Lax"
    app.config["SESSION_COOKIE_SECURE"] = env_flag("SESSION_COOKIE_SECURE", False)
    app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=12)
    configure_logging(app, LOG_DIR)
    app.logger.info("Using database at %s", app.config["DATABASE"])
    app.register_blueprint(auth_bp)
    if not app_password_configured():
        app.logger.warning("APP_PASSWORD_HASH/APP_PASSWORD is not configured; login is locked until .env is updated.")

    ensure_database(app.config["DATABASE"])
    app.register_blueprint(field_routes_bp)
    with closing(get_db(app.config["DATABASE"])) as bootstrap_conn:
        populate_missing_installation_groups(bootstrap_conn)
        populate_missing_group_metadata(bootstrap_conn)
        sync_all_contract_statuses(bootstrap_conn)
        ensure_integration_seed_data(bootstrap_conn)
        bootstrap_conn.commit()
    start_integration_scheduler(app)

    @app.before_request
    def before_request() -> None:
        g.db = get_db(app.config["DATABASE"])
        g.request_started_at = datetime.now()
        if request.method == "POST":
            sent_token = request.form.get("csrf_token") or request.headers.get("X-CSRF-Token", "")
            if not sent_token or not secrets.compare_digest(sent_token, csrf_token()):
                app.logger.warning("CSRF validation failed for %s %s", request.method, request.path)
                abort(400)
        if request.endpoint not in {"auth.login", "static"} and not session.get("authenticated"):
            return redirect(url_for("auth.login", next=request.full_path if request.query_string else request.path))

    @app.teardown_request
    def teardown_request(exception: BaseException | None) -> None:
        started_at = getattr(g, "request_started_at", None)
        elapsed_ms = ""
        if started_at:
            elapsed_ms = f" {(datetime.now() - started_at).total_seconds() * 1000:.0f}ms"
        if request.endpoint != "static":
            if exception:
                app.logger.exception("%s %s failed%s", request.method, request.path, elapsed_ms)
            else:
                app.logger.info("%s %s%s", request.method, request.path, elapsed_ms)
        db = g.pop("db", None)
        if db is not None:
            db.close()

    @app.context_processor
    def inject_globals() -> dict[str, Any]:
        return {
            "today_iso": date.today().isoformat(),
            "ticket_statuses": TICKET_STATUSES,
            "ticket_urgencies": TICKET_URGENCIES,
            "ticket_material_statuses": TICKET_MATERIAL_STATUSES,
            "ticket_work_types": TICKET_WORK_TYPES,
            "status_colors": STATUS_COLORS,
            "monitoring_sources": MONITORING_SOURCES,
            "asset_monitoring_statuses": ASSET_MONITORING_STATUSES,
            "renewal_statuses": RENEWAL_STATUSES,
            "integration_status_colors": INTEGRATION_STATUS_COLORS,
            "om_status_label": om_status_label,
            "format_date_pt": format_date_pt,
            "format_number": format_number,
            "compute_performance_percentage": compute_performance_percentage,
            "performance_bar_width": performance_bar_width,
            "performance_status_class": performance_status_class,
            "reference_diagnostic": reference_diagnostic,
            "csrf_token": csrf_token,
            "current_username": session.get("username"),
        }

    @app.errorhandler(400)
    def bad_request_error(error: Exception) -> tuple[str, int]:
        return render_template(
            "error.html",
            title="Pedido invalido",
            heading="Pedido invalido",
            message="A acao nao foi aceite. Atualiza a pagina e tenta novamente.",
        ), 400

    @app.errorhandler(404)
    def not_found_error(error: Exception) -> tuple[str, int]:
        return render_template(
            "error.html",
            title="Pagina nao encontrada",
            heading="Pagina nao encontrada",
            message="Nao encontrei esta pagina ou registo.",
        ), 404

    @app.errorhandler(500)
    def internal_error(error: Exception) -> tuple[str, int]:
        current_app.logger.exception("Unhandled application error")
        return render_template(
            "error.html",
            title="Erro interno",
            heading="Erro interno",
            message="Aconteceu um erro inesperado. Consulta os logs para o detalhe tecnico.",
        ), 500

    @app.errorhandler(413)
    def request_too_large_error(error: Exception) -> tuple[str, int]:
        return render_template(
            "error.html",
            title="Ficheiro demasiado grande",
            heading="Ficheiro demasiado grande",
            message="O ficheiro enviado excede o limite configurado para uploads.",
        ), 413

    @app.route("/operation")
    def operation() -> str:
        search = request.args.get("search", "").strip()
        om_only = request.args.get("om_only", "yes").strip()
        calendar_month = normalize_calendar_month(request.args.get("calendar_month", ""))
        calendar_start, calendar_end, previous_month, next_month = calendar_month_bounds(calendar_month)

        intervention_conditions = ["t.status != 'Fechado'"]
        intervention_params: list[Any] = []
        if om_only == "yes":
            intervention_conditions.append("a.active_contract = 'yes'")
        if search:
            wildcard = f"%{search}%"
            intervention_conditions.append(
                "(a.project_name LIKE ? OR a.location LIKE ? OR a.address LIKE ? OR t.title LIKE ? OR COALESCE(t.next_action, '') LIKE ?)"
            )
            intervention_params.extend([wildcard, wildcard, wildcard, wildcard, wildcard])
        intervention_where = " AND ".join(intervention_conditions)
        interventions = query_all(
            g.db,
            f"""
            SELECT
                t.*,
                a.project_name,
                a.installation_group,
                a.location,
                a.address,
                a.active_contract,
                a.latitude,
                a.longitude,
                a.coordinates_confidence,
                lm.status AS latest_status,
                lm.record_date AS latest_status_date
            FROM tickets t
            JOIN assets a ON a.id = t.asset_id
            LEFT JOIN latest_monitoring_view lm ON lm.asset_id = a.id
            WHERE {intervention_where}
            ORDER BY
                CASE t.urgency WHEN 'Critica' THEN 1 WHEN 'Alta' THEN 2 WHEN 'Media' THEN 3 ELSE 4 END,
                CASE WHEN COALESCE(t.planned_date, '') = '' THEN 0 ELSE 1 END,
                COALESCE(t.planned_date, '9999-12-31') ASC,
                t.updated_at DESC
            """,
            intervention_params,
        )

        problem_conditions = [
            "COALESCE(a.monitoring_status, 'active') != 'disabled'",
            "lm.status IN ('Erro', 'Desconectada')",
            "COALESCE(t.open_tickets, 0) = 0",
        ]
        problem_params: list[Any] = []
        if om_only == "yes":
            problem_conditions.append("a.active_contract = 'yes'")
        if search:
            wildcard = f"%{search}%"
            problem_conditions.append("(a.project_name LIKE ? OR a.location LIKE ? OR a.address LIKE ?)")
            problem_params.extend([wildcard, wildcard, wildcard])
        problems_without_action = enrich_operational_rows(
            g.db,
            query_all(
                g.db,
                f"""
                SELECT
                    a.id AS asset_id,
                    a.project_name,
                    a.installation_group,
                    a.location,
                    a.address,
                    a.active_contract,
                    a.latitude,
                    a.longitude,
                    a.coordinates_confidence,
                    lm.status,
                    lm.record_date,
                    0 AS open_tickets
                FROM assets a
                JOIN latest_monitoring_view lm ON lm.asset_id = a.id
                LEFT JOIN (
                    SELECT asset_id, COUNT(*) AS open_tickets
                    FROM tickets
                    WHERE status != 'Fechado'
                    GROUP BY asset_id
                ) t ON t.asset_id = a.id
                WHERE {" AND ".join(problem_conditions)}
                ORDER BY
                    CASE lm.status WHEN 'Erro' THEN 1 WHEN 'Desconectada' THEN 2 ELSE 3 END,
                    lm.record_date ASC,
                    a.project_name COLLATE NOCASE
                LIMIT 20
                """,
                problem_params,
            ),
        )

        planned_rows = query_all(
            g.db,
            f"""
            SELECT
                t.*,
                a.project_name,
                a.location,
                a.latitude,
                a.longitude,
                a.coordinates_confidence,
                lm.status AS latest_status
            FROM tickets t
            JOIN assets a ON a.id = t.asset_id
            LEFT JOIN latest_monitoring_view lm ON lm.asset_id = a.id
            WHERE t.status != 'Fechado'
              AND COALESCE(t.planned_date, '') BETWEEN ? AND ?
              {"AND a.active_contract = 'yes'" if om_only == "yes" else ""}
            ORDER BY t.planned_date ASC, t.urgency DESC, a.project_name COLLATE NOCASE
            """,
            [calendar_start.isoformat(), calendar_end.isoformat()],
        )
        planning_calendar = build_intervention_calendar(calendar_month, planned_rows)

        today = date.today().isoformat()
        week_end = (date.today() + timedelta(days=7)).isoformat()
        operation_stats = {
            "open": len(interventions),
            "critical": sum(1 for row in interventions if row["urgency"] == "Critica"),
            "unplanned": sum(1 for row in interventions if not row["planned_date"]),
            "blocked": sum(1 for row in interventions if row["material_status"] == "Bloqueado"),
            "this_week": sum(1 for row in interventions if today <= (row["planned_date"] or "") <= week_end),
            "without_action": len(problems_without_action),
            "ready_for_route": sum(1 for row in interventions if intervention_ready_for_route(row)),
        }

        return render_template(
            "operation.html",
            title="Operacao O&M",
            search=search,
            om_only=om_only,
            interventions=interventions,
            problems_without_action=problems_without_action,
            operation_stats=operation_stats,
            planning_calendar=planning_calendar,
            calendar_month=calendar_month,
            previous_month=previous_month,
            next_month=next_month,
            today=today,
            week_end=week_end,
        )

    @app.route("/")
    def dashboard() -> str:
        stats = fetch_dashboard_stats(g.db)
        availability_summary = get_dashboard_availability_summary(g.db)
        monitoring_by_day = query_all(
            g.db,
            """
            SELECT record_date, COUNT(*) AS total
            FROM monitoring_records
            GROUP BY record_date
            ORDER BY record_date DESC
            LIMIT 7
            """,
        )
        critical_assets = enrich_operational_rows(g.db, query_all(
            g.db,
            """
            SELECT
                a.id AS asset_id,
                a.project_name,
                a.active_contract,
                lm.status,
                lm.record_date,
                COUNT(t.id) AS open_tickets
            FROM assets a
            LEFT JOIN latest_monitoring_view lm ON lm.asset_id = a.id
            LEFT JOIN tickets t ON t.asset_id = a.id AND t.status != 'Fechado'
            WHERE a.active_contract = 'yes'
              AND COALESCE(a.monitoring_status, 'active') != 'disabled'
              AND (lm.status IN ('Erro', 'Desconectada') OR t.id IS NOT NULL)
            GROUP BY a.id, a.project_name, a.active_contract, lm.status, lm.record_date
            ORDER BY
                CASE lm.status WHEN 'Erro' THEN 1 WHEN 'Desconectada' THEN 2 ELSE 3 END,
                open_tickets DESC,
                a.project_name COLLATE NOCASE
            LIMIT 12
            """,
        ))
        critical_assets.sort(
            key=lambda row: (
                priority_rank(row["auto_priority"]),
                -int(row.get("problem_days") or 0),
                -int(row.get("recurrence_count") or 0),
                row["project_name"].lower(),
            )
        )
        potential_assets = query_all(
            g.db,
            """
            SELECT
                a.id,
                a.project_name,
                a.active_contract,
                lm.status,
                lm.record_date
            FROM assets a
            LEFT JOIN latest_monitoring_view lm ON lm.asset_id = a.id
            WHERE COALESCE(a.active_contract, '') != 'yes'
              AND COALESCE(a.monitoring_status, 'active') != 'disabled'
              AND lm.status IN ('Erro', 'Desconectada')
            ORDER BY
                CASE lm.status WHEN 'Erro' THEN 1 WHEN 'Desconectada' THEN 2 ELSE 3 END,
                a.project_name COLLATE NOCASE
            LIMIT 10
            """,
        )
        open_ticket_assets = query_all(
            g.db,
            """
            SELECT
                a.id,
                a.project_name,
                a.active_contract,
                COUNT(t.id) AS ticket_count,
                SUM(CASE WHEN t.urgency = 'Critica' THEN 1 ELSE 0 END) AS critical_count,
                MAX(t.updated_at) AS last_update
            FROM assets a
            JOIN tickets t ON t.asset_id = a.id
            WHERE t.status != 'Fechado' AND a.active_contract = 'yes'
            GROUP BY a.id, a.project_name, a.active_contract
            ORDER BY critical_count DESC, ticket_count DESC, a.project_name COLLATE NOCASE
            LIMIT 10
            """,
        )
        renewal_focus = query_all(
            g.db,
            """
            SELECT
                a.id,
                a.project_name,
                a.installation_group,
                a.company_name,
                COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) AS contract_end_date,
                COALESCE(oc.renewal_status, 'Por contactar') AS renewal_status
            FROM assets a
            LEFT JOIN om_contracts oc ON oc.asset_id = a.id
            WHERE (a.maintenance = 'yes' OR oc.id IS NOT NULL)
              AND COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) NOT IN ('', '-')
              AND (
                  COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) < ?
                  OR substr(COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')), 1, 4) = ?
              )
            ORDER BY COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) ASC, a.project_name COLLATE NOCASE
            LIMIT 10
            """,
            (date.today().isoformat(), str(date.today().year)),
        )
        executive_stats = build_executive_dashboard_stats(g.db)
        executive_priorities = critical_assets[:8]
        integration_summary = build_integration_summary(g.db)
        performance_risk_count = query_scalar(
            g.db,
            """
            SELECT COUNT(*)
            FROM availability_daily ad
            JOIN (
                SELECT asset_id, MAX(period_date || 'T' || printf('%09d', id)) AS marker
                FROM availability_daily
                GROUP BY asset_id
            ) latest
              ON latest.asset_id = ad.asset_id
             AND latest.marker = ad.period_date || 'T' || printf('%09d', ad.id)
            WHERE COALESCE(ad.unavailable_inverters, 0) > 0
               OR COALESCE(ad.no_communication_devices, 0) > 0
            """,
        )
        return render_template(
            "dashboard.html",
            stats=stats,
            executive_stats=executive_stats,
            executive_priorities=executive_priorities,
            integration_summary=integration_summary,
            availability_summary=availability_summary,
            monitoring_by_day=monitoring_by_day,
            critical_assets=critical_assets,
            potential_assets=potential_assets,
            open_ticket_assets=open_ticket_assets,
            renewal_focus=renewal_focus,
            performance_risk_count=performance_risk_count,
        )

    @app.route("/performance", methods=["GET", "POST"])
    def performance() -> str:
        if request.method == "POST":
            action = request.form.get("action", "sync_availability").strip()
            if action == "sync_availability":
                result = run_fusionsolar_device_availability_sync(
                    g.db,
                    INTEGRATION_PROVIDER_FUSIONSOLAR,
                    trigger_type="manual",
                )
                flash(
                    f"Disponibilidade sincronizada: {result['devices']} dispositivos, {result['assets']} centrais.",
                    "success",
                )
                return redirect(url_for("performance"))
            if action == "sync_inverter_time_availability":
                from_date = parse_date_value(request.form.get("from_date", ""))
                to_date = parse_date_value(request.form.get("to_date", ""))
                search = request.form.get("search", "").strip()
                om_only = request.form.get("om_only", "yes").strip()
                if not from_date or not to_date or from_date > to_date:
                    flash("Intervalo invalido para disponibilidade dos inversores.", "error")
                    return redirect(url_for("performance"))
                if to_date >= date.today():
                    flash("A disponibilidade temporal so pode ser calculada para dias fechados.", "error")
                    return redirect(url_for("performance"))
                if (to_date - from_date).days > 31:
                    flash("Calcula no maximo 32 dias de cada vez.", "error")
                    return redirect(url_for("performance"))
                job_id, created = create_background_job(
                    g.db,
                    "fusionsolar_inverter_availability_backfill",
                    {
                        "provider": INTEGRATION_PROVIDER_FUSIONSOLAR,
                        "from_date": from_date.isoformat(),
                        "to_date": to_date.isoformat(),
                    },
                )
                g.db.commit()
                if created:
                    schedule_background_job(current_app._get_current_object(), job_id)
                    flash(f"Calculo WAT enviado para background (job #{job_id}).", "success")
                else:
                    flash(f"Ja existe um calculo WAT pendente/em execucao (job #{job_id}).", "warning")
                return redirect(
                    url_for(
                        "performance",
                        period="custom",
                        from_date=from_date,
                        to_date=to_date,
                        search=search,
                        om_only=om_only,
                    )
                )
            flash("Acao de disponibilidade invalida.", "error")
            return redirect(url_for("performance"))

        asset_id = request.args.get("asset_id", "").strip()
        search = request.args.get("search", "").strip()
        om_only = request.args.get("om_only", "yes").strip()
        availability_period = request.args.get("period", "yesterday").strip()
        availability_from, availability_to = resolve_inverter_availability_period(
            availability_period,
            request.args.get("from_date", ""),
            request.args.get("to_date", ""),
        )

        inverter_time_report = get_inverter_availability_report(
            g.db,
            availability_from,
            availability_to,
            asset_id=int(asset_id) if asset_id.isdigit() else None,
            om_only=om_only == "yes",
            search=search,
        )
        inverter_chart_report = (
            get_inverter_availability_chart_report(
                g.db,
                int(asset_id),
                availability_from,
                availability_to,
            )
            if asset_id.isdigit()
            else None
        )
        return render_template(
            "performance.html",
            selected_asset_id=asset_id,
            search=search,
            om_only=om_only,
            inverter_time_report=inverter_time_report,
            inverter_chart_report=inverter_chart_report,
            availability_period=availability_period,
            availability_from=availability_from,
            availability_to=availability_to,
            availability_closed_max=date.today() - timedelta(days=1),
            background_jobs=fetch_latest_background_jobs(
                g.db,
                job_types=("fusionsolar_inverter_availability_backfill",),
            ),
            fusionsolar_api_warning=get_fusionsolar_performance_cooldown_reason(g.db),
        )

    @app.route("/performance/debug/<int:record_id>")
    def performance_debug(record_id: int) -> str:
        record = query_one(
            """
            SELECT pr.*, a.project_name
            FROM production_records pr
            JOIN assets a ON a.id = pr.asset_id
            WHERE pr.id = ?
            """,
            (record_id,),
        )
        if record is None:
            flash("Registo de performance nao encontrado.", "error")
            return redirect(url_for("performance"))

        raw_payload = record["payload_json"] or "{}"
        try:
            parsed_payload = json.loads(raw_payload)
            pretty_payload = json.dumps(parsed_payload, indent=2, ensure_ascii=False)
        except json.JSONDecodeError:
            pretty_payload = raw_payload

        return render_template(
            "performance_debug.html",
            record=record,
            pretty_payload=pretty_payload,
        )

    @app.route("/performance/backfill", methods=["GET", "POST"])
    def performance_backfill() -> str:
        current_year = date.today().year
        period_type = request.values.get("period_type", "day").strip()
        if period_type not in {"day", "month"}:
            period_type = "day"
        from_year = int(request.values.get("from_year", current_year - 1) or current_year - 1)
        to_year = int(request.values.get("to_year", current_year) or current_year)
        date_from_raw = request.values.get("date_from", "").strip()
        date_to_raw = request.values.get("date_to", "").strip()
        date_from = parse_date_value(date_from_raw)
        date_to = parse_date_value(date_to_raw)
        max_api_calls_raw = request.values.get("max_api_calls", request.values.get("max_days", str(FUSIONSOLAR_PERFORMANCE_MAX_API_CALLS))).strip()
        max_api_calls = int(max_api_calls_raw) if max_api_calls_raw.isdigit() else FUSIONSOLAR_PERFORMANCE_MAX_API_CALLS
        asset_id_raw = request.values.get("asset_id", "").strip()
        asset_id = int(asset_id_raw) if asset_id_raw.isdigit() else None
        estimated_station_count = 0
        estimated_api_calls = 0
        if request.method == "GET":
            estimated_assets = get_fusionsolar_performance_assets(g.db, INTEGRATION_PROVIDER_FUSIONSOLAR, asset_id)
            estimated_station_codes = [str(asset["external_id"] or "").strip() for asset in estimated_assets if str(asset["external_id"] or "").strip()]
            estimated_station_count = len(estimated_station_codes)
            estimated_chunks = len(chunked(estimated_station_codes, 100))
            estimated_periods = (
                len(iter_daily_backfill_months(from_year, to_year, today_value=date.today(), date_from=date_from, date_to=date_to))
                if period_type == "day"
                else len(iter_monthly_backfill_dates(from_year, to_year, today_value=date.today()))
            )
            estimated_api_calls = estimated_periods * estimated_chunks

        if request.method == "POST" and request.form.get("action") == "month_cycle":
            cycle_month = normalize_report_month(request.form.get("cycle_month", ""))
            cycle_asset_ids = [int(value) for value in request.form.getlist("cycle_asset_ids") if value.isdigit()]
            if not cycle_asset_ids:
                flash("Escolhe pelo menos uma instalacao para o ciclo.", "error")
                return redirect(url_for("performance_backfill", period_type=period_type))
            job_id, created = create_background_job(
                g.db,
                "fusionsolar_month_cycle",
                {
                    "provider": INTEGRATION_PROVIDER_FUSIONSOLAR,
                    "report_month": cycle_month,
                    "asset_ids": cycle_asset_ids,
                },
            )
            g.db.commit()
            if created:
                schedule_background_job(current_app._get_current_object(), job_id)
                flash(f"Ciclo mensal FusionSolar enviado para background (job #{job_id}).", "success")
            else:
                flash(f"Ja existe um ciclo mensal FusionSolar pendente/em execucao (job #{job_id}).", "warning")
            return redirect(url_for("performance_backfill", period_type=period_type, cycle_month=cycle_month))

        if request.method == "POST":
            job_id, created = create_background_job(
                g.db,
                "fusionsolar_production_backfill",
                {
                    "provider": INTEGRATION_PROVIDER_FUSIONSOLAR,
                    "period_type": period_type,
                    "from_year": from_year,
                    "to_year": to_year,
                    "asset_id": asset_id,
                    "date_from": date_from.isoformat() if date_from else "",
                    "date_to": date_to.isoformat() if date_to else "",
                    "max_api_calls": max_api_calls,
                },
            )
            g.db.commit()
            if created:
                schedule_background_job(current_app._get_current_object(), job_id)
                flash(f"Backfill historico enviado para background (job #{job_id}).", "success")
            else:
                flash(f"Ja existe um backfill historico pendente/em execucao (job #{job_id}).", "warning")
            return redirect(
                url_for(
                    "performance_backfill",
                    period_type=period_type,
                    from_year=from_year,
                    to_year=to_year,
                    date_from=date_from.isoformat() if date_from else date_from_raw,
                    date_to=date_to.isoformat() if date_to else date_to_raw,
                    max_api_calls=max_api_calls,
                    asset_id=asset_id or "",
                )
            )

        assets_for_backfill = get_fusionsolar_performance_assets(g.db, INTEGRATION_PROVIDER_FUSIONSOLAR)
        return render_template(
            "performance_backfill.html",
            period_type=period_type,
            from_year=from_year,
            to_year=to_year,
            date_from=date_from.isoformat() if date_from else date_from_raw,
            date_to=date_to.isoformat() if date_to else date_to_raw,
            max_api_calls=max_api_calls,
            estimated_api_calls=estimated_api_calls,
            estimated_station_count=estimated_station_count,
            selected_asset_id=asset_id,
            assets_for_backfill=assets_for_backfill,
            background_jobs=fetch_latest_background_jobs(g.db, job_types=BACKGROUND_JOB_TYPES_PERFORMANCE),
            fusionsolar_api_warning=get_fusionsolar_performance_cooldown_reason(g.db),
            current_year=current_year,
            cycle_month=request.args.get("cycle_month", date.today().strftime("%Y-%m")),
        )

    @app.route("/assets", methods=["GET", "POST"])
    def assets() -> str:
        if request.method == "POST":
            project_name = request.form.get("project_name", "").strip()
            installation_group = request.form.get("installation_group", "").strip()
            company_name = request.form.get("company_name", "").strip()
            location = request.form.get("location", "").strip()
            address = request.form.get("address", "").strip()
            kwp = request.form.get("kwp", "").strip()
            contract_type = request.form.get("contract_type", "").strip()
            maintenance = request.form.get("maintenance", "").strip()
            active_contract = request.form.get("active_contract", "").strip()
            start_contract = request.form.get("start_contract", "").strip()
            end_contract = request.form.get("end_contract", "").strip()
            contact_name = request.form.get("contact_name", "").strip()
            contact_email = request.form.get("contact_email", "").strip()
            contact_phone = request.form.get("contact_phone", "").strip()
            notes = request.form.get("notes", "").strip()

            if not project_name:
                flash("O nome da instalacao/central e obrigatorio.", "error")
                return redirect(url_for("assets"))

            existing = query_one("SELECT id FROM assets WHERE project_name = ?", (project_name,))
            if existing is not None:
                flash("Ja existe um asset com esse nome.", "error")
                return redirect(url_for("assets"))

            final_group = installation_group or infer_installation_group(project_name)
            inherited_payload = apply_group_defaults(
                g.db,
                {
                    "company_name": company_name,
                    "location": location,
                    "address": address,
                    "contract_type": contract_type,
                    "contact_name": contact_name,
                    "contact_email": contact_email,
                    "contact_phone": contact_phone,
                },
                final_group,
            )
            company_name = inherited_payload["company_name"]
            location = inherited_payload["location"]
            address = inherited_payload["address"]
            contract_type = inherited_payload["contract_type"]
            contact_name = inherited_payload["contact_name"]
            contact_email = inherited_payload["contact_email"]
            contact_phone = inherited_payload["contact_phone"]
            start_contract = normalize_date_value(start_contract)
            end_contract = normalize_date_value(end_contract)
            active_contract = derive_active_contract(end_contract, active_contract)
            cursor = g.db.execute(
                """
                INSERT INTO assets (
                    project_name, installation_group, company_name, location, address, kwp, contract_type,
                    maintenance, active_contract, start_contract, end_contract, contact_name,
                    contact_email, contact_phone, notes, alias_blob
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    project_name,
                    final_group,
                    company_name,
                    location,
                    address,
                    kwp,
                    contract_type,
                    maintenance,
                    active_contract,
                    start_contract,
                    end_contract,
                    contact_name,
                    contact_email,
                    contact_phone,
                    notes,
                    project_name,
                ),
            )
            asset_id = int(cursor.lastrowid)
            normalized_name = normalize_name(project_name)
            if normalized_name:
                g.db.execute(
                    "INSERT OR IGNORE INTO asset_aliases (asset_id, alias_name, normalized_alias, source) VALUES (?, ?, ?, ?)",
                    (asset_id, project_name, normalized_name, "manual-create"),
                )
            g.db.commit()
            rebuild_asset_alias_blob(g.db, asset_id)
            flash("Instalacao criada com sucesso.", "success")
            return redirect(url_for("asset_detail", asset_id=asset_id))

        search = request.args.get("search", "").strip()
        contract_filter = request.args.get("contract_type", "").strip()
        om_filter = request.args.get("active_contract", "").strip()

        conditions = []
        params: list[Any] = []
        if search:
            conditions.append(
                "(a.project_name LIKE ? OR a.company_name LIKE ? OR a.location LIKE ? OR a.alias_blob LIKE ?)"
            )
            wildcard = f"%{search}%"
            params.extend([wildcard, wildcard, wildcard, wildcard])
        if contract_filter:
            conditions.append("a.contract_type = ?")
            params.append(contract_filter)
        if om_filter:
            conditions.append("a.active_contract = ?")
            params.append(om_filter)

        where_sql = f"WHERE {' AND '.join(conditions)}" if conditions else ""
        assets_rows = query_all(
            g.db,
            f"""
            SELECT
                a.*,
                lm.status AS latest_status,
                lm.record_date AS latest_status_date,
                (
                    SELECT COUNT(*)
                    FROM tickets t
                    WHERE t.asset_id = a.id AND t.status != 'Fechado'
                ) AS open_tickets
            FROM assets a
            LEFT JOIN latest_monitoring_view lm ON lm.asset_id = a.id
            {where_sql}
            ORDER BY a.installation_group COLLATE NOCASE, a.project_name COLLATE NOCASE
            """,
            params,
        )
        contract_types = [
            row["contract_type"]
            for row in query_all(
                g.db,
                "SELECT DISTINCT contract_type FROM assets WHERE contract_type != '' ORDER BY contract_type",
            )
        ]
        return render_template(
            "assets.html",
            assets=assets_rows,
            contract_types=contract_types,
            installation_groups=list_installation_group_options(g.db),
            search=search,
            contract_filter=contract_filter,
            om_filter=om_filter,
        )

    @app.route("/asset/<int:asset_id>")
    def asset_detail(asset_id: int) -> str:
        asset = query_one("SELECT * FROM assets WHERE id = ?", (asset_id,))
        if asset is None:
            flash("Asset nao encontrado.", "error")
            return redirect(url_for("assets"))
        calendar_month = normalize_calendar_month(request.args.get("calendar_month", ""))
        calendar_start, calendar_end, previous_month, next_month = calendar_month_bounds(calendar_month)

        om_contract = query_one(
            """
            SELECT *
            FROM om_contracts
            WHERE asset_id = ?
            """,
            (asset_id,),
        )

        monitoring_history = query_all(
            g.db,
            """
            SELECT id, record_date, status, notes, source
            FROM monitoring_records
            WHERE asset_id = ?
            ORDER BY record_date DESC, id DESC
            LIMIT 100
            """,
            (asset_id,),
        )
        calendar_history = query_all(
            g.db,
            """
            SELECT id, asset_id, record_date, status, notes, source
            FROM monitoring_records
            WHERE asset_id = ?
              AND record_date <= ?
            ORDER BY record_date ASC, id ASC
            """,
            (asset_id, calendar_end.isoformat()),
        )
        asset_error_calendar = build_asset_error_calendar(calendar_month, calendar_history)
        tickets = query_all(
            g.db,
            """
            SELECT *
            FROM tickets
            WHERE asset_id = ?
            ORDER BY updated_at DESC, created_at DESC
            """,
            (asset_id,),
        )
        aliases = query_all(
            g.db,
            """
            SELECT id, alias_name
            FROM asset_aliases
            WHERE asset_id = ?
            ORDER BY alias_name COLLATE NOCASE
            """,
            (asset_id,),
        )
        latest_daily_performance = query_one(
            """
            SELECT *
            FROM production_records
            WHERE asset_id = ? AND period_type = 'day'
            ORDER BY period_date DESC, id DESC
            LIMIT 1
            """,
            (asset_id,),
        )
        latest_monthly_performance = query_one(
            """
            SELECT *
            FROM production_records
            WHERE asset_id = ? AND period_type = 'month' AND period_date < ?
            ORDER BY period_date DESC, id DESC
            LIMIT 1
            """,
            (asset_id, date.today().replace(day=1).isoformat()),
        )
        latest_mtd_performance = query_one(
            """
            SELECT *
            FROM production_records
            WHERE asset_id = ? AND period_type = 'mtd' AND period_date = ?
            ORDER BY id DESC
            LIMIT 1
            """,
            (asset_id, date.today().replace(day=1).isoformat()),
        )
        performance_settings = get_performance_settings(g.db, asset_id)
        latest_availability = get_latest_availability_by_asset(g.db, asset_id)
        latest_device_rows = get_latest_device_rows_for_asset(g.db, asset_id)
        expected_string_rows = query_all(
            g.db,
            """
            SELECT pd.id AS provider_device_id, pes.string_index, pes.expected, pes.source
            FROM provider_devices pd
            LEFT JOIN provider_device_expected_strings pes ON pes.provider_device_id = pd.id
            WHERE pd.asset_id = ? AND pd.enabled = 1
            ORDER BY pd.device_name COLLATE NOCASE, pd.id, pes.string_index
            """,
            (asset_id,),
        )
        expected_strings_by_device: dict[int, list[dict[str, Any]]] = {}
        for row in expected_string_rows:
            if row["string_index"] is None:
                continue
            expected_strings_by_device.setdefault(int(row["provider_device_id"]), []).append(row)
        visits_by_ticket = build_visits_by_ticket(
            query_all(
                g.db,
                """
                SELECT *
                FROM ticket_visits
                WHERE ticket_id IN (
                    SELECT id FROM tickets WHERE asset_id = ?
                )
                ORDER BY visit_date DESC, id DESC
                """,
                (asset_id,),
            )
        )
        current_installation_group = asset["installation_group"] or asset["project_name"]
        group_members = query_all(
            g.db,
            """
            SELECT
                a.id,
                a.project_name,
                a.location,
                a.active_contract,
                lm.status AS latest_status,
                lm.record_date AS latest_status_date
            FROM assets a
            LEFT JOIN latest_monitoring_view lm ON lm.asset_id = a.id
            WHERE COALESCE(NULLIF(TRIM(a.installation_group), ''), a.project_name) = ?
            ORDER BY a.project_name COLLATE NOCASE
            """,
            (current_installation_group,),
        )
        financial_model = build_asset_financial_model_context(g.db, asset_id=asset_id)
        return render_template(
            "asset_detail.html",
            asset=asset,
            current_installation_group=current_installation_group,
            group_members=group_members,
            installation_groups=list_installation_group_options(g.db),
            om_contract=om_contract,
            monitoring_history=monitoring_history,
            asset_error_calendar=asset_error_calendar,
            calendar_month=calendar_month,
            previous_month=previous_month,
            next_month=next_month,
            tickets=tickets,
            aliases=aliases,
            visits_by_ticket=visits_by_ticket,
            latest_daily_performance=latest_daily_performance,
            latest_mtd_performance=latest_mtd_performance,
            latest_monthly_performance=latest_monthly_performance,
            performance_settings=performance_settings,
            latest_availability=latest_availability,
            latest_device_rows=latest_device_rows,
            expected_strings_by_device=expected_strings_by_device,
            financial_model=financial_model,
        )

    @app.route("/asset/<int:asset_id>/financial-model/upload", methods=["POST"])
    def upload_asset_financial_model(asset_id: int):
        upload = request.files.get("file")
        base_year_raw = request.form.get("base_year", "").strip()
        if upload is None or not upload.filename:
            flash("Escolhe um ficheiro financeiro .xlsx ou .xlsm.", "error")
            return redirect(url_for("asset_detail", asset_id=asset_id))
        try:
            model_id = create_financial_model_preview(
                g.db,
                upload_dir=UPLOAD_DIR,
                file_storage=upload,
                asset_id=asset_id,
                base_year=int(base_year_raw) if base_year_raw.isdigit() else None,
            )
            g.db.commit()
            return redirect(url_for("financial_model_preview", asset_id=asset_id, model_id=model_id))
        except (FinancialModelError, ValueError) as exc:
            g.db.rollback()
            flash(f"Falha ao analisar modelo financeiro: {exc}", "error")
            return redirect(url_for("asset_detail", asset_id=asset_id))

    @app.route("/asset/<int:asset_id>/financial-model/<int:model_id>/preview")
    def financial_model_preview(asset_id: int, model_id: int):
        model = get_financial_model_for_asset(g.db, asset_id=asset_id, model_id=model_id)
        if model is None:
            abort(404)
        source = get_financial_model_source(g.db, model_id)
        monthly = list_financial_model_monthly(g.db, model_id=model_id)
        return render_template(
            "financial_model_preview.html",
            title="Preview modelo financeiro",
            asset=query_one("SELECT * FROM assets WHERE id = ?", (asset_id,)),
            model=model,
            source=source,
            monthly=monthly,
            validation=json.loads(model["validation_json"] or "{}"),
            warnings=json.loads(model["warnings_json"] or "[]"),
            details=parse_model_details_json(model),
        )

    @app.route("/asset/<int:asset_id>/financial-model/<int:model_id>/confirm", methods=["POST"])
    def confirm_asset_financial_model(asset_id: int, model_id: int):
        try:
            version = confirm_financial_model_import(
                g.db,
                model_id=model_id,
                asset_id=asset_id,
                override=bool(request.form.get("override")),
                override_reason=request.form.get("override_reason", ""),
            )
            g.db.commit()
            flash(f"Modelo financeiro confirmado como versao {version}.", "success")
        except (FinancialModelError, ValueError) as exc:
            g.db.rollback()
            flash(f"Falha ao confirmar modelo financeiro: {exc}", "error")
            return redirect(url_for("financial_model_preview", asset_id=asset_id, model_id=model_id))
        return redirect(url_for("asset_detail", asset_id=asset_id))

    @app.route("/asset/<int:asset_id>/financial-model/<int:model_id>/cancel", methods=["POST"])
    def cancel_asset_financial_model(asset_id: int, model_id: int):
        try:
            cancel_financial_model_preview(g.db, model_id=model_id, asset_id=asset_id)
            g.db.commit()
            flash("Preview de modelo financeiro cancelado.", "success")
        except FinancialModelError as exc:
            g.db.rollback()
            flash(f"Falha ao cancelar modelo financeiro: {exc}", "error")
        return redirect(url_for("asset_detail", asset_id=asset_id))

    @app.route("/asset/<int:asset_id>/financial-model/<int:model_id>/activate", methods=["POST"])
    def activate_asset_financial_model(asset_id: int, model_id: int):
        try:
            activate_financial_model(g.db, model_id=model_id, asset_id=asset_id)
            g.db.commit()
            flash("Modelo financeiro ativado.", "success")
        except FinancialModelError as exc:
            g.db.rollback()
            flash(f"Falha ao ativar modelo financeiro: {exc}", "error")
        return redirect(url_for("asset_detail", asset_id=asset_id))

    @app.route("/asset/<int:asset_id>/financial-model/<int:model_id>/archive", methods=["POST"])
    def archive_asset_financial_model(asset_id: int, model_id: int):
        try:
            archive_financial_model(g.db, model_id=model_id, asset_id=asset_id)
            g.db.commit()
            flash("Modelo financeiro arquivado.", "success")
        except FinancialModelError as exc:
            g.db.rollback()
            flash(f"Falha ao arquivar modelo financeiro: {exc}", "error")
        return redirect(url_for("asset_detail", asset_id=asset_id))

    @app.route("/asset/<int:asset_id>/financial-model/<int:model_id>/download")
    def download_asset_financial_model(asset_id: int, model_id: int):
        model = get_financial_model_for_asset(g.db, asset_id=asset_id, model_id=model_id)
        source = get_financial_model_source(g.db, model_id) if model else None
        if model is None or source is None:
            abort(404)
        try:
            path = resolve_financial_model_path(source)
            if model["file_sha256"] and financial_model_sha256_file(path) != model["file_sha256"]:
                abort(409)
        except FinancialModelError:
            abort(404)
        return send_file(path, as_attachment=True, download_name=source["original_filename"], max_age=0)

    @app.route("/asset/<int:asset_id>/financial-model/compare")
    def compare_asset_financial_models(asset_id: int):
        left_raw = request.args.get("left", "")
        right_raw = request.args.get("right", "")
        if not left_raw.isdigit() or not right_raw.isdigit():
            flash("Escolhe duas versoes para comparar.", "error")
            return redirect(url_for("asset_detail", asset_id=asset_id))
        try:
            comparison = compare_financial_models(g.db, asset_id=asset_id, left_id=int(left_raw), right_id=int(right_raw))
        except FinancialModelError:
            abort(404)
        return render_template(
            "financial_model_compare.html",
            title="Comparar modelos financeiros",
            asset=query_one("SELECT * FROM assets WHERE id = ?", (asset_id,)),
            comparison=comparison,
        )

    @app.route("/asset/<int:asset_id>/performance-settings", methods=["POST"])
    def update_asset_performance_settings(asset_id: int):
        asset = query_one("SELECT id FROM assets WHERE id = ?", (asset_id,))
        if asset is None:
            flash("Asset nao encontrado.", "error")
            return redirect(url_for("assets"))

        monthly_budget_json = request.form.get("monthly_budget_json", "").strip()
        if monthly_budget_json:
            try:
                payload = json.loads(monthly_budget_json)
                if not isinstance(payload, dict):
                    raise ValueError
            except (json.JSONDecodeError, ValueError):
                flash("OrÃ§amento mensal invÃ¡lido. Usa JSON com meses 01-12 e valores kWh/kWp.", "error")
                return redirect(url_for("asset_detail", asset_id=asset_id))

        def threshold(name: str, default: float) -> float:
            value = parse_float_value(request.form.get(name, ""))
            return value if value is not None else default

        now = datetime.now().isoformat(timespec="seconds")
        g.db.execute(
            """
            INSERT INTO performance_settings (
                asset_id, enabled, warning_deviation_pct, alert_deviation_pct, critical_deviation_pct,
                baseline_years, min_baseline_points, monthly_budget_json, notes, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(asset_id) DO UPDATE SET
                enabled = excluded.enabled,
                warning_deviation_pct = excluded.warning_deviation_pct,
                alert_deviation_pct = excluded.alert_deviation_pct,
                critical_deviation_pct = excluded.critical_deviation_pct,
                baseline_years = excluded.baseline_years,
                min_baseline_points = excluded.min_baseline_points,
                monthly_budget_json = excluded.monthly_budget_json,
                notes = excluded.notes,
                updated_at = excluded.updated_at
            """,
            (
                asset_id,
                1 if request.form.get("enabled") == "on" else 0,
                threshold("warning_deviation_pct", -10),
                threshold("alert_deviation_pct", -20),
                threshold("critical_deviation_pct", -30),
                int(parse_float_value(request.form.get("baseline_years", "")) or 2),
                int(parse_float_value(request.form.get("min_baseline_points", "")) or 1),
                monthly_budget_json,
                request.form.get("notes", "").strip(),
                now,
            ),
        )
        g.db.commit()
        flash("DefiniÃ§Ãµes de performance guardadas.", "success")
        return redirect(url_for("asset_detail", asset_id=asset_id))

    @app.route("/asset/<int:asset_id>/expected-strings", methods=["POST"])
    def update_asset_expected_strings(asset_id: int):
        device_id_raw = request.form.get("provider_device_id", "").strip()
        if not device_id_raw.isdigit():
            flash("Dispositivo invalido.", "error")
            return redirect(url_for("asset_detail", asset_id=asset_id))
        device_id = int(device_id_raw)
        device = query_one(
            "SELECT id FROM provider_devices WHERE id = ? AND asset_id = ?",
            (device_id, asset_id),
        )
        if device is None:
            flash("Dispositivo nao encontrado.", "error")
            return redirect(url_for("asset_detail", asset_id=asset_id))
        selected = {int(value) for value in request.form.getlist("expected_strings") if value.isdigit()}
        now = datetime.now().isoformat(timespec="seconds")
        for index in range(1, 37):
            existing = query_one(
                """
                SELECT *
                FROM provider_device_expected_strings
                WHERE provider_device_id = ? AND string_index = ?
                """,
                (device_id, index),
            )
            expected = 1 if index in selected else 0
            if existing:
                g.db.execute(
                    """
                    UPDATE provider_device_expected_strings
                    SET expected = ?, source = 'manual', updated_at = ?
                    WHERE id = ?
                    """,
                    (expected, now, existing["id"]),
                )
            elif expected:
                g.db.execute(
                    """
                    INSERT INTO provider_device_expected_strings (
                        provider_device_id, string_index, expected, source, observed_count,
                        created_at, updated_at
                    ) VALUES (?, ?, 1, 'manual', 0, ?, ?)
                    """,
                    (device_id, index, now, now),
                )
        g.db.commit()
        flash("Perfil de strings atualizado.", "success")
        return redirect(url_for("asset_detail", asset_id=asset_id))

    @app.route("/asset/<int:asset_id>/installation-group", methods=["POST"])
    def update_asset_installation_group(asset_id: int):
        asset = query_one("SELECT id, project_name FROM assets WHERE id = ?", (asset_id,))
        if asset is None:
            flash("Asset nao encontrado.", "error")
            return redirect(url_for("assets"))

        group_name = request.form.get("installation_group", "").strip()
        if not group_name:
            group_name = infer_installation_group(asset["project_name"])

        g.db.execute("UPDATE assets SET installation_group = ? WHERE id = ?", (group_name, asset_id))
        apply_group_defaults_to_asset(g.db, asset_id, group_name)
        g.db.commit()
        flash("Grupo de instalacao atualizado.", "success")
        return redirect(url_for("asset_detail", asset_id=asset_id))

    @app.route("/asset/<int:asset_id>/update", methods=["POST"])
    def update_asset(asset_id: int):
        asset = query_one("SELECT id FROM assets WHERE id = ?", (asset_id,))
        if asset is None:
            flash("Asset nao encontrado.", "error")
            return redirect(url_for("assets"))

        payload = {
            "project_name": request.form.get("project_name", "").strip(),
            "installation_group": request.form.get("installation_group", "").strip(),
            "company_name": request.form.get("company_name", "").strip(),
            "location": request.form.get("location", "").strip(),
            "address": request.form.get("address", "").strip(),
            "contract_type": request.form.get("contract_type", "").strip(),
            "maintenance": request.form.get("maintenance", "").strip(),
            "active_contract": request.form.get("active_contract", "").strip(),
            "start_contract": request.form.get("start_contract", "").strip(),
            "end_contract": request.form.get("end_contract", "").strip(),
            "contact_name": request.form.get("contact_name", "").strip(),
            "contact_email": request.form.get("contact_email", "").strip(),
            "contact_phone": request.form.get("contact_phone", "").strip(),
            "notes": request.form.get("notes", "").strip(),
            "monitoring_enabled": 1 if request.form.get("monitoring_enabled") == "on" else 0,
            "alerts_enabled": 1 if request.form.get("alerts_enabled") == "on" else 0,
            "selected_for_alerts": 1 if request.form.get("selected_for_alerts") == "on" else 0,
            "monitoring_status": request.form.get("monitoring_status", "active").strip() or "active",
            "silenced_until": request.form.get("silenced_until", "").strip(),
            "silence_reason": request.form.get("silence_reason", "").strip(),
        }
        if not payload["project_name"]:
            flash("O nome da central e obrigatorio.", "error")
            return redirect(url_for("asset_detail", asset_id=asset_id))
        if not payload["installation_group"]:
            payload["installation_group"] = infer_installation_group(payload["project_name"])
        payload = apply_group_defaults(g.db, payload, payload["installation_group"], exclude_asset_id=asset_id)
        payload["start_contract"] = normalize_date_value(payload["start_contract"])
        payload["end_contract"] = normalize_date_value(payload["end_contract"])
        payload["active_contract"] = derive_active_contract(payload["end_contract"], payload["active_contract"])
        if payload["monitoring_status"] not in ASSET_MONITORING_STATUSES:
            payload["monitoring_status"] = "active"
        if payload["monitoring_status"] != "silenced":
            payload["silenced_until"] = ""

        g.db.execute(
            """
            UPDATE assets
            SET project_name = ?, installation_group = ?, company_name = ?, location = ?, address = ?,
                contract_type = ?, maintenance = ?, active_contract = ?, start_contract = ?, end_contract = ?,
                contact_name = ?, contact_email = ?, contact_phone = ?, notes = ?,
                monitoring_enabled = ?, alerts_enabled = ?, selected_for_alerts = ?, monitoring_status = ?, silenced_until = ?, silence_reason = ?
            WHERE id = ?
            """,
            (
                payload["project_name"],
                payload["installation_group"],
                payload["company_name"],
                payload["location"],
                payload["address"],
                payload["contract_type"],
                payload["maintenance"],
                payload["active_contract"],
                payload["start_contract"],
                payload["end_contract"],
                payload["contact_name"],
                payload["contact_email"],
                payload["contact_phone"],
                payload["notes"],
                payload["monitoring_enabled"],
                payload["alerts_enabled"],
                payload["selected_for_alerts"],
                payload["monitoring_status"],
                payload["silenced_until"],
                payload["silence_reason"],
                asset_id,
            ),
        )
        g.db.commit()
        flash("Asset atualizado.", "success")
        return redirect(url_for("asset_detail", asset_id=asset_id))

    @app.route("/asset/<int:asset_id>/delete", methods=["POST"])
    def delete_asset(asset_id: int):
        asset = query_one("SELECT id, project_name FROM assets WHERE id = ?", (asset_id,))
        if asset is None:
            flash("Asset nao encontrado.", "error")
            return redirect(url_for("assets"))

        contract = query_one("SELECT pdf_path FROM om_contracts WHERE asset_id = ?", (asset_id,))
        if contract and contract["pdf_path"]:
            contract_path = resolve_runtime_file_path_within(contract["pdf_path"], CONTRACTS_DIR)
            if contract_path is not None and contract_path.exists():
                contract_path.unlink()

        g.db.execute("DELETE FROM assets WHERE id = ?", (asset_id,))
        g.db.commit()
        flash(f"Asset '{asset['project_name']}' apagado.", "success")
        return redirect(url_for("assets"))

    @app.route("/asset/<int:asset_id>/contract", methods=["POST"])
    def update_asset_contract(asset_id: int):
        asset = query_one("SELECT id, project_name FROM assets WHERE id = ?", (asset_id,))
        if asset is None:
            flash("Asset nao encontrado.", "error")
            return redirect(url_for("assets"))

        start_date = normalize_date_value(request.form.get("contract_start_date", "").strip())
        end_date = normalize_date_value(request.form.get("contract_end_date", "").strip())
        annual_value_raw = request.form.get("annual_value", "").strip()
        contract_notes = request.form.get("contract_notes", "").strip()
        uploaded_file = request.files.get("contract_pdf")

        annual_value = None
        if annual_value_raw:
            normalized_value = annual_value_raw.replace(" ", "").replace(",", ".")
            try:
                annual_value = float(normalized_value)
            except ValueError:
                flash("O valor anual do contrato nao e valido.", "error")
                return redirect(url_for("asset_detail", asset_id=asset_id))

        existing_contract = query_one(
            """
            SELECT *
            FROM om_contracts
            WHERE asset_id = ?
            """,
            (asset_id,),
        )

        stored_path = existing_contract["pdf_path"] if existing_contract else ""
        original_filename = existing_contract["original_filename"] if existing_contract else ""
        if uploaded_file and uploaded_file.filename:
            suffix = Path(uploaded_file.filename).suffix.lower()
            if suffix != ".pdf":
                flash("O contrato tem de ser um ficheiro PDF.", "error")
                return redirect(url_for("asset_detail", asset_id=asset_id))
            header = uploaded_file.stream.read(5)
            uploaded_file.stream.seek(0)
            if header != b"%PDF-":
                flash("O contrato enviado nao parece ser um PDF valido.", "error")
                return redirect(url_for("asset_detail", asset_id=asset_id))
            CONTRACTS_DIR.mkdir(parents=True, exist_ok=True)
            safe_stem = normalize_name(asset["project_name"]).replace(" ", "-") or f"asset-{asset_id}"
            filename = f"{asset_id}_{safe_stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            target_path = (CONTRACTS_DIR / filename).resolve()
            if not path_is_within(target_path, CONTRACTS_DIR):
                current_app.logger.error("Rejected contract upload outside contracts directory for asset %s", asset_id)
                abort(400)
            uploaded_file.save(target_path)
            if stored_path:
                old_path = resolve_runtime_file_path_within(stored_path, CONTRACTS_DIR)
                if old_path is not None and old_path.exists() and old_path != target_path:
                    old_path.unlink()
            stored_path = store_runtime_relative_path(target_path)
            original_filename = Path(uploaded_file.filename).name[:255]

        if existing_contract:
            g.db.execute(
                """
                UPDATE om_contracts
                SET contract_start_date = ?, contract_end_date = ?, annual_value = ?, notes = ?,
                    pdf_path = ?, original_filename = ?, updated_at = ?
                WHERE asset_id = ?
                """,
                (
                    start_date,
                    end_date,
                    annual_value,
                    contract_notes,
                    stored_path,
                    original_filename,
                    datetime.now().isoformat(timespec="seconds"),
                    asset_id,
                ),
            )
        else:
            g.db.execute(
                """
                INSERT INTO om_contracts (
                    asset_id, contract_start_date, contract_end_date, annual_value, notes, pdf_path,
                    original_filename, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    asset_id,
                    start_date,
                    end_date,
                    annual_value,
                    contract_notes,
                    stored_path,
                    original_filename,
                    datetime.now().isoformat(timespec="seconds"),
                    datetime.now().isoformat(timespec="seconds"),
                ),
            )
        sync_asset_contract_status(g.db, asset_id, start_date, end_date)
        g.db.commit()
        flash("Contrato O&M atualizado.", "success")
        return redirect(url_for("asset_detail", asset_id=asset_id))

    @app.route("/asset/<int:asset_id>/contract/open")
    def open_asset_contract(asset_id: int):
        contract = query_one(
            """
            SELECT pdf_path
            FROM om_contracts
            WHERE asset_id = ?
            """,
            (asset_id,),
        )
        if contract is None or not contract["pdf_path"]:
            flash("Esta central ainda nao tem contrato associado.", "error")
            return redirect(url_for("asset_detail", asset_id=asset_id))

        contract_path = resolve_runtime_file_path_within(contract["pdf_path"], CONTRACTS_DIR)
        if contract_path is None:
            current_app.logger.warning("Blocked contract path outside contracts directory for asset %s", asset_id)
            abort(404)
        if not contract_path.exists():
            flash("O ficheiro do contrato nao foi encontrado no projeto.", "error")
            return redirect(url_for("asset_detail", asset_id=asset_id))

        return send_file(contract_path, mimetype="application/pdf", as_attachment=False, max_age=0)

    @app.route("/asset/<int:asset_id>/alias", methods=["POST"])
    def add_alias(asset_id: int):
        alias_name = request.form.get("alias_name", "").strip()
        if not alias_name:
            flash("Indica um nome alternativo para guardar o alias.", "error")
            return redirect(url_for("asset_detail", asset_id=asset_id))

        normalized = normalize_name(alias_name)
        if not normalized:
            flash("O alias indicado nao e valido.", "error")
            return redirect(url_for("asset_detail", asset_id=asset_id))

        existing = query_one("SELECT id FROM asset_aliases WHERE normalized_alias = ?", (normalized,))
        if existing:
            flash("Esse alias ja esta associado a outra instalacao.", "error")
            return redirect(url_for("asset_detail", asset_id=asset_id))

        g.db.execute(
            "INSERT INTO asset_aliases (asset_id, alias_name, normalized_alias, source) VALUES (?, ?, ?, ?)",
            (asset_id, alias_name, normalized, "manual"),
        )
        g.db.commit()
        rebuild_asset_alias_blob(g.db, asset_id)
        flash("Alias guardado com sucesso.", "success")
        return redirect(url_for("asset_detail", asset_id=asset_id))

    @app.route("/asset/<int:asset_id>/alias/<int:alias_id>/update", methods=["POST"])
    def update_alias(asset_id: int, alias_id: int):
        alias_name = request.form.get("alias_name", "").strip()
        if not alias_name:
            flash("Indica um nome valido para o alias.", "error")
            return redirect(url_for("asset_detail", asset_id=asset_id))

        alias_row = query_one(
            "SELECT * FROM asset_aliases WHERE id = ? AND asset_id = ?",
            (alias_id, asset_id),
        )
        if alias_row is None:
            flash("Alias nao encontrado.", "error")
            return redirect(url_for("asset_detail", asset_id=asset_id))

        normalized = normalize_name(alias_name)
        existing = query_one("SELECT id, asset_id FROM asset_aliases WHERE normalized_alias = ?", (normalized,))
        if existing and (existing["id"] != alias_id or existing["asset_id"] != asset_id):
            flash("Esse alias ja esta associado a outra instalacao.", "error")
            return redirect(url_for("asset_detail", asset_id=asset_id))

        g.db.execute(
            """
            UPDATE asset_aliases
            SET alias_name = ?, normalized_alias = ?, source = ?
            WHERE id = ? AND asset_id = ?
            """,
            (alias_name, normalized, "manual-edit", alias_id, asset_id),
        )
        g.db.commit()
        rebuild_asset_alias_blob(g.db, asset_id)
        flash("Alias atualizado.", "success")
        return redirect(url_for("asset_detail", asset_id=asset_id))

    @app.route("/asset/<int:asset_id>/alias/<int:alias_id>/delete", methods=["POST"])
    def delete_alias(asset_id: int, alias_id: int):
        alias_row = query_one(
            "SELECT * FROM asset_aliases WHERE id = ? AND asset_id = ?",
            (alias_id, asset_id),
        )
        if alias_row is None:
            flash("Alias nao encontrado.", "error")
            return redirect(url_for("asset_detail", asset_id=asset_id))

        g.db.execute("DELETE FROM asset_aliases WHERE id = ? AND asset_id = ?", (alias_id, asset_id))
        g.db.commit()
        rebuild_asset_alias_blob(g.db, asset_id)
        flash("Alias apagado.", "success")
        return redirect(url_for("asset_detail", asset_id=asset_id))

    @app.route("/monitoring", methods=["GET", "POST"])
    def monitoring() -> str:
        if request.method == "POST":
            record_date = request.form.get("record_date", date.today().isoformat())
            pasted_table = request.form.get("pasted_table", "")
            default_notes = request.form.get("default_notes", "").strip()
            platform_source = request.form.get("platform_source", "Manual / Outro").strip() or "Manual / Outro"
            import_scope = request.form.get("import_scope", "complete").strip() or "complete"
            result = import_daily_monitoring(
                g.db,
                pasted_table,
                record_date,
                default_notes,
                platform_source,
                import_scope,
            )
            flash(
                f"Importacao concluida: {result.imported} registos, {result.matched} associados automaticamente, {result.unmatched} por mapear, {result.auto_resolved} resolvidos automaticamente.",
                "success" if result.imported else "warning",
            )
            if result.batch_id is not None:
                return redirect(url_for("monitoring", batch_id=result.batch_id))
            return redirect(url_for("monitoring"))

        search = request.args.get("search", "").strip()
        asset_filter = request.args.get("asset_id", "").strip()
        status_filter = request.args.get("status", "").strip()
        source_filter = request.args.get("source", "").strip()
        issue_only = request.args.get("issue_only", "no").strip()
        start_date = request.args.get("start_date", "").strip()
        end_date = request.args.get("end_date", "").strip()
        om_only = request.args.get("om_only", "yes").strip()
        batch_id = request.args.get("batch_id", "").strip()

        latest_conditions = []
        latest_params: list[Any] = []
        if search:
            wildcard = f"%{search}%"
            latest_conditions.append("(a.project_name LIKE ? OR a.alias_blob LIKE ? OR a.company_name LIKE ?)")
            latest_params.extend([wildcard, wildcard, wildcard])
        if asset_filter:
            latest_conditions.append("a.id = ?")
            latest_params.append(asset_filter)
        if status_filter:
            latest_conditions.append("lm.status = ?")
            latest_params.append(status_filter)
        elif issue_only == "yes":
            latest_conditions.append("lm.status IN ('Erro', 'Desconectada')")
        if source_filter:
            latest_conditions.append(
                "EXISTS (SELECT 1 FROM monitoring_records src WHERE src.asset_id = a.id AND src.record_date = lm.record_date AND src.status = lm.status AND src.source = ?)"
            )
            latest_params.append(source_filter)
        if om_only == "yes":
            latest_conditions.append("a.active_contract = 'yes'")
        if start_date:
            latest_conditions.append("lm.record_date >= ?")
            latest_params.append(start_date)
        if end_date:
            latest_conditions.append("lm.record_date <= ?")
            latest_params.append(end_date)

        latest_where_sql = f"WHERE {' AND '.join(latest_conditions)}" if latest_conditions else ""
        latest_rows = enrich_operational_rows(g.db, query_all(
            g.db,
            f"""
            SELECT
                a.id AS asset_id,
                a.project_name,
                a.installation_group,
                a.location,
                a.contract_type,
                a.active_contract,
                lm.status,
                lm.record_date,
                lm.notes,
                (
                    SELECT COUNT(*)
                    FROM monitoring_records mr
                    WHERE mr.asset_id = a.id
                ) AS history_count
            FROM latest_monitoring_view lm
            JOIN assets a ON a.id = lm.asset_id
            {latest_where_sql}
            ORDER BY
                CASE a.active_contract WHEN 'yes' THEN 1 ELSE 2 END,
                CASE lm.status
                    WHEN 'Erro' THEN 1
                    WHEN 'Desconectada' THEN 2
                    ELSE 3
                END,
                a.project_name COLLATE NOCASE
            """,
            latest_params,
        ))
        latest_availability_rows = {
            int(row["asset_id"]): row
            for row in query_all(
                g.db,
                """
                SELECT ad.*
                FROM availability_daily ad
                JOIN (
                    SELECT asset_id, MAX(period_date || 'T' || printf('%09d', id)) AS marker
                    FROM availability_daily
                    GROUP BY asset_id
                ) latest
                  ON latest.asset_id = ad.asset_id
                 AND latest.marker = ad.period_date || 'T' || printf('%09d', ad.id)
                """,
            )
        }
        for row in latest_rows:
            row["availability"] = latest_availability_rows.get(int(row["asset_id"]))

        filter_sql = []
        params: list[Any] = []
        if search:
            wildcard = f"%{search}%"
            filter_sql.append("(a.project_name LIKE ? OR a.alias_blob LIKE ? OR a.company_name LIKE ?)")
            params.extend([wildcard, wildcard, wildcard])
        if asset_filter:
            filter_sql.append("a.id = ?")
            params.append(asset_filter)
        if status_filter:
            filter_sql.append("mr.status = ?")
            params.append(status_filter)
        elif issue_only == "yes":
            filter_sql.append("mr.status IN ('Erro', 'Desconectada')")
        if source_filter:
            filter_sql.append("mr.source = ?")
            params.append(source_filter)
        if om_only == "yes":
            filter_sql.append("a.active_contract = 'yes'")
        if start_date:
            filter_sql.append("mr.record_date >= ?")
            params.append(start_date)
        if end_date:
            filter_sql.append("mr.record_date <= ?")
            params.append(end_date)

        where_sql = f"WHERE {' AND '.join(filter_sql)}" if filter_sql else ""
        history_rows = query_all(
            g.db,
            f"""
            SELECT
                mr.id,
                mr.record_date,
                mr.status,
                mr.notes,
                mr.source,
                mr.batch_id,
                a.id AS asset_id,
                a.project_name,
                a.installation_group,
                a.location,
                a.contract_type,
                a.active_contract,
                mib.imported_at
            FROM monitoring_records mr
            JOIN assets a ON a.id = mr.asset_id
            LEFT JOIN monitoring_import_batches mib ON mib.id = mr.batch_id
            {where_sql}
            ORDER BY mr.record_date DESC, a.project_name COLLATE NOCASE, mr.id DESC
            LIMIT 250
            """,
            params,
        )

        selected_asset = None
        asset_history = []
        asset_problem_periods = []
        selected_group_assets = []
        if asset_filter:
            selected_asset = query_one(
                """
                SELECT
                    a.*,
                    lm.status AS latest_status,
                    lm.record_date AS latest_status_date
                FROM assets a
                LEFT JOIN latest_monitoring_view lm ON lm.asset_id = a.id
                WHERE a.id = ?
                """,
                (asset_filter,),
            )
            asset_history = query_all(
                g.db,
                """
                SELECT mr.id, mr.record_date, mr.status, mr.notes, mr.source, mr.batch_id, mib.imported_at
                FROM monitoring_records mr
                LEFT JOIN monitoring_import_batches mib ON mib.id = mr.batch_id
                WHERE mr.asset_id = ?
                ORDER BY mr.record_date DESC, mr.id DESC
                LIMIT 200
                """,
                (asset_filter,),
            )
            asset_problem_periods = build_problem_periods(g.db, int(asset_filter))
            selected_group_assets = query_all(
                g.db,
                """
                SELECT
                    a.id,
                    a.project_name,
                    a.installation_group,
                    lm.status,
                    lm.record_date
                FROM assets a
                LEFT JOIN latest_monitoring_view lm ON lm.asset_id = a.id
                WHERE a.installation_group = ?
                ORDER BY a.project_name COLLATE NOCASE
                """,
                (selected_asset["installation_group"] or selected_asset["project_name"],),
            )

        unresolved = query_all(
            g.db,
            """
            SELECT mu.*, mib.imported_at
            FROM monitoring_unmatched mu
            LEFT JOIN monitoring_import_batches mib ON mib.id = mu.batch_id
            ORDER BY record_date DESC, original_name COLLATE NOCASE
            LIMIT 50
            """
        )
        import_batches = query_all(
            g.db,
            """
            SELECT
                mib.*,
                (
                    SELECT COUNT(*)
                    FROM monitoring_records mr
                    WHERE mr.batch_id = mib.id
                ) AS resolved_rows
            FROM monitoring_import_batches mib
            ORDER BY mib.imported_at DESC, mib.id DESC
            LIMIT 30
            """
        )
        assets_for_mapping = query_all(
            g.db,
            "SELECT id, project_name FROM assets ORDER BY project_name COLLATE NOCASE",
        )
        monitoring_statuses = [
            row["status"]
            for row in query_all(
                g.db,
                "SELECT DISTINCT status FROM monitoring_records ORDER BY status",
            )
        ]
        monitoring_stats = {
            "history_records": len(history_rows),
            "centrals_in_filter": len(latest_rows),
            "installations_in_filter": len(group_latest_rows_by_installation(latest_rows)),
            "current_errors": sum(1 for row in latest_rows if row["status"] == "Erro"),
            "current_disconnected": sum(1 for row in latest_rows if row["status"] == "Desconectada"),
            "current_active_om": sum(1 for row in latest_rows if row["active_contract"] == "yes"),
            "critical_priority": sum(1 for row in latest_rows if row.get("auto_priority") == "Critica"),
            "high_priority": sum(1 for row in latest_rows if row.get("auto_priority") == "Alta"),
            "recurring_90d": sum(1 for row in latest_rows if int(row.get("recurrence_count") or 0) >= 2),
        }
        grouped_latest_rows = group_latest_rows_by_installation(latest_rows)
        batch_insight = build_batch_insight(g.db, int(batch_id)) if batch_id else None

        return render_template(
            "monitoring.html",
            latest_rows=latest_rows,
            grouped_latest_rows=grouped_latest_rows,
            unresolved=unresolved,
            assets_for_mapping=assets_for_mapping,
            monitoring_statuses=monitoring_statuses,
            monitoring_stats=monitoring_stats,
            history_rows=history_rows,
            selected_asset=selected_asset,
            selected_group_assets=selected_group_assets,
            asset_history=asset_history,
            asset_problem_periods=asset_problem_periods,
            import_batches=import_batches,
            search=search,
            asset_filter=asset_filter,
            status_filter=status_filter,
            source_filter=source_filter,
            issue_only=issue_only,
            start_date=start_date,
            end_date=end_date,
            om_only=om_only,
            batch_insight=batch_insight,
        )

    @app.route("/monitoring/record/<int:record_id>/update", methods=["POST"])
    def update_monitoring_record(record_id: int):
        record = query_one("SELECT asset_id FROM monitoring_records WHERE id = ?", (record_id,))
        if record is None:
            flash("Registo de monitorizacao nao encontrado.", "error")
            return redirect(url_for("monitoring"))

        record_date = request.form.get("record_date", "").strip()
        status = normalize_status(request.form.get("status", "").strip())
        notes = request.form.get("notes", "").strip()
        if not record_date or not status:
            flash("Data e estado sao obrigatorios para atualizar o registo.", "error")
            return redirect(url_for("monitoring", asset_id=record["asset_id"]))

        g.db.execute(
            """
            UPDATE monitoring_records
            SET record_date = ?, status = ?, notes = ?
            WHERE id = ?
            """,
            (record_date, status, notes, record_id),
        )
        g.db.commit()
        flash("Registo de monitorizacao atualizado.", "success")
        return redirect(url_for("monitoring", asset_id=record["asset_id"]))

    @app.route("/monitoring/record/<int:record_id>/delete", methods=["POST"])
    def delete_monitoring_record(record_id: int):
        record = query_one("SELECT asset_id FROM monitoring_records WHERE id = ?", (record_id,))
        if record is None:
            flash("Registo de monitorizacao nao encontrado.", "error")
            return redirect(url_for("monitoring"))

        g.db.execute("DELETE FROM monitoring_records WHERE id = ?", (record_id,))
        g.db.commit()
        flash("Registo de monitorizacao apagado.", "success")
        return redirect(url_for("monitoring", asset_id=record["asset_id"]))

    @app.route("/monitoring/unmatched/<int:row_id>/resolve", methods=["POST"])
    def resolve_unmatched(row_id: int):
        asset_id = int(request.form["asset_id"])
        unmatched = query_one("SELECT * FROM monitoring_unmatched WHERE id = ?", (row_id,))
        if unmatched is None:
            flash("Linha pendente nao encontrada.", "error")
            return redirect(url_for("monitoring"))

        alias_name = unmatched["original_name"]
        normalized = normalize_name(alias_name)
        existing = query_one("SELECT id FROM asset_aliases WHERE normalized_alias = ?", (normalized,))
        if existing and existing["id"]:
            flash("Esse nome ja esta usado como alias.", "error")
            return redirect(url_for("monitoring"))

        g.db.execute(
            "INSERT INTO asset_aliases (asset_id, alias_name, normalized_alias, source) VALUES (?, ?, ?, ?)",
            (asset_id, alias_name, normalized, "resolved"),
        )
        g.db.execute(
            """
            INSERT INTO monitoring_records (asset_id, status, record_date, notes, source, batch_id)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                asset_id,
                unmatched["status"],
                unmatched["record_date"],
                unmatched["notes"],
                "resolved-unmatched",
                unmatched["batch_id"],
            ),
        )
        g.db.execute("DELETE FROM monitoring_unmatched WHERE id = ?", (row_id,))
        g.db.commit()
        rebuild_asset_alias_blob(g.db, asset_id)
        flash("Linha associada e importada com sucesso.", "success")
        return redirect(url_for("monitoring", asset_id=asset_id))

    @app.route("/tickets", methods=["GET", "POST"])
    def tickets() -> str:
        if request.method == "POST":
            asset_id = int(request.form["asset_id"])
            title = request.form.get("title", "").strip()
            urgency = request.form.get("urgency", "Media")
            status = request.form.get("status", "Aberto")
            installation_ref = request.form.get("installation_ref", "").strip()
            notes = request.form.get("notes", "").strip()
            next_action = request.form.get("next_action", "").strip()
            planned_date = normalize_optional_date(request.form.get("planned_date"))
            due_date = normalize_optional_date(request.form.get("due_date"))
            estimated_minutes = parse_positive_int(request.form.get("estimated_minutes"), default=60)
            assigned_to = request.form.get("assigned_to", "").strip()
            material_status = normalize_choice(
                request.form.get("material_status", "Nao definido"),
                TICKET_MATERIAL_STATUSES,
                "Nao definido",
            )
            work_type = normalize_choice(request.form.get("work_type", "Diagnostico"), TICKET_WORK_TYPES, "Diagnostico")
            planning_notes = request.form.get("planning_notes", "").strip()
            if not title:
                flash("A intervencao precisa de um titulo.", "error")
                return redirect(url_for("tickets"))

            g.db.execute(
                """
                INSERT INTO tickets (
                    asset_id, title, urgency, status, installation_ref, notes, next_action,
                    planned_date, due_date, estimated_minutes, assigned_to, material_status,
                    work_type, planning_notes, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    asset_id,
                    title,
                    urgency,
                    status,
                    installation_ref,
                    notes,
                    next_action,
                    planned_date,
                    due_date,
                    estimated_minutes,
                    assigned_to,
                    material_status,
                    work_type,
                    planning_notes,
                    datetime.now().isoformat(timespec="seconds"),
                    datetime.now().isoformat(timespec="seconds"),
                ),
            )
            g.db.commit()
            flash("Intervencao criada.", "success")
            return redirect(url_for("tickets", asset_id=asset_id))

        search = request.args.get("search", "").strip()
        asset_filter = request.args.get("asset_id", "").strip()
        status_filter = request.args.get("status", "").strip()
        urgency_filter = request.args.get("urgency", "").strip()
        scope = request.args.get("scope", "").strip()
        om_only = request.args.get("om_only", "yes").strip()
        calendar_month = normalize_calendar_month(request.args.get("calendar_month", ""))

        conditions = []
        params: list[Any] = []
        if search:
            wildcard = f"%{search}%"
            conditions.append(
                "(a.project_name LIKE ? OR a.alias_blob LIKE ? OR t.title LIKE ? OR COALESCE(t.notes, '') LIKE ?)"
            )
            params.extend([wildcard, wildcard, wildcard, wildcard])
        if asset_filter:
            conditions.append("a.id = ?")
            params.append(asset_filter)
        if status_filter:
            conditions.append("t.status = ?")
            params.append(status_filter)
        if urgency_filter:
            conditions.append("t.urgency = ?")
            params.append(urgency_filter)
        if scope == "open":
            conditions.append("t.status != 'Fechado'")
        if om_only == "yes":
            conditions.append("a.active_contract = 'yes'")

        where_sql = f"WHERE {' AND '.join(conditions)}" if conditions else ""
        ticket_rows = query_all(
            g.db,
            f"""
            SELECT
                t.*,
                a.project_name,
                a.location,
                a.active_contract,
                a.contract_type
            FROM tickets t
            JOIN assets a ON a.id = t.asset_id
            {where_sql}
            ORDER BY
                CASE a.active_contract WHEN 'yes' THEN 1 ELSE 2 END,
                a.project_name COLLATE NOCASE,
                CASE t.status
                    WHEN 'Aberto' THEN 1
                    WHEN 'Em analise' THEN 2
                    WHEN 'Agendado' THEN 3
                    WHEN 'Em visita' THEN 4
                    WHEN 'Resolvido' THEN 5
                    ELSE 6
                END,
                CASE t.urgency
                    WHEN 'Critica' THEN 1
                    WHEN 'Alta' THEN 2
                    WHEN 'Media' THEN 3
                    ELSE 4
                END,
                t.updated_at DESC,
                t.id DESC
            """,
            params,
        )

        assets_rows = query_all(
            g.db,
            "SELECT id, project_name FROM assets ORDER BY project_name COLLATE NOCASE",
        )
        visits_by_ticket = build_visits_by_ticket(
            query_all(
                g.db,
                """
                SELECT *
                FROM ticket_visits
                ORDER BY visit_date DESC, id DESC
                """,
            )
        )
        grouped_tickets = group_tickets_by_asset(ticket_rows)
        calendar_start, calendar_end, previous_month, next_month = calendar_month_bounds(calendar_month)
        calendar_conditions = [
            "mr.status IN ('Erro', 'Desconectada')",
            "mr.record_date BETWEEN ? AND ?",
        ]
        calendar_params: list[Any] = [calendar_start.isoformat(), calendar_end.isoformat()]
        if search:
            wildcard = f"%{search}%"
            calendar_conditions.append(
                "(a.project_name LIKE ? OR a.alias_blob LIKE ? OR COALESCE(mr.notes, '') LIKE ? OR COALESCE(mr.source, '') LIKE ?)"
            )
            calendar_params.extend([wildcard, wildcard, wildcard, wildcard])
        if asset_filter:
            calendar_conditions.append("a.id = ?")
            calendar_params.append(asset_filter)
        if om_only == "yes":
            calendar_conditions.append("a.active_contract = 'yes'")
        calendar_where_sql = f"WHERE {' AND '.join(calendar_conditions)}"

        calendar_rows = query_all(
            g.db,
            f"""
            SELECT
                mr.id,
                mr.asset_id,
                mr.status,
                mr.record_date,
                mr.notes,
                mr.source,
                a.project_name
            FROM monitoring_records mr
            JOIN assets a ON a.id = mr.asset_id
            {calendar_where_sql}
            ORDER BY
                mr.record_date,
                CASE mr.status
                    WHEN 'Erro' THEN 1
                    WHEN 'Desconectada' THEN 2
                    ELSE 3
                END,
                a.project_name COLLATE NOCASE,
                mr.id DESC
            """,
            calendar_params,
        )
        error_calendar = build_error_calendar(calendar_month, calendar_rows)

        selected_asset = None
        central_history = []
        central_summary = None
        if asset_filter:
            selected_asset = query_one("SELECT * FROM assets WHERE id = ?", (asset_filter,))
            central_history = query_all(
                g.db,
                """
                SELECT
                    t.*,
                    (
                        SELECT COUNT(*)
                        FROM ticket_visits tv
                        WHERE tv.ticket_id = t.id
                    ) AS visit_count
                FROM tickets t
                WHERE t.asset_id = ?
                ORDER BY t.updated_at DESC, t.id DESC
                """,
                (asset_filter,),
            )
            central_summary = {
                "total": len(central_history),
                "open": sum(1 for ticket in central_history if ticket["status"] != "Fechado"),
                "critical": sum(1 for ticket in central_history if ticket["urgency"] == "Critica" and ticket["status"] != "Fechado"),
                "visits": sum(ticket["visit_count"] for ticket in central_history),
            }

        ticket_stats = {
            "centrals": len(grouped_tickets),
            "tickets": len(ticket_rows),
            "open": sum(1 for ticket in ticket_rows if ticket["status"] != "Fechado"),
            "critical": sum(1 for ticket in ticket_rows if ticket["urgency"] == "Critica" and ticket["status"] != "Fechado"),
        }

        return render_template(
            "tickets.html",
            tickets=ticket_rows,
            grouped_tickets=grouped_tickets,
            assets=assets_rows,
            visits_by_ticket=visits_by_ticket,
            selected_asset=selected_asset,
            central_history=central_history,
            central_summary=central_summary,
            ticket_stats=ticket_stats,
            error_calendar=error_calendar,
            calendar_month=calendar_month,
            previous_month=previous_month,
            next_month=next_month,
            search=search,
            asset_filter=asset_filter,
            status_filter=status_filter,
            urgency_filter=urgency_filter,
            scope=scope,
            om_only=om_only,
        )

    @app.route("/tickets/<int:ticket_id>/update", methods=["POST"])
    def update_ticket(ticket_id: int):
        ticket = query_one("SELECT asset_id FROM tickets WHERE id = ?", (ticket_id,))
        status = request.form.get("status", "Aberto")
        urgency = request.form.get("urgency", "Media")
        next_action = request.form.get("next_action", "").strip()
        notes = request.form.get("notes", "").strip()
        planned_date = normalize_optional_date(request.form.get("planned_date"))
        due_date = normalize_optional_date(request.form.get("due_date"))
        estimated_minutes = parse_positive_int(request.form.get("estimated_minutes"), default=60)
        assigned_to = request.form.get("assigned_to", "").strip()
        material_status = normalize_choice(
            request.form.get("material_status", "Nao definido"),
            TICKET_MATERIAL_STATUSES,
            "Nao definido",
        )
        work_type = normalize_choice(request.form.get("work_type", "Diagnostico"), TICKET_WORK_TYPES, "Diagnostico")
        planning_notes = request.form.get("planning_notes", "").strip()
        g.db.execute(
            """
            UPDATE tickets
            SET status = ?, urgency = ?, next_action = ?, notes = ?,
                planned_date = ?, due_date = ?, estimated_minutes = ?, assigned_to = ?,
                material_status = ?, work_type = ?, planning_notes = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                status,
                urgency,
                next_action,
                notes,
                planned_date,
                due_date,
                estimated_minutes,
                assigned_to,
                material_status,
                work_type,
                planning_notes,
                datetime.now().isoformat(timespec="seconds"),
                ticket_id,
            ),
        )
        g.db.commit()
        flash("Intervencao atualizada.", "success")
        if ticket:
            return redirect(url_for("tickets", asset_id=ticket["asset_id"]))
        return redirect(url_for("tickets"))

    @app.route("/tickets/<int:ticket_id>/visit", methods=["POST"])
    def add_visit(ticket_id: int):
        ticket = query_one("SELECT asset_id FROM tickets WHERE id = ?", (ticket_id,))
        visit_date = request.form.get("visit_date", date.today().isoformat())
        technician = request.form.get("technician", "").strip()
        result = request.form.get("result", "").strip()
        notes = request.form.get("notes", "").strip()
        next_action = request.form.get("next_action", "").strip()

        g.db.execute(
            """
            INSERT INTO ticket_visits (ticket_id, visit_date, technician, result, notes, next_action)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (ticket_id, visit_date, technician, result, notes, next_action),
        )
        if next_action:
            g.db.execute(
                "UPDATE tickets SET next_action = ?, updated_at = ? WHERE id = ?",
                (next_action, datetime.now().isoformat(timespec="seconds"), ticket_id),
            )
        g.db.commit()
        flash("Visita registada.", "success")
        if ticket:
            return redirect(url_for("tickets", asset_id=ticket["asset_id"]))
        return redirect(url_for("tickets"))

    @app.route("/tickets/<int:ticket_id>/delete", methods=["POST"])
    def delete_ticket(ticket_id: int):
        ticket = query_one("SELECT asset_id FROM tickets WHERE id = ?", (ticket_id,))
        if ticket is None:
            flash("Intervencao nao encontrada.", "error")
            return redirect(url_for("tickets"))

        g.db.execute("DELETE FROM tickets WHERE id = ?", (ticket_id,))
        g.db.commit()
        flash("Intervencao apagada.", "success")
        return redirect(url_for("tickets", asset_id=ticket["asset_id"]))

    @app.route("/exports", methods=["GET", "POST"])
    def exports() -> str:
        def period_redirect_params(period: ReportingPeriod) -> dict[str, str]:
            params = {
                "report_month": period.start.strftime("%Y-%m"),
                "period_type": period.period_type.value,
                "report_year": str(period.start.year),
            }
            if period.period_type == ReportPeriodType.QUARTERLY:
                params["report_quarter"] = str(((period.start.month - 1) // 3) + 1)
            if period.period_type == ReportPeriodType.SEMIANNUAL:
                params["report_semester"] = "1" if period.start.month == 1 else "2"
            return params

        if request.method == "POST":
            asset_id_raw = request.form.get("asset_id", "").strip()
            try:
                report_period = period_from_form(request.form)
            except ReportingPeriodError as exc:
                flash(str(exc), "error")
                redirect_values = request.form.to_dict(flat=True)
                if asset_id_raw:
                    redirect_values["asset_id"] = asset_id_raw
                return redirect(url_for("exports", **redirect_values))
            if report_period.start > date.today():
                flash("Periodo futuro sem dados disponiveis.", "error")
                redirect_values = request.form.to_dict(flat=True)
                if asset_id_raw:
                    redirect_values["asset_id"] = asset_id_raw
                return redirect(url_for("exports", **redirect_values))

            report_month = report_period.start.strftime("%Y-%m")
            force_api = request.form.get("force_api") == "on"
            include_availability = request.form.get("include_availability") == "on"
            report_assets = get_fusionsolar_report_assets(g.db)
            try:
                selection = validate_report_asset_selection(report_assets, asset_id_raw)
                action = request.form.get("action", "generate_report").strip()
                source = parse_billing_values_source(request.form.get("billing_values_source", "saved"))
                if action not in {"generate_report", "save_billing_config"}:
                    raise BillingValidationError("Acao invalida.")
                manual_config = (
                    parse_billing_config_form(dict(request.form), selection.report_type)
                    if action == "save_billing_config" or source == "manual"
                    else None
                )
            except BillingValidationError as exc:
                flash(str(exc), "error")
                redirect_values = period_redirect_params(report_period)
                if asset_id_raw:
                    redirect_values["asset_id"] = asset_id_raw
                if request.form.get("include_availability") == "on":
                    redirect_values["include_availability"] = "on"
                return redirect(url_for("exports", **redirect_values))
            asset_id = selection.asset_id
            if action == "save_billing_config":
                if manual_config is None:
                    flash("Configuracao de cobranca invalida.", "error")
                    return redirect(url_for("exports", asset_id=asset_id, **period_redirect_params(report_period)))
                upsert_asset_billing_config(g.db, asset_id=asset_id, config=manual_config)
                g.db.commit()
                flash("Configuracao de cobranca guardada.", "success")
                return redirect(url_for("exports", asset_id=asset_id, **period_redirect_params(report_period)))
            billing_config = (
                get_asset_billing_config(g.db, asset_id, selection.report_type)
                if source == "saved"
                else manual_config
            )
            if billing_config is None:
                flash("Configuracao de cobranca invalida.", "error")
                return redirect(url_for("exports", asset_id=asset_id, **period_redirect_params(report_period)))
            try:
                report = build_fusionsolar_customer_production_report(
                    g.db,
                    asset_id=asset_id,
                    report_month=report_month,
                    electricity_price=float(billing_config.electricity_price_eur_kwh),
                    sell_price=float(billing_config.export_price_eur_kwh),
                    solcor_price_per_kwh=float(billing_config.solcor_price_per_kwh),
                    billing_config=billing_config,
                    force_api=force_api,
                    period=report_period,
                )
                if include_availability:
                    add_customer_report_availability(g.db, report, asset_id=asset_id, period=report_period)
                return export_customer_production_pdf(report)
            except Exception as exc:
                redirect_params = {
                    "asset_id": str(asset_id),
                    "include_availability": "on" if include_availability else "",
                    "electricity_price": str(billing_config.electricity_price_eur_kwh),
                    "sell_price": str(billing_config.export_price_eur_kwh),
                    "solcor_price_per_kwh": str(billing_config.solcor_price_per_kwh),
                    "fixed_monthly_fee_eur": str(billing_config.fixed_monthly_fee_eur),
                    "billing_mode": billing_config.billing_mode.value,
                    "billing_energy_base": billing_config.billing_energy_base.value,
                    **period_redirect_params(report_period),
                }
                if is_fusionsolar_rate_limit_error(exc):
                    flash(mark_fusionsolar_performance_rate_limited(g.db), "warning")
                    g.db.commit()
                    return redirect(url_for("exports", **redirect_params))
                if "FusionSolar API temporariamente limitada" in str(exc):
                    flash(str(exc), "warning")
                    return redirect(url_for("exports", **redirect_params))
                current_app.logger.exception("Falha ao gerar relatorio de producao")
                flash(f"Falha ao gerar relatorio de producao: {exc}", "error")
                return redirect(url_for("exports", **redirect_params))

        selected_asset_id = request.args.get("asset_id", "").strip()
        period_type = request.args.get("period_type") or request.args.get("report_period_type") or ReportPeriodType.MONTHLY.value
        raw_report_month = request.args.get("report_month", "")
        report_month = reporting_normalize_report_month(raw_report_month)
        try:
            report_period = period_from_form(
                {
                    "period_type": period_type,
                    "report_month": report_month,
                    "report_year": request.args.get("report_year") or report_month[:4],
                    "report_quarter": request.args.get("report_quarter") or "1",
                    "report_semester": request.args.get("report_semester") or "1",
                }
            )
        except ReportingPeriodError as exc:
            flash(str(exc), "error")
            report_period = monthly_period(report_month) if period_type == ReportPeriodType.MONTHLY.value and raw_report_month else monthly_period(date.today().strftime("%Y-%m"))
        report_assets = get_fusionsolar_report_assets(g.db)
        if selected_asset_id:
            try:
                validate_report_asset_selection(report_assets, selected_asset_id)
            except BillingValidationError as exc:
                flash(str(exc), "error")
                return redirect(url_for("exports", **period_redirect_params(report_period)))
        selected_report_type = next(
            (
                asset["report_type"]
                for asset in report_assets
                if selected_asset_id.isdigit() and int(asset["asset_id"]) == int(selected_asset_id)
            ),
            "",
        )
        selected_billing_config = (
            get_asset_billing_config(g.db, int(selected_asset_id), ReportType(selected_report_type))
            if selected_asset_id.isdigit() and selected_report_type
            else BillingConfig(report_type=ReportType.EPC)
        )
        selected_billing_config_exists = (
            get_asset_billing_config_row(g.db, int(selected_asset_id)) is not None
            if selected_asset_id.isdigit()
            else False
        )
        billing_form = billing_config_to_form_values(selected_billing_config)

        return render_template(
            "exports.html",
            report_assets=report_assets,
            selected_asset_id=selected_asset_id,
            selected_report_type=selected_report_type,
            period_type=report_period.period_type.value,
            report_month=report_period.start.strftime("%Y-%m"),
            report_year=str(report_period.start.year),
            report_quarter=str(((report_period.start.month - 1) // 3) + 1),
            report_semester="1" if report_period.start.month == 1 else "2",
            electricity_price=request.args.get("electricity_price", billing_form["electricity_price"]),
            sell_price=request.args.get("sell_price", billing_form["sell_price"]),
            solcor_price_per_kwh=request.args.get("solcor_price_per_kwh", billing_form["solcor_price_per_kwh"]),
            fixed_monthly_fee_eur=billing_form["fixed_monthly_fee_eur"],
            billing_mode=billing_form["billing_mode"],
            billing_energy_base=billing_form["billing_energy_base"],
            include_availability=request.args.get("include_availability") == "on",
            selected_billing_config_exists=selected_billing_config_exists,
            fusionsolar_api_warning=get_fusionsolar_performance_cooldown_reason(g.db),
        )

    @app.route("/portfolio-manager")
    def portfolio_manager() -> str:
        groups = list_portfolios(g.db, include_archived=True)
        selected_portfolio_id = int(request.args.get("portfolio_id", groups[0]["id"] if groups else 0) or 0)
        selected = get_portfolio(g.db, selected_portfolio_id) if selected_portfolio_id else None
        active_tab = request.args.get("tab", "installations").strip()
        if active_tab not in {"installations", "mappings", "aliases", "import", "settings"}:
            active_tab = "installations"
        search = request.args.get("search", "").strip()
        asset_filter = request.args.get("asset_filter", "available").strip()
        alias_search = request.args.get("alias_search", "").strip()
        alias_filter = request.args.get("alias_filter", "all").strip()
        import_id = int(request.args.get("import_id", "0") or 0)
        if import_id and request.args.get("tab") is None:
            active_tab = "import"
        members = list_portfolio_members(g.db, selected_portfolio_id) if selected else []
        assets = portfolio_list_available_assets(
            g.db,
            portfolio_id=selected_portfolio_id or None,
            search=search,
            asset_filter=asset_filter,
        )
        all_assets = portfolio_list_available_assets(g.db, portfolio_id=selected_portfolio_id or None, asset_filter="all")
        portfolio_counts = {
            int(row["portfolio_id"]): int(row["count"] or 0)
            for row in query_all(
                g.db,
                "SELECT portfolio_id, COUNT(*) AS count FROM portfolio_assets WHERE active = 1 GROUP BY portfolio_id",
            )
        }
        available_total = query_scalar(
            g.db,
            """
            SELECT COUNT(*)
            FROM assets a
            WHERE NOT EXISTS (
                SELECT 1 FROM portfolio_assets pa
                WHERE pa.portfolio_id = ? AND pa.asset_id = a.id AND pa.active = 1
            )
            """,
            (selected_portfolio_id,),
        ) if selected else 0
        pending_count = sum(1 for member in members if not member["asset_id"] or member["mapping_status"] == "mapping_pending")
        conflict_count = sum(1 for member in members if member["mapping_status"] == "mapping_conflict")
        aliases = []
        if active_tab == "aliases":
            aliases = portfolio_list_aliases(g.db, include_inactive=alias_filter != "active")
            if alias_search:
                alias_search_lower = alias_search.lower()
                aliases = [alias for alias in aliases if alias_search_lower in str(alias["alias_name"] or "").lower() or alias_search_lower in str(alias["project_name"] or "").lower()]
            if alias_filter == "inactive":
                aliases = [alias for alias in aliases if not alias["active"]]
        conflicts = detect_portfolio_conflicts(g.db, selected_portfolio_id or None)
        import_run = get_import_run(g.db, import_id) if import_id else None
        import_preview = import_preview_from_json(import_run["preview_json"] or "{}") if import_run else None
        mapping_context = portfolio_mapping_context(g.db)
        mapping_decisions = {
            int(member["id"]): portfolio_suggest_mapping(
                g.db,
                external_name=member["external_name"] or "",
                nif=member["nif"] or "",
                context=mapping_context,
            )
            for member in members
            if not member["asset_id"] or member["mapping_status"] in {"mapping_pending", "mapping_conflict", "mapping_suggested"}
        }
        return render_template(
            "portfolio_manager.html",
            title="Gestor de portfolios",
            groups=groups,
            selected=selected,
            active_tab=active_tab,
            selected_portfolio_id=selected_portfolio_id,
            portfolio_counts=portfolio_counts,
            available_total=available_total,
            pending_count=pending_count,
            conflict_count=conflict_count,
            members=members,
            assets=assets,
            all_assets=all_assets,
            aliases=aliases,
            conflicts=conflicts,
            search=search,
            asset_filter=asset_filter,
            alias_search=alias_search,
            alias_filter=alias_filter,
            import_run=import_run,
            import_preview=import_preview,
            mapping_decisions=mapping_decisions,
        )

    @app.route("/portfolio-manager/create", methods=["POST"])
    def portfolio_manager_create():
        try:
            portfolio_id = create_portfolio(g.db, name=request.form.get("name", ""), description=request.form.get("description", ""), notes=request.form.get("notes", ""))
            g.db.commit()
            flash("Portfolio criado.", "success")
            return _portfolio_manager_redirect(portfolio_id, tab="installations")
        except Exception as exc:
            g.db.rollback()
            flash(f"Falha ao criar portfolio: {_portfolio_error_message(exc)}", "error")
            return _portfolio_manager_redirect(0)

    @app.route("/portfolio-manager/update", methods=["POST"])
    def portfolio_manager_update():
        portfolio_id = int(request.form.get("portfolio_id", "0") or 0)
        try:
            update_portfolio(
                g.db,
                portfolio_id=portfolio_id,
                name=request.form.get("name", ""),
                description=request.form.get("description", ""),
                notes=request.form.get("notes", ""),
                display_order=int(request.form.get("display_order", "0") or 0) or None,
            )
            g.db.commit()
            flash("Portfolio atualizado.", "success")
        except Exception as exc:
            g.db.rollback()
            flash(f"Falha ao atualizar portfolio: {_portfolio_error_message(exc)}", "error")
        return _portfolio_manager_redirect(portfolio_id, tab="settings")

    @app.route("/portfolio-manager/duplicate", methods=["POST"])
    def portfolio_manager_duplicate():
        portfolio_id = int(request.form.get("portfolio_id", "0") or 0)
        try:
            new_id = duplicate_portfolio(g.db, portfolio_id=portfolio_id, new_name=request.form.get("new_name", ""))
            g.db.commit()
            flash("Portfolio duplicado.", "success")
            return _portfolio_manager_redirect(new_id, tab="installations")
        except Exception as exc:
            g.db.rollback()
            flash(f"Falha ao duplicar portfolio: {_portfolio_error_message(exc)}", "error")
            return _portfolio_manager_redirect(portfolio_id, tab="settings")

    @app.route("/portfolio-manager/archive", methods=["POST"])
    def portfolio_manager_archive():
        portfolio_id = int(request.form.get("portfolio_id", "0") or 0)
        try:
            archive_portfolio(g.db, portfolio_id)
            g.db.commit()
            flash("Portfolio arquivado.", "success")
        except Exception as exc:
            g.db.rollback()
            flash(f"Falha ao arquivar portfolio: {_portfolio_error_message(exc)}", "error")
        return _portfolio_manager_redirect(portfolio_id, tab="settings")

    @app.route("/portfolio-manager/reactivate", methods=["POST"])
    def portfolio_manager_reactivate():
        portfolio_id = int(request.form.get("portfolio_id", "0") or 0)
        try:
            reactivate_portfolio(g.db, portfolio_id)
            g.db.commit()
            flash("Portfolio reativado.", "success")
        except Exception as exc:
            g.db.rollback()
            flash(f"Falha ao reativar portfolio: {_portfolio_error_message(exc)}", "error")
        return _portfolio_manager_redirect(portfolio_id, tab="settings")

    @app.route("/portfolio-manager/delete", methods=["POST"])
    def portfolio_manager_delete():
        portfolio_id = int(request.form.get("portfolio_id", "0") or 0)
        try:
            delete_portfolio(g.db, portfolio_id, confirm_name=request.form.get("confirm_name", ""))
            g.db.commit()
            flash("Portfolio apagado.", "success")
            return _portfolio_manager_redirect(0)
        except Exception as exc:
            g.db.rollback()
            flash(f"Falha ao apagar portfolio: {_portfolio_error_message(exc)}", "error")
            return _portfolio_manager_redirect(portfolio_id, tab="settings")

    @app.route("/portfolio-manager/members/add", methods=["POST"])
    def portfolio_manager_members_add():
        portfolio_id = int(request.form.get("portfolio_id", "0") or 0)
        try:
            asset_ids = _form_int_list("asset_ids")
            reactivated = 0
            if asset_ids:
                for asset_id in asset_ids:
                    inactive = g.db.execute(
                        "SELECT 1 FROM portfolio_assets WHERE portfolio_id = ? AND asset_id = ? AND active = 0",
                        (portfolio_id, asset_id),
                    ).fetchone()
                    portfolio_add_member(g.db, portfolio_id=portfolio_id, asset_id=asset_id)
                    if inactive:
                        reactivated += 1
            else:
                asset_raw = request.form.get("asset_id", "").strip()
                asset_id = int(asset_raw) if asset_raw.isdigit() else None
                inactive = (
                    g.db.execute(
                        "SELECT 1 FROM portfolio_assets WHERE portfolio_id = ? AND asset_id = ? AND active = 0",
                        (portfolio_id, asset_id),
                    ).fetchone()
                    if asset_id
                    else None
                )
                portfolio_add_member(
                    g.db,
                    portfolio_id=portfolio_id,
                    asset_id=asset_id,
                    external_name=request.form.get("external_name", ""),
                    nif=request.form.get("nif", ""),
                    sub_account=request.form.get("sub_account", ""),
                    notes=request.form.get("notes", ""),
                )
                if inactive:
                    reactivated += 1
            g.db.commit()
            flash("Instalacao reativada no portfolio." if reactivated else "Instalacao adicionada ao portfolio.", "success")
        except Exception as exc:
            g.db.rollback()
            flash(f"Falha ao adicionar instalacao: {_portfolio_error_message(exc)}", "error")
        return _portfolio_manager_redirect(portfolio_id, tab="installations")

    @app.route("/portfolio-manager/members/apply", methods=["POST"])
    def portfolio_manager_members_apply():
        portfolio_id = int(request.form.get("portfolio_id", "0") or 0)
        try:
            asset_ids = _form_int_list("asset_ids")
            asset_names = {
                asset_id: request.form.get(f"asset_name_{asset_id}", "")
                for asset_id in asset_ids
            }
            result = sync_portfolio_asset_members(g.db, portfolio_id=portfolio_id, asset_ids=asset_ids, asset_names=asset_names)
            g.db.commit()
            flash(f"Alteracoes aplicadas: {result['selected']} instalacoes selecionadas.", "success")
        except Exception as exc:
            g.db.rollback()
            flash(f"Falha ao aplicar alteracoes: {_portfolio_error_message(exc)}", "error")
        return _portfolio_manager_redirect(portfolio_id, tab="installations")

    @app.route("/portfolio-manager/members/remove", methods=["POST"])
    def portfolio_manager_members_remove():
        portfolio_id = int(request.form.get("portfolio_id", "0") or 0)
        try:
            remove_members(g.db, portfolio_id=portfolio_id, member_ids=_form_int_list("member_ids"))
            g.db.commit()
            flash("Instalacoes removidas do portfolio.", "success")
        except Exception as exc:
            g.db.rollback()
            flash(f"Falha ao remover instalacoes: {_portfolio_error_message(exc)}", "error")
        return _portfolio_manager_redirect(portfolio_id, tab="installations")

    @app.route("/portfolio-manager/members/copy", methods=["POST"])
    def portfolio_manager_members_copy():
        source_id = int(request.form.get("portfolio_id", "0") or 0)
        target_id = int(request.form.get("target_portfolio_id", "0") or 0)
        try:
            copy_members(g.db, source_portfolio_id=source_id, target_portfolio_id=target_id, member_ids=_form_int_list("member_ids"), move=False)
            g.db.commit()
            flash("Instalacoes copiadas.", "success")
        except Exception as exc:
            g.db.rollback()
            flash(f"Falha ao copiar instalacoes: {_portfolio_error_message(exc)}", "error")
        return _portfolio_manager_redirect(source_id, tab="installations")

    @app.route("/portfolio-manager/members/move", methods=["POST"])
    def portfolio_manager_members_move():
        source_id = int(request.form.get("portfolio_id", "0") or 0)
        target_id = int(request.form.get("target_portfolio_id", "0") or 0)
        try:
            copy_members(g.db, source_portfolio_id=source_id, target_portfolio_id=target_id, member_ids=_form_int_list("member_ids"), move=True)
            g.db.commit()
            flash("Instalacoes movidas.", "success")
            return _portfolio_manager_redirect(target_id, tab="installations")
        except Exception as exc:
            g.db.rollback()
            flash(f"Falha ao mover instalacoes: {_portfolio_error_message(exc)}", "error")
            return _portfolio_manager_redirect(source_id, tab="installations")

    @app.route("/portfolio-manager/members/reorder", methods=["POST"])
    def portfolio_manager_members_reorder():
        portfolio_id = int(request.form.get("portfolio_id", "0") or 0)
        try:
            ordered = _form_int_list("ordered_ids")
            if not ordered:
                member_id = int(request.form.get("member_id", "0") or 0)
                move_member_up_down(g.db, portfolio_id=portfolio_id, member_id=member_id, direction=request.form.get("direction", "down"))
            else:
                reorder_members(g.db, portfolio_id=portfolio_id, ordered_ids=ordered)
            g.db.commit()
            flash("Ordem guardada.", "success")
        except Exception as exc:
            g.db.rollback()
            flash(f"Falha ao reordenar: {_portfolio_error_message(exc)}", "error")
        return _portfolio_manager_redirect(portfolio_id, tab="installations")

    @app.route("/portfolio-manager/members/update", methods=["POST"])
    def portfolio_manager_members_update():
        portfolio_id = int(request.form.get("portfolio_id", "0") or 0)
        member_id = int(request.form.get("member_id", "0") or 0)
        asset_raw = request.form.get("asset_id", "").strip()
        try:
            portfolio_update_member(
                g.db,
                member_id=member_id,
                portfolio_id=portfolio_id,
                asset_id=int(asset_raw) if asset_raw.isdigit() else None,
                external_name=request.form.get("external_name", ""),
                nif=request.form.get("nif", ""),
                sub_account=request.form.get("sub_account", ""),
                notes=request.form.get("notes", ""),
                active=request.form.get("active", "on") == "on",
                create_alias=request.form.get("create_alias") == "on",
            )
            g.db.commit()
            flash("Entrada atualizada.", "success")
        except Exception as exc:
            g.db.rollback()
            flash(f"Falha ao atualizar entrada: {_portfolio_error_message(exc)}", "error")
        return _portfolio_manager_redirect(portfolio_id, tab=request.form.get("tab") or "installations", focus=f"member-{member_id}")

    @app.route("/portfolio-manager/mappings/suggest", methods=["POST"])
    def portfolio_manager_mappings_suggest():
        portfolio_id = int(request.form.get("portfolio_id", "0") or 0)
        try:
            result = auto_map_portfolio_assets(g.db, portfolio_id=portfolio_id or None)
            g.db.commit()
            shared_nif_note = f" {result.get('shared_nif', 0)} com NIF partilhado para confirmar." if result.get("shared_nif") else ""
            duplicate_asset_note = f" {result.get('duplicate_asset', 0)} apontam para uma central ja presente neste portfolio." if result.get("duplicate_asset") else ""
            flash(f"Sugestoes calculadas: {result['mapped']} auto, {result['pending']} pendentes, {result['conflicts']} conflitos.{shared_nif_note}{duplicate_asset_note}", "success")
        except Exception as exc:
            g.db.rollback()
            flash(f"Falha ao sugerir mappings: {_portfolio_error_message(exc)}", "error")
        return _portfolio_manager_redirect(portfolio_id, tab="mappings")

    @app.route("/portfolio-manager/mappings/confirm", methods=["POST"])
    def portfolio_manager_mappings_confirm():
        portfolio_id = int(request.form.get("portfolio_id", "0") or 0)
        try:
            confirm_mapping(
                g.db,
                member_id=int(request.form.get("member_id", "0") or 0),
                portfolio_id=portfolio_id,
                asset_id=int(request.form.get("asset_id", "0") or 0),
                create_alias=request.form.get("create_alias", "on") == "on",
            )
            g.db.commit()
            flash("Mapping confirmado.", "success")
        except Exception as exc:
            g.db.rollback()
            flash(f"Falha ao confirmar mapping: {_portfolio_error_message(exc)}", "error")
        return _portfolio_manager_redirect(portfolio_id, tab=request.form.get("tab") or "mappings", focus=f"member-{request.form.get('member_id', '0')}")

    @app.route("/portfolio-manager/mappings/unmap", methods=["POST"])
    def portfolio_manager_mappings_unmap():
        portfolio_id = int(request.form.get("portfolio_id", "0") or 0)
        try:
            unmap_member(g.db, member_id=int(request.form.get("member_id", "0") or 0), portfolio_id=portfolio_id)
            g.db.commit()
            flash("Mapping removido.", "success")
        except Exception as exc:
            g.db.rollback()
            flash(f"Falha ao remover mapping: {_portfolio_error_message(exc)}", "error")
        return _portfolio_manager_redirect(portfolio_id, tab=request.form.get("tab") or "installations", focus=f"member-{request.form.get('member_id', '0')}")

    @app.route("/assets/<int:asset_id>/aliases/add", methods=["POST"])
    def portfolio_alias_add(asset_id: int):
        try:
            upsert_alias(g.db, asset_id=asset_id, alias_name=request.form.get("alias_name", ""), source="manual", notes=request.form.get("notes", ""))
            portfolio_rebuild_asset_alias_blob(g.db, asset_id)
            g.db.commit()
            flash("Alias guardado.", "success")
        except Exception as exc:
            g.db.rollback()
            flash(f"Falha ao guardar alias: {_portfolio_error_message(exc)}", "error")
        return _alias_return(asset_id)

    @app.route("/assets/<int:asset_id>/aliases/update", methods=["POST"])
    def portfolio_alias_update(asset_id: int):
        try:
            portfolio_update_alias(g.db, asset_id=asset_id, alias_id=int(request.form.get("alias_id", "0") or 0), alias_name=request.form.get("alias_name", ""), notes=request.form.get("notes", ""))
            portfolio_rebuild_asset_alias_blob(g.db, asset_id)
            g.db.commit()
            flash("Alias atualizado.", "success")
        except Exception as exc:
            g.db.rollback()
            flash(f"Falha ao atualizar alias: {_portfolio_error_message(exc)}", "error")
        return _alias_return(asset_id)

    @app.route("/assets/<int:asset_id>/aliases/toggle", methods=["POST"])
    def portfolio_alias_toggle(asset_id: int):
        try:
            toggle_alias(g.db, asset_id=asset_id, alias_id=int(request.form.get("alias_id", "0") or 0), active=request.form.get("active") == "1")
            portfolio_rebuild_asset_alias_blob(g.db, asset_id)
            g.db.commit()
            flash("Alias atualizado.", "success")
        except Exception as exc:
            g.db.rollback()
            flash(f"Falha ao alterar alias: {_portfolio_error_message(exc)}", "error")
        return _alias_return(asset_id)

    @app.route("/assets/<int:asset_id>/aliases/delete", methods=["POST"])
    def portfolio_alias_delete(asset_id: int):
        try:
            portfolio_delete_alias(g.db, asset_id=asset_id, alias_id=int(request.form.get("alias_id", "0") or 0))
            portfolio_rebuild_asset_alias_blob(g.db, asset_id)
            g.db.commit()
            flash("Alias apagado.", "success")
        except Exception as exc:
            g.db.rollback()
            flash(f"Falha ao apagar alias: {_portfolio_error_message(exc)}", "error")
        return _alias_return(asset_id)

    @app.route("/portfolio-manager/import", methods=["POST"])
    def portfolio_manager_import():
        portfolio_id = int(request.form.get("portfolio_id", "0") or 0) or None
        upload = request.files.get("file")
        if upload is None or not upload.filename:
            flash("Escolhe um ficheiro CSV ou XLSX.", "error")
            return redirect(url_for("portfolio_manager", portfolio_id=portfolio_id or 0))
        try:
            import_id = create_import_preview(g.db, portfolio_id=portfolio_id, original_filename=Path(upload.filename).name, data=upload.read())
            g.db.commit()
            flash("Preview de importacao criado.", "success")
            return _portfolio_manager_redirect(portfolio_id or 0, tab="import", import_id=import_id)
        except Exception as exc:
            g.db.rollback()
            flash(f"Falha ao importar configuracao: {_portfolio_error_message(exc)}", "error")
            return _portfolio_manager_redirect(portfolio_id or 0, tab="import")

    @app.route("/portfolio-manager/import/<int:import_id>/apply", methods=["POST"])
    def portfolio_manager_import_apply(import_id: int):
        try:
            run = get_import_run(g.db, import_id)
            selected_rows = _form_int_list("row_numbers")
            overrides = {
                int(key.removeprefix("asset_override_")): int(value)
                for key, value in request.form.items()
                if key.startswith("asset_override_") and str(value).isdigit()
            }
            apply_import_run(g.db, import_id, selected_rows=selected_rows or None, asset_overrides=overrides)
            g.db.commit()
            flash("Importacao aplicada.", "success")
            return _portfolio_manager_redirect(run["portfolio_id"] if run else 0, tab="import")
        except Exception as exc:
            g.db.rollback()
            flash(f"Falha ao aplicar importacao: {_portfolio_error_message(exc)}", "error")
            return _portfolio_manager_redirect(None, tab="import", import_id=import_id)

    @app.route("/portfolio-manager/export")
    def portfolio_manager_export():
        workbook = export_configuration_workbook(g.db)
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name="portfolio_configuration.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.route("/portfolios", methods=["GET", "POST"])
    def portfolios() -> str:
        if request.method == "POST":
            action = request.form.get("action", "").strip()
            report_month_redirect = normalize_report_month(request.form.get("report_month", ""))
            portfolio_id_redirect = request.form.get("portfolio_id", "").strip()
            try:
                if action == "auto_map":
                    portfolio_id = int(request.form.get("portfolio_id", "0") or 0)
                    result = auto_map_portfolio_assets(g.db, portfolio_id=portfolio_id or None)
                    g.db.commit()
                    shared_nif_note = f" {result.get('shared_nif', 0)} com NIF partilhado para confirmar." if result.get("shared_nif") else ""
                    duplicate_asset_note = f" {result.get('duplicate_asset', 0)} apontam para uma central ja presente neste portfolio." if result.get("duplicate_asset") else ""
                    flash(
                        f"Auto-mapping concluido: {result['mapped']} mapeadas, {result['pending']} pendentes, {result['conflicts']} conflitos.{shared_nif_note}{duplicate_asset_note}",
                        "success",
                    )
                    return redirect(url_for("portfolios", tab="config", portfolio_id=portfolio_id, report_month=report_month_redirect))
                if action == "seed_external":
                    seed_external_portfolio_rows(g.db)
                    g.db.commit()
                    flash("Lista externa do portfolio importada/atualizada.", "success")
                    return redirect(url_for("portfolios", tab="config", portfolio_id=portfolio_id_redirect, report_month=report_month_redirect))
                if action == "add_asset":
                    portfolio_id = int(request.form.get("portfolio_id", "0") or 0)
                    asset_id_raw = request.form.get("asset_id", "").strip()
                    external_name = request.form.get("external_name", "").strip()
                    nif = request.form.get("nif", "").strip()
                    sub_account = request.form.get("sub_account", "").strip()
                    mapping = (
                        {"asset_id": int(asset_id_raw), "mapping_status": "manual", "mapping_confidence": 1.0}
                        if asset_id_raw
                        else map_external_portfolio_entity(g.db, nif=nif, external_name=external_name)
                    )
                    if not mapping["asset_id"]:
                        flash("Nao foi possivel mapear automaticamente. Escolhe uma instalacao manualmente.", "error")
                    else:
                        g.db.execute(
                            """
                            INSERT INTO portfolio_assets (
                                portfolio_id, asset_id, external_name, nif, sub_account, active, mapping_status, mapping_confidence, notes
                            ) VALUES (?, ?, ?, ?, ?, 1, ?, ?, ?)
                            ON CONFLICT(portfolio_id, asset_id) DO UPDATE SET
                                external_name = excluded.external_name,
                                nif = excluded.nif,
                                sub_account = excluded.sub_account,
                                active = 1,
                                mapping_status = excluded.mapping_status,
                                mapping_confidence = excluded.mapping_confidence,
                                notes = excluded.notes
                            """,
                            (
                                portfolio_id,
                                mapping["asset_id"],
                                external_name,
                                nif,
                                sub_account,
                                mapping["mapping_status"],
                                mapping["mapping_confidence"],
                                request.form.get("notes", "").strip(),
                            ),
                        )
                        g.db.commit()
                        flash("Instalacao adicionada ao portfolio.", "success")
                    portfolio_id_redirect = str(portfolio_id)
                elif action == "update_asset":
                    portfolio_asset_id = int(request.form.get("portfolio_asset_id", "0") or 0)
                    asset_id_raw = request.form.get("asset_id", "").strip()
                    g.db.execute(
                        """
                        UPDATE portfolio_assets
                        SET asset_id = ?, external_name = ?, nif = ?, sub_account = ?, active = ?,
                            mapping_status = ?, mapping_confidence = ?, notes = ?
                        WHERE id = ?
                        """,
                        (
                            int(asset_id_raw) if asset_id_raw.isdigit() else None,
                            request.form.get("external_name", "").strip(),
                            request.form.get("nif", "").strip(),
                            request.form.get("sub_account", "").strip(),
                            1 if request.form.get("active") == "on" else 0,
                            "manual" if asset_id_raw.isdigit() else "mapping_pending",
                            1.0 if asset_id_raw.isdigit() else 0.0,
                            request.form.get("notes", "").strip(),
                            portfolio_asset_id,
                        ),
                    )
                    g.db.commit()
                    flash("Mapeamento atualizado.", "success")
                elif action == "save_tariff":
                    asset_id = int(request.form.get("asset_id", "0") or 0)
                    tariff_id_raw = request.form.get("tariff_id", "").strip()
                    invoice_file_id_raw = request.form.get("invoice_file_id", "").strip()
                    saved_tariff_id = save_asset_tariff(
                        g.db,
                        tariff_id=int(tariff_id_raw) if tariff_id_raw.isdigit() else None,
                        asset_id=asset_id,
                        tariff_type=request.form.get("tariff_type", "simple"),
                        cycle_type=request.form.get("cycle_type", ""),
                        simple_price_eur_kwh=request.form.get("simple_price_eur_kwh"),
                        ponta_price_eur_kwh=request.form.get("ponta_price_eur_kwh"),
                        cheia_price_eur_kwh=request.form.get("cheia_price_eur_kwh"),
                        vazio_price_eur_kwh=request.form.get("vazio_price_eur_kwh"),
                        super_vazio_price_eur_kwh=request.form.get("super_vazio_price_eur_kwh"),
                        invoice_file_id=int(invoice_file_id_raw) if invoice_file_id_raw.isdigit() else None,
                        valid_from=request.form.get("valid_from", "").strip(),
                        valid_to=request.form.get("valid_to", "").strip(),
                        notes=request.form.get("notes", "").strip(),
                    )
                    template_name = request.form.get("tri_hourly_template", "").strip()
                    if template_name:
                        apply_tri_hourly_template(g.db, tariff_id=saved_tariff_id, template_name=template_name)
                    g.db.commit()
                    flash("Tarifa guardada.", "success")
                    return redirect(url_for("portfolios", tab="config", portfolio_id=portfolio_id_redirect, report_month=report_month_redirect, focus=f"tariff-{saved_tariff_id}"))
                elif action == "duplicate_tariff":
                    asset_id = int(request.form.get("asset_id", "0") or 0)
                    new_tariff_id = duplicate_asset_tariff(
                        g.db,
                        tariff_id=int(request.form.get("tariff_id", "0") or 0),
                        asset_id=asset_id,
                    )
                    g.db.commit()
                    flash("Tarifa duplicada. RevÃª as datas de validade antes de usar.", "success")
                    return redirect(url_for("portfolios", tab="config", portfolio_id=portfolio_id_redirect, report_month=report_month_redirect, focus=f"tariff-{new_tariff_id}"))
                elif action == "delete_tariff":
                    asset_id = int(request.form.get("asset_id", "0") or 0)
                    delete_asset_tariff(
                        g.db,
                        tariff_id=int(request.form.get("tariff_id", "0") or 0),
                        asset_id=asset_id,
                        confirm=request.form.get("confirm_delete") == "on",
                    )
                    g.db.commit()
                    flash("Tarifa apagada.", "success")
                    return redirect(url_for("portfolios", tab="config", portfolio_id=portfolio_id_redirect, report_month=report_month_redirect, focus=f"asset-{asset_id}"))
                elif action == "add_tariff_rule":
                    tariff_id = int(request.form.get("tariff_id", "0") or 0)
                    add_tariff_period_rule(
                        g.db,
                        tariff_id=tariff_id,
                        weekday_type=request.form.get("weekday_type", "all"),
                        start_time=request.form.get("start_time", ""),
                        end_time=request.form.get("end_time", ""),
                        period_name=request.form.get("period_name", ""),
                    )
                    g.db.commit()
                    flash("Regra horaria adicionada.", "success")
                    return redirect(url_for("portfolios", tab="config", portfolio_id=portfolio_id_redirect, report_month=report_month_redirect, focus=f"tariff-{tariff_id}"))
                elif action == "update_tariff_rule":
                    tariff_id = int(request.form.get("tariff_id", "0") or 0)
                    update_tariff_period_rule(
                        g.db,
                        rule_id=int(request.form.get("rule_id", "0") or 0),
                        tariff_id=tariff_id,
                        weekday_type=request.form.get("weekday_type", "all"),
                        start_time=request.form.get("start_time", ""),
                        end_time=request.form.get("end_time", ""),
                        period_name=request.form.get("period_name", ""),
                    )
                    g.db.commit()
                    flash("Regra horaria atualizada.", "success")
                    return redirect(url_for("portfolios", tab="config", portfolio_id=portfolio_id_redirect, report_month=report_month_redirect, focus=f"tariff-{tariff_id}"))
                elif action == "delete_tariff_rule":
                    asset_id = int(request.form.get("asset_id", "0") or 0)
                    tariff_id = int(request.form.get("tariff_id", "0") or 0)
                    deleted = delete_tariff_period_rule(
                        g.db,
                        rule_id=int(request.form.get("rule_id", "0") or 0),
                        asset_id=asset_id or None,
                    )
                    g.db.commit()
                    flash("Regra horaria removida." if deleted else "Regra horaria nao encontrada.", "success" if deleted else "error")
                    return redirect(url_for("portfolios", tab="config", portfolio_id=portfolio_id_redirect, report_month=report_month_redirect, focus=f"tariff-{tariff_id or asset_id}"))
            except Exception as exc:
                g.db.rollback()
                flash(f"Falha ao guardar portfolio: {exc}", "error")
            return redirect(url_for("portfolios", tab="config", portfolio_id=portfolio_id_redirect, report_month=report_month_redirect))

        groups = query_all(g.db, "SELECT * FROM portfolio_groups ORDER BY name COLLATE NOCASE")
        selected_portfolio_id = int(request.args.get("portfolio_id", groups[0]["id"] if groups else 0) or 0)
        report_month = normalize_report_month(request.args.get("report_month", ""))
        active_tab = request.args.get("tab", "report").strip()
        warning_filter = request.args.get("warning_filter", "all").strip()
        config_filter = request.args.get("config_filter", "all").strip()
        report_rows_all = build_portfolio_report_rows(g.db, selected_portfolio_id, report_month) if selected_portfolio_id else []
        portfolio_name = next((row["name"] for row in groups if int(row["id"]) == selected_portfolio_id), "")
        for row in report_rows_all:
            row["portfolio"] = portfolio_name
        total_row = aggregate_portfolio_total(report_rows_all) if report_rows_all else None
        if total_row:
            total_row["portfolio"] = portfolio_name
        report_rows = filter_report_rows(report_rows_all, warning_filter)
        kpis = build_portfolio_kpis(report_rows_all, total_row)
        portfolio_assets = query_all(
            g.db,
            """
            SELECT pa.*, pg.name AS portfolio_name, a.project_name, a.nif AS asset_nif, a.mounting_date,
                   sf.original_filename AS latest_helioscope_file,
                   inv.original_filename AS latest_invoice_file
            FROM portfolio_assets pa
            JOIN portfolio_groups pg ON pg.id = pa.portfolio_id
            LEFT JOIN assets a ON a.id = pa.asset_id
            LEFT JOIN source_files sf ON sf.id = (
                SELECT id FROM source_files
                WHERE asset_id = pa.asset_id AND file_type IN ('financial_model', 'helioscope')
                ORDER BY uploaded_at DESC, id DESC LIMIT 1
            )
            LEFT JOIN source_files inv ON inv.id = (
                SELECT id FROM source_files
                WHERE asset_id = pa.asset_id AND file_type = 'invoice'
                ORDER BY uploaded_at DESC, id DESC LIMIT 1
            )
            ORDER BY pg.name COLLATE NOCASE, COALESCE(pa.external_name, a.project_name) COLLATE NOCASE
            """,
        )
        assets = query_all(g.db, "SELECT id, project_name, nif FROM assets ORDER BY project_name COLLATE NOCASE")
        invoices = query_all(g.db, "SELECT id, asset_id, original_filename FROM source_files WHERE file_type = 'invoice' ORDER BY uploaded_at DESC")
        invoice_documents = list_portfolio_invoice_documents(g.db, selected_portfolio_id) if selected_portfolio_id else []
        tariff_rules = query_all(g.db, "SELECT * FROM tariff_period_rules ORDER BY tariff_id, weekday_type, start_time")
        rules_by_tariff: dict[int, list[sqlite3.Row]] = {}
        for rule in tariff_rules:
            rules_by_tariff.setdefault(int(rule["tariff_id"]), []).append(rule)
        rules_by_tariff_json = {
            tariff_id: [
                {
                    "weekday_type": rule["weekday_type"],
                    "start_time": rule["start_time"],
                    "end_time": rule["end_time"],
                    "period_name": rule["period_name"],
                }
                for rule in rules
            ]
            for tariff_id, rules in rules_by_tariff.items()
        }
        report_start, report_end = month_bounds(report_month)
        config_assets = [dict(row) for row in portfolio_assets if int(row["portfolio_id"]) == selected_portfolio_id]
        tariffs_by_asset: dict[int, list[sqlite3.Row]] = {}
        for item in config_assets:
            asset_id = int(item["asset_id"] or 0)
            if not asset_id:
                item["tariff_summary"] = "Sem mapping"
                continue
            tariffs = list_asset_tariffs(g.db, asset_id)
            tariffs_by_asset[asset_id] = tariffs
            applicable = list_tariffs_at(g.db, asset_id=asset_id, moment=report_start)
            warnings = detect_tariff_validity_warnings(g.db, asset_id=asset_id, start=report_start, end=report_end)
            item["tariff_warnings"] = warnings
            item["tariff_warning_labels"] = [TARIFF_WARNING_MESSAGES.get(warning, warning) for warning in warnings]
            if len(applicable) == 1:
                item["tariff_summary"] = applicable[0]["tariff_type"]
                item["applicable_tariff_id"] = int(applicable[0]["id"])
            elif len(applicable) > 1:
                item["tariff_summary"] = "Conflito de tarifas"
                item["applicable_tariff_id"] = None
            else:
                item["tariff_summary"] = "Sem tarifa aplicavel"
                item["applicable_tariff_id"] = None
        if config_filter == "pending":
            config_assets = [row for row in config_assets if not row["asset_id"] or row["mapping_status"] == "mapping_pending"]
        elif config_filter == "conflicts":
            config_assets = [row for row in config_assets if row["mapping_status"] == "mapping_conflict"]
        return render_template(
            "portfolios.html",
            title="Portfolios",
            groups=groups,
            selected_portfolio_id=selected_portfolio_id,
            report_month=report_month,
            active_tab=active_tab if active_tab in {"report", "config"} else "report",
            warning_filter=warning_filter,
            config_filter=config_filter,
            rows=report_rows,
            total_row=total_row,
            kpis=kpis,
            portfolio_assets=portfolio_assets,
            config_assets=config_assets,
            assets=assets,
            invoices=invoices,
            invoice_documents=invoice_documents,
            tariff_rules=tariff_rules,
            rules_by_tariff=rules_by_tariff,
            rules_by_tariff_json=rules_by_tariff_json,
            tariffs_by_asset=tariffs_by_asset,
            report_start=report_start,
            report_end=report_end,
        )

    @app.route("/portfolios/export-mapping")
    def export_portfolio_mapping():
        portfolio_id = int(request.args.get("portfolio_id", "0") or 0)
        group = g.db.execute("SELECT * FROM portfolio_groups WHERE id = ?", (portfolio_id,)).fetchone()
        rows = query_all(
            g.db,
            """
            SELECT pg.name AS portfolio_name, pa.sub_account, pa.nif, pa.external_name,
                   a.project_name AS local_installation, pa.mapping_confidence, pa.mapping_status,
                   sf.original_filename AS helioscope_file, inv.original_filename AS invoice_file,
                   at.tariff_type
            FROM portfolio_assets pa
            JOIN portfolio_groups pg ON pg.id = pa.portfolio_id
            LEFT JOIN assets a ON a.id = pa.asset_id
            LEFT JOIN source_files sf ON sf.id = (
                SELECT id FROM source_files
                WHERE asset_id = pa.asset_id AND file_type IN ('financial_model', 'helioscope')
                ORDER BY uploaded_at DESC, id DESC LIMIT 1
            )
            LEFT JOIN source_files inv ON inv.id = (
                SELECT id FROM source_files
                WHERE asset_id = pa.asset_id AND file_type = 'invoice'
                ORDER BY uploaded_at DESC, id DESC LIMIT 1
            )
            LEFT JOIN asset_tariffs at ON at.id = (
                SELECT id FROM asset_tariffs
                WHERE asset_id = pa.asset_id
                ORDER BY COALESCE(valid_from, '') DESC, id DESC LIMIT 1
            )
            WHERE pa.portfolio_id = ?
            ORDER BY pa.sub_account
            """,
            (portfolio_id,),
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Mapping"
        headers = ["Portfolio", "Subconta", "NIF", "Nome externo", "Instalacao local", "Confianca", "Estado", "Helioscope", "Fatura", "Tarifa"]
        sheet.append(headers)
        for row in rows:
            sheet.append([
                row["portfolio_name"],
                row["sub_account"],
                row["nif"],
                row["external_name"],
                row["local_installation"] or "",
                row["mapping_confidence"],
                row["mapping_status"],
                row["helioscope_file"] or "",
                row["invoice_file"] or "",
                row["tariff_type"] or "",
            ])
        for column in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[column[0].column_letter].width = min(max(width + 2, 12), 42)
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        name = (group["name"] if group else f"portfolio_{portfolio_id}").replace(" ", "_")
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"{name}_mapping.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.route("/portfolios/upload-helioscope", methods=["POST"])
    def upload_portfolio_helioscope():
        asset_id = int(request.form.get("asset_id", "0") or 0)
        portfolio_id_raw = request.form.get("portfolio_id", "").strip()
        report_month = normalize_report_month(request.form.get("report_month", ""))
        upload = request.files.get("file")
        if not asset_id or upload is None or not upload.filename:
            flash("Escolhe a instalacao e o ficheiro financeiro/Helioscope.", "error")
            return redirect(url_for("portfolios"))
        try:
            result = import_helioscope_file(
                g.db,
                upload_dir=UPLOAD_DIR,
                file_storage=upload,
                asset_id=asset_id,
                portfolio_id=int(portfolio_id_raw) if portfolio_id_raw.isdigit() else None,
            )
            g.db.commit()
            flash(f"Helioscope importado: {result['months']} meses.", "success")
        except Exception as exc:
            g.db.rollback()
            flash(f"Falha ao importar Helioscope: {exc}", "error")
        return redirect(url_for("portfolios", tab="config", portfolio_id=portfolio_id_raw, report_month=report_month))

    @app.route("/portfolios/financial-import/<int:source_file_id>")
    def portfolio_financial_import_review(source_file_id: int):
        source = g.db.execute(
            """
            SELECT sf.*, a.project_name, pg.name AS portfolio_name
            FROM source_files sf
            JOIN assets a ON a.id = sf.asset_id
            LEFT JOIN portfolio_groups pg ON pg.id = sf.portfolio_id
            WHERE sf.id = ? AND sf.file_type IN ('financial_model', 'helioscope')
            """,
            (source_file_id,),
        ).fetchone()
        if source is None:
            flash("Import financeiro nao encontrado.", "error")
            return redirect(url_for("portfolios"))
        try:
            summary = json.loads(source["notes"] or "{}")
        except json.JSONDecodeError:
            summary = {}
        monthly = query_all(
            g.db,
            """
            SELECT month, expected_kwh
            FROM helioscope_expected_production
            WHERE source_file_id = ?
            ORDER BY month
            """,
            (source_file_id,),
        )
        interval_count = query_scalar(
            g.db,
            "SELECT COUNT(*) FROM helioscope_expected_interval_production WHERE source_file_id = ?",
            (source_file_id,),
        ) or 0
        return render_template(
            "financial_import_review.html",
            title="Revisao import financeiro",
            source=source,
            summary=summary,
            monthly=monthly,
            interval_count=interval_count,
            portfolio_id=request.args.get("portfolio_id", ""),
            report_month=request.args.get("report_month", ""),
        )

    @app.route("/portfolios/upload-invoice", methods=["POST"])
    def upload_portfolio_invoice():
        asset_id = int(request.form.get("asset_id", "0") or 0)
        portfolio_id_raw = request.form.get("portfolio_id", "").strip()
        report_month = normalize_report_month(request.form.get("report_month", ""))
        upload = request.files.get("file")
        if not asset_id or upload is None or not upload.filename:
            flash("Escolhe a instalacao e o ficheiro fonte da fatura.", "error")
            return redirect(url_for("portfolios"))
        stored_path: Path | None = None
        created_new_file = False
        try:
            document_id, warnings = store_invoice_upload(
                g.db,
                upload_dir=UPLOAD_DIR,
                file_storage=upload,
                asset_id=asset_id,
                portfolio_id=int(portfolio_id_raw) if portfolio_id_raw.isdigit() else None,
            )
            created_new_file = "duplicate_invoice" not in warnings
            if created_new_file:
                invoice = get_invoice_document(g.db, document_id)
                stored_path = Path(invoice["stored_path"]) if invoice is not None else None
            g.db.commit()
            suffix = " Avisos: " + ", ".join(warnings) if warnings else ""
            flash(f"Fatura guardada para revisao (doc {document_id}).{suffix}", "success")
        except Exception as exc:
            g.db.rollback()
            if created_new_file and stored_path is not None:
                stored_path.unlink(missing_ok=True)
            flash(f"Falha ao guardar fatura: {exc}", "error")
        return redirect(url_for("portfolios", tab="config", portfolio_id=portfolio_id_raw, report_month=report_month))

    @app.route("/invoices/<int:invoice_document_id>/extract", methods=["POST"])
    def extract_invoice_document_route(invoice_document_id: int):
        invoice = get_invoice_document(g.db, invoice_document_id)
        if invoice is None:
            abort(404)
        try:
            result = extract_invoice_file(Path(invoice["stored_path"]))
            persist_invoice_extraction_result(g.db, invoice, result)
            g.db.commit()
            flash("Extracao assistida concluida. RevÃª os valores antes de usar.", "success")
        except Exception as exc:
            g.db.rollback()
            flash(f"Falha na extracao: {exc}", "error")
        return redirect(url_for("review_invoice_document", invoice_document_id=invoice_document_id))

    @app.route("/invoices/<int:invoice_document_id>/review", methods=["GET", "POST"])
    def review_invoice_document(invoice_document_id: int):
        invoice = get_invoice_document(g.db, invoice_document_id)
        if invoice is None:
            abort(404)
        asset = g.db.execute("SELECT id, project_name, nif FROM assets WHERE id = ?", (invoice["asset_id"],)).fetchone()
        if request.method == "POST":
            action = request.form.get("action", "").strip()
            try:
                if action == "reject_invoice":
                    reject_invoice_document(g.db, invoice_document_id)
                    flash("Fatura rejeitada.", "success")
                elif action == "archive_invoice":
                    archive_invoice_document(g.db, invoice_document_id)
                    flash("Fatura arquivada.", "success")
                elif action in {"save_invoice_review", "confirm_invoice", "confirm_with_warnings"}:
                    values = invoice_values_from_form(request.form)
                    validation = validate_invoice_values(values, asset_nif=asset["nif"] if asset else None)
                    if not validation.valid:
                        raise ValueError("; ".join(validation.errors))
                    requires_override = warnings_require_override(validation.warnings)
                    if action == "confirm_invoice" and requires_override:
                        raise ValueError("Confirmacao exige override explicito para estes avisos: " + ", ".join(validation.warnings))
                    update_invoice_review(
                        g.db,
                        invoice_document_id=invoice_document_id,
                        values=values,
                        warnings=validation.warnings,
                        status=InvoiceStatus.CONFIRMED.value if action in {"confirm_invoice", "confirm_with_warnings"} else InvoiceStatus.REVIEW_REQUIRED.value,
                    )
                    if action == "confirm_with_warnings":
                        flash("Fatura confirmada com avisos preservados.", "warning")
                    elif action == "confirm_invoice":
                        flash("Fatura confirmada.", "success")
                    else:
                        flash("Revisao guardada.", "success")
                elif action == "use_invoice_tariff":
                    apply_invoice_to_tariff(g.db, invoice_document_id)
                    flash("Tarifa criada a partir da fatura confirmada.", "success")
                else:
                    raise ValueError("Acao invalida.")
                g.db.commit()
            except Exception as exc:
                g.db.rollback()
                flash(f"Falha na revisao da fatura: {exc}", "error")
            return redirect(url_for("review_invoice_document", invoice_document_id=invoice_document_id))
        runs = list_invoice_extraction_runs(g.db, invoice_document_id)
        return render_template("invoice_review.html", invoice=invoice, asset=asset, runs=runs, title="Revisao de fatura")

    @app.route("/portfolio-reports")
    def portfolio_reports() -> str:
        groups = query_all(g.db, "SELECT * FROM portfolio_groups ORDER BY name COLLATE NOCASE")
        selected_portfolio_id = int(request.args.get("portfolio_id", groups[0]["id"] if groups else 0) or 0)
        report_month = normalize_report_month(request.args.get("report_month", ""))
        portfolio_name = next((row["name"] for row in groups if int(row["id"]) == selected_portfolio_id), "")
        period_type = request.args.get("period_type", "monthly").strip() or "monthly"
        report_year = request.args.get("report_year") or report_month[:4]
        report_quarter = request.args.get("report_quarter") or str(((int(report_month[5:7]) - 1) // 3) + 1)
        report_semester = request.args.get("report_semester") or ("1" if int(report_month[5:7]) <= 6 else "2")
        comparison = request.args.get("comparison", "").strip()
        profiles = list_portfolio_report_profiles(g.db, selected_portfolio_id)
        selected_profile_id = int(request.args.get("profile_id", "0") or 0)
        profile = get_portfolio_report_profile(g.db, selected_profile_id) if selected_profile_id else None
        if profile is None:
            profile = get_default_portfolio_report_profile(g.db, selected_portfolio_id)
            selected_profile_id = int(profile.id or 0)
        portfolio_result = None
        rows = []
        total_row = None
        if selected_portfolio_id:
            try:
                portfolio_result = prepare_portfolio_report(
                    g.db,
                    portfolio_id=selected_portfolio_id,
                    portfolio_name=portfolio_name,
                    profile=profile,
                    period_type=period_type,
                    report_month=report_month,
                    year=report_year,
                    quarter=report_quarter,
                    semester=report_semester,
                    comparison=comparison,
                    profile_version=latest_portfolio_report_profile_version(g.db, profile.id),
                )
                portfolio_result = ensure_portfolio_result_data_requests(
                    g.db,
                    portfolio_result,
                    request_source="portfolio_report_page",
                )
            except Exception as exc:
                flash(f"Falha ao preparar relatorio configuravel: {exc}", "error")
                portfolio_result = None
            if portfolio_result is None:
                rows = build_portfolio_report_rows(g.db, selected_portfolio_id, report_month)
                for row in rows:
                    row["portfolio"] = portfolio_name
                total_row = aggregate_portfolio_total(rows) if rows else None
                if total_row:
                    total_row["portfolio"] = portfolio_name
        runs = list_portfolio_report_history(g.db, selected_portfolio_id, limit=20)
        return render_template(
            "portfolio_reports.html",
            title="Relatorios de portfolio",
            groups=groups,
            selected_portfolio_id=selected_portfolio_id,
            report_month=report_month,
            report_year=report_year,
            report_quarter=report_quarter,
            report_semester=report_semester,
            period_type=period_type,
            comparison=comparison,
            profiles=profiles,
            selected_profile_id=selected_profile_id,
            selected_profile=profile,
            metric_catalog=METRIC_CATALOG,
            portfolio_result=portfolio_result,
            rows=rows,
            total_row=total_row,
            runs=runs,
        )

    @app.route("/portfolio-reports/profiles", methods=["POST"])
    def save_portfolio_report_profile_route():
        action = request.form.get("action", "save").strip()
        portfolio_id = int(request.form.get("portfolio_id", "0") or 0) or None
        profile_id = int(request.form.get("profile_id", "0") or 0)
        try:
            if action == "archive" and profile_id:
                archive_portfolio_report_profile(g.db, profile_id)
                flash("Perfil arquivado.", "success")
            elif action == "set_default" and profile_id:
                set_default_portfolio_report_profile(g.db, profile_id)
                flash("Perfil predefinido atualizado.", "success")
            elif action == "duplicate" and profile_id:
                duplicate_portfolio_report_profile(g.db, profile_id, request.form.get("name", "").strip() or "Copia")
                flash("Perfil duplicado.", "success")
            else:
                base = get_portfolio_report_profile(g.db, profile_id) if profile_id else get_default_portfolio_report_profile(g.db, portfolio_id)
                selected_metrics = request.form.getlist("metric_key") or [column.metric_key for column in base.columns if column.visible]
                config = profile_to_config(base)
                config["name"] = request.form.get("name", "").strip() or base.name
                config["description"] = request.form.get("description", "").strip()
                config["portfolio_id"] = portfolio_id
                config["period_type"] = request.form.get("default_period_type", "monthly")
                config["comparison"] = request.form.get("default_comparison", "")
                selected_metrics = sorted(
                    selected_metrics,
                    key=lambda metric_key: int(request.form.get(f"order_{metric_key}", "9999") or 9999),
                )
                config["columns"] = [
                    {
                        "metric_key": metric_key,
                        "label": request.form.get(f"label_{metric_key}", "").strip() or metric_key,
                        "decimals": request.form.get(f"decimals_{metric_key}", ""),
                        "visible": True,
                        "display_order": index * 10,
                    }
                    for index, metric_key in enumerate(selected_metrics, start=1)
                ]
                profile = profile_from_config(config, profile_id=profile_id or None, portfolio_id=portfolio_id)
                profile_id = save_portfolio_report_profile(g.db, profile)
                flash("Perfil guardado.", "success")
            g.db.commit()
        except Exception as exc:
            g.db.rollback()
            flash(f"Falha ao guardar perfil: {exc}", "error")
        return redirect(url_for("portfolio_reports", portfolio_id=portfolio_id or 0, profile_id=profile_id))

    @app.route("/portfolio-reports/generate", methods=["POST"])
    def generate_portfolio_report():
        portfolio_id = int(request.form.get("portfolio_id", "0") or 0)
        report_month = normalize_report_month(request.form.get("report_month", ""))
        period_type = request.form.get("period_type", "monthly").strip() or "monthly"
        profile_id = int(request.form.get("profile_id", "0") or 0)
        comparison = request.form.get("comparison", "").strip()
        return_to = request.form.get("return_to", "").strip()
        try:
            group = g.db.execute("SELECT * FROM portfolio_groups WHERE id = ?", (portfolio_id,)).fetchone()
            if not group:
                raise ValueError("Portfolio invalido.")
            profile = get_portfolio_report_profile(g.db, profile_id) if profile_id else get_default_portfolio_report_profile(g.db, portfolio_id)
            result = prepare_portfolio_report(
                g.db,
                portfolio_id=portfolio_id,
                portfolio_name=group["name"],
                profile=profile,
                period_type=period_type,
                report_month=report_month,
                year=request.form.get("report_year") or report_month[:4],
                quarter=request.form.get("report_quarter"),
                semester=request.form.get("report_semester"),
                comparison=comparison,
                profile_version=latest_portfolio_report_profile_version(g.db, profile.id),
            )
            result = ensure_portfolio_result_data_requests(
                g.db,
                result,
                request_source="portfolio_report_snapshot",
            )
            report_id = snapshot_portfolio_result(g.db, result, request.form.get("notes", "").strip())
            g.db.commit()
            flash(f"Relatorio snapshot #{report_id} gerado.", "success")
        except Exception as exc:
            g.db.rollback()
            flash(f"Falha ao gerar relatorio: {exc}", "error")
        if return_to == "portfolios":
            return redirect(url_for("portfolios", portfolio_id=portfolio_id, report_month=report_month, tab="report"))
        return redirect(url_for("portfolio_reports", portfolio_id=portfolio_id, report_month=report_month, profile_id=profile_id, period_type=period_type, comparison=comparison))

    @app.route("/portfolio-reports/export")
    def export_portfolio_report():
        snapshot_id = int(request.args.get("snapshot_id", "0") or 0)
        if snapshot_id:
            result = get_portfolio_snapshot_result(g.db, snapshot_id)
            if result:
                workbook = export_portfolio_result_workbook(result)
                buffer = io.BytesIO()
                workbook.save(buffer)
                buffer.seek(0)
                return send_file(
                    buffer,
                    as_attachment=True,
                    download_name=f"portfolio_snapshot_{snapshot_id}.xlsx",
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
        portfolio_id = int(request.args.get("portfolio_id", "0") or 0)
        report_month = normalize_report_month(request.args.get("report_month", ""))
        group = g.db.execute("SELECT * FROM portfolio_groups WHERE id = ?", (portfolio_id,)).fetchone()
        profile_id = int(request.args.get("profile_id", "0") or 0)
        if group and (profile_id or request.args.get("period_type") or request.args.get("comparison")):
            profile = get_portfolio_report_profile(g.db, profile_id) if profile_id else get_default_portfolio_report_profile(g.db, portfolio_id)
            result = prepare_portfolio_report(
                g.db,
                portfolio_id=portfolio_id,
                portfolio_name=group["name"],
                profile=profile,
                period_type=request.args.get("period_type", "monthly"),
                report_month=report_month,
                year=request.args.get("report_year") or report_month[:4],
                quarter=request.args.get("report_quarter"),
                semester=request.args.get("report_semester"),
                comparison=request.args.get("comparison", ""),
                profile_version=latest_portfolio_report_profile_version(g.db, profile.id),
            )
            result = ensure_portfolio_result_data_requests(
                g.db,
                result,
                request_source="portfolio_report_export",
            )
            workbook = export_portfolio_result_workbook(result)
        else:
            rows = build_portfolio_report_rows(g.db, portfolio_id, report_month) if group else []
            for row in rows:
                row["portfolio"] = group["name"] if group else ""
            if rows:
                total = aggregate_portfolio_total(rows)
                total["portfolio"] = group["name"] if group else ""
                rows.append(total)
            workbook = export_portfolio_report_workbook(rows)
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"portfolio_{portfolio_id}_{report_month}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.route("/report-templates", methods=["GET", "POST"])
    def report_templates():
        if request.method == "POST":
            action = request.form.get("action", "save").strip()
            template_id = int(request.form.get("template_id", "0") or 0)
            portfolio_id = int(request.form.get("portfolio_id", "0") or 0) or None
            try:
                if action == "archive" and template_id:
                    archive_report_template(g.db, template_id)
                    flash("Template arquivado.", "success")
                elif action == "duplicate" and template_id:
                    existing = get_template(g.db, template_id)
                    if existing:
                        validate_template_scope(existing, existing.report_type, portfolio_id=existing.portfolio_id, client_key=existing.client_key, allow_inactive=True)
                    duplicate_report_template(g.db, template_id, request.form.get("name", "").strip() or "Copia")
                    flash("Template duplicado.", "success")
                elif action == "set_default" and template_id:
                    existing = get_template(g.db, template_id)
                    if existing:
                        validate_template_scope(existing, existing.report_type, portfolio_id=existing.portfolio_id, client_key=existing.client_key)
                    set_default_template(g.db, template_id)
                    flash("Template default atualizado.", "success")
                else:
                    base = get_template(g.db, template_id) if template_id else default_template("Portfolio executivo" if request.form.get("report_type") == "portfolio" else "Individual padrao", portfolio_id=portfolio_id)
                    if base.id:
                        validate_template_scope(base, base.report_type, portfolio_id=base.portfolio_id, client_key=base.client_key, allow_inactive=True)
                    logo_path = base.branding.logo_path
                    logo_file = request.files.get("logo")
                    old_logo_path = logo_path
                    if logo_file and logo_file.filename:
                        logo_path = store_report_logo(logo_file, old_logo_path=old_logo_path)
                    config = template_to_config(base)
                    config.update(
                        name=request.form.get("name", "").strip() or base.name,
                        report_type=request.form.get("report_type", base.report_type),
                        description=request.form.get("description", "").strip(),
                        portfolio_id=portfolio_id,
                        orientation=request.form.get("orientation", base.orientation),
                        title=request.form.get("title", "").strip() or base.title,
                        subtitle=request.form.get("subtitle", "").strip(),
                        filename_pattern=request.form.get("filename_pattern", "").strip() or base.filename_pattern,
                        branding={
                            **config.get("branding", {}),
                            "company_name": request.form.get("company_name", "").strip() or base.branding.company_name,
                            "client_name": request.form.get("client_name", "").strip(),
                            "primary_color": request.form.get("primary_color", base.branding.primary_color),
                            "secondary_color": request.form.get("secondary_color", base.branding.secondary_color),
                            "footer": request.form.get("footer", "").strip(),
                            "contacts": request.form.get("contacts", "").strip(),
                            "disclaimer": request.form.get("disclaimer", "").strip(),
                            "logo_path": logo_path,
                        },
                    )
                    enabled_sections = set(request.form.getlist("section_key"))
                    config["sections"] = [
                        {**section, "enabled": section["key"] in enabled_sections, "display_order": int(request.form.get(f"order_{section['key']}", section["display_order"]) or section["display_order"])}
                        for section in config["sections"]
                    ]
                    template = template_from_config(config, template_id=template_id or None, portfolio_id=portfolio_id)
                    template_id = save_template(g.db, template, is_default=1 if request.form.get("is_default") == "on" else 0)
                    flash("Template guardado.", "success")
                g.db.commit()
            except Exception as exc:
                g.db.rollback()
                LOGGER.warning("report_template_action_failed action=%s template_id=%s error_code=%s", action, template_id, type(exc).__name__)
                flash("Falha no template. Verifica os dados submetidos.", "error")
            return redirect(url_for("report_templates", template_id=template_id))
        templates = list_templates(g.db, include_inactive=True)
        selected_id = int(request.args.get("template_id", templates[0]["id"] if templates else 0) or 0)
        selected_template = get_template(g.db, selected_id) if selected_id else default_template("Portfolio executivo")
        groups = query_all(g.db, "SELECT * FROM portfolio_groups ORDER BY name COLLATE NOCASE")
        return render_template("report_templates.html", title="Templates de relatorio", templates=templates, selected_template=selected_template, groups=groups)

    @app.route("/report-generation", methods=["GET", "POST"])
    def report_generation():
        output_dir = UPLOAD_DIR / "generated_reports"
        output_dir.mkdir(parents=True, exist_ok=True)
        if request.method == "POST":
            report_type = request.form.get("report_type", "portfolio")
            template_id = int(request.form.get("template_id", "0") or 0)
            portfolio_id = int(request.form.get("portfolio_id", "0") or 0) or None
            snapshot_id = int(request.form.get("snapshot_id", "0") or 0) or None
            raw_formats = request.form.getlist("formats") or ["pdf"]
            run_id = None
            try:
                formats = validate_formats(report_type, raw_formats)
                main_formats = tuple(item for item in formats if item != "zip")
                template = get_template(g.db, template_id) if template_id else get_default_template(g.db, report_type, portfolio_id)
                if template is None:
                    raise ValueError("Template invalido.")
                snapshot_result = None
                if snapshot_id:
                    snapshot_result = get_portfolio_snapshot_result(g.db, snapshot_id)
                    if snapshot_result is None or (portfolio_id and snapshot_result.portfolio_id != portfolio_id):
                        raise ValueError("Snapshot invalido.")
                    portfolio_id = snapshot_result.portfolio_id
                    reject_snapshot_period_overrides(request.form, snapshot_result)
                client_key = resolve_generation_client_key(g.db, report_type, request.form, portfolio_id)
                validate_template_scope(template, report_type, portfolio_id=portfolio_id, client_key=client_key)
                if report_type == "portfolio" and not snapshot_id and not portfolio_id:
                    raise ValueError("Portfolio obrigatorio.")
                jobs = build_snapshot_generation_jobs(snapshot_result, main_formats) if snapshot_result else build_generation_jobs(request.form, report_type, main_formats)
                if not jobs:
                    raise ValueError("Pedido sem outputs principais.")
                if len(jobs) > MAX_TOTAL_OUTPUTS:
                    raise ValueError("Demasiados outputs no mesmo run.")
                first_period = jobs[0]["period"]
                run_id = create_generation_run(
                    g.db,
                    template_id=template.id,
                    template_version=latest_template_version(g.db, template.id),
                    report_type=report_type,
                    portfolio_id=portfolio_id,
                    asset_id=jobs[0].get("asset_id"),
                    snapshot_id=snapshot_id,
                    period_type=first_period["period_type"],
                    period_start=first_period["period_start"],
                    period_end=first_period["period_end"],
                    requested_count=len(jobs),
                )
                LOGGER.info("report_generation_run_created run_id=%s report_type=%s template_id=%s portfolio_id=%s snapshot_id=%s requested_count=%s", run_id, report_type, template.id, portfolio_id, snapshot_id, len(jobs))
                g.db.commit()
                completed = 0
                failed = 0
                skipped = 0
                warnings: list[str] = []
                completed_files = []
                if report_type == "portfolio":
                    for job in jobs:
                        try:
                            LOGGER.info("report_generation_item_started run_id=%s format=%s portfolio_id=%s snapshot_id=%s", run_id, job["format"], portfolio_id, snapshot_id)
                            result = snapshot_result or build_portfolio_generation_result(g.db, request.form, portfolio_id, snapshot_id, job["period"])
                            if snapshot_id:
                                portfolio_id = result.portfolio_id
                            rendered = render_portfolio_pdf(result, template) if job["format"] == "pdf" else render_portfolio_excel(result, template)
                            completed_files.append(register_rendered_generation_file(g.db, output_dir, run_id, rendered, snapshot_id=snapshot_id))
                            completed += 1
                            warnings.extend(rendered.warnings)
                            LOGGER.info("report_generation_item_completed run_id=%s format=%s portfolio_id=%s snapshot_id=%s", run_id, job["format"], rendered.portfolio_id, snapshot_id)
                        except Exception as exc:
                            failed += 1
                            LOGGER.warning("report_generation_item_failed run_id=%s format=%s portfolio_id=%s snapshot_id=%s error_code=%s", run_id, job.get("format"), portfolio_id, snapshot_id, type(exc).__name__)
                            add_failed_generation_file(g.db, run_id, job, str(exc), portfolio_id=portfolio_id, snapshot_id=snapshot_id)
                else:
                    for job in jobs:
                        try:
                            LOGGER.info("report_generation_item_started run_id=%s format=%s asset_id=%s", run_id, job["format"], job.get("asset_id"))
                            report = build_individual_generation_report(
                                g.db,
                                job["asset_id"],
                                job["period"],
                                include_wat=any(
                                    section.enabled
                                    and section.key == "availability"
                                    for section in template.sections
                                ),
                            )
                            rendered = render_individual_pdf(report, template) if job["format"] == "pdf" else render_individual_excel(report, template)
                            completed_files.append(register_rendered_generation_file(g.db, output_dir, run_id, rendered))
                            completed += 1
                            warnings.extend(rendered.warnings)
                            LOGGER.info("report_generation_item_completed run_id=%s format=%s asset_id=%s", run_id, job["format"], rendered.asset_id)
                        except Exception as exc:
                            failed += 1
                            LOGGER.warning("report_generation_item_failed run_id=%s format=%s asset_id=%s error_code=%s", run_id, job.get("format"), job.get("asset_id"), type(exc).__name__)
                            add_failed_generation_file(g.db, run_id, job, str(exc), asset_id=job.get("asset_id"))
                if "zip" in formats and completed_files:
                    zip_file = render_zip(completed_files, filename=f"run_{run_id}.zip")
                    register_rendered_generation_file(g.db, output_dir, run_id, zip_file)
                status = "completed" if completed and not failed else ("partial" if completed and failed else "failed")
                finish_generation_run(g.db, run_id, status=status, completed_count=completed, failed_count=failed, skipped_count=skipped, warnings=sorted(set(warnings)), error_message="" if completed else "Todos os outputs falharam.")
                g.db.commit()
                flash(f"Run #{run_id} terminado: {completed} concluÃ­dos, {failed} falhados.", "success" if status == "completed" else "warning")
            except Exception as exc:
                g.db.rollback()
                if run_id:
                    finish_generation_run(g.db, run_id, status="failed", completed_count=0, failed_count=0, error_message=str(exc))
                    g.db.commit()
                LOGGER.warning("report_generation_failed run_id=%s report_type=%s template_id=%s error_code=%s", run_id, report_type, template_id, type(exc).__name__)
                flash("Falha na geraÃ§Ã£o. RevÃª o pedido e consulta os logs para detalhe tÃ©cnico.", "error")
            return redirect(url_for("report_generation"))
        groups = query_all(g.db, "SELECT * FROM portfolio_groups ORDER BY name COLLATE NOCASE")
        assets = query_all(g.db, "SELECT id, project_name FROM assets WHERE active_contract = 'yes' OR active_contract IS NULL ORDER BY project_name COLLATE NOCASE LIMIT 200")
        templates = list_templates(g.db)
        profiles = list_portfolio_report_profiles(g.db)
        page = max(int(request.args.get("page", "1") or 1), 1)
        offset = (page - 1) * 20
        runs = list_generation_runs(g.db, limit=20, offset=offset)
        files = list_generated_files(g.db, limit=50, offset=(page - 1) * 50)
        return render_template("report_generation.html", title="Geracao de relatorios", groups=groups, assets=assets, templates=templates, profiles=profiles, runs=runs, files=files, page=page)

    @app.route("/reporting-health")
    def reporting_health():
        findings = reconcile_generated_reports(g.db, cleanup=False)
        stale_runs = query_scalar(g.db, "SELECT COUNT(*) FROM report_generation_runs WHERE status = 'running' AND created_at < datetime('now', '-2 hours')")
        payload = {
            "database": "ok",
            "storage": "ok" if UPLOAD_DIR.exists() else "missing",
            "defaults": len(list_templates(g.db)) > 0,
            "storage_findings": len([item for item in findings if item.status != "ok"]),
            "stale_runs": stale_runs,
        }
        return payload

    @app.route("/report-generation/preview")
    def report_generation_preview():
        portfolio_id = int(request.args.get("portfolio_id", "0") or 0)
        template_id = int(request.args.get("template_id", "0") or 0)
        template = get_template(g.db, template_id) if template_id else get_default_template(g.db, "portfolio", portfolio_id)
        if template is None:
            abort(404)
        validate_template_scope(template, "portfolio", portfolio_id=portfolio_id, client_key=resolve_report_client_key(g.db, portfolio_id=portfolio_id))
        snapshot_id = int(request.args.get("snapshot_id", "0") or 0)
        if snapshot_id:
            result = get_portfolio_snapshot_result(g.db, snapshot_id)
            if result is None or (portfolio_id and result.portfolio_id != portfolio_id):
                abort(404)
            return render_portfolio_html(result, template)
        group = g.db.execute("SELECT * FROM portfolio_groups WHERE id = ?", (portfolio_id,)).fetchone()
        if not group:
            abort(404)
        profile_id = int(request.args.get("profile_id", "0") or 0)
        profile = get_portfolio_report_profile(g.db, profile_id) if profile_id else get_default_portfolio_report_profile(g.db, portfolio_id)
        report_month = normalize_report_month(request.args.get("report_month", ""))
        period_job = parse_generation_periods(request.args)[0]
        result = prepare_portfolio_report(
            g.db,
            portfolio_id=portfolio_id,
            portfolio_name=group["name"],
            profile=profile,
            period_type=period_job["period_type"],
            report_month=report_month,
            year=period_job.get("report_year") or report_month[:4],
            quarter=period_job.get("report_quarter"),
            semester=period_job.get("report_semester"),
            comparison=request.args.get("comparison", ""),
        )
        result = ensure_portfolio_result_data_requests(
            g.db,
            result,
            request_source="portfolio_report_preview",
        )
        return render_portfolio_html(result, template)

    @app.route("/report-generation/files/<int:file_id>")
    def download_generated_report(file_id: int):
        row = get_generated_file(g.db, file_id)
        if row is None:
            abort(404)
        path = resolve_runtime_file_path_within(row["relative_path"], UPLOAD_DIR / "generated_reports")
        if path is None or not path.exists():
            abort(404)
        if path.is_symlink() or not path_is_within(path.resolve(), (UPLOAD_DIR / "generated_reports").resolve()):
            abort(404)
        if path.stat().st_size != int(row["size_bytes"] or 0):
            abort(404)
        if row["sha256"]:
            import hashlib

            if hashlib.sha256(path.read_bytes()).hexdigest() != row["sha256"]:
                abort(404)
        mimetype = "application/pdf" if row["format"] == "pdf" else ("application/zip" if row["format"] == "zip" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        return send_file(path, as_attachment=True, download_name=row["filename"], mimetype=mimetype)

    @app.route("/integrations", methods=["GET", "POST"])
    def integrations() -> str:
        provider = INTEGRATION_PROVIDER_FUSIONSOLAR
        if request.method == "POST":
            action = request.form.get("action", "").strip()
            if action == "save_config":
                auto_sync_enabled = 1 if request.form.get("auto_sync_enabled") == "on" else 0
                enabled = 1 if request.form.get("enabled") == "on" else 0
                production_sync_enabled = 1 if request.form.get("production_sync_enabled") == "on" else 0
                diagnostics_sync_enabled = 1 if request.form.get("diagnostics_sync_enabled") == "on" else 0
                state_sync_interval_hours = normalize_positive_int(
                    request.form.get("state_sync_interval_hours"),
                    DEFAULT_STATE_SYNC_INTERVAL_HOURS,
                    minimum=1,
                    maximum=24,
                )
                production_sync_time = normalize_clock_time(
                    request.form.get("production_sync_time"),
                    DEFAULT_FUSIONSOLAR_PRODUCTION_SYNC_TIME,
                )
                diagnostics_sync_time = normalize_clock_time(
                    request.form.get("diagnostics_sync_time"),
                    DEFAULT_FUSIONSOLAR_DIAGNOSTICS_SYNC_TIME,
                )
                submitted_password = request.form.get("password", "").strip()
                g.db.execute(
                    """
                    UPDATE integration_configs
                    SET username = ?,
                        password = CASE WHEN ? != '' THEN ? ELSE password END,
                        base_url = ?, login_endpoint = ?, plants_endpoint = ?,
                        real_time_endpoint = ?, device_list_endpoint = ?, device_real_time_endpoint = ?,
                        device_history_endpoint = ?, alarms_endpoint = ?,
                        day_kpi_endpoint = ?, month_kpi_endpoint = ?,
                        enabled = ?, auto_sync_enabled = ?,
                        production_sync_enabled = ?, diagnostics_sync_enabled = ?,
                        state_sync_interval_hours = ?, production_sync_time = ?, diagnostics_sync_time = ?,
                        sync_hours = ?, updated_at = ?
                    WHERE provider = ?
                    """,
                    (
                        request.form.get("username", "").strip(),
                        submitted_password,
                        submitted_password,
                        request.form.get("base_url", "").strip(),
                        request.form.get("login_endpoint", "").strip(),
                        request.form.get("plants_endpoint", "").strip(),
                        request.form.get("real_time_endpoint", "").strip(),
                        request.form.get("device_list_endpoint", "").strip(),
                        request.form.get("device_real_time_endpoint", "").strip(),
                        request.form.get("device_history_endpoint", "").strip(),
                        request.form.get("alarms_endpoint", "").strip(),
                        request.form.get("day_kpi_endpoint", "").strip(),
                        request.form.get("month_kpi_endpoint", "").strip(),
                        enabled,
                        auto_sync_enabled,
                        production_sync_enabled,
                        diagnostics_sync_enabled,
                        state_sync_interval_hours,
                        production_sync_time,
                        diagnostics_sync_time,
                        f"{production_sync_time},{diagnostics_sync_time}",
                        datetime.now().isoformat(timespec="seconds"),
                        provider,
                    ),
                )
                g.db.commit()
                refresh_integration_scheduler(app)
                flash("Configuracao FusionSolar guardada.", "success")
                return redirect(url_for("integrations"))

            if action == "test_connection":
                messages = []
                failures = []
                for item_provider in INTEGRATION_PROVIDER_OPTIONS:
                    item_config = get_integration_config(g.db, item_provider)
                    if item_config is None or not item_config["enabled"]:
                        continue
                    try:
                        result = run_provider_check(g.db, item_provider, dry_run=True)
                        messages.append(
                            f"{item_provider}: {result['station_count']} centrais, {result['realtime_count']} respostas realtime"
                        )
                    except Exception as exc:
                        failures.append(f"{item_provider}: {exc}")
                if messages:
                    flash("Ligacao validada: " + " | ".join(messages), "success")
                if failures:
                    flash("Falha no teste: " + " | ".join(failures), "error")
                if not messages and not failures:
                    flash("Nao ha integracoes ativas para testar.", "warning")
                return redirect(url_for("integrations"))

            if action == "test_fusionsolar_connection":
                try:
                    result = run_fusionsolar_check(g.db, provider, dry_run=True)
                    flash(
                        f"Ligacao FusionSolar validada: {result['station_count']} centrais, {result['realtime_count']} respostas realtime e {result['alarm_count']} alarmes ativos.",
                        "success",
                    )
                except Exception as exc:
                    flash(f"Falha no teste de ligacao FusionSolar: {exc}", "error")
                return redirect(url_for("integrations"))

            if action == "save_sigenergy_config":
                sig_provider = INTEGRATION_PROVIDER_SIGENERGY
                auto_sync_enabled = 1 if request.form.get("auto_sync_enabled") == "on" else 0
                enabled = 1 if request.form.get("enabled") == "on" else 0
                state_sync_interval_hours = normalize_positive_int(
                    request.form.get("state_sync_interval_hours"),
                    DEFAULT_STATE_SYNC_INTERVAL_HOURS,
                    minimum=1,
                    maximum=24,
                )
                submitted_secret = request.form.get("password", "").strip()
                env_secret_configured = bool(os.environ.get("SIGENERGY_APP_SECRET", "").strip())
                g.db.execute(
                    """
                    UPDATE integration_configs
                    SET username = ?,
                        password = CASE WHEN ? = 0 AND ? != '' THEN ? ELSE password END,
                        base_url = ?, login_endpoint = ?, plants_endpoint = ?,
                        energy_flow_endpoint = ?, onboard_endpoint = ?, enabled = ?, auto_sync_enabled = ?,
                        state_sync_interval_hours = ?, sync_hours = ?, region = ?, system_ids = ?, snapshot_retention_days = ?, updated_at = ?
                    WHERE provider = ?
                    """,
                    (
                        request.form.get("username", "").strip(),
                        1 if env_secret_configured else 0,
                        submitted_secret,
                        submitted_secret,
                        request.form.get("base_url", DEFAULT_SIGENERGY_BASE_URL).strip(),
                        request.form.get("login_endpoint", DEFAULT_SIGENERGY_AUTH_ENDPOINT).strip(),
                        request.form.get("plants_endpoint", DEFAULT_SIGENERGY_SYSTEMS_ENDPOINT).strip(),
                        request.form.get("energy_flow_endpoint", DEFAULT_SIGENERGY_ENERGY_FLOW_ENDPOINT).strip(),
                        request.form.get("onboard_endpoint", DEFAULT_SIGENERGY_ONBOARD_ENDPOINT).strip(),
                        enabled,
                        auto_sync_enabled,
                        state_sync_interval_hours,
                        f"{state_sync_interval_hours}:00",
                        request.form.get("region", DEFAULT_SIGENERGY_REGION).strip() or DEFAULT_SIGENERGY_REGION,
                        request.form.get("system_ids", "").strip(),
                        parse_int_value(request.form.get("snapshot_retention_days", str(DEFAULT_SIGENERGY_SNAPSHOT_RETENTION_DAYS))) or DEFAULT_SIGENERGY_SNAPSHOT_RETENTION_DAYS,
                        datetime.now().isoformat(timespec="seconds"),
                        sig_provider,
                    ),
                )
                g.db.commit()
                refresh_integration_scheduler(app)
                flash("Configuracao Sigenergy guardada.", "success")
                return redirect(url_for("integrations") + "#integrations-sigenergy")

            if action == "test_sigenergy_connection":
                try:
                    result = run_sigenergy_check(g.db, INTEGRATION_PROVIDER_SIGENERGY, dry_run=True)
                    flash(
                        f"Ligacao Sigenergy validada: {result['station_count']} instalacoes, {result['realtime_count']} energy flow, {result['failed_realtime_count']} falhas.",
                        "success" if not result["failed_realtime_count"] else "warning",
                    )
                except Exception as exc:
                    flash(f"Falha no teste de ligacao Sigenergy: {exc}", "error")
                return redirect(url_for("integrations") + "#integrations-sigenergy")

            if action == "onboard_sigenergy_system":
                try:
                    config = get_integration_config(g.db, INTEGRATION_PROVIDER_SIGENERGY)
                    if config is None:
                        raise ValueError("Configuracao Sigenergy nao encontrada.")
                    result = create_sigenergy_onboarding_request(
                        g.db,
                        config,
                        request.form.get("system_id", ""),
                        requested_by=str(session.get("username") or ""),
                    )
                    status_label = result.get("status", "failed")
                    flash(
                        f"Pedido Sigenergy registado para {result['system_id']}: {status_label}. {result.get('message', '')}",
                        "success" if status_label != "failed" else "error",
                    )
                except Exception as exc:
                    flash(f"Falha no pedido de acesso Sigenergy: {sigenergy_service.sanitize_sigenergy_error(exc)}", "error")
                return redirect(url_for("integrations") + "#integrations-sigenergy")

            if action == "refresh_sigenergy_onboarding":
                try:
                    result = run_sigenergy_check(g.db, INTEGRATION_PROVIDER_SIGENERGY, dry_run=True)
                    approved = reconcile_sigenergy_onboarding_requests(g.db, result.get("available_system_ids", []))
                    g.db.commit()
                    flash(f"Pedidos Sigenergy atualizados. {approved} aprovados encontrados.", "success")
                except Exception as exc:
                    flash(f"Falha ao atualizar pedidos Sigenergy: {sigenergy_service.sanitize_sigenergy_error(exc)}", "error")
                return redirect(url_for("integrations") + "#integrations-sigenergy")

            if action == "sync_sigenergy_now":
                try:
                    job_id, created = create_background_job(
                        g.db,
                        "sigenergy_state_sync",
                        {"provider": INTEGRATION_PROVIDER_SIGENERGY, "trigger_type": "manual_background"},
                    )
                    g.db.commit()
                    schedule_background_job(app, job_id)
                    flash(
                        f"Sync Sigenergy enviado para background (job #{job_id})." if created else f"Ja existe sync Sigenergy pendente/em execucao (job #{job_id}).",
                        "success" if created else "warning",
                    )
                except Exception as exc:
                    flash(f"Falha ao agendar sync Sigenergy: {sigenergy_service.sanitize_sigenergy_error(exc)}", "error")
                return redirect(url_for("integrations") + "#integrations-sigenergy")

            if action == "sync_now":
                try:
                    queued: list[str] = []
                    queued_job_ids: list[int] = []
                    for item_provider, job_type in (
                        (INTEGRATION_PROVIDER_FUSIONSOLAR, "fusionsolar_state_sync"),
                        (INTEGRATION_PROVIDER_SIGENERGY, "sigenergy_state_sync"),
                    ):
                        item_config = get_integration_config(g.db, item_provider)
                        if item_config is None or not item_config["enabled"]:
                            continue
                        job_id, created = create_background_job(
                            g.db,
                            job_type,
                            {"provider": item_provider, "trigger_type": "manual_background"},
                        )
                        queued.append(f"{item_provider} job #{job_id}" + ("" if created else " existente"))
                        queued_job_ids.append(job_id)
                    g.db.commit()
                    for job_id in queued_job_ids:
                        schedule_background_job(app, job_id)
                    if queued:
                        flash("Sync enviado para background: " + " | ".join(queued), "success")
                    else:
                        flash("Nao ha integracoes ativas para sincronizar.", "warning")
                except Exception as exc:
                    flash(f"Falha ao agendar sincronizacao: {exc}", "error")
                return redirect(url_for("integrations"))

            if action == "sync_provider_state_now":
                provider_value = request.form.get("provider", "").strip()
                if provider_value not in INTEGRATION_PROVIDER_OPTIONS:
                    flash("Integracao invalida.", "error")
                    return redirect(url_for("integrations"))
                job_type = "sigenergy_state_sync" if provider_value == INTEGRATION_PROVIDER_SIGENERGY else "fusionsolar_state_sync"
                try:
                    job_id, created = create_background_job(
                        g.db,
                        job_type,
                        {"provider": provider_value, "trigger_type": "manual_background"},
                    )
                    g.db.commit()
                    schedule_background_job(app, job_id)
                    cooldown = get_provider_cooldown_reason(g.db, provider_value, API_AREA_STATE)
                    suffix = " Vai aguardar pelo fim do cooldown." if cooldown else ""
                    flash(
                        f"Sync de estado {provider_value} enviado para background (job #{job_id}).{suffix}"
                        if created
                        else f"Ja existe sync de estado {provider_value} pendente/em execucao (job #{job_id}).{suffix}",
                        "success" if created else "warning",
                    )
                except Exception as exc:
                    flash(f"Falha ao agendar sync de estado {provider_value}: {exc}", "error")
                return redirect(url_for("integrations"))

            if action == "sync_fusionsolar_production_now":
                target_date = current_lisbon_date() - timedelta(days=1)
                try:
                    job_id, created = create_background_job(
                        g.db,
                        "fusionsolar_production_sync",
                        {
                            "provider": INTEGRATION_PROVIDER_FUSIONSOLAR,
                            "target_date": target_date.isoformat(),
                            "period_type": "day",
                            "trigger_type": "manual_background",
                        },
                    )
                    g.db.commit()
                    schedule_background_job(app, job_id)
                    cooldown = get_provider_cooldown_reason(g.db, INTEGRATION_PROVIDER_FUSIONSOLAR, API_AREA_PRODUCTION)
                    suffix = " Vai aguardar pelo fim do cooldown." if cooldown else ""
                    flash(
                        f"Sync de producao FusionSolar enviado para background (job #{job_id}, dia {target_date.isoformat()}).{suffix}"
                        if created
                        else f"Ja existe sync de producao FusionSolar pendente/em execucao (job #{job_id}).{suffix}",
                        "success" if created else "warning",
                    )
                except Exception as exc:
                    flash(f"Falha ao agendar producao FusionSolar: {exc}", "error")
                return redirect(url_for("integrations") + "#integrations-fusionsolar")

            if action == "sync_fusionsolar_diagnostics_now":
                target_date = current_lisbon_date() - timedelta(days=1)
                try:
                    job_id, created = create_background_job(
                        g.db,
                        "fusionsolar_inverter_availability_backfill",
                        {
                            "provider": INTEGRATION_PROVIDER_FUSIONSOLAR,
                            "from_date": target_date.isoformat(),
                            "to_date": target_date.isoformat(),
                            "trigger_type": "manual_background",
                        },
                    )
                    g.db.commit()
                    schedule_background_job(app, job_id)
                    cooldown = get_provider_cooldown_reason(g.db, INTEGRATION_PROVIDER_FUSIONSOLAR, API_AREA_DIAGNOSTICS)
                    suffix = " Vai aguardar pelo fim do cooldown." if cooldown else ""
                    flash(
                        f"Sync de diagnostics FusionSolar enviado para background (job #{job_id}, dia {target_date.isoformat()}).{suffix}"
                        if created
                        else f"Ja existe sync de diagnostics FusionSolar pendente/em execucao (job #{job_id}).{suffix}",
                        "success" if created else "warning",
                    )
                except Exception as exc:
                    flash(f"Falha ao agendar diagnostics FusionSolar: {exc}", "error")
                return redirect(url_for("integrations") + "#integrations-fusionsolar")

            if action == "clear_api_cooldown":
                provider_value = request.form.get("provider", "").strip()
                api_area = request.form.get("api_area", "").strip() or None
                confirm_value = request.form.get("confirm_clear_cooldown", "").strip()
                if provider_value not in INTEGRATION_PROVIDER_OPTIONS or api_area not in {API_AREA_STATE, API_AREA_PRODUCTION, API_AREA_DIAGNOSTICS, None}:
                    flash("Cooldown invalido.", "error")
                    return redirect(url_for("integrations"))
                if confirm_value != "on":
                    flash("Confirma a limpeza do cooldown antes de continuar.", "warning")
                    return redirect(url_for("integrations"))
                clear_api_cooldown(g.db, provider_value, api_area)
                g.db.commit()
                flash(
                    f"Cooldown {provider_value}{' / ' + api_area if api_area else ''} limpo manualmente.",
                    "success",
                )
                return redirect(url_for("integrations"))

            if action == "test_telegram":
                ok, message = test_telegram_connection()
                flash(message, "success" if ok else "error")
                return redirect(url_for("integrations"))

            if action == "save_alert_settings":
                set_alert_setting(g.db, "TELEGRAM_ALERTS_ENABLED", "true" if request.form.get("telegram_alerts_enabled") == "on" else "false")
                alert_scope = request.form.get("alert_scope", "only_o&m").strip()
                if alert_scope not in ALERT_SCOPE_OPTIONS:
                    alert_scope = "only_o&m"
                set_alert_setting(g.db, "ALERT_SCOPE", alert_scope)
                for key in [
                    "SEND_NEW_ERROR_ALERTS",
                    "SEND_OFFLINE_ALERTS",
                    "SEND_RESOLVED_ALERTS",
                    "SEND_PERSISTENT_ALERTS",
                    "SEND_RECURRENT_ALERTS",
                    "DAYTIME_OFFLINE_ONLY",
                    "IGNORE_HISTORICAL_ALERTS",
                ]:
                    set_alert_setting(g.db, key, "true" if request.form.get(key) == "on" else "false")
                for key in [
                    "MINIMUM_ALERT_SEVERITY",
                    "NEW_ERROR_COOLDOWN_MINUTES",
                    "OFFLINE_COOLDOWN_MINUTES",
                    "RESOLVED_COOLDOWN_MINUTES",
                    "PERSISTENT_COOLDOWN_HOURS",
                    "RECURRENT_COOLDOWN_HOURS",
                ]:
                    set_alert_setting(g.db, key, request.form.get(key, ALERT_SETTING_DEFAULTS.get(key, "")).strip())
                g.db.commit()
                flash("Filtros de alertas guardados.", "success")
                return redirect(url_for("integrations"))

            if action == "set_alert_baseline":
                baseline_at = datetime.now().isoformat(timespec="seconds")
                set_alert_setting(g.db, "ALERT_BASELINE_AT", baseline_at)
                g.db.execute(
                    "INSERT INTO alert_baseline (baseline_at, created_by, notes) VALUES (?, ?, ?)",
                    (baseline_at, session.get("username"), "Baseline definido pela UI."),
                )
                g.db.commit()
                flash("Estado atual definido como baseline de alertas.", "success")
                return redirect(url_for("integrations"))

            if action == "add_alert_blacklist":
                asset_id_raw = request.form.get("asset_id", "").strip()
                reason = request.form.get("reason", "").strip()
                asset_id = int(asset_id_raw) if asset_id_raw else None
                asset_name = request.form.get("asset_name", "").strip()
                if asset_id:
                    asset = query_one("SELECT project_name FROM assets WHERE id = ?", (asset_id,))
                    asset_name = asset["project_name"] if asset else asset_name
                if asset_id or asset_name:
                    g.db.execute(
                        "INSERT INTO alert_blacklist (asset_id, asset_name, reason, created_at, active) VALUES (?, ?, ?, ?, 1)",
                        (asset_id, asset_name, reason, datetime.now().isoformat(timespec="seconds")),
                    )
                    g.db.commit()
                    flash("Instalacao adicionada a blacklist de alertas.", "success")
                return redirect(url_for("integrations"))

            if action == "remove_alert_blacklist":
                blacklist_id = int(request.form["blacklist_id"])
                g.db.execute("UPDATE alert_blacklist SET active = 0 WHERE id = ?", (blacklist_id,))
                g.db.commit()
                flash("Instalacao removida da blacklist.", "success")
                return redirect(url_for("integrations"))

            if action == "quick_alert_action":
                asset_id = int(request.form["asset_id"])
                quick_action = request.form.get("quick_action", "")
                if quick_action == "disable_alerts":
                    g.db.execute("UPDATE assets SET alerts_enabled = 0 WHERE id = ?", (asset_id,))
                elif quick_action == "enable_alerts":
                    g.db.execute("UPDATE assets SET alerts_enabled = 1, monitoring_enabled = 1, monitoring_status = 'active' WHERE id = ?", (asset_id,))
                elif quick_action == "blacklist":
                    asset = query_one("SELECT project_name FROM assets WHERE id = ?", (asset_id,))
                    g.db.execute(
                        "INSERT INTO alert_blacklist (asset_id, asset_name, reason, created_at, active) VALUES (?, ?, ?, ?, 1)",
                        (asset_id, asset["project_name"] if asset else "", "Adicionado por acao rapida.", datetime.now().isoformat(timespec="seconds")),
                    )
                elif quick_action == "unblacklist":
                    g.db.execute("UPDATE alert_blacklist SET active = 0 WHERE asset_id = ?", (asset_id,))
                elif quick_action == "out_of_scope":
                    g.db.execute("UPDATE assets SET monitoring_status = 'out_of_scope' WHERE id = ?", (asset_id,))
                elif quick_action == "silence_24h":
                    g.db.execute(
                        "UPDATE assets SET monitoring_status = 'silenced', silenced_until = ?, silence_reason = ? WHERE id = ?",
                        ((datetime.now() + timedelta(hours=24)).isoformat(timespec="minutes"), "Silenciado 24h pela UI.", asset_id),
                    )
                elif quick_action == "reactivate":
                    g.db.execute("UPDATE assets SET monitoring_status = 'active', silenced_until = '', silence_reason = '' WHERE id = ?", (asset_id,))
                g.db.commit()
                flash("Filtro da instalacao atualizado.", "success")
                return redirect(url_for("integrations"))

            if action == "bulk_alert_action":
                bulk_action = request.form.get("bulk_action", "")
                if bulk_action == "blacklist_non_oem":
                    rows = query_all(g.db, "SELECT id, project_name FROM assets WHERE COALESCE(maintenance, '') NOT IN ('yes', 'true', '1', 'sim')")
                    for row in rows:
                        g.db.execute(
                            "INSERT INTO alert_blacklist (asset_id, asset_name, reason, created_at, active) VALUES (?, ?, ?, ?, 1)",
                            (row["id"], row["project_name"], "Sem Maintenance=yes.", datetime.now().isoformat(timespec="seconds")),
                        )
                elif bulk_action == "disable_no_active_contract":
                    g.db.execute("UPDATE assets SET alerts_enabled = 0 WHERE COALESCE(active_contract, '') != 'yes'")
                elif bulk_action == "enable_only_oem":
                    g.db.execute("UPDATE assets SET alerts_enabled = CASE WHEN COALESCE(maintenance, '') = 'yes' THEN 1 ELSE 0 END")
                    set_alert_setting(g.db, "ALERT_SCOPE", "only_o&m")
                elif bulk_action == "set_baseline":
                    baseline_at = datetime.now().isoformat(timespec="seconds")
                    set_alert_setting(g.db, "ALERT_BASELINE_AT", baseline_at)
                    g.db.execute(
                        "INSERT INTO alert_baseline (baseline_at, created_by, notes) VALUES (?, ?, ?)",
                        (baseline_at, session.get("username"), "Baseline definido por acao em massa."),
                    )
                g.db.commit()
                flash("Acao em massa aplicada.", "success")
                return redirect(url_for("integrations"))

            if action == "resolve_unresolved":
                unresolved_id = int(request.form["unresolved_id"])
                asset_id = int(request.form["asset_id"])
                resolve_fusionsolar_unresolved(g.db, unresolved_id, asset_id)
                flash("Entrada API associada ao asset.", "success")
                return redirect(url_for("integrations") + "#integrations-link-audit")

            if action == "update_fusionsolar_mapping":
                integration_id = int(request.form["integration_id"])
                asset_id = int(request.form["asset_id"])
                update_fusionsolar_mapping_asset(g.db, integration_id, asset_id)
                flash("Mapeamento FusionSolar atualizado.", "success")
                return redirect(url_for("integrations") + "#integrations-link-audit")

            if action == "create_asset_from_unresolved":
                unresolved_id = int(request.form["unresolved_id"])
                asset_id = create_asset_from_unresolved(g.db, unresolved_id)
                flash("Asset criado a partir da entrada API por resolver.", "success")
                return redirect(url_for("asset_detail", asset_id=asset_id))

            if action == "ignore_unresolved":
                unresolved_id = int(request.form["unresolved_id"])
                ignore_fusionsolar_unresolved(g.db, unresolved_id)
                flash("Entrada API marcada como ignorada.", "success")
                return redirect(url_for("integrations"))

        config = get_integration_config(g.db, provider)
        integration_configs = [get_integration_config(g.db, item_provider) for item_provider in INTEGRATION_PROVIDER_OPTIONS]
        integration_configs = [item for item in integration_configs if item is not None]
        sync_runs = query_all(
            g.db,
            """
            SELECT *
            FROM integration_sync_runs
            WHERE provider IN (?, ?)
            ORDER BY started_at DESC, id DESC
            LIMIT 20
            """,
            (INTEGRATION_PROVIDER_FUSIONSOLAR, INTEGRATION_PROVIDER_SIGENERGY),
        )
        unresolved_rows = query_all(
            g.db,
            """
            SELECT *
            FROM integration_unresolved
            WHERE provider IN (?, ?) AND resolution_status = 'pending'
            ORDER BY created_at DESC, id DESC
            LIMIT 100
            """,
            (INTEGRATION_PROVIDER_FUSIONSOLAR, INTEGRATION_PROVIDER_SIGENERGY),
        )
        mapped_assets = query_all(
            g.db,
            """
            SELECT ai.*, a.project_name, a.installation_group
            FROM asset_integrations ai
            JOIN assets a ON a.id = ai.asset_id
            WHERE ai.provider IN (?, ?)
            ORDER BY a.installation_group COLLATE NOCASE, a.project_name COLLATE NOCASE
            """,
            (INTEGRATION_PROVIDER_FUSIONSOLAR, INTEGRATION_PROVIDER_SIGENERGY),
        )
        sigenergy_config = get_integration_config(g.db, INTEGRATION_PROVIDER_SIGENERGY)
        sigenergy_system_rows = query_all(
            g.db,
            """
            SELECT
                s.*,
                ai.asset_id,
                a.project_name,
                ai.external_name AS mapped_external_name,
                ai.last_error
            FROM (
                SELECT *,
                       ROW_NUMBER() OVER (PARTITION BY provider, external_id ORDER BY collected_at DESC, id DESC) AS rn
                FROM integration_realtime_snapshots
                WHERE provider = ?
            ) s
            LEFT JOIN asset_integrations ai
              ON ai.provider = s.provider AND ai.external_id = s.external_id AND ai.enabled = 1
            LEFT JOIN assets a ON a.id = ai.asset_id
            WHERE s.rn = 1
            ORDER BY COALESCE(a.project_name, ai.external_name, s.external_id) COLLATE NOCASE
            """,
            (INTEGRATION_PROVIDER_SIGENERGY,),
        )
        sigenergy_onboarding_rows = query_all(
            g.db,
            """
            SELECT *
            FROM sigenergy_onboarding_requests
            ORDER BY updated_at DESC, id DESC
            LIMIT 50
            """,
        )
        sigenergy_last_run = g.db.execute(
            """
            SELECT *
            FROM integration_sync_runs
            WHERE provider = ?
            ORDER BY started_at DESC, id DESC
            LIMIT 1
            """,
            (INTEGRATION_PROVIDER_SIGENERGY,),
        ).fetchone()
        link_audit_rows = get_fusionsolar_link_audit_rows(g.db, provider)
        link_audit_counts = {
            "ok": sum(1 for row in link_audit_rows if row["verdict"] == "OK"),
            "attention": sum(1 for row in link_audit_rows if row["verdict"] == "Atencao"),
            "review": sum(1 for row in link_audit_rows if row["verdict"] == "Rever"),
            "unresolved": sum(1 for row in link_audit_rows if row["verdict"] == "Por resolver"),
        }
        assets_for_mapping = query_all(g.db, "SELECT id, project_name FROM assets ORDER BY project_name COLLATE NOCASE")
        alert_filter_assets = query_all(
            g.db,
            """
            SELECT
                a.id,
                a.project_name,
                a.maintenance,
                a.active_contract,
                a.alerts_enabled,
                a.monitoring_enabled,
                a.monitoring_status,
                a.selected_for_alerts,
                lm.status AS latest_status,
                MAX(ta.sent_at) AS last_alert_sent,
                CASE WHEN ab.id IS NULL THEN 0 ELSE 1 END AS blacklisted
            FROM assets a
            LEFT JOIN latest_monitoring_view lm ON lm.asset_id = a.id
            LEFT JOIN telegram_alerts ta ON ta.asset_id = a.id AND ta.status = 'sent'
            LEFT JOIN alert_blacklist ab ON ab.asset_id = a.id AND ab.active = 1
            GROUP BY a.id
            ORDER BY a.project_name COLLATE NOCASE
            LIMIT 200
            """,
        )
        alert_blacklist_rows = query_all(
            g.db,
            """
            SELECT ab.*, a.project_name
            FROM alert_blacklist ab
            LEFT JOIN assets a ON a.id = ab.asset_id
            WHERE ab.active = 1
            ORDER BY ab.created_at DESC, ab.id DESC
            LIMIT 100
            """,
        )
        return render_template(
            "integrations.html",
            provider=provider,
            config=config,
            integration_configs=integration_configs,
            integration_api_controls=build_integration_api_controls(g.db),
            sync_runs=sync_runs,
            unresolved_rows=unresolved_rows,
            mapped_assets=mapped_assets,
            sigenergy_config=sigenergy_config,
            sigenergy_system_rows=sigenergy_system_rows,
            sigenergy_onboarding_rows=sigenergy_onboarding_rows,
            sigenergy_last_run=sigenergy_last_run,
            link_audit_rows=link_audit_rows,
            link_audit_counts=link_audit_counts,
            assets_for_mapping=assets_for_mapping,
            telegram_config=get_telegram_config(),
            alert_settings=get_alert_settings(g.db),
            alert_scope_options=ALERT_SCOPE_OPTIONS,
            alert_filter_assets=alert_filter_assets,
            alert_blacklist_rows=alert_blacklist_rows,
            api_call_states=list_api_call_states(g.db),
            production_api_queue_states=build_production_api_queue_observability(
                g.db
            ),
            background_jobs=fetch_latest_background_jobs(g.db, job_types=BACKGROUND_JOB_TYPES_PERFORMANCE),
            fusionsolar_api_warning=get_fusionsolar_performance_cooldown_reason(g.db),
        )

    @app.route("/telegram-alerts")
    def telegram_alerts() -> str:
        status_filter = request.args.get("status", "").strip()
        asset_filter = request.args.get("asset_id", "").strip()
        alert_type_filter = request.args.get("alert_type", "").strip()
        blocked_reason_filter = request.args.get("blocked_reason", "").strip()
        conditions = []
        params: list[Any] = []
        if status_filter:
            conditions.append("ta.status = ?")
            params.append(status_filter)
        if asset_filter:
            conditions.append("ta.asset_id = ?")
            params.append(asset_filter)
        if alert_type_filter:
            conditions.append("ta.alert_type = ?")
            params.append(alert_type_filter)
        if blocked_reason_filter:
            conditions.append("ta.blocked_reason = ?")
            params.append(blocked_reason_filter)
        where_sql = f"WHERE {' AND '.join(conditions)}" if conditions else ""
        rows = query_all(
            g.db,
            f"""
            SELECT ta.*, a.project_name
            FROM telegram_alerts ta
            LEFT JOIN assets a ON a.id = ta.asset_id
            {where_sql}
            ORDER BY ta.sent_at DESC, ta.id DESC
            LIMIT 250
            """,
            params,
        )
        alert_types = [row["alert_type"] for row in query_all(g.db, "SELECT DISTINCT alert_type FROM telegram_alerts ORDER BY alert_type")]
        blocked_reasons = [row["blocked_reason"] for row in query_all(g.db, "SELECT DISTINCT blocked_reason FROM telegram_alerts WHERE blocked_reason IS NOT NULL AND blocked_reason != '' ORDER BY blocked_reason")]
        assets_for_mapping = query_all(
            g.db,
            """
            SELECT DISTINCT a.id, a.project_name
            FROM assets a
            JOIN telegram_alerts ta ON ta.asset_id = a.id
            ORDER BY a.project_name COLLATE NOCASE
            """,
        )
        return render_template(
            "telegram_alerts.html",
            alerts=rows,
            status_filter=status_filter,
            asset_filter=asset_filter,
            alert_type_filter=alert_type_filter,
            blocked_reason_filter=blocked_reason_filter,
            alert_types=alert_types,
            blocked_reasons=blocked_reasons,
            assets_for_mapping=assets_for_mapping,
        )

    @app.route("/renewals", methods=["GET", "POST"])
    def renewals() -> str:
        focus = request.args.get("focus", "").strip()
        if request.method == "POST":
            asset_id = int(request.form["asset_id"])
            renewal_status = request.form.get("renewal_status", "Por contactar").strip() or "Por contactar"
            last_contact_date = request.form.get("last_contact_date", "").strip()
            renewal_notes = request.form.get("renewal_notes", "").strip()
            annual_value_raw = request.form.get("annual_value", "").strip()
            contract_end_date = normalize_date_value(request.form.get("contract_end_date", "").strip())
            contract_start_date = normalize_date_value(request.form.get("contract_start_date", "").strip())

            annual_value = None
            if annual_value_raw:
                normalized_value = annual_value_raw.replace(" ", "").replace(",", ".")
                try:
                    annual_value = float(normalized_value)
                except ValueError:
                    flash("O valor anual nao e valido.", "error")
                    return redirect(url_for("renewals"))

            existing_contract = query_one("SELECT id FROM om_contracts WHERE asset_id = ?", (asset_id,))
            if existing_contract:
                g.db.execute(
                    """
                    UPDATE om_contracts
                    SET renewal_status = ?, last_contact_date = ?, renewal_notes = ?,
                        annual_value = COALESCE(?, annual_value),
                        contract_start_date = CASE WHEN ? != '' THEN ? ELSE contract_start_date END,
                        contract_end_date = CASE WHEN ? != '' THEN ? ELSE contract_end_date END,
                        updated_at = ?
                    WHERE asset_id = ?
                    """,
                    (
                        renewal_status,
                        last_contact_date,
                        renewal_notes,
                        annual_value,
                        contract_start_date,
                        contract_start_date,
                        contract_end_date,
                        contract_end_date,
                        datetime.now().isoformat(timespec="seconds"),
                        asset_id,
                    ),
                )
            else:
                g.db.execute(
                    """
                    INSERT INTO om_contracts (
                        asset_id, contract_start_date, contract_end_date, annual_value, renewal_status, last_contact_date,
                        renewal_notes, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        asset_id,
                        contract_start_date,
                        contract_end_date,
                        annual_value,
                        renewal_status,
                        last_contact_date,
                        renewal_notes,
                        datetime.now().isoformat(timespec="seconds"),
                        datetime.now().isoformat(timespec="seconds"),
                    ),
                )

            if renewal_status == "Renovado":
                g.db.execute(
                    """
                    UPDATE assets
                    SET maintenance = 'yes',
                        active_contract = 'yes',
                        start_contract = CASE WHEN ? != '' THEN ? ELSE start_contract END,
                        end_contract = CASE WHEN ? != '' THEN ? ELSE end_contract END
                    WHERE id = ?
                    """,
                    (
                        contract_start_date,
                        contract_start_date,
                        contract_end_date,
                        contract_end_date,
                        asset_id,
                    ),
                )
                sync_asset_contract_status(g.db, asset_id, contract_start_date, contract_end_date)
            g.db.commit()
            flash("Follow-up de renovacao atualizado.", "success")
            return redirect(url_for("renewals"))

        today_iso = date.today().isoformat()
        year_end = f"{date.today().year}-12-31"
        renewal_rows = query_all(
            g.db,
            """
            SELECT
                a.id AS asset_id,
                a.project_name,
                a.installation_group,
                a.company_name,
                a.location,
                a.address,
                a.contact_name,
                a.contact_email,
                a.contact_phone,
                a.active_contract,
                COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) AS contract_end_date,
                COALESCE(NULLIF(oc.contract_start_date, ''), NULLIF(a.start_contract, '')) AS contract_start_date,
                oc.annual_value,
                oc.pdf_path,
                oc.renewal_status,
                oc.last_contact_date,
                oc.renewal_notes,
                CASE
                    WHEN COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) < ? THEN 'Expirado'
                    WHEN julianday(COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, ''))) - julianday(?) <= 30 THEN '0-30 dias'
                    WHEN julianday(COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, ''))) - julianday(?) <= 90 THEN '31-90 dias'
                    ELSE 'Este ano'
                END AS renewal_bucket
            FROM assets a
            LEFT JOIN om_contracts oc ON oc.asset_id = a.id
            WHERE (a.maintenance = 'yes' OR oc.id IS NOT NULL)
              AND COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) NOT IN ('', '-')
            ORDER BY COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) ASC, a.project_name COLLATE NOCASE
            """,
            (today_iso, today_iso, today_iso),
        )
        expired_contracts = [
            row for row in renewal_rows
            if row["contract_end_date"] < today_iso
        ]
        ending_this_year = [
            row for row in renewal_rows
            if today_iso <= row["contract_end_date"] <= year_end
        ]
        if focus == "expired":
            ending_this_year = []
        elif focus == "year":
            expired_contracts = []
        elif focus == "90":
            expired_contracts = []
            ending_this_year = [row for row in ending_this_year if row["renewal_bucket"] in {"0-30 dias", "31-90 dias"}]
        renewal_metrics = {
            "expired": len(expired_contracts),
            "next_30_days": sum(1 for row in renewal_rows if row["renewal_bucket"] == "0-30 dias"),
            "next_90_days": sum(1 for row in renewal_rows if row["renewal_bucket"] == "31-90 dias"),
            "this_year": len(ending_this_year),
        }
        return render_template(
            "renewals.html",
            expired_contracts=expired_contracts,
            ending_this_year=ending_this_year,
            renewal_metrics=renewal_metrics,
            focus=focus,
            today_iso=today_iso,
        )

    @app.route("/settings", methods=["GET", "POST"])
    def settings() -> str:
        if request.method == "POST":
            excel_path = request.form.get("excel_path", "").strip()
            if not excel_path:
                flash("Indica o caminho do Excel.", "error")
                return redirect(url_for("settings"))
            excel_file = Path(excel_path)
            if not excel_file.exists() or not excel_file.is_file():
                flash("O ficheiro Excel indicado nao existe ou nao esta acessivel.", "error")
                return redirect(url_for("settings"))
            if excel_file.suffix.lower() not in {".xlsx", ".xlsm"}:
                flash("Indica um ficheiro Excel valido (.xlsx ou .xlsm).", "error")
                return redirect(url_for("settings"))

            app.config["EXCEL_PATH"] = excel_path
            backup_path = create_database_backup(Path(app.config["DATABASE"]), BACKUP_DIR)
            try:
                imported = import_excel_data(g.db, excel_file)
            except Exception as exc:
                flash(f"Falha ao importar o Excel: {exc}", "error")
                flash(f"A base de dados ficou salvaguardada no backup {backup_path.name}.", "warning")
                return redirect(url_for("settings"))
            flash(
                f"Importacao concluida. {imported['assets']} assets, {imported['monitoring']} linhas de monitorizacao e {imported['tickets']} tickets importados.",
                "success",
            )
            flash(
                f"Backup automatico criado antes da reimportacao: {backup_path.name}",
                "warning",
            )
            return redirect(url_for("settings"))

        db_info = {
            "assets": query_scalar(g.db, "SELECT COUNT(*) FROM assets"),
            "monitoring": query_scalar(g.db, "SELECT COUNT(*) FROM monitoring_records"),
            "tickets": query_scalar(g.db, "SELECT COUNT(*) FROM tickets"),
            "aliases": query_scalar(g.db, "SELECT COUNT(*) FROM asset_aliases"),
        }
        excel_path = app.config["EXCEL_PATH"]
        return render_template("settings.html", db_info=db_info, excel_path=excel_path)

    return app


def ensure_database(path: str) -> None:
    with closing(get_db(path)) as conn:
        configure_database_for_runtime(conn)
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS assets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_number TEXT,
                project_name TEXT NOT NULL,
                installation_group TEXT,
                company_name TEXT,
                nif TEXT,
                address TEXT,
                location TEXT,
                panels TEXT,
                kwp TEXT,
                contract_type TEXT,
                sell_to TEXT,
                duration TEXT,
                start_contract TEXT,
                maintenance TEXT,
                coverage_type TEXT,
                access_type TEXT,
                maintenance_comment TEXT,
                status_detail TEXT,
                contact_name TEXT,
                contact_role TEXT,
                contact_email TEXT,
                contact_phone TEXT,
                end_contract TEXT,
                active_contract TEXT,
                notes TEXT,
                asset_type TEXT,
                source_payload TEXT,
                alias_blob TEXT DEFAULT '',
                monitoring_enabled INTEGER DEFAULT 1,
                alerts_enabled INTEGER DEFAULT 1,
                monitoring_status TEXT DEFAULT 'active',
                silenced_until TEXT,
                silence_reason TEXT
            );

            CREATE TABLE IF NOT EXISTS asset_aliases (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL,
                alias_name TEXT NOT NULL,
                normalized_alias TEXT NOT NULL UNIQUE,
                source TEXT,
                active INTEGER DEFAULT 1,
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS monitoring_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL,
                status TEXT NOT NULL,
                record_date TEXT NOT NULL,
                notes TEXT,
                source TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS monitoring_unmatched (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                original_name TEXT NOT NULL,
                normalized_name TEXT NOT NULL,
                status TEXT NOT NULL,
                record_date TEXT NOT NULL,
                notes TEXT
            );

            CREATE TABLE IF NOT EXISTS monitoring_import_batches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                record_date TEXT NOT NULL,
                imported_at TEXT NOT NULL,
                source TEXT NOT NULL,
                default_notes TEXT,
                raw_input TEXT,
                imported_count INTEGER DEFAULT 0,
                matched_count INTEGER DEFAULT 0,
                unmatched_count INTEGER DEFAULT 0,
                auto_resolved_count INTEGER DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS export_templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                dataset TEXT NOT NULL,
                export_format TEXT NOT NULL,
                columns_json TEXT NOT NULL,
                filters_json TEXT NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS om_contracts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL UNIQUE,
                contract_start_date TEXT,
                contract_end_date TEXT,
                annual_value REAL,
                notes TEXT,
                pdf_path TEXT,
                original_filename TEXT,
                renewal_status TEXT,
                last_contact_date TEXT,
                renewal_notes TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS integration_configs (
                provider TEXT PRIMARY KEY,
                username TEXT,
                password TEXT,
                base_url TEXT,
                login_endpoint TEXT,
                plants_endpoint TEXT,
                real_time_endpoint TEXT,
                device_list_endpoint TEXT,
                device_real_time_endpoint TEXT,
                device_history_endpoint TEXT,
                alarms_endpoint TEXT,
                energy_flow_endpoint TEXT,
                onboard_endpoint TEXT,
                day_kpi_endpoint TEXT,
                month_kpi_endpoint TEXT,
                enabled INTEGER DEFAULT 0,
                auto_sync_enabled INTEGER DEFAULT 0,
                sync_hours TEXT,
                production_sync_enabled INTEGER DEFAULT 1,
                diagnostics_sync_enabled INTEGER DEFAULT 1,
                state_sync_interval_hours INTEGER DEFAULT 1,
                production_sync_time TEXT,
                diagnostics_sync_time TEXT,
                region TEXT,
                system_ids TEXT,
                snapshot_retention_days INTEGER,
                last_sync_at TEXT,
                last_sync_status TEXT,
                last_error TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS asset_integrations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL,
                provider TEXT NOT NULL,
                external_id TEXT,
                external_name TEXT,
                enabled INTEGER DEFAULT 1,
                last_sync_at TEXT,
                last_status TEXT,
                last_error TEXT,
                UNIQUE(provider, external_id),
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS integration_sync_runs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                provider TEXT NOT NULL,
                started_at TEXT NOT NULL,
                finished_at TEXT,
                trigger_type TEXT,
                status TEXT,
                matched_count INTEGER DEFAULT 0,
                unresolved_count INTEGER DEFAULT 0,
                auto_resolved_count INTEGER DEFAULT 0,
                error_message TEXT,
                summary_json TEXT
            );

            CREATE TABLE IF NOT EXISTS integration_unresolved (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                provider TEXT NOT NULL,
                sync_run_id INTEGER,
                external_id TEXT,
                external_name TEXT NOT NULL,
                normalized_name TEXT NOT NULL,
                external_status TEXT,
                payload_json TEXT,
                suggested_asset_id INTEGER,
                resolution_status TEXT DEFAULT 'pending',
                created_at TEXT NOT NULL,
                resolved_at TEXT,
                resolution_notes TEXT,
                FOREIGN KEY (sync_run_id) REFERENCES integration_sync_runs(id) ON DELETE CASCADE,
                FOREIGN KEY (suggested_asset_id) REFERENCES assets(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS integration_realtime_snapshots (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER,
                provider TEXT NOT NULL,
                external_id TEXT NOT NULL,
                collected_at TEXT NOT NULL,
                external_status TEXT,
                normalized_status TEXT,
                pv_power_kw REAL,
                load_power_kw REAL,
                grid_power_kw_raw REAL,
                battery_power_kw REAL,
                battery_soc_pct REAL,
                ev_power_kw REAL,
                ac_power_kw REAL,
                heat_pump_power_kw REAL,
                pv_capacity_kw REAL,
                battery_capacity_kwh REAL,
                payload_json TEXT,
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS sigenergy_onboarding_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                system_id TEXT NOT NULL,
                requested_at TEXT NOT NULL,
                requested_by TEXT,
                status TEXT NOT NULL,
                provider_code TEXT,
                provider_message TEXT,
                last_checked_at TEXT,
                approved_at TEXT,
                attempt_count INTEGER NOT NULL DEFAULT 1,
                last_error TEXT,
                response_json TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS provider_devices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL,
                provider TEXT NOT NULL,
                station_code TEXT NOT NULL,
                external_device_id TEXT,
                dev_dn TEXT,
                sn TEXT,
                device_name TEXT,
                dev_type_id INTEGER,
                model TEXT,
                rated_power_kw REAL,
                enabled INTEGER DEFAULT 1,
                last_seen_at TEXT,
                payload_json TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(provider, external_device_id),
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS device_realtime_snapshots (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                provider_device_id INTEGER NOT NULL,
                asset_id INTEGER NOT NULL,
                provider TEXT NOT NULL,
                station_code TEXT NOT NULL,
                collected_at TEXT NOT NULL,
                inverter_state INTEGER,
                active_power_kw REAL,
                day_energy_kwh REAL,
                availability_status TEXT NOT NULL,
                communication_status TEXT NOT NULL,
                string_available_count INTEGER,
                string_total_count INTEGER,
                pv_current_json TEXT,
                pv_voltage_json TEXT,
                payload_json TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (provider_device_id) REFERENCES provider_devices(id) ON DELETE CASCADE,
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS availability_daily (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL,
                provider TEXT NOT NULL,
                period_date TEXT NOT NULL,
                inverter_availability_pct REAL,
                capacity_availability_pct REAL,
                communication_availability_pct REAL,
                string_availability_pct REAL,
                available_inverters INTEGER,
                total_inverters INTEGER,
                unavailable_inverters INTEGER,
                no_communication_devices INTEGER,
                available_strings INTEGER,
                total_strings INTEGER,
                unavailable_strings INTEGER,
                affected_power_kw REAL,
                unavailable_minutes INTEGER,
                payload_json TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(asset_id, provider, period_date),
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS inverter_power_samples (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL,
                provider TEXT NOT NULL,
                external_station_id TEXT NOT NULL,
                inverter_id TEXT NOT NULL,
                inverter_name TEXT,
                inverter_power_kw REAL,
                sample_time TEXT NOT NULL,
                active_power_kw REAL,
                raw_payload TEXT,
                created_at TEXT NOT NULL,
                UNIQUE(provider, inverter_id, sample_time),
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS inverter_availability_daily (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL,
                provider TEXT NOT NULL,
                availability_date TEXT NOT NULL,
                inverter_id TEXT NOT NULL,
                inverter_name TEXT,
                inverter_power_kw REAL,
                valid_slots INTEGER NOT NULL,
                available_slots INTEGER NOT NULL,
                unavailable_slots INTEGER NOT NULL,
                availability_pct REAL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(provider, inverter_id, availability_date),
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS plant_availability_daily (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL,
                provider TEXT NOT NULL,
                availability_date TEXT NOT NULL,
                valid_slots INTEGER NOT NULL,
                weighted_availability_pct REAL,
                inverter_count INTEGER NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(provider, asset_id, availability_date),
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS provider_device_expected_strings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                provider_device_id INTEGER NOT NULL,
                string_index INTEGER NOT NULL,
                expected INTEGER NOT NULL DEFAULT 1,
                source TEXT NOT NULL,
                observed_count INTEGER DEFAULT 0,
                first_observed_at TEXT,
                last_observed_at TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(provider_device_id, string_index),
                FOREIGN KEY (provider_device_id) REFERENCES provider_devices(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS production_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL,
                provider TEXT NOT NULL DEFAULT 'FusionSolar',
                external_id TEXT,
                period_type TEXT NOT NULL,
                period_date TEXT NOT NULL,
                production_kwh REAL,
                specific_yield REAL,
                expected_kwh REAL,
                expected_specific_yield REAL,
                deviation_pct REAL,
                performance_status TEXT,
                expected_source TEXT,
                data_quality TEXT,
                notes TEXT,
                selected_production_key TEXT,
                selected_production_raw_value TEXT,
                reference_diagnostic_json TEXT,
                payload_json TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(asset_id, provider, period_type, period_date),
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS portfolio_groups (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                notes TEXT
            );

            CREATE TABLE IF NOT EXISTS portfolio_assets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                portfolio_id INTEGER NOT NULL,
                asset_id INTEGER,
                external_name TEXT,
                nif TEXT,
                sub_account TEXT,
                active INTEGER DEFAULT 1,
                mapping_status TEXT,
                mapping_confidence REAL,
                notes TEXT,
                UNIQUE(portfolio_id, asset_id),
                FOREIGN KEY (portfolio_id) REFERENCES portfolio_groups(id) ON DELETE CASCADE,
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS source_files (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL,
                portfolio_id INTEGER,
                file_type TEXT NOT NULL,
                original_filename TEXT NOT NULL,
                stored_path TEXT NOT NULL,
                uploaded_at TEXT NOT NULL,
                notes TEXT,
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE,
                FOREIGN KEY (portfolio_id) REFERENCES portfolio_groups(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS invoice_documents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_file_id INTEGER NOT NULL UNIQUE,
                asset_id INTEGER NOT NULL,
                status TEXT NOT NULL,
                sha256 TEXT,
                mime_type TEXT,
                size_bytes INTEGER,
                supplier_name TEXT,
                supplier_nif TEXT,
                customer_name TEXT,
                customer_nif TEXT,
                invoice_number TEXT,
                issue_date TEXT,
                billing_period_start TEXT,
                billing_period_end TEXT,
                currency TEXT,
                total_amount TEXT,
                total_energy_kwh TEXT,
                tariff_type_candidate TEXT,
                simple_price_eur_kwh TEXT,
                ponta_price_eur_kwh TEXT,
                cheia_price_eur_kwh TEXT,
                vazio_price_eur_kwh TEXT,
                super_vazio_price_eur_kwh TEXT,
                extraction_method TEXT,
                extraction_confidence TEXT,
                warnings_json TEXT,
                reviewed_at TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (source_file_id) REFERENCES source_files(id) ON DELETE CASCADE,
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS invoice_extraction_runs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_document_id INTEGER NOT NULL,
                parser_name TEXT NOT NULL,
                parser_version TEXT NOT NULL,
                status TEXT NOT NULL,
                extracted_values_json TEXT,
                confidence_json TEXT,
                evidence_json TEXT,
                warnings_json TEXT,
                error_message TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (invoice_document_id) REFERENCES invoice_documents(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS helioscope_expected_production (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL,
                source_file_id INTEGER NOT NULL,
                base_year INTEGER,
                month INTEGER NOT NULL,
                expected_kwh REAL NOT NULL,
                imported_at TEXT NOT NULL,
                notes TEXT,
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE,
                FOREIGN KEY (source_file_id) REFERENCES source_files(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS helioscope_expected_interval_production (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL,
                source_file_id INTEGER NOT NULL,
                period_start TEXT NOT NULL,
                period_end TEXT NOT NULL,
                expected_kwh REAL NOT NULL,
                imported_at TEXT NOT NULL,
                notes TEXT,
                UNIQUE(asset_id, period_start),
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE,
                FOREIGN KEY (source_file_id) REFERENCES source_files(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS asset_tariffs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL,
                tariff_type TEXT NOT NULL,
                cycle_type TEXT,
                simple_price_eur_kwh REAL,
                ponta_price_eur_kwh REAL,
                cheia_price_eur_kwh REAL,
                vazio_price_eur_kwh REAL,
                super_vazio_price_eur_kwh REAL,
                invoice_file_id INTEGER,
                valid_from TEXT,
                valid_to TEXT,
                notes TEXT,
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE,
                FOREIGN KEY (invoice_file_id) REFERENCES source_files(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS tariff_period_rules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tariff_id INTEGER NOT NULL,
                weekday_type TEXT NOT NULL,
                start_time TEXT NOT NULL,
                end_time TEXT NOT NULL,
                period_name TEXT NOT NULL,
                FOREIGN KEY (tariff_id) REFERENCES asset_tariffs(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS production_hourly_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL,
                provider TEXT NOT NULL,
                period_start TEXT NOT NULL,
                period_end TEXT NOT NULL,
                production_kwh REAL,
                payload_json TEXT,
                imported_at TEXT NOT NULL,
                UNIQUE(asset_id, provider, period_start),
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS portfolio_report_runs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                portfolio_id INTEGER NOT NULL,
                report_month TEXT NOT NULL,
                created_at TEXT NOT NULL,
                notes TEXT,
                FOREIGN KEY (portfolio_id) REFERENCES portfolio_groups(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS portfolio_report_rows (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                report_id INTEGER NOT NULL,
                asset_id INTEGER,
                actual_production_kwh REAL,
                production_ponta_kwh REAL,
                production_cheia_kwh REAL,
                production_vazio_kwh REAL,
                production_super_vazio_kwh REAL,
                helioscope_expected_kwh REAL,
                adjusted_expected_kwh REAL,
                degradation_factor REAL,
                deviation_kwh REAL,
                deviation_pct REAL,
                availability_pct REAL,
                estimated_value_eur REAL,
                data_status TEXT,
                warnings_json TEXT,
                FOREIGN KEY (report_id) REFERENCES portfolio_report_runs(id) ON DELETE CASCADE,
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS performance_settings (
                asset_id INTEGER PRIMARY KEY,
                enabled INTEGER DEFAULT 1,
                warning_deviation_pct REAL DEFAULT -10,
                alert_deviation_pct REAL DEFAULT -20,
                critical_deviation_pct REAL DEFAULT -30,
                baseline_years INTEGER DEFAULT 2,
                min_baseline_points INTEGER DEFAULT 1,
                monthly_budget_json TEXT DEFAULT '',
                notes TEXT,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS telegram_alerts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER,
                alert_type TEXT NOT NULL,
                alert_key TEXT NOT NULL,
                message TEXT NOT NULL,
                sent_at TEXT NOT NULL,
                status TEXT NOT NULL,
                error_message TEXT NULL,
                blocked_reason TEXT NULL
            );

            CREATE TABLE IF NOT EXISTS alert_settings (
                key TEXT PRIMARY KEY,
                value TEXT
            );

            CREATE TABLE IF NOT EXISTS app_state (
                key TEXT PRIMARY KEY,
                value TEXT,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS background_jobs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                job_type TEXT NOT NULL,
                status TEXT NOT NULL,
                params_json TEXT,
                result_json TEXT,
                error_message TEXT,
                created_at TEXT NOT NULL,
                started_at TEXT,
                finished_at TEXT,
                next_attempt_at TEXT
            );

            CREATE TABLE IF NOT EXISTS alert_blacklist (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NULL,
                asset_name TEXT NULL,
                reason TEXT NULL,
                created_at TEXT NOT NULL,
                active INTEGER DEFAULT 1,
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS alert_baseline (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                baseline_at TEXT NOT NULL,
                created_by TEXT NULL,
                notes TEXT NULL
            );

            CREATE TABLE IF NOT EXISTS tickets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL,
                title TEXT NOT NULL,
                urgency TEXT NOT NULL,
                status TEXT NOT NULL,
                installation_ref TEXT,
                notes TEXT,
                next_action TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS ticket_visits (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ticket_id INTEGER NOT NULL,
                visit_date TEXT NOT NULL,
                technician TEXT,
                result TEXT,
                notes TEXT,
                next_action TEXT,
                FOREIGN KEY (ticket_id) REFERENCES tickets(id) ON DELETE CASCADE
            );

            CREATE VIEW IF NOT EXISTS latest_monitoring_view AS
            SELECT mr.asset_id, mr.status, mr.record_date, mr.notes
            FROM monitoring_records mr
            JOIN (
                SELECT asset_id, MAX(record_date || 'T' || printf('%09d', id)) AS marker
                FROM monitoring_records
                GROUP BY asset_id
            ) latest
              ON latest.asset_id = mr.asset_id
             AND latest.marker = mr.record_date || 'T' || printf('%09d', mr.id);
            """
        )
        ensure_database_indexes(conn)
        ensure_column(conn, "monitoring_records", "batch_id INTEGER")
        ensure_column(conn, "monitoring_unmatched", "batch_id INTEGER")
        ensure_column(conn, "assets", "installation_group TEXT")
        ensure_column(conn, "assets", "monitoring_enabled INTEGER DEFAULT 1")
        ensure_column(conn, "assets", "alerts_enabled INTEGER DEFAULT 1")
        ensure_column(conn, "assets", "monitoring_status TEXT DEFAULT 'active'")
        ensure_column(conn, "assets", "silenced_until TEXT")
        ensure_column(conn, "assets", "silence_reason TEXT")
        ensure_column(conn, "assets", "mounting_date TEXT")
        ensure_column(conn, "asset_aliases", "active INTEGER DEFAULT 1")
        ensure_column(conn, "tickets", "planned_date TEXT")
        ensure_column(conn, "tickets", "due_date TEXT")
        ensure_column(conn, "tickets", "estimated_minutes INTEGER DEFAULT 60")
        ensure_column(conn, "tickets", "assigned_to TEXT")
        ensure_column(conn, "tickets", "material_status TEXT DEFAULT 'Nao definido'")
        ensure_column(conn, "tickets", "planning_notes TEXT")
        ensure_column(conn, "tickets", "work_type TEXT")
        ensure_column(conn, "assets", "selected_for_alerts INTEGER DEFAULT 0")
        ensure_column(conn, "telegram_alerts", "blocked_reason TEXT")
        ensure_column(conn, "om_contracts", "renewal_status TEXT")
        ensure_column(conn, "om_contracts", "last_contact_date TEXT")
        ensure_column(conn, "om_contracts", "renewal_notes TEXT")
        ensure_column(conn, "integration_configs", "real_time_endpoint TEXT")
        ensure_column(conn, "integration_configs", "device_list_endpoint TEXT")
        ensure_column(conn, "integration_configs", "device_real_time_endpoint TEXT")
        ensure_column(conn, "integration_configs", "device_history_endpoint TEXT")
        ensure_column(conn, "integration_configs", "energy_flow_endpoint TEXT")
        ensure_column(conn, "integration_configs", "onboard_endpoint TEXT")
        ensure_column(conn, "integration_configs", "region TEXT")
        ensure_column(conn, "integration_configs", "system_ids TEXT")
        ensure_column(conn, "integration_configs", "snapshot_retention_days INTEGER")
        ensure_column(conn, "integration_configs", "day_kpi_endpoint TEXT")
        ensure_column(conn, "integration_configs", "month_kpi_endpoint TEXT")
        ensure_column(conn, "integration_configs", "production_sync_enabled INTEGER DEFAULT 1")
        ensure_column(conn, "integration_configs", "diagnostics_sync_enabled INTEGER DEFAULT 1")
        ensure_column(conn, "integration_configs", "state_sync_interval_hours INTEGER DEFAULT 1")
        ensure_column(conn, "integration_configs", "production_sync_time TEXT")
        ensure_column(conn, "integration_configs", "diagnostics_sync_time TEXT")
        ensure_column(conn, "background_jobs", "next_attempt_at TEXT")
        ensure_column(conn, "background_jobs", "wait_reason TEXT")
        ensure_api_call_state_schema(conn)
        ensure_api_queue_schema(conn)
        ensure_sampled_availability_schema(conn)
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS report_data_request_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                request_source TEXT NOT NULL,
                provider TEXT NOT NULL,
                metric TEXT NOT NULL,
                period_start TEXT NOT NULL,
                period_end TEXT NOT NULL,
                asset_ids_json TEXT NOT NULL,
                background_job_id INTEGER NOT NULL,
                reused_existing_job INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                FOREIGN KEY (background_job_id) REFERENCES background_jobs(id)
                    ON DELETE CASCADE
            )
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_background_jobs_status_next_attempt
            ON background_jobs(status, next_attempt_at)
            """
        )
        ensure_column(conn, "production_records", "selected_production_key TEXT")
        ensure_column(conn, "production_records", "selected_production_raw_value TEXT")
        ensure_column(conn, "production_records", "reference_diagnostic_json TEXT")
        ensure_column(conn, "production_hourly_records", "self_use_kwh REAL")
        ensure_column(conn, "production_hourly_records", "export_kwh REAL")
        ensure_column(conn, "production_hourly_records", "consumption_kwh REAL")
        ensure_column(conn, "production_hourly_records", "grid_import_kwh REAL")
        ensure_column(conn, "production_hourly_records", "data_quality TEXT")
        ensure_column(conn, "production_hourly_records", "source_fields_json TEXT")
        ensure_column(conn, "source_files", "sha256 TEXT")
        ensure_column(conn, "source_files", "mime_type TEXT")
        ensure_column(conn, "source_files", "size_bytes INTEGER")
        ensure_column(conn, "source_files", "archived_at TEXT")
        disable_removed_inverter_devices(conn)
        populate_missing_inverter_rated_power(conn)
        populate_missing_installation_groups(conn)
        populate_missing_group_metadata(conn)
        ensure_predefined_export_templates(conn)
        ensure_portfolio_management_schema(conn)
        ensure_portfolio_reporting_schema(conn)
        ensure_report_template_schema(conn)
        ensure_financial_model_schema(conn)
        ensure_portfolio_seed_data(conn)
        ensure_alert_settings_defaults(conn)
        ensure_billing_config_schema(conn)
        conn.execute(
            """
            UPDATE integration_configs
            SET energy_flow_endpoint = CASE
                    WHEN COALESCE(energy_flow_endpoint, '') = '' THEN COALESCE(NULLIF(alarms_endpoint, ''), ?)
                    ELSE energy_flow_endpoint
                END,
                plants_endpoint = CASE
                    WHEN COALESCE(plants_endpoint, '') IN ('', '/openapi/system/list') THEN ?
                    ELSE plants_endpoint
                END,
                onboard_endpoint = CASE
                    WHEN COALESCE(onboard_endpoint, '') = '' THEN ?
                    ELSE onboard_endpoint
                END,
                snapshot_retention_days = CASE
                    WHEN snapshot_retention_days IS NULL OR snapshot_retention_days <= 0 THEN ?
                    ELSE snapshot_retention_days
                END
            WHERE provider = ?
            """,
            (
                DEFAULT_SIGENERGY_ENERGY_FLOW_ENDPOINT,
                DEFAULT_SIGENERGY_SYSTEMS_ENDPOINT,
                DEFAULT_SIGENERGY_ONBOARD_ENDPOINT,
                DEFAULT_SIGENERGY_SNAPSHOT_RETENTION_DAYS,
                INTEGRATION_PROVIDER_SIGENERGY,
            ),
        )
        conn.commit()


def ensure_database_indexes(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE INDEX IF NOT EXISTS idx_monitoring_records_asset_date_id
            ON monitoring_records(asset_id, record_date DESC, id DESC);

        CREATE INDEX IF NOT EXISTS idx_monitoring_records_record_date_source
            ON monitoring_records(record_date, source);

        CREATE INDEX IF NOT EXISTS idx_monitoring_records_status_record_date
            ON monitoring_records(status, record_date);

        CREATE INDEX IF NOT EXISTS idx_production_records_provider_period_asset
            ON production_records(provider, period_type, period_date, asset_id);

        CREATE INDEX IF NOT EXISTS idx_production_records_performance_status
            ON production_records(performance_status);

        CREATE INDEX IF NOT EXISTS idx_asset_integrations_provider_external_id
            ON asset_integrations(provider, external_id);

        CREATE INDEX IF NOT EXISTS idx_asset_integrations_provider_enabled_asset
            ON asset_integrations(provider, enabled, asset_id);

        CREATE INDEX IF NOT EXISTS idx_tickets_asset_status
            ON tickets(asset_id, status);

        CREATE INDEX IF NOT EXISTS idx_integration_unresolved_provider_resolution_created
            ON integration_unresolved(provider, resolution_status, created_at);

        CREATE INDEX IF NOT EXISTS idx_telegram_alerts_alert_key_status
            ON telegram_alerts(alert_key, status);

        CREATE INDEX IF NOT EXISTS idx_alert_blacklist_asset_active
            ON alert_blacklist(asset_id, active);

        CREATE INDEX IF NOT EXISTS idx_background_jobs_type_status_created
            ON background_jobs(job_type, status, created_at);

        CREATE INDEX IF NOT EXISTS idx_provider_devices_provider_station
            ON provider_devices(provider, station_code);

        CREATE INDEX IF NOT EXISTS idx_provider_devices_provider_external_id
            ON provider_devices(provider, external_device_id);

        CREATE INDEX IF NOT EXISTS idx_provider_devices_asset_enabled
            ON provider_devices(asset_id, enabled);

        CREATE INDEX IF NOT EXISTS idx_device_realtime_snapshots_asset_collected
            ON device_realtime_snapshots(asset_id, collected_at DESC);

        CREATE INDEX IF NOT EXISTS idx_device_realtime_snapshots_device_collected
            ON device_realtime_snapshots(provider_device_id, collected_at DESC);

        CREATE INDEX IF NOT EXISTS idx_integration_realtime_provider_external_collected
            ON integration_realtime_snapshots(provider, external_id, collected_at);

        CREATE INDEX IF NOT EXISTS idx_integration_realtime_asset_collected
            ON integration_realtime_snapshots(asset_id, collected_at);

        CREATE INDEX IF NOT EXISTS idx_sigenergy_onboarding_system
            ON sigenergy_onboarding_requests(system_id);

        CREATE INDEX IF NOT EXISTS idx_sigenergy_onboarding_status
            ON sigenergy_onboarding_requests(status, updated_at);

        CREATE INDEX IF NOT EXISTS idx_availability_daily_asset_provider_period
            ON availability_daily(asset_id, provider, period_date DESC);

        CREATE INDEX IF NOT EXISTS idx_inverter_power_samples_asset_time
            ON inverter_power_samples(asset_id, sample_time);

        CREATE INDEX IF NOT EXISTS idx_inverter_availability_daily_asset_date
            ON inverter_availability_daily(asset_id, availability_date);

        CREATE INDEX IF NOT EXISTS idx_plant_availability_daily_date_asset
            ON plant_availability_daily(availability_date, asset_id);

        CREATE INDEX IF NOT EXISTS idx_provider_device_expected_strings_device_index
            ON provider_device_expected_strings(provider_device_id, string_index);

        CREATE INDEX IF NOT EXISTS idx_portfolio_assets_portfolio_active
            ON portfolio_assets(portfolio_id, active);

        CREATE INDEX IF NOT EXISTS idx_helioscope_expected_asset_month
            ON helioscope_expected_production(asset_id, month, imported_at);

        CREATE INDEX IF NOT EXISTS idx_helioscope_expected_interval_asset_period
            ON helioscope_expected_interval_production(asset_id, period_start);

        CREATE INDEX IF NOT EXISTS idx_asset_tariffs_asset_validity
            ON asset_tariffs(asset_id, valid_from, valid_to);

        CREATE INDEX IF NOT EXISTS idx_production_hourly_asset_period
            ON production_hourly_records(asset_id, provider, period_start);

        CREATE INDEX IF NOT EXISTS idx_portfolio_report_runs_portfolio_month
            ON portfolio_report_runs(portfolio_id, report_month, created_at);
        """
    )


def ensure_portfolio_seed_data(conn: sqlite3.Connection) -> None:
    for name in ("Solcorelios I", "Solcorelios II"):
        conn.execute(
            "INSERT OR IGNORE INTO portfolio_groups (name, notes) VALUES (?, '')",
            (name,),
        )
    seed_external_portfolio_rows(conn)


def encode_job_params(params: dict[str, Any]) -> str:
    return json.dumps(params, ensure_ascii=True, sort_keys=True)


def decode_job_params(params_json: str | None) -> dict[str, Any]:
    try:
        value = json.loads(params_json or "{}")
    except (TypeError, json.JSONDecodeError):
        return {}
    return value if isinstance(value, dict) else {}


def background_job_params_with_scope(job_type: str, params: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(params or {})
    provider, api_area = background_job_api_scope(job_type, normalized)
    if provider and "provider" not in normalized:
        normalized["provider"] = provider
    if api_area and "api_area" not in normalized:
        normalized["api_area"] = api_area
    return normalized


def create_background_job(
    conn: sqlite3.Connection,
    job_type: str,
    params: dict[str, Any],
    prevent_duplicate: bool = True,
) -> tuple[int, bool]:
    normalized_params = background_job_params_with_scope(job_type, params)
    if prevent_duplicate:
        if job_type == "fusionsolar_month_close":
            active_month_close_jobs = conn.execute(
                """
                SELECT id, params_json
                FROM background_jobs
                WHERE job_type = ?
                  AND status IN ('pending', 'running', 'waiting_rate_limit', 'waiting_api_slot')
                ORDER BY id DESC
                """,
                (job_type,),
            ).fetchall()
            requested_month = str(normalized_params.get("report_month") or "")
            requested_provider = str(
                normalized_params.get("provider") or INTEGRATION_PROVIDER_FUSIONSOLAR
            )
            for active_job in active_month_close_jobs:
                active_params = decode_job_params(active_job["params_json"])
                if (
                    str(active_params.get("report_month") or "") == requested_month
                    and str(active_params.get("provider") or INTEGRATION_PROVIDER_FUSIONSOLAR)
                    == requested_provider
                ):
                    return int(active_job["id"]), False
        normalized_params_json = encode_job_params(normalized_params)
        existing = conn.execute(
            """
            SELECT id
            FROM background_jobs
            WHERE job_type = ?
              AND params_json = ?
              AND status IN ('pending', 'running', 'waiting_rate_limit', 'waiting_api_slot')
            ORDER BY id DESC
            LIMIT 1
            """,
            (job_type, normalized_params_json),
        ).fetchone()
        if existing is not None:
            return int(existing["id"]), False

    now = serialize_background_job_timestamp()
    cursor = conn.execute(
        """
        INSERT INTO background_jobs (job_type, status, params_json, created_at)
        VALUES (?, 'pending', ?, ?)
        """,
        (job_type, encode_job_params(normalized_params), now),
    )
    return int(cursor.lastrowid), True


def mark_background_job_running(conn: sqlite3.Connection, job_id: int) -> bool:
    now_value = background_job_utc_now()
    if conn.in_transaction:
        conn.commit()
    conn.execute("BEGIN IMMEDIATE")
    try:
        job = conn.execute(
            "SELECT status, next_attempt_at FROM background_jobs WHERE id = ?",
            (job_id,),
        ).fetchone()
        if job is None:
            conn.commit()
            return False
        status = str(job["status"] or "")
        may_start = status == "pending" or (
            status in {"waiting_rate_limit", "waiting_api_slot"}
            and background_job_timestamp_is_due(
                job["next_attempt_at"],
                now=now_value,
            )
        )
        if not may_start:
            conn.commit()
            return False
        cursor = conn.execute(
            """
            UPDATE background_jobs
            SET status = 'running', started_at = ?, error_message = NULL,
                next_attempt_at = NULL, wait_reason = NULL
            WHERE id = ? AND status = ?
            """,
            (
                serialize_background_job_timestamp(now_value),
                job_id,
                status,
            ),
        )
        conn.commit()
        return cursor.rowcount == 1
    except Exception:
        conn.rollback()
        raise


def mark_background_job_success(conn: sqlite3.Connection, job_id: int, result: dict[str, Any]) -> None:
    now = serialize_background_job_timestamp()
    conn.execute(
        """
        UPDATE background_jobs
        SET status = 'success', result_json = ?, error_message = NULL,
            finished_at = ?, next_attempt_at = NULL, wait_reason = NULL
        WHERE id = ?
        """,
        (json.dumps(result, ensure_ascii=True, sort_keys=True), now, job_id),
    )
    conn.commit()


def mark_background_job_failed(conn: sqlite3.Connection, job_id: int, error_message: str) -> None:
    now = serialize_background_job_timestamp()
    conn.execute(
        """
        UPDATE background_jobs
        SET status = 'failed', error_message = ?, finished_at = ?,
            next_attempt_at = NULL, wait_reason = NULL
        WHERE id = ?
        """,
        (error_message[:2000], now, job_id),
    )
    conn.commit()


def mark_background_job_waiting_rate_limit(
    conn: sqlite3.Connection,
    job_id: int,
    *,
    next_attempt_at: datetime,
    error_message: str,
    result: dict[str, Any] | None = None,
) -> None:
    now = serialize_background_job_timestamp()
    conn.execute(
        """
        UPDATE background_jobs
        SET status = 'waiting_rate_limit',
            error_message = ?,
            result_json = ?,
            finished_at = ?,
            next_attempt_at = ?,
            wait_reason = 'cooldown_407'
        WHERE id = ?
        """,
        (
            error_message[:2000],
            json.dumps(result or {}, ensure_ascii=True, sort_keys=True),
            now,
            serialize_background_job_timestamp(next_attempt_at),
            job_id,
        ),
    )
    conn.commit()


def mark_background_job_waiting_api_slot(
    conn: sqlite3.Connection,
    job_id: int,
    *,
    next_attempt_at: datetime,
    wait_reason: str,
    error_message: str,
    result: dict[str, Any] | None = None,
) -> None:
    now = serialize_background_job_timestamp()
    conn.execute(
        """
        UPDATE background_jobs
        SET status = 'waiting_api_slot',
            error_message = ?,
            result_json = ?,
            finished_at = ?,
            next_attempt_at = ?,
            wait_reason = ?
        WHERE id = ?
        """,
        (
            error_message[:2000],
            json.dumps(result or {}, ensure_ascii=True, sort_keys=True),
            now,
            serialize_background_job_timestamp(next_attempt_at),
            wait_reason,
            job_id,
        ),
    )
    conn.commit()


def postpone_pending_production_jobs_after_407(
    conn: sqlite3.Connection,
    *,
    cooldown_until: datetime,
    exclude_job_id: int | None = None,
) -> int:
    placeholders = ", ".join("?" for _ in FUSIONSOLAR_BACKGROUND_JOB_TYPES)
    params: list[Any] = [
        serialize_background_job_timestamp(cooldown_until),
        "cooldown_407",
        "Conta FusionSolar em cooldown global apos 407.",
        *FUSIONSOLAR_BACKGROUND_JOB_TYPES,
    ]
    exclude_sql = ""
    if exclude_job_id is not None:
        exclude_sql = "AND id != ?"
        params.append(exclude_job_id)
    cursor = conn.execute(
        f"""
        UPDATE background_jobs
        SET status = 'waiting_api_slot',
            next_attempt_at = ?,
            wait_reason = ?,
            error_message = ?,
            finished_at = NULL
        WHERE job_type IN ({placeholders})
          AND status IN ('pending', 'waiting_api_slot', 'waiting_rate_limit')
          {exclude_sql}
        """,
        params,
    )
    conn.commit()
    return cursor.rowcount


def mark_stale_running_background_jobs_failed(
    conn: sqlite3.Connection,
    stale_after_minutes: int = BACKGROUND_JOB_STALE_RUNNING_MINUTES,
) -> int:
    now_value = background_job_utc_now()
    cutoff = now_value - timedelta(minutes=stale_after_minutes)
    rows = conn.execute(
        """
        SELECT id, started_at, created_at
        FROM background_jobs
        WHERE status = 'running'
        """
    ).fetchall()
    stale_ids: list[int] = []
    for row in rows:
        started_at = parse_background_job_timestamp(
            row["started_at"] or row["created_at"]
        )
        if started_at is not None and started_at < cutoff:
            stale_ids.append(int(row["id"]))
    if not stale_ids:
        return 0
    placeholders = ", ".join("?" for _ in stale_ids)
    cursor = conn.execute(
        f"""
        UPDATE background_jobs
        SET status = 'failed',
            error_message = ?,
            finished_at = ?
        WHERE status = 'running'
          AND id IN ({placeholders})
        """,
        (
            f"Job marked failed on startup after being running for more than {stale_after_minutes} minutes.",
            serialize_background_job_timestamp(now_value),
            *stale_ids,
        ),
    )
    conn.commit()
    return cursor.rowcount


def reactivate_due_rate_limited_background_jobs(conn: sqlite3.Connection) -> int:
    now_value = background_job_utc_now()
    rows = conn.execute(
        """
        SELECT id, next_attempt_at
        FROM background_jobs
        WHERE status IN ('waiting_rate_limit', 'waiting_api_slot')
        """
    ).fetchall()
    due_ids = [
        int(row["id"])
        for row in rows
        if background_job_timestamp_is_due(
            row["next_attempt_at"],
            now=now_value,
        )
    ]
    if not due_ids:
        return 0
    placeholders = ", ".join("?" for _ in due_ids)
    cursor = conn.execute(
        f"""
        UPDATE background_jobs
        SET status = 'pending',
            error_message = NULL,
            finished_at = NULL
        WHERE status IN ('waiting_rate_limit', 'waiting_api_slot')
          AND id IN ({placeholders})
        """,
        due_ids,
    )
    conn.commit()
    return cursor.rowcount


def fetch_pending_background_job_ids(conn: sqlite3.Connection) -> list[int]:
    rows = query_all(
        conn,
        """
        SELECT id
        FROM background_jobs
        WHERE status = 'pending'
        ORDER BY id ASC
        """,
    )
    return [int(row["id"]) for row in rows]


def fetch_future_waiting_background_jobs(
    conn: sqlite3.Connection,
    *,
    now: datetime | None = None,
) -> list[tuple[int, datetime]]:
    now_value = as_background_job_utc(now or background_job_utc_now())
    rows = conn.execute(
        """
        SELECT id, next_attempt_at
        FROM background_jobs
        WHERE status IN ('waiting_rate_limit', 'waiting_api_slot')
          AND next_attempt_at IS NOT NULL
        ORDER BY id ASC
        """
    ).fetchall()
    scheduled: list[tuple[int, datetime]] = []
    for row in rows:
        next_attempt_at = parse_background_job_timestamp(
            row["next_attempt_at"]
        )
        if next_attempt_at is not None and next_attempt_at > now_value:
            scheduled.append((int(row["id"]), next_attempt_at))
    return scheduled


def fetch_latest_background_jobs(
    conn: sqlite3.Connection,
    limit: int = 10,
    job_types: tuple[str, ...] | None = None,
) -> list[dict[str, Any]]:
    params: list[Any] = []
    where_sql = ""
    if job_types:
        placeholders = ", ".join("?" for _ in job_types)
        where_sql = f"WHERE job_type IN ({placeholders})"
        params.extend(job_types)
    params.append(limit)
    rows = conn.execute(
        f"""
        SELECT id, job_type, status, params_json, result_json, error_message,
               created_at, started_at, finished_at, next_attempt_at, wait_reason
        FROM background_jobs
        {where_sql}
        ORDER BY id DESC
        LIMIT ?
        """,
        params,
    ).fetchall()
    jobs: list[dict[str, Any]] = []
    for row in rows:
        job = dict(row)
        for timestamp_field in (
            "created_at",
            "started_at",
            "finished_at",
            "next_attempt_at",
        ):
            job[timestamp_field] = background_job_timestamp_to_lisbon(
                row[timestamp_field]
            )
        params_payload = decode_job_params(row["params_json"])
        provider, api_area = background_job_api_scope(str(row["job_type"]), params_payload)
        job["provider"] = params_payload.get("provider") or provider or ""
        job["api_area"] = params_payload.get("api_area") or api_area or ""
        result_summary = ""
        if row["result_json"]:
            try:
                result = json.loads(row["result_json"])
            except json.JSONDecodeError:
                result = {}
            if isinstance(result, dict):
                parts = []
                if result.get("records_updated") is not None:
                    parts.append(f"registos: {result['records_updated']}")
                if result.get("monthly_records_updated") is not None:
                    parts.append(f"mensais: {result['monthly_records_updated']}")
                if result.get("month"):
                    parts.append(f"mes: {result['month']}")
                if result.get("assets_affected") is not None:
                    parts.append(f"instalacoes afetadas: {result['assets_affected']}")
                if isinstance(result.get("states_before"), dict):
                    parts.append(
                        "antes: "
                        + ", ".join(f"{key}={value}" for key, value in result["states_before"].items())
                    )
                if isinstance(result.get("states_after"), dict):
                    parts.append(
                        "depois: "
                        + ", ".join(f"{key}={value}" for key, value in result["states_after"].items())
                    )
                if result.get("api_calls_used") is not None:
                    parts.append(f"chamadas API: {result['api_calls_used']}")
                if result.get("wait_cycles"):
                    parts.append(f"esperas: {result['wait_cycles']}")
                if result.get("resume_hint"):
                    parts.append(f"retomar: {result['resume_hint']}")
                if result.get("stopped_reason"):
                    parts.append(str(result["stopped_reason"]))
                if result.get("cooldown_until"):
                    cooldown_label = background_job_timestamp_to_lisbon(
                        result["cooldown_until"]
                    )
                    parts.append(
                        f"nova tentativa: {cooldown_label or result['cooldown_until']}"
                    )
                elif result.get("status"):
                    parts.append(f"estado: {result['status']}")
                result_summary = " | ".join(parts)
        job["result_summary"] = result_summary
        jobs.append(job)
    return jobs


def query_one(sql: str, params: tuple[Any, ...] = ()) -> sqlite3.Row | None:
    return g.db.execute(sql, params).fetchone()


def ensure_alert_settings_defaults(conn: sqlite3.Connection) -> None:
    for key, value in ALERT_SETTING_DEFAULTS.items():
        conn.execute(
            "INSERT OR IGNORE INTO alert_settings (key, value) VALUES (?, ?)",
            (key, value),
        )


def row_get(row: sqlite3.Row | dict[str, Any], key: str, default: Any = None) -> Any:
    if isinstance(row, sqlite3.Row):
        return row[key] if key in row.keys() else default
    return row.get(key, default)


def normalize_bool(value: Any, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value != 0
    normalized = str(value).strip().lower()
    if not normalized:
        return default
    return normalized in {"1", "true", "yes", "on", "sim", "y"}


def normalize_positive_int(value: Any, default: int, minimum: int = 1, maximum: int | None = None) -> int:
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        parsed = default
    parsed = max(minimum, parsed)
    if maximum is not None:
        parsed = min(maximum, parsed)
    return parsed


def normalize_clock_time(value: Any, default: str) -> str:
    raw = str(value or "").strip()
    match = re.fullmatch(r"(\d{1,2}):(\d{2})", raw)
    if not match:
        return default
    hour = int(match.group(1))
    minute = int(match.group(2))
    if hour > 23 or minute > 59:
        return default
    return f"{hour:02d}:{minute:02d}"


def split_clock_time(value: Any, default: str) -> tuple[int, int]:
    normalized = normalize_clock_time(value, default)
    hour, minute = normalized.split(":")
    return int(hour), int(minute)


def get_alert_setting(conn: sqlite3.Connection, key: str, default: str | None = None) -> str:
    ensure_alert_settings_defaults(conn)
    value = query_scalar(conn, "SELECT value FROM alert_settings WHERE key = ?", (key,))
    if value is None:
        return ALERT_SETTING_DEFAULTS.get(key, default or "")
    return str(value)


def get_alert_settings(conn: sqlite3.Connection) -> dict[str, str]:
    ensure_alert_settings_defaults(conn)
    settings = dict(ALERT_SETTING_DEFAULTS)
    for row in query_all(conn, "SELECT key, value FROM alert_settings"):
        settings[row["key"]] = row["value"] or ""
    return settings


def alert_setting_bool(conn: sqlite3.Connection, key: str, default: bool = False) -> bool:
    fallback = "true" if default else "false"
    return normalize_bool(get_alert_setting(conn, key, fallback), default)


def set_alert_setting(conn: sqlite3.Connection, key: str, value: Any) -> None:
    conn.execute(
        """
        INSERT INTO alert_settings (key, value)
        VALUES (?, ?)
        ON CONFLICT(key) DO UPDATE SET value = excluded.value
        """,
        (key, str(value)),
    )


def telegram_env_allows_alerts() -> bool:
    value = os.environ.get("TELEGRAM_ALERTS_ENABLED")
    if value is None:
        return True
    return normalize_bool(value, False)


def fetch_dashboard_stats(conn: sqlite3.Connection) -> dict[str, Any]:
    latest_status_counts = {
        row["status"]: row["total"]
        for row in query_all(
            conn,
            """
            SELECT lm.status, COUNT(*) AS total
            FROM latest_monitoring_view lm
            JOIN assets a ON a.id = lm.asset_id
            WHERE COALESCE(a.monitoring_status, 'active') != 'disabled'
            GROUP BY lm.status
            """,
        )
    }
    return {
        "assets": query_scalar(conn, "SELECT COUNT(*) FROM assets"),
        "active_om_assets": query_scalar(conn, "SELECT COUNT(*) FROM assets WHERE active_contract = 'yes'"),
        "pipeline_assets": query_scalar(conn, "SELECT COUNT(*) FROM assets WHERE COALESCE(active_contract, '') != 'yes'"),
        "monitoring_today": query_scalar(
            conn,
            "SELECT COUNT(*) FROM monitoring_records WHERE record_date = ?",
            (date.today().isoformat(),),
        ),
        "open_tickets": query_scalar(conn, "SELECT COUNT(*) FROM tickets WHERE status != 'Fechado'"),
        "open_tickets_active_om": query_scalar(
            conn,
            """
            SELECT COUNT(*)
            FROM tickets t
            JOIN assets a ON a.id = t.asset_id
            WHERE t.status != 'Fechado' AND a.active_contract = 'yes'
            """,
        ),
        "critical_tickets": query_scalar(
            conn,
            "SELECT COUNT(*) FROM tickets WHERE urgency = 'Critica' AND status != 'Fechado'",
        ),
        "critical_active_issues": query_scalar(
            conn,
            """
            SELECT COUNT(*)
            FROM latest_monitoring_view lm
            JOIN assets a ON a.id = lm.asset_id
            WHERE a.active_contract = 'yes' AND lm.status IN ('Erro', 'Desconectada')
            """,
        ),
        "expired_renewals": query_scalar(
            conn,
            """
            SELECT COUNT(*)
            FROM assets a
            LEFT JOIN om_contracts oc ON oc.asset_id = a.id
            WHERE (a.maintenance = 'yes' OR oc.id IS NOT NULL)
              AND COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) NOT IN ('', '-')
              AND COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) < ?
            """,
            (date.today().isoformat(),),
        ),
        "renewals_next_90_days": query_scalar(
            conn,
            """
            SELECT COUNT(*)
            FROM assets a
            LEFT JOIN om_contracts oc ON oc.asset_id = a.id
            WHERE (a.maintenance = 'yes' OR oc.id IS NOT NULL)
              AND COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) NOT IN ('', '-')
              AND COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) BETWEEN ? AND ?
            """,
            (date.today().isoformat(), (date.today() + timedelta(days=90)).isoformat()),
        ),
        "renewals_this_year": query_scalar(
            conn,
            """
            SELECT COUNT(*)
            FROM assets a
            LEFT JOIN om_contracts oc ON oc.asset_id = a.id
            WHERE (a.maintenance = 'yes' OR oc.id IS NOT NULL)
              AND COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) NOT IN ('', '-')
              AND COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) BETWEEN ? AND ?
            """,
            (date.today().isoformat(), f"{date.today().year}-12-31"),
        ),
        "status_counts": latest_status_counts,
    }


def build_executive_dashboard_stats(conn: sqlite3.Connection) -> dict[str, Any]:
    active_problem_rows = query_all(
        conn,
        """
        SELECT
            a.id AS asset_id,
            a.project_name,
            a.active_contract,
            lm.status,
            lm.record_date
        FROM latest_monitoring_view lm
        JOIN assets a ON a.id = lm.asset_id
        WHERE a.active_contract = 'yes'
          AND lm.status IN ('Erro', 'Desconectada')
        """,
    )
    enriched = enrich_operational_rows(conn, active_problem_rows)
    critical_or_high = [row for row in enriched if row["auto_priority"] in {"Critica", "Alta"}]
    recurring = [row for row in enriched if int(row.get("recurrence_count") or 0) >= 2]
    long_running = [row for row in enriched if int(row.get("problem_days") or 0) >= 7]
    avg_days = round(
        sum(int(row.get("problem_days") or 0) for row in enriched) / len(enriched),
        1,
    ) if enriched else 0
    return {
        "active_om_problems": len(enriched),
        "critical_or_high": len(critical_or_high),
        "recurring_90d": len(recurring),
        "long_running_7d": len(long_running),
        "avg_problem_days": avg_days,
    }


def build_integration_summary(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    sources = ["FusionSolar", "Sigenergy", "Manual / Outro"]
    summary: list[dict[str, Any]] = []
    for source in sources:
        source_aliases = {
            "FusionSolar": ["FusionSolar", "FusionSolar API", "fusion-solar-sync"],
            "Sigenergy": ["Sigenergy"],
            "Manual / Outro": ["Manual / Outro", "manual-paste", "auto-resolved"],
        }[source]
        placeholders = ",".join("?" for _ in source_aliases)
        last_batch = conn.execute(
            f"""
            SELECT imported_at, record_date, imported_count, matched_count, unmatched_count
            FROM monitoring_import_batches
            WHERE source IN ({placeholders})
            ORDER BY imported_at DESC, id DESC
            LIMIT 1
            """,
            source_aliases,
        ).fetchone()
        summary.append(
            {
                "source": source,
                "last_imported_at": last_batch["imported_at"] if last_batch else "",
                "last_record_date": last_batch["record_date"] if last_batch else "",
                "imported_count": last_batch["imported_count"] if last_batch else 0,
                "matched_count": last_batch["matched_count"] if last_batch else 0,
                "unmatched_count": last_batch["unmatched_count"] if last_batch else 0,
            }
        )
    return summary


def normalize_name(value: str) -> str:
    lowered = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii").lower()
    cleaned = "".join(char if char.isalnum() else " " for char in lowered)
    return " ".join(cleaned.split())


def infer_installation_group(project_name: str) -> str:
    name = (project_name or "").strip()
    if not name:
        return ""
    stripped = re.sub(r"\s*\([^)]*\)\s*$", "", name).strip()
    return stripped or name


def classify_fusionsolar_link(external_name: str, project_name: str, installation_group: str | None = "") -> tuple[str, str]:
    external_norm = normalize_name(external_name or "")
    project_norm = normalize_name(project_name or "")
    group_norm = normalize_name(installation_group or "")
    local_names = [name for name in (project_norm, group_norm) if name]

    if not external_norm or not local_names:
        return "Rever", "Faltam nomes para comparar."
    if external_norm in local_names:
        return "OK", "Nome FusionSolar igual a central/instalacao local."
    for local_name in local_names:
        shorter, longer = sorted((external_norm, local_name), key=len)
        if len(shorter) >= 6 and shorter in longer:
            return "Atencao", "Nome parcialmente semelhante; confirma manualmente."
    return "Rever", "Nome FusionSolar diferente da central local associada."


def update_fusionsolar_mapping_asset(conn: sqlite3.Connection, integration_id: int, asset_id: int) -> None:
    integration = conn.execute(
        "SELECT id FROM asset_integrations WHERE id = ? AND provider = ?",
        (integration_id, INTEGRATION_PROVIDER_FUSIONSOLAR),
    ).fetchone()
    if integration is None:
        raise ValueError("Mapeamento FusionSolar nao encontrado.")
    asset = conn.execute("SELECT id FROM assets WHERE id = ?", (asset_id,)).fetchone()
    if asset is None:
        raise ValueError("Central local nao encontrada.")
    conn.execute(
        """
        UPDATE asset_integrations
        SET asset_id = ?, last_error = ''
        WHERE id = ?
        """,
        (asset_id, integration_id),
    )
    conn.commit()


def get_fusionsolar_link_audit_rows(conn: sqlite3.Connection, provider: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    mapped_rows = query_all(
        conn,
        """
        SELECT
            ai.id,
            ai.asset_id,
            ai.external_id,
            ai.external_name,
            ai.last_status,
            ai.last_sync_at,
            a.project_name,
            a.installation_group
        FROM asset_integrations ai
        JOIN assets a ON a.id = ai.asset_id
        WHERE ai.provider = ? AND ai.enabled = 1
        """,
        (provider,),
    )
    duplicate_counts: dict[int, int] = {}
    duplicate_names: dict[int, list[str]] = {}
    for row in mapped_rows:
        asset_id = int(row["asset_id"])
        duplicate_counts[asset_id] = duplicate_counts.get(asset_id, 0) + 1
        duplicate_names.setdefault(asset_id, []).append(row["external_name"] or "")

    for row in mapped_rows:
        asset_id = int(row["asset_id"])
        verdict, reason = classify_fusionsolar_link(
            row["external_name"] or "",
            row["project_name"] or "",
            row["installation_group"] or "",
        )
        duplicate_count = duplicate_counts.get(asset_id, 0)
        if duplicate_count > 1:
            verdict = "Atencao" if verdict == "OK" else verdict
            reason = f"{reason} Ha {duplicate_count} entradas FusionSolar ligadas a esta central local."
        rows.append(
            {
                "integration_id": row["id"],
                "unresolved_id": None,
                "external_id": row["external_id"] or "",
                "external_name": row["external_name"] or "",
                "asset_id": asset_id,
                "project_name": row["project_name"] or "",
                "installation_group": row["installation_group"] or "",
                "last_status": row["last_status"] or "",
                "last_sync_at": row["last_sync_at"] or "",
                "verdict": verdict,
                "reason": reason,
                "duplicate_count": duplicate_count,
                "duplicate_names": ", ".join(name for name in duplicate_names.get(asset_id, []) if name),
            }
        )

    unresolved_rows = query_all(
        conn,
        """
        SELECT id, external_id, external_name, external_status, created_at
        FROM integration_unresolved
        WHERE provider = ? AND resolution_status = 'pending'
        """,
        (provider,),
    )
    for row in unresolved_rows:
        rows.append(
            {
                "integration_id": None,
                "unresolved_id": row["id"],
                "external_id": row["external_id"] or "",
                "external_name": row["external_name"] or "",
                "asset_id": None,
                "project_name": "",
                "installation_group": "",
                "last_status": row["external_status"] or "",
                "last_sync_at": row["created_at"] or "",
                "verdict": "Por resolver",
                "reason": "Ainda nao esta associada a nenhuma central local.",
                "duplicate_count": 0,
                "duplicate_names": "",
            }
        )

    unmapped_local_rows = query_all(
        conn,
        """
        SELECT a.id, a.project_name, a.installation_group
        FROM assets a
        WHERE a.active_contract = 'yes'
          AND NOT EXISTS (
              SELECT 1
              FROM asset_integrations ai
              WHERE ai.provider = ? AND ai.asset_id = a.id AND ai.enabled = 1
          )
        ORDER BY a.project_name COLLATE NOCASE
        """,
        (provider,),
    )
    for row in unmapped_local_rows:
        rows.append(
            {
                "integration_id": None,
                "unresolved_id": None,
                "external_id": "",
                "external_name": "",
                "asset_id": int(row["id"]),
                "project_name": row["project_name"] or "",
                "installation_group": row["installation_group"] or "",
                "last_status": "",
                "last_sync_at": "",
                "verdict": "Rever",
                "reason": "Central local O&M sem entrada devolvida pelo FusionSolar. Verifica a autorizacao da planta na conta northbound.",
                "duplicate_count": 0,
                "duplicate_names": "",
            }
        )

    priority = {"Rever": 0, "Atencao": 1, "Por resolver": 2, "OK": 3}
    return sorted(rows, key=lambda item: (priority.get(item["verdict"], 9), (item["external_name"] or item["project_name"]).lower()))


def parse_date_value(value: str | None) -> date | None:
    if value in (None, "", "-"):
        return None
    raw_value = str(value).strip()
    for date_format in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw_value, date_format).date()
        except ValueError:
            continue
    return None


def normalize_date_value(value: str | None) -> str:
    parsed = parse_date_value(value)
    return parsed.isoformat() if parsed else (str(value).strip() if value else "")


def derive_active_contract(end_date: str | None, current_value: str = "") -> str:
    parsed_end_date = parse_date_value(end_date)
    if parsed_end_date is None:
        return current_value
    return "yes" if parsed_end_date >= date.today() else "no"


def om_status_label(value: str | None) -> str:
    return "O&M ativo" if value == "yes" else "Sem contrato"


def format_date_pt(value: str | None) -> str:
    parsed = parse_date_value(value)
    return parsed.strftime("%d/%m/%Y") if parsed else (value or "-")


def format_number(value: Any, max_decimals: int = 2) -> str:
    parsed = parse_float_value(value)
    if parsed is None:
        return "-"
    formatted = f"{parsed:.{max_decimals}f}"
    return formatted.rstrip("0").rstrip(".")


def record_value(record: sqlite3.Row | dict[str, Any], key: str) -> Any:
    if isinstance(record, sqlite3.Row):
        return record[key] if key in record.keys() else None
    return record.get(key)


def compute_performance_percentage(record: sqlite3.Row | dict[str, Any]) -> float | None:
    specific_yield = parse_float_value(record_value(record, "specific_yield"))
    expected_specific_yield = parse_float_value(record_value(record, "expected_specific_yield"))
    if specific_yield is None or expected_specific_yield is None or expected_specific_yield <= 0:
        return None
    return (specific_yield / expected_specific_yield) * 100


def performance_bar_width(record: sqlite3.Row | dict[str, Any]) -> str:
    percentage = compute_performance_percentage(record)
    if percentage is None:
        return "0%"
    return f"{max(0, min(percentage, 100)):.1f}%"


def performance_status_class(status: str | None) -> str:
    normalized = unicodedata.normalize("NFKD", status or "Sem dados")
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "-", ascii_value.lower()).strip("-")


def reference_diagnostic(record: sqlite3.Row | dict[str, Any]) -> dict[str, Any]:
    raw = record_value(record, "reference_diagnostic_json")
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
    except (TypeError, json.JSONDecodeError):
        return {}
    return parsed if isinstance(parsed, dict) else {}


def normalize_monitoring_source(value: str | None) -> str:
    normalized = normalize_name(value or "")
    if normalized in {"fusion solar", "fusionsolar", "fusion solar api", "fusion solar sync", "fusion-solar-sync"}:
        return "FusionSolar"
    if normalized in {"sigenergy", "sig energy"}:
        return "Sigenergy"
    if normalized in {"manual", "manual outro", "manual paste", "auto resolved"}:
        return "Manual / Outro"
    return (value or "Manual / Outro").strip() or "Manual / Outro"


def days_between(start_value: str | None, end_value: str | None = None) -> int:
    start = parse_date_value(start_value)
    end = parse_date_value(end_value) or date.today()
    if start is None:
        return 0
    return max((end - start).days + 1, 0)


def auto_priority(status: str | None, problem_days: int, recurrence_count: int, open_tickets: int, active_contract: str | None) -> str:
    if active_contract != "yes":
        return "Baixa"
    score = 0
    if status == "Erro":
        score += 4
    elif status == "Desconectada":
        score += 3
    if problem_days >= 7:
        score += 3
    elif problem_days >= 3:
        score += 2
    elif problem_days >= 1:
        score += 1
    if recurrence_count >= 3:
        score += 2
    elif recurrence_count >= 2:
        score += 1
    if open_tickets:
        score += 1
    if score >= 7:
        return "Critica"
    if score >= 5:
        return "Alta"
    if score >= 2:
        return "Media"
    return "Baixa"


def contract_end_sql() -> str:
    return "COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, ''))"


def apply_group_defaults(
    conn: sqlite3.Connection,
    payload: dict[str, str],
    installation_group: str,
    exclude_asset_id: int | None = None,
) -> dict[str, str]:
    if not installation_group:
        return payload
    available_fields = [field for field in GROUP_INHERITED_FIELDS if field in payload]
    if not available_fields:
        return payload

    conditions = ["installation_group = ?"]
    params: list[Any] = [installation_group]
    if exclude_asset_id is not None:
        conditions.append("id != ?")
        params.append(exclude_asset_id)

    sources = conn.execute(
        f"""
        SELECT {", ".join(available_fields)}
        FROM assets
        WHERE {" AND ".join(conditions)}
          AND ({ " OR ".join(f"NULLIF({field}, '') IS NOT NULL" for field in available_fields) })
        ORDER BY id ASC
        """,
        params,
    ).fetchall()

    for source in sources:
        for field in available_fields:
            if not payload.get(field) and source[field]:
                payload[field] = source[field]
        if all(payload.get(field) for field in available_fields):
            break
    return payload


def list_installation_group_options(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    return query_all(
        conn,
        """
        SELECT
            COALESCE(NULLIF(TRIM(installation_group), ''), project_name) AS name,
            COUNT(*) AS member_count
        FROM assets
        WHERE COALESCE(NULLIF(TRIM(installation_group), ''), project_name) != ''
        GROUP BY name
        ORDER BY name COLLATE NOCASE
        """,
    )


def apply_group_defaults_to_asset(conn: sqlite3.Connection, asset_id: int, installation_group: str) -> None:
    asset = conn.execute("SELECT * FROM assets WHERE id = ?", (asset_id,)).fetchone()
    if asset is None:
        return
    payload = {field: asset[field] or "" for field in GROUP_INHERITED_FIELDS}
    updated_payload = apply_group_defaults(conn, payload, installation_group, exclude_asset_id=asset_id)
    changed_fields = [field for field in GROUP_INHERITED_FIELDS if (asset[field] or "") != updated_payload.get(field, "")]
    if not changed_fields:
        return
    assignments = ", ".join(f"{field} = ?" for field in changed_fields)
    values = [updated_payload[field] for field in changed_fields]
    conn.execute(f"UPDATE assets SET {assignments} WHERE id = ?", values + [asset_id])


def populate_missing_group_metadata(conn: sqlite3.Connection) -> None:
    rows = conn.execute(
        """
        SELECT id, installation_group
        FROM assets
        WHERE installation_group IS NOT NULL AND TRIM(installation_group) != ''
        ORDER BY installation_group COLLATE NOCASE, id
        """
    ).fetchall()
    for row in rows:
        apply_group_defaults_to_asset(conn, row["id"], row["installation_group"])


def sync_asset_contract_status(
    conn: sqlite3.Connection,
    asset_id: int,
    start_date: str | None = None,
    end_date: str | None = None,
) -> None:
    asset = conn.execute("SELECT active_contract, start_contract, end_contract FROM assets WHERE id = ?", (asset_id,)).fetchone()
    if asset is None:
        return
    contract = conn.execute(
        "SELECT contract_start_date, contract_end_date FROM om_contracts WHERE asset_id = ?",
        (asset_id,),
    ).fetchone()
    final_start = normalize_date_value(start_date or (contract["contract_start_date"] if contract else "") or asset["start_contract"])
    final_end = normalize_date_value(end_date or (contract["contract_end_date"] if contract else "") or asset["end_contract"])
    active_contract = derive_active_contract(final_end, asset["active_contract"] or "")
    conn.execute(
        """
        UPDATE assets
        SET maintenance = CASE WHEN ? = 'yes' THEN 'yes' ELSE maintenance END,
            active_contract = ?,
            start_contract = CASE WHEN ? != '' THEN ? ELSE start_contract END,
            end_contract = CASE WHEN ? != '' THEN ? ELSE end_contract END
        WHERE id = ?
        """,
        (active_contract, active_contract, final_start, final_start, final_end, final_end, asset_id),
    )


def sync_all_contract_statuses(conn: sqlite3.Connection) -> None:
    rows = conn.execute(
        """
        SELECT a.id, a.start_contract, a.end_contract, oc.contract_start_date, oc.contract_end_date
        FROM assets a
        LEFT JOIN om_contracts oc ON oc.asset_id = a.id
        WHERE COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) IS NOT NULL
        """
    ).fetchall()
    for row in rows:
        sync_asset_contract_status(
            conn,
            row["id"],
            row["contract_start_date"] or row["start_contract"],
            row["contract_end_date"] or row["end_contract"],
        )


def populate_missing_installation_groups(conn: sqlite3.Connection) -> None:
    rows = conn.execute(
        """
        SELECT id, project_name
        FROM assets
        WHERE installation_group IS NULL OR TRIM(installation_group) = ''
        """
    ).fetchall()
    for row in rows:
        conn.execute(
            "UPDATE assets SET installation_group = ? WHERE id = ?",
            (infer_installation_group(row["project_name"]), row["id"]),
        )


def status_rank(status: str) -> int:
    order = {
        "Erro": 1,
        "Desconectada": 2,
        "Alerta": 3,
        "Aberto": 4,
        "Em analise": 5,
        "Agendado": 6,
        "Em visita": 7,
        "Resolvido": 8,
        "Operacional": 9,
        "ok": 8,
    }
    return order.get(status or "", 99)


def group_latest_rows_by_installation(rows: list[sqlite3.Row]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for row in rows:
        installation_group = row["installation_group"] or row["project_name"]
        bucket = grouped.setdefault(
            installation_group,
            {
                "installation_group": installation_group,
                "location": row["location"],
                "active_contract": row["active_contract"],
                "record_date": row["record_date"],
                "members": [],
            },
        )
        bucket["members"].append(row)
        if not bucket["location"] and row["location"]:
            bucket["location"] = row["location"]
        if row["active_contract"] == "yes":
            bucket["active_contract"] = "yes"
        if (row["record_date"] or "") > (bucket["record_date"] or ""):
            bucket["record_date"] = row["record_date"]

    grouped_rows = []
    for bucket in grouped.values():
        members = sorted(bucket["members"], key=lambda item: (status_rank(item["status"]), item["project_name"].lower()))
        bucket["members"] = members
        bucket["group_status"] = members[0]["status"] if members else ""
        bucket["member_count"] = len(members)
        bucket["history_count"] = sum(int(member["history_count"]) for member in members)
        grouped_rows.append(bucket)

    grouped_rows.sort(
        key=lambda item: (
            0 if item["active_contract"] == "yes" else 1,
            status_rank(item["group_status"]),
            item["installation_group"].lower(),
        )
    )
    return grouped_rows


def normalize_status(value: str) -> str:
    lookup = {
        "erro": "Erro",
        "desconectada": "Desconectada",
        "operacional": "Operacional",
        "resolvido": "Resolvido",
        "aberto": "Aberto",
        "em analise": "Em analise",
        "agendado": "Agendado",
        "em visita": "Em visita",
        "fechado": "Fechado",
        "on": "Resolvido",
        "off": "Aberto",
        "faulty": "Em analise",
    }
    normalized = normalize_name(value)
    return lookup.get(normalized, value.strip())


def parse_iso_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    normalized = str(value).strip().replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(normalized)
    except ValueError:
        try:
            return datetime.fromisoformat(normalized + "T00:00:00")
        except ValueError:
            return None


def is_daytime_for_alert(now: datetime) -> bool:
    return 8 <= now.hour < 19


def html_line(value: Any) -> str:
    return html.escape(str(value or "-"))


def get_latest_monitoring_row(conn: sqlite3.Connection, asset_id: int) -> sqlite3.Row | None:
    return conn.execute(
        """
        SELECT id, status, record_date, source, created_at
        FROM monitoring_records
        WHERE asset_id = ?
        ORDER BY record_date DESC, id DESC
        LIMIT 1
        """,
        (asset_id,),
    ).fetchone()


def alert_already_sent(conn: sqlite3.Connection, alert_key: str) -> bool:
    row = conn.execute(
        "SELECT 1 FROM telegram_alerts WHERE alert_key = ? AND status = 'sent' LIMIT 1",
        (alert_key,),
    ).fetchone()
    return row is not None


def alert_recently_sent(
    conn: sqlite3.Connection,
    asset_id: int | None,
    alert_type: str,
    now: datetime,
    *,
    minutes: int = 0,
    hours: int = 0,
) -> bool:
    cooldown = timedelta(minutes=minutes, hours=hours)
    if cooldown.total_seconds() <= 0 or asset_id is None:
        return False
    since = (now - cooldown).isoformat(timespec="seconds")
    row = conn.execute(
        """
        SELECT 1
        FROM telegram_alerts
        WHERE asset_id = ? AND alert_type = ? AND status = 'sent' AND sent_at >= ?
        LIMIT 1
        """,
        (asset_id, alert_type, since),
    ).fetchone()
    return row is not None


def get_alert_type_setting(alert_type: str) -> str:
    mapping = {
        "novo_erro": "SEND_NEW_ERROR_ALERTS",
        "nova_desconexao": "SEND_OFFLINE_ALERTS",
        "desconexao_persistente_2h": "SEND_OFFLINE_ALERTS",
        "resolvido": "SEND_RESOLVED_ALERTS",
        "erro_persistente_24h": "SEND_PERSISTENT_ALERTS",
        "recorrente_7d": "SEND_RECURRENT_ALERTS",
    }
    return mapping.get(alert_type, "")


def is_asset_blacklisted(conn: sqlite3.Connection, asset: sqlite3.Row | dict[str, Any]) -> bool:
    asset_id = row_get(asset, "id", row_get(asset, "asset_id"))
    asset_name = str(row_get(asset, "project_name", "") or "").strip()
    row = conn.execute(
        """
        SELECT 1
        FROM alert_blacklist
        WHERE active = 1
          AND (
            (asset_id IS NOT NULL AND asset_id = ?)
            OR (asset_name IS NOT NULL AND lower(asset_name) = lower(?))
          )
        LIMIT 1
        """,
        (asset_id, asset_name),
    ).fetchone()
    return row is not None


def is_asset_in_oem_scope(asset: sqlite3.Row | dict[str, Any], alert_scope: str = "only_o&m") -> bool:
    if alert_scope == "all_assets":
        return True
    maintenance = normalize_bool(
        row_get(asset, "maintenance", row_get(asset, "Maintenance", row_get(asset, "contract_signed"))),
        False,
    )
    active_contract = str(row_get(asset, "active_contract", row_get(asset, "Active Contract", "")) or "").strip().lower()
    active_contract_ok = active_contract in {"yes", "true", "1", "ativo", "active", "sim"}
    if alert_scope == "only_o&m":
        return maintenance
    if alert_scope == "only_active_contracts":
        return maintenance and active_contract_ok
    if alert_scope == "only_selected_assets":
        return normalize_bool(row_get(asset, "selected_for_alerts", 0), False)
    return maintenance


def get_alert_baseline_at(conn: sqlite3.Connection) -> datetime | None:
    value = get_alert_setting(conn, "ALERT_BASELINE_AT", "")
    return parse_iso_datetime(value)


def is_before_alert_baseline(conn: sqlite3.Connection, value: str | None) -> bool:
    baseline_at = get_alert_baseline_at(conn)
    checked_at = parse_iso_datetime(value)
    return bool(baseline_at and checked_at and checked_at < baseline_at)


def alert_cooldown_active(
    conn: sqlite3.Connection,
    asset_id: int | None,
    alert_type: str,
    now: datetime,
) -> bool:
    if alert_type == "nova_desconexao":
        return alert_recently_sent(
            conn,
            asset_id,
            alert_type,
            now,
            minutes=int(get_alert_setting(conn, "OFFLINE_COOLDOWN_MINUTES", "120") or 120),
        )
    if alert_type == "resolvido":
        return alert_recently_sent(
            conn,
            asset_id,
            alert_type,
            now,
            minutes=int(get_alert_setting(conn, "RESOLVED_COOLDOWN_MINUTES", "0") or 0),
        )
    if alert_type == "novo_erro":
        return alert_recently_sent(
            conn,
            asset_id,
            alert_type,
            now,
            minutes=int(get_alert_setting(conn, "NEW_ERROR_COOLDOWN_MINUTES", "0") or 0),
        )
    if alert_type in {"erro_persistente_24h", "desconexao_persistente_2h"}:
        return alert_recently_sent(
            conn,
            asset_id,
            alert_type,
            now,
            hours=int(get_alert_setting(conn, "PERSISTENT_COOLDOWN_HOURS", "24") or 24),
        )
    if alert_type == "recorrente_7d":
        return alert_recently_sent(
            conn,
            asset_id,
            alert_type,
            now,
            hours=int(get_alert_setting(conn, "RECURRENT_COOLDOWN_HOURS", "24") or 24),
        )
    return False


def alert_decision(
    conn: sqlite3.Connection,
    asset: sqlite3.Row | dict[str, Any],
    alert_type: str,
    alert_key: str,
    now: datetime,
) -> tuple[bool, str]:
    if not alert_setting_bool(conn, "TELEGRAM_ALERTS_ENABLED", True) or not telegram_env_allows_alerts():
        return False, "global_disabled"
    if not is_telegram_configured():
        return False, "telegram_not_configured"
    if int(row_get(asset, "monitoring_enabled", 1) if row_get(asset, "monitoring_enabled", 1) is not None else 1) == 0:
        return False, "monitoring_disabled"
    if int(row_get(asset, "alerts_enabled", 1) if row_get(asset, "alerts_enabled", 1) is not None else 1) == 0:
        return False, "disabled"
    if is_asset_blacklisted(conn, asset):
        return False, "blacklist"
    monitoring_status = str(row_get(asset, "monitoring_status", "active") or "active")
    if monitoring_status in {"maintenance", "out_of_scope", "disabled"}:
        return False, monitoring_status
    if monitoring_status == "silenced":
        silenced_until = parse_iso_datetime(row_get(asset, "silenced_until"))
        if silenced_until and now < silenced_until:
            return False, "silenced"
    if not is_asset_in_oem_scope(asset, get_alert_setting(conn, "ALERT_SCOPE", "only_o&m")):
        return False, "out_of_scope"
    alert_setting = get_alert_type_setting(alert_type)
    if alert_setting and not alert_setting_bool(conn, alert_setting, True):
        return False, "alert_type_disabled"
    if alert_already_sent(conn, alert_key) or alert_cooldown_active(conn, row_get(asset, "id", row_get(asset, "asset_id")), alert_type, now):
        return False, "cooldown"
    return True, ""


def should_send_alert(conn: sqlite3.Connection, asset: sqlite3.Row | dict[str, Any], alert_type: str, alert_key: str, now: datetime) -> bool:
    return alert_decision(conn, asset, alert_type, alert_key, now)[0]


def record_telegram_alert(
    conn: sqlite3.Connection,
    asset_id: int | None,
    alert_type: str,
    alert_key: str,
    message: str,
    status: str,
    error_message: str = "",
    blocked_reason: str = "",
    sent_at: datetime | None = None,
) -> None:
    conn.execute(
        """
        INSERT INTO telegram_alerts (asset_id, alert_type, alert_key, message, sent_at, status, error_message, blocked_reason)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            asset_id,
            alert_type,
            alert_key,
            message,
            (sent_at or datetime.now()).isoformat(timespec="seconds"),
            status,
            error_message,
            blocked_reason,
        ),
    )


def send_and_record_telegram_alert(
    conn: sqlite3.Connection,
    asset_id: int | None,
    alert_type: str,
    alert_key: str,
    message: str,
) -> bool:
    if alert_already_sent(conn, alert_key):
        return False
    if not alert_setting_bool(conn, "TELEGRAM_ALERTS_ENABLED", True) or not telegram_env_allows_alerts():
        record_telegram_alert(conn, asset_id, alert_type, alert_key, message, "blocked", "Telegram desativado.", "global_disabled")
        return False
    if not is_telegram_configured():
        record_telegram_alert(conn, asset_id, alert_type, alert_key, message, "blocked", "Telegram por configurar.", "telegram_not_configured")
        return False
    try:
        ok = send_telegram_message(message)
        record_telegram_alert(
            conn,
            asset_id,
            alert_type,
            alert_key,
            message,
            "sent" if ok else "failed",
            "" if ok else "Telegram API devolveu falha ou nao respondeu.",
        )
        return ok
    except Exception as exc:
        current_app.logger.warning("Telegram alert failed without breaking import: %s", exc)
        record_telegram_alert(conn, asset_id, alert_type, alert_key, message, "failed", str(exc))
        return False


def find_problem_start(conn: sqlite3.Connection, asset_id: int, problem_status: str) -> str:
    rows = query_all(
        conn,
        """
        SELECT status, record_date, created_at
        FROM monitoring_records
        WHERE asset_id = ?
        ORDER BY record_date DESC, id DESC
        """,
        (asset_id,),
    )
    first = ""
    for row in rows:
        if row["status"] != problem_status:
            break
        first = row["record_date"] or row["created_at"]
    return first


def count_problem_occurrences_since(conn: sqlite3.Connection, asset_id: int, since_date: str) -> int:
    return int(
        query_scalar(
            conn,
            """
            SELECT COUNT(*)
            FROM monitoring_records
            WHERE asset_id = ?
              AND status IN ('Erro', 'Desconectada')
              AND record_date >= ?
            """,
            (asset_id, since_date),
        )
        or 0
    )


def build_state_change_message(event: dict[str, Any]) -> tuple[str, str]:
    asset_name = html_line(event["project_name"])
    previous_status = html_line(event["previous_status"])
    current_status = html_line(event["current_status"])
    happened_at = html_line(event["happened_at"])
    alarm_lines = ""
    if event.get("primary_alarm_name") or event.get("alarm_summary"):
        alarm_lines = "\n\n"
        if event.get("primary_alarm_name"):
            alarm_lines += f"Tipo de erro: {html_line(event.get('primary_alarm_name'))}\n"
        if event.get("primary_alarm_device"):
            alarm_lines += f"Aparelho: {html_line(event.get('primary_alarm_device'))}\n"
        if event.get("primary_alarm_severity"):
            alarm_lines += f"Severidade: {html_line(event.get('primary_alarm_severity'))}\n"
        if event.get("alarm_summary"):
            alarm_lines += f"Alarmes ativos: {html_line(event.get('alarm_summary'))}"
    if event["alert_type"] == "novo_erro":
        return (
            "novo_erro",
            f"ðŸš¨ <b>ALERTA â€” Novo erro</b>\n\nInstalacao: {asset_name}\nEstado anterior: {previous_status}\nEstado atual: {current_status}\nHora: {happened_at}{alarm_lines}",
        )
    if event["alert_type"] == "nova_desconexao":
        return (
            "nova_desconexao",
            f"âš ï¸ <b>ALERTA â€” Instalacao desconectada</b>\n\nInstalacao: {asset_name}\nHora: {happened_at}\nNota: desconexao em periodo de producao{alarm_lines}",
        )
    return (
        "resolvido",
        f"âœ… <b>RESOLVIDO</b>\n\nInstalacao: {asset_name}\nEstado anterior: {previous_status}\nEstado atual: {current_status}\nDuracao aproximada: {html_line(event.get('duration') or '-')}",
    )


def build_monitoring_alert_event(
    conn: sqlite3.Connection,
    *,
    asset_id: int,
    previous_status: str,
    current_status: str,
    happened_at: str,
    alarm_context: dict[str, Any] | None = None,
) -> dict[str, Any] | None:
    previous_status = normalize_status(previous_status or "")
    current_status = normalize_status(current_status or "")
    if previous_status == current_status:
        return None

    alert_type = ""
    if previous_status in OK_MONITORING_STATUSES and current_status == "Erro":
        alert_type = "novo_erro"
    elif previous_status in OK_MONITORING_STATUSES and current_status == "Desconectada":
        if alert_setting_bool(conn, "DAYTIME_OFFLINE_ONLY", True) and not is_daytime_for_alert(parse_iso_datetime(happened_at) or datetime.now()):
            return None
        alert_type = "nova_desconexao"
    elif previous_status in PROBLEM_MONITORING_STATUSES and current_status in OK_MONITORING_STATUSES:
        alert_type = "resolvido"
    if not alert_type:
        return None

    asset = conn.execute("SELECT project_name FROM assets WHERE id = ?", (asset_id,)).fetchone()
    if asset is None:
        return None

    duration = ""
    if alert_type == "resolvido":
        latest_problem = conn.execute(
            """
            SELECT created_at, record_date
            FROM monitoring_records
            WHERE asset_id = ? AND status = ?
            ORDER BY record_date DESC, id DESC
            LIMIT 1
            """,
            (asset_id, previous_status),
        ).fetchone()
        if latest_problem:
            started_at = parse_iso_datetime(latest_problem["created_at"] or latest_problem["record_date"])
            ended_at = parse_iso_datetime(happened_at)
            if started_at and ended_at and ended_at >= started_at:
                hours = max(1, round((ended_at - started_at).total_seconds() / 3600))
                duration = f"{hours}h"

    return {
        "asset_id": asset_id,
        "project_name": asset["project_name"],
        "previous_status": previous_status,
        "current_status": current_status,
        "happened_at": happened_at,
        "alert_type": alert_type,
        "duration": duration,
        "primary_alarm_name": (alarm_context or {}).get("primary_alarm_name", ""),
        "primary_alarm_device": (alarm_context or {}).get("primary_alarm_device", ""),
        "primary_alarm_severity": (alarm_context or {}).get("primary_alarm_severity", ""),
        "primary_alarm_raised_at": (alarm_context or {}).get("primary_alarm_raised_at", ""),
        "alarm_summary": (alarm_context or {}).get("alarm_summary", ""),
    }


def process_monitoring_alerts(
    conn: sqlite3.Connection,
    events: list[dict[str, Any]],
    batch_id: int | None,
    now: datetime | None = None,
) -> None:
    now = now or datetime.now()
    if not events:
        process_persistent_monitoring_alerts(conn, now)
        return

    ready_alerts: list[dict[str, Any]] = []
    blocked_counts: dict[str, int] = {}
    for event in events:
        asset = conn.execute("SELECT * FROM assets WHERE id = ?", (event["asset_id"],)).fetchone()
        if asset is None:
            continue
        alert_type, message = build_state_change_message(event)
        alert_key = f"{event['asset_id']}:{alert_type}:batch:{batch_id}:to:{event['current_status']}"
        allowed, reason = alert_decision(conn, asset, alert_type, alert_key, now)
        if allowed:
            ready_alerts.append({"event": event, "alert_type": alert_type, "alert_key": alert_key, "message": message})
        else:
            blocked_counts[reason] = blocked_counts.get(reason, 0) + 1
            record_telegram_alert(conn, event["asset_id"], alert_type, alert_key, message, "blocked", "", reason)

    disconnected_alerts = [item for item in ready_alerts if item["alert_type"] == "nova_desconexao"]
    if len(disconnected_alerts) > 5 and len(ready_alerts) <= 10:
        ready_alerts = [item for item in ready_alerts if item["alert_type"] != "nova_desconexao"]
        alert_key = f"geral_desconexoes_batch_{batch_id or now.isoformat(timespec='seconds')}"
        message = (
            "âš ï¸ <b>ALERTA GERAL â€” Multiplas desconexoes</b>\n\n"
            f"{len(disconnected_alerts)} instalacoes ficaram Desconectadas nesta atualizacao.\n"
            "Possivel problema de comunicacao/plataforma/importacao."
        )
        ready_alerts.append({"event": {"asset_id": None}, "alert_type": "geral_multiplas_desconexoes", "alert_key": alert_key, "message": message})

    if len(ready_alerts) > 10:
        for item in ready_alerts:
            record_telegram_alert(conn, item["event"].get("asset_id"), item["alert_type"], item["alert_key"], item["message"], "blocked", "", "batch_aggregated")
        message = (
            "âš ï¸ <b>Muitos alertas filtrados</b>\n\n"
            f"Foram detetados {len(events)} eventos de monitorizacao.\n"
            "Enviados: 1\n"
            f"Bloqueados por filtros: {sum(blocked_counts.values())}\n"
            f"Blacklisted: {blocked_counts.get('blacklist', 0)}\n"
            f"Fora de O&amp;M: {blocked_counts.get('out_of_scope', 0)}\n\n"
            "Ver detalhes na pagina Alertas Telegram."
        )
        send_and_record_telegram_alert(conn, None, "batch_many_alerts", f"batch_many_alerts:{batch_id or now.isoformat(timespec='seconds')}", message)
    else:
        for item in ready_alerts:
            send_and_record_telegram_alert(conn, item["event"].get("asset_id"), item["alert_type"], item["alert_key"], item["message"])

    process_persistent_monitoring_alerts(conn, now)


def process_persistent_monitoring_alerts(conn: sqlite3.Connection, now: datetime | None = None) -> None:
    now = now or datetime.now()
    latest_rows = query_all(
        conn,
        """
        SELECT a.*, lm.status, lm.record_date
        FROM assets a
        JOIN latest_monitoring_view lm ON lm.asset_id = a.id
        WHERE lm.status IN ('Erro', 'Desconectada')
        """,
    )
    for asset in latest_rows:
        problem_start = find_problem_start(conn, int(asset["id"]), asset["status"])
        problem_start_dt = parse_iso_datetime(problem_start)
        if not problem_start_dt:
            continue
        if alert_setting_bool(conn, "IGNORE_HISTORICAL_ALERTS", True) and is_before_alert_baseline(conn, problem_start):
            continue
        age = now - problem_start_dt
        if asset["status"] == "Erro" and age >= timedelta(hours=24):
            alert_key = f"{asset['id']}:erro_persistente_24h:{problem_start}"
            message = (
                "ðŸš¨ <b>ERRO PERSISTENTE</b>\n\n"
                f"Instalacao: {html_line(asset['project_name'])}\n"
                "Estado: Erro\n"
                "Duracao: &gt;24h\n"
                f"Primeira detecao: {html_line(problem_start)}"
            )
            allowed, reason = alert_decision(conn, asset, "erro_persistente_24h", alert_key, now)
            if allowed:
                send_and_record_telegram_alert(conn, int(asset["id"]), "erro_persistente_24h", alert_key, message)
            elif not alert_already_sent(conn, alert_key):
                record_telegram_alert(conn, int(asset["id"]), "erro_persistente_24h", alert_key, message, "blocked", "", reason)
        offline_daytime_only = alert_setting_bool(conn, "DAYTIME_OFFLINE_ONLY", True)
        if asset["status"] == "Desconectada" and age >= timedelta(hours=2) and (not offline_daytime_only or is_daytime_for_alert(now)):
            alert_key = f"{asset['id']}:desconexao_persistente_2h:{problem_start}"
            message = (
                "âš ï¸ <b>DESCONEXAO PERSISTENTE</b>\n\n"
                f"Instalacao: {html_line(asset['project_name'])}\n"
                "Estado: Desconectada\n"
                "Duracao: &gt;2h em periodo de producao"
            )
            allowed, reason = alert_decision(conn, asset, "desconexao_persistente_2h", alert_key, now)
            if allowed:
                send_and_record_telegram_alert(conn, int(asset["id"]), "desconexao_persistente_2h", alert_key, message)
            elif not alert_already_sent(conn, alert_key):
                record_telegram_alert(conn, int(asset["id"]), "desconexao_persistente_2h", alert_key, message, "blocked", "", reason)

        since_date = (now.date() - timedelta(days=7)).isoformat()
        baseline_at = get_alert_baseline_at(conn)
        if baseline_at:
            since_date = max(since_date, baseline_at.date().isoformat())
        occurrences = count_problem_occurrences_since(conn, int(asset["id"]), since_date)
        if occurrences >= 3:
            alert_key = f"{asset['id']}:recorrente_7d:{now.date().isoformat()}"
            message = (
                "ðŸ” <b>ERRO RECORRENTE</b>\n\n"
                f"Instalacao: {html_line(asset['project_name'])}\n"
                f"Ocorrencias nos ultimos 7 dias: {occurrences}\n"
                f"Ultimo estado: {html_line(asset['status'])}"
            )
            allowed, reason = alert_decision(conn, asset, "recorrente_7d", alert_key, now)
            if allowed:
                send_and_record_telegram_alert(conn, int(asset["id"]), "recorrente_7d", alert_key, message)
            elif not alert_already_sent(conn, alert_key):
                record_telegram_alert(conn, int(asset["id"]), "recorrente_7d", alert_key, message, "blocked", "", reason)


def format_summary_list(rows: list[sqlite3.Row], empty: str = "-") -> str:
    if not rows:
        return empty
    lines = []
    for row in rows[:8]:
        name = html_line(row["project_name"])
        status = html_line(row["status"] if "status" in row.keys() else "")
        lines.append(f"- {name}: {status}" if status != "-" else f"- {name}")
    if len(rows) > 8:
        lines.append(f"- ... mais {len(rows) - 8}")
    return "\n".join(lines)


def send_daily_telegram_summary(conn: sqlite3.Connection, now: datetime | None = None) -> bool:
    now = now or datetime.now()
    if not telegram_daily_summary_enabled() or not is_telegram_configured():
        return False
    yesterday = (now.date() - timedelta(days=1)).isoformat()
    alert_key = f"daily_summary:{now.date().isoformat()}"
    if alert_already_sent(conn, alert_key):
        return False

    current_error = query_scalar(conn, "SELECT COUNT(*) FROM latest_monitoring_view WHERE status = 'Erro'") or 0
    current_disconnected = query_scalar(conn, "SELECT COUNT(*) FROM latest_monitoring_view WHERE status = 'Desconectada'") or 0
    new_rows = query_all(
        conn,
        """
        SELECT a.project_name, mr.status
        FROM monitoring_records mr
        JOIN assets a ON a.id = mr.asset_id
        WHERE mr.record_date >= ? AND mr.status IN ('Erro', 'Desconectada')
        ORDER BY mr.record_date DESC, mr.id DESC
        LIMIT 20
        """,
        (yesterday,),
    )
    resolved_rows = query_all(
        conn,
        """
        SELECT a.project_name, mr.status
        FROM monitoring_records mr
        JOIN assets a ON a.id = mr.asset_id
        WHERE mr.record_date >= ? AND mr.status IN ('Resolvido', 'Operacional')
        ORDER BY mr.record_date DESC, mr.id DESC
        LIMIT 20
        """,
        (yesterday,),
    )
    persistent_rows = query_all(
        conn,
        """
        SELECT a.project_name, lm.status
        FROM latest_monitoring_view lm
        JOIN assets a ON a.id = lm.asset_id
        WHERE lm.status IN ('Erro', 'Desconectada')
        ORDER BY a.project_name COLLATE NOCASE
        LIMIT 20
        """,
    )
    muted_rows = query_all(
        conn,
        """
        SELECT project_name, monitoring_status AS status
        FROM assets
        WHERE monitoring_status IN ('silenced', 'maintenance')
        ORDER BY project_name COLLATE NOCASE
        LIMIT 20
        """,
    )
    recurring_rows = []
    for row in persistent_rows:
        asset = conn.execute("SELECT id FROM assets WHERE project_name = ?", (row["project_name"],)).fetchone()
        if asset and count_problem_occurrences_since(conn, int(asset["id"]), (now.date() - timedelta(days=7)).isoformat()) >= 3:
            recurring_rows.append(row)

    message = (
        f"<b>Resumo O&amp;M - {html_line(now.date().isoformat())}</b>\n\n"
        "Ativos:\n"
        f"- Erro: {current_error}\n"
        f"- Desconectadas: {current_disconnected}\n\n"
        "Novos desde ontem:\n"
        f"{format_summary_list(new_rows)}\n\n"
        "Persistentes:\n"
        f"{format_summary_list(persistent_rows)}\n\n"
        "Resolvidos desde ontem:\n"
        f"{format_summary_list(resolved_rows)}\n\n"
        "Recorrentes:\n"
        f"{format_summary_list(recurring_rows)}\n\n"
        "Instalacoes silenciadas/manutencao:\n"
        f"{format_summary_list(muted_rows)}"
    )
    return send_and_record_telegram_alert(conn, None, "daily_summary", alert_key, message)


def excel_date_to_iso(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value in (None, "", "-"):
        return ""
    return str(value)


def row_value(row: tuple[Any, ...], index: int) -> str:
    value = row[index] if index < len(row) else None
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def get_sheet(workbook, expected_name: str):
    normalized_expected = normalize_name(expected_name)
    for sheet_name in workbook.sheetnames:
        if normalize_name(sheet_name) == normalized_expected:
            return workbook[sheet_name]
    return None


def find_asset_id(conn: sqlite3.Connection, name: str) -> int | None:
    normalized = normalize_name(name)
    row = conn.execute(
        """
        SELECT asset_id
        FROM asset_aliases
        WHERE normalized_alias = ?
        """,
        (normalized,),
    ).fetchone()
    return int(row["asset_id"]) if row else None


def rebuild_asset_alias_blob(conn: sqlite3.Connection, asset_id: int) -> None:
    aliases = [row["alias_name"] for row in query_all(conn, "SELECT alias_name FROM asset_aliases WHERE asset_id = ?", (asset_id,))]
    conn.execute("UPDATE assets SET alias_blob = ? WHERE id = ?", (" | ".join(aliases), asset_id))
    conn.commit()


def import_excel_data(conn: sqlite3.Connection, excel_path: Path) -> dict[str, int]:
    workbook = load_workbook(excel_path, data_only=True)
    excel_batch_id = create_monitoring_batch(
        conn,
        record_date=date.today().isoformat(),
        default_notes="Sincronizacao a partir do Excel.",
        raw_input="",
        source="excel-import",
    )

    assets_by_name: dict[str, int] = {}
    project_sheet = get_sheet(workbook, "Project Overview")
    if project_sheet is None:
        raise ValueError("Folha 'Project Overview' nao encontrada no Excel.")

    for row in project_sheet.iter_rows(min_row=2, values_only=True):
        project_name = row_value(row, 1)
        if not project_name:
            continue
        payload = {
            "project_number": row_value(row, 0),
            "project_name": project_name,
            "company_name": row_value(row, 2),
            "nif": row_value(row, 3),
            "address": row_value(row, 4),
            "location": row_value(row, 5),
            "panels": row_value(row, 6),
            "kwp": row_value(row, 7),
            "contract_type": row_value(row, 8),
            "sell_to": row_value(row, 9),
            "duration": row_value(row, 10),
            "start_contract": row_value(row, 12),
            "maintenance": row_value(row, 13),
            "coverage_type": row_value(row, 14),
            "access_type": row_value(row, 15),
            "maintenance_comment": row_value(row, 16),
            "status_detail": row_value(row, 17),
            "contact_name": row_value(row, 18),
            "contact_role": row_value(row, 19),
            "contact_email": row_value(row, 20),
            "contact_phone": row_value(row, 21),
            "end_contract": row_value(row, 22),
            "active_contract": row_value(row, 24),
            "notes": row_value(row, 25),
            "asset_type": row_value(row, 27),
        }
        asset_id = upsert_asset_from_excel(conn, payload)
        assets_by_name[project_name] = asset_id
        alias_names = {project_name, payload["company_name"]}
        for alias_name in alias_names:
            if alias_name:
                normalized = normalize_name(alias_name)
                if normalized:
                    conn.execute(
                        "INSERT OR IGNORE INTO asset_aliases (asset_id, alias_name, normalized_alias, source) VALUES (?, ?, ?, ?)",
                        (asset_id, alias_name, normalized, "excel"),
                    )

    monitoring_imported = 0
    monitoring_sheet = get_sheet(workbook, "Monotorizacao")
    if monitoring_sheet is not None:
        for row in monitoring_sheet.iter_rows(min_row=3, values_only=True):
            display_name = row_value(row, 0)
            status = normalize_status(row_value(row, 1))
            record_date = excel_date_to_iso(row[4] if len(row) > 4 else None)
            notes = row_value(row, 5)
            original_name = row_value(row, 6) or display_name
            if not status or not original_name:
                continue
            asset_id = find_asset_id(conn, original_name) or find_asset_id(conn, display_name)
            if asset_id:
                record_date_value = record_date or date.today().isoformat()
                existing = conn.execute(
                    """
                    SELECT 1
                    FROM monitoring_records
                    WHERE asset_id = ? AND status = ? AND record_date = ? AND source = 'excel'
                    LIMIT 1
                    """,
                    (asset_id, status, record_date_value),
                ).fetchone()
                if not existing:
                    conn.execute(
                        """
                        INSERT INTO monitoring_records (asset_id, status, record_date, notes, source, batch_id)
                        VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (asset_id, status, record_date_value, notes, "excel", excel_batch_id),
                    )
                    monitoring_imported += 1
                for alias_candidate in {display_name, original_name}:
                    normalized = normalize_name(alias_candidate)
                    if normalized:
                        conn.execute(
                            "INSERT OR IGNORE INTO asset_aliases (asset_id, alias_name, normalized_alias, source) VALUES (?, ?, ?, ?)",
                            (asset_id, alias_candidate, normalized, "excel-monitoring"),
                        )

    tickets_imported = 0
    corrective_sheet = get_sheet(workbook, "Corretivas")
    if corrective_sheet is not None:
        carried_asset_name = ""
        carried_notes = ""
        for row in corrective_sheet.iter_rows(min_row=2, values_only=True):
            asset_name = row_value(row, 0) or carried_asset_name
            installation_ref = row_value(row, 1)
            contract_type = row_value(row, 2)
            created_at = excel_date_to_iso(row[3])
            status = normalize_status(row_value(row, 4))
            notes = row_value(row, 5)
            next_action = row_value(row, 6)

            if row_value(row, 0):
                carried_asset_name = row_value(row, 0)
            if notes:
                carried_notes = notes
            elif not row_value(row, 0) and carried_notes:
                notes = carried_notes

            if not asset_name or not (status or notes or next_action):
                continue

            asset_id = find_asset_id(conn, asset_name)
            if not asset_id:
                continue

            urgency = "Alta" if status in {"Aberto", "Em analise"} else "Media"
            title = next_action.splitlines()[0][:120] if next_action else f"Corretiva - {asset_name}"
            created_at_value = created_at or date.today().isoformat()
            existing = conn.execute(
                """
                SELECT 1
                FROM tickets
                WHERE asset_id = ? AND title = ? AND created_at = ?
                LIMIT 1
                """,
                (asset_id, title, created_at_value),
            ).fetchone()
            if not existing:
                conn.execute(
                    """
                    INSERT INTO tickets (
                        asset_id, title, urgency, status, installation_ref, notes, next_action, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        asset_id,
                        title,
                        urgency,
                        status or "Aberto",
                        installation_ref or contract_type,
                        notes,
                        next_action,
                        created_at_value,
                        created_at_value,
                    ),
                )
                tickets_imported += 1

    for asset_id_row in query_all(conn, "SELECT id FROM assets"):
        rebuild_asset_alias_blob(conn, int(asset_id_row["id"]))

    conn.execute(
        """
        UPDATE monitoring_import_batches
        SET imported_count = ?, matched_count = ?, unmatched_count = 0, auto_resolved_count = 0
        WHERE id = ?
        """,
        (monitoring_imported, monitoring_imported, excel_batch_id),
    )
    conn.commit()
    return {"assets": len(assets_by_name), "monitoring": monitoring_imported, "tickets": tickets_imported}


def upsert_asset_from_excel(conn: sqlite3.Connection, payload: dict[str, str]) -> int:
    project_name = payload["project_name"]
    existing = conn.execute(
        """
        SELECT id, installation_group
        FROM assets
        WHERE project_name = ?
        LIMIT 1
        """,
        (project_name,),
    ).fetchone()
    asset_id = int(existing["id"]) if existing else (find_asset_id(conn, project_name) or 0)
    installation_group = (
        (existing["installation_group"] if existing and existing["installation_group"] else "")
        or infer_installation_group(payload["project_name"])
    )
    payload["start_contract"] = normalize_date_value(payload["start_contract"])
    payload["end_contract"] = normalize_date_value(payload["end_contract"])
    payload["active_contract"] = derive_active_contract(payload["end_contract"], payload["active_contract"])
    payload = apply_group_defaults(conn, payload, installation_group, exclude_asset_id=asset_id or None)

    values = (
        payload["project_number"],
        payload["project_name"],
        installation_group,
        payload["company_name"],
        payload["nif"],
        payload["address"],
        payload["location"],
        payload["panels"],
        payload["kwp"],
        payload["contract_type"],
        payload["sell_to"],
        payload["duration"],
        payload["start_contract"],
        payload["maintenance"],
        payload["coverage_type"],
        payload["access_type"],
        payload["maintenance_comment"],
        payload["status_detail"],
        payload["contact_name"],
        payload["contact_role"],
        payload["contact_email"],
        payload["contact_phone"],
        payload["end_contract"],
        payload["active_contract"],
        payload["notes"],
        payload["asset_type"],
        json.dumps(payload, ensure_ascii=True),
    )

    if asset_id:
        conn.execute(
            """
            UPDATE assets
            SET
                project_number = ?, project_name = ?, installation_group = ?, company_name = ?, nif = ?, address = ?, location = ?,
                panels = ?, kwp = ?, contract_type = ?, sell_to = ?, duration = ?, start_contract = ?,
                maintenance = ?, coverage_type = ?, access_type = ?, maintenance_comment = ?, status_detail = ?,
                contact_name = ?, contact_role = ?, contact_email = ?, contact_phone = ?, end_contract = ?,
                active_contract = ?, notes = ?, asset_type = ?, source_payload = ?
            WHERE id = ?
            """,
            values + (asset_id,),
        )
        return asset_id

    cursor = conn.execute(
        """
        INSERT INTO assets (
            project_number, project_name, installation_group, company_name, nif, address, location, panels, kwp,
            contract_type, sell_to, duration, start_contract, maintenance, coverage_type,
            access_type, maintenance_comment, status_detail, contact_name, contact_role,
            contact_email, contact_phone, end_contract, active_contract, notes, asset_type,
            source_payload, alias_blob
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        values + (project_name,),
    )
    return int(cursor.lastrowid)


@dataclass
class MonitoringImportResult:
    imported: int = 0
    matched: int = 0
    unmatched: int = 0
    auto_resolved: int = 0
    batch_id: int | None = None


def parse_monitoring_lines(pasted_table: str) -> list[tuple[str, str]]:
    lines = []
    for raw_line in pasted_table.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        lowered = normalize_name(line)
        if lowered in {"instalacao estado", "instalacao", "estado"}:
            continue
        if "\t" in line:
            parts = [part.strip() for part in line.split("\t") if part.strip()]
            if len(parts) >= 2:
                first_part = normalize_name(parts[0])
                second_part = normalize_name(parts[1])
                if first_part in {"instalacao", "instalacao estado"} or second_part == "estado":
                    continue
                lines.append((parts[0], normalize_status(parts[1])))
                continue
        for marker in (" Erro", " Desconectada", " Operacional", " Resolvido"):
            if line.endswith(marker):
                lines.append((line[: -len(marker)].strip(), normalize_status(marker.strip())))
                break
        else:
            parts = line.rsplit(" ", 1)
            if len(parts) == 2:
                lines.append((parts[0].strip(), normalize_status(parts[1].strip())))
    return lines


def import_daily_monitoring(
    conn: sqlite3.Connection,
    pasted_table: str,
    record_date: str,
    default_notes: str,
    platform_source: str,
    import_scope: str = "complete",
) -> MonitoringImportResult:
    result = MonitoringImportResult()
    platform_source = normalize_monitoring_source(platform_source)
    parsed_lines = parse_monitoring_lines(pasted_table)
    if not parsed_lines:
        return result
    batch_id = create_monitoring_batch(conn, record_date, default_notes, pasted_table, platform_source)
    result.batch_id = batch_id
    imported_asset_ids: set[int] = set()
    alert_events: list[dict[str, Any]] = []
    now = datetime.now()
    for original_name, status in parsed_lines:
        asset_id = find_asset_id(conn, original_name)
        if asset_id:
            duplicate = conn.execute(
                """
                SELECT 1
                FROM monitoring_records
                WHERE asset_id = ? AND status = ? AND record_date = ? AND source = ?
                LIMIT 1
                """,
                (asset_id, status, record_date, platform_source),
            ).fetchone()
            if duplicate:
                continue
            previous = get_latest_monitoring_row(conn, asset_id)
            result.imported += 1
            imported_asset_ids.add(asset_id)
            conn.execute(
                """
                INSERT INTO monitoring_records (asset_id, status, record_date, notes, source, batch_id)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (asset_id, status, record_date, default_notes, platform_source, batch_id),
            )
            result.matched += 1
            event = build_monitoring_alert_event(
                conn,
                asset_id=asset_id,
                previous_status=previous["status"] if previous else "",
                current_status=status,
                happened_at=now.isoformat(timespec="seconds"),
            )
            if event:
                alert_events.append(event)
        else:
            normalized_name = normalize_name(original_name)
            duplicate_unmatched = conn.execute(
                """
                SELECT 1
                FROM monitoring_unmatched
                WHERE normalized_name = ? AND status = ? AND record_date = ?
                LIMIT 1
                """,
                (normalized_name, status, record_date),
            ).fetchone()
            if duplicate_unmatched:
                continue
            result.imported += 1
            conn.execute(
                """
                INSERT INTO monitoring_unmatched (original_name, normalized_name, status, record_date, notes, batch_id)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (original_name, normalized_name, status, record_date, default_notes, batch_id),
            )
            result.unmatched += 1

    if import_scope == "complete":
        latest_problem_assets = query_all(
            conn,
            """
            SELECT lm.asset_id
            FROM latest_monitoring_view lm
            WHERE lm.status IN ('Erro', 'Desconectada')
            """,
        )
        for row in latest_problem_assets:
            asset_id = int(row["asset_id"])
            if asset_id in imported_asset_ids:
                continue
            existing_today = conn.execute(
                """
                SELECT 1
                FROM monitoring_records
                WHERE asset_id = ? AND record_date = ?
                LIMIT 1
                """,
                (asset_id, record_date),
            ).fetchone()
            if existing_today:
                continue
            previous = get_latest_monitoring_row(conn, asset_id)
            conn.execute(
                """
                INSERT INTO monitoring_records (asset_id, status, record_date, notes, source, batch_id)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    asset_id,
                    "Resolvido",
                    record_date,
                    "Resolvido automaticamente por nao constar na lista diaria.",
                    platform_source,
                    batch_id,
                ),
            )
            result.auto_resolved += 1
            event = build_monitoring_alert_event(
                conn,
                asset_id=asset_id,
                previous_status=previous["status"] if previous else "",
                current_status="Resolvido",
                happened_at=now.isoformat(timespec="seconds"),
            )
            if event:
                alert_events.append(event)
    conn.execute(
        """
        UPDATE monitoring_import_batches
        SET imported_count = ?, matched_count = ?, unmatched_count = ?, auto_resolved_count = ?
        WHERE id = ?
        """,
        (result.imported, result.matched, result.unmatched, result.auto_resolved, batch_id),
    )
    process_monitoring_alerts(conn, alert_events, batch_id, now)
    conn.commit()
    return result


def create_monitoring_batch(
    conn: sqlite3.Connection,
    record_date: str,
    default_notes: str,
    raw_input: str,
    source: str,
) -> int:
    source = normalize_monitoring_source(source)
    cursor = conn.execute(
        """
        INSERT INTO monitoring_import_batches (record_date, imported_at, source, default_notes, raw_input)
        VALUES (?, ?, ?, ?, ?)
        """,
        (record_date, datetime.now().isoformat(timespec="seconds"), source, default_notes, raw_input),
    )
    return int(cursor.lastrowid)


def build_problem_periods(conn: sqlite3.Connection, asset_id: int) -> list[dict[str, Any]]:
    rows = query_all(
        conn,
        """
        SELECT mr.record_date, mr.status, mr.notes, mr.source, mib.imported_at
        FROM monitoring_records mr
        LEFT JOIN monitoring_import_batches mib ON mib.id = mr.batch_id
        WHERE mr.asset_id = ?
        ORDER BY mr.record_date ASC, mr.id ASC
        """,
        (asset_id,),
    )
    problem_statuses = {"Erro", "Desconectada"}
    periods: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None

    for row in rows:
        status = row["status"]
        if status in problem_statuses:
            if current is None:
                current = {
                    "status": status,
                    "started_on": row["record_date"],
                    "started_at": row["imported_at"],
                    "start_notes": row["notes"],
                    "last_problem_on": row["record_date"],
                    "last_problem_status": status,
                }
            else:
                current["last_problem_on"] = row["record_date"]
                current["last_problem_status"] = status
            continue

        if current is not None:
            current["resolved_on"] = row["record_date"]
            current["resolved_at"] = row["imported_at"]
            current["resolution_status"] = status
            current["resolution_notes"] = row["notes"]
            periods.append(current)
            current = None

    if current is not None:
        current["resolved_on"] = None
        current["resolved_at"] = None
        current["resolution_status"] = "Ainda ativo"
        current["resolution_notes"] = ""
        periods.append(current)

    periods.reverse()
    return periods


def build_problem_metric_map(conn: sqlite3.Connection, asset_ids: list[int]) -> dict[int, dict[str, Any]]:
    metrics: dict[int, dict[str, Any]] = {}
    cutoff = (date.today() - timedelta(days=90)).isoformat()
    for asset_id in asset_ids:
        periods = build_problem_periods(conn, asset_id)
        active_period = next((period for period in periods if period["resolved_on"] is None), None)
        recurrence_count = sum(
            1
            for period in periods
            if (period.get("started_on") or "") >= cutoff
        )
        metrics[asset_id] = {
            "problem_started_on": active_period["started_on"] if active_period else "",
            "problem_days": days_between(active_period["started_on"]) if active_period else 0,
            "recurrence_count": recurrence_count,
            "last_problem_status": active_period["last_problem_status"] if active_period else "",
        }
    return metrics


def enrich_operational_rows(conn: sqlite3.Connection, rows: list[sqlite3.Row | dict[str, Any]]) -> list[dict[str, Any]]:
    enriched = [dict(row) for row in rows]
    asset_ids = [int(row["asset_id"] if "asset_id" in row else row["id"]) for row in enriched]
    metric_map = build_problem_metric_map(conn, asset_ids)
    ticket_counts: dict[int, int] = {}
    if asset_ids:
        ticket_counts = {
            int(row["asset_id"]): int(row["open_tickets"])
            for row in query_all(
                conn,
                f"""
                SELECT asset_id, COUNT(*) AS open_tickets
                FROM tickets
                WHERE status != 'Fechado' AND asset_id IN ({",".join("?" for _ in asset_ids)})
                GROUP BY asset_id
                """,
                asset_ids,
            )
        }
    for row in enriched:
        asset_id = int(row["asset_id"] if "asset_id" in row else row["id"])
        row.setdefault("id", asset_id)
        metrics = metric_map.get(asset_id, {})
        row.update(metrics)
        row["open_tickets"] = int(row.get("open_tickets") or ticket_counts.get(asset_id, 0) or 0)
        row["auto_priority"] = auto_priority(
            row.get("status"),
            int(row.get("problem_days") or 0),
            int(row.get("recurrence_count") or 0),
            int(row.get("open_tickets") or 0),
            row.get("active_contract"),
        )
    return enriched


def priority_rank(priority: str) -> int:
    return {"Critica": 1, "Alta": 2, "Media": 3, "Baixa": 4}.get(priority, 5)


def build_batch_insight(conn: sqlite3.Connection, batch_id: int) -> dict[str, Any] | None:
    batch = conn.execute(
        """
        SELECT *
        FROM monitoring_import_batches
        WHERE id = ?
        """,
        (batch_id,),
    ).fetchone()
    if batch is None:
        return None

    rows = query_all(
        conn,
        """
        SELECT
            mr.id,
            mr.asset_id,
            mr.status,
            mr.record_date,
            a.project_name,
            a.active_contract
        FROM monitoring_records mr
        JOIN assets a ON a.id = mr.asset_id
        WHERE mr.batch_id = ?
        ORDER BY mr.id ASC
        """,
        (batch_id,),
    )

    problem_statuses = {"Erro", "Desconectada"}
    new_problem_assets: list[dict[str, Any]] = []
    persistent_problem_assets: list[dict[str, Any]] = []
    resolved_assets: list[dict[str, Any]] = []

    for row in rows:
        previous = conn.execute(
            """
            SELECT mr.status, mr.record_date
            FROM monitoring_records mr
            WHERE mr.asset_id = ? AND mr.id < ?
            ORDER BY mr.id DESC
            LIMIT 1
           .»ã_Ê×¬¢h­µçBˆ™]\›ˆ™\\™WØÝ\ÝÛY\—Ü™\Ü
ˆ™\ÜˆÛÛÛÜ—ÜšXÙWÜ\—ÚÝÚ\ÛÛÛÜ—ÜšXÙWÜ\—ÚÝÚˆš[[™×ØÛÛ™šYÏXš[[™×ØÛÛ™šYËˆ
B‚‚™Yˆ[œÝ\™WÜ™\ÜÙ]WÜ™\]Y\ÝÊˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆ\ÜÙ]ÚYÎˆ\ÝÚ[Kˆ\š[Ùˆ™\Ü[™Ô\š[Ùˆ[˜ÛYWÝØ]ˆ›ÛÛˆ™\]Y\ÝÜÛÝ\˜ÙNˆÝ‹ˆ™Y™\™[˜ÙWÙ]Nˆ]KŠHOˆXÝÜÝ‹[žWN‚ˆÝ\œ™[Û[ÛH™Y™\™[˜ÙWÙ]Kœ™\XÙJ^OLJBˆ›Ü›X[^™YØ\ÜÙ]ÚYÈHÛÜY
Ú[
˜[YJH›Üˆ˜[YH[ˆ\ÜÙ]ÚYÈYˆ˜[Y_JBˆ›Ø—ÚYÎˆ\ÝÚ[HH×BˆØ\›š[™ÜÎˆ\ÝÜÝ—HH×Bˆ™]\ÙYØÛÝ[Hˆ]Y]YYØÛÝ[Hˆ›Üˆ[ÛÜÝ\[ˆ\š[Ùš[˜ÛYYÛ[ÛÎ‚ˆ[ÛÛX™[H[ÛÜÝ\œÝ™[YJ‰VKI[HŠBˆYˆ[ÛÜÝ\HÝ\œ™[Û[Û‚ˆÛÛ[YBˆZ\ÜÚ[™×Ü›ÙXÝ[Û—ÚYÈHÂˆ\ÜÙ]ÚYˆ›Üˆ\ÜÙ]ÚY[ˆ›Ü›X[^™YØ\ÜÙ]ÚYÂˆYˆ]˜[X]WÛØØ[Û[ÛWÜ›ÙXÝ[Û—Ü]X[]JˆÛÛ›‹ˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆ›ÝšY\RS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹ˆ[ÛÜÝ\[[ÛÜÝ\ˆ™Y™\™[˜ÙWÙ]O\™Y™\™[˜ÙWÙ]Kˆ
KœÝ]\ÂˆOH˜ÛÛ\]H‚ˆBˆYˆZ\ÜÚ[™×Ü›ÙXÝ[Û—ÚYÎ‚ˆ›Ø—ÚYÜ™X]YHÜ™X]WÛÜ—Ü™]\ÙWÜ™\ÜÙ]WÚ›ØŠˆÛÛ›‹ˆ›Ø—Ý\OH™\Ú[ÛœÛÛ\—Ü™\ÜÜ›ÙXÝ[Û—Ü™\]Y\Ý‹ˆ›ÝšY\RS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹ˆY]šXÏHœ›ÙXÝ[Ûˆ‹ˆ\š[ÙÜÝ\[[ÛÜÝ\ˆ\š[ÙÙ[™[[ÛÙ[™
[ÛÜÝ\
Kˆ\ÜÙ]ÚYÏ[Z\ÜÚ[™×Ü›ÙXÝ[Û—ÚYËˆ™\]Y\ÝÜÛÝ\˜ÙO\™\]Y\ÝÜÛÝ\˜ÙKˆ^˜WÜ\˜[\Ï^Èœ™\ÜÛ[ÛŽˆ[ÛÛX™[Kˆ
Bˆ›Ø—ÚYË˜\[™
›Ø—ÚY
Bˆ™]\ÙYØÛÝ[
ÏH[
›ÝÜ™X]Y
Bˆ]Y]YYØÛÝ[
ÏH[
Ü™X]Y
BˆØ\›š[™ÜË˜\[™
ˆˆœ›ÙXÝ[Û—ØÛÛXÝ[Û—Ü[™[™ÎžÛ[ÛÛX™[Nš›Ø—ÞÚ›Ø—ÚYH‚ˆ
B‚ˆYˆ[˜ÛYWÝØ]‚ˆØ]Ù[™H[ÛÙ[™
[ÛÜÝ\
BˆZ\ÜÚ[™×ÝØ]ÚYÈHÂˆ\ÜÙ]ÚYˆ›Üˆ\ÜÙ]ÚY[ˆ›Ü›X[^™YØ\ÜÙ]ÚYÂˆYˆ›Ý™X[ÝØ]Ü\š[ÙÚ\×ØÛÛ\]JˆÛÛ›‹ˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆœ›ÛWÙ]O[[ÛÜÝ\ˆ×Ù]O]Ø]Ù[™ˆ
BˆBˆYˆZ\ÜÚ[™×ÝØ]ÚYÎ‚ˆ›Ø—ÚYÜ™X]YHÜ™X]WÛÜ—Ü™]\ÙWÜ™\ÜÙ]WÚ›ØŠˆÛÛ›‹ˆ›Ø—Ý\OH™\Ú[ÛœÛÛ\—Ü™\ÜÝØ]Ü™\]Y\Ý‹ˆ›ÝšY\RS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹ˆY]šXÏHØ]‹ˆ\š[ÙÜÝ\[[ÛÜÝ\ˆ\š[ÙÙ[™]Ø]Ù[™ˆ\ÜÙ]ÚYÏ[Z\ÜÚ[™×ÝØ]ÚYËˆ™\]Y\ÝÜÛÝ\˜ÙO\™\]Y\ÝÜÛÝ\˜ÙKˆ^˜WÜ\˜[\Ï^Âˆ™œ›ÛWÙ]HŽˆ[ÛÜÝ\š\ÛÙ›Ü›X]

Kˆ×Ù]HŽˆØ]Ù[™š\ÛÙ›Ü›X]

KˆKˆ
Bˆ›Ø—ÚYË˜\[™
›Ø—ÚY
Bˆ™]\ÙYØÛÝ[
ÏH[
›ÝÜ™X]Y
Bˆ]Y]YYØÛÝ[
ÏH[
Ü™X]Y
BˆØ\›š[™ÜË˜\[™
ˆˆØ]ØÛÛXÝ[Û—Ü[™[™ÎžÛ[ÛÛX™[Nš›Ø—ÞÚ›Ø—ÚYH‚ˆ
BˆÛÛ›‹˜ÛÛ[Z]

Bˆ™]\›ˆÂˆš›Ø—ÚYÈŽˆÛÜY
Ù]
›Ø—ÚYÊJKˆØ\›š[™ÜÈŽˆØ\›š[™ÜËˆœ™]\ÙYØÛÝ[Žˆ™]\ÙYØÛÝ[ˆœ]Y]YYØÛÝ[Žˆ]Y]YYØÛÝ[ˆB‚‚™Yˆ[œ]Y]YWÜ™\ÜÜ›ÙXÝ[Û—Ü™\]Y\ÝÊˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆ\ÜÙ]ÚYˆ[ˆ\š[Ùˆ™\Ü[™Ô\š[Ùˆ[Û×Ü™\]Z\š[™×Ù˜[˜XÚÎˆ\ÝÜÝ—Kˆ™Y™\™[˜ÙWÙ]Nˆ]KŠHOˆ\ÝÚ[N‚ˆYˆ›Ý[Û×Ü™\]Z\š[™×Ù˜[˜XÚÎ‚ˆ™]\›ˆ×Bˆ™\]Y\ÝÈH[œÝ\™WÜ™\ÜÙ]WÜ™\]Y\ÝÊˆÛÛ›‹ˆ\ÜÙ]ÚYÏVØ\ÜÙ]ÚYKˆ\š[Ù\\š[Ùˆ[˜ÛYWÝØ]Q˜[ÙKˆ™\]Y\ÝÜÛÝ\˜ÙOHš[™]šYX[Ü™\Ü‹ˆ™Y™\™[˜ÙWÙ]O\™Y™\™[˜ÙWÙ]Kˆ
Bˆ™]\›ˆ\Ý
™\]Y\ÝÖÈš›Ø—ÚYÈ—JB‚‚™Yˆ[ÛÙ[™
[ÛÜÝ\ˆ]JHOˆ]N‚ˆ™]\›ˆ
ˆ[ÛÜÝ\œ™\XÙJ^OLŽ
H
È[YY[J^\ÏM
Bˆ
Kœ™\XÙJ^OLJHH[YY[J^\ÏLJB‚‚™Yˆ™X[ÝØ]Ü\š[ÙÚ\×ØÛÛ\]JˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆ\ÜÙ]ÚYˆ[ˆœ›ÛWÙ]Nˆ]Kˆ×Ù]Nˆ]KŠHOˆ›ÛÛ‚ˆ^XÝYÙ^\ÈH
×Ù]HHœ›ÛWÙ]JK™^\È
ÈBˆ›ÝÈHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕˆÓÕS•
TÕSÕ]˜Z[Xš[]WÙ]JHTÈÛÝ™\™YÙ^\ËˆÕSJˆÐTÑBˆÒSˆ˜[YÜÛÝÈˆˆS‘ÙZYÚYØ]˜Z[Xš[]WÜÝTÈ“Õ•SˆSˆSÑHBˆS‘ˆ
HTÈ[˜[YÙ^\Âˆ”“ÓH[Ø]˜Z[Xš[]WÙZ[BˆÒT‘H›ÝšY\ˆHÈS‘\ÜÙ]ÚYHÂˆS‘]˜Z[Xš[]WÙ]H‘UÑQSˆÈS‘Âˆˆˆ‹ˆ
ˆS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹ˆ\ÜÙ]ÚYˆœ›ÛWÙ]Kš\ÛÙ›Ü›X]

Kˆ×Ù]Kš\ÛÙ›Ü›X]

Kˆ
Kˆ
K™™]ÚÛ™J
Bˆ™]\›ˆ
ˆ[
›ÝÖÈ˜ÛÝ™\™YÙ^\È—HÜˆ
HOH^XÝYÙ^\Âˆ[™[
›ÝÖÈš[˜[YÙ^\È—HÜˆ
HOHˆ
B‚‚™YˆÜ™X]WÛÜ—Ü™]\ÙWÜ™\ÜÙ]WÚ›ØŠˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆ›Ø—Ý\NˆÝ‹ˆ›ÝšY\ŽˆÝ‹ˆY]šXÎˆÝ‹ˆ\š[ÙÜÝ\ˆ]Kˆ\š[ÙÙ[™ˆ]Kˆ\ÜÙ]ÚYÎˆ\ÝÚ[Kˆ™\]Y\ÝÜÛÝ\˜ÙNˆÝ‹ˆ^˜WÜ\˜[\ÎˆXÝÜÝ‹[žWKŠHOˆ\VÚ[›ÛÛN‚ˆ›Ü›X[^™YÚYÈHÛÜY
Ú[
˜[YJH›Üˆ˜[YH[ˆ\ÜÙ]ÚYßJBˆXÝ]™HHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕY›Ø—Ý\K\˜[\×ÚœÛÛ‚ˆ”“ÓH˜XÚÙÜ›Ý[™Ú›ØœÂˆÒT‘H›Ø—Ý\HSˆ
ËÊBˆS‘Ý]\ÈSˆ
ˆ	Ü[™[™ÉË	Ü[›š[™ÉË	ÝØZ][™×Ø\WÜÛÝ	Ë	ÝØZ][™×Ü˜]WÛ[Z]	Âˆ
BˆÔ‘Tˆ–HYˆˆˆ‹ˆ
ˆ›Ø—Ý\Kˆ
ˆ™\Ú[ÛœÛÛ\—Û[ÛØÛÜÙH‚ˆYˆY]šXÈOHœ›ÙXÝ[Ûˆ‚ˆ[ÙH™\Ú[ÛœÛÛ\—Ú[™\\—Ø]˜Z[Xš[]WØ˜XÚÙš[‚ˆ
Kˆ
Kˆ
K™™]Ú[

Bˆ›Ø—ÚYˆ[›Û™HH›Û™Bˆ›Üˆ›ÝÈ[ˆXÝ]™N‚ˆ\˜[\ÈHXÛÙWÚ›Ø—Ü\˜[\Ê›ÝÖÈœ\˜[\×ÚœÛÛˆ—JBˆ^\Ý[™×ÜÝ\HÝŠˆ\˜[\Ë™Ù]
œ™\ÜÛ[ÛŠBˆÜˆ\˜[\Ë™Ù]
™œ›ÛWÙ]HŠBˆÜˆˆ‚ˆ
BˆYˆY]šXÈOHœ›ÙXÝ[ÛˆŽ‚ˆ\š[ÙÛX]Ú\ÈH^\Ý[™×ÜÝ\OH\š[ÙÜÝ\œÝ™[YJ‰VKI[HŠBˆ[ÙN‚ˆ\š[ÙÛX]Ú\ÈH
ˆ^\Ý[™×ÜÝ\OH\š[ÙÜÝ\š\ÛÙ›Ü›X]

Bˆ[™ÝŠ\˜[\Ë™Ù]
×Ù]HŠHÜˆˆŠHOH\š[ÙÙ[™š\ÛÙ›Ü›X]

Bˆ
BˆYˆ›Ý\š[ÙÛX]Ú\Î‚ˆÛÛ[YBˆ^\Ý[™×ÚYÈHÂˆ[
˜[YJBˆ›Üˆ˜[YH[ˆ\˜[\Ë™Ù]
˜\ÜÙ]ÚYÈŠHÜˆ×BˆYˆÝŠ˜[YJKš\ÙYÚ]

BˆBˆYˆ›Ý^\Ý[™×ÚYÈÜˆÙ]
›Ü›X[^™YÚYÊKš\ÜÝXœÙ]
^\Ý[™×ÚYÊN‚ˆ›Ø—ÚYH[
›ÝÖÈšY—JBˆœ™XZÂˆÜ™X]YH˜[ÙBˆYˆ›Ø—ÚY\È›Û™N‚ˆ›Ø—ÚYÜ™X]YHÜ™X]WØ˜XÚÙÜ›Ý[™Ú›ØŠˆÛÛ›‹ˆ›Ø—Ý\KˆÂˆœ›ÝšY\ˆŽˆ›ÝšY\‹ˆ›Y]šXÈŽˆY]šXËˆœ\š[ÙÜÝ\Žˆ\š[ÙÜÝ\š\ÛÙ›Ü›X]

Kˆœ\š[ÙÙ[™Žˆ\š[ÙÙ[™š\ÛÙ›Ü›X]

Kˆ˜\ÜÙ]ÚYÈŽˆ›Ü›X[^™YÚYËˆšYÙÙ\—Ý\HŽˆœ™\ÜÙ˜Y‹ˆ
Š™^˜WÜ\˜[\ËˆKˆ
BˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆS”ÑT•S•È™\ÜÙ]WÜ™\]Y\ÝÙ]™[È
ˆ™\]Y\ÝÜÛÝ\˜ÙK›ÝšY\‹Y]šXË\š[ÙÜÝ\\š[ÙÙ[™ˆ\ÜÙ]ÚY×ÚœÛÛ‹˜XÚÙÜ›Ý[™Ú›Ø—ÚY™]\ÙYÙ^\Ý[™×Ú›Ø‹Ü™X]YØ]ˆ
HSQTÈ
ËËËËËËËËÊBˆˆˆ‹ˆ
ˆ™\]Y\ÝÜÛÝ\˜ÙKˆ›ÝšY\‹ˆY]šXËˆ\š[ÙÜÝ\š\ÛÙ›Ü›X]

Kˆ\š[ÙÙ[™š\ÛÙ›Ü›X]

KˆœÛÛ‹™[\Ê›Ü›X[^™YÚYÊKˆ›Ø—ÚYˆ[
›ÝÜ™X]Y
KˆÙ\šX[^™WØ˜XÚÙÜ›Ý[™Ú›Ø—Ý[Y\Ý[\

Kˆ
Kˆ
Bˆ™]\›ˆ›Ø—ÚYÜ™X]Y‚‚™Yˆ[œÝ\™WÜÜ›Û[×Ü™\Ý[Ù]WÜ™\]Y\ÝÊˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ™\Ý[ˆ[žKˆ
‹ˆ™\]Y\ÝÜÛÝ\˜ÙNˆÝ‹ŠHOˆ[žN‚ˆ\ÜÙ]ÚYÈHÂˆ[
›ÝË˜\ÜÙ]ÚY
Bˆ›Üˆ›ÝÈ[ˆ™\Ý[œ›ÝÜÂˆYˆ›ÝË˜\ÜÙ]ÚY\È›Ý›Û™BˆBˆ[˜ÛYWÝØ]H[žJˆÛÛ[[‹š\ÚX›H[™ÛÛ[[‹›Y]šX×ÚÙ^HOH˜]˜Z[Xš[]WÜÝ‚ˆ›ÜˆÛÛ[[ˆ[ˆ™\Ý[œ›Ùš[K˜ÛÛ[[œÂˆ
Bˆ™\]Y\ÝÈH[œÝ\™WÜ™\ÜÙ]WÜ™\]Y\ÝÊˆÛÛ›‹ˆ\ÜÙ]ÚYÏX\ÜÙ]ÚYËˆ\š[Ù\™\Ý[œ\š[Ùˆ[˜ÛYWÝØ]Z[˜ÛYWÝØ]ˆ™\]Y\ÝÜÛÝ\˜ÙO\™\]Y\ÝÜÛÝ\˜ÙKˆ™Y™\™[˜ÙWÙ]OXÝ\œ™[Û\Ø›Û—Ù]J
Kˆ
Bˆ›ÝÜÈH™\Ý[œ›ÝÜÂˆÝ[[X\žHH™\Ý[œÝ[[X\žBˆYˆ[˜ÛYWÝØ]‚ˆØ[š]^™YÜ›ÝÜÈH×Bˆ›Üˆ›ÝÈ[ˆ™\Ý[œ›ÝÜÎ‚ˆØ]ØÛÛ\]HH
ˆ›ÝË˜\ÜÙ]ÚY\È›Ý›Û™Bˆ[™™X[ÝØ]Ü\š[ÙÚ\×ØÛÛ\]JˆÛÛ›‹ˆ\ÜÙ]ÚYZ[
›ÝË˜\ÜÙ]ÚY
Kˆœ›ÛWÙ]O\™\Ý[œ\š[ÙœÝ\ˆ×Ù]O\™\Ý[œ\š[Ù™[™ˆ
Bˆ
BˆYˆØ]ØÛÛ\]N‚ˆØ[š]^™YÜ›ÝÜË˜\[™
›ÝÊBˆÛÛ[YBˆØ[š]^™YÜ›ÝÜË˜\[™
ˆ™\XÙJˆ›ÝËˆ˜[Y\Ï^ÊŠœ›ÝË˜[Y\Ë˜]˜Z[Xš[]WÜÝŽˆ›Û™_KˆØ\›š[™ÜÏ]\JˆÛÜY
ˆÙ]
›ÝËØ\›š[™ÜÊBˆÈØ]ØÛÛXÝ[Û—Ü[™[™ÈŸBˆ
Bˆ
Kˆ
Bˆ
Bˆ›ÝÜÈH\JØ[š]^™YÜ›ÝÜÊBˆÝ[[X\žHHYÙÜ™YØ]WÜ›ÝÜÊ›ÝÜË™\Ý[˜ÛÛ[[œÊBˆY]Y]HHÂˆ
Š™XÝ
™\Ý[›Y]Y]JKˆ™]WÜ™\]Y\ÝÚ›Ø—ÚYÈŽˆ™\]Y\ÝÖÈš›Ø—ÚYÈ—Kˆ™]WÜ™\]Y\ÝÝØ\›š[™ÜÈŽˆ™\]Y\ÝÖÈØ\›š[™ÜÈ—Kˆ™]WÜ™\]Y\ÝÜ™]\ÙYØÛÝ[Žˆ™\]Y\ÝÖÈœ™]\ÙYØÛÝ[—KˆBˆ™]\›ˆ™\XÙJˆ™\Ý[ˆ›ÝÜÏ\›ÝÜËˆÝ[[X\žO\Ý[[X\žKˆØ\›š[™ÜÏ]\JˆÛÜY
Ù]
™\Ý[Ø\›š[™ÜÊHÙ]
™\]Y\ÝÖÈØ\›š[™ÜÈ—JJBˆ
KˆY]Y]O[Y]Y]Kˆ
B‚‚™YˆZ[Ù\Ú[ÛœÛÛ\—ØÝ\ÝÛY\—Ü›ÙXÝ[Û—Ü™\Ü
ˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆ\ÜÙ]ÚYˆ[ˆ™\ÜÛ[ÛˆÝ‹ˆ[XÝšXÚ]WÜšXÙNˆ›Ø]ˆÙ[ÜšXÙNˆ›Ø]ˆÛÛÛÜ—ÜšXÙWÜ\—ÚÝÚˆ›Ø]HŒˆš[[™×ØÛÛ™šYÎˆš[[™ÐÛÛ™šYÈ›Û™HH›Û™Kˆ›Ü˜ÙWØ\Nˆ›ÛÛH˜[ÙKˆ\š[Ùˆ™\Ü[™Ô\š[Ù›Û™HH›Û™Kˆ™Y™\™[˜ÙWÙ]Nˆ]H›Û™HH›Û™KŠHOˆXÝÜÝ‹[žWN‚ˆ\š[ÙHÜ™\ÛÛ™WØÝ\ÝÛY\—Ü™\Ü[™×Ü\š[Ù
™\ÜÛ[Û\š[Ù
Bˆ™Y™\™[˜ÙWÙ]HH™Y™\™[˜ÙWÙ]HÜˆ]KÙ^J
BˆØØ[Ü™\ÜHZ[ÛØØ[ØÝ\ÝÛY\—Ü›ÙXÝ[Û—Ü™\Ü
ˆÛÛ›‹ˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆ™\ÜÛ[Û\™\ÜÛ[Ûˆ[XÝšXÚ]WÜšXÙOY[XÝšXÚ]WÜšXÙKˆÙ[ÜšXÙO\Ù[ÜšXÙKˆÛÛÛÜ—ÜšXÙWÜ\—ÚÝÚ\ÛÛÛÜ—ÜšXÙWÜ\—ÚÝÚˆš[[™×ØÛÛ™šYÏXš[[™×ØÛÛ™šYËˆ\š[Ù\\š[Ùˆ™Y™\™[˜ÙWÙ]O\™Y™\™[˜ÙWÙ]Kˆ
BˆYˆØØ[Ü™\Ü\È›Û™N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ“˜[È›ÚHÜÜÚ]™[™\\˜\ˆÈ™[]Üš[ÈÛÛHÜÈYÜÈØØZ\ËˆŠBˆ]WÜ™\]Y\ÝÈH[œÝ\™WÜ™\ÜÙ]WÜ™\]Y\ÝÊˆÛÛ›‹ˆ\ÜÙ]ÚYÏVØ\ÜÙ]ÚYKˆ\š[Ù\\š[Ùˆ[˜ÛYWÝØ]Q˜[ÙKˆ™\]Y\ÝÜÛÝ\˜ÙOHš[™]šYX[Ü™\Ü‹ˆ™Y™\™[˜ÙWÙ]O\™Y™\™[˜ÙWÙ]Kˆ
Bˆ™Yœ™\ÚÚ›Ø—ÚYÈH\Ý
]WÜ™\]Y\ÝÖÈš›Ø—ÚYÈ—JBˆØØ[Ü™\ÜÈœ›ÙXÝ[Û—Ü™Yœ™\ÚÚ›Ø—ÚYÈ—HH™Yœ™\ÚÚ›Ø—ÚYÂˆØØ[Ü™\ÜÈœ›ÙXÝ[Û—Ü™Yœ™\ÚÜ]Y]YY—HH›ÛÛ
™Yœ™\ÚÚ›Ø—ÚYÊBˆØØ[Ü™\ÜÈ™]WÜ™\]Y\ÝÝØ\›š[™ÜÈ—HH\Ý
ˆ]WÜ™\]Y\ÝÖÈØ\›š[™ÜÈ—Bˆ
BˆØØ[Ü™\ÜœÙ]Y˜][
œ™\ÜÛ›Ý\È‹×JK™^[™
ˆ]WÜ™\]Y\ÝÖÈØ\›š[™ÜÈ—Bˆ
BˆØØ[Ü™\ÜÈ™›Ü˜ÙWØ\WÚYÛ›Ü™Y—HH›ÛÛ
›Ü˜ÙWØ\JBˆ™]\›ˆØØ[Ü™\Ü‚‚™YˆYØÝ\ÝÛY\—Ü™\ÜØ]˜Z[Xš[]JˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ™\ÜˆXÝÜÝ‹[žWKˆ
‹ˆ\ÜÙ]ÚYˆ[ˆ\š[Ùˆ™\Ü[™Ô\š[Ùˆ[œÝ\™WÜ™\]Y\ÝÎˆ›ÛÛHYKŠHOˆ›ÛÛ‚ˆYˆ[œÝ\™WÜ™\]Y\ÝÎ‚ˆ™\]Y\ÝÈH[œÝ\™WÜ™\ÜÙ]WÜ™\]Y\ÝÊˆÛÛ›‹ˆ\ÜÙ]ÚYÏVØ\ÜÙ]ÚYKˆ\š[Ù\\š[Ùˆ[˜ÛYWÝØ]UYKˆ™\]Y\ÝÜÛÝ\˜ÙOHš[™]šYX[Ü™\ÜØ]˜Z[Xš[]H‹ˆ™Y™\™[˜ÙWÙ]OXÝ\œ™[Û\Ø›Û—Ù]J
Kˆ
Bˆ™\ÜœÙ]Y˜][
™]WÜ™\]Y\ÝÝØ\›š[™ÜÈ‹×JK™^[™
ˆ™\]Y\ÝÖÈØ\›š[™ÜÈ—Bˆ
Bˆ™\ÜœÙ]Y˜][
œ™\ÜÛ›Ý\È‹×JK™^[™
™\]Y\ÝÖÈØ\›š[™ÜÈ—JBˆYˆ›Ý™X[ÝØ]Ü\š[ÙÚ\×ØÛÛ\]JˆÛÛ›‹ˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆœ›ÛWÙ]O\\š[ÙœÝ\ˆ×Ù]O\\š[Ù™[™ˆ
N‚ˆ™\ÜÈ˜]˜Z[Xš[]WÙ\œ›Üˆ—HHØ]ØÛÛXÝ[Û—Ü[™[™È‚ˆ™\ÜœÙ]Y˜][
œ™\ÜÛ›Ý\È‹×JK˜\[™
ˆ•ÐU[™\ÜÛš]™[8 %YÜÈ[H™XÛÛNÈ™[]Üš[È[H˜\ØÝ[šËˆ‚ˆ
Bˆ™\ÜœÜ
˜]˜Z[Xš[]WÜÝ‹›Û™JBˆ™\ÜÈš[˜ÛYWØ]˜Z[Xš[]WÚÜH—HH˜[ÙBˆ™]\›ˆ˜[ÙBˆžN‚ˆ]˜Z[Xš[]HHÙ]Û[ÛWØ]˜Z[Xš[]JÛÛ›‹\ÜÙ]ÚY\š[ÙœÝ\\š[Ù™[™
Bˆ^Ù\^Ù\[ÛŽ‚ˆÑÑÑT‹™^Ù\[ÛŠˆ‘˜Z[YÈØYÝ\ÝÛY\ˆ™\Ü]˜Z[Xš[]H\ÜÙ]ÚYI\È\š[ÙÜÝ\I\È\š[ÙÙ[™I\È‹ˆ\ÜÙ]ÚYˆ\š[ÙœÝ\š\ÛÙ›Ü›X]

Kˆ\š[Ù™[™š\ÛÙ›Ü›X]

Kˆ
Bˆ™\ÜÈ˜]˜Z[Xš[]WÙ\œ›Üˆ—HH˜]˜Z[Xš[]WÛÛÚÝ\Ù˜Z[Y‚ˆ™\ÜœÙ]Y˜][
œ™\ÜÛ›Ý\È‹×JK˜\[™
‘\œ›È[È›ØÝ\˜\ˆ\ÜÛšXš[YYH
	JKˆŠBˆ™]\›ˆ˜[ÙBˆYˆ]˜Z[Xš[]H\È›Û™N‚ˆ™\ÜÈ˜]˜Z[Xš[]WÙ\œ›Üˆ—HHØ]ØÛÛXÝ[Û—Ü[™[™È‚ˆ™\ÜœÙ]Y˜][
œ™\ÜÛ›Ý\È‹×JK˜\[™
ˆ•ÐU[™\ÜÛš]™[8 %YÜÈ[H™XÛÛNÈ™[]Üš[È[H˜\ØÝ[šËˆ‚ˆ
Bˆ™]\›ˆ˜[ÙBˆ™\ÜÈš[˜ÛYWØ]˜Z[Xš[]WÚÜH—HHYBˆ™\ÜÈ˜]˜Z[Xš[]WÜÝ—HH]˜Z[Xš[]Bˆ™]\›ˆYB‚‚™Yˆ^ÜØÝ\ÝÛY\—Ü›ÙXÝ[Û—ÜŠ™\ÜˆXÝÜÝ‹[žWJN‚ˆ—Øž]\ÈHZ[ØÝ\ÝÛY\—Ü™\ÜÜŠ™\ÜÙÛ×Ü]PTÑWÑTˆÈœÝ]XÈˆÈœÛÛÛÜ‹[ÙÛËœ™ÈŠBˆY™™\ˆH[Ëž]\ÒSÊ—Øž]\ÊBˆØY™WÛ˜[YHH›Ü›X[^™WÛ˜[YJ™\ÜÈ˜\ÜÙ]—VÈœ›Ú™XÝÛ˜[YH—JKœ™\XÙJˆ‹—ÈŠHÜˆœ™[]Üš[È‚ˆ[Ù[HÝŠ™\Ü™Ù]
œ™\ÜÝ\HŠHÜˆ™\ÈŠK\\Š
Bˆ\š[ÙÜÛYÈHÝŠ™\Ü™Ù]
œ\š[ÙÛX™[ŠHÜˆ™\ÜÈ›[ÛÜÝ\—KœÝ™[YJ‰[KIVHŠJKœ™\XÙJˆ‹—ÈŠKœ™\XÙJ‹È‹‹HŠBˆš[[˜[YHHˆ”™[]Üš[×ÞÛ[Ù[WÞÜØY™WÛ˜[Y_WÞÜ\š[ÙÜÛYßKœˆ‚ˆ™]\›ˆÙ[™Ùš[JY™™\‹\×Ø]XÚY[UYKÝÛ›ØYÛ˜[YOYš[[˜[YKZ[Y]\OH˜\XØ][Û‹ÜˆŠB‚‚™YˆÝÜ™WÜ™\ÜÛÙÛÊš[WÜÝÜ˜YÙK
‹ÛÛÙÛ×Ü]ˆÝˆHˆŠHOˆÝŽ‚ˆš[[˜[YHHØY™WÝ\ØYÙš[[˜[YJš[WÜÝÜ˜YÙK™š[[˜[YHÜˆˆŠBˆ^[œÚ[ÛˆHš[[˜[YKœœÜ]
‹ˆ‹JVËLWK›ÝÙ\Š
HYˆ‹ˆˆ[ˆš[[˜[YH[ÙHˆ‚ˆYˆ^[œÚ[Ûˆ›Ý[ˆÈœ™È‹šœÈ‹šœYÈŸN‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠš[˜[YÛÙÛ×Ù^[œÚ[ÛˆŠBˆ^[ØYHš[WÜÝÜ˜YÙKœ™XY

BˆYˆ›Ý^[ØYÜˆ[Š^[ØY
HˆLLˆ
ˆL‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠš[˜[YÛÙÛ×ÜÚ^™HŠBˆÚYZYÚH[XYÙWÙ[Y[œÚ[ÛœÊ^[ØY^[œÚ[ÛŠBˆYˆÚYHÜˆZYÚHÜˆÚYˆÜˆZYÚˆLŒ‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠš[˜[YÛÙÛ×Ù[Y[œÚ[ÛœÈŠBˆÙÛ×Ù\ˆHTÐQÑTˆÈœ™\ÜÛÙÛÜÈ‚ˆÙÛ×Ù\‹›ZÙ\Š\™[ÏUYK^\ÝÛÚÏUYJBˆ\™Ù]HÙÛ×Ù\ˆÈˆžÙ]][YK››ÝÊ
KœÝ™[YJ	ÉVI[IY	R	SITÉÊ_WÞÙš[[˜[Y_H‚ˆ\™Ù]Üš]WØž]\Ê^[ØY
Bˆ™]\›ˆÝÜ™WÜ[[YWÜ™[]]™WÜ]
\™Ù]
B‚‚™YˆØY™WÝ\ØYÙš[[˜[YJš[[˜[YNˆÝŠHOˆÝŽ‚ˆÛX[™YHÙXÝ\™WÙš[[˜[YJš[[˜[YJBˆYˆ›ÝÛX[™YÜˆ[žJ][H[ˆÛX[™Y›Üˆ][H[ˆÈ‹‹ˆ‹‹È‹—ŸJN‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠš[˜[YÝ\ØYÙš[[˜[YHŠBˆ™]\›ˆÛX[™Y‚‚™Yˆ[XYÙWÙ[Y[œÚ[ÛœÊ^[ØYˆž]\Ë^[œÚ[ÛŽˆÝŠHOˆ\VÚ[[N‚ˆYˆ^[œÚ[ÛˆOHœ™ÈŽ‚ˆYˆ›Ý^[ØYœÝ\ÝÚ]
ˆ—T‘×——XWˆŠHÜˆ[Š^[ØY
H‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠš[˜[YÛÙÛ×ØÛÛ[ŠBˆÚYZYÚHÝXÝ[œXÚÊ’RH‹^[ØYÌMŽŒJBˆ™]\›ˆ[
ÚY
K[
ZYÚ
BˆYˆ^[œÚ[Ûˆ[ˆÈšœÈ‹šœYÈŸN‚ˆYˆ›Ý^[ØYœÝ\ÝÚ]
ˆ—™—ŠN‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠš[˜[YÛÙÛ×ØÛÛ[ŠBˆ[™^H‚ˆÚ[H[™^
ÈH[Š^[ØY
N‚ˆYˆ^[ØYÚ[™^HOH‘Ž‚ˆ[™^
ÏHBˆÛÛ[YBˆX\šÙ\ˆH^[ØYÚ[™^
ÈWBˆ[™^
ÏH‚ˆYˆX\šÙ\ˆ[ˆÌ_N‚ˆÛÛ[YBˆ[™ÝH[™œ›ÛWØž]\Ê^[ØYÚ[™^ˆ[™^
È—K˜šYÈŠBˆYˆ[™ÝŽ‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠš[˜[YÛÙÛ×ØÛÛ[ŠBˆYˆX\šÙ\ˆ[ˆÌÌÌŸH[™[™^
ÈÈ[Š^[ØY
N‚ˆZYÚH[™œ›ÛWØž]\Ê^[ØYÚ[™^
ÈÈˆ[™^
ÈWK˜šYÈŠBˆÚYH[™œ›ÛWØž]\Ê^[ØYÚ[™^
ÈHˆ[™^
È×K˜šYÈŠBˆ™]\›ˆ[
ÚY
K[
ZYÚ
Bˆ[™^
ÏH[™Ýˆ˜Z\ÙH˜[YQ\œ›ÜŠš[˜[YÛÙÛ×ØÛÛ[ŠB‚‚™YˆZ[ÙÙ[™\˜][Û—Ú›ØœÊ›Ü›Nˆ[žK™\ÜÝ\NˆÝ‹XZ[—Ù›Ü›X]Îˆ\VÜÝ‹‹‹—JHOˆ\ÝÙXÝÜÝ‹[žWWN‚ˆ\š[ÙÈH\œÙWÙÙ[™\˜][Û—Ü\š[ÙÊ›Ü›JBˆ›ØœÎˆ\ÝÙXÝÜÝ‹[žWWHH×BˆYˆ™\ÜÝ\HOHš[™]šYX[Ž‚ˆ\ÜÙ]ÚYÈHÚ[
][JH›Üˆ][H[ˆ›Ü›K™Ù]\Ý
˜\ÜÙ]ÚYÈŠHYˆÝŠ][JKš\ÙYÚ]

WBˆYˆ›Ý\ÜÙ]ÚYÎ‚ˆ\ÜÙ]ÚYÈHÚ[
›Ü›K™Ù]
˜\ÜÙ]ÚY‹ŒŠHÜˆ
WBˆ\ÜÙ]ÚYÈHÛÜY
Ú][H›Üˆ][H[ˆ\ÜÙ]ÚYÈYˆ][_JBˆYˆ›Ý\ÜÙ]ÚYÎ‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ’[œÝ[péðèÛÈØœšYØ]0ìÜšXKˆŠBˆYˆ[Š\ÜÙ]ÚYÊHˆPVÐUÒÐTÔÑUÎ‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ‘[X\ÚXY\È[œÝ[XÛÙ\È›ÈY\Û[È[‹ˆŠBˆ›Üˆ\ÜÙ]ÚY[ˆ\ÜÙ]ÚYÎ‚ˆ›Üˆ\š[Ù[ˆ\š[ÙÎ‚ˆ›Üˆ›][ˆXZ[—Ù›Ü›X]Î‚ˆ›ØœË˜\[™
È˜\ÜÙ]ÚYŽˆ\ÜÙ]ÚYœ\š[ÙŽˆ\š[Ù™›Ü›X]Žˆ›]JBˆ[ÙN‚ˆ›Üˆ\š[Ù[ˆ\š[ÙÎ‚ˆ›Üˆ›][ˆXZ[—Ù›Ü›X]Î‚ˆ›ØœË˜\[™
Èœ\š[ÙŽˆ\š[Ù™›Ü›X]Žˆ›]JBˆYˆ[Š›ØœÊHˆPVÕÕSÓÕUUÎ‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ‘[X\ÚXYÜÈÝ]]È›ÈY\Û[È[‹ˆŠBˆ™]\›ˆ›ØœÂ‚‚™YˆZ[ÜÛ˜\ÚÝÙÙ[™\˜][Û—Ú›ØœÊÛ˜\ÚÝÜ™\Ý[XZ[—Ù›Ü›X]Îˆ\VÜÝ‹‹‹—JHOˆ\ÝÙXÝÜÝ‹[žWWN‚ˆ\š[ÙHÂˆœ\š[ÙÝ\HŽˆÛ˜\ÚÝÜ™\Ý[œ\š[Ùœ\š[ÙÝ\K˜[YKˆœ™\ÜÛ[ÛŽˆÛ˜\ÚÝÜ™\Ý[œ\š[ÙœÝ\œÝ™[YJ‰VKI[HŠKˆœ\š[ÙÜÝ\ŽˆÛ˜\ÚÝÜ™\Ý[œ\š[ÙœÝ\š\ÛÙ›Ü›X]

Kˆœ\š[ÙÙ[™ŽˆÛ˜\ÚÝÜ™\Ý[œ\š[Ù™[™š\ÛÙ›Ü›X]

KˆBˆ™]\›ˆÞÈœ\š[ÙŽˆ\š[Ù™›Ü›X]Žˆ›]H›Üˆ›][ˆXZ[—Ù›Ü›X]×B‚‚™Yˆ™Z™XÝÜÛ˜\ÚÝÜ\š[ÙÛÝ™\œšY\Ê›Ü›Nˆ[žKÛ˜\ÚÝÜ™\Ý[
HOˆ›Û™N‚ˆ˜]×Û[ÛÈHÜ\›Üˆ][H[ˆ›Ü›K™Ù]\Ý
œ™\ÜÛ[ÛÈŠH›Üˆ\[ˆ™KœÜ]
ˆ–×Ë×JÈ‹ÝŠ][JJHYˆ\BˆYˆ[ŠÙ]
˜]×Û[ÛÊJHˆN‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠœÛ˜\ÚÝÜ™Z™XÝ×Û][\WÜ\š[ÙÈŠBˆÝX›Z]YÝ\HHÝŠ›Ü›K™Ù]
œ\š[ÙÝ\HŠHÜˆÛ˜\ÚÝÜ™\Ý[œ\š[Ùœ\š[ÙÝ\K˜[YJBˆYˆÝX›Z]YÝ\HOHÛ˜\ÚÝÜ™\Ý[œ\š[Ùœ\š[ÙÝ\K˜[YN‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠœÛ˜\ÚÝÜ\š[ÙÛZ\ÛX]ÚŠBˆÝX›Z]YÛ[ÛH›Ü›K™Ù]
œ™\ÜÛ[ÛŠBˆYˆÝX›Z]YÛ[Û[™›Ü›X[^™WÜ™\ÜÛ[Û
ÝX›Z]YÛ[Û
HOHÛ˜\ÚÝÜ™\Ý[œ\š[ÙœÝ\œÝ™[YJ‰VKI[HŠN‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠœÛ˜\ÚÝÜ\š[ÙÛZ\ÛX]ÚŠB‚‚™Yˆ™\ÛÛ™WÜ™\ÜØÛY[ÚÙ^JÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹
‹Ü›Û[×ÚYˆ[›Û™HH›Û™K\ÜÙ]ÚYˆ[›Û™HH›Û™JHOˆÝŽ‚ˆYˆÜ›Û[×ÚY‚ˆ›ÝÈHÛÛ›‹™^XÝ]J”ÑSPÕ˜[YH”“ÓHÜ›Û[×ÙÜ›Ý\ÈÒT‘HYHÈ‹
Ü›Û[×ÚY
JK™™]ÚÛ™J
Bˆ™]\›ˆ›Ü›X[^™WÛ˜[YJ›ÝÖÈ›˜[YH—JHYˆ›ÝÈ[ÙHˆ‚ˆYˆ\ÜÙ]ÚY‚ˆ›ÝÈHÛÛ›‹™^XÝ]J”ÑSPÕšY‹›Ú™XÝÛ˜[YH”“ÓH\ÜÙ]ÈÒT‘HYHÈ‹
\ÜÙ]ÚY
JK™™]ÚÛ™J
Bˆ™]\›ˆ›Ü›X[^™WÛ˜[YJ›ÝÖÈ›šYˆ—HÜˆ›ÝÖÈœ›Ú™XÝÛ˜[YH—JHYˆ›ÝÈ[ÙHˆ‚ˆ™]\›ˆˆ‚‚‚™Yˆ™\ÛÛ™WÙÙ[™\˜][Û—ØÛY[ÚÙ^JÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹™\ÜÝ\NˆÝ‹›Ü›Nˆ[žKÜ›Û[×ÚYˆ[›Û™JHOˆÝŽ‚ˆYˆ™\ÜÝ\HOHœÜ›Û[ÈŽ‚ˆ™]\›ˆ™\ÛÛ™WÜ™\ÜØÛY[ÚÙ^JÛÛ›‹Ü›Û[×ÚY\Ü›Û[×ÚY
Bˆ\ÜÙ]ÚYÈHÚ[
][JH›Üˆ][H[ˆ›Ü›K™Ù]\Ý
˜\ÜÙ]ÚYÈŠHYˆÝŠ][JKš\ÙYÚ]

WBˆYˆ›Ý\ÜÙ]ÚYÈ[™ÝŠ›Ü›K™Ù]
˜\ÜÙ]ÚY‹ˆŠJKš\ÙYÚ]

N‚ˆ\ÜÙ]ÚYÈHÚ[
›Ü›K™Ù]
˜\ÜÙ]ÚYŠJWBˆÙ^\ÈHÜ™\ÛÛ™WÜ™\ÜØÛY[ÚÙ^JÛÛ›‹\ÜÙ]ÚYX\ÜÙ]ÚY
H›Üˆ\ÜÙ]ÚY[ˆ\ÜÙ]ÚYÈYˆ\ÜÙ]ÚYBˆÙ^\Ë™\ØØ\™
ˆŠBˆYˆ[ŠÙ^\ÊHˆN‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ›Z^YØÛY[Ø˜]ÚŠBˆ™]\›ˆ™^
]\ŠÙ^\ÊKˆŠB‚‚™Yˆ\œÙWÙÙ[™\˜][Û—Ü\š[ÙÊ›Ü›Nˆ[žJHOˆ\ÝÙXÝÜÝ‹Ý—WN‚ˆ\š[ÙÝ\HHÝŠ›Ü›K™Ù]
œ\š[ÙÝ\H‹›[ÛHŠHÜˆ›[ÛHŠBˆ˜]×Û[ÛÈH›Ü›K™Ù]\Ý
œ™\ÜÛ[ÛÈŠBˆYˆ˜]×Û[ÛÎ‚ˆ˜]×Û[ÛÈHÜ\›Üˆ][H[ˆ˜]×Û[ÛÈ›Üˆ\[ˆ™KœÜ]
ˆ–×Ë×JÈ‹ÝŠ][JJHYˆ\Bˆ[Yˆ›Ü›K™Ù]
œ™\ÜÛ[ÛÈŠN‚ˆ˜]×Û[ÛÈH™KœÜ]
ˆ–×Ë×JÈ‹ÝŠ›Ü›K™Ù]
œ™\ÜÛ[ÛÈŠJJBˆYˆ\š[ÙÝ\HOH›[ÛHŽ‚ˆ[ÛÈHÛ›Ü›X[^™WÜ™\ÜÛ[Û
][JH›Üˆ][H[ˆ˜]×Û[ÛÈYˆÝŠ][JKœÝš\

WBˆYˆ›Ý[ÛÎ‚ˆ[ÛÈHÛ›Ü›X[^™WÜ™\ÜÛ[Û
›Ü›K™Ù]
œ™\ÜÛ[Û‹ˆŠJWBˆ[ÛÈH\Ý
XÝ™œ›ÛZÙ^\Ê[ÛÊJBˆYˆ[Š[ÛÊHˆPVÐUÒÔT’SÑÎ‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ‘[X\ÚXYÜÈ\š[ÙÜÈ›ÈY\Û[È[‹ˆŠBˆ\š[ÙÈH×Bˆ›Üˆ[Û[ˆ[ÛÎ‚ˆ\š[ÙHZ[Ü\š[Ù
›[ÛH‹™\ÜÛ[Û[[Û
Bˆ\š[ÙË˜\[™
Èœ\š[ÙÝ\HŽˆ›[ÛH‹œ™\ÜÛ[ÛŽˆ[Ûœ\š[ÙÜÝ\Žˆ\š[ÙœÝ\š\ÛÙ›Ü›X]

Kœ\š[ÙÙ[™Žˆ\š[Ù™[™š\ÛÙ›Ü›X]

_JBˆ™]\›ˆ\š[ÙÂˆYˆ[Š˜]×Û[ÛÊHˆN‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ”\š[ÙÜÈ\Ý]\˜YÜÈ][\ÜÈ]™[HÙ\ˆÝX›Y]YÜÈÙ\\˜Y[Y[KˆŠBˆ\š[ÙHZ[Ü\š[Ù
ˆ\š[ÙÝ\KˆYX\Y›Ü›K™Ù]
œ™\ÜÞYX\ˆŠHÜˆ›Ü›X[^™WÜ™\ÜÛ[Û
›Ü›K™Ù]
œ™\ÜÛ[Û‹ˆŠJVÎKˆ]X\\Y›Ü›K™Ù]
œ™\ÜÜ]X\\ˆŠKˆÙ[Y\Ý\Y›Ü›K™Ù]
œ™\ÜÜÙ[Y\Ý\ˆŠKˆ
Bˆ™]\›ˆÂˆÂˆœ\š[ÙÝ\HŽˆ\š[ÙÝ\Kˆœ™\ÜÛ[ÛŽˆ\š[ÙœÝ\œÝ™[YJ‰VKI[HŠKˆœ™\ÜÞYX\ˆŽˆÝŠ\š[ÙœÝ\žYX\ŠKˆœ™\ÜÜ]X\\ˆŽˆÝŠ

\š[ÙœÝ\›[ÛHJHËÈÊH
ÈJKˆœ™\ÜÜÙ[Y\Ý\ˆŽˆŒHˆYˆ\š[ÙœÝ\›[ÛOHH[ÙHŒˆ‹ˆœ\š[ÙÜÝ\Žˆ\š[ÙœÝ\š\ÛÙ›Ü›X]

Kˆœ\š[ÙÙ[™Žˆ\š[Ù™[™š\ÛÙ›Ü›X]

KˆBˆB‚‚™YˆZ[ÜÜ›Û[×ÙÙ[™\˜][Û—Ü™\Ý[
ÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹›Ü›Nˆ[žKÜ›Û[×ÚYˆ[›Û™KÛ˜\ÚÝÚYˆ[›Û™K\š[ÙÚ›ØŽˆXÝÜÝ‹Ý—JN‚ˆYˆÛ˜\ÚÝÚY‚ˆ™\Ý[HÙ]ÜÜ›Û[×ÜÛ˜\ÚÝÜ™\Ý[
ÛÛ›‹Û˜\ÚÝÚY
BˆYˆ™\Ý[\È›Û™HÜˆ
Ü›Û[×ÚY[™™\Ý[œÜ›Û[×ÚYOHÜ›Û[×ÚY
N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ”Û˜\ÚÝ[˜[YËˆŠBˆ™]\›ˆ™\Ý[ˆYˆ›ÝÜ›Û[×ÚY‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ”Ü›Û[ÈØœšYØ]Üš[ËˆŠBˆÜ›Ý\HÛÛ›‹™^XÝ]J”ÑSPÕ
ˆ”“ÓHÜ›Û[×ÙÜ›Ý\ÈÒT‘HYHÈ‹
Ü›Û[×ÚY
JK™™]ÚÛ™J
BˆYˆ›ÝÜ›Ý\‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ”Ü›Û[È[˜[YËˆŠBˆ›Ùš[WÚYH[
›Ü›K™Ù]
œ›Ùš[WÚY‹ŒŠHÜˆ
Bˆ›Ùš[HHÙ]ÜÜ›Û[×Ü™\ÜÜ›Ùš[JÛÛ›‹›Ùš[WÚY
HYˆ›Ùš[WÚY[ÙHÙ]ÙY˜][ÜÜ›Û[×Ü™\ÜÜ›Ùš[JÛÛ›‹Ü›Û[×ÚY
Bˆ™\Ý[H™\\™WÜÜ›Û[×Ü™\Ü
ˆÛÛ›‹ˆÜ›Û[×ÚY\Ü›Û[×ÚYˆÜ›Û[×Û˜[YOYÜ›Ý\È›˜[YH—Kˆ›Ùš[O\›Ùš[Kˆ\š[ÙÝ\O\\š[ÙÚ›Ø–Èœ\š[ÙÝ\H—Kˆ™\ÜÛ[Û\\š[ÙÚ›Ø‹™Ù]
œ™\ÜÛ[ÛŠKˆYX\\\š[ÙÚ›Ø‹™Ù]
œ™\ÜÞYX\ˆŠHÜˆ\š[ÙÚ›Ø–Èœ\š[ÙÜÝ\—VÎKˆ]X\\\\š[ÙÚ›Ø‹™Ù]
œ™\ÜÜ]X\\ˆŠKˆÙ[Y\Ý\\\š[ÙÚ›Ø‹™Ù]
œ™\ÜÜÙ[Y\Ý\ˆŠKˆÛÛ\\š\ÛÛY›Ü›K™Ù]
˜ÛÛ\\š\ÛÛˆ‹ˆŠKˆ›Ùš[WÝ™\œÚ[Û[]\ÝÜÜ›Û[×Ü™\ÜÜ›Ùš[WÝ™\œÚ[ÛŠÛÛ›‹›Ùš[KšY
Kˆ
Bˆ™]\›ˆ[œÝ\™WÜÜ›Û[×Ü™\Ý[Ù]WÜ™\]Y\ÝÊˆÛÛ›‹ˆ™\Ý[ˆ™\]Y\ÝÜÛÝ\˜ÙOHœÜ›Û[×Ü™\ÜÙÙ[™\˜][Ûˆ‹ˆ
B‚‚™YˆZ[Ú[™]šYX[ÙÙ[™\˜][Û—Ü™\Ü
ˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ\ÜÙ]ÚYˆ[ˆ\š[ÙÚ›ØŽˆXÝÜÝ‹Ý—Kˆ
‹ˆ[˜ÛYWÝØ]ˆ›ÛÛH˜[ÙKŠHOˆXÝÜÝ‹[žWN‚ˆ\š[ÙHZ[Ü\š[Ù
ˆ\š[ÙÚ›Ø–Èœ\š[ÙÝ\H—Kˆ™\ÜÛ[Û\\š[ÙÚ›Ø‹™Ù]
œ™\ÜÛ[ÛŠKˆYX\\\š[ÙÚ›Ø‹™Ù]
œ™\ÜÞYX\ˆŠHÜˆ\š[ÙÚ›Ø–Èœ\š[ÙÜÝ\—VÎKˆ]X\\\\š[ÙÚ›Ø‹™Ù]
œ™\ÜÜ]X\\ˆŠKˆÙ[Y\Ý\\\š[ÙÚ›Ø‹™Ù]
œ™\ÜÜÙ[Y\Ý\ˆŠKˆ
Bˆš[[™×ØÛÛ™šYÈHÙ]Ø\ÜÙ]Øš[[™×ØÛÛ™šYÊÛÛ›‹\ÜÙ]ÚY™\Ü\K‘TÊBˆ™\ÜHZ[ÛØØ[ØÝ\ÝÛY\—Ü›ÙXÝ[Û—Ü™\Ü
ˆÛÛ›‹ˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆ™\ÜÛ[Û\\š[ÙœÝ\œÝ™[YJ‰VKI[HŠKˆ[XÝšXÚ]WÜšXÙOY›Ø]
š[[™×ØÛÛ™šYË™[XÝšXÚ]WÜšXÙWÙ]\—ÚÝÚ
KˆÙ[ÜšXÙOY›Ø]
š[[™×ØÛÛ™šYË™^ÜÜšXÙWÙ]\—ÚÝÚ
Kˆš[[™×ØÛÛ™šYÏXš[[™×ØÛÛ™šYËˆ\š[Ù\\š[Ùˆ
BˆYˆ™\Ü\È›Û™N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠˆ”Ù[HYÜÈ\˜HH[œÝ[XØ[ÈØ\ÜÙ]ÚYKˆŠBˆ™\ÜÈ˜\ÜÙ]ÚY—HH\ÜÙ]ÚYˆ™\ÜÈ™[™Ú[™WÝ™\œÚ[Ûˆ—HHš[™]šYX[\™\Ü]ŒH‚ˆ™\]Y\ÝÈH[œÝ\™WÜ™\ÜÙ]WÜ™\]Y\ÝÊˆÛÛ›‹ˆ\ÜÙ]ÚYÏVØ\ÜÙ]ÚYKˆ\š[Ù\\š[Ùˆ[˜ÛYWÝØ]Z[˜ÛYWÝØ]ˆ™\]Y\ÝÜÛÝ\˜ÙOHš[™]šYX[Ü™\ÜÙÙ[™\˜][Ûˆ‹ˆ™Y™\™[˜ÙWÙ]OXÝ\œ™[Û\Ø›Û—Ù]J
Kˆ
Bˆ™\ÜÈ™]WÜ™\]Y\ÝÚ›Ø—ÚYÈ—HH\Ý
™\]Y\ÝÖÈš›Ø—ÚYÈ—JBˆ™\ÜÈ™]WÜ™\]Y\ÝÝØ\›š[™ÜÈ—HH\Ý
™\]Y\ÝÖÈØ\›š[™ÜÈ—JBˆ™\ÜœÙ]Y˜][
œ™\ÜÛ›Ý\È‹×JK™^[™
™\]Y\ÝÖÈØ\›š[™ÜÈ—JBˆYˆ[˜ÛYWÝØ]‚ˆYØÝ\ÝÛY\—Ü™\ÜØ]˜Z[Xš[]JˆÛÛ›‹ˆ™\Üˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆ\š[Ù\\š[Ùˆ[œÝ\™WÜ™\]Y\ÝÏQ˜[ÙKˆ
Bˆ™]\›ˆ™\Ü‚‚™Yˆ™YÚ\Ý\—Ü™[™\™YÙÙ[™\˜][Û—Ùš[JÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹Ý]]Ù\Žˆ][—ÚYˆ[š[K
‹Û˜\ÚÝÚYˆ[›Û™HH›Û™JN‚ˆ]ÈHÝÜ™WÜ™[™\™YÙš[JÝ]]Ù\‹[—ÚYš[JBˆYÙÙ[™\˜]YÙš[JˆÛÛ›‹ˆ[—ÚY\[—ÚYˆ›]Yš[K™›]ˆš[[˜[YO\]›˜[YKˆ™[]]™WÜ]\ÝÜ™WÜ[[YWÜ™[]]™WÜ]
]
KˆÚLMYš[KœÚLM‹ˆÚ^™WØž]\ÏYš[KœÚ^™WØž]\ËˆÜ›Û[×ÚYYš[KœÜ›Û[×ÚYˆ\ÜÙ]ÚYYš[K˜\ÜÙ]ÚYˆÛ˜\ÚÝÚY\Û˜\ÚÝÚYÜˆš[KœÛ˜\ÚÝÚYˆ\š[ÙÝ\OYš[Kœ\š[ÙÝ\Kˆ\š[ÙÜÝ\Yš[Kœ\š[ÙÜÝ\ˆ\š[ÙÙ[™Yš[Kœ\š[ÙÙ[™ˆ\×Ø]^[X\žOLHYˆš[Kš\×Ø]^[X\žH[ÙHˆØ\›š[™ÜÏ[\Ý
š[KØ\›š[™ÜÊKˆ
Bˆ™]\›ˆš[B‚‚™YˆYÙ˜Z[YÙÙ[™\˜][Û—Ùš[JÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹[—ÚYˆ[›ØŽˆXÝÜÝ‹[žWK\œ›ÜŽˆÝ‹
‹Ü›Û[×ÚYˆ[›Û™HH›Û™K\ÜÙ]ÚYˆ[›Û™HH›Û™KÛ˜\ÚÝÚYˆ[›Û™HH›Û™JHOˆ›Û™N‚ˆ\š[ÙH›Ø‹™Ù]
œ\š[ÙŠHÜˆßBˆYÙÙ[™\˜]YÙš[JˆÛÛ›‹ˆ[—ÚY\[—ÚYˆ›]\ÝŠ›Ø‹™Ù]
™›Ü›X]ŠHÜˆˆŠKˆš[[˜[YOH™˜Z[Y‹ˆ™[]]™WÜ]Hˆ‹ˆÚLMHˆ‹ˆÚ^™WØž]\ÏLˆÜ›Û[×ÚY\Ü›Û[×ÚYˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆÛ˜\ÚÝÚY\Û˜\ÚÝÚYˆ\š[ÙÝ\O\\š[Ù™Ù]
œ\š[ÙÝ\H‹ˆŠKˆ\š[ÙÜÝ\\\š[Ù™Ù]
œ\š[ÙÜÝ\‹ˆŠKˆ\š[ÙÙ[™\\š[Ù™Ù]
œ\š[ÙÙ[™‹ˆŠKˆÝ]\ÏH™˜Z[Y‹ˆ\œ›Ü—ÛY\ÜØYÙOY\œ›Ü–ÎLKˆ
B‚‚™Yˆ^ÜÜ›ÝÜ×Ùš[Jˆ›ÝÜÎˆ\ÝÙXÝÜÝ‹[žWWKˆXY\œÎˆ\ÝÝ\VÜÝ‹Ý—WKˆš[[˜[YNˆÝ‹ˆ^ÜÙ›Ü›X]ˆÝ‹ŠN‚ˆYˆ^ÜÙ›Ü›X]OHœˆŽ‚ˆY™™\ˆH[Ëž]\ÒSÊ
BˆØÈHÚ[\QØÕ[\]JY™™\‹YÙ\Ú^™O[[™ØØ\JM
KYX\™Ú[LšYÚX\™Ú[LÜX\™Ú[L›ÝÛSX\™Ú[L
BˆÝ[\ÈHÙ]Ø[\TÝ[TÚY]

Bˆ]HHÖÚXY\–ÌWH›ÜˆXY\ˆ[ˆXY\œ×WBˆ›Üˆ›ÝÈ[ˆ›ÝÜÎ‚ˆ]K˜\[™
ÜÝŠ›ÝË™Ù]
XY\–ÌKˆŠHÜˆ‹HŠH›ÜˆXY\ˆ[ˆXY\œ×JBˆX›HHX›J]K™\X]›ÝÜÏLJBˆX›KœÙ]Ý[JˆX›TÝ[JˆÂˆ
PÒÑÔ“ÕS‘‹

K
LK
KÛÛÜœË’^ÛÛÜŠˆÌY˜XÈŠJKˆ
•VÓÓÔˆ‹

K
LK
KÛÛÜœËÚ]JKˆ
‘Ô’Q‹

K
LKLJKKÛÛÜœË’^ÛÛÜŠˆØÎYÙÈŠJKˆ
‘“Ó•SQH‹

K
LK
K’[™]XØKP›ÛŠKˆ
”“ÕÐPÒÑÔ“ÕS‘È‹
JK
LKLJKØÛÛÜœËÚ]\Û[ÚÙKÛÛÜœË’^ÛÛÜŠˆÙYYŒˆŠWJKˆ
‘“Ó•ÒV‘H‹

K
LKLJK
KˆBˆ
Bˆ
BˆØË˜Z[
Ô\˜YÜ˜\
š[[˜[YKÝ[\ÖÈ’XY[™Ìˆ—JKÜXÙ\ŠKLŠKX›WJBˆY™™\‹œÙYZÊ
Bˆ™]\›ˆÙ[™Ùš[JY™™\‹\×Ø]XÚY[UYKÝÛ›ØYÛ˜[YOYˆžÙš[[˜[Y_Kœˆ‹Z[Y]\OH˜\XØ][Û‹ÜˆŠB‚ˆÛÜšØ›ÛÚÈHÛÜšØ›ÛÚÊ
BˆÛÜšÜÚY]HÛÜšØ›ÛÚË˜XÝ]™BˆÛÜšÜÚY]]HH‘^Ü‚ˆÛÜšÜÚY]˜\[™
ÚXY\–ÌWH›ÜˆXY\ˆ[ˆXY\œ×JBˆ›Üˆ›ÝÈ[ˆ›ÝÜÎ‚ˆÛÜšÜÚY]˜\[™
Ü›ÝË™Ù]
XY\–ÌKˆŠH›ÜˆXY\ˆ[ˆXY\œ×JBˆÝ]]H[Ëž]\ÒSÊ
BˆÛÜšØ›ÛÚËœØ]™JÝ]]
BˆÝ]]œÙYZÊ
Bˆ™]\›ˆÙ[™Ùš[JˆÝ]]ˆ\×Ø]XÚY[UYKˆÝÛ›ØYÛ˜[YOYˆžÙš[[˜[Y_KžÞ‹ˆZ[Y]\OH˜\XØ][Û‹Ý›™›Ü[ž[›Ü›X]Ë[Ù™šXÙYØÝ[Y[œÜ™XYÚY][œÚY]‹ˆ
B‚‚™YˆÙ›Ü›WÚ[Û\Ý
šY[Û˜[YNˆÝŠHOˆ\ÝÚ[N‚ˆ˜[Y\ÈH™\]Y\Ý™›Ü›K™Ù]\Ý
šY[Û˜[YJBˆYˆ[Š˜[Y\ÊHOHH[™‹ˆ[ˆ˜[Y\ÖÌN‚ˆ˜[Y\ÈHÚ][KœÝš\

H›Üˆ][H[ˆ˜[Y\ÖÌKœÜ]
‹ŠWBˆ™]\›ˆÚ[
˜[YJH›Üˆ˜[YH[ˆ˜[Y\ÈYˆÝŠ˜[YJKœÝš\

Kš\ÙYÚ]

WB‚‚™YˆØ[X\×Ü™]\›Š\ÜÙ]ÚYˆ[
HOˆ[žN‚ˆ™^Ý\›H™\]Y\Ý™›Ü›K™Ù]
›™^‹ˆŠKœÝš\

BˆYˆ™^Ý\›œÝ\ÝÚ]
‹ÈŠN‚ˆ™]\›ˆ™Y\™XÝ
™^Ý\›
BˆYˆ™\]Y\Ýœ™Y™\œ™\ˆ[™‹ÜÜ›Û[Ë[X[˜YÙ\ˆˆ[ˆ™\]Y\Ýœ™Y™\œ™\Ž‚ˆ™]\›ˆ™Y\™XÝ
™\]Y\Ýœ™Y™\œ™\ŠBˆ™]\›ˆ™Y\™XÝ
\›Ù›ÜŠ˜\ÜÙ]Ù]Z[‹\ÜÙ]ÚYX\ÜÙ]ÚY
JB‚‚™YˆÝÜ™WÚ[›ÚXÙWÝ\ØY
ˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆ\ØYÙ\Žˆ]ˆš[WÜÝÜ˜YÙNˆ[žKˆ\ÜÙ]ÚYˆ[ˆÜ›Û[×ÚYˆ[›Û™KŠHOˆ\VÚ[\VÜÝ‹‹‹—WN‚ˆYˆÛÛ›‹™^XÝ]J”ÑSPÕY”“ÓH\ÜÙ]ÈÒT‘HYHÈ‹
\ÜÙ]ÚY
JK™™]ÚÛ™J
H\È›Û™N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ’[œÝ[XØ[È[™^\Ý[KˆŠBˆÜšYÚ[˜[Ùš[[˜[YHH]
š[WÜÝÜ˜YÙK™š[[˜[YHÜˆš[›ÚXÙHŠK›˜[YBˆØY™WÛ˜[YHHÙXÝ\™WÙš[[˜[YJÜšYÚ[˜[Ùš[[˜[YJHÜˆš[›ÚXÙH‚ˆYˆ›Ý\×ÜÝ\ÜYÚ[›ÚXÙWÙ^[œÚ[ÛŠØY™WÛ˜[YJN‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ‘›Ü›X]ÈH˜]\˜H˜[ÈÝ\ÜYËˆŠBˆ\™Ù]Ù\ˆH\ØYÙ\ˆÈœÜ›Û[×ÜÛÝ\˜Ù\ÈˆÈÝŠ\ÜÙ]ÚY
Bˆ\™Ù]Ù\‹›ZÙ\Š\™[ÏUYK^\ÝÛÚÏUYJBˆÝ[\H]][YK››ÝÊ
KœÝ™[YJ‰VI[IYÉR	SIT×ÉYˆŠBˆ\™Ù]H\™Ù]Ù\ˆÈˆžÜÝ[\WÞÜØY™WÛ˜[Y_H‚ˆØY™WÜ]H]
ØY™WÛ˜[YJBˆ[\Ý\™Ù]H\™Ù]Ù\ˆÈˆ‹žÜÝ[\WÞÜØY™WÜ]œÝ[_K\ÜØY™WÜ]œÝY™š^H‚ˆš[˜[ØÜ™X]YH˜[ÙBˆžN‚ˆš[WÜÝÜ˜YÙKœØ]™J[\Ý\™Ù]
BˆÚ^™WØž]\ÈH[\Ý\™Ù]œÝ]

KœÝÜÚ^™BˆYˆÚ^™WØž]\ÈH‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ‘šXÚZ\›È˜^š[ËˆŠBˆYˆÚ^™WØž]\ÈˆX^Ý\ØYØž]\Ê
N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ‘šXÚZ\›È^ÙYHÈ[Z]HH\ØYˆŠBˆ˜[Y]WÚ[›ÚXÙWÙš[WØÛÛ[
[\Ý\™Ù]
Bˆ^Ù\^Ù\[ÛŽ‚ˆ[\Ý\™Ù][›[šÊZ\ÜÚ[™×ÛÚÏUYJBˆ˜Z\ÙBˆžN‚ˆYÙ\ÝHÚLM—Ùš[J[\Ý\™Ù]
BˆØ[YWØ\ÜÙ]Hš[™Ú[›ÚXÙWØžWÚ\Ú
ÛÛ›‹\ÜÙ]ÚYX\ÜÙ]ÚYÚLMYYÙ\Ý
BˆYˆØ[YWØ\ÜÙ]\È›Ý›Û™N‚ˆ[\Ý\™Ù][›[šÊZ\ÜÚ[™×ÛÚÏUYJBˆ™]\›ˆ[
Ø[YWØ\ÜÙ]ÈšY—JK
™\XØ]WÚ[›ÚXÙH‹
BˆØ\›š[™ÜÎˆ\ÝÜÝ—HH×BˆYˆš[™Ú[›ÚXÙWØžWÚ\Ú
ÛÛ›‹\ÜÙ]ÚYS›Û™KÚLMYYÙ\Ý
H\È›Ý›Û™N‚ˆØ\›š[™ÜË˜\[™
œÜÜÚX›WÙ\XØ]WÚ[›ÚXÙHŠBˆ[\Ý\™Ù]œ™\XÙJ\™Ù]
Bˆš[˜[ØÜ™X]YHYBˆZ[YWÝ\HHZ[Y]\\Ë™ÝY\Ü×Ý\JØY™WÛ˜[YJVÌHÜˆÝŠÙ]]Šš[WÜÝÜ˜YÙK›Z[Y]\H‹ˆŠHÜˆˆŠBˆÛÝ\˜ÙWÚYHÜ™X]WÜÛÝ\˜ÙWÙš[WÜ™XÛÜ™
ˆÛÛ›‹ˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆÜ›Û[×ÚY\Ü›Û[×ÚYˆš[WÝ\OHš[›ÚXÙH‹ˆÜšYÚ[˜[Ùš[[˜[YO[ÜšYÚ[˜[Ùš[[˜[YKˆÝÜ™YÜ]\ÝŠ\™Ù]
KˆÚLMYYÙ\ÝˆZ[YWÝ\O[Z[YWÝ\KˆÚ^™WØž]\Ï\Ú^™WØž]\Ëˆ
BˆØÝ[Y[ÚYHÜ™X]WÚ[›ÚXÙWÙØÝ[Y[
ˆÛÛ›‹ˆÛÝ\˜ÙWÙš[WÚY\ÛÝ\˜ÙWÚYˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆÚLMYYÙ\ÝˆZ[YWÝ\O[Z[YWÝ\KˆÚ^™WØž]\Ï\Ú^™WØž]\ËˆÝ]\ÏR[›ÚXÙTÝ]\Ë”‘U’QU×Ô‘TURT‘Q˜[YHYˆØ\›š[™ÜÈ[ÙH[›ÚXÙTÝ]\Ë•TÐQQ˜[YKˆØ\›š[™ÜÏ]\JØ\›š[™ÜÊKˆ
Bˆ™]\›ˆØÝ[Y[ÚY\JØ\›š[™ÜÊBˆ^Ù\^Ù\[ÛŽ‚ˆ[\Ý\™Ù][›[šÊZ\ÜÚ[™×ÛÚÏUYJBˆYˆš[˜[ØÜ™X]Y‚ˆ\™Ù][›[šÊZ\ÜÚ[™×ÛÚÏUYJBˆ˜Z\ÙB‚‚™Yˆ\œÚ\ÝÚ[›ÚXÙWÙ^˜XÝ[Û—Ü™\Ý[
ÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹[›ÚXÙNˆÜ[]LË”›ÝË™\Ý[ˆ[žJHOˆ›Û™N‚ˆÜ™X]WÚ[›ÚXÙWÙ^˜XÝ[Û—Ü[ŠÛÛ›‹[›ÚXÙWÙØÝ[Y[ÚYZ[
[›ÚXÙVÈšY—JK™\Ý[\™\Ý[
BˆØ[™Y]\ÈH\Ý
™\Ý[˜Ø[™Y]\ÊBˆYˆ™\Ý[\šY™—ØØ[™Y]K\šY™—Ý\H\È›Ý›Û™N‚ˆØ[™Y]\Ë˜\[™
ˆ[›ÚXÙPØ[™Y]Jˆ\šY™—Ý\WØØ[™Y]H‹ˆ™\Ý[\šY™—ØØ[™Y]K\šY™—Ý\K˜[YKˆ™\Ý[˜ÛÛ™šY[˜ÙKˆ]šY[˜ÙOHš[™™\šYÈÜÈ™XÛÜÈ^˜ZYÜÈ‹ˆÛÝ\˜ÙO\™\Ý[›Y]Ùˆ
Bˆ
Bˆ˜[Y\ÈHØØ[™Y]K™šY[Û˜[YNˆØ[™Y]K˜[YH›ÜˆØ[™Y]H[ˆØ[™Y]\ßBˆ\ÜÙ]HÛÛ›‹™^XÝ]J”ÑSPÕšYˆ”“ÓH\ÜÙ]ÈÒT‘HYHÈ‹
[›ÚXÙVÈ˜\ÜÙ]ÚY—K
JK™™]ÚÛ™J
Bˆ˜[Y][ÛˆH˜[Y]WÚ[›ÚXÙWÝ˜[Y\Ê˜[Y\Ë\ÜÙ]ÛšYX\ÜÙ]È›šYˆ—HYˆ\ÜÙ][ÙH›Û™JBˆØ\›š[™ÜÈH\JÛÜY
Êœ™\Ý[Ø\›š[™ÜË
˜[Y][Û‹Ø\›š[™ÜßJJBˆÝ]\ÈH[›ÚXÙTÝ]\Ë‘VPÕSÓ—ÑRSQ˜[YHYˆ™\Ý[™\œ›ÜœÈÜˆ˜[Y][Û‹™\œ›ÜœÈ[ÙH˜[Y][Û‹œÝ]\Ë˜[YBˆ\]WÚ[›ÚXÙWÙœ›ÛWØØ[™Y]\ÊˆÛÛ›‹ˆ[›ÚXÙWÙØÝ[Y[ÚYZ[
[›ÚXÙVÈšY—JKˆØ[™Y]\Ï]\JØ[™Y]\ÊKˆÝ]\Ï\Ý]\ËˆÛÛ™šY[˜ÙO\™\Ý[˜ÛÛ™šY[˜ÙKˆØ\›š[™ÜÏ]Ø\›š[™ÜËˆ
B‚‚™Yˆ[›ÚXÙWÝ˜[Y\×Ùœ›ÛWÙ›Ü›J›Ü›Nˆ[žJHOˆXÝÜÝ‹Ý—N‚ˆšY[ÈH
ˆœÝ\Y\—Û˜[YH‹ˆœÝ\Y\—ÛšYˆ‹ˆ˜Ý\ÝÛY\—Û˜[YH‹ˆ˜Ý\ÝÛY\—ÛšYˆ‹ˆš[›ÚXÙWÛ[X™\ˆ‹ˆš\ÜÝYWÙ]H‹ˆ˜š[[™×Ü\š[ÙÜÝ\‹ˆ˜š[[™×Ü\š[ÙÙ[™‹ˆ˜Ý\œ™[˜ÞH‹ˆÝ[Ø[[Ý[‹ˆÝ[Ù[™\™ÞWÚÝÚ‹ˆ\šY™—Ý\WØØ[™Y]H‹ˆœÚ[\WÜšXÙWÙ]\—ÚÝÚ‹ˆœÛWÜšXÙWÙ]\—ÚÝÚ‹ˆ˜ÚZXWÜšXÙWÙ]\—ÚÝÚ‹ˆ˜^š[×ÜšXÙWÙ]\—ÚÝÚ‹ˆœÝ\\—Ý˜^š[×ÜšXÙWÙ]\—ÚÝÚ‹ˆ
Bˆ™]\›ˆÙšY[ˆÝŠ›Ü›K™Ù]
šY[ˆŠJKœÝš\

H›ÜˆšY[[ˆšY[ßB‚‚™Yˆ\WÚ[›ÚXÙWÝ×Ý\šY™ŠÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹[›ÚXÙWÙØÝ[Y[ÚYˆ[
HOˆ[‚ˆ[›ÚXÙHHÙ]Ú[›ÚXÙWÙØÝ[Y[
ÛÛ›‹[›ÚXÙWÙØÝ[Y[ÚY
BˆYˆ[›ÚXÙH\È›Û™N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ‘˜]\˜H[™^\Ý[KˆŠBˆYˆ[›ÚXÙVÈœÝ]\È—HOH[›ÚXÙTÝ]\ËÓÓ‘’T“QQ˜[YN‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠH˜]\˜H[HH\Ý\ˆÛÛ™š\›XYH[\ÈHÙ\ˆ\ØYH˜H\šY˜KˆŠBˆ\šY™—Ý\HHÝŠ[›ÚXÙVÈ\šY™—Ý\WØØ[™Y]H—HÜˆœÚ[\HŠBˆYˆ\šY™—Ý\HOH\šY™•\K”ÒSTK˜[YN‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ•\šY˜\È][KZÜ˜\šX\È^YÙ[H™YÜ˜\ÈÜ˜\šX\ÈX[XZ\È[\ÈHÝX\™\‹ˆŠBˆ™]\›ˆØ]™WØ\ÜÙ]Ý\šY™ŠˆÛÛ›‹ˆ\ÜÙ]ÚYZ[
[›ÚXÙVÈ˜\ÜÙ]ÚY—JKˆ\šY™—Ý\O]\šY™—Ý\KˆÚ[\WÜšXÙWÙ]\—ÚÝÚZ[›ÚXÙVÈœÚ[\WÜšXÙWÙ]\—ÚÝÚ—Kˆ[›ÚXÙWÙš[WÚYZ[
[›ÚXÙVÈœÛÝ\˜ÙWÙš[WÚY—JKˆ˜[YÙœ›ÛO[›Ü›X[^™WÙ]J[›ÚXÙVÈ˜š[[™×Ü\š[ÙÜÝ\—JKš\ÛÙ›Ü›X]

HYˆ[›ÚXÙVÈ˜š[[™×Ü\š[ÙÜÝ\—H[ÙHˆ‹ˆ˜[YÝÏ[›Ü›X[^™WÙ]J[›ÚXÙVÈ˜š[[™×Ü\š[ÙÙ[™—JKš\ÛÙ›Ü›X]

HYˆ[›ÚXÙVÈ˜š[[™×Ü\š[ÙÙ[™—H[ÙHˆ‹ˆ›Ý\ÏYˆÜšXYHH\\ˆH˜]\˜HÚ[›ÚXÙVÉÚ[›ÚXÙWÛ[X™\‰×HÜˆ[›ÚXÙWÙØÝ[Y[ÚYH‹ˆ
B‚‚™YˆZ[Ýš\Ú]×ØžWÝXÚÙ]
š\Ú]Îˆ\ÝÜÜ[]LË”›Ý×JHOˆXÝÚ[\ÝÜÜ[]LË”›Ý×WN‚ˆš\Ú]×ØžWÝXÚÙ]ˆXÝÚ[\ÝÜÜ[]LË”›Ý×WHHßBˆ›Üˆš\Ú][ˆš\Ú]Î‚ˆš\Ú]×ØžWÝXÚÙ]œÙ]Y˜][
š\Ú]ÈXÚÙ]ÚY—K×JK˜\[™
š\Ú]
Bˆ™]\›ˆš\Ú]×ØžWÝXÚÙ]‚‚™Yˆ›Ü›X[^™WØØ[[™\—Û[Û
˜[YNˆÝˆ›Û™JHOˆÝŽ‚ˆYˆ˜[YN‚ˆžN‚ˆ™]\›ˆ]][YKœÝœ[YJ˜[YKœÝš\

K‰VKI[HŠKœÝ™[YJ‰VKI[HŠBˆ^Ù\˜[YQ\œ›ÜŽ‚ˆ\ÜÂˆ™]\›ˆ]KÙ^J
KœÝ™[YJ‰VKI[HŠB‚‚™Yˆ›Ü›X[^™WÛÜ[Û˜[Ù]J˜[YNˆÝˆ›Û™JHOˆÝŽ‚ˆYˆ›Ý˜[YN‚ˆ™]\›ˆˆ‚ˆžN‚ˆ™]\›ˆ]][YKœÝœ[YJ˜[YKœÝš\

K‰VKI[KIYŠK™]J
Kš\ÛÙ›Ü›X]

Bˆ^Ù\˜[YQ\œ›ÜŽ‚ˆ™]\›ˆˆ‚‚‚™Yˆ\œÙWÜÜÚ]]™WÚ[
˜[YNˆÝˆ›Û™KY˜][ˆ[H
HOˆ[‚ˆžN‚ˆ\œÙYH[
›Ø]
˜[YHÜˆˆŠJBˆ^Ù\
\Q\œ›Ü‹˜[YQ\œ›ÜŠN‚ˆ™]\›ˆY˜][ˆ™]\›ˆX^
\œÙY
B‚‚™Yˆ›Ü›X[^™WØÚÚXÙJ˜[YNˆÝˆ›Û™KÚÚXÙ\Îˆ\ÝÜÝ—KY˜][ˆÝŠHOˆÝŽ‚ˆ˜[YHH
˜[YHÜˆˆŠKœÝš\

Bˆ™]\›ˆ˜[YHYˆ˜[YH[ˆÚÚXÙ\È[ÙHY˜][‚‚™YˆØ[[™\—Û[ÛØ›Ý[™Ê[ÛÝ˜[YNˆÝŠHOˆ\VÙ]K]KÝ‹Ý—N‚ˆ[ÛÜÝ\H]][YKœÝœ[YJ[ÛÝ˜[YK‰VKI[HŠK™]J
Kœ™\XÙJ^OLJBˆË\ÝÙ^HHØ[[™\‹›[Û˜[™ÙJ[ÛÜÝ\žYX\‹[ÛÜÝ\›[Û
Bˆ[ÛÙ[™H[ÛÜÝ\œ™\XÙJ^O[\ÝÙ^JBˆ™]š[Ý\×Û[ÛÙ]HH
[ÛÜÝ\H[YY[J^\ÏLJJKœ™\XÙJ^OLJBˆ™^Û[ÛÙ]HH
[ÛÙ[™
È[YY[J^\ÏLJJKœ™\XÙJ^OLJBˆ™]\›ˆ[ÛÜÝ\[ÛÙ[™™]š[Ý\×Û[ÛÙ]KœÝ™[YJ‰VKI[HŠK™^Û[ÛÙ]KœÝ™[YJ‰VKI[HŠB‚‚™YˆZ[Ù\œ›Ü—ØØ[[™\Š[ÛÝ˜[YNˆÝ‹™XÛÜ™Îˆ\ÝÜÜ[]LË”›Ý×JHOˆXÝÜÝ‹[žWN‚ˆ[ÛÜÝ\[ÛÙ[™ËÈHØ[[™\—Û[ÛØ›Ý[™Ê[ÛÝ˜[YJBˆ™XÛÜ™×ØžWÙ^NˆXÝÜÝ‹\ÝÜÜ[]LË”›Ý×WHHßBˆ›Üˆ™XÛÜ™[ˆ™XÛÜ™Î‚ˆ™XÛÜ™×ØžWÙ^KœÙ]Y˜][
™XÛÜ™Èœ™XÛÜ™Ù]H—K×JK˜\[™
™XÛÜ™
B‚ˆÙYZÜÈH×BˆÙYZÈH×Bˆ›ÜˆÈ[ˆ˜[™ÙJ[ÛÜÝ\ÙYZÙ^J
JN‚ˆÙYZË˜\[™
È™]HŽˆ›Û™Kœ™XÛÜ™ÈŽˆ×_JB‚ˆÝ\œ™[Ù^HH[ÛÜÝ\ˆÚ[HÝ\œ™[Ù^HH[ÛÙ[™‚ˆ\Û×Ù^HHÝ\œ™[Ù^Kš\ÛÙ›Ü›X]

BˆÙYZË˜\[™
È™]HŽˆÝ\œ™[Ù^Kœ™XÛÜ™ÈŽˆ™XÛÜ™×ØžWÙ^K™Ù]
\Û×Ù^K×J_JBˆYˆ[ŠÙYZÊHOHÎ‚ˆÙYZÜË˜\[™
ÙYZÊBˆÙYZÈH×BˆÝ\œ™[Ù^H
ÏH[YY[J^\ÏLJB‚ˆYˆÙYZÎ‚ˆÚ[H[ŠÙYZÊHÎ‚ˆÙYZË˜\[™
È™]HŽˆ›Û™Kœ™XÛÜ™ÈŽˆ×_JBˆÙYZÜË˜\[™
ÙYZÊB‚ˆ™]\›ˆÂˆ›X™[ŽˆˆžÓSÓ•ÓSQT×ÔÛ[ÛÜÝ\›[Û_HÛ[ÛÜÝ\žYX\ŸH‹ˆÙYZÜÈŽˆÙYZÜËˆœ™XÛÜ™ØÛÝ[ŽˆÝ[J[Š›ÝÜÊH›Üˆ›ÝÜÈ[ˆ™XÛÜ™×ØžWÙ^K˜[Y\Ê
JKˆB‚‚™YˆZ[Ú[\™[[Û—ØØ[[™\Š[ÛÝ˜[YNˆÝ‹™XÛÜ™Îˆ\ÝÜÜ[]LË”›Ý×JHOˆXÝÜÝ‹[žWN‚ˆ[ÛÜÝ\[ÛÙ[™ËÈHØ[[™\—Û[ÛØ›Ý[™Ê[ÛÝ˜[YJBˆ™XÛÜ™×ØžWÙ^NˆXÝÜÝ‹\ÝÜÜ[]LË”›Ý×WHHßBˆ›Üˆ™XÛÜ™[ˆ™XÛÜ™Î‚ˆ[›™YÙ]HH™XÛÜ™Èœ[›™YÙ]H—BˆYˆ[›™YÙ]N‚ˆ™XÛÜ™×ØžWÙ^KœÙ]Y˜][
[›™YÙ]K×JK˜\[™
™XÛÜ™
B‚ˆÙYZÜÈH×BˆÙYZÈH×Bˆ›ÜˆÈ[ˆ˜[™ÙJ[ÛÜÝ\ÙYZÙ^J
JN‚ˆÙYZË˜\[™
È™]HŽˆ›Û™Kœ™XÛÜ™ÈŽˆ×_JB‚ˆÝ\œ™[Ù^HH[ÛÜÝ\ˆÚ[HÝ\œ™[Ù^HH[ÛÙ[™‚ˆ\Û×Ù^HHÝ\œ™[Ù^Kš\ÛÙ›Ü›X]

BˆÙYZË˜\[™
È™]HŽˆÝ\œ™[Ù^Kœ™XÛÜ™ÈŽˆ™XÛÜ™×ØžWÙ^K™Ù]
\Û×Ù^K×J_JBˆYˆ[ŠÙYZÊHOHÎ‚ˆÙYZÜË˜\[™
ÙYZÊBˆÙYZÈH×BˆÝ\œ™[Ù^H
ÏH[YY[J^\ÏLJB‚ˆYˆÙYZÎ‚ˆÚ[H[ŠÙYZÊHÎ‚ˆÙYZË˜\[™
È™]HŽˆ›Û™Kœ™XÛÜ™ÈŽˆ×_JBˆÙYZÜË˜\[™
ÙYZÊB‚ˆ™]\›ˆÂˆ›X™[ŽˆˆžÓSÓ•ÓSQT×ÔÛ[ÛÜÝ\›[Û_HÛ[ÛÜÝ\žYX\ŸH‹ˆÙYZÜÈŽˆÙYZÜËˆœ™XÛÜ™ØÛÝ[ŽˆÝ[J[Š›ÝÜÊH›Üˆ›ÝÜÈ[ˆ™XÛÜ™×ØžWÙ^K˜[Y\Ê
JKˆB‚‚™Yˆ[\™[[Û—Ü™XYWÙ›Ü—Ü›Ý]J›ÝÎˆÜ[]LË”›ÝÈXÝÜÝ‹[žWJHOˆ›ÛÛ‚ˆYˆ›ÝÖÈœÝ]\È—HOH‘™XÚYÈŽ‚ˆ™]\›ˆ˜[ÙBˆYˆ›ÝÖÈ›X]\šX[ÜÝ]\È—HOH›Ü]YXYÈŽ‚ˆ™]\›ˆ˜[ÙBˆYˆ›ÝÖÈ›]]YH—H\È›Û™HÜˆ›ÝÖÈ›Û™Ú]YH—H\È›Û™N‚ˆ™]\›ˆ˜[ÙBˆYˆ›ÝÖÈ˜ÛÛÜ™[˜]\×ØÛÛ™šY[˜ÙH—H[ˆÈœÝ\ÜXÝ‹œ™]šY]ÈŸN‚ˆ™]\›ˆ˜[ÙBˆ™]\›ˆYB‚‚™YˆZ[Ø\ÜÙ]Ù\œ›Ü—ØØ[[™\Š[ÛÝ˜[YNˆÝ‹™XÛÜ™Îˆ\ÝÜÜ[]LË”›Ý×JHOˆXÝÜÝ‹[žWN‚ˆ[ÛÜÝ\[ÛÙ[™ËÈHØ[[™\—Û[ÛØ›Ý[™Ê[ÛÝ˜[YJBˆ]™[×ØžWÙ^NˆXÝÜÝ‹\ÝÙXÝÜÝ‹[žWWWHHßBˆ™]š[Ý\×Ü›Ø›[HH˜[ÙB‚ˆ›Üˆ™XÛÜ™[ˆ™XÛÜ™Î‚ˆ™XÛÜ™Ù]HH™XÛÜ™Èœ™XÛÜ™Ù]H—BˆÝ]\ÈH™XÛÜ™ÈœÝ]\È—Bˆ\×Ü›Ø›[HHÝ]\È[ˆ“Ð“SWÓSÓ’UÔ’S‘×ÔÕUTÑTÂˆ]™[Ý\HHˆ‚ˆ]™[ÛX™[Hˆ‚‚ˆYˆ\×Ü›Ø›[H[™›Ý™]š[Ý\×Ü›Ø›[N‚ˆ]™[Ý\HHœÝ\‚ˆ]™[ÛX™[H\\™XÙ]H‚ˆ[Yˆ\×Ü›Ø›[N‚ˆ]™[Ý\HH˜XÝ]™H‚ˆ]™[ÛX™[H“X[[K\ÙH‚ˆ[Yˆ™]š[Ý\×Ü›Ø›[N‚ˆ]™[Ý\HH™[™‚ˆ]™[ÛX™[H‘\Ø\\™XÙ]H‚‚ˆYˆ[ÛÜÝ\š\ÛÙ›Ü›X]

HH™XÛÜ™Ù]HH[ÛÙ[™š\ÛÙ›Ü›X]

H[™]™[Ý\N‚ˆ]™[×ØžWÙ^KœÙ]Y˜][
™XÛÜ™Ù]K×JK˜\[™
ˆÂˆœÝ]\ÈŽˆÝ]\Ëˆ›X™[Žˆ]™[ÛX™[ˆ\HŽˆ]™[Ý\Kˆ››Ý\ÈŽˆ™XÛÜ™È››Ý\È—KˆœÛÝ\˜ÙHŽˆ™XÛÜ™ÈœÛÝ\˜ÙH—Kˆœ™XÛÜ™ÚYŽˆ™XÛÜ™ÈšY—KˆBˆ
B‚ˆ™]š[Ý\×Ü›Ø›[HH\×Ü›Ø›[B‚ˆÙYZÜÈH×BˆÙYZÈH×Bˆ›ÜˆÈ[ˆ˜[™ÙJ[ÛÜÝ\ÙYZÙ^J
JN‚ˆÙYZË˜\[™
È™]HŽˆ›Û™K™]™[ÈŽˆ×_JB‚ˆÝ\œ™[Ù^HH[ÛÜÝ\ˆÚ[HÝ\œ™[Ù^HH[ÛÙ[™‚ˆ\Û×Ù^HHÝ\œ™[Ù^Kš\ÛÙ›Ü›X]

BˆÙYZË˜\[™
È™]HŽˆÝ\œ™[Ù^K™]™[ÈŽˆ]™[×ØžWÙ^K™Ù]
\Û×Ù^K×J_JBˆYˆ[ŠÙYZÊHOHÎ‚ˆÙYZÜË˜\[™
ÙYZÊBˆÙYZÈH×BˆÝ\œ™[Ù^H
ÏH[YY[J^\ÏLJB‚ˆYˆÙYZÎ‚ˆÚ[H[ŠÙYZÊHÎ‚ˆÙYZË˜\[™
È™]HŽˆ›Û™K™]™[ÈŽˆ×_JBˆÙYZÜË˜\[™
ÙYZÊB‚ˆ™]\›ˆÂˆ›X™[ŽˆˆžÓSÓ•ÓSQT×ÔÛ[ÛÜÝ\›[Û_HÛ[ÛÜÝ\žYX\ŸH‹ˆÙYZÜÈŽˆÙYZÜËˆ™]™[ØÛÝ[ŽˆÝ[J[Š›ÝÜÊH›Üˆ›ÝÜÈ[ˆ]™[×ØžWÙ^K˜[Y\Ê
JKˆB‚‚™YˆÜ›Ý\ÝXÚÙ]×ØžWØ\ÜÙ]
XÚÙ]Îˆ\ÝÜÜ[]LË”›Ý×JHOˆ\ÝÙXÝÜÝ‹[žWWN‚ˆÜ›Ý\YˆXÝÚ[XÝÜÝ‹[žWWHHßBˆ›ÜˆXÚÙ][ˆXÚÙ]Î‚ˆ\ÜÙ]ÚYH[
XÚÙ]È˜\ÜÙ]ÚY—JBˆXÚÙ]HÜ›Ý\YœÙ]Y˜][
ˆ\ÜÙ]ÚYˆÂˆ˜\ÜÙ]ÚYŽˆ\ÜÙ]ÚYˆœ›Ú™XÝÛ˜[YHŽˆXÚÙ]Èœ›Ú™XÝÛ˜[YH—Kˆ›ØØ][ÛˆŽˆXÚÙ]È›ØØ][Ûˆ—Kˆ˜XÝ]™WØÛÛ˜XÝŽˆXÚÙ]È˜XÝ]™WØÛÛ˜XÝ—Kˆ˜ÛÛ˜XÝÝ\HŽˆXÚÙ]È˜ÛÛ˜XÝÝ\H—KˆXÚÙ]ÈŽˆ×KˆKˆ
BˆXÚÙ]ÈXÚÙ]È—K˜\[™
XÚÙ]
B‚ˆÜ™\™YH×Bˆ›Üˆ\ÜÙ]ÚYXÚÙ][ˆÜ›Ý\Yš][\Ê
N‚ˆXÚÙ]×Û\ÝHXÚÙ]ÈXÚÙ]È—BˆXÚÙ]È›Ü[—ØÛÝ[—HHÝ[JH›ÜˆXÚÙ][ˆXÚÙ]×Û\ÝYˆXÚÙ]ÈœÝ]\È—HOH‘™XÚYÈŠBˆXÚÙ]È˜Üš]XØ[ØÛÝ[—HHÝ[JˆH›ÜˆXÚÙ][ˆXÚÙ]×Û\ÝYˆXÚÙ]È\™Ù[˜ÞH—HOHÜš]XØHˆ[™XÚÙ]ÈœÝ]\È—HOH‘™XÚYÈ‚ˆ
BˆXÚÙ]È›\ÝÝ\]H—HHX^
XÚÙ]È\]YØ]—H›ÜˆXÚÙ][ˆXÚÙ]×Û\Ý
BˆÜ™\™Y˜\[™
XÚÙ]
B‚ˆÜ™\™YœÛÜ
ˆÙ^O[[X™H][Nˆ
ˆYˆ][VÈ˜XÝ]™WØÛÛ˜XÝ—HOHžY\Èˆ[ÙHKˆZ][VÈ˜Üš]XØ[ØÛÝ[—KˆZ][VÈ›Ü[—ØÛÝ[—Kˆ][VÈœ›Ú™XÝÛ˜[YH—K›ÝÙ\Š
Kˆ
Bˆ
Bˆ™]\›ˆÜ™\™Y‚‚™Yˆ[œÝ\™WÜ™YYš[™YÙ^ÜÝ[\]\ÊÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[ÛŠHOˆ›Û™N‚ˆ™YYš[™YÝ[\]\ÈHÂˆÂˆ›˜[YHŽˆ”™]™[]˜HHÉ“H]]›È‹ˆ™]\Ù]Žˆ˜\ÜÙ]È‹ˆ™^ÜÙ›Ü›X]ŽˆžÞ‹ˆ˜ÛÛ[[œÈŽˆÂˆœ›Ú™XÝÛ˜[YH‹ˆ›ØØ][Ûˆ‹ˆ˜Y™\ÜÈ‹ˆ˜ÛÛXÝÜÛ™H‹ˆ˜ÛÛXÝÛ˜[YH‹ˆ˜XØÙ\Ü×Ý\H‹ˆ˜ÛÝ™\˜YÙWÝ\H‹ˆKˆ™š[\œÈŽˆÂˆ›ÛWÛÛ›HŽˆžY\È‹ˆKˆKˆÂˆ›˜[YHŽˆ”™[]Üš[È[Ûš]Üš^˜XØ[ÈHX\š[È‹ˆ™]\Ù]Žˆ›[Ûš]Üš[™×Ü™\Ü‹ˆ™^ÜÙ›Ü›X]Žˆœˆ‹ˆ˜ÛÛ[[œÈŽˆÂˆœ\š[Ù‹ˆœ›Ú™XÝÛ˜[YH‹ˆ˜Ý\œ™[ÜÝ]\È‹ˆ™\œ›Ü—Ü™XÛÜ™È‹ˆ™\Ý[˜ÝÙ\œ›ÜœÈ‹ˆ™\œ›Ü—Ý\\È‹ˆ›Ü[—ÝXÚÙ]È‹ˆš\Ú]×Ü\š[Ù‹ˆ›\ÝÝš\Ú]Ù]H‹ˆ›]\ÝÛ›Ý\È‹ˆKˆ™š[\œÈŽˆÂˆœ\š[ÙŽˆ™^H‹ˆ›ÛWÛÛ›HŽˆžY\È‹ˆKˆKˆÂˆ›˜[YHŽˆ”™[]Üš[È[Ûš]Üš^˜XØ[ÈHÙ[X[˜[‹ˆ™]\Ù]Žˆ›[Ûš]Üš[™×Ü™\Ü‹ˆ™^ÜÙ›Ü›X]Žˆœˆ‹ˆ˜ÛÛ[[œÈŽˆÂˆœ\š[Ù‹ˆœ›Ú™XÝÛ˜[YH‹ˆ˜Ý\œ™[ÜÝ]\È‹ˆ™\œ›Ü—Ü™XÛÜ™È‹ˆ™\Ý[˜ÝÙ\œ›ÜœÈ‹ˆ™\œ›Ü—Ý\\È‹ˆ›Ü[—ÝXÚÙ]È‹ˆš\Ú]×Ü\š[Ù‹ˆ›\ÝÝš\Ú]Ù]H‹ˆ›]\ÝÛ›Ý\È‹ˆKˆ™š[\œÈŽˆÂˆœ\š[ÙŽˆÙYZÈ‹ˆ›ÛWÛÛ›HŽˆžY\È‹ˆKˆKˆÂˆ›˜[YHŽˆ”™[]Üš[È[Ûš]Üš^˜XØ[ÈHY[œØ[‹ˆ™]\Ù]Žˆ›[Ûš]Üš[™×Ü™\Ü‹ˆ™^ÜÙ›Ü›X]Žˆœˆ‹ˆ˜ÛÛ[[œÈŽˆÂˆœ\š[Ù‹ˆœ›Ú™XÝÛ˜[YH‹ˆ˜Ý\œ™[ÜÝ]\È‹ˆ™\œ›Ü—Ü™XÛÜ™È‹ˆ™\Ý[˜ÝÙ\œ›ÜœÈ‹ˆ™\œ›Ü—Ý\\È‹ˆ›Ü[—ÝXÚÙ]È‹ˆš\Ú]×Ü\š[Ù‹ˆ›\ÝÝš\Ú]Ù]H‹ˆ›]\ÝÛ›Ý\È‹ˆKˆ™š[\œÈŽˆÂˆœ\š[ÙŽˆ›[Û‹ˆ›ÛWÛÛ›HŽˆžY\È‹ˆKˆKˆÂˆ›˜[YHŽˆ”™[]Üš[È›ÙXØ[ÈHY[œØ[‹ˆ™]\Ù]Žˆœ›ÙXÝ[Û—Ü™\Ü‹ˆ™^ÜÙ›Ü›X]ŽˆžÞ‹ˆ˜ÛÛ[[œÈŽˆÂˆœ\š[Ù‹ˆœ›Ú™XÝÛ˜[YH‹ˆ›ØØ][Ûˆ‹ˆœ›ÝšY\ˆ‹ˆœ›ÙXÝ[Û—ÚÝÚ‹ˆœÜXÚYšX×ÞZY[‹ˆ™^XÝYÚÝÚ‹ˆ™]šX][Û—ÜÝ‹ˆœ\™›Ü›X[˜ÙWÜÝ]\È‹ˆ™]WÜÚ[È‹ˆ™]WÜÛÝ\˜ÙH‹ˆ›\ÝÝ\]H‹ˆ››Ý\È‹ˆKˆ™š[\œÈŽˆÂˆœ\š[ÙŽˆ›[Û‹ˆ›ÛWÛÛ›HŽˆžY\È‹ˆœÛÝ\˜ÙHŽˆ‘\Ú[Û”ÛÛ\ˆ‹ˆKˆKˆÂˆ›˜[YHŽˆ”™[]Üš[È›ÙXØ[ÈH[X[‹ˆ™]\Ù]Žˆœ›ÙXÝ[Û—Ü™\Ü‹ˆ™^ÜÙ›Ü›X]ŽˆžÞ‹ˆ˜ÛÛ[[œÈŽˆÂˆœ\š[Ù‹ˆœ›Ú™XÝÛ˜[YH‹ˆ›ØØ][Ûˆ‹ˆœ›ÝšY\ˆ‹ˆœ›ÙXÝ[Û—ÚÝÚ‹ˆœÜXÚYšX×ÞZY[‹ˆ™^XÝYÚÝÚ‹ˆ™]šX][Û—ÜÝ‹ˆœ\™›Ü›X[˜ÙWÜÝ]\È‹ˆ™]WÜÚ[È‹ˆ™]WÜÛÝ\˜ÙH‹ˆ›\ÝÝ\]H‹ˆ››Ý\È‹ˆKˆ™š[\œÈŽˆÂˆœ\š[ÙŽˆžYX\ˆ‹ˆ›ÛWÛÛ›HŽˆžY\È‹ˆœÛÝ\˜ÙHŽˆ‘\Ú[Û”ÛÛ\ˆ‹ˆKˆKˆB‚ˆ›Üˆ[\]H[ˆ™YYš[™YÝ[\]\Î‚ˆ^\Ý[™ÈHÛÛ›‹™^XÝ]Jˆ”ÑSPÕY”“ÓH^ÜÝ[\]\ÈÒT‘H˜[YHHÈSRUH‹ˆ
[\]VÈ›˜[YH—K
Kˆ
K™™]ÚÛ™J
BˆYˆ^\Ý[™Î‚ˆÛÛ[YBˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆS”ÑT•S•È^ÜÝ[\]\È
˜[YK]\Ù]^ÜÙ›Ü›X]ÛÛ[[œ×ÚœÛÛ‹š[\œ×ÚœÛÛ‹Ü™X]YØ]
BˆSQTÈ
ËËËËËÊBˆˆˆ‹ˆ
ˆ[\]VÈ›˜[YH—Kˆ[\]VÈ™]\Ù]—Kˆ[\]VÈ™^ÜÙ›Ü›X]—KˆœÛÛ‹™[\Ê[\]VÈ˜ÛÛ[[œÈ—K[œÝ\™WØ\ØÚZOUYJKˆœÛÛ‹™[\Ê[\]VÈ™š[\œÈ—K[œÝ\™WØ\ØÚZOUYJKˆ]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆ
Kˆ
B‚‚™YˆÙ]Ù\Ú[ÛœÛÛ\—Ù[—ØÛÛ™šYÊ
HOˆXÝÜÝ‹Ý—N‚ˆ™]\›ˆÂˆ\Ù\›˜[YHŽˆÜË™[š\›Û‹™Ù]
‘•TÒSÓ”ÓÓT—ÕTÑT“SQH‹ˆŠKœÝš\

Kˆœ\ÜÝÛÜ™ŽˆÜË™[š\›Û‹™Ù]
‘•TÒSÓ”ÓÓT—ÔTÔÕÓÔ‘‹ˆŠKœÝš\

Kˆ˜˜\ÙWÝ\›ŽˆÜË™[š\›Û‹™Ù]
‘•TÒSÓ”ÓÓT—ÐTÑWÕT“‹ˆŠKœÝš\

Kˆ›ÙÚ[—Ù[™Ú[ŽˆÜË™[š\›Û‹™Ù]
‘•TÒSÓ”ÓÓT—ÓÑÒS—ÑS‘ÒS•‹QUSÑ•TÒSÓ”ÓÓT—ÓÑÒS—ÑS‘ÒS•
KœÝš\

Kˆœ[×Ù[™Ú[ŽˆÜË™[š\›Û‹™Ù]
‘•TÒSÓ”ÓÓT—ÔÕUSÓ”×ÑS‘ÒS•‹QUSÑ•TÒSÓ”ÓÓT—ÔÕUSÓ”×ÑS‘ÒS•
KœÝš\

Kˆœ™X[Ý[YWÙ[™Ú[ŽˆÜË™[š\›Û‹™Ù]
ˆ‘•TÒSÓ”ÓÓT—Ô‘PSSQWÑS‘ÒS•‹ˆQUSÑ•TÒSÓ”ÓÓT—Ô‘PSSQWÑS‘ÒS•ˆ
KœÝš\

Kˆ™]šXÙWÛ\ÝÙ[™Ú[ŽˆÜË™[š\›Û‹™Ù]
ˆ‘•TÒSÓ”ÓÓT—ÑU’PÑT×ÑS‘ÒS•‹ˆQUSÑ•TÒSÓ”ÓÓT—ÑU’PÑT×ÑS‘ÒS•ˆ
KœÝš\

Kˆ™]šXÙWÜ™X[Ý[YWÙ[™Ú[ŽˆÜË™[š\›Û‹™Ù]
ˆ‘•TÒSÓ”ÓÓT—ÑU’PÑWÔ‘PSSQWÑS‘ÒS•‹ˆQUSÑ•TÒSÓ”ÓÓT—ÑU’PÑWÔ‘PSSQWÑS‘ÒS•ˆ
KœÝš\

Kˆ™]šXÙWÚ\ÝÜžWÙ[™Ú[ŽˆÜË™[š\›Û‹™Ù]
ˆ‘•TÒSÓ”ÓÓT—ÑU’PÑWÒTÕÔ–WÑS‘ÒS•‹ˆQUSÑ•TÒSÓ”ÓÓT—ÑU’PÑWÒTÕÔ–WÑS‘ÒS•ˆ
KœÝš\

Kˆ˜[\›\×Ù[™Ú[ŽˆÜË™[š\›Û‹™Ù]
‘•TÒSÓ”ÓÓT—ÐST“T×ÑS‘ÒS•‹QUSÑ•TÒSÓ”ÓÓT—ÐST“T×ÑS‘ÒS•
KœÝš\

Kˆ™^WÚÜWÙ[™Ú[ŽˆÜË™[š\›Û‹™Ù]
ˆ‘•TÒSÓ”ÓÓT—ÑVWÒÔWÑS‘ÒS•‹ˆQUSÑ•TÒSÓ”ÓÓT—ÑVWÒÔWÑS‘ÒS•ˆ
KœÝš\

Kˆ›[ÛÚÜWÙ[™Ú[ŽˆÜË™[š\›Û‹™Ù]
ˆ‘•TÒSÓ”ÓÓT—ÓSÓ•ÒÔWÑS‘ÒS•‹ˆQUSÑ•TÒSÓ”ÓÓT—ÓSÓ•ÒÔWÑS‘ÒS•ˆ
KœÝš\

KˆœÞ[˜×ÚÝ\œÈŽˆÜË™[š\›Û‹™Ù]
‘•TÒSÓ”ÓÓT—ÔÖS×ÒÕT”È‹QUSÑ•TÒSÓ”ÓÓT—ÔÖS×ÒÕT”ÊKœÝš\

KˆœÝ]WÜÞ[˜×Ú[\˜[ÚÝ\œÈŽˆÜË™[š\›Û‹™Ù]
‘•TÒSÓ”ÓÓT—ÔÕUWÔÖS×ÒS•T•SÒÕT”È‹ˆŠKœÝš\

Kˆœ›ÙXÝ[Û—ÜÞ[˜×Ù[˜X›YŽˆÜË™[š\›Û‹™Ù]
‘•TÒSÓ”ÓÓT—Ô“ÑPÕSÓ—ÔÖS×ÑSP“Q‹ˆŠKœÝš\

Kˆœ›ÙXÝ[Û—ÜÞ[˜×Ý[YHŽˆÜË™[š\›Û‹™Ù]
‘•TÒSÓ”ÓÓT—Ô“ÑPÕSÓ—ÔÖS×ÕSQH‹ˆŠKœÝš\

Kˆ™XYÛ›ÜÝXÜ×ÜÞ[˜×Ù[˜X›YŽˆÜË™[š\›Û‹™Ù]
‘•TÒSÓ”ÓÓT—ÑPQÓ“ÔÕPÔ×ÔÖS×ÑSP“Q‹ˆŠKœÝš\

Kˆ™XYÛ›ÜÝXÜ×ÜÞ[˜×Ý[YHŽˆÜË™[š\›Û‹™Ù]
‘•TÒSÓ”ÓÓT—ÑPQÓ“ÔÕPÔ×ÔÖS×ÕSQH‹ˆŠKœÝš\

KˆB‚‚™YˆÙ]ÜÚYÙ[™\™ÞWÙ[—ØÛÛ™šYÊ
HOˆXÝÜÝ‹Ý—N‚ˆ™]\›ˆÂˆ\Ù\›˜[YHŽˆÜË™[š\›Û‹™Ù]
”ÒQÑS‘T‘ÖWÐTÒÑVH‹ˆŠKœÝš\

Kˆœ\ÜÝÛÜ™ŽˆÜË™[š\›Û‹™Ù]
”ÒQÑS‘T‘ÖWÐTÔÑPÔ‘U‹ˆŠKœÝš\

Kˆ˜˜\ÙWÝ\›ŽˆÜË™[š\›Û‹™Ù]
”ÒQÑS‘T‘ÖWÐTÑWÕT“‹QUSÔÒQÑS‘T‘ÖWÐTÑWÕT“
KœÝš\

Kˆ›ÙÚ[—Ù[™Ú[ŽˆÜË™[š\›Û‹™Ù]
”ÒQÑS‘T‘ÖWÐUUÑS‘ÒS•‹QUSÔÒQÑS‘T‘ÖWÐUUÑS‘ÒS•
KœÝš\

Kˆœ[×Ù[™Ú[ŽˆÜË™[š\›Û‹™Ù]
”ÒQÑS‘T‘ÖWÔÖTÕST×ÑS‘ÒS•‹QUSÔÒQÑS‘T‘ÖWÔÖTÕST×ÑS‘ÒS•
KœÝš\

Kˆ™[™\™ÞWÙ›Ý×Ù[™Ú[ŽˆÜË™[š\›Û‹™Ù]
”ÒQÑS‘T‘ÖWÑS‘T‘ÖWÑ“Õ×ÑS‘ÒS•‹QUSÔÒQÑS‘T‘ÖWÑS‘T‘ÖWÑ“Õ×ÑS‘ÒS•
KœÝš\

Kˆ›Û˜›Ø\™Ù[™Ú[ŽˆÜË™[š\›Û‹™Ù]
”ÒQÑS‘T‘ÖWÓÓ“ÐT‘ÑS‘ÒS•‹QUSÔÒQÑS‘T‘ÖWÓÓ“ÐT‘ÑS‘ÒS•
KœÝš\

Kˆ™^WÚÜWÙ[™Ú[Žˆˆ‹ˆ›[ÛÚÜWÙ[™Ú[Žˆˆ‹ˆœÞ[˜×ÚÝ\œÈŽˆÜË™[š\›Û‹™Ù]
”ÒQÑS‘T‘ÖWÔÖS×ÒÕT”È‹QUSÔÒQÑS‘T‘ÖWÔÖS×ÒÕT”ÊKœÝš\

KˆœÝ]WÜÞ[˜×Ú[\˜[ÚÝ\œÈŽˆÜË™[š\›Û‹™Ù]
”ÒQÑS‘T‘ÖWÔÕUWÔÖS×ÒS•T•SÒÕT”È‹ˆŠKœÝš\

Kˆœ™YÚ[ÛˆŽˆÜË™[š\›Û‹™Ù]
”ÒQÑS‘T‘ÖWÔ‘QÒSÓˆ‹QUSÔÒQÑS‘T‘ÖWÔ‘QÒSÓŠKœÝš\

HÜˆQUSÔÒQÑS‘T‘ÖWÔ‘QÒSÓ‹ˆœÞ\Ý[WÚYÈŽˆÜË™[š\›Û‹™Ù]
”ÒQÑS‘T‘ÖWÔÖTÕSWÒQÈ‹ÜË™[š\›Û‹™Ù]
”ÒQÑS‘T‘ÖWÔÖTÕSWÒQ‹ˆŠJKœÝš\

KˆœÛ˜\ÚÝÜ™][[Û—Ù^\ÈŽˆÜË™[š\›Û‹™Ù]
”ÒQÑS‘T‘ÖWÔÓTÒÕÔ‘US•SÓ—ÑVTÈ‹ÝŠQUSÔÒQÑS‘T‘ÖWÔÓTÒÕÔ‘US•SÓ—ÑVTÊJKœÝš\

Kˆ™[˜X›YŽˆÜË™[š\›Û‹™Ù]
”ÒQÑS‘T‘ÖWÑSP“Q‹ˆŠKœÝš\

KˆB‚‚™Yˆ[œÝ\™WÚ[YÜ˜][Û—ÜÙYYÙ]JÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[ÛŠHOˆ›Û™N‚ˆ[—ØÛÛ™šYÈHÙ]Ù\Ú[ÛœÛÛ\—Ù[—ØÛÛ™šYÊ
Bˆ^\Ý[™ÈHÛÛ›‹™^XÝ]Jˆ”ÑSPÕ
ˆ”“ÓH[YÜ˜][Û—ØÛÛ™šYÜÈÒT‘H›ÝšY\ˆHÈ‹ˆ
S•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹
Kˆ
K™™]ÚÛ™J
BˆYˆ^\Ý[™Î‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUH[YÜ˜][Û—ØÛÛ™šYÜÂˆÑU\Ù\›˜[YHHÐTÑHÒSˆÓÐSTÐÑJ\Ù\›˜[YK	ÉÊHH	ÉÈSˆÈSÑH\Ù\›˜[YHS‘ˆ˜\ÙWÝ\›HÐTÑHÒSˆÓÐSTÐÑJ˜\ÙWÝ\›	ÉÊHH	ÉÈSˆÈSÑH˜\ÙWÝ\›S‘ˆÙÚ[—Ù[™Ú[HÐTÑHÒSˆÓÐSTÐÑJÙÚ[—Ù[™Ú[	ÉÊHH	ÉÈSˆÈSÑHÙÚ[—Ù[™Ú[S‘ˆ[×Ù[™Ú[HÐTÑHÒSˆÓÐSTÐÑJ[×Ù[™Ú[	ÉÊHH	ÉÈSˆÈSÑH[×Ù[™Ú[S‘ˆ™X[Ý[YWÙ[™Ú[HÐTÑHÒSˆÓÐSTÐÑJ™X[Ý[YWÙ[™Ú[	ÉÊHH	ÉÈSˆÈSÑH™X[Ý[YWÙ[™Ú[S‘ˆ]šXÙWÛ\ÝÙ[™Ú[HÐTÑHÒSˆÓÐSTÐÑJ]šXÙWÛ\ÝÙ[™Ú[	ÉÊHH	ÉÈSˆÈSÑH]šXÙWÛ\ÝÙ[™Ú[S‘ˆ]šXÙWÜ™X[Ý[YWÙ[™Ú[HÐTÑHÒSˆÓÐSTÐÑJ]šXÙWÜ™X[Ý[YWÙ[™Ú[	ÉÊHH	ÉÈSˆÈSÑH]šXÙWÜ™X[Ý[YWÙ[™Ú[S‘ˆ]šXÙWÚ\ÝÜžWÙ[™Ú[HÐTÑHÒSˆÓÐSTÐÑJ]šXÙWÚ\ÝÜžWÙ[™Ú[	ÉÊHH	ÉÈSˆÈSÑH]šXÙWÚ\ÝÜžWÙ[™Ú[S‘ˆ[\›\×Ù[™Ú[HÐTÑHÒSˆÓÐSTÐÑJ[\›\×Ù[™Ú[	ÉÊHH	ÉÈSˆÈSÑH[\›\×Ù[™Ú[S‘ˆ^WÚÜWÙ[™Ú[HÐTÑHÒSˆÓÐSTÐÑJ^WÚÜWÙ[™Ú[	ÉÊHH	ÉÈSˆÈSÑH^WÚÜWÙ[™Ú[S‘ˆ[ÛÚÜWÙ[™Ú[HÐTÑHÒSˆÓÐSTÐÑJ[ÛÚÜWÙ[™Ú[	ÉÊHH	ÉÈSˆÈSÑH[ÛÚÜWÙ[™Ú[S‘ˆÞ[˜×ÚÝ\œÈHÐTÑHÒSˆÓÐSTÐÑJÞ[˜×ÚÝ\œË	ÉÊHH	ÉÈSˆÈSÑHÞ[˜×ÚÝ\œÈS‘ˆ›ÙXÝ[Û—ÜÞ[˜×Ù[˜X›YHÐTÑHÒSˆ›ÙXÝ[Û—ÜÞ[˜×Ù[˜X›YTÈ•SSˆHSÑH›ÙXÝ[Û—ÜÞ[˜×Ù[˜X›YS‘ˆXYÛ›ÜÝXÜ×ÜÞ[˜×Ù[˜X›YHÐTÑHÒSˆXYÛ›ÜÝXÜ×ÜÞ[˜×Ù[˜X›YTÈ•SSˆHSÑHXYÛ›ÜÝXÜ×ÜÞ[˜×Ù[˜X›YS‘ˆÝ]WÜÞ[˜×Ú[\˜[ÚÝ\œÈHÐTÑHÒSˆÝ]WÜÞ[˜×Ú[\˜[ÚÝ\œÈTÈ•SÔˆÝ]WÜÞ[˜×Ú[\˜[ÚÝ\œÈHSˆÈSÑHÝ]WÜÞ[˜×Ú[\˜[ÚÝ\œÈS‘ˆ›ÙXÝ[Û—ÜÞ[˜×Ý[YHHÐTÑHÒSˆÓÐSTÐÑJ›ÙXÝ[Û—ÜÞ[˜×Ý[YK	ÉÊHH	ÉÈSˆÈSÑH›ÙXÝ[Û—ÜÞ[˜×Ý[YHS‘ˆXYÛ›ÜÝXÜ×ÜÞ[˜×Ý[YHHÐTÑHÒSˆÓÐSTÐÑJXYÛ›ÜÝXÜ×ÜÞ[˜×Ý[YK	ÉÊHH	ÉÈSˆÈSÑHXYÛ›ÜÝXÜ×ÜÞ[˜×Ý[YHS‘ˆ\]YØ]HÂˆÒT‘H›ÝšY\ˆHÂˆˆˆ‹ˆ
ˆ[—ØÛÛ™šYÖÈ\Ù\›˜[YH—Kˆ[—ØÛÛ™šYÖÈ˜˜\ÙWÝ\›—Kˆ[—ØÛÛ™šYÖÈ›ÙÚ[—Ù[™Ú[—Kˆ[—ØÛÛ™šYÖÈœ[×Ù[™Ú[—Kˆ[—ØÛÛ™šYÖÈœ™X[Ý[YWÙ[™Ú[—Kˆ[—ØÛÛ™šYÖÈ™]šXÙWÛ\ÝÙ[™Ú[—Kˆ[—ØÛÛ™šYÖÈ™]šXÙWÜ™X[Ý[YWÙ[™Ú[—Kˆ[—ØÛÛ™šYÖÈ™]šXÙWÚ\ÝÜžWÙ[™Ú[—Kˆ[—ØÛÛ™šYÖÈ˜[\›\×Ù[™Ú[—Kˆ[—ØÛÛ™šYÖÈ™^WÚÜWÙ[™Ú[—Kˆ[—ØÛÛ™šYÖÈ›[ÛÚÜWÙ[™Ú[—Kˆ[—ØÛÛ™šYÖÈœÞ[˜×ÚÝ\œÈ—KˆQUSÔÕUWÔÖS×ÒS•T•SÒÕT”ËˆQUSÑ•TÒSÓ”ÓÓT—Ô“ÑPÕSÓ—ÔÖS×ÕSQKˆQUSÑ•TÒSÓ”ÓÓT—ÑPQÓ“ÔÕPÔ×ÔÖS×ÕSQKˆ]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹ˆ
Kˆ
Bˆ[ÙN‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆS”ÑT•S•È[YÜ˜][Û—ØÛÛ™šYÜÈ
ˆ›ÝšY\‹\Ù\›˜[YK\ÜÝÛÜ™˜\ÙWÝ\›ÙÚ[—Ù[™Ú[[×Ù[™Ú[™X[Ý[YWÙ[™Ú[ˆ]šXÙWÛ\ÝÙ[™Ú[]šXÙWÜ™X[Ý[YWÙ[™Ú[]šXÙWÚ\ÝÜžWÙ[™Ú[[\›\×Ù[™Ú[ˆ^WÚÜWÙ[™Ú[[ÛÚÜWÙ[™Ú[ˆ[˜X›Y]]×ÜÞ[˜×Ù[˜X›YÞ[˜×ÚÝ\œË›ÙXÝ[Û—ÜÞ[˜×Ù[˜X›YXYÛ›ÜÝXÜ×ÜÞ[˜×Ù[˜X›YˆÝ]WÜÞ[˜×Ú[\˜[ÚÝ\œË›ÙXÝ[Û—ÜÞ[˜×Ý[YKXYÛ›ÜÝXÜ×ÜÞ[˜×Ý[YKÜ™X]YØ]\]YØ]ˆ
HSQTÈ
ËËËËËËËËËËËËËËËËËËËËËËÊBˆˆˆ‹ˆ
ˆS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹ˆ[—ØÛÛ™šYÖÈ\Ù\›˜[YH—Kˆˆ‹ˆ[—ØÛÛ™šYÖÈ˜˜\ÙWÝ\›—Kˆ[—ØÛÛ™šYÖÈ›ÙÚ[—Ù[™Ú[—Kˆ[—ØÛÛ™šYÖÈœ[×Ù[™Ú[—Kˆ[—ØÛÛ™šYÖÈœ™X[Ý[YWÙ[™Ú[—Kˆ[—ØÛÛ™šYÖÈ™]šXÙWÛ\ÝÙ[™Ú[—Kˆ[—ØÛÛ™šYÖÈ™]šXÙWÜ™X[Ý[YWÙ[™Ú[—Kˆ[—ØÛÛ™šYÖÈ™]šXÙWÚ\ÝÜžWÙ[™Ú[—Kˆ[—ØÛÛ™šYÖÈ˜[\›\×Ù[™Ú[—Kˆ[—ØÛÛ™šYÖÈ™^WÚÜWÙ[™Ú[—Kˆ[—ØÛÛ™šYÖÈ›[ÛÚÜWÙ[™Ú[—Kˆˆˆ[—ØÛÛ™šYÖÈœÞ[˜×ÚÝ\œÈ—KˆKˆKˆQUSÔÕUWÔÖS×ÒS•T•SÒÕT”ËˆQUSÑ•TÒSÓ”ÓÓT—Ô“ÑPÕSÓ—ÔÖS×ÕSQKˆQUSÑ•TÒSÓ”ÓÓT—ÑPQÓ“ÔÕPÔ×ÔÖS×ÕSQKˆ]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆ]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆ
Kˆ
B‚ˆÚYÙ[™\™ÞWÙ[ˆHÙ]ÜÚYÙ[™\™ÞWÙ[—ØÛÛ™šYÊ
BˆÚYÙ[™\™ÞWÙ^\Ý[™ÈHÛÛ›‹™^XÝ]Jˆ”ÑSPÕ
ˆ”“ÓH[YÜ˜][Û—ØÛÛ™šYÜÈÒT‘H›ÝšY\ˆHÈ‹ˆ
S•QÔUSÓ—Ô“Õ’QT—ÔÒQÑS‘T‘ÖK
Kˆ
K™™]ÚÛ™J
BˆÚYÙ[™\™ÞWÙ[˜X›YHHYˆÚYÙ[™\™ÞWÙ[–È™[˜X›Y—K›ÝÙ\Š
H[ˆÈŒH‹YH‹žY\È‹œÚ[H‹›ÛˆŸH[ÙHˆYˆÚYÙ[™\™ÞWÙ^\Ý[™Î‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUH[YÜ˜][Û—ØÛÛ™šYÜÂˆÑU\Ù\›˜[YHHÐTÑHÒSˆÓÐSTÐÑJ\Ù\›˜[YK	ÉÊHH	ÉÈSˆÈSÑH\Ù\›˜[YHS‘ˆ˜\ÙWÝ\›HÐTÑHÒSˆÓÐSTÐÑJ˜\ÙWÝ\›	ÉÊHH	ÉÈSˆÈSÑH˜\ÙWÝ\›S‘ˆÙÚ[—Ù[™Ú[HÐTÑHÒSˆÓÐSTÐÑJÙÚ[—Ù[™Ú[	ÉÊHH	ÉÈSˆÈSÑHÙÚ[—Ù[™Ú[S‘ˆ[×Ù[™Ú[HÐTÑHÒSˆÓÐSTÐÑJ[×Ù[™Ú[	ÉÊHH	ÉÈSˆÈSÑH[×Ù[™Ú[S‘ˆ[™\™ÞWÙ›Ý×Ù[™Ú[HÐTÑHÒSˆÓÐSTÐÑJ[™\™ÞWÙ›Ý×Ù[™Ú[	ÉÊHH	ÉÈSˆÈSÑH[™\™ÞWÙ›Ý×Ù[™Ú[S‘ˆÛ˜›Ø\™Ù[™Ú[HÐTÑHÒSˆÓÐSTÐÑJÛ˜›Ø\™Ù[™Ú[	ÉÊHH	ÉÈSˆÈSÑHÛ˜›Ø\™Ù[™Ú[S‘ˆÞ[˜×ÚÝ\œÈHÐTÑHÒSˆÓÐSTÐÑJÞ[˜×ÚÝ\œË	ÉÊHH	ÉÈSˆÈSÑHÞ[˜×ÚÝ\œÈS‘ˆÝ]WÜÞ[˜×Ú[\˜[ÚÝ\œÈHÐTÑHÒSˆÝ]WÜÞ[˜×Ú[\˜[ÚÝ\œÈTÈ•SÔˆÝ]WÜÞ[˜×Ú[\˜[ÚÝ\œÈHSˆÈSÑHÝ]WÜÞ[˜×Ú[\˜[ÚÝ\œÈS‘ˆ›ÙXÝ[Û—ÜÞ[˜×Ù[˜X›YHˆXYÛ›ÜÝXÜ×ÜÞ[˜×Ù[˜X›YHˆ™YÚ[ÛˆHÐTÑHÒSˆÓÐSTÐÑJ™YÚ[Û‹	ÉÊHH	ÉÈSˆÈSÑH™YÚ[ÛˆS‘ˆÞ\Ý[WÚYÈHÐTÑHÒSˆÓÐSTÐÑJÞ\Ý[WÚYË	ÉÊHH	ÉÈSˆÈSÑHÞ\Ý[WÚYÈS‘ˆÛ˜\ÚÝÜ™][[Û—Ù^\ÈHÐTÑHÒSˆÛ˜\ÚÝÜ™][[Û—Ù^\ÈTÈ•SÔˆÛ˜\ÚÝÜ™][[Û—Ù^\ÈHSˆÈSÑHÛ˜\ÚÝÜ™][[Û—Ù^\ÈS‘ˆ[˜X›YHÐTÑHÒSˆÈHHSˆHSÑH[˜X›YS‘ˆ\]YØ]HÂˆÒT‘H›ÝšY\ˆHÂˆˆˆ‹ˆ
ˆÚYÙ[™\™ÞWÙ[–È\Ù\›˜[YH—KˆÚYÙ[™\™ÞWÙ[–È˜˜\ÙWÝ\›—KˆÚYÙ[™\™ÞWÙ[–È›ÙÚ[—Ù[™Ú[—KˆÚYÙ[™\™ÞWÙ[–Èœ[×Ù[™Ú[—KˆÚYÙ[™\™ÞWÙ[–È™[™\™ÞWÙ›Ý×Ù[™Ú[—KˆÚYÙ[™\™ÞWÙ[–È›Û˜›Ø\™Ù[™Ú[—KˆÚYÙ[™\™ÞWÙ[–ÈœÞ[˜×ÚÝ\œÈ—KˆQUSÔÕUWÔÖS×ÒS•T•SÒÕT”ËˆÚYÙ[™\™ÞWÙ[–Èœ™YÚ[Ûˆ—KˆÚYÙ[™\™ÞWÙ[–ÈœÞ\Ý[WÚYÈ—Kˆ[
ÚYÙ[™\™ÞWÙ[–ÈœÛ˜\ÚÝÜ™][[Û—Ù^\È—HÜˆQUSÔÒQÑS‘T‘ÖWÔÓTÒÕÔ‘US•SÓ—ÑVTÊKˆÚYÙ[™\™ÞWÙ[˜X›Yˆ]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆS•QÔUSÓ—Ô“Õ’QT—ÔÒQÑS‘T‘ÖKˆ
Kˆ
Bˆ™]\›‚‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆS”ÑT•S•È[YÜ˜][Û—ØÛÛ™šYÜÈ
ˆ›ÝšY\‹\Ù\›˜[YK\ÜÝÛÜ™˜\ÙWÝ\›ÙÚ[—Ù[™Ú[[×Ù[™Ú[[™\™ÞWÙ›Ý×Ù[™Ú[Û˜›Ø\™Ù[™Ú[ˆ^WÚÜWÙ[™Ú[[ÛÚÜWÙ[™Ú[™YÚ[Û‹Þ\Ý[WÚYËÛ˜\ÚÝÜ™][[Û—Ù^\Ëˆ[˜X›Y]]×ÜÞ[˜×Ù[˜X›YÞ[˜×ÚÝ\œË›ÙXÝ[Û—ÜÞ[˜×Ù[˜X›YXYÛ›ÜÝXÜ×ÜÞ[˜×Ù[˜X›YˆÝ]WÜÞ[˜×Ú[\˜[ÚÝ\œËÜ™X]YØ]\]YØ]ˆ
HSQTÈ
ËËËËËËËËËËËËËËËËËËËËÊBˆˆˆ‹ˆ
ˆS•QÔUSÓ—Ô“Õ’QT—ÔÒQÑS‘T‘ÖKˆÚYÙ[™\™ÞWÙ[–È\Ù\›˜[YH—Kˆˆ‹ˆÚYÙ[™\™ÞWÙ[–È˜˜\ÙWÝ\›—KˆÚYÙ[™\™ÞWÙ[–È›ÙÚ[—Ù[™Ú[—KˆÚYÙ[™\™ÞWÙ[–Èœ[×Ù[™Ú[—KˆÚYÙ[™\™ÞWÙ[–È™[™\™ÞWÙ›Ý×Ù[™Ú[—KˆÚYÙ[™\™ÞWÙ[–È›Û˜›Ø\™Ù[™Ú[—Kˆˆ‹ˆˆ‹ˆÚYÙ[™\™ÞWÙ[–Èœ™YÚ[Ûˆ—KˆÚYÙ[™\™ÞWÙ[–ÈœÞ\Ý[WÚYÈ—Kˆ[
ÚYÙ[™\™ÞWÙ[–ÈœÛ˜\ÚÝÜ™][[Û—Ù^\È—HÜˆQUSÔÒQÑS‘T‘ÖWÔÓTÒÕÔ‘US•SÓ—ÑVTÊKˆÚYÙ[™\™ÞWÙ[˜X›YˆÚYÙ[™\™ÞWÙ[˜X›YˆÚYÙ[™\™ÞWÙ[–ÈœÞ[˜×ÚÝ\œÈ—KˆˆˆQUSÔÕUWÔÖS×ÒS•T•SÒÕT”Ëˆ]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆ]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆ
Kˆ
B‚‚™YˆÙ]Ú[YÜ˜][Û—ØÛÛ™šYÊÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹›ÝšY\ŽˆÝŠHOˆXÝÜÝ‹[žWH›Û™N‚ˆ›ÝÈHÛÛ›‹™^XÝ]J”ÑSPÕ
ˆ”“ÓH[YÜ˜][Û—ØÛÛ™šYÜÈÒT‘H›ÝšY\ˆHÈ‹
›ÝšY\‹
JK™™]ÚÛ™J
BˆYˆ›ÝÈ\È›Û™N‚ˆ™]\›ˆ›Û™BˆÛÛ™šYÈHXÝ
›ÝÊBˆYˆ›ÝšY\ˆOHS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓTŽ‚ˆ[—ØÛÛ™šYÈHÙ]Ù\Ú[ÛœÛÛ\—Ù[—ØÛÛ™šYÊ
BˆÛÛ™šYÖÈ™[—ÛÝ™\œšY\È—HHßBˆ›ÜˆÙ^K˜[YH[ˆ[—ØÛÛ™šYËš][\Ê
N‚ˆYˆ˜[YH[™Ù^H[ˆÛÛ™šYÎ‚ˆÛÛ™šYÖÚÙ^WHH˜[YBˆÛÛ™šYÖÈ™[—ÛÝ™\œšY\È—VÚÙ^WHHYBˆÛÛ™šYÖÈœÝ]WÜÞ[˜×Ú[\˜[ÚÝ\œÈ—HH›Ü›X[^™WÜÜÚ]]™WÚ[
ˆÛÛ™šYË™Ù]
œÝ]WÜÞ[˜×Ú[\˜[ÚÝ\œÈŠKˆQUSÔÕUWÔÖS×ÒS•T•SÒÕT”ËˆZ[š[][OLKˆX^[][OLˆ
BˆÛÛ™šYÖÈœ›ÙXÝ[Û—ÜÞ[˜×Ý[YH—HH›Ü›X[^™WØÛØÚ×Ý[YJˆÛÛ™šYË™Ù]
œ›ÙXÝ[Û—ÜÞ[˜×Ý[YHŠKˆQUSÑ•TÒSÓ”ÓÓT—Ô“ÑPÕSÓ—ÔÖS×ÕSQKˆ
BˆÛÛ™šYÖÈ™XYÛ›ÜÝXÜ×ÜÞ[˜×Ý[YH—HH›Ü›X[^™WØÛØÚ×Ý[YJˆÛÛ™šYË™Ù]
™XYÛ›ÜÝXÜ×ÜÞ[˜×Ý[YHŠKˆQUSÑ•TÒSÓ”ÓÓT—ÑPQÓ“ÔÕPÔ×ÔÖS×ÕSQKˆ
BˆÛÛ™šYÖÈœ›ÙXÝ[Û—ÜÞ[˜×Ù[˜X›Y—HHHYˆ›Ü›X[^™WØ›ÛÛ
ÛÛ™šYË™Ù]
œ›ÙXÝ[Û—ÜÞ[˜×Ù[˜X›YŠKYJH[ÙHˆÛÛ™šYÖÈ™XYÛ›ÜÝXÜ×ÜÞ[˜×Ù[˜X›Y—HHHYˆ›Ü›X[^™WØ›ÛÛ
ÛÛ™šYË™Ù]
™XYÛ›ÜÝXÜ×ÜÞ[˜×Ù[˜X›YŠKYJH[ÙHˆÛÛ™šYÖÈœ\ÜÝÛÜ™ØÛÛ™šYÝ\™Y—HH›ÛÛ
ÛÛ™šYË™Ù]
œ\ÜÝÛÜ™ŠJBˆÛÛ™šYÖÈœ\ÜÝÛÜ™ÜÛÝ\˜ÙH—HH™[ˆˆYˆ[—ØÛÛ™šYÖÈœ\ÜÝÛÜ™—H[ÙH
™]X˜\ÙHˆYˆÛÛ™šYË™Ù]
œ\ÜÝÛÜ™ŠH[ÙHˆŠBˆYˆ›ÝšY\ˆOHS•QÔUSÓ—Ô“Õ’QT—ÔÒQÑS‘T‘ÖN‚ˆ[—ØÛÛ™šYÈHÙ]ÜÚYÙ[™\™ÞWÙ[—ØÛÛ™šYÊ
Bˆ[—ÚÙ^WÛX\HÂˆ\Ù\›˜[YHŽˆ”ÒQÑS‘T‘ÖWÐTÒÑVH‹ˆœ\ÜÝÛÜ™Žˆ”ÒQÑS‘T‘ÖWÐTÔÑPÔ‘U‹ˆ˜˜\ÙWÝ\›Žˆ”ÒQÑS‘T‘ÖWÐTÑWÕT“‹ˆ›ÙÚ[—Ù[™Ú[Žˆ”ÒQÑS‘T‘ÖWÐUUÑS‘ÒS•‹ˆœ[×Ù[™Ú[Žˆ”ÒQÑS‘T‘ÖWÔÖTÕST×ÑS‘ÒS•‹ˆ™[™\™ÞWÙ›Ý×Ù[™Ú[Žˆ”ÒQÑS‘T‘ÖWÑS‘T‘ÖWÑ“Õ×ÑS‘ÒS•‹ˆ›Û˜›Ø\™Ù[™Ú[Žˆ”ÒQÑS‘T‘ÖWÓÓ“ÐT‘ÑS‘ÒS•‹ˆœÞ[˜×ÚÝ\œÈŽˆ”ÒQÑS‘T‘ÖWÔÖS×ÒÕT”È‹ˆœÝ]WÜÞ[˜×Ú[\˜[ÚÝ\œÈŽˆ”ÒQÑS‘T‘ÖWÔÕUWÔÖS×ÒS•T•SÒÕT”È‹ˆœÛ˜\ÚÝÜ™][[Û—Ù^\ÈŽˆ”ÒQÑS‘T‘ÖWÔÓTÒÕÔ‘US•SÓ—ÑVTÈ‹ˆBˆÛÛ™šYÖÈ™[—ÛÝ™\œšY\È—HHßBˆ›ÜˆÙ^K[—ÚÙ^H[ˆ[—ÚÙ^WÛX\š][\Ê
N‚ˆ[—Ý˜[YHHÜË™[š\›Û‹™Ù]
[—ÚÙ^KˆŠKœÝš\

BˆYˆ[—Ý˜[YH[™Ù^H[ˆÛÛ™šYÎ‚ˆÛÛ™šYÖÚÙ^WHH[—Ý˜[YBˆÛÛ™šYÖÈ™[—ÛÝ™\œšY\È—VÚÙ^WHHYBˆÛÛ™šYÖÈ˜˜\ÙWÝ\›—HHÛÛ™šYË™Ù]
˜˜\ÙWÝ\›ŠHÜˆQUSÔÒQÑS‘T‘ÖWÐTÑWÕT“ˆÛÛ™šYÖÈ›ÙÚ[—Ù[™Ú[—HHÛÛ™šYË™Ù]
›ÙÚ[—Ù[™Ú[ŠHÜˆQUSÔÒQÑS‘T‘ÖWÐUUÑS‘ÒS•ˆÛÛ™šYÖÈœ[×Ù[™Ú[—HHÛÛ™šYË™Ù]
œ[×Ù[™Ú[ŠHÜˆQUSÔÒQÑS‘T‘ÖWÔÖTÕST×ÑS‘ÒS•ˆÛÛ™šYÖÈ™[™\™ÞWÙ›Ý×Ù[™Ú[—HHÛÛ™šYË™Ù]
™[™\™ÞWÙ›Ý×Ù[™Ú[ŠHÜˆQUSÔÒQÑS‘T‘ÖWÑS‘T‘ÖWÑ“Õ×ÑS‘ÒS•ˆÛÛ™šYÖÈ›Û˜›Ø\™Ù[™Ú[—HHÛÛ™šYË™Ù]
›Û˜›Ø\™Ù[™Ú[ŠHÜˆQUSÔÒQÑS‘T‘ÖWÓÓ“ÐT‘ÑS‘ÒS•ˆÛÛ™šYÖÈœÞ[˜×ÚÝ\œÈ—HHÛÛ™šYË™Ù]
œÞ[˜×ÚÝ\œÈŠHÜˆQUSÔÒQÑS‘T‘ÖWÔÖS×ÒÕT”ÂˆÛÛ™šYÖÈœÝ]WÜÞ[˜×Ú[\˜[ÚÝ\œÈ—HH›Ü›X[^™WÜÜÚ]]™WÚ[
ˆÛÛ™šYË™Ù]
œÝ]WÜÞ[˜×Ú[\˜[ÚÝ\œÈŠKˆQUSÔÕUWÔÖS×ÒS•T•SÒÕT”ËˆZ[š[][OLKˆX^[][OLˆ
BˆÛÛ™šYÖÈœ›ÙXÝ[Û—ÜÞ[˜×Ù[˜X›Y—HHˆÛÛ™šYÖÈ™XYÛ›ÜÝXÜ×ÜÞ[˜×Ù[˜X›Y—HHˆÛÛ™šYÖÈœÛ˜\ÚÝÜ™][[Û—Ù^\È—HH[
ÛÛ™šYË™Ù]
œÛ˜\ÚÝÜ™][[Û—Ù^\ÈŠHÜˆQUSÔÒQÑS‘T‘ÖWÔÓTÒÕÔ‘US•SÓ—ÑVTÊBˆYˆÜË™[š\›Û‹™Ù]
”ÒQÑS‘T‘ÖWÔ‘QÒSÓˆ‹ˆŠKœÝš\

N‚ˆÛÛ™šYÖÈœ™YÚ[Ûˆ—HH[—ØÛÛ™šYÖÈœ™YÚ[Ûˆ—BˆÛÛ™šYÖÈ™[—ÛÝ™\œšY\È—VÈœ™YÚ[Ûˆ—HHYBˆ[ÙN‚ˆÛÛ™šYÖÈœ™YÚ[Ûˆ—HHÛÛ™šYË™Ù]
œ™YÚ[ÛˆŠHÜˆQUSÔÒQÑS‘T‘ÖWÔ‘QÒSÓ‚ˆYˆ[—ØÛÛ™šYÖÈœÞ\Ý[WÚYÈ—N‚ˆÛÛ™šYÖÈœÞ\Ý[WÚYÈ—HH[—ØÛÛ™šYÖÈœÞ\Ý[WÚYÈ—BˆÛÛ™šYÖÈ™[—ÛÝ™\œšY\È—VÈœÞ\Ý[WÚYÈ—HHYBˆ[ÙN‚ˆÛÛ™šYÖÈœÞ\Ý[WÚYÈ—HHÛÛ™šYË™Ù]
œÞ\Ý[WÚYÈŠHÜˆˆ‚ˆÛÛ™šYÖÈœ\ÜÝÛÜ™ØÛÛ™šYÝ\™Y—HH›ÛÛ
ÛÛ™šYË™Ù]
œ\ÜÝÛÜ™ŠJBˆÛÛ™šYÖÈœ\ÜÝÛÜ™ÜÛÝ\˜ÙH—HH™[ˆˆYˆ[—ØÛÛ™šYÖÈœ\ÜÝÛÜ™—H[ÙH
™]X˜\ÙHˆYˆÛÛ™šYË™Ù]
œ\ÜÝÛÜ™ŠH[ÙHˆŠBˆYˆ[—ØÛÛ™šYÖÈ™[˜X›Y—K›ÝÙ\Š
H[ˆÈŒH‹YH‹žY\È‹œÚ[H‹›ÛˆŸN‚ˆÛÛ™šYÖÈ™[˜X›Y—HHBˆ™]\›ˆÛÛ™šYÂ‚‚™YˆÝ\Ú[YÜ˜][Û—ÜØÚY[\Š\ˆ›\ÚÊHOˆ›Û™N‚ˆÛØ˜[ÐÒQST‚ˆYˆÐÒQSTˆ\È›Ý›Û™N‚ˆ™]\›‚ˆÐÒQSTˆH˜XÚÙÜ›Ý[™ØÚY[\Š[Y^›Û™OH‘]\›ÜKÓ\Ø›ÛˆŠBˆÐÒQST‹œÝ\

Bˆ™Yœ™\ÚÚ[YÜ˜][Û—ÜØÚY[\Š\
Bˆ™YÚ\Ý\—Ø˜XÚÙÜ›Ý[™Ú›Ø—Ü™XXÝ]˜][Û—ÜØÚY[\Š\
BˆØÚY[WÜ[™[™×Ø˜XÚÙÜ›Ý[™Ú›ØœÊ\
B‚‚™Yˆ™Yœ™\ÚÚ[YÜ˜][Û—ÜØÚY[\Š\ˆ›\ÚÊHOˆ›Û™N‚ˆÛØ˜[ÐÒQST‚ˆYˆÐÒQSTˆ\È›Û™N‚ˆ™]\›‚ˆ›Üˆ›Øˆ[ˆ\Ý
ÐÒQST‹™Ù]Ú›ØœÊ
JN‚ˆYˆ
ˆ›Ø‹šYœÝ\ÝÚ]
š[YÜ˜][Û‹\Þ[˜ËHŠBˆÜˆ›Ø‹šYœÝ\ÝÚ]
™\Ú[ÛœÛÛ\‹\Þ[˜ËHŠBˆÜˆ›Ø‹šYˆ[ˆÂˆ[YÜ˜[KYZ[K\Ý[[X\žH‹ˆ™\Ú[ÛœÛÛ\‹\›ÙXÝ[Û‹YZ[H‹ˆ™\Ú[ÛœÛÛ\‹]Ø]YZ[H‹ˆš[YÜ˜][Û‹\Ý]KY\Ú[ÛœÛÛ\‹ZÝ\›H‹ˆš[YÜ˜][Û‹\›ÙXÝ[Û‹Y\Ú[ÛœÛÛ\‹YZ[H‹ˆš[YÜ˜][Û‹YXYÛ›ÜÝXÜËY\Ú[ÛœÛÛ\‹YZ[H‹ˆš[YÜ˜][Û‹\›ÙXÝ[Û‹Y\Ú[ÛœÛÛ\‹[[ÛXÛÜÙH‹ˆš[YÜ˜][Û‹Y\Ú[ÛœÛÛ\‹\™X[[YKXÛX[\‹ˆš[YÜ˜][Û‹\Ý]K\ÚYÙ[™\™ÞKZÝ\›H‹ˆ˜˜XÚÙÜ›Ý[™Z›ØœË\™XXÝ]˜]K\˜]K[[Z]‹ˆBˆ
N‚ˆÐÒQST‹œ™[[Ý™WÚ›ØŠ›Ø‹šY
B‚ˆÚ]ÛÜÚ[™ÊÙ]ÙŠ\˜ÛÛ™šYÖÈ‘UPTÑH—JJH\ÈÛÛ›Ž‚ˆÛÛ™šYÜÈHÙÙ]Ú[YÜ˜][Û—ØÛÛ™šYÊÛÛ›‹›ÝšY\ŠH›Üˆ›ÝšY\ˆ[ˆS•QÔUSÓ—Ô“Õ’QT—ÓÔSÓ”×BˆYˆ[YÜ˜[WÙZ[WÜÝ[[X\žWÙ[˜X›Y

N‚ˆÐÒQST‹˜YÚ›ØŠˆ[˜Ï\[—ÜØÚY[YÝ[YÜ˜[WÙZ[WÜÝ[[X\žKˆšYÙÙ\H˜Ü›Ûˆ‹ˆÝ\NKˆZ[]OLˆ\™ÜÏVØ\KˆYH[YÜ˜[KYZ[K\Ý[[X\žH‹ˆ™\XÙWÙ^\Ý[™ÏUYKˆX^Ú[œÝ[˜Ù\ÏLKˆÛØ[\ØÙOUYKˆZ\Ùš\™WÙÜ˜XÙWÝ[YOLNˆ
Bˆ›ÜˆÛÛ™šYÈ[ˆÛÛ™šYÜÎ‚ˆYˆÛÛ™šYÈ\È›Û™HÜˆ›ÝÛÛ™šYÖÈ™[˜X›Y—N‚ˆÛÛ[YBˆ›ÝšY\ˆHÝŠÛÛ™šYÖÈœ›ÝšY\ˆ—JBˆYˆ›ÝšY\ˆOHS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓTŽ‚ˆ™YÚ\Ý\—Ù\Ú[ÛœÛÛ\—ÜØÚY[\—Ú›ØœÊ\ÛÛ™šYÊBˆÛÛ[YBˆYˆ›ÝšY\ˆOHS•QÔUSÓ—Ô“Õ’QT—ÔÒQÑS‘T‘ÖN‚ˆ™YÚ\Ý\—ÜÚYÙ[™\™ÞWÜØÚY[\—Ú›ØœÊ\ÛÛ™šYÊBˆ™YÚ\Ý\—Ø˜XÚÙÜ›Ý[™Ú›Ø—Ü™XXÝ]˜][Û—ÜØÚY[\Š\
B‚‚™YˆYÜØÚY[\—Ú›ØŠ
ŠšÝØ\™ÜÎˆ[žJHOˆ›Û™N‚ˆYˆÐÒQSTˆ\È›Û™N‚ˆ™]\›‚ˆÝØ\™ÜËœÙ]Y˜][
œ™\XÙWÙ^\Ý[™È‹YJBˆÝØ\™ÜËœÙ]Y˜][
›X^Ú[œÝ[˜Ù\È‹JBˆÝØ\™ÜËœÙ]Y˜][
˜ÛØ[\ØÙH‹YJBˆÝØ\™ÜËœÙ]Y˜][
›Z\Ùš\™WÙÜ˜XÙWÝ[YH‹N
BˆÐÒQST‹˜YÚ›ØŠ
ŠšÝØ\™ÜÊB‚‚™Yˆ™YÚ\Ý\—ÚÝ\›WÜÝ]WÜÞ[˜Ê\ˆ›\ÚËÛÛ™šYÎˆXÝÜÝ‹[žWK›Ø—ÚYˆÝŠHOˆ›Û™N‚ˆYˆ›ÝÛÛ™šYË™Ù]
˜]]×ÜÞ[˜×Ù[˜X›YŠN‚ˆ™]\›‚ˆ[\˜[ÚÝ\œÈH›Ü›X[^™WÜÜÚ]]™WÚ[
ˆÛÛ™šYË™Ù]
œÝ]WÜÞ[˜×Ú[\˜[ÚÝ\œÈŠKˆQUSÔÕUWÔÖS×ÒS•T•SÒÕT”ËˆZ[š[][OLKˆX^[][OLˆ
Bˆ›ÝšY\ˆHÝŠÛÛ™šYÖÈœ›ÝšY\ˆ—JBˆ\›ÙÙÙ\‹š[™›Ê”™YÚ\Ý\š[™È	\ÈÝ]HÞ[˜È]™\žH	\ÈÝ\ŠÊH‹›ÝšY\‹[\˜[ÚÝ\œÊBˆYÜØÚY[\—Ú›ØŠˆ[˜Ï\[—ÜØÚY[YÚÝ\›WÜÝ]WÜÞ[˜ËˆšYÙÙ\Hš[\˜[‹ˆÝ\œÏZ[\˜[ÚÝ\œËˆ\™ÜÏVØ\›ÝšY\—KˆYZ›Ø—ÚYˆ
B‚‚™Yˆ™YÚ\Ý\—Ù\Ú[ÛœÛÛ\—ÜØÚY[\—Ú›ØœÊ\ˆ›\ÚËÛÛ™šYÎˆXÝÜÝ‹[žWJHOˆ›Û™N‚ˆ™YÚ\Ý\—ÚÝ\›WÜÝ]WÜÞ[˜Ê\ÛÛ™šYËš[YÜ˜][Û‹\Ý]KY\Ú[ÛœÛÛ\‹ZÝ\›HŠBˆYˆÛÛ™šYË™Ù]
œ›ÙXÝ[Û—ÜÞ[˜×Ù[˜X›YŠN‚ˆÝ\‹Z[]HHÜ]ØÛØÚ×Ý[YJÛÛ™šYË™Ù]
œ›ÙXÝ[Û—ÜÞ[˜×Ý[YHŠKQUSÑ•TÒSÓ”ÓÓT—Ô“ÑPÕSÓ—ÔÖS×ÕSQJBˆYÜØÚY[\—Ú›ØŠˆ[˜Ï\[—ÜØÚY[YÙ\Ú[ÛœÛÛ\—Ü›ÙXÝ[Û—ÜÞ[˜ËˆšYÙÙ\H˜Ü›Ûˆ‹ˆÝ\ZÝ\‹ˆZ[]O[Z[]Kˆ\™ÜÏVØ\KˆYHš[YÜ˜][Û‹\›ÙXÝ[Û‹Y\Ú[ÛœÛÛ\‹YZ[H‹ˆ
BˆYˆÛÛ™šYË™Ù]
™XYÛ›ÜÝXÜ×ÜÞ[˜×Ù[˜X›YŠN‚ˆÝ\‹Z[]HHÜ]ØÛØÚ×Ý[YJÛÛ™šYË™Ù]
™XYÛ›ÜÝXÜ×ÜÞ[˜×Ý[YHŠKQUSÑ•TÒSÓ”ÓÓT—ÑPQÓ“ÔÕPÔ×ÔÖS×ÕSQJBˆYÜØÚY[\—Ú›ØŠˆ[˜Ï\[—ÜØÚY[YÙ\Ú[ÛœÛÛ\—ÙXYÛ›ÜÝXÜ×ÜÞ[˜ËˆšYÙÙ\H˜Ü›Ûˆ‹ˆÝ\ZÝ\‹ˆZ[]O[Z[]Kˆ\™ÜÏVØ\KˆYHš[YÜ˜][Û‹YXYÛ›ÜÝXÜËY\Ú[ÛœÛÛ\‹YZ[H‹ˆ
BˆYˆÛÛ™šYË™Ù]
œ›ÙXÝ[Û—ÜÞ[˜×Ù[˜X›YŠN‚ˆYÜØÚY[\—Ú›ØŠˆ[˜Ï\[—ÜØÚY[YÙ\Ú[ÛœÛÛ\—Û[ÛØÛÜÙKˆšYÙÙ\H˜Ü›Ûˆ‹ˆ^OHŒKMH‹ˆÝ\L‹ˆZ[]OLˆ\™ÜÏVØ\KˆYHš[YÜ˜][Û‹\›ÙXÝ[Û‹Y\Ú[ÛœÛÛ\‹[[ÛXÛÜÙH‹ˆ
BˆYÜØÚY[\—Ú›ØŠˆ[˜Ï\[—ÜØÚY[YÙ\Ú[ÛœÛÛ\—Ü™X[[YWØÛX[\ˆšYÙÙ\H˜Ü›Ûˆ‹ˆÝ\LËˆZ[]OLÌˆ\™ÜÏVØ\KˆYHš[YÜ˜][Û‹Y\Ú[ÛœÛÛ\‹\™X[[YKXÛX[\‹ˆ
B‚‚™Yˆ™YÚ\Ý\—ÜÚYÙ[™\™ÞWÜØÚY[\—Ú›ØœÊ\ˆ›\ÚËÛÛ™šYÎˆXÝÜÝ‹[žWJHOˆ›Û™N‚ˆ™YÚ\Ý\—ÚÝ\›WÜÝ]WÜÞ[˜Ê\ÛÛ™šYËš[YÜ˜][Û‹\Ý]K\ÚYÙ[™\™ÞKZÝ\›HŠB‚‚™Yˆ™YÚ\Ý\—Ø˜XÚÙÜ›Ý[™Ú›Ø—Ü™XXÝ]˜][Û—ÜØÚY[\Š\ˆ›\ÚÊHOˆ›Û™N‚ˆYÜØÚY[\—Ú›ØŠˆ[˜Ï\[—ÜØÚY[YØ˜XÚÙÜ›Ý[™Ú›Ø—Ü™XXÝ]˜][Û‹ˆšYÙÙ\Hš[\˜[‹ˆZ[]\ÏMKˆ\™ÜÏVØ\KˆYH˜˜XÚÙÜ›Ý[™Z›ØœË\™XXÝ]˜]K\˜]K[[Z]‹ˆ
B‚‚™Yˆ[—ÜØÚY[YØ˜XÚÙÜ›Ý[™Ú›Ø—Ü™XXÝ]˜][ÛŠ\ˆ›\ÚÊHOˆ›Û™N‚ˆÝ[[X\žHHØÚY[WÜ[™[™×Ø˜XÚÙÜ›Ý[™Ú›ØœÊ\
BˆYˆÝ[[X\žK™Ù]
œ˜]WÛ[Z]Ü™XXÝ]˜]YŠHÜˆÝ[[X\žK™Ù]
œ[™[™×ÜØÚY[YŠN‚ˆ\›ÙÙÙ\‹š[™›Ê˜XÚÙÜ›Ý[™›Øˆ™XXÝ]˜][ÛˆÝ[[X\žNˆ	\È‹Ý[[X\žJB‚‚™Yˆ[—ÜØÚY[YÝ[YÜ˜[WÙZ[WÜÝ[[X\žJ\ˆ›\ÚÊHOˆ›Û™N‚ˆÚ]\˜\ØÛÛ^

N‚ˆÚ]ÛÜÚ[™ÊÙ]ÙŠ\˜ÛÛ™šYÖÈ‘UPTÑH—JJH\ÈÛÛ›Ž‚ˆžN‚ˆÙ[™ÙZ[WÝ[YÜ˜[WÜÝ[[X\žJÛÛ›ŠBˆÛÛ›‹˜ÛÛ[Z]

Bˆ^Ù\^Ù\[ÛŽ‚ˆÝ\œ™[Ø\›ÙÙÙ\‹™^Ù\[ÛŠ”ØÚY[Y[YÜ˜[HZ[HÝ[[X\žH˜Z[YŠB‚‚™Yˆ[—ÜØÚY[YÚ[YÜ˜][Û—ÜÞ[˜Ê\ˆ›\ÚË›ÝšY\ŽˆÝŠHOˆ›Û™N‚ˆÚ]\˜\ØÛÛ^

N‚ˆÚ]ÛÜÚ[™ÊÙ]ÙŠ\˜ÛÛ™šYÖÈ‘UPTÑH—JJH\ÈÛÛ›Ž‚ˆžN‚ˆÝ\œ™[Ø\›ÙÙÙ\‹š[™›Ê”ØÚY[Y	\ÈÞ[˜ÈÝ\Y‹›ÝšY\ŠBˆ™\Ý[H[—Ú[YÜ˜][Û—ÜÞ[˜ÊÛÛ›‹›ÝšY\‹šYÙÙ\—Ý\OHœØÚY[YŠBˆÝ\œ™[Ø\›ÙÙÙ\‹š[™›Ê”ØÚY[Y	\ÈÞ[˜ÈÛÛ\]Yˆ	\È‹›ÝšY\‹™\Ý[
Bˆ^Ù\^Ù\[ÛŽ‚ˆÝ\œ™[Ø\›ÙÙÙ\‹™^Ù\[ÛŠ”ØÚY[Y	\ÈÞ[˜È˜Z[Y‹›ÝšY\ŠB‚‚™Yˆ[—ÜØÚY[YÚÝ\›WÜÝ]WÜÞ[˜Ê\ˆ›\ÚË›ÝšY\ŽˆÝŠHOˆ›Û™N‚ˆÚ]\˜\ØÛÛ^

N‚ˆÚ]ÛÜÚ[™ÊÙ]ÙŠ\˜ÛÛ™šYÖÈ‘UPTÑH—JJH\ÈÛÛ›Ž‚ˆžN‚ˆ›Ø—Ý\HHœÚYÙ[™\™ÞWÜÝ]WÜÞ[˜ÈˆYˆ›ÝšY\ˆOHS•QÔUSÓ—Ô“Õ’QT—ÔÒQÑS‘T‘ÖH[ÙH™\Ú[ÛœÛÛ\—ÜÝ]WÜÞ[˜È‚ˆ›Ø—ÚYÜ™X]YHÜ™X]WØ˜XÚÙÜ›Ý[™Ú›ØŠˆÛÛ›‹ˆ›Ø—Ý\KˆÈœ›ÝšY\ˆŽˆ›ÝšY\‹šYÙÙ\—Ý\HŽˆœØÚY[YÜÝ]HŸKˆ
BˆÛÛ›‹˜ÛÛ[Z]

BˆYˆÜ™X]Y‚ˆØÚY[WØ˜XÚÙÜ›Ý[™Ú›ØŠ\›Ø—ÚY
BˆÝ\œ™[Ø\›ÙÙÙ\‹š[™›Ê”ØÚY[Y	\ÈÝ]HÞ[˜È]Y]YYˆ›Ø—ÚYI\ÈÜ™X]YI\È‹›ÝšY\‹›Ø—ÚYÜ™X]Y
Bˆ^Ù\^Ù\[ÛŽ‚ˆÝ\œ™[Ø\›ÙÙÙ\‹™^Ù\[ÛŠ”ØÚY[Y	\ÈÝ]HÞ[˜È˜Z[Y‹›ÝšY\ŠB‚‚™Yˆ[—ÜØÚY[YÙ\Ú[ÛœÛÛ\—ÜÞ[˜Ê\ˆ›\ÚÊHOˆ›Û™N‚ˆ[—ÜØÚY[YÚÝ\›WÜÝ]WÜÞ[˜Ê\S•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓTŠB‚‚™YˆÝ\œ™[Û\Ø›Û—Ù]J
HOˆ]N‚ˆ™]\›ˆ]][YK››ÝÊTÐ“Ó—ÕSQV“Ó‘JK™]J
B‚‚™Yˆ[—ÜØÚY[YÙ\Ú[ÛœÛÛ\—Ü›ÙXÝ[Û—ÜÞ[˜Ê\ˆ›\ÚÊHOˆ›Û™N‚ˆØÚY[\—Ù]HHÝ\œ™[Û\Ø›Û—Ù]J
Bˆ\™Ù]Ù]HHØÚY[\—Ù]HH[YY[J^\ÏLJBˆÚ]\˜\ØÛÛ^

N‚ˆÚ]ÛÜÚ[™ÊÙ]ÙŠ\˜ÛÛ™šYÖÈ‘UPTÑH—JJH\ÈÛÛ›Ž‚ˆÝ\œ™[Ø\›ÙÙÙ\‹š[™›Êˆ”ØÚY[Y\Ú[Û”ÛÛ\ˆ›ÙXÝ[Ûˆ™\\š[™È™]š[Ý\ÈÛÜÙY^NˆØÚY[\—Ù]OI\È\™Ù]Ù]OI\È‹ˆØÚY[\—Ù]Kˆ\™Ù]Ù]Kˆ
BˆÛÛ™šYÈHÙ]Ú[YÜ˜][Û—ØÛÛ™šYÊÛÛ›‹S•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓTŠBˆYˆÛÛ™šYÈ\È›Û™HÜˆ›ÝÛÛ™šYÖÈ™[˜X›Y—N‚ˆÝ\œ™[Ø\›ÙÙÙ\‹š[™›Êˆ”ØÚY[Y\Ú[Û”ÛÛ\ˆ›ÙXÝ[ÛˆÚÚ\Y™XØ]\ÙH[YÜ˜][Ûˆ\È\ØX›Yˆ\™Ù]Ù]OI\È‹ˆ\™Ù]Ù]Kˆ
Bˆ™]\›‚ˆYˆ›ÝÛÛ™šYË™Ù]
œ›ÙXÝ[Û—ÜÞ[˜×Ù[˜X›YŠN‚ˆÝ\œ™[Ø\›ÙÙÙ\‹š[™›Êˆ”ØÚY[Y\Ú[Û”ÛÛ\ˆ›ÙXÝ[ÛˆÚÚ\Y™XØ]\ÙH›ÙXÝ[ÛˆÞ[˜È\È\ØX›Yˆ\™Ù]Ù]OI\È‹ˆ\™Ù]Ù]Kˆ
Bˆ™]\›‚ˆ›Ø—ÚYÜ™X]YHÜ™X]WØ˜XÚÙÜ›Ý[™Ú›ØŠˆÛÛ›‹ˆ™\Ú[ÛœÛÛ\—Ü›ÙXÝ[Û—ÜÞ[˜È‹ˆÂˆœ›ÝšY\ˆŽˆS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹ˆ\™Ù]Ù]HŽˆ\™Ù]Ù]Kš\ÛÙ›Ü›X]

Kˆœ\š[ÙÝ\HŽˆ™^H‹ˆšYÙÙ\—Ý\HŽˆœØÚY[Y‹ˆKˆ
BˆÛÛ›‹˜ÛÛ[Z]

BˆYˆÜ™X]Y‚ˆØÚY[WØ˜XÚÙÜ›Ý[™Ú›ØŠ\›Ø—ÚY
BˆÝ\œ™[Ø\›ÙÙÙ\‹š[™›Êˆ”ØÚY[Y\Ú[Û”ÛÛ\ˆ›ÙXÝ[Ûˆ]Y]YYˆ›Ø—ÚYI\È\™Ù]Ù]OI\È‹ˆ›Ø—ÚYˆ\™Ù]Ù]Kˆ
Bˆ[ÙN‚ˆÝ\œ™[Ø\›ÙÙÙ\‹š[™›Êˆ”ØÚY[Y\Ú[Û”ÛÛ\ˆ›ÙXÝ[Ûˆ™]\ÙY^\Ý[™È[™[™ËÜ[›š[™È›ØŽˆ›Ø—ÚYI\È\™Ù]Ù]OI\È‹ˆ›Ø—ÚYˆ\™Ù]Ù]Kˆ
B‚‚™Yˆ[—ÜØÚY[YÙ\Ú[ÛœÛÛ\—ÝØ]Ø˜XÚÙš[
\ˆ›\ÚÊHOˆ›Û™N‚ˆØÚY[\—Ù]HHÝ\œ™[Û\Ø›Û—Ù]J
Bˆ\™Ù]Ù]HHØÚY[\—Ù]HH[YY[J^\ÏLJBˆÚ]\˜\ØÛÛ^

N‚ˆÚ]ÛÜÚ[™ÊÙ]ÙŠ\˜ÛÛ™šYÖÈ‘UPTÑH—JJH\ÈÛÛ›Ž‚ˆÝ\œ™[Ø\›ÙÙÙ\‹š[™›Êˆ”ØÚY[Y\Ú[Û”ÛÛ\ˆÐU™\\š[™È™]š[Ý\ÈÛÜÙY^NˆØÚY[\—Ù]OI\È\™Ù]Ù]OI\È‹ˆØÚY[\—Ù]Kˆ\™Ù]Ù]Kˆ
BˆÛÛ™šYÈHÙ]Ú[YÜ˜][Û—ØÛÛ™šYÊÛÛ›‹S•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓTŠBˆYˆÛÛ™šYÈ\È›Û™HÜˆ›ÝÛÛ™šYÖÈ™[˜X›Y—N‚ˆÝ\œ™[Ø\›ÙÙÙ\‹š[™›Êˆ”ØÚY[Y\Ú[Û”ÛÛ\ˆÐUÚÚ\Y™XØ]\ÙH[YÜ˜][Ûˆ\È\ØX›Yˆ\™Ù]Ù]OI\È‹ˆ\™Ù]Ù]Kˆ
Bˆ™]\›‚ˆYˆ›ÝÛÛ™šYË™Ù]
™XYÛ›ÜÝXÜ×ÜÞ[˜×Ù[˜X›YŠN‚ˆÝ\œ™[Ø\›ÙÙÙ\‹š[™›Êˆ”ØÚY[Y\Ú[Û”ÛÛ\ˆÐUÚÚ\Y™XØ]\ÙHXYÛ›ÜÝXÜÈÞ[˜È\È\ØX›Yˆ\™Ù]Ù]OI\È‹ˆ\™Ù]Ù]Kˆ
Bˆ™]\›‚ˆ›Ø—ÚYÜ™X]YHÜ™X]WØ˜XÚÙÜ›Ý[™Ú›ØŠˆÛÛ›‹ˆ™\Ú[ÛœÛÛ\—Ú[™\\—Ø]˜Z[Xš[]WØ˜XÚÙš[‹ˆÂˆœ›ÝšY\ˆŽˆS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹ˆ™œ›ÛWÙ]HŽˆ\™Ù]Ù]Kš\ÛÙ›Ü›X]

Kˆ×Ù]HŽˆ\™Ù]Ù]Kš\ÛÙ›Ü›X]

KˆšYÙÙ\—Ý\HŽˆœØÚY[Y‹ˆKˆ
BˆÛÛ›‹˜ÛÛ[Z]

BˆYˆÜ™X]Y‚ˆØÚY[WØ˜XÚÙÜ›Ý[™Ú›ØŠ\›Ø—ÚY
BˆÝ\œ™[Ø\›ÙÙÙ\‹š[™›Êˆ”ØÚY[Y\Ú[Û”ÛÛ\ˆÐU]Y]YYˆ›Ø—ÚYI\È\™Ù]Ù]OI\È›Ø—Ý\OI\È‹ˆ›Ø—ÚYˆ\™Ù]Ù]Kˆ™\Ú[ÛœÛÛ\—Ú[™\\—Ø]˜Z[Xš[]WØ˜XÚÙš[‹ˆ
Bˆ[ÙN‚ˆÝ\œ™[Ø\›ÙÙÙ\‹š[™›Êˆ”ØÚY[Y\Ú[Û”ÛÛ\ˆÐU™]\ÙY^\Ý[™È[™[™ËÜ[›š[™È›ØŽˆ›Ø—ÚYI\È\™Ù]Ù]OI\È‹ˆ›Ø—ÚYˆ\™Ù]Ù]Kˆ
B‚‚™Yˆ[—ÜØÚY[YÙ\Ú[ÛœÛÛ\—ÙXYÛ›ÜÝXÜ×ÜÞ[˜Ê\ˆ›\ÚÊHOˆ›Û™N‚ˆ[—ÜØÚY[YÙ\Ú[ÛœÛÛ\—ÝØ]Ø˜XÚÙš[
\
B‚‚™Yˆ[—ÜØÚY[YÙ\Ú[ÛœÛÛ\—Ü™X[[YWØÛX[\
\ˆ›\ÚÊHOˆ›Û™N‚ˆÚ]\˜\ØÛÛ^

N‚ˆÚ]ÛÜÚ[™ÊÙ]ÙŠ\˜ÛÛ™šYÖÈ‘UPTÑH—JJH\ÈÛÛ›Ž‚ˆ›Ø—ÚYÜ™X]YHÜ™X]WØ˜XÚÙÜ›Ý[™Ú›ØŠˆÛÛ›‹ˆ™\Ú[ÛœÛÛ\—Ü™X[[YWÛX]\šX[^™WØÛX[\‹ˆÂˆœ›ÝšY\ˆŽˆS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹ˆœ™Y™\™[˜ÙWÙ]HŽˆÝ\œ™[Û\Ø›Û—Ù]J
Kš\ÛÙ›Ü›X]

KˆšYÙÙ\—Ý\HŽˆœØÚY[YÜ™][[Ûˆ‹ˆKˆ
BˆÛÛ›‹˜ÛÛ[Z]

BˆYˆÜ™X]Y‚ˆØÚY[WØ˜XÚÙÜ›Ý[™Ú›ØŠ\›Ø—ÚY
BˆÝ\œ™[Ø\›ÙÙÙ\‹š[™›Êˆ”ØÚY[Y\Ú[Û”ÛÛ\ˆ™X[[YHX]\šX[^˜][Û‹ØÛX[\]Y]YYˆ‚ˆš›Ø—ÚYI\ÈÜ™X]YI\È‹ˆ›Ø—ÚYˆÜ™X]Yˆ
B‚‚™Yˆ[—ÜØÚY[YÙ\Ú[ÛœÛÛ\—Û[ÛØÛÜÙJ\ˆ›\ÚÊHOˆ›Û™N‚ˆØÚY[\—Ù]HHÝ\œ™[Û\Ø›Û—Ù]J
BˆYˆØÚY[\—Ù]K™^H›Ý[ˆ˜[™ÙJKŠN‚ˆ\›ÙÙÙ\‹š[™›Êˆ”ØÚY[Y\Ú[Û”ÛÛ\ˆ[ÛÛÜÙHÚÚ\YÝ]ÚYH^\ÈKMNˆØÚY[\—Ù]OI\È‹ˆØÚY[\—Ù]Kˆ
Bˆ™]\›‚ˆ™]š[Ý\×Ù^HHØÚY[\—Ù]Kœ™\XÙJ^OLJHH[YY[J^\ÏLJBˆ™\ÜÛ[ÛH™]š[Ý\×Ù^KœÝ™[YJ‰VKI[HŠBˆÚ]\˜\ØÛÛ^

N‚ˆÚ]ÛÜÚ[™ÊÙ]ÙŠ\˜ÛÛ™šYÖÈ‘UPTÑH—JJH\ÈÛÛ›Ž‚ˆÛÛ™šYÈHÙ]Ú[YÜ˜][Û—ØÛÛ™šYÊÛÛ›‹S•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓTŠBˆYˆÛÛ™šYÈ\È›Û™HÜˆ›ÝÛÛ™šYÖÈ™[˜X›Y—HÜˆ›ÝÛÛ™šYË™Ù]
œ›ÙXÝ[Û—ÜÞ[˜×Ù[˜X›YŠN‚ˆÝ\œ™[Ø\›ÙÙÙ\‹š[™›Êˆ”ØÚY[Y\Ú[Û”ÛÛ\ˆ[ÛÛÜÙHÚÚ\Y™XØ]\ÙH›ÙXÝ[Ûˆ[YÜ˜][Ûˆ\È\ØX›Yˆ[ÛI\È‹ˆ™\ÜÛ[Ûˆ
Bˆ™]\›‚ˆ›Ø—ÚYÜ™X]YHÜ™X]WØ˜XÚÙÜ›Ý[™Ú›ØŠˆÛÛ›‹ˆ™\Ú[ÛœÛÛ\—Û[ÛØÛÜÙH‹ˆÂˆœ›ÝšY\ˆŽˆS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹ˆœ™\ÜÛ[ÛŽˆ™\ÜÛ[ÛˆšYÙÙ\—Ý\HŽˆœØÚY[YÛ[ÛØÛÜÙH‹ˆKˆ
BˆÛÛ›‹˜ÛÛ[Z]

BˆYˆÜ™X]Y‚ˆØÚY[WØ˜XÚÙÜ›Ý[™Ú›ØŠ\›Ø—ÚY
BˆÝ\œ™[Ø\›ÙÙÙ\‹š[™›Êˆ”ØÚY[Y\Ú[Û”ÛÛ\ˆ[ÛÛÜÙH]Y]YYˆ›Ø—ÚYI\ÈÜ™X]YI\È[ÛI\ÈØÚY[\—Ù]OI\È‹ˆ›Ø—ÚYˆÜ™X]Yˆ™\ÜÛ[ÛˆØÚY[\—Ù]Kˆ
B‚‚™YˆØÚY[WØ˜XÚÙÜ›Ý[™Ú›ØŠ\ˆ›\ÚË›Ø—ÚYˆ[[—Ù]Nˆ]][YH›Û™HH›Û™JHOˆ›ÛÛ‚ˆYˆÐÒQSTˆ\È›Û™N‚ˆ\›ÙÙÙ\‹™\œ›ÜŠ˜XÚÙÜ›Ý[™›Øˆ	\ÈØ\È]Y]YY]TØÚY[\ˆ\È›Ý[›š[™È‹›Ø—ÚY
Bˆ™]\›ˆ˜[ÙBˆØÚY[YØ]H\×Ø˜XÚÙÜ›Ý[™Ú›Ø—Ý]Êˆ[—Ù]HÜˆ˜XÚÙÜ›Ý[™Ú›Ø—Ý]×Û›ÝÊ
Bˆ
BˆÐÒQST‹˜YÚ›ØŠˆ[˜Ï\[—Ø˜XÚÙÜ›Ý[™Ú›Ø‹ˆšYÙÙ\H™]H‹ˆ[—Ù]O\ØÚY[YØ]ˆ\™ÜÏVØ\›Ø—ÚYKˆYYˆ˜˜XÚÙÜ›Ý[™Z›Ø‹^Ú›Ø—ÚYH‹ˆ™\XÙWÙ^\Ý[™ÏUYKˆX^Ú[œÝ[˜Ù\ÏLKˆ
Bˆ™]\›ˆYB‚‚™YˆØÚY[WÜ[™[™×Ø˜XÚÙÜ›Ý[™Ú›ØœÊ\ˆ›\ÚÊHOˆXÝÜÝ‹[žWN‚ˆÚ]ÛÜÚ[™ÊÙ]ÙŠ\˜ÛÛ™šYÖÈ‘UPTÑH—JJH\ÈÛÛ›Ž‚ˆ^\™YÛX\Ù\×Ü™XÛÝ™\™YH™XÛÝ™\—Ù^\™YÛX\Ù\ÊÛÛ›ŠBˆYˆ^\™YÛX\Ù\×Ü™XÛÝ™\™Y‚ˆ\›ÙÙÙ\‹š[™›Êˆ”™XÛÝ™\™Y	\È^\™Y›ÙXÝ[ÛˆTHX\Ù\È‹ˆ^\™YÛX\Ù\×Ü™XÛÝ™\™Yˆ
Bˆ™XÛÝ™\™YØÛÝ[HX\š×ÜÝ[WÜ[›š[™×Ø˜XÚÙÜ›Ý[™Ú›Øœ×Ù˜Z[Y
ÛÛ›ŠBˆYˆ™XÛÝ™\™YØÛÝ[‚ˆ\›ÙÙÙ\‹Ø\›š[™Ê“X\šÙY	\ÈÝ[H[›š[™È˜XÚÙÜ›Ý[™›ØœÈ\È˜Z[YÛˆÝ\\‹™XÛÝ™\™YØÛÝ[
Bˆ™XXÝ]˜]YØÛÝ[H™XXÝ]˜]WÙYWÜ˜]WÛ[Z]YØ˜XÚÙÜ›Ý[™Ú›ØœÊÛÛ›ŠBˆYˆ™XXÝ]˜]YØÛÝ[‚ˆ\›ÙÙÙ\‹š[™›Ê”™XXÝ]˜]Y	\È˜XÚÙÜ›Ý[™›ØœÈY\ˆTHÛÛÛÝÛˆ‹™XXÝ]˜]YØÛÝ[
Bˆ[™[™×Ú›Ø—ÚYÈH™]ÚÜ[™[™×Ø˜XÚÙÜ›Ý[™Ú›Ø—ÚYÊÛÛ›ŠBˆ]\™WÝØZ][™×Ú›ØœÈH™]ÚÙ]\™WÝØZ][™×Ø˜XÚÙÜ›Ý[™Ú›ØœÊÛÛ›ŠBˆ˜Z[YÚ›Ø—ÚYÎˆ\ÝÚ[HH×Bˆ›Üˆ›Ø—ÚY[ˆ[™[™×Ú›Ø—ÚYÎ‚ˆYˆ›ÝØÚY[WØ˜XÚÙÜ›Ý[™Ú›ØŠ\›Ø—ÚY
N‚ˆ˜Z[YÚ›Ø—ÚYË˜\[™
›Ø—ÚY
Bˆ˜Z[YÝØZ][™×Ú›Ø—ÚYÎˆ\ÝÚ[HH×Bˆ›Üˆ›Ø—ÚY™^Ø][\Ø][ˆ]\™WÝØZ][™×Ú›ØœÎ‚ˆYˆ›ÝØÚY[WØ˜XÚÙÜ›Ý[™Ú›ØŠˆ\ˆ›Ø—ÚYˆ[—Ù]O[™^Ø][\Ø]ˆ
N‚ˆ˜Z[YÝØZ][™×Ú›Ø—ÚYË˜\[™
›Ø—ÚY
BˆØÚY[YØÛÝ[H[Š[™[™×Ú›Ø—ÚYÊHH[Š˜Z[YÚ›Ø—ÚYÊBˆØZ][™×ÜØÚY[YØÛÝ[H
ˆ[Š]\™WÝØZ][™×Ú›ØœÊHH[Š˜Z[YÝØZ][™×Ú›Ø—ÚYÊBˆ
BˆYˆØÚY[YØÛÝ[‚ˆ\›ÙÙÙ\‹š[™›Ê”ØÚY[Y	\È[™[™È˜XÚÙÜ›Ý[™›ØœÈÛˆÝ\\‹ØÚY[YØÛÝ[
BˆYˆ˜Z[YÚ›Ø—ÚYÎ‚ˆ\›ÙÙÙ\‹Ø\›š[™ÊÛÝ[›ÝØÚY[H[™[™È˜XÚÙÜ›Ý[™›ØœÈÛˆÝ\\ˆ	\È‹˜Z[YÚ›Ø—ÚYÊBˆYˆ˜Z[YÝØZ][™×Ú›Ø—ÚYÎ‚ˆ\›ÙÙÙ\‹Ø\›š[™ÊˆÛÝ[›Ý™\ÝÜ™HØZ][™È˜XÚÙÜ›Ý[™›ØœÈÛˆÝ\\ˆ	\È‹ˆ˜Z[YÝØZ][™×Ú›Ø—ÚYËˆ
Bˆ™]\›ˆÂˆœÝ[WÜ[›š[™×Ù˜Z[YŽˆ™XÛÝ™\™YØÛÝ[ˆœ˜]WÛ[Z]Ü™XXÝ]˜]YŽˆ™XXÝ]˜]YØÛÝ[ˆœ[™[™×Ù›Ý[™Žˆ[Š[™[™×Ú›Ø—ÚYÊKˆœ[™[™×ÜØÚY[YŽˆØÚY[YØÛÝ[ˆœ[™[™×ÜØÚY[WÙ˜Z[YÚYÈŽˆ˜Z[YÚ›Ø—ÚYËˆØZ][™×Ù›Ý[™Žˆ[Š]\™WÝØZ][™×Ú›ØœÊKˆØZ][™×ÜØÚY[YŽˆØZ][™×ÜØÚY[YØÛÝ[ˆØZ][™×ÜØÚY[WÙ˜Z[YÚYÈŽˆ˜Z[YÝØZ][™×Ú›Ø—ÚYËˆB‚‚™Yˆ\×Ý˜[œÚY[ÜÜ[]WÛØÚÊ^Îˆ˜\ÙQ^Ù\[ÛŠHOˆ›ÛÛ‚ˆY\ÜØYÙHHÝŠ^ÊK›ÝÙ\Š
Bˆ™]\›ˆ\Ú[œÝ[˜ÙJ^ËÜ[]LË“Ü\˜][Û˜[\œ›ÜŠH[™
ˆ™]X˜\ÙH\ÈØÚÙYˆ[ˆY\ÜØYÙBˆÜˆ™]X˜\ÙHX›H\ÈØÚÙYˆ[ˆY\ÜØYÙBˆÜˆ™]X˜\ÙH\È\ÞHˆ[ˆY\ÜØYÙBˆ
B‚‚™YˆY™\—Ø˜XÚÙÜ›Ý[™Ú›Ø—ØY\—Ù]X˜\ÙWÛØÚÊˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ\ˆ›\ÚËˆ›Ø—ÚYˆ[ˆ\˜[\ÎˆXÝÜÝ‹[žWKŠHOˆ›ÛÛ‚ˆ][\ÈH[
\˜[\Ë™Ù]
—ÜÜ[]WÛØÚ×Ø][\ŠHÜˆ
H
ÈBˆYˆ][\ÈˆŽ‚ˆ™]\›ˆ˜[ÙBˆ[^WÜÙXÛÛ™ÈHZ[ŠH
ˆ
ˆ
Šˆ
][\ÈHJJKÌ
Bˆ™^Ø][\Ø]H˜XÚÙÜ›Ý[™Ú›Ø—Ý]×Û›ÝÊ
H
È[YY[JˆÙXÛÛ™ÏY[^WÜÙXÛÛ™Âˆ
Bˆ\˜[\ÈHÊŠœ\˜[\Ë—ÜÜ[]WÛØÚ×Ø][\Žˆ][\ßBˆÛÛ›‹œ›Û˜XÚÊ
BˆÛÛ›‹™^XÝ]Jˆ•TUH˜XÚÙÜ›Ý[™Ú›ØœÈÑU\˜[\×ÚœÛÛˆHÈÒT‘HYHÈ‹ˆ
[˜ÛÙWÚ›Ø—Ü\˜[\Ê\˜[\ÊK›Ø—ÚY
Kˆ
BˆX\š×Ø˜XÚÙÜ›Ý[™Ú›Ø—ÝØZ][™×Ø\WÜÛÝ
ˆÛÛ›‹ˆ›Ø—ÚYˆ™^Ø][\Ø][™^Ø][\Ø]ˆØZ]Ü™X\ÛÛH™]X˜\ÙWÛØÚÙY‹ˆ\œ›Ü—ÛY\ÜØYÙOJˆ”ÔS]H[\Ü˜\šX[Y[HØÝ\YNÈ›Ý˜H[]]˜H]]ÛX]XØH‚ˆˆžØ][\ßKÍ‹ˆ‚ˆ
Kˆ™\Ý[^ÂˆœÝ]\ÈŽˆØZ][™×Ø\WÜÛÝ‹ˆØZ]Ü™X\ÛÛˆŽˆ™]X˜\ÙWÛØÚÙY‹ˆœ™]žWØ][\Žˆ][\Ëˆ›™^Ø][\Ø]ŽˆÙ\šX[^™WØ˜XÚÙÜ›Ý[™Ú›Ø—Ý[Y\Ý[\
ˆ™^Ø][\Ø]ˆ
KˆœÝÜYÜ™X\ÛÛˆŽˆ™]X˜\ÙWÛØÚÙY‹ˆKˆ
BˆØÚY[WØ˜XÚÙÜ›Ý[™Ú›ØŠ\›Ø—ÚY[—Ù]O[™^Ø][\Ø]
Bˆ™]\›ˆYB‚‚™Yˆ[—Ø˜XÚÙÜ›Ý[™Ú›ØŠ\ˆ›\ÚË›Ø—ÚYˆ[
HOˆ›Û™N‚ˆÚ]\˜\ØÛÛ^

N‚ˆÚ]ÛÜÚ[™ÊÙ]ÙŠ\˜ÛÛ™šYÖÈ‘UPTÑH—JJH\ÈÛÛ›Ž‚ˆ›ØˆHÛÛ›‹™^XÝ]J”ÑSPÕ
ˆ”“ÓH˜XÚÙÜ›Ý[™Ú›ØœÈÒT‘HYHÈ‹
›Ø—ÚY
JK™™]ÚÛ™J
BˆYˆ›Øˆ\È›Û™N‚ˆÝ\œ™[Ø\›ÙÙÙ\‹™\œ›ÜŠ˜XÚÙÜ›Ý[™›Øˆ	\È›Ý›Ý[™‹›Ø—ÚY
Bˆ™]\›‚ˆžN‚ˆYˆ›ÝX\š×Ø˜XÚÙÜ›Ý[™Ú›Ø—Ü[›š[™ÊÛÛ›‹›Ø—ÚY
N‚ˆÝ\œ™[Ø\›ÙÙÙ\‹š[™›Ê˜XÚÙÜ›Ý[™›Øˆ	\ÈÚÚ\Y™XØ]\ÙH]\È›ÈÛ™Ù\ˆ[™[™È‹›Ø—ÚY
Bˆ™]\›‚ˆ^Ù\Ü[]LË“Ü\˜][Û˜[\œ›Üˆ\È^Î‚ˆYˆ›Ý\×Ý˜[œÚY[ÜÜ[]WÛØÚÊ^ÊN‚ˆ˜Z\ÙBˆ\˜[\ÈHXÛÙWÚ›Ø—Ü\˜[\Ê›Ø–Èœ\˜[\×ÚœÛÛˆ—JBˆYˆ›ÝY™\—Ø˜XÚÙÜ›Ý[™Ú›Ø—ØY\—Ù]X˜\ÙWÛØÚÊˆÛÛ›‹ˆ\ˆ›Ø—ÚYˆ\˜[\Ëˆ
N‚ˆ˜Z\ÙBˆ™]\›‚ˆžN‚ˆ\˜[\ÈHœÛÛ‹›ØYÊ›Ø–Èœ\˜[\×ÚœÛÛˆ—HÜˆžßHŠBˆÝ\œ™[Ø\›ÙÙÙ\‹š[™›Ê˜XÚÙÜ›Ý[™›Øˆ	\ÈÝ\Yˆ	\È‹›Ø—ÚY›Ø–Èš›Ø—Ý\H—JBˆÚ]›ÙXÝ[Û—ÚÜWØØ[ØÛÛ^
ˆ›Ø—ÚYZ›Ø—ÚYˆ›Ø—Ý\O\ÝŠ›Ø–Èš›Ø—Ý\H—JKˆ
N‚ˆ™\Ý[H[—Ø˜XÚÙÜ›Ý[™Ú›Ø—Ü^[ØY
ÛÛ›‹ÝŠ›Ø–Èš›Ø—Ý\H—JK\˜[\ÊBˆX\š×Ø˜XÚÙÜ›Ý[™Ú›Ø—ÜÝXØÙ\ÜÊÛÛ›‹›Ø—ÚY™\Ý[
BˆÝ\œ™[Ø\›ÙÙÙ\‹š[™›Ê˜XÚÙÜ›Ý[™›Øˆ	\ÈÛÛ\]Yˆ	\È‹›Ø—ÚY›Ø–Èš›Ø—Ý\H—JBˆ^Ù\\TÛÝ[˜]˜Z[X›Q\œ›Üˆ\È^Î‚ˆ™\Ý[HÂˆ
ŠŠˆ^Ëš›Ø—Ü™\Ý[ˆYˆ\Ú[œÝ[˜ÙJ^Ëš›Ø—Ü™\Ý[XÝ
Bˆ[ÙHßBˆ
KˆœÝ]\ÈŽˆØZ][™×Ø\WÜÛÝ‹ˆœ›ÝšY\ˆŽˆ^Ëœ›ÝšY\‹ˆ˜\WØ\™XHŽˆ^Ë˜\WØ\™XKˆØZ]Ü™X\ÛÛˆŽˆ^ËØZ]Ü™X\ÛÛ‹ˆ›™^Ø][\Ø]ŽˆÙ\šX[^™WØ˜XÚÙÜ›Ý[™Ú›Ø—Ý[Y\Ý[\
ˆ^Ë›™^Ø][\Ø]ˆ
KˆœÝÜYÜ™X\ÛÛˆŽˆ^Ë›Y\ÜØYÙKˆBˆX\š×Ø˜XÚÙÜ›Ý[™Ú›Ø—ÝØZ][™×Ø\WÜÛÝ
ˆÛÛ›‹ˆ›Ø—ÚYˆ™^Ø][\Ø]Y^Ë›™^Ø][\Ø]ˆØZ]Ü™X\ÛÛY^ËØZ]Ü™X\ÛÛ‹ˆ\œ›Ü—ÛY\ÜØYÙOY^Ë›Y\ÜØYÙKˆ™\Ý[\™\Ý[ˆ
BˆØÚY[WØ˜XÚÙÜ›Ý[™Ú›ØŠ\›Ø—ÚY[—Ù]OY^Ë›™^Ø][\Ø]
Bˆ^Ù\\T˜]S[Z]\œ›Üˆ\È^Î‚ˆÝ\œ™[Ø\›ÙÙÙ\‹š[™›Êˆ˜XÚÙÜ›Ý[™›Øˆ	\ÈØZ][™È›Üˆ	\È	\È˜]H[Z][[	\È‹ˆ›Ø—ÚYˆ^Ëœ›ÝšY\‹ˆ^Ë˜\™XKˆ^Ë˜ÛÛÛÝÛ—Ý[[ˆ
Bˆ\X[Ü™\Ý[HÙ]]Š^Ëš›Ø—Ü™\Ý[‹ßJBˆ™\Ý[HÂˆ
ŠŠ\X[Ü™\Ý[Yˆ\Ú[œÝ[˜ÙJ\X[Ü™\Ý[XÝ
H[ÙHßJKˆœÝ]\ÈŽˆØZ][™×Ü˜]WÛ[Z]‹ˆœ›ÝšY\ˆŽˆ^Ëœ›ÝšY\‹ˆ˜\WØ\™XHŽˆ^Ë˜\™XKˆ˜ÛÛÛÝÛ—Ý[[ŽˆÙ\šX[^™WØ˜XÚÙÜ›Ý[™Ú›Ø—Ý[Y\Ý[\
ˆ^Ë˜ÛÛÛÝÛ—Ý[[ˆ
Kˆ›™^Ø][\Ø]ŽˆÙ\šX[^™WØ˜XÚÙÜ›Ý[™Ú›Ø—Ý[Y\Ý[\
ˆ^Ë˜ÛÛÛÝÛ—Ý[[ˆ
KˆœÝÜYÜ™X\ÛÛˆŽˆ^Ë›Y\ÜØYÙKˆBˆYˆÝŠ›Ø–Èš›Ø—Ý\H—JHOH™\Ú[ÛœÛÛ\—Û[ÛØÛÜÙHŽ‚ˆ™\ÜÛ[ÛHÝŠ\˜[\Ë™Ù]
œ™\ÜÛ[ÛŠHÜˆˆŠBˆ™\Ý[œÙ]Y˜][
›[Û‹™\ÜÛ[Û
BˆYˆ›Ý™\Ý[™Ù]
œÝ]\×Ø™Y›Ü™HŠH[™™\ÜÛ[Û‚ˆžN‚ˆ[ÛÜÝ\H]][YKœÝœ[YJ™\ÜÛ[Û‰VKI[HŠK™]J
Bˆ™Y™\™[˜ÙWÙ]HHÝ\œ™[Û\Ø›Û—Ù]J
Bˆ\ÜÙ]ÈHÙ]Ù\Ú[ÛœÛÛ\—Ü\™›Ü›X[˜ÙWØ\ÜÙ]ÊˆÛÛ›‹ˆÝŠ\˜[\Ë™Ù]
œ›ÝšY\ˆŠHÜˆS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓTŠKˆ
Bˆ]X[]Y\ÈHÂˆ[
\ÜÙ]È˜\ÜÙ]ÚY—JNˆ]˜[X]WÛØØ[Û[ÛWÜ›ÙXÝ[Û—Ü]X[]JˆÛÛ›‹ˆ\ÜÙ]ÚYZ[
\ÜÙ]È˜\ÜÙ]ÚY—JKˆ›ÝšY\\ÝŠˆ\˜[\Ë™Ù]
œ›ÝšY\ˆŠBˆÜˆS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‚ˆ
Kˆ[ÛÜÝ\[[ÛÜÝ\ˆ™Y™\™[˜ÙWÙ]O\™Y™\™[˜ÙWÙ]Kˆ
Bˆ›Üˆ\ÜÙ][ˆ\ÜÙ]ÂˆBˆÝ]\ÈHÛÝ[Û[ÛWÜ›ÙXÝ[Û—ÜÝ]\Ê]X[]Y\ÊBˆ™\Ý[ÈœÝ]\×Ø™Y›Ü™H—HHÝ]\Âˆ™\Ý[ÈœÝ]\×ØY\ˆ—HHÝ]\Âˆ™\Ý[È˜\ÜÙ]×Ù]˜[X]Y—HH[Š]X[]Y\ÊBˆ™\Ý[È˜\ÜÙ]×ØY™™XÝY—HHÝ[Jˆ]X[]KœÝ]\È[ˆÈœ\X[‹›Z\ÜÚ[™È‹˜ÛÛ™›XÝŸBˆ›Üˆ]X[]H[ˆ]X[]Y\Ë˜[Y\Ê
Bˆ
Bˆ^Ù\^Ù\[ÛŽ‚ˆ™\Ý[œÙ]Y˜][
œÝ]\×Ø™Y›Ü™H‹ßJBˆ™\Ý[œÙ]Y˜][
œÝ]\×ØY\ˆ‹ßJBˆ[ÙN‚ˆ™\Ý[œÙ]Y˜][
œÝ]\×Ø™Y›Ü™H‹ßJBˆ™\Ý[œÙ]Y˜][
œÝ]\×ØY\ˆ‹ßJBˆ™\Ý[œÙ]Y˜][
˜\WØØ[×Ý\ÙY‹
Bˆ™\Ý[œÙ]Y˜][
˜\WØØ[×Ø][\Y‹
Bˆ™\Ý[œÙ]Y˜][
˜\ÜÙ]×ØY™™XÝY‹
BˆX\š×Ø˜XÚÙÜ›Ý[™Ú›Ø—ÝØZ][™×Ü˜]WÛ[Z]
ˆÛÛ›‹ˆ›Ø—ÚYˆ™^Ø][\Ø]Y^Ë˜ÛÛÛÝÛ—Ý[[ˆ\œ›Ü—ÛY\ÜØYÙOY^Ë›Y\ÜØYÙKˆ™\Ý[\™\Ý[ˆ
BˆØÚY[WØ˜XÚÙÜ›Ý[™Ú›ØŠ\›Ø—ÚY[—Ù]OY^Ë˜ÛÛÛÝÛ—Ý[[
Bˆ^Ù\Ü[]LË“Ü\˜][Û˜[\œ›Üˆ\È^Î‚ˆYˆ\×Ý˜[œÚY[ÜÜ[]WÛØÚÊ^ÊH[™Y™\—Ø˜XÚÙÜ›Ý[™Ú›Ø—ØY\—Ù]X˜\ÙWÛØÚÊˆÛÛ›‹ˆ\ˆ›Ø—ÚYˆ\˜[\Ëˆ
N‚ˆÝ\œ™[Ø\›ÙÙÙ\‹š[™›Êˆ˜XÚÙÜ›Ý[™›Øˆ	\ÈY™\œ™YY\ˆ˜[œÚY[ÔS]HØÚÈ‹ˆ›Ø—ÚYˆ
Bˆ™]\›‚ˆÝ\œ™[Ø\›ÙÙÙ\‹™^Ù\[ÛŠˆ˜XÚÙÜ›Ý[™›Øˆ	\È˜Z[YY\ˆÔS]HØÚÈ™]šY\Îˆ	\È‹ˆ›Ø—ÚYˆ›Ø–Èš›Ø—Ý\H—Kˆ
BˆX\š×Ø˜XÚÙÜ›Ý[™Ú›Ø—Ù˜Z[Y
ÛÛ›‹›Ø—ÚYÝŠ^ÊJBˆ^Ù\^Ù\[Ûˆ\È^Î‚ˆÝ\œ™[Ø\›ÙÙÙ\‹™^Ù\[ÛŠ˜XÚÙÜ›Ý[™›Øˆ	\È˜Z[Yˆ	\È‹›Ø—ÚY›Ø–Èš›Ø—Ý\H—JBˆX\š×Ø˜XÚÙÜ›Ý[™Ú›Ø—Ù˜Z[Y
ÛÛ›‹›Ø—ÚYÝŠ^ÊJB‚‚™Yˆ[—Ø˜XÚÙÜ›Ý[™Ú›Ø—Ü^[ØY
ÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹›Ø—Ý\NˆÝ‹\˜[\ÎˆXÝÜÝ‹[žWJHOˆXÝÜÝ‹[žWN‚ˆ›ÝšY\‹\WØ\™XHH˜XÚÙÜ›Ý[™Ú›Ø—Ø\WÜØÛÜJ›Ø—Ý\K\˜[\ÊBˆYˆ›ÝšY\ˆ[™\WØ\™XN‚ˆ™\]Z\™WÛ›ÝÚ[—ØÛÛÛÝÛŠÛÛ›‹›ÝšY\‹\WØ\™XJBˆ™XÛÜ™Ø\WØ][\
ÛÛ›‹›ÝšY\‹\WØ\™XJB‚ˆYˆ›Ø—Ý\HOH™\Ú[ÛœÛÛ\—ÜÝ]WÜÞ[˜ÈŽ‚ˆ™\Ý[H[—Ù\Ú[ÛœÛÛ\—ÜÞ[˜ÊˆÛÛ›‹ˆÝŠ\˜[\Ë™Ù]
œ›ÝšY\ˆŠHÜˆS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓTŠKˆšYÙÙ\—Ý\O\ÝŠ\˜[\Ë™Ù]
šYÙÙ\—Ý\HŠHÜˆ›X[X[Ø˜XÚÙÜ›Ý[™ŠKˆ
Bˆ™XÛÜ™Ø\WÜÝXØÙ\ÜÊÛÛ›‹S•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹TWÐT‘PWÔÕUJBˆ™]\›ˆ™\Ý[‚ˆYˆ›Ø—Ý\HOHœÚYÙ[™\™ÞWÜÝ]WÜÞ[˜ÈŽ‚ˆ™\Ý[H[—ÜÚYÙ[™\™ÞWÜÞ[˜ÊˆÛÛ›‹ˆÝŠ\˜[\Ë™Ù]
œ›ÝšY\ˆŠHÜˆS•QÔUSÓ—Ô“Õ’QT—ÔÒQÑS‘T‘ÖJKˆšYÙÙ\—Ý\O\ÝŠ\˜[\Ë™Ù]
šYÙÙ\—Ý\HŠHÜˆ›X[X[Ø˜XÚÙÜ›Ý[™ŠKˆ
Bˆ™XÛÜ™Ø\WÜÝXØÙ\ÜÊÛÛ›‹S•QÔUSÓ—Ô“Õ’QT—ÔÒQÑS‘T‘ÖKTWÐT‘PWÔÕUJBˆ™]\›ˆ™\Ý[‚ˆYˆ›Ø—Ý\HOH™\Ú[ÛœÛÛ\—Ü›ÙXÝ[Û—ÜÞ[˜ÈŽ‚ˆ\™Ù]Ù]HH\œÙWÙ]WÝ˜[YJÝŠ\˜[\Ë™Ù]
\™Ù]Ù]HŠHÜˆˆŠJBˆYˆ\™Ù]Ù]H\È›Û™N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ‘]H[˜[YH\˜HÞ[˜ÈH›ÙXØ[ËˆŠBˆ™\Ý[H[—Ù\Ú[ÛœÛÛ\—Ü›ÙXÝ[Û—ÜÞ[˜ÊˆÛÛ›‹ˆ›ÝšY\\ÝŠ\˜[\Ë™Ù]
œ›ÝšY\ˆŠHÜˆS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓTŠKˆ\™Ù]Ù]O]\™Ù]Ù]Kˆ\š[ÙÝ\O\ÝŠ\˜[\Ë™Ù]
œ\š[ÙÝ\HŠHÜˆ™^HŠKˆ
Bˆ™XÛÜ™Ø\WÜÝXØÙ\ÜÊÛÛ›‹S•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹TWÐT‘PWÔ“ÑPÕSÓŠBˆ™]\›ˆ™\Ý[‚ˆYˆ›Ø—Ý\HOHœ\™›Ü›X[˜ÙWÜ™Y™\™[˜ÙWÜ™XØ[Ý[][ÛˆŽ‚ˆ\š[ÙÙ]HH\œÙWÙ]WÝ˜[YJÝŠ\˜[\Ë™Ù]
œ\š[ÙÙ]HŠHÜˆˆŠJBˆYˆ\š[ÙÙ]H\È›Û™N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ‘]H[˜[YH\˜H™XØ[Ý[ÈH™Y™\™[˜ÚX\ËˆŠBˆ\ÜÙ]ÚYÝ˜[YHH\˜[\Ë™Ù]
˜\ÜÙ]ÚYŠBˆ™]\›ˆ™XØ[Ý[]WÜ\™›Ü›X[˜ÙWÜ™Y™\™[˜Ù\ÊˆÛÛ›‹ˆ\š[ÙÝ\O\ÝŠ\˜[\Ë™Ù]
œ\š[ÙÝ\HŠHÜˆ™^HŠKˆ\š[ÙÙ]O\\š[ÙÙ]Kˆ\ÜÙ]ÚYZ[
\ÜÙ]ÚYÝ˜[YJHYˆ\ÜÙ]ÚYÝ˜[YH[ÙH›Û™Kˆ›ÝšY\\ÝŠ\˜[\Ë™Ù]
œ›ÝšY\ˆŠHÜˆS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓTŠKˆ
B‚ˆYˆ›Ø—Ý\HOH™\Ú[ÛœÛÛ\—Ü›ÙXÝ[Û—Ø˜XÚÙš[Ž‚ˆ]WÙœ›ÛHH\œÙWÙ]WÝ˜[YJÝŠ\˜[\Ë™Ù]
™]WÙœ›ÛHŠHÜˆˆŠJBˆ]WÝÈH\œÙWÙ]WÝ˜[YJÝŠ\˜[\Ë™Ù]
™]WÝÈŠHÜˆˆŠJBˆ\ÜÙ]ÚYÝ˜[YHH\˜[\Ë™Ù]
˜\ÜÙ]ÚYŠBˆ™\Ý[H[—Ù\Ú[ÛœÛÛ\—Ü›ÙXÝ[Û—Ø˜XÚÙš[
ˆÛÛ›‹ˆ›ÝšY\\ÝŠ\˜[\Ë™Ù]
œ›ÝšY\ˆŠHÜˆS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓTŠKˆ\š[ÙÝ\O\ÝŠ\˜[\Ë™Ù]
œ\š[ÙÝ\HŠHÜˆ™^HŠKˆœ›ÛWÞYX\Z[
\˜[\ÖÈ™œ›ÛWÞYX\ˆ—JKˆ×ÞYX\Z[
\˜[\ÖÈ×ÞYX\ˆ—JKˆ\ÜÙ]ÚYZ[
\ÜÙ]ÚYÝ˜[YJHYˆ\ÜÙ]ÚYÝ˜[YH[ÙH›Û™Kˆ]WÙœ›ÛOY]WÙœ›ÛKˆ]WÝÏY]WÝËˆX^Ø\WØØ[ÏZ[
\˜[\Ë™Ù]
›X^Ø\WØØ[ÈŠHÜˆ•TÒSÓ”ÓÓT—ÔT‘“Ô“PSÑWÓPVÐTWÐÐSÊKˆ
Bˆ™XÛÜ™Ø\WÜÝXØÙ\ÜÊÛÛ›‹S•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹TWÐT‘PWÔ“ÑPÕSÓŠBˆ™]\›ˆ™\Ý[‚ˆYˆ›Ø—Ý\H[ˆÂˆ™\Ú[ÛœÛÛ\—Ú[™\\—Ø]˜Z[Xš[]WØ˜XÚÙš[‹ˆ™\Ú[ÛœÛÛ\—Ü™\ÜÝØ]Ü™\]Y\Ý‹ˆN‚ˆœ›ÛWÙ]HH\œÙWÙ]WÝ˜[YJÝŠ\˜[\Ë™Ù]
™œ›ÛWÙ]HŠHÜˆˆŠJBˆ×Ù]HH\œÙWÙ]WÝ˜[YJÝŠ\˜[\Ë™Ù]
×Ù]HŠHÜˆˆŠJBˆYˆœ›ÛWÙ]H\È›Û™HÜˆ×Ù]H\È›Û™HÜˆœ›ÛWÙ]Hˆ×Ù]N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ’[\˜[È[˜[YÈ\˜H˜XÚÙš[ÐUˆŠBˆ™\Ý[H[—Ù\Ú[ÛœÛÛ\—Ú[™\\—Ø]˜Z[Xš[]WØ˜XÚÙš[
ˆÛÛ›‹ˆœ›ÛWÙ]OYœ›ÛWÙ]Kˆ×Ù]O]×Ù]Kˆ
Bˆ™XÛÜ™Ø\WÜÝXØÙ\ÜÊÛÛ›‹S•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹TWÐT‘PWÑPQÓ“ÔÕPÔÊBˆ™]\›ˆ™\Ý[‚ˆYˆ›Ø—Ý\HOH™\Ú[ÛœÛÛ\—Ü™X[[YWÛX]\šX[^™WØÛX[\Ž‚ˆ™][[Û—Ù^\ÈH\œÙWÙ[—ÜÜÚ]]™WÚ[
ˆ‘•TÒSÓ”ÓÓT—Ô‘PSSQWÔÓTÒÕÔ‘US•SÓ—ÑVTÈ‹ˆQUSÑ•TÒSÓ”ÓÓT—Ô‘PSSQWÔÓTÒÕÔ‘US•SÓ—ÑVTËˆ
Bˆ™]\›ˆÛX[\Ü™X[[YWÜÛ˜\ÚÝÜ^[ØYÊˆÛÛ›‹ˆ›ÝšY\RS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹ˆ™][[Û—Ù^\Ï\™][[Û—Ù^\Ëˆ™Y™\™[˜ÙWÙ]OXÝ\œ™[Û\Ø›Û—Ù]J
Kˆ
B‚ˆYˆ›Ø—Ý\HOH™\Ú[ÛœÛÛ\—Û[ÛØÞXÛHŽ‚ˆ˜]×Ø\ÜÙ]ÚYÈH\˜[\Ë™Ù]
˜\ÜÙ]ÚYÈŠHÜˆ×Bˆ\ÜÙ]ÚYÈHÚ[
˜[YJH›Üˆ˜[YH[ˆ˜]×Ø\ÜÙ]ÚYÈYˆÝŠ˜[YJKš\ÙYÚ]

WBˆ™\Ý[H[—Ù\Ú[ÛœÛÛ\—Û[ÛØÞXÛJˆÛÛ›‹ˆ›ÝšY\\ÝŠ\˜[\Ë™Ù]
œ›ÝšY\ˆŠHÜˆS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓTŠKˆ™\ÜÛ[Û\ÝŠ\˜[\Ë™Ù]
œ™\ÜÛ[ÛŠHÜˆ]KÙ^J
KœÝ™[YJ‰VKI[HŠJKˆ\ÜÙ]ÚYÏX\ÜÙ]ÚYËˆ
Bˆ™XÛÜ™Ø\WÜÝXØÙ\ÜÊÛÛ›‹S•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹TWÐT‘PWÔ“ÑPÕSÓŠBˆ™]\›ˆ™\Ý[‚ˆYˆ›Ø—Ý\HOH™\Ú[ÛœÛÛ\—Û[ÛØÛÜÙHŽ‚ˆ™\Ý[H[—Ù\Ú[ÛœÛÛ\—Û[ÛØÛÜÙJˆÛÛ›‹ˆ›ÝšY\\ÝŠ\˜[\Ë™Ù]
œ›ÝšY\ˆŠHÜˆS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓTŠKˆ™\ÜÛ[Û\ÝŠ\˜[\Ë™Ù]
œ™\ÜÛ[ÛŠHÜˆˆŠKˆ™Y™\™[˜ÙWÙ]OXÝ\œ™[Û\Ø›Û—Ù]J
Kˆ
Bˆ™XÛÜ™Ø\WÜÝXØÙ\ÜÊÛÛ›‹S•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹TWÐT‘PWÔ“ÑPÕSÓŠBˆ™]\›ˆ™\Ý[‚ˆYˆ›Ø—Ý\HOH™\Ú[ÛœÛÛ\—Ü™\ÜÜ›ÙXÝ[Û—Ü™\]Y\ÝŽ‚ˆ˜]×Ø\ÜÙ]ÚYÈH\˜[\Ë™Ù]
˜\ÜÙ]ÚYÈŠHÜˆ×Bˆ\ÜÙ]ÚYÈHÚ[
˜[YJH›Üˆ˜[YH[ˆ˜]×Ø\ÜÙ]ÚYÈYˆÝŠ˜[YJKš\ÙYÚ]

WBˆ™\Ý[H[—Ù\Ú[ÛœÛÛ\—Û[ÛØÛÜÙJˆÛÛ›‹ˆ›ÝšY\\ÝŠ\˜[\Ë™Ù]
œ›ÝšY\ˆŠHÜˆS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓTŠKˆ™\ÜÛ[Û\ÝŠ\˜[\Ë™Ù]
œ™\ÜÛ[ÛŠHÜˆˆŠKˆ\ÜÙ]ÚYÏX\ÜÙ]ÚYËˆ™Y™\™[˜ÙWÙ]OXÝ\œ™[Û\Ø›Û—Ù]J
Kˆ
Bˆ™XÛÜ™Ø\WÜÝXØÙ\ÜÊÛÛ›‹S•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹TWÐT‘PWÔ“ÑPÕSÓŠBˆ™]\›ˆ™\Ý[‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠˆ•\ÈH›Øˆ\ØÛÛšXÚYÎˆÚ›Ø—Ý\_HŠB‚‚™Yˆ˜XÚÙÜ›Ý[™Ú›Ø—Ø\WÜØÛÜJ›Ø—Ý\NˆÝ‹\˜[\ÎˆXÝÜÝ‹[žWJHOˆ\VÜÝ‹Ý—N‚ˆ›ÝšY\ˆHÝŠ\˜[\Ë™Ù]
œ›ÝšY\ˆŠHÜˆˆŠBˆYˆ›Ø—Ý\HOH™\Ú[ÛœÛÛ\—ÜÝ]WÜÞ[˜ÈŽ‚ˆ™]\›ˆ›ÝšY\ˆÜˆS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹TWÐT‘PWÔÕUBˆYˆ›Ø—Ý\H[ˆÂˆ™\Ú[ÛœÛÛ\—Ü›ÙXÝ[Û—ÜÞ[˜È‹ˆ™\Ú[ÛœÛÛ\—Ü›ÙXÝ[Û—Ø˜XÚÙš[‹ˆ™\Ú[ÛœÛÛ\—Û[ÛØÞXÛH‹ˆ™\Ú[ÛœÛÛ\—Û[ÛØÛÜÙH‹ˆ™\Ú[ÛœÛÛ\—Ü™\ÜÜ›ÙXÝ[Û—Ü™\]Y\Ý‹ˆN‚ˆ™]\›ˆ›ÝšY\ˆÜˆS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹TWÐT‘PWÔ“ÑPÕSÓ‚ˆYˆ›Ø—Ý\H[ˆÂˆ™\Ú[ÛœÛÛ\—Ú[™\\—Ø]˜Z[Xš[]WØ˜XÚÙš[‹ˆ™\Ú[ÛœÛÛ\—Ü™\ÜÝØ]Ü™\]Y\Ý‹ˆN‚ˆ™]\›ˆ›ÝšY\ˆÜˆS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹TWÐT‘PWÑPQÓ“ÔÕPÔÂˆYˆ›Ø—Ý\HOHœÚYÙ[™\™ÞWÜÝ]WÜÞ[˜ÈŽ‚ˆ™]\›ˆ›ÝšY\ˆÜˆS•QÔUSÓ—Ô“Õ’QT—ÔÒQÑS‘T‘ÖKTWÐT‘PWÔÕUBˆ™]\›ˆˆ‹ˆ‚‚‚™Yˆ\Ú[ÛœÛÛ\—Ø\™XWÙ›Ü—Ý\›
\›ˆÝŠHOˆÝŽ‚ˆÝÙ\™YH\››ÝÙ\Š
BˆYˆ™Ù]Ü\Ý][Û™^Hˆ[ˆÝÙ\™YÜˆ™Ù]Ü\Ý][Û›[Ûˆ[ˆÝÙ\™Y‚ˆ™]\›ˆTWÐT‘PWÔ“ÑPÕSÓ‚ˆYˆ[žJÚÙ[ˆ[ˆÝÙ\™Y›ÜˆÚÙ[ˆ[ˆ
™Ù]]›\Ý‹™Ù]]œ™X[ÜH‹™Ù]]š\ÝÜžZÜH‹™Ù][\›[\ÝŠJN‚ˆ™]\›ˆTWÐT‘PWÑPQÓ“ÔÕPÔÂˆ™]\›ˆTWÐT‘PWÔÕUB‚‚™Yˆ˜Z\ÙWÙ\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]
^Îˆ\Ú[Û”ÛÛ\”˜]S[Z]\œ›Ü‹\WØ\™XNˆÝŠHOˆ›Û™N‚ˆÛÛ›ˆHØÝ\œ™[Ø\Ù—ØÛÛ›™XÝ[ÛŠ
BˆžN‚ˆ[[HX\š×Ù\Ú[ÛœÛÛ\—Ø\WØÛÛÛÝÛŠÛÛ›‹\WØ\™XK™X\ÛÛ\ÝŠ^ÊJBˆ™X\ÛÛˆHÙ]Ü›ÝšY\—ØÛÛÛÝÛ—Ü™X\ÛÛŠÛÛ›‹S•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹\WØ\™XJHYˆÛÛ›ˆ[ÙHÝŠ^ÊBˆ˜Z\ÙH\T˜]S[Z]\œ›ÜŠS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹\WØ\™XK[[™X\ÛÛŠHœ›ÛH^Âˆš[˜[N‚ˆYˆÛÛ›ˆ\È›Ý›Û™N‚ˆÛÛ›‹˜ÛÜÙJ
B‚‚™YˆÙ]Ü›ÝšY\—ØÛÛÛÝÛ—Ü™X\ÛÛŠˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ›ÝšY\ŽˆÝ‹ˆ\™XNˆÝ‹ˆ›Ý×Ý˜[YNˆ]][YH›Û™HH›Û™KŠHOˆÝŽ‚ˆ[[HXÝ]™WØÛÛÛÝÛ—Ý[[
ÛÛ›‹›ÝšY\‹\™XK›Ý×Ý˜[YHÜˆ]][YK››ÝÊ
JBˆYˆ[[\È›Û™N‚ˆ™]\›ˆˆ‚ˆ™[XZ[š[™×ÜÙXÛÛ™ÈH[

[[H
›Ý×Ý˜[YHÜˆ]][YK››ÝÊ
JJKÝ[ÜÙXÛÛ™Ê
JBˆ™[XZ[š[™×ÛZ[]\ÈHX^
K
™[XZ[š[™×ÜÙXÛÛ™È
ÈNJHËÈŒ
Bˆ™]\›ˆ
ˆˆžÜ›ÝšY\ŸH[H\Ü\˜HÜˆ[Z]HHTKˆ‚ˆˆ“›Ý˜H[]]˜H\ÜÛš]™[\ÜÈÝ[[š\ÛÙ›Ü›X]
[Y\ÜXÏIÛZ[]\ÉÊ_H
Ü™[XZ[š[™×ÛZ[]\ßHZ[ŠKˆ‚ˆ
B‚‚™Yˆ\ÝØ\WØØ[ÜÝ]\ÊÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[ÛŠHOˆ\ÝÙXÝÜÝ‹[žWWN‚ˆ[œÝ\™WØ\WØØ[ÜÝ]WÜØÚ[XJÛÛ›ŠBˆ›ÝÜÈH]Y\žWØ[
ˆÛÛ›‹ˆˆˆ‚ˆÑSPÕ›ÝšY\‹\WØ\™XKÛÛÛÝÛ—Ý[[\ÝÙ\œ›Ü‹\ÝÜÝXØÙ\Ü×Ø]ˆ\ÝØ][\Ø]\ÝØ[\Ø]\]YØ]ˆ”“ÓH\WØØ[ÜÝ]BˆÔ‘Tˆ–H›ÝšY\‹\WØ\™XBˆˆˆ‹ˆ
Bˆ›Ý×Ý˜[YHH]][YK››ÝÊ
BˆÝ]\Îˆ\ÝÙXÝÜÝ‹[žWWHH×Bˆ›Üˆ›ÝÈ[ˆ›ÝÜÎ‚ˆÝ]HHXÝ
›ÝÊBˆ[[HXÝ]™WØÛÛÛÝÛ—Ý[[
ÛÛ›‹ÝŠ›ÝÖÈœ›ÝšY\ˆ—JKÝŠ›ÝÖÈ˜\WØ\™XH—JK›Ý×Ý˜[YJBˆÝ]VÈ˜ÛÛÛÝÛ—ØXÝ]™H—HH›ÛÛ
[[
BˆÝ]VÈ˜ÛÛÛÝÛ—ÛY\ÜØYÙH—HHÙ]Ü›ÝšY\—ØÛÛÛÝÛ—Ü™X\ÛÛŠˆÛÛ›‹ˆÝŠ›ÝÖÈœ›ÝšY\ˆ—JKˆÝŠ›ÝÖÈ˜\WØ\™XH—JKˆ›Ý×Ý˜[YKˆ
BˆÝ]\Ë˜\[™
Ý]JBˆ™]\›ˆÝ]\Â‚‚™YˆZ[Ü›ÙXÝ[Û—Ø\WÜ]Y]YWÛØœÙ\˜Xš[]JˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ŠHOˆ\ÝÙXÝÜÝ‹[žWWN‚ˆ\Ú[ÛœÛÛ\—ØÛÛ™šYÈHÙ]Ú[YÜ˜][Û—ØÛÛ™šYÊˆÛÛ›‹ˆS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹ˆ
BˆYˆ\Ú[ÛœÛÛ\—ØÛÛ™šYÈ\È›Ý›Û™N‚ˆ[™Ú[ÈHÙ]Ù\Ú[ÛœÛÛ\—Ù[™Ú[ØÛÛ™šYÊ\Ú[ÛœÛÛ\—ØÛÛ™šYÊBˆ\Ú[Û—ØXØÛÝ[ÚÙ^HH\Ú[ÛœÛÛ\—Ü›ÙXÝ[Û—ØXØÛÝ[ÚÙ^Jˆ\Ú[ÛœÛÛ\—ØÛÛ™šYËˆ[™Ú[ÖÈ™^WÚÜWÙ[™Ú[—Kˆ
Bˆ[œÝ\™WØ\WÜ]Y]YWÜÝ]JˆÛÛ›‹ˆ›ÝšY\RS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹›ÝÙ\Š
KˆXØÛÝ[ÚÙ^WÝ˜[YOY\Ú[Û—ØXØÛÝ[ÚÙ^Kˆ\WØ\™XOT“ÑPÕSÓ—ÒÔWÐT‘PKˆÛXÞOY\Ú[ÛœÛÛ\—Ü›ÙXÝ[Û—ÚÜWÜÛXÞJ
Kˆ
Bˆ[œÝ\™WØ\WÜ]Y]YWÜÝ]JˆÛÛ›‹ˆ›ÝšY\RS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹›ÝÙ\Š
KˆXØÛÝ[ÚÙ^WÝ˜[YOY\Ú[Û—ØXØÛÝ[ÚÙ^Kˆ\WØ\™XOUÐUÒTÕÔ–WÐT‘PKˆÛXÞOY\Ú[ÛœÛÛ\—ÝØ]ÜÛXÞJ
Kˆ
Bˆ\Ú[Û—ØXØÛÝ[ÜÝ]HHÙ]ØXØÛÝ[Ü]Y]YWÜÝ]JˆÛÛ›‹ˆ›ÝšY\RS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹›ÝÙ\Š
KˆXØÛÝ[ÚÙ^WÝ˜[YOY\Ú[Û—ØXØÛÝ[ÚÙ^Kˆ
Bˆ[ÙN‚ˆ\Ú[Û—ØXØÛÝ[ÜÝ]HHßBˆÚYÙ[™\™ÞWØÛÛ™šYÈHÙ]Ú[YÜ˜][Û—ØÛÛ™šYÊˆÛÛ›‹ˆS•QÔUSÓ—Ô“Õ’QT—ÔÒQÑS‘T‘ÖKˆ
BˆYˆÚYÙ[™\™ÞWØÛÛ™šYÈ\È›Ý›Û™N‚ˆÚYÙ[™\™ÞWÙ[™Ú[HÝŠÚYÙ[™\™ÞWØÛÛ™šYÖÈ™[™\™ÞWÙ›Ý×Ù[™Ú[—HÜˆˆŠBˆ[œÝ\™WØ\WÜ]Y]YWÜÝ]JˆÛÛ›‹ˆ›ÝšY\RS•QÔUSÓ—Ô“Õ’QT—ÔÒQÑS‘T‘ÖK›ÝÙ\Š
KˆXØÛÝ[ÚÙ^WÝ˜[YO\›ÙXÝ[Û—Ø\WØXØÛÝ[ÚÙ^Jˆ›ÝšY\RS•QÔUSÓ—Ô“Õ’QT—ÔÒQÑS‘T‘ÖKˆ\Ù\›˜[YO\ÝŠÚYÙ[™\™ÞWØÛÛ™šYÖÈ\Ù\›˜[YH—HÜˆˆŠKˆ˜\ÙWÝ\›\ÝŠÚYÙ[™\™ÞWØÛÛ™šYÖÈ˜˜\ÙWÝ\›—HÜˆˆŠKˆ[™Ú[\ÚYÙ[™\™ÞWÙ[™Ú[ˆ
Kˆ\WØ\™XOT“ÑPÕSÓ—ÒÔWÐT‘PKˆÛXÞO\ÚYÙ[™\™ÞWÜ›ÙXÝ[Û—ÚÜWÜÛXÞJ
Kˆ
BˆØZ][™×Ü›ÝÜÈHÛÛ›‹™^XÝ]Jˆˆˆˆ‚ˆÑSPÕ›Ø—Ý\KÓÕS•

ŠHTÈÝ[ˆ”“ÓH˜XÚÙÜ›Ý[™Ú›ØœÂˆÒT‘H›Ø—Ý\HSˆ
È‹‹š›Ú[ŠÈˆ›ÜˆÈ[ˆ•TÒSÓ”ÓÓT—ÐPÒÑÔ“ÕS‘Ò“Ð—ÕTTÊ_JBˆS‘Ý]\ÈSˆ
	ÝØZ][™×Ø\WÜÛÝ	Ë	ÝØZ][™×Ü˜]WÛ[Z]	ÊBˆÔ“ÕT–H›Ø—Ý\Bˆˆˆ‹ˆ•TÒSÓ”ÓÓT—ÐPÒÑÔ“ÕS‘Ò“Ð—ÕTTËˆ
K™™]Ú[

BˆØZ][™×ØžWÝ\HHÂˆÝŠ][VÈš›Ø—Ý\H—JNˆ[
][VÈÝ[—JBˆ›Üˆ][H[ˆØZ][™×Ü›ÝÜÂˆBˆØZ][™×ØžWÜš[Üš]NˆXÝÚ[[HHßBˆ›Üˆ›Ø—Ý\KÛÝ[[ˆØZ][™×ØžWÝ\Kš][\Ê
N‚ˆš[Üš]HH›ÙXÝ[Û—Ú›Ø—Üš[Üš]J›Ø—Ý\JBˆØZ][™×ØžWÜš[Üš]VÜš[Üš]WHH
ˆØZ][™×ØžWÜš[Üš]K™Ù]
š[Üš]K
H
ÈÛÝ[ˆ
Bˆ™]\ÙYÜ™\ÜÜ™\]Y\ÝÈH[
ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕÓÕS•

ŠBˆ”“ÓH™\ÜÙ]WÜ™\]Y\ÝÙ]™[ÂˆÒT‘H™]\ÙYÙ^\Ý[™×Ú›ØˆHBˆˆˆ‚ˆ
K™™]ÚÛ™J
VÌBˆ
BˆØØ[WÜ™XØ[Ý[]YÙ^\ÈH[
ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕÓÕS•

ŠBˆ”“ÓH[Ø]˜Z[Xš[]WÜØ[\YÙZ[BˆÒT‘HÛÝ\˜ÙHH	Ü™X[[YWÜØ[\Y	Âˆˆˆ‚ˆ
K™™]ÚÛ™J
VÌBˆ
BˆØ[\YÜÝ]WÜ›ÝÜÈHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕÛÝ™\˜YÙWÜÝ]\ËÓÕS•

ŠHTÈÝ[ˆ”“ÓH[Ø]˜Z[Xš[]WÜØ[\YÙZ[BˆÔ“ÕT–HÛÝ™\˜YÙWÜÝ]\ÂˆÔ‘Tˆ–HÛÝ™\˜YÙWÜÝ]\Âˆˆˆ‚ˆ
K™™]Ú[

BˆØ[\YÜÝ]\ÈHÂˆÝŠ][VÈ˜ÛÝ™\˜YÙWÜÝ]\È—JNˆ[
][VÈÝ[—JBˆ›Üˆ][H[ˆØ[\YÜÝ]WÜ›ÝÜÂˆBˆ™X[ÝØ]Ù^\ÈH[
ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕÓÕS•

ŠBˆ”“ÓH[Ø]˜Z[Xš[]WÙZ[BˆÒT‘H˜[YÜÛÝÈˆS‘ÙZYÚYØ]˜Z[Xš[]WÜÝTÈ“Õ•Sˆˆˆ‚ˆ
K™™]ÚÛ™J
VÌBˆ
Bˆ›ÝÜÈH\ÝØ\WÜ]Y]YWÜÝ]\ÊÛÛ›ŠBˆ›Üˆ›ÝÈ[ˆ›ÝÜÎ‚ˆYˆ›ÝÖÈœ›ÝšY\ˆ—HOHS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹›ÝÙ\Š
N‚ˆ™[]˜[Ý\\ÈH
ˆ•TÒSÓ”ÓÓT—ÕÐUÒ“Ð—ÕTTÂˆYˆ›ÝÖÈ˜\WØ\™XH—HOHÐUÒTÕÔ–WÐT‘PBˆ[ÙH•TÒSÓ”ÓÓT—Ô“ÑPÕSÓ—Ò“Ð—ÕTTÂˆ
Bˆ›ÝÖÈØZ][™×Ú›ØœÈ—HHÝ[JˆØZ][™×ØžWÝ\K™Ù]
›Ø—Ý\K
Bˆ›Üˆ›Ø—Ý\H[ˆ™[]˜[Ý\\Âˆ
Bˆ›ÝÖÈØZ][™×Ú›Øœ×ØžWÜš[Üš]H—HHXÝ
ˆÛÜY
ØZ][™×ØžWÜš[Üš]Kš][\Ê
JBˆ
Bˆ›ÝÖÈ˜XØÛÝ[Û\ÝÍ×Ø]—HH\Ú[Û—ØXØÛÝ[ÜÝ]K™Ù]
ˆ›\ÝÍ×Ø]‚ˆ
Bˆ›ÝÖÈ˜XØÛÝ[ØÛÛÛÝÛ—Ý[[—HH\Ú[Û—ØXØÛÝ[ÜÝ]K™Ù]
ˆ˜ÛÛÛÝÛ—Ý[[‚ˆ
Bˆ›ÝÖÈœ™\ÜÜ™\]Y\Ý×Ü™]\ÙY—HH™]\ÙYÜ™\ÜÜ™\]Y\ÝÂˆ›ÝÖÈ™^\×Ü™XØ[Ý[]YÙœ›ÛWÙˆ—HHØØ[WÜ™XØ[Ý[]YÙ^\Âˆ›ÝÖÈœØ[\YØ]˜Z[Xš[]WÜÝ]\È—HHØ[\YÜÝ]\Âˆ›ÝÖÈœ™X[ÝØ]Ù^\È—HH™X[ÝØ]Ù^\Âˆ[ÙN‚ˆ›ÝÖÈØZ][™×Ú›ØœÈ—HHˆ›ÝÖÈØZ][™×Ú›Øœ×ØžWÜš[Üš]H—HHßBˆ›ÝÖÈ˜XØÛÝ[Û\ÝÍ×Ø]—HH›Û™Bˆ›ÝÖÈ˜XØÛÝ[ØÛÛÛÝÛ—Ý[[—HH›Û™Bˆ›ÝÖÈœ™\ÜÜ™\]Y\Ý×Ü™]\ÙY—HHˆ›ÝÖÈ™^\×Ü™XØ[Ý[]YÙœ›ÛWÙˆ—HHˆ›ÝÖÈœØ[\YØ]˜Z[Xš[]WÜÝ]\È—HHßBˆ›ÝÖÈœ™X[ÝØ]Ù^\È—HHˆØ[™Y]\ÈHÂˆ˜[YBˆ›Üˆ˜[YH[ˆ
ˆ›ÝË™Ù]
›™^Ø[ÝÙYØ]ŠKˆ›ÝË™Ù]
˜ÛÛÛÝÛ—Ý[[ŠKˆ›ÝË™Ù]
›X\ÙWÝ[[ŠKˆ›ÝË™Ù]
˜XØÛÝ[ØÛÛÛÝÛ—Ý[[ŠKˆ
BˆYˆ˜[YBˆBˆ›ÝÖÈœ™\Ý[YWÙ›Ü™XØ\Ý—HHX^
Ø[™Y]\ÊHYˆØ[™Y]\È[ÙHˆ‚ˆ™]\›ˆ›ÝÜÂ‚‚™YˆÛX\—Ø\WØÛÛÛÝÛŠÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹›ÝšY\ŽˆÝ‹\WØ\™XNˆÝˆ›Û™HH›Û™JHOˆ›Û™N‚ˆ[œÝ\™WØ\WØØ[ÜÝ]WÜØÚ[XJÛÛ›ŠBˆYˆ\WØ\™XN‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUH\WØØ[ÜÝ]BˆÑUÛÛÛÝÛ—Ý[[H	ÉË\ÝÙ\œ›ÜˆH	ÉË\]YØ]HÂˆÒT‘H›ÝšY\ˆHÈS‘\WØ\™XHHÂˆˆˆ‹ˆ
]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠK›ÝšY\‹\WØ\™XJKˆ
Bˆ[ÙN‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUH\WØØ[ÜÝ]BˆÑUÛÛÛÝÛ—Ý[[H	ÉË\ÝÙ\œ›ÜˆH	ÉË\]YØ]HÂˆÒT‘H›ÝšY\ˆHÂˆˆˆ‹ˆ
]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠK›ÝšY\ŠKˆ
BˆYˆ›ÝšY\ˆOHS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓTŽ‚ˆÛX\—Ù\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]ØÛÛÛÝÛŠÛÛ›ŠB‚‚™Yˆ]\ÝØ˜XÚÙÜ›Ý[™Ú›Ø—Ù›Ü—Ý\JÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹›Ø—Ý\NˆÝŠHOˆXÝÜÝ‹[žWH›Û™N‚ˆ›ÝÈHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕY›Ø—Ý\KÝ]\Ë\˜[\×ÚœÛÛ‹™\Ý[ÚœÛÛ‹\œ›Ü—ÛY\ÜØYÙKˆÜ™X]YØ]Ý\YØ]š[š\ÚYØ]™^Ø][\Ø]ˆ”“ÓH˜XÚÙÜ›Ý[™Ú›ØœÂˆÒT‘H›Ø—Ý\HHÂˆÔ‘Tˆ–HYTÐÂˆSRUBˆˆˆ‹ˆ
›Ø—Ý\K
Kˆ
K™™]ÚÛ™J
BˆYˆ›ÝÈ\È›Û™N‚ˆ™]\›ˆ›Û™Bˆ›ØˆHXÝ
›ÝÊBˆ›Üˆ[Y\Ý[\ÙšY[[ˆ
ˆ˜Ü™X]YØ]‹ˆœÝ\YØ]‹ˆ™š[š\ÚYØ]‹ˆ›™^Ø][\Ø]‹ˆ
N‚ˆ›Ø–Ý[Y\Ý[\ÙšY[HH˜XÚÙÜ›Ý[™Ú›Ø—Ý[Y\Ý[\Ý×Û\Ø›ÛŠˆ›ÝÖÝ[Y\Ý[\ÙšY[Bˆ
Bˆ™]\›ˆ›Ø‚‚‚™YˆØÚY[\—Û™^Ü[—ÛX™[
›Ø—ÚYˆÝŠHOˆÝŽ‚ˆYˆÐÒQSTˆ\È›Û™N‚ˆ™]\›ˆˆ‚ˆ›ØˆHÐÒQST‹™Ù]Ú›ØŠ›Ø—ÚY
Bˆ™^Ü[ˆHÙ]]Š›Ø‹›™^Ü[—Ý[YH‹›Û™JHYˆ›Øˆ[ÙH›Û™BˆYˆ›Ý™^Ü[Ž‚ˆ™]\›ˆˆ‚ˆžN‚ˆ™]\›ˆ™^Ü[‹˜\Ý[Y^›Û™JTÐ“Ó—ÕSQV“Ó‘JKš\ÛÙ›Ü›X]
[Y\ÜXÏH›Z[]\ÈŠBˆ^Ù\^Ù\[ÛŽ‚ˆ™]\›ˆÝŠ™^Ü[ŠB‚‚™YˆZ[Ú[YÜ˜][Û—Ø\WØÛÛ›ÛÊÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[ÛŠHOˆ\ÝÙXÝÜÝ‹[žWWN‚ˆÛÛ™šYÜÈHÂˆ›ÝšY\ŽˆÙ]Ú[YÜ˜][Û—ØÛÛ™šYÊÛÛ›‹›ÝšY\ŠBˆ›Üˆ›ÝšY\ˆ[ˆS•QÔUSÓ—Ô“Õ’QT—ÓÔSÓ”ÂˆBˆÝ]WÜ›ÝÜÈHÂˆ
›ÝÖÈœ›ÝšY\ˆ—K›ÝÖÈ˜\WØ\™XH—JNˆ›ÝÂˆ›Üˆ›ÝÈ[ˆ\ÝØ\WØØ[ÜÝ]\ÊÛÛ›ŠBˆBˆ›Ø—Ý\\ÈHÂˆ
S•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹TWÐT‘PWÔÕUJNˆ™\Ú[ÛœÛÛ\—ÜÝ]WÜÞ[˜È‹ˆ
S•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹TWÐT‘PWÔ“ÑPÕSÓŠNˆ™\Ú[ÛœÛÛ\—Ü›ÙXÝ[Û—ÜÞ[˜È‹ˆ
S•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹TWÐT‘PWÑPQÓ“ÔÕPÔÊNˆ™\Ú[ÛœÛÛ\—Ú[™\\—Ø]˜Z[Xš[]WØ˜XÚÙš[‹ˆ
S•QÔUSÓ—Ô“Õ’QT—ÔÒQÑS‘T‘ÖKTWÐT‘PWÔÕUJNˆœÚYÙ[™\™ÞWÜÝ]WÜÞ[˜È‹ˆBˆØÚY[\—Ú›ØœÈHÂˆ
S•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹TWÐT‘PWÔÕUJNˆš[YÜ˜][Û‹\Ý]KY\Ú[ÛœÛÛ\‹ZÝ\›H‹ˆ
S•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹TWÐT‘PWÔ“ÑPÕSÓŠNˆš[YÜ˜][Û‹\›ÙXÝ[Û‹Y\Ú[ÛœÛÛ\‹YZ[H‹ˆ
S•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹TWÐT‘PWÑPQÓ“ÔÕPÔÊNˆš[YÜ˜][Û‹YXYÛ›ÜÝXÜËY\Ú[ÛœÛÛ\‹YZ[H‹ˆ
S•QÔUSÓ—Ô“Õ’QT—ÔÒQÑS‘T‘ÖKTWÐT‘PWÔÕUJNˆš[YÜ˜][Û‹\Ý]K\ÚYÙ[™\™ÞKZÝ\›H‹ˆBˆÛÛ›ÛÎˆ\ÝÙXÝÜÝ‹[žWWHH×Bˆ›Üˆ›ÝšY\ˆ[ˆS•QÔUSÓ—Ô“Õ’QT—ÓÔSÓ”Î‚ˆÛÛ™šYÈHÛÛ™šYÜË™Ù]
›ÝšY\ŠBˆYˆÛÛ™šYÈ\È›Û™N‚ˆÛÛ[YBˆÜ™Y[X[×ÛÚÈH›ÛÛ
ÛÛ™šYË™Ù]
\Ù\›˜[YHŠJH[™›ÛÛ
ÛÛ™šYË™Ù]
œ\ÜÝÛÜ™ØÛÛ™šYÝ\™YŠJBˆ›ÝšY\—Ø\™X\ÈHÐTWÐT‘PWÔÕUWBˆYˆ›ÝšY\ˆOHS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓTŽ‚ˆ›ÝšY\—Ø\™X\Ë™^[™
ÐTWÐT‘PWÔ“ÑPÕSÓ‹TWÐT‘PWÑPQÓ“ÔÕPÔ×JBˆ\™X\ÎˆXÝÜÝ‹XÝÜÝ‹[žWWHHßBˆ›Üˆ\™XH[ˆ›ÝšY\—Ø\™X\Î‚ˆÝ]HHÝ]WÜ›ÝÜË™Ù]

›ÝšY\‹\™XJKßJBˆ›Ø—Ý\HH›Ø—Ý\\Ë™Ù]

›ÝšY\‹\™XJKˆŠBˆ]\ÝÚ›ØˆH]\ÝØ˜XÚÙÜ›Ý[™Ú›Ø—Ù›Ü—Ý\JÛÛ›‹›Ø—Ý\JHYˆ›Ø—Ý\H[ÙH›Û™BˆÛÛÛÝÛ—ØXÝ]™HH›ÛÛ
Ý]K™Ù]
˜ÛÛÛÝÛ—ØXÝ]™HŠJBˆYˆ›ÝÛÛ™šYË™Ù]
™[˜X›YŠN‚ˆ\WÜÝ]\ÈH‘\ØX›Y‚ˆ[Yˆ›ÝÜ™Y[X[×ÛÚÎ‚ˆ\WÜÝ]\ÈH”Ù[HÜ™Y[˜ÚXZ\È‚ˆ[YˆÛÛÛÝÛ—ØXÝ]™N‚ˆ\WÜÝ]\ÈH‘[HÛÛÛÝÛˆ‚ˆ[Yˆ]\ÝÚ›Øˆ[™]\ÝÚ›Ø‹™Ù]
œÝ]\ÈŠHOH™˜Z[YŽ‚ˆ\WÜÝ]\ÈH‘˜[ÝH[[XH[]]˜H‚ˆ[ÙN‚ˆ\WÜÝ]\ÈH“ÒÈ‚ˆ\™X\ÖØ\™XWHHÂˆœÝ]HŽˆÝ]Kˆ›]\ÝÚ›ØˆŽˆ]\ÝÚ›Ø‹ˆ˜\WÜÝ]\ÈŽˆ\WÜÝ]\Ëˆ›\ÝØ][\Ø]ŽˆÝ]K™Ù]
›\ÝØ][\Ø]ŠHÜˆ
]\ÝÚ›ØˆÜˆßJK™Ù]
œÝ\YØ]ŠHÜˆˆ‹ˆ›\ÝÜÝXØÙ\Ü×Ø]ŽˆÝ]K™Ù]
›\ÝÜÝXØÙ\Ü×Ø]ŠHÜˆˆ‹ˆ›™^Ø][\Ø]ŽˆÝ]K™Ù]
˜ÛÛÛÝÛ—Ý[[ŠHYˆÛÛÛÝÛ—ØXÝ]™H[ÙH
]\ÝÚ›ØˆÜˆßJK™Ù]
›™^Ø][\Ø]ŠHÜˆˆ‹ˆ›™^ÜØÚY[YØ]ŽˆØÚY[\—Û™^Ü[—ÛX™[
ØÚY[\—Ú›ØœË™Ù]

›ÝšY\‹\™XJKˆŠJKˆ›\ÝÙ\œ›ÜˆŽˆÝ]K™Ù]
›\ÝÙ\œ›ÜˆŠHÜˆ
]\ÝÚ›ØˆÜˆßJK™Ù]
™\œ›Ü—ÛY\ÜØYÙHŠHÜˆˆ‹ˆ˜ÛÛÛÝÛ—ØXÝ]™HŽˆÛÛÛÝÛ—ØXÝ]™Kˆ˜ÛÛÛÝÛ—ÛY\ÜØYÙHŽˆÝ]K™Ù]
˜ÛÛÛÝÛ—ÛY\ÜØYÙHŠHÜˆˆ‹ˆBˆÛÛ›ÛË˜\[™
ˆÂˆœ›ÝšY\ˆŽˆ›ÝšY\‹ˆ˜ÛÛ™šYÈŽˆÛÛ™šYËˆ™[˜X›YŽˆ›ÛÛ
ÛÛ™šYË™Ù]
™[˜X›YŠJKˆ˜Ü™Y[X[×ÛÚÈŽˆÜ™Y[X[×ÛÚËˆœÝ]WÙ[˜X›YŽˆ›ÛÛ
ÛÛ™šYË™Ù]
˜]]×ÜÞ[˜×Ù[˜X›YŠJKˆœ›ÙXÝ[Û—Ù[˜X›YŽˆ›ÛÛ
ÛÛ™šYË™Ù]
œ›ÙXÝ[Û—ÜÞ[˜×Ù[˜X›YŠJHYˆ›ÝšY\ˆOHS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓTˆ[ÙH˜[ÙKˆ™XYÛ›ÜÝXÜ×Ù[˜X›YŽˆ›ÛÛ
ÛÛ™šYË™Ù]
™XYÛ›ÜÝXÜ×ÜÞ[˜×Ù[˜X›YŠJHYˆ›ÝšY\ˆOHS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓTˆ[ÙH˜[ÙKˆ˜]]×ÜÞ[˜×Ù[˜X›YŽˆ›ÛÛ
ÛÛ™šYË™Ù]
˜]]×ÜÞ[˜×Ù[˜X›YŠJKˆœÝ]WÚ[\˜[ÚÝ\œÈŽˆÛÛ™šYË™Ù]
œÝ]WÜÞ[˜×Ú[\˜[ÚÝ\œÈŠHÜˆQUSÔÕUWÔÖS×ÒS•T•SÒÕT”Ëˆœ›ÙXÝ[Û—ÜÞ[˜×Ý[YHŽˆÛÛ™šYË™Ù]
œ›ÙXÝ[Û—ÜÞ[˜×Ý[YHŠHÜˆQUSÑ•TÒSÓ”ÓÓT—Ô“ÑPÕSÓ—ÔÖS×ÕSQKˆ™XYÛ›ÜÝXÜ×ÜÞ[˜×Ý[YHŽˆÛÛ™šYË™Ù]
™XYÛ›ÜÝXÜ×ÜÞ[˜×Ý[YHŠHÜˆQUSÑ•TÒSÓ”ÓÓT—ÑPQÓ“ÔÕPÔ×ÔÖS×ÕSQKˆ˜\™X\ÈŽˆ\™X\ËˆœÝ]WØ\™XHŽˆ\™X\Ë™Ù]
TWÐT‘PWÔÕUKßJKˆœ›ÙXÝ[Û—Ø\™XHŽˆ\™X\Ë™Ù]
TWÐT‘PWÔ“ÑPÕSÓ‹ßJKˆ™XYÛ›ÜÝXÜ×Ø\™XHŽˆ\™X\Ë™Ù]
TWÐT‘PWÑPQÓ“ÔÕPÔËßJKˆBˆ
Bˆ™]\›ˆÛÛ›ÛÂ‚‚™Yˆ›ÝYžWØ\WÜ˜]WÛ[Z]
ˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ›ÝšY\ŽˆÝ‹ˆ\WØ\™XNˆÝ‹ˆÛÛÛÝÛ—Ý[[ˆ]][YKˆY\ÜØYÙNˆÝ‹ŠHOˆ›Û™N‚ˆÝ]HHÙ]Ø\WØØ[ÜÝ]JÛÛ›‹›ÝšY\‹\WØ\™XJBˆ\ÝØ[\Ø]H\œÙWÙ]][YWÝ˜[YJÝŠÝ]K™Ù]
›\ÝØ[\Ø]ŠHÜˆˆŠJBˆYˆ\ÝØ[\Ø][™\ÝØ[\Ø]H]][YK››ÝÊ
HH[YY[JZ[]\ÏQQUSÑ•TÒSÓ”ÓÓT—ÔUWÓSRUÓRS•UTÊN‚ˆ™]\›‚ˆ›ÝÈH]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠBˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUH\WØØ[ÜÝ]BˆÑU\ÝØ[\Ø]HË\]YØ]HÂˆÒT‘H›ÝšY\ˆHÈS‘\WØ\™XHHÂˆˆˆ‹ˆ
›ÝË›ÝË›ÝšY\‹\WØ\™XJKˆ
Bˆ[\ÚÙ^HHˆ˜\WÜ˜]WÛ[Z]žÜ›ÝšY\ŸNžØ\WØ\™X_NžØÛÛÛÝÛ—Ý[[œÝ™[YJ	ÉVI[IY	R	SIÊ_H‚ˆ[YÜ˜[WÛY\ÜØYÙHH
ˆˆžÚ[™\ØØ\J›ÝšY\Š_HTH[HÛÛÛÝÛØ——ˆ‚ˆˆ\™XNˆÚ[™\ØØ\J\WØ\™XJ_Wˆ‚ˆˆ]NˆØÛÛÛÝÛ—Ý[[š\ÛÙ›Ü›X]
[Y\ÜXÏIÛZ[]\ÉÊ_Wˆ‚ˆˆžÚ[™\ØØ\JY\ÜØYÙJ_H‚ˆ
BˆÙ[™Ø[™Ü™XÛÜ™Ý[YÜ˜[WØ[\
ÛÛ›‹›Û™K˜\WÜ˜]WÛ[Z]‹[\ÚÙ^K[YÜ˜[WÛY\ÜØYÙJB‚‚™YˆÙ]Ù\Ú[ÛœÛÛ\—Ù[™Ú[ØÛÛ™šYÊÛÛ™šYÎˆÜ[]LË”›ÝÈXÝÜÝ‹[žWJHOˆXÝÜÝ‹Ý—N‚ˆ™]\›ˆÂˆ˜˜\ÙWÝ\›ŽˆÝŠÛÛ™šYÖÈ˜˜\ÙWÝ\›—HÜˆˆŠKœÝš\

Kˆ›ÙÚ[—Ù[™Ú[ŽˆÝŠÛÛ™šYÖÈ›ÙÚ[—Ù[™Ú[—HÜˆQUSÑ•TÒSÓ”ÓÓT—ÓÑÒS—ÑS‘ÒS•
KœÝš\

HÜˆQUSÑ•TÒSÓ”ÓÓT—ÓÑÒS—ÑS‘ÒS•ˆœ[×Ù[™Ú[ŽˆÝŠÛÛ™šYÖÈœ[×Ù[™Ú[—HÜˆQUSÑ•TÒSÓ”ÓÓT—ÔÕUSÓ”×ÑS‘ÒS•
KœÝš\

HÜˆQUSÑ•TÒSÓ”ÓÓT—ÔÕUSÓ”×ÑS‘ÒS•ˆœ™X[Ý[YWÙ[™Ú[ŽˆÝŠÛÛ™šYÖÈœ™X[Ý[YWÙ[™Ú[—HÜˆQUSÑ•TÒSÓ”ÓÓT—Ô‘PSSQWÑS‘ÒS•
KœÝš\

HÜˆQUSÑ•TÒSÓ”ÓÓT—Ô‘PSSQWÑS‘ÒS•ˆ™]šXÙWÛ\ÝÙ[™Ú[ŽˆÝŠÛÛ™šYÖÈ™]šXÙWÛ\ÝÙ[™Ú[—HÜˆQUSÑ•TÒSÓ”ÓÓT—ÑU’PÑT×ÑS‘ÒS•
KœÝš\

HÜˆQUSÑ•TÒSÓ”ÓÓT—ÑU’PÑT×ÑS‘ÒS•ˆ™]šXÙWÜ™X[Ý[YWÙ[™Ú[ŽˆÝŠÛÛ™šYÖÈ™]šXÙWÜ™X[Ý[YWÙ[™Ú[—HÜˆQUSÑ•TÒSÓ”ÓÓT—ÑU’PÑWÔ‘PSSQWÑS‘ÒS•
KœÝš\

HÜˆQUSÑ•TÒSÓ”ÓÓT—ÑU’PÑWÔ‘PSSQWÑS‘ÒS•ˆ™]šXÙWÚ\ÝÜžWÙ[™Ú[ŽˆÝŠÛÛ™šYÖÈ™]šXÙWÚ\ÝÜžWÙ[™Ú[—HÜˆQUSÑ•TÒSÓ”ÓÓT—ÑU’PÑWÒTÕÔ–WÑS‘ÒS•
KœÝš\

HÜˆQUSÑ•TÒSÓ”ÓÓT—ÑU’PÑWÒTÕÔ–WÑS‘ÒS•ˆ˜[\›\×Ù[™Ú[ŽˆÝŠÛÛ™šYÖÈ˜[\›\×Ù[™Ú[—HÜˆQUSÑ•TÒSÓ”ÓÓT—ÐST“T×ÑS‘ÒS•
KœÝš\

HÜˆQUSÑ•TÒSÓ”ÓÓT—ÐST“T×ÑS‘ÒS•ˆ™^WÚÜWÙ[™Ú[ŽˆÝŠÛÛ™šYÖÈ™^WÚÜWÙ[™Ú[—HÜˆQUSÑ•TÒSÓ”ÓÓT—ÑVWÒÔWÑS‘ÒS•
KœÝš\

HÜˆQUSÑ•TÒSÓ”ÓÓT—ÑVWÒÔWÑS‘ÒS•ˆ›[ÛÚÜWÙ[™Ú[ŽˆÝŠÛÛ™šYÖÈ›[ÛÚÜWÙ[™Ú[—HÜˆQUSÑ•TÒSÓ”ÓÓT—ÓSÓ•ÒÔWÑS‘ÒS•
KœÝš\

HÜˆQUSÑ•TÒSÓ”ÓÓT—ÓSÓ•ÒÔWÑS‘ÒS•ˆB‚‚™YˆZ[Ù\Ú[ÛœÛÛ\—Ù[™Ú[ÊÛÛ™šYÎˆÜ[]LË”›ÝÈXÝÜÝ‹[žWJHOˆ\Ú[Û”ÛÛ\‘[™Ú[Î‚ˆ[™Ú[ÈHÙ]Ù\Ú[ÛœÛÛ\—Ù[™Ú[ØÛÛ™šYÊÛÛ™šYÊBˆ™]\›ˆ\Ú[Û”ÛÛ\‘[™Ú[Ê
Š™[™Ú[ÊB‚‚™YˆZ[Ù\Ú[ÛœÛÛ\—ØÛY[
ÛÛ™šYÎˆÜ[]LË”›ÝÈXÝÜÝ‹[žWJHOˆ\Ú[Û”ÛÛ\ÛY[‚ˆ™]\›ˆ\Ú[ÛœÛÛ\—ØÛY[Û[Ù[K˜ÛY[Ùœ›ÛWØÛÛ™šYÊˆXÝ
ÛÛ™šYÊKˆZ[Ù\Ú[ÛœÛÛ\—Ù[™Ú[ÊÛÛ™šYÊKˆÙ\ÜÚ[Û—Ù˜XÝÜžO\™\]Y\ÝË”Ù\ÜÚ[Û‹ˆÙ\ÜÚ[Û—ØØXÚOQ•TÒSÓ”ÓÓT—ÔÑTÔÒSÓ—ÐÐPÒKˆÙ\ÜÚ[Û—ÛØÚÏQ•TÒSÓ”ÓÓT—ÔÑTÔÒSÓ—ÓÐÒËˆÙ\ÜÚ[Û—ØØXÚWÛZ[]\ÏY\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[Û—ØØXÚWÛZ[]\Ê
Kˆ[Ý×ÜÛY\[›Ý\×Ü™\]Y\ÝØÛÛ^

KˆÛY\\][YKœÛY\ˆ
B‚‚™Yˆ^˜XÝÙ\Ú[ÛœÛÛ\—ÞÜ™—ÝÚÙ[Š™\ÜÛœÙNˆ™\]Y\ÝË”™\ÜÛœÙKÙ\ÜÚ[ÛŽˆ™\]Y\ÝË”Ù\ÜÚ[ÛŠHOˆÝŽ‚ˆ™]\›ˆ\Ú[ÛœÛÛ\—ØÛY[Û[Ù[K™^˜XÝÞÜ™—ÝÚÙ[Š™\ÜÛœÙKÙ\ÜÚ[ÛŠB‚‚™YˆÙ]Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[ÛŠÛÛ™šYÎˆÜ[]LË”›ÝÈXÝÜÝ‹[žWK
‹›Ü˜ÙWÛÙÚ[Žˆ›ÛÛH˜[ÙJHOˆ\VÜ™\]Y\ÝË”Ù\ÜÚ[Û‹Ý—N‚ˆÛÛÛÝÛ—Ü™X\ÛÛˆHÙ]Ù\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]ØÛÛÛÝÛ—Ü™X\ÛÛŠ
BˆYˆÛÛÛÝÛ—Ü™X\ÛÛŽ‚ˆÑÑÑT‹š[™›Ê‘\Ú[Û”ÛÛ\ˆÙÚ[ˆ›ØÚÙYžHXÝ]™H˜]H[Z]ÛÛÛÝÛˆŠBˆ˜Z\ÙH˜[YQ\œ›ÜŠÛÛÛÝÛ—Ü™X\ÛÛŠBˆžN‚ˆ™]\›ˆZ[Ù\Ú[ÛœÛÛ\—ØÛY[
ÛÛ™šYÊK›ÙÚ[Š›Ü˜ÙWÛÙÚ[Y›Ü˜ÙWÛÙÚ[ŠBˆ^Ù\\Ú[Û”ÛÛ\”˜]S[Z]\œ›Üˆ\È^Î‚ˆÛÛ›ˆHØÝ\œ™[Ø\Ù—ØÛÛ›™XÝ[ÛŠ
BˆžN‚ˆ[[HX\š×Ù\Ú[ÛœÛÛ\—Ø\WØÛÛÛÝÛŠÛÛ›‹TWÐT‘PWÔÕUK™X\ÛÛ\ÝŠ^ÊJBˆ™X\ÛÛˆHÙ]Ü›ÝšY\—ØÛÛÛÝÛ—Ü™X\ÛÛŠÛÛ›‹S•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹TWÐT‘PWÔÕUJHYˆÛÛ›ˆ[ÙHÝŠ^ÊBˆ˜Z\ÙH\T˜]S[Z]\œ›ÜŠS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹TWÐT‘PWÔÕUK[[™X\ÛÛŠHœ›ÛH^Âˆš[˜[N‚ˆYˆÛÛ›ˆ\È›Ý›Û™N‚ˆÛÛ›‹˜ÛÜÙJ
B‚‚™Yˆ[˜[Y]WÙ\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[ÛŠÛÛ™šYÎˆÜ[]LË”›ÝÈXÝÜÝ‹[žWJHOˆ›Û™N‚ˆZ[Ù\Ú[ÛœÛÛ\—ØÛY[
ÛÛ™šYÊKš[˜[Y]WÜÙ\ÜÚ[ÛŠ
B‚‚™Yˆ\Ú[ÛœÛÛ\—Ü™]žWÛÜ[ÛœÊ
HOˆXÝÜÝ‹[žWN‚ˆ™]\›ˆÈ˜[Ý×ÜÛY\Žˆ›Ý\×Ü™\]Y\ÝØÛÛ^

KœÛY\\ˆŽˆ[YKœÛY\B‚‚™YˆÜÝÙ\Ú[ÛœÛÛ\—ÚœÛÛŠˆÙ\ÜÚ[ÛŽˆ™\]Y\ÝË”Ù\ÜÚ[Û‹ˆ\›ˆÝ‹ˆ^[ØYˆXÝÜÝ‹[žWKˆ
‹ˆ^XÝYÛY\ÜØYÙNˆÝ‹ŠHOˆXÝÜÝ‹[žWN‚ˆ\WØ\™XHH\Ú[ÛœÛÛ\—Ø\™XWÙ›Ü—Ý\›
\›
BˆÛÛ›ˆHØÝ\œ™[Ø\Ù—ØÛÛ›™XÝ[ÛŠ
BˆžN‚ˆYˆÛÛ›ˆ\È›Ý›Û™N‚ˆ™\]Z\™WÛ›ÝÚ[—ØÛÛÛÝÛŠÛÛ›‹S•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹\WØ\™XJBˆ™]\›ˆ\Ú[ÛœÛÛ\—ØÛY[Û[Ù[KœÜÝÙ\Ú[ÛœÛÛ\—ÚœÛÛŠˆÙ\ÜÚ[Û‹ˆ\›ˆ^[ØYˆ^XÝYÛY\ÜØYÙOY^XÝYÛY\ÜØYÙKˆ™\]Z\™WÙ]OQ˜[ÙKˆ
Š™\Ú[ÛœÛÛ\—Ü™]žWÛÜ[ÛœÊ
Kˆ
Bˆ^Ù\\Ú[Û”ÛÛ\”˜]S[Z]\œ›Üˆ\È^Î‚ˆ[[HX\š×Ù\Ú[ÛœÛÛ\—Ø\WØÛÛÛÝÛŠÛÛ›‹\WØ\™XK™X\ÛÛ\ÝŠ^ÊJBˆ™X\ÛÛˆHÙ]Ü›ÝšY\—ØÛÛÛÝÛ—Ü™X\ÛÛŠÛÛ›‹S•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹\WØ\™XJHYˆÛÛ›ˆ[ÙHÝŠ^ÊBˆ˜Z\ÙH\T˜]S[Z]\œ›ÜŠS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹\WØ\™XK[[™X\ÛÛŠHœ›ÛH^Âˆš[˜[N‚ˆYˆÛÛ›ˆ\È›Ý›Û™N‚ˆÛÛ›‹˜ÛÜÙJ
B‚‚™YˆÚ[šÙY
˜[Y\Îˆ\ÝÜÝ—KÚ^™Nˆ[
HOˆ\ÝÛ\ÝÜÝ—WN‚ˆ™]\›ˆÝ˜[Y\ÖÚ[™^ˆ[™^
ÈÚ^™WH›Üˆ[™^[ˆ˜[™ÙJ[Š˜[Y\ÊKÚ^™JWB‚‚™Yˆ™]ÚÙ\Ú[ÛœÛÛ\—ÜÝ][ÛœÊÙ\ÜÚ[ÛŽˆ™\]Y\ÝË”Ù\ÜÚ[Û‹
‹˜\ÙWÝ\›ˆÝ‹[™Ú[ˆÝŠHOˆ\ÝÙXÝÜÝ‹[žWWN‚ˆžN‚ˆ™]\›ˆ^XÝ]WÜ]Y]YYÙ\Ú[ÛœÛÛ\—ØXØÛÝ[ØØ[
ˆ[X™Nˆ\Ú[ÛœÛÛ\—ØÛY[Û[Ù[K™™]ÚÜÝ][ÛœÊˆÙ\ÜÚ[Û‹ˆ˜\ÙWÝ\›X˜\ÙWÝ\›ˆ[™Ú[Y[™Ú[ˆ
Š™\Ú[ÛœÛÛ\—Ü™]žWÛÜ[ÛœÊ
Kˆ
Bˆ
Bˆ^Ù\\Ú[Û”ÛÛ\”˜]S[Z]\œ›Üˆ\È^Î‚ˆ˜Z\ÙWÙ\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]
^ËTWÐT‘PWÔÕUJB‚‚™Yˆ™]ÚÙ\Ú[ÛœÛÛ\—Ü™X[[YWÛX\
ˆÙ\ÜÚ[ÛŽˆ™\]Y\ÝË”Ù\ÜÚ[Û‹ˆ
‹ˆ˜\ÙWÝ\›ˆÝ‹ˆ[™Ú[ˆÝ‹ˆÝ][Û—ØÛÙ\Îˆ\ÝÜÝ—KŠHOˆXÝÜÝ‹XÝÜÝ‹[žWWN‚ˆžN‚ˆ™\Ý[ˆXÝÜÝ‹XÝÜÝ‹[žWWHHßBˆ›ÜˆÝ][Û—ÙÜ›Ý\[ˆÚ[šÙY
Ý][Û—ØÛÙ\ËL
N‚ˆ™\Ý[\]Jˆ^XÝ]WÜ]Y]YYÙ\Ú[ÛœÛÛ\—ØXØÛÝ[ØØ[
ˆ[X™HÜ›Ý\\Ý][Û—ÙÜ›Ý\ˆ\Ú[ÛœÛÛ\—ØÛY[Û[Ù[K™™]ÚÜ™X[[YWÛX\
ˆÙ\ÜÚ[Û‹ˆ˜\ÙWÝ\›X˜\ÙWÝ\›ˆ[™Ú[Y[™Ú[ˆÝ][Û—ØÛÙ\ÏYÜ›Ý\ˆ
Š™\Ú[ÛœÛÛ\—Ü™]žWÛÜ[ÛœÊ
Kˆ
Bˆ
Bˆ
Bˆ™]\›ˆ™\Ý[ˆ^Ù\\Ú[Û”ÛÛ\”˜]S[Z]\œ›Üˆ\È^Î‚ˆ˜Z\ÙWÙ\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]
^ËTWÐT‘PWÔÕUJB‚‚™Yˆ™]ÚÙ\Ú[ÛœÛÛ\—Ù]šXÙWÛ\Ý
ˆÙ\ÜÚ[ÛŽˆ™\]Y\ÝË”Ù\ÜÚ[Û‹ˆ
‹ˆ˜\ÙWÝ\›ˆÝ‹ˆ[™Ú[ˆÝ‹ˆÝ][Û—ØÛÙ\Îˆ\ÝÜÝ—KŠHOˆ\ÝÙXÝÜÝ‹[žWWN‚ˆžN‚ˆ]šXÙ\Îˆ\ÝÙXÝÜÝ‹[žWWHH×Bˆ›ÜˆÝ][Û—ÙÜ›Ý\[ˆÚ[šÙY
Ý][Û—ØÛÙ\ËL
N‚ˆ]šXÙ\Ë™^[™
ˆ^XÝ]WÜ]Y]YYÙ\Ú[ÛœÛÛ\—ÙXYÛ›ÜÝXÜ×ØØ[
ˆ[X™HÜ›Ý\\Ý][Û—ÙÜ›Ý\ˆ\Ú[ÛœÛÛ\—ØÛY[Û[Ù[K™™]ÚÙ]šXÙWÛ\Ý
ˆÙ\ÜÚ[Û‹ˆ˜\ÙWÝ\›X˜\ÙWÝ\›ˆ[™Ú[Y[™Ú[ˆÝ][Û—ØÛÙ\ÏYÜ›Ý\ˆ
Š™\Ú[ÛœÛÛ\—Ü™]žWÛÜ[ÛœÊ
Kˆ
Bˆ
Bˆ
Bˆ™]\›ˆ]šXÙ\Âˆ^Ù\\Ú[Û”ÛÛ\”˜]S[Z]\œ›Üˆ\È^Î‚ˆ˜Z\ÙWÙ\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]
^ËTWÐT‘PWÑPQÓ“ÔÕPÔÊB‚‚™Yˆ™]ÚÙ\Ú[ÛœÛÛ\—Ù]šXÙWÜ™X[[YWÛX\
ˆÙ\ÜÚ[ÛŽˆ™\]Y\ÝË”Ù\ÜÚ[Û‹ˆ
‹ˆ˜\ÙWÝ\›ˆÝ‹ˆ[™Ú[ˆÝ‹ˆ]šXÙ\Îˆ\ÝÙXÝÜÝ‹[žWWKŠHOˆXÝÜÝ‹XÝÜÝ‹[žWWN‚ˆžN‚ˆ™\Ý[ˆXÝÜÝ‹XÝÜÝ‹[žWWHHßBˆ]šXÙ\×ØžWÝ\NˆXÝÚ[\ÝÙXÝÜÝ‹[žWWWHHßBˆ›Üˆ]šXÙH[ˆ]šXÙ\Î‚ˆYˆ]šXÙK™Ù]
™]—Ý\WÚYŠH\È›Ý›Û™N‚ˆ]šXÙ\×ØžWÝ\KœÙ]Y˜][
ˆ[
]šXÙVÈ™]—Ý\WÚY—JKˆ×Kˆ
K˜\[™
]šXÙJBˆ›Üˆ\YÙ]šXÙ\È[ˆ]šXÙ\×ØžWÝ\K˜[Y\Ê
N‚ˆ›Üˆ]šXÙWÙÜ›Ý\[ˆÂˆ\YÙ]šXÙ\ÖÚ[™^ˆ[™^
ÈLBˆ›Üˆ[™^[ˆ˜[™ÙJ[Š\YÙ]šXÙ\ÊKL
BˆN‚ˆ™\Ý[\]Jˆ^XÝ]WÜ]Y]YYÙ\Ú[ÛœÛÛ\—ÙXYÛ›ÜÝXÜ×ØØ[
ˆ[X™HÜ›Ý\Y]šXÙWÙÜ›Ý\ˆ\Ú[ÛœÛÛ\—ØÛY[Û[Ù[K™™]ÚÙ]šXÙWÜ™X[[YWÛX\
ˆÙ\ÜÚ[Û‹ˆ˜\ÙWÝ\›X˜\ÙWÝ\›ˆ[™Ú[Y[™Ú[ˆ]šXÙ\ÏYÜ›Ý\ˆ
Š™\Ú[ÛœÛÛ\—Ü™]žWÛÜ[ÛœÊ
Kˆ
Bˆ
Bˆ
Bˆ™]\›ˆ™\Ý[ˆ^Ù\\Ú[Û”ÛÛ\”˜]S[Z]\œ›Üˆ\È^Î‚ˆ˜Z\ÙWÙ\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]
^ËTWÐT‘PWÑPQÓ“ÔÕPÔÊB‚‚™Yˆ™]ÚÙ\Ú[ÛœÛÛ\—Ù]šXÙWÚ\ÝÜžJˆÙ\ÜÚ[ÛŽˆ™\]Y\ÝË”Ù\ÜÚ[Û‹ˆ
‹ˆ˜\ÙWÝ\›ˆÝ‹ˆ[™Ú[ˆÝ‹ˆ]šXÙ\Îˆ\ÝÙXÝÜÝ‹[žWWKˆ\™Ù]Ù]Nˆ]KˆØ[Ù[^WÜÙXÛÛ™Îˆ›Ø]HˆÛY\\Žˆ[žHH[YKœÛY\ŠHOˆ\ÝÙXÝÜÝ‹[žWWN‚ˆžN‚ˆ™]\›ˆ^XÝ]WÜ]Y]YYÙ\Ú[ÛœÛÛ\—ÙXYÛ›ÜÝXÜ×ØØ[
ˆ[X™Nˆ\Ú[ÛœÛÛ\—ØÛY[Û[Ù[K™™]ÚÙ]šXÙWÚ\ÝÜžJˆÙ\ÜÚ[Û‹ˆ˜\ÙWÝ\›X˜\ÙWÝ\›ˆ[™Ú[Y[™Ú[ˆ]šXÙ\ÏY]šXÙ\Ëˆ\™Ù]Ù]O]\™Ù]Ù]KˆØ[Ù[^WÜÙXÛÛ™ÏXØ[Ù[^WÜÙXÛÛ™ËˆÛY\\\ÛY\\‹ˆ[Ý×ÜÛY\Q˜[ÙKˆ›Ü›X[^™\[›Ü›X[^™WÙ\Ú[ÛœÛÛ\—Ù]šXÙWÚ\ÝÜžWÜ›ÝÜËˆ
Bˆ
Bˆ^Ù\\Ú[Û”ÛÛ\”˜]S[Z]\œ›Üˆ\È^Î‚ˆ˜Z\ÙWÙ\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]
^ËTWÐT‘PWÑPQÓ“ÔÕPÔÊB‚‚™Yˆ›Ü›X[^™WÙ\Ú[ÛœÛÛ\—Ù]šXÙWÚ\ÝÜžWÜ›ÝÜÊˆ]Nˆ[žKˆ]šXÙ\Îˆ\ÝÙXÝÜÝ‹[žWWKŠHOˆ\ÝÙXÝÜÝ‹[žWWN‚ˆ]šXÙWØžWÚYˆXÝÜÝ‹XÝÜÝ‹[žWWHHßBˆ›Üˆ]šXÙH[ˆ]šXÙ\Î‚ˆ›ÜˆÙ^H[ˆ
™^\›˜[Ù]šXÙWÚY‹™]—Ùˆ‹œÛˆŠN‚ˆ˜[YHHÝŠ]šXÙK™Ù]
Ù^JHÜˆˆŠKœÝš\

BˆYˆ˜[YN‚ˆ]šXÙWØžWÚYÝ˜[YWHH]šXÙBˆ˜[˜XÚ×Ù]šXÙHH]šXÙ\ÖÌHYˆ[Š]šXÙ\ÊHOHH[ÙH›Û™Bˆ›Ü›X[^™Yˆ\ÝÙXÝÜÝ‹[žWWHH×B‚ˆYˆš\Ú]
˜[YNˆ[žK[š\š]YÙ]šXÙNˆXÝÜÝ‹[žWH›Û™HH›Û™JHOˆ›Û™N‚ˆYˆ\Ú[œÝ[˜ÙJ˜[YK\Ý
N‚ˆ›Üˆ][H[ˆ˜[YN‚ˆš\Ú]
][K[š\š]YÙ]šXÙJBˆ™]\›‚ˆYˆ›Ý\Ú[œÝ[˜ÙJ˜[YKXÝ
N‚ˆ™]\›‚ˆ˜]×Ù]šXÙWÚYHš\œÝÛ›Û—Ù[\J˜[YKÈ™]’Y‹™]šXÙRY‹™]‘ˆ‹™]šXÙQˆ‹™\ÛÛÙH‹œÛˆ—JBˆ]šXÙHH]šXÙWØžWÚY™Ù]
ÝŠ˜]×Ù]šXÙWÚYÜˆˆŠKœÝš\

JHÜˆ[š\š]YÙ]šXÙHÜˆ˜[˜XÚ×Ù]šXÙBˆ]WÛX\H˜[YK™Ù]
™]R][SX\ŠHYˆ\Ú[œÝ[˜ÙJ˜[YK™Ù]
™]R][SX\ŠKXÝ
H[ÙH˜[YBˆØ[\WÝ[YHH\œÙWÙ]][YWÝ˜[YJš\œÝÛ›Û—Ù[\J˜[YKÈ˜ÛÛXÝ[YH‹œØ[\U[YH‹[YH‹[Y\Ý[\—JJBˆYˆØ[\WÝ[YH\È›Û™H[™]WÛX\\È›Ý˜[YN‚ˆØ[\WÝ[YHH\œÙWÙ]][YWÝ˜[YJš\œÝÛ›Û—Ù[\J]WÛX\È˜ÛÛXÝ[YH‹œØ[\U[YH‹[YH‹[Y\Ý[\—JJBˆXÝ]™WÜÝÙ\ˆHš\œÝÛ›Û—Ù[\J]WÛX\È˜XÝ]™WÜÝÙ\ˆ‹˜XÝ]™TÝÙ\ˆ‹˜XÝ]™WÜÝÙ\—ÚÝÈ‹œÝÙ\ˆ—JBˆYˆ]šXÙH[™Ø[\WÝ[YH\È›Ý›Û™H[™XÝ]™WÜÝÙ\ˆ›Ý[ˆ
›Û™KˆŠN‚ˆ›Ü›X[^™Y˜\[™
ˆÂˆ
Š™]šXÙKˆœØ[\WÝ[YHŽˆØ[\WÝ[YKˆ˜XÝ]™WÜÝÙ\—ÚÝÈŽˆ›Ü›X[^™WÜÝÙ\—Ý×ÚÝÊXÝ]™WÜÝÙ\ŠKˆœ˜]×Ü^[ØYŽˆ˜[YKˆBˆ
Bˆ™]\›‚ˆ™\ÝYÙ›Ý[™H˜[ÙBˆ›ÜˆÙ^H[ˆ
›\Ý‹™]H‹š\ÝÜžQ]H‹™]S\Ý‹œ™XÛÜ™È‹™]R][SX\ŠN‚ˆ™\ÝYH˜[YK™Ù]
Ù^JBˆYˆ\Ú[œÝ[˜ÙJ™\ÝY
\ÝXÝ
JH[™™\ÝY\È›Ý˜[YN‚ˆ™\ÝYÙ›Ý[™HYBˆš\Ú]
™\ÝY]šXÙJBˆYˆ›Ý™\ÝYÙ›Ý[™‚ˆ›ÜˆÙ^K™\ÝY[ˆ˜[YKš][\Ê
N‚ˆX\YÙ]šXÙHH]šXÙWØžWÚY™Ù]
ÝŠÙ^JJBˆYˆX\YÙ]šXÙH[™\Ú[œÝ[˜ÙJ™\ÝY
\ÝXÝ
JN‚ˆš\Ú]
™\ÝYX\YÙ]šXÙJB‚ˆš\Ú]
]JBˆ™]\›ˆ›Ü›X[^™Y‚‚™Yˆ™]ÚÙ\Ú[ÛœÛÛ\—Ø[\›WÛX\
ˆÙ\ÜÚ[ÛŽˆ™\]Y\ÝË”Ù\ÜÚ[Û‹ˆ
‹ˆ˜\ÙWÝ\›ˆÝ‹ˆ[™Ú[ˆÝ‹ˆÝ][Û—ØÛÙ\Îˆ\ÝÜÝ—KŠHOˆXÝÜÝ‹\ÝÙXÝÜÝ‹[žWWWN‚ˆžN‚ˆ™\Ý[ˆXÝÜÝ‹\ÝÙXÝÜÝ‹[žWWWHHßBˆ›ÜˆÝ][Û—ÙÜ›Ý\[ˆÚ[šÙY
Ý][Û—ØÛÙ\ËL
N‚ˆ\X[H^XÝ]WÜ]Y]YYÙ\Ú[ÛœÛÛ\—ÙXYÛ›ÜÝXÜ×ØØ[
ˆ[X™HÜ›Ý\\Ý][Û—ÙÜ›Ý\ˆ\Ú[ÛœÛÛ\—ØÛY[Û[Ù[K™™]ÚØ[\›WÛX\
ˆÙ\ÜÚ[Û‹ˆ˜\ÙWÝ\›X˜\ÙWÝ\›ˆ[™Ú[Y[™Ú[ˆÝ][Û—ØÛÙ\ÏYÜ›Ý\ˆ[™ÝXYÙOQQUSÑ•TÒSÓ”ÓÓT—ÐST“T×ÓS‘ÕPQÑKˆ
Š™\Ú[ÛœÛÛ\—Ü™]žWÛÜ[ÛœÊ
Kˆ
Bˆ
Bˆ›ÜˆÝ][Û—ØÛÙK›ÝÜÈ[ˆ\X[š][\Ê
N‚ˆ™\Ý[œÙ]Y˜][
Ý][Û—ØÛÙK×JK™^[™
›ÝÜÊBˆ™]\›ˆ™\Ý[ˆ^Ù\\Ú[Û”ÛÛ\”˜]S[Z]\œ›Üˆ\È^Î‚ˆ˜Z\ÙWÙ\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]
^ËTWÐT‘PWÑPQÓ“ÔÕPÔÊB‚‚™YˆÛÛXÝÝ[YWÛ\ÊÛÛXÝÙ]Nˆ]JHOˆ[‚ˆ™]\›ˆÛÛXÝÝ[YWÜÝ\ÛÙ—Ù^WÛ\ÊÛÛXÝÙ]JB‚‚™YˆÛÛXÝÝ[YWÛ›ÛÛ—Û\ÊÛÛXÝÙ]Nˆ]JHOˆ[‚ˆ™]\›ˆÛÛXÝÝ[YWÛ›ÛÛ—ÛÙ—Û[ÛÛ\ÊÛÛXÝÙ]JB‚‚™Yˆ›Ü›X[^™WÙ\Ú[ÛœÛÛ\—ÚÜWÜ›ÝÜÊ]Nˆ[žJHOˆ\ÝÙXÝÜÝ‹[žWWN‚ˆ™]\›ˆ›Ü›X[^™WØÛY[ÚÜWÜ›ÝÜÊ]JB‚‚™Yˆ\œÙWÙ\Ú[ÛœÛÛ\—ØÛÛXÝÙ]J›ÝÎˆXÝÜÝ‹[žWK˜[˜XÚ×Ù]Nˆ]H›Û™HH›Û™JHOˆ]H›Û™N‚ˆ™]\›ˆ\œÙWØÛY[ØÛÛXÝÙ]J›ÝË˜[˜XÚ×Ù]JB‚‚™Yˆ™]ÚÙ\Ú[ÛœÛÛ\—ÚÜWÛX\
ˆÙ\ÜÚ[ÛŽˆ™\]Y\ÝË”Ù\ÜÚ[Û‹ˆ
‹ˆ˜\ÙWÝ\›ˆÝ‹ˆ[™Ú[ˆÝ‹ˆÝ][Û—ØÛÙ\Îˆ\ÝÜÝ—KˆÛÛXÝÙ]Nˆ]Kˆ^XÝYÛY\ÜØYÙNˆÝ‹ŠHOˆXÝÜÝ‹XÝÜÝ‹[žWWN‚ˆYˆ[ŠÝ][Û—ØÛÙ\ÊHˆL‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠØYHÚ[XYHÔH\Ú[Û”ÛÛ\ˆXÙZ]H›ÈX^[[ÈL[œÝ[XÛÙ\ËˆŠBˆžN‚ˆ™]\›ˆ^XÝ]WÜ]Y]YYÙ\Ú[ÛœÛÛ\—ÚÜWØØ[
ˆ[X™Nˆ\Ú[ÛœÛÛ\—ØÛY[Û[Ù[K™™]ÚÚÜWÛX\
ˆÙ\ÜÚ[Û‹ˆ˜\ÙWÝ\›X˜\ÙWÝ\›ˆ[™Ú[Y[™Ú[ˆÝ][Û—ØÛÙ\Ï\Ý][Û—ØÛÙ\ËˆÛÛXÝÙ]OXÛÛXÝÙ]Kˆ^XÝYÛY\ÜØYÙOY^XÝYÛY\ÜØYÙKˆ
Š™\Ú[ÛœÛÛ\—Ü™]žWÛÜ[ÛœÊ
Kˆ
Kˆ[™Ú[Y[™Ú[ˆ
Bˆ^Ù\\Ú[Û”ÛÛ\”˜]S[Z]\œ›Üˆ\È^Î‚ˆ˜Z\ÙWÙ\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]
^ËTWÐT‘PWÔ“ÑPÕSÓŠB‚‚™Yˆ™]ÚÙ\Ú[ÛœÛÛ\—ÚÜWÜ›ÝÜÊˆÙ\ÜÚ[ÛŽˆ™\]Y\ÝË”Ù\ÜÚ[Û‹ˆ
‹ˆ˜\ÙWÝ\›ˆÝ‹ˆ[™Ú[ˆÝ‹ˆÝ][Û—ØÛÙ\Îˆ\ÝÜÝ—KˆÛÛXÝÙ]Nˆ]Kˆ^XÝYÛY\ÜØYÙNˆÝ‹ŠHOˆ\ÝÙXÝÜÝ‹[žWWN‚ˆYˆ[ŠÝ][Û—ØÛÙ\ÊHˆL‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠØYHÚ[XYHÔH\Ú[Û”ÛÛ\ˆXÙZ]H›ÈX^[[ÈL[œÝ[XÛÙ\ËˆŠBˆžN‚ˆ™]\›ˆ^XÝ]WÜ]Y]YYÙ\Ú[ÛœÛÛ\—ÚÜWØØ[
ˆ[X™Nˆ\Ú[ÛœÛÛ\—ØÛY[Û[Ù[K™™]ÚÚÜWÜ›ÝÜÊˆÙ\ÜÚ[Û‹ˆ˜\ÙWÝ\›X˜\ÙWÝ\›ˆ[™Ú[Y[™Ú[ˆÝ][Û—ØÛÙ\Ï\Ý][Û—ØÛÙ\ËˆÛÛXÝÙ]OXÛÛXÝÙ]Kœ™\XÙJ^OLJKˆ^XÝYÛY\ÜØYÙOY^XÝYÛY\ÜØYÙKˆ
Š™\Ú[ÛœÛÛ\—Ü™]žWÛÜ[ÛœÊ
Kˆ
Kˆ[™Ú[Y[™Ú[ˆ
Bˆ^Ù\\Ú[Û”ÛÛ\”˜]S[Z]\œ›Üˆ\È^Î‚ˆ˜Z\ÙWÙ\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]
^ËTWÐT‘PWÔ“ÑPÕSÓŠB‚‚™Yˆ™]ÚÙ\Ú[ÛœÛÛ\—ÚÜWÙ^WÜ›ÝÜÊˆÙ\ÜÚ[ÛŽˆ™\]Y\ÝË”Ù\ÜÚ[Û‹ˆ˜\ÙWÝ\›ˆÝ‹ˆ[™Ú[ˆÝ‹ˆÝ][Û—ØÛÙ\Îˆ\ÝÜÝ—KˆÛÛXÝÙ]Nˆ]KŠHOˆ\ÝÙXÝÜÝ‹[žWWN‚ˆ™]\›ˆ™]ÚÙ\Ú[ÛœÛÛ\—ÚÜWÜ›ÝÜÊˆÙ\ÜÚ[Û‹ˆ˜\ÙWÝ\›X˜\ÙWÝ\›ˆ[™Ú[Y[™Ú[ˆÝ][Û—ØÛÙ\Ï\Ý][Û—ØÛÙ\ËˆÛÛXÝÙ]OXÛÛXÝÙ]Kˆ^XÝYÛY\ÜØYÙOH‘˜[H[ÈØ\ˆÜÈÔ\ÈX\š[ÜÈ\Ú[Û”ÛÛ\‹ˆ‹ˆ
B‚‚™Yˆ™]ÚÙ\Ú[ÛœÛÛ\—ÚÜWÙ^WÛX\
ˆÙ\ÜÚ[ÛŽˆ™\]Y\ÝË”Ù\ÜÚ[Û‹ˆ˜\ÙWÝ\›ˆÝ‹ˆ[™Ú[ˆÝ‹ˆÝ][Û—ØÛÙ\Îˆ\ÝÜÝ—KˆÛÛXÝÙ]Nˆ]KŠHOˆXÝÜÝ‹XÝÜÝ‹[žWWN‚ˆ™]\›ˆ™]ÚÙ\Ú[ÛœÛÛ\—ÚÜWÛX\
ˆÙ\ÜÚ[Û‹ˆ˜\ÙWÝ\›X˜\ÙWÝ\›ˆ[™Ú[Y[™Ú[ˆÝ][Û—ØÛÙ\Ï\Ý][Û—ØÛÙ\ËˆÛÛXÝÙ]OXÛÛXÝÙ]Kˆ^XÝYÛY\ÜØYÙOH‘˜[H[ÈØ\ˆÜÈÔ\ÈX\š[ÜÈ\Ú[Û”ÛÛ\‹ˆ‹ˆ
B‚‚™Yˆ™]ÚÙ\Ú[ÛœÛÛ\—ÚÜWÛ[ÛÛX\
ˆÙ\ÜÚ[ÛŽˆ™\]Y\ÝË”Ù\ÜÚ[Û‹ˆ˜\ÙWÝ\›ˆÝ‹ˆ[™Ú[ˆÝ‹ˆÝ][Û—ØÛÙ\Îˆ\ÝÜÝ—KˆÛÛXÝÙ]Nˆ]KŠHOˆXÝÜÝ‹XÝÜÝ‹[žWWN‚ˆ[ÛÙ]HHÛÛXÝÙ]Kœ™\XÙJ^OLJBˆ™]\›ˆ™]ÚÙ\Ú[ÛœÛÛ\—ÚÜWÛX\
ˆÙ\ÜÚ[Û‹ˆ˜\ÙWÝ\›X˜\ÙWÝ\›ˆ[™Ú[Y[™Ú[ˆÝ][Û—ØÛÙ\Ï\Ý][Û—ØÛÙ\ËˆÛÛXÝÙ]O[[ÛÙ]Kˆ^XÝYÛY\ÜØYÙOH‘˜[H[ÈØ\ˆÜÈÔ\ÈY[œØZ\È\Ú[Û”ÛÛ\‹ˆ‹ˆ
B‚‚™Yˆ\œÙWÚÝÜÝ˜[YJ˜[YNˆ[žJHOˆ›Ø]›Û™N‚ˆYˆ˜[YH\È›Û™N‚ˆ™]\›ˆ›Û™Bˆ˜]ÈHÝŠ˜[YJKœÝš\

Kœ™\XÙJ‹‹‹ˆŠBˆYˆ›Ý˜]ÈÜˆ˜]ÈOH‹HŽ‚ˆ™]\›ˆ›Û™Bˆ˜]ÈH™KœÝXŠˆ–×ŒNK—WH‹ˆ‹˜]ÊBˆYˆ˜]È[ˆ
ˆ‹‹H‹‹ˆŠN‚ˆ™]\›ˆ›Û™BˆžN‚ˆ\œÙYH›Ø]
˜]ÊBˆ^Ù\˜[YQ\œ›ÜŽ‚ˆ™]\›ˆ›Û™Bˆ™]\›ˆ\œÙYYˆ\œÙYˆ[ÙH›Û™B‚‚™Yˆ\œÙWÙ›Ø]Ý˜[YJ˜[YNˆ[žJHOˆ›Ø]›Û™N‚ˆYˆ˜[YH\È›Û™HÜˆ˜[YHOHˆŽ‚ˆ™]\›ˆ›Û™BˆžN‚ˆ™]\›ˆ›Ø]
ÝŠ˜[YJKœÝš\

Kœ™\XÙJ‹‹‹ˆŠJBˆ^Ù\
\Q\œ›Ü‹˜[YQ\œ›ÜŠN‚ˆ™]\›ˆ›Û™B‚‚™Yˆ\œÙWÚ[Ý˜[YJ˜[YNˆ[žJHOˆ[›Û™N‚ˆ\œÙYH\œÙWÙ›Ø]Ý˜[YJ˜[YJBˆ™]\›ˆ[
\œÙY
HYˆ\œÙY\È›Ý›Û™H[ÙH›Û™B‚‚™YˆÙ[XÝÜ›ÙXÝ[Û—Ý˜[YJ]WÚ][WÛX\ˆXÝÜÝ‹[žWH›Û™JHOˆ\VÙ›Ø]›Û™KÝ‹Ý—N‚ˆ]HH]WÚ][WÛX\ÜˆßBˆ›ÜˆÙ^H[ˆ
”–ZY[‹š[™\\–ZY[‹š[™\\—ÜÝÙ\ˆŠN‚ˆ˜]×Ý˜[YHH]K™Ù]
Ù^JBˆ˜[YHH\œÙWÙ›Ø]Ý˜[YJ˜]×Ý˜[YJBˆYˆ˜[YH\È›Ý›Û™N‚ˆ™]\›ˆ˜[YKÙ^KÝŠ˜]×Ý˜[YJBˆ™]\›ˆ›Û™Kˆ‹ˆ‚‚‚™YˆÙ[XÝÜ›ÙXÝ[Û—ÚÝÚ
]WÚ][WÛX\ˆXÝÜÝ‹[žWH›Û™JHOˆ›Ø]›Û™N‚ˆ™]\›ˆÙ[XÝÜ›ÙXÝ[Û—Ý˜[YJ]WÚ][WÛX\
VÌB‚‚™YˆZ[ÛZ\ÜÚ[™×Ü›ÙXÝ[Û—Û›ÝJˆ]WÚ][WÛX\ˆXÝÜÝ‹[žWH›Û™Kˆ
‹ˆÝ][Û—ØÛÙNˆÝ‹ˆ\š[ÙÝ\NˆÝ‹ˆ\š[ÙÙ]Nˆ]KŠHOˆÝŽ‚ˆ]˜Z[X›WÚÙ^\ÈHÛÜY
ÝŠÙ^JH›ÜˆÙ^H[ˆ
]WÚ][WÛX\ÜˆßJKšÙ^\Ê
JBˆÙ^\×Ý^H‹‹š›Ú[Š]˜Z[X›WÚÙ^\ÊHYˆ]˜Z[X›WÚÙ^\È[ÙH››Û™H‚ˆ™]\›ˆ
ˆˆ“›È›ÙXÝ[ÛˆÙ^H›Ý[™ˆ]˜Z[X›HÙ^\ÎˆÚÙ^\×Ý^Kˆ‚ˆˆœÝ][ÛÛÙO^ÜÝ][Û—ØÛÙHÜˆ	ËIßNÈ\š[ÙÝ\O^Ü\š[ÙÝ\_NÈ\š[ÙÙ]O^Ü\š[ÙÙ]Kš\ÛÙ›Ü›X]

_H‚ˆ
B‚‚‘•TÒSÓ”ÓÓT—ÔUWÓSRUÐÓÓÓÕÓ—ÒÑVHH™\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]ØÛÛÛÝÛ—Ý[[‚‘•TÒSÓ”ÓÓT—ÓQÐPÖWÔT‘“Ô“PSÑWÐÓÓÓÕÓ—ÒÑVHH™\Ú[ÛœÛÛ\—Ü\™›Ü›X[˜ÙWØÛÛÛÝÛ—Ý[[‚‘•TÒSÓ”ÓÓT—ÔT‘“Ô“PSÑWÐÓÓÓÕÓ—ÒÑVHH•TÒSÓ”ÓÓT—ÔUWÓSRUÐÓÓÓÕÓ—ÒÑVB‘•TÒSÓ”ÓÓT—ÔUWÓSRUÓTÕÐST•ÒÑVHH™\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]Û\ÝØ[\ÚÙ^H‚‘•TÒSÓ”ÓÓT—ÔUWÓSRUÓTÕÐST•ÐUÒÑVHH™\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]Û\ÝØ[\Ø]‚‘•TÒSÓ”ÓÓT—ÔÕUSÓ—ÒS•‘S•Ô–WÑUWÒÑVHH™\Ú[ÛœÛÛ\—ÜÝ][Û—Ú[™[ÜžWÙ]H‚‚‚™YˆÙ\Ú[ÛœÛÛ\—Ù˜Z[ØÛÙJ^[ØYˆXÝÜÝ‹[žWH›Û™JHOˆ[›Û™N‚ˆYˆ›Ý\Ú[œÝ[˜ÙJ^[ØYXÝ
N‚ˆ™]\›ˆ›Û™BˆžN‚ˆ™]\›ˆ[
^[ØY™Ù]
™˜Z[ÛÙHŠJBˆ^Ù\
\Q\œ›Ü‹˜[YQ\œ›ÜŠN‚ˆ™]\›ˆ›Û™B‚‚™Yˆ\×Ù\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]Ü^[ØY
^[ØYˆXÝÜÝ‹[žWH›Û™JHOˆ›ÛÛ‚ˆ™]\›ˆÙ\Ú[ÛœÛÛ\—Ù˜Z[ØÛÙJ^[ØY
HOHÂ‚‚™Yˆ\×Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[Û—Ù^\™YÜ^[ØY
^[ØYˆXÝÜÝ‹[žWH›Û™JHOˆ›ÛÛ‚ˆYˆÙ\Ú[ÛœÛÛ\—Ù˜Z[ØÛÙJ^[ØY
HOHÌN‚ˆ™]\›ˆYBˆ™]\›ˆ•TÑT—ÓUTÕÔ‘SÑÒSˆˆ[ˆÝŠ
^[ØYÜˆßJK™Ù]
›Y\ÜØYÙHŠHÜˆˆŠB‚‚™Yˆ\×Ù\Ú[ÛœÛÛ\—Ú[˜[YØÜ™Y[X[×Ü^[ØY
^[ØYˆXÝÜÝ‹[žWH›Û™JHOˆ›ÛÛ‚ˆÛÙHHÙ\Ú[ÛœÛÛ\—Ù˜Z[ØÛÙJ^[ØY
BˆY\ÜØYÙHHÝŠ
^[ØYÜˆßJK™Ù]
›Y\ÜØYÙHŠHÜˆˆŠK›ÝÙ\Š
Bˆ™]\›ˆÛÙH[ˆÌŒKÌ‹ÌËÌHÜˆœ\ÜÝÛÜ™ˆ[ˆY\ÜØYÙHÜˆ˜Ü™Y[X[ˆ[ˆY\ÜØYÙHÜˆ\Ù\ˆ˜[YHˆ[ˆY\ÜØYÙB‚‚™Yˆ\×Ù\Ú[ÛœÛÛ\—ÚÙ\œ›ÜŠ^Îˆ^Ù\[ÛˆÝŠHOˆ›ÛÛ‚ˆ™]\›ˆ\Ú[œÝ[˜ÙJ^Ë\Ú[Û”ÛÛ\\Q\œ›ÜŠH[™^Ë™\œ›Ü—Ý\HOHš‚‚‚™Yˆ\×Ù\Ú[ÛœÛÛ\—Ú[˜[YÚœÛÛ—Ù\œ›ÜŠ^Îˆ^Ù\[ÛˆÝŠHOˆ›ÛÛ‚ˆ™]\›ˆ\Ú[œÝ[˜ÙJ^Ë\Ú[Û”ÛÛ\\Q\œ›ÜŠH[™^Ë™\œ›Ü—Ý\HOHš[˜[YÚœÛÛˆ‚‚‚™Yˆ\×Ù\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]Ù\œ›ÜŠ^Îˆ^Ù\[ÛˆÝŠHOˆ›ÛÛ‚ˆYˆ\Ú[œÝ[˜ÙJ^Ë\T˜]S[Z]\œ›ÜŠH[™^Ëœ›ÝšY\ˆOHS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓTŽ‚ˆ™]\›ˆYBˆYˆ\Ú[œÝ[˜ÙJ^Ë\Ú[Û”ÛÛ\”˜]S[Z]\œ›ÜŠN‚ˆ™]\›ˆYBˆYˆ\Ú[œÝ[˜ÙJ^Ë\Ú[Û”ÛÛ\\Q\œ›ÜŠH[™\×Ù\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]Ü^[ØY
^Ëœ^[ØY
N‚ˆ™]\›ˆYBˆY\ÜØYÙHHÝŠ^ÊBˆ™]\›ˆ™˜Z[ÛÙOMÈˆ[ˆY\ÜØYÙHÜˆ™\œ›ÜˆÛÙHÈˆ[ˆY\ÜØYÙHÜˆ˜ÛÙYÛÈÈˆ[ˆY\ÜØYÙHÜˆ˜ðàð¬ÙYÛÈÈˆ[ˆY\ÜØYÙB‚‚™Yˆ\×Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[Û—Ù^\™YÙ\œ›ÜŠ^Îˆ^Ù\[ÛˆÝŠHOˆ›ÛÛ‚ˆYˆ\Ú[œÝ[˜ÙJ^Ë\Ú[Û”ÛÛ\”Ù\ÜÚ[Û‘^\™Y\œ›ÜŠN‚ˆ™]\›ˆYBˆYˆ\Ú[œÝ[˜ÙJ^Ë\Ú[Û”ÛÛ\\Q\œ›ÜŠH[™\×Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[Û—Ù^\™YÜ^[ØY
^Ëœ^[ØY
N‚ˆ™]\›ˆYBˆY\ÜØYÙHHÝŠ^ÊBˆ™]\›ˆ™˜Z[ÛÙOLÌHˆ[ˆY\ÜØYÙHÜˆ•TÑT—ÓUTÕÔ‘SÑÒSˆˆ[ˆY\ÜØYÙB‚‚™YˆÙ]Ø\ÜÝ]WÝ˜[YJÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹Ù^NˆÝŠHOˆÝŽ‚ˆ›ÝÈHÛÛ›‹™^XÝ]J”ÑSPÕ˜[YH”“ÓH\ÜÝ]HÒT‘HÙ^HHÈ‹
Ù^K
JK™™]ÚÛ™J
Bˆ™]\›ˆÝŠ›ÝÖÈ˜[YH—HÜˆˆŠHYˆ›ÝÈ[ÙHˆ‚‚‚™YˆÙ]Ø\ÜÝ]WÝ˜[YJÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹Ù^NˆÝ‹˜[YNˆÝŠHOˆ›Û™N‚ˆ›ÝÈH]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠBˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆS”ÑT•S•È\ÜÝ]H
Ù^K˜[YK\]YØ]
BˆSQTÈ
ËËÊBˆÓˆÓÓ‘“PÕ
Ù^JHÈTUHÑU˜[YHH^ÛYY˜[YK\]YØ]H^ÛYY\]YØ]ˆˆˆ‹ˆ
Ù^K˜[YK›ÝÊKˆ
B‚‚™Yˆ›ÝYžWÙ\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]
ÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ÛÛÛÝÛ—Ý[[ˆ]][YKY\ÜØYÙNˆÝŠHOˆ›Û™N‚ˆ›Ý×Ý˜[YHH]][YK››ÝÊ
Bˆ™XÙ[ØÝ]Ù™ˆH
›Ý×Ý˜[YHH[YY[JZ[]\ÏY\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]ÛZ[]\Ê
JJKš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠBˆ™XÙ[Ø[\HÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕBˆ”“ÓH[YÜ˜[WØ[\ÂˆÒT‘H[\Ý\HH	Ù\Ú[ÛœÛÛ\—Ø\WÛ[Z]	ÂˆS‘Ý]\ÈSˆ
	ÜÙ[	Ë	Ø›ØÚÙY	Ë	Ù˜Z[Y	ÊBˆS‘Ù[Ø]HÂˆSRUBˆˆˆ‹ˆ
™XÙ[ØÝ]Ù™‹
Kˆ
K™™]ÚÛ™J
BˆYˆ™XÙ[Ø[\\È›Ý›Û™N‚ˆ™]\›‚ˆ\ÝØ[\Ø]HÙ]Ø\ÜÝ]WÝ˜[YJÛÛ›‹•TÒSÓ”ÓÓT—ÔUWÓSRUÓTÕÐST•ÐUÒÑVJBˆYˆ\ÝØ[\Ø]‚ˆžN‚ˆYˆ]][YK™œ›ÛZ\ÛÙ›Ü›X]
\ÝØ[\Ø]
HH›Ý×Ý˜[YHH[YY[JZ[]\ÏY\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]ÛZ[]\Ê
JN‚ˆ™]\›‚ˆ^Ù\˜[YQ\œ›ÜŽ‚ˆ\ÜÂ‚ˆ[\ÚÙ^HHˆ™\Ú[ÛœÛÛ\—Ø\WÛ[Z]žÛ›Ý×Ý˜[YKœÝ™[YJ	ÉVI[IY	R	SIÊ_H‚ˆÙ]Ø\ÜÝ]WÝ˜[YJÛÛ›‹•TÒSÓ”ÓÓT—ÔUWÓSRUÓTÕÐST•ÒÑVK[\ÚÙ^JBˆÙ]Ø\ÜÝ]WÝ˜[YJÛÛ›‹•TÒSÓ”ÓÓT—ÔUWÓSRUÓTÕÐST•ÐUÒÑVK›Ý×Ý˜[YKš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠJBˆÛÛ›‹˜ÛÛ[Z]

Bˆ[YÜ˜[WÛY\ÜØYÙHH
ˆ‘\Ú[Û”ÛÛ\ˆTH[Z]YOØ——ˆ‚ˆHTH]›Û™]H[Z]HHÚ[XY\ËÝÚÙ[œÈ\ÙÛÝYÈ\˜HÔ\ÈH›ÙpéðèÛË—ˆ‚ˆˆžÛY\ÜØYÙ_W—ˆ‚ˆ“ÜÈ™[]0ìÜš[ÜÈH›ÙpéðèÛÈH˜XÚÙš[ÈÙ[H˜[\ˆ]0êHÈ[Z]H™[›Ý˜\‹ˆ‚ˆ
BˆÙ[™Ø[™Ü™XÛÜ™Ý[YÜ˜[WØ[\
ˆÛÛ›‹ˆ›Û™Kˆ™\Ú[ÛœÛÛ\—Ø\WÛ[Z]‹ˆ[\ÚÙ^Kˆ[YÜ˜[WÛY\ÜØYÙKˆ
B‚‚™Yˆ\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]ÛZ[]\Ê
HOˆ[‚ˆ˜]×Ý˜[YHHÜË™[š\›Û‹™Ù]
‘•TÒSÓ”ÓÓT—ÔUWÓSRUÓRS•UTÈ‹ÝŠQUSÑ•TÒSÓ”ÓÓT—ÔUWÓSRUÓRS•UTÊJKœÝš\

BˆžN‚ˆ™]\›ˆX^
K[
˜]×Ý˜[YJJBˆ^Ù\˜[YQ\œ›ÜŽ‚ˆ™]\›ˆQUSÑ•TÒSÓ”ÓÓT—ÔUWÓSRUÓRS•UTÂ‚‚™Yˆ\Ú[ÛœÛÛ\—Ü›ÙXÝ[Û—ÚÜWÜÛXÞJ
HOˆ\T]Y]YTÛXÞN‚ˆZ[WØYÙ]H\œÙWÙ[—ÜÜÚ]]™WÚ[
ˆ‘•TÒSÓ”ÓÓT—Ô“ÑPÕSÓ—ÒÔWÑRSWÐ•QÑU‹ˆQUSÑ•TÒSÓ”ÓÓT—Ô“ÑPÕSÓ—ÒÔWÑRSWÐ•QÑUˆ
BˆZ[WÜ™\Ù\™YØØ[ÈHZ[Šˆ\œÙWÙ[—Û›Û›™YØ]]™WÚ[
ˆ‘•TÒSÓ”ÓÓT—Ô“ÑPÕSÓ—ÒÔWÑRSWÔ‘TÑT•‘QÐÐSÈ‹ˆQUSÑ•TÒSÓ”ÓÓT—Ô“ÑPÕSÓ—ÒÔWÑRSWÔ‘TÑT•‘QÐÐSËˆ
KˆZ[WØYÙ]ˆ
Bˆ[ÛØÛÜÙWÜ™\Ù\™YØØ[ÈHZ[Šˆ\œÙWÙ[—Û›Û›™YØ]]™WÚ[
ˆ‘•TÒSÓ”ÓÓT—Ô“ÑPÕSÓ—ÒÔWÓSÓ•ÐÓÔÑWÔ‘TÑT•‘QÐÐSÈ‹ˆQUSÑ•TÒSÓ”ÓÓT—Ô“ÑPÕSÓ—ÒÔWÓSÓ•ÐÓÔÑWÔ‘TÑT•‘QÐÐSËˆ
KˆX^
Z[WØYÙ]HZ[WÜ™\Ù\™YØØ[Ë
Kˆ
Bˆ™]\›ˆ\T]Y]YTÛXÞJˆZ[—Ú[\˜[ÜÙXÛÛ™Ï\\œÙWÙ[—ÜÜÚ]]™WÚ[
ˆ‘•TÒSÓ”ÓÓT—Ô“ÑPÕSÓ—ÒÔWÓRS—ÒS•T•SÔÑPÓÓ‘È‹ˆQUSÑ•TÒSÓ”ÓÓT—Ô“ÑPÕSÓ—ÒÔWÓRS—ÒS•T•SÔÑPÓÓ‘Ëˆ
KˆZ[WØYÙ]YZ[WØYÙ]ˆ™\Ù\™YØØ[×ØžWÜš[Üš]OJˆ
KZ[WÜ™\Ù\™YØØ[ÊKˆ
‹[ÛØÛÜÙWÜ™\Ù\™YØØ[ÊKˆ
Kˆ
B‚‚™YˆÚYÙ[™\™ÞWÜ›ÙXÝ[Û—ÚÜWÜÛXÞJ
HOˆ\T]Y]YTÛXÞN‚ˆ™]\›ˆ\T]Y]YTÛXÞJˆZ[—Ú[\˜[ÜÙXÛÛ™Ï\\œÙWÙ[—ÛÜ[Û˜[ÜÜÚ]]™WÚ[
ˆ”ÒQÑS‘T‘ÖWÔ“ÑPÕSÓ—ÓRS—ÒS•T•SÔÑPÓÓ‘È‚ˆ
KˆZ[WØYÙ]\\œÙWÙ[—ÛÜ[Û˜[ÜÜÚ]]™WÚ[
ˆ”ÒQÑS‘T‘ÖWÔ“ÑPÕSÓ—ÑRSWÐ•QÑU‚ˆ
Kˆ
B‚‚™Yˆ\Ú[ÛœÛÛ\—ÝØ]ÜÛXÞJ
HOˆ\T]Y]YTÛXÞN‚ˆ™]\›ˆ\T]Y]YTÛXÞJˆZ[—Ú[\˜[ÜÙXÛÛ™ÏS›Û™KˆZ[WØYÙ]\\œÙWÙ[—ÜÜÚ]]™WÚ[
ˆ‘•TÒSÓ”ÓÓT—ÕÐUÑRSWÐ•QÑU‹ˆQUSÑ•TÒSÓ”ÓÓT—ÕÐUÑRSWÐ•QÑUˆ
Kˆ
B‚‚™Yˆ\œÙWÙ[—ÜÜÚ]]™WÚ[
˜[YNˆÝ‹Y˜][ˆ[
HOˆ[‚ˆžN‚ˆ™]\›ˆX^
K[
ÜË™[š\›Û‹™Ù]
˜[YKÝŠY˜][
JKœÝš\

JJBˆ^Ù\˜[YQ\œ›ÜŽ‚ˆ™]\›ˆY˜][‚‚™Yˆ\œÙWÙ[—Û›Û›™YØ]]™WÚ[
˜[YNˆÝ‹Y˜][ˆ[
HOˆ[‚ˆžN‚ˆ™]\›ˆX^
[
ÜË™[š\›Û‹™Ù]
˜[YKÝŠY˜][
JKœÝš\

JJBˆ^Ù\˜[YQ\œ›ÜŽ‚ˆ™]\›ˆY˜][‚‚™Yˆ\œÙWÙ[—ÛÜ[Û˜[ÜÜÚ]]™WÚ[
˜[YNˆÝŠHOˆ[›Û™N‚ˆ˜]×Ý˜[YHHÜË™[š\›Û‹™Ù]
˜[YKˆŠKœÝš\

BˆYˆ›Ý˜]×Ý˜[YN‚ˆ™]\›ˆ›Û™BˆžN‚ˆ™]\›ˆX^
K[
˜]×Ý˜[YJJBˆ^Ù\˜[YQ\œ›ÜŽ‚ˆ™]\›ˆ›Û™B‚‚™Yˆ›ÙXÝ[Û—Ú›Ø—Üš[Üš]J›Ø—Ý\NˆÝŠHOˆ[‚ˆYˆ›Ø—Ý\HOH™\Ú[ÛœÛÛ\—Ü›ÙXÝ[Û—ÜÞ[˜ÈŽ‚ˆ™]\›ˆBˆYˆ›Ø—Ý\HOH™\Ú[ÛœÛÛ\—Û[ÛØÛÜÙHŽ‚ˆ™]\›ˆ‚ˆYˆ›Ø—Ý\H[ˆÈ™\Ú[ÛœÛÛ\—Ü›ÙXÝ[Û—Ø˜XÚÙš[‹™\Ú[ÛœÛÛ\—Û[ÛØÞXÛHŸN‚ˆ™]\›ˆÂˆYˆ›Ø—Ý\HOH™\Ú[ÛœÛÛ\—Ü™\ÜÜ›ÙXÝ[Û—Ü™\]Y\ÝŽ‚ˆ™]\›ˆˆYˆ›Ø—Ý\H[ˆ•TÒSÓ”ÓÓT—ÕÐUÒ“Ð—ÕTTÈÜˆ›Ø—Ý\HOH™\Ú[ÛœÛÛ\—ÜÝ]WÜÞ[˜ÈŽ‚ˆ™]\›ˆBˆ™]\›ˆ‚‚™YˆYÚ\—Üš[Üš]WÜ›ÙXÝ[Û—Ú›Ø—Ý\\Êš[Üš]Nˆ[
HOˆ\VÜÝ‹‹‹—N‚ˆ™]\›ˆ\Jˆ›Ø—Ý\Bˆ›Üˆ›Ø—Ý\H[ˆ•TÒSÓ”ÓÓT—Ô“ÑPÕSÓ—Ò“Ð—ÕTTÂˆYˆ›ÙXÝ[Û—Ú›Ø—Üš[Üš]J›Ø—Ý\JHš[Üš]Bˆ
B‚‚ÛÛ^X[˜YÙ\‚™Yˆ›ÙXÝ[Û—ÚÜWØØ[ØÛÛ^
ˆ
‹ˆ›Ø—ÚYˆ[ˆ›Ø—Ý\NˆÝ‹ŠHOˆ[žN‚ˆÚÙ[ˆH“ÑPÕSÓ—ÒÔWÐÐSÐÓÓ•VœÙ]
ˆÂˆš›Ø—ÚYŽˆ›Ø—ÚYˆš›Ø—Ý\HŽˆ›Ø—Ý\Kˆœš[Üš]HŽˆ›ÙXÝ[Û—Ú›Ø—Üš[Üš]J›Ø—Ý\JKˆ›X\ÙWÛÝÛ™\ˆŽˆˆ˜˜XÚÙÜ›Ý[™Z›Ø‹^Ú›Ø—ÚYH‹ˆBˆ
BˆžN‚ˆZY[ˆš[˜[N‚ˆ“ÑPÕSÓ—ÒÔWÐÐSÐÓÓ•Vœ™\Ù]
ÚÙ[ŠB‚‚™Yˆ\Ú[ÛœÛÛ\—Ü›ÙXÝ[Û—ØXØÛÝ[ÚÙ^JˆÛÛ™šYÎˆÜ[]LË”›ÝÈXÝÜÝ‹[žWKˆÙ[™Ú[ˆÝ‹ŠHOˆÝŽ‚ˆ™]\›ˆ\Ú[ÛœÛÛ\—ØXØÛÝ[ÚÙ^JÛÛ™šYÊB‚‚™Yˆ\Ú[ÛœÛÛ\—ØXØÛÝ[ÚÙ^JˆÛÛ™šYÎˆÜ[]LË”›ÝÈXÝÜÝ‹[žWKŠHOˆÝŽ‚ˆ™]\›ˆ›ÙXÝ[Û—Ø\WØXØÛÝ[ÚÙ^Jˆ›ÝšY\RS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹ˆ\Ù\›˜[YO\ÝŠÛÛ™šYË™Ù]
\Ù\›˜[YHŠHYˆ\Ú[œÝ[˜ÙJÛÛ™šYËXÝ
H[ÙHÛÛ™šYÖÈ\Ù\›˜[YH—HÜˆˆŠKˆ˜\ÙWÝ\›\ÝŠÛÛ™šYË™Ù]
˜˜\ÙWÝ\›ŠHYˆ\Ú[œÝ[˜ÙJÛÛ™šYËXÝ
H[ÙHÛÛ™šYÖÈ˜˜\ÙWÝ\›—HÜˆˆŠKˆ[™Ú[H˜XØÛÝ[‹ˆ
B‚‚™Yˆ^XÝ]WÜ]Y]YYÙ\Ú[ÛœÛÛ\—ÚÜWØØ[
ˆØ[˜XÚÎˆ[žKˆ
‹ˆ[™Ú[ˆÝ‹ŠHOˆ[žN‚ˆ™]\›ˆ^XÝ]WÜ]Y]YYÙ\Ú[ÛœÛÛ\—ØØ[
ˆØ[˜XÚËˆ\WØ\™XOT“ÑPÕSÓ—ÒÔWÐT‘PKˆÛXÞOY\Ú[ÛœÛÛ\—Ü›ÙXÝ[Û—ÚÜWÜÛXÞJ
Kˆ[™›Ü˜ÙWÜ›ÙXÝ[Û—Üš[Üš]OUYKˆ
B‚‚™Yˆ^XÝ]WÜ]Y]YYÙ\Ú[ÛœÛÛ\—ÙXYÛ›ÜÝXÜ×ØØ[
Ø[˜XÚÎˆ[žJHOˆ[žN‚ˆ™]\›ˆ^XÝ]WÜ]Y]YYÙ\Ú[ÛœÛÛ\—ØØ[
ˆØ[˜XÚËˆ\WØ\™XOUÐUÒTÕÔ–WÐT‘PKˆÛXÞOY\Ú[ÛœÛÛ\—ÝØ]ÜÛXÞJ
Kˆ[™›Ü˜ÙWÜ›ÙXÝ[Û—Üš[Üš]OUYKˆ
B‚‚™Yˆ^XÝ]WÜ]Y]YYÙ\Ú[ÛœÛÛ\—ØXØÛÝ[ØØ[
Ø[˜XÚÎˆ[žJHOˆ[žN‚ˆ™]\›ˆ^XÝ]WÜ]Y]YYÙ\Ú[ÛœÛÛ\—ØØ[
ˆØ[˜XÚËˆ\WØ\™XOPTWÐT‘PWÔÕUKˆÛXÞOS›Û™Kˆ[™›Ü˜ÙWÜ›ÙXÝ[Û—Üš[Üš]OUYKˆ
B‚‚™Yˆ^XÝ]WÜ]Y]YYÙ\Ú[ÛœÛÛ\—ØØ[
ˆØ[˜XÚÎˆ[žKˆ
‹ˆ\WØ\™XNˆÝ‹ˆÛXÞNˆ\T]Y]YTÛXÞH›Û™Kˆ[™›Ü˜ÙWÜ›ÙXÝ[Û—Üš[Üš]Nˆ›ÛÛŠHOˆ[žN‚ˆÛÛ^H“ÑPÕSÓ—ÒÔWÐÐSÐÓÓ•V™Ù]

BˆYˆÛÛ^\È›Û™HÜˆ›Ý\×Ø\ØÛÛ^

N‚ˆ™]\›ˆØ[˜XÚÊ
BˆÚ]ÛÜÚ[™ÊÙ]ÙŠÝ\œ™[Ø\˜ÛÛ™šYÖÈ‘UPTÑH—JJH\È]Y]YWØÛÛ›Ž‚ˆÛÛ™šYÈHÙ]Ú[YÜ˜][Û—ØÛÛ™šYÊ]Y]YWØÛÛ›‹S•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓTŠBˆYˆÛÛ™šYÈ\È›Û™N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠÛÛ™šYÝ\˜XØ[È\Ú[Û”ÛÛ\ˆ˜[È[˜ÛÛ˜YKˆŠBˆXØÛÝ[ÚÙ^WÝ˜[YHH\Ú[ÛœÛÛ\—ØXØÛÝ[ÚÙ^JÛÛ™šYÊBˆYÚ\—Üš[Üš]WÝ\\ÈH
ˆYÚ\—Üš[Üš]WÜ›ÙXÝ[Û—Ú›Ø—Ý\\Ê[
ÛÛ^Èœš[Üš]H—JJBˆYˆ[™›Ü˜ÙWÜ›ÙXÝ[Û—Üš[Üš]Bˆ[ÙH

Bˆ
BˆYˆYÚ\—Üš[Üš]WÝ\\Î‚ˆXÙZÛ\œÈH‹‹š›Ú[ŠÈˆ›ÜˆÈ[ˆYÚ\—Üš[Üš]WÝ\\ÊBˆYÚ\—Üš[Üš]WÚ›ØˆH]Y]YWØÛÛ›‹™^XÝ]Jˆˆˆˆ‚ˆÑSPÕY™^Ø][\Ø]ˆ”“ÓH˜XÚÙÜ›Ý[™Ú›ØœÂˆÒT‘H›Ø—Ý\HSˆ
ÜXÙZÛ\œßJBˆS‘Ý]\ÈSˆ
	Ü[™[™ÉË	Ü[›š[™ÉË	ÝØZ][™×Ø\WÜÛÝ	Ë	ÝØZ][™×Ü˜]WÛ[Z]	ÊBˆS‘YOHÂˆÔ‘Tˆ–HYˆSRUBˆˆˆ‹ˆ

šYÚ\—Üš[Üš]WÝ\\Ë[
ÛÛ^Èš›Ø—ÚY—JJKˆ
K™™]ÚÛ™J
BˆYˆYÚ\—Üš[Üš]WÚ›Øˆ\È›Ý›Û™N‚ˆ™^Ø][\Ø]H]][YK››ÝÊTÐ“Ó—ÕSQV“Ó‘JH
È[YY[JˆÙXÛÛ™ÏY\Ú[ÛœÛÛ\—Ü›ÙXÝ[Û—ÚÜWÜÛXÞJ
K›Z[—Ú[\˜[ÜÙXÛÛ™ÂˆÜˆBˆ
Bˆ˜Z\ÙH\TÛÝ[˜]˜Z[X›Q\œ›ÜŠˆ›ÝšY\RS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹ˆXØÛÝ[ÚÙ^OXXØÛÝ[ÚÙ^WÝ˜[YKˆ\WØ\™XOX\WØ\™XKˆ™^Ø][\Ø][™^Ø][\Ø]ˆØZ]Ü™X\ÛÛHœš[Üš]WÜ]Y]YH‹ˆ
BˆXØÛÝ[Ü™\Ù\˜][ÛˆH™\Ù\™WØXØÛÝ[ÛX\ÙJˆ]Y]YWØÛÛ›‹ˆ›ÝšY\RS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹›ÝÙ\Š
KˆXØÛÝ[ÚÙ^WÝ˜[YOXXØÛÝ[ÚÙ^WÝ˜[YKˆX\ÙWÛÝÛ™\\ÝŠÛÛ^È›X\ÙWÛÝÛ™\ˆ—JKˆ
BˆYˆ›ÝXØÛÝ[Ü™\Ù\˜][Û‹™Ü˜[Y‚ˆ˜Z\ÙH\TÛÝ[˜]˜Z[X›Q\œ›ÜŠˆ›ÝšY\RS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹ˆXØÛÝ[ÚÙ^OXXØÛÝ[ÚÙ^WÝ˜[YKˆ\WØ\™XOX\WØ\™XKˆ™^Ø][\Ø]XXØÛÝ[Ü™\Ù\˜][Û‹›™^Ø][\Ø]ˆØZ]Ü™X\ÛÛXXØÛÝ[Ü™\Ù\˜][Û‹ØZ]Ü™X\ÛÛ‹ˆ
Bˆ\™XWÜ™\Ù\™YH˜[ÙBˆžN‚ˆYˆÛXÞH\È›Ý›Û™N‚ˆ™\Ù\˜][ÛˆH™\Ù\™WØ\WÜÛÝ
ˆ]Y]YWØÛÛ›‹ˆ›ÝšY\RS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹›ÝÙ\Š
KˆXØÛÝ[ÚÙ^WÝ˜[YOXXØÛÝ[ÚÙ^WÝ˜[YKˆ\WØ\™XOX\WØ\™XKˆX\ÙWÛÝÛ™\\ÝŠÛÛ^È›X\ÙWÛÝÛ™\ˆ—JKˆš[Üš]OZ[
ÛÛ^Èœš[Üš]H—JKˆÛXÞO\ÛXÞKˆ
BˆYˆ›Ý™\Ù\˜][Û‹™Ü˜[Y‚ˆ˜Z\ÙH\TÛÝ[˜]˜Z[X›Q\œ›ÜŠˆ›ÝšY\RS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹ˆXØÛÝ[ÚÙ^OXXØÛÝ[ÚÙ^WÝ˜[YKˆ\WØ\™XOX\WØ\™XKˆ™^Ø][\Ø]\™\Ù\˜][Û‹›™^Ø][\Ø]ˆØZ]Ü™X\ÛÛ\™\Ù\˜][Û‹ØZ]Ü™X\ÛÛ‹ˆ
Bˆ\™XWÜ™\Ù\™YHYBˆ™]\›ˆØ[˜XÚÊ
Bˆ^Ù\^Ù\[Ûˆ\È^Î‚ˆYˆ\×Ù\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]Ù\œ›ÜŠ^ÊN‚ˆÛÛÛÝÛ—Ý[[H]][YK››ÝÊTÐ“Ó—ÕSQV“Ó‘JH
È[YY[JˆZ[]\ÏY\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]ÛZ[]\Ê
Bˆ
Bˆ™XÛÜ™ØXØÛÝ[ÍÊˆ]Y]YWØÛÛ›‹ˆ›ÝšY\RS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹›ÝÙ\Š
KˆXØÛÝ[ÚÙ^WÝ˜[YOXXØÛÝ[ÚÙ^WÝ˜[YKˆÛÛÛÝÛ—Ý[[XÛÛÛÝÛ—Ý[[ˆ
BˆYˆÛXÞH\È›Ý›Û™N‚ˆ™XÛÜ™Ü›ÙXÝ[Û—Ø\WÍÊˆ]Y]YWØÛÛ›‹ˆ›ÝšY\RS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹›ÝÙ\Š
KˆXØÛÝ[ÚÙ^WÝ˜[YOXXØÛÝ[ÚÙ^WÝ˜[YKˆ\WØ\™XOX\WØ\™XKˆÛÛÛÝÛ—Ý[[XÛÛÛÝÛ—Ý[[ˆ
BˆÜÝÛ™WÜ[™[™×Ü›ÙXÝ[Û—Ú›Øœ×ØY\—ÍÊˆ]Y]YWØÛÛ›‹ˆÛÛÛÝÛ—Ý[[XÛÛÛÝÛ—Ý[[ˆ^ÛYWÚ›Ø—ÚYZ[
ÛÛ^Èš›Ø—ÚY—JKˆ
Bˆ˜Z\ÙBˆš[˜[N‚ˆYˆ\™XWÜ™\Ù\™Y‚ˆ™[X\ÙWØ\WÛX\ÙJˆ]Y]YWØÛÛ›‹ˆ›ÝšY\RS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹›ÝÙ\Š
KˆXØÛÝ[ÚÙ^WÝ˜[YOXXØÛÝ[ÚÙ^WÝ˜[YKˆ\WØ\™XOX\WØ\™XKˆX\ÙWÛÝÛ™\\ÝŠÛÛ^È›X\ÙWÛÝÛ™\ˆ—JKˆ
Bˆ™[X\ÙWØXØÛÝ[ÛX\ÙJˆ]Y]YWØÛÛ›‹ˆ›ÝšY\RS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹›ÝÙ\Š
KˆXØÛÝ[ÚÙ^WÝ˜[YOXXØÛÝ[ÚÙ^WÝ˜[YKˆX\ÙWÛÝÛ™\\ÝŠÛÛ^È›X\ÙWÛÝÛ™\ˆ—JKˆ
B‚‚™Yˆ\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[Û—ØØXÚWÛZ[]\Ê
HOˆ[‚ˆ˜]×Ý˜[YHHÜË™[š\›Û‹™Ù]
‘•TÒSÓ”ÓÓT—ÔÑTÔÒSÓ—ÐÐPÒWÓRS•UTÈ‹ÝŠQUSÑ•TÒSÓ”ÓÓT—ÔÑTÔÒSÓ—ÐÐPÒWÓRS•UTÊJKœÝš\

BˆžN‚ˆ™]\›ˆX^
K[
˜]×Ý˜[YJJBˆ^Ù\˜[YQ\œ›ÜŽ‚ˆ™]\›ˆQUSÑ•TÒSÓ”ÓÓT—ÔÑTÔÒSÓ—ÐÐPÒWÓRS•UTÂ‚‚™YˆØÝ\œ™[Ø\Ù—ØÛÛ›™XÝ[ÛŠ
HOˆÜ[]LËÛÛ›™XÝ[Ûˆ›Û™N‚ˆYˆ›Ý\×Ø\ØÛÛ^

N‚ˆ™]\›ˆ›Û™Bˆ]X˜\ÙHHÝ\œ™[Ø\˜ÛÛ™šYË™Ù]
‘UPTÑHŠBˆ™]\›ˆÙ]ÙŠ]X˜\ÙJHYˆ]X˜\ÙH[ÙH›Û™B‚‚™YˆÙ]Ù\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]Ý[[
ˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Ûˆ›Û™HH›Û™Kˆ›Ý×Ý˜[YNˆ]][YH›Û™HH›Û™KŠHOˆ]][YH›Û™N‚ˆÛÛÛÝÛ—Ý[[H•TÒSÓ”ÓÓT—ÔT‘“Ô“PSÑWÔUWÓSRUÕS•SˆÛÜÙWØÛÛ›ˆH˜[ÙBˆYˆÛÛ›ˆ\È›Û™N‚ˆÛÛ›ˆHØÝ\œ™[Ø\Ù—ØÛÛ›™XÝ[ÛŠ
BˆÛÜÙWØÛÛ›ˆHÛÛ›ˆ\È›Ý›Û™BˆžN‚ˆYˆÛÛ›ˆ\È›Ý›Û™N‚ˆ›Üˆ\™XH[ˆ
TWÐT‘PWÔÕUKTWÐT‘PWÔ“ÑPÕSÓ‹TWÐT‘PWÑPQÓ“ÔÕPÔÊN‚ˆ\™XWÝ[[HXÝ]™WØÛÛÛÝÛ—Ý[[
ÛÛ›‹S•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹\™XJBˆYˆ\™XWÝ[[[™
ÛÛÛÝÛ—Ý[[\È›Û™HÜˆ\™XWÝ[[ˆÛÛÛÝÛ—Ý[[
N‚ˆÛÛÛÝÛ—Ý[[H\™XWÝ[[ˆ›ÜˆÙ^H[ˆ
•TÒSÓ”ÓÓT—ÔUWÓSRUÐÓÓÓÕÓ—ÒÑVK•TÒSÓ”ÓÓT—ÓQÐPÖWÔT‘“Ô“PSÑWÐÓÓÓÕÓ—ÒÑVJN‚ˆ˜]×Ý˜[YHHÙ]Ø\ÜÝ]WÝ˜[YJÛÛ›‹Ù^JBˆYˆ˜]×Ý˜[YN‚ˆžN‚ˆ\œÚ\ÝYÝ[[H]][YK™œ›ÛZ\ÛÙ›Ü›X]
˜]×Ý˜[YJBˆYˆÛÛÛÝÛ—Ý[[\È›Û™HÜˆ\œÚ\ÝYÝ[[ˆÛÛÛÝÛ—Ý[[‚ˆÛÛÛÝÛ—Ý[[H\œÚ\ÝYÝ[[ˆ^Ù\˜[YQ\œ›ÜŽ‚ˆ\ÜÂˆš[˜[N‚ˆYˆÛÜÙWØÛÛ›ˆ[™ÛÛ›ˆ\È›Ý›Û™N‚ˆÛÛ›‹˜ÛÜÙJ
BˆYˆÛÛÛÝÛ—Ý[[[™›Ý×Ý˜[YH[™ÛÛÛÝÛ—Ý[[H›Ý×Ý˜[YN‚ˆ™]\›ˆ›Û™Bˆ™]\›ˆÛÛÛÝÛ—Ý[[‚‚™YˆÙ]Ù\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]ØÛÛÛÝÛ—Ü™X\ÛÛŠˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Ûˆ›Û™HH›Û™Kˆ›Ý×Ý˜[YNˆ]][YH›Û™HH›Û™KŠHOˆÝŽ‚ˆ›Ý×Ý˜[YHH›Ý×Ý˜[YHÜˆ]][YK››ÝÊ
BˆÛÛÛÝÛ—Ý[[HÙ]Ù\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]Ý[[
ÛÛ›‹›Ý×Ý˜[YJBˆYˆÛÛÛÝÛ—Ý[[[™ÛÛÛÝÛ—Ý[[ˆ›Ý×Ý˜[YN‚ˆ™[XZ[š[™×ÜÙXÛÛ™ÈH[

ÛÛÛÝÛ—Ý[[H›Ý×Ý˜[YJKÝ[ÜÙXÛÛ™Ê
JBˆ™[XZ[š[™×ÛZ[]\ÈHX^
K
™[XZ[š[™×ÜÙXÛÛ™È
ÈNJHËÈŒ
Bˆ™]\›ˆ
ˆ‘\Ú[Û”ÛÛ\ˆ[\Ü˜\šX[Y[H[Z]YÈ[HTKˆ‚ˆˆ“›Ý˜H[]]˜H\ÜÛš]™[\ÜÈØÛÛÛÝÛ—Ý[[œÝ™[YJ	ÉR‰SIÊ_H
Ü™[XZ[š[™×ÛZ[]\ßHZ[ŠKˆ‚ˆ
Bˆ™]\›ˆˆ‚‚‚™YˆÙ]Ù\Ú[ÛœÛÛ\—Ü\™›Ü›X[˜ÙWØÛÛÛÝÛ—Ü™X\ÛÛŠˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Ûˆ›Û™HH›Û™Kˆ›Ý×Ý˜[YNˆ]][YH›Û™HH›Û™KŠHOˆÝŽ‚ˆ™]\›ˆÙ]Ù\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]ØÛÛÛÝÛ—Ü™X\ÛÛŠÛÛ›‹›Ý×Ý˜[YJB‚‚™YˆX\š×Ù\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]Y
ˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Ûˆ›Û™HH›Û™Kˆ›Ý×Ý˜[YNˆ]][YH›Û™HH›Û™KŠHOˆÝŽ‚ˆÛØ˜[•TÒSÓ”ÓÓT—ÔT‘“Ô“PSÑWÔUWÓSRUÕS•Sˆ›Ý×Ý˜[YHH›Ý×Ý˜[YHÜˆ]][YK››ÝÊ
Bˆ•TÒSÓ”ÓÓT—ÔT‘“Ô“PSÑWÔUWÓSRUÕS•SH›Ý×Ý˜[YH
È[YY[JZ[]\ÏY\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]ÛZ[]\Ê
JBˆÛÜÙWØÛÛ›ˆH˜[ÙBˆYˆÛÛ›ˆ\È›Û™N‚ˆÛÛ›ˆHØÝ\œ™[Ø\Ù—ØÛÛ›™XÝ[ÛŠ
BˆÛÜÙWØÛÛ›ˆHÛÛ›ˆ\È›Ý›Û™BˆYˆÛÛ›ˆ\È›Ý›Û™N‚ˆÙ]Ø\ÜÝ]WÝ˜[YJˆÛÛ›‹ˆ•TÒSÓ”ÓÓT—ÔUWÓSRUÐÓÓÓÕÓ—ÒÑVKˆ•TÒSÓ”ÓÓT—ÔT‘“Ô“PSÑWÔUWÓSRUÕS•Sš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆ
BˆY\ÜØYÙHH
ˆ‘\Ú[Û”ÛÛ\ˆ[\Ü˜\šX[Y[H[Z]YÈ[HTKˆ‚ˆˆ“›Ý˜H[]]˜H\ÜÛš]™[\ÜÈÑ•TÒSÓ”ÓÓT—ÔT‘“Ô“PSÑWÔUWÓSRUÕS•SœÝ™[YJ	ÉR‰SIÊ_Kˆ‚ˆ
BˆYˆÛÛ›ˆ\È›Ý›Û™N‚ˆ›ÝYžWÙ\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]
ÛÛ›‹•TÒSÓ”ÓÓT—ÔT‘“Ô“PSÑWÔUWÓSRUÕS•SY\ÜØYÙJBˆÛÛ›‹˜ÛÛ[Z]

BˆYˆÛÜÙWØÛÛ›ˆ[™ÛÛ›ˆ\È›Ý›Û™N‚ˆÛÛ›‹˜ÛÜÙJ
BˆÑÑÑT‹Ø\›š[™Êˆ‘\Ú[Û”ÛÛ\ˆ˜]H[Z]ÛÛÛÝÛˆXÝ]˜]Y[[	\È‹ˆ•TÒSÓ”ÓÓT—ÔT‘“Ô“PSÑWÔUWÓSRUÕS•Sš\ÛÙ›Ü›X]
[Y\ÜXÏH›Z[]\ÈŠKˆ
Bˆ™]\›ˆY\ÜØYÙB‚‚™YˆX\š×Ù\Ú[ÛœÛÛ\—Ø\WØÛÛÛÝÛŠˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Ûˆ›Û™Kˆ\WØ\™XNˆÝ‹ˆ
‹ˆ™X\ÛÛŽˆÝˆH‘\Ú[Û”ÛÛ\ˆ˜]H[Z]‹ˆ›Ý×Ý˜[YNˆ]][YH›Û™HH›Û™KŠHOˆ]][YN‚ˆÛØ˜[•TÒSÓ”ÓÓT—ÔT‘“Ô“PSÑWÔUWÓSRUÕS•Sˆ›Ý×Ý˜[YHH›Ý×Ý˜[YHÜˆ]][YK››ÝÊ
Bˆ[[H›Ý×Ý˜[YH
È[YY[JZ[]\ÏY\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]ÛZ[]\Ê
JBˆ•TÒSÓ”ÓÓT—ÔT‘“Ô“PSÑWÔUWÓSRUÕS•SH[[ˆÛÜÙWØÛÛ›ˆH˜[ÙBˆYˆÛÛ›ˆ\È›Û™N‚ˆÛÛ›ˆHØÝ\œ™[Ø\Ù—ØÛÛ›™XÝ[ÛŠ
BˆÛÜÙWØÛÛ›ˆHÛÛ›ˆ\È›Ý›Û™BˆžN‚ˆYˆÛÛ›ˆ\È›Ý›Û™N‚ˆX\š×Ø\WØÛÛÛÝÛŠˆÛÛ›‹ˆS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹ˆ\WØ\™XKˆ™X\ÛÛ‹ˆÛÛÛÝÛ—Ý[[][[ˆ›ÝÏ[›Ý×Ý˜[YKˆ
BˆÙ]Ø\ÜÝ]WÝ˜[YJÛÛ›‹•TÒSÓ”ÓÓT—ÔUWÓSRUÐÓÓÓÕÓ—ÒÑVK[[š\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠJBˆ›ÝYžWÙ\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]
ÛÛ›‹[[Ù]Ü›ÝšY\—ØÛÛÛÝÛ—Ü™X\ÛÛŠÛÛ›‹S•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹\WØ\™XJJBˆÛÛ›‹˜ÛÛ[Z]

Bˆš[˜[N‚ˆYˆÛÜÙWØÛÛ›ˆ[™ÛÛ›ˆ\È›Ý›Û™N‚ˆÛÛ›‹˜ÛÜÙJ
BˆÑÑÑT‹Ø\›š[™Ê‘\Ú[Û”ÛÛ\ˆ	\ÈÛÛÛÝÛˆXÝ]˜]Y[[	\È‹\WØ\™XK[[š\ÛÙ›Ü›X]
[Y\ÜXÏH›Z[]\ÈŠJBˆ™]\›ˆ[[‚‚™YˆX\š×Ù\Ú[ÛœÛÛ\—Ü\™›Ü›X[˜ÙWÜ˜]WÛ[Z]Y
ˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Ûˆ›Û™HH›Û™Kˆ›Ý×Ý˜[YNˆ]][YH›Û™HH›Û™KŠHOˆÝŽ‚ˆ™]\›ˆX\š×Ù\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]Y
ÛÛ›‹›Ý×Ý˜[YJB‚‚™YˆÛX\—Ù\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]ØÛÛÛÝÛŠÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Ûˆ›Û™HH›Û™JHOˆ›Û™N‚ˆÛØ˜[•TÒSÓ”ÓÓT—ÔT‘“Ô“PSÑWÔUWÓSRUÕS•Sˆ•TÒSÓ”ÓÓT—ÔT‘“Ô“PSÑWÔUWÓSRUÕS•SH›Û™BˆYˆÛÛ›ˆ\È›Ý›Û™N‚ˆÙ]Ø\ÜÝ]WÝ˜[YJÛÛ›‹•TÒSÓ”ÓÓT—ÔUWÓSRUÐÓÓÓÕÓ—ÒÑVKˆŠB‚‚™Yˆ\Ú[ÛœÛÛ\—ØÛÛÛÝÛ—ÜÛY\ÜÙXÛÛ™ÊˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Ûˆ›Û™HH›Û™Kˆ›Ý×Ý˜[YNˆ]][YH›Û™HH›Û™KŠHOˆ[‚ˆ›Ý×Ý˜[YHH›Ý×Ý˜[YHÜˆ]][YK››ÝÊ
BˆÛÛÛÝÛ—Ý[[HÙ]Ù\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]Ý[[
ÛÛ›‹›Ý×Ý˜[YJBˆYˆÛÛÛÝÛ—Ý[[[™ÛÛÛÝÛ—Ý[[ˆ›Ý×Ý˜[YN‚ˆ™]\›ˆX^
K[

ÛÛÛÝÛ—Ý[[H›Ý×Ý˜[YJKÝ[ÜÙXÛÛ™Ê
JH
ÈJBˆ™]\›ˆ‚‚™YˆØ[Ý[]WÜÜXÚYšX×ÞZY[
›ÙXÝ[Û—ÚÝÚˆ›Ø]›Û™KÝÜˆ›Ø]›Û™JHOˆ›Ø]›Û™N‚ˆYˆ›ÙXÝ[Û—ÚÝÚ\È›Û™HÜˆ›ÝÝÜ‚ˆ™]\›ˆ›Û™Bˆ™]\›ˆ›ÙXÝ[Û—ÚÝÚÈÝÜ‚‚™YˆÛ\ÜÚYžWÜ\™›Ü›X[˜ÙWÜÝ]\Êˆ›ÙXÝ[Û—ÚÝÚˆ›Ø]›Û™KˆÝÜˆ›Ø]›Û™Kˆ^XÝYÚÝÚˆ›Ø]›Û™Kˆ
‹ˆØ\›š[™×Ù]šX][Û—ÜÝˆ›Ø]HLLˆ[\Ù]šX][Û—ÜÝˆ›Ø]HLŒˆÜš]XØ[Ù]šX][Û—ÜÝˆ›Ø]HLÌŠHOˆ\VÜÝ‹Ý‹›Ø]›Û™WN‚ˆYˆ›ÙXÝ[Û—ÚÝÚ\È›Û™N‚ˆ™]\›ˆ”Ù[HYÜÈ‹›Z\ÜÚ[™×Ü›ÙXÝ[Ûˆ‹›Û™BˆYˆ›ÝÝÜ‚ˆ™]\›ˆ”Ù[H™Y™\°ê›˜ÚXH‹›Z\ÜÚ[™×ÚÝÜ‹›Û™BˆYˆ^XÝYÚÝÚ\È›Û™HÜˆ^XÝYÚÝÚH‚ˆ™]\›ˆ”Ù[H™Y™\°ê›˜ÚXH‹›ÚÈ‹›Û™B‚ˆ]šX][Û—ÜÝH

›ÙXÝ[Û—ÚÝÚH^XÝYÚÝÚ
HÈ^XÝYÚÝÚ
H
ˆLˆYˆ]šX][Û—ÜÝHØ\›š[™×Ù]šX][Û—ÜÝ‚ˆ™]\›ˆ“ÒÈ‹›ÚÈ‹]šX][Û—ÜÝˆYˆ]šX][Û—ÜÝH[\Ù]šX][Û—ÜÝ‚ˆ™]\›ˆ][°éðèÛÈ‹›ÚÈ‹]šX][Û—ÜÝˆYˆ]šX][Û—ÜÝHÜš]XØ[Ù]šX][Û—ÜÝ‚ˆ™]\›ˆ[\H‹›ÚÈ‹]šX][Û—ÜÝˆ™]\›ˆÜ°ë]XÛÈ‹›ÚÈ‹]šX][Û—ÜÝ‚‚™YˆÙ]Ü\™›Ü›X[˜ÙWÜÙ][™ÜÊÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹\ÜÙ]ÚYˆ[
HOˆXÝÜÝ‹[žWN‚ˆ›ÝÈHÛÛ›‹™^XÝ]J”ÑSPÕ
ˆ”“ÓH\™›Ü›X[˜ÙWÜÙ][™ÜÈÒT‘H\ÜÙ]ÚYHÈ‹
\ÜÙ]ÚY
JK™™]ÚÛ™J
BˆY˜][ÈHÂˆ˜\ÜÙ]ÚYŽˆ\ÜÙ]ÚYˆ™[˜X›YŽˆKˆØ\›š[™×Ù]šX][Û—ÜÝŽˆLLŒˆ˜[\Ù]šX][Û—ÜÝŽˆLŒŒˆ˜Üš]XØ[Ù]šX][Û—ÜÝŽˆLÌŒˆ˜˜\Ù[[™WÞYX\œÈŽˆ‹ˆ›Z[—Ø˜\Ù[[™WÜÚ[ÈŽˆKˆ›[ÛWØYÙ]ÚœÛÛˆŽˆˆ‹ˆ››Ý\ÈŽˆˆ‹ˆ\]YØ]Žˆˆ‹ˆBˆYˆ›ÝÈ\È›Û™N‚ˆ™]\›ˆY˜][ÂˆY˜][Ë\]JXÝ
›ÝÊJBˆ™]\›ˆY˜][Â‚‚™YˆÙ]Û[ÛWØYÙ]ÜÜXÚYšX×ÞZY[
Ù][™ÜÎˆXÝÜÝ‹[žWK\š[ÙÙ]Nˆ]JHOˆ›Ø]›Û™N‚ˆ˜]ÈHÝŠÙ][™ÜË™Ù]
›[ÛWØYÙ]ÚœÛÛˆŠHÜˆˆŠKœÝš\

BˆYˆ›Ý˜]Î‚ˆ™]\›ˆ›Û™BˆžN‚ˆ^[ØYHœÛÛ‹›ØYÊ˜]ÊBˆ^Ù\œÛÛ‹’”ÓÓ‘XÛÙQ\œ›ÜŽ‚ˆ™]\›ˆ›Û™BˆYˆ›Ý\Ú[œÝ[˜ÙJ^[ØYXÝ
N‚ˆ™]\›ˆ›Û™Bˆ™]\›ˆ\œÙWÙ›Ø]Ý˜[YJ^[ØY™Ù]
ˆžÜ\š[ÙÙ]K›[ÛŒ™HŠJB‚‚™YˆØ[Ý[]WÚ\ÝÜšXØ[Ø˜\Ù[[™JˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆ\ÜÙ]ÚYˆ[ˆ›ÝšY\ŽˆÝ‹ˆ\š[ÙÝ\NˆÝ‹ˆ\š[ÙÙ]Nˆ]Kˆ˜\Ù[[™WÞYX\œÎˆ[ˆZ[—Ø˜\Ù[[™WÜÚ[Îˆ[ŠHOˆ\VÙ›Ø]›Û™K›Ø]›Û™KÝ‹Ý—N‚ˆ™\Ý[HØ[Ý[]WÙ^XÝYÜ›ÙXÝ[Û—ÝÚ]ÙXYÛ›ÜÝXÊˆÛÛ›‹ˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆ›ÝšY\\›ÝšY\‹ˆ\š[ÙÝ\O\\š[ÙÝ\Kˆ\š[ÙÙ]O\\š[ÙÙ]KˆÝÜS›Û™KˆÙ][™ÜÏ^Âˆ˜˜\Ù[[™WÞYX\œÈŽˆ˜\Ù[[™WÞYX\œËˆ›Z[—Ø˜\Ù[[™WÜÚ[ÈŽˆZ[—Ø˜\Ù[[™WÜÚ[Ëˆ›[ÛWØYÙ]ÚœÛÛˆŽˆˆ‹ˆKˆ
Bˆ™]\›ˆ
ˆ™\Ý[È™^XÝYÚÝÚ—Kˆ™\Ý[È™^XÝYÜÜXÚYšX×ÞZY[—Kˆ™\Ý[È™^XÝYÜÛÝ\˜ÙH—Kˆ™\Ý[Èœ]X[]H—Kˆ
B‚‚™YˆØ[YWÙ]WÜ™]š[Ý\×ÞYX\œÊ\š[ÙÙ]Nˆ]K˜\Ù[[™WÞYX\œÎˆ[
HOˆ\ÝÙ]WN‚ˆØ[™Y]\Îˆ\ÝÙ]WHH×Bˆ›ÜˆYX\—ÛÙ™œÙ][ˆ˜[™ÙJKX^
[
˜\Ù[[™WÞYX\œÈÜˆJKJH
ÈJN‚ˆžN‚ˆ™]š[Ý\ÈH\š[ÙÙ]Kœ™\XÙJYX\\\š[ÙÙ]KžYX\ˆHYX\—ÛÙ™œÙ]
Bˆ^Ù\˜[YQ\œ›ÜŽ‚ˆ™]š[Ý\ÈH\š[ÙÙ]Kœ™\XÙJYX\\\š[ÙÙ]KžYX\ˆHYX\—ÛÙ™œÙ]^OLŽ
BˆØ[™Y]\Ë˜\[™
™]š[Ý\ÊBˆ™]\›ˆØ[™Y]\Â‚‚™YˆØYÝ˜[YØ˜\Ù[[™WÜ›ÝÜÊˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆ\ÜÙ]ÚYˆ[ˆ›ÝšY\ŽˆÝ‹ˆ\š[ÙÝ\NˆÝ‹ˆØ[™Y]WÙ]\Îˆ\ÝÙ]WKŠHOˆ\ÝÜÜ[]LË”›Ý×N‚ˆYˆ›ÝØ[™Y]WÙ]\Î‚ˆ™]\›ˆ×BˆXÙZÛ\œÈH‹‹š›Ú[ŠÈˆ›ÜˆÈ[ˆØ[™Y]WÙ]\ÊBˆ™]\›ˆ]Y\žWØ[
ˆÛÛ›‹ˆˆˆˆ‚ˆÑSPÕ›ÙXÝ[Û—ÚÝÚÜXÚYšX×ÞZY[\š[ÙÙ]Bˆ”“ÓH›ÙXÝ[Û—Ü™XÛÜ™ÂˆÒT‘H\ÜÙ]ÚYHÈS‘›ÝšY\ˆHÈS‘\š[ÙÝ\HHÂˆS‘\š[ÙÙ]HSˆ
ÜXÙZÛ\œßJBˆS‘›ÙXÝ[Û—ÚÝÚTÈ“Õ•SˆS‘ÜXÚYšX×ÞZY[TÈ“Õ•SˆS‘ÓÐSTÐÑJ]WÜ]X[]K	ÉÊHOH	ÛZ\ÜÚ[™×Ü›ÙXÝ[Û‰ÂˆÔ‘Tˆ–H\š[ÙÙ]HTÐÂˆˆˆ‹ˆØ\ÜÙ]ÚY›ÝšY\‹\š[ÙÝ\K
–Ú][Kš\ÛÙ›Ü›X]

H›Üˆ][H[ˆØ[™Y]WÙ]\×WKˆ
B‚‚™YˆØ[Ý[]WÛ]Ø˜\Ù[[™JˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆ\ÜÙ]ÚYˆ[ˆ›ÝšY\ŽˆÝ‹ˆ\š[ÙÙ]Nˆ]Kˆ˜\Ù[[™WÞYX\œÎˆ[ˆZ[—Ø˜\Ù[[™WÜÚ[Îˆ[ˆÙ^WÝ˜[YNˆ]H›Û™HH›Û™KŠHOˆXÝÜÝ‹[žWN‚ˆÙ^WÝ˜[YHHÙ^WÝ˜[YHÜˆ]KÙ^J
Bˆ\š[ÙÜÝ\H\š[ÙÙ]Kœ™\XÙJ^OLJBˆYˆ\š[ÙÜÝ\žYX\ˆOHÙ^WÝ˜[YKžYX\ˆ[™\š[ÙÜÝ\›[ÛOHÙ^WÝ˜[YK›[Û‚ˆ\š[ÙÙ[™HÙ^WÝ˜[YBˆ[ÙN‚ˆ\š[ÙÙ[™H\š[ÙÜÝ\œ™\XÙJ^OXØ[[™\‹›[Û˜[™ÙJ\š[ÙÜÝ\žYX\‹\š[ÙÜÝ\›[Û
VÌWJBˆ^WÜÜ[ˆH\š[ÙÙ[™™^BˆØ[™Y]WÜ˜[™Ù\Îˆ\ÝÜÝ—HH×BˆYX\›WÝ˜[Y\Îˆ\ÝÝ\VÙ›Ø]›Ø]WHH×Bˆ›ÜˆYX\—ÛÙ™œÙ][ˆ˜[™ÙJKX^
[
˜\Ù[[™WÞYX\œÈÜˆJKJH
ÈJN‚ˆÝ\H\š[ÙÜÝ\œ™\XÙJYX\\\š[ÙÜÝ\žYX\ˆHYX\—ÛÙ™œÙ]
Bˆ[™Ù^HHZ[Š^WÜÜ[‹Ø[[™\‹›[Û˜[™ÙJÝ\žYX\‹Ý\›[Û
VÌWJBˆ[™HÝ\œ™\XÙJ^OY[™Ù^JBˆØ[™Y]WÜ˜[™Ù\Ë˜\[™
ˆžÜÝ\š\ÛÙ›Ü›X]

_K‹žÙ[™š\ÛÙ›Ü›X]

_HŠBˆ›ÝÜÈH]Y\žWØ[
ˆÛÛ›‹ˆˆˆ‚ˆÑSPÕ›ÙXÝ[Û—ÚÝÚÜXÚYšX×ÞZY[ˆ”“ÓH›ÙXÝ[Û—Ü™XÛÜ™ÂˆÒT‘H\ÜÙ]ÚYHÈS‘›ÝšY\ˆHÈS‘\š[ÙÝ\HH	Ù^IÂˆS‘\š[ÙÙ]H‘UÑQSˆÈS‘ÂˆS‘›ÙXÝ[Û—ÚÝÚTÈ“Õ•SˆS‘ÜXÚYšX×ÞZY[TÈ“Õ•SˆS‘ÓÐSTÐÑJ]WÜ]X[]K	ÉÊHOH	ÛZ\ÜÚ[™×Ü›ÙXÝ[Û‰ÂˆÔ‘Tˆ–H\š[ÙÙ]HTÐÂˆˆˆ‹ˆ
\ÜÙ]ÚY›ÝšY\‹Ý\š\ÛÙ›Ü›X]

K[™š\ÛÙ›Ü›X]

JKˆ
BˆYˆ[Š›ÝÜÊHOH[™Ù^N‚ˆYX\›WÝ˜[Y\Ë˜\[™
ˆ
ˆÝ[J›Ø]
›ÝÖÈœ›ÙXÝ[Û—ÚÝÚ—JH›Üˆ›ÝÈ[ˆ›ÝÜÊKˆÝ[J›Ø]
›ÝÖÈœÜXÚYšX×ÞZY[—JH›Üˆ›ÝÈ[ˆ›ÝÜÊKˆ
Bˆ
BˆYˆ[ŠYX\›WÝ˜[Y\ÊHX^
[
Z[—Ø˜\Ù[[™WÜÚ[ÈÜˆJKJN‚ˆ™]\›ˆÂˆ™^XÝYÚÝÚŽˆ›Û™Kˆ™^XÝYÜÜXÚYšX×ÞZY[Žˆ›Û™Kˆ™^XÝYÜÛÝ\˜ÙHŽˆ››Û™H‹ˆœ]X[]HŽˆœ\X[Ú\ÝÜžHˆYˆYX\›WÝ˜[Y\È[ÙH›ÚÈ‹ˆ™XYÛ›ÜÝXÈŽˆÂˆš\ÝÜšXØ[Ü™XÛÜ™×Ù›Ý[™Žˆ[ŠYX\›WÝ˜[Y\ÊKˆ˜˜\Ù[[™WÞYX\œÈŽˆ˜\Ù[[™WÞYX\œËˆ›Z[—Ø˜\Ù[[™WÜÚ[ÈŽˆZ[—Ø˜\Ù[[™WÜÚ[Ëˆ˜Ø[™Y]WÚ\ÝÜšXØ[Ù]\ÈŽˆØ[™Y]WÜ˜[™Ù\Ëˆ™^XÝYÜÛÝ\˜ÙWØ][\YŽˆš\ÝÜšXØ[ÜØ[YWÜ\š[Ù‹ˆ››×Ü™Y™\™[˜ÙWÜ™X\ÛÛˆŽˆ“U™Y™\™[˜ÙH™\]Z\™\È\ÝÜšXØ[Z[H™XÛÜ™È›ÜˆØ[YH\š[Ù‹ˆœ\š[ÙÜÝ\Žˆ\š[ÙÜÝ\š\ÛÙ›Ü›X]

Kˆœ\š[ÙÙ[™Žˆ\š[ÙÙ[™š\ÛÙ›Ü›X]

KˆKˆBˆ^XÝYÚÝÚHÝ[J][VÌH›Üˆ][H[ˆYX\›WÝ˜[Y\ÊHÈ[ŠYX\›WÝ˜[Y\ÊBˆ^XÝYÜÜXÚYšX×ÞZY[HÝ[J][VÌWH›Üˆ][H[ˆYX\›WÝ˜[Y\ÊHÈ[ŠYX\›WÝ˜[Y\ÊBˆ™]\›ˆÂˆ™^XÝYÚÝÚŽˆ^XÝYÚÝÚˆ™^XÝYÜÜXÚYšX×ÞZY[Žˆ^XÝYÜÜXÚYšX×ÞZY[ˆ™^XÝYÜÛÝ\˜ÙHŽˆš\ÝÜšXØ[ÜØ[YWÜ\š[Ù‹ˆœ]X[]HŽˆ›ÚÈ‹ˆ™XYÛ›ÜÝXÈŽˆÂˆš\ÝÜšXØ[Ü™XÛÜ™×Ù›Ý[™Žˆ[ŠYX\›WÝ˜[Y\ÊKˆ˜˜\Ù[[™WÞYX\œÈŽˆ˜\Ù[[™WÞYX\œËˆ›Z[—Ø˜\Ù[[™WÜÚ[ÈŽˆZ[—Ø˜\Ù[[™WÜÚ[Ëˆ˜Ø[™Y]WÚ\ÝÜšXØ[Ù]\ÈŽˆØ[™Y]WÜ˜[™Ù\Ëˆ™^XÝYÜÛÝ\˜ÙWØ][\YŽˆš\ÝÜšXØ[ÜØ[YWÜ\š[Ù‹ˆ››×Ü™Y™\™[˜ÙWÜ™X\ÛÛˆŽˆˆ‹ˆœ\š[ÙÜÝ\Žˆ\š[ÙÜÝ\š\ÛÙ›Ü›X]

Kˆœ\š[ÙÙ[™Žˆ\š[ÙÙ[™š\ÛÙ›Ü›X]

KˆKˆB‚‚™YˆØ[Ý[]WÙ^XÝYÜ›ÙXÝ[Û—ÝÚ]ÙXYÛ›ÜÝXÊˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆ\ÜÙ]ÚYˆ[ˆ›ÝšY\ŽˆÝ‹ˆ\š[ÙÝ\NˆÝ‹ˆ\š[ÙÙ]Nˆ]KˆÝÜˆ›Ø]›Û™KˆÙ][™ÜÎˆXÝÜÝ‹[žWKˆ\ÜÙ]Û˜[YNˆÝˆHˆ‹ˆÙ^WÝ˜[YNˆ]H›Û™HH›Û™KŠHOˆXÝÜÝ‹[žWN‚ˆ˜\Ù[[™WÞYX\œÈH[
Ù][™ÜË™Ù]
˜˜\Ù[[™WÞYX\œÈŠHÜˆŠBˆZ[—Ø˜\Ù[[™WÜÚ[ÈH[
Ù][™ÜË™Ù]
›Z[—Ø˜\Ù[[™WÜÚ[ÈŠHÜˆJBˆYˆ\š[ÙÝ\HOH›]Ž‚ˆ™\Ý[HØ[Ý[]WÛ]Ø˜\Ù[[™JˆÛÛ›‹ˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆ›ÝšY\\›ÝšY\‹ˆ\š[ÙÙ]O\\š[ÙÙ]Kˆ˜\Ù[[™WÞYX\œÏX˜\Ù[[™WÞYX\œËˆZ[—Ø˜\Ù[[™WÜÚ[Ï[Z[—Ø˜\Ù[[™WÜÚ[ËˆÙ^WÝ˜[YO]Ù^WÝ˜[YKˆ
Bˆ[ÙN‚ˆ\ÝÜšXØ[Ý\HH›[ÛˆYˆ\š[ÙÝ\HOH›[Ûˆ[ÙH™^H‚ˆØ[™Y]\ÈHØ[YWÙ]WÜ™]š[Ý\×ÞYX\œÊ\š[ÙÙ]Kœ™\XÙJ^OLJHYˆ\š[ÙÝ\HOH›[Ûˆ[ÙH\š[ÙÙ]K˜\Ù[[™WÞYX\œÊBˆ›ÝÜÈHØYÝ˜[YØ˜\Ù[[™WÜ›ÝÜÊˆÛÛ›‹ˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆ›ÝšY\\›ÝšY\‹ˆ\š[ÙÝ\OZ\ÝÜšXØ[Ý\KˆØ[™Y]WÙ]\ÏXØ[™Y]\Ëˆ
BˆXYÛ›ÜÝXÈHÂˆš\ÝÜšXØ[Ü™XÛÜ™×Ù›Ý[™Žˆ[Š›ÝÜÊKˆ˜˜\Ù[[™WÞYX\œÈŽˆ˜\Ù[[™WÞYX\œËˆ›Z[—Ø˜\Ù[[™WÜÚ[ÈŽˆZ[—Ø˜\Ù[[™WÜÚ[Ëˆ˜Ø[™Y]WÚ\ÝÜšXØ[Ù]\ÈŽˆÚ][Kš\ÛÙ›Ü›X]

H›Üˆ][H[ˆØ[™Y]\×Kˆ™^XÝYÜÛÝ\˜ÙWØ][\YŽˆš\ÝÜšXØ[ÜØ[YWÜ\š[Ù‹ˆ››×Ü™Y™\™[˜ÙWÜ™X\ÛÛˆŽˆˆ‹ˆBˆYˆ[Š›ÝÜÊHZ[—Ø˜\Ù[[™WÜÚ[Î‚ˆX™[H›[ÛHˆYˆ\š[ÙÝ\HOH›[Ûˆ[ÙH™Z[H‚ˆXYÛ›ÜÝXÖÈ››×Ü™Y™\™[˜ÙWÜ™X\ÛÛˆ—HH
ˆˆ“›È\ÝÜšXØ[ÛX™[H™XÛÜ™È›Ý[™›ÜˆØ[YHÉÛ[Û	ÈYˆ\š[ÙÝ\HOH	Û[Û	È[ÙH	Ù^IßH[ˆ™]š[Ý\ÈYX\œÈ‚ˆYˆ›Ý›ÝÜÂˆ[ÙHˆ“Û›HÛ[Š›ÝÜÊ_H˜\Ù[[™HÚ[È›Ý[™Z[š[][H\ÈÛZ[—Ø˜\Ù[[™WÜÚ[ßH‚ˆ
Bˆ™\Ý[HÂˆ™^XÝYÚÝÚŽˆ›Û™Kˆ™^XÝYÜÜXÚYšX×ÞZY[Žˆ›Û™Kˆ™^XÝYÜÛÝ\˜ÙHŽˆ››Û™H‹ˆœ]X[]HŽˆœ\X[Ú\ÝÜžHˆYˆ›ÝÜÈ[ÙH›ÚÈ‹ˆ™XYÛ›ÜÝXÈŽˆXYÛ›ÜÝXËˆBˆ[ÙN‚ˆ›ÙXÝ[Û—Ý˜[Y\ÈHÙ›Ø]
›ÝÖÈœ›ÙXÝ[Û—ÚÝÚ—JH›Üˆ›ÝÈ[ˆ›ÝÜ×BˆÜXÚYšX×Ý˜[Y\ÈHÙ›Ø]
›ÝÖÈœÜXÚYšX×ÞZY[—JH›Üˆ›ÝÈ[ˆ›ÝÜ×Bˆ™\Ý[HÂˆ™^XÝYÚÝÚŽˆÝ[J›ÙXÝ[Û—Ý˜[Y\ÊHÈ[Š›ÙXÝ[Û—Ý˜[Y\ÊKˆ™^XÝYÜÜXÚYšX×ÞZY[ŽˆÝ[JÜXÚYšX×Ý˜[Y\ÊHÈ[ŠÜXÚYšX×Ý˜[Y\ÊKˆ™^XÝYÜÛÝ\˜ÙHŽˆš\ÝÜšXØ[ÜØ[YWÜ\š[Ù‹ˆœ]X[]HŽˆ›ÚÈ‹ˆ™XYÛ›ÜÝXÈŽˆXYÛ›ÜÝXËˆB‚ˆYˆ™\Ý[È™^XÝYÚÝÚ—H\È›Û™N‚ˆYÙ]ÜÜXÚYšXÈHÙ]Û[ÛWØYÙ]ÜÜXÚYšX×ÞZY[
Ù][™ÜË\š[ÙÙ]JBˆYˆYÙ]ÜÜXÚYšXÈ\È›Ý›Û™H[™ÝÜ‚ˆYˆ\š[ÙÝ\HOH™^HŽ‚ˆYÙ]ÜÜXÚYšXÈHYÙ]ÜÜXÚYšXÈÈØ[[™\‹›[Û˜[™ÙJ\š[ÙÙ]KžYX\‹\š[ÙÙ]K›[Û
VÌWBˆ™\Ý[HÂˆ™^XÝYÚÝÚŽˆYÙ]ÜÜXÚYšXÈ
ˆÝÜˆ™^XÝYÜÜXÚYšX×ÞZY[ŽˆYÙ]ÜÜXÚYšXËˆ™^XÝYÜÛÝ\˜ÙHŽˆ›[ÛWØYÙ]‹ˆœ]X[]HŽˆ›ÚÈ‹ˆ™XYÛ›ÜÝXÈŽˆÂˆ
Šœ™\Ý[È™XYÛ›ÜÝXÈ—Kˆ™^XÝYÜÛÝ\˜ÙWØ][\YŽˆ›[ÛWØYÙ]‹ˆ››×Ü™Y™\™[˜ÙWÜ™X\ÛÛˆŽˆˆ‹ˆKˆBˆYˆÝÜ\È›Û™N‚ˆ™\Ý[È™XYÛ›ÜÝXÈ—VÈ››×Ü™Y™\™[˜ÙWÜ™X\ÛÛˆ—HH“Z\ÜÚ[™ÈÕÜ‚ˆYˆ™\Ý[È™^XÝYÜÜXÚYšX×ÞZY[—H\È›Û™H[™›Ý™\Ý[È™XYÛ›ÜÝXÈ—K™Ù]
››×Ü™Y™\™[˜ÙWÜ™X\ÛÛˆŠN‚ˆ™\Ý[È™XYÛ›ÜÝXÈ—VÈ››×Ü™Y™\™[˜ÙWÜ™X\ÛÛˆ—HH“Z\ÜÚ[™È^XÝYÜÜXÚYšX×ÞZY[‚‚ˆÙÙÙ\ˆHÝ\œ™[Ø\›ÙÙÙ\ˆYˆ\×Ø\ØÛÛ^

H[ÙHÙÙÚ[™Ë™Ù]ÙÙÙ\Š×Û˜[YW×ÊBˆÙÙÙ\‹š[™›Êˆ”\™›Ü›X[˜ÙH™Y™\™[˜ÙHØ[Ý[][ÛŽˆ\ÜÙ]ÚYI\È\ÜÙ]Û˜[YOI\È\š[ÙÝ\OI\È\š[ÙÙ]OI\È˜\Ù[[™WÞYX\œÏI\ÈØ[™Y]WÙ]\ÏI\È˜[YØ˜\Ù[[™WÜ™XÛÜ™ÏI\È^XÝYÜÜXÚYšX×ÞZY[I\È^XÝYÜÛÝ\˜ÙOI\È›×Ü™Y™\™[˜ÙWÜ™X\ÛÛI\È‹ˆ\ÜÙ]ÚYˆ\ÜÙ]Û˜[YKˆ\š[ÙÝ\Kˆ\š[ÙÙ]Kš\ÛÙ›Ü›X]

Kˆ˜\Ù[[™WÞYX\œËˆ™\Ý[È™XYÛ›ÜÝXÈ—K™Ù]
˜Ø[™Y]WÚ\ÝÜšXØ[Ù]\ÈŠKˆ™\Ý[È™XYÛ›ÜÝXÈ—K™Ù]
š\ÝÜšXØ[Ü™XÛÜ™×Ù›Ý[™ŠKˆ™\Ý[È™^XÝYÜÜXÚYšX×ÞZY[—Kˆ™\Ý[È™^XÝYÜÛÝ\˜ÙH—Kˆ™\Ý[È™XYÛ›ÜÝXÈ—K™Ù]
››×Ü™Y™\™[˜ÙWÜ™X\ÛÛˆ‹ˆŠKˆ
Bˆ™]\›ˆ™\Ý[‚‚™YˆØ[Ý[]WÙ^XÝYÜ›ÙXÝ[ÛŠˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆ\ÜÙ]ÚYˆ[ˆ›ÝšY\ŽˆÝ‹ˆ\š[ÙÝ\NˆÝ‹ˆ\š[ÙÙ]Nˆ]KˆÝÜˆ›Ø]›Û™KˆÙ][™ÜÎˆXÝÜÝ‹[žWKŠHOˆ\VÙ›Ø]›Û™K›Ø]›Û™KÝ‹Ý—N‚ˆ™\Ý[HØ[Ý[]WÙ^XÝYÜ›ÙXÝ[Û—ÝÚ]ÙXYÛ›ÜÝXÊˆÛÛ›‹ˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆ›ÝšY\\›ÝšY\‹ˆ\š[ÙÝ\O\\š[ÙÝ\Kˆ\š[ÙÙ]O\\š[ÙÙ]KˆÝÜZÝÜˆÙ][™ÜÏ\Ù][™ÜËˆ
Bˆ™]\›ˆ™\Ý[È™^XÝYÚÝÚ—K™\Ý[È™^XÝYÜÜXÚYšX×ÞZY[—K™\Ý[È™^XÝYÜÛÝ\˜ÙH—K™\Ý[Èœ]X[]H—B‚‚™Yˆ\Ù\Ü›ÙXÝ[Û—Ü™XÛÜ™
ˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆ\ÜÙ]ÚYˆ[ˆ›ÝšY\ŽˆÝ‹ˆ^\›˜[ÚYˆÝ‹ˆ\š[ÙÝ\NˆÝ‹ˆ\š[ÙÙ]Nˆ]Kˆ›ÙXÝ[Û—ÚÝÚˆ›Ø]›Û™KˆÜXÚYšX×ÞZY[ˆ›Ø]›Û™Kˆ^XÝYÚÝÚˆ›Ø]›Û™Kˆ^XÝYÜÜXÚYšX×ÞZY[ˆ›Ø]›Û™Kˆ]šX][Û—ÜÝˆ›Ø]›Û™Kˆ\™›Ü›X[˜ÙWÜÝ]\ÎˆÝ‹ˆ^XÝYÜÛÝ\˜ÙNˆÝ‹ˆ]WÜ]X[]NˆÝ‹ˆ›Ý\ÎˆÝ‹ˆ^[ØYÚœÛÛŽˆÝ‹ˆÙ[XÝYÜ›ÙXÝ[Û—ÚÙ^NˆÝˆHˆ‹ˆÙ[XÝYÜ›ÙXÝ[Û—Ü˜]×Ý˜[YNˆÝˆHˆ‹ˆ™Y™\™[˜ÙWÙXYÛ›ÜÝX×ÚœÛÛŽˆÝˆHˆ‹ŠHOˆÝŽ‚ˆ›ÝÈH]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠBˆ^\Ý[™ÈHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕYˆ”“ÓH›ÙXÝ[Û—Ü™XÛÜ™ÂˆÒT‘H\ÜÙ]ÚYHÈS‘›ÝšY\ˆHÈS‘\š[ÙÝ\HHÈS‘\š[ÙÙ]HHÂˆSRUBˆˆˆ‹ˆ
\ÜÙ]ÚY›ÝšY\‹\š[ÙÝ\K\š[ÙÙ]Kš\ÛÙ›Ü›X]

JKˆ
K™™]ÚÛ™J
BˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆS”ÑT•S•È›ÙXÝ[Û—Ü™XÛÜ™È
ˆ\ÜÙ]ÚY›ÝšY\‹^\›˜[ÚY\š[ÙÝ\K\š[ÙÙ]K›ÙXÝ[Û—ÚÝÚÜXÚYšX×ÞZY[ˆ^XÝYÚÝÚ^XÝYÜÜXÚYšX×ÞZY[]šX][Û—ÜÝ\™›Ü›X[˜ÙWÜÝ]\Ë^XÝYÜÛÝ\˜ÙKˆ]WÜ]X[]K›Ý\ËÙ[XÝYÜ›ÙXÝ[Û—ÚÙ^KÙ[XÝYÜ›ÙXÝ[Û—Ü˜]×Ý˜[YK™Y™\™[˜ÙWÙXYÛ›ÜÝX×ÚœÛÛ‹ˆ^[ØYÚœÛÛ‹Ü™X]YØ]\]YØ]ˆ
HSQTÈ
ËËËËËËËËËËËËËËËËËËËÊBˆÓˆÓÓ‘“PÕ
\ÜÙ]ÚY›ÝšY\‹\š[ÙÝ\K\š[ÙÙ]JHÈTUHÑUˆ^\›˜[ÚYH^ÛYY™^\›˜[ÚYˆ›ÙXÝ[Û—ÚÝÚH^ÛYYœ›ÙXÝ[Û—ÚÝÚˆÜXÚYšX×ÞZY[H^ÛYYœÜXÚYšX×ÞZY[ˆ^XÝYÚÝÚH^ÛYY™^XÝYÚÝÚˆ^XÝYÜÜXÚYšX×ÞZY[H^ÛYY™^XÝYÜÜXÚYšX×ÞZY[ˆ]šX][Û—ÜÝH^ÛYY™]šX][Û—ÜÝˆ\™›Ü›X[˜ÙWÜÝ]\ÈH^ÛYYœ\™›Ü›X[˜ÙWÜÝ]\Ëˆ^XÝYÜÛÝ\˜ÙHH^ÛYY™^XÝYÜÛÝ\˜ÙKˆ]WÜ]X[]HH^ÛYY™]WÜ]X[]Kˆ›Ý\ÈH^ÛYY››Ý\ËˆÙ[XÝYÜ›ÙXÝ[Û—ÚÙ^HH^ÛYYœÙ[XÝYÜ›ÙXÝ[Û—ÚÙ^KˆÙ[XÝYÜ›ÙXÝ[Û—Ü˜]×Ý˜[YHH^ÛYYœÙ[XÝYÜ›ÙXÝ[Û—Ü˜]×Ý˜[YKˆ™Y™\™[˜ÙWÙXYÛ›ÜÝX×ÚœÛÛˆH^ÛYYœ™Y™\™[˜ÙWÙXYÛ›ÜÝX×ÚœÛÛ‹ˆ^[ØYÚœÛÛˆH^ÛYYœ^[ØYÚœÛÛ‹ˆ\]YØ]H^ÛYY\]YØ]ˆˆˆ‹ˆ
ˆ\ÜÙ]ÚYˆ›ÝšY\‹ˆ^\›˜[ÚYÜˆ›Û™Kˆ\š[ÙÝ\Kˆ\š[ÙÙ]Kš\ÛÙ›Ü›X]

Kˆ›ÙXÝ[Û—ÚÝÚˆÜXÚYšX×ÞZY[ˆ^XÝYÚÝÚˆ^XÝYÜÜXÚYšX×ÞZY[ˆ]šX][Û—ÜÝˆ\™›Ü›X[˜ÙWÜÝ]\Ëˆ^XÝYÜÛÝ\˜ÙKˆ]WÜ]X[]Kˆ›Ý\ËˆÙ[XÝYÜ›ÙXÝ[Û—ÚÙ^HÜˆ›Û™KˆÙ[XÝYÜ›ÙXÝ[Û—Ü˜]×Ý˜[YHÜˆ›Û™Kˆ™Y™\™[˜ÙWÙXYÛ›ÜÝX×ÚœÛÛˆÜˆ›Û™Kˆ^[ØYÚœÛÛ‹ˆ›ÝËˆ›ÝËˆ
Kˆ
Bˆ™]\›ˆ\]YˆYˆ^\Ý[™È[ÙHš[œÙ\Y‚‚‚™YˆÝÜ™WÜ›ÙXÝ[Û—ÚÜWÜ™XÛÜ™
ˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆ\ÜÙ]Ü›ÝÎˆÜ[]LË”›ÝÈXÝÜÝ‹[žWKˆ›ÝšY\ŽˆÝ‹ˆ^\›˜[ÚYˆÝ‹ˆ\š[ÙÝ\NˆÝ‹ˆ\š[ÙÙ]Nˆ]KˆÜWÜ›ÝÎˆXÝÜÝ‹[žWKˆ›Ý\×Ü™Yš^ˆÝˆHˆ‹ŠHOˆXÝÜÝ‹[žWN‚ˆ]WÚ][WÛX\HÜWÜ›ÝË™Ù]
™]R][SX\ŠHYˆ\Ú[œÝ[˜ÙJÜWÜ›ÝËXÝ
H[ÙHßBˆYˆ›Ý\Ú[œÝ[˜ÙJ]WÚ][WÛX\XÝ
N‚ˆ]WÚ][WÛX\HßBˆ›ÙXÝ[Û—ÚÝÚÙ[XÝYÚÙ^KÙ[XÝYÜ˜]×Ý˜[YHHÙ[XÝÜ›ÙXÝ[Û—Ý˜[YJ]WÚ][WÛX\
BˆÝÜH\œÙWÚÝÜÝ˜[YJ\ÜÙ]Ü›ÝÖÈšÝÜ—JBˆÜXÚYšX×ÞZY[HØ[Ý[]WÜÜXÚYšX×ÞZY[
›ÙXÝ[Û—ÚÝÚÝÜ
Bˆ\ÜÙ]ÚYH[
\ÜÙ]Ü›ÝÖÈ˜\ÜÙ]ÚY—HYˆ˜\ÜÙ]ÚYˆ[ˆ\ÜÙ]Ü›ÝËšÙ^\Ê
H[ÙH\ÜÙ]Ü›ÝÖÈšY—JBˆÙ][™ÜÈHÙ]Ü\™›Ü›X[˜ÙWÜÙ][™ÜÊÛÛ›‹\ÜÙ]ÚY
Bˆ›ÜˆÙ^H[ˆÙ][™ÜÎ‚ˆžN‚ˆ›Ý×Ý˜[YHH\ÜÙ]Ü›ÝÖÚÙ^WBˆ^Ù\
Ù^Q\œ›Ü‹[™^\œ›ÜŠN‚ˆÛÛ[YBˆYˆ›Ý×Ý˜[YH\È›Ý›Û™N‚ˆÙ][™ÜÖÚÙ^WHH›Ý×Ý˜[YB‚ˆ™Y™\™[˜ÙWÜ™\Ý[HØ[Ý[]WÙ^XÝYÜ›ÙXÝ[Û—ÝÚ]ÙXYÛ›ÜÝXÊˆÛÛ›‹ˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆ›ÝšY\\›ÝšY\‹ˆ\š[ÙÝ\O\\š[ÙÝ\Kˆ\š[ÙÙ]O\\š[ÙÙ]KˆÝÜZÝÜˆÙ][™ÜÏ\Ù][™ÜËˆ\ÜÙ]Û˜[YO\ÝŠ\ÜÙ]Ü›ÝÖÈœ›Ú™XÝÛ˜[YH—HYˆœ›Ú™XÝÛ˜[YHˆ[ˆ\ÜÙ]Ü›ÝËšÙ^\Ê
H[ÙHˆŠKˆ
Bˆ^XÝYÚÝÚH™Y™\™[˜ÙWÜ™\Ý[È™^XÝYÚÝÚ—Bˆ^XÝYÜÜXÚYšX×ÞZY[H™Y™\™[˜ÙWÜ™\Ý[È™^XÝYÜÜXÚYšX×ÞZY[—Bˆ^XÝYÜÛÝ\˜ÙHH™Y™\™[˜ÙWÜ™\Ý[È™^XÝYÜÛÝ\˜ÙH—Bˆ˜\Ù[[™WÜ]X[]HH™Y™\™[˜ÙWÜ™\Ý[Èœ]X[]H—Bˆ\™›Ü›X[˜ÙWÜÝ]\Ë]WÜ]X[]K]šX][Û—ÜÝHÛ\ÜÚYžWÜ\™›Ü›X[˜ÙWÜÝ]\Êˆ›ÙXÝ[Û—ÚÝÚˆÝÜˆ^XÝYÚÝÚˆØ\›š[™×Ù]šX][Û—ÜÝY›Ø]
Ù][™ÜË™Ù]
Ø\›š[™×Ù]šX][Û—ÜÝŠHÜˆLL
Kˆ[\Ù]šX][Û—ÜÝY›Ø]
Ù][™ÜË™Ù]
˜[\Ù]šX][Û—ÜÝŠHÜˆLŒ
KˆÜš]XØ[Ù]šX][Û—ÜÝY›Ø]
Ù][™ÜË™Ù]
˜Üš]XØ[Ù]šX][Û—ÜÝŠHÜˆLÌ
Kˆ
BˆYˆ]WÜ]X[]HOH›ÚÈˆ[™˜\Ù[[™WÜ]X[]HOHœ\X[Ú\ÝÜžHˆ[™^XÝYÜÛÝ\˜ÙHOH››Û™HŽ‚ˆ]WÜ]X[]HHœ\X[Ú\ÝÜžH‚‚ˆ›Ý\×Ü\ÈHÛ›Ý\×Ü™Yš^HYˆ›Ý\×Ü™Yš^[ÙH×BˆYˆ›ÙXÝ[Û—ÚÝÚ\È›Û™N‚ˆ›Ý\×Ü\Ë˜\[™
ˆZ[ÛZ\ÜÚ[™×Ü›ÙXÝ[Û—Û›ÝJˆ]WÚ][WÛX\ˆÝ][Û—ØÛÙOY^\›˜[ÚYˆ\š[ÙÝ\O\\š[ÙÝ\Kˆ\š[ÙÙ]O\\š[ÙÙ]Kˆ
Bˆ
BˆYˆÝÜ\È›Û™N‚ˆ›Ý\×Ü\Ë˜\[™
šÕÜØØ[[H˜[HÝH[˜[YËˆŠBˆYˆ^XÝYÜÛÝ\˜ÙHOH››Û™HŽ‚ˆ›Ý\×Ü\Ë˜\[™
”Ù[H\Ý0ìÜšXÛÈÝHÜ°éØ[Y[ÈY[œØ[\˜H™Y™\°ê›˜ÚXKˆŠB‚ˆYˆ›ÙXÝ[Û—ÚÝÚ\È›Û™N‚ˆ^\Ý[™×Ý˜[YHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕYˆ”“ÓH›ÙXÝ[Û—Ü™XÛÜ™ÂˆÒT‘H\ÜÙ]ÚYHÈS‘›ÝšY\ˆHÈS‘\š[ÙÝ\HHÈS‘\š[ÙÙ]HHÂˆS‘›ÙXÝ[Û—ÚÝÚTÈ“Õ•SˆS‘ÓÐSTÐÑJ]WÜ]X[]K	ÉÊHOH	ÛZ\ÜÚ[™×Ü›ÙXÝ[Û‰ÂˆSRUBˆˆˆ‹ˆ
\ÜÙ]ÚY›ÝšY\‹\š[ÙÝ\K\š[ÙÙ]Kš\ÛÙ›Ü›X]

JKˆ
K™™]ÚÛ™J
BˆYˆ^\Ý[™×Ý˜[Y‚ˆ™]\›ˆÂˆ\Ù\ÜÝ]\ÈŽˆœÚÚ\YÙ^\Ý[™×Ý˜[Y‹ˆœ›ÙXÝ[Û—ÚÝÚŽˆ›Û™KˆœÜXÚYšX×ÞZY[Žˆ›Û™Kˆœ\™›Ü›X[˜ÙWÜÝ]\ÈŽˆ\™›Ü›X[˜ÙWÜÝ]\Ëˆ™]WÜ]X[]HŽˆ]WÜ]X[]KˆB‚ˆ\Ù\ÜÝ]\ÈH\Ù\Ü›ÙXÝ[Û—Ü™XÛÜ™
ˆÛÛ›‹ˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆ›ÝšY\\›ÝšY\‹ˆ^\›˜[ÚYY^\›˜[ÚYˆ\š[ÙÝ\O\\š[ÙÝ\Kˆ\š[ÙÙ]O\\š[ÙÙ]Kˆ›ÙXÝ[Û—ÚÝÚ\›ÙXÝ[Û—ÚÝÚˆÜXÚYšX×ÞZY[\ÜXÚYšX×ÞZY[ˆ^XÝYÚÝÚY^XÝYÚÝÚˆ^XÝYÜÜXÚYšX×ÞZY[Y^XÝYÜÜXÚYšX×ÞZY[ˆ]šX][Û—ÜÝY]šX][Û—ÜÝˆ\™›Ü›X[˜ÙWÜÝ]\Ï\\™›Ü›X[˜ÙWÜÝ]\Ëˆ^XÝYÜÛÝ\˜ÙOY^XÝYÜÛÝ\˜ÙKˆ]WÜ]X[]OY]WÜ]X[]Kˆ›Ý\ÏHˆ‹š›Ú[Š›Ý\×Ü\ÊKˆ^[ØYÚœÛÛ\ÝŠÜWÜ›ÝË™Ù]
œ^[ØYÚœÛÛˆŠHÜˆœÛÛ‹™[\ÊÜWÜ›ÝË[œÝ\™WØ\ØÚZOUYJJKˆÙ[XÝYÜ›ÙXÝ[Û—ÚÙ^O\Ù[XÝYÚÙ^KˆÙ[XÝYÜ›ÙXÝ[Û—Ü˜]×Ý˜[YO\Ù[XÝYÜ˜]×Ý˜[YKˆ™Y™\™[˜ÙWÙXYÛ›ÜÝX×ÚœÛÛZœÛÛ‹™[\Ê™Y™\™[˜ÙWÜ™\Ý[È™XYÛ›ÜÝXÈ—K[œÝ\™WØ\ØÚZOUYJKˆ
Bˆ™]\›ˆÂˆ\Ù\ÜÝ]\ÈŽˆ\Ù\ÜÝ]\Ëˆœ›ÙXÝ[Û—ÚÝÚŽˆ›ÙXÝ[Û—ÚÝÚˆœ\™›Ü›X[˜ÙWÜÝ]\ÈŽˆ\™›Ü›X[˜ÙWÜÝ]\Ëˆ™]WÜ]X[]HŽˆ]WÜ]X[]KˆB‚‚™Yˆš\œÝÛ›Û—Ù[\J^[ØYˆXÝÜÝ‹[žWKÙ^\Îˆ\ÝÜÝ—JHOˆÝŽ‚ˆ›ÜˆÙ^H[ˆÙ^\Î‚ˆ˜[YHH^[ØY™Ù]
Ù^JBˆYˆ˜[YH›Ý[ˆ
›Û™KˆŠN‚ˆ™]\›ˆÝŠ˜[YJKœÝš\

Bˆ™]\›ˆˆ‚‚‚™Yˆ\œÙWÙ\Ú[ÛœÛÛ\—Ü—Ú[œ]Ê›ÝÎˆXÝÜÝ‹[žWJHOˆ\VÙXÝÜÝ‹[žWKXÝÜÝ‹[žWWN‚ˆÝ\œ™[ÎˆXÝÜÝ‹[žWHHßBˆ›ÛYÙ\ÎˆXÝÜÝ‹[žWHHßBˆÛÝ\˜ÙHH›ÝË™Ù]
™]R][SX\ŠHYˆ\Ú[œÝ[˜ÙJ›ÝË™Ù]
™]R][SX\ŠKXÝ
H[ÙH›ÝÂˆ›Üˆ[™^[ˆ˜[™ÙJKÍÊN‚ˆÝ\œ™[ÚÙ^HHˆœžÚ[™^WÚH‚ˆ›ÛYÙWÚÙ^HHˆœžÚ[™^WÝH‚ˆYˆÝ\œ™[ÚÙ^H[ˆÛÝ\˜ÙH[™ÛÝ\˜ÙVØÝ\œ™[ÚÙ^WH›Ý[ˆ
›Û™KˆŠN‚ˆÝ\œ™[ÖØÝ\œ™[ÚÙ^WHHÛÝ\˜ÙVØÝ\œ™[ÚÙ^WBˆYˆ›ÛYÙWÚÙ^H[ˆÛÝ\˜ÙH[™ÛÝ\˜ÙVÝ›ÛYÙWÚÙ^WH›Ý[ˆ
›Û™KˆŠN‚ˆ›ÛYÙ\ÖÝ›ÛYÙWÚÙ^WHHÛÝ\˜ÙVÝ›ÛYÙWÚÙ^WBˆ™]\›ˆÝ\œ™[Ë›ÛYÙ\Â‚‚™YˆØ[Ý[]WÜ—Ú[œ]ÚX[
ˆÝ\œ™[ÎˆXÝÜÝ‹[žWKˆ›ÛYÙ\ÎˆXÝÜÝ‹[žWKˆ
‹ˆ^XÝYÜÝš[™×Ú[™^\ÎˆÙ]Ú[KŠHOˆXÝÜÝ‹[žWN‚ˆ^XÝYÚ[œ]ÈHÛÜY
^XÝYÜÝš[™×Ú[™^\ÊBˆ]˜Z[X›WÚ[œ]ÈHˆ[˜]˜Z[X›WÚ[œ]ÈHˆ›ÛYÙWÝ˜[Y\ÎˆXÝÜÝ‹›Ø]HHßBˆ›Üˆ[™^[ˆ^XÝYÚ[œ]Î‚ˆ›ÛYÙHH\œÙWÙ›Ø]Ý˜[YJ›ÛYÙ\Ë™Ù]
ˆœžÚ[™^WÝHŠJBˆ›ÛYÙWÝ˜[Y\ÖÜÝŠ[™^
WHH›ÛYÙHÜˆŒˆYˆ›ÛYÙH\È›Ý›Û™H[™›ÛYÙHˆQUSÔÕ’S‘×Ô‘TÑS•Õ“ÓQÑWÕ‘TÒÓ‚ˆ]˜Z[X›WÚ[œ]È
ÏHBˆ[ÙN‚ˆ[˜]˜Z[X›WÚ[œ]È
ÏHBˆÝ[Ú[œ]ÈH[Š^XÝYÚ[œ]ÊBˆ™]\›ˆÂˆ˜]˜Z[X›WÜÝš[™ÜÈŽˆ]˜Z[X›WÚ[œ]ËˆÝ[ÜÝš[™ÜÈŽˆÝ[Ú[œ]Ëˆ[˜]˜Z[X›WÜÝš[™ÜÈŽˆ[˜]˜Z[X›WÚ[œ]ËˆœÝš[™×Ø]˜Z[Xš[]WÜÝŽˆ›Ý[™
]˜Z[X›WÚ[œ]ÈÈÝ[Ú[œ]È
ˆLŠHYˆÝ[Ú[œ]È[ÙH›Û™Kˆœ—Ú[œ]ÙXYÛ›ÜÝXÜÈŽˆÂˆ™^XÝYÚ[œ]ÈŽˆ^XÝYÚ[œ]Ëˆ›ÛYÙ\×ÝˆŽˆ›ÛYÙWÝ˜[Y\ËˆKˆB‚‚™YˆX\›—Ù^XÝYÜÝš[™Ü×Ùœ›ÛWÝ›ÛYÙJˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ›ÝšY\—Ù]šXÙWÚYˆ[ˆ›ÛYÙ\ÎˆXÝÜÝ‹[žWKˆØœÙ\™YØ]ˆÝ‹ŠHOˆÙ]Ú[N‚ˆX\›™YÚ[™^\ÎˆÙ]Ú[HHÙ]

Bˆ›ÜˆÙ^K˜]×Ý›ÛYÙH[ˆ›ÛYÙ\Ëš][\Ê
N‚ˆ›ÛYÙHH\œÙWÙ›Ø]Ý˜[YJ˜]×Ý›ÛYÙJBˆYˆ›ÛYÙH\È›Û™HÜˆ›ÛYÙHHQUSÔÕ’S‘×Ô‘TÑS•Õ“ÓQÑWÕ‘TÒÓ‚ˆÛÛ[YBˆ[™^H\œÙWÚ[Ý˜[YJÙ^Kœ™[[Ý™\™Yš^
œˆŠKœ™[[Ý™\ÝY™š^
—ÝHŠJBˆYˆ[™^\È›Û™N‚ˆÛÛ[YBˆ^\Ý[™ÈHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕ
‚ˆ”“ÓH›ÝšY\—Ù]šXÙWÙ^XÝYÜÝš[™ÜÂˆÒT‘H›ÝšY\—Ù]šXÙWÚYHÈS‘Ýš[™×Ú[™^HÂˆˆˆ‹ˆ
›ÝšY\—Ù]šXÙWÚY[™^
Kˆ
K™™]ÚÛ™J
Bˆ›ÝÈH]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠBˆYˆ^\Ý[™È\È›Û™N‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆS”ÑT•S•È›ÝšY\—Ù]šXÙWÙ^XÝYÜÝš[™ÜÈ
ˆ›ÝšY\—Ù]šXÙWÚYÝš[™×Ú[™^^XÝYÛÝ\˜ÙKØœÙ\™YØÛÝ[ˆš\œÝÛØœÙ\™YØ]\ÝÛØœÙ\™YØ]Ü™X]YØ]\]YØ]ˆ
HSQTÈ
ËË	Ø]]ÉËKËËËÊBˆˆˆ‹ˆ
›ÝšY\—Ù]šXÙWÚY[™^ØœÙ\™YØ]ØœÙ\™YØ]›ÝË›ÝÊKˆ
BˆÛÛ[YBˆØœÙ\™YØÛÝ[H[
^\Ý[™ÖÈ›ØœÙ\™YØÛÝ[—HÜˆ
H
ÈBˆ^XÝYH[
^\Ý[™ÖÈ™^XÝY—HÜˆ
BˆÛÝ\˜ÙHH^\Ý[™ÖÈœÛÝ\˜ÙH—BˆYˆÛÝ\˜ÙHOH˜]]Èˆ[™ØœÙ\™YØÛÝ[HQUSÔÕ’S‘×ÐUU×ÓPT“—ÓÐ”ÑT•USÓ”Î‚ˆ^XÝYHBˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUH›ÝšY\—Ù]šXÙWÙ^XÝYÜÝš[™ÜÂˆÑU^XÝYHËØœÙ\™YØÛÝ[HË\ÝÛØœÙ\™YØ]HË\]YØ]HÂˆÒT‘HYHÂˆˆˆ‹ˆ
^XÝYØœÙ\™YØÛÝ[ØœÙ\™YØ]›ÝË^\Ý[™ÖÈšY—JKˆ
B‚ˆ›ÝÜÈH]Y\žWØ[
ˆÛÛ›‹ˆˆˆ‚ˆÑSPÕÝš[™×Ú[™^ˆ”“ÓH›ÝšY\—Ù]šXÙWÙ^XÝYÜÝš[™ÜÂˆÒT‘H›ÝšY\—Ù]šXÙWÚYHÈS‘^XÝYHBˆˆˆ‹ˆ
›ÝšY\—Ù]šXÙWÚY
Kˆ
BˆX\›™YÚ[™^\Ë\]J[
›ÝÖÈœÝš[™×Ú[™^—JH›Üˆ›ÝÈ[ˆ›ÝÜÊBˆ™]\›ˆX\›™YÚ[™^\Â‚‚™Yˆ›Ü›X[^™WÙ\Ú[ÛœÛÛ\—Ù]šXÙWÚY[]J›ÝÎˆXÝÜÝ‹[žWJHOˆXÝÜÝ‹[žWN‚ˆ]—Ý\WÚYH\œÙWÚ[Ý˜[YJš\œÝÛ›Û—Ù[\J›ÝËÈ™]•\RY‹™]—Ý\WÚY‹™]šXÙU\RY—JJBˆ[Ù[Hš\œÝÛ›Û—Ù[\J›ÝËÈ›[Ù[‹™]“[Ù[‹™]šXÙS[Ù[‹š[•\H—JBˆ˜]YÜÝÙ\—ÚÝÈH›Ü›X[^™WÜÝÙ\—Ý×ÚÝÊš\œÝÛ›Û—Ù[\J›ÝËÈœ˜]YÝÙ\ˆ‹œ˜]YÜÝÙ\ˆ‹˜Ø\XÚ]H‹››ÛZ[˜[ÝÙ\ˆ—JJBˆ™]\›ˆÂˆœÝ][Û—ØÛÙHŽˆš\œÝÛ›Û—Ù[\J›ÝËÈœÝ][ÛÛÙH‹œ[ÛÙH—JKˆ™^\›˜[Ù]šXÙWÚYŽˆš\œÝÛ›Û—Ù[\J›ÝËÈ™]’Y‹šY‹™]‘ˆ‹™]šXÙQˆ‹™\ÛÛÙH‹œÛˆ—JKˆ™]—ÙˆŽˆš\œÝÛ›Û—Ù[\J›ÝËÈ™]‘ˆ‹™]šXÙQˆ—JKˆœÛˆŽˆš\œÝÛ›Û—Ù[\J›ÝËÈ™\ÛÛÙH‹œÛˆ—JKˆ™]šXÙWÛ˜[YHŽˆš\œÝÛ›Û—Ù[\J›ÝËÈ™]“˜[YH‹™]šXÙS˜[YH‹›˜[YH—JKˆ™]—Ý\WÚYŽˆ]—Ý\WÚYˆ›[Ù[Žˆ[Ù[ˆœ˜]YÜÝÙ\—ÚÝÈŽˆ˜]YÜÝÙ\—ÚÝÈYˆ˜]YÜÝÙ\—ÚÝÈ\È›Ý›Û™H[ÙH[™™\—Ú[™\\—ÜÝÙ\—Ùœ›ÛWÛ[Ù[
[Ù[
KˆB‚‚™Yˆ›Ü›X[^™WÜÝÙ\—Ý×ÚÝÊ˜[YNˆ[žJHOˆ›Ø]›Û™N‚ˆ\œÙYH\œÙWÙ›Ø]Ý˜[YJ˜[YJBˆYˆ\œÙY\È›Û™N‚ˆ™]\›ˆ›Û™Bˆ™]\›ˆ\œÙYÈLYˆ\œÙYˆL[ÙH\œÙY‚‚™Yˆ[™™\—Ú[™\\—ÜÝÙ\—Ùœ›ÛWÛ[Ù[
[Ù[ˆÝˆ›Û™JHOˆ›Ø]›Û™N‚ˆ›Ü›X[^™YHÝŠ[Ù[ÜˆˆŠK\\Š
Kœ™\XÙJˆ‹ˆŠBˆX]ÚH™KœÙX\˜Ú
ˆŠÎ”ÕSŒŒ_ŠJ
ÊÎ–Ë‹W
ÊOÊJÎ’ÕÊÎ‹_	
JH‹›Ü›X[^™Y
BˆYˆ›ÝX]Ú‚ˆ™]\›ˆ›Û™Bˆ™]\›ˆ\œÙWÙ›Ø]Ý˜[YJX]Ú™Ü›Ý\
JKœ™\XÙJ‹‹‹ˆŠJB‚‚™YˆÜ[]WÛZ\ÜÚ[™×Ú[™\\—Ü˜]YÜÝÙ\ŠÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[ÛŠHOˆ[‚ˆ›ÝÜÈHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕY^\›˜[Ù]šXÙWÚY[Ù[^[ØYÚœÛÛ‚ˆ”“ÓH›ÝšY\—Ù]šXÙ\ÂˆÒT‘H˜]YÜÝÙ\—ÚÝÈTÈ•SÔˆ˜]YÜÝÙ\—ÚÝÈHˆˆˆ‚ˆ
K™™]Ú[

Bˆ\]YHˆ›Üˆ›ÝÈ[ˆ›ÝÜÎ‚ˆ[Ù[H›ÝÖÈ›[Ù[—BˆYˆ›Ý[Ù[[™›ÝÖÈœ^[ØYÚœÛÛˆ—N‚ˆžN‚ˆ^[ØYHœÛÛ‹›ØYÊ›ÝÖÈœ^[ØYÚœÛÛˆ—JBˆ[Ù[Hš\œÝÛ›Û—Ù[\J^[ØYÈ›[Ù[‹™]“[Ù[‹™]šXÙS[Ù[‹š[•\H—JBˆ^Ù\
\Q\œ›Ü‹˜[YQ\œ›Ü‹œÛÛ‹’”ÓÓ‘XÛÙQ\œ›ÜŠN‚ˆ[Ù[H›Û™Bˆ˜]YÜÝÙ\—ÚÝÈH[™™\—Ú[™\\—ÜÝÙ\—Ùœ›ÛWÛ[Ù[
[Ù[
BˆYˆ˜]YÜÝÙ\—ÚÝÈ\È›Û™N‚ˆÛÛ[YBˆÛÛ›‹™^XÝ]Jˆ•TUH›ÝšY\—Ù]šXÙ\ÈÑU˜]YÜÝÙ\—ÚÝÈHÈÒT‘HYHÈ‹ˆ
˜]YÜÝÙ\—ÚÝË›ÝÖÈšY—JKˆ
BˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUH[™\\—Ø]˜Z[Xš[]WÙZ[BˆÑU[™\\—ÜÝÙ\—ÚÝÈHÂˆÒT‘H›ÝšY\ˆHÈS‘[™\\—ÚYHÂˆS‘
[™\\—ÜÝÙ\—ÚÝÈTÈ•SÔˆ[™\\—ÜÝÙ\—ÚÝÈH
Bˆˆˆ‹ˆ
˜]YÜÝÙ\—ÚÝËS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹›ÝÖÈ™^\›˜[Ù]šXÙWÚY—JKˆ
BˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUH[™\\—ÜÝÙ\—ÜØ[\\ÂˆÑU[™\\—ÜÝÙ\—ÚÝÈHÂˆÒT‘H›ÝšY\ˆHÈS‘[™\\—ÚYHÂˆS‘
[™\\—ÜÝÙ\—ÚÝÈTÈ•SÔˆ[™\\—ÜÝÙ\—ÚÝÈH
Bˆˆˆ‹ˆ
˜]YÜÝÙ\—ÚÝËS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹›ÝÖÈ™^\›˜[Ù]šXÙWÚY—JKˆ
Bˆ\]Y
ÏHBˆ™]\›ˆ\]Y‚‚™Yˆ\ØX›WÜ™[[Ý™YÚ[™\\—Ù]šXÙ\ÊÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[ÛŠHOˆ[‚ˆ›ÝÜÈHÛÛ›‹™^XÝ]Jˆ”ÑSPÕY]šXÙWÛ˜[YH”“ÓH›ÝšY\—Ù]šXÙ\ÈÒT‘H[˜X›YHH‚ˆ
K™™]Ú[

Bˆ™[[Ý™YÚYÈHÚ[
›ÝÖÈšY—JH›Üˆ›ÝÈ[ˆ›ÝÜÈYˆ\×Ü™[[Ý™YÚ[™\\—Û˜[YJ›ÝÖÈ™]šXÙWÛ˜[YH—JWBˆYˆ™[[Ý™YÚYÎ‚ˆÛÛ›‹™^XÝ][X[žJˆ•TUH›ÝšY\—Ù]šXÙ\ÈÑU[˜X›YHÒT‘HYHÈ‹ˆÊ]šXÙWÚY
H›Üˆ]šXÙWÚY[ˆ™[[Ý™YÚY×Kˆ
Bˆ›Üˆ]šXÙWÚY[ˆ™[[Ý™YÚYÎ‚ˆ™XÛÜ™Ù]šXÙWØÛÛ™šYÝ\˜][ÛŠˆÛÛ›‹ˆ›ÝšY\—Ù]šXÙWÚYY]šXÙWÚYˆXÝ]™OQ˜[ÙKˆY™™XÝ]™WÙ]OXÝ\œ™[Û\Ø›Û—Ù]J
Kˆ
Bˆ™]\›ˆ[Š™[[Ý™YÚYÊB‚‚™Yˆ\×Ü™[[Ý™YÚ[™\\—Û˜[YJ]šXÙWÛ˜[YNˆÝˆ›Û™JHOˆ›ÛÛ‚ˆ›Ü›X[^™YH›Ü›X[^™WÛ˜[YJÝŠ]šXÙWÛ˜[YHÜˆˆŠJBˆ™]\›ˆ[žJX\šÙ\ˆ[ˆ›Ü›X[^™Y›ÜˆX\šÙ\ˆ[ˆ
œ™[[ÝšYÈ‹œ™[[Ý™YŠJB‚‚™Yˆ\×Ú[™\\—Ø]˜Z[X›JXÝ]™WÜÝÙ\—ÚÝÎˆ›Ø]›Û™JHOˆ›ÛÛ‚ˆ™]\›ˆ™\Ü[™×Ú\×Ú[™\\—Ø]˜Z[X›JXÝ]™WÜÝÙ\—ÚÝÊB‚‚™Yˆ[™\\—Ø]˜Z[Xš[]WÜÛÝ
Ø[\WÝ[YNˆ]][YJHOˆ]][YN‚ˆ™]\›ˆ™\Ü[™×Ú[™\\—Ø]˜Z[Xš[]WÜÛÝ
ˆØ[\WÝ[YKˆÛÝÛZ[]\ÏRS•‘T•T—ÐURSP’SUWÔÓÕÓRS•UTËˆ
B‚‚™Yˆ\WÚ[™\\—ÙYÙWÝÛ\˜[˜ÙJˆ˜[YÜÛÝÎˆÙ]Ù]][YWKˆÛ\˜[˜ÙWÛZ[]\Îˆ[HS•‘T•T—ÐURSP’SUWÑQÑWÕÓTSÑWÓRS•UTËŠHOˆÙ]Ù]][YWN‚ˆ™]\›ˆ™\Ü[™×Ø\WÚ[™\\—ÙYÙWÝÛ\˜[˜ÙJˆ˜[YÜÛÝËˆÛ\˜[˜ÙWÛZ[]\Ï]Û\˜[˜ÙWÛZ[]\Ëˆ
B‚‚™YˆØ[Ý[]WÚ[™\\—ÙZ[WØ]˜Z[Xš[]JˆØ[\\Îˆ\ÝÙXÝÜÝ‹[žWWKˆ˜[YÜÛÝÎˆÙ]Ù]][YWH›Û™HH›Û™KˆYÙWÝÛ\˜[˜ÙWÛZ[]\Îˆ[HS•‘T•T—ÐURSP’SUWÑQÑWÕÓTSÑWÓRS•UTËŠHOˆXÝÜÝ‹[žWN‚ˆ™]\›ˆ™\Ü[™×ØØ[Ý[]WÚ[™\\—ÙZ[WØ]˜Z[Xš[]JˆØ[\\Ëˆ˜[YÜÛÝËˆÛÝÛZ[]\ÏRS•‘T•T—ÐURSP’SUWÔÓÕÓRS•UTËˆYÙWÝÛ\˜[˜ÙWÛZ[]\ÏYYÙWÝÛ\˜[˜ÙWÛZ[]\Ëˆ
B‚‚™YˆØ[Ý[]WÝÙZYÚYÜ[Ø]˜Z[Xš[]J[™\\—Ü›ÝÜÎˆ\ÝÙXÝÜÝ‹[žWWJHOˆ›Ø]›Û™N‚ˆ™]\›ˆ™\Ü[™×ØØ[Ý[]WÝÙZYÚYÜ[Ø]˜Z[Xš[]J[™\\—Ü›ÝÜÊB‚‚™Yˆ™\ÛÛ™WÚ[™\\—Ø]˜Z[Xš[]WÜ\š[Ù
ˆ\š[ÙˆÝ‹ˆ˜]×Ùœ›ÛWÙ]NˆÝˆHˆ‹ˆ˜]×Ý×Ù]NˆÝˆHˆ‹ŠHOˆ\VÙ]K]WN‚ˆY\Ý\™^HH]KÙ^J
HH[YY[J^\ÏLJBˆYˆ\š[ÙOH˜Ý\œ™[Û[ÛŽ‚ˆ™]\›ˆY\Ý\™^Kœ™\XÙJ^OLJKY\Ý\™^BˆYˆ\š[ÙOHœ™]š[Ý\×Û[ÛŽ‚ˆÝ\œ™[Û[ÛÜÝ\H]KÙ^J
Kœ™\XÙJ^OLJBˆ™]š[Ý\×Û[ÛÙ[™HÝ\œ™[Û[ÛÜÝ\H[YY[J^\ÏLJBˆ™]\›ˆ™]š[Ý\×Û[ÛÙ[™œ™\XÙJ^OLJK™]š[Ý\×Û[ÛÙ[™ˆYˆ\š[ÙOH˜Ý\ÝÛHŽ‚ˆœ›ÛWÙ]HH\œÙWÙ]WÝ˜[YJ˜]×Ùœ›ÛWÙ]JHÜˆY\Ý\™^Bˆ×Ù]HHZ[Š\œÙWÙ]WÝ˜[YJ˜]×Ý×Ù]JHÜˆY\Ý\™^KY\Ý\™^JBˆ™]\›ˆœ›ÛWÙ]K×Ù]Bˆ™]\›ˆY\Ý\™^KY\Ý\™^B‚‚™YˆÙ]Ú[™\\—Ø]˜Z[Xš[]WÜ™\Ü
ˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆœ›ÛWÙ]Nˆ]Kˆ×Ù]Nˆ]Kˆ
‹ˆ\ÜÙ]ÚYˆ[›Û™HH›Û™KˆÛWÛÛ›Nˆ›ÛÛHYKˆÙX\˜ÚˆÝˆHˆ‹ŠHOˆXÝÜÝ‹[žWN‚ˆYˆœ›ÛWÙ]Hˆ×Ù]N‚ˆ™]\›ˆÂˆ˜]™\˜YÙWÜÝŽˆ›Û™Kˆœ[ÈŽˆ×Kˆš[™\\œÈŽˆ×KˆÛÜœÝÜ[Žˆ›Û™Kˆ›Ý×Ø]˜Z[Xš[]WØÛÝ[ŽˆˆBˆÛÛ™][ÛœÈHÂˆšXYœ›ÝšY\ˆHÈ‹ˆšXY˜]˜Z[Xš[]WÙ]H‘UÑQSˆÈS‘È‹ˆBˆ\˜[\Îˆ\ÝÐ[žWHHÂˆS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹ˆœ›ÛWÙ]Kš\ÛÙ›Ü›X]

Kˆ×Ù]Kš\ÛÙ›Ü›X]

KˆBˆYˆ\ÜÙ]ÚY\È›Ý›Û™N‚ˆÛÛ™][ÛœË˜\[™
šXY˜\ÜÙ]ÚYHÈŠBˆ\˜[\Ë˜\[™
\ÜÙ]ÚY
BˆYˆÛWÛÛ›N‚ˆÛÛ™][ÛœË˜\[™
˜K˜XÝ]™WØÛÛ˜XÝH	ÞY\ÉÈŠBˆYˆÙX\˜Ú‚ˆÛÛ™][ÛœË˜\[™
ˆŠKœ›Ú™XÝÛ˜[YHRÑHÈÔˆK›ØØ][ÛˆRÑHÈÔˆK˜ÛÛ\[žWÛ˜[YHRÑHÈÔˆK˜[X\×Ø›ØˆRÑHÊH‚ˆ
BˆÚ[Ø\™Hˆ‰^ÜÙX\˜ÚIH‚ˆ\˜[\Ë™^[™
ÝÚ[Ø\™Ú[Ø\™Ú[Ø\™Ú[Ø\™JBˆ›ÝÜÈHÛÛ›‹™^XÝ]Jˆˆˆˆ‚ˆÑSPÕˆXY˜\ÜÙ]ÚYˆKœ›Ú™XÝÛ˜[YKˆXYš[™\\—ÚYˆXYš[™\\—Û˜[YKˆXYš[™\\—ÜÝÙ\—ÚÝËˆPV
™]šXÙWÛ˜[YJHTÈ›ÝšY\—Ù]šXÙWÛ˜[YKˆPV
œ˜]YÜÝÙ\—ÚÝÊHTÈ›ÝšY\—ÜÝÙ\—ÚÝËˆPV
›[Ù[
HTÈ›ÝšY\—Û[Ù[ˆPV
™[˜X›Y
HTÈ›ÝšY\—Ù[˜X›YˆÕSJXY˜[YÜÛÝÊHTÈ˜[YÜÛÝËˆÕSJXY˜]˜Z[X›WÜÛÝÊHTÈ]˜Z[X›WÜÛÝËˆÕSJXY[˜]˜Z[X›WÜÛÝÊHTÈ[˜]˜Z[X›WÜÛÝÂˆ”“ÓH[™\\—Ø]˜Z[Xš[]WÙZ[HXYˆ“ÒSˆ\ÜÙ]ÈHÓˆKšYHXY˜\ÜÙ]ÚYˆQ•“ÒSˆ›ÝšY\—Ù]šXÙ\ÈˆÓˆœ›ÝšY\ˆHXYœ›ÝšY\ˆS‘™^\›˜[Ù]šXÙWÚYHXYš[™\\—ÚYˆÒT‘HÉÈS‘	Ëš›Ú[ŠÛÛ™][ÛœÊ_BˆÔ“ÕT–HXY˜\ÜÙ]ÚYKœ›Ú™XÝÛ˜[YKXYš[™\\—ÚYXYš[™\\—Û˜[YKXYš[™\\—ÜÝÙ\—ÚÝÂˆU’S‘ÈÕSJXY˜[YÜÛÝÊHˆˆˆˆ‹ˆ\˜[\Ëˆ
K™™]Ú[

Bˆ[™\\—Ü›ÝÜÎˆ\ÝÙXÝÜÝ‹[žWWHH×Bˆ›Üˆ›ÝÈ[ˆ›ÝÜÎ‚ˆ][HHXÝ
›ÝÊBˆ]šXÙWÛ˜[YHH][K™Ù]
œ›ÝšY\—Ù]šXÙWÛ˜[YHŠHÜˆ][K™Ù]
š[™\\—Û˜[YHŠBˆYˆ][K™Ù]
œ›ÝšY\—Ù[˜X›YŠHOHÜˆ\×Ü™[[Ý™YÚ[™\\—Û˜[YJ]šXÙWÛ˜[YJN‚ˆÛÛ[YBˆ][VÈš[™\\—Û˜[YH—HH]šXÙWÛ˜[YBˆ][VÈš[™\\—ÜÝÙ\—ÚÝÈ—HH
ˆ\œÙWÙ›Ø]Ý˜[YJ][K™Ù]
š[™\\—ÜÝÙ\—ÚÝÈŠJBˆÜˆ\œÙWÙ›Ø]Ý˜[YJ][K™Ù]
œ›ÝšY\—ÜÝÙ\—ÚÝÈŠJBˆÜˆ[™™\—Ú[™\\—ÜÝÙ\—Ùœ›ÛWÛ[Ù[
][K™Ù]
œ›ÝšY\—Û[Ù[ŠJBˆ
Bˆ][VÈ˜]˜Z[Xš[]WÜÝ—HH›Ý[™
][VÈ˜]˜Z[X›WÜÛÝÈ—HÈ][VÈ˜[YÜÛÝÈ—H
ˆLŠBˆ[™\\—Ü›ÝÜË˜\[™
][JB‚ˆ[×ØžWÚYˆXÝÚ[XÝÜÝ‹[žWWHHßBˆ›Üˆ›ÝÈ[ˆ[™\\—Ü›ÝÜÎ‚ˆ[H[×ØžWÚYœÙ]Y˜][
ˆ[
›ÝÖÈ˜\ÜÙ]ÚY—JKˆÈ˜\ÜÙ]ÚYŽˆ›ÝÖÈ˜\ÜÙ]ÚY—Kœ›Ú™XÝÛ˜[YHŽˆ›ÝÖÈœ›Ú™XÝÛ˜[YH—Kš[™\\œÈŽˆ×_Kˆ
Bˆ[Èš[™\\œÈ—K˜\[™
›ÝÊBˆ[Ü›ÝÜÎˆ\ÝÙXÝÜÝ‹[žWWHH×Bˆ›Üˆ[[ˆ[×ØžWÚY˜[Y\Ê
N‚ˆ[Ü›ÝÜË˜\[™
ˆÂˆ˜\ÜÙ]ÚYŽˆ[È˜\ÜÙ]ÚY—Kˆœ›Ú™XÝÛ˜[YHŽˆ[Èœ›Ú™XÝÛ˜[YH—Kˆš[™\\—ØÛÝ[Žˆ[Š[Èš[™\\œÈ—JKˆ˜]˜Z[Xš[]WÜÝŽˆØ[Ý[]WÝÙZYÚYÜ[Ø]˜Z[Xš[]J[Èš[™\\œÈ—JKˆØ\›š[™ÜÈŽˆ™\Ü[™×Ú[˜[YÜÝÙ\—ÝØ\›š[™ÜÊˆ[Èš[™\\œÈ—KˆÝÙ\—ÚÙ^OHš[™\\—ÜÝÙ\—ÚÝÈ‹ˆØ\›š[™×ØÛÙOH›Z\ÜÚ[™×Ú[™\\—ÜÝÙ\ˆ‹ˆ
KˆBˆ
Bˆ[Ü›ÝÜËœÛÜ
Ù^O[[X™H›ÝÎˆ
›ÝÖÈ˜]˜Z[Xš[]WÜÝ—H\È›Û™K›ÝÖÈ˜]˜Z[Xš[]WÜÝ—HÜˆ›ÝÖÈœ›Ú™XÝÛ˜[YH—JJBˆ[™\\—Ü›ÝÜËœÛÜ
Ù^O[[X™H›ÝÎˆ
›ÝÖÈ˜]˜Z[Xš[]WÜÝ—K›ÝÖÈœ›Ú™XÝÛ˜[YH—K›ÝÖÈš[™\\—Û˜[YH—HÜˆ›ÝÖÈš[™\\—ÚY—JJBˆ›Üˆ˜[šË›ÝÈ[ˆ[[Y\˜]J[Ü›ÝÜËÝ\LJN‚ˆ›ÝÖÈœ˜[šÈ—HH˜[šÂˆ›Üˆ˜[šË›ÝÈ[ˆ[[Y\˜]J[™\\—Ü›ÝÜËÝ\LJN‚ˆ›ÝÖÈœ˜[šÈ—HH˜[šÂˆ\˜Ù[YÙ\ÈHÙ›Ø]
›ÝÖÈ˜]˜Z[Xš[]WÜÝ—JH›Üˆ›ÝÈ[ˆ[Ü›ÝÜÈYˆ›ÝÖÈ˜]˜Z[Xš[]WÜÝ—H\È›Ý›Û™WBˆ™]\›ˆÂˆ˜]™\˜YÙWÜÝŽˆ›Ý[™
Ý[J\˜Ù[YÙ\ÊHÈ[Š\˜Ù[YÙ\ÊKŠHYˆ\˜Ù[YÙ\È[ÙH›Û™Kˆœ[ÈŽˆ[Ü›ÝÜËˆš[™\\œÈŽˆ[™\\—Ü›ÝÜËˆÛÜœÝÜ[Žˆ[Ü›ÝÜÖÌHYˆ[Ü›ÝÜÈ[ÙH›Û™Kˆ›Ý×Ø]˜Z[Xš[]WØÛÝ[ŽˆÝ[JˆH›Üˆ›ÝÈ[ˆ[™\\—Ü›ÝÜÈYˆ›Ø]
›ÝÖÈ˜]˜Z[Xš[]WÜÝ—JHÕ×ÒS•‘T•T—ÐURSP’SUWÔÕˆ
KˆB‚‚™YˆÙ]Û[ÛWÝØ]Ü™\ÜÙ]JˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆœ›ÛWÙ]Nˆ]Kˆ×Ù]Nˆ]Kˆ\ÜÙ]ÚYˆ[›Û™HH›Û™KŠHOˆXÝÜÝ‹[žWN‚ˆYˆœ›ÛWÙ]Hˆ×Ù]N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ“È[\˜[ÈÐUH[˜[YËˆŠB‚ˆ›ÝšY\ˆHS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‚ˆœ›ÛWÚ\ÛÈHœ›ÛWÙ]Kš\ÛÙ›Ü›X]

Bˆ×Ú\ÛÈH×Ù]Kš\ÛÙ›Ü›X]

BˆØ[\WÙœ›ÛHH]][YK˜ÛÛXš[™Jœ›ÛWÙ]K]][YK›Z[‹[YJ
JKš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠBˆØ[\WÝÈH]][YK˜ÛÛXš[™J×Ù]H
È[YY[J^\ÏLJK]][YK›Z[‹[YJ
JKš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠBˆ^XÝYÙ^\ÈH
×Ù]HHœ›ÛWÙ]JK™^\È
ÈB‚ˆ\ÜÙ]Ü\˜[\Îˆ\ÝÐ[žWHH×BˆYˆ\ÜÙ]ÚY\È›Ý›Û™N‚ˆ\ÜÙ]Ùš[\ˆH˜KšYHÈ‚ˆ\ÜÙ]Ü\˜[\Ë˜\[™
\ÜÙ]ÚY
Bˆ[ÙN‚ˆ\ÜÙ]Ùš[\ˆHˆˆ‚ˆVTÕÈ
ˆÑSPÕH”“ÓH›ÝšY\—Ù]šXÙ\ÈˆÒT‘H˜\ÜÙ]ÚYHKšYS‘œ›ÝšY\ˆHÈS‘™[˜X›YHHS‘™]—Ý\WÚYSˆ
KÎ
Bˆ
BˆÔˆVTÕÈ
ˆÑSPÕH”“ÓH[™\\—Ø]˜Z[Xš[]WÙZ[HXYˆÒT‘HXY˜\ÜÙ]ÚYHKšYS‘XYœ›ÝšY\ˆHÈS‘XY˜]˜Z[Xš[]WÙ]H‘UÑQSˆÈS‘Âˆ
BˆÔˆVTÕÈ
ˆÑSPÕH”“ÓH[Ø]˜Z[Xš[]WÙZ[HYˆÒT‘HY˜\ÜÙ]ÚYHKšYS‘Yœ›ÝšY\ˆHÈS‘Y˜]˜Z[Xš[]WÙ]H‘UÑQSˆÈS‘Âˆ
BˆÔˆVTÕÈ
ˆÑSPÕH”“ÓH[™\\—ÜÝÙ\—ÜØ[\\È\ÂˆÒT‘H\Ë˜\ÜÙ]ÚYHKšYS‘\Ëœ›ÝšY\ˆHÈS‘\ËœØ[\WÝ[YHHÈS‘\ËœØ[\WÝ[YHÂˆ
Bˆˆˆ‚ˆ\ÜÙ]Ü\˜[\Ë™^[™
ˆÜ›ÝšY\‹›ÝšY\‹œ›ÛWÚ\ÛË×Ú\ÛË›ÝšY\‹œ›ÛWÚ\ÛË×Ú\ÛË›ÝšY\‹Ø[\WÙœ›ÛKØ[\WÝ×Bˆ
Bˆ\ÜÙ]ÈHÛÛ›‹™^XÝ]Jˆˆ”ÑSPÕKšYKœ›Ú™XÝÛ˜[YH”“ÓH\ÜÙ]ÈHÒT‘HØ\ÜÙ]Ùš[\ŸHÔ‘Tˆ–HKœ›Ú™XÝÛ˜[YHÓÓUH“ÐÐTÑH‹ˆ\ÜÙ]Ü\˜[\Ëˆ
K™™]Ú[

B‚ˆ[Îˆ\ÝÙXÝÜÝ‹[žWWHH×Bˆ›Üˆ\ÜÙ][ˆ\ÜÙ]Î‚ˆÝ\œ™[Ø\ÜÙ]ÚYH[
\ÜÙ]ÈšY—JBˆÛÛ™šYÝ\™YÙ]šXÙ\ÈH]Y\žWØ[
ˆÛÛ›‹ˆˆˆ‚ˆÑSPÕ^\›˜[Ù]šXÙWÚY]šXÙWÛ˜[YBˆ”“ÓH›ÝšY\—Ù]šXÙ\ÂˆÒT‘H\ÜÙ]ÚYHÈS‘›ÝšY\ˆHÈS‘[˜X›YHHS‘]—Ý\WÚYSˆ
KÎ
Bˆˆˆ‹ˆ
Ý\œ™[Ø\ÜÙ]ÚY›ÝšY\ŠKˆ
BˆÛÛ™šYÝ\™YÚ[™\\—ØÛÝ[HÝ[JˆH›Üˆ›ÝÈ[ˆÛÛ™šYÝ\™YÙ]šXÙ\ÈYˆ›Ý\×Ü™[[Ý™YÚ[™\\—Û˜[YJ›ÝÖÈ™]šXÙWÛ˜[YH—JBˆ
Bˆ[™\\—Ü›ÝÜÈH]Y\žWØ[
ˆÛÛ›‹ˆˆˆ‚ˆÑSPÕˆXYš[™\\—ÚYˆÓÐSTÐÑJPV
™]šXÙWÛ˜[YJKPV
XYš[™\\—Û˜[YJKXYš[™\\—ÚY
HTÈ[™\\—Û˜[YKˆPV
XYš[™\\—ÜÝÙ\—ÚÝÊHTÈÝÜ™YÜÝÙ\—ÚÝËˆPV
œ˜]YÜÝÙ\—ÚÝÊHTÈ›ÝšY\—ÜÝÙ\—ÚÝËˆPV
›[Ù[
HTÈ›ÝšY\—Û[Ù[ˆPV
™[˜X›Y
HTÈ›ÝšY\—Ù[˜X›YˆÓÕS•
TÕSÕXY˜]˜Z[Xš[]WÙ]JHTÈ]WÙ^\ËˆÕSJXY˜[YÜÛÝÊHTÈ˜[YÜÛÝËˆÕSJXY˜]˜Z[X›WÜÛÝÊHTÈ]˜Z[X›WÜÛÝËˆÕSJXY[˜]˜Z[X›WÜÛÝÊHTÈ[˜]˜Z[X›WÜÛÝÂˆ”“ÓH[™\\—Ø]˜Z[Xš[]WÙZ[HXYˆQ•“ÒSˆ›ÝšY\—Ù]šXÙ\ÈˆÓˆœ›ÝšY\ˆHXYœ›ÝšY\ˆS‘™^\›˜[Ù]šXÙWÚYHXYš[™\\—ÚYˆÒT‘HXY˜\ÜÙ]ÚYHÈS‘XYœ›ÝšY\ˆHÈS‘XY˜]˜Z[Xš[]WÙ]H‘UÑQSˆÈS‘ÂˆÔ“ÕT–HXYš[™\\—ÚYˆÔ‘Tˆ–H[™\\—Û˜[YHÓÓUH“ÐÐTÑKXYš[™\\—ÚYˆˆˆ‹ˆ
Ý\œ™[Ø\ÜÙ]ÚY›ÝšY\‹œ›ÛWÚ\ÛË×Ú\ÛÊKˆ
Bˆ™\ÜÚ[™\\œÎˆ\ÝÙXÝÜÝ‹[žWWHH×Bˆ›ÜˆÝÜ™YÜ›ÝÈ[ˆ[™\\—Ü›ÝÜÎ‚ˆ›ÝÈHXÝ
ÝÜ™YÜ›ÝÊBˆYˆ›ÝË™Ù]
œ›ÝšY\—Ù[˜X›YŠHOHÜˆ\×Ü™[[Ý™YÚ[™\\—Û˜[YJ›ÝË™Ù]
š[™\\—Û˜[YHŠJN‚ˆÛÛ[YBˆ˜[YÜÛÝÈH[
›ÝË™Ù]
˜[YÜÛÝÈŠHÜˆ
Bˆ]˜Z[X›WÜÛÝÈH[
›ÝË™Ù]
˜]˜Z[X›WÜÛÝÈŠHÜˆ
Bˆ™\ÜÚ[™\\œË˜\[™
ˆÂˆš[™\\—ÚYŽˆÝŠ›ÝÖÈš[™\\—ÚY—JKˆš[™\\—Û˜[YHŽˆÝŠ›ÝË™Ù]
š[™\\—Û˜[YHŠHÜˆ›ÝÖÈš[™\\—ÚY—JKˆš[™\\—ÜÝÙ\—ÚÝÈŽˆ
ˆ\œÙWÙ›Ø]Ý˜[YJ›ÝË™Ù]
œÝÜ™YÜÝÙ\—ÚÝÈŠJBˆÜˆ\œÙWÙ›Ø]Ý˜[YJ›ÝË™Ù]
œ›ÝšY\—ÜÝÙ\—ÚÝÈŠJBˆÜˆ[™™\—Ú[™\\—ÜÝÙ\—Ùœ›ÛWÛ[Ù[
›ÝË™Ù]
œ›ÝšY\—Û[Ù[ŠJBˆ
Kˆ™]WÙ^\ÈŽˆ[
›ÝË™Ù]
™]WÙ^\ÈŠHÜˆ
Kˆ˜[YÜÛÝÈŽˆ˜[YÜÛÝËˆ˜]˜Z[X›WÜÛÝÈŽˆ]˜Z[X›WÜÛÝËˆ[˜]˜Z[X›WÜÛÝÈŽˆ[
›ÝË™Ù]
[˜]˜Z[X›WÜÛÝÈŠHÜˆ
Kˆ˜]˜Z[Xš[]WÜÝŽˆ›Ý[™
]˜Z[X›WÜÛÝÈÈ˜[YÜÛÝÈ
ˆLŠHYˆ˜[YÜÛÝÈ[ÙH›Û™KˆBˆ
B‚ˆ[Ù^\ÈHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕ]˜Z[Xš[]WÙ]K˜[YÜÛÝËÙZYÚYØ]˜Z[Xš[]WÜÝ[™\\—ØÛÝ[ˆ”“ÓH[Ø]˜Z[Xš[]WÙZ[BˆÒT‘H\ÜÙ]ÚYHÈS‘›ÝšY\ˆHÈS‘]˜Z[Xš[]WÙ]H‘UÑQSˆÈS‘ÂˆÔ‘Tˆ–H]˜Z[Xš[]WÙ]Bˆˆˆ‹ˆ
Ý\œ™[Ø\ÜÙ]ÚY›ÝšY\‹œ›ÛWÚ\ÛË×Ú\ÛÊKˆ
K™™]Ú[

BˆØ[\WÜÝ[[X\žHHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕÓÕS•

ŠHTÈØ[\WØÛÝ[ÓÕS•
TÕSÕÝXœÝŠØ[\WÝ[YKKL
JHTÈØ[\WÙ^\Âˆ”“ÓH[™\\—ÜÝÙ\—ÜØ[\\ÂˆÒT‘H\ÜÙ]ÚYHÈS‘›ÝšY\ˆHÈS‘Ø[\WÝ[YHHÈS‘Ø[\WÝ[YHÂˆˆˆ‹ˆ
Ý\œ™[Ø\ÜÙ]ÚY›ÝšY\‹Ø[\WÙœ›ÛKØ[\WÝÊKˆ
K™™]ÚÛ™J
B‚ˆØ[\WØÛÝ[H[
Ø[\WÜÝ[[X\žVÈœØ[\WØÛÝ[—HÜˆ
BˆØ[\WÙ^\ÈH[
Ø[\WÜÝ[[X\žVÈœØ[\WÙ^\È—HÜˆ
Bˆ[Ù^WØÛÝ[H[Š[Ù^\ÊBˆ\×Ø[žWÙ]HH›ÛÛ
™\ÜÚ[™\\œÈÜˆ[Ù^WØÛÝ[ÜˆØ[\WØÛÝ[
BˆYˆ›Ý\×Ø[žWÙ]N‚ˆ]WÜÝ]\ÈHœÙ[HYÜÈ‚ˆ[ÙN‚ˆZ[WØÛÝ[ÈHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕ]˜Z[Xš[]WÙ]KÓÕS•

ŠHTÈ[™\\—ØÛÝ[ˆÕSJÐTÑHÒSˆ˜[YÜÛÝÈˆS‘]˜Z[Xš[]WÜÝTÈ“Õ•SSˆHSÑHS‘
HTÈ˜[YÚ[™\\œÂˆ”“ÓH[™\\—Ø]˜Z[Xš[]WÙZ[BˆÒT‘H\ÜÙ]ÚYHÈS‘›ÝšY\ˆHÈS‘]˜Z[Xš[]WÙ]H‘UÑQSˆÈS‘ÂˆÔ“ÕT–H]˜Z[Xš[]WÙ]Bˆˆˆ‹ˆ
Ý\œ™[Ø\ÜÙ]ÚY›ÝšY\‹œ›ÛWÚ\ÛË×Ú\ÛÊKˆ
K™™]Ú[

Bˆ^XÝYÚ[™\\—ØÛÝ[HÛÛ™šYÝ\™YÚ[™\\—ØÛÝ[ÜˆX^
ˆ
[
›ÝÖÈš[™\\—ØÛÝ[—HÜˆ
H›Üˆ›ÝÈ[ˆZ[WØÛÝ[ÊKˆY˜][Lˆ
BˆÛÛ\]WÚ[™\\—Ù^\ÈH
ˆ[ŠZ[WØÛÝ[ÊHOH^XÝYÙ^\Âˆ[™^XÝYÚ[™\\—ØÛÝ[ˆˆ[™[
ˆ[
›ÝÖÈš[™\\—ØÛÝ[—HÜˆ
HOH^XÝYÚ[™\\—ØÛÝ[ˆ[™[
›ÝÖÈ˜[YÚ[™\\œÈ—HÜˆ
HOH^XÝYÚ[™\\—ØÛÝ[ˆ›Üˆ›ÝÈ[ˆZ[WØÛÝ[Âˆ
Bˆ
BˆÛÛ\]WÜ[Ù^\ÈH
ˆ[Ù^WØÛÝ[OH^XÝYÙ^\Âˆ[™[
ˆ[
›ÝÖÈ˜[YÜÛÝÈ—HÜˆ
Hˆˆ[™›ÝÖÈÙZYÚYØ]˜Z[Xš[]WÜÝ—H\È›Ý›Û™Bˆ[™[
›ÝÖÈš[™\\—ØÛÝ[—HÜˆ
HOH^XÝYÚ[™\\—ØÛÝ[ˆ›Üˆ›ÝÈ[ˆ[Ù^\Âˆ
Bˆ
Bˆ]WÜÝ]\ÈH
ˆ›ÚÈ‚ˆYˆÛÛ\]WÚ[™\\—Ù^\È[™ÛÛ\]WÜ[Ù^\È[™Ø[\WÙ^\ÈOH^XÝYÙ^\Âˆ[ÙHœ\˜ÚX[‚ˆ
B‚ˆ˜[šÙYÚ[™\\œÈHÜ›ÝÈ›Üˆ›ÝÈ[ˆ™\ÜÚ[™\\œÈYˆ›ÝÖÈ˜]˜Z[Xš[]WÜÝ—H\È›Ý›Û™WBˆ˜[šÙYÚ[™\\œËœÛÜ
ˆÙ^O[[X™H›ÝÎˆ
›Ø]
›ÝÖÈ˜]˜Z[Xš[]WÜÝ—JK›ÝÖÈš[™\\—Û˜[YH—K›ÝÙ\Š
K›ÝÖÈš[™\\—ÚY—JBˆ
BˆÛÜœÝÚ[™\\ˆH˜[šÙYÚ[™\\œÖÌHYˆ˜[šÙYÚ[™\\œÈ[ÙH›Û™BˆÙZYÚYÝØ]HØ[Ý[]WÝÙZYÚYÜ[Ø]˜Z[Xš[]J™\ÜÚ[™\\œÊBˆØ\›š[™ÜÈH™\Ü[™×Ú[˜[YÜÝÙ\—ÝØ\›š[™ÜÊˆ™\ÜÚ[™\\œËˆÝÙ\—ÚÙ^OHš[™\\—ÜÝÙ\—ÚÝÈ‹ˆØ\›š[™×ØÛÙOH›Z\ÜÚ[™×Ú[™\\—ÜÝÙ\ˆ‹ˆ
Bˆ[Ë˜\[™
ˆÂˆ˜\ÜÙ]ÚYŽˆÝ\œ™[Ø\ÜÙ]ÚYˆœ›Ú™XÝÛ˜[YHŽˆÝŠ\ÜÙ]Èœ›Ú™XÝÛ˜[YH—JKˆÙZYÚYÝØ]ÜÝŽˆÙZYÚYÝØ]ˆš[™\\—ØÛÝ[Žˆ[Š™\ÜÚ[™\\œÊHÜˆÛÛ™šYÝ\™YÚ[™\\—ØÛÝ[ˆš[™\\œ×Ø™[Ý×ÎLØÛÝ[ŽˆÝ[JˆBˆ›Üˆ›ÝÈ[ˆ™\ÜÚ[™\\œÂˆYˆ›ÝÖÈ˜]˜Z[Xš[]WÜÝ—H\È›Ý›Û™Bˆ[™›Ø]
›ÝÖÈ˜]˜Z[Xš[]WÜÝ—JHÕ×ÒS•‘T•T—ÐURSP’SUWÔÕˆ
KˆÛÜœÝÚ[™\\ˆŽˆÛÜœÝÚ[™\\–Èš[™\\—Û˜[YH—HYˆÛÜœÝÚ[™\\ˆ[ÙH›Û™KˆÛÜœÝÚ[™\\—ÚYŽˆÛÜœÝÚ[™\\–Èš[™\\—ÚY—HYˆÛÜœÝÚ[™\\ˆ[ÙH›Û™KˆÛÜœÝÚ[™\\—ÝØ]ÜÝŽˆÛÜœÝÚ[™\\–È˜]˜Z[Xš[]WÜÝ—HYˆÛÜœÝÚ[™\\ˆ[ÙH›Û™Kˆ˜[YÜÛÝÈŽˆÝ[J[
›ÝÖÈ˜[YÜÛÝÈ—JH›Üˆ›ÝÈ[ˆ™\ÜÚ[™\\œÊKˆ[˜]˜Z[X›WÜÛÝÈŽˆÝ[J[
›ÝÖÈ[˜]˜Z[X›WÜÛÝÈ—JH›Üˆ›ÝÈ[ˆ™\ÜÚ[™\\œÊKˆ™]WÜÝ]\ÈŽˆ]WÜÝ]\ËˆØ\›š[™ÜÈŽˆØ\›š[™ÜËˆBˆ
B‚ˆ™]\›ˆÂˆ™œ›ÛWÙ]HŽˆœ›ÛWÚ\ÛËˆ×Ù]HŽˆ×Ú\ÛËˆœ[ÈŽˆ[ËˆB‚‚™YˆÙ]ÙZ[WÝØ]Ü™\ÜÙ]JÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹\™Ù]Ù]Nˆ]JHOˆXÝÜÝ‹[žWN‚ˆ™\ÜHÙ]Û[ÛWÝØ]Ü™\ÜÙ]JÛÛ›‹\™Ù]Ù]K\™Ù]Ù]JBˆ™]\›ˆÂˆ\™Ù]Ù]HŽˆ\™Ù]Ù]Kš\ÛÙ›Ü›X]

Kˆœ[ÈŽˆ™\ÜÈœ[È—KˆB‚‚™YˆÙ]Ú[™\\—Ø]˜Z[Xš[]WØÚ\Ü™\Ü
ˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ\ÜÙ]ÚYˆ[ˆœ›ÛWÙ]Nˆ]Kˆ×Ù]Nˆ]KŠHOˆXÝÜÝ‹[žWH›Û™N‚ˆ\ÜÙ]HÛÛ›‹™^XÝ]Jˆ”ÑSPÕY›Ú™XÝÛ˜[YH”“ÓH\ÜÙ]ÈÒT‘HYHÈ‹ˆ
\ÜÙ]ÚY
Kˆ
K™™]ÚÛ™J
BˆYˆ\ÜÙ]\È›Û™N‚ˆ™]\›ˆ›Û™Bˆ]šXÙWÜ›ÝÜÈHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕˆ^\›˜[Ù]šXÙWÚYTÈ[™\\—ÚYˆ]šXÙWÛ˜[YHTÈ[™\\—Û˜[YKˆ˜]YÜÝÙ\—ÚÝÈTÈ[™\\—ÜÝÙ\—ÚÝËˆ[Ù[ˆ”“ÓH›ÝšY\—Ù]šXÙ\ÂˆÒT‘H\ÜÙ]ÚYHÈS‘›ÝšY\ˆHÈS‘[˜X›YHHS‘]—Ý\WÚYSˆ
KÎ
BˆÔ‘Tˆ–H]šXÙWÛ˜[YHÓÓUH“ÐÐTÑK^\›˜[Ù]šXÙWÚYˆˆˆ‹ˆ
\ÜÙ]ÚYS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓTŠKˆ
K™™]Ú[

BˆZ[WÜ›ÝÜÈHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕ[™\\—ÚY[™\\—Û˜[YK[™\\—ÜÝÙ\—ÚÝË]˜Z[Xš[]WÜÝˆ”“ÓH[™\\—Ø]˜Z[Xš[]WÙZ[BˆÒT‘H\ÜÙ]ÚYHÈS‘›ÝšY\ˆHÈS‘]˜Z[Xš[]WÙ]H‘UÑQSˆÈS‘ÂˆÔ‘Tˆ–H]˜Z[Xš[]WÙ]Bˆˆˆ‹ˆ
\ÜÙ]ÚYS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹œ›ÛWÙ]Kš\ÛÙ›Ü›X]

K×Ù]Kš\ÛÙ›Ü›X]

JKˆ
K™™]Ú[

Bˆ[™\\œ×ØžWÚYˆXÝÜÝ‹XÝÜÝ‹[žWWHHßBˆ›Üˆ›ÝÈ[ˆ]šXÙWÜ›ÝÜÎ‚ˆ[™\\—ÚYHÝŠ›ÝÖÈš[™\\—ÚY—JBˆYˆ\×Ü™[[Ý™YÚ[™\\—Û˜[YJ›ÝÖÈš[™\\—Û˜[YH—JN‚ˆÛÛ[YBˆ[™\\œ×ØžWÚYÚ[™\\—ÚYHHÂˆš[™\\—ÚYŽˆ[™\\—ÚYˆš[™\\—Û˜[YHŽˆ›ÝÖÈš[™\\—Û˜[YH—Kˆš[™\\—ÜÝÙ\—ÚÝÈŽˆ\œÙWÙ›Ø]Ý˜[YJ›ÝÖÈš[™\\—ÜÝÙ\—ÚÝÈ—JBˆÜˆ[™™\—Ú[™\\—ÜÝÙ\—Ùœ›ÛWÛ[Ù[
›ÝÖÈ›[Ù[—JKˆ˜]˜Z[Xš[]WÝ˜[Y\ÈŽˆ×KˆBˆ›Üˆ›ÝÈ[ˆZ[WÜ›ÝÜÎ‚ˆ[™\\—ÚYHÝŠ›ÝÖÈš[™\\—ÚY—JBˆYˆ\×Ü™[[Ý™YÚ[™\\—Û˜[YJ›ÝÖÈš[™\\—Û˜[YH—JN‚ˆÛÛ[YBˆ[™\\ˆH[™\\œ×ØžWÚYœÙ]Y˜][
ˆ[™\\—ÚYˆÂˆš[™\\—ÚYŽˆ[™\\—ÚYˆš[™\\—Û˜[YHŽˆ›ÝÖÈš[™\\—Û˜[YH—Kˆš[™\\—ÜÝÙ\—ÚÝÈŽˆ›ÝÖÈš[™\\—ÜÝÙ\—ÚÝÈ—Kˆ˜]˜Z[Xš[]WÝ˜[Y\ÈŽˆ×KˆKˆ
BˆYˆ›ÝÖÈ˜]˜Z[Xš[]WÜÝ—H\È›Ý›Û™N‚ˆ[™\\–È˜]˜Z[Xš[]WÝ˜[Y\È—K˜\[™
›Ø]
›ÝÖÈ˜]˜Z[Xš[]WÜÝ—JJB‚ˆ\š[ÙÜÝ\H]][YK˜ÛÛXš[™Jœ›ÛWÙ]K]][YK›Z[‹[YJ
JBˆ\š[ÙÙ[™H]][YK˜ÛÛXš[™J×Ù]H
È[YY[J^\ÏLJK]][YK›Z[‹[YJ
JBˆØ[\WÜ›ÝÜÈHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕ[™\\—ÚYØ[\WÝ[YKXÝ]™WÜÝÙ\—ÚÝÂˆ”“ÓH[™\\—ÜÝÙ\—ÜØ[\\ÂˆÒT‘H\ÜÙ]ÚYHÈS‘›ÝšY\ˆHÈS‘Ø[\WÝ[YHHÈS‘Ø[\WÝ[YHÂˆÔ‘Tˆ–H[™\\—ÚYØ[\WÝ[YBˆˆˆ‹ˆ
ˆ\ÜÙ]ÚYˆS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹ˆ\š[ÙÜÝ\š\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆ\š[ÙÙ[™š\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆ
Kˆ
K™™]Ú[

BˆØ[\\×ØžWÚ[™\\ŽˆXÝÜÝ‹\ÝÝ\VÙ]][YK›Ø]WWHHßBˆ›Üˆ›ÝÈ[ˆØ[\WÜ›ÝÜÎ‚ˆØ[\WÝ[YHH\œÙWÙ]][YWÝ˜[YJ›ÝÖÈœØ[\WÝ[YH—JBˆXÝ]™WÜÝÙ\ˆH\œÙWÙ›Ø]Ý˜[YJ›ÝÖÈ˜XÝ]™WÜÝÙ\—ÚÝÈ—JBˆYˆØ[\WÝ[YH\È›Û™HÜˆXÝ]™WÜÝÙ\ˆ\È›Û™N‚ˆÛÛ[YBˆØ[\\×ØžWÚ[™\\‹œÙ]Y˜][
ÝŠ›ÝÖÈš[™\\—ÚY—JK×JK˜\[™

Ø[\WÝ[YKX^
XÝ]™WÜÝÙ\‹Œ
JJB‚ˆ[™\\œÈH\Ý
[™\\œ×ØžWÚY˜[Y\Ê
JBˆ›Üˆ[™\\ˆ[ˆ[™\\œÎ‚ˆ\˜Ù[YÙ\ÈH[™\\‹œÜ
˜]˜Z[Xš[]WÝ˜[Y\ÈŠBˆ[™\\–È˜]™\˜YÙWÜÝ—HH›Ý[™
Ý[J\˜Ù[YÙ\ÊHÈ[Š\˜Ù[YÙ\ÊKŠHYˆ\˜Ù[YÙ\È[ÙH›Û™Bˆ[™\\–È˜Ú\—HHZ[Ú[™\\—ÜÝÙ\—ØÚ\
ˆØ[\\×ØžWÚ[™\\‹™Ù]
[™\\–Èš[™\\—ÚY—K×JKˆ\š[ÙÜÝ\ˆ\š[ÙÙ[™ˆ\œÙWÙ›Ø]Ý˜[YJ[™\\–Èš[™\\—ÜÝÙ\—ÚÝÈ—JKˆ
Bˆ[™\\œËœÛÜ
Ù^O[[X™H›ÝÎˆ
›ÝÖÈ˜]™\˜YÙWÜÝ—H\È›Û™K›ÝÖÈ˜]™\˜YÙWÜÝ—HÜˆ›ÝÖÈš[™\\—Û˜[YH—HÜˆ›ÝÖÈš[™\\—ÚY—JJBˆ™]\›ˆÂˆ˜\ÜÙ]ÚYŽˆ[
\ÜÙ]ÈšY—JKˆœ›Ú™XÝÛ˜[YHŽˆ\ÜÙ]Èœ›Ú™XÝÛ˜[YH—Kˆš[™\\œÈŽˆ[™\\œËˆB‚‚™YˆZ[Ú[™\\—ÜÝÙ\—ØÚ\
ˆØ[\\Îˆ\ÝÝ\VÙ]][YK›Ø]WKˆ\š[ÙÜÝ\ˆ]][YKˆ\š[ÙÙ[™ˆ]][YKˆ˜]YÜÝÙ\—ÚÝÎˆ›Ø]›Û™KŠHOˆXÝÜÝ‹[žWN‚ˆÚ\ÝÚYHÍŒŒˆÚ\ÚZYÚHŒŒˆÝÛYHL‹ŒˆÝÜšYÚHM‹ŒˆÝÝÜHM‹ŒˆÝØ›ÝÛHHÍ‹ŒˆÝÝÚYHÚ\ÝÚYHÝÛYHÝÜšYÚˆÝÚZYÚHÚ\ÚZYÚHÝÝÜHÝØ›ÝÛBˆÝ[ÜÙXÛÛ™ÈHX^

\š[ÙÙ[™H\š[ÙÜÝ\
KÝ[ÜÙXÛÛ™Ê
KKŒ
BˆÛÜYÜØ[\\ÈHÛÜY
Ø[\\ËÙ^O[[X™H][Nˆ][VÌJBˆYˆ[ŠÛÜYÜØ[\\ÊHˆŒ‚ˆÝ\HX^
[ŠÛÜYÜØ[\\ÊHËÈŒJBˆ™YXÙYHÛÜYÜØ[\\ÖÎŽœÝ\BˆYˆ™YXÙYËLWHOHÛÜYÜØ[\\ÖËLWN‚ˆ™YXÙY˜\[™
ÛÜYÜØ[\\ÖËLWJBˆÛÜYÜØ[\\ÈH™YXÙYˆØœÙ\™YÛX^HX^

ÝÙ\ˆ›ÜˆËÝÙ\ˆ[ˆÛÜYÜØ[\\ÊKY˜][LŒ
BˆWÛX^HX^
ØœÙ\™YÛX^˜]YÜÝÙ\—ÚÝÈÜˆŒKŒ
BˆYˆWÛX^ˆL‚ˆWÛX^H›Ø]

[
WÛX^
ÈŽNNJHËÈJH
ˆJBˆ[ÙN‚ˆWÛX^H›Ø]
[
WÛX^
ÈŽNNJJBˆÛÛÜ™[˜]\Îˆ\ÝÝ\VÙ›Ø]›Ø]WHH×Bˆ›ÜˆØ[\WÝ[YKÝÙ\ˆ[ˆÛÜYÜØ[\\Î‚ˆ[\ÙYHZ[ŠX^

Ø[\WÝ[YHH\š[ÙÜÝ\
KÝ[ÜÙXÛÛ™Ê
KŒ
KÝ[ÜÙXÛÛ™ÊBˆHÝÛY
È[\ÙYÈÝ[ÜÙXÛÛ™È
ˆÝÝÚYˆHHÝÝÜ
È
HHZ[ŠÝÙ\ˆÈWÛX^KŒ
JH
ˆÝÚZYÚˆÛÛÜ™[˜]\Ë˜\[™

›Ý[™
ŠK›Ý[™
KŠJJBˆ[™WÜ]Hˆ‹š›Ú[Šˆ
“HˆYˆ[™^OH[ÙH“ŠH
ÈˆˆÞHÞ_H‚ˆ›Üˆ[™^
JH[ˆ[[Y\˜]JÛÛÜ™[˜]\ÊBˆ
Bˆ˜\Ù[[™HHÝÝÜ
ÈÝÚZYÚˆ\™XWÜ]Hˆ‚ˆYˆÛÛÜ™[˜]\Î‚ˆ\™XWÜ]H
ˆˆ“HØÛÛÜ™[˜]\ÖÌVÌ_HØ˜\Ù[[™_H‚ˆ
Èˆ‹š›Ú[Šˆ“ÞHÞ_Hˆ›ÜˆH[ˆÛÛÜ™[˜]\ÊBˆ
ÈˆˆØÛÛÜ™[˜]\ÖËLWVÌ_HØ˜\Ù[[™_Hˆ‚ˆ
BˆÝXÚÜÈH×Bˆ›Üˆ[™^[ˆ˜[™ÙJJN‚ˆ˜][ÈH[™^ÈˆXÚ×Ý[YHH\š[ÙÜÝ\
È[YY[JÙXÛÛ™Ï]Ý[ÜÙXÛÛ™È
ˆ˜][ÊBˆÝXÚÜË˜\[™
ˆÂˆžŽˆ›Ý[™
ÝÛY
ÈÝÝÚY
ˆ˜][ËŠKˆ›X™[ŽˆXÚ×Ý[YKœÝ™[YJ‰R‰SHŠHYˆ
\š[ÙÙ[™H\š[ÙÜÝ\
K™^\ÈHH[ÙHXÚ×Ý[YKœÝ™[YJ‰YÉ[HŠKˆBˆ
BˆWÝXÚÜÈHÂˆÂˆžHŽˆ›Ý[™
ÝÝÜ
ÈÝÚZYÚ
ˆ[™^ÈŠKˆ›X™[Žˆ›Ý[™
WÛX^
ˆ
HH[™^È
KJKˆBˆ›Üˆ[™^[ˆ˜[™ÙJJBˆBˆ™]\›ˆÂˆÚYŽˆ[
Ú\ÝÚY
KˆšZYÚŽˆ[
Ú\ÚZYÚ
KˆœÝÛYŽˆÝÛYˆœÝÜšYÚŽˆÚ\ÝÚYHÝÜšYÚˆœÝÝÜŽˆÝÝÜˆœÝØ›ÝÛHŽˆ˜\Ù[[™Kˆ›[™WÜ]Žˆ[™WÜ]ˆ˜\™XWÜ]Žˆ\™XWÜ]ˆžÝXÚÜÈŽˆÝXÚÜËˆžWÝXÚÜÈŽˆWÝXÚÜËˆœØ[\WØÛÝ[Žˆ[ŠØ[\\ÊKˆ›X^ÜÝÙ\—ÚÝÈŽˆ›Ý[™
WÛX^JKˆB‚‚™Yˆ\œÙWÙ]][YWÝ˜[YJ˜[YNˆ[žJHOˆ]][YH›Û™N‚ˆYˆ˜[YH[ˆ
›Û™KˆŠN‚ˆ™]\›ˆ›Û™BˆYˆ\Ú[œÝ[˜ÙJ˜[YK
[›Ø]
JN‚ˆ[Y\Ý[\H›Ø]
˜[YJBˆYˆ[Y\Ý[\ˆLÌÌÌ‚ˆ[Y\Ý[\ÏHLˆžN‚ˆ™]\›ˆ]][YK™œ›Û][Y\Ý[\
[Y\Ý[\
Bˆ^Ù\
ÔÑ\œ›Ü‹Ý™\™›ÝÑ\œ›Ü‹˜[YQ\œ›ÜŠN‚ˆ™]\›ˆ›Û™Bˆ˜]ÈHÝŠ˜[YJKœÝš\

BˆžN‚ˆ[Y\Ý[\H›Ø]
˜]ÊBˆYˆ[Y\Ý[\ˆLÌÌÌ‚ˆ[Y\Ý[\ÏHLˆ™]\›ˆ]][YK™œ›Û][Y\Ý[\
[Y\Ý[\
Bˆ^Ù\
ÔÑ\œ›Ü‹Ý™\™›ÝÑ\œ›Ü‹˜[YQ\œ›ÜŠN‚ˆ\ÜÂˆ˜]ÈH˜]Ëœ™\XÙJ–ˆ‹ŠÌŒŠBˆžN‚ˆ\œÙYH]][YK™œ›ÛZ\ÛÙ›Ü›X]
˜]ÊBˆ™]\›ˆ\œÙYœ™\XÙJš[™›ÏS›Û™JHYˆ\œÙYš[™›È[ÙH\œÙYˆ^Ù\˜[YQ\œ›ÜŽ‚ˆ™]\›ˆ›Û™B‚‚™YˆØ[Ý[]WØ\ÜÙ]Ø]˜Z[Xš[]J]šXÙWÜ›ÝÜÎˆ\ÝÙXÝÜÝ‹[žWWJHOˆXÝÜÝ‹[žWN‚ˆ[˜X›YÜ›ÝÜÈHÜ›ÝÈ›Üˆ›ÝÈ[ˆ]šXÙWÜ›ÝÜÈYˆ[
›ÝË™Ù]
™[˜X›Y‹JHÜˆ
HOHWBˆÝ[H[Š[˜X›YÜ›ÝÜÊBˆ]˜Z[X›HHÝ[JH›Üˆ›ÝÈ[ˆ[˜X›YÜ›ÝÜÈYˆ›ÝË™Ù]
˜]˜Z[Xš[]WÜÝ]\ÈŠHOH˜]˜Z[X›HŠBˆ[˜]˜Z[X›HHÝ[JˆH›Üˆ›ÝÈ[ˆ[˜X›YÜ›ÝÜÈYˆ›ÝË™Ù]
˜]˜Z[Xš[]WÜÝ]\ÈŠH[ˆÈ[˜]˜Z[X›H‹››×ØÛÛ[][šXØ][ÛˆŸBˆ
Bˆ›×ØÛÛ[][šXØ][ÛˆHÝ[JH›Üˆ›ÝÈ[ˆ[˜X›YÜ›ÝÜÈYˆ›ÝË™Ù]
˜]˜Z[Xš[]WÜÝ]\ÈŠHOH››×ØÛÛ[][šXØ][ÛˆŠBˆ™XÙ[HÝ[JH›Üˆ›ÝÈ[ˆ[˜X›YÜ›ÝÜÈYˆ›ÝË™Ù]
˜ÛÛ[][šXØ][Û—ÜÝ]\ÈŠHOHœ™XÙ[ŠBˆ˜]YÝ˜[Y\ÈHÜ›ÝË™Ù]
œ˜]YÜÝÙ\—ÚÝÈŠH›Üˆ›ÝÈ[ˆ[˜X›YÜ›ÝÜ×BˆÛ›ÝÛ—Ü˜]YHÙ›Ø]
˜[YJH›Üˆ˜[YH[ˆ˜]YÝ˜[Y\ÈYˆ˜[YH\È›Ý›Û™WBˆY™™XÝYÜ›ÝÜÈHÂˆ›ÝÈ›Üˆ›ÝÈ[ˆ[˜X›YÜ›ÝÜÈYˆ›ÝË™Ù]
˜]˜Z[Xš[]WÜÝ]\ÈŠH[ˆÈ[˜]˜Z[X›H‹››×ØÛÛ[][šXØ][ÛˆŸBˆBˆY™™XÝYÜÝÙ\—ÚÝÈH
ˆÝ[J›Ø]
›ÝÖÈœ˜]YÜÝÙ\—ÚÝÈ—JH›Üˆ›ÝÈ[ˆY™™XÝYÜ›ÝÜÊBˆYˆY™™XÝYÜ›ÝÜÈ[™[
›ÝË™Ù]
œ˜]YÜÝÙ\—ÚÝÈŠH\È›Ý›Û™H›Üˆ›ÝÈ[ˆY™™XÝYÜ›ÝÜÊBˆ[ÙH›Û™Bˆ
BˆØ\XÚ]WØ]˜Z[Xš[]WÜÝH›Û™BˆYˆÝ[[™[ŠÛ›ÝÛ—Ü˜]Y
HOHÝ[[™Ý[JÛ›ÝÛ—Ü˜]Y
Hˆ‚ˆØ\XÚ]WØ]˜Z[Xš[]WÜÝH›Ý[™
ˆÝ[J›Ø]
›ÝÖÈœ˜]YÜÝÙ\—ÚÝÈ—JH›Üˆ›ÝÈ[ˆ[˜X›YÜ›ÝÜÈYˆ›ÝË™Ù]
˜]˜Z[Xš[]WÜÝ]\ÈŠHOH˜]˜Z[X›HŠBˆÈÝ[JÛ›ÝÛ—Ü˜]Y
Bˆ
ˆLˆ‹ˆ
Bˆ™]\›ˆÂˆš[™\\—Ø]˜Z[Xš[]WÜÝŽˆ›Ý[™
]˜Z[X›HÈÝ[
ˆLŠHYˆÝ[[ÙH›Û™Kˆ˜Ø\XÚ]WØ]˜Z[Xš[]WÜÝŽˆØ\XÚ]WØ]˜Z[Xš[]WÜÝˆ˜ÛÛ[][šXØ][Û—Ø]˜Z[Xš[]WÜÝŽˆ›Ý[™
™XÙ[ÈÝ[
ˆLŠHYˆÝ[[ÙH›Û™Kˆ˜]˜Z[X›WÚ[™\\œÈŽˆ]˜Z[X›KˆÝ[Ú[™\\œÈŽˆÝ[ˆ[˜]˜Z[X›WÚ[™\\œÈŽˆ[˜]˜Z[X›Kˆ››×ØÛÛ[][šXØ][Û—Ù]šXÙ\ÈŽˆ›×ØÛÛ[][šXØ][Û‹ˆ˜Y™™XÝYÜÝÙ\—ÚÝÈŽˆY™™XÝYÜÝÙ\—ÚÝËˆ˜]˜Z[X›WÜÝš[™ÜÈŽˆÝ[J[
›ÝË™Ù]
˜]˜Z[X›WÜÝš[™ÜÈŠHÜˆ
H›Üˆ›ÝÈ[ˆ[˜X›YÜ›ÝÜÊKˆÝ[ÜÝš[™ÜÈŽˆÝ[J[
›ÝË™Ù]
Ý[ÜÝš[™ÜÈŠHÜˆ
H›Üˆ›ÝÈ[ˆ[˜X›YÜ›ÝÜÊKˆ[˜]˜Z[X›WÜÝš[™ÜÈŽˆÝ[J[
›ÝË™Ù]
[˜]˜Z[X›WÜÝš[™ÜÈŠHÜˆ
H›Üˆ›ÝÈ[ˆ[˜X›YÜ›ÝÜÊKˆœÝš[™×Ø]˜Z[Xš[]WÜÝŽˆ
ˆ›Ý[™
ˆÝ[J[
›ÝË™Ù]
˜]˜Z[X›WÜÝš[™ÜÈŠHÜˆ
H›Üˆ›ÝÈ[ˆ[˜X›YÜ›ÝÜÊBˆÈÝ[J[
›ÝË™Ù]
Ý[ÜÝš[™ÜÈŠHÜˆ
H›Üˆ›ÝÈ[ˆ[˜X›YÜ›ÝÜÊBˆ
ˆLˆ‹ˆ
BˆYˆÝ[J[
›ÝË™Ù]
Ý[ÜÝš[™ÜÈŠHÜˆ
H›Üˆ›ÝÈ[ˆ[˜X›YÜ›ÝÜÊBˆ[ÙH›Û™Bˆ
KˆB‚‚™Yˆ›Ü›X]Ù\Ú[ÛœÛÛ\—Ø[\›WÝ[YJ˜[YNˆ[žJHOˆÝŽ‚ˆYˆ˜[YH[ˆ
›Û™KˆŠN‚ˆ™]\›ˆˆ‚ˆ˜]ÈHÝŠ˜[YJKœÝš\

BˆYˆ˜]Ëš\ÙYÚ]

N‚ˆ[Y\Ý[\H[
˜]ÊBˆYˆ[Y\Ý[\ˆLÌÌÌ‚ˆ[Y\Ý[\H[
[Y\Ý[\ÈL
BˆžN‚ˆ™]\›ˆ]][YK™œ›Û][Y\Ý[\
[Y\Ý[\
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠBˆ^Ù\
˜[YQ\œ›Ü‹ÔÑ\œ›ÜŠN‚ˆ™]\›ˆ˜]Âˆ™]\›ˆ˜]Â‚‚™Yˆ›Ü›X[^™WÙ\Ú[ÛœÛÛ\—Ø[\›J›ÝÎˆXÝÜÝ‹[žWJHOˆXÝÜÝ‹Ý—N‚ˆ[\›WÛ˜[YHHš\œÝÛ›Û—Ù[\Jˆ›ÝËˆÂˆ˜[\›S˜[YH‹ˆ˜[\›WÛ˜[YH‹ˆ›˜[YH‹ˆ˜[\›U\H‹ˆ˜[\›U\S˜[YH‹ˆ™˜][˜[YH‹ˆ™]™[˜[YH‹ˆ˜Ø]\ÙH‹ˆKˆ
Bˆ]šXÙWÛ˜[YHHš\œÝÛ›Û—Ù[\Jˆ›ÝËˆÂˆ™]“˜[YH‹ˆ™]šXÙS˜[YH‹ˆ™]šXÙWÛ˜[YH‹ˆ™\]Z\Y[˜[YH‹ˆš[™\\“˜[YH‹ˆ™][X\È‹ˆ™]•\S˜[YH‹ˆ™]‘ˆ‹ˆ™]šXÙQˆ‹ˆKˆ
Bˆ™]\›ˆÂˆ˜[\›WÛ˜[YHŽˆ[\›WÛ˜[YHÜˆš\œÝÛ›Û—Ù[\J›ÝËÈ˜[\›RY‹˜[\›WÚY‹šY—JHÜˆ[\›YH]]›È‹ˆ™]šXÙWÛ˜[YHŽˆ]šXÙWÛ˜[YHÜˆ\\™[È˜[ÈY[YšXØYÈ‹ˆœÙ]™\š]HŽˆš\œÝÛ›Û—Ù[\J›ÝËÈ›]ˆ‹›]™[‹œÙ]™\š]H‹˜[\›S]™[—JKˆœ˜Z\ÙYØ]Žˆ›Ü›X]Ù\Ú[ÛœÛÛ\—Ø[\›WÝ[YJš\œÝÛ›Û—Ù[\J›ÝËÈœ˜Z\ÙU[YH‹œÝ\[YH‹›ØØÝ\•[YH‹š\[•[YH—JJKˆœÝ]\ÈŽˆš\œÝÛ›Û—Ù[\J›ÝËÈœÝ]\È‹˜[\›TÝ]\È‹œÝ]H—JKˆB‚‚™YˆÝ[[X\š^™WÙ\Ú[ÛœÛÛ\—Ø[\›\Ê[\›\Îˆ\ÝÙXÝÜÝ‹[žWWK[Z]ˆ[HÊHOˆXÝÜÝ‹[žWN‚ˆ›Ü›X[^™YHÛ›Ü›X[^™WÙ\Ú[ÛœÛÛ\—Ø[\›J[\›JH›Üˆ[\›H[ˆ[\›\ÈYˆ\Ú[œÝ[˜ÙJ[\›KXÝ
WBˆš[X\žHH›Ü›X[^™YÌHYˆ›Ü›X[^™Y[ÙHßBˆÝ[[X\žWÜ\ÈH×Bˆ›Üˆ[\›H[ˆ›Ü›X[^™YÎ›[Z]N‚ˆX™[H[\›VÈ˜[\›WÛ˜[YH—BˆYˆ[\›VÈ™]šXÙWÛ˜[YH—N‚ˆX™[HˆžÛX™[HØ[\›VÉÙ]šXÙWÛ˜[YI×_H‚ˆYˆ[\›VÈœÙ]™\š]H—N‚ˆX™[HˆžÛX™[H
Ù]‹ˆØ[\›VÉÜÙ]™\š]I×_JH‚ˆÝ[[X\žWÜ\Ë˜\[™
X™[
BˆYˆ[Š›Ü›X[^™Y
Hˆ[Z]‚ˆÝ[[X\žWÜ\Ë˜\[™
ˆŠÞÛ[Š›Ü›X[^™Y
HH[Z]H[\›Y\ÈŠBˆ™]\›ˆÂˆœš[X\žWØ[\›WÛ˜[YHŽˆš[X\žK™Ù]
˜[\›WÛ˜[YH‹ˆŠKˆœš[X\žWØ[\›WÙ]šXÙHŽˆš[X\žK™Ù]
™]šXÙWÛ˜[YH‹ˆŠKˆœš[X\žWØ[\›WÜÙ]™\š]HŽˆš[X\žK™Ù]
œÙ]™\š]H‹ˆŠKˆœš[X\žWØ[\›WÜ˜Z\ÙYØ]Žˆš[X\žK™Ù]
œ˜Z\ÙYØ]‹ˆŠKˆ˜[\›WÜÝ[[X\žHŽˆŽÈ‹š›Ú[ŠÝ[[X\žWÜ\ÊKˆ››Ü›X[^™YØ[\›\ÈŽˆ›Ü›X[^™YˆB‚‚™Yˆ\Ú[ÛœÛÛ\—Ø[\›WÜÙ]™\š]WÜ˜[šÊ˜[YNˆ[žJHOˆ[›Û™N‚ˆ›Ü›X[^™YH›Ü›X[^™WÛ˜[YJÝŠ˜[YHÜˆˆŠJBˆYˆ›Ý›Ü›X[^™Y‚ˆ™]\›ˆ›Û™BˆYˆ›Ü›X[^™Y[ˆÈŒH‹ŒKŒ‹˜Üš]XØ[‹˜Üš]XØH‹˜Üš]XÛÈŸN‚ˆ™]\›ˆBˆYˆ›Ü›X[^™Y[ˆÈŒˆ‹Œ‹Œ‹›XZ›Üˆ‹˜[H‹›XZ[ÜˆŸN‚ˆ™]\›ˆ‚ˆYˆ›Ü›X[^™Y[ˆÈŒÈ‹ŒËŒ‹›Z[›Üˆ‹›Y[›Üˆ‹˜˜Z^HŸN‚ˆ™]\›ˆÂˆYˆ›Ü›X[^™Y[ˆÈ‹Œ‹Ø\›š[™È‹˜]š\ÛÈ‹˜[\HŸN‚ˆ™]\›ˆˆ™]\›ˆ›Û™B‚‚™Yˆ\š]™WÙ\Ú[ÛœÛÛ\—Û[Ûš]Üš[™×ÜÝ]\ÊX[Ü˜]Îˆ[žK[\›\Îˆ\ÝÙXÝÜÝ‹[žWWJHOˆÝŽ‚ˆÝ]\ÈHX\Ù\Ú[ÛœÛÛ\—ÜÝ]\ÊX[Ü˜]ÊBˆYˆÝ]\ÈOH‘\ØÛÛ™XÝYHŽ‚ˆ™]\›ˆÝ]\Â‚ˆ[\›WÜ˜[šÜÈHÂˆ˜[šÂˆ›Üˆ[\›H[ˆ[\›\ÂˆYˆ\Ú[œÝ[˜ÙJ[\›KXÝ
Bˆ›Üˆ˜[šÈ[ˆÙ\Ú[ÛœÛÛ\—Ø[\›WÜÙ]™\š]WÜ˜[šÊš\œÝÛ›Û—Ù[\J[\›KÈ›]ˆ‹›]™[‹œÙ]™\š]H‹˜[\›S]™[—JJWBˆYˆ˜[šÈ\È›Ý›Û™BˆBˆYˆ[žJ˜[šÈHˆ›Üˆ˜[šÈ[ˆ[\›WÜ˜[šÜÊN‚ˆ™]\›ˆ‘\œ›È‚ˆYˆ[\›WÜ˜[šÜÎ‚ˆ™]\›ˆ[\H‚ˆ™]\›ˆÝ]\Â‚‚™Yˆ›Ü›X[^™WÙ\Ú[ÛœÛÛ\—Ü[Ü›ÝÊˆÝ][Û—Ü›ÝÎˆXÝÜÝ‹[žWKˆ™X[[YWÜ›ÝÎˆXÝÜÝ‹[žWH›Û™HH›Û™Kˆ[\›\Îˆ\ÝÙXÝÜÝ‹[žWWH›Û™HH›Û™KŠHOˆXÝÜÝ‹[žWN‚ˆ^\›˜[ÚYHÝŠÝ][Û—Ü›ÝË™Ù]
œ[ÛÙHŠHÜˆÝ][Û—Ü›ÝË™Ù]
œÝ][ÛÛÙHŠHÜˆˆŠKœÝš\

Bˆ^\›˜[Û˜[YHHÝŠÝ][Û—Ü›ÝË™Ù]
œ[˜[YHŠHÜˆÝ][Û—Ü›ÝË™Ù]
œÝ][Û“˜[YHŠHÜˆˆŠKœÝš\

BˆYˆ›Ý^\›˜[Û˜[YN‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠH™\ÜÜÝH\Ú[Û”ÛÛ\ˆ˜[È›Ý^H›ÛYHHÙ[˜[[XH\È[š\ËˆŠB‚ˆ]WÚ][WÛX\H
™X[[YWÜ›ÝÈÜˆßJK™Ù]
™]R][SX\ŠHÜˆßBˆX[Ü˜]ÈH]WÚ][WÛX\™Ù]
œ™X[ÚX[ÜÝ]HŠBˆ˜]×ÜÝ]\ÈH\ØÜšX™WÙ\Ú[ÛœÛÛ\—ÚX[ÜÝ]JX[Ü˜]ÊB‚ˆXÝ]™WØ[\›\ÈH[\›\ÈÜˆ×BˆÝ]\ÈH\š]™WÙ\Ú[ÛœÛÛ\—Û[Ûš]Üš[™×ÜÝ]\ÊX[Ü˜]ËXÝ]™WØ[\›\ÊBˆ[\›WÜÝ[[X\žHHÝ[[X\š^™WÙ\Ú[ÛœÛÛ\—Ø[\›\ÊXÝ]™WØ[\›\ÊBˆ[\›WÛ]™[ÈHÛÜY
ÜÝŠ][K™Ù]
›]ˆŠJH›Üˆ][H[ˆXÝ]™WØ[\›\ÈYˆ][K™Ù]
›]ˆŠH\È›Ý›Û™_JBˆ›Ý\×Ü\ÈHÙˆšX[ÜÝ]O^Ü˜]×ÜÝ]\ßH—BˆYˆXÝ]™WØ[\›\Î‚ˆ›Ý\×Ü\Ë˜\[™
ˆ˜XÝ]™WØ[\›\Ï^Û[ŠXÝ]™WØ[\›\Ê_HŠBˆYˆ[\›WÛ]™[Î‚ˆ›Ý\×Ü\Ë˜\[™
ˆ›]™[Ï^ÉË	Ëš›Ú[Š[\›WÛ]™[Ê_HŠBˆYˆ[\›WÜÝ[[X\žVÈ˜[\›WÜÝ[[X\žH—N‚ˆ›Ý\×Ü\Ë˜\[™
ˆ˜[\›WÙ]Z[Ï^Ø[\›WÜÝ[[X\žVÉØ[\›WÜÝ[[X\žI×_HŠB‚ˆ™]\›ˆÂˆ™^\›˜[ÚYŽˆ^\›˜[ÚYˆ™^\›˜[Û˜[YHŽˆ^\›˜[Û˜[YKˆœÝ]\ÈŽˆÝ]\Ëˆœ˜]×ÜÝ]\ÈŽˆ˜]×ÜÝ]\ËˆšX[ÜÝ]HŽˆ˜]×ÜÝ]\Ëˆ˜[\›WØÛÝ[Žˆ[ŠXÝ]™WØ[\›\ÊKˆ˜[\›WÛ]™[ÈŽˆ‹‹š›Ú[Š[\›WÛ]™[ÊKˆœš[X\žWØ[\›WÛ˜[YHŽˆ[\›WÜÝ[[X\žVÈœš[X\žWØ[\›WÛ˜[YH—Kˆœš[X\žWØ[\›WÙ]šXÙHŽˆ[\›WÜÝ[[X\žVÈœš[X\žWØ[\›WÙ]šXÙH—Kˆœš[X\žWØ[\›WÜÙ]™\š]HŽˆ[\›WÜÝ[[X\žVÈœš[X\žWØ[\›WÜÙ]™\š]H—Kˆœš[X\žWØ[\›WÜ˜Z\ÙYØ]Žˆ[\›WÜÝ[[X\žVÈœš[X\žWØ[\›WÜ˜Z\ÙYØ]—Kˆ˜[\›WÜÝ[[X\žHŽˆ[\›WÜÝ[[X\žVÈ˜[\›WÜÝ[[X\žH—Kˆ››Ý\ÈŽˆŽÈ‹š›Ú[Š›Ý\×Ü\ÊKˆœ^[ØYŽˆÂˆœÝ][ÛˆŽˆÝ][Û—Ü›ÝËˆœ™X[[YHŽˆ™X[[YWÜ›ÝÈÜˆßKˆ˜[\›\ÈŽˆXÝ]™WØ[\›\Ëˆ››Ü›X[^™YØ[\›\ÈŽˆ[\›WÜÝ[[X\žVÈ››Ü›X[^™YØ[\›\È—KˆKˆB‚‚™Yˆš[™ÜÝYÙÙ\ÝYØ\ÜÙ]ÚY
ÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹^\›˜[Û˜[YNˆÝŠHOˆ[›Û™N‚ˆ›Ü›X[^™YÛ˜[YHH›Ü›X[^™WÛ˜[YJ^\›˜[Û˜[YJBˆ^XÝHš[™Ø\ÜÙ]ÚY
ÛÛ›‹^\›˜[Û˜[YJBˆYˆ^XÝ‚ˆ™]\›ˆ^XÝˆØ[™Y]HHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕYˆ”“ÓH\ÜÙ]ÂˆÒT‘H‘TPÑJÕÑTŠ›Ú™XÝÛ˜[YJK	È	Ë	ÉÊHRÑHÂˆÔ‘Tˆ–H›Ú™XÝÛ˜[YHÓÓUH“ÐÐTÑBˆSRUBˆˆˆ‹ˆ
ˆ‰^Û›Ü›X[^™YÛ˜[YKœ™\XÙJ	È	Ë	ÉÊ_IH‹
Kˆ
K™™]ÚÛ™J
Bˆ™]\›ˆ[
Ø[™Y]VÈšY—JHYˆØ[™Y]H[ÙH›Û™B‚‚™Yˆ\œÙWÜ›ÝšY\—Ü^[ØYÙ]J^[ØYˆXÝÜÝ‹[žWJHOˆ[žN‚ˆ]HH^[ØY™Ù]
™]HŠBˆYˆ\Ú[œÝ[˜ÙJ]KÝŠN‚ˆžN‚ˆ™]\›ˆœÛÛ‹›ØYÊ]JBˆ^Ù\œÛÛ‹’”ÓÓ‘XÛÙQ\œ›ÜŽ‚ˆ™]\›ˆ]Bˆ™]\›ˆ]B‚‚™YˆÙ]ÜÚYÙ[™\™ÞWÙ[™Ú[ØÛÛ™šYÊÛÛ™šYÎˆÜ[]LË”›ÝÈXÝÜÝ‹[žWJHOˆXÝÜÝ‹Ý—N‚ˆÛÛ™šY×ÛX\HXÝ
ÛÛ™šYÊBˆYØXÞWÙ[™\™ÞWÙ›ÝÈHÛÛ™šY×ÛX\™Ù]
™[™\™ÞWÙ›Ý×Ù[™Ú[ŠHÜˆÛÛ™šY×ÛX\™Ù]
˜[\›\×Ù[™Ú[ŠBˆ™]\›ˆÂˆ˜˜\ÙWÝ\›ŽˆÝŠÛÛ™šY×ÛX\™Ù]
˜˜\ÙWÝ\›ŠHÜˆQUSÔÒQÑS‘T‘ÖWÐTÑWÕT“
KœÝš\

HÜˆQUSÔÒQÑS‘T‘ÖWÐTÑWÕT“ˆ›ÙÚ[—Ù[™Ú[ŽˆÝŠÛÛ™šY×ÛX\™Ù]
›ÙÚ[—Ù[™Ú[ŠHÜˆQUSÔÒQÑS‘T‘ÖWÐUUÑS‘ÒS•
KœÝš\

HÜˆQUSÔÒQÑS‘T‘ÖWÐUUÑS‘ÒS•ˆœÞ\Ý[\×Ù[™Ú[ŽˆÝŠÛÛ™šY×ÛX\™Ù]
œ[×Ù[™Ú[ŠHÜˆQUSÔÒQÑS‘T‘ÖWÔÖTÕST×ÑS‘ÒS•
KœÝš\

HÜˆQUSÔÒQÑS‘T‘ÖWÔÖTÕST×ÑS‘ÒS•ˆ™[™\™ÞWÙ›Ý×Ù[™Ú[ŽˆÝŠYØXÞWÙ[™\™ÞWÙ›ÝÈÜˆQUSÔÒQÑS‘T‘ÖWÑS‘T‘ÖWÑ“Õ×ÑS‘ÒS•
KœÝš\

HÜˆQUSÔÒQÑS‘T‘ÖWÑS‘T‘ÖWÑ“Õ×ÑS‘ÒS•ˆœ™YÚ[ÛˆŽˆÝŠÛÛ™šY×ÛX\™Ù]
œ™YÚ[ÛˆŠHÜˆQUSÔÒQÑS‘T‘ÖWÔ‘QÒSÓŠKœÝš\

HÜˆQUSÔÒQÑS‘T‘ÖWÔ‘QÒSÓ‹ˆB‚‚™YˆÚYÙ[™\™ÞWØÛÛ™šYÝ\™YÜÞ\Ý[WÚYÊÛÛ™šYÎˆÜ[]LË”›ÝÈXÝÜÝ‹[žWJHOˆ\ÝÜÝ—N‚ˆ˜]×Ý˜[YHHÝŠXÝ
ÛÛ™šYÊK™Ù]
œÞ\Ý[WÚYÈŠHÜˆˆŠKœÝš\

Bˆ™]\›ˆÚ][KœÝš\

H›Üˆ][H[ˆ™KœÜ]
ˆ–Ë××JÈ‹˜]×Ý˜[YJHYˆ][KœÝš\

WB‚‚™YˆZ[ÜÚYÙ[™\™ÞWÜÙ\šXÙWØÛÛ™šYÊÛÛ™šYÎˆÜ[]LË”›ÝÈXÝÜÝ‹[žWJHOˆXÝÜÝ‹[žWN‚ˆÛÛ™šY×ÛX\HXÝ
ÛÛ™šYÊBˆ[™Ú[ÈHÙ]ÜÚYÙ[™\™ÞWÙ[™Ú[ØÛÛ™šYÊÛÛ™šY×ÛX\
Bˆ™]\›ˆÂˆ\Ù\›˜[YHŽˆÝŠÛÛ™šY×ÛX\™Ù]
\Ù\›˜[YHŠHÜˆˆŠKœÝš\

Kˆœ\ÜÝÛÜ™ŽˆÝŠÛÛ™šY×ÛX\™Ù]
œ\ÜÝÛÜ™ŠHÜˆˆŠKœÝš\

Kˆ˜˜\ÙWÝ\›Žˆ[™Ú[ÖÈ˜˜\ÙWÝ\›—Kˆ›ÙÚ[—Ù[™Ú[Žˆ[™Ú[ÖÈ›ÙÚ[—Ù[™Ú[—KˆœÞ\Ý[\×Ù[™Ú[Žˆ[™Ú[ÖÈœÞ\Ý[\×Ù[™Ú[—Kˆœ[×Ù[™Ú[Žˆ[™Ú[ÖÈœÞ\Ý[\×Ù[™Ú[—Kˆ™[™\™ÞWÙ›Ý×Ù[™Ú[Žˆ[™Ú[ÖÈ™[™\™ÞWÙ›Ý×Ù[™Ú[—Kˆ›Û˜›Ø\™Ù[™Ú[ŽˆÝŠÛÛ™šY×ÛX\™Ù]
›Û˜›Ø\™Ù[™Ú[ŠHÜˆQUSÔÒQÑS‘T‘ÖWÓÓ“ÐT‘ÑS‘ÒS•
KœÝš\

HÜˆQUSÔÒQÑS‘T‘ÖWÓÓ“ÐT‘ÑS‘ÒS•ˆœ™YÚ[ÛˆŽˆ[™Ú[ÖÈœ™YÚ[Ûˆ—KˆœÞ\Ý[WÚYÈŽˆÝŠÛÛ™šY×ÛX\™Ù]
œÞ\Ý[WÚYÈŠHÜˆˆŠKœÝš\

KˆB‚‚™Yˆ›Ü›X[^™WÜÚYÙ[™\™ÞWÜÞ\Ý[WÜ›ÝÜÊ]Nˆ[žJHOˆ\ÝÙXÝÜÝ‹[žWWN‚ˆYˆ\Ú[œÝ[˜ÙJ]K\Ý
N‚ˆ™]\›ˆÜ›ÝÈ›Üˆ›ÝÈ[ˆ]HYˆ\Ú[œÝ[˜ÙJ›ÝËXÝ
WBˆYˆ\Ú[œÝ[˜ÙJ]KXÝ
N‚ˆ›ÜˆÙ^H[ˆ
›\Ý‹œ™XÛÜ™È‹œÞ\Ý[\È‹š][\È‹œÞ\Ý[S\Ý‹œ›ÝÜÈŠN‚ˆ›ÝÜÈH]K™Ù]
Ù^JBˆYˆ\Ú[œÝ[˜ÙJ›ÝÜË\Ý
N‚ˆ™]\›ˆÜ›ÝÈ›Üˆ›ÝÈ[ˆ›ÝÜÈYˆ\Ú[œÝ[˜ÙJ›ÝËXÝ
WBˆYˆ[žJÙ^H[ˆ]H›ÜˆÙ^H[ˆ
œÞ\Ý[RY‹šY‹œÞ\Ý[S˜[YH‹›˜[YHŠJN‚ˆ™]\›ˆÙ]WBˆ™]\›ˆ×B‚‚™YˆX\ÜÚYÙ[™\™ÞWÜÝ]\Ê˜]×ÜÝ]\Îˆ[žK[™\™ÞWÙ›ÝÎˆXÝÜÝ‹[žWH›Û™HH›Û™JHOˆÝŽ‚ˆ[[™\™ÞWÙ›ÝÂˆ™]\›ˆÚYÙ[™\™ÞWÜÙ\šXÙK›X\ÜÚYÙ[™\™ÞWÜÝ]\Ê˜]×ÜÝ]\ÊB‚‚™Yˆ›Ü›X]ÜÚYÙ[™\™ÞWÚÝÊ˜[YNˆ[žJHOˆÝŽ‚ˆYˆ˜[YH[ˆ
›Û™KˆŠN‚ˆ™]\›ˆ“‹ÐH‚ˆžN‚ˆ™]\›ˆˆžÙ›Ø]
˜[YJN™ßHÕÈ‚ˆ^Ù\
\Q\œ›Ü‹˜[YQ\œ›ÜŠN‚ˆ™]\›ˆ“‹ÐH‚‚‚™Yˆ›Ü›X]ÜÚYÙ[™\™ÞWÜÝ
˜[YNˆ[žJHOˆÝŽ‚ˆYˆ˜[YH[ˆ
›Û™KˆŠN‚ˆ™]\›ˆ“‹ÐH‚ˆžN‚ˆ™]\›ˆˆžÙ›Ø]
˜[YJN™ßIH‚ˆ^Ù\
\Q\œ›Ü‹˜[YQ\œ›ÜŠN‚ˆ™]\›ˆ“‹ÐH‚‚‚™YˆZ[ÜÚYÙ[™\™ÞWÛ[Ûš]Üš[™×Û›Ý\Ê›ÝÎˆXÝÜÝ‹[žWJHOˆÝŽ‚ˆ˜]\žWØØ\XÚ]HH›ÝË™Ù]
˜˜]\žWØØ\XÚ]WÚÝÚŠBˆžN‚ˆ\×Ø˜]\žHH˜]\žWØØ\XÚ]H›Ý[ˆ
›Û™KˆŠH[™›Ø]
˜]\žWØØ\XÚ]HÜˆ
Hˆˆ^Ù\
\Q\œ›Ü‹˜[YQ\œ›ÜŠN‚ˆ\×Ø˜]\žHH˜[ÙBˆ˜]\žWÝ˜[YHH›Ü›X]ÜÚYÙ[™\™ÞWÚÝÊ›ÝË™Ù]
˜˜]\žWÜÝÙ\—ÚÝÈŠJHYˆ\×Ø˜]\žH[ÙH“‹ÐH‚ˆÛØ×Ý˜[YHH›Ü›X]ÜÚYÙ[™\™ÞWÜÝ
›ÝË™Ù]
˜˜]\žWÜÛØ×ÜÝŠJHYˆ\×Ø˜]\žH[ÙH“‹ÐH‚ˆ™]\›ˆˆ‹š›Ú[ŠˆÂˆˆ”ŽˆÙ›Ü›X]ÜÚYÙ[™\™ÞWÚÝÊ›ÝË™Ù]
	Ü—ÜÝÙ\—ÚÝÉÊJ_H‹ˆˆØ\™ØNˆÙ›Ü›X]ÜÚYÙ[™\™ÞWÚÝÊ›ÝË™Ù]
	ÛØYÜÝÙ\—ÚÝÉÊJ_H‹ˆˆ”™YNˆÙ›Ü›X]ÜÚYÙ[™\™ÞWÚÝÊ›ÝË™Ù]
	ÙÜšYÜÝÙ\—ÚÝ×Ü˜]ÉÊJ_H‹ˆˆ˜]\šXNˆØ˜]\žWÝ˜[Y_H‹ˆˆ”ÓÐÎˆÜÛØ×Ý˜[Y_H‹ˆBˆ
B‚‚”ÒQÑS‘T‘ÖWÔÖTÕSWÒQÔUT“ˆH™K˜ÛÛ\[Jˆ—–ÐKV˜K^ŒNWËW^ÌKIŠB”ÒQÑS‘T‘ÖWÐPÕU‘WÓÓ“ÐT‘S‘×ÔÕUTÑTÈHÈœ™\]Y\ÝY‹˜[™XYWÜ™\]Y\ÝY‹˜[™XYWÜ™\]Y\ÝYÛÜ—ÛÛ˜›Ø\™YŸB‚‚™Yˆ›Ü›X[^™WÜÚYÙ[™\™ÞWÜÞ\Ý[WÚYÙ›Ü—ØÛÛ\\™JÞ\Ý[WÚYˆÝŠHOˆÝŽ‚ˆ™]\›ˆÞ\Ý[WÚYœÝš\

K›ÝÙ\Š
B‚‚™Yˆ˜[Y]WÜÚYÙ[™\™ÞWÜÞ\Ý[WÚY
˜]×ÜÞ\Ý[WÚYˆÝŠHOˆÝŽ‚ˆÞ\Ý[WÚYH
˜]×ÜÞ\Ý[WÚYÜˆˆŠKœÝš\

BˆYˆ›ÝÞ\Ý[WÚY‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ”™Y[˜ÚHÈÞ\Ý[HQÚYÙ[™\™ÞKˆŠBˆYˆ‹ˆ[ˆÞ\Ý[WÚYÜˆŽÈˆ[ˆÞ\Ý[WÚYÜˆ[žJÚ\‹š\ÜÜXÙJ
H›ÜˆÚ\ˆ[ˆÞ\Ý[WÚY
N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ‘[šXH\[˜\È[HÞ\Ý[HQÜˆYYËˆŠBˆYˆ›ÝÒQÑS‘T‘ÖWÔÖTÕSWÒQÔUT“‹™[X]Ú
Þ\Ý[WÚY
N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ“ÈÞ\Ý[HQ]™H\ˆ]HØ\˜XÝ\™\ÈH\Ø\ˆ\[˜\È]˜\Ë[Y\›ÜËY™[ˆÝH[™\œØÛÜ™KˆŠBˆ™]\›ˆÞ\Ý[WÚY‚‚™Yˆ\Ù\ÜÚYÙ[™\™ÞWÛÛ˜›Ø\™[™×Ü™\]Y\Ý
ˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆÞ\Ý[WÚYˆÝ‹ˆ™\]Y\ÝYØžNˆÝ‹ˆ™\Ý[ˆXÝÜÝ‹[žWKŠHOˆ[‚ˆ›Ü›X[^™YH›Ü›X[^™WÜÚYÙ[™\™ÞWÜÞ\Ý[WÚYÙ›Ü—ØÛÛ\\™JÞ\Ý[WÚY
Bˆ^\Ý[™ÈHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕ
‚ˆ”“ÓHÚYÙ[™\™ÞWÛÛ˜›Ø\™[™×Ü™\]Y\ÝÂˆÒT‘HÕÑTŠÞ\Ý[WÚY
HHÈS‘Ý]\ÈSˆ
	Ü™\]Y\ÝY	Ë	Ø[™XYWÜ™\]Y\ÝY	Ë	Ø[™XYWÜ™\]Y\ÝYÛÜ—ÛÛ˜›Ø\™Y	ÊBˆÔ‘Tˆ–HYTÐÂˆSRUBˆˆˆ‹ˆ
›Ü›X[^™Y
Kˆ
K™™]ÚÛ™J
Bˆ›ÝÈH]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠBˆ™\ÜÛœÙWÚœÛÛˆHœÛÛ‹™[\ÊÚYÙ[™\™ÞWÜÙ\šXÙKœØ[š]^™WÜ^[ØY
™\Ý[™Ù]
œ™\ÜÛœÙHŠHÜˆ™\Ý[
K[œÝ\™WØ\ØÚZOUYJBˆYˆ^\Ý[™Î‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUHÚYÙ[™\™ÞWÛÛ˜›Ø\™[™×Ü™\]Y\ÝÂˆÑU][\ØÛÝ[H][\ØÛÝ[
ÈK›ÝšY\—ØÛÙHHË›ÝšY\—ÛY\ÜØYÙHHËˆ\ÝÙ\œ›ÜˆHË™\ÜÛœÙWÚœÛÛˆHË\]YØ]HÂˆÒT‘HYHÂˆˆˆ‹ˆ
ˆ™\Ý[™Ù]
œ›ÝšY\—ØÛÙH‹ˆŠKˆ™\Ý[™Ù]
›Y\ÜØYÙH‹ˆŠKˆˆˆYˆ™\Ý[™Ù]
œÝ]\ÈŠHOH™˜Z[Yˆ[ÙH™\Ý[™Ù]
›Y\ÜØYÙH‹ˆŠKˆ™\ÜÛœÙWÚœÛÛ‹ˆ›ÝËˆ^\Ý[™ÖÈšY—Kˆ
Kˆ
Bˆ™]\›ˆ[
^\Ý[™ÖÈšY—JBˆÝ\œÛÜˆHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆS”ÑT•S•ÈÚYÙ[™\™ÞWÛÛ˜›Ø\™[™×Ü™\]Y\ÝÈ
ˆÞ\Ý[WÚY™\]Y\ÝYØ]™\]Y\ÝYØžKÝ]\Ë›ÝšY\—ØÛÙK›ÝšY\—ÛY\ÜØYÙKˆ\ÝØÚXÚÙYØ]\›Ý™YØ]][\ØÛÝ[\ÝÙ\œ›Ü‹™\ÜÛœÙWÚœÛÛ‹Ü™X]YØ]\]YØ]ˆ
HSQTÈ
ËËËËËË•S•SKËËËÊBˆˆˆ‹ˆ
ˆÞ\Ý[WÚYˆ›ÝËˆ™\]Y\ÝYØžKˆ™\Ý[™Ù]
œÝ]\È‹™˜Z[YŠKˆ™\Ý[™Ù]
œ›ÝšY\—ØÛÙH‹ˆŠKˆ™\Ý[™Ù]
›Y\ÜØYÙH‹ˆŠKˆˆˆYˆ™\Ý[™Ù]
œÝ]\ÈŠHOH™˜Z[Yˆ[ÙH™\Ý[™Ù]
›Y\ÜØYÙH‹ˆŠKˆ™\ÜÛœÙWÚœÛÛ‹ˆ›ÝËˆ›ÝËˆ
Kˆ
Bˆ™]\›ˆ[
Ý\œÛÜ‹›\Ý›ÝÚY
B‚‚™YˆÜ™X]WÜÚYÙ[™\™ÞWÛÛ˜›Ø\™[™×Ü™\]Y\Ý
ˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆÛÛ™šYÎˆXÝÜÝ‹[žWKˆÞ\Ý[WÚYˆÝ‹ˆ™\]Y\ÝYØžNˆÝˆHˆ‹ŠHOˆXÝÜÝ‹[žWN‚ˆÞ\Ý[WÚYH˜[Y]WÜÚYÙ[™\™ÞWÜÞ\Ý[WÚY
Þ\Ý[WÚY
BˆÙ\šXÙWØÛÛ™šYÈHZ[ÜÚYÙ[™\™ÞWÜÙ\šXÙWØÛÛ™šYÊÛÛ™šYÊBˆ™\Ý[HÚYÙ[™\™ÞWÜÙ\šXÙK›Û˜›Ø\™ÜÞ\Ý[JÙ\šXÙWØÛÛ™šYËÞ\Ý[WÚYÙ\ÜÚ[Û\™\]Y\ÝË”Ù\ÜÚ[ÛŠ
JBˆ™\]Y\ÝÚYH\Ù\ÜÚYÙ[™\™ÞWÛÛ˜›Ø\™[™×Ü™\]Y\Ý
ÛÛ›‹Þ\Ý[WÚY\Þ\Ý[WÚY™\]Y\ÝYØžO\™\]Y\ÝYØžK™\Ý[\™\Ý[
BˆÛÛ›‹˜ÛÛ[Z]

Bˆ™]\›ˆÊŠœ™\Ý[œ™\]Y\ÝÚYŽˆ™\]Y\ÝÚYB‚‚™Yˆ™XÛÛ˜Ú[WÜÚYÙ[™\™ÞWÛÛ˜›Ø\™[™×Ü™\]Y\ÝÊÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹]˜Z[X›WÜÞ\Ý[WÚYÎˆ\ÝÜÝ—JHOˆ[‚ˆ›Ü›X[^™YØ]˜Z[X›HHÛ›Ü›X[^™WÜÚYÙ[™\™ÞWÜÞ\Ý[WÚYÙ›Ü—ØÛÛ\\™JÞ\Ý[WÚY
H›ÜˆÞ\Ý[WÚY[ˆ]˜Z[X›WÜÞ\Ý[WÚYÈYˆÞ\Ý[WÚYBˆYˆ›Ý›Ü›X[^™YØ]˜Z[X›N‚ˆ™]\›ˆˆ[™[™ÈH]Y\žWØ[
ˆÛÛ›‹ˆˆˆ‚ˆÑSPÕ
‚ˆ”“ÓHÚYÙ[™\™ÞWÛÛ˜›Ø\™[™×Ü™\]Y\ÝÂˆÒT‘HÝ]\ÈSˆ
	Ü™\]Y\ÝY	Ë	Ø[™XYWÜ™\]Y\ÝY	Ë	Ø[™XYWÜ™\]Y\ÝYÛÜ—ÛÛ˜›Ø\™Y	ÊBˆˆˆ‹ˆ
Bˆ›ÝÈH]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠBˆ\›Ý™YØÛÝ[Hˆ›Üˆ›ÝÈ[ˆ[™[™Î‚ˆYˆ›Ü›X[^™WÜÚYÙ[™\™ÞWÜÞ\Ý[WÚYÙ›Ü—ØÛÛ\\™J›ÝÖÈœÞ\Ý[WÚY—JH[ˆ›Ü›X[^™YØ]˜Z[X›N‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUHÚYÙ[™\™ÞWÛÛ˜›Ø\™[™×Ü™\]Y\ÝÂˆÑUÝ]\ÈH	Ø\›Ý™Y	Ë\ÝØÚXÚÙYØ]HË\›Ý™YØ]HÓÐSTÐÑJ\›Ý™YØ]ÊK\]YØ]HË\ÝÙ\œ›ÜˆH	ÉÂˆÒT‘HYHÂˆˆˆ‹ˆ
›ÝË›ÝË›ÝË›ÝÖÈšY—JKˆ
Bˆ\›Ý™YØÛÝ[
ÏHBˆ[ÙN‚ˆÛÛ›‹™^XÝ]Jˆ•TUHÚYÙ[™\™ÞWÛÛ˜›Ø\™[™×Ü™\]Y\ÝÈÑU\ÝØÚXÚÙYØ]HË\]YØ]HÈÒT‘HYHÈ‹ˆ
›ÝË›ÝË›ÝÖÈšY—JKˆ
Bˆ™]\›ˆ\›Ý™YØÛÝ[‚‚™YˆÛX[\ÜÚYÙ[™\™ÞWÜÛ˜\ÚÝÊÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹›ÝšY\ŽˆÝ‹™][[Û—Ù^\Îˆ[
HOˆ[‚ˆ™][[Û—Ù^\ÈHX^
[
™][[Û—Ù^\ÈÜˆQUSÔÒQÑS‘T‘ÖWÔÓTÒÕÔ‘US•SÓ—ÑVTÊKJBˆÙ^WÚÙ^HHˆœÚYÙ[™\™ÞWÜÛ˜\ÚÝØÛX[\ÞÜ›ÝšY\ŸWÞØÝ\œ™[Û\Ø›Û—Ù]J
Kš\ÛÙ›Ü›X]

_H‚ˆYˆ]Y\žWÜØØ[\ŠÛÛ›‹”ÑSPÕ˜[YH”“ÓH\ÜÝ]HÒT‘HÙ^HHÈ‹
Ù^WÚÙ^K
JN‚ˆ™]\›ˆˆÝ]Ù™ˆH
]][YK››ÝÊ
HH[YY[J^\Ï\™][[Û—Ù^\ÊJKš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠBˆÝ\œÛÜˆHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆSUH”“ÓH[YÜ˜][Û—Ü™X[[YWÜÛ˜\ÚÝÂˆÒT‘H›ÝšY\ˆHÂˆS‘ÛÛXÝYØ]ÂˆS‘Y“ÕSˆ
ˆÑSPÕPV
Y
Bˆ”“ÓH[YÜ˜][Û—Ü™X[[YWÜÛ˜\ÚÝÂˆÒT‘H›ÝšY\ˆHÂˆÔ“ÕT–H^\›˜[ÚYˆ
Bˆˆˆ‹ˆ
›ÝšY\‹Ý]Ù™‹›ÝšY\ŠKˆ
BˆÛÛ›‹™^XÝ]Jˆ’S”ÑT•Ôˆ‘TPÑHS•È\ÜÝ]H
Ù^K˜[YK\]YØ]
HSQTÈ
Ë	ÙÛ™IËÊH‹ˆ
Ù^WÚÙ^K]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠJKˆ
Bˆ™]\›ˆ[
Ý\œÛÜ‹œ›ÝØÛÝ[Üˆ
B‚‚™Yˆ›Ü›X[^™WÜÚYÙ[™\™ÞWÜÞ\Ý[WÜ›ÝÊˆÞ\Ý[WÜ›ÝÎˆXÝÜÝ‹[žWKˆ™X[[YWÜ›ÝÎˆXÝÜÝ‹[žWH›Û™HH›Û™Kˆ[™\™ÞWÙ›ÝÎˆXÝÜÝ‹[žWH›Û™HH›Û™KŠHOˆXÝÜÝ‹[žWN‚ˆ^\›˜[ÚYHš\œÝÛ›Û—Ù[\JÞ\Ý[WÜ›ÝËÈœÞ\Ý[RY‹šY‹œÝ][Û’Y‹œ[Y—JBˆYˆ›Ý^\›˜[ÚY‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠH™\ÜÜÝHÚYÙ[™\™ÞH˜[È›Ý^HÞ\Ý[RY[XH\È[š\ËˆŠBˆ^\›˜[Û˜[YHHš\œÝÛ›Û—Ù[\JÞ\Ý[WÜ›ÝËÈœÞ\Ý[S˜[YH‹›˜[YH‹œÝ][Û“˜[YH‹œ[˜[YH—JHÜˆ^\›˜[ÚYˆ›ÝÈH[™\™ÞWÙ›ÝÈÜˆßBˆ™X[[YHH™X[[YWÜ›ÝÈÜˆßBˆ˜]×ÜÝ]\ÈHš\œÝÛ›Û—Ù[\JÞ\Ý[WÜ›ÝËÈœÝ]\È‹œÞ\Ý[TÝ]\È‹œ[›š[™ÔÝ]\È‹œÝ]H—JHÜˆš\œÝÛ›Û—Ù[\J™X[[YKÈœÝ]\È‹œÞ\Ý[TÝ]\È‹œ[›š[™ÔÝ]\È‹œÝ]H—JBˆÞ\Ý[WÛ›Ü›X[^™YHÚYÙ[™\™ÞWÜÙ\šXÙK››Ü›X[^™WÜÞ\Ý[JÞ\Ý[WÜ›ÝÊBˆ›Ý×Û›Ü›X[^™YHÚYÙ[™\™ÞWÜÙ\šXÙK››Ü›X[^™WÙ[™\™ÞWÙ›ÝÊ›ÝÊBˆÝ]\ÈHX\ÜÚYÙ[™\™ÞWÜÝ]\Ê˜]×ÜÝ]\Ë›ÝÊBˆ›Ý\×Ü\ÈHZ[ÜÚYÙ[™\™ÞWÛ[Ûš]Üš[™×Û›Ý\ÊÊŠ™›Ý×Û›Ü›X[^™Y
ŠœÞ\Ý[WÛ›Ü›X[^™YJBˆXY×Ü\ÈHÙˆœÞ\Ý[WÜÝ]\Ï^Ü˜]×ÜÝ]\ÈÜˆ	Ý[šÛ›ÝÛ‰ßH—Bˆ›ÜˆÙ^H[ˆ
œ”ÝÙ\ˆ‹™ÜšYÝÙ\ˆ‹˜˜]\žTÝÙ\ˆ‹˜˜]\žTÛØÈ‹›ØYÝÙ\ˆŠN‚ˆYˆÙ^H[ˆ›ÝÈ[™›ÝÖÚÙ^WH›Ý[ˆ
›Û™KˆŠN‚ˆXY×Ü\Ë˜\[™
ˆžÚÙ^_O^Ù›ÝÖÚÙ^W_HŠBˆ™]\›ˆÂˆ™^\›˜[ÚYŽˆ^\›˜[ÚYˆ™^\›˜[Û˜[YHŽˆ^\›˜[Û˜[YKˆœÝ]\ÈŽˆÝ]\Ëˆœ˜]×ÜÝ]\ÈŽˆ˜]×ÜÝ]\ÈÜˆ[šÛ›ÝÛˆ‹ˆ››Ý\ÈŽˆˆžÛ›Ý\×Ü\ßHÉÎÈ	Ëš›Ú[ŠXY×Ü\Ê_H‹ˆ™™]ÚÜÝ]\ÈŽˆ›ÚÈ‹ˆ™™]ÚÙ\œ›ÜˆŽˆˆ‹ˆ
ŠœÞ\Ý[WÛ›Ü›X[^™Yˆ
Š™›Ý×Û›Ü›X[^™Yˆœ^[ØYŽˆÂˆœÞ\Ý[HŽˆÞ\Ý[WÜ›ÝËˆœ™X[[YHŽˆ™X[[YKˆ™[™\™ÞWÙ›ÝÈŽˆ›ÝËˆKˆB‚‚™Yˆ[—ÜÚYÙ[™\™ÞWØÚXÚÊÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹›ÝšY\ŽˆÝ‹žWÜ[Žˆ›ÛÛH˜[ÙJHOˆXÝÜÝ‹[žWN‚ˆÛÛ™šYÈHÙ]Ú[YÜ˜][Û—ØÛÛ™šYÊÛÛ›‹›ÝšY\ŠBˆYˆÛÛ™šYÈ\È›Û™N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠÛÛ™šYÝ\˜XØ[ÈÚYÙ[™\™ÞH˜[È[˜ÛÛ˜YKˆŠBˆ[™Ú[ÈHÙ]ÜÚYÙ[™\™ÞWÙ[™Ú[ØÛÛ™šYÊÛÛ™šYÊBˆÙ\šXÙWØÛÛ™šYÈHZ[ÜÚYÙ[™\™ÞWÜÙ\šXÙWØÛÛ™šYÊÛÛ™šYÊBˆÙ\ÜÚ[ÛˆH™\]Y\ÝË”Ù\ÜÚ[ÛŠ
BˆÛY[HÚYÙ[™\™ÞWÜÙ\šXÙK”ÚYÙ[™\™ÞPÛY[
Ù\šXÙWØÛÛ™šYËÙ\ÜÚ[Û\Ù\ÜÚ[ÛŠBˆÞ\Ý[\ÈH™]žWØ\WØØ[
ÛY[›\ÝÜÞ\Ý[\Ë[Ý×ÜÛY\[›Ý\×Ü™\]Y\ÝØÛÛ^

KÛY\\][YKœÛY\
Bˆ]˜Z[X›WÜÞ\Ý[WÚYÈHÙš\œÝÛ›Û—Ù[\J›ÝËÈœÞ\Ý[RY‹šY‹œÝ][Û’Y‹œ[Y—JH›Üˆ›ÝÈ[ˆÞ\Ý[\ÈYˆš\œÝÛ›Û—Ù[\J›ÝËÈœÞ\Ý[RY‹šY‹œÝ][Û’Y‹œ[Y—JWBˆ›Ü›X[^™YÜ›ÝÜÎˆ\ÝÙXÝÜÝ‹[žWWHH×Bˆ[™\™ÞWÙ›Ý×ØÛÝ[Hˆ[™\™ÞWÙ›Ý×Ù\œ›ÜœÎˆ\ÝÜÝ—HH×Bˆ›ÜˆÞ\Ý[WÜ›ÝÈ[ˆÞ\Ý[\Î‚ˆÞ\Ý[WÚYHš\œÝÛ›Û—Ù[\JÞ\Ý[WÜ›ÝËÈœÞ\Ý[RY‹šY‹œÝ][Û’Y‹œ[Y—JBˆYˆ›ÝÞ\Ý[WÚY‚ˆÛÛ[YBˆ[™\™ÞWÙ›ÝÎˆXÝÜÝ‹[žWHHßBˆ›ÝÈH›Ü›X[^™WÜÚYÙ[™\™ÞWÜÞ\Ý[WÜ›ÝÊÞ\Ý[WÜ›ÝËßK[™\™ÞWÙ›ÝÊBˆžN‚ˆ[™\™ÞWÙ›ÝÈH™]žWØ\WØØ[
ˆ[X™HÞ\Ý[WÚY\Þ\Ý[WÚYˆÛY[™Ù]Ù[™\™ÞWÙ›ÝÊÞ\Ý[WÚY
Kˆ[Ý×ÜÛY\[›Ý\×Ü™\]Y\ÝØÛÛ^

KˆÛY\\][YKœÛY\ˆ
Bˆ[™\™ÞWÙ›Ý×ØÛÝ[
ÏHBˆ›ÝÈH›Ü›X[^™WÜÚYÙ[™\™ÞWÜÞ\Ý[WÜ›ÝÊÞ\Ý[WÜ›ÝËßK[™\™ÞWÙ›ÝÊBˆ^Ù\\T˜]S[Z]\œ›ÜŽ‚ˆ˜Z\ÙBˆ^Ù\^Ù\[Ûˆ\È^Î‚ˆØ[š]^™YÙ\œ›ÜˆHÚYÙ[™\™ÞWÜÙ\šXÙKœØ[š]^™WÜÚYÙ[™\™ÞWÙ\œ›ÜŠ^ÊBˆ[™\™ÞWÙ›Ý×Ù\œ›ÜœË˜\[™
ˆžÜÞ\Ý[WÚYNˆÜØ[š]^™YÙ\œ›ÜŸHŠBˆ›ÝÖÈœÝ]\È—HH”Ù[HYÜÈ‚ˆ›ÝÖÈ››Ü›X[^™YÜÝ]\È—HH”Ù[HYÜÈ‚ˆ›ÝÖÈ››Ý\È—HHˆ‘[™\™ÞH›ÝÈ[™\ÜÛš]™[ˆÜØ[š]^™YÙ\œ›ÜŸH‚ˆ›ÝÖÈ™™]ÚÜÝ]\È—HH™\œ›Üˆ‚ˆ›ÝÖÈ™™]ÚÙ\œ›Üˆ—HHØ[š]^™YÙ\œ›Ü‚ˆ›ÝÖÈœ^[ØY—VÈ™™]ÚÙ\œ›Üˆ—HHØ[š]^™YÙ\œ›Ü‚ˆ›Ü›X[^™YÜ›ÝÜË˜\[™
›ÝÊBˆ[YKœÛY\
ŒŠB‚ˆYˆ›ÝžWÜ[Ž‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUH[YÜ˜][Û—ØÛÛ™šYÜÂˆÑU\ÝÜÞ[˜×ÜÝ]\ÈHË\ÝÙ\œ›ÜˆHË\]YØ]HÂˆÒT‘H›ÝšY\ˆHÂˆˆˆ‹ˆ
œÝXØÙ\ÜÈ‹ˆ‹]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠK›ÝšY\ŠKˆ
BˆÛÛ›‹˜ÛÛ[Z]

Bˆ™]\›ˆÂˆœ›ÝÜÈŽˆ›Ü›X[^™YÜ›ÝÜËˆœÞ\Ý[\ÈŽˆ›Ü›X[^™YÜ›ÝÜËˆœÝ][Û—ØÛÝ[Žˆ[ŠÞ\Ý[\ÊKˆœ™X[[YWØÛÝ[Žˆ[™\™ÞWÙ›Ý×ØÛÝ[ˆ™˜Z[YÜ™X[[YWØÛÝ[Žˆ[Š[™\™ÞWÙ›Ý×Ù\œ›ÜœÊKˆ˜[\›WØÛÝ[Žˆˆ˜[\›WÙ\œ›ÜˆŽˆŽÈ‹š›Ú[Š[™\™ÞWÙ›Ý×Ù\œ›ÜœÊKˆ™[™\™ÞWÙ›Ý×ØÛÝ[Žˆ[™\™ÞWÙ›Ý×ØÛÝ[ˆ˜˜\ÙWÝ\›Žˆ[™Ú[ÖÈ˜˜\ÙWÝ\›—Kˆ˜]˜Z[X›WÜÞ\Ý[WÚYÈŽˆ]˜Z[X›WÜÞ\Ý[WÚYËˆB‚‚™Yˆ[—Ü›ÝšY\—ØÚXÚÊÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹›ÝšY\ŽˆÝ‹žWÜ[Žˆ›ÛÛH˜[ÙJHOˆXÝÜÝ‹[žWN‚ˆYˆ›ÝšY\ˆOHS•QÔUSÓ—Ô“Õ’QT—ÔÒQÑS‘T‘ÖN‚ˆ™]\›ˆ[—ÜÚYÙ[™\™ÞWØÚXÚÊÛÛ›‹›ÝšY\‹žWÜ[YžWÜ[ŠBˆ™]\›ˆ[—Ù\Ú[ÛœÛÛ\—ØÚXÚÊÛÛ›‹›ÝšY\‹žWÜ[YžWÜ[ŠB‚‚™Yˆš[™ÜÚYÙ[™\™ÞWØ\ÜÙ]ÚY
ÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹›ÝšY\ŽˆÝ‹^\›˜[ÚYˆÝ‹^\›˜[Û˜[YNˆÝŠHOˆ[›Û™N‚ˆX\YHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕ\ÜÙ]ÚYˆ”“ÓH\ÜÙ]Ú[YÜ˜][ÛœÂˆÒT‘H›ÝšY\ˆHÈS‘^\›˜[ÚYHÈS‘[˜X›YHBˆSRUBˆˆˆ‹ˆ
›ÝšY\‹^\›˜[ÚY
Kˆ
K™™]ÚÛ™J
BˆYˆX\Y‚ˆ™]\›ˆ[
X\YÈ˜\ÜÙ]ÚY—JB‚ˆ›Ü›X[^™YH›Ü›X[^™WÛ˜[YJ^\›˜[Û˜[YJBˆYˆ›Ý›Ü›X[^™Y‚ˆ™]\›ˆ›Û™B‚ˆ›Ú™XÝÛX]Ú\ÈHÂˆ[
›ÝÖÈšY—JBˆ›Üˆ›ÝÈ[ˆ]Y\žWØ[
ÛÛ›‹”ÑSPÕY›Ú™XÝÛ˜[YH”“ÓH\ÜÙ]ÈÒT‘H›Ú™XÝÛ˜[YHTÈ“Õ•SŠBˆYˆ›Ü›X[^™WÛ˜[YJ›ÝÖÈœ›Ú™XÝÛ˜[YH—HÜˆˆŠHOH›Ü›X[^™YˆBˆYˆ[ŠÙ]
›Ú™XÝÛX]Ú\ÊJHOHN‚ˆ™]\›ˆ›Ú™XÝÛX]Ú\ÖÌBˆYˆ[ŠÙ]
›Ú™XÝÛX]Ú\ÊJHˆN‚ˆ™]\›ˆ›Û™B‚ˆ[X\×ÛX]Ú\ÈHÂˆ[
›ÝÖÈ˜\ÜÙ]ÚY—JBˆ›Üˆ›ÝÈ[ˆ]Y\žWØ[
ˆÛÛ›‹ˆ”ÑSPÕ\ÜÙ]ÚY”“ÓH\ÜÙ]Ø[X\Ù\ÈÒT‘H›Ü›X[^™YØ[X\ÈHÈS‘ÓÐSTÐÑJXÝ]™KJHHH‹ˆ
›Ü›X[^™Y
Kˆ
BˆBˆYˆ[ŠÙ]
[X\×ÛX]Ú\ÊJHOHN‚ˆ™]\›ˆ[X\×ÛX]Ú\ÖÌBˆYˆ[ŠÙ]
[X\×ÛX]Ú\ÊJHˆN‚ˆ™]\›ˆ›Û™B‚ˆÜ›Ý\ÛX]Ú\ÈHÂˆ[
›ÝÖÈšY—JBˆ›Üˆ›ÝÈ[ˆ]Y\žWØ[
ÛÛ›‹”ÑSPÕY[œÝ[][Û—ÙÜ›Ý\”“ÓH\ÜÙ]ÈÒT‘HÓÐSTÐÑJ[œÝ[][Û—ÙÜ›Ý\	ÉÊHOH	ÉÈŠBˆYˆ›Ü›X[^™WÛ˜[YJ›ÝÖÈš[œÝ[][Û—ÙÜ›Ý\—HÜˆˆŠHOH›Ü›X[^™YˆBˆYˆ[ŠÙ]
Ü›Ý\ÛX]Ú\ÊJHOHN‚ˆ™]\›ˆÜ›Ý\ÛX]Ú\ÖÌBˆ™]\›ˆ›Û™B‚‚™Yˆ[œÙ\Ú[YÜ˜][Û—Ü™X[[YWÜÛ˜\ÚÝ
ˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆ\ÜÙ]ÚYˆ[›Û™Kˆ›ÝšY\ŽˆÝ‹ˆ›ÝÎˆXÝÜÝ‹[žWKˆÛÛXÝYØ]ˆÝ‹ŠHOˆ›Û™N‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆS”ÑT•S•È[YÜ˜][Û—Ü™X[[YWÜÛ˜\ÚÝÈ
ˆ\ÜÙ]ÚY›ÝšY\‹^\›˜[ÚYÛÛXÝYØ]^\›˜[ÜÝ]\Ë›Ü›X[^™YÜÝ]\Ëˆ—ÜÝÙ\—ÚÝËØYÜÝÙ\—ÚÝËÜšYÜÝÙ\—ÚÝ×Ü˜]Ë˜]\žWÜÝÙ\—ÚÝË˜]\žWÜÛØ×ÜÝˆ]—ÜÝÙ\—ÚÝËX×ÜÝÙ\—ÚÝËX]Ü[\ÜÝÙ\—ÚÝË—ØØ\XÚ]WÚÝË˜]\žWØØ\XÚ]WÚÝÚ^[ØYÚœÛÛ‚ˆ
HSQTÈ
ËËËËËËËËËËËËËËËËÊBˆˆˆ‹ˆ
ˆ\ÜÙ]ÚYˆ›ÝšY\‹ˆ›ÝÖÈ™^\›˜[ÚY—KˆÛÛXÝYØ]ˆ›ÝË™Ù]
œ˜]×ÜÝ]\ÈŠKˆ›ÝË™Ù]
œÝ]\ÈŠKˆ›ÝË™Ù]
œ—ÜÝÙ\—ÚÝÈŠKˆ›ÝË™Ù]
›ØYÜÝÙ\—ÚÝÈŠKˆ›ÝË™Ù]
™ÜšYÜÝÙ\—ÚÝ×Ü˜]ÈŠKˆ›ÝË™Ù]
˜˜]\žWÜÝÙ\—ÚÝÈŠKˆ›ÝË™Ù]
˜˜]\žWÜÛØ×ÜÝŠKˆ›ÝË™Ù]
™]—ÜÝÙ\—ÚÝÈŠKˆ›ÝË™Ù]
˜X×ÜÝÙ\—ÚÝÈŠKˆ›ÝË™Ù]
šX]Ü[\ÜÝÙ\—ÚÝÈŠKˆ›ÝË™Ù]
œ—ØØ\XÚ]WÚÝÈŠKˆ›ÝË™Ù]
˜˜]\žWØØ\XÚ]WÚÝÚŠKˆœÛÛ‹™[\ÊˆÂˆ
ŠŠ›ÝË™Ù]
œ^[ØYŠHÜˆßJKˆ™™]ÚÜÝ]\ÈŽˆ›ÝË™Ù]
™™]ÚÜÝ]\È‹›ÚÈŠKˆ™™]ÚÙ\œ›ÜˆŽˆ›ÝË™Ù]
™™]ÚÙ\œ›Üˆ‹ˆŠKˆKˆ[œÝ\™WØ\ØÚZOUYKˆ
Kˆ
Kˆ
B‚‚™Yˆ[—ÜÚYÙ[™\™ÞWÜÞ[˜ÊÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹›ÝšY\ŽˆÝˆH”ÚYÙ[™\™ÞH‹šYÙÙ\—Ý\NˆÝˆH›X[X[ŠHOˆXÝÜÝ‹[žWN‚ˆÚ]ÒQÑS‘T‘ÖWÔÖS×ÓÐÒÎ‚ˆÛÛ™šYÈHÙ]Ú[YÜ˜][Û—ØÛÛ™šYÊÛÛ›‹›ÝšY\ŠBˆYˆÛÛ™šYÈ\È›Û™N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠÛÛ™šYÝ\˜XØ[ÈÚYÙ[™\™ÞH˜[È[˜ÛÛ˜YKˆŠBˆYˆ›ÝÛÛ™šYÖÈ™[˜X›Y—N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠH[YÜ˜XØ[ÈÚYÙ[™\™ÞH\ÝH\Ø]]˜YKˆŠB‚ˆ[—ÚYHÜ™X]WÚ[YÜ˜][Û—Ü[ŠÛÛ›‹›ÝšY\‹šYÙÙ\—Ý\JBˆ˜]ÚÚYHÜ™X]WÛ[Ûš]Üš[™×Ø˜]Ú
ˆÛÛ›‹ˆ™XÛÜ™Ù]OXÝ\œ™[Û\Ø›Û—Ù]J
Kš\ÛÙ›Ü›X]

KˆY˜][Û›Ý\ÏYˆ”Þ[˜ÈÜ›ÝšY\ŸH
ÝšYÙÙ\—Ý\_JH‹ˆ˜]×Ú[œ]Hˆ‹ˆÛÝ\˜ÙO\›ÝšY\‹ˆ
BˆžN‚ˆ™\Ý[H[—Ü›ÝšY\—ØÚXÚÊÛÛ›‹›ÝšY\‹žWÜ[UYJBˆ›ÝÜÈH™\Ý[Èœ›ÝÜÈ—Bˆ™XÛÛ˜Ú[WÜÚYÙ[™\™ÞWÛÛ˜›Ø\™[™×Ü™\]Y\ÝÊÛÛ›‹™\Ý[™Ù]
˜]˜Z[X›WÜÞ\Ý[WÚYÈ‹×JJBˆX]ÚYHˆ[œ™\ÛÛ™YHˆÞ[˜ÙYØ\ÜÙ]ÚYÎˆÙ]Ú[HHÙ]

Bˆ[\Ù]™[Îˆ\ÝÙXÝÜÝ‹[žWWHH×Bˆ›ÝÈH]][YK››ÝÊ
BˆÛÛXÝYØ]H›ÝËš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠB‚ˆ›Üˆ›ÝÈ[ˆ›ÝÜÎ‚ˆ\ÜÙ]ÚYHš[™ÜÚYÙ[™\™ÞWØ\ÜÙ]ÚY
ÛÛ›‹›ÝšY\‹›ÝÖÈ™^\›˜[ÚY—K›ÝÖÈ™^\›˜[Û˜[YH—JBˆ[œÙ\Ú[YÜ˜][Û—Ü™X[[YWÜÛ˜\ÚÝ
ÛÛ›‹\ÜÙ]ÚYX\ÜÙ]ÚY›ÝšY\\›ÝšY\‹›ÝÏ\›ÝËÛÛXÝYØ]XÛÛXÝYØ]
BˆYˆ›ÝË™Ù]
™™]ÚÜÝ]\ÈŠHOH™\œ›ÜˆŽ‚ˆYˆ\ÜÙ]ÚY‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUH\ÜÙ]Ú[YÜ˜][ÛœÂˆÑU\ÝÙ\œ›ÜˆHË\ÝÜÞ[˜×Ø]HÂˆÒT‘H›ÝšY\ˆHÈS‘^\›˜[ÚYHÂˆˆˆ‹ˆ
›ÝË™Ù]
™™]ÚÙ\œ›Üˆ‹ˆŠKÛÛXÝYØ]›ÝšY\‹›ÝÖÈ™^\›˜[ÚY—JKˆ
BˆÛÛ[YBˆYˆ\ÜÙ]ÚY‚ˆÞ[˜ÙYØ\ÜÙ]ÚYË˜Y
\ÜÙ]ÚY
Bˆ™]š[Ý\ÈHÙ]Û]\ÝÛ[Ûš]Üš[™×Ü›ÝÊÛÛ›‹\ÜÙ]ÚY
Bˆ\XØ]WÛ]\ÝH
ˆ™]š[Ý\È\È›Ý›Û™Bˆ[™™]š[Ý\ÖÈœÝ]\È—HOH›ÝÖÈœÝ]\È—Bˆ[™™]š[Ý\ÖÈœ™XÛÜ™Ù]H—HOHÝ\œ™[Û\Ø›Û—Ù]J
Kš\ÛÙ›Ü›X]

Bˆ[™™]š[Ý\ÖÈœÛÝ\˜ÙH—HOH›ÝšY\‚ˆ
BˆYˆ›Ý\XØ]WÛ]\Ý‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆS”ÑT•S•È[Ûš]Üš[™×Ü™XÛÜ™È
\ÜÙ]ÚYÝ]\Ë™XÛÜ™Ù]K›Ý\ËÛÝ\˜ÙK˜]ÚÚY
BˆSQTÈ
ËËËËËÊBˆˆˆ‹ˆ
ˆ\ÜÙ]ÚYˆ›ÝÖÈœÝ]\È—KˆÝ\œ™[Û\Ø›Û—Ù]J
Kš\ÛÙ›Ü›X]

Kˆ›ÝÖÈ››Ý\È—Kˆ›ÝšY\‹ˆ˜]ÚÚYˆ
Kˆ
Bˆ]™[HZ[Û[Ûš]Üš[™×Ø[\Ù]™[
ˆÛÛ›‹ˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆ™]š[Ý\×ÜÝ]\Ï\™]š[Ý\ÖÈœÝ]\È—HYˆ™]š[Ý\È[ÙHˆ‹ˆÝ\œ™[ÜÝ]\Ï\›ÝÖÈœÝ]\È—Kˆ\[™YØ]XÛÛXÝYØ]ˆ[\›WØÛÛ^\›ÝËˆ
BˆYˆ]™[‚ˆ[\Ù]™[Ë˜\[™
]™[
BˆÜ™X]WÛÜ—Ý\]WØ\ÜÙ]Ú[YÜ˜][ÛŠÛÛ›‹\ÜÙ]ÚY›ÝšY\‹›ÝÖÈ™^\›˜[ÚY—K›ÝÖÈ™^\›˜[Û˜[YH—K›ÝÖÈœÝ]\È—JBˆX]ÚY
ÏHBˆ[ÙN‚ˆ\Ù\Ú[YÜ˜][Û—Ý[œ™\ÛÛ™Y
ˆÛÛ›‹ˆ›ÝšY\\›ÝšY\‹ˆ[—ÚY\[—ÚYˆ^\›˜[ÚY\›ÝÖÈ™^\›˜[ÚY—Kˆ^\›˜[Û˜[YO\›ÝÖÈ™^\›˜[Û˜[YH—KˆÝ]\Ï\›ÝÖÈœÝ]\È—Kˆ^[ØY\›ÝÖÈœ^[ØY—Kˆ
Bˆ[œ™\ÛÛ™Y
ÏHB‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUH[Ûš]Üš[™×Ú[\ÜØ˜]Ú\ÂˆÑU[\ÜYØÛÝ[HËX]ÚYØÛÝ[HË[›X]ÚYØÛÝ[HË]]×Ü™\ÛÛ™YØÛÝ[HˆÒT‘HYHÂˆˆˆ‹ˆ
X]ÚY
È[œ™\ÛÛ™YX]ÚY[œ™\ÛÛ™Y˜]ÚÚY
Kˆ
Bˆ˜Z[YÜ™X[[YWØÛÝ[H[
™\Ý[™Ù]
™˜Z[YÜ™X[[YWØÛÝ[‹
HÜˆ
BˆÞ[˜×ÜÝ]\ÈHœ\X[ˆYˆ˜Z[YÜ™X[[YWØÛÝ[[ÙHœÝXØÙ\ÜÈ‚ˆÞ[˜×Ù\œ›ÜˆHÚYÙ[™\™ÞWÜÙ\šXÙKœØ[š]^™WÜÚYÙ[™\™ÞWÙ\œ›ÜŠ™\Ý[™Ù]
˜[\›WÙ\œ›Üˆ‹ˆŠJBˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUH[YÜ˜][Û—ØÛÛ™šYÜÂˆÑU\ÝÜÞ[˜×Ø]HË\ÝÜÞ[˜×ÜÝ]\ÈHË\ÝÙ\œ›ÜˆHË\]YØ]HÂˆÒT‘H›ÝšY\ˆHÂˆˆˆ‹ˆ
ÛÛXÝYØ]Þ[˜×ÜÝ]\ËÞ[˜×Ù\œ›Ü‹ÛÛXÝYØ]›ÝšY\ŠKˆ
Bˆš[˜[^™WÚ[YÜ˜][Û—Ü[ŠˆÛÛ›‹ˆ[—ÚYˆÝ]\Ï\Þ[˜×ÜÝ]\ËˆX]ÚYØÛÝ[[X]ÚYˆ[œ™\ÛÛ™YØÛÝ[][œ™\ÛÛ™Yˆ]]×Ü™\ÛÛ™YØÛÝ[LˆÝ[[X\žWÚœÛÛ^Âˆœ›ÝšY\—Ü›ÝÜÈŽˆ[Š›ÝÜÊKˆœÝ][Û—Ü›ÝÜÈŽˆ™\Ý[™Ù]
œÝ][Û—ØÛÝ[‹[Š›ÝÜÊJKˆœ™X[[YWÜ›ÝÜÈŽˆ™\Ý[™Ù]
œ™X[[YWØÛÝ[‹
Kˆ™˜Z[YÜ™X[[YWÜ›ÝÜÈŽˆ˜Z[YÜ™X[[YWØÛÝ[ˆ™[™\™ÞWÙ›Ý×Ù\œ›ÜˆŽˆÞ[˜×Ù\œ›Ü‹ˆKˆ
BˆžN‚ˆÛX[\ÜÚYÙ[™\™ÞWÜÛ˜\ÚÝÊÛÛ›‹›ÝšY\‹[
ÛÛ™šYË™Ù]
œÛ˜\ÚÝÜ™][[Û—Ù^\ÈŠHÜˆQUSÔÒQÑS‘T‘ÖWÔÓTÒÕÔ‘US•SÓ—ÑVTÊJBˆ^Ù\^Ù\[Ûˆ\ÈÛX[\Ù^Î‚ˆÑÑÑT‹Ø\›š[™Ê”ÚYÙ[™\™ÞHÛ˜\ÚÝÛX[\˜Z[Yˆ	\È‹ÚYÙ[™\™ÞWÜÙ\šXÙKœØ[š]^™WÜÚYÙ[™\™ÞWÙ\œ›ÜŠÛX[\Ù^ÊJBˆ›ØÙ\Ü×Û[Ûš]Üš[™×Ø[\ÊÛÛ›‹[\Ù]™[Ë˜]ÚÚY›ÝÊBˆÛÛ›‹˜ÛÛ[Z]

Bˆ™]\›ˆÈ›X]ÚYŽˆX]ÚY[œ™\ÛÛ™YŽˆ[œ™\ÛÛ™Y˜]]×Ü™\ÛÛ™YŽˆœÛ˜\ÚÝÈŽˆ[Š›ÝÜÊKœÝ]\ÈŽˆÞ[˜×ÜÝ]\ßBˆ^Ù\\T˜]S[Z]\œ›Üˆ\È^Î‚ˆ[[HX\š×Ø\WØÛÛÛÝÛŠˆÛÛ›‹ˆ›ÝšY\‹ˆTWÐT‘PWÔÕUKˆ^Ë›Y\ÜØYÙKˆÛÛÛÝÛ—Ý[[Y^Ë˜ÛÛÛÝÛ—Ý[[ˆ
Bˆ›ÝYžWØ\WÜ˜]WÛ[Z]
ÛÛ›‹›ÝšY\‹TWÐT‘PWÔÕUK[[^Ë›Y\ÜØYÙJBˆØ[š]^™YÙ\œ›ÜˆHÚYÙ[™\™ÞWÜÙ\šXÙKœØ[š]^™WÜÚYÙ[™\™ÞWÙ\œ›ÜŠ^Ë›Y\ÜØYÙJBˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUH[YÜ˜][Û—ØÛÛ™šYÜÂˆÑU\ÝÜÞ[˜×ÜÝ]\ÈH	ÝØZ][™×Ü˜]WÛ[Z]	Ë\ÝÙ\œ›ÜˆHË\]YØ]HÂˆÒT‘H›ÝšY\ˆHÂˆˆˆ‹ˆ
Ø[š]^™YÙ\œ›Ü‹]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠK›ÝšY\ŠKˆ
Bˆš[˜[^™WÚ[YÜ˜][Û—Ü[ŠˆÛÛ›‹ˆ[—ÚYˆÝ]\ÏHØZ][™×Ü˜]WÛ[Z]‹ˆX]ÚYØÛÝ[Lˆ[œ™\ÛÛ™YØÛÝ[Lˆ]]×Ü™\ÛÛ™YØÛÝ[Lˆ\œ›Ü—ÛY\ÜØYÙO\Ø[š]^™YÙ\œ›Ü‹ˆ
BˆÛÛ›‹˜ÛÛ[Z]

Bˆ˜Z\ÙH\T˜]S[Z]\œ›ÜŠ›ÝšY\‹TWÐT‘PWÔÕUK[[Ù]Ü›ÝšY\—ØÛÛÛÝÛ—Ü™X\ÛÛŠÛÛ›‹›ÝšY\‹TWÐT‘PWÔÕUJJHœ›ÛH^Âˆ^Ù\^Ù\[Ûˆ\È^Î‚ˆØ[š]^™YÙ\œ›ÜˆHÚYÙ[™\™ÞWÜÙ\šXÙKœØ[š]^™WÜÚYÙ[™\™ÞWÙ\œ›ÜŠ^ÊBˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUH[YÜ˜][Û—ØÛÛ™šYÜÂˆÑU\ÝÜÞ[˜×ÜÝ]\ÈH	Ù\œ›Ü‰Ë\ÝÙ\œ›ÜˆHË\]YØ]HÂˆÒT‘H›ÝšY\ˆHÂˆˆˆ‹ˆ
Ø[š]^™YÙ\œ›Ü‹]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠK›ÝšY\ŠKˆ
Bˆš[˜[^™WÚ[YÜ˜][Û—Ü[ŠˆÛÛ›‹ˆ[—ÚYˆÝ]\ÏH™\œ›Üˆ‹ˆX]ÚYØÛÝ[Lˆ[œ™\ÛÛ™YØÛÝ[Lˆ]]×Ü™\ÛÛ™YØÛÝ[Lˆ\œ›Ü—ÛY\ÜØYÙO\Ø[š]^™YÙ\œ›Ü‹ˆ
BˆÛÛ›‹˜ÛÛ[Z]

Bˆ˜Z\ÙB‚‚™YˆØYÛØØ[Ù\Ú[ÛœÛÛ\—ÜÝ][ÛœÊˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ›ÝšY\ŽˆÝˆHS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹ŠHOˆ\ÝÙXÝÜÝ‹[žWWN‚ˆ›ÝÜÈHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕZK™^\›˜[ÚYZK™^\›˜[Û˜[YKKœ›Ú™XÝÛ˜[YBˆ”“ÓH\ÜÙ]Ú[YÜ˜][ÛœÈZBˆ“ÒSˆ\ÜÙ]ÈHÓˆKšYHZK˜\ÜÙ]ÚYˆÒT‘HZKœ›ÝšY\ˆHÂˆS‘ZK™[˜X›YHBˆS‘ÓÐSTÐÑJZK™^\›˜[ÚY	ÉÊHOH	ÉÂˆS‘ÓÐSTÐÑJK›[Ûš]Üš[™×ÜÝ]\Ë	ØXÝ]™IÊHOH	Ù\ØX›Y	ÂˆÔ‘Tˆ–HKœ›Ú™XÝÛ˜[YHÓÓUH“ÐÐTÑBˆˆˆ‹ˆ
›ÝšY\‹
Kˆ
K™™]Ú[

Bˆ™]\›ˆÂˆÂˆœ[ÛÙHŽˆÝŠ›ÝÖÈ™^\›˜[ÚY—JKˆœ[˜[YHŽˆÝŠˆ›ÝÖÈ™^\›˜[Û˜[YH—HÜˆ›ÝÖÈœ›Ú™XÝÛ˜[YH—HÜˆ›ÝÖÈ™^\›˜[ÚY—Bˆ
KˆBˆ›Üˆ›ÝÈ[ˆ›ÝÜÂˆB‚‚™Yˆ[—Ù\Ú[ÛœÛÛ\—ØÚXÚÊˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ›ÝšY\ŽˆÝ‹ˆžWÜ[Žˆ›ÛÛH˜[ÙKˆ[˜ÛYWÙXYÛ›ÜÝXÜÎˆ›ÛÛHYKˆ™Y™\—ÛØØ[ÜÝ][Û—Ú[™[ÜžNˆ›ÛÛH˜[ÙKŠHOˆXÝÜÝ‹[žWN‚ˆÛÛ™šYÈHÙ]Ú[YÜ˜][Û—ØÛÛ™šYÊÛÛ›‹›ÝšY\ŠBˆYˆÛÛ™šYÈ\È›Û™N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠÛÛ™šYÝ\˜XØ[È\Ú[Û”ÛÛ\ˆ˜[È[˜ÛÛ˜YKˆŠBˆ[™Ú[ÈHÙ]Ù\Ú[ÛœÛÛ\—Ù[™Ú[ØÛÛ™šYÊÛÛ™šYÊBˆYˆ›Ý[™Ú[ÖÈ˜˜\ÙWÝ\›—N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ‘˜[HH˜\ÙHT“È\Ú[Û”ÛÛ\‹ˆŠB‚ˆ\ÝÙ\œ›ÜŽˆ^Ù\[Ûˆ›Û™HH›Û™Bˆ›Üˆ][\[ˆ˜[™ÙJŠN‚ˆžN‚ˆÙ\ÜÚ[Û‹ÈHÙ]Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[ÛŠÛÛ™šYË›Ü˜ÙWÛÙÚ[X][\OHJBˆ[™[ÜžWÙ]HHÙ]Ø\ÜÝ]WÝ˜[YJˆÛÛ›‹ˆ•TÒSÓ”ÓÓT—ÔÕUSÓ—ÒS•‘S•Ô–WÑUWÒÑVKˆ
Bˆ\ÙWÛØØ[Ú[™[ÜžHH
ˆ™Y™\—ÛØØ[ÜÝ][Û—Ú[™[ÜžBˆ[™[™[ÜžWÙ]HOHÝ\œ™[Û\Ø›Û—Ù]J
Kš\ÛÙ›Ü›X]

Bˆ
BˆÝ][ÛœÈH
ˆØYÛØØ[Ù\Ú[ÛœÛÛ\—ÜÝ][ÛœÊÛÛ›‹›ÝšY\ŠBˆYˆ\ÙWÛØØ[Ú[™[ÜžBˆ[ÙH×Bˆ
BˆÝ][Û—Û\ÝØ\WØØ[ÈHˆYˆ›ÝÝ][ÛœÎ‚ˆÝ][ÛœÈH™]ÚÙ\Ú[ÛœÛÛ\—ÜÝ][ÛœÊˆÙ\ÜÚ[Û‹ˆ˜\ÙWÝ\›Y[™Ú[ÖÈ˜˜\ÙWÝ\›—Kˆ[™Ú[Y[™Ú[ÖÈœ[×Ù[™Ú[—Kˆ
BˆÝ][Û—Û\ÝØ\WØØ[ÈHX^
KX]˜ÙZ[
[ŠÝ][ÛœÊHÈL
JBˆÙ]Ø\ÜÝ]WÝ˜[YJˆÛÛ›‹ˆ•TÒSÓ”ÓÓT—ÔÕUSÓ—ÒS•‘S•Ô–WÑUWÒÑVKˆÝ\œ™[Û\Ø›Û—Ù]J
Kš\ÛÙ›Ü›X]

Kˆ
BˆÛÛ›‹˜ÛÛ[Z]

BˆYˆ›ÝÝ][ÛœÎ‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠHTH\Ú[Û”ÛÛ\ˆ˜[È]›Û™]HÙ[˜Z\È\˜H\ÝHÛÛKˆŠB‚ˆÝ][Û—ØÛÙ\ÈHÂˆÝŠ›ÝË™Ù]
œ[ÛÙHŠHÜˆ›ÝË™Ù]
œÝ][ÛÛÙHŠHÜˆˆŠKœÝš\

Bˆ›Üˆ›ÝÈ[ˆÝ][ÛœÂˆYˆÝŠ›ÝË™Ù]
œ[ÛÙHŠHÜˆ›ÝË™Ù]
œÝ][ÛÛÙHŠHÜˆˆŠKœÝš\

BˆBˆ™X[[YWÛX\H™]ÚÙ\Ú[ÛœÛÛ\—Ü™X[[YWÛX\
ˆÙ\ÜÚ[Û‹ˆ˜\ÙWÝ\›Y[™Ú[ÖÈ˜˜\ÙWÝ\›—Kˆ[™Ú[Y[™Ú[ÖÈœ™X[Ý[YWÙ[™Ú[—KˆÝ][Û—ØÛÙ\Ï\Ý][Û—ØÛÙ\Ëˆ
Bˆ[\›WÛX\ˆXÝÜÝ‹\ÝÙXÝÜÝ‹[žWWWHHßBˆ[\›WÙ\œ›ÜˆHˆ‚ˆYˆ[˜ÛYWÙXYÛ›ÜÝXÜÎ‚ˆžN‚ˆ[\›WÛX\H™]ÚÙ\Ú[ÛœÛÛ\—Ø[\›WÛX\
ˆÙ\ÜÚ[Û‹ˆ˜\ÙWÝ\›Y[™Ú[ÖÈ˜˜\ÙWÝ\›—Kˆ[™Ú[Y[™Ú[ÖÈ˜[\›\×Ù[™Ú[—KˆÝ][Û—ØÛÙ\Ï\Ý][Û—ØÛÙ\Ëˆ
Bˆ^Ù\\T˜]S[Z]\œ›ÜŽ‚ˆ˜Z\ÙBˆ^Ù\^Ù\[Ûˆ\È^Î‚ˆ[\›WÙ\œ›ÜˆHÝŠ^ÊBˆ›Ü›X[^™YÜ›ÝÜÈHÂˆ›Ü›X[^™WÙ\Ú[ÛœÛÛ\—Ü[Ü›ÝÊˆÝ][Û—Ü›ÝËˆ™X[[YWÛX\™Ù]
ÝŠÝ][Û—Ü›ÝË™Ù]
œ[ÛÙHŠHÜˆÝ][Û—Ü›ÝË™Ù]
œÝ][ÛÛÙHŠHÜˆˆŠKœÝš\

JKˆ[\›WÛX\™Ù]
ÝŠÝ][Û—Ü›ÝË™Ù]
œ[ÛÙHŠHÜˆÝ][Û—Ü›ÝË™Ù]
œÝ][ÛÛÙHŠHÜˆˆŠKœÝš\

K×JKˆ
Bˆ›ÜˆÝ][Û—Ü›ÝÈ[ˆÝ][ÛœÂˆBˆœ™XZÂˆ^Ù\^Ù\[Ûˆ\È^Î‚ˆ\ÝÙ\œ›ÜˆH^ÂˆYˆ›Ý\×Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[Û—Ù^\™YÙ\œ›ÜŠ^ÊHÜˆ][\OHN‚ˆ˜Z\ÙBˆÑÑÑT‹š[™›Ê‘\Ú[Û”ÛÛ\ˆÙ\ÜÚ[Ûˆ^\™YÈ[˜[Y][™ÈØXÚH[™™]žZ[™ÈÙÚ[ˆÛ˜ÙHŠBˆ[˜[Y]WÙ\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[ÛŠÛÛ™šYÊBˆ[ÙN‚ˆ˜Z\ÙH\ÝÙ\œ›ÜˆÜˆ˜[YQ\œ›ÜŠ‘˜[H\ØÛÛšXÚYH›È\Ú[Û”ÛÛ\‹ˆŠB‚ˆYˆ›ÝžWÜ[Ž‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUH[YÜ˜][Û—ØÛÛ™šYÜÂˆÑU\ÝÜÞ[˜×ÜÝ]\ÈHË\ÝÙ\œ›ÜˆHË\]YØ]HÂˆÒT‘H›ÝšY\ˆHÂˆˆˆ‹ˆ
œÝXØÙ\ÜÈ‹ˆ‹]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠK›ÝšY\ŠKˆ
BˆÛÛ›‹˜ÛÛ[Z]

Bˆ™]\›ˆÂˆœ›ÝÜÈŽˆ›Ü›X[^™YÜ›ÝÜËˆœÝ][Û—ØÛÝ[Žˆ[ŠÝ][ÛœÊKˆœ™X[[YWØÛÝ[Žˆ[Š™X[[YWÛX\
Kˆ˜[\›WØÛÝ[ŽˆÝ[J[Š][\ÊH›Üˆ][\È[ˆ[\›WÛX\˜[Y\Ê
JKˆ˜[\›WÙ\œ›ÜˆŽˆ[\›WÙ\œ›Ü‹ˆœÝ][Û—Ú[™[ÜžWÜÛÝ\˜ÙHŽˆ
ˆ›ØØ[ˆYˆ\ÙWÛØØ[Ú[™[ÜžH[ÙH˜\H‚ˆ
Kˆ˜\WØØ[×Ý\ÙYŽˆ
ˆÝ][Û—Û\ÝØ\WØØ[Âˆ
ÈX]˜ÙZ[
[ŠÝ][Û—ØÛÙ\ÊHÈL
Bˆ
È
ˆX]˜ÙZ[
[ŠÝ][Û—ØÛÙ\ÊHÈL
BˆYˆ[˜ÛYWÙXYÛ›ÜÝXÜÂˆ[ÙHˆ
Bˆ
KˆB‚‚™Yˆ[—Ù\Ú[ÛœÛÛ\—Ü›ÙXÝ[Û—ÜÞ[˜ÊˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ›ÝšY\ŽˆÝˆH‘\Ú[Û”ÛÛ\ˆ‹ˆ\™Ù]Ù]Nˆ]H›Û™HH›Û™Kˆ\š[ÙÝ\NˆÝˆH™^H‹ŠHOˆXÝÜÝ‹[žWN‚ˆYˆ\š[ÙÝ\H›Ý[ˆÈ™^H‹›[ÛŸN‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ”\š[ÙÈ[˜[YÈ\˜H\™›Ü›X[˜ÙKˆŠBˆYˆ\™Ù]Ù]H\È›Û™N‚ˆ\™Ù]Ù]HH]KÙ^J
HH[YY[J^\ÏLJBˆYˆ\š[ÙÝ\HOH›[ÛŽ‚ˆ\™Ù]Ù]HH\™Ù]Ù]Kœ™\XÙJ^OLJB‚ˆYˆ›Ý•TÒSÓ”ÓÓT—ÔÖS×ÓÐÒË˜XÜ]Z\™J›ØÚÚ[™ÏQ˜[ÙJN‚ˆY\ÜØYÙHH”Ú[˜Ü›Ûš^˜XØ[È\Ú[Û”ÛÛ\ˆYÛ›Ü˜YHÜœ]YH˜H^\ÝHÝ]˜H[HÝ\œÛËˆ‚ˆÑÑÑT‹š[™›ÊY\ÜØYÙJBˆ™]\›ˆÈœ›ØÙ\ÜÙYŽˆ›Z\ÜÚ[™×Ù]HŽˆ››×Ü™Y™\™[˜ÙHŽˆœ\š[ÙÙ]HŽˆ\™Ù]Ù]Kš\ÛÙ›Ü›X]

KœÝÜYÜ™X\ÛÛˆŽˆY\ÜØYÙ_B‚ˆÚ]™[X\ÙWÙ\Ú[ÛœÛÛ\—ÜÞ[˜×ÛØÚÊ
N‚ˆÛÛ™šYÈHÙ]Ú[YÜ˜][Û—ØÛÛ™šYÊÛÛ›‹›ÝšY\ŠBˆYˆÛÛ™šYÈ\È›Û™N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠÛÛ™šYÝ\˜XØ[È\Ú[Û”ÛÛ\ˆ˜[È[˜ÛÛ˜YKˆŠBˆYˆ›ÝÛÛ™šYÖÈ™[˜X›Y—N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠH[YÜ˜XØ[È\Ú[Û”ÛÛ\ˆ\ÝH\Ø]]˜YKˆŠBˆ[™Ú[ÈHÙ]Ù\Ú[ÛœÛÛ\—Ù[™Ú[ØÛÛ™šYÊÛÛ™šYÊBˆYˆ›Ý[™Ú[ÖÈ˜˜\ÙWÝ\›—N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ‘˜[HH˜\ÙHT“È\Ú[Û”ÛÛ\‹ˆŠB‚ˆX\YØ\ÜÙ]ÈH]Y\žWØ[
ˆÛÛ›‹ˆˆˆ‚ˆÑSPÕˆKšYTÈ\ÜÙ]ÚYˆKœ›Ú™XÝÛ˜[YKˆKšÝÜˆZK™^\›˜[ÚYˆÓÐSTÐÑJË™[˜X›YJHTÈ\™›Ü›X[˜ÙWÙ[˜X›YˆËØ\›š[™×Ù]šX][Û—ÜÝˆË˜[\Ù]šX][Û—ÜÝˆË˜Üš]XØ[Ù]šX][Û—ÜÝˆË˜˜\Ù[[™WÞYX\œËˆË›Z[—Ø˜\Ù[[™WÜÚ[ËˆË›[ÛWØYÙ]ÚœÛÛ‚ˆ”“ÓH\ÜÙ]Ú[YÜ˜][ÛœÈZBˆ“ÒSˆ\ÜÙ]ÈHÓˆKšYHZK˜\ÜÙ]ÚYˆQ•“ÒSˆ\™›Ü›X[˜ÙWÜÙ][™ÜÈÈÓˆË˜\ÜÙ]ÚYHKšYˆÒT‘HZKœ›ÝšY\ˆHÂˆS‘ZK™[˜X›YHBˆS‘ÓÐSTÐÑJZK™^\›˜[ÚY	ÉÊHOH	ÉÂˆS‘ÓÐSTÐÑJK›[Ûš]Üš[™×ÜÝ]\Ë	ØXÝ]™IÊHOH	Ù\ØX›Y	ÂˆS‘ÓÐSTÐÑJË™[˜X›YJHHBˆÔ‘Tˆ–HKœ›Ú™XÝÛ˜[YHÓÓUH“ÐÐTÑBˆˆˆ‹ˆ
›ÝšY\‹
Kˆ
BˆYˆ›ÝX\YØ\ÜÙ]Î‚ˆ™]\›ˆÈœ›ØÙ\ÜÙYŽˆ›Z\ÜÚ[™×Ù]HŽˆ››×Ü™Y™\™[˜ÙHŽˆœ\š[ÙÙ]HŽˆ\™Ù]Ù]Kš\ÛÙ›Ü›X]

_B‚ˆ[™XYWÝ˜[YØ\ÜÙ]ÚYÈHÂˆ[
][VÈ˜\ÜÙ]ÚY—JBˆ›Üˆ][H[ˆ]Y\žWØ[
ˆÛÛ›‹ˆˆˆ‚ˆÑSPÕ\ÜÙ]ÚYˆ”“ÓH›ÙXÝ[Û—Ü™XÛÜ™ÂˆÒT‘H›ÝšY\ˆHÈS‘\š[ÙÝ\HHÈS‘\š[ÙÙ]HHÂˆS‘›ÙXÝ[Û—ÚÝÚTÈ“Õ•Sˆˆˆ‹ˆ
›ÝšY\‹\š[ÙÝ\K\™Ù]Ù]Kš\ÛÙ›Ü›X]

JKˆ
BˆBˆX\YØ\ÜÙ]ÈHÂˆ›ÝÂˆ›Üˆ›ÝÈ[ˆX\YØ\ÜÙ]ÂˆYˆ[
›ÝÖÈ˜\ÜÙ]ÚY—JH›Ý[ˆ[™XYWÝ˜[YØ\ÜÙ]ÚYÂˆBˆÝ][Û—ØÛÙ\ÈHÂˆÝŠ›ÝÖÈ™^\›˜[ÚY—JKœÝš\

Bˆ›Üˆ›ÝÈ[ˆX\YØ\ÜÙ]ÂˆYˆÝŠ›ÝÖÈ™^\›˜[ÚY—HÜˆˆŠKœÝš\

BˆBˆYˆ›ÝÝ][Û—ØÛÙ\Î‚ˆ™]\›ˆÂˆœ›ØÙ\ÜÙYŽˆˆ›Z\ÜÚ[™×Ù]HŽˆˆ››×Ü™Y™\™[˜ÙHŽˆˆœ\š[ÙÙ]HŽˆ\™Ù]Ù]Kš\ÛÙ›Ü›X]

KˆœÚÚ\YØÛÛ\]HŽˆ[Š[™XYWÝ˜[YØ\ÜÙ]ÚYÊKˆBˆÙ\ÜÚ[Û—ÛØš‹ÈHÙ]Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[ÛŠÛÛ™šYÊBˆ[™Ú[Ý\ÙYH
ˆ[™Ú[ÖÈ›[ÛÚÜWÙ[™Ú[—BˆYˆ\š[ÙÝ\HOH›[Û‚ˆ[ÙH[™Ú[ÖÈ™^WÚÜWÙ[™Ú[—Bˆ
BˆÜWÛX\ˆXÝÜÝ‹XÝÜÝ‹[žWWHHßBˆ™\]Y\ÝYÜÝ][Û—ØÛÙ\ÎˆÙ]ÜÝ—HHÙ]

BˆY™\œ™YÜÛÝÙ\œ›ÜŽˆ\TÛÝ[˜]˜Z[X›Q\œ›Üˆ›Û™HH›Û™Bˆ›ÜˆÝ][Û—ÙÜ›Ý\[ˆÚ[šÙY
Ý][Û—ØÛÙ\ËL
N‚ˆ™]žWØY\—Ü™[ÙÚ[ˆH˜[ÙBˆÚ[HYN‚ˆžN‚ˆYˆ\š[ÙÝ\HOH›[ÛŽ‚ˆÚ[š×ÛX\H™]ÚÙ\Ú[ÛœÛÛ\—ÚÜWÛ[ÛÛX\
ˆÙ\ÜÚ[Û—ÛØš‹ˆ[™Ú[ÖÈ˜˜\ÙWÝ\›—Kˆ[™Ú[Ý\ÙYˆÝ][Û—ÙÜ›Ý\ˆ\™Ù]Ù]Kˆ
Bˆ[ÙN‚ˆÚ[š×ÛX\H™]ÚÙ\Ú[ÛœÛÛ\—ÚÜWÙ^WÛX\
ˆÙ\ÜÚ[Û—ÛØš‹ˆ[™Ú[ÖÈ˜˜\ÙWÝ\›—Kˆ[™Ú[Ý\ÙYˆÝ][Û—ÙÜ›Ý\ˆ\™Ù]Ù]Kˆ
BˆÜWÛX\\]JÚ[š×ÛX\
Bˆ™\]Y\ÝYÜÝ][Û—ØÛÙ\Ë\]JÝ][Û—ÙÜ›Ý\
Bˆœ™XZÂˆ^Ù\\TÛÝ[˜]˜Z[X›Q\œ›Üˆ\È^Î‚ˆY™\œ™YÜÛÝÙ\œ›ÜˆH^Âˆœ™XZÂˆ^Ù\^Ù\[Ûˆ\È^Î‚ˆYˆ\×Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[Û—Ù^\™YÙ\œ›ÜŠ^ÊH[™›Ý™]žWØY\—Ü™[ÙÚ[Ž‚ˆÑÑÑT‹š[™›Ê‘\Ú[Û”ÛÛ\ˆ›ÙXÝ[ÛˆÙ\ÜÚ[Ûˆ^\™YÈ[˜[Y][™ÈØXÚH[™™]žZ[™ÈÙÚ[ˆÛ˜ÙHŠBˆ™]žWØY\—Ü™[ÙÚ[ˆHYBˆ[˜[Y]WÙ\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[ÛŠÛÛ™šYÊBˆÙ\ÜÚ[Û—ÛØš‹ÈHÙ]Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[ÛŠÛÛ™šYË›Ü˜ÙWÛÙÚ[UYJBˆÛÛ[YBˆYˆ\Ú[œÝ[˜ÙJ^Ë
\T˜]S[Z]\œ›Ü‹\TÛÝ[˜]˜Z[X›Q\œ›ÜŠJN‚ˆ˜Z\ÙBˆYˆ\×Ù\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]Ù\œ›ÜŠ^ÊN‚ˆ™X\ÛÛˆHX\š×Ù\Ú[ÛœÛÛ\—Ü\™›Ü›X[˜ÙWÜ˜]WÛ[Z]Y
ÛÛ›ŠBˆÛÛ›‹˜ÛÛ[Z]

Bˆ™]\›ˆÂˆœ›ØÙ\ÜÙYŽˆˆ›Z\ÜÚ[™×Ù]HŽˆˆ››×Ü™Y™\™[˜ÙHŽˆˆœ\š[ÙÙ]HŽˆ\™Ù]Ù]Kš\ÛÙ›Ü›X]

KˆœÝÜYÜ™X\ÛÛˆŽˆ™X\ÛÛ‹ˆBˆ˜Z\ÙBˆYˆY™\œ™YÜÛÝÙ\œ›Üˆ\È›Ý›Û™N‚ˆœ™XZÂˆ›ØÙ\ÜÙYHˆZ\ÜÚ[™×Ù]HHˆ›×Ü™Y™\™[˜ÙHHˆÚ]Ü›ÙXÝ[ÛˆHˆ›Üˆ›ÝÈ[ˆX\YØ\ÜÙ]Î‚ˆ\ÜÙ]ÚYH[
›ÝÖÈ˜\ÜÙ]ÚY—JBˆ^\›˜[ÚYHÝŠ›ÝÖÈ™^\›˜[ÚY—HÜˆˆŠKœÝš\

BˆYˆ^\›˜[ÚY›Ý[ˆ™\]Y\ÝYÜÝ][Û—ØÛÙ\Î‚ˆÛÛ[YBˆÜWÜ›ÝÈHÜWÛX\™Ù]
^\›˜[ÚYßJBˆ]WÚ][WÛX\HÜWÜ›ÝË™Ù]
™]R][SX\ŠHYˆ\Ú[œÝ[˜ÙJÜWÜ›ÝËXÝ
H[ÙHßBˆYˆ›Ý\Ú[œÝ[˜ÙJ]WÚ][WÛX\XÝ
N‚ˆ]WÚ][WÛX\HßBˆ›ÙXÝ[Û—ÚÝÚÙ[XÝYÚÙ^KÙ[XÝYÜ˜]×Ý˜[YHHÙ[XÝÜ›ÙXÝ[Û—Ý˜[YJ]WÚ][WÛX\
BˆYˆ›ÙXÝ[Û—ÚÝÚ\È›Ý›Û™N‚ˆÚ]Ü›ÙXÝ[Ûˆ
ÏHBˆÝÜH\œÙWÚÝÜÝ˜[YJ›ÝÖÈšÝÜ—JBˆÜXÚYšX×ÞZY[HØ[Ý[]WÜÜXÚYšX×ÞZY[
›ÙXÝ[Û—ÚÝÚÝÜ
BˆÙ][™ÜÈHÙ]Ü\™›Ü›X[˜ÙWÜÙ][™ÜÊÛÛ›‹\ÜÙ]ÚY
BˆÙ][™ÜË\]JÚÙ^Nˆ›ÝÖÚÙ^WH›ÜˆÙ^H[ˆ›ÝËšÙ^\Ê
HYˆÙ^H[ˆÙ][™ÜÈ[™›ÝÖÚÙ^WH\È›Ý›Û™_JB‚ˆ^XÝYÚÝÚ^XÝYÜÜXÚYšX×ÞZY[^XÝYÜÛÝ\˜ÙK˜\Ù[[™WÜ]X[]HHØ[Ý[]WÙ^XÝYÜ›ÙXÝ[ÛŠˆÛÛ›‹ˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆ›ÝšY\\›ÝšY\‹ˆ\š[ÙÝ\O\\š[ÙÝ\Kˆ\š[ÙÙ]O]\™Ù]Ù]KˆÝÜZÝÜˆÙ][™ÜÏ\Ù][™ÜËˆ
Bˆ\™›Ü›X[˜ÙWÜÝ]\Ë]WÜ]X[]K]šX][Û—ÜÝHÛ\ÜÚYžWÜ\™›Ü›X[˜ÙWÜÝ]\Êˆ›ÙXÝ[Û—ÚÝÚˆÝÜˆ^XÝYÚÝÚˆØ\›š[™×Ù]šX][Û—ÜÝY›Ø]
Ù][™ÜË™Ù]
Ø\›š[™×Ù]šX][Û—ÜÝŠHÜˆLL
Kˆ[\Ù]šX][Û—ÜÝY›Ø]
Ù][™ÜË™Ù]
˜[\Ù]šX][Û—ÜÝŠHÜˆLŒ
KˆÜš]XØ[Ù]šX][Û—ÜÝY›Ø]
Ù][™ÜË™Ù]
˜Üš]XØ[Ù]šX][Û—ÜÝŠHÜˆLÌ
Kˆ
BˆYˆ]WÜ]X[]HOH›ÚÈˆ[™˜\Ù[[™WÜ]X[]HOHœ\X[Ú\ÝÜžHˆ[™^XÝYÜÛÝ\˜ÙHOH››Û™HŽ‚ˆ]WÜ]X[]HHœ\X[Ú\ÝÜžH‚ˆYˆ\™›Ü›X[˜ÙWÜÝ]\ÈOH”Ù[HYÜÈŽ‚ˆZ\ÜÚ[™×Ù]H
ÏHBˆYˆ\™›Ü›X[˜ÙWÜÝ]\ÈOH”Ù[H™Y™\°ê›˜ÚXHŽ‚ˆ›×Ü™Y™\™[˜ÙH
ÏHB‚ˆ›Ý\×Ü\ÈH×BˆYˆ›ÙXÝ[Û—ÚÝÚ\È›Û™N‚ˆ›Ý\×Ü\Ë˜\[™
ˆZ[ÛZ\ÜÚ[™×Ü›ÙXÝ[Û—Û›ÝJˆ]WÚ][WÛX\ˆÝ][Û—ØÛÙOY^\›˜[ÚYˆ\š[ÙÝ\O\\š[ÙÝ\Kˆ\š[ÙÙ]O]\™Ù]Ù]Kˆ
Bˆ
BˆYˆÝÜ\È›Û™N‚ˆ›Ý\×Ü\Ë˜\[™
šÕÜØØ[[H˜[HÝH[˜[YËˆŠBˆYˆ^XÝYÜÛÝ\˜ÙHOH››Û™HŽ‚ˆ›Ý\×Ü\Ë˜\[™
”Ù[H\Ý0ìÜšXÛÈÝHÜ°éØ[Y[ÈY[œØ[\˜H™Y™\°ê›˜ÚXKˆŠBˆ\Ù\Ü›ÙXÝ[Û—Ü™XÛÜ™
ˆÛÛ›‹ˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆ›ÝšY\\›ÝšY\‹ˆ^\›˜[ÚYY^\›˜[ÚYˆ\š[ÙÝ\O\\š[ÙÝ\Kˆ\š[ÙÙ]O]\™Ù]Ù]Kˆ›ÙXÝ[Û—ÚÝÚ\›ÙXÝ[Û—ÚÝÚˆÜXÚYšX×ÞZY[\ÜXÚYšX×ÞZY[ˆ^XÝYÚÝÚY^XÝYÚÝÚˆ^XÝYÜÜXÚYšX×ÞZY[Y^XÝYÜÜXÚYšX×ÞZY[ˆ]šX][Û—ÜÝY]šX][Û—ÜÝˆ\™›Ü›X[˜ÙWÜÝ]\Ï\\™›Ü›X[˜ÙWÜÝ]\Ëˆ^XÝYÜÛÝ\˜ÙOY^XÝYÜÛÝ\˜ÙKˆ]WÜ]X[]OY]WÜ]X[]Kˆ›Ý\ÏHˆ‹š›Ú[Š›Ý\×Ü\ÊKˆ^[ØYÚœÛÛ\ÝŠÜWÜ›ÝË™Ù]
œ^[ØYÚœÛÛˆŠHÜˆœÛÛ‹™[\ÊÜWÜ›ÝË[œÝ\™WØ\ØÚZOUYJJKˆÙ[XÝYÜ›ÙXÝ[Û—ÚÙ^O\Ù[XÝYÚÙ^KˆÙ[XÝYÜ›ÙXÝ[Û—Ü˜]×Ý˜[YO\Ù[XÝYÜ˜]×Ý˜[YKˆ
Bˆ›ØÙ\ÜÙY
ÏHB‚ˆ™XØ[Ý[]WÜ\™›Ü›X[˜ÙWÜ™Y™\™[˜Ù\ÊˆÛÛ›‹ˆ\š[ÙÝ\O\\š[ÙÝ\Kˆ\š[ÙÙ]O]\™Ù]Ù]Kˆ›ÝšY\\›ÝšY\‹ˆ
BˆÙÙÙ\ˆHÝ\œ™[Ø\›ÙÙÙ\ˆYˆ\×Ø\ØÛÛ^

H[ÙHÙÙÚ[™Ë™Ù]ÙÙÙ\Š×Û˜[YW×ÊBˆÙÙÙ\‹š[™›Êˆ‘\Ú[Û”ÛÛ\ˆ\™›Ü›X[˜ÙHÞ[˜Îˆ™\]Y\ÝYÜÝ][Û—ØÛÙ\ÏI\È[™Ú[I\È\š[ÙÝ\OI\È\™Ù]Ù]OI\È\WÜ›ÝÜÏI\ÈÚ]Ü›ÙXÝ[ÛI\ÈZ\ÜÚ[™×Ü›ÙXÝ[ÛI\È‹ˆ[ŠÝ][Û—ØÛÙ\ÊKˆ[™Ú[Ý\ÙYˆ\š[ÙÝ\Kˆ\™Ù]Ù]Kš\ÛÙ›Ü›X]

Kˆ[ŠÜWÛX\
KˆÚ]Ü›ÙXÝ[Û‹ˆZ\ÜÚ[™×Ù]Kˆ
BˆÛÛ›‹˜ÛÛ[Z]

BˆYˆY™\œ™YÜÛÝÙ\œ›Üˆ\È›Ý›Û™N‚ˆ˜Z\ÙHY™\œ™YÜÛÝÙ\œ›Ü‚ˆ™]\›ˆÂˆœ›ØÙ\ÜÙYŽˆ›ØÙ\ÜÙYˆ›Z\ÜÚ[™×Ù]HŽˆZ\ÜÚ[™×Ù]Kˆ››×Ü™Y™\™[˜ÙHŽˆ›×Ü™Y™\™[˜ÙKˆœ\š[ÙÙ]HŽˆ\™Ù]Ù]Kš\ÛÙ›Ü›X]

KˆB‚‚™Yˆ]\—ÙZ[WØ˜XÚÙš[Ù]\Êœ›ÛWÞYX\Žˆ[×ÞYX\Žˆ[
‹Ù^WÝ˜[YNˆ]H›Û™HH›Û™JHOˆ\ÝÙ]WN‚ˆÙ^WÝ˜[YHHÙ^WÝ˜[YHÜˆ]KÙ^J
BˆÝ\œÛÜˆH]Jœ›ÛWÞYX\‹KJBˆ[™Ù]HHZ[Š]J×ÞYX\‹L‹ÌJKÙ^WÝ˜[YHH[YY[J^\ÏLJJBˆ^\Îˆ\ÝÙ]WHH×BˆÚ[HÝ\œÛÜˆH[™Ù]N‚ˆ^\Ë˜\[™
Ý\œÛÜŠBˆÝ\œÛÜˆ
ÏH[YY[J^\ÏLJBˆ™]\›ˆ^\Â‚‚™Yˆ]\—ÙZ[WØ˜XÚÙš[Û[ÛÊˆœ›ÛWÞYX\Žˆ[ˆ×ÞYX\Žˆ[ˆ
‹ˆÙ^WÝ˜[YNˆ]H›Û™HH›Û™Kˆ]WÙœ›ÛNˆ]H›Û™HH›Û™Kˆ]WÝÎˆ]H›Û™HH›Û™KŠHOˆ\ÝÙ]WN‚ˆÙ^WÝ˜[YHHÙ^WÝ˜[YHÜˆ]KÙ^J
BˆÝ\Ù]HH]WÙœ›ÛHÜˆ]Jœ›ÛWÞYX\‹KJBˆ[™Ù]HH]WÝÈÜˆ]J×ÞYX\‹L‹ÌJBˆ[™Ù]HHZ[Š[™Ù]KÙ^WÝ˜[YHH[YY[J^\ÏLJJBˆYˆÝ\Ù]Hˆ[™Ù]N‚ˆ™]\›ˆ×BˆÝ\œÛÜˆHÝ\Ù]Kœ™\XÙJ^OLJBˆ[™Û[ÛH[™Ù]Kœ™\XÙJ^OLJBˆ[ÛÎˆ\ÝÙ]WHH×BˆÚ[HÝ\œÛÜˆH[™Û[Û‚ˆ[ÛË˜\[™
Ý\œÛÜŠBˆYX\ˆHÝ\œÛÜ‹žYX\ˆ
È
HYˆÝ\œÛÜ‹›[ÛOHLˆ[ÙH
Bˆ[ÛHHYˆÝ\œÛÜ‹›[ÛOHLˆ[ÙHÝ\œÛÜ‹›[Û
ÈBˆÝ\œÛÜˆH]JYX\‹[ÛJBˆ™]\›ˆ[ÛÂ‚‚™Yˆ]WÚ[—Ø˜XÚÙš[ÝÚ[™ÝÊˆØ[™Y]Nˆ]Kˆ
‹ˆœ›ÛWÞYX\Žˆ[ˆ×ÞYX\Žˆ[ˆÙ^WÝ˜[YNˆ]Kˆ]WÙœ›ÛNˆ]H›Û™HH›Û™Kˆ]WÝÎˆ]H›Û™HH›Û™KŠHOˆ›ÛÛ‚ˆÝ\Ù]HH]WÙœ›ÛHÜˆ]Jœ›ÛWÞYX\‹KJBˆ[™Ù]HH]WÝÈÜˆ]J×ÞYX\‹L‹ÌJBˆ[™Ù]HHZ[Š[™Ù]KÙ^WÝ˜[YHH[YY[J^\ÏLJJBˆ™]\›ˆÝ\Ù]HHØ[™Y]HH[™Ù]B‚‚™Yˆ]\—Û[ÛWØ˜XÚÙš[Ù]\Êœ›ÛWÞYX\Žˆ[×ÞYX\Žˆ[
‹Ù^WÝ˜[YNˆ]H›Û™HH›Û™JHOˆ\ÝÙ]WN‚ˆÙ^WÝ˜[YHHÙ^WÝ˜[YHÜˆ]KÙ^J
BˆÝ\œ™[Û[ÛHÙ^WÝ˜[YKœ™\XÙJ^OLJBˆ[ÛÎˆ\ÝÙ]WHH×Bˆ›ÜˆYX\ˆ[ˆ˜[™ÙJœ›ÛWÞYX\‹×ÞYX\ˆ
ÈJN‚ˆ›Üˆ[Û[ˆ˜[™ÙJKLÊN‚ˆ\š[ÙÙ]HH]JYX\‹[ÛJBˆYˆ\š[ÙÙ]HHÝ\œ™[Û[Û‚ˆÛÛ[YBˆ[ÛË˜\[™
\š[ÙÙ]JBˆ™]\›ˆ[ÛÂ‚‚™YˆÙ]Ù\Ú[ÛœÛÛ\—Ü\™›Ü›X[˜ÙWØ\ÜÙ]ÊˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ›ÝšY\ŽˆÝ‹ˆ\ÜÙ]ÚYˆ[›Û™HH›Û™Kˆ\ÜÙ]ÚYÎˆ\ÝÚ[H›Û™HH›Û™KŠHOˆ\ÝÜÜ[]LË”›Ý×N‚ˆÛÛ™][ÛœÈHÂˆ˜ZKœ›ÝšY\ˆHÈ‹ˆ˜ZK™[˜X›YHH‹ˆÓÐSTÐÑJZK™^\›˜[ÚY	ÉÊHOH	ÉÈ‹ˆÓÐSTÐÑJK›[Ûš]Üš[™×ÜÝ]\Ë	ØXÝ]™IÊHOH	Ù\ØX›Y	È‹ˆÓÐSTÐÑJË™[˜X›YJHHH‹ˆBˆ\˜[\Îˆ\ÝÐ[žWHHÜ›ÝšY\—BˆYˆ\ÜÙ]ÚY‚ˆÛÛ™][ÛœË˜\[™
˜KšYHÈŠBˆ\˜[\Ë˜\[™
\ÜÙ]ÚY
BˆYˆ\ÜÙ]ÚYÎ‚ˆXÙZÛ\œÈH‹‹š›Ú[ŠÈˆ›ÜˆÈ[ˆ\ÜÙ]ÚYÊBˆÛÛ™][ÛœË˜\[™
ˆ˜KšYSˆ
ÜXÙZÛ\œßJHŠBˆ\˜[\Ë™^[™
\ÜÙ]ÚYÊBˆ™]\›ˆ]Y\žWØ[
ˆÛÛ›‹ˆˆˆˆ‚ˆÑSPÕˆKšYTÈ\ÜÙ]ÚYˆKœ›Ú™XÝÛ˜[YKˆKšÝÜˆZK™^\›˜[ÚYˆÓÐSTÐÑJË™[˜X›YJHTÈ\™›Ü›X[˜ÙWÙ[˜X›YˆËØ\›š[™×Ù]šX][Û—ÜÝˆË˜[\Ù]šX][Û—ÜÝˆË˜Üš]XØ[Ù]šX][Û—ÜÝˆË˜˜\Ù[[™WÞYX\œËˆË›Z[—Ø˜\Ù[[™WÜÚ[ËˆË›[ÛWØYÙ]ÚœÛÛ‚ˆ”“ÓH\ÜÙ]Ú[YÜ˜][ÛœÈZBˆ“ÒSˆ\ÜÙ]ÈHÓˆKšYHZK˜\ÜÙ]ÚYˆQ•“ÒSˆ\™›Ü›X[˜ÙWÜÙ][™ÜÈÈÓˆË˜\ÜÙ]ÚYHKšYˆÒT‘HÈˆS‘‹š›Ú[ŠÛÛ™][ÛœÊ_BˆÔ‘Tˆ–HKœ›Ú™XÝÛ˜[YHÓÓUH“ÐÐTÑBˆˆˆ‹ˆ\˜[\Ëˆ
B‚‚™Yˆ™XØ[Ý[]WÜ›ÙXÝ[Û—Ù^XÝ][ÛœÊˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆ›ÝšY\ŽˆÝ‹ˆ\ÜÙ]ÚYÎˆ\ÝÚ[Kˆ\š[ÙÝ\NˆÝˆ›Û™HH›Û™KŠHOˆ[‚ˆYˆ›Ý\ÜÙ]ÚYÎ‚ˆ™]\›ˆˆÝ[Hˆ›Üˆ\ÜÙ]ÚY[ˆ\ÜÙ]ÚYÎ‚ˆÝ[[X\žHH™XØ[Ý[]WÜ\™›Ü›X[˜ÙWÜ™Y™\™[˜Ù\ÊˆÛÛ›‹ˆ\š[ÙÝ\O\\š[ÙÝ\Kˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆ›ÝšY\\›ÝšY\‹ˆ
BˆÝ[
ÏHÝ[[X\žVÈœ™XÛÜ™×Ü›ØÙ\ÜÙY—Bˆ™]\›ˆÝ[‚‚™Yˆ™XØ[Ý[]WÜ\™›Ü›X[˜ÙWÜ™Y™\™[˜Ù\ÊˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ\š[ÙÝ\NˆÝˆ›Û™HH›Û™Kˆ\š[ÙÙ]Nˆ]HÝˆ›Û™HH›Û™Kˆ\ÜÙ]ÚYˆ[›Û™HH›Û™Kˆ›ÝšY\ŽˆÝˆH‘\Ú[Û”ÛÛ\ˆ‹ˆÙ^WÝ˜[YNˆ]H›Û™HH›Û™KŠHOˆXÝÜÝ‹[N‚ˆÛÛ™][ÛœÈHÈœ‹œ›ÝšY\ˆHÈ—Bˆ\˜[\Îˆ\ÝÐ[žWHHÜ›ÝšY\—BˆYˆ\š[ÙÝ\N‚ˆÛÛ™][ÛœË˜\[™
œ‹œ\š[ÙÝ\HHÈŠBˆ\˜[\Ë˜\[™
\š[ÙÝ\JBˆYˆ\š[ÙÙ]N‚ˆ›Ü›X[^™YÙ]HH\š[ÙÙ]Kš\ÛÙ›Ü›X]

HYˆ\Ú[œÝ[˜ÙJ\š[ÙÙ]K]JH[ÙHÝŠ\š[ÙÙ]JBˆÛÛ™][ÛœË˜\[™
œ‹œ\š[ÙÙ]HHÈŠBˆ\˜[\Ë˜\[™
›Ü›X[^™YÙ]JBˆYˆ\ÜÙ]ÚY‚ˆÛÛ™][ÛœË˜\[™
œ‹˜\ÜÙ]ÚYHÈŠBˆ\˜[\Ë˜\[™
\ÜÙ]ÚY
Bˆ›ÝÜÈH]Y\žWØ[
ˆÛÛ›‹ˆˆˆˆ‚ˆÑSPÕˆ‹Š‹ˆKœ›Ú™XÝÛ˜[YKˆKšÝÜˆËØ\›š[™×Ù]šX][Û—ÜÝˆË˜[\Ù]šX][Û—ÜÝˆË˜Üš]XØ[Ù]šX][Û—ÜÝˆË˜˜\Ù[[™WÞYX\œËˆË›Z[—Ø˜\Ù[[™WÜÚ[ËˆË›[ÛWØYÙ]ÚœÛÛ‚ˆ”“ÓH›ÙXÝ[Û—Ü™XÛÜ™È‚ˆ“ÒSˆ\ÜÙ]ÈHÓˆKšYH‹˜\ÜÙ]ÚYˆQ•“ÒSˆ\™›Ü›X[˜ÙWÜÙ][™ÜÈÈÓˆË˜\ÜÙ]ÚYH‹˜\ÜÙ]ÚYˆÒT‘HÈˆS‘‹š›Ú[ŠÛÛ™][ÛœÊ_BˆÔ‘Tˆ–H‹œ\š[ÙÙ]HTÐË‹šYTÐÂˆˆˆ‹ˆ\˜[\Ëˆ
BˆÝ[[X\žHHÂˆœ™XÛÜ™×Ü›ØÙ\ÜÙYŽˆˆœ™Y™\™[˜Ù\×ØÜ™X]YŽˆˆœÝ[ÝÚ]Ý]Ü™Y™\™[˜ÙHŽˆˆ›Z\ÜÚ[™×ÚÝÜŽˆˆ›Z\ÜÚ[™×Ü›ÙXÝ[ÛˆŽˆˆBˆ›Üˆ›ÝÈ[ˆ›ÝÜÎ‚ˆ\™Ù]Ù]HH\œÙWÙ]WÝ˜[YJ›ÝÖÈœ\š[ÙÙ]H—JBˆYˆ\™Ù]Ù]H\È›Û™N‚ˆÛÛ[YBˆÝÜH\œÙWÚÝÜÝ˜[YJ›ÝÖÈšÝÜ—JBˆÙ][™ÜÈHÙ]Ü\™›Ü›X[˜ÙWÜÙ][™ÜÊÛÛ›‹[
›ÝÖÈ˜\ÜÙ]ÚY—JJBˆÙ][™ÜË\]JÚÙ^Nˆ›ÝÖÚÙ^WH›ÜˆÙ^H[ˆ›ÝËšÙ^\Ê
HYˆÙ^H[ˆÙ][™ÜÈ[™›ÝÖÚÙ^WH\È›Ý›Û™_JBˆ™Y™\™[˜ÙWÜ™\Ý[HØ[Ý[]WÙ^XÝYÜ›ÙXÝ[Û—ÝÚ]ÙXYÛ›ÜÝXÊˆÛÛ›‹ˆ\ÜÙ]ÚYZ[
›ÝÖÈ˜\ÜÙ]ÚY—JKˆ›ÝšY\\›ÝšY\‹ˆ\š[ÙÝ\O\›ÝÖÈœ\š[ÙÝ\H—Kˆ\š[ÙÙ]O]\™Ù]Ù]KˆÝÜZÝÜˆÙ][™ÜÏ\Ù][™ÜËˆ\ÜÙ]Û˜[YO\›ÝÖÈœ›Ú™XÝÛ˜[YH—KˆÙ^WÝ˜[YO]Ù^WÝ˜[YKˆ
Bˆ^XÝYÚÝÚH™Y™\™[˜ÙWÜ™\Ý[È™^XÝYÚÝÚ—Bˆ^XÝYÜÜXÚYšX×ÞZY[H™Y™\™[˜ÙWÜ™\Ý[È™^XÝYÜÜXÚYšX×ÞZY[—Bˆ^XÝYÜÛÝ\˜ÙHH™Y™\™[˜ÙWÜ™\Ý[È™^XÝYÜÛÝ\˜ÙH—Bˆ\™›Ü›X[˜ÙWÜÝ]\Ë]WÜ]X[]K]šX][Û—ÜÝHÛ\ÜÚYžWÜ\™›Ü›X[˜ÙWÜÝ]\Êˆ›ÝÖÈœ›ÙXÝ[Û—ÚÝÚ—KˆÝÜˆ^XÝYÚÝÚˆØ\›š[™×Ù]šX][Û—ÜÝY›Ø]
Ù][™ÜË™Ù]
Ø\›š[™×Ù]šX][Û—ÜÝŠHÜˆLL
Kˆ[\Ù]šX][Û—ÜÝY›Ø]
Ù][™ÜË™Ù]
˜[\Ù]šX][Û—ÜÝŠHÜˆLŒ
KˆÜš]XØ[Ù]šX][Û—ÜÝY›Ø]
Ù][™ÜË™Ù]
˜Üš]XØ[Ù]šX][Û—ÜÝŠHÜˆLÌ
Kˆ
BˆYˆ]WÜ]X[]HOH›ÚÈˆ[™™Y™\™[˜ÙWÜ™\Ý[Èœ]X[]H—HOHœ\X[Ú\ÝÜžHˆ[™^XÝYÜÛÝ\˜ÙHOH››Û™HŽ‚ˆ]WÜ]X[]HHœ\X[Ú\ÝÜžH‚ˆ›Ý\ÈH›ÝÖÈ››Ý\È—HÜˆˆ‚ˆ™X\ÛÛˆH™Y™\™[˜ÙWÜ™\Ý[È™XYÛ›ÜÝXÈ—K™Ù]
››×Ü™Y™\™[˜ÙWÜ™X\ÛÛˆŠHÜˆˆ‚ˆYˆ™X\ÛÛˆ[™™X\ÛÛˆ›Ý[ˆ›Ý\Î‚ˆ›Ý\ÈHˆžÛ›Ý\ßHÜ™X\ÛÛŸH‹œÝš\

BˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUH›ÙXÝ[Û—Ü™XÛÜ™ÂˆÑU^XÝYÚÝÚHË^XÝYÜÜXÚYšX×ÞZY[HË]šX][Û—ÜÝHËˆ\™›Ü›X[˜ÙWÜÝ]\ÈHË^XÝYÜÛÝ\˜ÙHHË]WÜ]X[]HHËˆ™Y™\™[˜ÙWÙXYÛ›ÜÝX×ÚœÛÛˆHË›Ý\ÈHË\]YØ]HÂˆÒT‘HYHÂˆˆˆ‹ˆ
ˆ^XÝYÚÝÚˆ^XÝYÜÜXÚYšX×ÞZY[ˆ]šX][Û—ÜÝˆ\™›Ü›X[˜ÙWÜÝ]\Ëˆ^XÝYÜÛÝ\˜ÙKˆ]WÜ]X[]KˆœÛÛ‹™[\Ê™Y™\™[˜ÙWÜ™\Ý[È™XYÛ›ÜÝXÈ—K[œÝ\™WØ\ØÚZOUYJKˆ›Ý\Ëˆ]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆ›ÝÖÈšY—Kˆ
Kˆ
BˆÝ[[X\žVÈœ™XÛÜ™×Ü›ØÙ\ÜÙY—H
ÏHBˆYˆ^XÝYÚÝÚ\È›Ý›Û™H[™^XÝYÜÜXÚYšX×ÞZY[\È›Ý›Û™N‚ˆÝ[[X\žVÈœ™Y™\™[˜Ù\×ØÜ™X]Y—H
ÏHBˆ[ÙN‚ˆÝ[[X\žVÈœÝ[ÝÚ]Ý]Ü™Y™\™[˜ÙH—H
ÏHBˆYˆÝÜ\È›Û™N‚ˆÝ[[X\žVÈ›Z\ÜÚ[™×ÚÝÜ—H
ÏHBˆYˆ›ÝÖÈœ›ÙXÝ[Û—ÚÝÚ—H\È›Û™N‚ˆÝ[[X\žVÈ›Z\ÜÚ[™×Ü›ÙXÝ[Ûˆ—H
ÏHBˆÛÛ›‹˜ÛÛ[Z]

Bˆ™]\›ˆÝ[[X\žB‚‚™YˆÜ[—Ù\Ú[ÛœÛÛ\—Ü›ÙXÝ[Û—Ø˜XÚÙš[ÛYØXÞJˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆ›ÝšY\ŽˆÝˆH‘\Ú[Û”ÛÛ\ˆ‹ˆ\š[ÙÝ\NˆÝˆH™^H‹ˆœ›ÛWÞYX\Žˆ[ˆ×ÞYX\Žˆ[ˆ\ÜÙ]ÚYˆ[›Û™HH›Û™KˆÙ^WÝ˜[YNˆ]H›Û™HH›Û™Kˆ]WÙœ›ÛNˆ]H›Û™HH›Û™Kˆ]WÝÎˆ]H›Û™HH›Û™KˆX^Ù^\Îˆ[›Û™HH›Û™KŠHOˆXÝÜÝ‹[žWN‚ˆYˆ\š[ÙÝ\H›Ý[ˆÈ™^H‹›[ÛŸN‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ•\ÈH\°ë[ÙÈ[°è[YËˆŠBˆYˆœ›ÛWÞYX\ˆˆ×ÞYX\Ž‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ[›È[šXÚX[˜[ÈÙHÙ\ˆÝ\\š[Üˆ[È[›Èš[˜[ˆŠBˆYˆ
×ÞYX\ˆHœ›ÛWÞYX\ˆ
ÈJHˆÎ‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ’[\˜[ÈÝ\\š[ÜˆHÈ[›ÜËˆ™Y^ˆÈ\°ë[ÙÈ\˜H^XÝ]\ˆÈ˜XÚÙš[ˆŠB‚ˆÙ^WÝ˜[YHHÙ^WÝ˜[YHÜˆ]KÙ^J
BˆÛÛ™šYÈHÙ]Ú[YÜ˜][Û—ØÛÛ™šYÊÛÛ›‹›ÝšY\ŠBˆYˆÛÛ™šYÈ\È›Û™N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠÛÛ™šYÝ\˜XØ[È\Ú[Û”ÛÛ\ˆ˜[È[˜ÛÛ˜YKˆŠBˆYˆ›ÝÛÛ™šYÖÈ™[˜X›Y—N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠH[YÜ˜XØ[È\Ú[Û”ÛÛ\ˆ\ÝH\Ø]]˜YKˆŠBˆ[™Ú[ÈHÙ]Ù\Ú[ÛœÛÛ\—Ù[™Ú[ØÛÛ™šYÊÛÛ™šYÊBˆ\ÜÙ]ÈHÙ]Ù\Ú[ÛœÛÛ\—Ü\™›Ü›X[˜ÙWØ\ÜÙ]ÊÛÛ›‹›ÝšY\‹\ÜÙ]ÚY
Bˆ]\ÈH
ˆ]\—ÙZ[WØ˜XÚÙš[Ù]\Êœ›ÛWÞYX\‹×ÞYX\‹Ù^WÝ˜[YO]Ù^WÝ˜[YJBˆYˆ\š[ÙÝ\HOH™^H‚ˆ[ÙH]\—Û[ÛWØ˜XÚÙš[Ù]\Êœ›ÛWÞYX\‹×ÞYX\‹Ù^WÝ˜[YO]Ù^WÝ˜[YJBˆ
BˆYˆ\š[ÙÝ\HOH™^HŽ‚ˆYˆ]WÙœ›ÛH\È›Ý›Û™N‚ˆ]\ÈHØØ[™Y]H›ÜˆØ[™Y]H[ˆ]\ÈYˆØ[™Y]HH]WÙœ›ÛWBˆYˆ]WÝÈ\È›Ý›Û™N‚ˆ]\ÈHØØ[™Y]H›ÜˆØ[™Y]H[ˆ]\ÈYˆØ[™Y]HH]WÝ×BˆYˆX^Ù^\È\È›Ý›Û™H[™X^Ù^\Èˆ‚ˆ]\ÈH]\ÖÎ›X^Ù^\×B‚ˆÝ[[X\žHHÂˆ˜\ÜÙ]×Ü›ØÙ\ÜÙYŽˆˆœ™XÛÜ™×Ý\]YŽˆˆ›Z\ÜÚ[™×Ü›ÙXÝ[ÛˆŽˆˆ˜\WÙ\œ›ÜœÈŽˆˆ›]Ü™XÛÜ™×Ý\]YŽˆˆ˜˜\Ù[[™\×Ü™XØ[Ý[]YŽˆˆœ™Y™\™[˜Ù\×ØÜ™X]YŽˆˆœÝ[ÝÚ]Ý]Ü™Y™\™[˜ÙHŽˆˆBˆÙÙÙ\ˆHÝ\œ™[Ø\›ÙÙÙ\ˆYˆ\×Ø\ØÛÛ^

H[ÙHÙÙÚ[™Ë™Ù]ÙÙÙ\Š×Û˜[YW×ÊBˆÙ\ÜÚ[Û—ÛØš‹ÈHÙ]Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[ÛŠÛÛ™šYÊB‚ˆ›Üˆ\ÜÙ][ˆ\ÜÙ]Î‚ˆÝ[[X\žVÈ˜\ÜÙ]×Ü›ØÙ\ÜÙY—H
ÏHBˆ^\›˜[ÚYHÝŠ\ÜÙ]È™^\›˜[ÚY—HÜˆˆŠKœÝš\

Bˆ›Üˆ\š[ÙÙ]H[ˆ]\Î‚ˆžN‚ˆYˆ\š[ÙÝ\HOH›[ÛŽ‚ˆÜWÛX\H™]ÚÙ\Ú[ÛœÛÛ\—ÚÜWÛ[ÛÛX\
ˆÙ\ÜÚ[Û—ÛØš‹ˆ[™Ú[ÖÈ˜˜\ÙWÝ\›—Kˆ[™Ú[ÖÈ›[ÛÚÜWÙ[™Ú[—KˆÙ^\›˜[ÚYKˆ\š[ÙÙ]Kˆ
Bˆ[ÙN‚ˆÜWÛX\H™]ÚÙ\Ú[ÛœÛÛ\—ÚÜWÙ^WÛX\
ˆÙ\ÜÚ[Û—ÛØš‹ˆ[™Ú[ÖÈ˜˜\ÙWÝ\›—Kˆ[™Ú[ÖÈ™^WÚÜWÙ[™Ú[—KˆÙ^\›˜[ÚYKˆ\š[ÙÙ]Kˆ
Bˆ™\Ý[HÝÜ™WÜ›ÙXÝ[Û—ÚÜWÜ™XÛÜ™
ˆÛÛ›‹ˆ\ÜÙ]Ü›ÝÏX\ÜÙ]ˆ›ÝšY\\›ÝšY\‹ˆ^\›˜[ÚYY^\›˜[ÚYˆ\š[ÙÝ\O\\š[ÙÝ\Kˆ\š[ÙÙ]O\\š[ÙÙ]KˆÜWÜ›ÝÏZÜWÛX\™Ù]
^\›˜[ÚYßJKˆ›Ý\×Ü™Yš^H˜XÚÙš[\Ý0ìÜšXÛËˆ‹ˆ
BˆÝ[[X\žVÈœ™XÛÜ™×Ý\]Y—H
ÏHBˆYˆ™\Ý[Èœ›ÙXÝ[Û—ÚÝÚ—H\È›Û™N‚ˆÝ[[X\žVÈ›Z\ÜÚ[™×Ü›ÙXÝ[Ûˆ—H
ÏHBˆ^Ù\^Ù\[Ûˆ\È^Î‚ˆÝ[[X\žVÈ˜\WÙ\œ›ÜœÈ—H
ÏHBˆYˆ\Ú[œÝ[˜ÙJ^Ë\TÛÝ[˜]˜Z[X›Q\œ›ÜŠN‚ˆ˜Z\ÙBˆÙÙÙ\‹Ø\›š[™Êˆ‘\Ú[Û”ÛÛ\ˆ\™›Ü›X[˜ÙH˜XÚÙš[˜Z[Yˆ\ÜÙ]ÚYI\ÈÝ][ÛÛÙOI\È\š[ÙÝ\OI\È\š[ÙÙ]OI\È\œ›ÜI\È‹ˆ\ÜÙ]È˜\ÜÙ]ÚY—Kˆ^\›˜[ÚYˆ\š[ÙÝ\Kˆ\š[ÙÙ]Kš\ÛÙ›Ü›X]

Kˆ^Ëˆ
BˆÛÛ[YB‚ˆÙ[XÝYØ\ÜÙ]ÚYÈHÚ[
\ÜÙ]È˜\ÜÙ]ÚY—JH›Üˆ\ÜÙ][ˆ\ÜÙ]×Bˆ™XØ[×Ý\™Ù]ÈH]\ÈYˆ\š[ÙÝ\HOH›[Ûˆ[ÙH
ÛX^
]\ÊWHYˆ]\È[ÙH×JBˆYˆ›ÝÝ[[X\žVÈœÝÜYÜ™X\ÛÛˆ—N‚ˆ›Üˆ\™Ù][ˆ™XØ[×Ý\™Ù]Î‚ˆ™XØ[ÈH™XØ[Ý[]WÜ\™›Ü›X[˜ÙWÜ™Y™\™[˜Ù\ÊˆÛÛ›‹ˆ\š[ÙÝ\O\\š[ÙÝ\Kˆ\š[ÙÙ]O]\™Ù]ˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆ›ÝšY\\›ÝšY\‹ˆÙ^WÝ˜[YO]Ù^WÝ˜[YKˆ
BˆÝ[[X\žVÈ˜˜\Ù[[™\×Ü™XØ[Ý[]Y—H
ÏH™XØ[ÖÈœ™XÛÜ™×Ü›ØÙ\ÜÙY—BˆÝ[[X\žVÈœ™Y™\™[˜Ù\×ØÜ™X]Y—H
ÏH™XØ[ÖÈœ™Y™\™[˜Ù\×ØÜ™X]Y—BˆÝ[[X\žVÈœÝ[ÝÚ]Ý]Ü™Y™\™[˜ÙH—H
ÏH™XØ[ÖÈœÝ[ÝÚ]Ý]Ü™Y™\™[˜ÙH—B‚ˆÝ\œ™[Û[ÛHÙ^WÝ˜[YKœ™\XÙJ^OLJBˆYˆÙ[XÝYØ\ÜÙ]ÚYÎ‚ˆ›Üˆ\ÜÙ][ˆ\ÜÙ]Î‚ˆ^\›˜[ÚYHÝŠ\ÜÙ]È™^\›˜[ÚY—HÜˆˆŠKœÝš\

BˆžN‚ˆÜWÛX\H™]ÚÙ\Ú[ÛœÛÛ\—ÚÜWÛ[ÛÛX\
ˆÙ\ÜÚ[Û—ÛØš‹ˆ[™Ú[ÖÈ˜˜\ÙWÝ\›—Kˆ[™Ú[ÖÈ›[ÛÚÜWÙ[™Ú[—KˆÙ^\›˜[ÚYKˆÝ\œ™[Û[Ûˆ
BˆÝÜ™WÜ›ÙXÝ[Û—ÚÜWÜ™XÛÜ™
ˆÛÛ›‹ˆ\ÜÙ]Ü›ÝÏX\ÜÙ]ˆ›ÝšY\\›ÝšY\‹ˆ^\›˜[ÚYY^\›˜[ÚYˆ\š[ÙÝ\OH›]‹ˆ\š[ÙÙ]OXÝ\œ™[Û[ÛˆÜWÜ›ÝÏZÜWÛX\™Ù]
^\›˜[ÚYßJKˆ›Ý\×Ü™Yš^H“U™XØ[Ý[YÈ\0ìÜÈ˜XÚÙš[ˆ‹ˆ
BˆÝ[[X\žVÈ›]Ü™XÛÜ™×Ý\]Y—H
ÏHBˆ^Ù\^Ù\[Ûˆ\È^Î‚ˆÝ[[X\žVÈ˜\WÙ\œ›ÜœÈ—H
ÏHBˆYˆ\Ú[œÝ[˜ÙJ^Ë\TÛÝ[˜]˜Z[X›Q\œ›ÜŠN‚ˆ˜Z\ÙBˆÙÙÙ\‹Ø\›š[™Êˆ‘\Ú[Û”ÛÛ\ˆU™XØ[Ý[][Ûˆ˜Z[YY\ˆ˜XÚÙš[ˆ\ÜÙ]ÚYI\ÈÝ][ÛÛÙOI\È\œ›ÜI\È‹ˆ\ÜÙ]È˜\ÜÙ]ÚY—Kˆ^\›˜[ÚYˆ^Ëˆ
Bˆ™XØ[ÈH™XØ[Ý[]WÜ\™›Ü›X[˜ÙWÜ™Y™\™[˜Ù\ÊˆÛÛ›‹ˆ\š[ÙÝ\OH›]‹ˆ\š[ÙÙ]OXÝ\œ™[Û[Ûˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆ›ÝšY\\›ÝšY\‹ˆÙ^WÝ˜[YO]Ù^WÝ˜[YKˆ
BˆÝ[[X\žVÈ˜˜\Ù[[™\×Ü™XØ[Ý[]Y—H
ÏH™XØ[ÖÈœ™XÛÜ™×Ü›ØÙ\ÜÙY—BˆÝ[[X\žVÈœ™Y™\™[˜Ù\×ØÜ™X]Y—H
ÏH™XØ[ÖÈœ™Y™\™[˜Ù\×ØÜ™X]Y—BˆÝ[[X\žVÈœÝ[ÝÚ]Ý]Ü™Y™\™[˜ÙH—H
ÏH™XØ[ÖÈœÝ[ÝÚ]Ý]Ü™Y™\™[˜ÙH—BˆÛÛ›‹˜ÛÛ[Z]

Bˆ™]\›ˆÝ[[X\žB‚‚™Yˆ[—Ù\Ú[ÛœÛÛ\—Ü›ÙXÝ[Û—Ø˜XÚÙš[
ˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆ›ÝšY\ŽˆÝˆH‘\Ú[Û”ÛÛ\ˆ‹ˆ\š[ÙÝ\NˆÝˆH™^H‹ˆœ›ÛWÞYX\Žˆ[ˆ×ÞYX\Žˆ[ˆ\ÜÙ]ÚYˆ[›Û™HH›Û™KˆÙ^WÝ˜[YNˆ]H›Û™HH›Û™Kˆ]WÙœ›ÛNˆ]H›Û™HH›Û™Kˆ]WÝÎˆ]H›Û™HH›Û™KˆX^Ù^\Îˆ[›Û™HH›Û™KˆX^Ø\WØØ[Îˆ[›Û™HH›Û™KˆÜWØØ[Ù[^WÜÙXÛÛ™Îˆ›Ø]›Û™HH›Û™KˆÛY\\Žˆ[žH›Û™HH›Û™KˆX^ÝØZ]ØÞXÛ\Îˆ[HŠHOˆXÝÜÝ‹[žWN‚ˆYˆ\š[ÙÝ\H›Ý[ˆÈ™^H‹›[ÛŸN‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ•\ÈH\š[ÙÈ[˜[YËˆŠBˆYˆœ›ÛWÞYX\ˆˆ×ÞYX\Ž‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ[›È[šXÚX[˜[ÈÙHÙ\ˆÝ\\š[Üˆ[È[›Èš[˜[ˆŠBˆYˆ
×ÞYX\ˆHœ›ÛWÞYX\ˆ
ÈJHˆÎ‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ’[\˜[ÈÝ\\š[ÜˆHÈ[›ÜËˆ™Y^ˆÈ\š[ÙÈ\˜H^XÝ]\ˆÈ˜XÚÙš[ˆŠB‚ˆÙ^WÝ˜[YHHÙ^WÝ˜[YHÜˆ]KÙ^J
BˆÛÛ™šYÈHÙ]Ú[YÜ˜][Û—ØÛÛ™šYÊÛÛ›‹›ÝšY\ŠBˆYˆÛÛ™šYÈ\È›Û™N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠÛÛ™šYÝ\˜XØ[È\Ú[Û”ÛÛ\ˆ˜[È[˜ÛÛ˜YKˆŠBˆYˆ›ÝÛÛ™šYÖÈ™[˜X›Y—N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠH[YÜ˜XØ[È\Ú[Û”ÛÛ\ˆ\ÝH\Ø]]˜YKˆŠBˆ[™Ú[ÈHÙ]Ù\Ú[ÛœÛÛ\—Ù[™Ú[ØÛÛ™šYÊÛÛ™šYÊBˆ\ÜÙ]ÈHÙ]Ù\Ú[ÛœÛÛ\—Ü\™›Ü›X[˜ÙWØ\ÜÙ]ÊÛÛ›‹›ÝšY\‹\ÜÙ]ÚY
Bˆ]\ÈH]\—Û[ÛWØ˜XÚÙš[Ù]\Êœ›ÛWÞYX\‹×ÞYX\‹Ù^WÝ˜[YO]Ù^WÝ˜[YJBˆYˆ\š[ÙÝ\HOH™^HŽ‚ˆ]\ÈH]\—ÙZ[WØ˜XÚÙš[Û[ÛÊˆœ›ÛWÞYX\‹ˆ×ÞYX\‹ˆÙ^WÝ˜[YO]Ù^WÝ˜[YKˆ]WÙœ›ÛOY]WÙœ›ÛKˆ]WÝÏY]WÝËˆ
B‚ˆÝ[[X\žHHÂˆ˜\ÜÙ]×Ü›ØÙ\ÜÙYŽˆˆœ™XÛÜ™×Ý\]YŽˆˆ›Z\ÜÚ[™×Ü›ÙXÝ[ÛˆŽˆˆ˜\WÙ\œ›ÜœÈŽˆˆ˜\WØØ[×Ý\ÙYŽˆˆ›[Û×Ü›ØÙ\ÜÙYŽˆˆ˜Ú[šÜ×Ü›ØÙ\ÜÙYŽˆˆ›]Ü™XÛÜ™×Ý\]YŽˆˆ˜˜\Ù[[™\×Ü™XØ[Ý[]YŽˆˆœ™Y™\™[˜Ù\×ØÜ™X]YŽˆˆœÝ[ÝÚ]Ý]Ü™Y™\™[˜ÙHŽˆˆœÝÜYÜ™X\ÛÛˆŽˆˆ‹ˆœ™\Ý[YWÚ[Žˆˆ‹ˆØZ]ØÞXÛ\ÈŽˆˆBˆÙÙÙ\ˆHÝ\œ™[Ø\›ÙÙÙ\ˆYˆ\×Ø\ØÛÛ^

H[ÙHÙÙÚ[™Ë™Ù]ÙÙÙ\Š×Û˜[YW×ÊBˆÝ][Û—ØÛÙ\ÈHÜÝŠ\ÜÙ]È™^\›˜[ÚY—HÜˆˆŠKœÝš\

H›Üˆ\ÜÙ][ˆ\ÜÙ]ÈYˆÝŠ\ÜÙ]È™^\›˜[ÚY—HÜˆˆŠKœÝš\

WBˆ\ÜÙ]×ØžWÙ^\›˜[ÚYHÂˆÝŠ\ÜÙ]È™^\›˜[ÚY—HÜˆˆŠKœÝš\

Nˆ\ÜÙ]ˆ›Üˆ\ÜÙ][ˆ\ÜÙ]ÂˆYˆÝŠ\ÜÙ]È™^\›˜[ÚY—HÜˆˆŠKœÝš\

BˆBˆÝ[[X\žVÈ˜\ÜÙ]×Ü›ØÙ\ÜÙY—HH[Š\ÜÙ]×ØžWÙ^\›˜[ÚY
BˆX^Ø\WØØ[ÈHX^Ø\WØØ[ÈYˆX^Ø\WØØ[È\È›Ý›Û™H[ÙHX^Ù^\ÂˆX^Ø\WØØ[ÈHX^Ø\WØØ[ÈYˆX^Ø\WØØ[È\È›Ý›Û™H[ÙH•TÒSÓ”ÓÓT—ÔT‘“Ô“PSÑWÓPVÐTWÐÐSÂˆÜWØØ[Ù[^WÜÙXÛÛ™ÈH
ˆ•TÒSÓ”ÓÓT—ÔT‘“Ô“PSÑWÒÔWÑSVWÔÑPÓÓ‘ÂˆYˆÜWØØ[Ù[^WÜÙXÛÛ™È\È›Û™Bˆ[ÙHÜWØØ[Ù[^WÜÙXÛÛ™Âˆ
BˆÛY\Ù[˜ÈHÛY\\ˆÜˆ[YKœÛY\ˆ›ØÙ\ÜÙYÙ]\Îˆ\ÝÙ]WHH×B‚ˆYˆØZ]ØY\—Ü˜]WÛ[Z]
™X\ÛÛŽˆÝ‹™\Ý[YWÚ[ˆ]H›Û™HH›Û™JHOˆ›ÛÛ‚ˆÝ[[X\žVÈØZ]ØÞXÛ\È—H
ÏHBˆÝ[[X\žVÈœ™\Ý[YWÚ[—HH™\Ý[YWÚ[š\ÛÙ›Ü›X]

HYˆ™\Ý[YWÚ[[ÙHÝ[[X\žVÈœ™\Ý[YWÚ[—BˆYˆÝ[[X\žVÈØZ]ØÞXÛ\È—HˆX^ÝØZ]ØÞXÛ\Î‚ˆÝ[[X\žVÈœÝÜYÜ™X\ÛÛˆ—HHˆ“[Z]H\Ú[Û”ÛÛ\ˆ™\]YÈ[X\ÚXY\È™^™\Ëˆ[[[È\ÝYÎˆÜ™X\ÛÛŸH‚ˆÙÙÙ\‹Ø\›š[™Êˆ‘\Ú[Û”ÛÛ\ˆ\™›Ü›X[˜ÙH˜XÚÙš[ÝÜYY\ˆ™\X]YÛÛÛÝÛœÎˆ\š[ÙÝ\OI\ÈÝ][Û—ØÛÝ[I\ÈØZ]ØÞXÛ\ÏI\È™X\ÛÛI\È‹ˆ\š[ÙÝ\Kˆ[ŠÝ][Û—ØÛÙ\ÊKˆÝ[[X\žVÈØZ]ØÞXÛ\È—Kˆ™X\ÛÛ‹ˆ
Bˆ™]\›ˆ˜[ÙBˆÛÛ›‹˜ÛÛ[Z]

BˆÙXÛÛ™ÈH\Ú[ÛœÛÛ\—ØÛÛÛÝÛ—ÜÛY\ÜÙXÛÛ™ÊÛÛ›ŠBˆÙÙÙ\‹Ø\›š[™Êˆ‘\Ú[Û”ÛÛ\ˆ\™›Ü›X[˜ÙH˜XÚÙš[ØZ][™È›ÜˆTHÛÛÛÝÛŽˆ\š[ÙÝ\OI\ÈÝ][Û—ØÛÝ[I\ÈÙXÛÛ™ÏI\ÈØZ]ØÞXÛOI\È‹ˆ\š[ÙÝ\Kˆ[ŠÝ][Û—ØÛÙ\ÊKˆÙXÛÛ™ËˆÝ[[X\žVÈØZ]ØÞXÛ\È—Kˆ
BˆÛY\Ù[˜ÊÙXÛÛ™ÊBˆ™]\›ˆYB‚ˆÛÛÛÝÛ—Ü™X\ÛÛˆHÙ]Ù\Ú[ÛœÛÛ\—Ü\™›Ü›X[˜ÙWØÛÛÛÝÛ—Ü™X\ÛÛŠÛÛ›ŠBˆYˆÛÛÛÝÛ—Ü™X\ÛÛˆ[™›ÝØZ]ØY\—Ü˜]WÛ[Z]
ÛÛÛÝÛ—Ü™X\ÛÛŠN‚ˆÝ[[X\žVÈ˜\WÙ\œ›ÜœÈ—HHBˆ™]\›ˆÝ[[X\žB‚ˆÙ\ÜÚ[Û—ÛØš‹ÈHÙ]Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[ÛŠÛÛ™šYÊB‚ˆYˆ™Yœ™\ÚÜÙ\ÜÚ[Û—ØY\—Ù^\žJ^Îˆ^Ù\[Û‹ÛÛ^ˆÝŠHOˆ›Û™N‚ˆ›Û›ØØ[Ù\ÜÚ[Û—ÛØš‚ˆ[˜[Y]WÙ\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[ÛŠÛÛ™šYÊBˆÙ\ÜÚ[Û—ÛØš‹ÈHÙ]Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[ÛŠÛÛ™šYË›Ü˜ÙWÛÙÚ[UYJBˆÙÙÙ\‹Ø\›š[™Ê‘\Ú[Û”ÛÛ\ˆÙ\ÜÚ[Ûˆ™Yœ™\ÚYY\ˆ^\™YÙÚ[ŽˆÛÛ^I\È\œ›ÜI\È‹ÛÛ^^ÊB‚ˆYˆÝÜ™WÚÜWÛX\
ˆ\š[ÙÙ]WÝ˜[YNˆ]KˆÜWÛX\Ý˜[YNˆXÝÜÝ‹XÝÜÝ‹[žWWKˆ™XÛÜ™Ý\NˆÝ‹ˆ›Ý\×Ü™Yš^ˆÝ‹ˆ^\›˜[ÚYÎˆ\ÝÜÝ—H›Û™HH›Û™Kˆ
HOˆ›Û™N‚ˆÙ[XÝYÙ^\›˜[ÚYÈH^\›˜[ÚYÈÜˆ\Ý
\ÜÙ]×ØžWÙ^\›˜[ÚY
Bˆ›Üˆ^\›˜[ÚYÝ˜[YH[ˆÙ[XÝYÙ^\›˜[ÚYÎ‚ˆ\ÜÙ]H\ÜÙ]×ØžWÙ^\›˜[ÚYÙ^\›˜[ÚYÝ˜[YWBˆ™\Ý[HÝÜ™WÜ›ÙXÝ[Û—ÚÜWÜ™XÛÜ™
ˆÛÛ›‹ˆ\ÜÙ]Ü›ÝÏX\ÜÙ]ˆ›ÝšY\\›ÝšY\‹ˆ^\›˜[ÚYY^\›˜[ÚYÝ˜[YKˆ\š[ÙÝ\O\™XÛÜ™Ý\Kˆ\š[ÙÙ]O\\š[ÙÙ]WÝ˜[YKˆÜWÜ›ÝÏZÜWÛX\Ý˜[YK™Ù]
^\›˜[ÚYÝ˜[YKßJKˆ›Ý\×Ü™Yš^[›Ý\×Ü™Yš^ˆ
BˆYˆ™XÛÜ™Ý\HOH›]Ž‚ˆÝ[[X\žVÈ›]Ü™XÛÜ™×Ý\]Y—H
ÏHBˆ[ÙN‚ˆÝ[[X\žVÈœ™XÛÜ™×Ý\]Y—H
ÏHBˆYˆ™\Ý[Èœ›ÙXÝ[Û—ÚÝÚ—H\È›Û™N‚ˆÝ[[X\žVÈ›Z\ÜÚ[™×Ü›ÙXÝ[Ûˆ—H
ÏHB‚ˆYˆØZ]Ø™Y›Ü™WÛ™^ØØ[

HOˆ›Û™N‚ˆYˆ
ˆ“ÑPÕSÓ—ÒÔWÐÐSÐÓÓ•V™Ù]

H\È›Û™Bˆ[™ÜWØØ[Ù[^WÜÙXÛÛ™Âˆ[™ÜWØØ[Ù[^WÜÙXÛÛ™Èˆˆ
N‚ˆÛY\Ù[˜ÊÜWØØ[Ù[^WÜÙXÛÛ™ÊB‚ˆYˆ\š[ÙÝ\HOH™^HŽ‚ˆ›Üˆ[ÛÝ˜[YH[ˆ]\Î‚ˆ[ÛÚYÜ™XÛÜ™ÈH˜[ÙBˆ[ÛÜÝ][Û—ØÛÙ\ÈHÂˆ^\›˜[ÚYˆ›Üˆ^\›˜[ÚY\ÜÙ][ˆ\ÜÙ]×ØžWÙ^\›˜[ÚYš][\Ê
BˆYˆ]˜[X]WÛØØ[Û[ÛWÜ›ÙXÝ[Û—Ü]X[]JˆÛÛ›‹ˆ\ÜÙ]ÚYZ[
\ÜÙ]È˜\ÜÙ]ÚY—JKˆ›ÝšY\\›ÝšY\‹ˆ[ÛÜÝ\[[ÛÝ˜[YKˆ™Y™\™[˜ÙWÙ]O]Ù^WÝ˜[YKˆ
KœÝ]\ÂˆOH˜ÛÛ\]H‚ˆBˆÝ][Û—ØÚ[šÜÈHÚ[šÙY
[ÛÜÝ][Û—ØÛÙ\ËL
Bˆ›ÜˆÚ[š×Ú[™^Ý][Û—ÙÜ›Ý\[ˆ[[Y\˜]JÝ][Û—ØÚ[šÜËÝ\LJN‚ˆÙ\ÜÚ[Û—Ü™]žWÝ\ÙYH˜[ÙBˆÚ[HYN‚ˆYˆÝ[[X\žVÈ˜\WØØ[×Ý\ÙY—HHX^Ø\WØØ[Î‚ˆÝ[[X\žVÈœÝÜYÜ™X\ÛÛˆ—HH
ˆˆ“[Z]HØØ[HÛX^Ø\WØØ[ßHÚ[XY\ÈTH][™ÚYËˆ‚ˆˆ”™]ÛXHH\\ˆHÛ[ÛÝ˜[YKš\ÛÙ›Ü›X]

_Kˆ‚ˆ
BˆÝ[[X\žVÈœ™\Ý[YWÚ[—HH[ÛÝ˜[YKš\ÛÙ›Ü›X]

BˆÙÙÙ\‹š[™›Êˆ‘\Ú[Û”ÛÛ\ˆ\™›Ü›X[˜ÙH˜XÚÙš[ÝÜYžHX^Ø[Îˆ\š[ÙÝ\OI\È[ÛI\È\WØØ[×Ý\ÙYI\ÈX^Ø\WØØ[ÏI\È‹ˆ\š[ÙÝ\Kˆ[ÛÝ˜[YKš\ÛÙ›Ü›X]

KˆÝ[[X\žVÈ˜\WØØ[×Ý\ÙY—KˆX^Ø\WØØ[Ëˆ
Bˆœ™XZÂˆYˆÝ[[X\žVÈ˜\WØØ[×Ý\ÙY—Hˆ‚ˆØZ]Ø™Y›Ü™WÛ™^ØØ[

BˆžN‚ˆÙÙÙ\‹š[™›Êˆ‘\Ú[Û”ÛÛ\ˆZ[H\™›Ü›X[˜ÙH˜XÚÙš[™\]Y\Ýˆ\š[ÙÝ\OI\È[ÛI\ÈÝ][Û—ØÛÝ[I\ÈÚ[š×Ú[™^I\È\WØØ[×Ý\ÙYI\È‹ˆ\š[ÙÝ\Kˆ[ÛÝ˜[YKš\ÛÙ›Ü›X]

Kˆ[ŠÝ][Û—ÙÜ›Ý\
KˆÚ[š×Ú[™^ˆÝ[[X\žVÈ˜\WØØ[×Ý\ÙY—Kˆ
Bˆ›ÝÜÈH™]ÚÙ\Ú[ÛœÛÛ\—ÚÜWÙ^WÜ›ÝÜÊˆÙ\ÜÚ[Û—ÛØš‹ˆ[™Ú[ÖÈ˜˜\ÙWÝ\›—Kˆ[™Ú[ÖÈ™^WÚÜWÙ[™Ú[—KˆÝ][Û—ÙÜ›Ý\ˆ[ÛÝ˜[YKˆ
BˆÝ[[X\žVÈ˜\WØØ[×Ý\ÙY—H
ÏHBˆÝ[[X\žVÈ˜Ú[šÜ×Ü›ØÙ\ÜÙY—H
ÏHBˆ›Üˆ›ÝÈ[ˆ›ÝÜÎ‚ˆ^\›˜[ÚYÝ˜[YHHÝŠ›ÝË™Ù]
œÝ][ÛÛÙHŠHÜˆ›ÝË™Ù]
œ[ÛÙHŠHÜˆˆŠKœÝš\

Bˆ\ÜÙ]H\ÜÙ]×ØžWÙ^\›˜[ÚY™Ù]
^\›˜[ÚYÝ˜[YJBˆYˆ\ÜÙ]\È›Û™N‚ˆÛÛ[YBˆ›Ý×Ù]HH\œÙWÙ\Ú[ÛœÛÛ\—ØÛÛXÝÙ]J›ÝË[ÛÝ˜[YJBˆYˆ›Ý×Ù]H\È›Û™HÜˆ›Ý×Ù]Kœ™\XÙJ^OLJHOH[ÛÝ˜[YN‚ˆÛÛ[YBˆYˆ›Ý]WÚ[—Ø˜XÚÙš[ÝÚ[™ÝÊˆ›Ý×Ù]Kˆœ›ÛWÞYX\Yœ›ÛWÞYX\‹ˆ×ÞYX\]×ÞYX\‹ˆÙ^WÝ˜[YO]Ù^WÝ˜[YKˆ]WÙœ›ÛOY]WÙœ›ÛKˆ]WÝÏY]WÝËˆ
N‚ˆÛÛ[YBˆ™\Ý[HÝÜ™WÜ›ÙXÝ[Û—ÚÜWÜ™XÛÜ™
ˆÛÛ›‹ˆ\ÜÙ]Ü›ÝÏX\ÜÙ]ˆ›ÝšY\\›ÝšY\‹ˆ^\›˜[ÚYY^\›˜[ÚYÝ˜[YKˆ\š[ÙÝ\OH™^H‹ˆ\š[ÙÙ]O\›Ý×Ù]KˆÜWÜ›ÝÏ\›ÝËˆ›Ý\×Ü™Yš^H˜XÚÙš[\ÝÜšXÛÈY[œØ[X\š[Ëˆ‹ˆ
BˆYˆ™\Ý[È\Ù\ÜÝ]\È—HOHœÚÚ\YÙ^\Ý[™×Ý˜[YŽ‚ˆÝ[[X\žVÈœ™XÛÜ™×Ý\]Y—H
ÏHBˆ›ØÙ\ÜÙYÙ]\Ë˜\[™
›Ý×Ù]JBˆ[ÛÚYÜ™XÛÜ™ÈHYBˆYˆ™\Ý[Èœ›ÙXÝ[Û—ÚÝÚ—H\È›Û™N‚ˆÝ[[X\žVÈ›Z\ÜÚ[™×Ü›ÙXÝ[Ûˆ—H
ÏHBˆÙÙÙ\‹š[™›Êˆ‘\Ú[Û”ÛÛ\ˆZ[H\™›Ü›X[˜ÙH˜XÚÙš[™\ÜÛœÙNˆ[ÛI\ÈÚ[š×Ú[™^I\È›ÝÜÏI\È™XÛÜ™×Ý\]YI\È\WØØ[×Ý\ÙYI\È‹ˆ[ÛÝ˜[YKš\ÛÙ›Ü›X]

KˆÚ[š×Ú[™^ˆ[Š›ÝÜÊKˆÝ[[X\žVÈœ™XÛÜ™×Ý\]Y—KˆÝ[[X\žVÈ˜\WØØ[×Ý\ÙY—Kˆ
Bˆœ™XZÂˆ^Ù\^Ù\[Ûˆ\È^Î‚ˆÝ[[X\žVÈ˜\WÙ\œ›ÜœÈ—H
ÏHBˆYˆ\Ú[œÝ[˜ÙJ^Ë
\T˜]S[Z]\œ›Ü‹\TÛÝ[˜]˜Z[X›Q\œ›ÜŠJN‚ˆ˜Z\ÙBˆYˆ\×Ù\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]Ù\œ›ÜŠ^ÊN‚ˆ™X\ÛÛˆHX\š×Ù\Ú[ÛœÛÛ\—Ü\™›Ü›X[˜ÙWÜ˜]WÛ[Z]Y
ÛÛ›ŠBˆÙÙÙ\‹Ø\›š[™Êˆ‘\Ú[Û”ÛÛ\ˆ\™›Ü›X[˜ÙH˜XÚÙš[˜]H[Z]Yˆ\š[ÙÝ\OI\È[ÛI\ÈÝ][Û—ØÛÝ[I\ÈÚ[š×Ú[™^I\È\WØØ[×Ý\ÙYI\È\œ›ÜI\È‹ˆ\š[ÙÝ\Kˆ[ÛÝ˜[YKš\ÛÙ›Ü›X]

Kˆ[ŠÝ][Û—ÙÜ›Ý\
KˆÚ[š×Ú[™^ˆÝ[[X\žVÈ˜\WØØ[×Ý\ÙY—Kˆ^Ëˆ
BˆYˆØZ]ØY\—Ü˜]WÛ[Z]
™X\ÛÛ‹[ÛÝ˜[YJN‚ˆÛÛ[YBˆœ™XZÂˆYˆ\×Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[Û—Ù^\™YÙ\œ›ÜŠ^ÊH[™›ÝÙ\ÜÚ[Û—Ü™]žWÝ\ÙY‚ˆÙ\ÜÚ[Û—Ü™]žWÝ\ÙYHYBˆ™Yœ™\ÚÜÙ\ÜÚ[Û—ØY\—Ù^\žJ^Ëˆ™Z[NžÛ[ÛÝ˜[YKš\ÛÙ›Ü›X]

_N˜Ú[šÎžØÚ[š×Ú[™^HŠBˆÛÛ[YBˆÙÙÙ\‹Ø\›š[™Êˆ‘\Ú[Û”ÛÛ\ˆZ[H\™›Ü›X[˜ÙH˜XÚÙš[Ú[šÈ˜Z[Yˆ\š[ÙÝ\OI\È[ÛI\ÈÝ][Û—ØÛÝ[I\ÈÚ[š×Ú[™^I\È\œ›ÜI\È‹ˆ\š[ÙÝ\Kˆ[ÛÝ˜[YKš\ÛÙ›Ü›X]

Kˆ[ŠÝ][Û—ÙÜ›Ý\
KˆÚ[š×Ú[™^ˆ^Ëˆ
Bˆœ™XZÂˆYˆÝ[[X\žVÈœÝÜYÜ™X\ÛÛˆ—N‚ˆœ™XZÂˆYˆ[ÛÚYÜ™XÛÜ™Î‚ˆÝ[[X\žVÈ›[Û×Ü›ØÙ\ÜÙY—H
ÏHBˆYˆÝ[[X\žVÈœÝÜYÜ™X\ÛÛˆ—N‚ˆœ™XZÂ‚ˆ™XØ[×Ý\™Ù]ÈHÛÜY
Ù]
›ØÙ\ÜÙYÙ]\ÊJBˆ›Üˆ\™Ù][ˆ™XØ[×Ý\™Ù]Î‚ˆ™XØ[ÈH™XØ[Ý[]WÜ\™›Ü›X[˜ÙWÜ™Y™\™[˜Ù\ÊˆÛÛ›‹ˆ\š[ÙÝ\OH™^H‹ˆ\š[ÙÙ]O]\™Ù]ˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆ›ÝšY\\›ÝšY\‹ˆÙ^WÝ˜[YO]Ù^WÝ˜[YKˆ
BˆÝ[[X\žVÈ˜˜\Ù[[™\×Ü™XØ[Ý[]Y—H
ÏH™XØ[ÖÈœ™XÛÜ™×Ü›ØÙ\ÜÙY—BˆÝ[[X\žVÈœ™Y™\™[˜Ù\×ØÜ™X]Y—H
ÏH™XØ[ÖÈœ™Y™\™[˜Ù\×ØÜ™X]Y—BˆÝ[[X\žVÈœÝ[ÝÚ]Ý]Ü™Y™\™[˜ÙH—H
ÏH™XØ[ÖÈœÝ[ÝÚ]Ý]Ü™Y™\™[˜ÙH—BˆÙ[XÝYØ\ÜÙ]ÚYÈHÚ[
\ÜÙ]È˜\ÜÙ]ÚY—JH›Üˆ\ÜÙ][ˆ\ÜÙ]×BˆÝ\œ™[Û[ÛHÙ^WÝ˜[YKœ™\XÙJ^OLJBˆYˆÙ[XÝYØ\ÜÙ]ÚYÈ[™›ÝÝ[[X\žVÈœÝÜYÜ™X\ÛÛˆ—N‚ˆÙ\ÜÚ[Û—Ü™]žWÝ\ÙYH˜[ÙBˆÚ[HYN‚ˆžN‚ˆYˆÝ[[X\žVÈ˜\WØØ[×Ý\ÙY—Hˆ‚ˆØZ]Ø™Y›Ü™WÛ™^ØØ[

Bˆ›ÜˆÝ][Û—ÙÜ›Ý\[ˆÚ[šÙY
Ý][Û—ØÛÙ\ËL
N‚ˆÜWÛX\H™]ÚÙ\Ú[ÛœÛÛ\—ÚÜWÛ[ÛÛX\
ˆÙ\ÜÚ[Û—ÛØš‹ˆ[™Ú[ÖÈ˜˜\ÙWÝ\›—Kˆ[™Ú[ÖÈ›[ÛÚÜWÙ[™Ú[—KˆÝ][Û—ÙÜ›Ý\ˆÝ\œ™[Û[Ûˆ
BˆÝ[[X\žVÈ˜\WØØ[×Ý\ÙY—H
ÏHBˆÝÜ™WÚÜWÛX\
ˆÝ\œ™[Û[ÛˆÜWÛX\ˆ›]‹ˆ“U™XØ[Ý[YÈ\ÜÈ˜XÚÙš[ˆ‹ˆÝ][Û—ÙÜ›Ý\ˆ
Bˆœ™XZÂˆ^Ù\^Ù\[Ûˆ\È^Î‚ˆÝ[[X\žVÈ˜\WÙ\œ›ÜœÈ—H
ÏHBˆYˆ\Ú[œÝ[˜ÙJ^Ë
\T˜]S[Z]\œ›Ü‹\TÛÝ[˜]˜Z[X›Q\œ›ÜŠJN‚ˆ˜Z\ÙBˆYˆ\×Ù\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]Ù\œ›ÜŠ^ÊN‚ˆ™X\ÛÛˆHX\š×Ù\Ú[ÛœÛÛ\—Ü\™›Ü›X[˜ÙWÜ˜]WÛ[Z]Y
ÛÛ›ŠBˆYˆØZ]ØY\—Ü˜]WÛ[Z]
™X\ÛÛ‹Ý\œ™[Û[Û
N‚ˆÛÛ[YBˆYˆ\×Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[Û—Ù^\™YÙ\œ›ÜŠ^ÊH[™›ÝÙ\ÜÚ[Û—Ü™]žWÝ\ÙY‚ˆÙ\ÜÚ[Û—Ü™]žWÝ\ÙYHYBˆ™Yœ™\ÚÜÙ\ÜÚ[Û—ØY\—Ù^\žJ^Ëˆ™Z[K[]žØÝ\œ™[Û[Ûš\ÛÙ›Ü›X]

_HŠBˆÛÛ[YBˆÙÙÙ\‹Ø\›š[™Êˆ‘\Ú[Û”ÛÛ\ˆU™XØ[Ý[][Ûˆ˜Z[YY\ˆ˜XÚÙš[ˆÝ][Û—ØÛÝ[I\È\œ›ÜI\È‹ˆ[ŠÝ][Û—ØÛÙ\ÊKˆ^Ëˆ
Bˆœ™XZÂˆ™XØ[ÈH™XØ[Ý[]WÜ\™›Ü›X[˜ÙWÜ™Y™\™[˜Ù\ÊˆÛÛ›‹ˆ\š[ÙÝ\OH›]‹ˆ\š[ÙÙ]OXÝ\œ™[Û[Ûˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆ›ÝšY\\›ÝšY\‹ˆÙ^WÝ˜[YO]Ù^WÝ˜[YKˆ
BˆÝ[[X\žVÈ˜˜\Ù[[™\×Ü™XØ[Ý[]Y—H
ÏH™XØ[ÖÈœ™XÛÜ™×Ü›ØÙ\ÜÙY—BˆÝ[[X\žVÈœ™Y™\™[˜Ù\×ØÜ™X]Y—H
ÏH™XØ[ÖÈœ™Y™\™[˜Ù\×ØÜ™X]Y—BˆÝ[[X\žVÈœÝ[ÝÚ]Ý]Ü™Y™\™[˜ÙH—H
ÏH™XØ[ÖÈœÝ[ÝÚ]Ý]Ü™Y™\™[˜ÙH—BˆÛÛ›‹˜ÛÛ[Z]

Bˆ™]\›ˆÝ[[X\žB‚ˆ›Üˆ\š[ÙÙ]WÝ˜[YH[ˆ]\Î‚ˆÝÜ™YØÝ\œ™[Ù]HH˜[ÙBˆÙ\ÜÚ[Û—Ü™]žWÝ\ÙYH˜[ÙBˆÚ[HYN‚ˆžN‚ˆYˆÝ[[X\žVÈ˜\WØØ[×Ý\ÙY—HHX^Ø\WØØ[Î‚ˆÝ[[X\žVÈœÝÜYÜ™X\ÛÛˆ—HH
ˆˆ“[Z]HØØ[HÛX^Ø\WØØ[ßHÚ[XY\ÈTH][™ÚYËˆ‚ˆˆ”™]ÛXHH\\ˆHÜ\š[ÙÙ]WÝ˜[YKš\ÛÙ›Ü›X]

_Kˆ‚ˆ
BˆÝ[[X\žVÈœ™\Ý[YWÚ[—HH\š[ÙÙ]WÝ˜[YKš\ÛÙ›Ü›X]

Bˆœ™XZÂˆYˆÝ[[X\žVÈ˜\WØØ[×Ý\ÙY—Hˆ‚ˆØZ]Ø™Y›Ü™WÛ™^ØØ[

Bˆ›ÜˆÝ][Û—ÙÜ›Ý\[ˆÚ[šÙY
Ý][Û—ØÛÙ\ËL
N‚ˆÜWÛX\H™]ÚÙ\Ú[ÛœÛÛ\—ÚÜWÛ[ÛÛX\
ˆÙ\ÜÚ[Û—ÛØš‹ˆ[™Ú[ÖÈ˜˜\ÙWÝ\›—Kˆ[™Ú[ÖÈ›[ÛÚÜWÙ[™Ú[—KˆÝ][Û—ÙÜ›Ý\ˆ\š[ÙÙ]WÝ˜[YKˆ
BˆÝ[[X\žVÈ˜\WØØ[×Ý\ÙY—H
ÏHBˆÝÜ™WÚÜWÛX\
ˆ\š[ÙÙ]WÝ˜[YKˆÜWÛX\ˆ\š[ÙÝ\Kˆ˜XÚÙš[\ÝÜšXÛËˆ‹ˆÝ][Û—ÙÜ›Ý\ˆ
BˆÝÜ™YØÝ\œ™[Ù]HHYBˆœ™XZÂˆ^Ù\^Ù\[Ûˆ\È^Î‚ˆYˆ\Ú[œÝ[˜ÙJ^Ë
\T˜]S[Z]\œ›Ü‹\TÛÝ[˜]˜Z[X›Q\œ›ÜŠJN‚ˆ˜Z\ÙBˆYˆ\×Ù\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]Ù\œ›ÜŠ^ÊN‚ˆÝ[[X\žVÈ˜\WÙ\œ›ÜœÈ—H
ÏHBˆ™X\ÛÛˆHX\š×Ù\Ú[ÛœÛÛ\—Ü\™›Ü›X[˜ÙWÜ˜]WÛ[Z]Y
ÛÛ›ŠBˆÙÙÙ\‹Ø\›š[™Êˆ‘\Ú[Û”ÛÛ\ˆ\™›Ü›X[˜ÙH˜XÚÙš[˜]H[Z]Yˆ\š[ÙÝ\OI\È\š[ÙÙ]OI\ÈÝ][Û—ØÛÝ[I\È\œ›ÜI\È‹ˆ\š[ÙÝ\Kˆ\š[ÙÙ]WÝ˜[YKš\ÛÙ›Ü›X]

Kˆ[ŠÝ][Û—ØÛÙ\ÊKˆ^Ëˆ
BˆYˆØZ]ØY\—Ü˜]WÛ[Z]
™X\ÛÛ‹\š[ÙÙ]WÝ˜[YJN‚ˆÛÛ[YBˆœ™XZÂˆYˆ\×Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[Û—Ù^\™YÙ\œ›ÜŠ^ÊH[™›ÝÙ\ÜÚ[Û—Ü™]žWÝ\ÙY‚ˆÙ\ÜÚ[Û—Ü™]žWÝ\ÙYHYBˆÝ[[X\žVÈ˜\WÙ\œ›ÜœÈ—H
ÏHBˆ™Yœ™\ÚÜÙ\ÜÚ[Û—ØY\—Ù^\žJ^Ëˆ›[ÛYÜ›Ý\žÜ\š[ÙÙ]WÝ˜[YKš\ÛÙ›Ü›X]

_HŠBˆÛÛ[YBˆÙÙÙ\‹Ø\›š[™Êˆ‘\Ú[Û”ÛÛ\ˆÜ›Ý\Y\™›Ü›X[˜ÙH˜XÚÙš[˜Z[Y™]žZ[™È\ˆ\ÜÙ]ˆ\š[ÙÝ\OI\È\š[ÙÙ]OI\ÈÝ][Û—ØÛÝ[I\È\œ›ÜI\È‹ˆ\š[ÙÝ\Kˆ\š[ÙÙ]WÝ˜[YKš\ÛÙ›Ü›X]

Kˆ[ŠÝ][Û—ØÛÙ\ÊKˆ^Ëˆ
Bˆœ™XZÂˆYˆÝ[[X\žVÈœÝÜYÜ™X\ÛÛˆ—N‚ˆœ™XZÂˆYˆÝÜ™YØÝ\œ™[Ù]N‚ˆ›ØÙ\ÜÙYÙ]\Ë˜\[™
\š[ÙÙ]WÝ˜[YJBˆÛÛ[YB‚ˆÝÜ™YØ[žWÙ›Ü—Ù]HH˜[ÙBˆ›Üˆ^\›˜[ÚYÝ˜[YK\ÜÙ][ˆ\ÜÙ]×ØžWÙ^\›˜[ÚYš][\Ê
N‚ˆÙ\ÜÚ[Û—Ü™]žWÝ\ÙYH˜[ÙBˆÚ[HYN‚ˆžN‚ˆYˆÝ[[X\žVÈ˜\WØØ[×Ý\ÙY—HHX^Ø\WØØ[Î‚ˆÝ[[X\žVÈœÝÜYÜ™X\ÛÛˆ—HH
ˆˆ“[Z]HØØ[HÛX^Ø\WØØ[ßHÚ[XY\ÈTH][™ÚYËˆ‚ˆˆ”™]ÛXHH\\ˆHÜ\š[ÙÙ]WÝ˜[YKš\ÛÙ›Ü›X]

_Kˆ‚ˆ
BˆÝ[[X\žVÈœ™\Ý[YWÚ[—HH\š[ÙÙ]WÝ˜[YKš\ÛÙ›Ü›X]

Bˆœ™XZÂˆYˆÝ[[X\žVÈ˜\WØØ[×Ý\ÙY—Hˆ‚ˆØZ]Ø™Y›Ü™WÛ™^ØØ[

BˆÚ[™ÛWÛX\H™]ÚÙ\Ú[ÛœÛÛ\—ÚÜWÛ[ÛÛX\
ˆÙ\ÜÚ[Û—ÛØš‹ˆ[™Ú[ÖÈ˜˜\ÙWÝ\›—Kˆ[™Ú[ÖÈ›[ÛÚÜWÙ[™Ú[—KˆÙ^\›˜[ÚYÝ˜[YWKˆ\š[ÙÙ]WÝ˜[YKˆ
BˆÝ[[X\žVÈ˜\WØØ[×Ý\ÙY—H
ÏHBˆÝÜ™WÚÜWÛX\
\š[ÙÙ]WÝ˜[YKÚ[™ÛWÛX\\š[ÙÝ\K˜XÚÙš[\ÝÜšXÛËˆŠBˆÝÜ™YØ[žWÙ›Ü—Ù]HHYBˆœ™XZÂˆ^Ù\^Ù\[Ûˆ\È\ÜÙ]Ù^Î‚ˆÝ[[X\žVÈ˜\WÙ\œ›ÜœÈ—H
ÏHBˆYˆ\Ú[œÝ[˜ÙJ\ÜÙ]Ù^Ë
\T˜]S[Z]\œ›Ü‹\TÛÝ[˜]˜Z[X›Q\œ›ÜŠJN‚ˆ˜Z\ÙBˆYˆ\×Ù\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]Ù\œ›ÜŠ\ÜÙ]Ù^ÊN‚ˆ™X\ÛÛˆHX\š×Ù\Ú[ÛœÛÛ\—Ü\™›Ü›X[˜ÙWÜ˜]WÛ[Z]Y
ÛÛ›ŠBˆÙÙÙ\‹Ø\›š[™Êˆ‘\Ú[Û”ÛÛ\ˆ\™›Ü›X[˜ÙH˜XÚÙš[˜]H[Z]Yˆ\ÜÙ]ÚYI\ÈÝ][ÛÛÙOI\È\š[ÙÝ\OI\È\š[ÙÙ]OI\È\œ›ÜI\È‹ˆ\ÜÙ]È˜\ÜÙ]ÚY—Kˆ^\›˜[ÚYÝ˜[YKˆ\š[ÙÝ\Kˆ\š[ÙÙ]WÝ˜[YKš\ÛÙ›Ü›X]

Kˆ\ÜÙ]Ù^Ëˆ
BˆYˆØZ]ØY\—Ü˜]WÛ[Z]
™X\ÛÛ‹\š[ÙÙ]WÝ˜[YJN‚ˆÛÛ[YBˆœ™XZÂˆYˆ\×Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[Û—Ù^\™YÙ\œ›ÜŠ\ÜÙ]Ù^ÊH[™›ÝÙ\ÜÚ[Û—Ü™]žWÝ\ÙY‚ˆÙ\ÜÚ[Û—Ü™]žWÝ\ÙYHYBˆ™Yœ™\ÚÜÙ\ÜÚ[Û—ØY\—Ù^\žJ\ÜÙ]Ù^Ëˆ›[ÛX\ÜÙ]žÜ\š[ÙÙ]WÝ˜[YKš\ÛÙ›Ü›X]

_NžÙ^\›˜[ÚYÝ˜[Y_HŠBˆÛÛ[YBˆÙÙÙ\‹Ø\›š[™Êˆ‘\Ú[Û”ÛÛ\ˆ\™›Ü›X[˜ÙH˜XÚÙš[˜Z[Yˆ\ÜÙ]ÚYI\ÈÝ][ÛÛÙOI\È\š[ÙÝ\OI\È\š[ÙÙ]OI\È\œ›ÜI\È‹ˆ\ÜÙ]È˜\ÜÙ]ÚY—Kˆ^\›˜[ÚYÝ˜[YKˆ\š[ÙÝ\Kˆ\š[ÙÙ]WÝ˜[YKš\ÛÙ›Ü›X]

Kˆ\ÜÙ]Ù^Ëˆ
Bˆœ™XZÂˆYˆÝ[[X\žVÈœÝÜYÜ™X\ÛÛˆ—N‚ˆœ™XZÂˆYˆÝ[[X\žVÈœÝÜYÜ™X\ÛÛˆ—N‚ˆœ™XZÂˆYˆÝÜ™YØ[žWÙ›Ü—Ù]N‚ˆ›ØÙ\ÜÙYÙ]\Ë˜\[™
\š[ÙÙ]WÝ˜[YJB‚ˆÙ[XÝYØ\ÜÙ]ÚYÈHÚ[
\ÜÙ]È˜\ÜÙ]ÚY—JH›Üˆ\ÜÙ][ˆ\ÜÙ]×Bˆ™XØ[×Ý\™Ù]ÈH›ØÙ\ÜÙYÙ]\Âˆ›Üˆ\™Ù][ˆ™XØ[×Ý\™Ù]Î‚ˆ™XØ[ÈH™XØ[Ý[]WÜ\™›Ü›X[˜ÙWÜ™Y™\™[˜Ù\ÊˆÛÛ›‹ˆ\š[ÙÝ\O\\š[ÙÝ\Kˆ\š[ÙÙ]O]\™Ù]ˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆ›ÝšY\\›ÝšY\‹ˆÙ^WÝ˜[YO]Ù^WÝ˜[YKˆ
BˆÝ[[X\žVÈ˜˜\Ù[[™\×Ü™XØ[Ý[]Y—H
ÏH™XØ[ÖÈœ™XÛÜ™×Ü›ØÙ\ÜÙY—BˆÝ[[X\žVÈœ™Y™\™[˜Ù\×ØÜ™X]Y—H
ÏH™XØ[ÖÈœ™Y™\™[˜Ù\×ØÜ™X]Y—BˆÝ[[X\žVÈœÝ[ÝÚ]Ý]Ü™Y™\™[˜ÙH—H
ÏH™XØ[ÖÈœÝ[ÝÚ]Ý]Ü™Y™\™[˜ÙH—B‚ˆÝ\œ™[Û[ÛHÙ^WÝ˜[YKœ™\XÙJ^OLJBˆYˆÙ[XÝYØ\ÜÙ]ÚYÈ[™›ÝÝ[[X\žVÈœÝÜYÜ™X\ÛÛˆ—N‚ˆÙ\ÜÚ[Û—Ü™]žWÝ\ÙYH˜[ÙBˆÚ[HYN‚ˆžN‚ˆYˆÝ[[X\žVÈ˜\WØØ[×Ý\ÙY—Hˆ‚ˆØZ]Ø™Y›Ü™WÛ™^ØØ[

Bˆ›ÜˆÝ][Û—ÙÜ›Ý\[ˆÚ[šÙY
Ý][Û—ØÛÙ\ËL
N‚ˆÜWÛX\H™]ÚÙ\Ú[ÛœÛÛ\—ÚÜWÛ[ÛÛX\
ˆÙ\ÜÚ[Û—ÛØš‹ˆ[™Ú[ÖÈ˜˜\ÙWÝ\›—Kˆ[™Ú[ÖÈ›[ÛÚÜWÙ[™Ú[—KˆÝ][Û—ÙÜ›Ý\ˆÝ\œ™[Û[Ûˆ
BˆÝ[[X\žVÈ˜\WØØ[×Ý\ÙY—H
ÏHBˆÝÜ™WÚÜWÛX\
ˆÝ\œ™[Û[ÛˆÜWÛX\ˆ›]‹ˆ“U™XØ[Ý[YÈ\ÜÈ˜XÚÙš[ˆ‹ˆÝ][Û—ÙÜ›Ý\ˆ
Bˆœ™XZÂˆ^Ù\^Ù\[Ûˆ\È^Î‚ˆÝ[[X\žVÈ˜\WÙ\œ›ÜœÈ—H
ÏHBˆYˆ\Ú[œÝ[˜ÙJ^Ë
\T˜]S[Z]\œ›Ü‹\TÛÝ[˜]˜Z[X›Q\œ›ÜŠJN‚ˆ˜Z\ÙBˆYˆ\×Ù\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]Ù\œ›ÜŠ^ÊN‚ˆ™X\ÛÛˆHX\š×Ù\Ú[ÛœÛÛ\—Ü\™›Ü›X[˜ÙWÜ˜]WÛ[Z]Y
ÛÛ›ŠBˆYˆØZ]ØY\—Ü˜]WÛ[Z]
™X\ÛÛ‹Ý\œ™[Û[Û
N‚ˆÛÛ[YBˆYˆ\×Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[Û—Ù^\™YÙ\œ›ÜŠ^ÊH[™›ÝÙ\ÜÚ[Û—Ü™]žWÝ\ÙY‚ˆÙ\ÜÚ[Û—Ü™]žWÝ\ÙYHYBˆ™Yœ™\ÚÜÙ\ÜÚ[Û—ØY\—Ù^\žJ^Ëˆ›[Û[]žØÝ\œ™[Û[Ûš\ÛÙ›Ü›X]

_HŠBˆÛÛ[YBˆÙÙÙ\‹Ø\›š[™Êˆ‘\Ú[Û”ÛÛ\ˆU™XØ[Ý[][Ûˆ˜Z[YY\ˆ˜XÚÙš[ˆÝ][Û—ØÛÝ[I\È\œ›ÜI\È‹ˆ[ŠÝ][Û—ØÛÙ\ÊKˆ^Ëˆ
Bˆœ™XZÂˆ™XØ[ÈH™XØ[Ý[]WÜ\™›Ü›X[˜ÙWÜ™Y™\™[˜Ù\ÊˆÛÛ›‹ˆ\š[ÙÝ\OH›]‹ˆ\š[ÙÙ]OXÝ\œ™[Û[Ûˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆ›ÝšY\\›ÝšY\‹ˆÙ^WÝ˜[YO]Ù^WÝ˜[YKˆ
BˆÝ[[X\žVÈ˜˜\Ù[[™\×Ü™XØ[Ý[]Y—H
ÏH™XØ[ÖÈœ™XÛÜ™×Ü›ØÙ\ÜÙY—BˆÝ[[X\žVÈœ™Y™\™[˜Ù\×ØÜ™X]Y—H
ÏH™XØ[ÖÈœ™Y™\™[˜Ù\×ØÜ™X]Y—BˆÝ[[X\žVÈœÝ[ÝÚ]Ý]Ü™Y™\™[˜ÙH—H
ÏH™XØ[ÖÈœÝ[ÝÚ]Ý]Ü™Y™\™[˜ÙH—BˆÛÛ›‹˜ÛÛ[Z]

Bˆ™]\›ˆÝ[[X\žB‚‚™Yˆ]˜[X]WÛØØ[Û[ÛWÜ›ÙXÝ[Û—Ü]X[]JˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆ\ÜÙ]ÚYˆ[ˆ›ÝšY\ŽˆÝ‹ˆ[ÛÜÝ\ˆ]Kˆ™Y™\™[˜ÙWÙ]Nˆ]KŠHOˆ[žN‚ˆË\ÝÙ^HHØ[[™\‹›[Û˜[™ÙJ[ÛÜÝ\žYX\‹[ÛÜÝ\›[Û
Bˆ[ÛÙ[™H[ÛÜÝ\œ™\XÙJ^O[\ÝÙ^JBˆ[ÛWÜ™XÛÜ™ÈH]Y\žWØ[
ˆÛÛ›‹ˆˆˆ‚ˆÑSPÕ\š[ÙÙ]K›ÙXÝ[Û—ÚÝÚˆ”“ÓH›ÙXÝ[Û—Ü™XÛÜ™ÂˆÒT‘H\ÜÙ]ÚYHÈS‘›ÝšY\ˆHÈS‘\š[ÙÝ\HH	Û[Û	ÈS‘\š[ÙÙ]HHÂˆÔ‘Tˆ–HYˆˆˆ‹ˆ
\ÜÙ]ÚY›ÝšY\‹[ÛÜÝ\š\ÛÙ›Ü›X]

JKˆ
BˆZ[WÜ™XÛÜ™ÈH]Y\žWØ[
ˆÛÛ›‹ˆˆˆ‚ˆÑSPÕ\š[ÙÙ]K›ÙXÝ[Û—ÚÝÚˆ”“ÓH›ÙXÝ[Û—Ü™XÛÜ™ÂˆÒT‘H\ÜÙ]ÚYHÈS‘›ÝšY\ˆHÈS‘\š[ÙÝ\HH	Ù^IÂˆS‘\š[ÙÙ]H‘UÑQSˆÈS‘ÂˆÔ‘Tˆ–H\š[ÙÙ]KYˆˆˆ‹ˆ
\ÜÙ]ÚY›ÝšY\‹[ÛÜÝ\š\ÛÙ›Ü›X]

K[ÛÙ[™š\ÛÙ›Ü›X]

JKˆ
Bˆ™]\›ˆ]˜[X]WÛ[ÛWÜ›ÙXÝ[Û—Ü]X[]Jˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆ[ÛÜÝ\[[ÛÜÝ\ˆ™Y™\™[˜ÙWÙ]O\™Y™\™[˜ÙWÙ]Kˆ[ÛWÜ™XÛÜ™Ï[[ÛWÜ™XÛÜ™ËˆZ[WÜ™XÛÜ™ÏYZ[WÜ™XÛÜ™Ëˆ
B‚‚™YˆÛÝ[Û[ÛWÜ›ÙXÝ[Û—ÜÝ]\Ê]X[]Y\ÎˆXÝÚ[[žWJHOˆXÝÜÝ‹[N‚ˆÛÝ[ÈHÈ˜ÛÛ\]HŽˆœ\X[Žˆ›Z\ÜÚ[™ÈŽˆ˜ÛÛ™›XÝŽˆBˆ›Üˆ]X[]H[ˆ]X[]Y\Ë˜[Y\Ê
N‚ˆÝ]\ÈHÝŠ]X[]KœÝ]\ÊBˆÛÝ[ÖÜÝ]\×HHÛÝ[Ë™Ù]
Ý]\Ë
H
ÈBˆ™]\›ˆÛÝ[Â‚‚™Yˆ[—Ù\Ú[ÛœÛÛ\—Û[ÛØÛÜÙJˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆ›ÝšY\ŽˆÝˆH‘\Ú[Û”ÛÛ\ˆ‹ˆ™\ÜÛ[ÛˆÝ‹ˆ\ÜÙ]ÚYÎˆ\ÝÚ[H›Û™HH›Û™Kˆ™Y™\™[˜ÙWÙ]Nˆ]H›Û™HH›Û™KˆÜWØØ[Ù[^WÜÙXÛÛ™Îˆ›Ø]›Û™HH›Û™KˆÛY\\Žˆ[žH›Û™HH›Û™KŠHOˆXÝÜÝ‹[žWN‚ˆ™Y™\™[˜ÙWÙ]HH™Y™\™[˜ÙWÙ]HÜˆÝ\œ™[Û\Ø›Û—Ù]J
Bˆ[ÛÜÝ\H]][YKœÝœ[YJ›Ü›X[^™WÜ™\ÜÛ[Û
™\ÜÛ[Û
K‰VKI[HŠK™]J
BˆYˆ[ÛÜÝ\H™Y™\™[˜ÙWÙ]Kœ™\XÙJ^OLJN‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ“È™XÚÈY[œØ[]]ÛX]XÛÈÛÈÙHÛÛœÝ[\ˆY\Ù\ÈÚ]š\È[\š[Ü™\ËˆŠB‚ˆÛÛ™šYÈHÙ]Ú[YÜ˜][Û—ØÛÛ™šYÊÛÛ›‹›ÝšY\ŠBˆYˆÛÛ™šYÈ\È›Û™N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠÛÛ™šYÝ\˜XØ[È\Ú[Û”ÛÛ\ˆ˜[È[˜ÛÛ˜YKˆŠBˆYˆ›ÝÛÛ™šYÖÈ™[˜X›Y—N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠH[YÜ˜XØ[È\Ú[Û”ÛÛ\ˆ\ÝH\Ø]]˜YKˆŠBˆ[™Ú[ÈHÙ]Ù\Ú[ÛœÛÛ\—Ù[™Ú[ØÛÛ™šYÊÛÛ™šYÊBˆ\ÜÙ]ÈHÙ]Ù\Ú[ÛœÛÛ\—Ü\™›Ü›X[˜ÙWØ\ÜÙ]ÊˆÛÛ›‹ˆ›ÝšY\‹ˆ\ÜÙ]ÚYÏX\ÜÙ]ÚYËˆ
Bˆ\ÜÙ]×ØžWÚYHÚ[
\ÜÙ]È˜\ÜÙ]ÚY—JNˆ\ÜÙ]›Üˆ\ÜÙ][ˆ\ÜÙ]ßBˆ\ÜÙ]×ØžWÙ^\›˜[ÚYHÂˆÝŠ\ÜÙ]È™^\›˜[ÚY—HÜˆˆŠKœÝš\

Nˆ\ÜÙ]ˆ›Üˆ\ÜÙ][ˆ\ÜÙ]ÂˆYˆÝŠ\ÜÙ]È™^\›˜[ÚY—HÜˆˆŠKœÝš\

BˆB‚ˆYˆ]˜[X]WØ\ÜÙ]Ê
HOˆXÝÚ[[žWN‚ˆ™]\›ˆÂˆ\ÜÙ]ÚYˆ]˜[X]WÛØØ[Û[ÛWÜ›ÙXÝ[Û—Ü]X[]JˆÛÛ›‹ˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆ›ÝšY\\›ÝšY\‹ˆ[ÛÜÝ\[[ÛÜÝ\ˆ™Y™\™[˜ÙWÙ]O\™Y™\™[˜ÙWÙ]Kˆ
Bˆ›Üˆ\ÜÙ]ÚY[ˆ\ÜÙ]×ØžWÚYˆB‚ˆ]X[]Y\ÈH]˜[X]WØ\ÜÙ]Ê
Bˆ[™[™×Ø\ÜÙ]ÚYÈHÂˆ\ÜÙ]ÚYˆ›Üˆ\ÜÙ]ÚY]X[]H[ˆ]X[]Y\Ëš][\Ê
BˆYˆ]X[]KœÝ]\È[ˆÈœ\X[‹›Z\ÜÚ[™È‹˜ÛÛ™›XÝŸBˆBˆY™™XÝYØ\ÜÙ]ÚYÈHÙ]
[™[™×Ø\ÜÙ]ÚYÊBˆ›ØÙ\ÜÙYÙZ[WÝ\™Ù]ÎˆÙ]Ý\VÚ[]WWHHÙ]

BˆÝ[[X\žNˆXÝÜÝ‹[žWHHÂˆœÝ]\ÈŽˆœ[›š[™È‹ˆ›[ÛŽˆ[ÛÜÝ\œÝ™[YJ‰VKI[HŠKˆ˜\ÜÙ]×Ù]˜[X]YŽˆ[Š\ÜÙ]×ØžWÚY
Kˆ˜\ÜÙ]×ØY™™XÝYŽˆ[ŠY™™XÝYØ\ÜÙ]ÚYÊKˆ˜\ÜÙ]ÚY×ØY™™XÝYŽˆÛÜY
Y™™XÝYØ\ÜÙ]ÚYÊKˆœÝ]\×Ø™Y›Ü™HŽˆÛÝ[Û[ÛWÜ›ÙXÝ[Û—ÜÝ]\Ê]X[]Y\ÊKˆœÝ]\×ØY\—Û[ÛHŽˆÛÝ[Û[ÛWÜ›ÙXÝ[Û—ÜÝ]\Ê]X[]Y\ÊKˆœÝ]\×ØY\ˆŽˆÛÝ[Û[ÛWÜ›ÙXÝ[Û—ÜÝ]\Ê]X[]Y\ÊKˆ›[ÛWØ\ÜÙ]×Ü™\]Y\ÝYŽˆˆ™Z[WØ\ÜÙ]×Ü™\]Y\ÝYŽˆˆ›[ÛWÜ™XÛÜ™×Ý\]YŽˆˆœ™XÛÜ™×Ý\]YŽˆˆ›Z\ÜÚ[™×Ü›ÙXÝ[ÛˆŽˆˆ˜\WØØ[×Ø][\YŽˆˆ˜\WØØ[×Ý\ÙYŽˆˆ˜\WÙ\œ›ÜœÈŽˆˆœÝÜYÜ™X\ÛÛˆŽˆˆ‹ˆ›™^Ø][\Ø]Žˆˆ‹ˆBˆYˆ›Ý[™[™×Ø\ÜÙ]ÚYÎ‚ˆÝ[[X\žVÈœÝ]\È—HH˜ÛÛ\]Y‚ˆ™]\›ˆÝ[[X\žB‚ˆÛY\Ù[˜ÈHÛY\\ˆÜˆ[YKœÛY\ˆ[^WÜÙXÛÛ™ÈH
ˆ•TÒSÓ”ÓÓT—ÔT‘“Ô“PSÑWÒÔWÑSVWÔÑPÓÓ‘ÂˆYˆÜWØØ[Ù[^WÜÙXÛÛ™È\È›Û™Bˆ[ÙHÜWØØ[Ù[^WÜÙXÛÛ™Âˆ
BˆÙÙÙ\ˆHÝ\œ™[Ø\›ÙÙÙ\ˆYˆ\×Ø\ØÛÛ^

H[ÙHÙÙÚ[™Ë™Ù]ÙÙÙ\Š×Û˜[YW×ÊBˆÙ\ÜÚ[Û—ÛØšŽˆ[žHH›Û™B‚ˆYˆ™XØ[Ý[]WØY™™XÝYÜ™Y™\™[˜Ù\Ê
HOˆ›Û™N‚ˆ›Üˆ\ÜÙ]ÚY\™Ù]Ù]H[ˆÛÜY
›ØÙ\ÜÙYÙZ[WÝ\™Ù]ÊN‚ˆ™XØ[Ý[]WÜ\™›Ü›X[˜ÙWÜ™Y™\™[˜Ù\ÊˆÛÛ›‹ˆ\š[ÙÝ\OH™^H‹ˆ\š[ÙÙ]O]\™Ù]Ù]Kˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆ›ÝšY\\›ÝšY\‹ˆÙ^WÝ˜[YO\™Y™\™[˜ÙWÙ]Kˆ
Bˆ›Üˆ\ÜÙ]ÚY[ˆÛÜY
Y™™XÝYØ\ÜÙ]ÚYÊN‚ˆ™XØ[Ý[]WÜ\™›Ü›X[˜ÙWÜ™Y™\™[˜Ù\ÊˆÛÛ›‹ˆ\š[ÙÝ\OH›[Û‹ˆ\š[ÙÙ]O[[ÛÜÝ\ˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆ›ÝšY\\›ÝšY\‹ˆÙ^WÝ˜[YO\™Y™\™[˜ÙWÙ]Kˆ
B‚ˆYˆÝÜÙ›Ü—Ü˜]WÛ[Z]
^Îˆ^Ù\[ÛŠHOˆ›Û™N‚ˆYˆ\Ú[œÝ[˜ÙJ^Ë\T˜]S[Z]\œ›ÜŠN‚ˆ˜]WÙ\œ›ÜˆH^Âˆ[ÙN‚ˆ[[HX\š×Ù\Ú[ÛœÛÛ\—Ø\WØÛÛÛÝÛŠˆÛÛ›‹ˆTWÐT‘PWÔ“ÑPÕSÓ‹ˆ™X\ÛÛ\ÝŠ^ÊKˆ
Bˆ˜]WÙ\œ›ÜˆH\T˜]S[Z]\œ›ÜŠˆ›ÝšY\‹ˆTWÐT‘PWÔ“ÑPÕSÓ‹ˆ[[ˆÙ]Ü›ÝšY\—ØÛÛÛÝÛ—Ü™X\ÛÛŠÛÛ›‹›ÝšY\‹TWÐT‘PWÔ“ÑPÕSÓŠKˆ
Bˆ™XØ[Ý[]WØY™™XÝYÜ™Y™\™[˜Ù\Ê
BˆÝ\œ™[Ü]X[]Y\ÈH]˜[X]WØ\ÜÙ]Ê
BˆÝ[[X\žVÈœÝ]\È—HHØZ][™×Ü˜]WÛ[Z]‚ˆÝ[[X\žVÈœÝ]\×ØY\ˆ—HHÛÝ[Û[ÛWÜ›ÙXÝ[Û—ÜÝ]\ÊÝ\œ™[Ü]X[]Y\ÊBˆÝ[[X\žVÈœÝÜYÜ™X\ÛÛˆ—HH˜]WÙ\œ›Ü‹›Y\ÜØYÙBˆÝ[[X\žVÈ›™^Ø][\Ø]—HH˜]WÙ\œ›Ü‹˜ÛÛÛÝÛ—Ý[[š\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠBˆÝ[[X\žVÈ˜ÛÛÛÝÛ—Ý[[—HHÝ[[X\žVÈ›™^Ø][\Ø]—BˆÛÛ›‹˜ÛÛ[Z]

Bˆ˜]WÙ\œ›Ü‹š›Ø—Ü™\Ý[HXÝ
Ý[[X\žJBˆ˜Z\ÙH˜]WÙ\œ›Ü‚‚ˆžN‚ˆÙ\ÜÚ[Û—ÛØš‹ÈHÙ]Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[ÛŠÛÛ™šYÊBˆ^Ù\^Ù\[Ûˆ\È^Î‚ˆÝ[[X\žVÈ˜\WÙ\œ›ÜœÈ—H
ÏHBˆYˆ\×Ù\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]Ù\œ›ÜŠ^ÊN‚ˆÝÜÙ›Ü—Ü˜]WÛ[Z]
^ÊBˆ˜Z\ÙB‚ˆYˆØ[ÚÜWØ\JØ[˜XÚÎˆ[žKÛÛ^ˆÝŠHOˆ[žN‚ˆ›Û›ØØ[Ù\ÜÚ[Û—ÛØš‚ˆ™[ÙÚ[—Ý\ÙYH˜[ÙBˆÚ[HYN‚ˆYˆ
ˆ“ÑPÕSÓ—ÒÔWÐÐSÐÓÓ•V™Ù]

H\È›Û™Bˆ[™Ý[[X\žVÈ˜\WØØ[×Ý\ÙY—Bˆ[™[^WÜÙXÛÛ™Âˆ[™[^WÜÙXÛÛ™Èˆˆ
N‚ˆÛY\Ù[˜Ê[^WÜÙXÛÛ™ÊBˆÝ[[X\žVÈ˜\WØØ[×Ø][\Y—H
ÏHBˆžN‚ˆ™\Ý[HØ[˜XÚÊ
BˆÝ[[X\žVÈ˜\WØØ[×Ý\ÙY—H
ÏHBˆ™]\›ˆ™\Ý[ˆ^Ù\^Ù\[Ûˆ\È^Î‚ˆÝ[[X\žVÈ˜\WÙ\œ›ÜœÈ—H
ÏHBˆYˆ\×Ù\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]Ù\œ›ÜŠ^ÊN‚ˆÝÜÙ›Ü—Ü˜]WÛ[Z]
^ÊBˆYˆ\×Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[Û—Ù^\™YÙ\œ›ÜŠ^ÊH[™›Ý™[ÙÚ[—Ý\ÙY‚ˆ™[ÙÚ[—Ý\ÙYHYBˆ[˜[Y]WÙ\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[ÛŠÛÛ™šYÊBˆžN‚ˆÙ\ÜÚ[Û—ÛØš‹ÈHÙ]Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[ÛŠÛÛ™šYË›Ü˜ÙWÛÙÚ[UYJBˆ^Ù\^Ù\[Ûˆ\ÈÙÚ[—Ù^Î‚ˆÝ[[X\žVÈ˜\WÙ\œ›ÜœÈ—H
ÏHBˆYˆ\×Ù\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]Ù\œ›ÜŠÙÚ[—Ù^ÊN‚ˆÝÜÙ›Ü—Ü˜]WÛ[Z]
ÙÚ[—Ù^ÊBˆ˜Z\ÙBˆÙÙÙ\‹Ø\›š[™Êˆ‘\Ú[Û”ÛÛ\ˆ[ÛÛÜÙHÙ\ÜÚ[Ûˆ™Yœ™\ÚYÛ˜ÙNˆ[ÛI\ÈÛÛ^I\È\œ›ÜI\È‹ˆ[ÛÜÝ\ˆÛÛ^ˆ^Ëˆ
BˆÛÛ[YBˆ˜Z\ÙB‚ˆ[™[™×ÜÝ][Û—ØÛÙ\ÈHÂˆÝŠ\ÜÙ]×ØžWÚYØ\ÜÙ]ÚYVÈ™^\›˜[ÚY—HÜˆˆŠKœÝš\

Bˆ›Üˆ\ÜÙ]ÚY[ˆ[™[™×Ø\ÜÙ]ÚYÂˆBˆÝ[[X\žVÈ›[ÛWØ\ÜÙ]×Ü™\]Y\ÝY—HH[Š[™[™×ÜÝ][Û—ØÛÙ\ÊBˆ›ÜˆÚ[š×Ú[™^Ý][Û—ÙÜ›Ý\[ˆ[[Y\˜]JÚ[šÙY
[™[™×ÜÝ][Û—ØÛÙ\ËL
KÝ\LJN‚ˆÜWÛX\HØ[ÚÜWØ\Jˆ[X™HÝ][Û—ÙÜ›Ý\\Ý][Û—ÙÜ›Ý\ˆ™]ÚÙ\Ú[ÛœÛÛ\—ÚÜWÛ[ÛÛX\
ˆÙ\ÜÚ[Û—ÛØš‹ˆ[™Ú[ÖÈ˜˜\ÙWÝ\›—Kˆ[™Ú[ÖÈ›[ÛÚÜWÙ[™Ú[—KˆÝ][Û—ÙÜ›Ý\ˆ[ÛÜÝ\ˆ
Kˆˆ›[ÛNžØÚ[š×Ú[™^H‹ˆ
Bˆ›Üˆ^\›˜[ÚY[ˆÝ][Û—ÙÜ›Ý\‚ˆ\ÜÙ]H\ÜÙ]×ØžWÙ^\›˜[ÚYÙ^\›˜[ÚYBˆ™\Ý[HÝÜ™WÜ›ÙXÝ[Û—ÚÜWÜ™XÛÜ™
ˆÛÛ›‹ˆ\ÜÙ]Ü›ÝÏX\ÜÙ]ˆ›ÝšY\\›ÝšY\‹ˆ^\›˜[ÚYY^\›˜[ÚYˆ\š[ÙÝ\OH›[Û‹ˆ\š[ÙÙ]O[[ÛÜÝ\ˆÜWÜ›ÝÏZÜWÛX\™Ù]
^\›˜[ÚYßJKˆ›Ý\×Ü™Yš^H‘™XÚÈY[œØ[]]ÛX]XÛËˆ‹ˆ
BˆYˆ™\Ý[È\Ù\ÜÝ]\È—HOHœÚÚ\YÙ^\Ý[™×Ý˜[YŽ‚ˆÝ[[X\žVÈ›[ÛWÜ™XÛÜ™×Ý\]Y—H
ÏHBˆYˆ™\Ý[Èœ›ÙXÝ[Û—ÚÝÚ—H\È›Û™N‚ˆÝ[[X\žVÈ›Z\ÜÚ[™×Ü›ÙXÝ[Ûˆ—H
ÏHBˆÛÛ›‹˜ÛÛ[Z]

B‚ˆ]X[]Y\ÈH]˜[X]WØ\ÜÙ]Ê
BˆÝ[[X\žVÈœÝ]\×ØY\—Û[ÛH—HHÛÝ[Û[ÛWÜ›ÙXÝ[Û—ÜÝ]\Ê]X[]Y\ÊBˆZ[WØ\ÜÙ]ÚYÈHÂˆ\ÜÙ]ÚYˆ›Üˆ\ÜÙ]ÚY[ˆ[™[™×Ø\ÜÙ]ÚYÂˆYˆ]X[]Y\ÖØ\ÜÙ]ÚYKœÝ]\È[ˆÈœ\X[‹›Z\ÜÚ[™È‹˜ÛÛ™›XÝŸBˆBˆZ[WÜÝ][Û—ØÛÙ\ÈHÂˆÝŠ\ÜÙ]×ØžWÚYØ\ÜÙ]ÚYVÈ™^\›˜[ÚY—HÜˆˆŠKœÝš\

Bˆ›Üˆ\ÜÙ]ÚY[ˆZ[WØ\ÜÙ]ÚYÂˆBˆÝ[[X\žVÈ™Z[WØ\ÜÙ]×Ü™\]Y\ÝY—HH[ŠZ[WÜÝ][Û—ØÛÙ\ÊBˆË\ÝÙ^HHØ[[™\‹›[Û˜[™ÙJ[ÛÜÝ\žYX\‹[ÛÜÝ\›[Û
Bˆ[ÛÙ[™H[ÛÜÝ\œ™\XÙJ^O[\ÝÙ^JBˆ›ÜˆÚ[š×Ú[™^Ý][Û—ÙÜ›Ý\[ˆ[[Y\˜]JÚ[šÙY
Z[WÜÝ][Û—ØÛÙ\ËL
KÝ\LJN‚ˆ›ÝÜÈHØ[ÚÜWØ\Jˆ[X™HÝ][Û—ÙÜ›Ý\\Ý][Û—ÙÜ›Ý\ˆ™]ÚÙ\Ú[ÛœÛÛ\—ÚÜWÙ^WÜ›ÝÜÊˆÙ\ÜÚ[Û—ÛØš‹ˆ[™Ú[ÖÈ˜˜\ÙWÝ\›—Kˆ[™Ú[ÖÈ™^WÚÜWÙ[™Ú[—KˆÝ][Û—ÙÜ›Ý\ˆ[ÛÜÝ\ˆ
Kˆˆ™Z[NžØÚ[š×Ú[™^H‹ˆ
Bˆ›Üˆ›ÝÈ[ˆ›ÝÜÎ‚ˆ^\›˜[ÚYHÝŠ›ÝË™Ù]
œÝ][ÛÛÙHŠHÜˆ›ÝË™Ù]
œ[ÛÙHŠHÜˆˆŠKœÝš\

Bˆ\ÜÙ]H\ÜÙ]×ØžWÙ^\›˜[ÚY™Ù]
^\›˜[ÚY
Bˆ›Ý×Ù]HH\œÙWÙ\Ú[ÛœÛÛ\—ØÛÛXÝÙ]J›ÝË[ÛÜÝ\
BˆYˆ\ÜÙ]\È›Û™HÜˆ›Ý×Ù]H\È›Û™HÜˆ›Ý[ÛÜÝ\H›Ý×Ù]HH[ÛÙ[™‚ˆÛÛ[YBˆ™\Ý[HÝÜ™WÜ›ÙXÝ[Û—ÚÜWÜ™XÛÜ™
ˆÛÛ›‹ˆ\ÜÙ]Ü›ÝÏX\ÜÙ]ˆ›ÝšY\\›ÝšY\‹ˆ^\›˜[ÚYY^\›˜[ÚYˆ\š[ÙÝ\OH™^H‹ˆ\š[ÙÙ]O\›Ý×Ù]KˆÜWÜ›ÝÏ\›ÝËˆ›Ý\×Ü™Yš^H”™XÛÛ˜Ú[XXØ[ÈX\šXHÈ™XÚÈY[œØ[ˆ‹ˆ
BˆYˆ™\Ý[È\Ù\ÜÝ]\È—HOHœÚÚ\YÙ^\Ý[™×Ý˜[YŽ‚ˆÝ[[X\žVÈœ™XÛÜ™×Ý\]Y—H
ÏHBˆ›ØÙ\ÜÙYÙZ[WÝ\™Ù]Ë˜Y

[
\ÜÙ]È˜\ÜÙ]ÚY—JK›Ý×Ù]JJBˆYˆ™\Ý[Èœ›ÙXÝ[Û—ÚÝÚ—H\È›Û™N‚ˆÝ[[X\žVÈ›Z\ÜÚ[™×Ü›ÙXÝ[Ûˆ—H
ÏHBˆÛÛ›‹˜ÛÛ[Z]

B‚ˆ™XØ[Ý[]WØY™™XÝYÜ™Y™\™[˜Ù\Ê
Bˆ]X[]Y\ÈH]˜[X]WØ\ÜÙ]Ê
BˆÝ[[X\žVÈœÝ]\×ØY\ˆ—HHÛÝ[Û[ÛWÜ›ÙXÝ[Û—ÜÝ]\Ê]X[]Y\ÊBˆÝ[[X\žVÈœÝ]\È—HH˜ÛÛ\]Y‚ˆÛÛ›‹˜ÛÛ[Z]

Bˆ™]\›ˆÝ[[X\žB‚‚™Yˆ[—Ù\Ú[ÛœÛÛ\—Û[ÛØÞXÛJˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆ›ÝšY\ŽˆÝˆH‘\Ú[Û”ÛÛ\ˆ‹ˆ™\ÜÛ[ÛˆÝ‹ˆ\ÜÙ]ÚYÎˆ\ÝÚ[KˆÜWØØ[Ù[^WÜÙXÛÛ™Îˆ›Ø]›Û™HH›Û™KˆÛY\\Žˆ[žH›Û™HH›Û™KˆX^ÝØZ]ØÞXÛ\Îˆ[HŠHOˆXÝÜÝ‹[žWN‚ˆYˆ›Ý\ÜÙ]ÚYÎ‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ‘\ØÛÛH[ÈY[›ÜÈ[XH[œÝ[XØ[È\˜HÈÚXÛËˆŠBˆ[ÛÜÝ\H]][YKœÝœ[YJ›Ü›X[^™WÜ™\ÜÛ[Û
™\ÜÛ[Û
K‰VKI[HŠK™]J
BˆË[ÛÛ\ÝÙ^HHØ[[™\‹›[Û˜[™ÙJ[ÛÜÝ\žYX\‹[ÛÜÝ\›[Û
Bˆ[ÛÙ[™H[ÛÜÝ\œ™\XÙJ^O[[ÛÛ\ÝÙ^JBˆÙ^WÝ˜[YHH]KÙ^J
Bˆ]WÝÈHZ[Š[ÛÙ[™Ù^WÝ˜[YHH[YY[J^\ÏLJJBˆYˆ]WÝÈ[ÛÜÝ\‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ“ÈY\È\ØÛÛYÈZ[™H˜[È[HX\È™XÚYÜÈ\˜H[\Ü\‹ˆŠB‚ˆÛÛ™šYÈHÙ]Ú[YÜ˜][Û—ØÛÛ™šYÊÛÛ›‹›ÝšY\ŠBˆYˆÛÛ™šYÈ\È›Û™N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠÛÛ™šYÝ\˜XØ[È\Ú[Û”ÛÛ\ˆ˜[È[˜ÛÛ˜YKˆŠBˆYˆ›ÝÛÛ™šYÖÈ™[˜X›Y—N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠH[YÜ˜XØ[È\Ú[Û”ÛÛ\ˆ\ÝH\Ø]]˜YKˆŠBˆ[™Ú[ÈHÙ]Ù\Ú[ÛœÛÛ\—Ù[™Ú[ØÛÛ™šYÊÛÛ™šYÊBˆ\ÜÙ]ÈHÙ]Ù\Ú[ÛœÛÛ\—Ü\™›Ü›X[˜ÙWØ\ÜÙ]ÊÛÛ›‹›ÝšY\‹\ÜÙ]ÚYÏX\ÜÙ]ÚYÊBˆYˆ›Ý\ÜÙ]Î‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ“™[š[XH\È[œÝ[XÛÙ\È\ØÛÛY\È[HX\X[Y[È\Ú[Û”ÛÛ\ˆ]]›ËˆŠB‚ˆÝ][Û—ØÛÙ\ÈHÜÝŠ\ÜÙ]È™^\›˜[ÚY—HÜˆˆŠKœÝš\

H›Üˆ\ÜÙ][ˆ\ÜÙ]ÈYˆÝŠ\ÜÙ]È™^\›˜[ÚY—HÜˆˆŠKœÝš\

WBˆ\ÜÙ]×ØžWÙ^\›˜[ÚYHÜÝŠ\ÜÙ]È™^\›˜[ÚY—HÜˆˆŠKœÝš\

Nˆ\ÜÙ]›Üˆ\ÜÙ][ˆ\ÜÙ]ÈYˆÝŠ\ÜÙ]È™^\›˜[ÚY—HÜˆˆŠKœÝš\

_BˆÝ][Û—ØÚ[šÜÈHÚ[šÙY
Ý][Û—ØÛÙ\ËL
BˆÙ\ÜÚ[Û—ÛØš‹ÈHÙ]Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[ÛŠÛÛ™šYÊBˆÛY\Ù[˜ÈHÛY\\ˆÜˆ[YKœÛY\ˆÜWØØ[Ù[^WÜÙXÛÛ™ÈH•TÒSÓ”ÓÓT—ÔT‘“Ô“PSÑWÒÔWÑSVWÔÑPÓÓ‘ÈYˆÜWØØ[Ù[^WÜÙXÛÛ™È\È›Û™H[ÙHÜWØØ[Ù[^WÜÙXÛÛ™ÂˆÙÙÙ\ˆHÝ\œ™[Ø\›ÙÙÙ\ˆYˆ\×Ø\ØÛÛ^

H[ÙHÙÙÚ[™Ë™Ù]ÙÙÙ\Š×Û˜[YW×ÊBˆÝ[[X\žHHÂˆ˜\ÜÙ]×ÜÙ[XÝYŽˆ[Š\ÜÙ]ÚYÊKˆ˜\ÜÙ]×ÛX\YŽˆ[Š\ÜÙ]×ØžWÙ^\›˜[ÚY
Kˆ›[ÛŽˆ[ÛÜÝ\œÝ™[YJ‰VKI[HŠKˆ™Z[WØÚ[šÜ×ÝÝ[Žˆ[ŠÝ][Û—ØÚ[šÜÊKˆ™Z[WØÚ[šÜ×ØÛÛ\]YŽˆˆ›[ÛWØÚ[šÜ×ØÛÛ\]YŽˆˆœ™XÛÜ™×Ý\]YŽˆˆ›[ÛWÜ™XÛÜ™×Ý\]YŽˆˆ›Z\ÜÚ[™×Ü›ÙXÝ[ÛˆŽˆˆ˜\WØØ[×Ý\ÙYŽˆˆ˜\WÙ\œ›ÜœÈŽˆˆØZ]ØÞXÛ\ÈŽˆˆœÝ]\ÈŽˆœ[›š[™È‹ˆBˆ›ØÙ\ÜÙYÙ]\ÎˆÙ]Ù]WHHÙ]

B‚ˆYˆØZ]ØY\—Ü˜]WÛ[Z]
™X\ÛÛŽˆÝŠHOˆ›Û™N‚ˆÝ[[X\žVÈØZ]ØÞXÛ\È—H
ÏHBˆYˆÝ[[X\žVÈØZ]ØÞXÛ\È—HˆX^ÝØZ]ØÞXÛ\Î‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠˆ“[Z]H\Ú[Û”ÛÛ\ˆ™\]YÈ[X\ÚXY\È™^™\Ëˆ[[[È\ÝYÎˆÜ™X\ÛÛŸHŠBˆÛÛ›‹˜ÛÛ[Z]

BˆÙXÛÛ™ÈH\Ú[ÛœÛÛ\—ØÛÛÛÝÛ—ÜÛY\ÜÙXÛÛ™ÊÛÛ›ŠBˆÙÙÙ\‹Ø\›š[™Êˆ‘\Ú[Û”ÛÛ\ˆ[ÛÞXÛHØZ][™È›ÜˆTHÛÛÛÝÛŽˆ[ÛI\ÈÙXÛÛ™ÏI\ÈØZ]ØÞXÛOI\È‹ˆ[ÛÜÝ\š\ÛÙ›Ü›X]

KˆÙXÛÛ™ËˆÝ[[X\žVÈØZ]ØÞXÛ\È—Kˆ
BˆÛY\Ù[˜ÊÙXÛÛ™ÊB‚ˆYˆØZ]Ø™]ÙY[—ØØ[Ê
HOˆ›Û™N‚ˆYˆ
ˆ“ÑPÕSÓ—ÒÔWÐÐSÐÓÓ•V™Ù]

H\È›Û™Bˆ[™Ý[[X\žVÈ˜\WØØ[×Ý\ÙY—Hˆˆ[™ÜWØØ[Ù[^WÜÙXÛÛ™Âˆ[™ÜWØØ[Ù[^WÜÙXÛÛ™Èˆˆ
N‚ˆÛY\Ù[˜ÊÜWØØ[Ù[^WÜÙXÛÛ™ÊB‚ˆYˆ™Yœ™\ÚÜÙ\ÜÚ[Û—ØY\—Ù^\žJ^Îˆ^Ù\[Û‹ÛÛ^ˆÝŠHOˆ›Û™N‚ˆ›Û›ØØ[Ù\ÜÚ[Û—ÛØš‚ˆ[˜[Y]WÙ\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[ÛŠÛÛ™šYÊBˆÙ\ÜÚ[Û—ÛØš‹ÈHÙ]Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[ÛŠÛÛ™šYË›Ü˜ÙWÛÙÚ[UYJBˆÙÙÙ\‹Ø\›š[™Ê‘\Ú[Û”ÛÛ\ˆÙ\ÜÚ[Ûˆ™Yœ™\ÚYY\ˆ^\™YÙÚ[ŽˆÛÛ^I\È\œ›ÜI\È‹ÛÛ^^ÊB‚ˆ›ÜˆÚ[š×Ú[™^Ý][Û—ÙÜ›Ý\[ˆ[[Y\˜]JÝ][Û—ØÚ[šÜËÝ\LJN‚ˆÙ\ÜÚ[Û—Ü™]žWÝ\ÙYH˜[ÙBˆÚ[HYN‚ˆžN‚ˆØZ]Ø™]ÙY[—ØØ[Ê
Bˆ›ÝÜÈH™]ÚÙ\Ú[ÛœÛÛ\—ÚÜWÙ^WÜ›ÝÜÊˆÙ\ÜÚ[Û—ÛØš‹ˆ[™Ú[ÖÈ˜˜\ÙWÝ\›—Kˆ[™Ú[ÖÈ™^WÚÜWÙ[™Ú[—KˆÝ][Û—ÙÜ›Ý\ˆ[ÛÜÝ\ˆ
BˆÝ[[X\žVÈ˜\WØØ[×Ý\ÙY—H
ÏHBˆ›Üˆ›ÝÈ[ˆ›ÝÜÎ‚ˆ^\›˜[ÚYÝ˜[YHHÝŠ›ÝË™Ù]
œÝ][ÛÛÙHŠHÜˆ›ÝË™Ù]
œ[ÛÙHŠHÜˆˆŠKœÝš\

Bˆ\ÜÙ]H\ÜÙ]×ØžWÙ^\›˜[ÚY™Ù]
^\›˜[ÚYÝ˜[YJBˆYˆ\ÜÙ]\È›Û™N‚ˆÛÛ[YBˆ›Ý×Ù]HH\œÙWÙ\Ú[ÛœÛÛ\—ØÛÛXÝÙ]J›ÝË[ÛÜÝ\
BˆYˆ›Ý×Ù]H\È›Û™HÜˆ›Ý×Ù]H[ÛÜÝ\Üˆ›Ý×Ù]Hˆ]WÝÎ‚ˆÛÛ[YBˆ™\Ý[HÝÜ™WÜ›ÙXÝ[Û—ÚÜWÜ™XÛÜ™
ˆÛÛ›‹ˆ\ÜÙ]Ü›ÝÏX\ÜÙ]ˆ›ÝšY\\›ÝšY\‹ˆ^\›˜[ÚYY^\›˜[ÚYÝ˜[YKˆ\š[ÙÝ\OH™^H‹ˆ\š[ÙÙ]O\›Ý×Ù]KˆÜWÜ›ÝÏ\›ÝËˆ›Ý\×Ü™Yš^HÚXÛÈY[œØ[]]ÛX]XÛËˆ‹ˆ
BˆYˆ™\Ý[È\Ù\ÜÝ]\È—HOHœÚÚ\YÙ^\Ý[™×Ý˜[YŽ‚ˆÝ[[X\žVÈœ™XÛÜ™×Ý\]Y—H
ÏHBˆ›ØÙ\ÜÙYÙ]\Ë˜Y
›Ý×Ù]JBˆYˆ™\Ý[Èœ›ÙXÝ[Û—ÚÝÚ—H\È›Û™N‚ˆÝ[[X\žVÈ›Z\ÜÚ[™×Ü›ÙXÝ[Ûˆ—H
ÏHBˆÝ[[X\žVÈ™Z[WØÚ[šÜ×ØÛÛ\]Y—HHÚ[š×Ú[™^ˆÛÛ›‹˜ÛÛ[Z]

Bˆœ™XZÂˆ^Ù\^Ù\[Ûˆ\È^Î‚ˆÝ[[X\žVÈ˜\WÙ\œ›ÜœÈ—H
ÏHBˆYˆ\Ú[œÝ[˜ÙJ^Ë
\T˜]S[Z]\œ›Ü‹\TÛÝ[˜]˜Z[X›Q\œ›ÜŠJN‚ˆ˜Z\ÙBˆYˆ\×Ù\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]Ù\œ›ÜŠ^ÊN‚ˆØZ]ØY\—Ü˜]WÛ[Z]
X\š×Ù\Ú[ÛœÛÛ\—Ü\™›Ü›X[˜ÙWÜ˜]WÛ[Z]Y
ÛÛ›ŠJBˆÛÛ[YBˆYˆ\×Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[Û—Ù^\™YÙ\œ›ÜŠ^ÊH[™›ÝÙ\ÜÚ[Û—Ü™]žWÝ\ÙY‚ˆÙ\ÜÚ[Û—Ü™]žWÝ\ÙYHYBˆ™Yœ™\ÚÜÙ\ÜÚ[Û—ØY\—Ù^\žJ^Ëˆ›[ÛXÞXÛKY^NžÛ[ÛÜÝ\š\ÛÙ›Ü›X]

_N˜Ú[šÎžØÚ[š×Ú[™^HŠBˆÛÛ[YBˆ˜Z\ÙB‚ˆ›ÜˆÚ[š×Ú[™^Ý][Û—ÙÜ›Ý\[ˆ[[Y\˜]JÝ][Û—ØÚ[šÜËÝ\LJN‚ˆÙ\ÜÚ[Û—Ü™]žWÝ\ÙYH˜[ÙBˆÚ[HYN‚ˆžN‚ˆØZ]Ø™]ÙY[—ØØ[Ê
BˆÜWÛX\H™]ÚÙ\Ú[ÛœÛÛ\—ÚÜWÛ[ÛÛX\
ˆÙ\ÜÚ[Û—ÛØš‹ˆ[™Ú[ÖÈ˜˜\ÙWÝ\›—Kˆ[™Ú[ÖÈ›[ÛÚÜWÙ[™Ú[—KˆÝ][Û—ÙÜ›Ý\ˆ[ÛÜÝ\ˆ
BˆÝ[[X\žVÈ˜\WØØ[×Ý\ÙY—H
ÏHBˆ›Üˆ^\›˜[ÚYÝ˜[YH[ˆÝ][Û—ÙÜ›Ý\‚ˆ\ÜÙ]H\ÜÙ]×ØžWÙ^\›˜[ÚY™Ù]
^\›˜[ÚYÝ˜[YJBˆYˆ\ÜÙ]\È›Û™N‚ˆÛÛ[YBˆ™\Ý[HÝÜ™WÜ›ÙXÝ[Û—ÚÜWÜ™XÛÜ™
ˆÛÛ›‹ˆ\ÜÙ]Ü›ÝÏX\ÜÙ]ˆ›ÝšY\\›ÝšY\‹ˆ^\›˜[ÚYY^\›˜[ÚYÝ˜[YKˆ\š[ÙÝ\OH›[Û‹ˆ\š[ÙÙ]O[[ÛÜÝ\ˆÜWÜ›ÝÏZÜWÛX\™Ù]
^\›˜[ÚYÝ˜[YKßJKˆ›Ý\×Ü™Yš^HÚXÛÈY[œØ[]]ÛX]XÛËˆ‹ˆ
BˆYˆ™\Ý[È\Ù\ÜÝ]\È—HOHœÚÚ\YÙ^\Ý[™×Ý˜[YŽ‚ˆÝ[[X\žVÈ›[ÛWÜ™XÛÜ™×Ý\]Y—H
ÏHBˆYˆ™\Ý[Èœ›ÙXÝ[Û—ÚÝÚ—H\È›Û™N‚ˆÝ[[X\žVÈ›Z\ÜÚ[™×Ü›ÙXÝ[Ûˆ—H
ÏHBˆÝ[[X\žVÈ›[ÛWØÚ[šÜ×ØÛÛ\]Y—HHÚ[š×Ú[™^ˆÛÛ›‹˜ÛÛ[Z]

Bˆœ™XZÂˆ^Ù\^Ù\[Ûˆ\È^Î‚ˆÝ[[X\žVÈ˜\WÙ\œ›ÜœÈ—H
ÏHBˆYˆ\Ú[œÝ[˜ÙJ^Ë
\T˜]S[Z]\œ›Ü‹\TÛÝ[˜]˜Z[X›Q\œ›ÜŠJN‚ˆ˜Z\ÙBˆYˆ\×Ù\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]Ù\œ›ÜŠ^ÊN‚ˆØZ]ØY\—Ü˜]WÛ[Z]
X\š×Ù\Ú[ÛœÛÛ\—Ü\™›Ü›X[˜ÙWÜ˜]WÛ[Z]Y
ÛÛ›ŠJBˆÛÛ[YBˆYˆ\×Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[Û—Ù^\™YÙ\œ›ÜŠ^ÊH[™›ÝÙ\ÜÚ[Û—Ü™]žWÝ\ÙY‚ˆÙ\ÜÚ[Û—Ü™]žWÝ\ÙYHYBˆ™Yœ™\ÚÜÙ\ÜÚ[Û—ØY\—Ù^\žJ^Ëˆ›[ÛXÞXÛK[[ÛžÛ[ÛÜÝ\š\ÛÙ›Ü›X]

_N˜Ú[šÎžØÚ[š×Ú[™^HŠBˆÛÛ[YBˆ˜Z\ÙB‚ˆ›Üˆ\™Ù][ˆÛÜY
›ØÙ\ÜÙYÙ]\ÊN‚ˆ›ÜˆÙ[XÝYØ\ÜÙ]ÚY[ˆ\ÜÙ]ÚYÎ‚ˆ™XØ[Ý[]WÜ\™›Ü›X[˜ÙWÜ™Y™\™[˜Ù\ÊˆÛÛ›‹ˆ\š[ÙÝ\OH™^H‹ˆ\š[ÙÙ]O]\™Ù]ˆ\ÜÙ]ÚY\Ù[XÝYØ\ÜÙ]ÚYˆ›ÝšY\\›ÝšY\‹ˆÙ^WÝ˜[YO]Ù^WÝ˜[YKˆ
Bˆ›ÜˆÙ[XÝYØ\ÜÙ]ÚY[ˆ\ÜÙ]ÚYÎ‚ˆ™XØ[Ý[]WÜ\™›Ü›X[˜ÙWÜ™Y™\™[˜Ù\ÊˆÛÛ›‹ˆ\š[ÙÝ\OH›[Û‹ˆ\š[ÙÙ]O[[ÛÜÝ\ˆ\ÜÙ]ÚY\Ù[XÝYØ\ÜÙ]ÚYˆ›ÝšY\\›ÝšY\‹ˆÙ^WÝ˜[YO]Ù^WÝ˜[YKˆ
BˆÛÛ›‹˜ÛÛ[Z]

BˆÝ[[X\žVÈœÝ]\È—HH˜ÛÛ\]Y‚ˆ™]\›ˆÝ[[X\žB‚‚™YˆÜ™X]WÚ[YÜ˜][Û—Ü[ŠÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹›ÝšY\ŽˆÝ‹šYÙÙ\—Ý\NˆÝŠHOˆ[‚ˆÝ\œÛÜˆHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆS”ÑT•S•È[YÜ˜][Û—ÜÞ[˜×Ü[œÈ
›ÝšY\‹Ý\YØ]šYÙÙ\—Ý\KÝ]\ÊBˆSQTÈ
ËËËÊBˆˆˆ‹ˆ
›ÝšY\‹]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKšYÙÙ\—Ý\Kœ[›š[™ÈŠKˆ
Bˆ™]\›ˆ[
Ý\œÛÜ‹›\Ý›ÝÚY
B‚‚™Yˆš[˜[^™WÚ[YÜ˜][Û—Ü[ŠˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ[—ÚYˆ[ˆ
‹ˆÝ]\ÎˆÝ‹ˆX]ÚYØÛÝ[ˆ[ˆ[œ™\ÛÛ™YØÛÝ[ˆ[ˆ]]×Ü™\ÛÛ™YØÛÝ[ˆ[ˆ\œ›Ü—ÛY\ÜØYÙNˆÝˆHˆ‹ˆÝ[[X\žWÚœÛÛŽˆXÝÜÝ‹[žWH›Û™HH›Û™KŠHOˆ›Û™N‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUH[YÜ˜][Û—ÜÞ[˜×Ü[œÂˆÑUš[š\ÚYØ]HËÝ]\ÈHËX]ÚYØÛÝ[HË[œ™\ÛÛ™YØÛÝ[HËˆ]]×Ü™\ÛÛ™YØÛÝ[HË\œ›Ü—ÛY\ÜØYÙHHËÝ[[X\žWÚœÛÛˆHÂˆÒT‘HYHÂˆˆˆ‹ˆ
ˆ]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆÝ]\ËˆX]ÚYØÛÝ[ˆ[œ™\ÛÛ™YØÛÝ[ˆ]]×Ü™\ÛÛ™YØÛÝ[ˆ\œ›Ü—ÛY\ÜØYÙKˆœÛÛ‹™[\ÊÝ[[X\žWÚœÛÛˆÜˆßK[œÝ\™WØ\ØÚZOUYJKˆ[—ÚYˆ
Kˆ
B‚‚™YˆÜ™X]WÛÜ—Ý\]WØ\ÜÙ]Ú[YÜ˜][ÛŠˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ\ÜÙ]ÚYˆ[ˆ›ÝšY\ŽˆÝ‹ˆ^\›˜[ÚYˆÝ‹ˆ^\›˜[Û˜[YNˆÝ‹ˆÝ]\ÎˆÝ‹ŠHOˆ›Û™N‚ˆ^\Ý[™ÈH›Û™BˆYˆ^\›˜[ÚY‚ˆ^\Ý[™ÈHÛÛ›‹™^XÝ]Jˆ”ÑSPÕY”“ÓH\ÜÙ]Ú[YÜ˜][ÛœÈÒT‘H›ÝšY\ˆHÈS‘^\›˜[ÚYHÈ‹ˆ
›ÝšY\‹^\›˜[ÚY
Kˆ
K™™]ÚÛ™J
BˆYˆ^\Ý[™Î‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUH\ÜÙ]Ú[YÜ˜][ÛœÂˆÑU\ÜÙ]ÚYHË^\›˜[Û˜[YHHË[˜X›YHK\ÝÜÞ[˜×Ø]HË\ÝÜÝ]\ÈHË\ÝÙ\œ›ÜˆH	ÉÂˆÒT‘HYHÂˆˆˆ‹ˆ
\ÜÙ]ÚY^\›˜[Û˜[YK]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKÝ]\Ë^\Ý[™ÖÈšY—JKˆ
Bˆ™]\›‚‚ˆØ[™Y]HHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕYˆ”“ÓH\ÜÙ]Ú[YÜ˜][ÛœÂˆÒT‘H›ÝšY\ˆHÈS‘\ÜÙ]ÚYHÂˆSRUBˆˆˆ‹ˆ
›ÝšY\‹\ÜÙ]ÚY
Kˆ
K™™]ÚÛ™J
BˆYˆØ[™Y]N‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUH\ÜÙ]Ú[YÜ˜][ÛœÂˆÑU^\›˜[ÚYHË^\›˜[Û˜[YHHË[˜X›YHK\ÝÜÞ[˜×Ø]HË\ÝÜÝ]\ÈHË\ÝÙ\œ›ÜˆH	ÉÂˆÒT‘HYHÂˆˆˆ‹ˆ
^\›˜[ÚYÜˆ›Û™K^\›˜[Û˜[YK]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKÝ]\ËØ[™Y]VÈšY—JKˆ
Bˆ™]\›‚‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆS”ÑT•S•È\ÜÙ]Ú[YÜ˜][ÛœÈ
\ÜÙ]ÚY›ÝšY\‹^\›˜[ÚY^\›˜[Û˜[YK[˜X›Y\ÝÜÞ[˜×Ø]\ÝÜÝ]\Ë\ÝÙ\œ›ÜŠBˆSQTÈ
ËËËËKËË	ÉÊBˆˆˆ‹ˆ
ˆ\ÜÙ]ÚYˆ›ÝšY\‹ˆ^\›˜[ÚYÜˆ›Û™Kˆ^\›˜[Û˜[YKˆ]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆÝ]\Ëˆ
Kˆ
B‚‚™Yˆ\Ù\Ú[YÜ˜][Û—Ý[œ™\ÛÛ™Y
ˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆ›ÝšY\ŽˆÝ‹ˆ[—ÚYˆ[ˆ^\›˜[ÚYˆÝ‹ˆ^\›˜[Û˜[YNˆÝ‹ˆÝ]\ÎˆÝ‹ˆ^[ØYˆXÝÜÝ‹[žWKŠHOˆ›Û™N‚ˆ›Ü›X[^™YÛ˜[YHH›Ü›X[^™WÛ˜[YJ^\›˜[Û˜[YJBˆÝYÙÙ\ÝYØ\ÜÙ]ÚYHš[™ÜÝYÙÙ\ÝYØ\ÜÙ]ÚY
ÛÛ›‹^\›˜[Û˜[YJBˆ^\Ý[™ÈHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕYˆ”“ÓH[YÜ˜][Û—Ý[œ™\ÛÛ™YˆÒT‘H›ÝšY\ˆHÈS‘›Ü›X[^™YÛ˜[YHHÈS‘™\ÛÛ][Û—ÜÝ]\ÈH	Ü[™[™ÉÂˆSRUBˆˆˆ‹ˆ
›ÝšY\‹›Ü›X[^™YÛ˜[YJKˆ
K™™]ÚÛ™J
BˆYˆ^\Ý[™Î‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUH[YÜ˜][Û—Ý[œ™\ÛÛ™YˆÑUÞ[˜×Ü[—ÚYHË^\›˜[ÚYHË^\›˜[ÜÝ]\ÈHË^[ØYÚœÛÛˆHËÝYÙÙ\ÝYØ\ÜÙ]ÚYHËÜ™X]YØ]HÂˆÒT‘HYHÂˆˆˆ‹ˆ
ˆ[—ÚYˆ^\›˜[ÚYÜˆ›Û™KˆÝ]\ËˆœÛÛ‹™[\Ê^[ØY[œÝ\™WØ\ØÚZOUYJKˆÝYÙÙ\ÝYØ\ÜÙ]ÚYˆ]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆ^\Ý[™ÖÈšY—Kˆ
Kˆ
Bˆ™]\›‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆS”ÑT•S•È[YÜ˜][Û—Ý[œ™\ÛÛ™Y
ˆ›ÝšY\‹Þ[˜×Ü[—ÚY^\›˜[ÚY^\›˜[Û˜[YK›Ü›X[^™YÛ˜[YK^\›˜[ÜÝ]\Ëˆ^[ØYÚœÛÛ‹ÝYÙÙ\ÝYØ\ÜÙ]ÚY™\ÛÛ][Û—ÜÝ]\ËÜ™X]YØ]ˆ
HSQTÈ
ËËËËËËËË	Ü[™[™ÉËÊBˆˆˆ‹ˆ
ˆ›ÝšY\‹ˆ[—ÚYˆ^\›˜[ÚYÜˆ›Û™Kˆ^\›˜[Û˜[YKˆ›Ü›X[^™YÛ˜[YKˆÝ]\ËˆœÛÛ‹™[\Ê^[ØY[œÝ\™WØ\ØÚZOUYJKˆÝYÙÙ\ÝYØ\ÜÙ]ÚYˆ]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆ
Kˆ
B‚‚™YˆÙ]Û]\ÝØ]˜Z[Xš[]WØžWØ\ÜÙ]
ÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹\ÜÙ]ÚYˆ[
HOˆXÝÜÝ‹[žWH›Û™N‚ˆ›ÝÈHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕ
‚ˆ”“ÓH]˜Z[Xš[]WÙZ[BˆÒT‘H\ÜÙ]ÚYHÂˆÔ‘Tˆ–H\š[ÙÙ]HTÐËYTÐÂˆSRUBˆˆˆ‹ˆ
\ÜÙ]ÚY
Kˆ
K™™]ÚÛ™J
Bˆ™]\›ˆXÝ
›ÝÊHYˆ›ÝÈ[ÙH›Û™B‚‚™YˆÙ]Ù\Ú›Ø\™Ø]˜Z[Xš[]WÜÝ[[X\žJÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[ÛŠHOˆXÝÜÝ‹[žWN‚ˆ›ÝÜÈH]Y\žWØ[
ˆÛÛ›‹ˆˆˆ‚ˆÑSPÕYŠ‚ˆ”“ÓH]˜Z[Xš[]WÙZ[HYˆ“ÒSˆ
ˆÑSPÕ\ÜÙ]ÚYPV
\š[ÙÙ]H	Õ	Èš[Š	ÉLY	ËY
JHTÈX\šÙ\‚ˆ”“ÓH]˜Z[Xš[]WÙZ[BˆÔ“ÕT–H\ÜÙ]ÚYˆ
H]\ÝˆÓˆ]\Ý˜\ÜÙ]ÚYHY˜\ÜÙ]ÚYˆS‘]\Ý›X\šÙ\ˆHYœ\š[ÙÙ]H	Õ	Èš[Š	ÉLY	ËYšY
Bˆˆˆ‚ˆ
BˆÝÝ˜[Y\ÈHÜ›ÝÖÈš[™\\—Ø]˜Z[Xš[]WÜÝ—H›Üˆ›ÝÈ[ˆ›ÝÜÈYˆ›ÝÖÈš[™\\—Ø]˜Z[Xš[]WÜÝ—H\È›Ý›Û™WBˆY™™XÝYÝ˜[Y\ÈHÜ›ÝÖÈ˜Y™™XÝYÜÝÙ\—ÚÝÈ—H›Üˆ›ÝÈ[ˆ›ÝÜÈYˆ›ÝÖÈ˜Y™™XÝYÜÝÙ\—ÚÝÈ—H\È›Ý›Û™WBˆ™]\›ˆÂˆ˜]™\˜YÙWÚ[™\\—Ø]˜Z[Xš[]WÜÝŽˆ›Ý[™
Ý[JÝÝ˜[Y\ÊHÈ[ŠÝÝ˜[Y\ÊKŠHYˆÝÝ˜[Y\È[ÙH›Û™Kˆ[˜]˜Z[X›WÚ[™\\œÈŽˆÝ[J[
›ÝÖÈ[˜]˜Z[X›WÚ[™\\œÈ—HÜˆ
H›Üˆ›ÝÈ[ˆ›ÝÜÊKˆ››×ØÛÛ[][šXØ][Û—Ù]šXÙ\ÈŽˆÝ[J[
›ÝÖÈ››×ØÛÛ[][šXØ][Û—Ù]šXÙ\È—HÜˆ
H›Üˆ›ÝÈ[ˆ›ÝÜÊKˆ˜Y™™XÝYÜÝÙ\—ÚÝÈŽˆ›Ý[™
Ý[J›Ø]
˜[YJH›Üˆ˜[YH[ˆY™™XÝYÝ˜[Y\ÊKŠHYˆY™™XÝYÝ˜[Y\È[ÙH›Û™KˆœÝš[™×Ø]˜Z[Xš[]WÜÝŽˆ
ˆ›Ý[™
ˆÝ[J[
›ÝÖÈ˜]˜Z[X›WÜÝš[™ÜÈ—HÜˆ
H›Üˆ›ÝÈ[ˆ›ÝÜÊBˆÈÝ[J[
›ÝÖÈÝ[ÜÝš[™ÜÈ—HÜˆ
H›Üˆ›ÝÈ[ˆ›ÝÜÊBˆ
ˆLˆ‹ˆ
BˆYˆÝ[J[
›ÝÖÈÝ[ÜÝš[™ÜÈ—HÜˆ
H›Üˆ›ÝÈ[ˆ›ÝÜÊBˆ[ÙH›Û™Bˆ
Kˆ[˜]˜Z[X›WÜÝš[™ÜÈŽˆÝ[J[
›ÝÖÈ[˜]˜Z[X›WÜÝš[™ÜÈ—HÜˆ
H›Üˆ›ÝÈ[ˆ›ÝÜÊKˆB‚‚™YˆÙ]Û]\ÝÙ]šXÙWÜ›ÝÜ×Ù›Ü—Ø\ÜÙ]
ÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹\ÜÙ]ÚYˆ[
HOˆ\ÝÙXÝÜÝ‹[žWWN‚ˆ™]\›ˆ]Y\žWØ[
ˆÛÛ›‹ˆˆˆ‚ˆÑSPÕˆ™]šXÙWÛ˜[YKˆœ˜]YÜÝÙ\—ÚÝËˆ›\ÝÜÙY[—Ø]ˆœËŠ‚ˆ”“ÓH›ÝšY\—Ù]šXÙ\ÈˆQ•“ÒSˆ]šXÙWÜ™X[[YWÜÛ˜\ÚÝÈœÂˆÓˆœËšYH
ˆÑSPÕ]\ÝšYˆ”“ÓH]šXÙWÜ™X[[YWÜÛ˜\ÚÝÈ]\ÝˆÒT‘H]\Ýœ›ÝšY\—Ù]šXÙWÚYHšYˆÔ‘Tˆ–H]\Ý˜ÛÛXÝYØ]TÐË]\ÝšYTÐÂˆSRUBˆ
BˆÒT‘H˜\ÜÙ]ÚYHÈS‘™[˜X›YHBˆÔ‘Tˆ–H™]šXÙWÛ˜[YHÓÓUH“ÐÐTÑKšYˆˆˆ‹ˆ
\ÜÙ]ÚY
Kˆ
B‚‚™Yˆ\Ù\Ü›ÝšY\—Ù]šXÙJÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹\ÜÙ]ÚYˆ[›ÝšY\ŽˆÝ‹›ÝÎˆXÝÜÝ‹[žWJHOˆ[‚ˆ›ÝÈH]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠBˆ[˜X›YH[
›ÝË™Ù]
™[˜X›Y‹JJBˆ^\Ý[™ÈHÛÛ›‹™^XÝ]Jˆ”ÑSPÕY[˜X›Y”“ÓH›ÝšY\—Ù]šXÙ\ÈÒT‘H›ÝšY\ˆHÈS‘^\›˜[Ù]šXÙWÚYHÈ‹ˆ
›ÝšY\‹›ÝÖÈ™^\›˜[Ù]šXÙWÚY—JKˆ
K™™]ÚÛ™J
Bˆ^[ØYÚœÛÛˆHœÛÛ‹™[\Ê›ÝÖÈœ^[ØY—K[œÝ\™WØ\ØÚZOUYJBˆYˆ^\Ý[™Î‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUH›ÝšY\—Ù]šXÙ\ÂˆÑU\ÜÙ]ÚYHËÝ][Û—ØÛÙHHË]—ÙˆHËÛˆHË]šXÙWÛ˜[YHHË]—Ý\WÚYHËˆ[Ù[HË˜]YÜÝÙ\—ÚÝÈHË[˜X›YHË^[ØYÚœÛÛˆHË\]YØ]HÂˆÒT‘HYHÂˆˆˆ‹ˆ
ˆ\ÜÙ]ÚYˆ›ÝÖÈœÝ][Û—ØÛÙH—Kˆ›ÝÖÈ™]—Ùˆ—Kˆ›ÝÖÈœÛˆ—Kˆ›ÝÖÈ™]šXÙWÛ˜[YH—Kˆ›ÝÖÈ™]—Ý\WÚY—Kˆ›ÝÖÈ›[Ù[—Kˆ›ÝÖÈœ˜]YÜÝÙ\—ÚÝÈ—Kˆ[˜X›Yˆ^[ØYÚœÛÛ‹ˆ›ÝËˆ^\Ý[™ÖÈšY—Kˆ
Kˆ
Bˆ]šXÙWÚYH[
^\Ý[™ÖÈšY—JBˆ™XÛÜ™Ù]šXÙWØÛÛ™šYÝ\˜][ÛŠˆÛÛ›‹ˆ›ÝšY\—Ù]šXÙWÚYY]šXÙWÚYˆXÝ]™OX›ÛÛ
[˜X›Y
KˆY™™XÝ]™WÙ]OXÝ\œ™[Û\Ø›Û—Ù]J
Kˆ
Bˆ™]\›ˆ]šXÙWÚYˆÝ\œÛÜˆHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆS”ÑT•S•È›ÝšY\—Ù]šXÙ\È
ˆ\ÜÙ]ÚY›ÝšY\‹Ý][Û—ØÛÙK^\›˜[Ù]šXÙWÚY]—Ù‹Û‹]šXÙWÛ˜[YK]—Ý\WÚYˆ[Ù[˜]YÜÝÙ\—ÚÝË[˜X›Y^[ØYÚœÛÛ‹Ü™X]YØ]\]YØ]ˆ
HSQTÈ
ËËËËËËËËËËËËËÊBˆˆˆ‹ˆ
ˆ\ÜÙ]ÚYˆ›ÝšY\‹ˆ›ÝÖÈœÝ][Û—ØÛÙH—Kˆ›ÝÖÈ™^\›˜[Ù]šXÙWÚY—Kˆ›ÝÖÈ™]—Ùˆ—Kˆ›ÝÖÈœÛˆ—Kˆ›ÝÖÈ™]šXÙWÛ˜[YH—Kˆ›ÝÖÈ™]—Ý\WÚY—Kˆ›ÝÖÈ›[Ù[—Kˆ›ÝÖÈœ˜]YÜÝÙ\—ÚÝÈ—Kˆ[˜X›Yˆ^[ØYÚœÛÛ‹ˆ›ÝËˆ›ÝËˆ
Kˆ
Bˆ]šXÙWÚYH[
Ý\œÛÜ‹›\Ý›ÝÚY
Bˆ™XÛÜ™Ù]šXÙWØÛÛ™šYÝ\˜][ÛŠˆÛÛ›‹ˆ›ÝšY\—Ù]šXÙWÚYY]šXÙWÚYˆXÝ]™OX›ÛÛ
[˜X›Y
KˆY™™XÝ]™WÙ]OXÝ\œ™[Û\Ø›Û—Ù]J
Kˆ
Bˆ™]\›ˆ]šXÙWÚY‚‚™Yˆ™\\™WÙ\Ú[ÛœÛÛ\—Ú[™\\—Ú\ÝÜžWØÛÛ^
ˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆ›Ü˜ÙWÛÙÚ[Žˆ›ÛÛH˜[ÙKˆ\ÝÜžWØØ[Ù[^WÜÙXÛÛ™Îˆ›Ø]HˆÛY\\Žˆ[žHH[YKœÛY\ŠHOˆXÝÜÝ‹[žWN‚ˆ›ÝšY\ˆHS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‚ˆÛÛ™šYÈHÙ]Ú[YÜ˜][Û—ØÛÛ™šYÊÛÛ›‹›ÝšY\ŠBˆYˆÛÛ™šYÈ\È›Û™HÜˆ›ÝÛÛ™šYÖÈ™[˜X›Y—N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠÛÛ™šYÝ\˜XØ[È\Ú[Û”ÛÛ\ˆ[™\ÜÛš]™[ˆŠBˆ[™Ú[ÈHÙ]Ù\Ú[ÛœÛÛ\—Ù[™Ú[ØÛÛ™šYÊÛÛ™šYÊBˆX\[™ÜÈH]Y\žWØ[
ˆÛÛ›‹ˆˆˆ‚ˆÑSPÕ\ÜÙ]ÚY^\›˜[ÚYˆ”“ÓH\ÜÙ]Ú[YÜ˜][ÛœÂˆÒT‘H›ÝšY\ˆHÈS‘[˜X›YHHS‘ÓÐSTÐÑJ^\›˜[ÚY	ÉÊHOH	ÉÂˆˆˆ‹ˆ
›ÝšY\‹
Kˆ
BˆÝ][Û—Ý×Ø\ÜÙ]HÜÝŠ›ÝÖÈ™^\›˜[ÚY—JNˆ[
›ÝÖÈ˜\ÜÙ]ÚY—JH›Üˆ›ÝÈ[ˆX\[™ÜßBˆYˆ›ÝÝ][Û—Ý×Ø\ÜÙ]‚ˆ™]\›ˆÈœ›ÝšY\ˆŽˆ›ÝšY\‹˜ÛÛ™šYÈŽˆÛÛ™šYË™[™Ú[ÈŽˆ[™Ú[ËœÙ\ÜÚ[ÛˆŽˆ›Û™K™]šXÙ\ÈŽˆ×_BˆÙ\ÜÚ[Û‹ÈHÙ]Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[ÛŠÛÛ™šYË›Ü˜ÙWÛÙÚ[Y›Ü˜ÙWÛÙÚ[ŠBˆ˜]×Ù]šXÙ\ÈH™]ÚÙ\Ú[ÛœÛÛ\—Ù]šXÙWÛ\Ý
ˆÙ\ÜÚ[Û‹ˆ˜\ÙWÝ\›Y[™Ú[ÖÈ˜˜\ÙWÝ\›—Kˆ[™Ú[Y[™Ú[ÖÈ™]šXÙWÛ\ÝÙ[™Ú[—KˆÝ][Û—ØÛÙ\Ï\ÛÜY
Ý][Û—Ý×Ø\ÜÙ]
Kˆ
Bˆ]šXÙ\Îˆ\ÝÙXÝÜÝ‹[žWWHH×Bˆ›Üˆ˜]×Ü›ÝÈ[ˆ˜]×Ù]šXÙ\Î‚ˆ]šXÙHH›Ü›X[^™WÙ\Ú[ÛœÛÛ\—Ù]šXÙWÚY[]J˜]×Ü›ÝÊBˆYˆ]šXÙVÈ™]—Ý\WÚY—H›Ý[ˆ•TÒSÓ”ÓÓT—ÒS•‘T•T—ÑU’PÑWÕTWÒQÎ‚ˆÛÛ[YBˆ\ÜÙ]ÚYHÝ][Û—Ý×Ø\ÜÙ]™Ù]
ÝŠ]šXÙVÈœÝ][Û—ØÛÙH—HÜˆˆŠJBˆYˆ›Ý\ÜÙ]ÚYÜˆ›Ý]šXÙVÈ™^\›˜[Ù]šXÙWÚY—N‚ˆÛÛ[YBˆ]šXÙVÈ˜\ÜÙ]ÚY—HH\ÜÙ]ÚYˆ]šXÙVÈ™[˜X›Y—HHYˆ\×Ü™[[Ý™YÚ[™\\—Û˜[YJ]šXÙVÈ™]šXÙWÛ˜[YH—JH[ÙHBˆ]šXÙVÈœ^[ØY—HH˜]×Ü›ÝÂˆ]šXÙVÈœ›ÝšY\—Ù]šXÙWÚY—HH\Ù\Ü›ÝšY\—Ù]šXÙJÛÛ›‹\ÜÙ]ÚY›ÝšY\‹]šXÙJBˆYˆ›Ý]šXÙVÈ™[˜X›Y—N‚ˆÙÙÚ[™Ëš[™›Êˆ‘\Ú[Û”ÛÛ\ˆ[™\\ˆ^ÛYY™XØ]\ÙH]\ÈX\šÙY\È™[[Ý™Yˆ\ÜÙ]ÚYI\È[™\\—ÚYI\È˜[YOI\È‹ˆ\ÜÙ]ÚYˆ]šXÙVÈ™^\›˜[Ù]šXÙWÚY—Kˆ]šXÙVÈ™]šXÙWÛ˜[YH—Kˆ
BˆÛÛ[YBˆ]šXÙ\Ë˜\[™
]šXÙJBˆÛÛ›‹˜ÛÛ[Z]

Bˆ™]\›ˆÂˆœ›ÝšY\ˆŽˆ›ÝšY\‹ˆ˜ÛÛ™šYÈŽˆÛÛ™šYËˆ™[™Ú[ÈŽˆ[™Ú[ËˆœÙ\ÜÚ[ÛˆŽˆÙ\ÜÚ[Û‹ˆ™]šXÙ\ÈŽˆ]šXÙ\Ëˆš\ÝÜžWØØ[Ù[^WÜÙXÛÛ™ÈŽˆ\ÝÜžWØØ[Ù[^WÜÙXÛÛ™ËˆœÛY\\ˆŽˆÛY\\‹ˆB‚‚™YˆÞ[˜×Ù\Ú[ÛœÛÛ\—Ú[™\\—Ø]˜Z[Xš[]WÙ›Ü—Ù]JˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ\™Ù]Ù]Nˆ]Kˆ
‹ˆÛÛ^ˆXÝÜÝ‹[žWH›Û™HH›Û™KŠHOˆXÝÜÝ‹[žWN‚ˆYˆ\™Ù]Ù]HHÝ\œ™[Û\Ø›Û—Ù]J
N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠH\ÜÛšXš[YYH[\Ü˜[™\]Y\ˆ[HXH™XÚYËˆŠBˆÞ[˜×ØÛÛ^HÛÛ^Üˆ™\\™WÙ\Ú[ÛœÛÛ\—Ú[™\\—Ú\ÝÜžWØÛÛ^
ÛÛ›ŠBˆ]šXÙ\ÈHÞ[˜×ØÛÛ^È™]šXÙ\È—BˆYˆ›Ý]šXÙ\Î‚ˆ™]\›ˆÈ™]HŽˆ\™Ù]Ù]Kš\ÛÙ›Ü›X]

KœØ[\\ÈŽˆœ[ÈŽˆš[™\\œÈŽˆBˆžN‚ˆ\ÝÜžWÜ›ÝÜÈH™]ÚÙ\Ú[ÛœÛÛ\—Ù]šXÙWÚ\ÝÜžJˆÞ[˜×ØÛÛ^ÈœÙ\ÜÚ[Ûˆ—Kˆ˜\ÙWÝ\›\Þ[˜×ØÛÛ^È™[™Ú[È—VÈ˜˜\ÙWÝ\›—Kˆ[™Ú[\Þ[˜×ØÛÛ^È™[™Ú[È—VÈ™]šXÙWÚ\ÝÜžWÙ[™Ú[—Kˆ]šXÙ\ÏY]šXÙ\Ëˆ\™Ù]Ù]O]\™Ù]Ù]KˆØ[Ù[^WÜÙXÛÛ™ÏY›Ø]
Þ[˜×ØÛÛ^™Ù]
š\ÝÜžWØØ[Ù[^WÜÙXÛÛ™ÈŠHÜˆ
KˆÛY\\\Þ[˜×ØÛÛ^™Ù]
œÛY\\ˆŠHÜˆ[YKœÛY\ˆ
Bˆ^Ù\^Ù\[ÛŽ‚ˆÙÙÚ[™Ë™^Ù\[ÛŠˆ‘\Ú[Û”ÛÛ\ˆ]šXÙH\ÝÜžH™\]Y\Ý˜Z[Yˆ\™Ù]Ù]OI\È[™Ú[I\È[™\\—ØÛÝ[I\È‹ˆ\™Ù]Ù]KˆÞ[˜×ØÛÛ^È™[™Ú[È—VÈ™]šXÙWÚ\ÝÜžWÙ[™Ú[—Kˆ[Š]šXÙ\ÊKˆ
Bˆ˜Z\ÙB‚ˆ›ÝÈH]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠBˆ›ÜˆØ[\H[ˆ\ÝÜžWÜ›ÝÜÎ‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆS”ÑT•S•È[™\\—ÜÝÙ\—ÜØ[\\È
ˆ\ÜÙ]ÚY›ÝšY\‹^\›˜[ÜÝ][Û—ÚY[™\\—ÚY[™\\—Û˜[YK[™\\—ÜÝÙ\—ÚÝËˆØ[\WÝ[YKXÝ]™WÜÝÙ\—ÚÝË˜]×Ü^[ØYÜ™X]YØ]ˆ
HSQTÈ
ËËËËËËËËËÊBˆÓˆÓÓ‘“PÕ
›ÝšY\‹[™\\—ÚYØ[\WÝ[YJHÈTUHÑUˆ\ÜÙ]ÚYH^ÛYY˜\ÜÙ]ÚYˆ^\›˜[ÜÝ][Û—ÚYH^ÛYY™^\›˜[ÜÝ][Û—ÚYˆ[™\\—Û˜[YHH^ÛYYš[™\\—Û˜[YKˆ[™\\—ÜÝÙ\—ÚÝÈH^ÛYYš[™\\—ÜÝÙ\—ÚÝËˆXÝ]™WÜÝÙ\—ÚÝÈH^ÛYY˜XÝ]™WÜÝÙ\—ÚÝËˆ˜]×Ü^[ØYH^ÛYYœ˜]×Ü^[ØYˆˆˆ‹ˆ
ˆØ[\VÈ˜\ÜÙ]ÚY—KˆÞ[˜×ØÛÛ^Èœ›ÝšY\ˆ—KˆØ[\VÈœÝ][Û—ØÛÙH—KˆØ[\VÈ™^\›˜[Ù]šXÙWÚY—KˆØ[\VÈ™]šXÙWÛ˜[YH—KˆØ[\VÈœ˜]YÜÝÙ\—ÚÝÈ—KˆØ[\VÈœØ[\WÝ[YH—Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆØ[\VÈ˜XÝ]™WÜÝÙ\—ÚÝÈ—KˆœÛÛ‹™[\ÊØ[\VÈœ˜]×Ü^[ØY—K[œÝ\™WØ\ØÚZOUYJKˆ›ÝËˆ
Kˆ
B‚ˆ[×ÝÜš][ˆHˆ]šXÙ\×ØžWØ\ÜÙ]ˆXÝÚ[\ÝÙXÝÜÝ‹[žWWWHHßBˆ›Üˆ]šXÙH[ˆ]šXÙ\Î‚ˆ]šXÙ\×ØžWØ\ÜÙ]œÙ]Y˜][
[
]šXÙVÈ˜\ÜÙ]ÚY—JK×JK˜\[™
]šXÙJBˆ›Üˆ\ÜÙ]ÚY[Ù]šXÙ\È[ˆ]šXÙ\×ØžWØ\ÜÙ]š][\Ê
N‚ˆÝÜ™YÜØ[\\ÈH]Y\žWØ[
ˆÛÛ›‹ˆˆˆ‚ˆÑSPÕ[™\\—ÚYØ[\WÝ[YKXÝ]™WÜÝÙ\—ÚÝÂˆ”“ÓH[™\\—ÜÝÙ\—ÜØ[\\ÂˆÒT‘H\ÜÙ]ÚYHÈS‘›ÝšY\ˆHÈS‘Ø[\WÝ[YHHÈS‘Ø[\WÝ[YHÂˆˆˆ‹ˆ
ˆ\ÜÙ]ÚYˆÞ[˜×ØÛÛ^Èœ›ÝšY\ˆ—Kˆ]][YK˜ÛÛXš[™J\™Ù]Ù]K]][YK›Z[‹[YJ
JKš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆ]][YK˜ÛÛXš[™J\™Ù]Ù]H
È[YY[J^\ÏLJK]][YK›Z[‹[YJ
JKš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆ
Kˆ
BˆØ[\\×ØžWÚ[™\\ŽˆXÝÜÝ‹\ÝÙXÝÜÝ‹[žWWWHHßBˆ˜[YÜÛÝÎˆÙ]Ù]][YWHHÙ]

Bˆ›Üˆ›ÝÈ[ˆÝÜ™YÜØ[\\Î‚ˆØ[\WÝ[YHH\œÙWÙ]][YWÝ˜[YJ›ÝÖÈœØ[\WÝ[YH—JBˆYˆØ[\WÝ[YH\È›Û™N‚ˆÛÛ[YBˆØ[\HHÈœØ[\WÝ[YHŽˆØ[\WÝ[YK˜XÝ]™WÜÝÙ\—ÚÝÈŽˆ›ÝÖÈ˜XÝ]™WÜÝÙ\—ÚÝÈ—_BˆØ[\\×ØžWÚ[™\\‹œÙ]Y˜][
ÝŠ›ÝÖÈš[™\\—ÚY—JK×JK˜\[™
Ø[\JBˆYˆ\×Ú[™\\—Ø]˜Z[X›J›ÝÖÈ˜XÝ]™WÜÝÙ\—ÚÝÈ—JN‚ˆ˜[YÜÛÝË˜Y
[™\\—Ø]˜Z[Xš[]WÜÛÝ
Ø[\WÝ[YJJB‚ˆÛÛ›‹™^XÝ]Jˆ‘SUH”“ÓH[™\\—Ø]˜Z[Xš[]WÙZ[HÒT‘H\ÜÙ]ÚYHÈS‘›ÝšY\ˆHÈS‘]˜Z[Xš[]WÙ]HHÈ‹ˆ
\ÜÙ]ÚYÞ[˜×ØÛÛ^Èœ›ÝšY\ˆ—K\™Ù]Ù]Kš\ÛÙ›Ü›X]

JKˆ
Bˆ[™\\—Ü™\Ý[Îˆ\ÝÙXÝÜÝ‹[žWWHH×Bˆ›Üˆ]šXÙH[ˆ[Ù]šXÙ\Î‚ˆ]šXÙWÜØ[\\ÈHØ[\\×ØžWÚ[™\\‹™Ù]
ÝŠ]šXÙVÈ™^\›˜[Ù]šXÙWÚY—JK×JBˆ™\Ý[HØ[Ý[]WÚ[™\\—ÙZ[WØ]˜Z[Xš[]Jˆ]šXÙWÜØ[\\Ëˆ˜[YÜÛÝËˆ
Bˆ™\Ý[\]JˆÂˆ˜\ÜÙ]ÚYŽˆ\ÜÙ]ÚYˆš[™\\—ÚYŽˆ]šXÙVÈ™^\›˜[Ù]šXÙWÚY—Kˆš[™\\—Û˜[YHŽˆ]šXÙVÈ™]šXÙWÛ˜[YH—Kˆš[™\\—ÜÝÙ\—ÚÝÈŽˆ]šXÙVÈœ˜]YÜÝÙ\—ÚÝÈ—KˆBˆ
Bˆ[™\\—Ü™\Ý[Ë˜\[™
™\Ý[
BˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆS”ÑT•S•È[™\\—Ø]˜Z[Xš[]WÙZ[H
ˆ\ÜÙ]ÚY›ÝšY\‹]˜Z[Xš[]WÙ]K[™\\—ÚY[™\\—Û˜[YK[™\\—ÜÝÙ\—ÚÝËˆ˜[YÜÛÝË]˜Z[X›WÜÛÝË[˜]˜Z[X›WÜÛÝË]˜Z[Xš[]WÜÝÜ™X]YØ]\]YØ]ˆ
HSQTÈ
ËËËËËËËËËËËÊBˆÓˆÓÓ‘“PÕ
›ÝšY\‹[™\\—ÚY]˜Z[Xš[]WÙ]JHÈTUHÑUˆ\ÜÙ]ÚYH^ÛYY˜\ÜÙ]ÚYˆ[™\\—Û˜[YHH^ÛYYš[™\\—Û˜[YKˆ[™\\—ÜÝÙ\—ÚÝÈH^ÛYYš[™\\—ÜÝÙ\—ÚÝËˆ˜[YÜÛÝÈH^ÛYY˜[YÜÛÝËˆ]˜Z[X›WÜÛÝÈH^ÛYY˜]˜Z[X›WÜÛÝËˆ[˜]˜Z[X›WÜÛÝÈH^ÛYY[˜]˜Z[X›WÜÛÝËˆ]˜Z[Xš[]WÜÝH^ÛYY˜]˜Z[Xš[]WÜÝˆ\]YØ]H^ÛYY\]YØ]ˆˆˆ‹ˆ
ˆ\ÜÙ]ÚYˆÞ[˜×ØÛÛ^Èœ›ÝšY\ˆ—Kˆ\™Ù]Ù]Kš\ÛÙ›Ü›X]

Kˆ]šXÙVÈ™^\›˜[Ù]šXÙWÚY—Kˆ]šXÙVÈ™]šXÙWÛ˜[YH—Kˆ]šXÙVÈœ˜]YÜÝÙ\—ÚÝÈ—Kˆ™\Ý[È˜[YÜÛÝÈ—Kˆ™\Ý[È˜]˜Z[X›WÜÛÝÈ—Kˆ™\Ý[È[˜]˜Z[X›WÜÛÝÈ—Kˆ™\Ý[È˜]˜Z[Xš[]WÜÝ—Kˆ›ÝËˆ›ÝËˆ
Kˆ
BˆÙZYÚYÜÝHØ[Ý[]WÝÙZYÚYÜ[Ø]˜Z[Xš[]J[™\\—Ü™\Ý[ÊBˆÛ\˜]YÝ˜[YÜÛÝÈH\WÚ[™\\—ÙYÙWÝÛ\˜[˜ÙJ˜[YÜÛÝÊBˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆS”ÑT•S•È[Ø]˜Z[Xš[]WÙZ[H
ˆ\ÜÙ]ÚY›ÝšY\‹]˜Z[Xš[]WÙ]K˜[YÜÛÝËÙZYÚYØ]˜Z[Xš[]WÜÝˆ[™\\—ØÛÝ[Ü™X]YØ]\]YØ]ˆ
HSQTÈ
ËËËËËËËÊBˆÓˆÓÓ‘“PÕ
›ÝšY\‹\ÜÙ]ÚY]˜Z[Xš[]WÙ]JHÈTUHÑUˆ˜[YÜÛÝÈH^ÛYY˜[YÜÛÝËˆÙZYÚYØ]˜Z[Xš[]WÜÝH^ÛYYÙZYÚYØ]˜Z[Xš[]WÜÝˆ[™\\—ØÛÝ[H^ÛYYš[™\\—ØÛÝ[ˆ\]YØ]H^ÛYY\]YØ]ˆˆˆ‹ˆ
ˆ\ÜÙ]ÚYˆÞ[˜×ØÛÛ^Èœ›ÝšY\ˆ—Kˆ\™Ù]Ù]Kš\ÛÙ›Ü›X]

Kˆ[ŠÛ\˜]YÝ˜[YÜÛÝÊKˆÙZYÚYÜÝˆ[Š[™\\—Ü™\Ý[ÊKˆ›ÝËˆ›ÝËˆ
Kˆ
Bˆ[×ÝÜš][ˆ
ÏHBˆÙÙÚ[™Ëš[™›Êˆ‘\Ú[Û”ÛÛ\ˆ[™\\ˆ]˜Z[Xš[]HØ[Ý[]Yˆ\™Ù]Ù]OI\È\ÜÙ]ÚYI\È[™\\œÏI\È˜[YÜÛÝÏI\È]˜Z[Xš[]WÜÝI\È‹ˆ\™Ù]Ù]Kˆ\ÜÙ]ÚYˆ[Š[™\\—Ü™\Ý[ÊKˆ[ŠÛ\˜]YÝ˜[YÜÛÝÊKˆÙZYÚYÜÝˆ
BˆÛÛ›‹˜ÛÛ[Z]

Bˆ™]\›ˆÂˆ™]HŽˆ\™Ù]Ù]Kš\ÛÙ›Ü›X]

KˆœØ[\\ÈŽˆ[Š\ÝÜžWÜ›ÝÜÊKˆœ[ÈŽˆ[×ÝÜš][‹ˆš[™\\œÈŽˆ[Š]šXÙ\ÊKˆB‚‚™YˆÞ[˜×Ù\Ú[ÛœÛÛ\—Ú[™\\—Ø]˜Z[Xš[]WÜ˜[™ÙJˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆœ›ÛWÙ]Nˆ]Kˆ×Ù]Nˆ]KŠHOˆXÝÜÝ‹[žWN‚ˆÛÛ^H™\\™WÙ\Ú[ÛœÛÛ\—Ú[™\\—Ú\ÝÜžWØÛÛ^
ÛÛ›ŠBˆÝ[ÈHÈ™^\ÈŽˆœØ[\\ÈŽˆœ[ÈŽˆš[™\\œÈŽˆ[ŠÛÛ^È™]šXÙ\È—J_BˆÝ\œ™[Hœ›ÛWÙ]BˆÚ[HÝ\œ™[H×Ù]N‚ˆ™\Ý[HÞ[˜×Ù\Ú[ÛœÛÛ\—Ú[™\\—Ø]˜Z[Xš[]WÙ›Ü—Ù]JÛÛ›‹Ý\œ™[ÛÛ^XÛÛ^
BˆÝ[ÖÈ™^\È—H
ÏHBˆÝ[ÖÈœØ[\\È—H
ÏH[
™\Ý[ÈœØ[\\È—JBˆÝ[ÖÈœ[È—H
ÏH[
™\Ý[Èœ[È—JBˆÝ\œ™[
ÏH[YY[J^\ÏLJBˆ™]\›ˆÝ[Â‚‚™YˆØYÛØØ[Ù\Ú[ÛœÛÛ\—ÝØ]Ù]šXÙ\ÊˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆ\™Ù]Ù]Nˆ]Kˆ\ÜÙ]ÚYÎˆ\ÝÚ[H›Û™HH›Û™KŠHOˆ\ÝÙXÝÜÝ‹[žWWN‚ˆ[œÝ\™WÜØ[\YØ]˜Z[Xš[]WÜØÚ[XJÛÛ›ŠBˆÛÛ™][ÛœÈHÂˆšœ›ÝšY\ˆHÈ‹ˆš™^XÝYHH‹ˆš˜[YÙœ›ÛHHÈ‹ˆŠ˜[YÝÈTÈ•SÔˆ˜[YÝÈHÊH‹ˆœ™]—Ý\WÚYSˆ
KÎ
H‹ˆBˆ\˜[\Îˆ\ÝÐ[žWHHÂˆS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹ˆ\™Ù]Ù]Kš\ÛÙ›Ü›X]

Kˆ\™Ù]Ù]Kš\ÛÙ›Ü›X]

KˆBˆYˆ\ÜÙ]ÚYÎ‚ˆXÙZÛ\œÈH‹‹š›Ú[ŠÈˆ›ÜˆÈ[ˆ\ÜÙ]ÚYÊBˆÛÛ™][ÛœË˜\[™
ˆš˜\ÜÙ]ÚYSˆ
ÜXÙZÛ\œßJHŠBˆ\˜[\Ë™^[™
ÛÜY
Ù]
\ÜÙ]ÚYÊJJBˆ›ÝÜÈHÛÛ›‹™^XÝ]Jˆˆˆˆ‚ˆÑSPÕˆšYTÈ›ÝšY\—Ù]šXÙWÚYˆ˜\ÜÙ]ÚYˆœÝ][Û—ØÛÙKˆ™^\›˜[Ù]šXÙWÚYˆ™]—Ù‹ˆœÛ‹ˆ™]šXÙWÛ˜[YKˆ™]—Ý\WÚYˆ›[Ù[ˆœ˜]YÜÝÙ\—ÚÝËˆ\]YØ]ˆ”“ÓH›ÝšY\—Ù]šXÙWØÛÛ™šYÝ\˜][Û—Ú\ÝÜžHˆ“ÒSˆ›ÝšY\—Ù]šXÙ\ÈÓˆšYHœ›ÝšY\—Ù]šXÙWÚYˆÒT‘HÉÈS‘	Ëš›Ú[ŠÛÛ™][ÛœÊ_BˆÔ‘Tˆ–H˜\ÜÙ]ÚY™^\›˜[Ù]šXÙWÚYˆˆˆ‹ˆ\˜[\Ëˆ
K™™]Ú[

Bˆ™]\›ˆÙXÝ
›ÝÊH›Üˆ›ÝÈ[ˆ›ÝÜ×B‚‚™Yˆ\Ú[ÛœÛÛ\—Ù]šXÙWØØ][Ù×Ú\×ÜÝ[JˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆ›ÝÎˆ]][YH›Û™HH›Û™KŠHOˆ›ÛÛ‚ˆ›ÝÈHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕÓÕS•

ŠHTÈÝ[PV
\]YØ]
HTÈ\ÝÝ\]YØ]ˆ”“ÓH›ÝšY\—Ù]šXÙ\ÂˆÒT‘H›ÝšY\ˆHÈS‘]—Ý\WÚYSˆ
KÎ
Bˆˆˆ‹ˆ
S•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹
Kˆ
K™™]ÚÛ™J
BˆYˆ[
›ÝÖÈÝ[—HÜˆ
HOH‚ˆ™]\›ˆYBˆ\ÝÝ\]YH\œÙWØ˜XÚÙÜ›Ý[™Ú›Ø—Ý[Y\Ý[\
›ÝÖÈ›\ÝÝ\]YØ]—JBˆYˆ\ÝÝ\]Y\È›Û™N‚ˆ™]\›ˆYBˆ›Ý×Ý]ÈH\×Ø˜XÚÙÜ›Ý[™Ú›Ø—Ý]Ê›ÝÈÜˆ˜XÚÙÜ›Ý[™Ú›Ø—Ý]×Û›ÝÊ
JBˆ™]\›ˆ›Ý×Ý]ÈH\ÝÝ\]Yˆ[YY[JÝ\œÏL
B‚‚™YˆZ\ÜÚ[™×ÝØ]Ù]šXÙ\×Ù›Ü—Ù]JˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆ\™Ù]Ù]Nˆ]Kˆ]šXÙ\Îˆ\ÝÙXÝÜÝ‹[žWWKŠHOˆ\VÛ\ÝÙXÝÜÝ‹[žWWK[N‚ˆYˆ›Ý]šXÙ\Î‚ˆ™]\›ˆ×KˆÝ\H]][YK˜ÛÛXš[™J\™Ù]Ù]K]][YK›Z[‹[YJ
JKš\ÛÙ›Ü›X]
ˆ[Y\ÜXÏHœÙXÛÛ™È‚ˆ
Bˆ[™H]][YK˜ÛÛXš[™Jˆ\™Ù]Ù]H
È[YY[J^\ÏLJKˆ]][YK›Z[‹[YJ
Kˆ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠBˆ›ÝÜÈHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕ[™\\—ÚYØ[\WÝ[YKXÝ]™WÜÝÙ\—ÚÝÂˆ”“ÓH[™\\—ÜÝÙ\—ÜØ[\\ÂˆÒT‘H›ÝšY\ˆHÈS‘Ø[\WÝ[YHHÈS‘Ø[\WÝ[YHÂˆÔ‘Tˆ–HØ[\WÝ[YBˆˆˆ‹ˆ
S•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹Ý\[™
Kˆ
K™™]Ú[

Bˆ˜[YÜÛÝÎˆÙ]Ù]][YWHHÙ]

BˆØœÙ\™YØžWÚ[™\\ŽˆXÝÜÝ‹Ù]Ù]][YWWHHßBˆ›Üˆ›ÝÈ[ˆ›ÝÜÎ‚ˆØ[\WÝ[YHH\œÙWÙ]][YWÝ˜[YJ›ÝÖÈœØ[\WÝ[YH—JBˆYˆØ[\WÝ[YH\È›Û™N‚ˆÛÛ[YBˆÛÝH[™\\—Ø]˜Z[Xš[]WÜÛÝ
Ø[\WÝ[YJBˆØœÙ\™YØžWÚ[™\\‹œÙ]Y˜][
ˆÝŠ›ÝÖÈš[™\\—ÚY—JKˆÙ]

Kˆ
K˜Y
ÛÝ
BˆYˆ\×Ú[™\\—Ø]˜Z[X›J›ÝÖÈ˜XÝ]™WÜÝÙ\—ÚÝÈ—JN‚ˆ˜[YÜÛÝË˜Y
ÛÝ
BˆÛÛœÚY\™YÜÛÝÈH\WÚ[™\\—ÙYÙWÝÛ\˜[˜ÙJ˜[YÜÛÝÊBˆYˆ›ÝÛÛœÚY\™YÜÛÝÎ‚ˆ™]\›ˆ\Ý
]šXÙ\ÊKˆZ\ÜÚ[™ÈHÂˆ]šXÙBˆ›Üˆ]šXÙH[ˆ]šXÙ\ÂˆYˆ›ÝÛÛœÚY\™YÜÛÝËš\ÜÝXœÙ]
ˆØœÙ\™YØžWÚ[™\\‹™Ù]
ˆÝŠ]šXÙVÈ™^\›˜[Ù]šXÙWÚY—JKˆÙ]

Kˆ
Bˆ
BˆBˆ™]\›ˆZ\ÜÚ[™Ë[ŠÛÛœÚY\™YÜÛÝÊB‚‚™YˆÝÜ™WÙ\Ú[ÛœÛÛ\—ÝØ]Ú\ÝÜžWØ˜]Ú
ˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆØ[\\Îˆ\ÝÙXÝÜÝ‹[žWWKˆ›ÝšY\ŽˆÝˆHS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹ŠHOˆ[‚ˆ›ÝÈHÙ\šX[^™WØ˜XÚÙÜ›Ý[™Ú›Ø—Ý[Y\Ý[\

Bˆ›ÜˆØ[\H[ˆØ[\\Î‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆS”ÑT•S•È[™\\—ÜÝÙ\—ÜØ[\\È
ˆ\ÜÙ]ÚY›ÝšY\‹^\›˜[ÜÝ][Û—ÚY[™\\—ÚYˆ[™\\—Û˜[YK[™\\—ÜÝÙ\—ÚÝËØ[\WÝ[YKˆXÝ]™WÜÝÙ\—ÚÝË˜]×Ü^[ØYÜ™X]YØ]ˆ
HSQTÈ
ËËËËËËËËËÊBˆÓˆÓÓ‘“PÕ
›ÝšY\‹[™\\—ÚYØ[\WÝ[YJHÈTUHÑUˆ\ÜÙ]ÚYH^ÛYY˜\ÜÙ]ÚYˆ^\›˜[ÜÝ][Û—ÚYH^ÛYY™^\›˜[ÜÝ][Û—ÚYˆ[™\\—Û˜[YHH^ÛYYš[™\\—Û˜[YKˆ[™\\—ÜÝÙ\—ÚÝÈH^ÛYYš[™\\—ÜÝÙ\—ÚÝËˆXÝ]™WÜÝÙ\—ÚÝÈH^ÛYY˜XÝ]™WÜÝÙ\—ÚÝËˆ˜]×Ü^[ØYH^ÛYYœ˜]×Ü^[ØYˆˆˆ‹ˆ
ˆØ[\VÈ˜\ÜÙ]ÚY—Kˆ›ÝšY\‹ˆØ[\VÈœÝ][Û—ØÛÙH—KˆØ[\VÈ™^\›˜[Ù]šXÙWÚY—KˆØ[\VÈ™]šXÙWÛ˜[YH—KˆØ[\VÈœ˜]YÜÝÙ\—ÚÝÈ—KˆØ[\VÈœØ[\WÝ[YH—Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆØ[\VÈ˜XÝ]™WÜÝÙ\—ÚÝÈ—KˆœÛÛ‹™[\ÊØ[\VÈœ˜]×Ü^[ØY—K[œÝ\™WØ\ØÚZOUYJKˆ›ÝËˆ
Kˆ
BˆÛÛ›‹˜ÛÛ[Z]

Bˆ™]\›ˆ[ŠØ[\\ÊB‚‚™Yˆ\]WØÝ\œ™[Ø˜XÚÙÜ›Ý[™Ú›Ø—ØÚXÚÜÚ[
ˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ\]\ÎˆXÝÜÝ‹[žWKŠHOˆXÝÜÝ‹[žWN‚ˆÛÛ^H“ÑPÕSÓ—ÒÔWÐÐSÐÓÓ•V™Ù]

BˆYˆ›ÝÛÛ^‚ˆ™]\›ˆßBˆ›Ø—ÚYH[
ÛÛ^Èš›Ø—ÚY—JBˆ›ÝÈHÛÛ›‹™^XÝ]Jˆ”ÑSPÕ\˜[\×ÚœÛÛˆ”“ÓH˜XÚÙÜ›Ý[™Ú›ØœÈÒT‘HYHÈ‹ˆ
›Ø—ÚY
Kˆ
K™™]ÚÛ™J
Bˆ\˜[\ÈHXÛÙWÚ›Ø—Ü\˜[\Ê›ÝÖÈœ\˜[\×ÚœÛÛˆ—HYˆ›ÝÈ[ÙHžßHŠBˆ\˜[\Ë\]J\]\ÊBˆÛÛ›‹™^XÝ]Jˆ•TUH˜XÚÙÜ›Ý[™Ú›ØœÈÑU\˜[\×ÚœÛÛˆHÈÒT‘HYHÈ‹ˆ
[˜ÛÙWÚ›Ø—Ü\˜[\Ê\˜[\ÊK›Ø—ÚY
Kˆ
BˆÛÛ›‹˜ÛÛ[Z]

Bˆ™]\›ˆ\˜[\Â‚‚™Yˆ[—Ü™\Ý[XX›WÙ\Ú[ÛœÛÛ\—ÝØ]Ø˜XÚÙš[
ˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆœ›ÛWÙ]Nˆ]Kˆ×Ù]Nˆ]KŠHOˆXÝÜÝ‹[žWN‚ˆÛÛ^H“ÑPÕSÓ—ÒÔWÐÐSÐÓÓ•V™Ù]

HÜˆßBˆ›Ø—ÚYH[
ÛÛ^™Ù]
š›Ø—ÚYŠHÜˆ
Bˆ›ØˆHÛÛ›‹™^XÝ]Jˆ”ÑSPÕ\˜[\×ÚœÛÛˆ”“ÓH˜XÚÙÜ›Ý[™Ú›ØœÈÒT‘HYHÈ‹ˆ
›Ø—ÚY
Kˆ
K™™]ÚÛ™J
Bˆ\˜[\ÈHXÛÙWÚ›Ø—Ü\˜[\Ê›Ø–Èœ\˜[\×ÚœÛÛˆ—HYˆ›Øˆ[ÙHžßHŠBˆÝ\œÛÜ—Ù]HH
ˆ\œÙWÙ]WÝ˜[YJÝŠ\˜[\Ë™Ù]
˜Ý\œÛÜ—Ù]HŠHÜˆˆŠJBˆÜˆœ›ÛWÙ]Bˆ
Bˆ\ÜÙ]ÚYÈHÂˆ[
˜[YJBˆ›Üˆ˜[YH[ˆ\˜[\Ë™Ù]
˜\ÜÙ]ÚYÈŠHÜˆ×BˆYˆÝŠ˜[YJKš\ÙYÚ]

BˆBˆ][\YÚYÈHÂˆÝŠ˜[YJBˆ›Üˆ˜[YH[ˆ\˜[\Ë™Ù]
˜][\YÚ[™\\—ÚYÈŠHÜˆ×BˆYˆÝŠ˜[YJBˆBˆÝ[[X\žNˆXÝÜÝ‹[žWHHÂˆ™^\ÈŽˆˆœØ[\\ÈŽˆˆœ[ÈŽˆˆš[™\\œÈŽˆˆ˜\WØØ[×Ý\ÙYŽˆˆ™^\×Ü™XØ[Ý[]YÙœ›ÛWÙˆŽˆˆœ™\Ý[YWÚ[ŽˆÝ\œÛÜ—Ù]Kš\ÛÙ›Ü›X]

KˆœÝÜYÜ™X\ÛÛˆŽˆˆ‹ˆB‚ˆYˆ\Ú[ÛœÛÛ\—Ù]šXÙWØØ][Ù×Ú\×ÜÝ[JÛÛ›ŠN‚ˆ™Yœ™\ÚYH™\\™WÙ\Ú[ÛœÛÛ\—Ú[™\\—Ú\ÝÜžWØÛÛ^
ˆÛÛ›‹ˆ\ÝÜžWØØ[Ù[^WÜÙXÛÛ™ÏLˆÛY\\[[X™HÜÙXÛÛ™Îˆ›Û™Kˆ
BˆÝ][Û—ØÚ[šÜÈHX^
ˆKˆX]˜ÙZ[
ˆ[ŠˆÂˆÝŠ]šXÙVÈœÝ][Û—ØÛÙH—JBˆ›Üˆ]šXÙH[ˆ™Yœ™\ÚYÈ™]šXÙ\È—BˆYˆ]šXÙK™Ù]
œÝ][Û—ØÛÙHŠBˆBˆ
BˆÈLˆ
Kˆ
BˆÝ[[X\žVÈ˜\WØØ[×Ý\ÙY—H
ÏHÝ][Û—ØÚ[šÜÈ
ˆ‚‚ˆÝ\œ™[HÝ\œÛÜ—Ù]BˆÚ[HÝ\œ™[H×Ù]N‚ˆ]šXÙ\ÈHØYÛØØ[Ù\Ú[ÛœÛÛ\—ÝØ]Ù]šXÙ\ÊˆÛÛ›‹ˆ\™Ù]Ù]OXÝ\œ™[ˆ\ÜÙ]ÚYÏX\ÜÙ]ÚYÈÜˆ›Û™Kˆ
BˆÝ[[X\žVÈš[™\\œÈ—HH[Š]šXÙ\ÊBˆ™XØ[Ý[]WÜÝÜ™YÚ[™\\—Ø]˜Z[Xš[]JˆÛÛ›‹ˆÝ\œ™[ˆÝ\œ™[ˆ
BˆZ\ÜÚ[™×Ù]šXÙ\Ë˜[YÜÛÝÈHZ\ÜÚ[™×ÝØ]Ù]šXÙ\×Ù›Ü—Ù]JˆÛÛ›‹ˆ\™Ù]Ù]OXÝ\œ™[ˆ]šXÙ\ÏY]šXÙ\Ëˆ
BˆYˆ›ÝZ\ÜÚ[™×Ù]šXÙ\Î‚ˆÝ[[X\žVÈ™^\È—H
ÏHBˆÝ[[X\žVÈ™^\×Ü™XØ[Ý[]YÙœ›ÛWÙˆ—H
ÏHBˆÝ[[X\žVÈœ[È—H
ÏH[ŠˆÚ[
]šXÙVÈ˜\ÜÙ]ÚY—JH›Üˆ]šXÙH[ˆ]šXÙ\ßBˆ
BˆÝ\œ™[
ÏH[YY[J^\ÏLJBˆ][\YÚYÈHÙ]

Bˆ\]WØÝ\œ™[Ø˜XÚÙÜ›Ý[™Ú›Ø—ØÚXÚÜÚ[
ˆÛÛ›‹ˆÂˆ˜Ý\œÛÜ—Ù]HŽˆÝ\œ™[š\ÛÙ›Ü›X]

Kˆ˜][\YÚ[™\\—ÚYÈŽˆ×KˆKˆ
BˆÛÛ[YB‚ˆ[™[™×Ù]šXÙ\ÈHÂˆ]šXÙBˆ›Üˆ]šXÙH[ˆZ\ÜÚ[™×Ù]šXÙ\ÂˆYˆÝŠ]šXÙVÈ™^\›˜[Ù]šXÙWÚY—JH›Ý[ˆ][\YÚYÂˆBˆYˆ›Ý[™[™×Ù]šXÙ\Î‚ˆÝ[[X\žVÈ™^\È—H
ÏHBˆÝ[[X\žVÈœÝÜYÜ™X\ÛÛˆ—HH
ˆØ]Ú\ÝÜžWÚ[˜ÛÛ\]WØY\—Ü™Yœ™\Ú‚ˆYˆ˜[YÜÛÝÂˆ[ÙH››×ÛØœÙ\™YÛÜ\˜][™×ÝÚ[™ÝÈ‚ˆ
BˆÝ\œ™[
ÏH[YY[J^\ÏLJBˆ][\YÚYÈHÙ]

Bˆ\]WØÝ\œ™[Ø˜XÚÙÜ›Ý[™Ú›Ø—ØÚXÚÜÚ[
ˆÛÛ›‹ˆÂˆ˜Ý\œÛÜ—Ù]HŽˆÝ\œ™[š\ÛÙ›Ü›X]

Kˆ˜][\YÚ[™\\—ÚYÈŽˆ×KˆKˆ
BˆÛÛ[YB‚ˆ˜]ÚH[™[™×Ù]šXÙ\ÖÎŒLBˆÛÛ™šYÈHÙ]Ú[YÜ˜][Û—ØÛÛ™šYÊˆÛÛ›‹ˆS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹ˆ
BˆYˆÛÛ™šYÈ\È›Û™N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠÛÛ™šYÝ\˜XØ[È\Ú[Û”ÛÛ\ˆ[™\ÜÛš]™[ˆŠBˆ[™Ú[ÈHÙ]Ù\Ú[ÛœÛÛ\—Ù[™Ú[ØÛÛ™šYÊÛÛ™šYÊBˆÛÛ›‹˜ÛÛ[Z]

BˆÙ\ÜÚ[Û‹ÝÚÙ[ˆHÙ]Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[ÛŠÛÛ™šYÊBˆ\ÝÜžWÜ›ÝÜÈH™]ÚÙ\Ú[ÛœÛÛ\—Ù]šXÙWÚ\ÝÜžJˆÙ\ÜÚ[Û‹ˆ˜\ÙWÝ\›Y[™Ú[ÖÈ˜˜\ÙWÝ\›—Kˆ[™Ú[Y[™Ú[ÖÈ™]šXÙWÚ\ÝÜžWÙ[™Ú[—Kˆ]šXÙ\ÏX˜]Úˆ\™Ù]Ù]OXÝ\œ™[ˆØ[Ù[^WÜÙXÛÛ™ÏLˆÛY\\[[X™HÜÙXÛÛ™Îˆ›Û™Kˆ
BˆÝ[[X\žVÈ˜\WØØ[×Ý\ÙY—H
ÏHBˆÝ[[X\žVÈœØ[\\È—H
ÏHÝÜ™WÙ\Ú[ÛœÛÛ\—ÝØ]Ú\ÝÜžWØ˜]Ú
ˆÛÛ›‹ˆØ[\\ÏZ\ÝÜžWÜ›ÝÜËˆ
Bˆ][\YÚYË\]JˆÝŠ]šXÙVÈ™^\›˜[Ù]šXÙWÚY—JH›Üˆ]šXÙH[ˆ˜]Úˆ
Bˆ™XØ[Ý[]WÜÝÜ™YÚ[™\\—Ø]˜Z[Xš[]JˆÛÛ›‹ˆÝ\œ™[ˆÝ\œ™[ˆ
Bˆ\]WØÝ\œ™[Ø˜XÚÙÜ›Ý[™Ú›Ø—ØÚXÚÜÚ[
ˆÛÛ›‹ˆÂˆ˜Ý\œÛÜ—Ù]HŽˆÝ\œ™[š\ÛÙ›Ü›X]

Kˆ˜][\YÚ[™\\—ÚYÈŽˆÛÜY
][\YÚYÊKˆKˆ
BˆÝ[[X\žVÈœ™\Ý[YWÚ[—HHÝ\œ™[š\ÛÙ›Ü›X]

Bˆ˜Z\ÙH\TÛÝ[˜]˜Z[X›Q\œ›ÜŠˆ›ÝšY\RS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹ˆXØÛÝ[ÚÙ^OY\Ú[ÛœÛÛ\—ØXØÛÝ[ÚÙ^JÛÛ™šYÊKˆ\WØ\™XOUÐUÒTÕÔ–WÐT‘PKˆ™^Ø][\Ø]Y]][YK››ÝÊTÐ“Ó—ÕSQV“Ó‘JBˆ
È[YY[JÙXÛÛ™ÏLJKˆØZ]Ü™X\ÛÛH˜ÚXÚÜÚ[‹ˆY\ÜØYÙOH“ÝHÐUÝX\™YÎÈ›Øˆ™\\˜YÈ\˜HÈÝHÙYÝZ[Kˆ‹ˆ›Ø—Ü™\Ý[\Ý[[X\žKˆ
B‚ˆÝ[[X\žVÈœ™\Ý[YWÚ[—HHˆ‚ˆ™]\›ˆÝ[[X\žB‚‚™Yˆ[—Ù\Ú[ÛœÛÛ\—Ú[™\\—Ø]˜Z[Xš[]WØ˜XÚÙš[
ˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ
‹ˆœ›ÛWÙ]Nˆ]Kˆ×Ù]Nˆ]KˆÛY\\Žˆ[žHH[YKœÛY\ˆ\ÝÜžWØØ[Ù[^WÜÙXÛÛ™Îˆ›Ø]H•TÒSÓ”ÓÓT—ÔT‘“Ô“PSÑWÒÔWÑSVWÔÑPÓÓ‘ËˆX^ÝØZ]ØÞXÛ\Îˆ[HŠHOˆXÝÜÝ‹[žWN‚ˆYˆœ›ÛWÙ]Hˆ×Ù]HÜˆ×Ù]HHÝ\œ™[Û\Ø›Û—Ù]J
N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ“È˜XÚÙš[ÐU™\]Y\ˆ[H[\˜[È˜[YÈHX\È™XÚYÜËˆŠBˆYˆ“ÑPÕSÓ—ÒÔWÐÐSÐÓÓ•V™Ù]

H\È›Ý›Û™N‚ˆ™]\›ˆ[—Ü™\Ý[XX›WÙ\Ú[ÛœÛÛ\—ÝØ]Ø˜XÚÙš[
ˆÛÛ›‹ˆœ›ÛWÙ]OYœ›ÛWÙ]Kˆ×Ù]O]×Ù]Kˆ
BˆÝ[[X\žNˆXÝÜÝ‹[žWHHÂˆ™^\ÈŽˆˆœØ[\\ÈŽˆˆœ[ÈŽˆˆš[™\\œÈŽˆˆ˜\WÙ\œ›ÜœÈŽˆˆØZ]ØÞXÛ\ÈŽˆˆœ™\Ý[YWÚ[Žˆœ›ÛWÙ]Kš\ÛÙ›Ü›X]

KˆœÝÜYÜ™X\ÛÛˆŽˆˆ‹ˆB‚ˆYˆØZ]ØY\—Ü˜]WÛ[Z]
™X\ÛÛŽˆÝ‹™\Ý[YWÙ]Nˆ]JHOˆ›ÛÛ‚ˆÝ[[X\žVÈØZ]ØÞXÛ\È—H
ÏHBˆÝ[[X\žVÈœ™\Ý[YWÚ[—HH™\Ý[YWÙ]Kš\ÛÙ›Ü›X]

BˆYˆÝ[[X\žVÈØZ]ØÞXÛ\È—HˆX^ÝØZ]ØÞXÛ\Î‚ˆÝ[[X\žVÈœÝÜYÜ™X\ÛÛˆ—HHˆ“[Z]H\Ú[Û”ÛÛ\ˆ™\]YÈ[X\ÚXY\È™^™\Ëˆ[[[È\ÝYÎˆÜ™X\ÛÛŸH‚ˆÙÙÚ[™ËØ\›š[™Êˆ‘\Ú[Û”ÛÛ\ˆÐU˜XÚÙš[ÝÜYY\ˆ™\X]YÛÛÛÝÛœÎˆ™\Ý[YWÙ]OI\ÈØZ]ØÞXÛ\ÏI\È™X\ÛÛI\È‹ˆ™\Ý[YWÙ]KˆÝ[[X\žVÈØZ]ØÞXÛ\È—Kˆ™X\ÛÛ‹ˆ
Bˆ™]\›ˆ˜[ÙBˆÛÛ›‹˜ÛÛ[Z]

BˆÙXÛÛ™ÈH\Ú[ÛœÛÛ\—ØÛÛÛÝÛ—ÜÛY\ÜÙXÛÛ™ÊÛÛ›ŠBˆÙÙÚ[™ËØ\›š[™Êˆ‘\Ú[Û”ÛÛ\ˆÐU˜XÚÙš[ØZ][™È›ÜˆTHÛÛÛÝÛŽˆ™\Ý[YWÙ]OI\ÈÙXÛÛ™ÏI\ÈØZ]ØÞXÛOI\È‹ˆ™\Ý[YWÙ]KˆÙXÛÛ™ËˆÝ[[X\žVÈØZ]ØÞXÛ\È—Kˆ
BˆÛY\\ŠÙXÛÛ™ÊBˆÛX\—Ù\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]ØÛÛÛÝÛŠÛÛ›ŠBˆ™]\›ˆYB‚ˆYˆ™\\™WØÛÛ^
›Ü˜ÙWÛÙÚ[Žˆ›ÛÛH˜[ÙJHOˆXÝÜÝ‹[žWH›Û™N‚ˆÚ[HYN‚ˆžN‚ˆ™]\›ˆ™\\™WÙ\Ú[ÛœÛÛ\—Ú[™\\—Ú\ÝÜžWØÛÛ^
ˆÛÛ›‹ˆ›Ü˜ÙWÛÙÚ[Y›Ü˜ÙWÛÙÚ[‹ˆ\ÝÜžWØØ[Ù[^WÜÙXÛÛ™ÏZ\ÝÜžWØØ[Ù[^WÜÙXÛÛ™ËˆÛY\\\ÛY\\‹ˆ
Bˆ^Ù\^Ù\[Ûˆ\È^Î‚ˆÝ[[X\žVÈ˜\WÙ\œ›ÜœÈ—H
ÏHBˆYˆ\Ú[œÝ[˜ÙJ^Ë\T˜]S[Z]\œ›ÜŠN‚ˆ˜Z\ÙBˆYˆ\×Ù\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]Ù\œ›ÜŠ^ÊN‚ˆ™X\ÛÛˆHX\š×Ù\Ú[ÛœÛÛ\—Ü\™›Ü›X[˜ÙWÜ˜]WÛ[Z]Y
ÛÛ›ŠBˆYˆ›ÝØZ]ØY\—Ü˜]WÛ[Z]
™X\ÛÛ‹œ›ÛWÙ]JN‚ˆ™]\›ˆ›Û™Bˆ›Ü˜ÙWÛÙÚ[ˆHYBˆÛÛ[YBˆYˆ\×Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[Û—Ù^\™YÙ\œ›ÜŠ^ÊN‚ˆ›Ü˜ÙWÛÙÚ[ˆHYBˆÛÛ[YBˆ˜Z\ÙB‚ˆÛÛÛÝÛ—Ü™X\ÛÛˆHÙ]Ù\Ú[ÛœÛÛ\—Ü\™›Ü›X[˜ÙWØÛÛÛÝÛ—Ü™X\ÛÛŠÛÛ›ŠBˆYˆÛÛÛÝÛ—Ü™X\ÛÛˆ[™›ÝØZ]ØY\—Ü˜]WÛ[Z]
ÛÛÛÝÛ—Ü™X\ÛÛ‹œ›ÛWÙ]JN‚ˆ™]\›ˆÝ[[X\žBˆÛÛ^H™\\™WØÛÛ^

BˆYˆÛÛ^\È›Û™N‚ˆ™]\›ˆÝ[[X\žBˆÝ[[X\žVÈš[™\\œÈ—HH[ŠÛÛ^È™]šXÙ\È—JB‚ˆÝ\œ™[Hœ›ÛWÙ]BˆÚ[HÝ\œ™[H×Ù]N‚ˆÙ\ÜÚ[Û—Ü™]žWÝ\ÙYH˜[ÙBˆÚ[HYN‚ˆžN‚ˆÙÙÚ[™Ëš[™›Ê‘\Ú[Û”ÛÛ\ˆÐU˜XÚÙš[^HÝ\Yˆ\™Ù]Ù]OI\È‹Ý\œ™[
Bˆ™\Ý[HÞ[˜×Ù\Ú[ÛœÛÛ\—Ú[™\\—Ø]˜Z[Xš[]WÙ›Ü—Ù]JÛÛ›‹Ý\œ™[ÛÛ^XÛÛ^
BˆÝÜ™YØÛÝ[ÈHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕˆ
ÑSPÕÓÕS•

ŠH”“ÓH[™\\—ÜÝÙ\—ÜØ[\\ÂˆÒT‘H›ÝšY\ˆHÈS‘Ø[\WÝ[YHHÈS‘Ø[\WÝ[YHÊHTÈÝÙ\—ÜØ[\\Ëˆ
ÑSPÕÓÕS•

ŠH”“ÓH[™\\—Ø]˜Z[Xš[]WÙZ[BˆÒT‘H›ÝšY\ˆHÈS‘]˜Z[Xš[]WÙ]HHÊHTÈ[™\\—ÙZ[WÜ›ÝÜËˆ
ÑSPÕÓÕS•

ŠH”“ÓH[Ø]˜Z[Xš[]WÙZ[BˆÒT‘H›ÝšY\ˆHÈS‘]˜Z[Xš[]WÙ]HHÊHTÈ[ÙZ[WÜ›ÝÜÂˆˆˆ‹ˆ
ˆS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹ˆ]][YK˜ÛÛXš[™JÝ\œ™[]][YK›Z[‹[YJ
JKš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆ]][YK˜ÛÛXš[™JÝ\œ™[
È[YY[J^\ÏLJK]][YK›Z[‹[YJ
JKš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹ˆÝ\œ™[š\ÛÙ›Ü›X]

KˆS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‹ˆÝ\œ™[š\ÛÙ›Ü›X]

Kˆ
Kˆ
K™™]ÚÛ™J
BˆÝ[[X\žVÈ™^\È—H
ÏHBˆÝ[[X\žVÈœØ[\\È—H
ÏH[
™\Ý[ÈœØ[\\È—JBˆÝ[[X\žVÈœ[È—H
ÏH[
™\Ý[Èœ[È—JBˆÝ[[X\žVÈœ™\Ý[YWÚ[—HH
Ý\œ™[
È[YY[J^\ÏLJJKš\ÛÙ›Ü›X]

BˆÛÛ›‹˜ÛÛ[Z]

BˆÙÙÚ[™Ëš[™›Êˆ‘\Ú[Û”ÛÛ\ˆÐU˜XÚÙš[^HÛÛ\]Yˆ\™Ù]Ù]OI\È™]ÚYÜØ[\\ÏI\È‚ˆœÝÜ™YÜÝÙ\—ÜØ[\\ÏI\ÈÝÜ™YÚ[™\\—ÙZ[OI\ÈÝÜ™YÜ[ÙZ[OI\È[ÏI\È[™\\œÏI\È‹ˆÝ\œ™[ˆ™\Ý[ÈœØ[\\È—KˆÝÜ™YØÛÝ[ÖÈœÝÙ\—ÜØ[\\È—KˆÝÜ™YØÛÝ[ÖÈš[™\\—ÙZ[WÜ›ÝÜÈ—KˆÝÜ™YØÛÝ[ÖÈœ[ÙZ[WÜ›ÝÜÈ—Kˆ™\Ý[Èœ[È—Kˆ™\Ý[Èš[™\\œÈ—Kˆ
Bˆœ™XZÂˆ^Ù\^Ù\[Ûˆ\È^Î‚ˆÝ[[X\žVÈ˜\WÙ\œ›ÜœÈ—H
ÏHBˆYˆ\Ú[œÝ[˜ÙJ^Ë\T˜]S[Z]\œ›ÜŠN‚ˆ˜Z\ÙBˆYˆ\×Ù\Ú[ÛœÛÛ\—Ü˜]WÛ[Z]Ù\œ›ÜŠ^ÊN‚ˆ™X\ÛÛˆHX\š×Ù\Ú[ÛœÛÛ\—Ü\™›Ü›X[˜ÙWÜ˜]WÛ[Z]Y
ÛÛ›ŠBˆYˆ›ÝØZ]ØY\—Ü˜]WÛ[Z]
™X\ÛÛ‹Ý\œ™[
N‚ˆ™]\›ˆÝ[[X\žBˆÛÛ[YBˆYˆ\×Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[Û—Ù^\™YÙ\œ›ÜŠ^ÊH[™›ÝÙ\ÜÚ[Û—Ü™]žWÝ\ÙY‚ˆÙ\ÜÚ[Û—Ü™]žWÝ\ÙYHYBˆ[˜[Y]WÙ\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[ÛŠÛÛ^È˜ÛÛ™šYÈ—JBˆ™Yœ™\ÚYH™\\™WØÛÛ^
›Ü˜ÙWÛÙÚ[UYJBˆYˆ™Yœ™\ÚY\È›Û™N‚ˆ™]\›ˆÝ[[X\žBˆÛÛ^H™Yœ™\ÚYˆÛÛ[YBˆ˜Z\ÙBˆÝ\œ™[
ÏH[YY[J^\ÏLJBˆÝ[[X\žVÈœ™\Ý[YWÚ[—HHˆ‚ˆ™]\›ˆÝ[[X\žB‚‚™Yˆ™XØ[Ý[]WÜÝÜ™YÚ[™\\—Ø]˜Z[Xš[]JˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆœ›ÛWÙ]Nˆ]Kˆ×Ù]Nˆ]Kˆ
‹ˆ\ÜÙ]ÚYˆ[›Û™HH›Û™KŠHOˆXÝÜÝ‹[N‚ˆ›ÝšY\ˆHS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓT‚ˆÛÛ™][ÛœÈHÈœ›ÝšY\ˆHÈ‹œØ[\WÝ[YHHÈ‹œØ[\WÝ[YHÈ—Bˆ\˜[\Îˆ\ÝÐ[žWHHÂˆ›ÝšY\‹ˆ]][YK˜ÛÛXš[™Jœ›ÛWÙ]K]][YK›Z[‹[YJ
JKš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆ]][YK˜ÛÛXš[™J×Ù]H
È[YY[J^\ÏLJK]][YK›Z[‹[YJ
JKš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆBˆYˆ\ÜÙ]ÚY\È›Ý›Û™N‚ˆÛÛ™][ÛœË˜\[™
˜\ÜÙ]ÚYHÈŠBˆ\˜[\Ë˜\[™
\ÜÙ]ÚY
BˆØ[\WÙ]\ÈHÛÛ›‹™^XÝ]Jˆˆˆˆ‚ˆÑSPÕTÕSÕ\ÜÙ]ÚYÝXœÝŠØ[\WÝ[YKKL
HTÈØ[\WÙ]Bˆ”“ÓH[™\\—ÜÝÙ\—ÜØ[\\ÂˆÒT‘HÉÈS‘	Ëš›Ú[ŠÛÛ™][ÛœÊ_BˆÔ‘Tˆ–HØ[\WÙ]K\ÜÙ]ÚYˆˆˆ‹ˆ\˜[\Ëˆ
K™™]Ú[

Bˆ›ÝÈH]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠBˆÝ[ÈHÈ™^\ÈŽˆœ[ÈŽˆš[™\\œÈŽˆBˆ›Üˆ]WÜ›ÝÈ[ˆØ[\WÙ]\Î‚ˆÝ\œ™[Ø\ÜÙ]ÚYH[
]WÜ›ÝÖÈ˜\ÜÙ]ÚY—JBˆ\™Ù]Ù]HH\œÙWÙ]WÝ˜[YJ]WÜ›ÝÖÈœØ[\WÙ]H—JBˆYˆ\™Ù]Ù]H\È›Û™N‚ˆÛÛ[YBˆ]šXÙ\ÈH^XÝYÙ]šXÙ\×Ù›Ü—Ù]JˆÛÛ›‹ˆ\ÜÙ]ÚYXÝ\œ™[Ø\ÜÙ]ÚYˆ›ÝšY\\›ÝšY\‹ˆ\™Ù]Ù]O]\™Ù]Ù]Kˆ
BˆØ[\\ÈH]Y\žWØ[
ˆÛÛ›‹ˆˆˆ‚ˆÑSPÕ[™\\—ÚYØ[\WÝ[YKXÝ]™WÜÝÙ\—ÚÝÂˆ”“ÓH[™\\—ÜÝÙ\—ÜØ[\\ÂˆÒT‘H\ÜÙ]ÚYHÈS‘›ÝšY\ˆHÈS‘Ø[\WÝ[YHHÈS‘Ø[\WÝ[YHÂˆÔ‘Tˆ–HØ[\WÝ[YBˆˆˆ‹ˆ
ˆÝ\œ™[Ø\ÜÙ]ÚYˆ›ÝšY\‹ˆ]][YK˜ÛÛXš[™J\™Ù]Ù]K]][YK›Z[‹[YJ
JKš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆ]][YK˜ÛÛXš[™J\™Ù]Ù]H
È[YY[J^\ÏLJK]][YK›Z[‹[YJ
JKš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆ
Kˆ
BˆØ[\\×ØžWÚ[™\\ŽˆXÝÜÝ‹\ÝÙXÝÜÝ‹[žWWWHHßBˆ˜[YÜÛÝÎˆÙ]Ù]][YWHHÙ]

Bˆ›ÜˆØ[\WÜ›ÝÈ[ˆØ[\\Î‚ˆØ[\WÝ[YHH\œÙWÙ]][YWÝ˜[YJØ[\WÜ›ÝÖÈœØ[\WÝ[YH—JBˆYˆØ[\WÝ[YH\È›Û™N‚ˆÛÛ[YBˆØ[\HHÈœØ[\WÝ[YHŽˆØ[\WÝ[YK˜XÝ]™WÜÝÙ\—ÚÝÈŽˆØ[\WÜ›ÝÖÈ˜XÝ]™WÜÝÙ\—ÚÝÈ—_BˆØ[\\×ØžWÚ[™\\‹œÙ]Y˜][
ÝŠØ[\WÜ›ÝÖÈš[™\\—ÚY—JK×JK˜\[™
Ø[\JBˆYˆ\×Ú[™\\—Ø]˜Z[X›JØ[\WÜ›ÝÖÈ˜XÝ]™WÜÝÙ\—ÚÝÈ—JN‚ˆ˜[YÜÛÝË˜Y
[™\\—Ø]˜Z[Xš[]WÜÛÝ
Ø[\WÝ[YJJB‚ˆÛÛ›‹™^XÝ]Jˆ‘SUH”“ÓH[™\\—Ø]˜Z[Xš[]WÙZ[HÒT‘H\ÜÙ]ÚYHÈS‘›ÝšY\ˆHÈS‘]˜Z[Xš[]WÙ]HHÈ‹ˆ
Ý\œ™[Ø\ÜÙ]ÚY›ÝšY\‹\™Ù]Ù]Kš\ÛÙ›Ü›X]

JKˆ
BˆÛ\˜]YÝ˜[YÜÛÝÈH\WÚ[™\\—ÙYÙWÝÛ\˜[˜ÙJ˜[YÜÛÝÊBˆ[™\\—Ü™\Ý[Îˆ\ÝÙXÝÜÝ‹[žWWHH×Bˆ›Üˆ]šXÙH[ˆ]šXÙ\Î‚ˆYˆ\×Ü™[[Ý™YÚ[™\\—Û˜[YJ]šXÙVÈ™]šXÙWÛ˜[YH—JN‚ˆÛÛ[YBˆ[™\\—ÚYHÝŠ]šXÙVÈ™^\›˜[Ù]šXÙWÚY—HÜˆˆŠBˆYˆ›Ý[™\\—ÚY‚ˆÛÛ[YBˆ]šXÙWÜØ[\\ÈHØ[\\×ØžWÚ[™\\‹™Ù]
[™\\—ÚY×JBˆ[™\\—ÜÝÙ\—ÚÝÈH
ˆ\œÙWÙ›Ø]Ý˜[YJ]šXÙVÈœ˜]YÜÝÙ\—ÚÝÈ—JBˆÜˆ[™™\—Ú[™\\—ÜÝÙ\—Ùœ›ÛWÛ[Ù[
]šXÙVÈ›[Ù[—JBˆ
Bˆ™\Ý[HØ[Ý[]WÚ[™\\—ÙZ[WØ]˜Z[Xš[]J]šXÙWÜØ[\\Ë˜[YÜÛÝÊBˆØœÙ\™YÜÛÝÈHÂˆ[™\\—Ø]˜Z[Xš[]WÜÛÝ
Ø[\VÈœØ[\WÝ[YH—JBˆ›ÜˆØ[\H[ˆ]šXÙWÜØ[\\ÂˆBˆYˆ
ˆ›ÝÛ\˜]YÝ˜[YÜÛÝÂˆÜˆ›ÝÛ\˜]YÝ˜[YÜÛÝËš\ÜÝXœÙ]
ØœÙ\™YÜÛÝÊBˆ
N‚ˆ™\Ý[È˜]˜Z[Xš[]WÜÝ—HH›Û™Bˆ™\Ý[Èš[™\\—ÜÝÙ\—ÚÝÈ—HH[™\\—ÜÝÙ\—ÚÝÂˆ[™\\—Ü™\Ý[Ë˜\[™
™\Ý[
BˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆS”ÑT•S•È[™\\—Ø]˜Z[Xš[]WÙZ[H
ˆ\ÜÙ]ÚY›ÝšY\‹]˜Z[Xš[]WÙ]K[™\\—ÚY[™\\—Û˜[YK[™\\—ÜÝÙ\—ÚÝËˆ˜[YÜÛÝË]˜Z[X›WÜÛÝË[˜]˜Z[X›WÜÛÝË]˜Z[Xš[]WÜÝÜ™X]YØ]\]YØ]ˆ
HSQTÈ
ËËËËËËËËËËËÊBˆˆˆ‹ˆ
ˆÝ\œ™[Ø\ÜÙ]ÚYˆ›ÝšY\‹ˆ\™Ù]Ù]Kš\ÛÙ›Ü›X]

Kˆ[™\\—ÚYˆ]šXÙVÈ™]šXÙWÛ˜[YH—Kˆ[™\\—ÜÝÙ\—ÚÝËˆ™\Ý[È˜[YÜÛÝÈ—Kˆ™\Ý[È˜]˜Z[X›WÜÛÝÈ—Kˆ™\Ý[È[˜]˜Z[X›WÜÛÝÈ—Kˆ™\Ý[È˜]˜Z[Xš[]WÜÝ—Kˆ›ÝËˆ›ÝËˆ
Kˆ
BˆÙZYÚYÜÝH
ˆØ[Ý[]WÝÙZYÚYÜ[Ø]˜Z[Xš[]J[™\\—Ü™\Ý[ÊBˆYˆ[™\\—Ü™\Ý[Âˆ[™[
ˆ›ÝË™Ù]
˜]˜Z[Xš[]WÜÝŠH\È›Ý›Û™Bˆ›Üˆ›ÝÈ[ˆ[™\\—Ü™\Ý[Âˆ
Bˆ[ÙH›Û™Bˆ
BˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆS”ÑT•S•È[Ø]˜Z[Xš[]WÙZ[H
ˆ\ÜÙ]ÚY›ÝšY\‹]˜Z[Xš[]WÙ]K˜[YÜÛÝËÙZYÚYØ]˜Z[Xš[]WÜÝˆ[™\\—ØÛÝ[Ü™X]YØ]\]YØ]ˆ
HSQTÈ
ËËËËËËËÊBˆÓˆÓÓ‘“PÕ
›ÝšY\‹\ÜÙ]ÚY]˜Z[Xš[]WÙ]JHÈTUHÑUˆ˜[YÜÛÝÈH^ÛYY˜[YÜÛÝËˆÙZYÚYØ]˜Z[Xš[]WÜÝH^ÛYYÙZYÚYØ]˜Z[Xš[]WÜÝˆ[™\\—ØÛÝ[H^ÛYYš[™\\—ØÛÝ[ˆ\]YØ]H^ÛYY\]YØ]ˆˆˆ‹ˆ
ˆÝ\œ™[Ø\ÜÙ]ÚYˆ›ÝšY\‹ˆ\™Ù]Ù]Kš\ÛÙ›Ü›X]

Kˆ[ŠÛ\˜]YÝ˜[YÜÛÝÊKˆÙZYÚYÜÝˆ[Š[™\\—Ü™\Ý[ÊKˆ›ÝËˆ›ÝËˆ
Kˆ
BˆÝ[ÖÈœ[È—H
ÏHBˆÝ[ÖÈš[™\\œÈ—H
ÏH[Š[™\\—Ü™\Ý[ÊBˆÝ[ÖÈ™^\È—H
ÏHBˆÛÛ›‹˜ÛÛ[Z]

Bˆ™]\›ˆÝ[Â‚‚™Yˆ[—Ù\Ú[ÛœÛÛ\—Ù]šXÙWØ]˜Z[Xš[]WÜÞ[˜ÊˆÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹ˆ›ÝšY\ŽˆÝ‹ˆšYÙÙ\—Ý\NˆÝˆH›X[X[‹ŠHOˆXÝÜÝ‹[žWN‚ˆÛÛ™šYÈHÙ]Ú[YÜ˜][Û—ØÛÛ™šYÊÛÛ›‹›ÝšY\ŠBˆYˆÛÛ™šYÈ\È›Û™HÜˆ›ÝÛÛ™šYÖÈ™[˜X›Y—N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠˆÛÛ™šYÝ\˜XØ[ÈÜ›ÝšY\ŸH[™\ÜÛš]™[ˆŠBˆ[™Ú[ÈHÙ]Ù\Ú[ÛœÛÛ\—Ù[™Ú[ØÛÛ™šYÊÛÛ™šYÊBˆX\[™ÜÈH]Y\žWØ[
ˆÛÛ›‹ˆˆˆ‚ˆÑSPÕ\ÜÙ]ÚY^\›˜[ÚYˆ”“ÓH\ÜÙ]Ú[YÜ˜][ÛœÂˆÒT‘H›ÝšY\ˆHÈS‘[˜X›YHHS‘ÓÐSTÐÑJ^\›˜[ÚY	ÉÊHOH	ÉÂˆˆˆ‹ˆ
›ÝšY\‹
Kˆ
BˆÝ][Û—Ý×Ø\ÜÙ]HÜÝŠ›ÝÖÈ™^\›˜[ÚY—JNˆ[
›ÝÖÈ˜\ÜÙ]ÚY—JH›Üˆ›ÝÈ[ˆX\[™ÜßBˆÝ][Û—ØÛÙ\ÈHÛÜY
Ý][Û—Ý×Ø\ÜÙ]
BˆYˆ›ÝÝ][Û—ØÛÙ\Î‚ˆ™]\›ˆÈ™]šXÙ\ÈŽˆœÛ˜\ÚÝÈŽˆ˜\ÜÙ]ÈŽˆB‚ˆ\ÝÙ\œ›ÜŽˆ^Ù\[Ûˆ›Û™HH›Û™Bˆ›Üˆ][\[ˆ˜[™ÙJŠN‚ˆžN‚ˆÙ\ÜÚ[Û‹ÈHÙ]Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[ÛŠÛÛ™šYË›Ü˜ÙWÛÙÚ[X][\OHJBˆ]šXÙWÜ›ÝÜÈH™]ÚÙ\Ú[ÛœÛÛ\—Ù]šXÙWÛ\Ý
ˆÙ\ÜÚ[Û‹ˆ˜\ÙWÝ\›Y[™Ú[ÖÈ˜˜\ÙWÝ\›—Kˆ[™Ú[Y[™Ú[ÖÈ™]šXÙWÛ\ÝÙ[™Ú[—KˆÝ][Û—ØÛÙ\Ï\Ý][Û—ØÛÙ\Ëˆ
Bˆœ™XZÂˆ^Ù\^Ù\[Ûˆ\È^Î‚ˆ\ÝÙ\œ›ÜˆH^ÂˆYˆ›Ý\×Ù\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[Û—Ù^\™YÙ\œ›ÜŠ^ÊHÜˆ][\OHN‚ˆ˜Z\ÙBˆÑÑÑT‹š[™›Ê‘\Ú[Û”ÛÛ\ˆ]šXÙHÙ\ÜÚ[Ûˆ^\™YÈ[˜[Y][™ÈØXÚH[™™]žZ[™ÈÙÚ[ˆÛ˜ÙHŠBˆ[˜[Y]WÙ\Ú[ÛœÛÛ\—ÜÙ\ÜÚ[ÛŠÛÛ™šYÊBˆ[ÙN‚ˆ˜Z\ÙH\ÝÙ\œ›ÜˆÜˆ˜[YQ\œ›ÜŠ‘˜[H\ØÛÛšXÚYH›È\Ú[Û”ÛÛ\‹ˆŠB‚ˆ˜XÚÙYˆ\ÝÙXÝÜÝ‹[žWWHH×Bˆ›Üˆ˜]×Ü›ÝÈ[ˆ]šXÙWÜ›ÝÜÎ‚ˆ›Ü›X[^™YH›Ü›X[^™WÙ\Ú[ÛœÛÛ\—Ù]šXÙWÚY[]J˜]×Ü›ÝÊBˆYˆ›Ü›X[^™YÈ™]—Ý\WÚY—H›Ý[ˆ•TÒSÓ”ÓÓT—ÒS•‘T•T—ÑU’PÑWÕTWÒQÎ‚ˆÛÛ[YBˆYˆ›Ý›Ü›X[^™YÈœÝ][Û—ØÛÙH—HÜˆ›Ý›Ü›X[^™YÈ™^\›˜[Ù]šXÙWÚY—N‚ˆÛÛ[YBˆ\ÜÙ]ÚYHÝ][Û—Ý×Ø\ÜÙ]™Ù]
›Ü›X[^™YÈœÝ][Û—ØÛÙH—JBˆYˆ›Ý\ÜÙ]ÚY‚ˆÛÛ[YBˆ›Ü›X[^™YÈœ^[ØY—HH˜]×Ü›ÝÂˆ›Ü›X[^™YÈ™[˜X›Y—HHYˆ\×Ü™[[Ý™YÚ[™\\—Û˜[YJ›Ü›X[^™YÈ™]šXÙWÛ˜[YH—JH[ÙHBˆ›Ü›X[^™YÈœ›ÝšY\—Ù]šXÙWÚY—HH\Ù\Ü›ÝšY\—Ù]šXÙJÛÛ›‹\ÜÙ]ÚY›ÝšY\‹›Ü›X[^™Y
BˆYˆ›Ý›Ü›X[^™YÈ™[˜X›Y—N‚ˆÛÛ[YBˆ›Ü›X[^™YÈ˜\ÜÙ]ÚY—HH\ÜÙ]ÚYˆ˜XÚÙY˜\[™
›Ü›X[^™Y
B‚ˆ™X[[YWÛX\H™]ÚÙ\Ú[ÛœÛÛ\—Ù]šXÙWÜ™X[[YWÛX\
ˆÙ\ÜÚ[Û‹ˆ˜\ÙWÝ\›Y[™Ú[ÖÈ˜˜\ÙWÝ\›—Kˆ[™Ú[Y[™Ú[ÖÈ™]šXÙWÜ™X[Ý[YWÙ[™Ú[—Kˆ]šXÙ\Ï]˜XÚÙYˆ
BˆÛÛXÝYØ]ÙH]][YK››ÝÊ
BˆÛÛXÝYØ]HÙ\šX[^™WØ˜XÚÙÜ›Ý[™Ú›Ø—Ý[Y\Ý[\

BˆÛ˜\ÚÝ×ØžWØ\ÜÙ]ˆXÝÚ[\ÝÙXÝÜÝ‹[žWWWHHßBˆ›Üˆ]šXÙH[ˆ˜XÚÙY‚ˆ™X[[YHH™^
ˆ
™X[[YWÛX\ÚÙ^WH›ÜˆÙ^H[ˆ
]šXÙVÈ™^\›˜[Ù]šXÙWÚY—K]šXÙVÈ™]—Ùˆ—K]šXÙVÈœÛˆ—JHYˆÙ^H[™Ù^H[ˆ™X[[YWÛX\
KˆßKˆ
Bˆ]WÛX\H™X[[YK™Ù]
™]R][SX\ŠHYˆ\Ú[œÝ[˜ÙJ™X[[YK™Ù]
™]R][SX\ŠKXÝ
H[ÙH™X[[YBˆÙY[—ÙH\œÙWÙ]][YWÝ˜[YJš\œÝÛ›Û—Ù[\J™X[[YKÈ˜ÛÛXÝ[YH‹˜ÛÛXÝY]‹›\ÝÙY[ˆ‹›\ÝÜÙY[—Ø]—JJBˆ\×Ü™XÙ[Ù]HHYHYˆ™X[[YH[™ÙY[—Ù\È›Û™H[ÙH›ÛÛ
ˆÙY[—Ù[™ÛÛXÝYØ]ÙHÙY[—ÙH[YY[JZ[]\ÏQQUSÑU’PÑWÐÓÓSUS’PÐUSÓ—Õ‘TÒÓÓRS•UTÊBˆ
Bˆ]˜Z[Xš[]WÜÝ]\ÈHÛ\ÜÚYžWÙ\Ú[ÛœÛÛ\—Ú[™\\—Ø]˜Z[Xš[]JˆÈš[™\\—ÜÝ]HŽˆš\œÝÛ›Û—Ù[\J]WÛX\Èš[™\\—ÜÝ]H‹š[™\\”Ý]H—J_Kˆ\×Ü™XÙ[Ù]OZ\×Ü™XÙ[Ù]Kˆ
BˆÛÛ[][šXØ][Û—ÜÝ]\ÈHœ™XÙ[ˆYˆ\×Ü™XÙ[Ù]H[ÙHœÝ[H‚ˆÝ\œ™[Ë›ÛYÙ\ÈH\œÙWÙ\Ú[ÛœÛÛ\—Ü—Ú[œ]Ê™X[[YJBˆ^XÝYÜÝš[™×Ú[™^\ÈHX\›—Ù^XÝYÜÝš[™Ü×Ùœ›ÛWÝ›ÛYÙJˆÛÛ›‹ˆ]šXÙVÈœ›ÝšY\—Ù]šXÙWÚY—Kˆ›ÛYÙ\ËˆÛÛXÝYØ]ˆ
Bˆ—ÚX[HØ[Ý[]WÜ—Ú[œ]ÚX[
ˆÝ\œ™[Ëˆ›ÛYÙ\Ëˆ^XÝYÜÝš[™×Ú[™^\ÏY^XÝYÜÝš[™×Ú[™^\Ëˆ
BˆÛ˜\ÚÝHÂˆ
Š™]šXÙKˆ˜]˜Z[Xš[]WÜÝ]\ÈŽˆ]˜Z[Xš[]WÜÝ]\Ëˆ˜ÛÛ[][šXØ][Û—ÜÝ]\ÈŽˆÛÛ[][šXØ][Û—ÜÝ]\Ëˆœ˜]YÜÝÙ\—ÚÝÈŽˆ]šXÙVÈœ˜]YÜÝÙ\—ÚÝÈ—Kˆ
Šœ—ÚX[ˆBˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆS”ÑT•S•È]šXÙWÜ™X[[YWÜÛ˜\ÚÝÈ
ˆ›ÝšY\—Ù]šXÙWÚY\ÜÙ]ÚY›ÝšY\‹Ý][Û—ØÛÙKÛÛXÝYØ][™\\—ÜÝ]KˆXÝ]™WÜÝÙ\—ÚÝË^WÙ[™\™ÞWÚÝÚ]˜Z[Xš[]WÜÝ]\ËÛÛ[][šXØ][Û—ÜÝ]\ËˆÝš[™×Ø]˜Z[X›WØÛÝ[Ýš[™×ÝÝ[ØÛÝ[—ØÝ\œ™[ÚœÛÛ‹—Ý›ÛYÙWÚœÛÛ‹^[ØYÚœÛÛ‹Ü™X]YØ]ˆ
HSQTÈ
ËËËËËËËËËËËËËËËÊBˆˆˆ‹ˆ
ˆ]šXÙVÈœ›ÝšY\—Ù]šXÙWÚY—Kˆ]šXÙVÈ˜\ÜÙ]ÚY—Kˆ›ÝšY\‹ˆ]šXÙVÈœÝ][Û—ØÛÙH—KˆÛÛXÝYØ]ˆ\œÙWÚ[Ý˜[YJš\œÝÛ›Û—Ù[\J]WÛX\Èš[™\\—ÜÝ]H‹š[™\\”Ý]H—JJKˆ›Ü›X[^™WÜÝÙ\—Ý×ÚÝÊš\œÝÛ›Û—Ù[\J]WÛX\È˜XÝ]™WÜÝÙ\ˆ‹˜XÝ]™TÝÙ\ˆ—JJKˆ\œÙWÙ›Ø]Ý˜[YJš\œÝÛ›Û—Ù[\J]WÛX\È™^WØØ\‹™^Q[™\™ÞH‹™^WÙ[™\™ÞH—JJKˆ]˜Z[Xš[]WÜÝ]\ËˆÛÛ[][šXØ][Û—ÜÝ]\Ëˆ—ÚX[È˜]˜Z[X›WÜÝš[™ÜÈ—Kˆ—ÚX[ÈÝ[ÜÝš[™ÜÈ—KˆœÛÛ‹™[\ÊÝ\œ™[Ë[œÝ\™WØ\ØÚZOUYJHYˆÝ\œ™[È[ÙH›Û™KˆœÛÛ‹™[\Ê›ÛYÙ\Ë[œÝ\™WØ\ØÚZOUYJHYˆ›ÛYÙ\È[ÙH›Û™KˆœÛÛ‹™[\Ê™X[[YK[œÝ\™WØ\ØÚZOUYJKˆÛÛXÝYØ]ˆ
Kˆ
BˆÛÛ›‹™^XÝ]Jˆ•TUH›ÝšY\—Ù]šXÙ\ÈÑU\ÝÜÙY[—Ø]HÓÐSTÐÑJË\ÝÜÙY[—Ø]
K\]YØ]HÈÒT‘HYHÈ‹ˆ
ÙY[—Ùš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠHYˆÙY[—Ù[ÙHÛÛXÝYØ]Yˆ™X[[YH[ÙH›Û™KÛÛXÝYØ]]šXÙVÈœ›ÝšY\—Ù]šXÙWÚY—JKˆ
BˆÛ˜\ÚÝ×ØžWØ\ÜÙ]œÙ]Y˜][
]šXÙVÈ˜\ÜÙ]ÚY—K×JK˜\[™
Û˜\ÚÝ
B‚ˆ›Üˆ\ÜÙ]ÚY›ÝÜÈ[ˆÛ˜\ÚÝ×ØžWØ\ÜÙ]š][\Ê
N‚ˆÝ[[X\žHHØ[Ý[]WØ\ÜÙ]Ø]˜Z[Xš[]J›ÝÜÊBˆ›ÝÈH]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠBˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆS”ÑT•S•È]˜Z[Xš[]WÙZ[H
ˆ\ÜÙ]ÚY›ÝšY\‹\š[ÙÙ]K[™\\—Ø]˜Z[Xš[]WÜÝØ\XÚ]WØ]˜Z[Xš[]WÜÝˆÛÛ[][šXØ][Û—Ø]˜Z[Xš[]WÜÝÝš[™×Ø]˜Z[Xš[]WÜÝ]˜Z[X›WÚ[™\\œËÝ[Ú[™\\œË[˜]˜Z[X›WÚ[™\\œËˆ›×ØÛÛ[][šXØ][Û—Ù]šXÙ\Ë]˜Z[X›WÜÝš[™ÜËÝ[ÜÝš[™ÜË[˜]˜Z[X›WÜÝš[™ÜËY™™XÝYÜÝÙ\—ÚÝËˆ^[ØYÚœÛÛ‹Ü™X]YØ]\]YØ]ˆ
HSQTÈ
ËËËËËËËËËËËËËËËËËÊBˆÓˆÓÓ‘“PÕ
\ÜÙ]ÚY›ÝšY\‹\š[ÙÙ]JHÈTUHÑUˆ[™\\—Ø]˜Z[Xš[]WÜÝH^ÛYYš[™\\—Ø]˜Z[Xš[]WÜÝˆØ\XÚ]WØ]˜Z[Xš[]WÜÝH^ÛYY˜Ø\XÚ]WØ]˜Z[Xš[]WÜÝˆÛÛ[][šXØ][Û—Ø]˜Z[Xš[]WÜÝH^ÛYY˜ÛÛ[][šXØ][Û—Ø]˜Z[Xš[]WÜÝˆÝš[™×Ø]˜Z[Xš[]WÜÝH^ÛYYœÝš[™×Ø]˜Z[Xš[]WÜÝˆ]˜Z[X›WÚ[™\\œÈH^ÛYY˜]˜Z[X›WÚ[™\\œËˆÝ[Ú[™\\œÈH^ÛYYÝ[Ú[™\\œËˆ[˜]˜Z[X›WÚ[™\\œÈH^ÛYY[˜]˜Z[X›WÚ[™\\œËˆ›×ØÛÛ[][šXØ][Û—Ù]šXÙ\ÈH^ÛYY››×ØÛÛ[][šXØ][Û—Ù]šXÙ\Ëˆ]˜Z[X›WÜÝš[™ÜÈH^ÛYY˜]˜Z[X›WÜÝš[™ÜËˆÝ[ÜÝš[™ÜÈH^ÛYYÝ[ÜÝš[™ÜËˆ[˜]˜Z[X›WÜÝš[™ÜÈH^ÛYY[˜]˜Z[X›WÜÝš[™ÜËˆY™™XÝYÜÝÙ\—ÚÝÈH^ÛYY˜Y™™XÝYÜÝÙ\—ÚÝËˆ^[ØYÚœÛÛˆH^ÛYYœ^[ØYÚœÛÛ‹ˆ\]YØ]H^ÛYY\]YØ]ˆˆˆ‹ˆ
ˆ\ÜÙ]ÚYˆ›ÝšY\‹ˆÝ\œ™[Û\Ø›Û—Ù]J
Kš\ÛÙ›Ü›X]

KˆÝ[[X\žVÈš[™\\—Ø]˜Z[Xš[]WÜÝ—KˆÝ[[X\žVÈ˜Ø\XÚ]WØ]˜Z[Xš[]WÜÝ—KˆÝ[[X\žVÈ˜ÛÛ[][šXØ][Û—Ø]˜Z[Xš[]WÜÝ—KˆÝ[[X\žVÈœÝš[™×Ø]˜Z[Xš[]WÜÝ—KˆÝ[[X\žVÈ˜]˜Z[X›WÚ[™\\œÈ—KˆÝ[[X\žVÈÝ[Ú[™\\œÈ—KˆÝ[[X\žVÈ[˜]˜Z[X›WÚ[™\\œÈ—KˆÝ[[X\žVÈ››×ØÛÛ[][šXØ][Û—Ù]šXÙ\È—KˆÝ[[X\žVÈ˜]˜Z[X›WÜÝš[™ÜÈ—KˆÝ[[X\žVÈÝ[ÜÝš[™ÜÈ—KˆÝ[[X\žVÈ[˜]˜Z[X›WÜÝš[™ÜÈ—KˆÝ[[X\žVÈ˜Y™™XÝYÜÝÙ\—ÚÝÈ—KˆœÛÛ‹™[\ÊÝ[[X\žK[œÝ\™WØ\ØÚZOUYJKˆ›ÝËˆ›ÝËˆ
Kˆ
BˆØ[\YÜÝ]\ÎˆXÝÜÝ‹[HHßBˆ›Üˆ\ÜÙ]ÚY[ˆÛ˜\ÚÝ×ØžWØ\ÜÙ]‚ˆØ[\YHX]\šX[^™WÜØ[\YØ]˜Z[Xš[]WÙ^JˆÛÛ›‹ˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆ›ÝšY\\›ÝšY\‹ˆ\™Ù]Ù]OXÝ\œ™[Û\Ø›Û—Ù]J
Kˆ
BˆÝ]HHÝŠØ[\YÈ˜ÛÝ™\˜YÙWÜÝ]\È—JBˆØ[\YÜÝ]\ÖÜÝ]WHHØ[\YÜÝ]\Ë™Ù]
Ý]K
H
ÈBˆÛÛ›‹˜ÛÛ[Z]

Bˆ™]\›ˆÂˆ™]šXÙ\ÈŽˆ[Š˜XÚÙY
KˆœÛ˜\ÚÝÈŽˆÝ[J[Š›ÝÜÊH›Üˆ›ÝÜÈ[ˆÛ˜\ÚÝ×ØžWØ\ÜÙ]˜[Y\Ê
JKˆ˜\ÜÙ]ÈŽˆ[ŠÛ˜\ÚÝ×ØžWØ\ÜÙ]
KˆœØ[\YÙ^\×Ü™XØ[Ý[]YÛØØ[HŽˆ[ŠÛ˜\ÚÝ×ØžWØ\ÜÙ]
KˆœØ[\YÜÝ]\ÈŽˆØ[\YÜÝ]\ËˆB‚‚™Yˆ[—Ú[YÜ˜][Û—ÜÞ[˜ÊÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹›ÝšY\ŽˆÝ‹šYÙÙ\—Ý\NˆÝˆH›X[X[ŠHOˆXÝÜÝ‹[žWN‚ˆYˆ›ÝšY\ˆOHS•QÔUSÓ—Ô“Õ’QT—ÔÒQÑS‘T‘ÖN‚ˆ™]\›ˆ[—ÜÚYÙ[™\™ÞWÜÞ[˜ÊÛÛ›‹›ÝšY\‹šYÙÙ\—Ý\O]šYÙÙ\—Ý\JBˆ™]\›ˆ[—Ù\Ú[ÛœÛÛ\—ÜÞ[˜ÊÛÛ›‹›ÝšY\‹šYÙÙ\—Ý\O]šYÙÙ\—Ý\JB‚‚™Yˆ[—Ù\Ú[ÛœÛÛ\—ÜÞ[˜ÊÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹›ÝšY\ŽˆÝ‹šYÙÙ\—Ý\NˆÝˆH›X[X[ŠHOˆXÝÜÝ‹[žWN‚ˆYˆ›Ý•TÒSÓ”ÓÓT—ÔÖS×ÓÐÒË˜XÜ]Z\™J›ØÚÚ[™ÏQ˜[ÙJN‚ˆY\ÜØYÙHH”Ú[˜Ü›Ûš^˜XØ[È\Ú[Û”ÛÛ\ˆYÛ›Ü˜YHÜœ]YH˜H^\ÝHÝ]˜H[HÝ\œÛËˆ‚ˆÑÑÑT‹š[™›ÊY\ÜØYÙJBˆ™]\›ˆÈ›X]ÚYŽˆ[œ™\ÛÛ™YŽˆ˜]]×Ü™\ÛÛ™YŽˆœÝ]\ÈŽˆœÚÚ\Y‹œÝÜYÜ™X\ÛÛˆŽˆY\ÜØYÙ_B‚ˆÚ]™[X\ÙWÙ\Ú[ÛœÛÛ\—ÜÞ[˜×ÛØÚÊ
N‚ˆÛÛ™šYÈHÙ]Ú[YÜ˜][Û—ØÛÛ™šYÊÛÛ›‹›ÝšY\ŠBˆYˆÛÛ™šYÈ\È›Û™N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠˆÛÛ™šYÝ\˜XØ[ÈÜ›ÝšY\ŸH˜[È[˜ÛÛ˜YKˆŠBˆYˆ›ÝÛÛ™šYÖÈ™[˜X›Y—N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠˆH[YÜ˜XØ[ÈÜ›ÝšY\ŸH\ÝH\Ø]]˜YKˆŠB‚ˆ[—ÚYHÜ™X]WÚ[YÜ˜][Û—Ü[ŠÛÛ›‹›ÝšY\‹šYÙÙ\—Ý\JBˆ˜]ÚÚYHÜ™X]WÛ[Ûš]Üš[™×Ø˜]Ú
ˆÛÛ›‹ˆ™XÛÜ™Ù]OY]KÙ^J
Kš\ÛÙ›Ü›X]

KˆY˜][Û›Ý\ÏYˆ”Þ[˜ÈÜ›ÝšY\ŸH
ÝšYÙÙ\—Ý\_JH‹ˆ˜]×Ú[œ]Hˆ‹ˆÛÝ\˜ÙO\›ÝšY\‹ˆ
BˆÛÛ›‹˜ÛÛ[Z]

B‚ˆžN‚ˆYˆšYÙÙ\—Ý\HOHœØÚY[YÜÝ]HŽ‚ˆ™\Ý[H[—Ù\Ú[ÛœÛÛ\—ØÚXÚÊˆÛÛ›‹ˆ›ÝšY\‹ˆžWÜ[UYKˆ[˜ÛYWÙXYÛ›ÜÝXÜÏQ˜[ÙKˆ™Y™\—ÛØØ[ÜÝ][Û—Ú[™[ÜžOUYKˆ
Bˆ[ÙN‚ˆ™\Ý[H[—Ù\Ú[ÛœÛÛ\—ØÚXÚÊÛÛ›‹›ÝšY\‹žWÜ[UYJBˆ›ÝÜÈH™\Ý[Èœ›ÝÜÈ—BˆX]ÚYHˆ[œ™\ÛÛ™YHˆ]]×Ü™\ÛÛ™YHˆÞ[˜ÙYØ\ÜÙ]ÚYÎˆÙ]Ú[HHÙ]

Bˆ[\Ù]™[Îˆ\ÝÙXÝÜÝ‹[žWWHH×Bˆ›ÝÈH]][YK››ÝÊ
B‚ˆ›Üˆ›ÝÈ[ˆ›ÝÜÎ‚ˆ^\›˜[ÚYH›ÝÖÈ™^\›˜[ÚY—Bˆ^\›˜[Û˜[YHH›ÝÖÈ™^\›˜[Û˜[YH—BˆÝ]\ÈH›ÝÖÈœÝ]\È—B‚ˆX\YØ\ÜÙ]H›Û™BˆYˆ^\›˜[ÚY‚ˆX\YØ\ÜÙ]HÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕZK˜\ÜÙ]ÚYˆ”“ÓH\ÜÙ]Ú[YÜ˜][ÛœÈZBˆÒT‘HZKœ›ÝšY\ˆHÈS‘ZK™^\›˜[ÚYHÈS‘ZK™[˜X›YHBˆSRUBˆˆˆ‹ˆ
›ÝšY\‹^\›˜[ÚY
Kˆ
K™™]ÚÛ™J
BˆYˆX\YØ\ÜÙ]\È›Û™N‚ˆX\YØ\ÜÙ]HÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕZK˜\ÜÙ]ÚYˆ”“ÓH\ÜÙ]Ú[YÜ˜][ÛœÈZBˆÒT‘HZKœ›ÝšY\ˆHÈS‘ZK™^\›˜[Û˜[YHHÈS‘ZK™[˜X›YHBˆSRUBˆˆˆ‹ˆ
›ÝšY\‹^\›˜[Û˜[YJKˆ
K™™]ÚÛ™J
Bˆ\ÜÙ]ÚYH[
X\YØ\ÜÙ]È˜\ÜÙ]ÚY—JHYˆX\YØ\ÜÙ][ÙH
š[™Ø\ÜÙ]ÚY
ÛÛ›‹^\›˜[Û˜[YJHÜˆ
B‚ˆYˆ\ÜÙ]ÚY‚ˆÞ[˜ÙYØ\ÜÙ]ÚYË˜Y
\ÜÙ]ÚY
Bˆ™]š[Ý\ÈHÙ]Û]\ÝÛ[Ûš]Üš[™×Ü›ÝÊÛÛ›‹\ÜÙ]ÚY
Bˆ\XØ]WÛ]\ÝH
ˆ™]š[Ý\È\È›Ý›Û™Bˆ[™™]š[Ý\ÖÈœÝ]\È—HOHÝ]\Âˆ[™™]š[Ý\ÖÈœ™XÛÜ™Ù]H—HOH]KÙ^J
Kš\ÛÙ›Ü›X]

Bˆ[™™]š[Ý\ÖÈœÛÝ\˜ÙH—HOH›ÝšY\‚ˆ
BˆYˆ›Ý\XØ]WÛ]\Ý‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆS”ÑT•S•È[Ûš]Üš[™×Ü™XÛÜ™È
\ÜÙ]ÚYÝ]\Ë™XÛÜ™Ù]K›Ý\ËÛÝ\˜ÙK˜]ÚÚY
BˆSQTÈ
ËËËËËÊBˆˆˆ‹ˆ
ˆ\ÜÙ]ÚYˆÝ]\Ëˆ]KÙ^J
Kš\ÛÙ›Ü›X]

Kˆˆ”Þ[˜ÈÜ›ÝšY\ŸNˆÜ›ÝÖÉÛ›Ý\É×_H‹ˆ›ÝšY\‹ˆ˜]ÚÚYˆ
Kˆ
Bˆ]™[HZ[Û[Ûš]Üš[™×Ø[\Ù]™[
ˆÛÛ›‹ˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆ™]š[Ý\×ÜÝ]\Ï\™]š[Ý\ÖÈœÝ]\È—HYˆ™]š[Ý\È[ÙHˆ‹ˆÝ\œ™[ÜÝ]\Ï\Ý]\Ëˆ\[™YØ][›ÝËš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆ[\›WØÛÛ^\›ÝËˆ
BˆYˆ]™[‚ˆ[\Ù]™[Ë˜\[™
]™[
BˆÜ™X]WÛÜ—Ý\]WØ\ÜÙ]Ú[YÜ˜][ÛŠÛÛ›‹\ÜÙ]ÚY›ÝšY\‹^\›˜[ÚY^\›˜[Û˜[YKÝ]\ÊBˆX]ÚY
ÏHBˆ[ÙN‚ˆ\Ù\Ú[YÜ˜][Û—Ý[œ™\ÛÛ™Y
ˆÛÛ›‹ˆ›ÝšY\\›ÝšY\‹ˆ[—ÚY\[—ÚYˆ^\›˜[ÚYY^\›˜[ÚYˆ^\›˜[Û˜[YOY^\›˜[Û˜[YKˆÝ]\Ï\Ý]\Ëˆ^[ØY\›ÝÖÈœ^[ØY—Kˆ
Bˆ[œ™\ÛÛ™Y
ÏHB‚ˆX\YØ\ÜÙ]ÈH]Y\žWØ[
ˆÛÛ›‹ˆˆˆ‚ˆÑSPÕZK˜\ÜÙ]ÚYˆ”“ÓH\ÜÙ]Ú[YÜ˜][ÛœÈZBˆ“ÒSˆ]\ÝÛ[Ûš]Üš[™×ÝšY]ÈHÓˆK˜\ÜÙ]ÚYHZK˜\ÜÙ]ÚYˆÒT‘HZKœ›ÝšY\ˆHÈS‘ZK™[˜X›YHHS‘KœÝ]\ÈSˆ
	Ñ\œ›ÉË	Ñ\ØÛÛ™XÝYIÊBˆˆˆ‹ˆ
›ÝšY\‹
Kˆ
Bˆ›Üˆ›ÝÈ[ˆX\YØ\ÜÙ]Î‚ˆ\ÜÙ]ÚYH[
›ÝÖÈ˜\ÜÙ]ÚY—JBˆYˆ\ÜÙ]ÚY[ˆÞ[˜ÙYØ\ÜÙ]ÚYÎ‚ˆÛÛ[YBˆ^\Ý[™×ÝÙ^HHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆÑSPÕBˆ”“ÓH[Ûš]Üš[™×Ü™XÛÜ™ÂˆÒT‘H\ÜÙ]ÚYHÈS‘™XÛÜ™Ù]HHÈS‘ÛÝ\˜ÙHHÂˆSRUBˆˆˆ‹ˆ
\ÜÙ]ÚY]KÙ^J
Kš\ÛÙ›Ü›X]

K›ÝšY\ŠKˆ
K™™]ÚÛ™J
BˆYˆ^\Ý[™×ÝÙ^N‚ˆÛÛ[YBˆ™]š[Ý\ÈHÙ]Û]\ÝÛ[Ûš]Üš[™×Ü›ÝÊÛÛ›‹\ÜÙ]ÚY
BˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆS”ÑT•S•È[Ûš]Üš[™×Ü™XÛÜ™È
\ÜÙ]ÚYÝ]\Ë™XÛÜ™Ù]K›Ý\ËÛÝ\˜ÙK˜]ÚÚY
BˆSQTÈ
ËËËËËÊBˆˆˆ‹ˆ
ˆ\ÜÙ]ÚYˆ”™\ÛÛšYÈ‹ˆ]KÙ^J
Kš\ÛÙ›Ü›X]

Kˆˆ”™\ÛÛšYÈ]]ÛX]XØ[Y[HÜˆ]\Ù[˜ÚXH›ÈÞ[˜ÈÜ›ÝšY\ŸKˆ‹ˆ›ÝšY\‹ˆ˜]ÚÚYˆ
Kˆ
Bˆ]]×Ü™\ÛÛ™Y
ÏHBˆ]™[HZ[Û[Ûš]Üš[™×Ø[\Ù]™[
ˆÛÛ›‹ˆ\ÜÙ]ÚYX\ÜÙ]ÚYˆ™]š[Ý\×ÜÝ]\Ï\™]š[Ý\ÖÈœÝ]\È—HYˆ™]š[Ý\È[ÙHˆ‹ˆÝ\œ™[ÜÝ]\ÏH”™\ÛÛšYÈ‹ˆ\[™YØ][›ÝËš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆ
BˆYˆ]™[‚ˆ[\Ù]™[Ë˜\[™
]™[
B‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUH[YÜ˜][Û—ØÛÛ™šYÜÂˆÑU\ÝÜÞ[˜×Ø]HË\ÝÜÞ[˜×ÜÝ]\ÈH	ÜÝXØÙ\ÜÉË\ÝÙ\œ›ÜˆH	ÉË\]YØ]HÂˆÒT‘H›ÝšY\ˆHÂˆˆˆ‹ˆ
ˆ]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆ]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆ›ÝšY\‹ˆ
Kˆ
BˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUH[Ûš]Üš[™×Ú[\ÜØ˜]Ú\ÂˆÑU[\ÜYØÛÝ[HËX]ÚYØÛÝ[HË[›X]ÚYØÛÝ[HË]]×Ü™\ÛÛ™YØÛÝ[HÂˆÒT‘HYHÂˆˆˆ‹ˆ
X]ÚY
È[œ™\ÛÛ™YX]ÚY[œ™\ÛÛ™Y]]×Ü™\ÛÛ™Y˜]ÚÚY
Kˆ
Bˆš[˜[^™WÚ[YÜ˜][Û—Ü[ŠˆÛÛ›‹ˆ[—ÚYˆÝ]\ÏHœÝXØÙ\ÜÈ‹ˆX]ÚYØÛÝ[[X]ÚYˆ[œ™\ÛÛ™YØÛÝ[][œ™\ÛÛ™Yˆ]]×Ü™\ÛÛ™YØÛÝ[X]]×Ü™\ÛÛ™YˆÝ[[X\žWÚœÛÛ^Âˆœ›ÝšY\—Ü›ÝÜÈŽˆ[Š›ÝÜÊKˆ˜[\›WÜ›ÝÜÈŽˆ™\Ý[™Ù]
˜[\›WØÛÝ[‹
Kˆ˜[\›WÙ\œ›ÜˆŽˆ™\Ý[™Ù]
˜[\›WÙ\œ›Üˆ‹ˆŠKˆœÝ][Û—Ü›ÝÜÈŽˆ™\Ý[™Ù]
œÝ][Û—ØÛÝ[‹[Š›ÝÜÊJKˆœ™X[[YWÜ›ÝÜÈŽˆ™\Ý[™Ù]
œ™X[[YWØÛÝ[‹[Š›ÝÜÊJKˆœÝ][Û—Ú[™[ÜžWÜÛÝ\˜ÙHŽˆ™\Ý[™Ù]
ˆœÝ][Û—Ú[™[ÜžWÜÛÝ\˜ÙH‹ˆ˜\H‹ˆ
Kˆ˜\WØØ[×Ý\ÙYŽˆ[
™\Ý[™Ù]
˜\WØØ[×Ý\ÙYŠHÜˆ
KˆKˆ
Bˆ›ØÙ\Ü×Û[Ûš]Üš[™×Ø[\ÊÛÛ›‹[\Ù]™[Ë˜]ÚÚY›ÝÊBˆÛÛ›‹˜ÛÛ[Z]

Bˆ]šXÙWØ]˜Z[Xš[]NˆXÝÜÝ‹[žWH›Û™HH›Û™BˆYˆ›ÝšY\ˆOHS•QÔUSÓ—Ô“Õ’QT—Ñ•TÒSÓ”ÓÓTˆ[™šYÙÙ\—Ý\H›Ý[ˆÈœØÚY[YÜÝ]H‹›X[X[Ø˜XÚÙÜ›Ý[™ŸN‚ˆžN‚ˆ]šXÙWØ]˜Z[Xš[]HH[—Ù\Ú[ÛœÛÛ\—Ù]šXÙWØ]˜Z[Xš[]WÜÞ[˜ÊÛÛ›‹›ÝšY\‹šYÙÙ\—Ý\O]šYÙÙ\—Ý\JBˆ^Ù\^Ù\[Ûˆ\È^Î‚ˆÙÙÚ[™Ë™Ù]ÙÙÙ\Š×Û˜[YW×ÊKØ\›š[™Ê‘\Ú[Û”ÛÛ\ˆ]šXÙH]˜Z[Xš[]HÞ[˜È˜Z[Yˆ	\È‹^ÊBˆ™]\›ˆÂˆ›X]ÚYŽˆX]ÚYˆ[œ™\ÛÛ™YŽˆ[œ™\ÛÛ™Yˆ˜]]×Ü™\ÛÛ™YŽˆ]]×Ü™\ÛÛ™Yˆ™]šXÙWØ]˜Z[Xš[]HŽˆ]šXÙWØ]˜Z[Xš[]Kˆ˜\WØØ[×Ý\ÙYŽˆ[
™\Ý[™Ù]
˜\WØØ[×Ý\ÙYŠHÜˆ
KˆœÝ][Û—Ú[™[ÜžWÜÛÝ\˜ÙHŽˆ™\Ý[™Ù]
ˆœÝ][Û—Ú[™[ÜžWÜÛÝ\˜ÙH‹ˆ˜\H‹ˆ
KˆBˆ^Ù\\T˜]S[Z]\œ›Üˆ\È^Î‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUH[YÜ˜][Û—ØÛÛ™šYÜÂˆÑU\ÝÜÞ[˜×ÜÝ]\ÈH	ÝØZ][™×Ü˜]WÛ[Z]	Ë\ÝÙ\œ›ÜˆHË\]YØ]HÂˆÒT‘H›ÝšY\ˆHÂˆˆˆ‹ˆ
^Ë›Y\ÜØYÙK]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠK›ÝšY\ŠKˆ
Bˆš[˜[^™WÚ[YÜ˜][Û—Ü[ŠˆÛÛ›‹ˆ[—ÚYˆÝ]\ÏHØZ][™×Ü˜]WÛ[Z]‹ˆX]ÚYØÛÝ[Lˆ[œ™\ÛÛ™YØÛÝ[Lˆ]]×Ü™\ÛÛ™YØÛÝ[Lˆ\œ›Ü—ÛY\ÜØYÙOY^Ë›Y\ÜØYÙKˆ
BˆÛÛ›‹˜ÛÛ[Z]

Bˆ˜Z\ÙBˆ^Ù\^Ù\[Ûˆ\È^Î‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUH[YÜ˜][Û—ØÛÛ™šYÜÂˆÑU\ÝÜÞ[˜×ÜÝ]\ÈH	Ù\œ›Ü‰Ë\ÝÙ\œ›ÜˆHË\]YØ]HÂˆÒT‘H›ÝšY\ˆHÂˆˆˆ‹ˆ
ÝŠ^ÊK]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠK›ÝšY\ŠKˆ
Bˆš[˜[^™WÚ[YÜ˜][Û—Ü[ŠˆÛÛ›‹ˆ[—ÚYˆÝ]\ÏH™\œ›Üˆ‹ˆX]ÚYØÛÝ[Lˆ[œ™\ÛÛ™YØÛÝ[Lˆ]]×Ü™\ÛÛ™YØÛÝ[Lˆ\œ›Ü—ÛY\ÜØYÙO\ÝŠ^ÊKˆ
BˆÛÛ›‹˜ÛÛ[Z]

Bˆ˜Z\ÙB‚‚™Yˆ[—Ø[Ú[YÜ˜][Û—ÜÞ[˜ÜÊÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹šYÙÙ\—Ý\NˆÝˆH›X[X[ŠHOˆXÝÜÝ‹[žWN‚ˆ™\Ý[ÎˆXÝÜÝ‹[žWHHßBˆ\œ›ÜœÎˆXÝÜÝ‹Ý—HHßBˆ›Üˆ›ÝšY\ˆ[ˆS•QÔUSÓ—Ô“Õ’QT—ÓÔSÓ”Î‚ˆÛÛ™šYÈHÙ]Ú[YÜ˜][Û—ØÛÛ™šYÊÛÛ›‹›ÝšY\ŠBˆYˆÛÛ™šYÈ\È›Û™HÜˆ›ÝÛÛ™šYÖÈ™[˜X›Y—N‚ˆÛÛ[YBˆžN‚ˆ™\Ý[ÖÜ›ÝšY\—HH[—Ú[YÜ˜][Û—ÜÞ[˜ÊÛÛ›‹›ÝšY\‹šYÙÙ\—Ý\O]šYÙÙ\—Ý\JBˆ^Ù\^Ù\[Ûˆ\È^Î‚ˆ\œ›ÜœÖÜ›ÝšY\—HHÝŠ^ÊBˆYˆ›Ý™\Ý[È[™\œ›ÜœÎ‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠŽÈ‹š›Ú[ŠˆžÜ›ÝšY\ŸNˆÛY\ÜØYÙ_Hˆ›Üˆ›ÝšY\‹Y\ÜØYÙH[ˆ\œ›ÜœËš][\Ê
JJBˆ™]\›ˆÈœ™\Ý[ÈŽˆ™\Ý[Ë™\œ›ÜœÈŽˆ\œ›ÜœßB‚‚™Yˆ™\ÛÛ™WÙ\Ú[ÛœÛÛ\—Ý[œ™\ÛÛ™Y
ÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹[œ™\ÛÛ™YÚYˆ[\ÜÙ]ÚYˆ[
HOˆ›Û™N‚ˆ›ÝÈHÛÛ›‹™^XÝ]Jˆ”ÑSPÕ
ˆ”“ÓH[YÜ˜][Û—Ý[œ™\ÛÛ™YÒT‘HYHÈS‘™\ÛÛ][Û—ÜÝ]\ÈH	Ü[™[™ÉÈ‹ˆ
[œ™\ÛÛ™YÚY
Kˆ
K™™]ÚÛ™J
BˆYˆ›ÝÈ\È›Û™N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ‘[˜YH\Ú[Û”ÛÛ\ˆÜˆ™\ÛÛ™\ˆ˜[È[˜ÛÛ˜YKˆŠBˆÜ™X]WÛÜ—Ý\]WØ\ÜÙ]Ú[YÜ˜][ÛŠˆÛÛ›‹ˆ\ÜÙ]ÚYˆ›ÝÖÈœ›ÝšY\ˆ—KˆÝŠ›ÝÖÈ™^\›˜[ÚY—HÜˆˆŠKˆ›ÝÖÈ™^\›˜[Û˜[YH—Kˆ›ÝÖÈ™^\›˜[ÜÝ]\È—HÜˆ“Ü\˜XÚ[Û˜[‹ˆ
BˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUH[YÜ˜][Û—Ý[œ™\ÛÛ™YˆÑU™\ÛÛ][Û—ÜÝ]\ÈH	Ü™\ÛÛ™Y	Ë™\ÛÛ™YØ]HË™\ÛÛ][Û—Û›Ý\ÈHÂˆÒT‘HYHÂˆˆˆ‹ˆ
ˆ]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠKˆˆ\ÜÛØÚXYÈ[È\ÜÙ]Ø\ÜÙ]ÚYH‹ˆ[œ™\ÛÛ™YÚYˆ
Kˆ
BˆÛÛ›‹˜ÛÛ[Z]

B‚‚™YˆÜ™X]WØ\ÜÙ]Ùœ›ÛWÝ[œ™\ÛÛ™Y
ÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹[œ™\ÛÛ™YÚYˆ[
HOˆ[‚ˆ›ÝÈHÛÛ›‹™^XÝ]Jˆ”ÑSPÕ
ˆ”“ÓH[YÜ˜][Û—Ý[œ™\ÛÛ™YÒT‘HYHÈS‘™\ÛÛ][Û—ÜÝ]\ÈH	Ü[™[™ÉÈ‹ˆ
[œ™\ÛÛ™YÚY
Kˆ
K™™]ÚÛ™J
BˆYˆ›ÝÈ\È›Û™N‚ˆ˜Z\ÙH˜[YQ\œ›ÜŠ‘[˜YH\Ú[Û”ÛÛ\ˆÜˆ™\ÛÛ™\ˆ˜[È[˜ÛÛ˜YKˆŠB‚ˆ›Ú™XÝÛ˜[YHH›ÝÖÈ™^\›˜[Û˜[YH—Bˆ[œÝ[][Û—ÙÜ›Ý\H[™™\—Ú[œÝ[][Û—ÙÜ›Ý\
›Ú™XÝÛ˜[YJBˆÝ\œÛÜˆHÛÛ›‹™^XÝ]Jˆˆˆ‚ˆS”ÑT•S•È\ÜÙ]È
›Ú™XÝÛ˜[YK[œÝ[][Û—ÙÜ›Ý\XÝ]™WØÛÛ˜XÝ›Ý\Ë[X\×Ø›ØŠBˆSQTÈ
ËË	Û›ÉËËÊBˆˆˆ‹ˆ
ˆ›Ú™XÝÛ˜[YKˆ[œÝ[][Û—ÙÜ›Ý\ˆˆÜšXYÈH\\ˆÈ›ÝšY\ˆÜ›ÝÖÉÜ›ÝšY\‰×_Kˆ‹ˆ›Ú™XÝÛ˜[YKˆ
Kˆ
Bˆ\ÜÙ]ÚYH[
Ý\œÛÜ‹›\Ý›ÝÚY
Bˆ›Ü›X[^™YÛ˜[YHH›Ü›X[^™WÛ˜[YJ›Ú™XÝÛ˜[YJBˆYˆ›Ü›X[^™YÛ˜[YN‚ˆÛÛ›‹™^XÝ]Jˆ’S”ÑT•ÔˆQÓ“Ô‘HS•È\ÜÙ]Ø[X\Ù\È
\ÜÙ]ÚY[X\×Û˜[YK›Ü›X[^™YØ[X\ËÛÝ\˜ÙJHSQTÈ
ËËËÊH‹ˆ
\ÜÙ]ÚY›Ú™XÝÛ˜[YK›Ü›X[^™YÛ˜[YKš[YÜ˜][Û‹XÜ™X]HŠKˆ
Bˆ™\ÛÛ™WÙ\Ú[ÛœÛÛ\—Ý[œ™\ÛÛ™Y
ÛÛ›‹[œ™\ÛÛ™YÚY\ÜÙ]ÚY
Bˆ™XZ[Ø\ÜÙ]Ø[X\×Ø›ØŠÛÛ›‹\ÜÙ]ÚY
Bˆ™]\›ˆ\ÜÙ]ÚY‚‚™YˆYÛ›Ü™WÙ\Ú[ÛœÛÛ\—Ý[œ™\ÛÛ™Y
ÛÛ›ŽˆÜ[]LËÛÛ›™XÝ[Û‹[œ™\ÛÛ™YÚYˆ[
HOˆ›Û™N‚ˆÛÛ›‹™^XÝ]Jˆˆˆ‚ˆTUH[YÜ˜][Û—Ý[œ™\ÛÛ™YˆÑU™\ÛÛ][Û—ÜÝ]\ÈH	ÚYÛ›Ü™Y	Ë™\ÛÛ™YØ]HË™\ÛÛ][Û—Û›Ý\ÈH	ÒYÛ›Ü˜YÈX[X[Y[IÂˆÒT‘HYHÂˆˆˆ‹ˆ
]][YK››ÝÊ
Kš\ÛÙ›Ü›X]
[Y\ÜXÏHœÙXÛÛ™ÈŠK[œ™\ÛÛ™YÚY
Kˆ
BˆÛÛ›‹˜ÛÛ[Z]

B‚‚˜\HÜ™X]WØ\

B‚‚™Yˆ\œÙWØÛWØ\™ÜÊ
HOˆ\™Ü\œÙK“˜[Y\ÜXÙN‚ˆ\œÙ\ˆH\™Ü\œÙK\™Ý[Y[\œÙ\Š\ØÜš\[ÛH“[Ûš]Üš[™È›Ø\™ØØ[\ŠBˆ\œÙ\‹˜YØ\™Ý[Y[
‹KZÜÝ‹Y˜][HŒLËŒŒŒH‹[H’ÜÝÒTÛ™HH\˜ZH\ØÝ]\ˆŠBˆ\œÙ\‹˜YØ\™Ý[Y[
‹K\Ü‹\OZ[Y˜][ML[H”ÜHÛ™HH\˜ZH\œ˜[˜Ø\ˆŠBˆ\œÙ\‹˜YØ\™Ý[Y[
‹KYXYÈ‹XÝ[ÛHœÝÜ™WÝYH‹[H]]˜HXYÈÈ›\ÚÈŠBˆ™]\›ˆ\œÙ\‹œ\œÙWØ\™ÜÊ
B‚‚šYˆ×Û˜[YW×ÈOH—×ÛXZ[—×ÈŽ‚ˆ\™ÜÈH\œÙWØÛWØ\™ÜÊ
BˆYˆQUSÑVÑSÔU[™›Ý—ÔU™^\ÝÊ
N‚ˆÚ]ÛÜÚ[™ÊÙ]ÙŠÝŠ—ÔU
JJH\ÈÛÛ›Ž‚ˆ[\ÜÙ^Ù[Ù]JÛÛ›‹QUSÑVÑSÔU
Bˆ[YˆQUSÑVÑSÔU‚ˆÚ]ÛÜÚ[™ÊÙ]ÙŠÝŠ—ÔU
JJH\ÈÛÛ›Ž‚ˆYˆ]Y\žWÜØØ[\ŠÛÛ›‹”ÑSPÕÓÕS•

ŠH”“ÓH\ÜÙ]ÈŠHOH‚ˆ[\ÜÙ^Ù[Ù]JÛÛ›‹QUSÑVÑSÔU
Bˆ\œ[ŠÜÝX\™ÜËšÜÝÜX\™ÜËœÜXYÏX\™ÜË™XYÊB