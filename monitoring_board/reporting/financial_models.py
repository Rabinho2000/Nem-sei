from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


PARSER_NAME = "financial_model_workbook"
PARSER_VERSION = "3"


MONTH_LABELS = {
    1: ("1", "jan", "janeiro", "january"),
    2: ("2", "fev", "fevereiro", "feb", "february"),
    3: ("3", "mar", "marco", "marco", "march"),
    4: ("4", "abr", "abril", "apr", "april"),
    5: ("5", "mai", "maio", "may"),
    6: ("6", "jun", "junho", "june"),
    7: ("7", "jul", "julho", "july"),
    8: ("8", "ago", "agosto", "aug", "august"),
    9: ("9", "set", "setembro", "sep", "september"),
    10: ("10", "out", "outubro", "oct", "october"),
    11: ("11", "nov", "novembro", "november"),
    12: ("12", "dez", "dezembro", "dec", "december"),
}

METRIC_ALIASES = {
    "production": ("pv", "pv production", "production", "producao", "producao pv", "producao fotovoltaica", "producao solar"),
    "consumption": ("consumption", "consumo"),
    "self_use": ("sc", "self-consumption", "self consumption", "self-c", "autoconsumo"),
    "self_consumption_rate": ("self-consumption rate", "self consumption rate", "taxa de autoconsumo", "% ac"),
    "export": ("export", "exported energy", "excedente", "excess", "injecao na rede"),
    "grid_import": ("grid import", "import from grid", "importacao da rede", "energia importada", "buy from grid", "sum of grid", "soma de grid"),
}

REQUIRED_MONTHS = set(range(1, 13))


@dataclass(frozen=True)
class ParsedFinancialModel:
    detected_name: str
    detected_nif: str
    detected_kwp: float | None
    base_year: int | None
    sheet_name: str
    monthly: tuple[dict[str, Any], ...]
    warnings: tuple[str, ...]
    details: dict[str, Any]
    source_cells: dict[str, Any]
    parser_name: str = PARSER_NAME
    parser_version: str = PARSER_VERSION


class FinancialModelParseError(ValueError):
    pass


def normalize_text(value: Any) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    ascii_value = "".join(char for char in normalized if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", ascii_value.strip().lower())


def parse_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    raw = str(value).strip().replace(" ", "").replace("\u00a0", "")
    if "," in raw and "." in raw:
        raw = raw.replace(".", "").replace(",", ".")
    else:
        raw = raw.replace(",", ".")
    raw = re.sub(r"[^0-9.\-]", "", raw)
    if raw in {"", "-", "."}:
        return None
    try:
        return float(raw)
    except ValueError:
        return None


def parse_month(value: Any) -> int | None:
    raw = normalize_text(value)
    if not raw:
        return None
    if raw.endswith(".0"):
        raw = raw[:-2]
    for month, labels in MONTH_LABELS.items():
        if raw in labels:
            return month
    return None


def metric_key(value: Any) -> str | None:
    raw = normalize_text(value)
    if not raw:
        return None
    for key in ("self_consumption_rate", "self_use", "grid_import", "export", "production", "consumption"):
        if raw in METRIC_ALIASES[key]:
            return key
    for key in ("self_consumption_rate", "self_use", "grid_import", "export", "production", "consumption"):
        if any(alias in raw for alias in METRIC_ALIASES[key]):
            return key
    return None


def cell_ref(sheet_name: str, row: int, column: int) -> str:
    return f"{sheet_name}!{get_column_letter(column)}{row}"


def parse_financial_model_workbook(path: Path) -> ParsedFinancialModel:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True, keep_vba=False, keep_links=False)
    except (BadZipFile, OSError, ValueError) as exc:
        raise FinancialModelParseError("workbook_invalid") from exc
    try:
        financial_automatic = _financial_automatic_sheets(workbook)
        if financial_automatic is not None:
            project_sheet, savings_sheet = financial_automatic
            sheet_name = project_sheet.title
            monthly, warnings, source_cells = _parse_financial_automatic_monthly(project_sheet, savings_sheet)
        else:
            sheet = _select_monthly_sheet(workbook)
            sheet_name = sheet.title
            monthly, warnings, source_cells = _parse_monthly_sheet(sheet)
        detected_name, detected_nif, detected_kwp, base_year, metadata_cells = _parse_metadata(workbook)
        if financial_automatic is not None:
            project_sheet, savings_sheet = financial_automatic
            base_year, base_year_cell = _parse_financial_automatic_base_year(project_sheet, base_year)
            if base_year_cell:
                metadata_cells["base_year"] = base_year_cell
            details = _parse_financial_automatic_details(project_sheet, savings_sheet, monthly)
        else:
            details = _parse_details(workbook)
        warnings.extend(_annual_reconciliation_warnings(monthly, details))
        return ParsedFinancialModel(
            detected_name=detected_name,
            detected_nif=detected_nif,
            detected_kwp=detected_kwp,
            base_year=base_year,
            sheet_name=sheet_name,
            monthly=tuple(monthly),
            warnings=tuple(sorted(set(warnings))),
            details=details,
            source_cells={**source_cells, **metadata_cells},
        )
    finally:
        workbook.close()


def _annual_reconciliation_warnings(monthly: list[dict[str, Any]], details: dict[str, Any]) -> list[str]:
    references = {
        str(item.get("key") or ""): parse_number(item.get("value"))
        for item in details.get("upac_summary") or []
    }
    mappings = {
        "expected_production_kwh": "annual_pv_production_kwh",
        "expected_consumption_kwh": "annual_consumption_kwh",
        "expected_self_use_kwh": "annual_self_consumption_kwh",
        "expected_export_kwh": "annual_feed_in_kwh",
        "expected_grid_import_kwh": "annual_grid_import_kwh",
    }
    warnings = []
    for monthly_key, annual_key in mappings.items():
        values = [parse_number(row.get(monthly_key)) for row in monthly]
        present = [value for value in values if value is not None]
        reference = references.get(annual_key)
        if not present or reference is None:
            continue
        tolerance = max(1.0, abs(reference) * 0.001)
        if abs(sum(present) - reference) > tolerance:
            metric = monthly_key.removeprefix("expected_").removesuffix("_kwh")
            warnings.append(f"financial_model_annual_total_mismatch_{metric}")
    return warnings


def _financial_automatic_sheets(workbook: Any) -> tuple[Any, Any] | None:
    sheets = {normalize_text(sheet.title): sheet for sheet in workbook.worksheets}
    project_sheet = sheets.get("projeto")
    savings_sheet = sheets.get("savings yr1")
    if project_sheet is None or savings_sheet is None:
        return None
    if _find_financial_automatic_header(project_sheet, {"production", "self_use"}) and _find_financial_automatic_header(
        savings_sheet, {"consumption", "self_use", "export"}
    ):
        return project_sheet, savings_sheet
    return None


def _find_financial_automatic_header(sheet: Any, required: set[str]) -> tuple[int, int, dict[str, int]] | None:
    rows = sheet.iter_rows(
        min_row=1,
        max_row=min(sheet.max_row or 20, 20),
        max_col=min(sheet.max_column or 30, 30),
        values_only=True,
    )
    for row_index, row in enumerate(rows, start=1):
        month_column = None
        headers: dict[str, int] = {}
        for column, value in enumerate(row, start=1):
            if normalize_text(value) in {"month", "mes", "row labels"}:
                month_column = column
            key = _financial_automatic_metric_key(value)
            if key and key not in headers:
                headers[key] = column
        if month_column and required.issubset(headers):
            return row_index, month_column, headers
    return None


def _parse_financial_automatic_monthly(
    project_sheet: Any,
    savings_sheet: Any,
) -> tuple[list[dict[str, Any]], list[str], dict[str, Any]]:
    project_header = _find_financial_automatic_header(project_sheet, {"production", "self_use"})
    savings_header = _find_financial_automatic_header(savings_sheet, {"consumption", "self_use", "export"})
    if project_header is None or savings_header is None:
        raise FinancialModelParseError("financial_model_missing_month")
    project_header_row, project_month_column, project_headers = project_header
    savings_header_row, savings_month_column, savings_headers = savings_header

    project_rows: dict[int, int] = {}
    for row_index in range(project_header_row + 1, min(project_sheet.max_row or project_header_row + 20, project_header_row + 20) + 1):
        month = parse_month(project_sheet.cell(row_index, project_month_column).value)
        if month and month not in project_rows:
            project_rows[month] = row_index
    savings_rows: dict[int, int] = {}
    for row_index in range(savings_header_row + 1, min(savings_sheet.max_row or savings_header_row + 20, savings_header_row + 20) + 1):
        month = parse_month(savings_sheet.cell(row_index, savings_month_column).value)
        if month and month not in savings_rows:
            savings_rows[month] = row_index
    if set(project_rows) != REQUIRED_MONTHS or set(savings_rows) != REQUIRED_MONTHS:
        raise FinancialModelParseError("financial_model_missing_month")

    monthly: list[dict[str, Any]] = []
    source_cells: dict[str, Any] = {"monthly": {}}
    warnings: list[str] = []
    for month in range(1, 13):
        project_row = project_rows[month]
        savings_row = savings_rows[month]
        production_cell = project_sheet.cell(project_row, project_headers["production"])
        project_self_use_cell = project_sheet.cell(project_row, project_headers["self_use"])
        consumption_cell = savings_sheet.cell(savings_row, savings_headers["consumption"])
        savings_self_use_cell = savings_sheet.cell(savings_row, savings_headers["self_use"])
        export_cell = savings_sheet.cell(savings_row, savings_headers["export"])
        production = parse_number(production_cell.value)
        project_self_use = parse_number(project_self_use_cell.value)
        savings_self_use = parse_number(savings_self_use_cell.value)
        if production is None or savings_self_use is None:
            raise FinancialModelParseError("financial_model_missing_month")
        if project_self_use is not None and abs(project_self_use - savings_self_use) > 0.01:
            warnings.append("financial_model_self_use_source_mismatch")
        export = parse_number(export_cell.value)
        export_adjustment_cell = None
        battery_charge_column = savings_headers.get("battery_charge")
        if battery_charge_column and export is not None:
            export_adjustment_cell = savings_sheet.cell(savings_row, battery_charge_column)
            battery_charge = parse_number(export_adjustment_cell.value)
            if battery_charge is not None:
                export = max(export - battery_charge, 0)
                warnings.append("financial_model_battery_export_adjusted")
        row: dict[str, Any] = {
            "month": month,
            "expected_production_kwh": production,
            "expected_consumption_kwh": parse_number(consumption_cell.value),
            "expected_self_use_kwh": savings_self_use,
            "expected_export_kwh": export,
            "expected_grid_import_kwh": None,
            "source_fields": {
                "expected_production_kwh": {"cell": cell_ref(project_sheet.title, project_row, production_cell.column)},
                "expected_consumption_kwh": {"cell": cell_ref(savings_sheet.title, savings_row, consumption_cell.column)},
                "expected_self_use_kwh": {"cell": cell_ref(savings_sheet.title, savings_row, savings_self_use_cell.column)},
                "expected_export_kwh": {"cell": cell_ref(savings_sheet.title, savings_row, export_cell.column)},
            },
        }
        if export_adjustment_cell is not None:
            row["source_fields"]["expected_export_kwh"]["adjustment_cell"] = cell_ref(
                savings_sheet.title, savings_row, export_adjustment_cell.column
            )
            row["calculated_fields"] = {"expected_export_kwh": "export_minus_battery_charge"}
            row["warnings"] = ["financial_model_battery_export_adjusted"]
        grid_import_column = savings_headers.get("grid_import")
        if grid_import_column:
            grid_import_cell = savings_sheet.cell(savings_row, grid_import_column)
            row["expected_grid_import_kwh"] = parse_number(grid_import_cell.value)
            row["source_fields"]["expected_grid_import_kwh"] = {
                "cell": cell_ref(savings_sheet.title, savings_row, grid_import_column)
            }
        monthly.append(row)
        source_cells["monthly"][str(month)] = {
            field: source["cell"] for field, source in row["source_fields"].items()
        }
    return _finalize_monthly(monthly, source_cells, warnings)


def _financial_automatic_metric_key(value: Any) -> str | None:
    raw = normalize_text(value).replace("_", " ")
    if "ess" in raw and raw.startswith(("exc", "exced")):
        return "battery_charge"
    if "production" in raw or "producao" in raw or raw.startswith("pv"):
        return "production"
    if raw.startswith("cons"):
        return "consumption"
    if raw.startswith("ac") or "autoconsumo" in raw or "self consumption" in raw:
        return "self_use"
    if raw.startswith("exced") or "export" in raw:
        return "export"
    if "buy grid" in raw or "grid import" in raw or "importacao da rede" in raw:
        return "grid_import"
    return None


def _select_monthly_sheet(workbook: Any) -> Any:
    for preferred_name in ("prod month", "monthly production", "data pv proposal"):
        preferred = [sheet for sheet in workbook.worksheets if normalize_text(sheet.title) == preferred_name]
        if len(preferred) == 1 and _row_or_column_month_layout(_preview_rows(preferred[0], max_rows=30, max_cols=30)):
            return preferred[0]
    compatible = []
    for sheet in workbook.worksheets:
        rows = _preview_rows(sheet, max_rows=30, max_cols=20)
        if _row_or_column_month_layout(rows):
            compatible.append(sheet)
    if len(compatible) == 1:
        return compatible[0]
    if len(compatible) > 1:
        raise FinancialModelParseError("ambiguous_financial_model")
    raise FinancialModelParseError("financial_model_missing_month")


def _preview_rows(sheet: Any, *, max_rows: int, max_cols: int) -> list[tuple[Any, ...]]:
    return [
        tuple(row)
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row or max_rows, max_rows), max_col=min(sheet.max_column or max_cols, max_cols), values_only=True)
    ]


def _row_or_column_month_layout(rows: list[tuple[Any, ...]]) -> bool:
    for row in rows:
        if sum(1 for value in row if parse_month(value)) >= 10:
            return True
    max_columns = max((len(row) for row in rows), default=0)
    return any(
        sum(1 for row in rows if column < len(row) and parse_month(row[column])) >= 10
        for column in range(max_columns)
    )


def _parse_monthly_sheet(sheet: Any) -> tuple[list[dict[str, Any]], list[str], dict[str, Any]]:
    rows = list(sheet.iter_rows(values_only=False))
    candidates = [candidate for candidate in (_parse_metric_rows(sheet.title, rows), _parse_month_rows(sheet.title, rows)) if candidate]
    parsed = max(candidates, key=lambda candidate: _monthly_candidate_score(candidate[0]), default=None)
    if parsed is None:
        raise FinancialModelParseError("financial_model_missing_month")
    monthly, source_cells = parsed
    return _finalize_monthly(monthly, source_cells)


def _monthly_candidate_score(monthly: list[dict[str, Any]]) -> int:
    fields = (
        "expected_production_kwh",
        "expected_consumption_kwh",
        "expected_self_use_kwh",
        "expected_export_kwh",
        "expected_grid_import_kwh",
    )
    return sum(row.get(field) is not None for row in monthly for field in fields)


def _finalize_monthly(
    monthly: list[dict[str, Any]],
    source_cells: dict[str, Any],
    initial_warnings: list[str] | None = None,
) -> tuple[list[dict[str, Any]], list[str], dict[str, Any]]:
    warnings = list(initial_warnings or [])
    months = {row["month"] for row in monthly}
    if months != REQUIRED_MONTHS:
        raise FinancialModelParseError("financial_model_missing_month")
    if any(row.get("expected_production_kwh") is None for row in monthly):
        raise FinancialModelParseError("financial_model_missing_month")
    for row in monthly:
        calculated: dict[str, str] = dict(row.get("calculated_fields") or {})
        row_warnings: list[str] = list(row.get("warnings") or [])
        production = row.get("expected_production_kwh")
        consumption = row.get("expected_consumption_kwh")
        self_use = row.get("expected_self_use_kwh")
        if row.get("expected_export_kwh") is None and production is not None and self_use is not None:
            row["expected_export_kwh"] = max(production - self_use, 0)
            calculated["expected_export_kwh"] = "production_minus_self_use"
            row_warnings.append("financial_model_calculated_export")
        if row.get("expected_grid_import_kwh") is None and consumption is not None and self_use is not None:
            row["expected_grid_import_kwh"] = max(consumption - self_use, 0)
            calculated["expected_grid_import_kwh"] = "consumption_minus_self_use"
            row_warnings.append("financial_model_calculated_grid_import")
        row["expected_self_consumption_rate_pct"] = _ratio_pct(self_use, production)
        row["expected_self_sufficiency_rate_pct"] = _ratio_pct(self_use, consumption)
        row["calculated_fields"] = calculated
        row["warnings"] = row_warnings
        warnings.extend(row_warnings)
    return monthly, warnings, source_cells


def _parse_metric_rows(sheet_name: str, rows: list[tuple[Any, ...]]) -> tuple[list[dict[str, Any]], dict[str, Any]] | None:
    for row_index, row in enumerate(rows, start=1):
        month_columns = {parse_month(cell.value): col for col, cell in enumerate(row, start=1) if parse_month(cell.value)}
        if len(month_columns) < 12:
            continue
        metrics: dict[str, dict[int, tuple[float, str, str | None]]] = {}
        for metric_row_index in range(row_index + 1, min(row_index + 12, len(rows)) + 1):
            metric_row = rows[metric_row_index - 1]
            metric = metric_key(metric_row[0].value if metric_row else None)
            if not metric:
                continue
            if metric in metrics:
                raise FinancialModelParseError("ambiguous_financial_model")
            metrics[metric] = {}
            unit = _unit_for_label(metric_row[0].value)
            for month, col in month_columns.items():
                cell = metric_row[col - 1]
                value = parse_number(cell.value)
                if value is None:
                    continue
                converted, conversion = _normalize_energy_value(value, metric, unit)
                metrics[metric][month] = (converted, cell_ref(sheet_name, metric_row_index, col), conversion)
        return _build_monthly_from_metrics(metrics)
    return None


def _parse_month_rows(sheet_name: str, rows: list[tuple[Any, ...]]) -> tuple[list[dict[str, Any]], dict[str, Any]] | None:
    header_index = None
    headers: dict[str, int] = {}
    for index, row in enumerate(rows, start=1):
        detected = {metric_key(cell.value): col for col, cell in enumerate(row, start=1) if metric_key(cell.value)}
        if len(detected) >= 1:
            header_index = index
            headers = {key: col for key, col in detected.items() if key}
            break
    if header_index is None:
        return None
    candidate_month_columns: dict[int, int] = {}
    for column in range(1, max((len(row) for row in rows), default=0) + 1):
        candidate_month_columns[column] = sum(
            1
            for row in rows[header_index:]
            if column <= len(row) and parse_month(row[column - 1].value)
        )
    month_column = max(candidate_month_columns, key=candidate_month_columns.get, default=1)
    if candidate_month_columns.get(month_column, 0) < 10:
        return None
    metrics = {key: {} for key in headers}
    for row_index in range(header_index + 1, len(rows) + 1):
        row = rows[row_index - 1]
        month = parse_month(row[month_column - 1].value if len(row) >= month_column else None)
        if not month:
            continue
        for metric, col in headers.items():
            cell = row[col - 1]
            value = parse_number(cell.value)
            if value is None:
                continue
            metrics[metric][month] = (value, cell_ref(sheet_name, row_index, col), None)
    return _build_monthly_from_metrics(metrics)


def _build_monthly_from_metrics(metrics: dict[str, dict[int, tuple[float, str, str | None]]]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    monthly: list[dict[str, Any]] = []
    source_cells: dict[str, Any] = {"monthly": {}}
    mapping = {
        "production": "expected_production_kwh",
        "consumption": "expected_consumption_kwh",
        "self_use": "expected_self_use_kwh",
        "export": "expected_export_kwh",
        "grid_import": "expected_grid_import_kwh",
    }
    for month in range(1, 13):
        row: dict[str, Any] = {"month": month, "source_fields": {}}
        source_cells["monthly"][str(month)] = {}
        for metric, target in mapping.items():
            item = metrics.get(metric, {}).get(month)
            if item is None:
                row[target] = None
                continue
            value, ref, conversion = item
            row[target] = value
            row["source_fields"][target] = {"cell": ref, **({"conversion": conversion} if conversion else {})}
            source_cells["monthly"][str(month)][target] = ref
        monthly.append(row)
    return monthly, source_cells


def _unit_for_label(value: Any) -> str:
    raw = normalize_text(value)
    if "mwh" in raw:
        return "mwh"
    if "kwh" in raw or raw:
        return "kwh"
    return ""


def _normalize_energy_value(value: float, metric: str, unit: str) -> tuple[float, str | None]:
    if metric == "self_consumption_rate":
        return (value * 100 if value <= 1 else value), None
    if unit == "mwh":
        return value * 1000, "mwh_to_kwh"
    return value, None


def _ratio_pct(numerator: float | None, denominator: float | None) -> float | None:
    if numerator is None or denominator is None or denominator <= 0:
        return None
    return numerator / denominator * 100


def _parse_metadata(workbook: Any) -> tuple[str, str, float | None, int | None, dict[str, Any]]:
    cells: dict[str, Any] = {}
    name = ""
    nif = ""
    kwp = None
    base_year = None
    if "UPAC" in workbook.sheetnames:
        sheet = workbook["UPAC"]
        name = str(sheet["A4"].value or "").strip()
        kwp = parse_number(sheet["D4"].value)
        cells.update({"detected_name": "UPAC!A4", "detected_kwp": "UPAC!D4"})
    if "Projeto" in workbook.sheetnames:
        sheet = workbook["Projeto"]
        name = name or str(sheet["C5"].value or "").strip()
        kwp = kwp if kwp is not None else parse_number(sheet["H8"].value)
        cells.update({"detected_name": cells.get("detected_name") or "Projeto!C5", "detected_kwp": cells.get("detected_kwp") or "Projeto!H8"})
    for sheet in workbook.worksheets[:5]:
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row or 30, 30), max_col=min(sheet.max_column or 12, 12), values_only=False):
            for index, cell in enumerate(row):
                label = normalize_text(cell.value)
                if not label:
                    continue
                next_value = row[index + 1].value if index + 1 < len(row) else None
                if not nif and label == "nif":
                    nif = re.sub(r"\D+", "", str(next_value or ""))
                    cells["detected_nif"] = cell_ref(sheet.title, cell.row, cell.column + 1)
                if base_year is None and ("ano base" in label or label in {"year", "ano"}):
                    parsed = parse_number(next_value)
                    candidate = int(parsed) if parsed else None
                    if candidate is not None and 2000 <= candidate <= 2100:
                        base_year = candidate
                        cells["base_year"] = cell_ref(sheet.title, cell.row, cell.column + 1)
    return name, nif, kwp, base_year, cells


def _parse_financial_automatic_base_year(project_sheet: Any, current: int | None) -> tuple[int | None, str | None]:
    if current is not None:
        return current, None
    raw = normalize_text(project_sheet["G39"].value)
    match = re.fullmatch(r"(20\d{2})\s*/\s*\d+", raw)
    if not match:
        return None, None
    return int(match.group(1)), cell_ref(project_sheet.title, 39, 7)


def _parse_financial_automatic_details(
    project_sheet: Any,
    savings_sheet: Any,
    monthly: list[dict[str, Any]],
) -> dict[str, Any]:
    details: dict[str, Any] = {"format": "financial_automatic_as_sold"}
    upac_cells = (
        ("project_name", "Project name", "C5", ""),
        ("installed_capacity_kwp", "Installed capacity", "H8", "kWp"),
        ("installation_cost_eur_kwp", "Installation cost", "D26", "EUR/kWp"),
        ("installation_cost_total_eur", "Installation cost total", "E26", "EUR"),
        ("selling_price_eur_kwp", "Selling price", "D28", "EUR/kWp"),
        ("selling_price_total_eur", "Selling price total", "E28", "EUR"),
        ("first_year_degradation_pct", "First year degradation", "H22", "%"),
        ("degradation_pct", "Annual degradation", "H23", "%"),
        ("annual_consumption_kwh", "Consumption", "P5", "kWh"),
        ("annual_pv_production_kwh", "PV production", "P6", "kWh"),
        ("annual_self_consumption_kwh", "Self-consumption", "P7", "kWh"),
        ("annual_feed_in_kwh", "Feed-in", "P8", "kWh"),
        ("self_consumption_rate_pct", "Self-consumption rate", "P9", "%"),
        ("self_sufficiency_rate_pct", "Self-sufficiency rate", "P10", "%"),
        ("specific_yield_kwh_kwp", "Specific yield", "H14", "kWh/kWp"),
        ("avoided_tariff_eur_kwh", "Avoided tariff", "L32", "EUR/kWh"),
        ("surplus_sale_eur_kwh", "Surplus sale", "F46", "EUR/kWh"),
    )
    upac_summary = [
        _detail_item(key, label, project_sheet[cell].value, cell_ref(project_sheet.title, project_sheet[cell].row, project_sheet[cell].column), unit)
        for key, label, cell, unit in upac_cells
        if project_sheet[cell].value not in (None, "")
    ]
    annual_grid_import = parse_number(savings_sheet["F45"].value)
    annual_grid_import_cell = cell_ref(savings_sheet.title, 45, 6)
    if annual_grid_import is None:
        grid_import_values = [parse_number(row.get("expected_grid_import_kwh")) for row in monthly]
        present_grid_import = [value for value in grid_import_values if value is not None]
        annual_grid_import = sum(present_grid_import) if present_grid_import else None
        annual_grid_import_cell = "calculated:monthly_total"
    if annual_grid_import is not None:
        upac_summary.append(
            _detail_item(
                "annual_grid_import_kwh",
                "Buy from grid",
                annual_grid_import,
                annual_grid_import_cell,
                "kWh",
            )
        )
    tariff_scheme_values = _adjacent_values_for_labels(
        project_sheet,
        {"tipo de tarifa", "ciclo horario", "tarifa de ciclo", "nivel de tensao"},
    )
    if not tariff_scheme_values:
        tariff_scheme_values = [
            str(value).strip()
            for value in (project_sheet["C44"].value, project_sheet["C45"].value)
            if value not in (None, "")
        ]
    tariff_scheme = " / ".join(dict.fromkeys(tariff_scheme_values))
    if tariff_scheme:
        upac_summary.append(
            {
                "key": "tariff_scheme",
                "label": "Tariff scheme",
                "value": tariff_scheme,
                "unit": "",
                "source_cell": "labelled project tariff fields",
            }
        )
    details["upac_summary"] = upac_summary

    tariff_periods = []
    electricity_costs = []
    project_tariff_rows: dict[str, int] = {}
    tariff_layout = _find_project_tariff_layout(project_sheet)
    tariff_row_indexes = range(tariff_layout[0] + 1, tariff_layout[0] + 9) if tariff_layout else range(41, 45)
    period_column, energy_column, network_column, total_column = tariff_layout[1:] if tariff_layout else (5, 6, 7, 8)
    for row_index in tariff_row_indexes:
        label = project_sheet.cell(row_index, period_column).value
        if label in (None, ""):
            continue
        normalized_label = normalize_text(label)
        if normalized_label not in {"sv", "vazio", "cheia", "ponta"}:
            continue
        project_tariff_rows[normalized_label] = row_index
        total = project_sheet.cell(row_index, total_column).value
        if total not in (None, ""):
            tariff_periods.append(
                _detail_item(
                    normalize_text(label).replace(" ", "_"),
                    str(label),
                    total,
                    cell_ref(project_sheet.title, row_index, total_column),
                    "EUR/kWh",
                )
            )
        electricity_costs.append(
            {
                "period": str(label),
                "energy_eur_kwh": parse_number(project_sheet.cell(row_index, energy_column).value),
                "network_eur_kwh": parse_number(project_sheet.cell(row_index, network_column).value),
                "source_cells": {
                    "energy_eur_kwh": cell_ref(project_sheet.title, row_index, energy_column),
                    "network_eur_kwh": cell_ref(project_sheet.title, row_index, network_column),
                },
            }
        )
    details["tariff_periods"] = tariff_periods
    details["electricity_costs"] = electricity_costs

    savings_tariff_rows = {
        normalize_text(savings_sheet.cell(row_index, 2).value): row_index
        for row_index in range(41, 45)
        if savings_sheet.cell(row_index, 2).value not in (None, "")
    }
    savings_value_rows = {
        normalize_text(savings_sheet.cell(row_index, 2).value): row_index
        for row_index in range(48, 52)
        if savings_sheet.cell(row_index, 2).value not in (None, "")
    }
    total_self_use = parse_number(savings_sheet["E45"].value)
    invoice_periods = []
    invoice_prices = []
    for normalized_period, tariff_row in savings_tariff_rows.items():
        value_row = savings_value_rows.get(normalized_period)
        project_row = project_tariff_rows.get(normalized_period)
        label = str(savings_sheet.cell(tariff_row, 2).value)
        self_use = parse_number(savings_sheet.cell(tariff_row, 5).value)
        invoice_periods.append(
            {
                "period": label,
                "self_consumption_kwh": self_use,
                "savings_energy_eur": parse_number(savings_sheet.cell(value_row, 5).value) if value_row else None,
                "share_pct": _ratio_pct(self_use, total_self_use),
                "source_row": tariff_row,
            }
        )
        if project_row:
            energy_kwh = parse_number(savings_sheet.cell(tariff_row, 3).value)
            energy_price = parse_number(project_sheet.cell(project_row, energy_column).value)
            network_price = parse_number(project_sheet.cell(project_row, network_column).value)
            invoice_prices.append(
                {
                    "label": label,
                    "energy_kwh": energy_kwh,
                    "energy_eur_kwh": energy_price,
                    "network_eur_kwh": network_price,
                    "energy_cost_eur": energy_kwh * energy_price if energy_kwh is not None and energy_price is not None else None,
                    "network_cost_eur": energy_kwh * network_price if energy_kwh is not None and network_price is not None else None,
                    "source_row": tariff_row,
                }
            )
    details["invoice_periods"] = invoice_periods
    details["invoice_prices"] = invoice_prices
    invoice_totals = []
    savings_header = _find_financial_automatic_header(savings_sheet, {"consumption", "self_use", "export"})
    total_row = None
    money_headers: dict[str, int] = {}
    if savings_header:
        header_row, month_column, _ = savings_header
        header_values = next(
            savings_sheet.iter_rows(
                min_row=header_row,
                max_row=header_row,
                max_col=min(savings_sheet.max_column or 30, 30),
                values_only=True,
            ),
            (),
        )
        money_headers = {
            key: column
            for column, value in enumerate(header_values, start=1)
            if (key := _financial_money_key(value))
        }
        total_candidates = savings_sheet.iter_rows(
            min_row=header_row + 1,
            max_row=min(savings_sheet.max_row or header_row + 20, header_row + 20),
            min_col=month_column,
            max_col=month_column,
            values_only=True,
        )
        for row_index, (value,) in enumerate(total_candidates, start=header_row + 1):
            if normalize_text(value) in {"total", "grand total"}:
                total_row = row_index
                break
    for key, label, money_key in (
        ("estimated_invoice_total", "Estimated annual invoice", "invoice"),
        ("savings_energy_total", "Savings energy total", "self_use_savings"),
        ("surplus_revenue_total", "Surplus revenue total", "export_revenue"),
    ):
        column = money_headers.get(money_key)
        value = savings_sheet.cell(total_row, column).value if total_row and column else None
        if value not in (None, ""):
            invoice_totals.append(
                _detail_item(key, label, value, cell_ref(savings_sheet.title, total_row, column), "EUR")
            )
    total_benefit = savings_sheet["D52"].value
    if total_benefit not in (None, ""):
        invoice_totals.append(_detail_item("total_benefit", "Total benefit", total_benefit, cell_ref(savings_sheet.title, 52, 4), "EUR"))
    details["invoice_totals"] = invoice_totals
    details["proposal_rows"] = _proposal_rows_from_monthly(monthly)
    return {key: value for key, value in details.items() if value}


def _adjacent_values_for_labels(sheet: Any, labels: set[str]) -> list[str]:
    values = []
    rows = sheet.iter_rows(
        min_row=1,
        max_row=min(sheet.max_row or 70, 70),
        max_col=min(sheet.max_column or 20, 20),
        values_only=True,
    )
    for row in rows:
        for column, value in enumerate(row[:-1]):
            if normalize_text(value) not in labels:
                continue
            adjacent = row[column + 1]
            if adjacent not in (None, ""):
                values.append(str(adjacent).strip())
    return values


def _find_project_tariff_layout(sheet: Any) -> tuple[int, int, int, int, int] | None:
    rows = sheet.iter_rows(
        min_row=1,
        max_row=min(sheet.max_row or 70, 70),
        max_col=min(sheet.max_column or 20, 20),
        values_only=True,
    )
    for row_index, row in enumerate(rows, start=1):
        headers = {normalize_text(value): column for column, value in enumerate(row, start=1)}
        period_column = headers.get("periodo")
        energy_column = headers.get("energia")
        network_column = headers.get("redes")
        total_column = headers.get("total")
        if period_column and energy_column and network_column and total_column:
            return row_index, period_column, energy_column, network_column, total_column
    return None


def _financial_money_key(value: Any) -> str | None:
    raw = normalize_text(value)
    if raw.startswith(("faturas", "invoice", "bill")):
        return "invoice"
    if (raw.startswith("save") or "savings" in raw) and ("ac" in raw or "self" in raw):
        return "self_use_savings"
    if (raw.startswith("exc") or "export" in raw or "surplus" in raw) and ("eur" in raw or "€" in str(value or "")) and "kwh" not in raw:
        return "export_revenue"
    return None


def _proposal_rows_from_monthly(monthly: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for label, key in (
        ("Consumption (kWh)", "expected_consumption_kwh"),
        ("PV production (kWh)", "expected_production_kwh"),
        ("Self-consumption (kWh)", "expected_self_use_kwh"),
        ("Export (kWh)", "expected_export_kwh"),
        ("Buy from grid (kWh)", "expected_grid_import_kwh"),
    ):
        monthly_values = [row.get(key) for row in monthly]
        present = [float(value) for value in monthly_values if value is not None]
        rows.append(
            {
                "label": label,
                "monthly": monthly_values,
                "annual": sum(present) if present else None,
                "source_row": None,
            }
        )
    return rows


def _parse_details(workbook: Any) -> dict[str, Any]:
    details: dict[str, Any] = {}
    if "UPAC" in workbook.sheetnames:
        details["upac_summary"] = _parse_upac_summary(workbook["UPAC"])
        details["tariff_periods"] = _parse_upac_tariff_periods(workbook["UPAC"])
        details["electricity_costs"] = _parse_upac_electricity_costs(workbook["UPAC"])
    if "Data PV Proposal" in workbook.sheetnames:
        details["proposal_rows"] = _parse_proposal_rows(workbook["Data PV Proposal"])
    if "Detalhes da fatura" in workbook.sheetnames:
        invoice = _parse_invoice_details(workbook["Detalhes da fatura"])
        details.update(invoice)
    return {key: value for key, value in details.items() if value}


def _parse_upac_summary(sheet: Any) -> list[dict[str, Any]]:
    cells = (
        ("project_name", "Project name", "A4", ""),
        ("installed_capacity_kwp", "Installed capacity", "D4", "kWp"),
        ("selling_price_eur_kwp", "Selling price", "D5", "EUR/kWp"),
        ("selling_price_total_eur", "Selling price total", "D6", "EUR"),
        ("buying_price_eur_kwp", "Buying price", "D7", "EUR/kWp"),
        ("buying_price_total_eur", "Buying price total", "D8", "EUR"),
        ("contract_duration_years", "Contract duration", "D9", "years"),
        ("degradation_pct", "Degradation", "D11", "%"),
        ("tariff_scheme", "Tariff scheme", "H12", ""),
        ("omie_eur_kwh", "OMIE", "H19", "EUR/kWh"),
        ("annual_consumption_kwh", "Consumption", "K4", "kWh"),
        ("annual_pv_production_kwh", "PV production", "K5", "kWh"),
        ("annual_grid_import_kwh", "Buy from grid", "K6", "kWh"),
        ("annual_self_consumption_kwh", "Self-consumption", "K7", "kWh"),
        ("annual_feed_in_kwh", "Feed-in", "K8", "kWh"),
        ("self_consumption_rate_pct", "Self-consumption rate", "K9", "%"),
        ("self_sufficiency_rate_pct", "Self-sufficiency rate", "K10", "%"),
        ("specific_yield_kwh_kwp", "Specific yield", "K11", "kWh/kWp"),
        ("self_consumption_value_eur_kwh", "Value self-consumption", "K14", "EUR/kWh"),
        ("client_charge_eur_kwh", "Charge client", "K15", "EUR/kWh"),
        ("all_electricity_value_eur_kwh", "Value all electricity", "K19", "EUR/kWh"),
        ("all_electricity_charge_eur_kwh", "Charge client all electricity", "K20", "EUR/kWh"),
        ("savings_year_1_eur", "Savings year 1", "N4", "EUR"),
        ("savings_during_eur", "Savings during", "N6", "EUR"),
        ("savings_after_eur", "Savings after", "N7", "EUR"),
        ("npv_client_eur", "NPV client", "N9", "EUR"),
        ("iberdrola_eur", "Iberdrola", "N11", "EUR"),
        ("solcor_eur", "Solcor", "N12", "EUR"),
        ("current_energy_cost_eur", "Current cost energy", "Q3", "EUR"),
        ("current_power_cost_eur", "Current cost power", "Q4", "EUR"),
        ("current_total_cost_eur", "Current total cost", "Q5", "EUR"),
        ("savings_cost_ratio_pct", "Savings/cost ratio", "Q7", "%"),
        ("co2_ton_year", "CO2", "Q11", "ton/year"),
        ("trees", "Trees", "Q12", ""),
        ("km_year", "km/year", "Q13", ""),
    )
    return [_detail_item(key, label, sheet[cell].value, cell_ref(sheet.title, sheet[cell].row, sheet[cell].column), unit) for key, label, cell, unit in cells if sheet[cell].value not in (None, "")]


def _parse_upac_tariff_periods(sheet: Any) -> list[dict[str, Any]]:
    rows = []
    for row_index in range(13, 20):
        label = sheet.cell(row_index, 7).value
        value = sheet.cell(row_index, 8).value
        if label not in (None, "") and value not in (None, ""):
            rows.append(_detail_item(normalize_text(label).replace(" ", "_"), str(label), value, cell_ref(sheet.title, row_index, 8), "EUR/kWh"))
    return rows


def _parse_upac_electricity_costs(sheet: Any) -> list[dict[str, Any]]:
    rows = []
    for row_index in range(16, 20):
        label = sheet.cell(row_index, 13).value
        energy = sheet.cell(row_index, 14).value
        network = sheet.cell(row_index, 15).value
        if label in (None, ""):
            continue
        rows.append(
            {
                "period": str(label),
                "energy_eur_kwh": parse_number(energy),
                "network_eur_kwh": parse_number(network),
                "source_cells": {
                    "energy_eur_kwh": cell_ref(sheet.title, row_index, 14),
                    "network_eur_kwh": cell_ref(sheet.title, row_index, 15),
                },
            }
        )
    return rows


def _parse_proposal_rows(sheet: Any) -> list[dict[str, Any]]:
    rows = []
    month_columns = list(range(3, 15))
    for row_index in range(1, min(sheet.max_row or 0, 90) + 1):
        label = sheet.cell(row_index, 2).value
        if label in (None, ""):
            continue
        month_values = [parse_number(sheet.cell(row_index, column).value) for column in month_columns]
        annual = parse_number(sheet.cell(row_index, 16).value)
        if not any(value is not None for value in month_values) and annual is None:
            continue
        rows.append(
            {
                "label": str(label),
                "monthly": month_values,
                "annual": annual,
                "source_row": row_index,
            }
        )
    return rows


def _parse_invoice_details(sheet: Any) -> dict[str, Any]:
    periods = []
    for row_index in range(3, min(sheet.max_row or 0, 8) + 1):
        label = sheet.cell(row_index, 2).value
        if label in (None, "") or normalize_text(label) == "grand total":
            continue
        periods.append(
            {
                "period": str(label),
                "self_consumption_kwh": parse_number(sheet.cell(row_index, 3).value),
                "savings_energy_eur": parse_number(sheet.cell(row_index, 4).value),
                "share_pct": _as_pct(parse_number(sheet.cell(row_index, 5).value)),
                "source_row": row_index,
            }
        )
    prices = []
    for row_index in range(3, min(sheet.max_row or 0, 11) + 1):
        label = sheet.cell(row_index, 6).value
        if label in (None, ""):
            continue
        prices.append(
            {
                "label": str(label),
                "energy_kwh": parse_number(sheet.cell(row_index, 7).value),
                "energy_eur_kwh": parse_number(sheet.cell(row_index, 8).value),
                "network_eur_kwh": parse_number(sheet.cell(row_index, 9).value),
                "energy_cost_eur": parse_number(sheet.cell(row_index, 11).value),
                "network_cost_eur": parse_number(sheet.cell(row_index, 12).value),
                "source_row": row_index,
            }
        )
    totals = []
    for label, row_index, column, unit in (
        ("Savings energy total", 7, 4, "EUR"),
        ("Peak power savings", 8, 9, "EUR"),
        ("Revenue surplus", 10, 9, "EUR"),
        ("Total benefit", 12, 7, "EUR"),
        ("Value", 13, 7, "EUR/kWh"),
    ):
        value = sheet.cell(row_index, column).value
        if value not in (None, ""):
            totals.append(_detail_item(normalize_text(label).replace(" ", "_"), label, value, cell_ref(sheet.title, row_index, column), unit))
    return {
        "invoice_periods": periods,
        "invoice_prices": prices,
        "invoice_totals": totals,
    }


def _detail_item(key: str, label: str, value: Any, source_cell: str, unit: str) -> dict[str, Any]:
    parsed = parse_number(value)
    if parsed is not None and unit == "%":
        parsed = _as_pct(parsed)
    return {
        "key": key,
        "label": label,
        "value": parsed if parsed is not None else str(value),
        "unit": unit,
        "source_cell": source_cell,
    }


def _as_pct(value: float | None) -> float | None:
    if value is None:
        return None
    return value * 100 if abs(value) <= 1 else value
