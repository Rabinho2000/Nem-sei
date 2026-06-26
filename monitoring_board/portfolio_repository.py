from __future__ import annotations

import csv
import io
import json
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from monitoring_board.db import ensure_column, query_all
from monitoring_board.portfolio_management import (
    MappingDecision,
    PortfolioImportPreview,
    PortfolioImportRow,
    decide_mapping,
    normalize_name,
    normalize_nif,
    validate_alias,
    validate_portfolio_name,
)


ALIAS_SOURCES = {"manual", "integration", "portfolio_import", "mapping_confirmed", "legacy", "excel-monitoring", "resolved", "manual-edit"}
ORDER_STEP = 10


def now_text() -> str:
    return datetime.now().isoformat(timespec="seconds")


def ensure_portfolio_management_schema(conn: sqlite3.Connection) -> None:
    ensure_column(conn, "asset_aliases", "source TEXT DEFAULT 'legacy'")
    ensure_column(conn, "asset_aliases", "active INTEGER DEFAULT 1")
    ensure_column(conn, "asset_aliases", "created_at TEXT")
    ensure_column(conn, "asset_aliases", "updated_at TEXT")
    ensure_column(conn, "asset_aliases", "notes TEXT")
    ensure_column(conn, "portfolio_groups", "description TEXT")
    ensure_column(conn, "portfolio_groups", "active INTEGER DEFAULT 1")
    ensure_column(conn, "portfolio_groups", "archived_at TEXT")
    ensure_column(conn, "portfolio_groups", "display_order INTEGER DEFAULT 0")
    ensure_column(conn, "portfolio_groups", "created_at TEXT")
    ensure_column(conn, "portfolio_groups", "updated_at TEXT")
    ensure_column(conn, "portfolio_assets", "display_order INTEGER DEFAULT 0")
    ensure_column(conn, "portfolio_assets", "mapping_method TEXT")
    ensure_column(conn, "portfolio_assets", "mapped_at TEXT")
    ensure_column(conn, "portfolio_assets", "updated_at TEXT")
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS portfolio_mapping_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            portfolio_asset_id INTEGER,
            external_name TEXT,
            previous_asset_id INTEGER,
            selected_asset_id INTEGER,
            method TEXT,
            confidence REAL,
            alias_created INTEGER DEFAULT 0,
            notes TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY (portfolio_asset_id) REFERENCES portfolio_assets(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS portfolio_import_runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            portfolio_id INTEGER,
            original_filename TEXT NOT NULL,
            status TEXT NOT NULL,
            rows_total INTEGER NOT NULL DEFAULT 0,
            rows_valid INTEGER NOT NULL DEFAULT 0,
            rows_pending INTEGER NOT NULL DEFAULT 0,
            rows_conflict INTEGER NOT NULL DEFAULT 0,
            preview_json TEXT,
            created_at TEXT NOT NULL,
            applied_at TEXT,
            FOREIGN KEY (portfolio_id) REFERENCES portfolio_groups(id) ON DELETE SET NULL
        );
        CREATE INDEX IF NOT EXISTS idx_portfolio_assets_portfolio_order ON portfolio_assets(portfolio_id, display_order, id);
        CREATE INDEX IF NOT EXISTS idx_portfolio_assets_portfolio_sub ON portfolio_assets(portfolio_id, sub_account);
        CREATE INDEX IF NOT EXISTS idx_portfolio_mapping_events_member ON portfolio_mapping_events(portfolio_asset_id);
        """
    )
    current = now_text()
    conn.execute("UPDATE asset_aliases SET source = 'legacy' WHERE source IS NULL OR source = ''")
    conn.execute("UPDATE asset_aliases SET active = 1 WHERE active IS NULL")
    conn.execute("UPDATE asset_aliases SET created_at = ? WHERE created_at IS NULL OR created_at = ''", (current,))
    conn.execute("UPDATE asset_aliases SET updated_at = COALESCE(NULLIF(updated_at, ''), created_at, ?)", (current,))
    conn.execute("UPDATE portfolio_groups SET active = 1 WHERE active IS NULL")
    conn.execute("UPDATE portfolio_groups SET created_at = ? WHERE created_at IS NULL OR created_at = ''", (current,))
    conn.execute("UPDATE portfolio_groups SET updated_at = COALESCE(NULLIF(updated_at, ''), created_at, ?)", (current,))
    conn.execute("UPDATE portfolio_assets SET mapping_method = COALESCE(NULLIF(mapping_method, ''), mapping_status) WHERE mapping_method IS NULL OR mapping_method = ''")
    conn.execute("UPDATE portfolio_assets SET updated_at = ? WHERE updated_at IS NULL OR updated_at = ''", (current,))
    _backfill_display_order(conn)


def _backfill_display_order(conn: sqlite3.Connection) -> None:
    groups = query_all(conn, "SELECT id FROM portfolio_groups ORDER BY id")
    for group in groups:
        rows = query_all(conn, "SELECT id, display_order FROM portfolio_assets WHERE portfolio_id = ? ORDER BY COALESCE(display_order, 0), id", (group["id"],))
        changed = False
        for index, row in enumerate(rows, start=1):
            order = index * ORDER_STEP
            if not row["display_order"] or int(row["display_order"]) != order:
                conn.execute("UPDATE portfolio_assets SET display_order = ? WHERE id = ?", (order, row["id"]))
                changed = True
        if changed:
            conn.execute("UPDATE portfolio_groups SET updated_at = ? WHERE id = ?", (now_text(), group["id"]))


def list_portfolios(conn: sqlite3.Connection, *, include_archived: bool = False) -> list[sqlite3.Row]:
    where = "" if include_archived else "WHERE active = 1 AND archived_at IS NULL"
    return query_all(conn, f"SELECT * FROM portfolio_groups {where} ORDER BY display_order, name COLLATE NOCASE, id")


def get_portfolio(conn: sqlite3.Connection, portfolio_id: int) -> sqlite3.Row | None:
    return conn.execute("SELECT * FROM portfolio_groups WHERE id = ?", (portfolio_id,)).fetchone()


def create_portfolio(conn: sqlite3.Connection, *, name: str, description: str = "", notes: str = "") -> int:
    clean = validate_portfolio_name(name)
    order = int(conn.execute("SELECT COALESCE(MAX(display_order), 0) + ? FROM portfolio_groups", (ORDER_STEP,)).fetchone()[0] or ORDER_STEP)
    current = now_text()
    cursor = conn.execute(
        """
        INSERT INTO portfolio_groups (name, description, notes, active, display_order, created_at, updated_at)
        VALUES (?, ?, ?, 1, ?, ?, ?)
        """,
        (clean, description.strip(), notes.strip(), order, current, current),
    )
    return int(cursor.lastrowid)


def update_portfolio(conn: sqlite3.Connection, *, portfolio_id: int, name: str, description: str = "", notes: str = "", display_order: int | None = None) -> None:
    if get_portfolio(conn, portfolio_id) is None:
        raise ValueError("portfolio_not_found")
    conn.execute(
        """
        UPDATE portfolio_groups
        SET name = ?, description = ?, notes = ?, display_order = COALESCE(?, display_order), updated_at = ?
        WHERE id = ?
        """,
        (validate_portfolio_name(name), description.strip(), notes.strip(), display_order, now_text(), portfolio_id),
    )


def duplicate_portfolio(conn: sqlite3.Connection, *, portfolio_id: int, new_name: str) -> int:
    source = get_portfolio(conn, portfolio_id)
    if source is None:
        raise ValueError("portfolio_not_found")
    new_id = create_portfolio(conn, name=new_name, description=source["description"] or "", notes=source["notes"] or "")
    for row in list_portfolio_members(conn, portfolio_id):
        conn.execute(
            """
            INSERT INTO portfolio_assets (
                portfolio_id, asset_id, external_name, nif, sub_account, active, mapping_status,
                mapping_confidence, notes, display_order, mapping_method, mapped_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                new_id,
                row["asset_id"],
                row["external_name"],
                row["nif"],
                row["sub_account"],
                row["active"],
                row["mapping_status"],
                row["mapping_confidence"],
                row["notes"],
                row["display_order"],
                row["mapping_method"],
                row["mapped_at"],
                now_text(),
            ),
        )
    return new_id


def archive_portfolio(conn: sqlite3.Connection, portfolio_id: int) -> None:
    if get_portfolio(conn, portfolio_id) is None:
        raise ValueError("portfolio_not_found")
    conn.execute("UPDATE portfolio_groups SET active = 0, archived_at = ?, updated_at = ? WHERE id = ?", (now_text(), now_text(), portfolio_id))


def reactivate_portfolio(conn: sqlite3.Connection, portfolio_id: int) -> None:
    if get_portfolio(conn, portfolio_id) is None:
        raise ValueError("portfolio_not_found")
    conn.execute("UPDATE portfolio_groups SET active = 1, archived_at = NULL, updated_at = ? WHERE id = ?", (now_text(), portfolio_id))


def delete_portfolio(conn: sqlite3.Connection, portfolio_id: int, *, confirm_name: str) -> None:
    portfolio = get_portfolio(conn, portfolio_id)
    if portfolio is None:
        raise ValueError("portfolio_not_found")
    if confirm_name.strip() != portfolio["name"]:
        raise ValueError("delete_confirmation_mismatch")
    if conn.execute("SELECT 1 FROM portfolio_report_runs WHERE portfolio_id = ? LIMIT 1", (portfolio_id,)).fetchone():
        raise ValueError("portfolio_has_report_history")
    conn.execute("DELETE FROM portfolio_groups WHERE id = ?", (portfolio_id,))


def list_portfolio_members(conn: sqlite3.Connection, portfolio_id: int) -> list[sqlite3.Row]:
    return query_all(
        conn,
        """
        SELECT pa.*, a.project_name, a.nif AS asset_nif, a.alias_blob
        FROM portfolio_assets pa
        LEFT JOIN assets a ON a.id = pa.asset_id
        WHERE pa.portfolio_id = ?
        ORDER BY pa.display_order, pa.id
        """,
        (portfolio_id,),
    )


def list_available_assets(conn: sqlite3.Connection, *, portfolio_id: int | None = None, search: str = "", asset_filter: str = "available") -> list[sqlite3.Row]:
    params: list[Any] = []
    where: list[str] = []
    if search:
        like = f"%{search.strip()}%"
        where.append("(a.project_name LIKE ? OR COALESCE(a.nif, '') LIKE ? OR COALESCE(a.alias_blob, '') LIKE ?)")
        params.extend([like, like, like])
    if portfolio_id and asset_filter == "available":
        where.append(
            """
            NOT EXISTS (
                SELECT 1 FROM portfolio_assets current_pa
                WHERE current_pa.portfolio_id = ? AND current_pa.asset_id = a.id AND current_pa.active = 1
            )
            """
        )
        params.append(portfolio_id)
    elif portfolio_id and asset_filter == "in_current":
        where.append(
            """
            EXISTS (
                SELECT 1 FROM portfolio_assets current_pa
                WHERE current_pa.portfolio_id = ? AND current_pa.asset_id = a.id AND current_pa.active = 1
            )
            """
        )
        params.append(portfolio_id)
    elif asset_filter == "without_portfolio":
        where.append("NOT EXISTS (SELECT 1 FROM portfolio_assets any_pa WHERE any_pa.asset_id = a.id AND any_pa.active = 1)")
    elif portfolio_id and asset_filter == "other_portfolio":
        where.append("EXISTS (SELECT 1 FROM portfolio_assets other_pa WHERE other_pa.asset_id = a.id AND other_pa.active = 1 AND other_pa.portfolio_id != ?)")
        params.append(portfolio_id)
    elif asset_filter == "inactive":
        where.append("COALESCE(a.active_contract, '') NOT IN ('yes', 'Sim', 'sim', 'active')")
    elif asset_filter == "mapping_pending":
        where.append("EXISTS (SELECT 1 FROM portfolio_assets pending_pa WHERE pending_pa.asset_id = a.id AND pending_pa.mapping_status = 'mapping_pending')")
    elif asset_filter == "conflict":
        where.append("EXISTS (SELECT 1 FROM portfolio_assets conflict_pa WHERE conflict_pa.asset_id = a.id AND conflict_pa.mapping_status = 'mapping_conflict')")
    sql = """
        SELECT a.id, a.project_name, a.nif, a.active_contract, a.alias_blob,
               GROUP_CONCAT(DISTINCT pg.name) AS portfolios,
               GROUP_CONCAT(DISTINCT ai.provider || ':' || ai.external_name) AS integrations
        FROM assets a
        LEFT JOIN portfolio_assets pa ON pa.asset_id = a.id AND pa.active = 1
        LEFT JOIN portfolio_groups pg ON pg.id = pa.portfolio_id
        LEFT JOIN asset_integrations ai ON ai.asset_id = a.id AND ai.enabled = 1
    """
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " GROUP BY a.id ORDER BY a.project_name COLLATE NOCASE"
    return query_all(conn, sql, params)


def add_member(
    conn: sqlite3.Connection,
    *,
    portfolio_id: int,
    asset_id: int | None,
    external_name: str = "",
    nif: str = "",
    sub_account: str = "",
    notes: str = "",
    mapping_method: str = "manual",
    confidence: float = 1.0,
) -> int:
    _require_portfolio(conn, portfolio_id, active_only=True)
    if asset_id is not None:
        _require_asset(conn, asset_id)
        existing = conn.execute("SELECT id, active FROM portfolio_assets WHERE portfolio_id = ? AND asset_id = ?", (portfolio_id, asset_id)).fetchone()
        if existing:
            if int(existing["active"] or 0) == 0:
                conn.execute(
                    """
                    UPDATE portfolio_assets
                    SET active = 1, external_name = ?, nif = ?, sub_account = ?, notes = ?,
                        mapping_status = 'manual', mapping_confidence = ?, mapping_method = ?, mapped_at = ?, updated_at = ?
                    WHERE id = ? AND portfolio_id = ?
                    """,
                    (
                        external_name.strip(),
                        normalize_nif(nif),
                        sub_account.strip(),
                        notes.strip(),
                        confidence,
                        mapping_method,
                        now_text(),
                        now_text(),
                        existing["id"],
                        portfolio_id,
                    ),
                )
                return int(existing["id"])
            raise ValueError("member_already_exists")
    order = next_member_order(conn, portfolio_id)
    status = "manual" if asset_id is not None else "mapping_pending"
    cursor = conn.execute(
        """
        INSERT INTO portfolio_assets (
            portfolio_id, asset_id, external_name, nif, sub_account, active, mapping_status,
            mapping_confidence, notes, display_order, mapping_method, mapped_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, 1, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            portfolio_id,
            asset_id,
            external_name.strip(),
            normalize_nif(nif),
            sub_account.strip(),
            status,
            confidence if asset_id is not None else 0.0,
            notes.strip(),
            order,
            mapping_method if asset_id is not None else "unmapped",
            now_text() if asset_id is not None else None,
            now_text(),
        ),
    )
    return int(cursor.lastrowid)


def update_member(
    conn: sqlite3.Connection,
    *,
    member_id: int,
    portfolio_id: int,
    asset_id: int | None,
    external_name: str,
    nif: str,
    sub_account: str,
    notes: str,
    active: bool,
    create_alias: bool = False,
) -> None:
    member = _require_member(conn, member_id, portfolio_id)
    previous_asset_id = member["asset_id"]
    if asset_id is not None:
        _require_asset(conn, asset_id)
        duplicate = conn.execute(
            "SELECT * FROM portfolio_assets WHERE portfolio_id = ? AND asset_id = ? AND id != ? LIMIT 1",
            (portfolio_id, asset_id, member_id),
        ).fetchone()
        if duplicate:
            if member["asset_id"] is None:
                current = now_text()
                conn.execute(
                    """
                    UPDATE portfolio_assets
                    SET external_name = ?, nif = ?, sub_account = ?, active = ?,
                        mapping_status = 'manual', mapping_confidence = 1.0, mapping_method = 'manual',
                        mapped_at = ?, notes = ?, updated_at = ?
                    WHERE id = ? AND portfolio_id = ?
                    """,
                    (
                        external_name.strip(),
                        normalize_nif(nif),
                        sub_account.strip(),
                        1 if active else 0,
                        current,
                        notes.strip(),
                        current,
                        duplicate["id"],
                        portfolio_id,
                    ),
                )
                alias_created = False
                if create_alias and external_name.strip():
                    alias_created = upsert_alias(conn, asset_id=asset_id, alias_name=external_name, source="mapping_confirmed", notes="Criado ao confirmar mapping") is not None
                    rebuild_asset_alias_blob(conn, asset_id)
                record_mapping_event(
                    conn,
                    portfolio_asset_id=duplicate["id"],
                    external_name=external_name,
                    previous_asset_id=None,
                    selected_asset_id=asset_id,
                    method="manual_merge",
                    confidence=1.0,
                    alias_created=alias_created,
                    notes=f"Merge da entrada pendente #{member_id}",
                )
                conn.execute("DELETE FROM portfolio_assets WHERE id = ? AND portfolio_id = ?", (member_id, portfolio_id))
                normalize_member_order(conn, portfolio_id)
                return
            raise ValueError("member_already_exists")
    conn.execute(
        """
        UPDATE portfolio_assets
        SET asset_id = ?, external_name = ?, nif = ?, sub_account = ?, active = ?,
            mapping_status = ?, mapping_confidence = ?, mapping_method = ?, mapped_at = ?, notes = ?, updated_at = ?
        WHERE id = ? AND portfolio_id = ?
        """,
        (
            asset_id,
            external_name.strip(),
            normalize_nif(nif),
            sub_account.strip(),
            1 if active else 0,
            "manual" if asset_id is not None else "mapping_pending",
            1.0 if asset_id is not None else 0.0,
            "manual" if asset_id is not None else "unmapped",
            now_text() if asset_id is not None else None,
            notes.strip(),
            now_text(),
            member_id,
            portfolio_id,
        ),
    )
    alias_created = False
    if create_alias and asset_id is not None and external_name.strip():
        alias_created = upsert_alias(conn, asset_id=asset_id, alias_name=external_name, source="mapping_confirmed", notes="Criado ao confirmar mapping") is not None
    record_mapping_event(conn, portfolio_asset_id=member_id, external_name=external_name, previous_asset_id=previous_asset_id, selected_asset_id=asset_id, method="manual", confidence=1.0 if asset_id else 0.0, alias_created=alias_created)


def remove_members(conn: sqlite3.Connection, *, portfolio_id: int, member_ids: list[int]) -> None:
    ids = _validated_unique_ids(member_ids)
    for member_id in ids:
        _require_member(conn, member_id, portfolio_id)
    conn.executemany("DELETE FROM portfolio_assets WHERE id = ? AND portfolio_id = ?", [(member_id, portfolio_id) for member_id in ids])
    normalize_member_order(conn, portfolio_id)


def sync_portfolio_asset_members(
    conn: sqlite3.Connection,
    *,
    portfolio_id: int,
    asset_ids: list[int],
    asset_names: dict[int, str] | None = None,
) -> dict[str, int]:
    _require_portfolio(conn, portfolio_id, active_only=True)
    cleaned: list[int] = []
    seen: set[int] = set()
    for asset_id in asset_ids:
        parsed = int(asset_id)
        if parsed <= 0 or parsed in seen:
            continue
        seen.add(parsed)
        cleaned.append(parsed)
    if len(cleaned) > 100:
        raise ValueError("too_many_assets")
    for asset_id in cleaned:
        _require_asset(conn, asset_id)

    names = asset_names or {}
    existing_rows = query_all(
        conn,
        "SELECT id, asset_id FROM portfolio_assets WHERE portfolio_id = ? AND asset_id IS NOT NULL ORDER BY display_order, id",
        (portfolio_id,),
    )
    existing_by_asset = {int(row["asset_id"]): int(row["id"]) for row in existing_rows}
    selected = set(cleaned)
    remove_ids = [int(row["id"]) for row in existing_rows if int(row["asset_id"]) not in selected]
    if remove_ids:
        conn.executemany("DELETE FROM portfolio_assets WHERE id = ? AND portfolio_id = ?", [(member_id, portfolio_id) for member_id in remove_ids])

    added = 0
    current = now_text()
    for index, asset_id in enumerate(cleaned, start=1):
        member_id = existing_by_asset.get(asset_id)
        display_order = index * ORDER_STEP
        external_name = str(names.get(asset_id, "") or "").strip()
        if member_id:
            conn.execute(
                """
                UPDATE portfolio_assets
                SET active = 1, external_name = ?, display_order = ?, updated_at = ?
                WHERE id = ? AND portfolio_id = ?
                """,
                (external_name, display_order, current, member_id, portfolio_id),
            )
            continue
        conn.execute(
            """
            INSERT INTO portfolio_assets (
                portfolio_id, asset_id, external_name, nif, sub_account, active, mapping_status,
                mapping_confidence, notes, display_order, mapping_method, mapped_at, updated_at
            ) VALUES (?, ?, ?, '', '', 1, 'manual', 1.0, '', ?, 'manual', ?, ?)
            """,
            (portfolio_id, asset_id, external_name, display_order, current, current),
        )
        added += 1
    normalize_member_order(conn, portfolio_id)
    return {"selected": len(cleaned), "added": added, "removed": len(remove_ids)}


def copy_members(conn: sqlite3.Connection, *, source_portfolio_id: int, target_portfolio_id: int, member_ids: list[int], move: bool = False) -> None:
    _require_portfolio(conn, source_portfolio_id)
    _require_portfolio(conn, target_portfolio_id, active_only=True)
    ids = _validated_unique_ids(member_ids)
    copied: list[int] = []
    moved_source_ids: list[int] = []
    for member_id in ids:
        member = _require_member(conn, member_id, source_portfolio_id)
        if member["asset_id"] is not None and conn.execute("SELECT 1 FROM portfolio_assets WHERE portfolio_id = ? AND asset_id = ? AND active = 1", (target_portfolio_id, member["asset_id"])).fetchone():
            continue
        copied.append(
            add_member(
                conn,
                portfolio_id=target_portfolio_id,
                asset_id=member["asset_id"],
                external_name=member["external_name"] or "",
                nif=member["nif"] or "",
                sub_account=member["sub_account"] or "",
                notes=member["notes"] or "",
                mapping_method=member["mapping_method"] or "manual",
                confidence=float(member["mapping_confidence"] or 0),
            )
        )
        moved_source_ids.append(member_id)
    if move:
        if not moved_source_ids:
            raise ValueError("no_members_moved")
        remove_members(conn, portfolio_id=source_portfolio_id, member_ids=moved_source_ids)
    if not copied and not move:
        raise ValueError("no_members_copied")


def reorder_members(conn: sqlite3.Connection, *, portfolio_id: int, ordered_ids: list[int]) -> None:
    ids = _validated_unique_ids(ordered_ids)
    existing = [int(row["id"]) for row in query_all(conn, "SELECT id FROM portfolio_assets WHERE portfolio_id = ? ORDER BY display_order, id", (portfolio_id,))]
    if set(ids) != set(existing):
        raise ValueError("invalid_member_order")
    for index, member_id in enumerate(ids, start=1):
        conn.execute("UPDATE portfolio_assets SET display_order = ?, updated_at = ? WHERE id = ? AND portfolio_id = ?", (index * ORDER_STEP, now_text(), member_id, portfolio_id))


def move_member_up_down(conn: sqlite3.Connection, *, portfolio_id: int, member_id: int, direction: str) -> None:
    rows = [int(row["id"]) for row in list_portfolio_members(conn, portfolio_id)]
    if member_id not in rows:
        raise ValueError("member_not_found")
    index = rows.index(member_id)
    target = index - 1 if direction == "up" else index + 1
    if target < 0 or target >= len(rows):
        return
    rows[index], rows[target] = rows[target], rows[index]
    reorder_members(conn, portfolio_id=portfolio_id, ordered_ids=rows)


def normalize_member_order(conn: sqlite3.Connection, portfolio_id: int) -> None:
    rows = [int(row["id"]) for row in list_portfolio_members(conn, portfolio_id)]
    for index, member_id in enumerate(rows, start=1):
        conn.execute("UPDATE portfolio_assets SET display_order = ? WHERE id = ?", (index * ORDER_STEP, member_id))


def next_member_order(conn: sqlite3.Connection, portfolio_id: int) -> int:
    return int(conn.execute("SELECT COALESCE(MAX(display_order), 0) + ? FROM portfolio_assets WHERE portfolio_id = ?", (ORDER_STEP, portfolio_id)).fetchone()[0] or ORDER_STEP)


def list_aliases(conn: sqlite3.Connection, *, asset_id: int | None = None, include_inactive: bool = True) -> list[sqlite3.Row]:
    conditions: list[str] = []
    params: list[Any] = []
    if asset_id is not None:
        conditions.append("aa.asset_id = ?")
        params.append(asset_id)
    if not include_inactive:
        conditions.append("aa.active = 1")
    where = "WHERE " + " AND ".join(conditions) if conditions else ""
    return query_all(
        conn,
        f"""
        SELECT aa.*, a.project_name,
               (SELECT COUNT(*) FROM portfolio_assets pa WHERE pa.asset_id = aa.asset_id AND COALESCE(pa.external_name, '') != '') AS used_count
        FROM asset_aliases aa
        JOIN assets a ON a.id = aa.asset_id
        {where}
        ORDER BY aa.active DESC, aa.alias_name COLLATE NOCASE
        """,
        params,
    )


def upsert_alias(conn: sqlite3.Connection, *, asset_id: int, alias_name: str, source: str = "manual", notes: str = "") -> int | None:
    _require_asset(conn, asset_id)
    clean, normalized = validate_alias(alias_name)
    existing = conn.execute("SELECT * FROM asset_aliases WHERE normalized_alias = ?", (normalized,)).fetchone()
    current = now_text()
    if existing:
        if int(existing["asset_id"]) != asset_id:
            raise ValueError("alias_conflict")
        conn.execute(
            "UPDATE asset_aliases SET alias_name = ?, source = ?, active = 1, notes = ?, updated_at = ? WHERE id = ?",
            (clean, source if source in ALIAS_SOURCES else "manual", notes.strip(), current, existing["id"]),
        )
        return None
    cursor = conn.execute(
        """
        INSERT INTO asset_aliases (asset_id, alias_name, normalized_alias, source, active, created_at, updated_at, notes)
        VALUES (?, ?, ?, ?, 1, ?, ?, ?)
        """,
        (asset_id, clean, normalized, source if source in ALIAS_SOURCES else "manual", current, current, notes.strip()),
    )
    return int(cursor.lastrowid)


def update_alias(conn: sqlite3.Connection, *, asset_id: int, alias_id: int, alias_name: str, notes: str = "") -> None:
    alias = _require_alias(conn, asset_id, alias_id)
    clean, normalized = validate_alias(alias_name)
    existing = conn.execute("SELECT id, asset_id FROM asset_aliases WHERE normalized_alias = ? AND id != ?", (normalized, alias_id)).fetchone()
    if existing:
        raise ValueError("alias_conflict")
    conn.execute("UPDATE asset_aliases SET alias_name = ?, normalized_alias = ?, source = 'manual', notes = ?, updated_at = ? WHERE id = ?", (clean, normalized, notes.strip(), now_text(), alias["id"]))


def toggle_alias(conn: sqlite3.Connection, *, asset_id: int, alias_id: int, active: bool) -> None:
    _require_alias(conn, asset_id, alias_id)
    conn.execute("UPDATE asset_aliases SET active = ?, updated_at = ? WHERE id = ? AND asset_id = ?", (1 if active else 0, now_text(), alias_id, asset_id))


def delete_alias(conn: sqlite3.Connection, *, asset_id: int, alias_id: int) -> None:
    _require_alias(conn, asset_id, alias_id)
    conn.execute("DELETE FROM asset_aliases WHERE id = ? AND asset_id = ?", (alias_id, asset_id))


def rebuild_asset_alias_blob(conn: sqlite3.Connection, asset_id: int) -> None:
    aliases = query_all(conn, "SELECT alias_name FROM asset_aliases WHERE asset_id = ? AND active = 1 ORDER BY alias_name COLLATE NOCASE", (asset_id,))
    conn.execute("UPDATE assets SET alias_blob = ? WHERE id = ?", (" | ".join(row["alias_name"] for row in aliases), asset_id))


def mapping_context(conn: sqlite3.Connection) -> tuple[tuple[dict[str, Any], ...], tuple[dict[str, Any], ...]]:
    assets = tuple(dict(row) for row in query_all(conn, "SELECT id, project_name, nif FROM assets ORDER BY id"))
    aliases = tuple(dict(row) for row in query_all(conn, "SELECT id, asset_id, alias_name, normalized_alias, active FROM asset_aliases ORDER BY id"))
    extras: list[dict[str, Any]] = []
    for row in query_all(conn, "SELECT DISTINCT asset_id, external_name FROM portfolio_assets WHERE asset_id IS NOT NULL AND COALESCE(external_name, '') != ''"):
        extras.append({"id": None, "asset_id": row["asset_id"], "alias_name": row["external_name"], "normalized_alias": normalize_name(row["external_name"]), "active": True})
    return assets, tuple([*aliases, *extras])


def suggest_mapping(conn: sqlite3.Connection, *, external_name: str, nif: str = "") -> MappingDecision:
    assets, aliases = mapping_context(conn)
    return decide_mapping(external_name=external_name, nif=nif, assets=assets, aliases=aliases)


def auto_map_portfolio_assets(conn: sqlite3.Connection, portfolio_id: int | None = None) -> dict[str, int]:
    conditions = ["active = 1"]
    params: list[Any] = []
    if portfolio_id:
        conditions.append("portfolio_id = ?")
        params.append(portfolio_id)
    rows = query_all(conn, f"SELECT * FROM portfolio_assets WHERE {' AND '.join(conditions)}", params)
    mapped = pending = conflicts = 0
    for row in rows:
        decision = suggest_mapping(conn, external_name=row["external_name"] or "", nif=row["nif"] or "")
        if decision.auto_mappable and decision.asset_id is not None:
            mapped += 1
            conn.execute(
                """
                UPDATE portfolio_assets
                SET asset_id = ?, mapping_status = ?, mapping_method = ?, mapping_confidence = ?, mapped_at = ?, updated_at = ?
                WHERE id = ?
                """,
                (decision.asset_id, "mapped", decision.method, decision.score, now_text(), now_text(), row["id"]),
            )
        elif decision.status == "mapping_conflict":
            conflicts += 1
            conn.execute("UPDATE portfolio_assets SET mapping_status = 'mapping_conflict', mapping_method = 'conflict', mapping_confidence = ?, updated_at = ? WHERE id = ?", (decision.score, now_text(), row["id"]))
        else:
            pending += 1
            conn.execute("UPDATE portfolio_assets SET mapping_status = ?, mapping_method = ?, mapping_confidence = ?, updated_at = ? WHERE id = ?", (decision.status, decision.method, decision.score, now_text(), row["id"]))
    return {"mapped": mapped, "pending": pending, "conflicts": conflicts}


def confirm_mapping(conn: sqlite3.Connection, *, member_id: int, portfolio_id: int, asset_id: int, create_alias: bool = True) -> None:
    member = _require_member(conn, member_id, portfolio_id)
    _require_asset(conn, asset_id)
    duplicate = conn.execute(
        "SELECT * FROM portfolio_assets WHERE portfolio_id = ? AND asset_id = ? AND id != ? LIMIT 1",
        (portfolio_id, asset_id, member_id),
    ).fetchone()
    if duplicate:
        if member["asset_id"] is None:
            external_name = member["external_name"] or ""
            alias_created = False
            if create_alias and external_name:
                alias_created = upsert_alias(conn, asset_id=asset_id, alias_name=external_name, source="mapping_confirmed", notes="Criado ao confirmar mapping") is not None
                rebuild_asset_alias_blob(conn, asset_id)
            current = now_text()
            conn.execute(
                """
                UPDATE portfolio_assets
                SET external_name = ?, nif = ?, sub_account = ?, active = 1,
                    mapping_status = 'manual', mapping_method = 'manual', mapping_confidence = 1,
                    mapped_at = ?, notes = ?, updated_at = ?
                WHERE id = ? AND portfolio_id = ?
                """,
                (
                    external_name,
                    member["nif"] or "",
                    member["sub_account"] or "",
                    current,
                    member["notes"] or "",
                    current,
                    duplicate["id"],
                    portfolio_id,
                ),
            )
            record_mapping_event(conn, portfolio_asset_id=duplicate["id"], external_name=external_name, previous_asset_id=None, selected_asset_id=asset_id, method="manual_merge", confidence=1.0, alias_created=alias_created, notes=f"Merge da entrada pendente #{member_id}")
            conn.execute("DELETE FROM portfolio_assets WHERE id = ? AND portfolio_id = ?", (member_id, portfolio_id))
            normalize_member_order(conn, portfolio_id)
            return
        raise ValueError("member_already_exists")
    previous = member["asset_id"]
    external_name = member["external_name"] or ""
    alias_created = False
    if create_alias and external_name:
        alias_created = upsert_alias(conn, asset_id=asset_id, alias_name=external_name, source="mapping_confirmed", notes="Criado ao confirmar mapping") is not None
        rebuild_asset_alias_blob(conn, asset_id)
    conn.execute(
        """
        UPDATE portfolio_assets
        SET asset_id = ?, mapping_status = 'manual', mapping_method = 'manual', mapping_confidence = 1,
            mapped_at = ?, updated_at = ?
        WHERE id = ? AND portfolio_id = ?
        """,
        (asset_id, now_text(), now_text(), member_id, portfolio_id),
    )
    record_mapping_event(conn, portfolio_asset_id=member_id, external_name=external_name, previous_asset_id=previous, selected_asset_id=asset_id, method="manual", confidence=1.0, alias_created=alias_created)


def unmap_member(conn: sqlite3.Connection, *, member_id: int, portfolio_id: int) -> None:
    member = _require_member(conn, member_id, portfolio_id)
    conn.execute("UPDATE portfolio_assets SET asset_id = NULL, mapping_status = 'mapping_pending', mapping_method = 'unmapped', mapping_confidence = 0, mapped_at = NULL, updated_at = ? WHERE id = ?", (now_text(), member_id))
    record_mapping_event(conn, portfolio_asset_id=member_id, external_name=member["external_name"] or "", previous_asset_id=member["asset_id"], selected_asset_id=None, method="unmapped", confidence=0.0, alias_created=False)


def record_mapping_event(
    conn: sqlite3.Connection,
    *,
    portfolio_asset_id: int,
    external_name: str,
    previous_asset_id: int | None,
    selected_asset_id: int | None,
    method: str,
    confidence: float,
    alias_created: bool,
    notes: str = "",
) -> None:
    conn.execute(
        """
        INSERT INTO portfolio_mapping_events (
            portfolio_asset_id, external_name, previous_asset_id, selected_asset_id,
            method, confidence, alias_created, notes, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (portfolio_asset_id, external_name, previous_asset_id, selected_asset_id, method, confidence, 1 if alias_created else 0, notes, now_text()),
    )


def detect_portfolio_conflicts(conn: sqlite3.Connection, portfolio_id: int | None = None) -> list[dict[str, Any]]:
    conflicts: list[dict[str, Any]] = []
    params: list[Any] = []
    where = ""
    if portfolio_id:
        where = "WHERE portfolio_id = ?"
        params.append(portfolio_id)
    for row in query_all(conn, f"SELECT portfolio_id, sub_account, COUNT(*) AS count FROM portfolio_assets {where} GROUP BY portfolio_id, sub_account HAVING COALESCE(sub_account, '') != '' AND COUNT(*) > 1", params):
        conflicts.append({"code": "duplicate_sub_account", "portfolio_id": row["portfolio_id"], "value": row["sub_account"]})
    for row in query_all(conn, "SELECT normalized_alias, COUNT(DISTINCT asset_id) AS count FROM asset_aliases WHERE active = 1 GROUP BY normalized_alias HAVING count > 1"):
        conflicts.append({"code": "duplicate_alias", "value": row["normalized_alias"]})
    for row in query_all(conn, "SELECT nif, COUNT(DISTINCT id) AS count FROM assets WHERE COALESCE(nif, '') != '' GROUP BY nif HAVING count > 1"):
        conflicts.append({"code": "duplicate_asset_nif", "value": row["nif"]})
    return conflicts


def parse_import_file(filename: str, data: bytes) -> list[dict[str, Any]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        text = data.decode("utf-8-sig", errors="replace")
        return [dict(row) for row in csv.DictReader(io.StringIO(text))]
    if suffix == ".xlsx":
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True, keep_vba=False)
        if {"Portfolios", "Membros", "Aliases"} & set(workbook.sheetnames):
            parsed = []
            for sheet_name in ("Portfolios", "Membros", "Aliases"):
                if sheet_name not in workbook.sheetnames:
                    continue
                sheet = workbook[sheet_name]
                rows = list(sheet.iter_rows(values_only=True))
                if not rows:
                    continue
                headers = [str(value or "").strip() for value in rows[0]]
                for row in rows[1:]:
                    item = {headers[index]: row[index] if index < len(row) else "" for index in range(len(headers))}
                    item["__sheet"] = sheet_name
                    parsed.append(item)
            workbook.close()
            return parsed
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()
        if not rows:
            return []
        headers = [str(value or "").strip() for value in rows[0]]
        parsed: list[dict[str, Any]] = []
        for row in rows[1:]:
            parsed.append({headers[index]: row[index] if index < len(row) else "" for index in range(len(headers))})
        return parsed
    raise ValueError("unsupported_import_format")


def create_import_preview(conn: sqlite3.Connection, *, portfolio_id: int | None, original_filename: str, data: bytes) -> int:
    raw_rows = parse_import_file(original_filename, data)
    rows: list[PortfolioImportRow] = []
    seen_members: set[tuple[str, str, str]] = set()
    for index, raw in enumerate(raw_rows, start=2):
        normalized = _normalize_import_headers(raw)
        errors: list[str] = []
        warnings: list[str] = []
        sheet = str(raw.get("__sheet") or "")
        portfolio = str(normalized.get("portfolio") or "").strip()
        if not portfolio and portfolio_id is None:
            errors.append("missing_portfolio")
        external_name = str(normalized.get("external_name") or normalized.get("asset_name") or "").strip()
        asset_id = _int_or_none(normalized.get("asset_id"))
        if asset_id is not None and conn.execute("SELECT 1 FROM assets WHERE id = ?", (asset_id,)).fetchone() is None:
            errors.append("asset_not_found")
        if not any(str(value or "").strip() for key, value in raw.items() if key != "__sheet"):
            errors.append("empty_row")
        active_raw = normalized.get("active")
        if active_raw not in (None, "") and str(active_raw).strip().lower() not in {"1", "0", "true", "false", "yes", "no", "sim", "nao", "não", "ativo", "inativo", "active", "inactive"}:
            errors.append("invalid_active")
        if asset_id is not None and normalized.get("nif"):
            asset = conn.execute("SELECT nif FROM assets WHERE id = ?", (asset_id,)).fetchone()
            if asset and normalize_nif(asset["nif"]) and normalize_nif(asset["nif"]) != normalize_nif(normalized.get("nif")):
                errors.append("nif_mismatch")
        key = (portfolio, str(normalized.get("sub_account") or "").strip(), str(asset_id or ""))
        if key in seen_members and (key[1] or key[2]):
            errors.append("duplicate_import_member")
        seen_members.add(key)
        if portfolio:
            existing_portfolio = conn.execute("SELECT id FROM portfolio_groups WHERE name = ?", (portfolio,)).fetchone()
            effective_portfolio_id = int(existing_portfolio["id"]) if existing_portfolio else portfolio_id
            if effective_portfolio_id and normalized.get("sub_account"):
                duplicate_sub = conn.execute(
                    "SELECT 1 FROM portfolio_assets WHERE portfolio_id = ? AND sub_account = ? LIMIT 1",
                    (effective_portfolio_id, str(normalized.get("sub_account") or "").strip()),
                ).fetchone()
                if duplicate_sub and sheet != "Membros":
                    warnings.append("sub_account_exists")
            if effective_portfolio_id and asset_id:
                duplicate_member = conn.execute("SELECT 1 FROM portfolio_assets WHERE portfolio_id = ? AND asset_id = ? LIMIT 1", (effective_portfolio_id, asset_id)).fetchone()
                if duplicate_member:
                    warnings.append("member_exists")
        alias_name = str(normalized.get("alias") or raw.get("alias") or "").strip()
        if sheet == "Aliases":
            alias_name = str(raw.get("alias") or raw.get("Alias") or "").strip()
            asset_id = _int_or_none(raw.get("asset_id"))
            portfolio = portfolio or "_aliases"
            external_name = str(raw.get("asset_name") or "").strip()
        if alias_name:
            normalized_alias = normalize_name(alias_name)
            conflict = conn.execute("SELECT asset_id FROM asset_aliases WHERE normalized_alias = ?", (normalized_alias,)).fetchone()
            if conflict and asset_id and int(conflict["asset_id"]) != asset_id:
                errors.append("alias_conflict")
        decision = suggest_mapping(conn, external_name=external_name or str(normalized.get("asset_name") or ""), nif=str(normalized.get("nif") or ""))
        if asset_id is not None:
            decision = MappingDecision(asset_id, "manual", 1.0, "strong", "mapped", (), auto_mappable=False)
        if sheet == "Portfolios":
            action = "portfolio"
        elif sheet == "Aliases":
            action = "alias"
        else:
            action = "error" if errors else ("conflict" if decision.status == "mapping_conflict" else ("new_member" if external_name or asset_id else "no_change"))
        rows.append(
            PortfolioImportRow(
                row_number=index,
                portfolio=portfolio,
                sub_account=str(normalized.get("sub_account") or "").strip(),
                external_name=external_name,
                nif=normalize_nif(normalized.get("nif")),
                asset_name=str(normalized.get("asset_name") or "").strip(),
                asset_id=asset_id or decision.asset_id,
                alias=alias_name,
                notes=str(normalized.get("notes") or "").strip(),
                active=_bool_value(normalized.get("active"), default=True),
                action=action,
                errors=tuple(errors),
                warnings=tuple(sorted({*warnings, *decision.warnings})),
                decision=decision,
            )
        )
    preview = PortfolioImportPreview(
        rows=tuple(rows),
        rows_total=len(rows),
        rows_valid=sum(1 for row in rows if not row.errors),
        rows_pending=sum(1 for row in rows if row.action in {"new_member", "pending"}),
        rows_conflict=sum(1 for row in rows if row.action == "conflict"),
    )
    cursor = conn.execute(
        """
        INSERT INTO portfolio_import_runs (
            portfolio_id, original_filename, status, rows_total, rows_valid, rows_pending,
            rows_conflict, preview_json, created_at
        ) VALUES (?, ?, 'preview', ?, ?, ?, ?, ?, ?)
        """,
        (portfolio_id, original_filename, preview.rows_total, preview.rows_valid, preview.rows_pending, preview.rows_conflict, import_preview_to_json(preview), now_text()),
    )
    return int(cursor.lastrowid)


def get_import_run(conn: sqlite3.Connection, import_id: int) -> sqlite3.Row | None:
    return conn.execute("SELECT * FROM portfolio_import_runs WHERE id = ?", (import_id,)).fetchone()


def apply_import_run(conn: sqlite3.Connection, import_id: int, *, selected_rows: list[int] | None = None, asset_overrides: dict[int, int] | None = None) -> None:
    run = get_import_run(conn, import_id)
    if run is None:
        raise ValueError("import_not_found")
    if run["status"] == "applied":
        raise ValueError("import_already_applied")
    preview = import_preview_from_json(run["preview_json"] or "{}")
    selected = set(selected_rows or [row.row_number for row in preview.rows])
    overrides = asset_overrides or {}
    selected_preview_rows = [row for row in preview.rows if row.row_number in selected]
    for row in selected_preview_rows:
        if row.errors or row.action == "conflict":
            raise ValueError("import_has_errors_or_conflicts")
    for row in selected_preview_rows:
        row_asset_id = overrides.get(row.row_number, row.asset_id)
        portfolio_id = int(run["portfolio_id"] or 0)
        if not portfolio_id:
            portfolio = conn.execute("SELECT id FROM portfolio_groups WHERE name = ?", (row.portfolio,)).fetchone()
            portfolio_id = int(portfolio["id"]) if portfolio else create_portfolio(conn, name=row.portfolio)
        if row.action == "portfolio":
            update_portfolio(conn, portfolio_id=portfolio_id, name=row.portfolio, description="", notes=row.notes)
            continue
        if row.action == "alias":
            if row_asset_id is None:
                raise ValueError("alias_without_asset")
            upsert_alias(conn, asset_id=row_asset_id, alias_name=row.alias, source="portfolio_import", notes=row.notes)
            rebuild_asset_alias_blob(conn, row_asset_id)
            continue
        existing = conn.execute("SELECT id FROM portfolio_assets WHERE portfolio_id = ? AND asset_id = ? AND ? IS NOT NULL", (portfolio_id, row_asset_id, row_asset_id)).fetchone()
        if existing:
            portfolio_asset_id = int(existing["id"])
            conn.execute(
                """
                UPDATE portfolio_assets
                SET external_name = ?, nif = ?, sub_account = ?, active = ?, notes = ?,
                    mapping_status = ?, mapping_method = ?, mapping_confidence = ?, updated_at = ?
                WHERE id = ?
                """,
                (row.external_name or row.asset_name, row.nif, row.sub_account, 1 if row.active else 0, row.notes, "manual" if row_asset_id else "mapping_pending", "manual" if row_asset_id else "unmapped", 1.0 if row_asset_id else 0.0, now_text(), portfolio_asset_id),
            )
            member_id = portfolio_asset_id
        else:
            member_id = add_member(
                conn,
                portfolio_id=portfolio_id,
                asset_id=row_asset_id,
                external_name=row.external_name or row.asset_name,
                nif=row.nif,
                sub_account=row.sub_account,
                notes=row.notes,
                mapping_method="manual" if row_asset_id else "unmapped",
                confidence=1.0 if row_asset_id else 0.0,
            )
        if row.alias and row_asset_id:
            upsert_alias(conn, asset_id=row_asset_id, alias_name=row.alias, source="portfolio_import", notes="Criado por importacao de portfolio")
            rebuild_asset_alias_blob(conn, row_asset_id)
        record_mapping_event(conn, portfolio_asset_id=member_id, external_name=row.external_name, previous_asset_id=None, selected_asset_id=row_asset_id, method="portfolio_import", confidence=1.0 if row_asset_id else 0.0, alias_created=bool(row.alias and row_asset_id))
    conn.execute("UPDATE portfolio_import_runs SET status = 'applied', applied_at = ? WHERE id = ?", (now_text(), import_id))


def export_configuration_workbook(conn: sqlite3.Connection) -> Workbook:
    workbook = Workbook()
    portfolios = workbook.active
    portfolios.title = "Portfolios"
    portfolios.append(["portfolio", "description", "notes", "active", "display_order"])
    for row in list_portfolios(conn, include_archived=True):
        portfolios.append([row["name"], row["description"] or "", row["notes"] or "", row["active"], row["display_order"]])
    members = workbook.create_sheet("Membros")
    members.append(["portfolio", "sub_account", "external_name", "nif", "asset_name", "asset_id", "notes", "active"])
    for row in query_all(
        conn,
        """
        SELECT pg.name AS portfolio, pa.*, a.project_name
        FROM portfolio_assets pa
        JOIN portfolio_groups pg ON pg.id = pa.portfolio_id
        LEFT JOIN assets a ON a.id = pa.asset_id
        ORDER BY pg.display_order, pa.display_order, pa.id
        """,
    ):
        members.append([row["portfolio"], row["sub_account"] or "", row["external_name"] or "", row["nif"] or "", row["project_name"] or "", row["asset_id"] or "", row["notes"] or "", row["active"]])
    aliases = workbook.create_sheet("Aliases")
    aliases.append(["asset_id", "asset_name", "alias", "normalized_alias", "source", "active", "notes"])
    for row in list_aliases(conn):
        aliases.append([row["asset_id"], row["project_name"], row["alias_name"], row["normalized_alias"], row["source"], row["active"], row["notes"] or ""])
    pending = workbook.create_sheet("Pendentes")
    pending.append(["portfolio", "external_name", "nif", "sub_account", "mapping_status"])
    for row in query_all(conn, "SELECT pg.name AS portfolio, pa.* FROM portfolio_assets pa JOIN portfolio_groups pg ON pg.id = pa.portfolio_id WHERE pa.asset_id IS NULL OR pa.mapping_status IN ('mapping_pending', 'mapping_conflict')"):
        pending.append([row["portfolio"], row["external_name"] or "", row["nif"] or "", row["sub_account"] or "", row["mapping_status"] or ""])
    conflicts = workbook.create_sheet("Conflitos")
    conflicts.append(["code", "value", "portfolio_id"])
    for conflict in detect_portfolio_conflicts(conn):
        conflicts.append([conflict.get("code", ""), conflict.get("value", ""), conflict.get("portfolio_id", "")])
    for sheet in workbook.worksheets:
        for column in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[column[0].column_letter].width = min(max(width + 2, 12), 48)
    return workbook


def import_preview_to_json(preview: PortfolioImportPreview) -> str:
    rows: list[dict[str, Any]] = []
    for row in preview.rows:
        rows.append(
            {
                "row_number": row.row_number,
                "portfolio": row.portfolio,
                "sub_account": row.sub_account,
                "external_name": row.external_name,
                "nif": row.nif,
                "asset_name": row.asset_name,
                "asset_id": row.asset_id,
                "alias": row.alias,
                "notes": row.notes,
                "active": row.active,
                "action": row.action,
                "errors": list(row.errors),
                "warnings": list(row.warnings),
                "candidate_asset_id": row.decision.asset_id if row.decision else None,
                "candidate_method": row.decision.method if row.decision else "",
                "candidate_score": row.decision.score if row.decision else 0,
                "candidate_confidence": row.decision.confidence if row.decision else "",
            }
        )
    return json.dumps({"rows": rows}, ensure_ascii=True)


def import_preview_from_json(payload: str) -> PortfolioImportPreview:
    data = json.loads(payload or "{}")
    rows = tuple(
        PortfolioImportRow(
            row_number=int(item.get("row_number") or 0),
            portfolio=str(item.get("portfolio") or ""),
            sub_account=str(item.get("sub_account") or ""),
            external_name=str(item.get("external_name") or ""),
            nif=str(item.get("nif") or ""),
            asset_name=str(item.get("asset_name") or ""),
            asset_id=_int_or_none(item.get("asset_id")),
            alias=str(item.get("alias") or ""),
            notes=str(item.get("notes") or ""),
            active=bool(item.get("active", True)),
            action=str(item.get("action") or "pending"),
            errors=tuple(item.get("errors") or ()),
            warnings=tuple(item.get("warnings") or ()),
        )
        for item in data.get("rows", [])
    )
    return PortfolioImportPreview(
        rows=rows,
        rows_total=len(rows),
        rows_valid=sum(1 for row in rows if not row.errors),
        rows_pending=sum(1 for row in rows if row.action in {"new_member", "pending"}),
        rows_conflict=sum(1 for row in rows if row.action == "conflict"),
    )


def _normalize_import_headers(row: dict[str, Any]) -> dict[str, Any]:
    aliases = {
        "portfolio": {"portfolio", "portefolio", "grupo"},
        "sub_account": {"sub_account", "subconta", "sub conta"},
        "external_name": {"external_name", "nome externo", "instalacao externa", "instalação externa"},
        "nif": {"nif", "vat"},
        "asset_name": {"asset_name", "nome asset", "instalacao", "instalação"},
        "asset_id": {"asset_id", "id asset", "id"},
        "alias": {"alias", "nome alternativo"},
        "notes": {"notes", "notas"},
        "active": {"active", "ativo", "activa"},
    }
    result: dict[str, Any] = {}
    for header, value in row.items():
        normalized = normalize_name(header)
        for target, names in aliases.items():
            if normalized in {normalize_name(name) for name in names}:
                result[target] = value
                break
    return result


def _bool_value(value: Any, *, default: bool) -> bool:
    if value in (None, ""):
        return default
    return str(value).strip().lower() in {"1", "true", "yes", "sim", "ativo", "active"}


def _int_or_none(value: Any) -> int | None:
    raw = str(value or "").strip()
    return int(raw) if raw.isdigit() else None


def _validated_unique_ids(ids: list[int]) -> list[int]:
    cleaned = [int(item) for item in ids if int(item) > 0]
    if len(cleaned) != len(set(cleaned)):
        raise ValueError("duplicate_ids")
    if not cleaned:
        raise ValueError("empty_ids")
    return cleaned


def _require_portfolio(conn: sqlite3.Connection, portfolio_id: int, *, active_only: bool = False) -> sqlite3.Row:
    row = get_portfolio(conn, portfolio_id)
    if row is None:
        raise ValueError("portfolio_not_found")
    if active_only and (not row["active"] or row["archived_at"]):
        raise ValueError("portfolio_archived")
    return row


def _require_asset(conn: sqlite3.Connection, asset_id: int) -> sqlite3.Row:
    row = conn.execute("SELECT * FROM assets WHERE id = ?", (asset_id,)).fetchone()
    if row is None:
        raise ValueError("asset_not_found")
    return row


def _require_member(conn: sqlite3.Connection, member_id: int, portfolio_id: int) -> sqlite3.Row:
    row = conn.execute("SELECT * FROM portfolio_assets WHERE id = ? AND portfolio_id = ?", (member_id, portfolio_id)).fetchone()
    if row is None:
        raise ValueError("member_not_found")
    return row


def _require_alias(conn: sqlite3.Connection, asset_id: int, alias_id: int) -> sqlite3.Row:
    row = conn.execute("SELECT * FROM asset_aliases WHERE id = ? AND asset_id = ?", (alias_id, asset_id)).fetchone()
    if row is None:
        raise ValueError("alias_not_found")
    return row
