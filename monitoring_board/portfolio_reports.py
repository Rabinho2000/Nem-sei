from __future__ import annotations

import json
import re
import sqlite3
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from monitoring_board.portfolio_repository import (
    auto_map_portfolio_assets as repository_auto_map_portfolio_assets,
    suggest_mapping,
)
from monitoring_board.reporting.availability import calculate_weighted_portfolio_availability
from monitoring_board.reporting.billing import calculate_billing, decimal_from_value, detect_report_type_value
from monitoring_board.reporting.degradation import calculate_degradation_factor
from monitoring_board.reporting.models import EnergyBreakdown
from monitoring_board.reporting.periods import month_bounds
from monitoring_board.reporting.repositories import (
    detect_tariff_validity_warnings,
    get_asset_billing_config,
    get_latest_helioscope_expected,
    get_latest_tariff,
    get_daily_production_totals,
    get_monthly_availability,
    get_monthly_production_record,
    has_expired_tariff,
    list_hourly_production_records,
    list_portfolio_report_assets,
    list_tariffs_intersecting_period,
    list_tariff_period_rules,
    row_to_hourly_energy_record,
    row_to_tariff_config,
    save_asset_tariff,
    upsert_asset_billing_config,
)
from monitoring_board.services.financial_models import get_active_expected_for_month
from monitoring_board.reporting.tariffs import (
    classify_tariff_period as tariff_classify_tariff_period,
    result_to_legacy_dict,
    time_in_rule as tariff_time_in_rule,
    value_tariff_energy,
)


MONTH_LABELS = {
    1: ("jan", "janeiro", "january"),
    2: ("fev", "fevereiro", "feb", "february"),
    3: ("mar", "marco", "março", "march"),
    4: ("abr", "abril", "apr", "april"),
    5: ("mai", "maio", "may"),
    6: ("jun", "junho", "june"),
    7: ("jul", "julho", "july"),
    8: ("ago", "agosto", "aug", "august"),
    9: ("set", "setembro", "sep", "september"),
    10: ("out", "outubro", "oct", "october"),
    11: ("nov", "novembro", "november"),
    12: ("dez", "dezembro", "dec", "december"),
}
PERIOD_NAMES = ("ponta", "cheia", "vazio", "super_vazio")
FINANCIAL_SOURCE_TYPE = "financial_model"
WARNING_LABELS = {
    "ok": "OK",
    "missing_monthly_production": "Sem producao real FusionSolar",
    "missing_hourly_production": "Sem producao horaria FusionSolar",
    "missing_hourly_self_use": "Sem autoconsumo horario",
    "missing_helioscope_expected": "Sem Helioscope",
    "missing_financial_model": "Sem modelo financeiro",
    "ambiguous_financial_model": "Modelo financeiro ambiguo",
    "financial_model_missing_year": "Modelo sem ano base",
    "financial_model_missing_nif": "Modelo sem NIF",
    "financial_model_nif_mismatch": "NIF do modelo diferente",
    "financial_model_name_mismatch": "Nome do modelo diferente",
    "financial_model_kwp_mismatch": "Potencia do modelo diferente",
    "financial_model_missing_month": "Modelo com meses em falta",
    "financial_model_missing_cached_formula": "Formula sem valor guardado",
    "financial_model_unknown_unit": "Unidade desconhecida",
    "financial_model_calculated_export": "Excedente previsto calculado",
    "financial_model_calculated_grid_import": "Importacao prevista calculada",
    "missing_mounting_date": "Sem data de montagem",
    "invalid_mounting_date": "Sem data de montagem",
    "missing_tariff": "Sem tarifa",
    "missing_tariff_rules": "Sem tarifa",
    "missing_simple_tariff_price": "Sem tarifa",
    "missing_invoice": "Sem fatura",
    "review_required": "Fatura por rever",
    "extraction_failed": "Extracao falhou",
    "incompatible_invoice": "Fatura nao compativel",
    "missing_availability": "Sem availability",
    "mapping_pending": "Mapping pendente",
    "mapping_conflict": "Mapping conflito",
    "expired_tariff": "Tarifa expirada",
    "missing_installed_power": "Sem potencia instalada",
    "unclassified_hourly_production": "Dados incompletos",
    "unclassified_hourly_energy": "Dados incompletos",
    "inferred_hourly_self_use": "Autoconsumo inferido",
    "incomplete_tariff_coverage": "Cobertura tarifaria incompleta",
    "tariff_validity_gap": "Lacuna tarifaria",
    "overlapping_tariffs": "Tarifas sobrepostas",
}

PORTFOLIO_EXTERNAL_ROWS = {
    "Solcorelios I": [
        ("001", "500001022", "A COLMEIA DO MINHO SA"),
        ("002", "510912974", "FLORINEVE - PRODUCAO E COMERCIO DE FLORES LDA"),
        ("003", "510731570", "DIALOGOS DO BOSQUE UNIPESSOAL LDA"),
        ("004", "502265906", "A PIRES LOURENCO E FILHOS SA"),
        ("005", "505435748", "USINAGE MAQUINACAO E PORTA MOLDES LDA"),
        ("006", "500792640", "FUNDACAO ABEL E JOAO DE LACERDA"),
        ("007", "501754679", "MARMORES GRANJA LDA"),
        ("008", "514968044", "AH PINHAL DO REI LDA"),
        ("009", "503722170", "LUSOBATATA PRODUTOS IV GAMA LDA"),
        ("010", "505030896", "SOLIDUS - SOLUCOES PARA FERRAMENTAS E MOLDES LD"),
        ("011", "503995762", "MARMO J - EXPORTACAO, IMPORTACAO, MARMORES UNIPESSOAL LDA"),
        ("012", "501782265", "ROUFIMAR INDUSTRIA DE MARMORES, S.A."),
        ("013", "500751722", "VIARCO INDUSTRIA LAPIS LDA"),
        ("014", "513203524", "GADELHO DE CASTRO S A"),
        ("015", "502480335", "AMADOIS CALCADO LDA"),
        ("016", "500269203", "SOC TRANSPORTES POIARENSE LDA"),
        ("017", "501150617", "GRANETOS-MARMORES E GRANITOS,S.A."),
        ("018", "502882514", "MULTIAVES AVICOLA INTERNACIONAL LDA"),
        ("019", "513018590", "INTERVEDROS - SUPERMERCADOS LDA"),
        ("020", "501258060", "SERRALHARIA VIEIRA LDA"),
        ("021", "502786078", "COLEGIO DE NOSSA SENHORA DA APRESENTACAO"),
        ("022", "507685733", "RAS RECYCLING, Lda."),
        ("023", "501277676", "SICOBRITA - EXTRACCAO E BRITAGEM DE PEDRA S.A"),
        ("024", "501416382", "COOPERATIVA AGRO-PECUARIA MIRANDESA CRL"),
        ("025", "500876746", "FERRAZ E FERREIRA LDA"),
        ("026", "502605545", "ARDAVI MAQUINACAO E COMERCIALIZACAO ARTEFACTOS METALICOS LDA"),
        ("027", "502247185", "GRANITOS GALRAO NORTE LDA"),
        ("028", "503194387", "JETESETECAR EQUIPAMENTOS AUTO LDA"),
        ("029", "502099666", "DIAMANTINO COELHO FILHO SA"),
        ("030", "513239731", "PH Energia, Lda (Prod)"),
        ("031", "515346306", "SOLCORACTION, LDA"),
        ("032", "980763703", "Simples Energia de Espana, SL-Sucursal de Portugal"),
        ("033", "506178374", "RAP - INDUSTRIAL LDA"),
        ("034", "509090990", "LICOFRUTOS UNIPESSOAL LDA"),
        ("035", "510504965", "HOMEUPDATE LDA"),
        ("036", "501137092", "ASSOC HUMANITARIA DOS BOMBEIROS VOLUNTARIOS ARRUDA VINHOS"),
        ("037", "500280614", "TECNISATA INDUSTRIA METALOMECANICA SA"),
        ("038", "501493603", "ITECMO INDUSTRIA FABRICACAO MOLDES LDA"),
        ("039", "502930942", "ENCONTRUS SOC HOTELEIRA LDA"),
        ("040", "999999990", "Consumidor Final"),
        ("041", "501084819", "Real Sport Clube"),
        ("042", "510352855", "MARYASA - IMPORTACAO E EXPORTACAO UNIPESSOAL LDA"),
        ("043", "502680296", "PANCRISP-INDUSTRIA DE PANIFICACAO LDA"),
        ("044", "500091161", "MARMORES GALRAO - EDUARDO GALRAO JORGE & FILHOS S A"),
        ("045", "516250779", "RUMOS VIRTUOSOS, LDA"),
        ("046", "517295890", "SOLCORELIOS II, UNIPESSOAL, LDA"),
        ("047", "510091490", "Negocios.Doc - Produtos e Servicos de Gestao Integrada, Lda"),
    ],
    "Solcorelios II": [
        ("001", "", "Subconta 001 - importar mais tarde"),
        ("002", "", "Subconta 002 - importar mais tarde"),
        ("003", "", "Subconta 003 - importar mais tarde"),
        ("004", "", "Subconta 004 - importar mais tarde"),
        ("005", "", "Subconta 005 - importar mais tarde"),
        ("006", "502312254", "Casa da Divina Providencia e de Maria Auxiliadora"),
        ("007", "504646788", "APPACDM DE LISBOA - ASSOCIACAO PORTUGUESA DE PAIS"),
        ("008", "500878684", "ASESM - Associacao de Solidariedade e Educacao de Salir de ..."),
        ("009", "600083780", "Agrupamento de Escolas da Boa Agua"),
        ("010", "503780774", "O CASARAO HOTELARIA E TURISMO LDA"),
        ("011", "501426892", "Fundacao Irene Rolo"),
        ("012", "501130179", "Associacao Humanitaria de Bombeiro Voluntarios de Alcoentre"),
        ("013", "501379550", "FUTEBOL CLUBE DE ALVERCA"),
        ("014", "513983511", "EZU Energia, Lda"),
        ("015", "501512071", "ALVARSOL SOC LAVANDARIAS ALGARVE LDA"),
        ("016", "508619041", "NEUTRIPURO - LAVAGENS INDUSTRIAIS , LDA"),
        ("017", "508776597", "Antonio Evaristo Goncalves - Sociedade Agricola, Unipessoal ..."),
        ("018", "502962801", "LAREIRAS SOUSA LDA"),
        ("019", "501440623", "Associacao Humanitaria de Bombeiro Voluntarios de Montela..."),
        ("020", "501358331", "SANTOS E FERREIRA LDA"),
        ("021", "508546192", "LUIS BAPTISTA GONCALVES - SOCIEDADE AGRICOLA UNIPE..."),
        ("022", "501237089", "ASSOCIACAO HUMANITARIA DE BOMBEIROS DO PINHAL NO..."),
        ("023", "502377380", "Centro de Solidariedade Social Padre Jose Filipe Rodrigues"),
        ("024", "500827540", "MANUEL BARBOSA E FILHOS LDA"),
        ("025", "501137092", "ASSOC HUMANITARIA DOS BOMBEIROS VOLUNTARIOS ARR..."),
        ("026", "500877386", "Legado do Caixeiro Alentejano - Associacao Mutualista"),
        ("027", "501241230", "ASSOCIACAO HUMANITARIA DE BOMBEIROS VOLUNTARIOS ..."),
        ("028", "503194387", "JETESETECAR EQUIPAMENTOS AUTO LDA"),
        ("029", "509090990", "LICOFRUTOS UNIPESSOAL LDA"),
        ("030", "510091490", "Negocios.Doc - Produtos e Servicos de Gestao Integrada, Lda"),
        ("031", "501131981", "ASSOCIACAO HUMANITARIA DOS BOMBEIROS VOLUNTARIO..."),
        ("032", "500731179", "PROVINCIA PORTUGUESA INSTITUTO IRMAS SANTA DOROT..."),
    ],
}


def normalize_text(value: Any) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    ascii_value = "".join(char for char in normalized if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", ascii_value.strip().lower())


def normalize_nif(value: Any) -> str:
    return re.sub(r"\D+", "", str(value or ""))


def parse_month(value: Any) -> int | None:
    if isinstance(value, datetime):
        return value.month
    if isinstance(value, date):
        return value.month
    raw = normalize_text(value)
    if not raw:
        return None
    if raw.isdigit() and 1 <= int(raw) <= 12:
        return int(raw)
    for month, labels in MONTH_LABELS.items():
        if raw in labels:
            return month
    return None


def parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    raw = str(value).strip().replace(" ", "").replace(",", ".")
    raw = re.sub(r"[^0-9.\-]", "", raw)
    if raw in {"", "-", "."}:
        return None
    try:
        parsed = float(raw)
    except ValueError:
        return None
    return parsed if parsed >= 0 else None


@dataclass(frozen=True)
class FinancialTariffImport:
    tariff_type: str
    cycle_type: str
    simple_price_eur_kwh: float | None
    ponta_price_eur_kwh: float | None
    cheia_price_eur_kwh: float | None
    vazio_price_eur_kwh: float | None
    super_vazio_price_eur_kwh: float | None
    export_price_eur_kwh: float | None
    valid_from: str
    valid_to: str
    solcor_price_eur_kwh: float | None
    fixed_monthly_fee_eur: float | None


@dataclass(frozen=True)
class FinancialIntervalExpected:
    period_start: datetime
    period_end: datetime
    expected_kwh: float


@dataclass(frozen=True)
class FinancialModelImport:
    project_name: str
    installed_power_kwp: float | None
    monthly_expected: dict[int, float]
    interval_expected: tuple[FinancialIntervalExpected, ...]
    tariff: FinancialTariffImport | None
    financial_outputs: dict[str, Any]
    warnings: tuple[str, ...] = ()


def parse_helioscope_monthly_expected(path: Path) -> dict[int, float]:
    workbook = load_workbook(path, data_only=True, keep_links=False)
    try:
        candidates: list[dict[int, float]] = []
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            for row_index, row in enumerate(rows):
                month_columns: dict[int, int] = {}
                for col_index, value in enumerate(row):
                    month = parse_month(value)
                    if month:
                        month_columns[month] = col_index
                if len(month_columns) < 10:
                    continue
                for values_row in rows[row_index + 1 : min(row_index + 8, len(rows))]:
                    parsed: dict[int, float] = {}
                    for month, col_index in month_columns.items():
                        value = parse_float(values_row[col_index] if col_index < len(values_row) else None)
                        if value is not None:
                            parsed[month] = value
                    if len(parsed) >= 10:
                        candidates.append(parsed)
        if not candidates:
            raise ValueError("Nao foi possivel identificar valores mensais no ficheiro Helioscope.")
        complete_candidates = [candidate for candidate in candidates if len(candidate) == 12]
        if len(complete_candidates) > 1:
            unique_candidates = {tuple(round(candidate[month], 6) for month in range(1, 13)) for candidate in complete_candidates}
            if len(unique_candidates) > 1:
                raise ValueError("O ficheiro Helioscope contem varias series mensais plausiveis.")
        best = max(candidates, key=len)
        if len(best) != 12:
            raise ValueError("O ficheiro Helioscope nao contem 12 valores mensais confiaveis.")
        return {month: float(best[month]) for month in range(1, 13)}
    finally:
        workbook.close()


def parse_financial_model_file(path: Path) -> FinancialModelImport:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        warnings: list[str] = []
        monthly = _parse_financial_monthly_expected(workbook) or _parse_upac_monthly_expected(workbook)
        if not monthly:
            monthly = parse_helioscope_monthly_expected(path)
            warnings.append("financial_monthly_fallback")
        intervals = _parse_financial_interval_expected(workbook)
        if not intervals:
            warnings.append("missing_financial_interval_expected")
        tariff = _parse_financial_tariff(workbook) or _parse_upac_tariff(workbook)
        if tariff is None:
            warnings.append("missing_financial_tariff")
        outputs = _parse_financial_outputs(workbook) or _parse_upac_financial_outputs(workbook)
        project_name = str(_cell_value(workbook, "Projeto", "C5") or _cell_value(workbook, "UPAC", "A4") or "").strip()
        return FinancialModelImport(
            project_name=project_name,
            installed_power_kwp=parse_float(_cell_value(workbook, "Projeto", "H8")) or parse_float(_cell_value(workbook, "UPAC", "D4")),
            monthly_expected=monthly,
            interval_expected=tuple(intervals),
            tariff=tariff,
            financial_outputs=outputs,
            warnings=tuple(warnings),
        )
    finally:
        workbook.close()


def _cell_value(workbook: Any, sheet_name: str, address: str) -> Any:
    if sheet_name not in workbook.sheetnames:
        return None
    return workbook[sheet_name][address].value


def _parse_financial_monthly_expected(workbook: Any) -> dict[int, float]:
    if "Projeto" not in workbook.sheetnames:
        return {}
    sheet = workbook["Projeto"]
    parsed: dict[int, float] = {}
    for month_cell, expected_cell in sheet.iter_rows(min_row=6, max_row=17, min_col=10, max_col=11, values_only=True):
        month = parse_month(month_cell)
        expected = parse_float(expected_cell)
        if month and expected is not None:
            parsed[month] = expected
    return parsed if len(parsed) == 12 else {}


def _parse_upac_monthly_expected(workbook: Any) -> dict[int, float]:
    if "Prod month" not in workbook.sheetnames:
        return {}
    sheet = workbook["Prod month"]
    parsed: dict[int, float] = {}
    for row in sheet.iter_rows(min_row=1, max_row=24, values_only=True):
        month = parse_month(row[0] if row else None)
        expected = parse_float(row[2] if len(row) > 2 else None)
        if month and expected is not None:
            parsed[month] = expected
    return parsed if len(parsed) == 12 else {}


def _parse_financial_interval_expected(workbook: Any) -> list[FinancialIntervalExpected]:
    if "Helio&Cons" not in workbook.sheetnames:
        return []
    sheet = workbook["Helio&Cons"]
    records: list[FinancialIntervalExpected] = []
    for row in sheet.iter_rows(min_row=2, min_col=12, max_col=14, values_only=True):
        raw_expected = row[0]
        raw_start = row[2]
        expected_wh = parse_float(raw_expected)
        started_at = _parse_excel_datetime(raw_start)
        if expected_wh is None or started_at is None:
            continue
        records.append(
            FinancialIntervalExpected(
                period_start=started_at,
                period_end=started_at + timedelta(minutes=15),
                expected_kwh=expected_wh / 1000,
            )
        )
    return records


def _parse_excel_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if value is None or value == "":
        return None
    try:
        return datetime.fromisoformat(str(value).strip())
    except ValueError:
        return None


def _parse_financial_tariff(workbook: Any) -> FinancialTariffImport | None:
    if "Projeto" not in workbook.sheetnames:
        return None
    sheet = workbook["Projeto"]
    tariff_type = _map_financial_tariff_type(sheet["C44"].value)
    if tariff_type is None:
        return None
    totals = {
        "super_vazio": parse_float(sheet["H41"].value),
        "vazio": parse_float(sheet["H42"].value),
        "cheia": parse_float(sheet["H43"].value),
        "ponta": parse_float(sheet["H44"].value),
    }
    simple_price = totals["super_vazio"] if tariff_type == "simple" else None
    return FinancialTariffImport(
        tariff_type=tariff_type,
        cycle_type=_map_financial_cycle_type(sheet["C45"].value),
        simple_price_eur_kwh=simple_price,
        ponta_price_eur_kwh=totals["ponta"] if tariff_type in {"tri-hourly", "tetra-hourly"} else None,
        cheia_price_eur_kwh=totals["cheia"] if tariff_type in {"bi-hourly", "tri-hourly", "tetra-hourly"} else None,
        vazio_price_eur_kwh=totals["vazio"] if tariff_type in {"bi-hourly", "tri-hourly", "tetra-hourly"} else None,
        super_vazio_price_eur_kwh=totals["super_vazio"] if tariff_type == "tetra-hourly" else None,
        export_price_eur_kwh=parse_float(sheet["F46"].value),
        valid_from=_date_iso(sheet["C41"].value),
        valid_to=_date_iso(sheet["C42"].value),
        solcor_price_eur_kwh=parse_float(sheet["L33"].value),
        fixed_monthly_fee_eur=parse_float(sheet["M33"].value),
    )


def _parse_upac_tariff(workbook: Any) -> FinancialTariffImport | None:
    if "UPAC" not in workbook.sheetnames:
        return None
    sheet = workbook["UPAC"]
    prices = {
        "super_vazio": parse_float(sheet["H13"].value),
        "vazio": parse_float(sheet["H14"].value),
        "cheia": parse_float(sheet["H15"].value),
        "ponta": parse_float(sheet["H16"].value),
    }
    if not any(value is not None for value in prices.values()):
        return None
    if prices["super_vazio"] is not None:
        tariff_type = "tetra-hourly"
    elif prices["ponta"] is not None:
        tariff_type = "tri-hourly"
    elif prices["cheia"] is not None and prices["vazio"] is not None:
        tariff_type = "bi-hourly"
    else:
        tariff_type = "simple"
    return FinancialTariffImport(
        tariff_type=tariff_type,
        cycle_type=_map_financial_cycle_type(sheet["H12"].value),
        simple_price_eur_kwh=prices["vazio"] if tariff_type == "simple" else None,
        ponta_price_eur_kwh=prices["ponta"] if tariff_type in {"tri-hourly", "tetra-hourly"} else None,
        cheia_price_eur_kwh=prices["cheia"] if tariff_type in {"bi-hourly", "tri-hourly", "tetra-hourly"} else None,
        vazio_price_eur_kwh=prices["vazio"] if tariff_type in {"bi-hourly", "tri-hourly", "tetra-hourly"} else None,
        super_vazio_price_eur_kwh=prices["super_vazio"] if tariff_type == "tetra-hourly" else None,
        export_price_eur_kwh=parse_float(sheet["H19"].value),
        valid_from="",
        valid_to="",
        solcor_price_eur_kwh=parse_float(sheet["K15"].value),
        fixed_monthly_fee_eur=None,
    )


def _parse_financial_outputs(workbook: Any) -> dict[str, Any]:
    if "Projeto" not in workbook.sheetnames:
        return {}
    sheet = workbook["Projeto"]
    return {
        "sale_price_eur": parse_float(sheet["P28"].value),
        "first_year_net_benefit_eur": parse_float(sheet["L28"].value),
        "first_year_global_benefit_eur": parse_float(sheet["P29"].value),
        "global_roi_years": parse_float(sheet["P30"].value),
        "profit_margin_pct": parse_float(sheet["P32"].value),
        "avoided_tariff_eur_kwh": parse_float(sheet["L32"].value),
        "ppa_tariff_eur_kwh": parse_float(sheet["L33"].value),
        "monthly_payment_eur": parse_float(sheet["M33"].value),
    }


def _parse_upac_financial_outputs(workbook: Any) -> dict[str, Any]:
    if "UPAC" not in workbook.sheetnames:
        return {}
    sheet = workbook["UPAC"]
    return {
        "sale_price_eur": parse_float(sheet["D6"].value),
        "first_year_net_benefit_eur": parse_float(sheet["N4"].value),
        "first_year_global_benefit_eur": parse_float(sheet["N6"].value),
        "client_npv_eur": parse_float(sheet["N9"].value),
        "avoided_tariff_eur_kwh": parse_float(sheet["K14"].value),
        "ppa_tariff_eur_kwh": parse_float(sheet["K15"].value),
        "annual_pv_production_kwh": parse_float(sheet["K5"].value),
        "annual_self_consumption_kwh": parse_float(sheet["K7"].value),
        "annual_export_kwh": parse_float(sheet["K8"].value),
    }


def _map_financial_tariff_type(value: Any) -> str | None:
    normalized = normalize_text(value)
    if "tetra" in normalized:
        return "tetra-hourly"
    if "tri" in normalized:
        return "tri-hourly"
    if "bi" in normalized:
        return "bi-hourly"
    if "simples" in normalized or "simples" in normalized or "simple" in normalized:
        return "simple"
    return None


def _map_financial_cycle_type(value: Any) -> str:
    normalized = normalize_text(value)
    if "diario" in normalized:
        return "daily"
    if "semanal" in normalized:
        return "weekly"
    return ""


def _date_iso(value: Any) -> str:
    parsed = _parse_excel_datetime(value)
    return parsed.date().isoformat() if parsed else ""


def store_source_file(
    conn: sqlite3.Connection,
    *,
    upload_dir: Path,
    file_storage: Any,
    asset_id: int,
    portfolio_id: int | None,
    file_type: str,
    notes: str = "",
) -> int:
    original = Path(file_storage.filename or f"{file_type}.bin").name
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    target_dir = upload_dir / "portfolio_sources" / str(asset_id)
    target_dir.mkdir(parents=True, exist_ok=True)
    stored_path = target_dir / f"{timestamp}_{original}"
    file_storage.save(stored_path)
    cursor = conn.execute(
        """
        INSERT INTO source_files (asset_id, portfolio_id, file_type, original_filename, stored_path, uploaded_at, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (asset_id, portfolio_id, file_type, original, str(stored_path), datetime.now().isoformat(timespec="seconds"), notes),
    )
    return int(cursor.lastrowid)


def import_helioscope_file(
    conn: sqlite3.Connection,
    *,
    upload_dir: Path,
    file_storage: Any,
    asset_id: int,
    portfolio_id: int | None,
    base_year: int | None = None,
) -> dict[str, Any]:
    source_id = store_source_file(
        conn,
        upload_dir=upload_dir,
        file_storage=file_storage,
        asset_id=asset_id,
        portfolio_id=portfolio_id,
        file_type="helioscope",
    )
    stored = conn.execute("SELECT stored_path FROM source_files WHERE id = ?", (source_id,)).fetchone()
    monthly = parse_helioscope_monthly_expected(Path(stored["stored_path"]))
    year = base_year or date.today().year
    conn.execute("DELETE FROM helioscope_expected_production WHERE asset_id = ? AND source_file_id != ?", (asset_id, source_id))
    for month, expected_kwh in monthly.items():
        conn.execute(
            """
            INSERT INTO helioscope_expected_production (asset_id, source_file_id, base_year, month, expected_kwh, imported_at, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (asset_id, source_id, year, month, expected_kwh, datetime.now().isoformat(timespec="seconds"), ""),
        )
    return {"source_file_id": source_id, "months": len(monthly)}


def import_financial_model_file(
    conn: sqlite3.Connection,
    *,
    upload_dir: Path,
    file_storage: Any,
    asset_id: int,
    portfolio_id: int | None,
    base_year: int | None = None,
) -> dict[str, Any]:
    source_id = store_source_file(
        conn,
        upload_dir=upload_dir,
        file_storage=file_storage,
        asset_id=asset_id,
        portfolio_id=portfolio_id,
        file_type=FINANCIAL_SOURCE_TYPE,
    )
    stored = conn.execute("SELECT stored_path FROM source_files WHERE id = ?", (source_id,)).fetchone()
    parsed = parse_financial_model_file(Path(stored["stored_path"]))
    year = base_year or date.today().year
    imported_at = datetime.now().isoformat(timespec="seconds")
    conn.execute("DELETE FROM helioscope_expected_production WHERE asset_id = ? AND source_file_id != ?", (asset_id, source_id))
    for month, expected_kwh in parsed.monthly_expected.items():
        conn.execute(
            """
            INSERT INTO helioscope_expected_production (asset_id, source_file_id, base_year, month, expected_kwh, imported_at, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (asset_id, source_id, year, month, expected_kwh, imported_at, "financial_model"),
        )
    conn.execute("DELETE FROM helioscope_expected_interval_production WHERE asset_id = ?", (asset_id,))
    conn.executemany(
        """
        INSERT INTO helioscope_expected_interval_production (
            asset_id, source_file_id, period_start, period_end, expected_kwh, imported_at, notes
        ) VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (
                asset_id,
                source_id,
                item.period_start.isoformat(timespec="seconds"),
                item.period_end.isoformat(timespec="seconds"),
                item.expected_kwh,
                imported_at,
                "financial_model",
            )
            for item in parsed.interval_expected
        ],
    )
    tariff_action = "none"
    tariff_id = None
    if parsed.tariff is not None:
        tariff_id, tariff_action = _apply_financial_tariff(conn, asset_id, parsed.tariff)
        report_type = detect_report_type_value(conn.execute("SELECT * FROM assets WHERE id = ?", (asset_id,)).fetchone() or {})
        upsert_asset_billing_config(
            conn,
            asset_id=asset_id,
            config=_financial_billing_config(parsed.tariff, report_type),
        )
    summary = {
        "source_file_id": source_id,
        "project_name": parsed.project_name,
        "installed_power_kwp": parsed.installed_power_kwp,
        "months": len(parsed.monthly_expected),
        "intervals_15m": len(parsed.interval_expected),
        "tariff_id": tariff_id,
        "tariff_action": tariff_action,
        "financial_outputs": parsed.financial_outputs,
        "warnings": list(parsed.warnings),
    }
    conn.execute("UPDATE source_files SET notes = ? WHERE id = ?", (json.dumps(summary, ensure_ascii=True), source_id))
    return summary


def _apply_financial_tariff(conn: sqlite3.Connection, asset_id: int, tariff: FinancialTariffImport) -> tuple[int, str]:
    start = datetime.fromisoformat(tariff.valid_from).date() if tariff.valid_from else date.today()
    end = datetime.fromisoformat(tariff.valid_to).date() if tariff.valid_to else start
    existing = list_tariffs_intersecting_period(conn, asset_id=asset_id, start=start, end=end)
    tariff_id = int(existing[0]["id"]) if len(existing) == 1 else None
    action = "updated" if tariff_id is not None else "created"
    if len(existing) > 1:
        return int(existing[0]["id"]), "skipped_multiple_overlaps"
    saved = save_asset_tariff(
        conn,
        tariff_id=tariff_id,
        asset_id=asset_id,
        tariff_type=tariff.tariff_type,
        cycle_type=tariff.cycle_type,
        simple_price_eur_kwh=tariff.simple_price_eur_kwh,
        ponta_price_eur_kwh=tariff.ponta_price_eur_kwh,
        cheia_price_eur_kwh=tariff.cheia_price_eur_kwh,
        vazio_price_eur_kwh=tariff.vazio_price_eur_kwh,
        super_vazio_price_eur_kwh=tariff.super_vazio_price_eur_kwh,
        valid_from=tariff.valid_from,
        valid_to=tariff.valid_to,
        notes="Importado do modelo financeiro",
    )
    return saved, action


def _financial_billing_config(tariff: FinancialTariffImport, report_type: Any) -> Any:
    from monitoring_board.reporting.models import BillingConfig

    return BillingConfig(
        report_type=report_type,
        solcor_price_per_kwh=decimal_from_value(tariff.solcor_price_eur_kwh),
        fixed_monthly_fee_eur=decimal_from_value(tariff.fixed_monthly_fee_eur),
        electricity_price_eur_kwh=decimal_from_value(tariff.simple_price_eur_kwh or tariff.cheia_price_eur_kwh or 0),
        export_price_eur_kwh=decimal_from_value(tariff.export_price_eur_kwh),
    )


def map_external_portfolio_entity(conn: sqlite3.Connection, *, nif: str = "", external_name: str = "") -> dict[str, Any]:
    decision = suggest_mapping(conn, nif=nif, external_name=external_name)
    return {
        "asset_id": decision.asset_id if decision.auto_mappable else None,
        "mapping_status": decision.status if decision.auto_mappable else ("mapping_conflict" if decision.status == "mapping_conflict" else "unmapped"),
        "mapping_confidence": decision.score,
    }


def auto_map_portfolio_assets(conn: sqlite3.Connection, portfolio_id: int | None = None) -> dict[str, int]:
    return repository_auto_map_portfolio_assets(conn, portfolio_id)


def seed_external_portfolio_rows(conn: sqlite3.Connection) -> None:
    for portfolio_name, rows in PORTFOLIO_EXTERNAL_ROWS.items():
        group = conn.execute("SELECT id FROM portfolio_groups WHERE name = ?", (portfolio_name,)).fetchone()
        if group is None:
            continue
        portfolio_id = int(group["id"])
        for index, (sub_account, nif, external_name) in enumerate(rows, start=1):
            existing = conn.execute(
                """
                SELECT id
                FROM portfolio_assets
                WHERE portfolio_id = ? AND sub_account = ?
                LIMIT 1
                """,
                (portfolio_id, sub_account),
            ).fetchone()
            if existing:
                conn.execute(
                    """
                    UPDATE portfolio_assets
                    SET external_name = COALESCE(NULLIF(external_name, ''), ?),
                        nif = COALESCE(NULLIF(nif, ''), ?),
                        display_order = COALESCE(NULLIF(display_order, 0), ?),
                        active = 1
                    WHERE id = ?
                    """,
                    (external_name, nif, index * 10, existing["id"]),
                )
                continue
            mapping_status = "mapping_pending" if nif else "missing_source"
            notes = "Subconta em falta na fonte original; importar mais tarde." if not nif else ""
            conn.execute(
                """
                INSERT INTO portfolio_assets (
                    portfolio_id, asset_id, external_name, nif, sub_account, active,
                    mapping_status, mapping_confidence, notes, display_order, mapping_method, updated_at
                ) VALUES (?, NULL, ?, ?, ?, 1, ?, 0, ?, ?, 'unmapped', ?)
                """,
                (portfolio_id, external_name, nif, sub_account, mapping_status, notes, index * 10, datetime.now().isoformat(timespec="seconds")),
            )


def warning_label(code: str) -> str:
    return WARNING_LABELS.get(code, "Dados incompletos")


def label_warnings(warnings: list[str]) -> list[str]:
    if not warnings:
        return ["OK"]
    return [warning_label(code) for code in warnings]


def data_status_label(warnings: list[str]) -> str:
    if not warnings:
        return "OK"
    if "mapping_pending" in warnings:
        return "Mapping pendente"
    if "mapping_conflict" in warnings:
        return "Mapping conflito"
    if "expired_tariff" in warnings:
        return "Tarifa expirada"
    return "Dados incompletos"


def time_in_rule(sample_time, start, end) -> bool:
    return tariff_time_in_rule(sample_time, start, end)


def classify_tariff_period(moment: datetime, rules: list[sqlite3.Row | dict[str, Any]]) -> str | None:
    return tariff_classify_tariff_period(moment, rules)


def calculate_tariff_value(
    tariff: sqlite3.Row | dict[str, Any] | None,
    *,
    monthly_kwh: float | None,
    hourly_records: list[sqlite3.Row | dict[str, Any]],
    rules: list[sqlite3.Row | dict[str, Any]],
) -> dict[str, Any]:
    config = row_to_tariff_config(tariff, rules) if tariff is not None else None
    records = [row_to_hourly_energy_record(record) for record in hourly_records]
    result = value_tariff_energy(
        config,
        hourly_records=records,
        aggregate_self_use_kwh=None if monthly_kwh is None else decimal_from_value(monthly_kwh),
    )
    legacy = result_to_legacy_dict(result)
    if "missing_hourly_self_use" in legacy["warnings"] and hourly_records:
        legacy["warnings"].append("missing_hourly_production")
    return legacy


def build_portfolio_report_rows(conn: sqlite3.Connection, portfolio_id: int, report_month: str) -> list[dict[str, Any]]:
    start, end = month_bounds(report_month)
    assets = list_portfolio_report_assets(conn, portfolio_id)
    rows: list[dict[str, Any]] = []
    for asset in assets:
        asset_id = int(asset["asset_id"]) if asset["asset_id"] is not None else None
        warnings: list[str] = []
        mapping_status = str(asset["mapping_status"] or "")
        if asset_id is None:
            warnings.append("mapping_pending")
        if mapping_status == "mapping_conflict":
            warnings.append("mapping_conflict")
        prod = get_monthly_production_record(conn, asset_id, start)
        daily_totals = get_daily_production_totals(conn, asset_id, start, end) if asset_id is not None else None
        actual = float(prod["production_kwh"]) if prod and prod["production_kwh"] is not None else None
        if actual is None and daily_totals is not None:
            actual = daily_totals["production_kwh"]
        if actual is None:
            warnings.append("missing_monthly_production")
        financial_expected = get_active_expected_for_month(conn, asset_id=asset_id, year=start.year, month=start.month)
        expected_source = "financial_model" if financial_expected else "none"
        expected_kwh = float(financial_expected["expected_production_kwh"]) if financial_expected and financial_expected["expected_production_kwh"] is not None else None
        expected_consumption_kwh = float(financial_expected["expected_consumption_kwh"]) if financial_expected and financial_expected["expected_consumption_kwh"] is not None else None
        expected_self_use_kwh = float(financial_expected["expected_self_use_kwh"]) if financial_expected and financial_expected["expected_self_use_kwh"] is not None else None
        expected_export_kwh = float(financial_expected["expected_export_kwh"]) if financial_expected and financial_expected["expected_export_kwh"] is not None else None
        expected_grid_import_kwh = float(financial_expected["expected_grid_import_kwh"]) if financial_expected and financial_expected["expected_grid_import_kwh"] is not None else None
        expected_self_consumption_rate_pct = float(financial_expected["expected_self_consumption_rate_pct"]) if financial_expected and financial_expected["expected_self_consumption_rate_pct"] is not None else None
        expected_self_sufficiency_rate_pct = float(financial_expected["expected_self_sufficiency_rate_pct"]) if financial_expected and financial_expected["expected_self_sufficiency_rate_pct"] is not None else None
        expected_specific_yield = None
        if expected_kwh is not None and parse_float(asset["kwp"]):
            expected_specific_yield = expected_kwh / parse_float(asset["kwp"])
        if expected_kwh is None:
            expected = get_latest_helioscope_expected(conn, asset_id, start.month)
            expected_kwh = float(expected["expected_kwh"]) if expected else None
            expected_source = "helioscope" if expected else "none"
        if expected_kwh is None:
            warnings.append("missing_helioscope_expected")
        mount_raw = asset["mounting_date"] or asset["start_contract"]
        mounting = None
        if mount_raw:
            try:
                mounting = datetime.fromisoformat(str(mount_raw)[:10]).date()
            except ValueError:
                warnings.append("invalid_mounting_date")
        else:
            warnings.append("missing_mounting_date")
        factor = calculate_degradation_factor(mounting, start)
        adjusted = expected_kwh * factor if expected_kwh is not None else None
        deviation = actual - adjusted if actual is not None and adjusted is not None else None
        deviation_pct = (deviation / adjusted * 100) if deviation is not None and adjusted else None
        availability = get_monthly_availability(conn, asset_id, start, end) if asset_id is not None else None
        if availability is None:
            warnings.append("missing_availability")
        tariff_validity_warnings = detect_tariff_validity_warnings(conn, asset_id=asset_id, start=start, end=end) if asset_id is not None else ()
        warnings.extend(tariff_validity_warnings)
        tariff = None if "overlapping_tariffs" in tariff_validity_warnings else (get_latest_tariff(conn, asset_id, start) if asset_id is not None else None)
        if tariff is None and asset_id is not None and has_expired_tariff(conn, asset_id, start):
            warnings.append("expired_tariff")
        rules = list_tariff_period_rules(conn, int(tariff["id"])) if tariff else []
        hourly = list_hourly_production_records(
            conn,
            asset_id=asset_id,
            start_iso=start.isoformat(),
            end_iso=datetime(end.year, end.month, end.day, 23, 59, 59).isoformat(),
        )
        monthly_self_use = None
        if prod is not None and "self_use_kwh" in prod.keys() and prod["self_use_kwh"] is not None:
            monthly_self_use = float(prod["self_use_kwh"])
        tariff_result = calculate_tariff_value(tariff, monthly_kwh=monthly_self_use, hourly_records=hourly, rules=rules)
        warnings.extend(tariff_result["warnings"])
        invoice_status = "missing_invoice"
        if tariff and tariff["invoice_file_id"]:
            invoice_doc = conn.execute("SELECT status FROM invoice_documents WHERE source_file_id = ?", (tariff["invoice_file_id"],)).fetchone()
            if invoice_doc is None:
                invoice_status = "review_required"
            elif invoice_doc["status"] == "confirmed":
                invoice_status = "ok"
            elif invoice_doc["status"] == "extraction_failed":
                invoice_status = "extraction_failed"
            elif invoice_doc["status"] in {"rejected", "archived"}:
                invoice_status = "incompatible_invoice"
            else:
                invoice_status = "review_required"
        if invoice_status != "ok":
            warnings.append(invoice_status)
        data_status = "ok" if not warnings else ("missing_data" if any(w.startswith("missing") for w in warnings) else "warning")
        period_kwh = tariff_result["production_period_kwh"]
        self_use_period_kwh = tariff_result["self_use_period_kwh"]
        value_by_period = {
            item.period_name: float(item.value_eur)
            for item in tariff_result.get("breakdown", [])
        }
        energy_by_period = {
            item.period_name: float(item.energy_kwh)
            for item in tariff_result.get("breakdown", [])
        }
        simple_self_use = energy_by_period.get("simple") if tariff else None
        multi_self_use = sum(energy_by_period.get(period, 0.0) for period in PERIOD_NAMES)
        self_use_total = simple_self_use if simple_self_use is not None else (multi_self_use if hourly else monthly_self_use)
        if self_use_total is None and "missing_hourly_self_use" not in tariff_result["warnings"] and actual is not None:
            self_use_total = multi_self_use
        hourly_energy = [row_to_hourly_energy_record(record) for record in hourly]

        def monthly_field(key: str) -> float | None:
            if prod is not None and key in prod.keys() and prod[key] is not None:
                return float(prod[key])
            if daily_totals is not None and key in daily_totals and daily_totals[key] is not None:
                return float(daily_totals[key])
            return None

        def hourly_total(key: str) -> float | None:
            values = [getattr(record, key) for record in hourly_energy if getattr(record, key) is not None]
            return round(float(sum(values)), 6) if values else None

        export_kwh = monthly_field("export_kwh")
        if export_kwh is None:
            export_kwh = hourly_total("export_kwh")
        consumption_kwh = monthly_field("consumption_kwh")
        if consumption_kwh is None:
            consumption_kwh = hourly_total("consumption_kwh")
        grid_import_kwh = monthly_field("grid_import_kwh")
        if grid_import_kwh is None:
            grid_import_kwh = hourly_total("grid_import_kwh")
        if grid_import_kwh is None and consumption_kwh is not None and self_use_total is not None:
            grid_import_kwh = max(consumption_kwh - self_use_total, 0.0)
        if export_kwh is None and actual is not None and self_use_total is not None:
            export_kwh = max(actual - self_use_total, 0.0)
        if self_use_total is None:
            warnings.append("missing_self_use")
        if export_kwh is None:
            warnings.append("missing_export")
        if consumption_kwh is None:
            warnings.append("missing_consumption")
        report_type = detect_report_type_value(asset)
        billing_config = get_asset_billing_config(conn, asset_id, report_type) if asset_id is not None else None
        billing = None
        if actual is not None and self_use_total is not None and export_kwh is not None and consumption_kwh is not None and billing_config is not None:
            billing = calculate_billing(
                EnergyBreakdown(
                    production_kwh=decimal_from_value(actual),
                    self_use_kwh=decimal_from_value(self_use_total),
                    export_kwh=decimal_from_value(export_kwh),
                    consumption_kwh=decimal_from_value(consumption_kwh),
                ),
                billing_config,
                months_count=1,
            )
            warnings.extend(billing.warnings)
        else:
            warnings.append("missing_billing")
        rows.append(
            {
                "portfolio_id": portfolio_id,
                "asset_id": asset_id,
                "portfolio": "",
                "installation": asset["external_name"] or asset["project_name"] or "",
                "external_installation": asset["external_name"] or "",
                "local_installation": asset["project_name"] or "",
                "nif": asset["nif"] or asset["asset_nif"] or "",
                "sub_account": asset["sub_account"] or "",
                "installed_power_kwp": parse_float(asset["kwp"]),
                "actual_production_kwh": round(actual, 2) if actual is not None else None,
                "production_ponta_kwh": round(period_kwh["ponta"], 2),
                "production_cheia_kwh": round(period_kwh["cheia"], 2),
                "production_vazio_kwh": round(period_kwh["vazio"], 2),
                "production_super_vazio_kwh": round(period_kwh["super_vazio"], 2),
                "self_use_kwh": round(self_use_total, 2) if self_use_total is not None else None,
                "self_use_ponta_kwh": round(self_use_period_kwh["ponta"], 2),
                "self_use_cheia_kwh": round(self_use_period_kwh["cheia"], 2),
                "self_use_vazio_kwh": round(self_use_period_kwh["vazio"], 2),
                "self_use_super_vazio_kwh": round(self_use_period_kwh["super_vazio"], 2),
                "self_use_simple_kwh": round(simple_self_use, 2) if simple_self_use is not None else None,
                "self_use_value_ponta_eur": round(value_by_period.get("ponta", 0.0), 2),
                "self_use_value_cheia_eur": round(value_by_period.get("cheia", 0.0), 2),
                "self_use_value_vazio_eur": round(value_by_period.get("vazio", 0.0), 2),
                "self_use_value_super_vazio_eur": round(value_by_period.get("super_vazio", 0.0), 2),
                "self_use_value_simple_eur": round(value_by_period.get("simple", 0.0), 2) if "simple" in value_by_period else None,
                "export_kwh": round(export_kwh, 2) if export_kwh is not None else None,
                "consumption_kwh": round(consumption_kwh, 2) if consumption_kwh is not None else None,
                "grid_import_kwh": round(grid_import_kwh, 2) if grid_import_kwh is not None else None,
                "helioscope_expected_kwh": round(expected_kwh, 2) if expected_kwh is not None else None,
                "expected_production_kwh": round(expected_kwh, 2) if expected_kwh is not None else None,
                "expected_consumption_kwh": round(expected_consumption_kwh, 2) if expected_consumption_kwh is not None else None,
                "expected_self_use_kwh": round(expected_self_use_kwh, 2) if expected_self_use_kwh is not None else None,
                "expected_export_kwh": round(expected_export_kwh, 2) if expected_export_kwh is not None else None,
                "expected_grid_import_kwh": round(expected_grid_import_kwh, 2) if expected_grid_import_kwh is not None else None,
                "expected_self_consumption_rate_pct": round(expected_self_consumption_rate_pct, 2) if expected_self_consumption_rate_pct is not None else None,
                "expected_self_sufficiency_rate_pct": round(expected_self_sufficiency_rate_pct, 2) if expected_self_sufficiency_rate_pct is not None else None,
                "expected_specific_yield": round(expected_specific_yield, 2) if expected_specific_yield is not None else None,
                "expected_production_source": expected_source,
                "adjusted_expected_kwh": round(adjusted, 2) if adjusted is not None else None,
                "degradation_factor": round(factor, 6),
                "deviation_kwh": round(deviation, 2) if deviation is not None else None,
                "deviation_pct": round(deviation_pct, 2) if deviation_pct is not None else None,
                "availability_pct": availability,
                "tariff_type": tariff["tariff_type"] if tariff else "",
                "estimated_value_eur": tariff_result["estimated_value_eur"],
                "export_revenue_eur": round(float(billing.export_revenue_eur), 2) if billing else None,
                "esco_payment_eur": round(float(billing.solcor_payment_eur), 2) if billing else None,
                "fixed_fee_eur": round(float(billing.fixed_monthly_fee_eur), 2) if billing else None,
                "net_benefit_eur": round(float(billing.net_benefit_eur), 2) if billing else None,
                "invoice_status": invoice_status,
                "data_status": data_status,
                "warnings": sorted(set(warnings)),
                "warning_labels": label_warnings(sorted(set(warnings))),
                "status_label": data_status_label(sorted(set(warnings))),
            }
        )
    return rows


def aggregate_portfolio_total(rows: list[dict[str, Any]]) -> dict[str, Any]:
    def total(key: str) -> float | None:
        values = [row.get(key) for row in rows if row.get(key) is not None]
        return round(sum(float(value) for value in values), 2) if values else None

    adjusted_total = total("adjusted_expected_kwh")
    deviation = None
    deviation_pct = None
    actual_total = total("actual_production_kwh")
    if actual_total is not None and adjusted_total:
        deviation = round(actual_total - adjusted_total, 2)
        deviation_pct = round(deviation / adjusted_total * 100, 2)
    availability = calculate_weighted_portfolio_availability(rows)
    warnings = sorted({warning for row in rows for warning in row.get("warnings", [])})
    if rows and any(row.get("availability_pct") is not None and not row.get("installed_power_kwp") for row in rows):
        warnings = sorted({*warnings, "missing_installed_power"})
    return {
        "portfolio": "",
        "installation": "TOTAL",
        "external_installation": "TOTAL",
        "local_installation": "",
        "nif": "",
        "sub_account": "",
        "installed_power_kwp": total("installed_power_kwp"),
        "actual_production_kwh": actual_total,
        "production_ponta_kwh": total("production_ponta_kwh"),
        "production_cheia_kwh": total("production_cheia_kwh"),
        "production_vazio_kwh": total("production_vazio_kwh"),
        "production_super_vazio_kwh": total("production_super_vazio_kwh"),
        "self_use_ponta_kwh": total("self_use_ponta_kwh"),
        "self_use_cheia_kwh": total("self_use_cheia_kwh"),
        "self_use_vazio_kwh": total("self_use_vazio_kwh"),
        "self_use_super_vazio_kwh": total("self_use_super_vazio_kwh"),
        "self_use_value_ponta_eur": total("self_use_value_ponta_eur"),
        "self_use_value_cheia_eur": total("self_use_value_cheia_eur"),
        "self_use_value_vazio_eur": total("self_use_value_vazio_eur"),
        "self_use_value_super_vazio_eur": total("self_use_value_super_vazio_eur"),
        "helioscope_expected_kwh": total("helioscope_expected_kwh"),
        "expected_production_kwh": total("expected_production_kwh"),
        "expected_consumption_kwh": total("expected_consumption_kwh"),
        "expected_self_use_kwh": total("expected_self_use_kwh"),
        "expected_export_kwh": total("expected_export_kwh"),
        "expected_grid_import_kwh": total("expected_grid_import_kwh"),
        "expected_self_consumption_rate_pct": "",
        "expected_self_sufficiency_rate_pct": "",
        "expected_specific_yield": "",
        "expected_production_source": "",
        "adjusted_expected_kwh": adjusted_total,
        "degradation_factor": "",
        "deviation_kwh": deviation,
        "deviation_pct": deviation_pct,
        "availability_pct": availability,
        "tariff_type": "",
        "estimated_value_eur": total("estimated_value_eur"),
        "invoice_status": "",
        "data_status": "ok" if not warnings else "warning",
        "warnings": warnings,
        "warning_labels": label_warnings(warnings),
        "status_label": "OK" if not warnings else "Dados incompletos",
        "missing_data_count": sum(1 for row in rows if row.get("warnings")),
    }


def build_portfolio_kpis(rows: list[dict[str, Any]], total_row: dict[str, Any] | None) -> dict[str, Any]:
    total = total_row or aggregate_portfolio_total(rows)
    return {
        "actual_production_kwh": total.get("actual_production_kwh"),
        "adjusted_expected_kwh": total.get("adjusted_expected_kwh"),
        "deviation_kwh": total.get("deviation_kwh"),
        "deviation_pct": total.get("deviation_pct"),
        "availability_pct": total.get("availability_pct"),
        "estimated_value_eur": total.get("estimated_value_eur"),
        "missing_data_installations": sum(1 for row in rows if row.get("warnings")),
        "installations": len(rows),
    }


def filter_report_rows(rows: list[dict[str, Any]], warning_filter: str) -> list[dict[str, Any]]:
    if warning_filter == "warnings":
        return [row for row in rows if row.get("warnings")]
    if warning_filter == "helioscope":
        return [row for row in rows if "missing_helioscope_expected" in row.get("warnings", [])]
    if warning_filter == "invoice":
        return [row for row in rows if "missing_invoice" in row.get("warnings", [])]
    if warning_filter == "mapping":
        return [row for row in rows if any(item in row.get("warnings", []) for item in {"mapping_pending", "mapping_conflict"})]
    return rows


def snapshot_portfolio_report(conn: sqlite3.Connection, portfolio_id: int, report_month: str, notes: str = "") -> int:
    rows = build_portfolio_report_rows(conn, portfolio_id, report_month)
    cursor = conn.execute(
        "INSERT INTO portfolio_report_runs (portfolio_id, report_month, created_at, notes) VALUES (?, ?, ?, ?)",
        (portfolio_id, report_month, datetime.now().isoformat(timespec="seconds"), notes),
    )
    report_id = int(cursor.lastrowid)
    for row in rows:
        conn.execute(
            """
            INSERT INTO portfolio_report_rows (
                report_id, asset_id, actual_production_kwh, production_ponta_kwh, production_cheia_kwh,
                production_vazio_kwh, production_super_vazio_kwh, helioscope_expected_kwh,
                adjusted_expected_kwh, degradation_factor, deviation_kwh, deviation_pct,
                availability_pct, estimated_value_eur, data_status, warnings_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                report_id,
                row["asset_id"],
                row["actual_production_kwh"],
                row["production_ponta_kwh"],
                row["production_cheia_kwh"],
                row["production_vazio_kwh"],
                row["production_super_vazio_kwh"],
                row["helioscope_expected_kwh"],
                row["adjusted_expected_kwh"],
                row["degradation_factor"],
                row["deviation_kwh"],
                row["deviation_pct"],
                row["availability_pct"],
                row["estimated_value_eur"],
                row["data_status"],
                json.dumps(row["warnings"], ensure_ascii=True),
            ),
        )
    return report_id


def export_portfolio_report_workbook(rows: list[dict[str, Any]]) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Portfolio report"
    headers = [
        ("portfolio", "Portfolio"),
        ("sub_account", "Subconta"),
        ("nif", "NIF"),
        ("external_installation", "Instalacao externa"),
        ("local_installation", "Instalacao local"),
        ("actual_production_kwh", "Producao real mensal kWh"),
        ("production_ponta_kwh", "Producao ponta kWh"),
        ("production_cheia_kwh", "Producao cheia kWh"),
        ("production_vazio_kwh", "Producao vazio kWh"),
        ("production_super_vazio_kwh", "Producao super vazio kWh"),
        ("self_use_ponta_kwh", "Autoconsumo ponta kWh"),
        ("self_use_cheia_kwh", "Autoconsumo cheia kWh"),
        ("self_use_vazio_kwh", "Autoconsumo vazio kWh"),
        ("self_use_super_vazio_kwh", "Autoconsumo super vazio kWh"),
        ("self_use_value_ponta_eur", "Valor autoconsumo ponta EUR"),
        ("self_use_value_cheia_eur", "Valor autoconsumo cheia EUR"),
        ("self_use_value_vazio_eur", "Valor autoconsumo vazio EUR"),
        ("self_use_value_super_vazio_eur", "Valor autoconsumo super vazio EUR"),
        ("helioscope_expected_kwh", "Producao Helioscope base kWh"),
        ("expected_production_kwh", "Producao prevista base kWh"),
        ("expected_consumption_kwh", "Consumo previsto kWh"),
        ("expected_self_use_kwh", "Autoconsumo previsto kWh"),
        ("expected_export_kwh", "Excedente previsto kWh"),
        ("expected_grid_import_kwh", "Importacao rede prevista kWh"),
        ("expected_self_consumption_rate_pct", "Taxa autoconsumo prevista %"),
        ("expected_self_sufficiency_rate_pct", "Taxa autossuficiencia prevista %"),
        ("expected_specific_yield", "Specific yield previsto"),
        ("expected_production_source", "Origem producao prevista"),
        ("adjusted_expected_kwh", "Producao esperada ajustada kWh"),
        ("degradation_factor", "Fator degradacao"),
        ("deviation_kwh", "Desvio kWh"),
        ("deviation_pct", "Desvio %"),
        ("availability_pct", "Availability time %"),
        ("tariff_type", "Tipo de tarifa"),
        ("estimated_value_eur", "Valor estimado EUR"),
        ("status_label", "Estado dos dados"),
        ("warning_labels", "Avisos"),
    ]
    sheet.append([label for _, label in headers])
    for row in rows:
        sheet.append(["; ".join(row[key]) if key == "warning_labels" and isinstance(row.get(key), list) else row.get(key, "") for key, _ in headers])
    for column in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(width + 2, 12), 42)
    return workbook
