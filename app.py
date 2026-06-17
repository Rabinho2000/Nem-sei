from __future__ import annotations

import argparse
import base64
import calendar
import html
import io
import json
import logging
import os
import re
import secrets
import sqlite3
import threading
import time
import unicodedata
from contextlib import closing
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import requests
from apscheduler.schedulers.background import BackgroundScheduler
from flask import current_app
from flask import Flask, abort, flash, g, has_app_context, redirect, render_template, request, send_file, session, url_for
from monitoring_board.db import configure_database_for_runtime, create_database_backup, ensure_column, get_db, query_all, query_scalar
from monitoring_board.logging_config import configure_logging
from monitoring_board.routes.auth import auth_bp
from monitoring_board.routes.field_routes import field_routes_bp
from monitoring_board.runtime import (
    BACKUP_DIR,
    BASE_DIR,
    CONTRACTS_DIR,
    DB_PATH,
    DEFAULT_EXCEL_PATH,
    LOG_DIR,
    RUNTIME_PATHS,
    UPLOAD_DIR,
    build_runtime_paths,
    ensure_runtime_directories,
    env_flag,
    load_local_env,
    max_upload_bytes,
    path_is_within,
    resolve_runtime_file_path,
    resolve_runtime_file_path_within,
    store_runtime_relative_path,
)
from monitoring_board.security import app_password_configured, csrf_token, flask_secret_key
from monitoring_board.customer_reports import (
    build_customer_report_pdf,
    detect_report_type,
    prepare_customer_report,
)
from monitoring_board.services.fusionsolar import (
    build_provider_url,
    classify_fusionsolar_inverter_availability,
    describe_fusionsolar_health_state,
    map_fusionsolar_status,
    normalize_sync_hours,
)
from monitoring_board.services.telegram_service import (
    get_telegram_config,
    is_telegram_configured,
    send_telegram_message,
    telegram_daily_summary_enabled,
    test_telegram_connection,
)
from openpyxl import load_workbook
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


INTEGRATION_PROVIDER_FUSIONSOLAR = "FusionSolar"
INTEGRATION_PROVIDER_SIGENERGY = "Sigenergy"
INTEGRATION_PROVIDER_OPTIONS = [INTEGRATION_PROVIDER_FUSIONSOLAR, INTEGRATION_PROVIDER_SIGENERGY]
BACKGROUND_JOB_TYPES_PERFORMANCE = (
    "fusionsolar_production_sync",
    "fusionsolar_production_backfill",
    "fusionsolar_inverter_availability_backfill",
    "fusionsolar_month_cycle",
    "performance_reference_recalculation",
)
BACKGROUND_JOB_STALE_RUNNING_MINUTES = 30
DEFAULT_FUSIONSOLAR_SYNC_HOURS = "08:00,14:00"
DEFAULT_FUSIONSOLAR_LOGIN_ENDPOINT = "/thirdData/login"
DEFAULT_FUSIONSOLAR_STATIONS_ENDPOINT = "/thirdData/stations"
DEFAULT_FUSIONSOLAR_REALTIME_ENDPOINT = "/thirdData/getStationRealKpi"
DEFAULT_FUSIONSOLAR_DEVICES_ENDPOINT = "/thirdData/getDevList"
DEFAULT_FUSIONSOLAR_DEVICE_REALTIME_ENDPOINT = "/thirdData/getDevRealKpi"
DEFAULT_FUSIONSOLAR_DEVICE_HISTORY_ENDPOINT = "/thirdData/getDevHistoryKpi"
DEFAULT_FUSIONSOLAR_ALARMS_ENDPOINT = "/thirdData/getAlarmList"
DEFAULT_FUSIONSOLAR_DAY_KPI_ENDPOINT = "/thirdData/getKpiStationDay"
DEFAULT_FUSIONSOLAR_MONTH_KPI_ENDPOINT = "/thirdData/getKpiStationMonth"
DEFAULT_FUSIONSOLAR_ALARMS_LANGUAGE = "en_US"
DEFAULT_SIGENERGY_BASE_URL = "https://api-eu.sigencloud.com"
DEFAULT_SIGENERGY_AUTH_ENDPOINT = "/openapi/auth/login/key"
DEFAULT_SIGENERGY_SYSTEMS_ENDPOINT = "/openapi/system/list"
DEFAULT_SIGENERGY_REALTIME_ENDPOINT = "/openapi/system/realtime/data"
DEFAULT_SIGENERGY_ENERGY_FLOW_ENDPOINT = "/openapi/systems/{system_id}/energyFlow"
DEFAULT_SIGENERGY_REGION = "eu"
FUSIONSOLAR_PERFORMANCE_RATE_LIMIT_MINUTES = 60
FUSIONSOLAR_PERFORMANCE_KPI_DELAY_SECONDS = 65
FUSIONSOLAR_PERFORMANCE_MAX_API_CALLS = 20
FUSIONSOLAR_PERFORMANCE_RATE_LIMIT_UNTIL: datetime | None = None
DEFAULT_DEVICE_COMMUNICATION_THRESHOLD_MINUTES = 15
FUSIONSOLAR_INVERTER_DEVICE_TYPE_IDS = {1, 38}
INVERTER_AVAILABILITY_SLOT_MINUTES = 15
INVERTER_AVAILABILITY_EDGE_TOLERANCE_MINUTES = 30
LOW_INVERTER_AVAILABILITY_PCT = 90.0
LISBON_TIMEZONE = ZoneInfo("Europe/Lisbon")
DEFAULT_STRING_PRESENT_VOLTAGE_THRESHOLD = 100.0
DEFAULT_STRING_AUTO_LEARN_OBSERVATIONS = 2
SIGENERGY_TOKEN_CACHE: dict[str, dict[str, Any]] = {}
SIGENERGY_TOKEN_LOCK = threading.Lock()

STATUS_COLORS = {
    "Erro": "danger",
    "Desconectada": "warning",
    "Resolvido": "success",
    "Operacional": "success",
    "OK": "success",
    "Atenção": "warning",
    "Alerta": "warning",
    "Crítico": "danger",
    "Sem referência": "muted",
    "Sem dados": "muted",
    "Aberto": "danger",
    "Em analise": "warning",
    "Agendado": "accent",
    "Em visita": "accent",
    "Fechado": "muted",
}

TICKET_STATUSES = ["Aberto", "Em analise", "Agendado", "Em visita", "Resolvido", "Fechado"]
TICKET_URGENCIES = ["Baixa", "Media", "Alta", "Critica"]
TICKET_MATERIAL_STATUSES = ["Nao definido", "Sem material", "Necessario", "Pronto", "Bloqueado"]
TICKET_WORK_TYPES = ["Diagnostico", "Comunicacao", "Inversor", "String", "Estrutura", "Limpeza", "Preventiva", "Outro"]
MONTH_NAMES_PT = [
    "",
    "Janeiro",
    "Fevereiro",
    "Março",
    "Abril",
    "Maio",
    "Junho",
    "Julho",
    "Agosto",
    "Setembro",
    "Outubro",
    "Novembro",
    "Dezembro",
]
MONITORING_SOURCES = ["FusionSolar", "Sigenergy", "Manual / Outro"]
ASSET_MONITORING_STATUSES = ["active", "silenced", "maintenance", "out_of_scope", "disabled"]
OK_MONITORING_STATUSES = {"Operacional", "Resolvido", "OK"}
PROBLEM_MONITORING_STATUSES = {"Erro", "Desconectada"}
ALERT_SCOPE_OPTIONS = ["all_assets", "only_o&m", "only_active_contracts", "only_selected_assets"]
ALERT_SETTING_DEFAULTS = {
    "TELEGRAM_ALERTS_ENABLED": "true",
    "ALERT_SCOPE": "only_o&m",
    "SEND_NEW_ERROR_ALERTS": "true",
    "SEND_OFFLINE_ALERTS": "true",
    "SEND_RESOLVED_ALERTS": "true",
    "SEND_PERSISTENT_ALERTS": "true",
    "SEND_RECURRENT_ALERTS": "false",
    "DAYTIME_OFFLINE_ONLY": "true",
    "IGNORE_HISTORICAL_ALERTS": "true",
    "MINIMUM_ALERT_SEVERITY": "info",
    "NEW_ERROR_COOLDOWN_MINUTES": "0",
    "OFFLINE_COOLDOWN_MINUTES": "120",
    "RESOLVED_COOLDOWN_MINUTES": "0",
    "PERSISTENT_COOLDOWN_HOURS": "24",
    "RECURRENT_COOLDOWN_HOURS": "24",
    "ALERT_BASELINE_AT": "",
}
RENEWAL_STATUSES = ["Por contactar", "Email enviado", "Em negociacao", "Renovado", "Sem interesse"]
INTEGRATION_STATUS_COLORS = {
    "success": "success",
    "error": "danger",
    "warning": "warning",
    "pending": "accent",
}
EXPORT_DATASETS = {
    "assets": {
        "label": "Instalacoes / centrais",
        "columns": [
            ("project_name", "Central"),
            ("location", "Localizacao"),
            ("address", "Morada"),
            ("contact_phone", "Contacto"),
            ("contact_name", "Nome"),
            ("access_type", "Acesso"),
            ("coverage_type", "Tipo de cobertura"),
            ("contract_type", "Contrato"),
            ("active_contract", "O&M"),
            ("company_name", "Empresa"),
            ("contact_email", "Email"),
        ],
    },
    "monitoring": {
        "label": "Monitorizacao filtrada",
        "columns": [
            ("record_date", "Data"),
            ("imported_at", "Importado em"),
            ("project_name", "Central"),
            ("location", "Localizacao"),
            ("contract_type", "Contrato"),
            ("active_contract", "O&M"),
            ("status", "Estado"),
            ("notes", "Notas"),
            ("source", "Origem"),
        ],
    },
    "tickets": {
        "label": "Intervencoes O&M",
        "columns": [
            ("project_name", "Central"),
            ("location", "Localizacao"),
            ("contract_type", "Contrato"),
            ("active_contract", "O&M"),
            ("title", "Titulo"),
            ("status", "Estado"),
            ("urgency", "Urgencia"),
            ("installation_ref", "Referencia"),
            ("next_action", "Proxima acao"),
            ("work_type", "Tipo de trabalho"),
            ("material_status", "Material"),
            ("planned_date", "Data planeada"),
            ("due_date", "Data limite"),
            ("estimated_minutes", "Minutos previstos"),
            ("assigned_to", "Equipa"),
            ("planning_notes", "Notas planeamento"),
            ("notes", "Notas"),
            ("created_at", "Criado em"),
            ("updated_at", "Atualizado em"),
        ],
    },
    "executive_report": {
        "label": "Relatorio executivo O&M",
        "columns": [
            ("section", "Seccao"),
            ("priority", "Prioridade"),
            ("project_name", "Central"),
            ("status", "Estado"),
            ("problem_days", "Dias em problema"),
            ("recurrence_count", "Recorrencias 90d"),
            ("open_tickets", "Tickets abertos"),
            ("source", "Origem"),
            ("notes", "Notas"),
        ],
    },
    "monitoring_report": {
        "label": "Relatorio limpo de monitorizacao",
        "columns": [
            ("period", "Periodo"),
            ("project_name", "Instalacao"),
            ("location", "Localizacao"),
            ("current_status", "Estado atual"),
            ("last_record_date", "Ultima monitorizacao"),
            ("monitoring_records", "Registos no periodo"),
            ("error_records", "Erros no periodo"),
            ("distinct_errors", "Erros diferentes"),
            ("error_types", "Tipos de erro"),
            ("open_tickets", "Tickets abertos"),
            ("visits_period", "Visitas no periodo"),
            ("last_visit_date", "Ultima visita"),
            ("latest_notes", "Notas"),
        ],
    },
    "production_report": {
        "label": "Relatorio de producao mensal/anual",
        "columns": [
            ("period", "Periodo"),
            ("project_name", "Instalacao"),
            ("location", "Localizacao"),
            ("provider", "Origem API"),
            ("production_kwh", "Producao kWh"),
            ("specific_yield", "kWh/kWp"),
            ("expected_kwh", "Producao esperada kWh"),
            ("deviation_pct", "Desvio %"),
            ("performance_status", "Estado performance"),
            ("data_points", "Pontos de dados"),
            ("data_source", "Tipo de dados"),
            ("last_update", "Ultima atualizacao"),
            ("notes", "Notas"),
        ],
    },
}

GROUP_INHERITED_FIELDS = [
    "company_name",
    "location",
    "address",
    "contract_type",
    "contact_name",
    "contact_role",
    "contact_email",
    "contact_phone",
    "access_type",
    "coverage_type",
]

SCHEDULER: BackgroundScheduler | None = None
FUSIONSOLAR_SESSION_CACHE: dict[str, Any] = {}
FUSIONSOLAR_SESSION_LOCK = threading.Lock()
FUSIONSOLAR_SYNC_LOCK = threading.Lock()


def create_app() -> Flask:
    ensure_runtime_directories(RUNTIME_PATHS)
    app = Flask(__name__)
    app.config["SECRET_KEY"] = flask_secret_key()
    app.config["DATA_DIR"] = str(RUNTIME_PATHS.data_dir)
    app.config["DATABASE"] = str(DB_PATH)
    app.config["EXCEL_PATH"] = str(DEFAULT_EXCEL_PATH) if DEFAULT_EXCEL_PATH else ""
    app.config["MAX_CONTENT_LENGTH"] = max_upload_bytes()
    app.config["SESSION_COOKIE_HTTPONLY"] = True
    app.config["SESSION_COOKIE_SAMESITE"] = os.environ.get("SESSION_COOKIE_SAMESITE", "Lax").strip() or "Lax"
    app.config["SESSION_COOKIE_SECURE"] = env_flag("SESSION_COOKIE_SECURE", False)
    app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=12)
    configure_logging(app, LOG_DIR)
    app.logger.info("Using database at %s", app.config["DATABASE"])
    app.register_blueprint(auth_bp)
    if not app_password_configured():
        app.logger.warning("APP_PASSWORD_HASH/APP_PASSWORD is not configured; login is locked until .env is updated.")

    ensure_database(app.config["DATABASE"])
    app.register_blueprint(field_routes_bp)
    with closing(get_db(app.config["DATABASE"])) as bootstrap_conn:
        populate_missing_installation_groups(bootstrap_conn)
        populate_missing_group_metadata(bootstrap_conn)
        sync_all_contract_statuses(bootstrap_conn)
        ensure_integration_seed_data(bootstrap_conn)
        bootstrap_conn.commit()
    start_integration_scheduler(app)

    @app.before_request
    def before_request() -> None:
        g.db = get_db(app.config["DATABASE"])
        g.request_started_at = datetime.now()
        if request.method == "POST":
            sent_token = request.form.get("csrf_token") or request.headers.get("X-CSRF-Token", "")
            if not sent_token or not secrets.compare_digest(sent_token, csrf_token()):
                app.logger.warning("CSRF validation failed for %s %s", request.method, request.path)
                abort(400)
        if request.endpoint not in {"auth.login", "static"} and not session.get("authenticated"):
            return redirect(url_for("auth.login", next=request.full_path if request.query_string else request.path))

    @app.teardown_request
    def teardown_request(exception: BaseException | None) -> None:
        started_at = getattr(g, "request_started_at", None)
        elapsed_ms = ""
        if started_at:
            elapsed_ms = f" {(datetime.now() - started_at).total_seconds() * 1000:.0f}ms"
        if request.endpoint != "static":
            if exception:
                app.logger.exception("%s %s failed%s", request.method, request.path, elapsed_ms)
            else:
                app.logger.info("%s %s%s", request.method, request.path, elapsed_ms)
        db = g.pop("db", None)
        if db is not None:
            db.close()

    @app.context_processor
    def inject_globals() -> dict[str, Any]:
        return {
            "today_iso": date.today().isoformat(),
            "ticket_statuses": TICKET_STATUSES,
            "ticket_urgencies": TICKET_URGENCIES,
            "ticket_material_statuses": TICKET_MATERIAL_STATUSES,
            "ticket_work_types": TICKET_WORK_TYPES,
            "status_colors": STATUS_COLORS,
            "monitoring_sources": MONITORING_SOURCES,
            "asset_monitoring_statuses": ASSET_MONITORING_STATUSES,
            "renewal_statuses": RENEWAL_STATUSES,
            "integration_status_colors": INTEGRATION_STATUS_COLORS,
            "om_status_label": om_status_label,
            "format_date_pt": format_date_pt,
            "format_number": format_number,
            "compute_performance_percentage": compute_performance_percentage,
            "performance_bar_width": performance_bar_width,
            "performance_status_class": performance_status_class,
            "reference_diagnostic": reference_diagnostic,
            "csrf_token": csrf_token,
            "current_username": session.get("username"),
        }

    @app.errorhandler(400)
    def bad_request_error(error: Exception) -> tuple[str, int]:
        return render_template(
            "error.html",
            title="Pedido invalido",
            heading="Pedido invalido",
            message="A acao nao foi aceite. Atualiza a pagina e tenta novamente.",
        ), 400

    @app.errorhandler(404)
    def not_found_error(error: Exception) -> tuple[str, int]:
        return render_template(
            "error.html",
            title="Pagina nao encontrada",
            heading="Pagina nao encontrada",
            message="Nao encontrei esta pagina ou registo.",
        ), 404

    @app.errorhandler(500)
    def internal_error(error: Exception) -> tuple[str, int]:
        current_app.logger.exception("Unhandled application error")
        return render_template(
            "error.html",
            title="Erro interno",
            heading="Erro interno",
            message="Aconteceu um erro inesperado. Consulta os logs para o detalhe tecnico.",
        ), 500

    @app.errorhandler(413)
    def request_too_large_error(error: Exception) -> tuple[str, int]:
        return render_template(
            "error.html",
            title="Ficheiro demasiado grande",
            heading="Ficheiro demasiado grande",
            message="O ficheiro enviado excede o limite configurado para uploads.",
        ), 413

    @app.route("/operation")
    def operation() -> str:
        search = request.args.get("search", "").strip()
        om_only = request.args.get("om_only", "yes").strip()
        calendar_month = normalize_calendar_month(request.args.get("calendar_month", ""))
        calendar_start, calendar_end, previous_month, next_month = calendar_month_bounds(calendar_month)

        intervention_conditions = ["t.status != 'Fechado'"]
        intervention_params: list[Any] = []
        if om_only == "yes":
            intervention_conditions.append("a.active_contract = 'yes'")
        if search:
            wildcard = f"%{search}%"
            intervention_conditions.append(
                "(a.project_name LIKE ? OR a.location LIKE ? OR a.address LIKE ? OR t.title LIKE ? OR COALESCE(t.next_action, '') LIKE ?)"
            )
            intervention_params.extend([wildcard, wildcard, wildcard, wildcard, wildcard])
        intervention_where = " AND ".join(intervention_conditions)
        interventions = query_all(
            g.db,
            f"""
            SELECT
                t.*,
                a.project_name,
                a.installation_group,
                a.location,
                a.address,
                a.active_contract,
                a.latitude,
                a.longitude,
                a.coordinates_confidence,
                lm.status AS latest_status,
                lm.record_date AS latest_status_date
            FROM tickets t
            JOIN assets a ON a.id = t.asset_id
            LEFT JOIN latest_monitoring_view lm ON lm.asset_id = a.id
            WHERE {intervention_where}
            ORDER BY
                CASE t.urgency WHEN 'Critica' THEN 1 WHEN 'Alta' THEN 2 WHEN 'Media' THEN 3 ELSE 4 END,
                CASE WHEN COALESCE(t.planned_date, '') = '' THEN 0 ELSE 1 END,
                COALESCE(t.planned_date, '9999-12-31') ASC,
                t.updated_at DESC
            """,
            intervention_params,
        )

        problem_conditions = [
            "COALESCE(a.monitoring_status, 'active') != 'disabled'",
            "lm.status IN ('Erro', 'Desconectada')",
            "COALESCE(t.open_tickets, 0) = 0",
        ]
        problem_params: list[Any] = []
        if om_only == "yes":
            problem_conditions.append("a.active_contract = 'yes'")
        if search:
            wildcard = f"%{search}%"
            problem_conditions.append("(a.project_name LIKE ? OR a.location LIKE ? OR a.address LIKE ?)")
            problem_params.extend([wildcard, wildcard, wildcard])
        problems_without_action = enrich_operational_rows(
            g.db,
            query_all(
                g.db,
                f"""
                SELECT
                    a.id AS asset_id,
                    a.project_name,
                    a.installation_group,
                    a.location,
                    a.address,
                    a.active_contract,
                    a.latitude,
                    a.longitude,
                    a.coordinates_confidence,
                    lm.status,
                    lm.record_date,
                    0 AS open_tickets
                FROM assets a
                JOIN latest_monitoring_view lm ON lm.asset_id = a.id
                LEFT JOIN (
                    SELECT asset_id, COUNT(*) AS open_tickets
                    FROM tickets
                    WHERE status != 'Fechado'
                    GROUP BY asset_id
                ) t ON t.asset_id = a.id
                WHERE {" AND ".join(problem_conditions)}
                ORDER BY
                    CASE lm.status WHEN 'Erro' THEN 1 WHEN 'Desconectada' THEN 2 ELSE 3 END,
                    lm.record_date ASC,
                    a.project_name COLLATE NOCASE
                LIMIT 20
                """,
                problem_params,
            ),
        )

        planned_rows = query_all(
            g.db,
            f"""
            SELECT
                t.*,
                a.project_name,
                a.location,
                a.latitude,
                a.longitude,
                a.coordinates_confidence,
                lm.status AS latest_status
            FROM tickets t
            JOIN assets a ON a.id = t.asset_id
            LEFT JOIN latest_monitoring_view lm ON lm.asset_id = a.id
            WHERE t.status != 'Fechado'
              AND COALESCE(t.planned_date, '') BETWEEN ? AND ?
              {"AND a.active_contract = 'yes'" if om_only == "yes" else ""}
            ORDER BY t.planned_date ASC, t.urgency DESC, a.project_name COLLATE NOCASE
            """,
            [calendar_start.isoformat(), calendar_end.isoformat()],
        )
        planning_calendar = build_intervention_calendar(calendar_month, planned_rows)

        today = date.today().isoformat()
        week_end = (date.today() + timedelta(days=7)).isoformat()
        operation_stats = {
            "open": len(interventions),
            "critical": sum(1 for row in interventions if row["urgency"] == "Critica"),
            "unplanned": sum(1 for row in interventions if not row["planned_date"]),
            "blocked": sum(1 for row in interventions if row["material_status"] == "Bloqueado"),
            "this_week": sum(1 for row in interventions if today <= (row["planned_date"] or "") <= week_end),
            "without_action": len(problems_without_action),
            "ready_for_route": sum(1 for row in interventions if intervention_ready_for_route(row)),
        }

        return render_template(
            "operation.html",
            title="Operacao O&M",
            search=search,
            om_only=om_only,
            interventions=interventions,
            problems_without_action=problems_without_action,
            operation_stats=operation_stats,
            planning_calendar=planning_calendar,
            calendar_month=calendar_month,
            previous_month=previous_month,
            next_month=next_month,
            today=today,
            week_end=week_end,
        )

    @app.route("/")
    def dashboard() -> str:
        stats = fetch_dashboard_stats(g.db)
        availability_summary = get_dashboard_availability_summary(g.db)
        monitoring_by_day = query_all(
            g.db,
            """
            SELECT record_date, COUNT(*) AS total
            FROM monitoring_records
            GROUP BY record_date
            ORDER BY record_date DESC
            LIMIT 7
            """,
        )
        critical_assets = enrich_operational_rows(g.db, query_all(
            g.db,
            """
            SELECT
                a.id AS asset_id,
                a.project_name,
                a.active_contract,
                lm.status,
                lm.record_date,
                COUNT(t.id) AS open_tickets
            FROM assets a
            LEFT JOIN latest_monitoring_view lm ON lm.asset_id = a.id
            LEFT JOIN tickets t ON t.asset_id = a.id AND t.status != 'Fechado'
            WHERE a.active_contract = 'yes'
              AND COALESCE(a.monitoring_status, 'active') != 'disabled'
              AND (lm.status IN ('Erro', 'Desconectada') OR t.id IS NOT NULL)
            GROUP BY a.id, a.project_name, a.active_contract, lm.status, lm.record_date
            ORDER BY
                CASE lm.status WHEN 'Erro' THEN 1 WHEN 'Desconectada' THEN 2 ELSE 3 END,
                open_tickets DESC,
                a.project_name COLLATE NOCASE
            LIMIT 12
            """,
        ))
        critical_assets.sort(
            key=lambda row: (
                priority_rank(row["auto_priority"]),
                -int(row.get("problem_days") or 0),
                -int(row.get("recurrence_count") or 0),
                row["project_name"].lower(),
            )
        )
        potential_assets = query_all(
            g.db,
            """
            SELECT
                a.id,
                a.project_name,
                a.active_contract,
                lm.status,
                lm.record_date
            FROM assets a
            LEFT JOIN latest_monitoring_view lm ON lm.asset_id = a.id
            WHERE COALESCE(a.active_contract, '') != 'yes'
              AND COALESCE(a.monitoring_status, 'active') != 'disabled'
              AND lm.status IN ('Erro', 'Desconectada')
            ORDER BY
                CASE lm.status WHEN 'Erro' THEN 1 WHEN 'Desconectada' THEN 2 ELSE 3 END,
                a.project_name COLLATE NOCASE
            LIMIT 10
            """,
        )
        open_ticket_assets = query_all(
            g.db,
            """
            SELECT
                a.id,
                a.project_name,
                a.active_contract,
                COUNT(t.id) AS ticket_count,
                SUM(CASE WHEN t.urgency = 'Critica' THEN 1 ELSE 0 END) AS critical_count,
                MAX(t.updated_at) AS last_update
            FROM assets a
            JOIN tickets t ON t.asset_id = a.id
            WHERE t.status != 'Fechado' AND a.active_contract = 'yes'
            GROUP BY a.id, a.project_name, a.active_contract
            ORDER BY critical_count DESC, ticket_count DESC, a.project_name COLLATE NOCASE
            LIMIT 10
            """,
        )
        renewal_focus = query_all(
            g.db,
            """
            SELECT
                a.id,
                a.project_name,
                a.installation_group,
                a.company_name,
                COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) AS contract_end_date,
                COALESCE(oc.renewal_status, 'Por contactar') AS renewal_status
            FROM assets a
            LEFT JOIN om_contracts oc ON oc.asset_id = a.id
            WHERE (a.maintenance = 'yes' OR oc.id IS NOT NULL)
              AND COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) NOT IN ('', '-')
              AND (
                  COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) < ?
                  OR substr(COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')), 1, 4) = ?
              )
            ORDER BY COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) ASC, a.project_name COLLATE NOCASE
            LIMIT 10
            """,
            (date.today().isoformat(), str(date.today().year)),
        )
        executive_stats = build_executive_dashboard_stats(g.db)
        executive_priorities = critical_assets[:8]
        integration_summary = build_integration_summary(g.db)
        performance_risk_count = query_scalar(
            g.db,
            """
            SELECT COUNT(*)
            FROM availability_daily ad
            JOIN (
                SELECT asset_id, MAX(period_date || 'T' || printf('%09d', id)) AS marker
                FROM availability_daily
                GROUP BY asset_id
            ) latest
              ON latest.asset_id = ad.asset_id
             AND latest.marker = ad.period_date || 'T' || printf('%09d', ad.id)
            WHERE COALESCE(ad.unavailable_inverters, 0) > 0
               OR COALESCE(ad.no_communication_devices, 0) > 0
            """,
        )
        return render_template(
            "dashboard.html",
            stats=stats,
            executive_stats=executive_stats,
            executive_priorities=executive_priorities,
            integration_summary=integration_summary,
            availability_summary=availability_summary,
            monitoring_by_day=monitoring_by_day,
            critical_assets=critical_assets,
            potential_assets=potential_assets,
            open_ticket_assets=open_ticket_assets,
            renewal_focus=renewal_focus,
            performance_risk_count=performance_risk_count,
        )

    @app.route("/performance", methods=["GET", "POST"])
    def performance() -> str:
        if request.method == "POST":
            action = request.form.get("action", "sync_availability").strip()
            if action == "sync_availability":
                result = run_fusionsolar_device_availability_sync(
                    g.db,
                    INTEGRATION_PROVIDER_FUSIONSOLAR,
                    trigger_type="manual",
                )
                flash(
                    f"Disponibilidade sincronizada: {result['devices']} dispositivos, {result['assets']} centrais.",
                    "success",
                )
                return redirect(url_for("performance"))
            if action == "sync_inverter_time_availability":
                from_date = parse_date_value(request.form.get("from_date", ""))
                to_date = parse_date_value(request.form.get("to_date", ""))
                search = request.form.get("search", "").strip()
                om_only = request.form.get("om_only", "yes").strip()
                if not from_date or not to_date or from_date > to_date:
                    flash("Intervalo invalido para disponibilidade dos inversores.", "error")
                    return redirect(url_for("performance"))
                if to_date >= date.today():
                    flash("A disponibilidade temporal so pode ser calculada para dias fechados.", "error")
                    return redirect(url_for("performance"))
                if (to_date - from_date).days > 31:
                    flash("Calcula no maximo 32 dias de cada vez.", "error")
                    return redirect(url_for("performance"))
                job_id, created = create_background_job(
                    g.db,
                    "fusionsolar_inverter_availability_backfill",
                    {
                        "provider": INTEGRATION_PROVIDER_FUSIONSOLAR,
                        "from_date": from_date.isoformat(),
                        "to_date": to_date.isoformat(),
                    },
                )
                g.db.commit()
                if created:
                    schedule_background_job(current_app._get_current_object(), job_id)
                    flash(f"Calculo WAT enviado para background (job #{job_id}).", "success")
                else:
                    flash(f"Ja existe um calculo WAT pendente/em execucao (job #{job_id}).", "warning")
                return redirect(
                    url_for(
                        "performance",
                        period="custom",
                        from_date=from_date,
                        to_date=to_date,
                        search=search,
                        om_only=om_only,
                    )
                )
            flash("Acao de disponibilidade invalida.", "error")
            return redirect(url_for("performance"))

        asset_id = request.args.get("asset_id", "").strip()
        search = request.args.get("search", "").strip()
        om_only = request.args.get("om_only", "yes").strip()
        availability_period = request.args.get("period", "yesterday").strip()
        availability_from, availability_to = resolve_inverter_availability_period(
            availability_period,
            request.args.get("from_date", ""),
            request.args.get("to_date", ""),
        )

        inverter_time_report = get_inverter_availability_report(
            g.db,
            availability_from,
            availability_to,
            asset_id=int(asset_id) if asset_id.isdigit() else None,
            om_only=om_only == "yes",
            search=search,
        )
        inverter_chart_report = (
            get_inverter_availability_chart_report(
                g.db,
                int(asset_id),
                availability_from,
                availability_to,
            )
            if asset_id.isdigit()
            else None
        )
        return render_template(
            "performance.html",
            selected_asset_id=asset_id,
            search=search,
            om_only=om_only,
            inverter_time_report=inverter_time_report,
            inverter_chart_report=inverter_chart_report,
            availability_period=availability_period,
            availability_from=availability_from,
            availability_to=availability_to,
            availability_closed_max=date.today() - timedelta(days=1),
            background_jobs=fetch_latest_background_jobs(
                g.db,
                job_types=("fusionsolar_inverter_availability_backfill",),
            ),
            fusionsolar_api_warning=get_fusionsolar_performance_cooldown_reason(g.db),
        )

    @app.route("/performance/debug/<int:record_id>")
    def performance_debug(record_id: int) -> str:
        record = query_one(
            """
            SELECT pr.*, a.project_name
            FROM production_records pr
            JOIN assets a ON a.id = pr.asset_id
            WHERE pr.id = ?
            """,
            (record_id,),
        )
        if record is None:
            flash("Registo de performance nao encontrado.", "error")
            return redirect(url_for("performance"))

        raw_payload = record["payload_json"] or "{}"
        try:
            parsed_payload = json.loads(raw_payload)
            pretty_payload = json.dumps(parsed_payload, indent=2, ensure_ascii=False)
        except json.JSONDecodeError:
            pretty_payload = raw_payload

        return render_template(
            "performance_debug.html",
            record=record,
            pretty_payload=pretty_payload,
        )

    @app.route("/performance/backfill", methods=["GET", "POST"])
    def performance_backfill() -> str:
        current_year = date.today().year
        period_type = request.values.get("period_type", "day").strip()
        if period_type not in {"day", "month"}:
            period_type = "day"
        from_year = int(request.values.get("from_year", current_year - 1) or current_year - 1)
        to_year = int(request.values.get("to_year", current_year) or current_year)
        date_from_raw = request.values.get("date_from", "").strip()
        date_to_raw = request.values.get("date_to", "").strip()
        date_from = parse_date_value(date_from_raw)
        date_to = parse_date_value(date_to_raw)
        max_api_calls_raw = request.values.get("max_api_calls", request.values.get("max_days", str(FUSIONSOLAR_PERFORMANCE_MAX_API_CALLS))).strip()
        max_api_calls = int(max_api_calls_raw) if max_api_calls_raw.isdigit() else FUSIONSOLAR_PERFORMANCE_MAX_API_CALLS
        asset_id_raw = request.values.get("asset_id", "").strip()
        asset_id = int(asset_id_raw) if asset_id_raw.isdigit() else None
        estimated_station_count = 0
        estimated_api_calls = 0
        if request.method == "GET":
            estimated_assets = get_fusionsolar_performance_assets(g.db, INTEGRATION_PROVIDER_FUSIONSOLAR, asset_id)
            estimated_station_codes = [str(asset["external_id"] or "").strip() for asset in estimated_assets if str(asset["external_id"] or "").strip()]
            estimated_station_count = len(estimated_station_codes)
            estimated_chunks = len(chunked(estimated_station_codes, 100))
            estimated_periods = (
                len(iter_daily_backfill_months(from_year, to_year, today_value=date.today(), date_from=date_from, date_to=date_to))
                if period_type == "day"
                else len(iter_monthly_backfill_dates(from_year, to_year, today_value=date.today()))
            )
            estimated_api_calls = estimated_periods * estimated_chunks

        if request.method == "POST" and request.form.get("action") == "month_cycle":
            cycle_month = normalize_report_month(request.form.get("cycle_month", ""))
            cycle_asset_ids = [int(value) for value in request.form.getlist("cycle_asset_ids") if value.isdigit()]
            if not cycle_asset_ids:
                flash("Escolhe pelo menos uma instalacao para o ciclo.", "error")
                return redirect(url_for("performance_backfill", period_type=period_type))
            job_id, created = create_background_job(
                g.db,
                "fusionsolar_month_cycle",
                {
                    "provider": INTEGRATION_PROVIDER_FUSIONSOLAR,
                    "report_month": cycle_month,
                    "asset_ids": cycle_asset_ids,
                },
            )
            g.db.commit()
            if created:
                schedule_background_job(current_app._get_current_object(), job_id)
                flash(f"Ciclo mensal FusionSolar enviado para background (job #{job_id}).", "success")
            else:
                flash(f"Ja existe um ciclo mensal FusionSolar pendente/em execucao (job #{job_id}).", "warning")
            return redirect(url_for("performance_backfill", period_type=period_type, cycle_month=cycle_month))

        if request.method == "POST":
            job_id, created = create_background_job(
                g.db,
                "fusionsolar_production_backfill",
                {
                    "provider": INTEGRATION_PROVIDER_FUSIONSOLAR,
                    "period_type": period_type,
                    "from_year": from_year,
                    "to_year": to_year,
                    "asset_id": asset_id,
                    "date_from": date_from.isoformat() if date_from else "",
                    "date_to": date_to.isoformat() if date_to else "",
                    "max_api_calls": max_api_calls,
                },
            )
            g.db.commit()
            if created:
                schedule_background_job(current_app._get_current_object(), job_id)
                flash(f"Backfill historico enviado para background (job #{job_id}).", "success")
            else:
                flash(f"Ja existe um backfill historico pendente/em execucao (job #{job_id}).", "warning")
            return redirect(
                url_for(
                    "performance_backfill",
                    period_type=period_type,
                    from_year=from_year,
                    to_year=to_year,
                    date_from=date_from.isoformat() if date_from else date_from_raw,
                    date_to=date_to.isoformat() if date_to else date_to_raw,
                    max_api_calls=max_api_calls,
                    asset_id=asset_id or "",
                )
            )

        assets_for_backfill = get_fusionsolar_performance_assets(g.db, INTEGRATION_PROVIDER_FUSIONSOLAR)
        return render_template(
            "performance_backfill.html",
            period_type=period_type,
            from_year=from_year,
            to_year=to_year,
            date_from=date_from.isoformat() if date_from else date_from_raw,
            date_to=date_to.isoformat() if date_to else date_to_raw,
            max_api_calls=max_api_calls,
            estimated_api_calls=estimated_api_calls,
            estimated_station_count=estimated_station_count,
            selected_asset_id=asset_id,
            assets_for_backfill=assets_for_backfill,
            background_jobs=fetch_latest_background_jobs(g.db, job_types=BACKGROUND_JOB_TYPES_PERFORMANCE),
            fusionsolar_api_warning=get_fusionsolar_performance_cooldown_reason(g.db),
            current_year=current_year,
            cycle_month=request.args.get("cycle_month", date.today().strftime("%Y-%m")),
        )

    @app.route("/assets", methods=["GET", "POST"])
    def assets() -> str:
        if request.method == "POST":
            project_name = request.form.get("project_name", "").strip()
            installation_group = request.form.get("installation_group", "").strip()
            company_name = request.form.get("company_name", "").strip()
            location = request.form.get("location", "").strip()
            address = request.form.get("address", "").strip()
            kwp = request.form.get("kwp", "").strip()
            contract_type = request.form.get("contract_type", "").strip()
            maintenance = request.form.get("maintenance", "").strip()
            active_contract = request.form.get("active_contract", "").strip()
            start_contract = request.form.get("start_contract", "").strip()
            end_contract = request.form.get("end_contract", "").strip()
            contact_name = request.form.get("contact_name", "").strip()
            contact_email = request.form.get("contact_email", "").strip()
            contact_phone = request.form.get("contact_phone", "").strip()
            notes = request.form.get("notes", "").strip()

            if not project_name:
                flash("O nome da instalacao/central e obrigatorio.", "error")
                return redirect(url_for("assets"))

            existing = query_one("SELECT id FROM assets WHERE project_name = ?", (project_name,))
            if existing is not None:
                flash("Ja existe um asset com esse nome.", "error")
                return redirect(url_for("assets"))

            final_group = installation_group or infer_installation_group(project_name)
            inherited_payload = apply_group_defaults(
                g.db,
                {
                    "company_name": company_name,
                    "location": location,
                    "address": address,
                    "contract_type": contract_type,
                    "contact_name": contact_name,
                    "contact_email": contact_email,
                    "contact_phone": contact_phone,
                },
                final_group,
            )
            company_name = inherited_payload["company_name"]
            location = inherited_payload["location"]
            address = inherited_payload["address"]
            contract_type = inherited_payload["contract_type"]
            contact_name = inherited_payload["contact_name"]
            contact_email = inherited_payload["contact_email"]
            contact_phone = inherited_payload["contact_phone"]
            start_contract = normalize_date_value(start_contract)
            end_contract = normalize_date_value(end_contract)
            active_contract = derive_active_contract(end_contract, active_contract)
            cursor = g.db.execute(
                """
                INSERT INTO assets (
                    project_name, installation_group, company_name, location, address, kwp, contract_type,
                    maintenance, active_contract, start_contract, end_contract, contact_name,
                    contact_email, contact_phone, notes, alias_blob
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    project_name,
                    final_group,
                    company_name,
                    location,
                    address,
                    kwp,
                    contract_type,
                    maintenance,
                    active_contract,
                    start_contract,
                    end_contract,
                    contact_name,
                    contact_email,
                    contact_phone,
                    notes,
                    project_name,
                ),
            )
            asset_id = int(cursor.lastrowid)
            normalized_name = normalize_name(project_name)
            if normalized_name:
                g.db.execute(
                    "INSERT OR IGNORE INTO asset_aliases (asset_id, alias_name, normalized_alias, source) VALUES (?, ?, ?, ?)",
                    (asset_id, project_name, normalized_name, "manual-create"),
                )
            g.db.commit()
            rebuild_asset_alias_blob(g.db, asset_id)
            flash("Instalacao criada com sucesso.", "success")
            return redirect(url_for("asset_detail", asset_id=asset_id))

        search = request.args.get("search", "").strip()
        contract_filter = request.args.get("contract_type", "").strip()
        om_filter = request.args.get("active_contract", "").strip()

        conditions = []
        params: list[Any] = []
        if search:
            conditions.append(
                "(a.project_name LIKE ? OR a.company_name LIKE ? OR a.location LIKE ? OR a.alias_blob LIKE ?)"
            )
            wildcard = f"%{search}%"
            params.extend([wildcard, wildcard, wildcard, wildcard])
        if contract_filter:
            conditions.append("a.contract_type = ?")
            params.append(contract_filter)
        if om_filter:
            conditions.append("a.active_contract = ?")
            params.append(om_filter)

        where_sql = f"WHERE {' AND '.join(conditions)}" if conditions else ""
        assets_rows = query_all(
            g.db,
            f"""
            SELECT
                a.*,
                lm.status AS latest_status,
                lm.record_date AS latest_status_date,
                (
                    SELECT COUNT(*)
                    FROM tickets t
                    WHERE t.asset_id = a.id AND t.status != 'Fechado'
                ) AS open_tickets
            FROM assets a
            LEFT JOIN latest_monitoring_view lm ON lm.asset_id = a.id
            {where_sql}
            ORDER BY a.installation_group COLLATE NOCASE, a.project_name COLLATE NOCASE
            """,
            params,
        )
        contract_types = [
            row["contract_type"]
            for row in query_all(
                g.db,
                "SELECT DISTINCT contract_type FROM assets WHERE contract_type != '' ORDER BY contract_type",
            )
        ]
        return render_template(
            "assets.html",
            assets=assets_rows,
            contract_types=contract_types,
            search=search,
            contract_filter=contract_filter,
            om_filter=om_filter,
        )

    @app.route("/asset/<int:asset_id>")
    def asset_detail(asset_id: int) -> str:
        asset = query_one("SELECT * FROM assets WHERE id = ?", (asset_id,))
        if asset is None:
            flash("Asset nao encontrado.", "error")
            return redirect(url_for("assets"))
        calendar_month = normalize_calendar_month(request.args.get("calendar_month", ""))
        calendar_start, calendar_end, previous_month, next_month = calendar_month_bounds(calendar_month)

        om_contract = query_one(
            """
            SELECT *
            FROM om_contracts
            WHERE asset_id = ?
            """,
            (asset_id,),
        )

        monitoring_history = query_all(
            g.db,
            """
            SELECT id, record_date, status, notes, source
            FROM monitoring_records
            WHERE asset_id = ?
            ORDER BY record_date DESC, id DESC
            LIMIT 100
            """,
            (asset_id,),
        )
        calendar_history = query_all(
            g.db,
            """
            SELECT id, asset_id, record_date, status, notes, source
            FROM monitoring_records
            WHERE asset_id = ?
              AND record_date <= ?
            ORDER BY record_date ASC, id ASC
            """,
            (asset_id, calendar_end.isoformat()),
        )
        asset_error_calendar = build_asset_error_calendar(calendar_month, calendar_history)
        tickets = query_all(
            g.db,
            """
            SELECT *
            FROM tickets
            WHERE asset_id = ?
            ORDER BY updated_at DESC, created_at DESC
            """,
            (asset_id,),
        )
        aliases = query_all(
            g.db,
            """
            SELECT id, alias_name
            FROM asset_aliases
            WHERE asset_id = ?
            ORDER BY alias_name COLLATE NOCASE
            """,
            (asset_id,),
        )
        latest_daily_performance = query_one(
            """
            SELECT *
            FROM production_records
            WHERE asset_id = ? AND period_type = 'day'
            ORDER BY period_date DESC, id DESC
            LIMIT 1
            """,
            (asset_id,),
        )
        latest_monthly_performance = query_one(
            """
            SELECT *
            FROM production_records
            WHERE asset_id = ? AND period_type = 'month' AND period_date < ?
            ORDER BY period_date DESC, id DESC
            LIMIT 1
            """,
            (asset_id, date.today().replace(day=1).isoformat()),
        )
        latest_mtd_performance = query_one(
            """
            SELECT *
            FROM production_records
            WHERE asset_id = ? AND period_type = 'mtd' AND period_date = ?
            ORDER BY id DESC
            LIMIT 1
            """,
            (asset_id, date.today().replace(day=1).isoformat()),
        )
        performance_settings = get_performance_settings(g.db, asset_id)
        latest_availability = get_latest_availability_by_asset(g.db, asset_id)
        latest_device_rows = get_latest_device_rows_for_asset(g.db, asset_id)
        expected_string_rows = query_all(
            g.db,
            """
            SELECT pd.id AS provider_device_id, pes.string_index, pes.expected, pes.source
            FROM provider_devices pd
            LEFT JOIN provider_device_expected_strings pes ON pes.provider_device_id = pd.id
            WHERE pd.asset_id = ? AND pd.enabled = 1
            ORDER BY pd.device_name COLLATE NOCASE, pd.id, pes.string_index
            """,
            (asset_id,),
        )
        expected_strings_by_device: dict[int, list[dict[str, Any]]] = {}
        for row in expected_string_rows:
            if row["string_index"] is None:
                continue
            expected_strings_by_device.setdefault(int(row["provider_device_id"]), []).append(row)
        visits_by_ticket = build_visits_by_ticket(
            query_all(
                g.db,
                """
                SELECT *
                FROM ticket_visits
                WHERE ticket_id IN (
                    SELECT id FROM tickets WHERE asset_id = ?
                )
                ORDER BY visit_date DESC, id DESC
                """,
                (asset_id,),
            )
        )
        return render_template(
            "asset_detail.html",
            asset=asset,
            om_contract=om_contract,
            monitoring_history=monitoring_history,
            asset_error_calendar=asset_error_calendar,
            calendar_month=calendar_month,
            previous_month=previous_month,
            next_month=next_month,
            tickets=tickets,
            aliases=aliases,
            visits_by_ticket=visits_by_ticket,
            latest_daily_performance=latest_daily_performance,
            latest_mtd_performance=latest_mtd_performance,
            latest_monthly_performance=latest_monthly_performance,
            performance_settings=performance_settings,
            latest_availability=latest_availability,
            latest_device_rows=latest_device_rows,
            expected_strings_by_device=expected_strings_by_device,
        )

    @app.route("/asset/<int:asset_id>/performance-settings", methods=["POST"])
    def update_asset_performance_settings(asset_id: int):
        asset = query_one("SELECT id FROM assets WHERE id = ?", (asset_id,))
        if asset is None:
            flash("Asset nao encontrado.", "error")
            return redirect(url_for("assets"))

        monthly_budget_json = request.form.get("monthly_budget_json", "").strip()
        if monthly_budget_json:
            try:
                payload = json.loads(monthly_budget_json)
                if not isinstance(payload, dict):
                    raise ValueError
            except (json.JSONDecodeError, ValueError):
                flash("Orçamento mensal inválido. Usa JSON com meses 01-12 e valores kWh/kWp.", "error")
                return redirect(url_for("asset_detail", asset_id=asset_id))

        def threshold(name: str, default: float) -> float:
            value = parse_float_value(request.form.get(name, ""))
            return value if value is not None else default

        now = datetime.now().isoformat(timespec="seconds")
        g.db.execute(
            """
            INSERT INTO performance_settings (
                asset_id, enabled, warning_deviation_pct, alert_deviation_pct, critical_deviation_pct,
                baseline_years, min_baseline_points, monthly_budget_json, notes, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(asset_id) DO UPDATE SET
                enabled = excluded.enabled,
                warning_deviation_pct = excluded.warning_deviation_pct,
                alert_deviation_pct = excluded.alert_deviation_pct,
                critical_deviation_pct = excluded.critical_deviation_pct,
                baseline_years = excluded.baseline_years,
                min_baseline_points = excluded.min_baseline_points,
                monthly_budget_json = excluded.monthly_budget_json,
                notes = excluded.notes,
                updated_at = excluded.updated_at
            """,
            (
                asset_id,
                1 if request.form.get("enabled") == "on" else 0,
                threshold("warning_deviation_pct", -10),
                threshold("alert_deviation_pct", -20),
                threshold("critical_deviation_pct", -30),
                int(parse_float_value(request.form.get("baseline_years", "")) or 2),
                int(parse_float_value(request.form.get("min_baseline_points", "")) or 1),
                monthly_budget_json,
                request.form.get("notes", "").strip(),
                now,
            ),
        )
        g.db.commit()
        flash("Definições de performance guardadas.", "success")
        return redirect(url_for("asset_detail", asset_id=asset_id))

    @app.route("/asset/<int:asset_id>/expected-strings", methods=["POST"])
    def update_asset_expected_strings(asset_id: int):
        device_id_raw = request.form.get("provider_device_id", "").strip()
        if not device_id_raw.isdigit():
            flash("Dispositivo invalido.", "error")
            return redirect(url_for("asset_detail", asset_id=asset_id))
        device_id = int(device_id_raw)
        device = query_one(
            "SELECT id FROM provider_devices WHERE id = ? AND asset_id = ?",
            (device_id, asset_id),
        )
        if device is None:
            flash("Dispositivo nao encontrado.", "error")
            return redirect(url_for("asset_detail", asset_id=asset_id))
        selected = {int(value) for value in request.form.getlist("expected_strings") if value.isdigit()}
        now = datetime.now().isoformat(timespec="seconds")
        for index in range(1, 37):
            existing = query_one(
                """
                SELECT *
                FROM provider_device_expected_strings
                WHERE provider_device_id = ? AND string_index = ?
                """,
                (device_id, index),
            )
            expected = 1 if index in selected else 0
            if existing:
                g.db.execute(
                    """
                    UPDATE provider_device_expected_strings
                    SET expected = ?, source = 'manual', updated_at = ?
                    WHERE id = ?
                    """,
                    (expected, now, existing["id"]),
                )
            elif expected:
                g.db.execute(
                    """
                    INSERT INTO provider_device_expected_strings (
                        provider_device_id, string_index, expected, source, observed_count,
                        created_at, updated_at
                    ) VALUES (?, ?, 1, 'manual', 0, ?, ?)
                    """,
                    (device_id, index, now, now),
                )
        g.db.commit()
        flash("Perfil de strings atualizado.", "success")
        return redirect(url_for("asset_detail", asset_id=asset_id))

    @app.route("/asset/<int:asset_id>/installation-group", methods=["POST"])
    def update_asset_installation_group(asset_id: int):
        asset = query_one("SELECT id, project_name FROM assets WHERE id = ?", (asset_id,))
        if asset is None:
            flash("Asset nao encontrado.", "error")
            return redirect(url_for("assets"))

        group_name = request.form.get("installation_group", "").strip()
        if not group_name:
            group_name = infer_installation_group(asset["project_name"])

        g.db.execute("UPDATE assets SET installation_group = ? WHERE id = ?", (group_name, asset_id))
        apply_group_defaults_to_asset(g.db, asset_id, group_name)
        g.db.commit()
        flash("Grupo de instalacao atualizado.", "success")
        return redirect(url_for("asset_detail", asset_id=asset_id))

    @app.route("/asset/<int:asset_id>/update", methods=["POST"])
    def update_asset(asset_id: int):
        asset = query_one("SELECT id FROM assets WHERE id = ?", (asset_id,))
        if asset is None:
            flash("Asset nao encontrado.", "error")
            return redirect(url_for("assets"))

        payload = {
            "project_name": request.form.get("project_name", "").strip(),
            "installation_group": request.form.get("installation_group", "").strip(),
            "company_name": request.form.get("company_name", "").strip(),
            "location": request.form.get("location", "").strip(),
            "address": request.form.get("address", "").strip(),
            "contract_type": request.form.get("contract_type", "").strip(),
            "maintenance": request.form.get("maintenance", "").strip(),
            "active_contract": request.form.get("active_contract", "").strip(),
            "start_contract": request.form.get("start_contract", "").strip(),
            "end_contract": request.form.get("end_contract", "").strip(),
            "contact_name": request.form.get("contact_name", "").strip(),
            "contact_email": request.form.get("contact_email", "").strip(),
            "contact_phone": request.form.get("contact_phone", "").strip(),
            "notes": request.form.get("notes", "").strip(),
            "monitoring_enabled": 1 if request.form.get("monitoring_enabled") == "on" else 0,
            "alerts_enabled": 1 if request.form.get("alerts_enabled") == "on" else 0,
            "selected_for_alerts": 1 if request.form.get("selected_for_alerts") == "on" else 0,
            "monitoring_status": request.form.get("monitoring_status", "active").strip() or "active",
            "silenced_until": request.form.get("silenced_until", "").strip(),
            "silence_reason": request.form.get("silence_reason", "").strip(),
        }
        if not payload["project_name"]:
            flash("O nome da central e obrigatorio.", "error")
            return redirect(url_for("asset_detail", asset_id=asset_id))
        if not payload["installation_group"]:
            payload["installation_group"] = infer_installation_group(payload["project_name"])
        payload = apply_group_defaults(g.db, payload, payload["installation_group"], exclude_asset_id=asset_id)
        payload["start_contract"] = normalize_date_value(payload["start_contract"])
        payload["end_contract"] = normalize_date_value(payload["end_contract"])
        payload["active_contract"] = derive_active_contract(payload["end_contract"], payload["active_contract"])
        if payload["monitoring_status"] not in ASSET_MONITORING_STATUSES:
            payload["monitoring_status"] = "active"
        if payload["monitoring_status"] != "silenced":
            payload["silenced_until"] = ""

        g.db.execute(
            """
            UPDATE assets
            SET project_name = ?, installation_group = ?, company_name = ?, location = ?, address = ?,
                contract_type = ?, maintenance = ?, active_contract = ?, start_contract = ?, end_contract = ?,
                contact_name = ?, contact_email = ?, contact_phone = ?, notes = ?,
                monitoring_enabled = ?, alerts_enabled = ?, selected_for_alerts = ?, monitoring_status = ?, silenced_until = ?, silence_reason = ?
            WHERE id = ?
            """,
            (
                payload["project_name"],
                payload["installation_group"],
                payload["company_name"],
                payload["location"],
                payload["address"],
                payload["contract_type"],
                payload["maintenance"],
                payload["active_contract"],
                payload["start_contract"],
                payload["end_contract"],
                payload["contact_name"],
                payload["contact_email"],
                payload["contact_phone"],
                payload["notes"],
                payload["monitoring_enabled"],
                payload["alerts_enabled"],
                payload["selected_for_alerts"],
                payload["monitoring_status"],
                payload["silenced_until"],
                payload["silence_reason"],
                asset_id,
            ),
        )
        g.db.commit()
        flash("Asset atualizado.", "success")
        return redirect(url_for("asset_detail", asset_id=asset_id))

    @app.route("/asset/<int:asset_id>/delete", methods=["POST"])
    def delete_asset(asset_id: int):
        asset = query_one("SELECT id, project_name FROM assets WHERE id = ?", (asset_id,))
        if asset is None:
            flash("Asset nao encontrado.", "error")
            return redirect(url_for("assets"))

        contract = query_one("SELECT pdf_path FROM om_contracts WHERE asset_id = ?", (asset_id,))
        if contract and contract["pdf_path"]:
            contract_path = resolve_runtime_file_path_within(contract["pdf_path"], CONTRACTS_DIR)
            if contract_path is not None and contract_path.exists():
                contract_path.unlink()

        g.db.execute("DELETE FROM assets WHERE id = ?", (asset_id,))
        g.db.commit()
        flash(f"Asset '{asset['project_name']}' apagado.", "success")
        return redirect(url_for("assets"))

    @app.route("/asset/<int:asset_id>/contract", methods=["POST"])
    def update_asset_contract(asset_id: int):
        asset = query_one("SELECT id, project_name FROM assets WHERE id = ?", (asset_id,))
        if asset is None:
            flash("Asset nao encontrado.", "error")
            return redirect(url_for("assets"))

        start_date = normalize_date_value(request.form.get("contract_start_date", "").strip())
        end_date = normalize_date_value(request.form.get("contract_end_date", "").strip())
        annual_value_raw = request.form.get("annual_value", "").strip()
        contract_notes = request.form.get("contract_notes", "").strip()
        uploaded_file = request.files.get("contract_pdf")

        annual_value = None
        if annual_value_raw:
            normalized_value = annual_value_raw.replace(" ", "").replace(",", ".")
            try:
                annual_value = float(normalized_value)
            except ValueError:
                flash("O valor anual do contrato nao e valido.", "error")
                return redirect(url_for("asset_detail", asset_id=asset_id))

        existing_contract = query_one(
            """
            SELECT *
            FROM om_contracts
            WHERE asset_id = ?
            """,
            (asset_id,),
        )

        stored_path = existing_contract["pdf_path"] if existing_contract else ""
        original_filename = existing_contract["original_filename"] if existing_contract else ""
        if uploaded_file and uploaded_file.filename:
            suffix = Path(uploaded_file.filename).suffix.lower()
            if suffix != ".pdf":
                flash("O contrato tem de ser um ficheiro PDF.", "error")
                return redirect(url_for("asset_detail", asset_id=asset_id))
            header = uploaded_file.stream.read(5)
            uploaded_file.stream.seek(0)
            if header != b"%PDF-":
                flash("O contrato enviado nao parece ser um PDF valido.", "error")
                return redirect(url_for("asset_detail", asset_id=asset_id))
            CONTRACTS_DIR.mkdir(parents=True, exist_ok=True)
            safe_stem = normalize_name(asset["project_name"]).replace(" ", "-") or f"asset-{asset_id}"
            filename = f"{asset_id}_{safe_stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            target_path = (CONTRACTS_DIR / filename).resolve()
            if not path_is_within(target_path, CONTRACTS_DIR):
                current_app.logger.error("Rejected contract upload outside contracts directory for asset %s", asset_id)
                abort(400)
            uploaded_file.save(target_path)
            if stored_path:
                old_path = resolve_runtime_file_path_within(stored_path, CONTRACTS_DIR)
                if old_path is not None and old_path.exists() and old_path != target_path:
                    old_path.unlink()
            stored_path = store_runtime_relative_path(target_path)
            original_filename = Path(uploaded_file.filename).name[:255]

        if existing_contract:
            g.db.execute(
                """
                UPDATE om_contracts
                SET contract_start_date = ?, contract_end_date = ?, annual_value = ?, notes = ?,
                    pdf_path = ?, original_filename = ?, updated_at = ?
                WHERE asset_id = ?
                """,
                (
                    start_date,
                    end_date,
                    annual_value,
                    contract_notes,
                    stored_path,
                    original_filename,
                    datetime.now().isoformat(timespec="seconds"),
                    asset_id,
                ),
            )
        else:
            g.db.execute(
                """
                INSERT INTO om_contracts (
                    asset_id, contract_start_date, contract_end_date, annual_value, notes, pdf_path,
                    original_filename, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    asset_id,
                    start_date,
                    end_date,
                    annual_value,
                    contract_notes,
                    stored_path,
                    original_filename,
                    datetime.now().isoformat(timespec="seconds"),
                    datetime.now().isoformat(timespec="seconds"),
                ),
            )
        sync_asset_contract_status(g.db, asset_id, start_date, end_date)
        g.db.commit()
        flash("Contrato O&M atualizado.", "success")
        return redirect(url_for("asset_detail", asset_id=asset_id))

    @app.route("/asset/<int:asset_id>/contract/open")
    def open_asset_contract(asset_id: int):
        contract = query_one(
            """
            SELECT pdf_path
            FROM om_contracts
            WHERE asset_id = ?
            """,
            (asset_id,),
        )
        if contract is None or not contract["pdf_path"]:
            flash("Esta central ainda nao tem contrato associado.", "error")
            return redirect(url_for("asset_detail", asset_id=asset_id))

        contract_path = resolve_runtime_file_path_within(contract["pdf_path"], CONTRACTS_DIR)
        if contract_path is None:
            current_app.logger.warning("Blocked contract path outside contracts directory for asset %s", asset_id)
            abort(404)
        if not contract_path.exists():
            flash("O ficheiro do contrato nao foi encontrado no projeto.", "error")
            return redirect(url_for("asset_detail", asset_id=asset_id))

        return send_file(contract_path, mimetype="application/pdf", as_attachment=False, max_age=0)

    @app.route("/asset/<int:asset_id>/alias", methods=["POST"])
    def add_alias(asset_id: int):
        alias_name = request.form.get("alias_name", "").strip()
        if not alias_name:
            flash("Indica um nome alternativo para guardar o alias.", "error")
            return redirect(url_for("asset_detail", asset_id=asset_id))

        normalized = normalize_name(alias_name)
        if not normalized:
            flash("O alias indicado nao e valido.", "error")
            return redirect(url_for("asset_detail", asset_id=asset_id))

        existing = query_one("SELECT id FROM asset_aliases WHERE normalized_alias = ?", (normalized,))
        if existing:
            flash("Esse alias ja esta associado a outra instalacao.", "error")
            return redirect(url_for("asset_detail", asset_id=asset_id))

        g.db.execute(
            "INSERT INTO asset_aliases (asset_id, alias_name, normalized_alias, source) VALUES (?, ?, ?, ?)",
            (asset_id, alias_name, normalized, "manual"),
        )
        g.db.commit()
        rebuild_asset_alias_blob(g.db, asset_id)
        flash("Alias guardado com sucesso.", "success")
        return redirect(url_for("asset_detail", asset_id=asset_id))

    @app.route("/asset/<int:asset_id>/alias/<int:alias_id>/update", methods=["POST"])
    def update_alias(asset_id: int, alias_id: int):
        alias_name = request.form.get("alias_name", "").strip()
        if not alias_name:
            flash("Indica um nome valido para o alias.", "error")
            return redirect(url_for("asset_detail", asset_id=asset_id))

        alias_row = query_one(
            "SELECT * FROM asset_aliases WHERE id = ? AND asset_id = ?",
            (alias_id, asset_id),
        )
        if alias_row is None:
            flash("Alias nao encontrado.", "error")
            return redirect(url_for("asset_detail", asset_id=asset_id))

        normalized = normalize_name(alias_name)
        existing = query_one("SELECT id, asset_id FROM asset_aliases WHERE normalized_alias = ?", (normalized,))
        if existing and (existing["id"] != alias_id or existing["asset_id"] != asset_id):
            flash("Esse alias ja esta associado a outra instalacao.", "error")
            return redirect(url_for("asset_detail", asset_id=asset_id))

        g.db.execute(
            """
            UPDATE asset_aliases
            SET alias_name = ?, normalized_alias = ?, source = ?
            WHERE id = ? AND asset_id = ?
            """,
            (alias_name, normalized, "manual-edit", alias_id, asset_id),
        )
        g.db.commit()
        rebuild_asset_alias_blob(g.db, asset_id)
        flash("Alias atualizado.", "success")
        return redirect(url_for("asset_detail", asset_id=asset_id))

    @app.route("/asset/<int:asset_id>/alias/<int:alias_id>/delete", methods=["POST"])
    def delete_alias(asset_id: int, alias_id: int):
        alias_row = query_one(
            "SELECT * FROM asset_aliases WHERE id = ? AND asset_id = ?",
            (alias_id, asset_id),
        )
        if alias_row is None:
            flash("Alias nao encontrado.", "error")
            return redirect(url_for("asset_detail", asset_id=asset_id))

        g.db.execute("DELETE FROM asset_aliases WHERE id = ? AND asset_id = ?", (alias_id, asset_id))
        g.db.commit()
        rebuild_asset_alias_blob(g.db, asset_id)
        flash("Alias apagado.", "success")
        return redirect(url_for("asset_detail", asset_id=asset_id))

    @app.route("/monitoring", methods=["GET", "POST"])
    def monitoring() -> str:
        if request.method == "POST":
            record_date = request.form.get("record_date", date.today().isoformat())
            pasted_table = request.form.get("pasted_table", "")
            default_notes = request.form.get("default_notes", "").strip()
            platform_source = request.form.get("platform_source", "Manual / Outro").strip() or "Manual / Outro"
            import_scope = request.form.get("import_scope", "complete").strip() or "complete"
            result = import_daily_monitoring(
                g.db,
                pasted_table,
                record_date,
                default_notes,
                platform_source,
                import_scope,
            )
            flash(
                f"Importacao concluida: {result.imported} registos, {result.matched} associados automaticamente, {result.unmatched} por mapear, {result.auto_resolved} resolvidos automaticamente.",
                "success" if result.imported else "warning",
            )
            if result.batch_id is not None:
                return redirect(url_for("monitoring", batch_id=result.batch_id))
            return redirect(url_for("monitoring"))

        search = request.args.get("search", "").strip()
        asset_filter = request.args.get("asset_id", "").strip()
        status_filter = request.args.get("status", "").strip()
        source_filter = request.args.get("source", "").strip()
        issue_only = request.args.get("issue_only", "no").strip()
        start_date = request.args.get("start_date", "").strip()
        end_date = request.args.get("end_date", "").strip()
        om_only = request.args.get("om_only", "yes").strip()
        batch_id = request.args.get("batch_id", "").strip()

        latest_conditions = []
        latest_params: list[Any] = []
        if search:
            wildcard = f"%{search}%"
            latest_conditions.append("(a.project_name LIKE ? OR a.alias_blob LIKE ? OR a.company_name LIKE ?)")
            latest_params.extend([wildcard, wildcard, wildcard])
        if asset_filter:
            latest_conditions.append("a.id = ?")
            latest_params.append(asset_filter)
        if status_filter:
            latest_conditions.append("lm.status = ?")
            latest_params.append(status_filter)
        elif issue_only == "yes":
            latest_conditions.append("lm.status IN ('Erro', 'Desconectada')")
        if source_filter:
            latest_conditions.append(
                "EXISTS (SELECT 1 FROM monitoring_records src WHERE src.asset_id = a.id AND src.record_date = lm.record_date AND src.status = lm.status AND src.source = ?)"
            )
            latest_params.append(source_filter)
        if om_only == "yes":
            latest_conditions.append("a.active_contract = 'yes'")
        if start_date:
            latest_conditions.append("lm.record_date >= ?")
            latest_params.append(start_date)
        if end_date:
            latest_conditions.append("lm.record_date <= ?")
            latest_params.append(end_date)

        latest_where_sql = f"WHERE {' AND '.join(latest_conditions)}" if latest_conditions else ""
        latest_rows = enrich_operational_rows(g.db, query_all(
            g.db,
            f"""
            SELECT
                a.id AS asset_id,
                a.project_name,
                a.installation_group,
                a.location,
                a.contract_type,
                a.active_contract,
                lm.status,
                lm.record_date,
                lm.notes,
                (
                    SELECT COUNT(*)
                    FROM monitoring_records mr
                    WHERE mr.asset_id = a.id
                ) AS history_count
            FROM latest_monitoring_view lm
            JOIN assets a ON a.id = lm.asset_id
            {latest_where_sql}
            ORDER BY
                CASE a.active_contract WHEN 'yes' THEN 1 ELSE 2 END,
                CASE lm.status
                    WHEN 'Erro' THEN 1
                    WHEN 'Desconectada' THEN 2
                    ELSE 3
                END,
                a.project_name COLLATE NOCASE
            """,
            latest_params,
        ))
        latest_availability_rows = {
            int(row["asset_id"]): row
            for row in query_all(
                g.db,
                """
                SELECT ad.*
                FROM availability_daily ad
                JOIN (
                    SELECT asset_id, MAX(period_date || 'T' || printf('%09d', id)) AS marker
                    FROM availability_daily
                    GROUP BY asset_id
                ) latest
                  ON latest.asset_id = ad.asset_id
                 AND latest.marker = ad.period_date || 'T' || printf('%09d', ad.id)
                """,
            )
        }
        for row in latest_rows:
            row["availability"] = latest_availability_rows.get(int(row["asset_id"]))

        filter_sql = []
        params: list[Any] = []
        if search:
            wildcard = f"%{search}%"
            filter_sql.append("(a.project_name LIKE ? OR a.alias_blob LIKE ? OR a.company_name LIKE ?)")
            params.extend([wildcard, wildcard, wildcard])
        if asset_filter:
            filter_sql.append("a.id = ?")
            params.append(asset_filter)
        if status_filter:
            filter_sql.append("mr.status = ?")
            params.append(status_filter)
        elif issue_only == "yes":
            filter_sql.append("mr.status IN ('Erro', 'Desconectada')")
        if source_filter:
            filter_sql.append("mr.source = ?")
            params.append(source_filter)
        if om_only == "yes":
            filter_sql.append("a.active_contract = 'yes'")
        if start_date:
            filter_sql.append("mr.record_date >= ?")
            params.append(start_date)
        if end_date:
            filter_sql.append("mr.record_date <= ?")
            params.append(end_date)

        where_sql = f"WHERE {' AND '.join(filter_sql)}" if filter_sql else ""
        history_rows = query_all(
            g.db,
            f"""
            SELECT
                mr.id,
                mr.record_date,
                mr.status,
                mr.notes,
                mr.source,
                mr.batch_id,
                a.id AS asset_id,
                a.project_name,
                a.installation_group,
                a.location,
                a.contract_type,
                a.active_contract,
                mib.imported_at
            FROM monitoring_records mr
            JOIN assets a ON a.id = mr.asset_id
            LEFT JOIN monitoring_import_batches mib ON mib.id = mr.batch_id
            {where_sql}
            ORDER BY mr.record_date DESC, a.project_name COLLATE NOCASE, mr.id DESC
            LIMIT 250
            """,
            params,
        )

        selected_asset = None
        asset_history = []
        asset_problem_periods = []
        selected_group_assets = []
        if asset_filter:
            selected_asset = query_one(
                """
                SELECT
                    a.*,
                    lm.status AS latest_status,
                    lm.record_date AS latest_status_date
                FROM assets a
                LEFT JOIN latest_monitoring_view lm ON lm.asset_id = a.id
                WHERE a.id = ?
                """,
                (asset_filter,),
            )
            asset_history = query_all(
                g.db,
                """
                SELECT mr.id, mr.record_date, mr.status, mr.notes, mr.source, mr.batch_id, mib.imported_at
                FROM monitoring_records mr
                LEFT JOIN monitoring_import_batches mib ON mib.id = mr.batch_id
                WHERE mr.asset_id = ?
                ORDER BY mr.record_date DESC, mr.id DESC
                LIMIT 200
                """,
                (asset_filter,),
            )
            asset_problem_periods = build_problem_periods(g.db, int(asset_filter))
            selected_group_assets = query_all(
                g.db,
                """
                SELECT
                    a.id,
                    a.project_name,
                    a.installation_group,
                    lm.status,
                    lm.record_date
                FROM assets a
                LEFT JOIN latest_monitoring_view lm ON lm.asset_id = a.id
                WHERE a.installation_group = ?
                ORDER BY a.project_name COLLATE NOCASE
                """,
                (selected_asset["installation_group"] or selected_asset["project_name"],),
            )

        unresolved = query_all(
            g.db,
            """
            SELECT mu.*, mib.imported_at
            FROM monitoring_unmatched mu
            LEFT JOIN monitoring_import_batches mib ON mib.id = mu.batch_id
            ORDER BY record_date DESC, original_name COLLATE NOCASE
            LIMIT 50
            """
        )
        import_batches = query_all(
            g.db,
            """
            SELECT
                mib.*,
                (
                    SELECT COUNT(*)
                    FROM monitoring_records mr
                    WHERE mr.batch_id = mib.id
                ) AS resolved_rows
            FROM monitoring_import_batches mib
            ORDER BY mib.imported_at DESC, mib.id DESC
            LIMIT 30
            """
        )
        assets_for_mapping = query_all(
            g.db,
            "SELECT id, project_name FROM assets ORDER BY project_name COLLATE NOCASE",
        )
        monitoring_statuses = [
            row["status"]
            for row in query_all(
                g.db,
                "SELECT DISTINCT status FROM monitoring_records ORDER BY status",
            )
        ]
        monitoring_stats = {
            "history_records": len(history_rows),
            "centrals_in_filter": len(latest_rows),
            "installations_in_filter": len(group_latest_rows_by_installation(latest_rows)),
            "current_errors": sum(1 for row in latest_rows if row["status"] == "Erro"),
            "current_disconnected": sum(1 for row in latest_rows if row["status"] == "Desconectada"),
            "current_active_om": sum(1 for row in latest_rows if row["active_contract"] == "yes"),
            "critical_priority": sum(1 for row in latest_rows if row.get("auto_priority") == "Critica"),
            "high_priority": sum(1 for row in latest_rows if row.get("auto_priority") == "Alta"),
            "recurring_90d": sum(1 for row in latest_rows if int(row.get("recurrence_count") or 0) >= 2),
        }
        grouped_latest_rows = group_latest_rows_by_installation(latest_rows)
        batch_insight = build_batch_insight(g.db, int(batch_id)) if batch_id else None

        return render_template(
            "monitoring.html",
            latest_rows=latest_rows,
            grouped_latest_rows=grouped_latest_rows,
            unresolved=unresolved,
            assets_for_mapping=assets_for_mapping,
            monitoring_statuses=monitoring_statuses,
            monitoring_stats=monitoring_stats,
            history_rows=history_rows,
            selected_asset=selected_asset,
            selected_group_assets=selected_group_assets,
            asset_history=asset_history,
            asset_problem_periods=asset_problem_periods,
            import_batches=import_batches,
            search=search,
            asset_filter=asset_filter,
            status_filter=status_filter,
            source_filter=source_filter,
            issue_only=issue_only,
            start_date=start_date,
            end_date=end_date,
            om_only=om_only,
            batch_insight=batch_insight,
        )

    @app.route("/monitoring/record/<int:record_id>/update", methods=["POST"])
    def update_monitoring_record(record_id: int):
        record = query_one("SELECT asset_id FROM monitoring_records WHERE id = ?", (record_id,))
        if record is None:
            flash("Registo de monitorizacao nao encontrado.", "error")
            return redirect(url_for("monitoring"))

        record_date = request.form.get("record_date", "").strip()
        status = normalize_status(request.form.get("status", "").strip())
        notes = request.form.get("notes", "").strip()
        if not record_date or not status:
            flash("Data e estado sao obrigatorios para atualizar o registo.", "error")
            return redirect(url_for("monitoring", asset_id=record["asset_id"]))

        g.db.execute(
            """
            UPDATE monitoring_records
            SET record_date = ?, status = ?, notes = ?
            WHERE id = ?
            """,
            (record_date, status, notes, record_id),
        )
        g.db.commit()
        flash("Registo de monitorizacao atualizado.", "success")
        return redirect(url_for("monitoring", asset_id=record["asset_id"]))

    @app.route("/monitoring/record/<int:record_id>/delete", methods=["POST"])
    def delete_monitoring_record(record_id: int):
        record = query_one("SELECT asset_id FROM monitoring_records WHERE id = ?", (record_id,))
        if record is None:
            flash("Registo de monitorizacao nao encontrado.", "error")
            return redirect(url_for("monitoring"))

        g.db.execute("DELETE FROM monitoring_records WHERE id = ?", (record_id,))
        g.db.commit()
        flash("Registo de monitorizacao apagado.", "success")
        return redirect(url_for("monitoring", asset_id=record["asset_id"]))

    @app.route("/monitoring/unmatched/<int:row_id>/resolve", methods=["POST"])
    def resolve_unmatched(row_id: int):
        asset_id = int(request.form["asset_id"])
        unmatched = query_one("SELECT * FROM monitoring_unmatched WHERE id = ?", (row_id,))
        if unmatched is None:
            flash("Linha pendente nao encontrada.", "error")
            return redirect(url_for("monitoring"))

        alias_name = unmatched["original_name"]
        normalized = normalize_name(alias_name)
        existing = query_one("SELECT id FROM asset_aliases WHERE normalized_alias = ?", (normalized,))
        if existing and existing["id"]:
            flash("Esse nome ja esta usado como alias.", "error")
            return redirect(url_for("monitoring"))

        g.db.execute(
            "INSERT INTO asset_aliases (asset_id, alias_name, normalized_alias, source) VALUES (?, ?, ?, ?)",
            (asset_id, alias_name, normalized, "resolved"),
        )
        g.db.execute(
            """
            INSERT INTO monitoring_records (asset_id, status, record_date, notes, source, batch_id)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                asset_id,
                unmatched["status"],
                unmatched["record_date"],
                unmatched["notes"],
                "resolved-unmatched",
                unmatched["batch_id"],
            ),
        )
        g.db.execute("DELETE FROM monitoring_unmatched WHERE id = ?", (row_id,))
        g.db.commit()
        rebuild_asset_alias_blob(g.db, asset_id)
        flash("Linha associada e importada com sucesso.", "success")
        return redirect(url_for("monitoring", asset_id=asset_id))

    @app.route("/tickets", methods=["GET", "POST"])
    def tickets() -> str:
        if request.method == "POST":
            asset_id = int(request.form["asset_id"])
            title = request.form.get("title", "").strip()
            urgency = request.form.get("urgency", "Media")
            status = request.form.get("status", "Aberto")
            installation_ref = request.form.get("installation_ref", "").strip()
            notes = request.form.get("notes", "").strip()
            next_action = request.form.get("next_action", "").strip()
            planned_date = normalize_optional_date(request.form.get("planned_date"))
            due_date = normalize_optional_date(request.form.get("due_date"))
            estimated_minutes = parse_positive_int(request.form.get("estimated_minutes"), default=60)
            assigned_to = request.form.get("assigned_to", "").strip()
            material_status = normalize_choice(
                request.form.get("material_status", "Nao definido"),
                TICKET_MATERIAL_STATUSES,
                "Nao definido",
            )
            work_type = normalize_choice(request.form.get("work_type", "Diagnostico"), TICKET_WORK_TYPES, "Diagnostico")
            planning_notes = request.form.get("planning_notes", "").strip()
            if not title:
                flash("A intervencao precisa de um titulo.", "error")
                return redirect(url_for("tickets"))

            g.db.execute(
                """
                INSERT INTO tickets (
                    asset_id, title, urgency, status, installation_ref, notes, next_action,
                    planned_date, due_date, estimated_minutes, assigned_to, material_status,
                    work_type, planning_notes, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    asset_id,
                    title,
                    urgency,
                    status,
                    installation_ref,
                    notes,
                    next_action,
                    planned_date,
                    due_date,
                    estimated_minutes,
                    assigned_to,
                    material_status,
                    work_type,
                    planning_notes,
                    datetime.now().isoformat(timespec="seconds"),
                    datetime.now().isoformat(timespec="seconds"),
                ),
            )
            g.db.commit()
            flash("Intervencao criada.", "success")
            return redirect(url_for("tickets", asset_id=asset_id))

        search = request.args.get("search", "").strip()
        asset_filter = request.args.get("asset_id", "").strip()
        status_filter = request.args.get("status", "").strip()
        urgency_filter = request.args.get("urgency", "").strip()
        scope = request.args.get("scope", "").strip()
        om_only = request.args.get("om_only", "yes").strip()
        calendar_month = normalize_calendar_month(request.args.get("calendar_month", ""))

        conditions = []
        params: list[Any] = []
        if search:
            wildcard = f"%{search}%"
            conditions.append(
                "(a.project_name LIKE ? OR a.alias_blob LIKE ? OR t.title LIKE ? OR COALESCE(t.notes, '') LIKE ?)"
            )
            params.extend([wildcard, wildcard, wildcard, wildcard])
        if asset_filter:
            conditions.append("a.id = ?")
            params.append(asset_filter)
        if status_filter:
            conditions.append("t.status = ?")
            params.append(status_filter)
        if urgency_filter:
            conditions.append("t.urgency = ?")
            params.append(urgency_filter)
        if scope == "open":
            conditions.append("t.status != 'Fechado'")
        if om_only == "yes":
            conditions.append("a.active_contract = 'yes'")

        where_sql = f"WHERE {' AND '.join(conditions)}" if conditions else ""
        ticket_rows = query_all(
            g.db,
            f"""
            SELECT
                t.*,
                a.project_name,
                a.location,
                a.active_contract,
                a.contract_type
            FROM tickets t
            JOIN assets a ON a.id = t.asset_id
            {where_sql}
            ORDER BY
                CASE a.active_contract WHEN 'yes' THEN 1 ELSE 2 END,
                a.project_name COLLATE NOCASE,
                CASE t.status
                    WHEN 'Aberto' THEN 1
                    WHEN 'Em analise' THEN 2
                    WHEN 'Agendado' THEN 3
                    WHEN 'Em visita' THEN 4
                    WHEN 'Resolvido' THEN 5
                    ELSE 6
                END,
                CASE t.urgency
                    WHEN 'Critica' THEN 1
                    WHEN 'Alta' THEN 2
                    WHEN 'Media' THEN 3
                    ELSE 4
                END,
                t.updated_at DESC,
                t.id DESC
            """,
            params,
        )

        assets_rows = query_all(
            g.db,
            "SELECT id, project_name FROM assets ORDER BY project_name COLLATE NOCASE",
        )
        visits_by_ticket = build_visits_by_ticket(
            query_all(
                g.db,
                """
                SELECT *
                FROM ticket_visits
                ORDER BY visit_date DESC, id DESC
                """,
            )
        )
        grouped_tickets = group_tickets_by_asset(ticket_rows)
        calendar_start, calendar_end, previous_month, next_month = calendar_month_bounds(calendar_month)
        calendar_conditions = [
            "mr.status IN ('Erro', 'Desconectada')",
            "mr.record_date BETWEEN ? AND ?",
        ]
        calendar_params: list[Any] = [calendar_start.isoformat(), calendar_end.isoformat()]
        if search:
            wildcard = f"%{search}%"
            calendar_conditions.append(
                "(a.project_name LIKE ? OR a.alias_blob LIKE ? OR COALESCE(mr.notes, '') LIKE ? OR COALESCE(mr.source, '') LIKE ?)"
            )
            calendar_params.extend([wildcard, wildcard, wildcard, wildcard])
        if asset_filter:
            calendar_conditions.append("a.id = ?")
            calendar_params.append(asset_filter)
        if om_only == "yes":
            calendar_conditions.append("a.active_contract = 'yes'")
        calendar_where_sql = f"WHERE {' AND '.join(calendar_conditions)}"

        calendar_rows = query_all(
            g.db,
            f"""
            SELECT
                mr.id,
                mr.asset_id,
                mr.status,
                mr.record_date,
                mr.notes,
                mr.source,
                a.project_name
            FROM monitoring_records mr
            JOIN assets a ON a.id = mr.asset_id
            {calendar_where_sql}
            ORDER BY
                mr.record_date,
                CASE mr.status
                    WHEN 'Erro' THEN 1
                    WHEN 'Desconectada' THEN 2
                    ELSE 3
                END,
                a.project_name COLLATE NOCASE,
                mr.id DESC
            """,
            calendar_params,
        )
        error_calendar = build_error_calendar(calendar_month, calendar_rows)

        selected_asset = None
        central_history = []
        central_summary = None
        if asset_filter:
            selected_asset = query_one("SELECT * FROM assets WHERE id = ?", (asset_filter,))
            central_history = query_all(
                g.db,
                """
                SELECT
                    t.*,
                    (
                        SELECT COUNT(*)
                        FROM ticket_visits tv
                        WHERE tv.ticket_id = t.id
                    ) AS visit_count
                FROM tickets t
                WHERE t.asset_id = ?
                ORDER BY t.updated_at DESC, t.id DESC
                """,
                (asset_filter,),
            )
            central_summary = {
                "total": len(central_history),
                "open": sum(1 for ticket in central_history if ticket["status"] != "Fechado"),
                "critical": sum(1 for ticket in central_history if ticket["urgency"] == "Critica" and ticket["status"] != "Fechado"),
                "visits": sum(ticket["visit_count"] for ticket in central_history),
            }

        ticket_stats = {
            "centrals": len(grouped_tickets),
            "tickets": len(ticket_rows),
            "open": sum(1 for ticket in ticket_rows if ticket["status"] != "Fechado"),
            "critical": sum(1 for ticket in ticket_rows if ticket["urgency"] == "Critica" and ticket["status"] != "Fechado"),
        }

        return render_template(
            "tickets.html",
            tickets=ticket_rows,
            grouped_tickets=grouped_tickets,
            assets=assets_rows,
            visits_by_ticket=visits_by_ticket,
            selected_asset=selected_asset,
            central_history=central_history,
            central_summary=central_summary,
            ticket_stats=ticket_stats,
            error_calendar=error_calendar,
            calendar_month=calendar_month,
            previous_month=previous_month,
            next_month=next_month,
            search=search,
            asset_filter=asset_filter,
            status_filter=status_filter,
            urgency_filter=urgency_filter,
            scope=scope,
            om_only=om_only,
        )

    @app.route("/tickets/<int:ticket_id>/update", methods=["POST"])
    def update_ticket(ticket_id: int):
        ticket = query_one("SELECT asset_id FROM tickets WHERE id = ?", (ticket_id,))
        status = request.form.get("status", "Aberto")
        urgency = request.form.get("urgency", "Media")
        next_action = request.form.get("next_action", "").strip()
        notes = request.form.get("notes", "").strip()
        planned_date = normalize_optional_date(request.form.get("planned_date"))
        due_date = normalize_optional_date(request.form.get("due_date"))
        estimated_minutes = parse_positive_int(request.form.get("estimated_minutes"), default=60)
        assigned_to = request.form.get("assigned_to", "").strip()
        material_status = normalize_choice(
            request.form.get("material_status", "Nao definido"),
            TICKET_MATERIAL_STATUSES,
            "Nao definido",
        )
        work_type = normalize_choice(request.form.get("work_type", "Diagnostico"), TICKET_WORK_TYPES, "Diagnostico")
        planning_notes = request.form.get("planning_notes", "").strip()
        g.db.execute(
            """
            UPDATE tickets
            SET status = ?, urgency = ?, next_action = ?, notes = ?,
                planned_date = ?, due_date = ?, estimated_minutes = ?, assigned_to = ?,
                material_status = ?, work_type = ?, planning_notes = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                status,
                urgency,
                next_action,
                notes,
                planned_date,
                due_date,
                estimated_minutes,
                assigned_to,
                material_status,
                work_type,
                planning_notes,
                datetime.now().isoformat(timespec="seconds"),
                ticket_id,
            ),
        )
        g.db.commit()
        flash("Intervencao atualizada.", "success")
        if ticket:
            return redirect(url_for("tickets", asset_id=ticket["asset_id"]))
        return redirect(url_for("tickets"))

    @app.route("/tickets/<int:ticket_id>/visit", methods=["POST"])
    def add_visit(ticket_id: int):
        ticket = query_one("SELECT asset_id FROM tickets WHERE id = ?", (ticket_id,))
        visit_date = request.form.get("visit_date", date.today().isoformat())
        technician = request.form.get("technician", "").strip()
        result = request.form.get("result", "").strip()
        notes = request.form.get("notes", "").strip()
        next_action = request.form.get("next_action", "").strip()

        g.db.execute(
            """
            INSERT INTO ticket_visits (ticket_id, visit_date, technician, result, notes, next_action)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (ticket_id, visit_date, technician, result, notes, next_action),
        )
        if next_action:
            g.db.execute(
                "UPDATE tickets SET next_action = ?, updated_at = ? WHERE id = ?",
                (next_action, datetime.now().isoformat(timespec="seconds"), ticket_id),
            )
        g.db.commit()
        flash("Visita registada.", "success")
        if ticket:
            return redirect(url_for("tickets", asset_id=ticket["asset_id"]))
        return redirect(url_for("tickets"))

    @app.route("/tickets/<int:ticket_id>/delete", methods=["POST"])
    def delete_ticket(ticket_id: int):
        ticket = query_one("SELECT asset_id FROM tickets WHERE id = ?", (ticket_id,))
        if ticket is None:
            flash("Intervencao nao encontrada.", "error")
            return redirect(url_for("tickets"))

        g.db.execute("DELETE FROM tickets WHERE id = ?", (ticket_id,))
        g.db.commit()
        flash("Intervencao apagada.", "success")
        return redirect(url_for("tickets", asset_id=ticket["asset_id"]))

    @app.route("/exports", methods=["GET", "POST"])
    def exports() -> str:
        if request.method == "POST":
            asset_id_raw = request.form.get("asset_id", "").strip()
            report_month = normalize_report_month(request.form.get("report_month", ""))
            electricity_price_value = parse_float_value(request.form.get("electricity_price"))
            electricity_price = 0.20725 if electricity_price_value is None else max(electricity_price_value, 0.0)
            sell_price = max(parse_float_value(request.form.get("sell_price")) or 0.0, 0.0)
            solcor_price_per_kwh = max(parse_float_value(request.form.get("solcor_price_per_kwh")) or 0.0, 0.0)
            force_api = request.form.get("force_api") == "on"
            if not asset_id_raw.isdigit():
                flash("Escolhe uma instalacao FusionSolar para gerar o relatorio.", "error")
                return redirect(url_for("exports", report_month=report_month))
            try:
                report = build_fusionsolar_customer_production_report(
                    g.db,
                    asset_id=int(asset_id_raw),
                    report_month=report_month,
                    electricity_price=electricity_price,
                    sell_price=sell_price,
                    solcor_price_per_kwh=solcor_price_per_kwh,
                    force_api=force_api,
                )
                return export_customer_production_pdf(report)
            except Exception as exc:
                if is_fusionsolar_rate_limit_error(exc):
                    flash(mark_fusionsolar_performance_rate_limited(g.db), "warning")
                    g.db.commit()
                    return redirect(url_for("exports", asset_id=asset_id_raw, report_month=report_month))
                if "FusionSolar API temporariamente limitada" in str(exc):
                    flash(str(exc), "warning")
                    return redirect(url_for("exports", asset_id=asset_id_raw, report_month=report_month))
                current_app.logger.exception("Falha ao gerar relatorio de producao")
                flash(f"Falha ao gerar relatorio de producao: {exc}", "error")
                return redirect(url_for("exports", asset_id=asset_id_raw, report_month=report_month))

        selected_asset_id = request.args.get("asset_id", "").strip()
        report_month = normalize_report_month(request.args.get("report_month", ""))
        report_assets = get_fusionsolar_report_assets(g.db)
        selected_report_type = next(
            (
                asset["report_type"]
                for asset in report_assets
                if selected_asset_id.isdigit() and int(asset["asset_id"]) == int(selected_asset_id)
            ),
            "",
        )

        return render_template(
            "exports.html",
            report_assets=report_assets,
            selected_asset_id=selected_asset_id,
            selected_report_type=selected_report_type,
            report_month=report_month,
            electricity_price=request.args.get("electricity_price", "0.20725"),
            sell_price=request.args.get("sell_price", "0.00"),
            solcor_price_per_kwh=request.args.get("solcor_price_per_kwh", "0.00"),
            fusionsolar_api_warning=get_fusionsolar_performance_cooldown_reason(g.db),
        )

    @app.route("/integrations", methods=["GET", "POST"])
    def integrations() -> str:
        provider = INTEGRATION_PROVIDER_FUSIONSOLAR
        if request.method == "POST":
            action = request.form.get("action", "").strip()
            if action == "save_config":
                sync_hours = normalize_sync_hours(request.form.get("sync_hours", DEFAULT_FUSIONSOLAR_SYNC_HOURS))
                auto_sync_enabled = 1 if request.form.get("auto_sync_enabled") == "on" else 0
                enabled = 1 if request.form.get("enabled") == "on" else 0
                submitted_password = request.form.get("password", "").strip()
                g.db.execute(
                    """
                    UPDATE integration_configs
                    SET username = ?,
                        password = CASE WHEN ? != '' THEN ? ELSE password END,
                        base_url = ?, login_endpoint = ?, plants_endpoint = ?,
                        real_time_endpoint = ?, device_list_endpoint = ?, device_real_time_endpoint = ?,
                        device_history_endpoint = ?, alarms_endpoint = ?,
                        day_kpi_endpoint = ?, month_kpi_endpoint = ?,
                        enabled = ?, auto_sync_enabled = ?, sync_hours = ?, updated_at = ?
                    WHERE provider = ?
                    """,
                    (
                        request.form.get("username", "").strip(),
                        submitted_password,
                        submitted_password,
                        request.form.get("base_url", "").strip(),
                        request.form.get("login_endpoint", "").strip(),
                        request.form.get("plants_endpoint", "").strip(),
                        request.form.get("real_time_endpoint", "").strip(),
                        request.form.get("device_list_endpoint", "").strip(),
                        request.form.get("device_real_time_endpoint", "").strip(),
                        request.form.get("device_history_endpoint", "").strip(),
                        request.form.get("alarms_endpoint", "").strip(),
                        request.form.get("day_kpi_endpoint", "").strip(),
                        request.form.get("month_kpi_endpoint", "").strip(),
                        enabled,
                        auto_sync_enabled,
                        sync_hours,
                        datetime.now().isoformat(timespec="seconds"),
                        provider,
                    ),
                )
                g.db.commit()
                refresh_integration_scheduler(app)
                flash("Configuracao FusionSolar guardada.", "success")
                return redirect(url_for("integrations"))

            if action == "test_connection":
                messages = []
                failures = []
                for item_provider in INTEGRATION_PROVIDER_OPTIONS:
                    item_config = get_integration_config(g.db, item_provider)
                    if item_config is None or not item_config["enabled"]:
                        continue
                    try:
                        result = run_provider_check(g.db, item_provider, dry_run=True)
                        messages.append(
                            f"{item_provider}: {result['station_count']} centrais, {result['realtime_count']} respostas realtime"
                        )
                    except Exception as exc:
                        failures.append(f"{item_provider}: {exc}")
                if messages:
                    flash("Ligacao validada: " + " | ".join(messages), "success")
                if failures:
                    flash("Falha no teste: " + " | ".join(failures), "error")
                if not messages and not failures:
                    flash("Nao ha integracoes ativas para testar.", "warning")
                return redirect(url_for("integrations"))

            if action == "test_fusionsolar_connection":
                try:
                    result = run_fusionsolar_check(g.db, provider, dry_run=True)
                    flash(
                        f"Ligacao FusionSolar validada: {result['station_count']} centrais, {result['realtime_count']} respostas realtime e {result['alarm_count']} alarmes ativos.",
                        "success",
                    )
                except Exception as exc:
                    flash(f"Falha no teste de ligacao FusionSolar: {exc}", "error")
                return redirect(url_for("integrations"))

            if action == "sync_now":
                try:
                    result = run_all_integration_syncs(g.db, trigger_type="manual")
                    summaries = [
                        f"{item_provider}: {item_result['matched']} associados, {item_result['unresolved']} por resolver, {item_result['auto_resolved']} resolvidos"
                        for item_provider, item_result in result["results"].items()
                    ]
                    if summaries:
                        flash("Sync concluido: " + " | ".join(summaries), "success")
                    else:
                        flash("Nao ha integracoes ativas para sincronizar.", "warning")
                    if result["errors"]:
                        flash("Falhas no sync: " + " | ".join(f"{key}: {value}" for key, value in result["errors"].items()), "error")
                except Exception as exc:
                    flash(f"Falha ao sincronizar integracoes: {exc}", "error")
                return redirect(url_for("integrations"))

            if action == "test_telegram":
                ok, message = test_telegram_connection()
                flash(message, "success" if ok else "error")
                return redirect(url_for("integrations"))

            if action == "save_alert_settings":
                set_alert_setting(g.db, "TELEGRAM_ALERTS_ENABLED", "true" if request.form.get("telegram_alerts_enabled") == "on" else "false")
                alert_scope = request.form.get("alert_scope", "only_o&m").strip()
                if alert_scope not in ALERT_SCOPE_OPTIONS:
                    alert_scope = "only_o&m"
                set_alert_setting(g.db, "ALERT_SCOPE", alert_scope)
                for key in [
                    "SEND_NEW_ERROR_ALERTS",
                    "SEND_OFFLINE_ALERTS",
                    "SEND_RESOLVED_ALERTS",
                    "SEND_PERSISTENT_ALERTS",
                    "SEND_RECURRENT_ALERTS",
                    "DAYTIME_OFFLINE_ONLY",
                    "IGNORE_HISTORICAL_ALERTS",
                ]:
                    set_alert_setting(g.db, key, "true" if request.form.get(key) == "on" else "false")
                for key in [
                    "MINIMUM_ALERT_SEVERITY",
                    "NEW_ERROR_COOLDOWN_MINUTES",
                    "OFFLINE_COOLDOWN_MINUTES",
                    "RESOLVED_COOLDOWN_MINUTES",
                    "PERSISTENT_COOLDOWN_HOURS",
                    "RECURRENT_COOLDOWN_HOURS",
                ]:
                    set_alert_setting(g.db, key, request.form.get(key, ALERT_SETTING_DEFAULTS.get(key, "")).strip())
                g.db.commit()
                flash("Filtros de alertas guardados.", "success")
                return redirect(url_for("integrations"))

            if action == "set_alert_baseline":
                baseline_at = datetime.now().isoformat(timespec="seconds")
                set_alert_setting(g.db, "ALERT_BASELINE_AT", baseline_at)
                g.db.execute(
                    "INSERT INTO alert_baseline (baseline_at, created_by, notes) VALUES (?, ?, ?)",
                    (baseline_at, session.get("username"), "Baseline definido pela UI."),
                )
                g.db.commit()
                flash("Estado atual definido como baseline de alertas.", "success")
                return redirect(url_for("integrations"))

            if action == "add_alert_blacklist":
                asset_id_raw = request.form.get("asset_id", "").strip()
                reason = request.form.get("reason", "").strip()
                asset_id = int(asset_id_raw) if asset_id_raw else None
                asset_name = request.form.get("asset_name", "").strip()
                if asset_id:
                    asset = query_one("SELECT project_name FROM assets WHERE id = ?", (asset_id,))
                    asset_name = asset["project_name"] if asset else asset_name
                if asset_id or asset_name:
                    g.db.execute(
                        "INSERT INTO alert_blacklist (asset_id, asset_name, reason, created_at, active) VALUES (?, ?, ?, ?, 1)",
                        (asset_id, asset_name, reason, datetime.now().isoformat(timespec="seconds")),
                    )
                    g.db.commit()
                    flash("Instalacao adicionada a blacklist de alertas.", "success")
                return redirect(url_for("integrations"))

            if action == "remove_alert_blacklist":
                blacklist_id = int(request.form["blacklist_id"])
                g.db.execute("UPDATE alert_blacklist SET active = 0 WHERE id = ?", (blacklist_id,))
                g.db.commit()
                flash("Instalacao removida da blacklist.", "success")
                return redirect(url_for("integrations"))

            if action == "quick_alert_action":
                asset_id = int(request.form["asset_id"])
                quick_action = request.form.get("quick_action", "")
                if quick_action == "disable_alerts":
                    g.db.execute("UPDATE assets SET alerts_enabled = 0 WHERE id = ?", (asset_id,))
                elif quick_action == "enable_alerts":
                    g.db.execute("UPDATE assets SET alerts_enabled = 1, monitoring_enabled = 1, monitoring_status = 'active' WHERE id = ?", (asset_id,))
                elif quick_action == "blacklist":
                    asset = query_one("SELECT project_name FROM assets WHERE id = ?", (asset_id,))
                    g.db.execute(
                        "INSERT INTO alert_blacklist (asset_id, asset_name, reason, created_at, active) VALUES (?, ?, ?, ?, 1)",
                        (asset_id, asset["project_name"] if asset else "", "Adicionado por acao rapida.", datetime.now().isoformat(timespec="seconds")),
                    )
                elif quick_action == "unblacklist":
                    g.db.execute("UPDATE alert_blacklist SET active = 0 WHERE asset_id = ?", (asset_id,))
                elif quick_action == "out_of_scope":
                    g.db.execute("UPDATE assets SET monitoring_status = 'out_of_scope' WHERE id = ?", (asset_id,))
                elif quick_action == "silence_24h":
                    g.db.execute(
                        "UPDATE assets SET monitoring_status = 'silenced', silenced_until = ?, silence_reason = ? WHERE id = ?",
                        ((datetime.now() + timedelta(hours=24)).isoformat(timespec="minutes"), "Silenciado 24h pela UI.", asset_id),
                    )
                elif quick_action == "reactivate":
                    g.db.execute("UPDATE assets SET monitoring_status = 'active', silenced_until = '', silence_reason = '' WHERE id = ?", (asset_id,))
                g.db.commit()
                flash("Filtro da instalacao atualizado.", "success")
                return redirect(url_for("integrations"))

            if action == "bulk_alert_action":
                bulk_action = request.form.get("bulk_action", "")
                if bulk_action == "blacklist_non_oem":
                    rows = query_all(g.db, "SELECT id, project_name FROM assets WHERE COALESCE(maintenance, '') NOT IN ('yes', 'true', '1', 'sim')")
                    for row in rows:
                        g.db.execute(
                            "INSERT INTO alert_blacklist (asset_id, asset_name, reason, created_at, active) VALUES (?, ?, ?, ?, 1)",
                            (row["id"], row["project_name"], "Sem Maintenance=yes.", datetime.now().isoformat(timespec="seconds")),
                        )
                elif bulk_action == "disable_no_active_contract":
                    g.db.execute("UPDATE assets SET alerts_enabled = 0 WHERE COALESCE(active_contract, '') != 'yes'")
                elif bulk_action == "enable_only_oem":
                    g.db.execute("UPDATE assets SET alerts_enabled = CASE WHEN COALESCE(maintenance, '') = 'yes' THEN 1 ELSE 0 END")
                    set_alert_setting(g.db, "ALERT_SCOPE", "only_o&m")
                elif bulk_action == "set_baseline":
                    baseline_at = datetime.now().isoformat(timespec="seconds")
                    set_alert_setting(g.db, "ALERT_BASELINE_AT", baseline_at)
                    g.db.execute(
                        "INSERT INTO alert_baseline (baseline_at, created_by, notes) VALUES (?, ?, ?)",
                        (baseline_at, session.get("username"), "Baseline definido por acao em massa."),
                    )
                g.db.commit()
                flash("Acao em massa aplicada.", "success")
                return redirect(url_for("integrations"))

            if action == "resolve_unresolved":
                unresolved_id = int(request.form["unresolved_id"])
                asset_id = int(request.form["asset_id"])
                resolve_fusionsolar_unresolved(g.db, unresolved_id, asset_id)
                flash("Entrada API associada ao asset.", "success")
                return redirect(url_for("integrations") + "#integrations-link-audit")

            if action == "update_fusionsolar_mapping":
                integration_id = int(request.form["integration_id"])
                asset_id = int(request.form["asset_id"])
                update_fusionsolar_mapping_asset(g.db, integration_id, asset_id)
                flash("Mapeamento FusionSolar atualizado.", "success")
                return redirect(url_for("integrations") + "#integrations-link-audit")

            if action == "create_asset_from_unresolved":
                unresolved_id = int(request.form["unresolved_id"])
                asset_id = create_asset_from_unresolved(g.db, unresolved_id)
                flash("Asset criado a partir da entrada API por resolver.", "success")
                return redirect(url_for("asset_detail", asset_id=asset_id))

            if action == "ignore_unresolved":
                unresolved_id = int(request.form["unresolved_id"])
                ignore_fusionsolar_unresolved(g.db, unresolved_id)
                flash("Entrada API marcada como ignorada.", "success")
                return redirect(url_for("integrations"))

        config = get_integration_config(g.db, provider)
        integration_configs = [get_integration_config(g.db, item_provider) for item_provider in INTEGRATION_PROVIDER_OPTIONS]
        integration_configs = [item for item in integration_configs if item is not None]
        sync_runs = query_all(
            g.db,
            """
            SELECT *
            FROM integration_sync_runs
            WHERE provider IN (?, ?)
            ORDER BY started_at DESC, id DESC
            LIMIT 20
            """,
            (INTEGRATION_PROVIDER_FUSIONSOLAR, INTEGRATION_PROVIDER_SIGENERGY),
        )
        unresolved_rows = query_all(
            g.db,
            """
            SELECT *
            FROM integration_unresolved
            WHERE provider IN (?, ?) AND resolution_status = 'pending'
            ORDER BY created_at DESC, id DESC
            LIMIT 100
            """,
            (INTEGRATION_PROVIDER_FUSIONSOLAR, INTEGRATION_PROVIDER_SIGENERGY),
        )
        mapped_assets = query_all(
            g.db,
            """
            SELECT ai.*, a.project_name, a.installation_group
            FROM asset_integrations ai
            JOIN assets a ON a.id = ai.asset_id
            WHERE ai.provider IN (?, ?)
            ORDER BY a.installation_group COLLATE NOCASE, a.project_name COLLATE NOCASE
            """,
            (INTEGRATION_PROVIDER_FUSIONSOLAR, INTEGRATION_PROVIDER_SIGENERGY),
        )
        link_audit_rows = get_fusionsolar_link_audit_rows(g.db, provider)
        link_audit_counts = {
            "ok": sum(1 for row in link_audit_rows if row["verdict"] == "OK"),
            "attention": sum(1 for row in link_audit_rows if row["verdict"] == "Atencao"),
            "review": sum(1 for row in link_audit_rows if row["verdict"] == "Rever"),
            "unresolved": sum(1 for row in link_audit_rows if row["verdict"] == "Por resolver"),
        }
        assets_for_mapping = query_all(g.db, "SELECT id, project_name FROM assets ORDER BY project_name COLLATE NOCASE")
        alert_filter_assets = query_all(
            g.db,
            """
            SELECT
                a.id,
                a.project_name,
                a.maintenance,
                a.active_contract,
                a.alerts_enabled,
                a.monitoring_enabled,
                a.monitoring_status,
                a.selected_for_alerts,
                lm.status AS latest_status,
                MAX(ta.sent_at) AS last_alert_sent,
                CASE WHEN ab.id IS NULL THEN 0 ELSE 1 END AS blacklisted
            FROM assets a
            LEFT JOIN latest_monitoring_view lm ON lm.asset_id = a.id
            LEFT JOIN telegram_alerts ta ON ta.asset_id = a.id AND ta.status = 'sent'
            LEFT JOIN alert_blacklist ab ON ab.asset_id = a.id AND ab.active = 1
            GROUP BY a.id
            ORDER BY a.project_name COLLATE NOCASE
            LIMIT 200
            """,
        )
        alert_blacklist_rows = query_all(
            g.db,
            """
            SELECT ab.*, a.project_name
            FROM alert_blacklist ab
            LEFT JOIN assets a ON a.id = ab.asset_id
            WHERE ab.active = 1
            ORDER BY ab.created_at DESC, ab.id DESC
            LIMIT 100
            """,
        )
        return render_template(
            "integrations.html",
            provider=provider,
            config=config,
            integration_configs=integration_configs,
            sync_runs=sync_runs,
            unresolved_rows=unresolved_rows,
            mapped_assets=mapped_assets,
            link_audit_rows=link_audit_rows,
            link_audit_counts=link_audit_counts,
            assets_for_mapping=assets_for_mapping,
            telegram_config=get_telegram_config(),
            alert_settings=get_alert_settings(g.db),
            alert_scope_options=ALERT_SCOPE_OPTIONS,
            alert_filter_assets=alert_filter_assets,
            alert_blacklist_rows=alert_blacklist_rows,
            fusionsolar_api_warning=get_fusionsolar_performance_cooldown_reason(g.db),
        )

    @app.route("/telegram-alerts")
    def telegram_alerts() -> str:
        status_filter = request.args.get("status", "").strip()
        asset_filter = request.args.get("asset_id", "").strip()
        alert_type_filter = request.args.get("alert_type", "").strip()
        blocked_reason_filter = request.args.get("blocked_reason", "").strip()
        conditions = []
        params: list[Any] = []
        if status_filter:
            conditions.append("ta.status = ?")
            params.append(status_filter)
        if asset_filter:
            conditions.append("ta.asset_id = ?")
            params.append(asset_filter)
        if alert_type_filter:
            conditions.append("ta.alert_type = ?")
            params.append(alert_type_filter)
        if blocked_reason_filter:
            conditions.append("ta.blocked_reason = ?")
            params.append(blocked_reason_filter)
        where_sql = f"WHERE {' AND '.join(conditions)}" if conditions else ""
        rows = query_all(
            g.db,
            f"""
            SELECT ta.*, a.project_name
            FROM telegram_alerts ta
            LEFT JOIN assets a ON a.id = ta.asset_id
            {where_sql}
            ORDER BY ta.sent_at DESC, ta.id DESC
            LIMIT 250
            """,
            params,
        )
        alert_types = [row["alert_type"] for row in query_all(g.db, "SELECT DISTINCT alert_type FROM telegram_alerts ORDER BY alert_type")]
        blocked_reasons = [row["blocked_reason"] for row in query_all(g.db, "SELECT DISTINCT blocked_reason FROM telegram_alerts WHERE blocked_reason IS NOT NULL AND blocked_reason != '' ORDER BY blocked_reason")]
        assets_for_mapping = query_all(
            g.db,
            """
            SELECT DISTINCT a.id, a.project_name
            FROM assets a
            JOIN telegram_alerts ta ON ta.asset_id = a.id
            ORDER BY a.project_name COLLATE NOCASE
            """,
        )
        return render_template(
            "telegram_alerts.html",
            alerts=rows,
            status_filter=status_filter,
            asset_filter=asset_filter,
            alert_type_filter=alert_type_filter,
            blocked_reason_filter=blocked_reason_filter,
            alert_types=alert_types,
            blocked_reasons=blocked_reasons,
            assets_for_mapping=assets_for_mapping,
        )

    @app.route("/renewals", methods=["GET", "POST"])
    def renewals() -> str:
        focus = request.args.get("focus", "").strip()
        if request.method == "POST":
            asset_id = int(request.form["asset_id"])
            renewal_status = request.form.get("renewal_status", "Por contactar").strip() or "Por contactar"
            last_contact_date = request.form.get("last_contact_date", "").strip()
            renewal_notes = request.form.get("renewal_notes", "").strip()
            annual_value_raw = request.form.get("annual_value", "").strip()
            contract_end_date = normalize_date_value(request.form.get("contract_end_date", "").strip())
            contract_start_date = normalize_date_value(request.form.get("contract_start_date", "").strip())

            annual_value = None
            if annual_value_raw:
                normalized_value = annual_value_raw.replace(" ", "").replace(",", ".")
                try:
                    annual_value = float(normalized_value)
                except ValueError:
                    flash("O valor anual nao e valido.", "error")
                    return redirect(url_for("renewals"))

            existing_contract = query_one("SELECT id FROM om_contracts WHERE asset_id = ?", (asset_id,))
            if existing_contract:
                g.db.execute(
                    """
                    UPDATE om_contracts
                    SET renewal_status = ?, last_contact_date = ?, renewal_notes = ?,
                        annual_value = COALESCE(?, annual_value),
                        contract_start_date = CASE WHEN ? != '' THEN ? ELSE contract_start_date END,
                        contract_end_date = CASE WHEN ? != '' THEN ? ELSE contract_end_date END,
                        updated_at = ?
                    WHERE asset_id = ?
                    """,
                    (
                        renewal_status,
                        last_contact_date,
                        renewal_notes,
                        annual_value,
                        contract_start_date,
                        contract_start_date,
                        contract_end_date,
                        contract_end_date,
                        datetime.now().isoformat(timespec="seconds"),
                        asset_id,
                    ),
                )
            else:
                g.db.execute(
                    """
                    INSERT INTO om_contracts (
                        asset_id, contract_start_date, contract_end_date, annual_value, renewal_status, last_contact_date,
                        renewal_notes, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        asset_id,
                        contract_start_date,
                        contract_end_date,
                        annual_value,
                        renewal_status,
                        last_contact_date,
                        renewal_notes,
                        datetime.now().isoformat(timespec="seconds"),
                        datetime.now().isoformat(timespec="seconds"),
                    ),
                )

            if renewal_status == "Renovado":
                g.db.execute(
                    """
                    UPDATE assets
                    SET maintenance = 'yes',
                        active_contract = 'yes',
                        start_contract = CASE WHEN ? != '' THEN ? ELSE start_contract END,
                        end_contract = CASE WHEN ? != '' THEN ? ELSE end_contract END
                    WHERE id = ?
                    """,
                    (
                        contract_start_date,
                        contract_start_date,
                        contract_end_date,
                        contract_end_date,
                        asset_id,
                    ),
                )
                sync_asset_contract_status(g.db, asset_id, contract_start_date, contract_end_date)
            g.db.commit()
            flash("Follow-up de renovacao atualizado.", "success")
            return redirect(url_for("renewals"))

        today_iso = date.today().isoformat()
        year_end = f"{date.today().year}-12-31"
        renewal_rows = query_all(
            g.db,
            """
            SELECT
                a.id AS asset_id,
                a.project_name,
                a.installation_group,
                a.company_name,
                a.location,
                a.address,
                a.contact_name,
                a.contact_email,
                a.contact_phone,
                a.active_contract,
                COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) AS contract_end_date,
                COALESCE(NULLIF(oc.contract_start_date, ''), NULLIF(a.start_contract, '')) AS contract_start_date,
                oc.annual_value,
                oc.pdf_path,
                oc.renewal_status,
                oc.last_contact_date,
                oc.renewal_notes,
                CASE
                    WHEN COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) < ? THEN 'Expirado'
                    WHEN julianday(COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, ''))) - julianday(?) <= 30 THEN '0-30 dias'
                    WHEN julianday(COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, ''))) - julianday(?) <= 90 THEN '31-90 dias'
                    ELSE 'Este ano'
                END AS renewal_bucket
            FROM assets a
            LEFT JOIN om_contracts oc ON oc.asset_id = a.id
            WHERE (a.maintenance = 'yes' OR oc.id IS NOT NULL)
              AND COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) NOT IN ('', '-')
            ORDER BY COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) ASC, a.project_name COLLATE NOCASE
            """,
            (today_iso, today_iso, today_iso),
        )
        expired_contracts = [
            row for row in renewal_rows
            if row["contract_end_date"] < today_iso
        ]
        ending_this_year = [
            row for row in renewal_rows
            if today_iso <= row["contract_end_date"] <= year_end
        ]
        if focus == "expired":
            ending_this_year = []
        elif focus == "year":
            expired_contracts = []
        elif focus == "90":
            expired_contracts = []
            ending_this_year = [row for row in ending_this_year if row["renewal_bucket"] in {"0-30 dias", "31-90 dias"}]
        renewal_metrics = {
            "expired": len(expired_contracts),
            "next_30_days": sum(1 for row in renewal_rows if row["renewal_bucket"] == "0-30 dias"),
            "next_90_days": sum(1 for row in renewal_rows if row["renewal_bucket"] == "31-90 dias"),
            "this_year": len(ending_this_year),
        }
        return render_template(
            "renewals.html",
            expired_contracts=expired_contracts,
            ending_this_year=ending_this_year,
            renewal_metrics=renewal_metrics,
            focus=focus,
            today_iso=today_iso,
        )

    @app.route("/settings", methods=["GET", "POST"])
    def settings() -> str:
        if request.method == "POST":
            excel_path = request.form.get("excel_path", "").strip()
            if not excel_path:
                flash("Indica o caminho do Excel.", "error")
                return redirect(url_for("settings"))
            excel_file = Path(excel_path)
            if not excel_file.exists() or not excel_file.is_file():
                flash("O ficheiro Excel indicado nao existe ou nao esta acessivel.", "error")
                return redirect(url_for("settings"))
            if excel_file.suffix.lower() not in {".xlsx", ".xlsm"}:
                flash("Indica um ficheiro Excel valido (.xlsx ou .xlsm).", "error")
                return redirect(url_for("settings"))

            app.config["EXCEL_PATH"] = excel_path
            backup_path = create_database_backup(Path(app.config["DATABASE"]), BACKUP_DIR)
            try:
                imported = import_excel_data(g.db, excel_file)
            except Exception as exc:
                flash(f"Falha ao importar o Excel: {exc}", "error")
                flash(f"A base de dados ficou salvaguardada no backup {backup_path.name}.", "warning")
                return redirect(url_for("settings"))
            flash(
                f"Importacao concluida. {imported['assets']} assets, {imported['monitoring']} linhas de monitorizacao e {imported['tickets']} tickets importados.",
                "success",
            )
            flash(
                f"Backup automatico criado antes da reimportacao: {backup_path.name}",
                "warning",
            )
            return redirect(url_for("settings"))

        db_info = {
            "assets": query_scalar(g.db, "SELECT COUNT(*) FROM assets"),
            "monitoring": query_scalar(g.db, "SELECT COUNT(*) FROM monitoring_records"),
            "tickets": query_scalar(g.db, "SELECT COUNT(*) FROM tickets"),
            "aliases": query_scalar(g.db, "SELECT COUNT(*) FROM asset_aliases"),
        }
        excel_path = app.config["EXCEL_PATH"]
        return render_template("settings.html", db_info=db_info, excel_path=excel_path)

    return app


def ensure_database(path: str) -> None:
    with closing(get_db(path)) as conn:
        configure_database_for_runtime(conn)
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS assets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_number TEXT,
                project_name TEXT NOT NULL,
                installation_group TEXT,
                company_name TEXT,
                nif TEXT,
                address TEXT,
                location TEXT,
                panels TEXT,
                kwp TEXT,
                contract_type TEXT,
                sell_to TEXT,
                duration TEXT,
                start_contract TEXT,
                maintenance TEXT,
                coverage_type TEXT,
                access_type TEXT,
                maintenance_comment TEXT,
                status_detail TEXT,
                contact_name TEXT,
                contact_role TEXT,
                contact_email TEXT,
                contact_phone TEXT,
                end_contract TEXT,
                active_contract TEXT,
                notes TEXT,
                asset_type TEXT,
                source_payload TEXT,
                alias_blob TEXT DEFAULT '',
                monitoring_enabled INTEGER DEFAULT 1,
                alerts_enabled INTEGER DEFAULT 1,
                monitoring_status TEXT DEFAULT 'active',
                silenced_until TEXT,
                silence_reason TEXT
            );

            CREATE TABLE IF NOT EXISTS asset_aliases (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL,
                alias_name TEXT NOT NULL,
                normalized_alias TEXT NOT NULL UNIQUE,
                source TEXT,
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS monitoring_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL,
                status TEXT NOT NULL,
                record_date TEXT NOT NULL,
                notes TEXT,
                source TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS monitoring_unmatched (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                original_name TEXT NOT NULL,
                normalized_name TEXT NOT NULL,
                status TEXT NOT NULL,
                record_date TEXT NOT NULL,
                notes TEXT
            );

            CREATE TABLE IF NOT EXISTS monitoring_import_batches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                record_date TEXT NOT NULL,
                imported_at TEXT NOT NULL,
                source TEXT NOT NULL,
                default_notes TEXT,
                raw_input TEXT,
                imported_count INTEGER DEFAULT 0,
                matched_count INTEGER DEFAULT 0,
                unmatched_count INTEGER DEFAULT 0,
                auto_resolved_count INTEGER DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS export_templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                dataset TEXT NOT NULL,
                export_format TEXT NOT NULL,
                columns_json TEXT NOT NULL,
                filters_json TEXT NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS om_contracts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL UNIQUE,
                contract_start_date TEXT,
                contract_end_date TEXT,
                annual_value REAL,
                notes TEXT,
                pdf_path TEXT,
                original_filename TEXT,
                renewal_status TEXT,
                last_contact_date TEXT,
                renewal_notes TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS integration_configs (
                provider TEXT PRIMARY KEY,
                username TEXT,
                password TEXT,
                base_url TEXT,
                login_endpoint TEXT,
                plants_endpoint TEXT,
                real_time_endpoint TEXT,
                device_list_endpoint TEXT,
                device_real_time_endpoint TEXT,
                device_history_endpoint TEXT,
                alarms_endpoint TEXT,
                day_kpi_endpoint TEXT,
                month_kpi_endpoint TEXT,
                enabled INTEGER DEFAULT 0,
                auto_sync_enabled INTEGER DEFAULT 0,
                sync_hours TEXT,
                last_sync_at TEXT,
                last_sync_status TEXT,
                last_error TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS asset_integrations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL,
                provider TEXT NOT NULL,
                external_id TEXT,
                external_name TEXT,
                enabled INTEGER DEFAULT 1,
                last_sync_at TEXT,
                last_status TEXT,
                last_error TEXT,
                UNIQUE(provider, external_id),
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS integration_sync_runs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                provider TEXT NOT NULL,
                started_at TEXT NOT NULL,
                finished_at TEXT,
                trigger_type TEXT,
                status TEXT,
                matched_count INTEGER DEFAULT 0,
                unresolved_count INTEGER DEFAULT 0,
                auto_resolved_count INTEGER DEFAULT 0,
                error_message TEXT,
                summary_json TEXT
            );

            CREATE TABLE IF NOT EXISTS integration_unresolved (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                provider TEXT NOT NULL,
                sync_run_id INTEGER,
                external_id TEXT,
                external_name TEXT NOT NULL,
                normalized_name TEXT NOT NULL,
                external_status TEXT,
                payload_json TEXT,
                suggested_asset_id INTEGER,
                resolution_status TEXT DEFAULT 'pending',
                created_at TEXT NOT NULL,
                resolved_at TEXT,
                resolution_notes TEXT,
                FOREIGN KEY (sync_run_id) REFERENCES integration_sync_runs(id) ON DELETE CASCADE,
                FOREIGN KEY (suggested_asset_id) REFERENCES assets(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS provider_devices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL,
                provider TEXT NOT NULL,
                station_code TEXT NOT NULL,
                external_device_id TEXT,
                dev_dn TEXT,
                sn TEXT,
                device_name TEXT,
                dev_type_id INTEGER,
                model TEXT,
                rated_power_kw REAL,
                enabled INTEGER DEFAULT 1,
                last_seen_at TEXT,
                payload_json TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(provider, external_device_id),
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS device_realtime_snapshots (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                provider_device_id INTEGER NOT NULL,
                asset_id INTEGER NOT NULL,
                provider TEXT NOT NULL,
                station_code TEXT NOT NULL,
                collected_at TEXT NOT NULL,
                inverter_state INTEGER,
                active_power_kw REAL,
                day_energy_kwh REAL,
                availability_status TEXT NOT NULL,
                communication_status TEXT NOT NULL,
                string_available_count INTEGER,
                string_total_count INTEGER,
                pv_current_json TEXT,
                pv_voltage_json TEXT,
                payload_json TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (provider_device_id) REFERENCES provider_devices(id) ON DELETE CASCADE,
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS availability_daily (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL,
                provider TEXT NOT NULL,
                period_date TEXT NOT NULL,
                inverter_availability_pct REAL,
                capacity_availability_pct REAL,
                communication_availability_pct REAL,
                string_availability_pct REAL,
                available_inverters INTEGER,
                total_inverters INTEGER,
                unavailable_inverters INTEGER,
                no_communication_devices INTEGER,
                available_strings INTEGER,
                total_strings INTEGER,
                unavailable_strings INTEGER,
                affected_power_kw REAL,
                unavailable_minutes INTEGER,
                payload_json TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(asset_id, provider, period_date),
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS inverter_power_samples (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL,
                provider TEXT NOT NULL,
                external_station_id TEXT NOT NULL,
                inverter_id TEXT NOT NULL,
                inverter_name TEXT,
                inverter_power_kw REAL,
                sample_time TEXT NOT NULL,
                active_power_kw REAL,
                raw_payload TEXT,
                created_at TEXT NOT NULL,
                UNIQUE(provider, inverter_id, sample_time),
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS inverter_availability_daily (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL,
                provider TEXT NOT NULL,
                availability_date TEXT NOT NULL,
                inverter_id TEXT NOT NULL,
                inverter_name TEXT,
                inverter_power_kw REAL,
                valid_slots INTEGER NOT NULL,
                available_slots INTEGER NOT NULL,
                unavailable_slots INTEGER NOT NULL,
                availability_pct REAL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(provider, inverter_id, availability_date),
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS plant_availability_daily (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL,
                provider TEXT NOT NULL,
                availability_date TEXT NOT NULL,
                valid_slots INTEGER NOT NULL,
                weighted_availability_pct REAL,
                inverter_count INTEGER NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(provider, asset_id, availability_date),
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS provider_device_expected_strings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                provider_device_id INTEGER NOT NULL,
                string_index INTEGER NOT NULL,
                expected INTEGER NOT NULL DEFAULT 1,
                source TEXT NOT NULL,
                observed_count INTEGER DEFAULT 0,
                first_observed_at TEXT,
                last_observed_at TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(provider_device_id, string_index),
                FOREIGN KEY (provider_device_id) REFERENCES provider_devices(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS production_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL,
                provider TEXT NOT NULL DEFAULT 'FusionSolar',
                external_id TEXT,
                period_type TEXT NOT NULL,
                period_date TEXT NOT NULL,
                production_kwh REAL,
                specific_yield REAL,
                expected_kwh REAL,
                expected_specific_yield REAL,
                deviation_pct REAL,
                performance_status TEXT,
                expected_source TEXT,
                data_quality TEXT,
                notes TEXT,
                selected_production_key TEXT,
                selected_production_raw_value TEXT,
                reference_diagnostic_json TEXT,
                payload_json TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(asset_id, provider, period_type, period_date),
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS performance_settings (
                asset_id INTEGER PRIMARY KEY,
                enabled INTEGER DEFAULT 1,
                warning_deviation_pct REAL DEFAULT -10,
                alert_deviation_pct REAL DEFAULT -20,
                critical_deviation_pct REAL DEFAULT -30,
                baseline_years INTEGER DEFAULT 2,
                min_baseline_points INTEGER DEFAULT 1,
                monthly_budget_json TEXT DEFAULT '',
                notes TEXT,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS telegram_alerts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER,
                alert_type TEXT NOT NULL,
                alert_key TEXT NOT NULL,
                message TEXT NOT NULL,
                sent_at TEXT NOT NULL,
                status TEXT NOT NULL,
                error_message TEXT NULL,
                blocked_reason TEXT NULL
            );

            CREATE TABLE IF NOT EXISTS alert_settings (
                key TEXT PRIMARY KEY,
                value TEXT
            );

            CREATE TABLE IF NOT EXISTS app_state (
                key TEXT PRIMARY KEY,
                value TEXT,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS background_jobs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                job_type TEXT NOT NULL,
                status TEXT NOT NULL,
                params_json TEXT,
                result_json TEXT,
                error_message TEXT,
                created_at TEXT NOT NULL,
                started_at TEXT,
                finished_at TEXT
            );

            CREATE TABLE IF NOT EXISTS alert_blacklist (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NULL,
                asset_name TEXT NULL,
                reason TEXT NULL,
                created_at TEXT NOT NULL,
                active INTEGER DEFAULT 1,
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS alert_baseline (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                baseline_at TEXT NOT NULL,
                created_by TEXT NULL,
                notes TEXT NULL
            );

            CREATE TABLE IF NOT EXISTS tickets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL,
                title TEXT NOT NULL,
                urgency TEXT NOT NULL,
                status TEXT NOT NULL,
                installation_ref TEXT,
                notes TEXT,
                next_action TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS ticket_visits (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ticket_id INTEGER NOT NULL,
                visit_date TEXT NOT NULL,
                technician TEXT,
                result TEXT,
                notes TEXT,
                next_action TEXT,
                FOREIGN KEY (ticket_id) REFERENCES tickets(id) ON DELETE CASCADE
            );

            CREATE VIEW IF NOT EXISTS latest_monitoring_view AS
            SELECT mr.asset_id, mr.status, mr.record_date, mr.notes
            FROM monitoring_records mr
            JOIN (
                SELECT asset_id, MAX(record_date || 'T' || printf('%09d', id)) AS marker
                FROM monitoring_records
                GROUP BY asset_id
            ) latest
              ON latest.asset_id = mr.asset_id
             AND latest.marker = mr.record_date || 'T' || printf('%09d', mr.id);
            """
        )
        ensure_database_indexes(conn)
        ensure_column(conn, "monitoring_records", "batch_id INTEGER")
        ensure_column(conn, "monitoring_unmatched", "batch_id INTEGER")
        ensure_column(conn, "assets", "installation_group TEXT")
        ensure_column(conn, "assets", "monitoring_enabled INTEGER DEFAULT 1")
        ensure_column(conn, "assets", "alerts_enabled INTEGER DEFAULT 1")
        ensure_column(conn, "assets", "monitoring_status TEXT DEFAULT 'active'")
        ensure_column(conn, "assets", "silenced_until TEXT")
        ensure_column(conn, "assets", "silence_reason TEXT")
        ensure_column(conn, "tickets", "planned_date TEXT")
        ensure_column(conn, "tickets", "due_date TEXT")
        ensure_column(conn, "tickets", "estimated_minutes INTEGER DEFAULT 60")
        ensure_column(conn, "tickets", "assigned_to TEXT")
        ensure_column(conn, "tickets", "material_status TEXT DEFAULT 'Nao definido'")
        ensure_column(conn, "tickets", "planning_notes TEXT")
        ensure_column(conn, "tickets", "work_type TEXT")
        ensure_column(conn, "assets", "selected_for_alerts INTEGER DEFAULT 0")
        ensure_column(conn, "telegram_alerts", "blocked_reason TEXT")
        ensure_column(conn, "om_contracts", "renewal_status TEXT")
        ensure_column(conn, "om_contracts", "last_contact_date TEXT")
        ensure_column(conn, "om_contracts", "renewal_notes TEXT")
        ensure_column(conn, "integration_configs", "real_time_endpoint TEXT")
        ensure_column(conn, "integration_configs", "device_list_endpoint TEXT")
        ensure_column(conn, "integration_configs", "device_real_time_endpoint TEXT")
        ensure_column(conn, "integration_configs", "device_history_endpoint TEXT")
        ensure_column(conn, "integration_configs", "day_kpi_endpoint TEXT")
        ensure_column(conn, "integration_configs", "month_kpi_endpoint TEXT")
        ensure_column(conn, "production_records", "selected_production_key TEXT")
        ensure_column(conn, "production_records", "selected_production_raw_value TEXT")
        ensure_column(conn, "production_records", "reference_diagnostic_json TEXT")
        disable_removed_inverter_devices(conn)
        populate_missing_inverter_rated_power(conn)
        populate_missing_installation_groups(conn)
        populate_missing_group_metadata(conn)
        ensure_predefined_export_templates(conn)
        ensure_alert_settings_defaults(conn)
        conn.commit()


def ensure_database_indexes(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE INDEX IF NOT EXISTS idx_monitoring_records_asset_date_id
            ON monitoring_records(asset_id, record_date DESC, id DESC);

        CREATE INDEX IF NOT EXISTS idx_monitoring_records_record_date_source
            ON monitoring_records(record_date, source);

        CREATE INDEX IF NOT EXISTS idx_monitoring_records_status_record_date
            ON monitoring_records(status, record_date);

        CREATE INDEX IF NOT EXISTS idx_production_records_provider_period_asset
            ON production_records(provider, period_type, period_date, asset_id);

        CREATE INDEX IF NOT EXISTS idx_production_records_performance_status
            ON production_records(performance_status);

        CREATE INDEX IF NOT EXISTS idx_asset_integrations_provider_external_id
            ON asset_integrations(provider, external_id);

        CREATE INDEX IF NOT EXISTS idx_asset_integrations_provider_enabled_asset
            ON asset_integrations(provider, enabled, asset_id);

        CREATE INDEX IF NOT EXISTS idx_tickets_asset_status
            ON tickets(asset_id, status);

        CREATE INDEX IF NOT EXISTS idx_integration_unresolved_provider_resolution_created
            ON integration_unresolved(provider, resolution_status, created_at);

        CREATE INDEX IF NOT EXISTS idx_telegram_alerts_alert_key_status
            ON telegram_alerts(alert_key, status);

        CREATE INDEX IF NOT EXISTS idx_alert_blacklist_asset_active
            ON alert_blacklist(asset_id, active);

        CREATE INDEX IF NOT EXISTS idx_background_jobs_type_status_created
            ON background_jobs(job_type, status, created_at);

        CREATE INDEX IF NOT EXISTS idx_provider_devices_provider_station
            ON provider_devices(provider, station_code);

        CREATE INDEX IF NOT EXISTS idx_provider_devices_provider_external_id
            ON provider_devices(provider, external_device_id);

        CREATE INDEX IF NOT EXISTS idx_provider_devices_asset_enabled
            ON provider_devices(asset_id, enabled);

        CREATE INDEX IF NOT EXISTS idx_device_realtime_snapshots_asset_collected
            ON device_realtime_snapshots(asset_id, collected_at DESC);

        CREATE INDEX IF NOT EXISTS idx_device_realtime_snapshots_device_collected
            ON device_realtime_snapshots(provider_device_id, collected_at DESC);

        CREATE INDEX IF NOT EXISTS idx_availability_daily_asset_provider_period
            ON availability_daily(asset_id, provider, period_date DESC);

        CREATE INDEX IF NOT EXISTS idx_inverter_power_samples_asset_time
            ON inverter_power_samples(asset_id, sample_time);

        CREATE INDEX IF NOT EXISTS idx_inverter_availability_daily_asset_date
            ON inverter_availability_daily(asset_id, availability_date);

        CREATE INDEX IF NOT EXISTS idx_plant_availability_daily_date_asset
            ON plant_availability_daily(availability_date, asset_id);

        CREATE INDEX IF NOT EXISTS idx_provider_device_expected_strings_device_index
            ON provider_device_expected_strings(provider_device_id, string_index);
        """
    )


def encode_job_params(params: dict[str, Any]) -> str:
    return json.dumps(params, ensure_ascii=True, sort_keys=True)


def create_background_job(
    conn: sqlite3.Connection,
    job_type: str,
    params: dict[str, Any],
    prevent_duplicate: bool = True,
) -> tuple[int, bool]:
    if prevent_duplicate:
        existing = conn.execute(
            """
            SELECT id
            FROM background_jobs
            WHERE job_type = ? AND status IN ('pending', 'running')
            ORDER BY id DESC
            LIMIT 1
            """,
            (job_type,),
        ).fetchone()
        if existing is not None:
            return int(existing["id"]), False

    now = datetime.now().isoformat(timespec="seconds")
    cursor = conn.execute(
        """
        INSERT INTO background_jobs (job_type, status, params_json, created_at)
        VALUES (?, 'pending', ?, ?)
        """,
        (job_type, encode_job_params(params), now),
    )
    return int(cursor.lastrowid), True


def mark_background_job_running(conn: sqlite3.Connection, job_id: int) -> bool:
    now = datetime.now().isoformat(timespec="seconds")
    cursor = conn.execute(
        """
        UPDATE background_jobs
        SET status = 'running', started_at = ?, error_message = NULL
        WHERE id = ? AND status = 'pending'
        """,
        (now, job_id),
    )
    conn.commit()
    return cursor.rowcount == 1


def mark_background_job_success(conn: sqlite3.Connection, job_id: int, result: dict[str, Any]) -> None:
    now = datetime.now().isoformat(timespec="seconds")
    conn.execute(
        """
        UPDATE background_jobs
        SET status = 'success', result_json = ?, error_message = NULL, finished_at = ?
        WHERE id = ?
        """,
        (json.dumps(result, ensure_ascii=True, sort_keys=True), now, job_id),
    )
    conn.commit()


def mark_background_job_failed(conn: sqlite3.Connection, job_id: int, error_message: str) -> None:
    now = datetime.now().isoformat(timespec="seconds")
    conn.execute(
        """
        UPDATE background_jobs
        SET status = 'failed', error_message = ?, finished_at = ?
        WHERE id = ?
        """,
        (error_message[:2000], now, job_id),
    )
    conn.commit()


def mark_stale_running_background_jobs_failed(
    conn: sqlite3.Connection,
    stale_after_minutes: int = BACKGROUND_JOB_STALE_RUNNING_MINUTES,
) -> int:
    cutoff = datetime.now() - timedelta(minutes=stale_after_minutes)
    now = datetime.now().isoformat(timespec="seconds")
    cursor = conn.execute(
        """
        UPDATE background_jobs
        SET status = 'failed',
            error_message = ?,
            finished_at = ?
        WHERE status = 'running'
          AND COALESCE(started_at, created_at) < ?
        """,
        (
            f"Job marked failed on startup after being running for more than {stale_after_minutes} minutes.",
            now,
            cutoff.isoformat(timespec="seconds"),
        ),
    )
    conn.commit()
    return cursor.rowcount


def fetch_pending_background_job_ids(conn: sqlite3.Connection) -> list[int]:
    rows = query_all(
        conn,
        """
        SELECT id
        FROM background_jobs
        WHERE status = 'pending'
        ORDER BY id ASC
        """,
    )
    return [int(row["id"]) for row in rows]


def fetch_latest_background_jobs(
    conn: sqlite3.Connection,
    limit: int = 10,
    job_types: tuple[str, ...] | None = None,
) -> list[dict[str, Any]]:
    params: list[Any] = []
    where_sql = ""
    if job_types:
        placeholders = ", ".join("?" for _ in job_types)
        where_sql = f"WHERE job_type IN ({placeholders})"
        params.extend(job_types)
    params.append(limit)
    rows = conn.execute(
        f"""
        SELECT id, job_type, status, params_json, result_json, error_message, created_at, started_at, finished_at
        FROM background_jobs
        {where_sql}
        ORDER BY id DESC
        LIMIT ?
        """,
        params,
    ).fetchall()
    jobs: list[dict[str, Any]] = []
    for row in rows:
        job = dict(row)
        result_summary = ""
        if row["result_json"]:
            try:
                result = json.loads(row["result_json"])
            except json.JSONDecodeError:
                result = {}
            if isinstance(result, dict):
                parts = []
                if result.get("records_updated") is not None:
                    parts.append(f"registos: {result['records_updated']}")
                if result.get("monthly_records_updated") is not None:
                    parts.append(f"mensais: {result['monthly_records_updated']}")
                if result.get("api_calls_used") is not None:
                    parts.append(f"chamadas API: {result['api_calls_used']}")
                if result.get("wait_cycles"):
                    parts.append(f"esperas: {result['wait_cycles']}")
                if result.get("resume_hint"):
                    parts.append(f"retomar: {result['resume_hint']}")
                if result.get("stopped_reason"):
                    parts.append(str(result["stopped_reason"]))
                elif result.get("status"):
                    parts.append(f"estado: {result['status']}")
                result_summary = " | ".join(parts)
        job["result_summary"] = result_summary
        jobs.append(job)
    return jobs


def query_one(sql: str, params: tuple[Any, ...] = ()) -> sqlite3.Row | None:
    return g.db.execute(sql, params).fetchone()


def ensure_alert_settings_defaults(conn: sqlite3.Connection) -> None:
    for key, value in ALERT_SETTING_DEFAULTS.items():
        conn.execute(
            "INSERT OR IGNORE INTO alert_settings (key, value) VALUES (?, ?)",
            (key, value),
        )


def row_get(row: sqlite3.Row | dict[str, Any], key: str, default: Any = None) -> Any:
    if isinstance(row, sqlite3.Row):
        return row[key] if key in row.keys() else default
    return row.get(key, default)


def normalize_bool(value: Any, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value != 0
    normalized = str(value).strip().lower()
    if not normalized:
        return default
    return normalized in {"1", "true", "yes", "on", "sim", "y"}


def get_alert_setting(conn: sqlite3.Connection, key: str, default: str | None = None) -> str:
    ensure_alert_settings_defaults(conn)
    value = query_scalar(conn, "SELECT value FROM alert_settings WHERE key = ?", (key,))
    if value is None:
        return ALERT_SETTING_DEFAULTS.get(key, default or "")
    return str(value)


def get_alert_settings(conn: sqlite3.Connection) -> dict[str, str]:
    ensure_alert_settings_defaults(conn)
    settings = dict(ALERT_SETTING_DEFAULTS)
    for row in query_all(conn, "SELECT key, value FROM alert_settings"):
        settings[row["key"]] = row["value"] or ""
    return settings


def alert_setting_bool(conn: sqlite3.Connection, key: str, default: bool = False) -> bool:
    fallback = "true" if default else "false"
    return normalize_bool(get_alert_setting(conn, key, fallback), default)


def set_alert_setting(conn: sqlite3.Connection, key: str, value: Any) -> None:
    conn.execute(
        """
        INSERT INTO alert_settings (key, value)
        VALUES (?, ?)
        ON CONFLICT(key) DO UPDATE SET value = excluded.value
        """,
        (key, str(value)),
    )


def telegram_env_allows_alerts() -> bool:
    value = os.environ.get("TELEGRAM_ALERTS_ENABLED")
    if value is None:
        return True
    return normalize_bool(value, False)


def fetch_dashboard_stats(conn: sqlite3.Connection) -> dict[str, Any]:
    latest_status_counts = {
        row["status"]: row["total"]
        for row in query_all(
            conn,
            """
            SELECT lm.status, COUNT(*) AS total
            FROM latest_monitoring_view lm
            JOIN assets a ON a.id = lm.asset_id
            WHERE COALESCE(a.monitoring_status, 'active') != 'disabled'
            GROUP BY lm.status
            """,
        )
    }
    return {
        "assets": query_scalar(conn, "SELECT COUNT(*) FROM assets"),
        "active_om_assets": query_scalar(conn, "SELECT COUNT(*) FROM assets WHERE active_contract = 'yes'"),
        "pipeline_assets": query_scalar(conn, "SELECT COUNT(*) FROM assets WHERE COALESCE(active_contract, '') != 'yes'"),
        "monitoring_today": query_scalar(
            conn,
            "SELECT COUNT(*) FROM monitoring_records WHERE record_date = ?",
            (date.today().isoformat(),),
        ),
        "open_tickets": query_scalar(conn, "SELECT COUNT(*) FROM tickets WHERE status != 'Fechado'"),
        "open_tickets_active_om": query_scalar(
            conn,
            """
            SELECT COUNT(*)
            FROM tickets t
            JOIN assets a ON a.id = t.asset_id
            WHERE t.status != 'Fechado' AND a.active_contract = 'yes'
            """,
        ),
        "critical_tickets": query_scalar(
            conn,
            "SELECT COUNT(*) FROM tickets WHERE urgency = 'Critica' AND status != 'Fechado'",
        ),
        "critical_active_issues": query_scalar(
            conn,
            """
            SELECT COUNT(*)
            FROM latest_monitoring_view lm
            JOIN assets a ON a.id = lm.asset_id
            WHERE a.active_contract = 'yes' AND lm.status IN ('Erro', 'Desconectada')
            """,
        ),
        "expired_renewals": query_scalar(
            conn,
            """
            SELECT COUNT(*)
            FROM assets a
            LEFT JOIN om_contracts oc ON oc.asset_id = a.id
            WHERE (a.maintenance = 'yes' OR oc.id IS NOT NULL)
              AND COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) NOT IN ('', '-')
              AND COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) < ?
            """,
            (date.today().isoformat(),),
        ),
        "renewals_next_90_days": query_scalar(
            conn,
            """
            SELECT COUNT(*)
            FROM assets a
            LEFT JOIN om_contracts oc ON oc.asset_id = a.id
            WHERE (a.maintenance = 'yes' OR oc.id IS NOT NULL)
              AND COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) NOT IN ('', '-')
              AND COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) BETWEEN ? AND ?
            """,
            (date.today().isoformat(), (date.today() + timedelta(days=90)).isoformat()),
        ),
        "renewals_this_year": query_scalar(
            conn,
            """
            SELECT COUNT(*)
            FROM assets a
            LEFT JOIN om_contracts oc ON oc.asset_id = a.id
            WHERE (a.maintenance = 'yes' OR oc.id IS NOT NULL)
              AND COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) NOT IN ('', '-')
              AND COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) BETWEEN ? AND ?
            """,
            (date.today().isoformat(), f"{date.today().year}-12-31"),
        ),
        "status_counts": latest_status_counts,
    }


def build_executive_dashboard_stats(conn: sqlite3.Connection) -> dict[str, Any]:
    active_problem_rows = query_all(
        conn,
        """
        SELECT
            a.id AS asset_id,
            a.project_name,
            a.active_contract,
            lm.status,
            lm.record_date
        FROM latest_monitoring_view lm
        JOIN assets a ON a.id = lm.asset_id
        WHERE a.active_contract = 'yes'
          AND lm.status IN ('Erro', 'Desconectada')
        """,
    )
    enriched = enrich_operational_rows(conn, active_problem_rows)
    critical_or_high = [row for row in enriched if row["auto_priority"] in {"Critica", "Alta"}]
    recurring = [row for row in enriched if int(row.get("recurrence_count") or 0) >= 2]
    long_running = [row for row in enriched if int(row.get("problem_days") or 0) >= 7]
    avg_days = round(
        sum(int(row.get("problem_days") or 0) for row in enriched) / len(enriched),
        1,
    ) if enriched else 0
    return {
        "active_om_problems": len(enriched),
        "critical_or_high": len(critical_or_high),
        "recurring_90d": len(recurring),
        "long_running_7d": len(long_running),
        "avg_problem_days": avg_days,
    }


def build_integration_summary(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    sources = ["FusionSolar", "Sigenergy", "Manual / Outro"]
    summary: list[dict[str, Any]] = []
    for source in sources:
        source_aliases = {
            "FusionSolar": ["FusionSolar", "FusionSolar API", "fusion-solar-sync"],
            "Sigenergy": ["Sigenergy"],
            "Manual / Outro": ["Manual / Outro", "manual-paste", "auto-resolved"],
        }[source]
        placeholders = ",".join("?" for _ in source_aliases)
        last_batch = conn.execute(
            f"""
            SELECT imported_at, record_date, imported_count, matched_count, unmatched_count
            FROM monitoring_import_batches
            WHERE source IN ({placeholders})
            ORDER BY imported_at DESC, id DESC
            LIMIT 1
            """,
            source_aliases,
        ).fetchone()
        summary.append(
            {
                "source": source,
                "last_imported_at": last_batch["imported_at"] if last_batch else "",
                "last_record_date": last_batch["record_date"] if last_batch else "",
                "imported_count": last_batch["imported_count"] if last_batch else 0,
                "matched_count": last_batch["matched_count"] if last_batch else 0,
                "unmatched_count": last_batch["unmatched_count"] if last_batch else 0,
            }
        )
    return summary


def normalize_name(value: str) -> str:
    lowered = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii").lower()
    cleaned = "".join(char if char.isalnum() else " " for char in lowered)
    return " ".join(cleaned.split())


def infer_installation_group(project_name: str) -> str:
    name = (project_name or "").strip()
    if not name:
        return ""
    stripped = re.sub(r"\s*\([^)]*\)\s*$", "", name).strip()
    return stripped or name


def classify_fusionsolar_link(external_name: str, project_name: str, installation_group: str | None = "") -> tuple[str, str]:
    external_norm = normalize_name(external_name or "")
    project_norm = normalize_name(project_name or "")
    group_norm = normalize_name(installation_group or "")
    local_names = [name for name in (project_norm, group_norm) if name]

    if not external_norm or not local_names:
        return "Rever", "Faltam nomes para comparar."
    if external_norm in local_names:
        return "OK", "Nome FusionSolar igual a central/instalacao local."
    for local_name in local_names:
        shorter, longer = sorted((external_norm, local_name), key=len)
        if len(shorter) >= 6 and shorter in longer:
            return "Atencao", "Nome parcialmente semelhante; confirma manualmente."
    return "Rever", "Nome FusionSolar diferente da central local associada."


def update_fusionsolar_mapping_asset(conn: sqlite3.Connection, integration_id: int, asset_id: int) -> None:
    integration = conn.execute(
        "SELECT id FROM asset_integrations WHERE id = ? AND provider = ?",
        (integration_id, INTEGRATION_PROVIDER_FUSIONSOLAR),
    ).fetchone()
    if integration is None:
        raise ValueError("Mapeamento FusionSolar nao encontrado.")
    asset = conn.execute("SELECT id FROM assets WHERE id = ?", (asset_id,)).fetchone()
    if asset is None:
        raise ValueError("Central local nao encontrada.")
    conn.execute(
        """
        UPDATE asset_integrations
        SET asset_id = ?, last_error = ''
        WHERE id = ?
        """,
        (asset_id, integration_id),
    )
    conn.commit()


def get_fusionsolar_link_audit_rows(conn: sqlite3.Connection, provider: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    mapped_rows = query_all(
        conn,
        """
        SELECT
            ai.id,
            ai.asset_id,
            ai.external_id,
            ai.external_name,
            ai.last_status,
            ai.last_sync_at,
            a.project_name,
            a.installation_group
        FROM asset_integrations ai
        JOIN assets a ON a.id = ai.asset_id
        WHERE ai.provider = ? AND ai.enabled = 1
        """,
        (provider,),
    )
    duplicate_counts: dict[int, int] = {}
    duplicate_names: dict[int, list[str]] = {}
    for row in mapped_rows:
        asset_id = int(row["asset_id"])
        duplicate_counts[asset_id] = duplicate_counts.get(asset_id, 0) + 1
        duplicate_names.setdefault(asset_id, []).append(row["external_name"] or "")

    for row in mapped_rows:
        asset_id = int(row["asset_id"])
        verdict, reason = classify_fusionsolar_link(
            row["external_name"] or "",
            row["project_name"] or "",
            row["installation_group"] or "",
        )
        duplicate_count = duplicate_counts.get(asset_id, 0)
        if duplicate_count > 1:
            verdict = "Atencao" if verdict == "OK" else verdict
            reason = f"{reason} Ha {duplicate_count} entradas FusionSolar ligadas a esta central local."
        rows.append(
            {
                "integration_id": row["id"],
                "unresolved_id": None,
                "external_id": row["external_id"] or "",
                "external_name": row["external_name"] or "",
                "asset_id": asset_id,
                "project_name": row["project_name"] or "",
                "installation_group": row["installation_group"] or "",
                "last_status": row["last_status"] or "",
                "last_sync_at": row["last_sync_at"] or "",
                "verdict": verdict,
                "reason": reason,
                "duplicate_count": duplicate_count,
                "duplicate_names": ", ".join(name for name in duplicate_names.get(asset_id, []) if name),
            }
        )

    unresolved_rows = query_all(
        conn,
        """
        SELECT id, external_id, external_name, external_status, created_at
        FROM integration_unresolved
        WHERE provider = ? AND resolution_status = 'pending'
        """,
        (provider,),
    )
    for row in unresolved_rows:
        rows.append(
            {
                "integration_id": None,
                "unresolved_id": row["id"],
                "external_id": row["external_id"] or "",
                "external_name": row["external_name"] or "",
                "asset_id": None,
                "project_name": "",
                "installation_group": "",
                "last_status": row["external_status"] or "",
                "last_sync_at": row["created_at"] or "",
                "verdict": "Por resolver",
                "reason": "Ainda nao esta associada a nenhuma central local.",
                "duplicate_count": 0,
                "duplicate_names": "",
            }
        )

    unmapped_local_rows = query_all(
        conn,
        """
        SELECT a.id, a.project_name, a.installation_group
        FROM assets a
        WHERE a.active_contract = 'yes'
          AND NOT EXISTS (
              SELECT 1
              FROM asset_integrations ai
              WHERE ai.provider = ? AND ai.asset_id = a.id AND ai.enabled = 1
          )
        ORDER BY a.project_name COLLATE NOCASE
        """,
        (provider,),
    )
    for row in unmapped_local_rows:
        rows.append(
            {
                "integration_id": None,
                "unresolved_id": None,
                "external_id": "",
                "external_name": "",
                "asset_id": int(row["id"]),
                "project_name": row["project_name"] or "",
                "installation_group": row["installation_group"] or "",
                "last_status": "",
                "last_sync_at": "",
                "verdict": "Rever",
                "reason": "Central local O&M sem entrada devolvida pelo FusionSolar. Verifica a autorizacao da planta na conta northbound.",
                "duplicate_count": 0,
                "duplicate_names": "",
            }
        )

    priority = {"Rever": 0, "Atencao": 1, "Por resolver": 2, "OK": 3}
    return sorted(rows, key=lambda item: (priority.get(item["verdict"], 9), (item["external_name"] or item["project_name"]).lower()))


def parse_date_value(value: str | None) -> date | None:
    if value in (None, "", "-"):
        return None
    raw_value = str(value).strip()
    for date_format in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw_value, date_format).date()
        except ValueError:
            continue
    return None


def normalize_date_value(value: str | None) -> str:
    parsed = parse_date_value(value)
    return parsed.isoformat() if parsed else (str(value).strip() if value else "")


def derive_active_contract(end_date: str | None, current_value: str = "") -> str:
    parsed_end_date = parse_date_value(end_date)
    if parsed_end_date is None:
        return current_value
    return "yes" if parsed_end_date >= date.today() else "no"


def om_status_label(value: str | None) -> str:
    return "O&M ativo" if value == "yes" else "Sem contrato"


def format_date_pt(value: str | None) -> str:
    parsed = parse_date_value(value)
    return parsed.strftime("%d/%m/%Y") if parsed else (value or "-")


def format_number(value: Any, max_decimals: int = 2) -> str:
    parsed = parse_float_value(value)
    if parsed is None:
        return "-"
    formatted = f"{parsed:.{max_decimals}f}"
    return formatted.rstrip("0").rstrip(".")


def record_value(record: sqlite3.Row | dict[str, Any], key: str) -> Any:
    if isinstance(record, sqlite3.Row):
        return record[key] if key in record.keys() else None
    return record.get(key)


def compute_performance_percentage(record: sqlite3.Row | dict[str, Any]) -> float | None:
    specific_yield = parse_float_value(record_value(record, "specific_yield"))
    expected_specific_yield = parse_float_value(record_value(record, "expected_specific_yield"))
    if specific_yield is None or expected_specific_yield is None or expected_specific_yield <= 0:
        return None
    return (specific_yield / expected_specific_yield) * 100


def performance_bar_width(record: sqlite3.Row | dict[str, Any]) -> str:
    percentage = compute_performance_percentage(record)
    if percentage is None:
        return "0%"
    return f"{max(0, min(percentage, 100)):.1f}%"


def performance_status_class(status: str | None) -> str:
    normalized = unicodedata.normalize("NFKD", status or "Sem dados")
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "-", ascii_value.lower()).strip("-")


def reference_diagnostic(record: sqlite3.Row | dict[str, Any]) -> dict[str, Any]:
    raw = record_value(record, "reference_diagnostic_json")
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
    except (TypeError, json.JSONDecodeError):
        return {}
    return parsed if isinstance(parsed, dict) else {}


def normalize_monitoring_source(value: str | None) -> str:
    normalized = normalize_name(value or "")
    if normalized in {"fusion solar", "fusionsolar", "fusion solar api", "fusion solar sync", "fusion-solar-sync"}:
        return "FusionSolar"
    if normalized in {"sigenergy", "sig energy"}:
        return "Sigenergy"
    if normalized in {"manual", "manual outro", "manual paste", "auto resolved"}:
        return "Manual / Outro"
    return (value or "Manual / Outro").strip() or "Manual / Outro"


def days_between(start_value: str | None, end_value: str | None = None) -> int:
    start = parse_date_value(start_value)
    end = parse_date_value(end_value) or date.today()
    if start is None:
        return 0
    return max((end - start).days + 1, 0)


def auto_priority(status: str | None, problem_days: int, recurrence_count: int, open_tickets: int, active_contract: str | None) -> str:
    if active_contract != "yes":
        return "Baixa"
    score = 0
    if status == "Erro":
        score += 4
    elif status == "Desconectada":
        score += 3
    if problem_days >= 7:
        score += 3
    elif problem_days >= 3:
        score += 2
    elif problem_days >= 1:
        score += 1
    if recurrence_count >= 3:
        score += 2
    elif recurrence_count >= 2:
        score += 1
    if open_tickets:
        score += 1
    if score >= 7:
        return "Critica"
    if score >= 5:
        return "Alta"
    if score >= 2:
        return "Media"
    return "Baixa"


def contract_end_sql() -> str:
    return "COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, ''))"


def apply_group_defaults(
    conn: sqlite3.Connection,
    payload: dict[str, str],
    installation_group: str,
    exclude_asset_id: int | None = None,
) -> dict[str, str]:
    if not installation_group:
        return payload
    available_fields = [field for field in GROUP_INHERITED_FIELDS if field in payload]
    if not available_fields:
        return payload

    conditions = ["installation_group = ?"]
    params: list[Any] = [installation_group]
    if exclude_asset_id is not None:
        conditions.append("id != ?")
        params.append(exclude_asset_id)

    sources = conn.execute(
        f"""
        SELECT {", ".join(available_fields)}
        FROM assets
        WHERE {" AND ".join(conditions)}
          AND ({ " OR ".join(f"NULLIF({field}, '') IS NOT NULL" for field in available_fields) })
        ORDER BY id ASC
        """,
        params,
    ).fetchall()

    for source in sources:
        for field in available_fields:
            if not payload.get(field) and source[field]:
                payload[field] = source[field]
        if all(payload.get(field) for field in available_fields):
            break
    return payload


def apply_group_defaults_to_asset(conn: sqlite3.Connection, asset_id: int, installation_group: str) -> None:
    asset = conn.execute("SELECT * FROM assets WHERE id = ?", (asset_id,)).fetchone()
    if asset is None:
        return
    payload = {field: asset[field] or "" for field in GROUP_INHERITED_FIELDS}
    updated_payload = apply_group_defaults(conn, payload, installation_group, exclude_asset_id=asset_id)
    changed_fields = [field for field in GROUP_INHERITED_FIELDS if (asset[field] or "") != updated_payload.get(field, "")]
    if not changed_fields:
        return
    assignments = ", ".join(f"{field} = ?" for field in changed_fields)
    values = [updated_payload[field] for field in changed_fields]
    conn.execute(f"UPDATE assets SET {assignments} WHERE id = ?", values + [asset_id])


def populate_missing_group_metadata(conn: sqlite3.Connection) -> None:
    rows = conn.execute(
        """
        SELECT id, installation_group
        FROM assets
        WHERE installation_group IS NOT NULL AND TRIM(installation_group) != ''
        ORDER BY installation_group COLLATE NOCASE, id
        """
    ).fetchall()
    for row in rows:
        apply_group_defaults_to_asset(conn, row["id"], row["installation_group"])


def sync_asset_contract_status(
    conn: sqlite3.Connection,
    asset_id: int,
    start_date: str | None = None,
    end_date: str | None = None,
) -> None:
    asset = conn.execute("SELECT active_contract, start_contract, end_contract FROM assets WHERE id = ?", (asset_id,)).fetchone()
    if asset is None:
        return
    contract = conn.execute(
        "SELECT contract_start_date, contract_end_date FROM om_contracts WHERE asset_id = ?",
        (asset_id,),
    ).fetchone()
    final_start = normalize_date_value(start_date or (contract["contract_start_date"] if contract else "") or asset["start_contract"])
    final_end = normalize_date_value(end_date or (contract["contract_end_date"] if contract else "") or asset["end_contract"])
    active_contract = derive_active_contract(final_end, asset["active_contract"] or "")
    conn.execute(
        """
        UPDATE assets
        SET maintenance = CASE WHEN ? = 'yes' THEN 'yes' ELSE maintenance END,
            active_contract = ?,
            start_contract = CASE WHEN ? != '' THEN ? ELSE start_contract END,
            end_contract = CASE WHEN ? != '' THEN ? ELSE end_contract END
        WHERE id = ?
        """,
        (active_contract, active_contract, final_start, final_start, final_end, final_end, asset_id),
    )


def sync_all_contract_statuses(conn: sqlite3.Connection) -> None:
    rows = conn.execute(
        """
        SELECT a.id, a.start_contract, a.end_contract, oc.contract_start_date, oc.contract_end_date
        FROM assets a
        LEFT JOIN om_contracts oc ON oc.asset_id = a.id
        WHERE COALESCE(NULLIF(oc.contract_end_date, ''), NULLIF(a.end_contract, '')) IS NOT NULL
        """
    ).fetchall()
    for row in rows:
        sync_asset_contract_status(
            conn,
            row["id"],
            row["contract_start_date"] or row["start_contract"],
            row["contract_end_date"] or row["end_contract"],
        )


def populate_missing_installation_groups(conn: sqlite3.Connection) -> None:
    rows = conn.execute(
        """
        SELECT id, project_name
        FROM assets
        WHERE installation_group IS NULL OR TRIM(installation_group) = ''
        """
    ).fetchall()
    for row in rows:
        conn.execute(
            "UPDATE assets SET installation_group = ? WHERE id = ?",
            (infer_installation_group(row["project_name"]), row["id"]),
        )


def status_rank(status: str) -> int:
    order = {
        "Erro": 1,
        "Desconectada": 2,
        "Alerta": 3,
        "Aberto": 4,
        "Em analise": 5,
        "Agendado": 6,
        "Em visita": 7,
        "Resolvido": 8,
        "Operacional": 9,
        "ok": 8,
    }
    return order.get(status or "", 99)


def group_latest_rows_by_installation(rows: list[sqlite3.Row]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for row in rows:
        installation_group = row["installation_group"] or row["project_name"]
        bucket = grouped.setdefault(
            installation_group,
            {
                "installation_group": installation_group,
                "location": row["location"],
                "active_contract": row["active_contract"],
                "record_date": row["record_date"],
                "members": [],
            },
        )
        bucket["members"].append(row)
        if not bucket["location"] and row["location"]:
            bucket["location"] = row["location"]
        if row["active_contract"] == "yes":
            bucket["active_contract"] = "yes"
        if (row["record_date"] or "") > (bucket["record_date"] or ""):
            bucket["record_date"] = row["record_date"]

    grouped_rows = []
    for bucket in grouped.values():
        members = sorted(bucket["members"], key=lambda item: (status_rank(item["status"]), item["project_name"].lower()))
        bucket["members"] = members
        bucket["group_status"] = members[0]["status"] if members else ""
        bucket["member_count"] = len(members)
        bucket["history_count"] = sum(int(member["history_count"]) for member in members)
        grouped_rows.append(bucket)

    grouped_rows.sort(
        key=lambda item: (
            0 if item["active_contract"] == "yes" else 1,
            status_rank(item["group_status"]),
            item["installation_group"].lower(),
        )
    )
    return grouped_rows


def normalize_status(value: str) -> str:
    lookup = {
        "erro": "Erro",
        "desconectada": "Desconectada",
        "operacional": "Operacional",
        "resolvido": "Resolvido",
        "aberto": "Aberto",
        "em analise": "Em analise",
        "agendado": "Agendado",
        "em visita": "Em visita",
        "fechado": "Fechado",
        "on": "Resolvido",
        "off": "Aberto",
        "faulty": "Em analise",
    }
    normalized = normalize_name(value)
    return lookup.get(normalized, value.strip())


def parse_iso_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    normalized = str(value).strip().replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(normalized)
    except ValueError:
        try:
            return datetime.fromisoformat(normalized + "T00:00:00")
        except ValueError:
            return None


def is_daytime_for_alert(now: datetime) -> bool:
    return 8 <= now.hour < 19


def html_line(value: Any) -> str:
    return html.escape(str(value or "-"))


def get_latest_monitoring_row(conn: sqlite3.Connection, asset_id: int) -> sqlite3.Row | None:
    return conn.execute(
        """
        SELECT id, status, record_date, source, created_at
        FROM monitoring_records
        WHERE asset_id = ?
        ORDER BY record_date DESC, id DESC
        LIMIT 1
        """,
        (asset_id,),
    ).fetchone()


def alert_already_sent(conn: sqlite3.Connection, alert_key: str) -> bool:
    row = conn.execute(
        "SELECT 1 FROM telegram_alerts WHERE alert_key = ? AND status = 'sent' LIMIT 1",
        (alert_key,),
    ).fetchone()
    return row is not None


def alert_recently_sent(
    conn: sqlite3.Connection,
    asset_id: int | None,
    alert_type: str,
    now: datetime,
    *,
    minutes: int = 0,
    hours: int = 0,
) -> bool:
    cooldown = timedelta(minutes=minutes, hours=hours)
    if cooldown.total_seconds() <= 0 or asset_id is None:
        return False
    since = (now - cooldown).isoformat(timespec="seconds")
    row = conn.execute(
        """
        SELECT 1
        FROM telegram_alerts
        WHERE asset_id = ? AND alert_type = ? AND status = 'sent' AND sent_at >= ?
        LIMIT 1
        """,
        (asset_id, alert_type, since),
    ).fetchone()
    return row is not None


def get_alert_type_setting(alert_type: str) -> str:
    mapping = {
        "novo_erro": "SEND_NEW_ERROR_ALERTS",
        "nova_desconexao": "SEND_OFFLINE_ALERTS",
        "desconexao_persistente_2h": "SEND_OFFLINE_ALERTS",
        "resolvido": "SEND_RESOLVED_ALERTS",
        "erro_persistente_24h": "SEND_PERSISTENT_ALERTS",
        "recorrente_7d": "SEND_RECURRENT_ALERTS",
    }
    return mapping.get(alert_type, "")


def is_asset_blacklisted(conn: sqlite3.Connection, asset: sqlite3.Row | dict[str, Any]) -> bool:
    asset_id = row_get(asset, "id", row_get(asset, "asset_id"))
    asset_name = str(row_get(asset, "project_name", "") or "").strip()
    row = conn.execute(
        """
        SELECT 1
        FROM alert_blacklist
        WHERE active = 1
          AND (
            (asset_id IS NOT NULL AND asset_id = ?)
            OR (asset_name IS NOT NULL AND lower(asset_name) = lower(?))
          )
        LIMIT 1
        """,
        (asset_id, asset_name),
    ).fetchone()
    return row is not None


def is_asset_in_oem_scope(asset: sqlite3.Row | dict[str, Any], alert_scope: str = "only_o&m") -> bool:
    if alert_scope == "all_assets":
        return True
    maintenance = normalize_bool(
        row_get(asset, "maintenance", row_get(asset, "Maintenance", row_get(asset, "contract_signed"))),
        False,
    )
    active_contract = str(row_get(asset, "active_contract", row_get(asset, "Active Contract", "")) or "").strip().lower()
    active_contract_ok = active_contract in {"yes", "true", "1", "ativo", "active", "sim"}
    if alert_scope == "only_o&m":
        return maintenance
    if alert_scope == "only_active_contracts":
        return maintenance and active_contract_ok
    if alert_scope == "only_selected_assets":
        return normalize_bool(row_get(asset, "selected_for_alerts", 0), False)
    return maintenance


def get_alert_baseline_at(conn: sqlite3.Connection) -> datetime | None:
    value = get_alert_setting(conn, "ALERT_BASELINE_AT", "")
    return parse_iso_datetime(value)


def is_before_alert_baseline(conn: sqlite3.Connection, value: str | None) -> bool:
    baseline_at = get_alert_baseline_at(conn)
    checked_at = parse_iso_datetime(value)
    return bool(baseline_at and checked_at and checked_at < baseline_at)


def alert_cooldown_active(
    conn: sqlite3.Connection,
    asset_id: int | None,
    alert_type: str,
    now: datetime,
) -> bool:
    if alert_type == "nova_desconexao":
        return alert_recently_sent(
            conn,
            asset_id,
            alert_type,
            now,
            minutes=int(get_alert_setting(conn, "OFFLINE_COOLDOWN_MINUTES", "120") or 120),
        )
    if alert_type == "resolvido":
        return alert_recently_sent(
            conn,
            asset_id,
            alert_type,
            now,
            minutes=int(get_alert_setting(conn, "RESOLVED_COOLDOWN_MINUTES", "0") or 0),
        )
    if alert_type == "novo_erro":
        return alert_recently_sent(
            conn,
            asset_id,
            alert_type,
            now,
            minutes=int(get_alert_setting(conn, "NEW_ERROR_COOLDOWN_MINUTES", "0") or 0),
        )
    if alert_type in {"erro_persistente_24h", "desconexao_persistente_2h"}:
        return alert_recently_sent(
            conn,
            asset_id,
            alert_type,
            now,
            hours=int(get_alert_setting(conn, "PERSISTENT_COOLDOWN_HOURS", "24") or 24),
        )
    if alert_type == "recorrente_7d":
        return alert_recently_sent(
            conn,
            asset_id,
            alert_type,
            now,
            hours=int(get_alert_setting(conn, "RECURRENT_COOLDOWN_HOURS", "24") or 24),
        )
    return False


def alert_decision(
    conn: sqlite3.Connection,
    asset: sqlite3.Row | dict[str, Any],
    alert_type: str,
    alert_key: str,
    now: datetime,
) -> tuple[bool, str]:
    if not alert_setting_bool(conn, "TELEGRAM_ALERTS_ENABLED", True) or not telegram_env_allows_alerts():
        return False, "global_disabled"
    if not is_telegram_configured():
        return False, "telegram_not_configured"
    if int(row_get(asset, "monitoring_enabled", 1) if row_get(asset, "monitoring_enabled", 1) is not None else 1) == 0:
        return False, "monitoring_disabled"
    if int(row_get(asset, "alerts_enabled", 1) if row_get(asset, "alerts_enabled", 1) is not None else 1) == 0:
        return False, "disabled"
    if is_asset_blacklisted(conn, asset):
        return False, "blacklist"
    monitoring_status = str(row_get(asset, "monitoring_status", "active") or "active")
    if monitoring_status in {"maintenance", "out_of_scope", "disabled"}:
        return False, monitoring_status
    if monitoring_status == "silenced":
        silenced_until = parse_iso_datetime(row_get(asset, "silenced_until"))
        if silenced_until and now < silenced_until:
            return False, "silenced"
    if not is_asset_in_oem_scope(asset, get_alert_setting(conn, "ALERT_SCOPE", "only_o&m")):
        return False, "out_of_scope"
    alert_setting = get_alert_type_setting(alert_type)
    if alert_setting and not alert_setting_bool(conn, alert_setting, True):
        return False, "alert_type_disabled"
    if alert_already_sent(conn, alert_key) or alert_cooldown_active(conn, row_get(asset, "id", row_get(asset, "asset_id")), alert_type, now):
        return False, "cooldown"
    return True, ""


def should_send_alert(conn: sqlite3.Connection, asset: sqlite3.Row | dict[str, Any], alert_type: str, alert_key: str, now: datetime) -> bool:
    return alert_decision(conn, asset, alert_type, alert_key, now)[0]


def record_telegram_alert(
    conn: sqlite3.Connection,
    asset_id: int | None,
    alert_type: str,
    alert_key: str,
    message: str,
    status: str,
    error_message: str = "",
    blocked_reason: str = "",
    sent_at: datetime | None = None,
) -> None:
    conn.execute(
        """
        INSERT INTO telegram_alerts (asset_id, alert_type, alert_key, message, sent_at, status, error_message, blocked_reason)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            asset_id,
            alert_type,
            alert_key,
            message,
            (sent_at or datetime.now()).isoformat(timespec="seconds"),
            status,
            error_message,
            blocked_reason,
        ),
    )


def send_and_record_telegram_alert(
    conn: sqlite3.Connection,
    asset_id: int | None,
    alert_type: str,
    alert_key: str,
    message: str,
) -> bool:
    if alert_already_sent(conn, alert_key):
        return False
    if not alert_setting_bool(conn, "TELEGRAM_ALERTS_ENABLED", True) or not telegram_env_allows_alerts():
        record_telegram_alert(conn, asset_id, alert_type, alert_key, message, "blocked", "Telegram desativado.", "global_disabled")
        return False
    if not is_telegram_configured():
        record_telegram_alert(conn, asset_id, alert_type, alert_key, message, "blocked", "Telegram por configurar.", "telegram_not_configured")
        return False
    try:
        ok = send_telegram_message(message)
        record_telegram_alert(
            conn,
            asset_id,
            alert_type,
            alert_key,
            message,
            "sent" if ok else "failed",
            "" if ok else "Telegram API devolveu falha ou nao respondeu.",
        )
        return ok
    except Exception as exc:
        current_app.logger.warning("Telegram alert failed without breaking import: %s", exc)
        record_telegram_alert(conn, asset_id, alert_type, alert_key, message, "failed", str(exc))
        return False


def find_problem_start(conn: sqlite3.Connection, asset_id: int, problem_status: str) -> str:
    rows = query_all(
        conn,
        """
        SELECT status, record_date, created_at
        FROM monitoring_records
        WHERE asset_id = ?
        ORDER BY record_date DESC, id DESC
        """,
        (asset_id,),
    )
    first = ""
    for row in rows:
        if row["status"] != problem_status:
            break
        first = row["record_date"] or row["created_at"]
    return first


def count_problem_occurrences_since(conn: sqlite3.Connection, asset_id: int, since_date: str) -> int:
    return int(
        query_scalar(
            conn,
            """
            SELECT COUNT(*)
            FROM monitoring_records
            WHERE asset_id = ?
              AND status IN ('Erro', 'Desconectada')
              AND record_date >= ?
            """,
            (asset_id, since_date),
        )
        or 0
    )


def build_state_change_message(event: dict[str, Any]) -> tuple[str, str]:
    asset_name = html_line(event["project_name"])
    previous_status = html_line(event["previous_status"])
    current_status = html_line(event["current_status"])
    happened_at = html_line(event["happened_at"])
    alarm_lines = ""
    if event.get("primary_alarm_name") or event.get("alarm_summary"):
        alarm_lines = "\n\n"
        if event.get("primary_alarm_name"):
            alarm_lines += f"Tipo de erro: {html_line(event.get('primary_alarm_name'))}\n"
        if event.get("primary_alarm_device"):
            alarm_lines += f"Aparelho: {html_line(event.get('primary_alarm_device'))}\n"
        if event.get("primary_alarm_severity"):
            alarm_lines += f"Severidade: {html_line(event.get('primary_alarm_severity'))}\n"
        if event.get("alarm_summary"):
            alarm_lines += f"Alarmes ativos: {html_line(event.get('alarm_summary'))}"
    if event["alert_type"] == "novo_erro":
        return (
            "novo_erro",
            f"🚨 <b>ALERTA — Novo erro</b>\n\nInstalacao: {asset_name}\nEstado anterior: {previous_status}\nEstado atual: {current_status}\nHora: {happened_at}{alarm_lines}",
        )
    if event["alert_type"] == "nova_desconexao":
        return (
            "nova_desconexao",
            f"⚠️ <b>ALERTA — Instalacao desconectada</b>\n\nInstalacao: {asset_name}\nHora: {happened_at}\nNota: desconexao em periodo de producao{alarm_lines}",
        )
    return (
        "resolvido",
        f"✅ <b>RESOLVIDO</b>\n\nInstalacao: {asset_name}\nEstado anterior: {previous_status}\nEstado atual: {current_status}\nDuracao aproximada: {html_line(event.get('duration') or '-')}",
    )


def build_monitoring_alert_event(
    conn: sqlite3.Connection,
    *,
    asset_id: int,
    previous_status: str,
    current_status: str,
    happened_at: str,
    alarm_context: dict[str, Any] | None = None,
) -> dict[str, Any] | None:
    previous_status = normalize_status(previous_status or "")
    current_status = normalize_status(current_status or "")
    if previous_status == current_status:
        return None

    alert_type = ""
    if previous_status in OK_MONITORING_STATUSES and current_status == "Erro":
        alert_type = "novo_erro"
    elif previous_status in OK_MONITORING_STATUSES and current_status == "Desconectada":
        if alert_setting_bool(conn, "DAYTIME_OFFLINE_ONLY", True) and not is_daytime_for_alert(parse_iso_datetime(happened_at) or datetime.now()):
            return None
        alert_type = "nova_desconexao"
    elif previous_status in PROBLEM_MONITORING_STATUSES and current_status in OK_MONITORING_STATUSES:
        alert_type = "resolvido"
    if not alert_type:
        return None

    asset = conn.execute("SELECT project_name FROM assets WHERE id = ?", (asset_id,)).fetchone()
    if asset is None:
        return None

    duration = ""
    if alert_type == "resolvido":
        latest_problem = conn.execute(
            """
            SELECT created_at, record_date
            FROM monitoring_records
            WHERE asset_id = ? AND status = ?
            ORDER BY record_date DESC, id DESC
            LIMIT 1
            """,
            (asset_id, previous_status),
        ).fetchone()
        if latest_problem:
            started_at = parse_iso_datetime(latest_problem["created_at"] or latest_problem["record_date"])
            ended_at = parse_iso_datetime(happened_at)
            if started_at and ended_at and ended_at >= started_at:
                hours = max(1, round((ended_at - started_at).total_seconds() / 3600))
                duration = f"{hours}h"

    return {
        "asset_id": asset_id,
        "project_name": asset["project_name"],
        "previous_status": previous_status,
        "current_status": current_status,
        "happened_at": happened_at,
        "alert_type": alert_type,
        "duration": duration,
        "primary_alarm_name": (alarm_context or {}).get("primary_alarm_name", ""),
        "primary_alarm_device": (alarm_context or {}).get("primary_alarm_device", ""),
        "primary_alarm_severity": (alarm_context or {}).get("primary_alarm_severity", ""),
        "primary_alarm_raised_at": (alarm_context or {}).get("primary_alarm_raised_at", ""),
        "alarm_summary": (alarm_context or {}).get("alarm_summary", ""),
    }


def process_monitoring_alerts(
    conn: sqlite3.Connection,
    events: list[dict[str, Any]],
    batch_id: int | None,
    now: datetime | None = None,
) -> None:
    now = now or datetime.now()
    if not events:
        process_persistent_monitoring_alerts(conn, now)
        return

    ready_alerts: list[dict[str, Any]] = []
    blocked_counts: dict[str, int] = {}
    for event in events:
        asset = conn.execute("SELECT * FROM assets WHERE id = ?", (event["asset_id"],)).fetchone()
        if asset is None:
            continue
        alert_type, message = build_state_change_message(event)
        alert_key = f"{event['asset_id']}:{alert_type}:batch:{batch_id}:to:{event['current_status']}"
        allowed, reason = alert_decision(conn, asset, alert_type, alert_key, now)
        if allowed:
            ready_alerts.append({"event": event, "alert_type": alert_type, "alert_key": alert_key, "message": message})
        else:
            blocked_counts[reason] = blocked_counts.get(reason, 0) + 1
            record_telegram_alert(conn, event["asset_id"], alert_type, alert_key, message, "blocked", "", reason)

    disconnected_alerts = [item for item in ready_alerts if item["alert_type"] == "nova_desconexao"]
    if len(disconnected_alerts) > 5 and len(ready_alerts) <= 10:
        ready_alerts = [item for item in ready_alerts if item["alert_type"] != "nova_desconexao"]
        alert_key = f"geral_desconexoes_batch_{batch_id or now.isoformat(timespec='seconds')}"
        message = (
            "⚠️ <b>ALERTA GERAL — Multiplas desconexoes</b>\n\n"
            f"{len(disconnected_alerts)} instalacoes ficaram Desconectadas nesta atualizacao.\n"
            "Possivel problema de comunicacao/plataforma/importacao."
        )
        ready_alerts.append({"event": {"asset_id": None}, "alert_type": "geral_multiplas_desconexoes", "alert_key": alert_key, "message": message})

    if len(ready_alerts) > 10:
        for item in ready_alerts:
            record_telegram_alert(conn, item["event"].get("asset_id"), item["alert_type"], item["alert_key"], item["message"], "blocked", "", "batch_aggregated")
        message = (
            "⚠️ <b>Muitos alertas filtrados</b>\n\n"
            f"Foram detetados {len(events)} eventos de monitorizacao.\n"
            "Enviados: 1\n"
            f"Bloqueados por filtros: {sum(blocked_counts.values())}\n"
            f"Blacklisted: {blocked_counts.get('blacklist', 0)}\n"
            f"Fora de O&amp;M: {blocked_counts.get('out_of_scope', 0)}\n\n"
            "Ver detalhes na pagina Alertas Telegram."
        )
        send_and_record_telegram_alert(conn, None, "batch_many_alerts", f"batch_many_alerts:{batch_id or now.isoformat(timespec='seconds')}", message)
    else:
        for item in ready_alerts:
            send_and_record_telegram_alert(conn, item["event"].get("asset_id"), item["alert_type"], item["alert_key"], item["message"])

    process_persistent_monitoring_alerts(conn, now)


def process_persistent_monitoring_alerts(conn: sqlite3.Connection, now: datetime | None = None) -> None:
    now = now or datetime.now()
    latest_rows = query_all(
        conn,
        """
        SELECT a.*, lm.status, lm.record_date
        FROM assets a
        JOIN latest_monitoring_view lm ON lm.asset_id = a.id
        WHERE lm.status IN ('Erro', 'Desconectada')
        """,
    )
    for asset in latest_rows:
        problem_start = find_problem_start(conn, int(asset["id"]), asset["status"])
        problem_start_dt = parse_iso_datetime(problem_start)
        if not problem_start_dt:
            continue
        if alert_setting_bool(conn, "IGNORE_HISTORICAL_ALERTS", True) and is_before_alert_baseline(conn, problem_start):
            continue
        age = now - problem_start_dt
        if asset["status"] == "Erro" and age >= timedelta(hours=24):
            alert_key = f"{asset['id']}:erro_persistente_24h:{problem_start}"
            message = (
                "🚨 <b>ERRO PERSISTENTE</b>\n\n"
                f"Instalacao: {html_line(asset['project_name'])}\n"
                "Estado: Erro\n"
                "Duracao: &gt;24h\n"
                f"Primeira detecao: {html_line(problem_start)}"
            )
            allowed, reason = alert_decision(conn, asset, "erro_persistente_24h", alert_key, now)
            if allowed:
                send_and_record_telegram_alert(conn, int(asset["id"]), "erro_persistente_24h", alert_key, message)
            elif not alert_already_sent(conn, alert_key):
                record_telegram_alert(conn, int(asset["id"]), "erro_persistente_24h", alert_key, message, "blocked", "", reason)
        offline_daytime_only = alert_setting_bool(conn, "DAYTIME_OFFLINE_ONLY", True)
        if asset["status"] == "Desconectada" and age >= timedelta(hours=2) and (not offline_daytime_only or is_daytime_for_alert(now)):
            alert_key = f"{asset['id']}:desconexao_persistente_2h:{problem_start}"
            message = (
                "⚠️ <b>DESCONEXAO PERSISTENTE</b>\n\n"
                f"Instalacao: {html_line(asset['project_name'])}\n"
                "Estado: Desconectada\n"
                "Duracao: &gt;2h em periodo de producao"
            )
            allowed, reason = alert_decision(conn, asset, "desconexao_persistente_2h", alert_key, now)
            if allowed:
                send_and_record_telegram_alert(conn, int(asset["id"]), "desconexao_persistente_2h", alert_key, message)
            elif not alert_already_sent(conn, alert_key):
                record_telegram_alert(conn, int(asset["id"]), "desconexao_persistente_2h", alert_key, message, "blocked", "", reason)

        since_date = (now.date() - timedelta(days=7)).isoformat()
        baseline_at = get_alert_baseline_at(conn)
        if baseline_at:
            since_date = max(since_date, baseline_at.date().isoformat())
        occurrences = count_problem_occurrences_since(conn, int(asset["id"]), since_date)
        if occurrences >= 3:
            alert_key = f"{asset['id']}:recorrente_7d:{now.date().isoformat()}"
            message = (
                "🔁 <b>ERRO RECORRENTE</b>\n\n"
                f"Instalacao: {html_line(asset['project_name'])}\n"
                f"Ocorrencias nos ultimos 7 dias: {occurrences}\n"
                f"Ultimo estado: {html_line(asset['status'])}"
            )
            allowed, reason = alert_decision(conn, asset, "recorrente_7d", alert_key, now)
            if allowed:
                send_and_record_telegram_alert(conn, int(asset["id"]), "recorrente_7d", alert_key, message)
            elif not alert_already_sent(conn, alert_key):
                record_telegram_alert(conn, int(asset["id"]), "recorrente_7d", alert_key, message, "blocked", "", reason)


def format_summary_list(rows: list[sqlite3.Row], empty: str = "-") -> str:
    if not rows:
        return empty
    lines = []
    for row in rows[:8]:
        name = html_line(row["project_name"])
        status = html_line(row["status"] if "status" in row.keys() else "")
        lines.append(f"- {name}: {status}" if status != "-" else f"- {name}")
    if len(rows) > 8:
        lines.append(f"- ... mais {len(rows) - 8}")
    return "\n".join(lines)


def send_daily_telegram_summary(conn: sqlite3.Connection, now: datetime | None = None) -> bool:
    now = now or datetime.now()
    if not telegram_daily_summary_enabled() or not is_telegram_configured():
        return False
    yesterday = (now.date() - timedelta(days=1)).isoformat()
    alert_key = f"daily_summary:{now.date().isoformat()}"
    if alert_already_sent(conn, alert_key):
        return False

    current_error = query_scalar(conn, "SELECT COUNT(*) FROM latest_monitoring_view WHERE status = 'Erro'") or 0
    current_disconnected = query_scalar(conn, "SELECT COUNT(*) FROM latest_monitoring_view WHERE status = 'Desconectada'") or 0
    new_rows = query_all(
        conn,
        """
        SELECT a.project_name, mr.status
        FROM monitoring_records mr
        JOIN assets a ON a.id = mr.asset_id
        WHERE mr.record_date >= ? AND mr.status IN ('Erro', 'Desconectada')
        ORDER BY mr.record_date DESC, mr.id DESC
        LIMIT 20
        """,
        (yesterday,),
    )
    resolved_rows = query_all(
        conn,
        """
        SELECT a.project_name, mr.status
        FROM monitoring_records mr
        JOIN assets a ON a.id = mr.asset_id
        WHERE mr.record_date >= ? AND mr.status IN ('Resolvido', 'Operacional')
        ORDER BY mr.record_date DESC, mr.id DESC
        LIMIT 20
        """,
        (yesterday,),
    )
    persistent_rows = query_all(
        conn,
        """
        SELECT a.project_name, lm.status
        FROM latest_monitoring_view lm
        JOIN assets a ON a.id = lm.asset_id
        WHERE lm.status IN ('Erro', 'Desconectada')
        ORDER BY a.project_name COLLATE NOCASE
        LIMIT 20
        """,
    )
    muted_rows = query_all(
        conn,
        """
        SELECT project_name, monitoring_status AS status
        FROM assets
        WHERE monitoring_status IN ('silenced', 'maintenance')
        ORDER BY project_name COLLATE NOCASE
        LIMIT 20
        """,
    )
    recurring_rows = []
    for row in persistent_rows:
        asset = conn.execute("SELECT id FROM assets WHERE project_name = ?", (row["project_name"],)).fetchone()
        if asset and count_problem_occurrences_since(conn, int(asset["id"]), (now.date() - timedelta(days=7)).isoformat()) >= 3:
            recurring_rows.append(row)

    message = (
        f"<b>Resumo O&amp;M - {html_line(now.date().isoformat())}</b>\n\n"
        "Ativos:\n"
        f"- Erro: {current_error}\n"
        f"- Desconectadas: {current_disconnected}\n\n"
        "Novos desde ontem:\n"
        f"{format_summary_list(new_rows)}\n\n"
        "Persistentes:\n"
        f"{format_summary_list(persistent_rows)}\n\n"
        "Resolvidos desde ontem:\n"
        f"{format_summary_list(resolved_rows)}\n\n"
        "Recorrentes:\n"
        f"{format_summary_list(recurring_rows)}\n\n"
        "Instalacoes silenciadas/manutencao:\n"
        f"{format_summary_list(muted_rows)}"
    )
    return send_and_record_telegram_alert(conn, None, "daily_summary", alert_key, message)


def excel_date_to_iso(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value in (None, "", "-"):
        return ""
    return str(value)


def row_value(row: tuple[Any, ...], index: int) -> str:
    value = row[index] if index < len(row) else None
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def get_sheet(workbook, expected_name: str):
    normalized_expected = normalize_name(expected_name)
    for sheet_name in workbook.sheetnames:
        if normalize_name(sheet_name) == normalized_expected:
            return workbook[sheet_name]
    return None


def find_asset_id(conn: sqlite3.Connection, name: str) -> int | None:
    normalized = normalize_name(name)
    row = conn.execute(
        """
        SELECT asset_id
        FROM asset_aliases
        WHERE normalized_alias = ?
        """,
        (normalized,),
    ).fetchone()
    return int(row["asset_id"]) if row else None


def rebuild_asset_alias_blob(conn: sqlite3.Connection, asset_id: int) -> None:
    aliases = [row["alias_name"] for row in query_all(conn, "SELECT alias_name FROM asset_aliases WHERE asset_id = ?", (asset_id,))]
    conn.execute("UPDATE assets SET alias_blob = ? WHERE id = ?", (" | ".join(aliases), asset_id))
    conn.commit()


def import_excel_data(conn: sqlite3.Connection, excel_path: Path) -> dict[str, int]:
    workbook = load_workbook(excel_path, data_only=True)
    excel_batch_id = create_monitoring_batch(
        conn,
        record_date=date.today().isoformat(),
        default_notes="Sincronizacao a partir do Excel.",
        raw_input="",
        source="excel-import",
    )

    assets_by_name: dict[str, int] = {}
    project_sheet = get_sheet(workbook, "Project Overview")
    if project_sheet is None:
        raise ValueError("Folha 'Project Overview' nao encontrada no Excel.")

    for row in project_sheet.iter_rows(min_row=2, values_only=True):
        project_name = row_value(row, 1)
        if not project_name:
            continue
        payload = {
            "project_number": row_value(row, 0),
            "project_name": project_name,
            "company_name": row_value(row, 2),
            "nif": row_value(row, 3),
            "address": row_value(row, 4),
            "location": row_value(row, 5),
            "panels": row_value(row, 6),
            "kwp": row_value(row, 7),
            "contract_type": row_value(row, 8),
            "sell_to": row_value(row, 9),
            "duration": row_value(row, 10),
            "start_contract": row_value(row, 12),
            "maintenance": row_value(row, 13),
            "coverage_type": row_value(row, 14),
            "access_type": row_value(row, 15),
            "maintenance_comment": row_value(row, 16),
            "status_detail": row_value(row, 17),
            "contact_name": row_value(row, 18),
            "contact_role": row_value(row, 19),
            "contact_email": row_value(row, 20),
            "contact_phone": row_value(row, 21),
            "end_contract": row_value(row, 22),
            "active_contract": row_value(row, 24),
            "notes": row_value(row, 25),
            "asset_type": row_value(row, 27),
        }
        asset_id = upsert_asset_from_excel(conn, payload)
        assets_by_name[project_name] = asset_id
        alias_names = {project_name, payload["company_name"]}
        for alias_name in alias_names:
            if alias_name:
                normalized = normalize_name(alias_name)
                if normalized:
                    conn.execute(
                        "INSERT OR IGNORE INTO asset_aliases (asset_id, alias_name, normalized_alias, source) VALUES (?, ?, ?, ?)",
                        (asset_id, alias_name, normalized, "excel"),
                    )

    monitoring_imported = 0
    monitoring_sheet = get_sheet(workbook, "Monotorizacao")
    if monitoring_sheet is not None:
        for row in monitoring_sheet.iter_rows(min_row=3, values_only=True):
            display_name = row_value(row, 0)
            status = normalize_status(row_value(row, 1))
            record_date = excel_date_to_iso(row[4] if len(row) > 4 else None)
            notes = row_value(row, 5)
            original_name = row_value(row, 6) or display_name
            if not status or not original_name:
                continue
            asset_id = find_asset_id(conn, original_name) or find_asset_id(conn, display_name)
            if asset_id:
                record_date_value = record_date or date.today().isoformat()
                existing = conn.execute(
                    """
                    SELECT 1
                    FROM monitoring_records
                    WHERE asset_id = ? AND status = ? AND record_date = ? AND source = 'excel'
                    LIMIT 1
                    """,
                    (asset_id, status, record_date_value),
                ).fetchone()
                if not existing:
                    conn.execute(
                        """
                        INSERT INTO monitoring_records (asset_id, status, record_date, notes, source, batch_id)
                        VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (asset_id, status, record_date_value, notes, "excel", excel_batch_id),
                    )
                    monitoring_imported += 1
                for alias_candidate in {display_name, original_name}:
                    normalized = normalize_name(alias_candidate)
                    if normalized:
                        conn.execute(
                            "INSERT OR IGNORE INTO asset_aliases (asset_id, alias_name, normalized_alias, source) VALUES (?, ?, ?, ?)",
                            (asset_id, alias_candidate, normalized, "excel-monitoring"),
                        )

    tickets_imported = 0
    corrective_sheet = get_sheet(workbook, "Corretivas")
    if corrective_sheet is not None:
        carried_asset_name = ""
        carried_notes = ""
        for row in corrective_sheet.iter_rows(min_row=2, values_only=True):
            asset_name = row_value(row, 0) or carried_asset_name
            installation_ref = row_value(row, 1)
            contract_type = row_value(row, 2)
            created_at = excel_date_to_iso(row[3])
            status = normalize_status(row_value(row, 4))
            notes = row_value(row, 5)
            next_action = row_value(row, 6)

            if row_value(row, 0):
                carried_asset_name = row_value(row, 0)
            if notes:
                carried_notes = notes
            elif not row_value(row, 0) and carried_notes:
                notes = carried_notes

            if not asset_name or not (status or notes or next_action):
                continue

            asset_id = find_asset_id(conn, asset_name)
            if not asset_id:
                continue

            urgency = "Alta" if status in {"Aberto", "Em analise"} else "Media"
            title = next_action.splitlines()[0][:120] if next_action else f"Corretiva - {asset_name}"
            created_at_value = created_at or date.today().isoformat()
            existing = conn.execute(
                """
                SELECT 1
                FROM tickets
                WHERE asset_id = ? AND title = ? AND created_at = ?
                LIMIT 1
                """,
                (asset_id, title, created_at_value),
            ).fetchone()
            if not existing:
                conn.execute(
                    """
                    INSERT INTO tickets (
                        asset_id, title, urgency, status, installation_ref, notes, next_action, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        asset_id,
                        title,
                        urgency,
                        status or "Aberto",
                        installation_ref or contract_type,
                        notes,
                        next_action,
                        created_at_value,
                        created_at_value,
                    ),
                )
                tickets_imported += 1

    for asset_id_row in query_all(conn, "SELECT id FROM assets"):
        rebuild_asset_alias_blob(conn, int(asset_id_row["id"]))

    conn.execute(
        """
        UPDATE monitoring_import_batches
        SET imported_count = ?, matched_count = ?, unmatched_count = 0, auto_resolved_count = 0
        WHERE id = ?
        """,
        (monitoring_imported, monitoring_imported, excel_batch_id),
    )
    conn.commit()
    return {"assets": len(assets_by_name), "monitoring": monitoring_imported, "tickets": tickets_imported}


def upsert_asset_from_excel(conn: sqlite3.Connection, payload: dict[str, str]) -> int:
    project_name = payload["project_name"]
    existing = conn.execute(
        """
        SELECT id, installation_group
        FROM assets
        WHERE project_name = ?
        LIMIT 1
        """,
        (project_name,),
    ).fetchone()
    asset_id = int(existing["id"]) if existing else (find_asset_id(conn, project_name) or 0)
    installation_group = (
        (existing["installation_group"] if existing and existing["installation_group"] else "")
        or infer_installation_group(payload["project_name"])
    )
    payload["start_contract"] = normalize_date_value(payload["start_contract"])
    payload["end_contract"] = normalize_date_value(payload["end_contract"])
    payload["active_contract"] = derive_active_contract(payload["end_contract"], payload["active_contract"])
    payload = apply_group_defaults(conn, payload, installation_group, exclude_asset_id=asset_id or None)

    values = (
        payload["project_number"],
        payload["project_name"],
        installation_group,
        payload["company_name"],
        payload["nif"],
        payload["address"],
        payload["location"],
        payload["panels"],
        payload["kwp"],
        payload["contract_type"],
        payload["sell_to"],
        payload["duration"],
        payload["start_contract"],
        payload["maintenance"],
        payload["coverage_type"],
        payload["access_type"],
        payload["maintenance_comment"],
        payload["status_detail"],
        payload["contact_name"],
        payload["contact_role"],
        payload["contact_email"],
        payload["contact_phone"],
        payload["end_contract"],
        payload["active_contract"],
        payload["notes"],
        payload["asset_type"],
        json.dumps(payload, ensure_ascii=True),
    )

    if asset_id:
        conn.execute(
            """
            UPDATE assets
            SET
                project_number = ?, project_name = ?, installation_group = ?, company_name = ?, nif = ?, address = ?, location = ?,
                panels = ?, kwp = ?, contract_type = ?, sell_to = ?, duration = ?, start_contract = ?,
                maintenance = ?, coverage_type = ?, access_type = ?, maintenance_comment = ?, status_detail = ?,
                contact_name = ?, contact_role = ?, contact_email = ?, contact_phone = ?, end_contract = ?,
                active_contract = ?, notes = ?, asset_type = ?, source_payload = ?
            WHERE id = ?
            """,
            values + (asset_id,),
        )
        return asset_id

    cursor = conn.execute(
        """
        INSERT INTO assets (
            project_number, project_name, installation_group, company_name, nif, address, location, panels, kwp,
            contract_type, sell_to, duration, start_contract, maintenance, coverage_type,
            access_type, maintenance_comment, status_detail, contact_name, contact_role,
            contact_email, contact_phone, end_contract, active_contract, notes, asset_type,
            source_payload, alias_blob
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        values + (project_name,),
    )
    return int(cursor.lastrowid)


@dataclass
class MonitoringImportResult:
    imported: int = 0
    matched: int = 0
    unmatched: int = 0
    auto_resolved: int = 0
    batch_id: int | None = None


def parse_monitoring_lines(pasted_table: str) -> list[tuple[str, str]]:
    lines = []
    for raw_line in pasted_table.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        lowered = normalize_name(line)
        if lowered in {"instalacao estado", "instalacao", "estado"}:
            continue
        if "\t" in line:
            parts = [part.strip() for part in line.split("\t") if part.strip()]
            if len(parts) >= 2:
                first_part = normalize_name(parts[0])
                second_part = normalize_name(parts[1])
                if first_part in {"instalacao", "instalacao estado"} or second_part == "estado":
                    continue
                lines.append((parts[0], normalize_status(parts[1])))
                continue
        for marker in (" Erro", " Desconectada", " Operacional", " Resolvido"):
            if line.endswith(marker):
                lines.append((line[: -len(marker)].strip(), normalize_status(marker.strip())))
                break
        else:
            parts = line.rsplit(" ", 1)
            if len(parts) == 2:
                lines.append((parts[0].strip(), normalize_status(parts[1].strip())))
    return lines


def import_daily_monitoring(
    conn: sqlite3.Connection,
    pasted_table: str,
    record_date: str,
    default_notes: str,
    platform_source: str,
    import_scope: str = "complete",
) -> MonitoringImportResult:
    result = MonitoringImportResult()
    platform_source = normalize_monitoring_source(platform_source)
    parsed_lines = parse_monitoring_lines(pasted_table)
    if not parsed_lines:
        return result
    batch_id = create_monitoring_batch(conn, record_date, default_notes, pasted_table, platform_source)
    result.batch_id = batch_id
    imported_asset_ids: set[int] = set()
    alert_events: list[dict[str, Any]] = []
    now = datetime.now()
    for original_name, status in parsed_lines:
        asset_id = find_asset_id(conn, original_name)
        if asset_id:
            duplicate = conn.execute(
                """
                SELECT 1
                FROM monitoring_records
                WHERE asset_id = ? AND status = ? AND record_date = ? AND source = ?
                LIMIT 1
                """,
                (asset_id, status, record_date, platform_source),
            ).fetchone()
            if duplicate:
                continue
            previous = get_latest_monitoring_row(conn, asset_id)
            result.imported += 1
            imported_asset_ids.add(asset_id)
            conn.execute(
                """
                INSERT INTO monitoring_records (asset_id, status, record_date, notes, source, batch_id)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (asset_id, status, record_date, default_notes, platform_source, batch_id),
            )
            result.matched += 1
            event = build_monitoring_alert_event(
                conn,
                asset_id=asset_id,
                previous_status=previous["status"] if previous else "",
                current_status=status,
                happened_at=now.isoformat(timespec="seconds"),
            )
            if event:
                alert_events.append(event)
        else:
            normalized_name = normalize_name(original_name)
            duplicate_unmatched = conn.execute(
                """
                SELECT 1
                FROM monitoring_unmatched
                WHERE normalized_name = ? AND status = ? AND record_date = ?
                LIMIT 1
                """,
                (normalized_name, status, record_date),
            ).fetchone()
            if duplicate_unmatched:
                continue
            result.imported += 1
            conn.execute(
                """
                INSERT INTO monitoring_unmatched (original_name, normalized_name, status, record_date, notes, batch_id)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (original_name, normalized_name, status, record_date, default_notes, batch_id),
            )
            result.unmatched += 1

    if import_scope == "complete":
        latest_problem_assets = query_all(
            conn,
            """
            SELECT lm.asset_id
            FROM latest_monitoring_view lm
            WHERE lm.status IN ('Erro', 'Desconectada')
            """,
        )
        for row in latest_problem_assets:
            asset_id = int(row["asset_id"])
            if asset_id in imported_asset_ids:
                continue
            existing_today = conn.execute(
                """
                SELECT 1
                FROM monitoring_records
                WHERE asset_id = ? AND record_date = ?
                LIMIT 1
                """,
                (asset_id, record_date),
            ).fetchone()
            if existing_today:
                continue
            previous = get_latest_monitoring_row(conn, asset_id)
            conn.execute(
                """
                INSERT INTO monitoring_records (asset_id, status, record_date, notes, source, batch_id)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    asset_id,
                    "Resolvido",
                    record_date,
                    "Resolvido automaticamente por nao constar na lista diaria.",
                    platform_source,
                    batch_id,
                ),
            )
            result.auto_resolved += 1
            event = build_monitoring_alert_event(
                conn,
                asset_id=asset_id,
                previous_status=previous["status"] if previous else "",
                current_status="Resolvido",
                happened_at=now.isoformat(timespec="seconds"),
            )
            if event:
                alert_events.append(event)
    conn.execute(
        """
        UPDATE monitoring_import_batches
        SET imported_count = ?, matched_count = ?, unmatched_count = ?, auto_resolved_count = ?
        WHERE id = ?
        """,
        (result.imported, result.matched, result.unmatched, result.auto_resolved, batch_id),
    )
    process_monitoring_alerts(conn, alert_events, batch_id, now)
    conn.commit()
    return result


def create_monitoring_batch(
    conn: sqlite3.Connection,
    record_date: str,
    default_notes: str,
    raw_input: str,
    source: str,
) -> int:
    source = normalize_monitoring_source(source)
    cursor = conn.execute(
        """
        INSERT INTO monitoring_import_batches (record_date, imported_at, source, default_notes, raw_input)
        VALUES (?, ?, ?, ?, ?)
        """,
        (record_date, datetime.now().isoformat(timespec="seconds"), source, default_notes, raw_input),
    )
    return int(cursor.lastrowid)


def build_problem_periods(conn: sqlite3.Connection, asset_id: int) -> list[dict[str, Any]]:
    rows = query_all(
        conn,
        """
        SELECT mr.record_date, mr.status, mr.notes, mr.source, mib.imported_at
        FROM monitoring_records mr
        LEFT JOIN monitoring_import_batches mib ON mib.id = mr.batch_id
        WHERE mr.asset_id = ?
        ORDER BY mr.record_date ASC, mr.id ASC
        """,
        (asset_id,),
    )
    problem_statuses = {"Erro", "Desconectada"}
    periods: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None

    for row in rows:
        status = row["status"]
        if status in problem_statuses:
            if current is None:
                current = {
                    "status": status,
                    "started_on": row["record_date"],
                    "started_at": row["imported_at"],
                    "start_notes": row["notes"],
                    "last_problem_on": row["record_date"],
                    "last_problem_status": status,
                }
            else:
                current["last_problem_on"] = row["record_date"]
                current["last_problem_status"] = status
            continue

        if current is not None:
            current["resolved_on"] = row["record_date"]
            current["resolved_at"] = row["imported_at"]
            current["resolution_status"] = status
            current["resolution_notes"] = row["notes"]
            periods.append(current)
            current = None

    if current is not None:
        current["resolved_on"] = None
        current["resolved_at"] = None
        current["resolution_status"] = "Ainda ativo"
        current["resolution_notes"] = ""
        periods.append(current)

    periods.reverse()
    return periods


def build_problem_metric_map(conn: sqlite3.Connection, asset_ids: list[int]) -> dict[int, dict[str, Any]]:
    metrics: dict[int, dict[str, Any]] = {}
    cutoff = (date.today() - timedelta(days=90)).isoformat()
    for asset_id in asset_ids:
        periods = build_problem_periods(conn, asset_id)
        active_period = next((period for period in periods if period["resolved_on"] is None), None)
        recurrence_count = sum(
            1
            for period in periods
            if (period.get("started_on") or "") >= cutoff
        )
        metrics[asset_id] = {
            "problem_started_on": active_period["started_on"] if active_period else "",
            "problem_days": days_between(active_period["started_on"]) if active_period else 0,
            "recurrence_count": recurrence_count,
            "last_problem_status": active_period["last_problem_status"] if active_period else "",
        }
    return metrics


def enrich_operational_rows(conn: sqlite3.Connection, rows: list[sqlite3.Row | dict[str, Any]]) -> list[dict[str, Any]]:
    enriched = [dict(row) for row in rows]
    asset_ids = [int(row["asset_id"] if "asset_id" in row else row["id"]) for row in enriched]
    metric_map = build_problem_metric_map(conn, asset_ids)
    ticket_counts: dict[int, int] = {}
    if asset_ids:
        ticket_counts = {
            int(row["asset_id"]): int(row["open_tickets"])
            for row in query_all(
                conn,
                f"""
                SELECT asset_id, COUNT(*) AS open_tickets
                FROM tickets
                WHERE status != 'Fechado' AND asset_id IN ({",".join("?" for _ in asset_ids)})
                GROUP BY asset_id
                """,
                asset_ids,
            )
        }
    for row in enriched:
        asset_id = int(row["asset_id"] if "asset_id" in row else row["id"])
        row.setdefault("id", asset_id)
        metrics = metric_map.get(asset_id, {})
        row.update(metrics)
        row["open_tickets"] = int(row.get("open_tickets") or ticket_counts.get(asset_id, 0) or 0)
        row["auto_priority"] = auto_priority(
            row.get("status"),
            int(row.get("problem_days") or 0),
            int(row.get("recurrence_count") or 0),
            int(row.get("open_tickets") or 0),
            row.get("active_contract"),
        )
    return enriched


def priority_rank(priority: str) -> int:
    return {"Critica": 1, "Alta": 2, "Media": 3, "Baixa": 4}.get(priority, 5)


def build_batch_insight(conn: sqlite3.Connection, batch_id: int) -> dict[str, Any] | None:
    batch = conn.execute(
        """
        SELECT *
        FROM monitoring_import_batches
        WHERE id = ?
        """,
        (batch_id,),
    ).fetchone()
    if batch is None:
        return None

    rows = query_all(
        conn,
        """
        SELECT
            mr.id,
            mr.asset_id,
            mr.status,
            mr.record_date,
            a.project_name,
            a.active_contract
        FROM monitoring_records mr
        JOIN assets a ON a.id = mr.asset_id
        WHERE mr.batch_id = ?
        ORDER BY mr.id ASC
        """,
        (batch_id,),
    )

    problem_statuses = {"Erro", "Desconectada"}
    new_problem_assets: list[dict[str, Any]] = []
    persistent_problem_assets: list[dict[str, Any]] = []
    resolved_assets: list[dict[str, Any]] = []

    for row in rows:
        previous = conn.execute(
            """
            SELECT mr.status, mr.record_date
            FROM monitoring_records mr
            WHERE mr.asset_id = ? AND mr.id < ?
            ORDER BY mr.id DESC
            LIMIT 1
            """,
            (row["asset_id"], row["id"]),
        ).fetchone()
        previous_status = previous["status"] if previous else ""
        current_status = row["status"]
        item = {
            "asset_id": row["asset_id"],
            "project_name": row["project_name"],
            "current_status": current_status,
            "previous_status": previous_status or "-",
            "record_date": row["record_date"],
        }

        if row["active_contract"] == "yes" and current_status in problem_statuses:
            if previous_status not in problem_statuses:
                new_problem_assets.append(item)
            else:
                persistent_problem_assets.append(item)
        elif row["active_contract"] == "yes" and current_status in {"Resolvido", "Operacional"} and previous_status in problem_statuses:
            resolved_assets.append(item)

    return {
        "batch": batch,
        "new_problem_assets": new_problem_assets,
        "persistent_problem_assets": persistent_problem_assets,
        "resolved_assets": resolved_assets,
    }


def extract_export_filters(source: Any, dataset: str, for_query: bool = False) -> dict[str, str]:
    get_value = source.get
    filters = {
        "search": get_value("search", "").strip(),
        "asset_id": get_value("asset_id", "").strip(),
        "om_only": get_value("om_only", "yes").strip(),
    }
    if dataset == "monitoring":
        filters.update(
            {
                "status": get_value("status", "").strip(),
                "source": get_value("source", "").strip(),
                "start_date": get_value("start_date", "").strip(),
                "end_date": get_value("end_date", "").strip(),
            }
        )
    elif dataset in {"executive_report", "monitoring_report", "production_report"}:
        filters.update(
            {
                "period": get_value("period", "week").strip() or "week",
                "source": get_value("source", "").strip(),
                "report_month": get_value("report_month", "").strip(),
                "report_year": get_value("report_year", "").strip(),
            }
        )
    else:
        filters.update(
            {
                "status": get_value("status", "").strip(),
                "urgency": get_value("urgency", "").strip(),
            }
        )
    if for_query:
        if dataset == "production_report" and filters.get("period") not in {"month", "year"}:
            filters["period"] = "month"
        return filters
    if dataset == "production_report" and filters.get("period") not in {"month", "year"}:
        filters["period"] = "month"
    return {key: value for key, value in filters.items() if value}


def normalize_report_month(value: str | None) -> str:
    if value:
        try:
            return datetime.strptime(value.strip(), "%Y-%m").strftime("%Y-%m")
        except ValueError:
            pass
    return date.today().strftime("%Y-%m")


def normalize_report_year(value: str | None) -> int:
    if value and value.strip().isdigit():
        year = int(value.strip())
        if 2000 <= year <= date.today().year + 1:
            return year
    return date.today().year


def report_period_dates(period: str, report_month: str | None = None, report_year: str | None = None) -> tuple[str, str]:
    today = date.today()
    if period == "day":
        return today.isoformat(), today.isoformat()
    if period == "year":
        year = normalize_report_year(report_year)
        start = date(year, 1, 1)
        end = date(year, 12, 31)
        if year == today.year:
            end = today
        return start.isoformat(), end.isoformat()
    if period == "month":
        month_value = normalize_report_month(report_month)
        month_start = datetime.strptime(month_value, "%Y-%m").date()
        _, last_day = calendar.monthrange(month_start.year, month_start.month)
        month_end = month_start.replace(day=last_day)
        if month_start.year == today.year and month_start.month == today.month:
            month_end = today
        return month_start.isoformat(), month_end.isoformat()
    return (today - timedelta(days=7)).isoformat(), today.isoformat()


def report_period_label(period: str, report_month: str | None = None, report_year: str | None = None) -> str:
    start_date, end_date = report_period_dates(period, report_month, report_year)
    labels = {
        "day": "Diario",
        "week": "Semanal",
        "month": "Mensal",
        "year": "Anual",
    }
    return f"{labels.get(period, 'Semanal')} ({start_date} a {end_date})"


def report_sync_bounds(filters: dict[str, str]) -> tuple[int, int, date, date]:
    period = filters.get("period", "month")
    start_raw, end_raw = report_period_dates(period, filters.get("report_month"), filters.get("report_year"))
    start = datetime.strptime(start_raw, "%Y-%m-%d").date()
    end = datetime.strptime(end_raw, "%Y-%m-%d").date()
    return start.year, end.year, start, end


def build_monitoring_report_rows(
    conn: sqlite3.Connection,
    filters: dict[str, str],
    limit: int | None = None,
) -> list[dict[str, Any]]:
    period = filters.get("period", "week")
    start_date, end_date = report_period_dates(period, filters.get("report_month"), filters.get("report_year"))
    source_filter = filters.get("source", "")

    filter_sql = []
    params: list[Any] = []
    if filters.get("search"):
        wildcard = f"%{filters['search']}%"
        filter_sql.append("(a.project_name LIKE ? OR a.alias_blob LIKE ? OR a.company_name LIKE ? OR a.location LIKE ?)")
        params.extend([wildcard, wildcard, wildcard, wildcard])
    if filters.get("asset_id"):
        filter_sql.append("a.id = ?")
        params.append(filters["asset_id"])
    if filters.get("om_only", "yes") == "yes":
        filter_sql.append("a.active_contract = 'yes'")
    where_sql = f"WHERE {' AND '.join(filter_sql)}" if filter_sql else ""

    assets = query_all(
        conn,
        f"""
        SELECT
            a.id,
            a.project_name,
            a.location,
            a.active_contract,
            lm.status AS current_status,
            lm.record_date AS last_record_date,
            lm.notes AS latest_notes
        FROM assets a
        LEFT JOIN latest_monitoring_view lm ON lm.asset_id = a.id
        {where_sql}
        ORDER BY
            CASE a.active_contract WHEN 'yes' THEN 1 ELSE 2 END,
            a.project_name COLLATE NOCASE
        """,
        params,
    )

    rows: list[dict[str, Any]] = []
    monitoring_source_sql = "AND mr.source = ?" if source_filter else ""
    monitoring_source_params: list[Any] = [source_filter] if source_filter else []
    for asset in assets:
        asset_id = int(asset["id"])
        monitoring_metrics = conn.execute(
            f"""
            SELECT
                COUNT(*) AS monitoring_records,
                SUM(CASE WHEN mr.status IN ('Erro', 'Desconectada') THEN 1 ELSE 0 END) AS error_records,
                COUNT(DISTINCT CASE
                    WHEN mr.status IN ('Erro', 'Desconectada')
                    THEN mr.status || '|' || COALESCE(NULLIF(TRIM(mr.notes), ''), '-')
                END) AS distinct_errors,
                MAX(mr.record_date) AS last_record_date
            FROM monitoring_records mr
            WHERE mr.asset_id = ?
              AND mr.record_date BETWEEN ? AND ?
              {monitoring_source_sql}
            """,
            [asset_id, start_date, end_date] + monitoring_source_params,
        ).fetchone()
        error_rows = query_all(
            conn,
            f"""
            SELECT DISTINCT mr.status, COALESCE(NULLIF(TRIM(mr.notes), ''), '-') AS notes
            FROM monitoring_records mr
            WHERE mr.asset_id = ?
              AND mr.record_date BETWEEN ? AND ?
              AND mr.status IN ('Erro', 'Desconectada')
              {monitoring_source_sql}
            ORDER BY mr.status COLLATE NOCASE, notes COLLATE NOCASE
            """,
            [asset_id, start_date, end_date] + monitoring_source_params,
        )
        open_tickets = int(
            query_scalar(
                conn,
                "SELECT COUNT(*) FROM tickets WHERE asset_id = ? AND status != 'Fechado'",
                (asset_id,),
            )
            or 0
        )
        visit_metrics = conn.execute(
            """
            SELECT COUNT(*) AS visits_period, MAX(tv.visit_date) AS last_visit_date
            FROM ticket_visits tv
            JOIN tickets t ON t.id = tv.ticket_id
            WHERE t.asset_id = ?
              AND tv.visit_date BETWEEN ? AND ?
            """,
            (asset_id, start_date, end_date),
        ).fetchone()

        monitoring_count = int(monitoring_metrics["monitoring_records"] or 0)
        error_count = int(monitoring_metrics["error_records"] or 0)
        visits_period = int(visit_metrics["visits_period"] or 0)
        if monitoring_count == 0 and open_tickets == 0 and visits_period == 0:
            continue

        error_types = []
        for row in error_rows:
            label = row["status"]
            if row["notes"] and row["notes"] != "-":
                label = f"{label}: {row['notes']}"
            error_types.append(label)

        rows.append(
            {
                "period": report_period_label(period, filters.get("report_month"), filters.get("report_year")),
                "project_name": asset["project_name"],
                "location": asset["location"] or "-",
                "current_status": asset["current_status"] or "-",
                "last_record_date": monitoring_metrics["last_record_date"] or asset["last_record_date"] or "-",
                "monitoring_records": monitoring_count,
                "error_records": error_count,
                "distinct_errors": int(monitoring_metrics["distinct_errors"] or 0),
                "error_types": "; ".join(error_types) if error_types else "-",
                "open_tickets": open_tickets,
                "visits_period": visits_period,
                "last_visit_date": visit_metrics["last_visit_date"] or "-",
                "latest_notes": asset["latest_notes"] or "",
            }
        )

    rows.sort(
        key=lambda row: (
            0 if row["current_status"] in {"Erro", "Desconectada"} else 1,
            -int(row["error_records"] or 0),
            -int(row["open_tickets"] or 0),
            -int(row["visits_period"] or 0),
            row["project_name"].lower(),
        )
    )
    return rows[:limit] if limit else rows


def build_executive_report_rows(
    conn: sqlite3.Connection,
    filters: dict[str, str],
    limit: int | None = None,
) -> list[dict[str, Any]]:
    start_date, end_date = report_period_dates(
        filters.get("period", "week"),
        filters.get("report_month"),
        filters.get("report_year"),
    )
    source_filter = filters.get("source", "")
    source_sql = "AND mr.source = ?" if source_filter else ""
    source_params: list[Any] = [source_filter] if source_filter else []

    active_rows = query_all(
        conn,
        f"""
        SELECT
            a.id AS asset_id,
            a.project_name,
            a.active_contract,
            lm.status,
            lm.record_date,
            lm.notes,
            (
                SELECT mr.source
                FROM monitoring_records mr
                WHERE mr.asset_id = a.id
                  AND mr.record_date = lm.record_date
                  AND mr.status = lm.status
                  {source_sql}
                ORDER BY mr.id DESC
                LIMIT 1
            ) AS source
        FROM latest_monitoring_view lm
        JOIN assets a ON a.id = lm.asset_id
        WHERE a.active_contract = 'yes'
          AND lm.status IN ('Erro', 'Desconectada')
        """,
        source_params,
    )
    enriched = [
        row for row in enrich_operational_rows(conn, active_rows)
        if not source_filter or row.get("source")
    ]
    enriched.sort(
        key=lambda row: (
            priority_rank(row["auto_priority"]),
            -int(row.get("problem_days") or 0),
            -int(row.get("recurrence_count") or 0),
            row["project_name"].lower(),
        )
    )

    rows: list[dict[str, Any]] = []
    for row in enriched:
        rows.append(
            {
                "section": "Problemas ativos O&M",
                "priority": row["auto_priority"],
                "project_name": row["project_name"],
                "status": row["status"],
                "problem_days": row["problem_days"],
                "recurrence_count": row["recurrence_count"],
                "open_tickets": row["open_tickets"],
                "source": row.get("source") or "-",
                "notes": row.get("notes") or "",
            }
        )

    resolved_rows = query_all(
        conn,
        f"""
        SELECT
            a.project_name,
            mr.status,
            mr.record_date,
            mr.source,
            mr.notes
        FROM monitoring_records mr
        JOIN assets a ON a.id = mr.asset_id
        WHERE a.active_contract = 'yes'
          AND mr.status = 'Resolvido'
          AND mr.record_date BETWEEN ? AND ?
          {source_sql}
        ORDER BY mr.record_date DESC, a.project_name COLLATE NOCASE
        LIMIT 50
        """,
        [start_date, end_date] + source_params,
    )
    for row in resolved_rows:
        rows.append(
            {
                "section": "Resolvidos no periodo",
                "priority": "-",
                "project_name": row["project_name"],
                "status": row["status"],
                "problem_days": "",
                "recurrence_count": "",
                "open_tickets": "",
                "source": row["source"] or "-",
                "notes": row["notes"] or "",
            }
        )

    return rows[:limit] if limit else rows


def production_report_status(rows: list[sqlite3.Row]) -> str:
    if not rows:
        return "Sem dados"
    ordered = ["Critico", "Crítico", "Alerta", "Atencao", "Atenção", "Sem dados", "Sem referencia", "Sem referência", "OK"]
    statuses = {str(row["performance_status"] or "") for row in rows}
    for status in ordered:
        if status in statuses:
            return status
    return next(iter(statuses), "Sem dados") or "Sem dados"


def build_production_report_rows(
    conn: sqlite3.Connection,
    filters: dict[str, str],
    limit: int | None = None,
) -> list[dict[str, Any]]:
    period = filters.get("period", "month")
    if period not in {"month", "year"}:
        period = "month"
    start_date, end_date = report_period_dates(period, filters.get("report_month"), filters.get("report_year"))
    source_filter = filters.get("source", "")

    filter_sql = []
    params: list[Any] = []
    if filters.get("search"):
        wildcard = f"%{filters['search']}%"
        filter_sql.append("(a.project_name LIKE ? OR a.alias_blob LIKE ? OR a.company_name LIKE ? OR a.location LIKE ?)")
        params.extend([wildcard, wildcard, wildcard, wildcard])
    if filters.get("asset_id"):
        filter_sql.append("a.id = ?")
        params.append(filters["asset_id"])
    if filters.get("om_only", "yes") == "yes":
        filter_sql.append("a.active_contract = 'yes'")
    where_sql = f"WHERE {' AND '.join(filter_sql)}" if filter_sql else ""

    assets = query_all(
        conn,
        f"""
        SELECT a.id, a.project_name, a.location, a.kwp, a.active_contract
        FROM assets a
        {where_sql}
        ORDER BY
            CASE a.active_contract WHEN 'yes' THEN 1 ELSE 2 END,
            a.project_name COLLATE NOCASE
        """,
        params,
    )

    provider_sql = "AND provider = ?" if source_filter else ""
    provider_params: list[Any] = [source_filter] if source_filter else []
    rows: list[dict[str, Any]] = []
    for asset in assets:
        asset_id = int(asset["id"])
        monthly_records = query_all(
            conn,
            f"""
            SELECT *
            FROM production_records
            WHERE asset_id = ?
              AND period_type = 'month'
              AND period_date BETWEEN ? AND ?
              {provider_sql}
            ORDER BY period_date
            """,
            [asset_id, start_date, end_date] + provider_params,
        )
        daily_records = query_all(
            conn,
            f"""
            SELECT *
            FROM production_records
            WHERE asset_id = ?
              AND period_type = 'day'
              AND period_date BETWEEN ? AND ?
              {provider_sql}
            ORDER BY period_date
            """,
            [asset_id, start_date, end_date] + provider_params,
        )
        source_records = monthly_records if monthly_records else daily_records
        production_values = [float(row["production_kwh"]) for row in source_records if row["production_kwh"] is not None]
        expected_values = [float(row["expected_kwh"]) for row in source_records if row["expected_kwh"] is not None]
        production_kwh = sum(production_values) if production_values else None
        expected_kwh = sum(expected_values) if expected_values else None
        kwp = parse_kwp_value(asset["kwp"])
        specific_yield = calculate_specific_yield(production_kwh, kwp)
        deviation_pct = None
        if production_kwh is not None and expected_kwh:
            deviation_pct = ((production_kwh - expected_kwh) / expected_kwh) * 100
        providers = sorted({str(row["provider"] or "") for row in source_records if row["provider"]})
        last_update = max((str(row["updated_at"] or "") for row in source_records), default="")
        if not source_records and filters.get("hide_empty") == "yes":
            continue

        rows.append(
            {
                "period": report_period_label(period, filters.get("report_month"), filters.get("report_year")),
                "project_name": asset["project_name"],
                "location": asset["location"] or "-",
                "provider": ", ".join(providers) if providers else "-",
                "production_kwh": round(production_kwh, 2) if production_kwh is not None else "",
                "specific_yield": round(specific_yield, 2) if specific_yield is not None else "",
                "expected_kwh": round(expected_kwh, 2) if expected_kwh is not None else "",
                "deviation_pct": round(deviation_pct, 2) if deviation_pct is not None else "",
                "performance_status": production_report_status(source_records),
                "data_points": len(source_records),
                "data_source": "KPI mensal API" if monthly_records else ("KPI diario API" if daily_records else "Sem dados API"),
                "last_update": last_update or "-",
                "notes": "" if source_records else "Sem producao sincronizada para o periodo.",
            }
        )

    rows.sort(
        key=lambda row: (
            1 if row["data_points"] else 0,
            row["performance_status"] == "OK",
            row["project_name"].lower(),
        )
    )
    return rows[:limit] if limit else rows


def build_export_dataset(
    conn: sqlite3.Connection,
    dataset: str,
    filters: dict[str, str],
    columns: list[str],
    limit: int | None = None,
) -> tuple[list[dict[str, Any]], list[tuple[str, str]]]:
    headers = [column for column in EXPORT_DATASETS[dataset]["columns"] if column[0] in columns]
    if not headers:
        headers = EXPORT_DATASETS[dataset]["columns"]

    if dataset == "assets":
        filter_sql = []
        params: list[Any] = []
        if filters.get("search"):
            wildcard = f"%{filters['search']}%"
            filter_sql.append("(a.project_name LIKE ? OR a.alias_blob LIKE ? OR a.company_name LIKE ? OR a.location LIKE ?)")
            params.extend([wildcard, wildcard, wildcard, wildcard])
        if filters.get("asset_id"):
            filter_sql.append("a.id = ?")
            params.append(filters["asset_id"])
        if filters.get("om_only", "yes") == "yes":
            filter_sql.append("a.active_contract = 'yes'")
        where_sql = f"WHERE {' AND '.join(filter_sql)}" if filter_sql else ""
        limit_sql = f"LIMIT {limit}" if limit else ""
        rows = query_all(
            conn,
            f"""
            SELECT
                a.project_name,
                a.location,
                a.address,
                a.contact_phone,
                a.contact_name,
                a.access_type,
                a.coverage_type,
                a.contract_type,
                a.active_contract,
                a.company_name,
                a.contact_email
            FROM assets a
            {where_sql}
            ORDER BY
                CASE a.active_contract WHEN 'yes' THEN 1 ELSE 2 END,
                a.project_name COLLATE NOCASE
            {limit_sql}
            """,
            params,
        )
    elif dataset == "monitoring":
        filter_sql = []
        params: list[Any] = []
        if filters.get("search"):
            wildcard = f"%{filters['search']}%"
            filter_sql.append("(a.project_name LIKE ? OR a.alias_blob LIKE ? OR a.company_name LIKE ?)")
            params.extend([wildcard, wildcard, wildcard])
        if filters.get("asset_id"):
            filter_sql.append("a.id = ?")
            params.append(filters["asset_id"])
        if filters.get("status"):
            filter_sql.append("mr.status = ?")
            params.append(filters["status"])
        if filters.get("source"):
            filter_sql.append("mr.source = ?")
            params.append(filters["source"])
        if filters.get("om_only", "yes") == "yes":
            filter_sql.append("a.active_contract = 'yes'")
        if filters.get("start_date"):
            filter_sql.append("mr.record_date >= ?")
            params.append(filters["start_date"])
        if filters.get("end_date"):
            filter_sql.append("mr.record_date <= ?")
            params.append(filters["end_date"])
        where_sql = f"WHERE {' AND '.join(filter_sql)}" if filter_sql else ""
        limit_sql = f"LIMIT {limit}" if limit else ""
        rows = query_all(
            conn,
            f"""
            SELECT
                mr.record_date,
                mib.imported_at,
                a.project_name,
                a.location,
                a.contract_type,
                a.active_contract,
                mr.status,
                mr.notes,
                mr.source
            FROM monitoring_records mr
            JOIN assets a ON a.id = mr.asset_id
            LEFT JOIN monitoring_import_batches mib ON mib.id = mr.batch_id
            {where_sql}
            ORDER BY mr.record_date DESC, a.project_name COLLATE NOCASE, mr.id DESC
            {limit_sql}
            """,
            params,
        )
    elif dataset == "executive_report":
        rows = build_executive_report_rows(conn, filters, limit=limit)
    elif dataset == "monitoring_report":
        rows = build_monitoring_report_rows(conn, filters, limit=limit)
    elif dataset == "production_report":
        rows = build_production_report_rows(conn, filters, limit=limit)
    else:
        filter_sql = []
        params = []
        if filters.get("search"):
            wildcard = f"%{filters['search']}%"
            filter_sql.append("(a.project_name LIKE ? OR a.alias_blob LIKE ? OR t.title LIKE ? OR COALESCE(t.notes, '') LIKE ?)")
            params.extend([wildcard, wildcard, wildcard, wildcard])
        if filters.get("asset_id"):
            filter_sql.append("a.id = ?")
            params.append(filters["asset_id"])
        if filters.get("status"):
            filter_sql.append("t.status = ?")
            params.append(filters["status"])
        if filters.get("urgency"):
            filter_sql.append("t.urgency = ?")
            params.append(filters["urgency"])
        if filters.get("om_only", "yes") == "yes":
            filter_sql.append("a.active_contract = 'yes'")
        where_sql = f"WHERE {' AND '.join(filter_sql)}" if filter_sql else ""
        limit_sql = f"LIMIT {limit}" if limit else ""
        rows = query_all(
            conn,
            f"""
            SELECT
                a.project_name,
                a.location,
                a.contract_type,
                a.active_contract,
                t.title,
                t.status,
                t.urgency,
                t.installation_ref,
                t.next_action,
                t.notes,
                t.created_at,
                t.updated_at
            FROM tickets t
            JOIN assets a ON a.id = t.asset_id
            {where_sql}
            ORDER BY
                CASE a.active_contract WHEN 'yes' THEN 1 ELSE 2 END,
                a.project_name COLLATE NOCASE,
                t.updated_at DESC
            {limit_sql}
            """,
            params,
        )

    normalized_rows = [dict(row) for row in rows]
    return normalized_rows, headers


def get_fusionsolar_report_assets(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    rows = query_all(
        conn,
        """
        SELECT
            a.id AS asset_id,
            a.project_name,
            a.location,
            a.kwp,
            a.contract_type,
            a.asset_type,
            a.coverage_type,
            a.sell_to,
            ai.external_id,
            ai.external_name
        FROM asset_integrations ai
        JOIN assets a ON a.id = ai.asset_id
        WHERE ai.provider = ?
          AND ai.enabled = 1
          AND COALESCE(ai.external_id, '') != ''
        ORDER BY a.project_name COLLATE NOCASE
        """,
        (INTEGRATION_PROVIDER_FUSIONSOLAR,),
    )
    assets = [dict(row) for row in rows]
    for asset in assets:
        asset["report_type"] = detect_report_type(asset)
    return assets


def first_numeric(data: dict[str, Any], keys: list[str]) -> float | None:
    for key in keys:
        value = parse_float_value(data.get(key))
        if value is not None:
            return value
    return None


def normalize_customer_kpi_row(row: dict[str, Any], fallback_date: date) -> dict[str, Any]:
    data_item_map = row.get("dataItemMap") if isinstance(row, dict) else {}
    if not isinstance(data_item_map, dict):
        data_item_map = {}
    production = first_numeric(data_item_map, ["PVYield", "inverterYield", "inverter_power"])
    export = first_numeric(data_item_map, ["ongrid_power", "total_feed_in_to_grid"])
    self_use = first_numeric(data_item_map, ["selfUsePower", "selfProvide"])
    consumption = first_numeric(data_item_map, ["use_power", "day_use_energy"])
    if self_use is None and production is not None:
        self_use = max(production - (export or 0), 0)
    if export is None and production is not None and self_use is not None:
        export = max(production - self_use, 0)
    return {
        "date": parse_fusionsolar_collect_date(row, fallback_date) or fallback_date,
        "production_kwh": production or 0.0,
        "self_use_kwh": self_use or 0.0,
        "export_kwh": export or 0.0,
        "consumption_kwh": consumption,
        "raw": row,
    }


def parse_production_record_payload(row: sqlite3.Row) -> dict[str, Any]:
    try:
        payload = json.loads(row["payload_json"] or "{}")
    except (TypeError, json.JSONDecodeError):
        payload = {}
    return payload if isinstance(payload, dict) else {}


def extract_customer_tariff_values(row: dict[str, Any] | None) -> dict[str, float]:
    data_item_map = row.get("dataItemMap") if isinstance(row, dict) else {}
    if not isinstance(data_item_map, dict):
        data_item_map = {}
    aliases = {
        "self_use_cheia_kwh": ["self_use_cheia_kwh", "selfUsePeak", "selfUseCheia"],
        "self_use_ponta_kwh": ["self_use_ponta_kwh", "selfUseHighPeak", "selfUsePonta"],
        "self_use_vazio_kwh": ["self_use_vazio_kwh", "selfUseOffPeak", "selfUseVazio"],
        "self_use_super_vazio_kwh": ["self_use_super_vazio_kwh", "selfUseSuperOffPeak", "selfUseSuperVazio"],
    }
    return {key: first_numeric(data_item_map, keys) or 0.0 for key, keys in aliases.items()}


def normalize_customer_production_record(row: sqlite3.Row, fallback_date: date) -> dict[str, Any]:
    payload = parse_production_record_payload(row)
    normalized = normalize_customer_kpi_row(payload, fallback_date)
    record_date = parse_date_value(row["period_date"]) or fallback_date
    production_kwh = parse_float_value(row["production_kwh"])
    if production_kwh is not None:
        normalized["production_kwh"] = production_kwh
    normalized["date"] = record_date
    if normalized["production_kwh"] and not normalized["self_use_kwh"] and not normalized["export_kwh"]:
        normalized["self_use_kwh"] = normalized["production_kwh"]
    return normalized


def build_local_customer_production_report(
    conn: sqlite3.Connection,
    *,
    asset_id: int,
    report_month: str,
    electricity_price: float,
    sell_price: float,
    solcor_price_per_kwh: float = 0.0,
) -> dict[str, Any] | None:
    asset = conn.execute(
        """
        SELECT
            a.id AS asset_id,
            a.project_name,
            a.location,
            a.kwp,
            a.contract_type,
            a.asset_type,
            a.coverage_type,
            a.sell_to,
            ai.external_id,
            ai.external_name
        FROM asset_integrations ai
        JOIN assets a ON a.id = ai.asset_id
        WHERE a.id = ?
          AND ai.provider = ?
          AND ai.enabled = 1
          AND COALESCE(ai.external_id, '') != ''
        LIMIT 1
        """,
        (asset_id, INTEGRATION_PROVIDER_FUSIONSOLAR),
    ).fetchone()
    if asset is None:
        raise ValueError("A instalacao nao tem mapeamento FusionSolar ativo.")

    month_start = datetime.strptime(report_month, "%Y-%m").date()
    _, last_day = calendar.monthrange(month_start.year, month_start.month)
    month_end = month_start.replace(day=last_day)
    daily_records = query_all(
        conn,
        """
        SELECT *
        FROM production_records
        WHERE asset_id = ?
          AND provider = ?
          AND period_type = 'day'
          AND period_date BETWEEN ? AND ?
          AND production_kwh IS NOT NULL
        ORDER BY period_date
        """,
        (asset_id, INTEGRATION_PROVIDER_FUSIONSOLAR, month_start.isoformat(), month_end.isoformat()),
    )
    monthly_record = conn.execute(
        """
        SELECT *
        FROM production_records
        WHERE asset_id = ?
          AND provider = ?
          AND period_type = 'month'
          AND period_date = ?
          AND production_kwh IS NOT NULL
        LIMIT 1
        """,
        (asset_id, INTEGRATION_PROVIDER_FUSIONSOLAR, month_start.isoformat()),
    ).fetchone()
    if monthly_record is None and not daily_records:
        return None

    daily_rows = [normalize_customer_production_record(row, month_start) for row in daily_records]
    daily_rows.sort(key=lambda item: item["date"])
    monthly = normalize_customer_production_record(monthly_record, month_start) if monthly_record else None

    production_kwh = monthly["production_kwh"] if monthly else sum(item["production_kwh"] for item in daily_rows)
    self_use_kwh = monthly["self_use_kwh"] if monthly else sum(item["self_use_kwh"] for item in daily_rows)
    export_kwh = monthly["export_kwh"] if monthly else sum(item["export_kwh"] for item in daily_rows)
    if production_kwh and not self_use_kwh and not export_kwh:
        self_use_kwh = production_kwh
    consumption_kwh = (
        monthly["consumption_kwh"]
        if monthly and monthly.get("consumption_kwh") is not None
        else sum((item.get("consumption_kwh") or 0.0) for item in daily_rows)
    )
    report = {
        "asset": dict(asset),
        "station_code": str(asset["external_id"]),
        "month_start": month_start,
        "month_end": month_end,
        "month_label": f"{MONTH_NAMES_PT[month_start.month]} {month_start.year}",
        "daily_rows": daily_rows,
        "production_kwh": production_kwh,
        "self_use_kwh": self_use_kwh,
        "export_kwh": export_kwh,
        "consumption_kwh": consumption_kwh,
        "electricity_price": electricity_price,
        "sell_price": sell_price,
        "data_source": "Dados locais",
    }
    report.update(extract_customer_tariff_values(parse_production_record_payload(monthly_record) if monthly_record else None))
    return prepare_customer_report(report, solcor_price_per_kwh=solcor_price_per_kwh)


def build_fusionsolar_customer_production_report(
    conn: sqlite3.Connection,
    *,
    asset_id: int,
    report_month: str,
    electricity_price: float,
    sell_price: float,
    solcor_price_per_kwh: float = 0.0,
    force_api: bool = False,
) -> dict[str, Any]:
    if not force_api:
        local_report = build_local_customer_production_report(
            conn,
            asset_id=asset_id,
            report_month=report_month,
            electricity_price=electricity_price,
            sell_price=sell_price,
            solcor_price_per_kwh=solcor_price_per_kwh,
        )
        if local_report is not None:
            return local_report

    cooldown_reason = get_fusionsolar_performance_cooldown_reason(conn)
    if cooldown_reason:
        raise ValueError(cooldown_reason)

    asset = conn.execute(
        """
        SELECT
            a.id AS asset_id,
            a.project_name,
            a.location,
            a.kwp,
            a.contract_type,
            a.asset_type,
            a.coverage_type,
            a.sell_to,
            ai.external_id,
            ai.external_name
        FROM asset_integrations ai
        JOIN assets a ON a.id = ai.asset_id
        WHERE a.id = ?
          AND ai.provider = ?
          AND ai.enabled = 1
          AND COALESCE(ai.external_id, '') != ''
        LIMIT 1
        """,
        (asset_id, INTEGRATION_PROVIDER_FUSIONSOLAR),
    ).fetchone()
    if asset is None:
        raise ValueError("A instalacao nao tem mapeamento FusionSolar ativo.")

    config = get_integration_config(conn, INTEGRATION_PROVIDER_FUSIONSOLAR)
    if config is None or not config["enabled"]:
        raise ValueError("Configuracao FusionSolar ativa nao encontrada.")
    endpoints = get_fusionsolar_endpoint_config(config)
    session_obj, _ = get_fusionsolar_session(config)
    month_start = datetime.strptime(report_month, "%Y-%m").date()
    _, last_day = calendar.monthrange(month_start.year, month_start.month)
    month_end = month_start.replace(day=last_day)
    station_code = str(asset["external_id"])

    daily_rows_raw = fetch_fusionsolar_kpi_day_rows(
        session_obj,
        base_url=endpoints["base_url"],
        endpoint=endpoints["day_kpi_endpoint"],
        station_codes=[station_code],
        collect_date=month_start,
    )
    daily_rows = []
    for raw_row in daily_rows_raw:
        if str(raw_row.get("stationCode") or raw_row.get("plantCode") or "").strip() != station_code:
            continue
        normalized = normalize_customer_kpi_row(raw_row, month_start)
        if month_start <= normalized["date"] <= month_end:
            daily_rows.append(normalized)
    daily_rows.sort(key=lambda item: item["date"])

    monthly_map = fetch_fusionsolar_kpi_month_map(
        session_obj,
        base_url=endpoints["base_url"],
        endpoint=endpoints["month_kpi_endpoint"],
        station_codes=[station_code],
        collect_date=month_start,
    )
    monthly_row = monthly_map.get(station_code)
    monthly = normalize_customer_kpi_row(monthly_row, month_start) if monthly_row else None

    production_kwh = monthly["production_kwh"] if monthly else sum(item["production_kwh"] for item in daily_rows)
    self_use_kwh = monthly["self_use_kwh"] if monthly else sum(item["self_use_kwh"] for item in daily_rows)
    export_kwh = monthly["export_kwh"] if monthly else sum(item["export_kwh"] for item in daily_rows)
    if production_kwh and not self_use_kwh and not export_kwh:
        self_use_kwh = production_kwh
    consumption_kwh = (
        monthly["consumption_kwh"]
        if monthly and monthly.get("consumption_kwh") is not None
        else sum((item.get("consumption_kwh") or 0.0) for item in daily_rows)
    )
    report = {
        "asset": dict(asset),
        "station_code": station_code,
        "month_start": month_start,
        "month_end": month_end,
        "month_label": f"{MONTH_NAMES_PT[month_start.month]} {month_start.year}",
        "daily_rows": daily_rows,
        "production_kwh": production_kwh,
        "self_use_kwh": self_use_kwh,
        "export_kwh": export_kwh,
        "consumption_kwh": consumption_kwh,
        "electricity_price": electricity_price,
        "sell_price": sell_price,
        "data_source": "FusionSolar API",
    }
    report.update(extract_customer_tariff_values(monthly_row))
    return prepare_customer_report(report, solcor_price_per_kwh=solcor_price_per_kwh)


def export_customer_production_pdf(report: dict[str, Any]):
    pdf_bytes = build_customer_report_pdf(report, logo_path=BASE_DIR / "static" / "solcor-logo.png")
    buffer = io.BytesIO(pdf_bytes)
    safe_name = normalize_name(report["asset"]["project_name"]).replace(" ", "_") or "relatorio"
    model = str(report.get("report_type") or "epc").upper()
    filename = f"Relatorio_Mensal_{model}_{safe_name}_{report['month_start'].strftime('%m-%Y')}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/pdf")


def export_rows_file(
    rows: list[dict[str, Any]],
    headers: list[tuple[str, str]],
    filename: str,
    export_format: str,
):
    if export_format == "pdf":
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()
        data = [[header[1] for header in headers]]
        for row in rows:
            data.append([str(row.get(header[0], "") or "-") for header in headers])
        table = Table(data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f6b5c")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#c9d3d7")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#eef4f2")]),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                ]
            )
        )
        doc.build([Paragraph(filename, styles["Heading2"]), Spacer(1, 12), table])
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=f"{filename}.pdf", mimetype="application/pdf")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Export"
    worksheet.append([header[1] for header in headers])
    for row in rows:
        worksheet.append([row.get(header[0], "") for header in headers])
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"{filename}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def build_visits_by_ticket(visits: list[sqlite3.Row]) -> dict[int, list[sqlite3.Row]]:
    visits_by_ticket: dict[int, list[sqlite3.Row]] = {}
    for visit in visits:
        visits_by_ticket.setdefault(visit["ticket_id"], []).append(visit)
    return visits_by_ticket


def normalize_calendar_month(value: str | None) -> str:
    if value:
        try:
            return datetime.strptime(value.strip(), "%Y-%m").strftime("%Y-%m")
        except ValueError:
            pass
    return date.today().strftime("%Y-%m")


def normalize_optional_date(value: str | None) -> str:
    if not value:
        return ""
    try:
        return datetime.strptime(value.strip(), "%Y-%m-%d").date().isoformat()
    except ValueError:
        return ""


def parse_positive_int(value: str | None, default: int = 0) -> int:
    try:
        parsed = int(float(value or ""))
    except (TypeError, ValueError):
        return default
    return max(parsed, 0)


def normalize_choice(value: str | None, choices: list[str], default: str) -> str:
    value = (value or "").strip()
    return value if value in choices else default


def calendar_month_bounds(month_value: str) -> tuple[date, date, str, str]:
    month_start = datetime.strptime(month_value, "%Y-%m").date().replace(day=1)
    _, last_day = calendar.monthrange(month_start.year, month_start.month)
    month_end = month_start.replace(day=last_day)
    previous_month_date = (month_start - timedelta(days=1)).replace(day=1)
    next_month_date = (month_end + timedelta(days=1)).replace(day=1)
    return month_start, month_end, previous_month_date.strftime("%Y-%m"), next_month_date.strftime("%Y-%m")


def build_error_calendar(month_value: str, records: list[sqlite3.Row]) -> dict[str, Any]:
    month_start, month_end, _, _ = calendar_month_bounds(month_value)
    records_by_day: dict[str, list[sqlite3.Row]] = {}
    for record in records:
        records_by_day.setdefault(record["record_date"], []).append(record)

    weeks = []
    week = []
    for _ in range(month_start.weekday()):
        week.append({"date": None, "records": []})

    current_day = month_start
    while current_day <= month_end:
        iso_day = current_day.isoformat()
        week.append({"date": current_day, "records": records_by_day.get(iso_day, [])})
        if len(week) == 7:
            weeks.append(week)
            week = []
        current_day += timedelta(days=1)

    if week:
        while len(week) < 7:
            week.append({"date": None, "records": []})
        weeks.append(week)

    return {
        "label": f"{MONTH_NAMES_PT[month_start.month]} {month_start.year}",
        "weeks": weeks,
        "record_count": sum(len(rows) for rows in records_by_day.values()),
    }


def build_intervention_calendar(month_value: str, records: list[sqlite3.Row]) -> dict[str, Any]:
    month_start, month_end, _, _ = calendar_month_bounds(month_value)
    records_by_day: dict[str, list[sqlite3.Row]] = {}
    for record in records:
        planned_date = record["planned_date"]
        if planned_date:
            records_by_day.setdefault(planned_date, []).append(record)

    weeks = []
    week = []
    for _ in range(month_start.weekday()):
        week.append({"date": None, "records": []})

    current_day = month_start
    while current_day <= month_end:
        iso_day = current_day.isoformat()
        week.append({"date": current_day, "records": records_by_day.get(iso_day, [])})
        if len(week) == 7:
            weeks.append(week)
            week = []
        current_day += timedelta(days=1)

    if week:
        while len(week) < 7:
            week.append({"date": None, "records": []})
        weeks.append(week)

    return {
        "label": f"{MONTH_NAMES_PT[month_start.month]} {month_start.year}",
        "weeks": weeks,
        "record_count": sum(len(rows) for rows in records_by_day.values()),
    }


def intervention_ready_for_route(row: sqlite3.Row | dict[str, Any]) -> bool:
    if row["status"] == "Fechado":
        return False
    if row["material_status"] == "Bloqueado":
        return False
    if row["latitude"] is None or row["longitude"] is None:
        return False
    if row["coordinates_confidence"] in {"suspect", "review"}:
        return False
    return True


def build_asset_error_calendar(month_value: str, records: list[sqlite3.Row]) -> dict[str, Any]:
    month_start, month_end, _, _ = calendar_month_bounds(month_value)
    events_by_day: dict[str, list[dict[str, Any]]] = {}
    previous_problem = False

    for record in records:
        record_date = record["record_date"]
        status = record["status"]
        is_problem = status in PROBLEM_MONITORING_STATUSES
        event_type = ""
        event_label = ""

        if is_problem and not previous_problem:
            event_type = "start"
            event_label = "Apareceu"
        elif is_problem:
            event_type = "active"
            event_label = "Mantem-se"
        elif previous_problem:
            event_type = "end"
            event_label = "Desapareceu"

        if month_start.isoformat() <= record_date <= month_end.isoformat() and event_type:
            events_by_day.setdefault(record_date, []).append(
                {
                    "status": status,
                    "label": event_label,
                    "type": event_type,
                    "notes": record["notes"],
                    "source": record["source"],
                    "record_id": record["id"],
                }
            )

        previous_problem = is_problem

    weeks = []
    week = []
    for _ in range(month_start.weekday()):
        week.append({"date": None, "events": []})

    current_day = month_start
    while current_day <= month_end:
        iso_day = current_day.isoformat()
        week.append({"date": current_day, "events": events_by_day.get(iso_day, [])})
        if len(week) == 7:
            weeks.append(week)
            week = []
        current_day += timedelta(days=1)

    if week:
        while len(week) < 7:
            week.append({"date": None, "events": []})
        weeks.append(week)

    return {
        "label": f"{MONTH_NAMES_PT[month_start.month]} {month_start.year}",
        "weeks": weeks,
        "event_count": sum(len(rows) for rows in events_by_day.values()),
    }


def group_tickets_by_asset(tickets: list[sqlite3.Row]) -> list[dict[str, Any]]:
    grouped: dict[int, dict[str, Any]] = {}
    for ticket in tickets:
        asset_id = int(ticket["asset_id"])
        bucket = grouped.setdefault(
            asset_id,
            {
                "asset_id": asset_id,
                "project_name": ticket["project_name"],
                "location": ticket["location"],
                "active_contract": ticket["active_contract"],
                "contract_type": ticket["contract_type"],
                "tickets": [],
            },
        )
        bucket["tickets"].append(ticket)

    ordered = []
    for asset_id, bucket in grouped.items():
        tickets_list = bucket["tickets"]
        bucket["open_count"] = sum(1 for ticket in tickets_list if ticket["status"] != "Fechado")
        bucket["critical_count"] = sum(
            1 for ticket in tickets_list if ticket["urgency"] == "Critica" and ticket["status"] != "Fechado"
        )
        bucket["last_update"] = max(ticket["updated_at"] for ticket in tickets_list)
        ordered.append(bucket)

    ordered.sort(
        key=lambda item: (
            0 if item["active_contract"] == "yes" else 1,
            -item["critical_count"],
            -item["open_count"],
            item["project_name"].lower(),
        )
    )
    return ordered


def ensure_predefined_export_templates(conn: sqlite3.Connection) -> None:
    predefined_templates = [
        {
            "name": "Preventiva - O&M ativo",
            "dataset": "assets",
            "export_format": "xlsx",
            "columns": [
                "project_name",
                "location",
                "address",
                "contact_phone",
                "contact_name",
                "access_type",
                "coverage_type",
            ],
            "filters": {
                "om_only": "yes",
            },
        },
        {
            "name": "Relatorio monitorizacao - diario",
            "dataset": "monitoring_report",
            "export_format": "pdf",
            "columns": [
                "period",
                "project_name",
                "current_status",
                "error_records",
                "distinct_errors",
                "error_types",
                "open_tickets",
                "visits_period",
                "last_visit_date",
                "latest_notes",
            ],
            "filters": {
                "period": "day",
                "om_only": "yes",
            },
        },
        {
            "name": "Relatorio monitorizacao - semanal",
            "dataset": "monitoring_report",
            "export_format": "pdf",
            "columns": [
                "period",
                "project_name",
                "current_status",
                "error_records",
                "distinct_errors",
                "error_types",
                "open_tickets",
                "visits_period",
                "last_visit_date",
                "latest_notes",
            ],
            "filters": {
                "period": "week",
                "om_only": "yes",
            },
        },
        {
            "name": "Relatorio monitorizacao - mensal",
            "dataset": "monitoring_report",
            "export_format": "pdf",
            "columns": [
                "period",
                "project_name",
                "current_status",
                "error_records",
                "distinct_errors",
                "error_types",
                "open_tickets",
                "visits_period",
                "last_visit_date",
                "latest_notes",
            ],
            "filters": {
                "period": "month",
                "om_only": "yes",
            },
        },
        {
            "name": "Relatorio producao - mensal",
            "dataset": "production_report",
            "export_format": "xlsx",
            "columns": [
                "period",
                "project_name",
                "location",
                "provider",
                "production_kwh",
                "specific_yield",
                "expected_kwh",
                "deviation_pct",
                "performance_status",
                "data_points",
                "data_source",
                "last_update",
                "notes",
            ],
            "filters": {
                "period": "month",
                "om_only": "yes",
                "source": "FusionSolar",
            },
        },
        {
            "name": "Relatorio producao - anual",
            "dataset": "production_report",
            "export_format": "xlsx",
            "columns": [
                "period",
                "project_name",
                "location",
                "provider",
                "production_kwh",
                "specific_yield",
                "expected_kwh",
                "deviation_pct",
                "performance_status",
                "data_points",
                "data_source",
                "last_update",
                "notes",
            ],
            "filters": {
                "period": "year",
                "om_only": "yes",
                "source": "FusionSolar",
            },
        },
    ]

    for template in predefined_templates:
        existing = conn.execute(
            "SELECT id FROM export_templates WHERE name = ? LIMIT 1",
            (template["name"],),
        ).fetchone()
        if existing:
            continue
        conn.execute(
            """
            INSERT INTO export_templates (name, dataset, export_format, columns_json, filters_json, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                template["name"],
                template["dataset"],
                template["export_format"],
                json.dumps(template["columns"], ensure_ascii=True),
                json.dumps(template["filters"], ensure_ascii=True),
                datetime.now().isoformat(timespec="seconds"),
            ),
        )


def get_fusionsolar_env_config() -> dict[str, str]:
    return {
        "username": os.environ.get("FUSIONSOLAR_USERNAME", "").strip(),
        "password": os.environ.get("FUSIONSOLAR_PASSWORD", "").strip(),
        "base_url": os.environ.get("FUSIONSOLAR_BASE_URL", "").strip(),
        "login_endpoint": os.environ.get("FUSIONSOLAR_LOGIN_ENDPOINT", DEFAULT_FUSIONSOLAR_LOGIN_ENDPOINT).strip(),
        "plants_endpoint": os.environ.get("FUSIONSOLAR_STATIONS_ENDPOINT", DEFAULT_FUSIONSOLAR_STATIONS_ENDPOINT).strip(),
        "real_time_endpoint": os.environ.get(
            "FUSIONSOLAR_REALTIME_ENDPOINT",
            DEFAULT_FUSIONSOLAR_REALTIME_ENDPOINT,
        ).strip(),
        "device_list_endpoint": os.environ.get(
            "FUSIONSOLAR_DEVICES_ENDPOINT",
            DEFAULT_FUSIONSOLAR_DEVICES_ENDPOINT,
        ).strip(),
        "device_real_time_endpoint": os.environ.get(
            "FUSIONSOLAR_DEVICE_REALTIME_ENDPOINT",
            DEFAULT_FUSIONSOLAR_DEVICE_REALTIME_ENDPOINT,
        ).strip(),
        "device_history_endpoint": os.environ.get(
            "FUSIONSOLAR_DEVICE_HISTORY_ENDPOINT",
            DEFAULT_FUSIONSOLAR_DEVICE_HISTORY_ENDPOINT,
        ).strip(),
        "alarms_endpoint": os.environ.get("FUSIONSOLAR_ALARMS_ENDPOINT", DEFAULT_FUSIONSOLAR_ALARMS_ENDPOINT).strip(),
        "day_kpi_endpoint": os.environ.get(
            "FUSIONSOLAR_DAY_KPI_ENDPOINT",
            DEFAULT_FUSIONSOLAR_DAY_KPI_ENDPOINT,
        ).strip(),
        "month_kpi_endpoint": os.environ.get(
            "FUSIONSOLAR_MONTH_KPI_ENDPOINT",
            DEFAULT_FUSIONSOLAR_MONTH_KPI_ENDPOINT,
        ).strip(),
        "sync_hours": os.environ.get("FUSIONSOLAR_SYNC_HOURS", DEFAULT_FUSIONSOLAR_SYNC_HOURS).strip(),
    }


def get_sigenergy_env_config() -> dict[str, str]:
    return {
        "username": os.environ.get("SIGENERGY_APP_KEY", "").strip(),
        "password": os.environ.get("SIGENERGY_APP_SECRET", "").strip(),
        "base_url": os.environ.get("SIGENERGY_BASE_URL", DEFAULT_SIGENERGY_BASE_URL).strip(),
        "login_endpoint": os.environ.get("SIGENERGY_AUTH_ENDPOINT", DEFAULT_SIGENERGY_AUTH_ENDPOINT).strip(),
        "plants_endpoint": os.environ.get("SIGENERGY_SYSTEMS_ENDPOINT", DEFAULT_SIGENERGY_SYSTEMS_ENDPOINT).strip(),
        "real_time_endpoint": os.environ.get("SIGENERGY_REALTIME_ENDPOINT", DEFAULT_SIGENERGY_REALTIME_ENDPOINT).strip(),
        "alarms_endpoint": os.environ.get("SIGENERGY_ENERGY_FLOW_ENDPOINT", DEFAULT_SIGENERGY_ENERGY_FLOW_ENDPOINT).strip(),
        "day_kpi_endpoint": "",
        "month_kpi_endpoint": "",
        "sync_hours": os.environ.get("SIGENERGY_SYNC_HOURS", DEFAULT_FUSIONSOLAR_SYNC_HOURS).strip(),
        "region": os.environ.get("SIGENERGY_REGION", DEFAULT_SIGENERGY_REGION).strip() or DEFAULT_SIGENERGY_REGION,
        "system_ids": os.environ.get("SIGENERGY_SYSTEM_IDS", os.environ.get("SIGENERGY_SYSTEM_ID", "")).strip(),
        "enabled": os.environ.get("SIGENERGY_ENABLED", "").strip(),
    }


def ensure_integration_seed_data(conn: sqlite3.Connection) -> None:
    env_config = get_fusionsolar_env_config()
    existing = conn.execute(
        "SELECT * FROM integration_configs WHERE provider = ?",
        (INTEGRATION_PROVIDER_FUSIONSOLAR,),
    ).fetchone()
    if existing:
        conn.execute(
            """
            UPDATE integration_configs
            SET username = CASE WHEN COALESCE(username, '') = '' THEN ? ELSE username END,
                base_url = CASE WHEN COALESCE(base_url, '') = '' THEN ? ELSE base_url END,
                login_endpoint = CASE WHEN COALESCE(login_endpoint, '') = '' THEN ? ELSE login_endpoint END,
                plants_endpoint = CASE WHEN COALESCE(plants_endpoint, '') = '' THEN ? ELSE plants_endpoint END,
                real_time_endpoint = CASE WHEN COALESCE(real_time_endpoint, '') = '' THEN ? ELSE real_time_endpoint END,
                device_list_endpoint = CASE WHEN COALESCE(device_list_endpoint, '') = '' THEN ? ELSE device_list_endpoint END,
                device_real_time_endpoint = CASE WHEN COALESCE(device_real_time_endpoint, '') = '' THEN ? ELSE device_real_time_endpoint END,
                device_history_endpoint = CASE WHEN COALESCE(device_history_endpoint, '') = '' THEN ? ELSE device_history_endpoint END,
                alarms_endpoint = CASE WHEN COALESCE(alarms_endpoint, '') = '' THEN ? ELSE alarms_endpoint END,
                day_kpi_endpoint = CASE WHEN COALESCE(day_kpi_endpoint, '') = '' THEN ? ELSE day_kpi_endpoint END,
                month_kpi_endpoint = CASE WHEN COALESCE(month_kpi_endpoint, '') = '' THEN ? ELSE month_kpi_endpoint END,
                sync_hours = CASE WHEN COALESCE(sync_hours, '') = '' THEN ? ELSE sync_hours END,
                updated_at = ?
            WHERE provider = ?
            """,
            (
                env_config["username"],
                env_config["base_url"],
                env_config["login_endpoint"],
                env_config["plants_endpoint"],
                env_config["real_time_endpoint"],
                env_config["device_list_endpoint"],
                env_config["device_real_time_endpoint"],
                env_config["device_history_endpoint"],
                env_config["alarms_endpoint"],
                env_config["day_kpi_endpoint"],
                env_config["month_kpi_endpoint"],
                env_config["sync_hours"],
                datetime.now().isoformat(timespec="seconds"),
                INTEGRATION_PROVIDER_FUSIONSOLAR,
            ),
        )
    else:
        conn.execute(
            """
            INSERT INTO integration_configs (
                provider, username, password, base_url, login_endpoint, plants_endpoint, real_time_endpoint,
                device_list_endpoint, device_real_time_endpoint, device_history_endpoint, alarms_endpoint,
                day_kpi_endpoint, month_kpi_endpoint,
                enabled, auto_sync_enabled, sync_hours, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                INTEGRATION_PROVIDER_FUSIONSOLAR,
                env_config["username"],
                "",
                env_config["base_url"],
                env_config["login_endpoint"],
                env_config["plants_endpoint"],
                env_config["real_time_endpoint"],
                env_config["device_list_endpoint"],
                env_config["device_real_time_endpoint"],
                env_config["device_history_endpoint"],
                env_config["alarms_endpoint"],
                env_config["day_kpi_endpoint"],
                env_config["month_kpi_endpoint"],
                0,
                0,
                env_config["sync_hours"],
                datetime.now().isoformat(timespec="seconds"),
                datetime.now().isoformat(timespec="seconds"),
            ),
        )

    sigenergy_env = get_sigenergy_env_config()
    sigenergy_existing = conn.execute(
        "SELECT * FROM integration_configs WHERE provider = ?",
        (INTEGRATION_PROVIDER_SIGENERGY,),
    ).fetchone()
    sigenergy_enabled = 1 if sigenergy_env["enabled"].lower() in {"1", "true", "yes", "sim", "on"} else 0
    if sigenergy_existing:
        conn.execute(
            """
            UPDATE integration_configs
            SET username = CASE WHEN COALESCE(username, '') = '' THEN ? ELSE username END,
                base_url = CASE WHEN COALESCE(base_url, '') = '' THEN ? ELSE base_url END,
                login_endpoint = CASE WHEN COALESCE(login_endpoint, '') = '' THEN ? ELSE login_endpoint END,
                plants_endpoint = CASE WHEN COALESCE(plants_endpoint, '') = '' THEN ? ELSE plants_endpoint END,
                real_time_endpoint = CASE WHEN COALESCE(real_time_endpoint, '') = '' THEN ? ELSE real_time_endpoint END,
                alarms_endpoint = CASE WHEN COALESCE(alarms_endpoint, '') = '' THEN ? ELSE alarms_endpoint END,
                sync_hours = CASE WHEN COALESCE(sync_hours, '') = '' THEN ? ELSE sync_hours END,
                enabled = CASE WHEN ? = 1 THEN 1 ELSE enabled END,
                updated_at = ?
            WHERE provider = ?
            """,
            (
                sigenergy_env["username"],
                sigenergy_env["base_url"],
                sigenergy_env["login_endpoint"],
                sigenergy_env["plants_endpoint"],
                sigenergy_env["real_time_endpoint"],
                sigenergy_env["alarms_endpoint"],
                sigenergy_env["sync_hours"],
                sigenergy_enabled,
                datetime.now().isoformat(timespec="seconds"),
                INTEGRATION_PROVIDER_SIGENERGY,
            ),
        )
        return

    conn.execute(
        """
        INSERT INTO integration_configs (
            provider, username, password, base_url, login_endpoint, plants_endpoint, real_time_endpoint, alarms_endpoint,
            day_kpi_endpoint, month_kpi_endpoint,
            enabled, auto_sync_enabled, sync_hours, created_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            INTEGRATION_PROVIDER_SIGENERGY,
            sigenergy_env["username"],
            "",
            sigenergy_env["base_url"],
            sigenergy_env["login_endpoint"],
            sigenergy_env["plants_endpoint"],
            sigenergy_env["real_time_endpoint"],
            sigenergy_env["alarms_endpoint"],
            "",
            "",
            sigenergy_enabled,
            sigenergy_enabled,
            sigenergy_env["sync_hours"],
            datetime.now().isoformat(timespec="seconds"),
            datetime.now().isoformat(timespec="seconds"),
        ),
    )


def get_integration_config(conn: sqlite3.Connection, provider: str) -> dict[str, Any] | None:
    row = conn.execute("SELECT * FROM integration_configs WHERE provider = ?", (provider,)).fetchone()
    if row is None:
        return None
    config = dict(row)
    if provider == INTEGRATION_PROVIDER_FUSIONSOLAR:
        env_config = get_fusionsolar_env_config()
        for key, value in env_config.items():
            if value and key in config:
                config[key] = value
        config["password_configured"] = bool(config.get("password"))
        config["password_source"] = "env" if env_config["password"] else ("database" if config.get("password") else "")
    if provider == INTEGRATION_PROVIDER_SIGENERGY:
        env_config = get_sigenergy_env_config()
        for key, value in env_config.items():
            if value and key in config:
                config[key] = value
        config["region"] = env_config["region"]
        config["system_ids"] = env_config["system_ids"]
        config["password_configured"] = bool(config.get("password"))
        config["password_source"] = "env" if env_config["password"] else ("database" if config.get("password") else "")
        if env_config["enabled"].lower() in {"1", "true", "yes", "sim", "on"}:
            config["enabled"] = 1
    return config


def start_integration_scheduler(app: Flask) -> None:
    global SCHEDULER
    if SCHEDULER is not None:
        return
    SCHEDULER = BackgroundScheduler(timezone="Europe/Lisbon")
    SCHEDULER.start()
    refresh_integration_scheduler(app)
    schedule_pending_background_jobs(app)


def refresh_integration_scheduler(app: Flask) -> None:
    global SCHEDULER
    if SCHEDULER is None:
        return
    for job in list(SCHEDULER.get_jobs()):
        if (
            job.id.startswith("integration-sync-")
            or job.id.startswith("fusionsolar-sync-")
            or job.id in {"telegram-daily-summary", "fusionsolar-wat-daily"}
        ):
            SCHEDULER.remove_job(job.id)

    with closing(get_db(app.config["DATABASE"])) as conn:
        configs = [get_integration_config(conn, provider) for provider in INTEGRATION_PROVIDER_OPTIONS]
    if telegram_daily_summary_enabled():
        SCHEDULER.add_job(
            func=run_scheduled_telegram_daily_summary,
            trigger="cron",
            hour=9,
            minute=0,
            args=[app],
            id="telegram-daily-summary",
            replace_existing=True,
            max_instances=1,
            coalesce=True,
            misfire_grace_time=1800,
        )
    for config in configs:
        if config is None or not config["enabled"]:
            continue
        provider = str(config["provider"])
        if provider == INTEGRATION_PROVIDER_FUSIONSOLAR:
            SCHEDULER.add_job(
                func=run_scheduled_integration_sync,
                trigger="cron",
                minute=0,
                args=[app, provider],
                id="integration-sync-fusionsolar-hourly",
                replace_existing=True,
                max_instances=1,
                coalesce=True,
                misfire_grace_time=1800,
            )
            SCHEDULER.add_job(
                func=run_scheduled_fusionsolar_wat_backfill,
                trigger="cron",
                hour=0,
                minute=30,
                args=[app],
                id="fusionsolar-wat-daily",
                replace_existing=True,
                max_instances=1,
                coalesce=True,
                misfire_grace_time=1800,
            )
            continue
        if not config["auto_sync_enabled"]:
            continue
        for index, item in enumerate(normalize_sync_hours(config["sync_hours"] or DEFAULT_FUSIONSOLAR_SYNC_HOURS).split(","), start=1):
            hour, minute = item.split(":")
            SCHEDULER.add_job(
                func=run_scheduled_integration_sync,
                trigger="cron",
                hour=int(hour),
                minute=int(minute),
                args=[app, provider],
                id=f"integration-sync-{normalize_name(provider).replace(' ', '-')}-{index}",
                replace_existing=True,
                max_instances=1,
                coalesce=True,
                misfire_grace_time=1800,
            )


def run_scheduled_telegram_daily_summary(app: Flask) -> None:
    with app.app_context():
        with closing(get_db(app.config["DATABASE"])) as conn:
            try:
                send_daily_telegram_summary(conn)
                conn.commit()
            except Exception:
                current_app.logger.exception("Scheduled Telegram daily summary failed")


def run_scheduled_integration_sync(app: Flask, provider: str) -> None:
    with app.app_context():
        with closing(get_db(app.config["DATABASE"])) as conn:
            try:
                current_app.logger.info("Scheduled %s sync started", provider)
                result = run_integration_sync(conn, provider, trigger_type="scheduled")
                current_app.logger.info("Scheduled %s sync completed: %s", provider, result)
            except Exception:
                current_app.logger.exception("Scheduled %s sync failed", provider)


def run_scheduled_fusionsolar_sync(app: Flask) -> None:
    run_scheduled_integration_sync(app, INTEGRATION_PROVIDER_FUSIONSOLAR)


def current_lisbon_date() -> date:
    return datetime.now(LISBON_TIMEZONE).date()


def run_scheduled_fusionsolar_wat_backfill(app: Flask) -> None:
    scheduler_date = current_lisbon_date()
    target_date = scheduler_date - timedelta(days=1)
    with app.app_context():
        with closing(get_db(app.config["DATABASE"])) as conn:
            current_app.logger.info(
                "Scheduled FusionSolar WAT preparing previous closed day: scheduler_date=%s target_date=%s",
                scheduler_date,
                target_date,
            )
            config = get_integration_config(conn, INTEGRATION_PROVIDER_FUSIONSOLAR)
            if config is None or not config["enabled"]:
                current_app.logger.info(
                    "Scheduled FusionSolar WAT skipped because integration is disabled: target_date=%s",
                    target_date,
                )
                return
            job_id, created = create_background_job(
                conn,
                "fusionsolar_inverter_availability_backfill",
                {
                    "provider": INTEGRATION_PROVIDER_FUSIONSOLAR,
                    "from_date": target_date.isoformat(),
                    "to_date": target_date.isoformat(),
                    "trigger_type": "scheduled",
                },
            )
            conn.commit()
            if created:
                schedule_background_job(app, job_id)
                current_app.logger.info(
                    "Scheduled FusionSolar WAT queued: job_id=%s target_date=%s job_type=%s",
                    job_id,
                    target_date,
                    "fusionsolar_inverter_availability_backfill",
                )
            else:
                current_app.logger.info(
                    "Scheduled FusionSolar WAT reused existing pending/running job: job_id=%s target_date=%s",
                    job_id,
                    target_date,
                )


def schedule_background_job(app: Flask, job_id: int) -> bool:
    if SCHEDULER is None:
        app.logger.error("Background job %s was queued but APScheduler is not running", job_id)
        return False
    SCHEDULER.add_job(
        func=run_background_job,
        trigger="date",
        run_date=datetime.now(),
        args=[app, job_id],
        id=f"background-job-{job_id}",
        replace_existing=True,
        max_instances=1,
    )
    return True


def schedule_pending_background_jobs(app: Flask) -> dict[str, Any]:
    with closing(get_db(app.config["DATABASE"])) as conn:
        recovered_count = mark_stale_running_background_jobs_failed(conn)
        if recovered_count:
            app.logger.warning("Marked %s stale running background jobs as failed on startup", recovered_count)
        pending_job_ids = fetch_pending_background_job_ids(conn)
    failed_job_ids: list[int] = []
    for job_id in pending_job_ids:
        if not schedule_background_job(app, job_id):
            failed_job_ids.append(job_id)
    scheduled_count = len(pending_job_ids) - len(failed_job_ids)
    if scheduled_count:
        app.logger.info("Scheduled %s pending background jobs on startup", scheduled_count)
    if failed_job_ids:
        app.logger.warning("Could not schedule pending background jobs on startup: %s", failed_job_ids)
    return {
        "stale_running_failed": recovered_count,
        "pending_found": len(pending_job_ids),
        "pending_scheduled": scheduled_count,
        "pending_schedule_failed_ids": failed_job_ids,
    }


def run_background_job(app: Flask, job_id: int) -> None:
    with app.app_context():
        with closing(get_db(app.config["DATABASE"])) as conn:
            job = conn.execute("SELECT * FROM background_jobs WHERE id = ?", (job_id,)).fetchone()
            if job is None:
                current_app.logger.error("Background job %s not found", job_id)
                return
            if not mark_background_job_running(conn, job_id):
                current_app.logger.info("Background job %s skipped because it is no longer pending", job_id)
                return
            try:
                params = json.loads(job["params_json"] or "{}")
                current_app.logger.info("Background job %s started: %s", job_id, job["job_type"])
                result = run_background_job_payload(conn, str(job["job_type"]), params)
                mark_background_job_success(conn, job_id, result)
                current_app.logger.info("Background job %s completed: %s", job_id, job["job_type"])
            except Exception as exc:
                current_app.logger.exception("Background job %s failed: %s", job_id, job["job_type"])
                mark_background_job_failed(conn, job_id, str(exc))


def run_background_job_payload(conn: sqlite3.Connection, job_type: str, params: dict[str, Any]) -> dict[str, Any]:
    if job_type == "fusionsolar_production_sync":
        target_date = parse_date_value(str(params.get("target_date") or ""))
        if target_date is None:
            raise ValueError("Data invalida para sync de producao.")
        return run_fusionsolar_production_sync(
            conn,
            provider=str(params.get("provider") or INTEGRATION_PROVIDER_FUSIONSOLAR),
            target_date=target_date,
            period_type=str(params.get("period_type") or "day"),
        )

    if job_type == "performance_reference_recalculation":
        period_date = parse_date_value(str(params.get("period_date") or ""))
        if period_date is None:
            raise ValueError("Data invalida para recalculo de referencias.")
        asset_id_value = params.get("asset_id")
        return recalculate_performance_references(
            conn,
            period_type=str(params.get("period_type") or "day"),
            period_date=period_date,
            asset_id=int(asset_id_value) if asset_id_value else None,
            provider=str(params.get("provider") or INTEGRATION_PROVIDER_FUSIONSOLAR),
        )

    if job_type == "fusionsolar_production_backfill":
        date_from = parse_date_value(str(params.get("date_from") or ""))
        date_to = parse_date_value(str(params.get("date_to") or ""))
        asset_id_value = params.get("asset_id")
        return run_fusionsolar_production_backfill(
            conn,
            provider=str(params.get("provider") or INTEGRATION_PROVIDER_FUSIONSOLAR),
            period_type=str(params.get("period_type") or "day"),
            from_year=int(params["from_year"]),
            to_year=int(params["to_year"]),
            asset_id=int(asset_id_value) if asset_id_value else None,
            date_from=date_from,
            date_to=date_to,
            max_api_calls=int(params.get("max_api_calls") or FUSIONSOLAR_PERFORMANCE_MAX_API_CALLS),
        )

    if job_type == "fusionsolar_inverter_availability_backfill":
        from_date = parse_date_value(str(params.get("from_date") or ""))
        to_date = parse_date_value(str(params.get("to_date") or ""))
        if from_date is None or to_date is None or from_date > to_date:
            raise ValueError("Intervalo invalido para backfill WAT.")
        return run_fusionsolar_inverter_availability_backfill(
            conn,
            from_date=from_date,
            to_date=to_date,
        )

    if job_type == "fusionsolar_month_cycle":
        raw_asset_ids = params.get("asset_ids") or []
        asset_ids = [int(value) for value in raw_asset_ids if str(value).isdigit()]
        return run_fusionsolar_month_cycle(
            conn,
            provider=str(params.get("provider") or INTEGRATION_PROVIDER_FUSIONSOLAR),
            report_month=str(params.get("report_month") or date.today().strftime("%Y-%m")),
            asset_ids=asset_ids,
        )

    raise ValueError(f"Tipo de job desconhecido: {job_type}")


def get_fusionsolar_endpoint_config(config: sqlite3.Row | dict[str, Any]) -> dict[str, str]:
    return {
        "base_url": str(config["base_url"] or "").strip(),
        "login_endpoint": str(config["login_endpoint"] or DEFAULT_FUSIONSOLAR_LOGIN_ENDPOINT).strip() or DEFAULT_FUSIONSOLAR_LOGIN_ENDPOINT,
        "plants_endpoint": str(config["plants_endpoint"] or DEFAULT_FUSIONSOLAR_STATIONS_ENDPOINT).strip() or DEFAULT_FUSIONSOLAR_STATIONS_ENDPOINT,
        "real_time_endpoint": str(config["real_time_endpoint"] or DEFAULT_FUSIONSOLAR_REALTIME_ENDPOINT).strip() or DEFAULT_FUSIONSOLAR_REALTIME_ENDPOINT,
        "device_list_endpoint": str(config["device_list_endpoint"] or DEFAULT_FUSIONSOLAR_DEVICES_ENDPOINT).strip() or DEFAULT_FUSIONSOLAR_DEVICES_ENDPOINT,
        "device_real_time_endpoint": str(config["device_real_time_endpoint"] or DEFAULT_FUSIONSOLAR_DEVICE_REALTIME_ENDPOINT).strip() or DEFAULT_FUSIONSOLAR_DEVICE_REALTIME_ENDPOINT,
        "device_history_endpoint": str(config["device_history_endpoint"] or DEFAULT_FUSIONSOLAR_DEVICE_HISTORY_ENDPOINT).strip() or DEFAULT_FUSIONSOLAR_DEVICE_HISTORY_ENDPOINT,
        "alarms_endpoint": str(config["alarms_endpoint"] or DEFAULT_FUSIONSOLAR_ALARMS_ENDPOINT).strip() or DEFAULT_FUSIONSOLAR_ALARMS_ENDPOINT,
        "day_kpi_endpoint": str(config["day_kpi_endpoint"] or DEFAULT_FUSIONSOLAR_DAY_KPI_ENDPOINT).strip() or DEFAULT_FUSIONSOLAR_DAY_KPI_ENDPOINT,
        "month_kpi_endpoint": str(config["month_kpi_endpoint"] or DEFAULT_FUSIONSOLAR_MONTH_KPI_ENDPOINT).strip() or DEFAULT_FUSIONSOLAR_MONTH_KPI_ENDPOINT,
    }


def extract_fusionsolar_xsrf_token(response: requests.Response, session: requests.Session) -> str:
    for key, value in response.headers.items():
        if key.lower() == "xsrf-token" and value:
            return value.strip()
    for cookie_name in ("XSRF-TOKEN", "xsrf-token"):
        cookie_value = session.cookies.get(cookie_name)
        if cookie_value:
            return str(cookie_value).strip()
    raise ValueError("O login FusionSolar respondeu sem XSRF-TOKEN no header/cookies.")


def get_fusionsolar_session(config: sqlite3.Row | dict[str, Any], *, force_login: bool = False) -> tuple[requests.Session, str]:
    endpoints = get_fusionsolar_endpoint_config(config)
    base_url = endpoints["base_url"]
    username = str(config["username"] or "").strip()
    password = str(config["password"] or "").strip()

    if not username or not password:
        raise ValueError("Preenche username e password do FusionSolar.")
    if not base_url:
        raise ValueError("Preenche a Base URL do FusionSolar.")

    cache_key = f"{base_url}|{username}"
    now = datetime.now()

    with FUSIONSOLAR_SESSION_LOCK:
        cached = FUSIONSOLAR_SESSION_CACHE.get(cache_key)
        if cached and not force_login and cached["expires_at"] > now:
            return cached["session"], cached["xsrf_token"]

        session = requests.Session()
        login_response = session.post(
            build_provider_url(base_url, endpoints["login_endpoint"]),
            json={"userName": username, "systemCode": password},
            headers={"Content-Type": "application/json", "Accept": "application/json, */*"},
            timeout=30,
        )
        login_response.raise_for_status()
        payload = login_response.json()
        if payload.get("success") is not True or int(payload.get("failCode") or 0) != 0:
            message = payload.get("message") or "Login FusionSolar falhou."
            raise ValueError(f"{message} (failCode={payload.get('failCode')})")

        xsrf_token = extract_fusionsolar_xsrf_token(login_response, session)
        session.headers.update(
            {
                "Content-Type": "application/json",
                "Accept": "application/json, */*",
                "XSRF-TOKEN": xsrf_token,
            }
        )
        FUSIONSOLAR_SESSION_CACHE[cache_key] = {
            "session": session,
            "xsrf_token": xsrf_token,
            "expires_at": now + timedelta(minutes=25),
        }
        return session, xsrf_token


def invalidate_fusionsolar_session(config: sqlite3.Row | dict[str, Any]) -> None:
    endpoints = get_fusionsolar_endpoint_config(config)
    cache_key = f"{endpoints['base_url']}|{str(config['username'] or '').strip()}"
    with FUSIONSOLAR_SESSION_LOCK:
        FUSIONSOLAR_SESSION_CACHE.pop(cache_key, None)


def post_fusionsolar_json(
    session: requests.Session,
    url: str,
    payload: dict[str, Any],
    *,
    expected_message: str,
) -> dict[str, Any]:
    response = session.post(url, json=payload, timeout=30)
    response.raise_for_status()
    data = response.json()
    if data.get("success") is not True or int(data.get("failCode") or 0) != 0:
        message = data.get("message") or expected_message
        raise ValueError(f"{message} (failCode={data.get('failCode')})")
    return data


def chunked(values: list[str], size: int) -> list[list[str]]:
    return [values[index : index + size] for index in range(0, len(values), size)]


def fetch_fusionsolar_stations(session: requests.Session, *, base_url: str, endpoint: str) -> list[dict[str, Any]]:
    url = build_provider_url(base_url, endpoint)
    stations: list[dict[str, Any]] = []
    page_no = 1
    page_count = 1

    while page_no <= page_count:
        payload = post_fusionsolar_json(
            session,
            url,
            {"pageNo": page_no},
            expected_message="Falha ao obter a lista de centrais FusionSolar.",
        )
        page_data = payload.get("data") or {}
        page_list = page_data.get("list") or []
        if not isinstance(page_list, list):
            raise ValueError("A resposta FusionSolar da lista de centrais nao trouxe data.list.")
        stations.extend([item for item in page_list if isinstance(item, dict)])
        page_count = int(page_data.get("pageCount") or 1)
        page_no += 1

    return stations


def fetch_fusionsolar_realtime_map(
    session: requests.Session,
    *,
    base_url: str,
    endpoint: str,
    station_codes: list[str],
) -> dict[str, dict[str, Any]]:
    url = build_provider_url(base_url, endpoint)
    real_time_map: dict[str, dict[str, Any]] = {}

    for group in chunked(station_codes, 100):
        payload = post_fusionsolar_json(
            session,
            url,
            {"stationCodes": ",".join(group)},
            expected_message="Falha ao obter os dados realtime das centrais FusionSolar.",
        )
        data_rows = payload.get("data") or []
        if not isinstance(data_rows, list):
            raise ValueError("A resposta FusionSolar realtime nao trouxe uma lista em data.")
        for row in data_rows:
            if not isinstance(row, dict):
                continue
            station_code = str(row.get("stationCode") or "").strip()
            if station_code:
                real_time_map[station_code] = row

    return real_time_map


def fetch_fusionsolar_device_list(
    session: requests.Session,
    *,
    base_url: str,
    endpoint: str,
    station_codes: list[str],
) -> list[dict[str, Any]]:
    url = build_provider_url(base_url, endpoint)
    devices: list[dict[str, Any]] = []
    for group in chunked(station_codes, 100):
        payload = post_fusionsolar_json(
            session,
            url,
            {"stationCodes": ",".join(group)},
            expected_message="Falha ao obter a lista de dispositivos FusionSolar.",
        )
        data = payload.get("data") or []
        rows = data.get("list") if isinstance(data, dict) else data
        if not isinstance(rows, list):
            raise ValueError("A resposta FusionSolar de dispositivos nao trouxe uma lista em data.")
        devices.extend([row for row in rows if isinstance(row, dict)])
    return devices


def fetch_fusionsolar_device_realtime_map(
    session: requests.Session,
    *,
    base_url: str,
    endpoint: str,
    devices: list[dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    url = build_provider_url(base_url, endpoint)
    real_time_map: dict[str, dict[str, Any]] = {}
    devices_by_type: dict[int, list[str]] = {}
    for device in devices:
        dev_type_id = device.get("dev_type_id")
        external_device_id = str(device.get("external_device_id") or "").strip()
        if dev_type_id is None or not external_device_id:
            continue
        devices_by_type.setdefault(int(dev_type_id), []).append(external_device_id)
    for dev_type_id, device_ids in devices_by_type.items():
        for group in chunked(device_ids, 100):
            payload = post_fusionsolar_json(
                session,
                url,
                {"devIds": ",".join(group), "devTypeId": dev_type_id},
                expected_message="Falha ao obter os dados realtime dos dispositivos FusionSolar.",
            )
            rows = payload.get("data") or []
            if not isinstance(rows, list):
                raise ValueError("A resposta FusionSolar realtime de dispositivos nao trouxe uma lista em data.")
            for row in rows:
                if not isinstance(row, dict):
                    continue
                for key in ("devId", "id", "devDn", "deviceDn", "esnCode", "sn"):
                    value = str(row.get(key) or "").strip()
                    if value:
                        real_time_map[value] = row
    return real_time_map


def fetch_fusionsolar_device_history(
    session: requests.Session,
    *,
    base_url: str,
    endpoint: str,
    devices: list[dict[str, Any]],
    target_date: date,
    call_delay_seconds: float = 0,
    sleeper: Any = time.sleep,
) -> list[dict[str, Any]]:
    url = build_provider_url(base_url, endpoint)
    start_time = int(datetime.combine(target_date, datetime.min.time()).timestamp() * 1000)
    end_time = int(datetime.combine(target_date + timedelta(days=1), datetime.min.time()).timestamp() * 1000) - 1
    rows: list[dict[str, Any]] = []
    calls_made = 0
    devices_by_type: dict[int, list[dict[str, Any]]] = {}
    for device in devices:
        if device.get("dev_type_id") is None or not device.get("external_device_id"):
            continue
        devices_by_type.setdefault(int(device["dev_type_id"]), []).append(device)
    for dev_type_id, typed_devices in devices_by_type.items():
        for group in [typed_devices[index : index + 10] for index in range(0, len(typed_devices), 10)]:
            if calls_made and call_delay_seconds > 0:
                sleeper(call_delay_seconds)
            payload = post_fusionsolar_json(
                session,
                url,
                {
                    "devIds": ",".join(str(device["external_device_id"]) for device in group),
                    "devTypeId": dev_type_id,
                    "startTime": start_time,
                    "endTime": end_time,
                },
                expected_message="Falha ao obter o historico dos inversores FusionSolar.",
            )
            calls_made += 1
            rows.extend(normalize_fusionsolar_device_history_rows(payload.get("data"), group))
    return rows


def normalize_fusionsolar_device_history_rows(
    data: Any,
    devices: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    device_by_id: dict[str, dict[str, Any]] = {}
    for device in devices:
        for key in ("external_device_id", "dev_dn", "sn"):
            value = str(device.get(key) or "").strip()
            if value:
                device_by_id[value] = device
    fallback_device = devices[0] if len(devices) == 1 else None
    normalized: list[dict[str, Any]] = []

    def visit(value: Any, inherited_device: dict[str, Any] | None = None) -> None:
        if isinstance(value, list):
            for item in value:
                visit(item, inherited_device)
            return
        if not isinstance(value, dict):
            return
        raw_device_id = first_non_empty(value, ["devId", "deviceId", "devDn", "deviceDn", "esnCode", "sn"])
        device = device_by_id.get(str(raw_device_id or "").strip()) or inherited_device or fallback_device
        data_map = value.get("dataItemMap") if isinstance(value.get("dataItemMap"), dict) else value
        sample_time = parse_datetime_value(first_non_empty(value, ["collectTime", "sampleTime", "time", "timestamp"]))
        if sample_time is None and data_map is not value:
            sample_time = parse_datetime_value(first_non_empty(data_map, ["collectTime", "sampleTime", "time", "timestamp"]))
        active_power = first_non_empty(data_map, ["active_power", "activePower", "active_power_kw", "power"])
        if device and sample_time is not None and active_power not in (None, ""):
            normalized.append(
                {
                    **device,
                    "sample_time": sample_time,
                    "active_power_kw": normalize_power_to_kw(active_power),
                    "raw_payload": value,
                }
            )
            return
        nested_found = False
        for key in ("list", "data", "historyData", "dataList", "records", "dataItemMap"):
            nested = value.get(key)
            if isinstance(nested, (list, dict)) and nested is not value:
                nested_found = True
                visit(nested, device)
        if not nested_found:
            for key, nested in value.items():
                mapped_device = device_by_id.get(str(key))
                if mapped_device and isinstance(nested, (list, dict)):
                    visit(nested, mapped_device)

    visit(data)
    return normalized


def fetch_fusionsolar_alarm_map(
    session: requests.Session,
    *,
    base_url: str,
    endpoint: str,
    station_codes: list[str],
) -> dict[str, list[dict[str, Any]]]:
    url = build_provider_url(base_url, endpoint)
    alarm_map: dict[str, list[dict[str, Any]]] = {}
    now_ms = int(datetime.now().timestamp() * 1000)

    for group in chunked(station_codes, 100):
        payload = post_fusionsolar_json(
            session,
            url,
            {
                "stationCodes": ",".join(group),
                "beginTime": 0,
                "endTime": now_ms,
                "language": DEFAULT_FUSIONSOLAR_ALARMS_LANGUAGE,
            },
            expected_message="Falha ao obter alarmes ativos FusionSolar.",
        )
        alarm_rows = payload.get("data") or []
        if not isinstance(alarm_rows, list):
            raise ValueError("A resposta FusionSolar de alarmes nao trouxe uma lista em data.")
        for row in alarm_rows:
            if not isinstance(row, dict):
                continue
            station_code = str(row.get("stationCode") or "").strip()
            if not station_code:
                continue
            alarm_map.setdefault(station_code, []).append(row)

    return alarm_map


def collect_time_ms(collect_date: date) -> int:
    return int(datetime.combine(collect_date, datetime.min.time()).timestamp() * 1000)


def collect_time_noon_ms(collect_date: date) -> int:
    return int(datetime.combine(collect_date, datetime.min.time().replace(hour=12)).timestamp() * 1000)


def normalize_fusionsolar_kpi_rows(data: Any) -> list[dict[str, Any]]:
    if isinstance(data, list):
        return [row for row in data if isinstance(row, dict)]
    if isinstance(data, dict):
        rows = data.get("list")
        if isinstance(rows, list):
            return [row for row in rows if isinstance(row, dict)]
        return [data]
    return []


def parse_fusionsolar_collect_date(row: dict[str, Any], fallback_date: date | None = None) -> date | None:
    for key in ("collectTime", "collect_time", "time", "timestamp"):
        raw_value = row.get(key)
        if raw_value in (None, ""):
            continue
        try:
            timestamp = int(float(str(raw_value).strip()))
            if timestamp > 10_000_000_000:
                timestamp = timestamp // 1000
            return datetime.fromtimestamp(timestamp).date()
        except (TypeError, ValueError, OSError, OverflowError):
            parsed = parse_date_value(str(raw_value))
            if parsed:
                return parsed
    for key in ("collectDate", "date", "day", "periodDate"):
        parsed = parse_date_value(str(row.get(key) or "").strip())
        if parsed:
            return parsed
    return fallback_date


def fetch_fusionsolar_kpi_map(
    session: requests.Session,
    *,
    base_url: str,
    endpoint: str,
    station_codes: list[str],
    collect_date: date,
    expected_message: str,
) -> dict[str, dict[str, Any]]:
    url = build_provider_url(base_url, endpoint)
    kpi_map: dict[str, dict[str, Any]] = {}

    for group in chunked(station_codes, 100):
        payload = post_fusionsolar_json(
            session,
            url,
            {
                "stationCodes": ",".join(group),
                "collectTime": collect_time_ms(collect_date),
            },
            expected_message=expected_message,
        )
        for row in normalize_fusionsolar_kpi_rows(payload.get("data")):
            station_code = str(row.get("stationCode") or row.get("plantCode") or "").strip()
            if station_code:
                enriched = dict(row)
                enriched["payload_json"] = json.dumps(row, ensure_ascii=True)
                kpi_map[station_code] = enriched

    return kpi_map


def fetch_fusionsolar_kpi_rows(
    session: requests.Session,
    *,
    base_url: str,
    endpoint: str,
    station_codes: list[str],
    collect_date: date,
    expected_message: str,
) -> list[dict[str, Any]]:
    url = build_provider_url(base_url, endpoint)
    payload = post_fusionsolar_json(
        session,
        url,
        {
            "stationCodes": ",".join(station_codes),
            "collectTime": collect_time_noon_ms(collect_date.replace(day=1)),
        },
        expected_message=expected_message,
    )
    rows: list[dict[str, Any]] = []
    for row in normalize_fusionsolar_kpi_rows(payload.get("data")):
        enriched = dict(row)
        enriched["payload_json"] = json.dumps(row, ensure_ascii=True)
        rows.append(enriched)
    return rows


def fetch_fusionsolar_kpi_day_rows(
    session: requests.Session,
    base_url: str,
    endpoint: str,
    station_codes: list[str],
    collect_date: date,
) -> list[dict[str, Any]]:
    return fetch_fusionsolar_kpi_rows(
        session,
        base_url=base_url,
        endpoint=endpoint,
        station_codes=station_codes,
        collect_date=collect_date,
        expected_message="Falha ao obter os KPIs diarios FusionSolar.",
    )


def fetch_fusionsolar_kpi_day_map(
    session: requests.Session,
    base_url: str,
    endpoint: str,
    station_codes: list[str],
    collect_date: date,
) -> dict[str, dict[str, Any]]:
    return fetch_fusionsolar_kpi_map(
        session,
        base_url=base_url,
        endpoint=endpoint,
        station_codes=station_codes,
        collect_date=collect_date,
        expected_message="Falha ao obter os KPIs diarios FusionSolar.",
    )


def fetch_fusionsolar_kpi_month_map(
    session: requests.Session,
    base_url: str,
    endpoint: str,
    station_codes: list[str],
    collect_date: date,
) -> dict[str, dict[str, Any]]:
    month_date = collect_date.replace(day=1)
    return fetch_fusionsolar_kpi_map(
        session,
        base_url=base_url,
        endpoint=endpoint,
        station_codes=station_codes,
        collect_date=month_date,
        expected_message="Falha ao obter os KPIs mensais FusionSolar.",
    )


def parse_kwp_value(value: Any) -> float | None:
    if value is None:
        return None
    raw = str(value).strip().replace(",", ".")
    if not raw or raw == "-":
        return None
    raw = re.sub(r"[^0-9.\-]", "", raw)
    if raw in ("", "-", "."):
        return None
    try:
        parsed = float(raw)
    except ValueError:
        return None
    return parsed if parsed > 0 else None


def parse_float_value(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).strip().replace(",", "."))
    except (TypeError, ValueError):
        return None


def parse_int_value(value: Any) -> int | None:
    parsed = parse_float_value(value)
    return int(parsed) if parsed is not None else None


def select_production_value(data_item_map: dict[str, Any] | None) -> tuple[float | None, str, str]:
    data = data_item_map or {}
    for key in ("PVYield", "inverterYield", "inverter_power"):
        raw_value = data.get(key)
        value = parse_float_value(raw_value)
        if value is not None:
            return value, key, str(raw_value)
    return None, "", ""


def select_production_kwh(data_item_map: dict[str, Any] | None) -> float | None:
    return select_production_value(data_item_map)[0]


def build_missing_production_note(
    data_item_map: dict[str, Any] | None,
    *,
    station_code: str,
    period_type: str,
    period_date: date,
) -> str:
    available_keys = sorted(str(key) for key in (data_item_map or {}).keys())
    keys_text = ", ".join(available_keys) if available_keys else "none"
    return (
        f"No production key found. Available keys: {keys_text}. "
        f"stationCode={station_code or '-'}; period_type={period_type}; period_date={period_date.isoformat()}"
    )


def is_fusionsolar_rate_limit_error(exc: Exception | str) -> bool:
    message = str(exc)
    return "failCode=407" in message or "error code 407" in message or "código 407" in message


def is_fusionsolar_session_expired_error(exc: Exception | str) -> bool:
    message = str(exc)
    return "failCode=305" in message or "USER_MUST_RELOGIN" in message


FUSIONSOLAR_PERFORMANCE_COOLDOWN_KEY = "fusionsolar_performance_cooldown_until"
FUSIONSOLAR_RATE_LIMIT_LAST_ALERT_KEY = "fusionsolar_rate_limit_last_alert_key"
FUSIONSOLAR_RATE_LIMIT_LAST_ALERT_AT_KEY = "fusionsolar_rate_limit_last_alert_at"


def get_app_state_value(conn: sqlite3.Connection, key: str) -> str:
    row = conn.execute("SELECT value FROM app_state WHERE key = ?", (key,)).fetchone()
    return str(row["value"] or "") if row else ""


def set_app_state_value(conn: sqlite3.Connection, key: str, value: str) -> None:
    now = datetime.now().isoformat(timespec="seconds")
    conn.execute(
        """
        INSERT INTO app_state (key, value, updated_at)
        VALUES (?, ?, ?)
        ON CONFLICT(key) DO UPDATE SET value = excluded.value, updated_at = excluded.updated_at
        """,
        (key, value, now),
    )


def mark_fusionsolar_performance_rate_limited(
    conn: sqlite3.Connection | None = None,
    now_value: datetime | None = None,
) -> str:
    global FUSIONSOLAR_PERFORMANCE_RATE_LIMIT_UNTIL
    now_value = now_value or datetime.now()
    FUSIONSOLAR_PERFORMANCE_RATE_LIMIT_UNTIL = now_value + timedelta(minutes=FUSIONSOLAR_PERFORMANCE_RATE_LIMIT_MINUTES)
    if conn is not None:
        set_app_state_value(
            conn,
            FUSIONSOLAR_PERFORMANCE_COOLDOWN_KEY,
            FUSIONSOLAR_PERFORMANCE_RATE_LIMIT_UNTIL.isoformat(timespec="seconds"),
        )
    message = (
        "FusionSolar API temporariamente limitada. "
        f"Tenta novamente depois de {FUSIONSOLAR_PERFORMANCE_RATE_LIMIT_UNTIL.isoformat(timespec='minutes')}."
    )
    if conn is not None:
        notify_fusionsolar_rate_limit(conn, FUSIONSOLAR_PERFORMANCE_RATE_LIMIT_UNTIL, message)
    return message


def notify_fusionsolar_rate_limit(conn: sqlite3.Connection, cooldown_until: datetime, message: str) -> None:
    now_value = datetime.now()
    recent_cutoff = (now_value - timedelta(minutes=FUSIONSOLAR_PERFORMANCE_RATE_LIMIT_MINUTES)).isoformat(timespec="seconds")
    recent_alert = conn.execute(
        """
        SELECT 1
        FROM telegram_alerts
        WHERE alert_type = 'fusionsolar_api_limit'
          AND status IN ('sent', 'blocked', 'failed')
          AND sent_at >= ?
        LIMIT 1
        """,
        (recent_cutoff,),
    ).fetchone()
    if recent_alert is not None:
        return
    last_alert_at = get_app_state_value(conn, FUSIONSOLAR_RATE_LIMIT_LAST_ALERT_AT_KEY)
    if last_alert_at:
        try:
            if datetime.fromisoformat(last_alert_at) >= now_value - timedelta(minutes=FUSIONSOLAR_PERFORMANCE_RATE_LIMIT_MINUTES):
                return
        except ValueError:
            pass

    alert_key = f"fusionsolar_api_limit:{now_value.strftime('%Y%m%d%H%M')}"
    set_app_state_value(conn, FUSIONSOLAR_RATE_LIMIT_LAST_ALERT_KEY, alert_key)
    set_app_state_value(conn, FUSIONSOLAR_RATE_LIMIT_LAST_ALERT_AT_KEY, now_value.isoformat(timespec="seconds"))
    conn.commit()
    telegram_message = (
        "<b>FusionSolar API limitada</b>\n\n"
        "A API devolveu limite de chamadas/tokens esgotado para KPIs de produção.\n"
        f"{message}\n\n"
        "Os relatórios de produção e backfills podem falhar até o limite renovar."
    )
    send_and_record_telegram_alert(
        conn,
        None,
        "fusionsolar_api_limit",
        alert_key,
        telegram_message,
    )


def get_fusionsolar_performance_cooldown_reason(
    conn: sqlite3.Connection | None = None,
    now_value: datetime | None = None,
) -> str:
    now_value = now_value or datetime.now()
    cooldown_until = FUSIONSOLAR_PERFORMANCE_RATE_LIMIT_UNTIL
    if conn is not None:
        raw_value = get_app_state_value(conn, FUSIONSOLAR_PERFORMANCE_COOLDOWN_KEY)
        if raw_value:
            try:
                persisted_until = datetime.fromisoformat(raw_value)
                if cooldown_until is None or persisted_until > cooldown_until:
                    cooldown_until = persisted_until
            except ValueError:
                pass
    if cooldown_until and cooldown_until > now_value:
        remaining_seconds = int((cooldown_until - now_value).total_seconds())
        remaining_minutes = max(1, (remaining_seconds + 59) // 60)
        return (
            "FusionSolar API temporariamente limitada. "
            f"Tenta novamente depois de {cooldown_until.isoformat(timespec='minutes')} ({remaining_minutes} min)."
        )
    return ""


def fusionsolar_cooldown_sleep_seconds(
    conn: sqlite3.Connection | None = None,
    now_value: datetime | None = None,
) -> int:
    now_value = now_value or datetime.now()
    cooldown_until = FUSIONSOLAR_PERFORMANCE_RATE_LIMIT_UNTIL
    if conn is not None:
        raw_value = get_app_state_value(conn, FUSIONSOLAR_PERFORMANCE_COOLDOWN_KEY)
        if raw_value:
            try:
                persisted_until = datetime.fromisoformat(raw_value)
                if cooldown_until is None or persisted_until > cooldown_until:
                    cooldown_until = persisted_until
            except ValueError:
                pass
    if cooldown_until and cooldown_until > now_value:
        return max(1, int((cooldown_until - now_value).total_seconds()) + 5)
    return 0


def calculate_specific_yield(production_kwh: float | None, kwp: float | None) -> float | None:
    if production_kwh is None or not kwp:
        return None
    return production_kwh / kwp


def classify_performance_status(
    production_kwh: float | None,
    kwp: float | None,
    expected_kwh: float | None,
    *,
    warning_deviation_pct: float = -10,
    alert_deviation_pct: float = -20,
    critical_deviation_pct: float = -30,
) -> tuple[str, str, float | None]:
    if production_kwh is None:
        return "Sem dados", "missing_production", None
    if not kwp:
        return "Sem referência", "missing_kwp", None
    if expected_kwh is None or expected_kwh <= 0:
        return "Sem referência", "ok", None

    deviation_pct = ((production_kwh - expected_kwh) / expected_kwh) * 100
    if deviation_pct >= warning_deviation_pct:
        return "OK", "ok", deviation_pct
    if deviation_pct >= alert_deviation_pct:
        return "Atenção", "ok", deviation_pct
    if deviation_pct >= critical_deviation_pct:
        return "Alerta", "ok", deviation_pct
    return "Crítico", "ok", deviation_pct


def get_performance_settings(conn: sqlite3.Connection, asset_id: int) -> dict[str, Any]:
    row = conn.execute("SELECT * FROM performance_settings WHERE asset_id = ?", (asset_id,)).fetchone()
    defaults = {
        "asset_id": asset_id,
        "enabled": 1,
        "warning_deviation_pct": -10.0,
        "alert_deviation_pct": -20.0,
        "critical_deviation_pct": -30.0,
        "baseline_years": 2,
        "min_baseline_points": 1,
        "monthly_budget_json": "",
        "notes": "",
        "updated_at": "",
    }
    if row is None:
        return defaults
    defaults.update(dict(row))
    return defaults


def get_monthly_budget_specific_yield(settings: dict[str, Any], period_date: date) -> float | None:
    raw = str(settings.get("monthly_budget_json") or "").strip()
    if not raw:
        return None
    try:
        payload = json.loads(raw)
    except json.JSONDecodeError:
        return None
    if not isinstance(payload, dict):
        return None
    return parse_float_value(payload.get(f"{period_date.month:02d}"))


def calculate_historical_baseline(
    conn: sqlite3.Connection,
    *,
    asset_id: int,
    provider: str,
    period_type: str,
    period_date: date,
    baseline_years: int,
    min_baseline_points: int,
) -> tuple[float | None, float | None, str, str]:
    result = calculate_expected_production_with_diagnostic(
        conn,
        asset_id=asset_id,
        provider=provider,
        period_type=period_type,
        period_date=period_date,
        kwp=None,
        settings={
            "baseline_years": baseline_years,
            "min_baseline_points": min_baseline_points,
            "monthly_budget_json": "",
        },
    )
    return (
        result["expected_kwh"],
        result["expected_specific_yield"],
        result["expected_source"],
        result["quality"],
    )


def same_date_previous_years(period_date: date, baseline_years: int) -> list[date]:
    candidates: list[date] = []
    for year_offset in range(1, max(int(baseline_years or 1), 1) + 1):
        try:
            previous = period_date.replace(year=period_date.year - year_offset)
        except ValueError:
            previous = period_date.replace(year=period_date.year - year_offset, day=28)
        candidates.append(previous)
    return candidates


def load_valid_baseline_rows(
    conn: sqlite3.Connection,
    *,
    asset_id: int,
    provider: str,
    period_type: str,
    candidate_dates: list[date],
) -> list[sqlite3.Row]:
    if not candidate_dates:
        return []
    placeholders = ",".join("?" for _ in candidate_dates)
    return query_all(
        conn,
        f"""
        SELECT production_kwh, specific_yield, period_date
        FROM production_records
        WHERE asset_id = ? AND provider = ? AND period_type = ?
          AND period_date IN ({placeholders})
          AND production_kwh IS NOT NULL
          AND specific_yield IS NOT NULL
          AND COALESCE(data_quality, '') != 'missing_production'
        ORDER BY period_date DESC
        """,
        [asset_id, provider, period_type, *[item.isoformat() for item in candidate_dates]],
    )


def calculate_mtd_baseline(
    conn: sqlite3.Connection,
    *,
    asset_id: int,
    provider: str,
    period_date: date,
    baseline_years: int,
    min_baseline_points: int,
    today_value: date | None = None,
) -> dict[str, Any]:
    today_value = today_value or date.today()
    period_start = period_date.replace(day=1)
    if period_start.year == today_value.year and period_start.month == today_value.month:
        period_end = today_value
    else:
        period_end = period_start.replace(day=calendar.monthrange(period_start.year, period_start.month)[1])
    day_span = period_end.day
    candidate_ranges: list[str] = []
    yearly_values: list[tuple[float, float]] = []
    for year_offset in range(1, max(int(baseline_years or 1), 1) + 1):
        start = period_start.replace(year=period_start.year - year_offset)
        end_day = min(day_span, calendar.monthrange(start.year, start.month)[1])
        end = start.replace(day=end_day)
        candidate_ranges.append(f"{start.isoformat()}..{end.isoformat()}")
        rows = query_all(
            conn,
            """
            SELECT production_kwh, specific_yield
            FROM production_records
            WHERE asset_id = ? AND provider = ? AND period_type = 'day'
              AND period_date BETWEEN ? AND ?
              AND production_kwh IS NOT NULL
              AND specific_yield IS NOT NULL
              AND COALESCE(data_quality, '') != 'missing_production'
            ORDER BY period_date ASC
            """,
            (asset_id, provider, start.isoformat(), end.isoformat()),
        )
        if len(rows) == end_day:
            yearly_values.append(
                (
                    sum(float(row["production_kwh"]) for row in rows),
                    sum(float(row["specific_yield"]) for row in rows),
                )
            )
    if len(yearly_values) < max(int(min_baseline_points or 1), 1):
        return {
            "expected_kwh": None,
            "expected_specific_yield": None,
            "expected_source": "none",
            "quality": "partial_history" if yearly_values else "ok",
            "diagnostic": {
                "historical_records_found": len(yearly_values),
                "baseline_years": baseline_years,
                "min_baseline_points": min_baseline_points,
                "candidate_historical_dates": candidate_ranges,
                "expected_source_attempted": "historical_same_period",
                "no_reference_reason": "MTD reference requires historical daily records for same period",
                "period_start": period_start.isoformat(),
                "period_end": period_end.isoformat(),
            },
        }
    expected_kwh = sum(item[0] for item in yearly_values) / len(yearly_values)
    expected_specific_yield = sum(item[1] for item in yearly_values) / len(yearly_values)
    return {
        "expected_kwh": expected_kwh,
        "expected_specific_yield": expected_specific_yield,
        "expected_source": "historical_same_period",
        "quality": "ok",
        "diagnostic": {
            "historical_records_found": len(yearly_values),
            "baseline_years": baseline_years,
            "min_baseline_points": min_baseline_points,
            "candidate_historical_dates": candidate_ranges,
            "expected_source_attempted": "historical_same_period",
            "no_reference_reason": "",
            "period_start": period_start.isoformat(),
            "period_end": period_end.isoformat(),
        },
    }


def calculate_expected_production_with_diagnostic(
    conn: sqlite3.Connection,
    *,
    asset_id: int,
    provider: str,
    period_type: str,
    period_date: date,
    kwp: float | None,
    settings: dict[str, Any],
    asset_name: str = "",
    today_value: date | None = None,
) -> dict[str, Any]:
    baseline_years = int(settings.get("baseline_years") or 2)
    min_baseline_points = int(settings.get("min_baseline_points") or 1)
    if period_type == "mtd":
        result = calculate_mtd_baseline(
            conn,
            asset_id=asset_id,
            provider=provider,
            period_date=period_date,
            baseline_years=baseline_years,
            min_baseline_points=min_baseline_points,
            today_value=today_value,
        )
    else:
        historical_type = "month" if period_type == "month" else "day"
        candidates = same_date_previous_years(period_date.replace(day=1) if period_type == "month" else period_date, baseline_years)
        rows = load_valid_baseline_rows(
            conn,
            asset_id=asset_id,
            provider=provider,
            period_type=historical_type,
            candidate_dates=candidates,
        )
        diagnostic = {
            "historical_records_found": len(rows),
            "baseline_years": baseline_years,
            "min_baseline_points": min_baseline_points,
            "candidate_historical_dates": [item.isoformat() for item in candidates],
            "expected_source_attempted": "historical_same_period",
            "no_reference_reason": "",
        }
        if len(rows) < min_baseline_points:
            label = "monthly" if period_type == "month" else "daily"
            diagnostic["no_reference_reason"] = (
                f"No historical {label} records found for same {'month' if period_type == 'month' else 'day'} in previous years"
                if not rows
                else f"Only {len(rows)} baseline points found, minimum is {min_baseline_points}"
            )
            result = {
                "expected_kwh": None,
                "expected_specific_yield": None,
                "expected_source": "none",
                "quality": "partial_history" if rows else "ok",
                "diagnostic": diagnostic,
            }
        else:
            production_values = [float(row["production_kwh"]) for row in rows]
            specific_values = [float(row["specific_yield"]) for row in rows]
            result = {
                "expected_kwh": sum(production_values) / len(production_values),
                "expected_specific_yield": sum(specific_values) / len(specific_values),
                "expected_source": "historical_same_period",
                "quality": "ok",
                "diagnostic": diagnostic,
            }

    if result["expected_kwh"] is None:
        budget_specific = get_monthly_budget_specific_yield(settings, period_date)
        if budget_specific is not None and kwp:
            if period_type == "day":
                budget_specific = budget_specific / calendar.monthrange(period_date.year, period_date.month)[1]
            result = {
                "expected_kwh": budget_specific * kwp,
                "expected_specific_yield": budget_specific,
                "expected_source": "monthly_budget",
                "quality": "ok",
                "diagnostic": {
                    **result["diagnostic"],
                    "expected_source_attempted": "monthly_budget",
                    "no_reference_reason": "",
                },
            }
    if kwp is None:
        result["diagnostic"]["no_reference_reason"] = "Missing kWp"
    if result["expected_specific_yield"] is None and not result["diagnostic"].get("no_reference_reason"):
        result["diagnostic"]["no_reference_reason"] = "Missing expected_specific_yield"

    logger = current_app.logger if has_app_context() else logging.getLogger(__name__)
    logger.info(
        "Performance reference calculation: asset_id=%s asset_name=%s period_type=%s period_date=%s baseline_years=%s candidate_dates=%s valid_baseline_records=%s expected_specific_yield=%s expected_source=%s no_reference_reason=%s",
        asset_id,
        asset_name,
        period_type,
        period_date.isoformat(),
        baseline_years,
        result["diagnostic"].get("candidate_historical_dates"),
        result["diagnostic"].get("historical_records_found"),
        result["expected_specific_yield"],
        result["expected_source"],
        result["diagnostic"].get("no_reference_reason", ""),
    )
    return result


def calculate_expected_production(
    conn: sqlite3.Connection,
    *,
    asset_id: int,
    provider: str,
    period_type: str,
    period_date: date,
    kwp: float | None,
    settings: dict[str, Any],
) -> tuple[float | None, float | None, str, str]:
    result = calculate_expected_production_with_diagnostic(
        conn,
        asset_id=asset_id,
        provider=provider,
        period_type=period_type,
        period_date=period_date,
        kwp=kwp,
        settings=settings,
    )
    return result["expected_kwh"], result["expected_specific_yield"], result["expected_source"], result["quality"]


def upsert_production_record(
    conn: sqlite3.Connection,
    *,
    asset_id: int,
    provider: str,
    external_id: str,
    period_type: str,
    period_date: date,
    production_kwh: float | None,
    specific_yield: float | None,
    expected_kwh: float | None,
    expected_specific_yield: float | None,
    deviation_pct: float | None,
    performance_status: str,
    expected_source: str,
    data_quality: str,
    notes: str,
    payload_json: str,
    selected_production_key: str = "",
    selected_production_raw_value: str = "",
    reference_diagnostic_json: str = "",
) -> str:
    now = datetime.now().isoformat(timespec="seconds")
    existing = conn.execute(
        """
        SELECT id
        FROM production_records
        WHERE asset_id = ? AND provider = ? AND period_type = ? AND period_date = ?
        LIMIT 1
        """,
        (asset_id, provider, period_type, period_date.isoformat()),
    ).fetchone()
    conn.execute(
        """
        INSERT INTO production_records (
            asset_id, provider, external_id, period_type, period_date, production_kwh, specific_yield,
            expected_kwh, expected_specific_yield, deviation_pct, performance_status, expected_source,
            data_quality, notes, selected_production_key, selected_production_raw_value, reference_diagnostic_json,
            payload_json, created_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(asset_id, provider, period_type, period_date) DO UPDATE SET
            external_id = excluded.external_id,
            production_kwh = excluded.production_kwh,
            specific_yield = excluded.specific_yield,
            expected_kwh = excluded.expected_kwh,
            expected_specific_yield = excluded.expected_specific_yield,
            deviation_pct = excluded.deviation_pct,
            performance_status = excluded.performance_status,
            expected_source = excluded.expected_source,
            data_quality = excluded.data_quality,
            notes = excluded.notes,
            selected_production_key = excluded.selected_production_key,
            selected_production_raw_value = excluded.selected_production_raw_value,
            reference_diagnostic_json = excluded.reference_diagnostic_json,
            payload_json = excluded.payload_json,
            updated_at = excluded.updated_at
        """,
        (
            asset_id,
            provider,
            external_id or None,
            period_type,
            period_date.isoformat(),
            production_kwh,
            specific_yield,
            expected_kwh,
            expected_specific_yield,
            deviation_pct,
            performance_status,
            expected_source,
            data_quality,
            notes,
            selected_production_key or None,
            selected_production_raw_value or None,
            reference_diagnostic_json or None,
            payload_json,
            now,
            now,
        ),
    )
    return "updated" if existing else "inserted"


def store_production_kpi_record(
    conn: sqlite3.Connection,
    *,
    asset_row: sqlite3.Row | dict[str, Any],
    provider: str,
    external_id: str,
    period_type: str,
    period_date: date,
    kpi_row: dict[str, Any],
    notes_prefix: str = "",
) -> dict[str, Any]:
    data_item_map = kpi_row.get("dataItemMap") if isinstance(kpi_row, dict) else {}
    if not isinstance(data_item_map, dict):
        data_item_map = {}
    production_kwh, selected_key, selected_raw_value = select_production_value(data_item_map)
    kwp = parse_kwp_value(asset_row["kwp"])
    specific_yield = calculate_specific_yield(production_kwh, kwp)
    asset_id = int(asset_row["asset_id"] if "asset_id" in asset_row.keys() else asset_row["id"])
    settings = get_performance_settings(conn, asset_id)
    for key in settings:
        try:
            row_value = asset_row[key]
        except (KeyError, IndexError):
            continue
        if row_value is not None:
            settings[key] = row_value

    reference_result = calculate_expected_production_with_diagnostic(
        conn,
        asset_id=asset_id,
        provider=provider,
        period_type=period_type,
        period_date=period_date,
        kwp=kwp,
        settings=settings,
        asset_name=str(asset_row["project_name"] if "project_name" in asset_row.keys() else ""),
    )
    expected_kwh = reference_result["expected_kwh"]
    expected_specific_yield = reference_result["expected_specific_yield"]
    expected_source = reference_result["expected_source"]
    baseline_quality = reference_result["quality"]
    performance_status, data_quality, deviation_pct = classify_performance_status(
        production_kwh,
        kwp,
        expected_kwh,
        warning_deviation_pct=float(settings.get("warning_deviation_pct") or -10),
        alert_deviation_pct=float(settings.get("alert_deviation_pct") or -20),
        critical_deviation_pct=float(settings.get("critical_deviation_pct") or -30),
    )
    if data_quality == "ok" and baseline_quality == "partial_history" and expected_source == "none":
        data_quality = "partial_history"

    notes_parts = [notes_prefix] if notes_prefix else []
    if production_kwh is None:
        notes_parts.append(
            build_missing_production_note(
                data_item_map,
                station_code=external_id,
                period_type=period_type,
                period_date=period_date,
            )
        )
    if kwp is None:
        notes_parts.append("kWp local em falta ou invalido.")
    if expected_source == "none":
        notes_parts.append("Sem histórico ou orçamento mensal para referência.")

    if production_kwh is None:
        existing_valid = conn.execute(
            """
            SELECT id
            FROM production_records
            WHERE asset_id = ? AND provider = ? AND period_type = ? AND period_date = ?
              AND production_kwh IS NOT NULL
              AND COALESCE(data_quality, '') != 'missing_production'
            LIMIT 1
            """,
            (asset_id, provider, period_type, period_date.isoformat()),
        ).fetchone()
        if existing_valid:
            return {
                "upsert_status": "skipped_existing_valid",
                "production_kwh": None,
                "specific_yield": None,
                "performance_status": performance_status,
                "data_quality": data_quality,
            }

    upsert_status = upsert_production_record(
        conn,
        asset_id=asset_id,
        provider=provider,
        external_id=external_id,
        period_type=period_type,
        period_date=period_date,
        production_kwh=production_kwh,
        specific_yield=specific_yield,
        expected_kwh=expected_kwh,
        expected_specific_yield=expected_specific_yield,
        deviation_pct=deviation_pct,
        performance_status=performance_status,
        expected_source=expected_source,
        data_quality=data_quality,
        notes=" ".join(notes_parts),
        payload_json=str(kpi_row.get("payload_json") or json.dumps(kpi_row, ensure_ascii=True)),
        selected_production_key=selected_key,
        selected_production_raw_value=selected_raw_value,
        reference_diagnostic_json=json.dumps(reference_result["diagnostic"], ensure_ascii=True),
    )
    return {
        "upsert_status": upsert_status,
        "production_kwh": production_kwh,
        "performance_status": performance_status,
        "data_quality": data_quality,
    }


def first_non_empty(payload: dict[str, Any], keys: list[str]) -> str:
    for key in keys:
        value = payload.get(key)
        if value not in (None, ""):
            return str(value).strip()
    return ""


def parse_fusionsolar_pv_inputs(row: dict[str, Any]) -> tuple[dict[str, Any], dict[str, Any]]:
    currents: dict[str, Any] = {}
    voltages: dict[str, Any] = {}
    source = row.get("dataItemMap") if isinstance(row.get("dataItemMap"), dict) else row
    for index in range(1, 37):
        current_key = f"pv{index}_i"
        voltage_key = f"pv{index}_u"
        if current_key in source and source[current_key] not in (None, ""):
            currents[current_key] = source[current_key]
        if voltage_key in source and source[voltage_key] not in (None, ""):
            voltages[voltage_key] = source[voltage_key]
    return currents, voltages


def calculate_pv_input_health(
    currents: dict[str, Any],
    voltages: dict[str, Any],
    *,
    expected_string_indexes: set[int],
) -> dict[str, Any]:
    expected_inputs = sorted(expected_string_indexes)
    available_inputs = 0
    unavailable_inputs = 0
    voltage_values: dict[str, float] = {}
    for index in expected_inputs:
        voltage = parse_float_value(voltages.get(f"pv{index}_u"))
        voltage_values[str(index)] = voltage or 0.0
        if voltage is not None and voltage > DEFAULT_STRING_PRESENT_VOLTAGE_THRESHOLD:
            available_inputs += 1
        else:
            unavailable_inputs += 1
    total_inputs = len(expected_inputs)
    return {
        "available_strings": available_inputs,
        "total_strings": total_inputs,
        "unavailable_strings": unavailable_inputs,
        "string_availability_pct": round(available_inputs / total_inputs * 100, 2) if total_inputs else None,
        "pv_input_diagnostics": {
            "expected_inputs": expected_inputs,
            "voltages_v": voltage_values,
        },
    }


def learn_expected_strings_from_voltage(
    conn: sqlite3.Connection,
    provider_device_id: int,
    voltages: dict[str, Any],
    observed_at: str,
) -> set[int]:
    learned_indexes: set[int] = set()
    for key, raw_voltage in voltages.items():
        voltage = parse_float_value(raw_voltage)
        if voltage is None or voltage <= DEFAULT_STRING_PRESENT_VOLTAGE_THRESHOLD:
            continue
        index = parse_int_value(key.removeprefix("pv").removesuffix("_u"))
        if index is None:
            continue
        existing = conn.execute(
            """
            SELECT *
            FROM provider_device_expected_strings
            WHERE provider_device_id = ? AND string_index = ?
            """,
            (provider_device_id, index),
        ).fetchone()
        now = datetime.now().isoformat(timespec="seconds")
        if existing is None:
            conn.execute(
                """
                INSERT INTO provider_device_expected_strings (
                    provider_device_id, string_index, expected, source, observed_count,
                    first_observed_at, last_observed_at, created_at, updated_at
                ) VALUES (?, ?, 0, 'auto', 1, ?, ?, ?, ?)
                """,
                (provider_device_id, index, observed_at, observed_at, now, now),
            )
            continue
        observed_count = int(existing["observed_count"] or 0) + 1
        expected = int(existing["expected"] or 0)
        source = existing["source"]
        if source == "auto" and observed_count >= DEFAULT_STRING_AUTO_LEARN_OBSERVATIONS:
            expected = 1
        conn.execute(
            """
            UPDATE provider_device_expected_strings
            SET expected = ?, observed_count = ?, last_observed_at = ?, updated_at = ?
            WHERE id = ?
            """,
            (expected, observed_count, observed_at, now, existing["id"]),
        )

    rows = query_all(
        conn,
        """
        SELECT string_index
        FROM provider_device_expected_strings
        WHERE provider_device_id = ? AND expected = 1
        """,
        (provider_device_id,),
    )
    learned_indexes.update(int(row["string_index"]) for row in rows)
    return learned_indexes


def normalize_fusionsolar_device_identity(row: dict[str, Any]) -> dict[str, Any]:
    dev_type_id = parse_int_value(first_non_empty(row, ["devTypeId", "dev_type_id", "deviceTypeId"]))
    model = first_non_empty(row, ["model", "devModel", "deviceModel", "invType"])
    rated_power_kw = normalize_power_to_kw(first_non_empty(row, ["ratedPower", "rated_power", "capacity", "nominalPower"]))
    return {
        "station_code": first_non_empty(row, ["stationCode", "plantCode"]),
        "external_device_id": first_non_empty(row, ["devId", "id", "devDn", "deviceDn", "esnCode", "sn"]),
        "dev_dn": first_non_empty(row, ["devDn", "deviceDn"]),
        "sn": first_non_empty(row, ["esnCode", "sn"]),
        "device_name": first_non_empty(row, ["devName", "deviceName", "name"]),
        "dev_type_id": dev_type_id,
        "model": model,
        "rated_power_kw": rated_power_kw if rated_power_kw is not None else infer_inverter_power_from_model(model),
    }


def normalize_power_to_kw(value: Any) -> float | None:
    parsed = parse_float_value(value)
    if parsed is None:
        return None
    return parsed / 1000 if parsed > 1000 else parsed


def infer_inverter_power_from_model(model: str | None) -> float | None:
    normalized = str(model or "").upper().replace(" ", "")
    match = re.search(r"(?:SUN2000-|^)(\d+(?:[.,]\d+)?)(?:KTL|K(?:-|$))", normalized)
    if not match:
        return None
    return parse_float_value(match.group(1).replace(",", "."))


def populate_missing_inverter_rated_power(conn: sqlite3.Connection) -> int:
    rows = conn.execute(
        """
        SELECT id, external_device_id, model, payload_json
        FROM provider_devices
        WHERE rated_power_kw IS NULL OR rated_power_kw <= 0
        """
    ).fetchall()
    updated = 0
    for row in rows:
        model = row["model"]
        if not model and row["payload_json"]:
            try:
                payload = json.loads(row["payload_json"])
                model = first_non_empty(payload, ["model", "devModel", "deviceModel", "invType"])
            except (TypeError, ValueError, json.JSONDecodeError):
                model = None
        rated_power_kw = infer_inverter_power_from_model(model)
        if rated_power_kw is None:
            continue
        conn.execute(
            "UPDATE provider_devices SET rated_power_kw = ? WHERE id = ?",
            (rated_power_kw, row["id"]),
        )
        conn.execute(
            """
            UPDATE inverter_availability_daily
            SET inverter_power_kw = ?
            WHERE provider = ? AND inverter_id = ?
              AND (inverter_power_kw IS NULL OR inverter_power_kw <= 0)
            """,
            (rated_power_kw, INTEGRATION_PROVIDER_FUSIONSOLAR, row["external_device_id"]),
        )
        conn.execute(
            """
            UPDATE inverter_power_samples
            SET inverter_power_kw = ?
            WHERE provider = ? AND inverter_id = ?
              AND (inverter_power_kw IS NULL OR inverter_power_kw <= 0)
            """,
            (rated_power_kw, INTEGRATION_PROVIDER_FUSIONSOLAR, row["external_device_id"]),
        )
        updated += 1
    return updated


def disable_removed_inverter_devices(conn: sqlite3.Connection) -> int:
    rows = conn.execute(
        "SELECT id, device_name FROM provider_devices WHERE enabled = 1"
    ).fetchall()
    removed_ids = [int(row["id"]) for row in rows if is_removed_inverter_name(row["device_name"])]
    if removed_ids:
        conn.executemany(
            "UPDATE provider_devices SET enabled = 0 WHERE id = ?",
            [(device_id,) for device_id in removed_ids],
        )
    return len(removed_ids)


def is_removed_inverter_name(device_name: str | None) -> bool:
    normalized = normalize_name(str(device_name or ""))
    return any(marker in normalized for marker in ("removido", "removed"))


def is_inverter_available(active_power_kw: float | None) -> bool:
    return active_power_kw is not None and active_power_kw > 0


def inverter_availability_slot(sample_time: datetime) -> datetime:
    minute = sample_time.minute - (sample_time.minute % INVERTER_AVAILABILITY_SLOT_MINUTES)
    return sample_time.replace(minute=minute, second=0, microsecond=0)


def apply_inverter_edge_tolerance(
    valid_slots: set[datetime],
    tolerance_minutes: int = INVERTER_AVAILABILITY_EDGE_TOLERANCE_MINUTES,
) -> set[datetime]:
    slots_by_date: dict[date, list[datetime]] = {}
    for slot in valid_slots:
        slots_by_date.setdefault(slot.date(), []).append(slot)
    considered: set[datetime] = set()
    tolerance = timedelta(minutes=max(tolerance_minutes, 0))
    for day_slots in slots_by_date.values():
        ordered = sorted(day_slots)
        if not ordered:
            continue
        first_slot = ordered[0]
        last_slot = ordered[-1]
        considered.update(
            slot
            for slot in ordered
            if slot - first_slot >= tolerance and last_slot - slot >= tolerance
        )
    return considered


def calculate_inverter_daily_availability(
    samples: list[dict[str, Any]],
    valid_slots: set[datetime] | None = None,
    edge_tolerance_minutes: int = INVERTER_AVAILABILITY_EDGE_TOLERANCE_MINUTES,
) -> dict[str, Any]:
    available_slots = {
        inverter_availability_slot(sample["sample_time"])
        for sample in samples
        if isinstance(sample.get("sample_time"), datetime) and is_inverter_available(sample.get("active_power_kw"))
    }
    raw_valid_slots = set(valid_slots) if valid_slots is not None else set(available_slots)
    considered_slots = apply_inverter_edge_tolerance(raw_valid_slots, edge_tolerance_minutes)
    available_count = len(available_slots & considered_slots)
    valid_count = len(considered_slots)
    return {
        "valid_slots": valid_count,
        "available_slots": available_count,
        "unavailable_slots": max(valid_count - available_count, 0),
        "availability_pct": round(available_count / valid_count * 100, 2) if valid_count else None,
    }


def calculate_weighted_plant_availability(inverter_rows: list[dict[str, Any]]) -> float | None:
    rows = [row for row in inverter_rows if row.get("availability_pct") is not None]
    if not rows:
        return None
    powers = [parse_float_value(row.get("inverter_power_kw")) for row in rows]
    if all(power is not None and power > 0 for power in powers):
        total_power = sum(float(power) for power in powers if power is not None)
        return round(
            sum(float(row["availability_pct"]) * float(power) for row, power in zip(rows, powers) if power is not None)
            / total_power,
            2,
        )
    return round(sum(float(row["availability_pct"]) for row in rows) / len(rows), 2)


def resolve_inverter_availability_period(
    period: str,
    raw_from_date: str = "",
    raw_to_date: str = "",
) -> tuple[date, date]:
    yesterday = date.today() - timedelta(days=1)
    if period == "current_month":
        return yesterday.replace(day=1), yesterday
    if period == "previous_month":
        current_month_start = date.today().replace(day=1)
        previous_month_end = current_month_start - timedelta(days=1)
        return previous_month_end.replace(day=1), previous_month_end
    if period == "custom":
        from_date = parse_date_value(raw_from_date) or yesterday
        to_date = min(parse_date_value(raw_to_date) or yesterday, yesterday)
        return from_date, to_date
    return yesterday, yesterday


def get_inverter_availability_report(
    conn: sqlite3.Connection,
    from_date: date,
    to_date: date,
    *,
    asset_id: int | None = None,
    om_only: bool = True,
    search: str = "",
) -> dict[str, Any]:
    if from_date > to_date:
        return {
            "average_pct": None,
            "plants": [],
            "inverters": [],
            "worst_plant": None,
            "low_availability_count": 0,
        }
    conditions = [
        "iad.provider = ?",
        "iad.availability_date BETWEEN ? AND ?",
        """EXISTS (
            SELECT 1
            FROM inverter_power_samples ips
            WHERE ips.provider = iad.provider
              AND ips.asset_id = iad.asset_id
              AND ips.inverter_id = iad.inverter_id
              AND ips.sample_time >= ? AND ips.sample_time < ?
        )""",
    ]
    params: list[Any] = [
        INTEGRATION_PROVIDER_FUSIONSOLAR,
        from_date.isoformat(),
        to_date.isoformat(),
        datetime.combine(from_date, datetime.min.time()).isoformat(timespec="seconds"),
        datetime.combine(to_date + timedelta(days=1), datetime.min.time()).isoformat(timespec="seconds"),
    ]
    if asset_id is not None:
        conditions.append("iad.asset_id = ?")
        params.append(asset_id)
    if om_only:
        conditions.append("a.active_contract = 'yes'")
    if search:
        conditions.append(
            "(a.project_name LIKE ? OR a.location LIKE ? OR a.company_name LIKE ? OR a.alias_blob LIKE ?)"
        )
        wildcard = f"%{search}%"
        params.extend([wildcard, wildcard, wildcard, wildcard])
    rows = conn.execute(
        f"""
        SELECT
            iad.asset_id,
            a.project_name,
            iad.inverter_id,
            iad.inverter_name,
            iad.inverter_power_kw,
            MAX(pd.device_name) AS provider_device_name,
            MAX(pd.rated_power_kw) AS provider_power_kw,
            MAX(pd.model) AS provider_model,
            MAX(pd.enabled) AS provider_enabled,
            SUM(iad.valid_slots) AS valid_slots,
            SUM(iad.available_slots) AS available_slots,
            SUM(iad.unavailable_slots) AS unavailable_slots
        FROM inverter_availability_daily iad
        JOIN assets a ON a.id = iad.asset_id
        LEFT JOIN provider_devices pd
          ON pd.provider = iad.provider AND pd.external_device_id = iad.inverter_id
        WHERE {' AND '.join(conditions)}
        GROUP BY iad.asset_id, a.project_name, iad.inverter_id, iad.inverter_name, iad.inverter_power_kw
        HAVING SUM(iad.valid_slots) > 0
        """,
        params,
    ).fetchall()
    inverter_rows: list[dict[str, Any]] = []
    for row in rows:
        item = dict(row)
        device_name = item.get("provider_device_name") or item.get("inverter_name")
        if item.get("provider_enabled") == 0 or is_removed_inverter_name(device_name):
            continue
        item["inverter_name"] = device_name
        item["inverter_power_kw"] = (
            parse_float_value(item.get("inverter_power_kw"))
            or parse_float_value(item.get("provider_power_kw"))
            or infer_inverter_power_from_model(item.get("provider_model"))
        )
        item["availability_pct"] = round(item["available_slots"] / item["valid_slots"] * 100, 2)
        inverter_rows.append(item)

    plants_by_id: dict[int, dict[str, Any]] = {}
    for row in inverter_rows:
        plant = plants_by_id.setdefault(
            int(row["asset_id"]),
            {"asset_id": row["asset_id"], "project_name": row["project_name"], "inverters": []},
        )
        plant["inverters"].append(row)
    plant_rows: list[dict[str, Any]] = []
    for plant in plants_by_id.values():
        plant_rows.append(
            {
                "asset_id": plant["asset_id"],
                "project_name": plant["project_name"],
                "inverter_count": len(plant["inverters"]),
                "availability_pct": calculate_weighted_plant_availability(plant["inverters"]),
            }
        )
    plant_rows.sort(key=lambda row: (row["availability_pct"] is None, row["availability_pct"] or 0, row["project_name"]))
    inverter_rows.sort(key=lambda row: (row["availability_pct"], row["project_name"], row["inverter_name"] or row["inverter_id"]))
    for rank, row in enumerate(plant_rows, start=1):
        row["rank"] = rank
    for rank, row in enumerate(inverter_rows, start=1):
        row["rank"] = rank
    percentages = [float(row["availability_pct"]) for row in plant_rows if row["availability_pct"] is not None]
    return {
        "average_pct": round(sum(percentages) / len(percentages), 2) if percentages else None,
        "plants": plant_rows,
        "inverters": inverter_rows,
        "worst_plant": plant_rows[0] if plant_rows else None,
        "low_availability_count": sum(
            1 for row in inverter_rows if float(row["availability_pct"]) < LOW_INVERTER_AVAILABILITY_PCT
        ),
    }


def get_monthly_wat_report_data(
    conn: sqlite3.Connection,
    from_date: date,
    to_date: date,
    asset_id: int | None = None,
) -> dict[str, Any]:
    if from_date > to_date:
        raise ValueError("O intervalo WAT e invalido.")

    provider = INTEGRATION_PROVIDER_FUSIONSOLAR
    from_iso = from_date.isoformat()
    to_iso = to_date.isoformat()
    sample_from = datetime.combine(from_date, datetime.min.time()).isoformat(timespec="seconds")
    sample_to = datetime.combine(to_date + timedelta(days=1), datetime.min.time()).isoformat(timespec="seconds")
    expected_days = (to_date - from_date).days + 1

    asset_params: list[Any] = []
    if asset_id is not None:
        asset_filter = "a.id = ?"
        asset_params.append(asset_id)
    else:
        asset_filter = """
            EXISTS (
                SELECT 1 FROM provider_devices pd
                WHERE pd.asset_id = a.id AND pd.provider = ? AND pd.enabled = 1 AND pd.dev_type_id IN (1, 38)
            )
            OR EXISTS (
                SELECT 1 FROM inverter_availability_daily iad
                WHERE iad.asset_id = a.id AND iad.provider = ? AND iad.availability_date BETWEEN ? AND ?
            )
            OR EXISTS (
                SELECT 1 FROM plant_availability_daily pad
                WHERE pad.asset_id = a.id AND pad.provider = ? AND pad.availability_date BETWEEN ? AND ?
            )
            OR EXISTS (
                SELECT 1 FROM inverter_power_samples ips
                WHERE ips.asset_id = a.id AND ips.provider = ? AND ips.sample_time >= ? AND ips.sample_time < ?
            )
        """
        asset_params.extend(
            [provider, provider, from_iso, to_iso, provider, from_iso, to_iso, provider, sample_from, sample_to]
        )
    assets = conn.execute(
        f"SELECT a.id, a.project_name FROM assets a WHERE {asset_filter} ORDER BY a.project_name COLLATE NOCASE",
        asset_params,
    ).fetchall()

    plants: list[dict[str, Any]] = []
    for asset in assets:
        current_asset_id = int(asset["id"])
        configured_devices = query_all(
            conn,
            """
            SELECT external_device_id, device_name
            FROM provider_devices
            WHERE asset_id = ? AND provider = ? AND enabled = 1 AND dev_type_id IN (1, 38)
            """,
            (current_asset_id, provider),
        )
        configured_inverter_count = sum(
            1 for row in configured_devices if not is_removed_inverter_name(row["device_name"])
        )
        inverter_rows = query_all(
            conn,
            """
            SELECT
                iad.inverter_id,
                COALESCE(MAX(pd.device_name), MAX(iad.inverter_name), iad.inverter_id) AS inverter_name,
                MAX(iad.inverter_power_kw) AS stored_power_kw,
                MAX(pd.rated_power_kw) AS provider_power_kw,
                MAX(pd.model) AS provider_model,
                MAX(pd.enabled) AS provider_enabled,
                COUNT(DISTINCT iad.availability_date) AS data_days,
                SUM(iad.valid_slots) AS valid_slots,
                SUM(iad.available_slots) AS available_slots,
                SUM(iad.unavailable_slots) AS unavailable_slots
            FROM inverter_availability_daily iad
            LEFT JOIN provider_devices pd
              ON pd.provider = iad.provider AND pd.external_device_id = iad.inverter_id
            WHERE iad.asset_id = ? AND iad.provider = ? AND iad.availability_date BETWEEN ? AND ?
            GROUP BY iad.inverter_id
            ORDER BY inverter_name COLLATE NOCASE, iad.inverter_id
            """,
            (current_asset_id, provider, from_iso, to_iso),
        )
        report_inverters: list[dict[str, Any]] = []
        for stored_row in inverter_rows:
            row = dict(stored_row)
            if row.get("provider_enabled") == 0 or is_removed_inverter_name(row.get("inverter_name")):
                continue
            valid_slots = int(row.get("valid_slots") or 0)
            available_slots = int(row.get("available_slots") or 0)
            report_inverters.append(
                {
                    "inverter_id": str(row["inverter_id"]),
                    "inverter_name": str(row.get("inverter_name") or row["inverter_id"]),
                    "inverter_power_kw": (
                        parse_float_value(row.get("stored_power_kw"))
                        or parse_float_value(row.get("provider_power_kw"))
                        or infer_inverter_power_from_model(row.get("provider_model"))
                    ),
                    "data_days": int(row.get("data_days") or 0),
                    "valid_slots": valid_slots,
                    "available_slots": available_slots,
                    "unavailable_slots": int(row.get("unavailable_slots") or 0),
                    "availability_pct": round(available_slots / valid_slots * 100, 2) if valid_slots else None,
                }
            )

        plant_days = conn.execute(
            """
            SELECT availability_date, valid_slots, weighted_availability_pct, inverter_count
            FROM plant_availability_daily
            WHERE asset_id = ? AND provider = ? AND availability_date BETWEEN ? AND ?
            ORDER BY availability_date
            """,
            (current_asset_id, provider, from_iso, to_iso),
        ).fetchall()
        sample_summary = conn.execute(
            """
            SELECT COUNT(*) AS sample_count, COUNT(DISTINCT substr(sample_time, 1, 10)) AS sample_days
            FROM inverter_power_samples
            WHERE asset_id = ? AND provider = ? AND sample_time >= ? AND sample_time < ?
            """,
            (current_asset_id, provider, sample_from, sample_to),
        ).fetchone()

        sample_count = int(sample_summary["sample_count"] or 0)
        sample_days = int(sample_summary["sample_days"] or 0)
        plant_day_count = len(plant_days)
        has_any_data = bool(report_inverters or plant_day_count or sample_count)
        if not has_any_data:
            data_status = "sem dados"
        else:
            daily_counts = conn.execute(
                """
                SELECT availability_date, COUNT(*) AS inverter_count,
                       SUM(CASE WHEN valid_slots > 0 AND availability_pct IS NOT NULL THEN 1 ELSE 0 END) AS valid_inverters
                FROM inverter_availability_daily
                WHERE asset_id = ? AND provider = ? AND availability_date BETWEEN ? AND ?
                GROUP BY availability_date
                """,
                (current_asset_id, provider, from_iso, to_iso),
            ).fetchall()
            expected_inverter_count = configured_inverter_count or max(
                (int(row["inverter_count"] or 0) for row in daily_counts),
                default=0,
            )
            complete_inverter_days = (
                len(daily_counts) == expected_days
                and expected_inverter_count > 0
                and all(
                    int(row["inverter_count"] or 0) == expected_inverter_count
                    and int(row["valid_inverters"] or 0) == expected_inverter_count
                    for row in daily_counts
                )
            )
            complete_plant_days = (
                plant_day_count == expected_days
                and all(
                    int(row["valid_slots"] or 0) > 0
                    and row["weighted_availability_pct"] is not None
                    and int(row["inverter_count"] or 0) == expected_inverter_count
                    for row in plant_days
                )
            )
            data_status = (
                "ok"
                if complete_inverter_days and complete_plant_days and sample_days == expected_days
                else "parcial"
            )

        ranked_inverters = [row for row in report_inverters if row["availability_pct"] is not None]
        ranked_inverters.sort(
            key=lambda row: (float(row["availability_pct"]), row["inverter_name"].lower(), row["inverter_id"])
        )
        worst_inverter = ranked_inverters[0] if ranked_inverters else None
        weighted_wat = calculate_weighted_plant_availability(report_inverters)
        plants.append(
            {
                "asset_id": current_asset_id,
                "project_name": str(asset["project_name"]),
                "weighted_wat_pct": weighted_wat,
                "inverter_count": len(report_inverters) or configured_inverter_count,
                "inverters_below_90_count": sum(
                    1
                    for row in report_inverters
                    if row["availability_pct"] is not None
                    and float(row["availability_pct"]) < LOW_INVERTER_AVAILABILITY_PCT
                ),
                "worst_inverter": worst_inverter["inverter_name"] if worst_inverter else None,
                "worst_inverter_id": worst_inverter["inverter_id"] if worst_inverter else None,
                "worst_inverter_wat_pct": worst_inverter["availability_pct"] if worst_inverter else None,
                "valid_slots": sum(int(row["valid_slots"]) for row in report_inverters),
                "unavailable_slots": sum(int(row["unavailable_slots"]) for row in report_inverters),
                "data_status": data_status,
            }
        )

    return {
        "from_date": from_iso,
        "to_date": to_iso,
        "plants": plants,
    }


def get_daily_wat_report_data(conn: sqlite3.Connection, target_date: date) -> dict[str, Any]:
    report = get_monthly_wat_report_data(conn, target_date, target_date)
    return {
        "target_date": target_date.isoformat(),
        "plants": report["plants"],
    }


def get_inverter_availability_chart_report(
    conn: sqlite3.Connection,
    asset_id: int,
    from_date: date,
    to_date: date,
) -> dict[str, Any] | None:
    asset = conn.execute(
        "SELECT id, project_name FROM assets WHERE id = ?",
        (asset_id,),
    ).fetchone()
    if asset is None:
        return None
    device_rows = conn.execute(
        """
        SELECT
            external_device_id AS inverter_id,
            device_name AS inverter_name,
            rated_power_kw AS inverter_power_kw,
            model
        FROM provider_devices
        WHERE asset_id = ? AND provider = ? AND enabled = 1 AND dev_type_id IN (1, 38)
        ORDER BY device_name COLLATE NOCASE, external_device_id
        """,
        (asset_id, INTEGRATION_PROVIDER_FUSIONSOLAR),
    ).fetchall()
    daily_rows = conn.execute(
        """
        SELECT inverter_id, inverter_name, inverter_power_kw, availability_pct
        FROM inverter_availability_daily
        WHERE asset_id = ? AND provider = ? AND availability_date BETWEEN ? AND ?
        ORDER BY availability_date
        """,
        (asset_id, INTEGRATION_PROVIDER_FUSIONSOLAR, from_date.isoformat(), to_date.isoformat()),
    ).fetchall()
    inverters_by_id: dict[str, dict[str, Any]] = {}
    for row in device_rows:
        inverter_id = str(row["inverter_id"])
        if is_removed_inverter_name(row["inverter_name"]):
            continue
        inverters_by_id[inverter_id] = {
            "inverter_id": inverter_id,
            "inverter_name": row["inverter_name"],
            "inverter_power_kw": parse_float_value(row["inverter_power_kw"])
            or infer_inverter_power_from_model(row["model"]),
            "availability_values": [],
        }
    for row in daily_rows:
        inverter_id = str(row["inverter_id"])
        if is_removed_inverter_name(row["inverter_name"]):
            continue
        inverter = inverters_by_id.setdefault(
            inverter_id,
            {
                "inverter_id": inverter_id,
                "inverter_name": row["inverter_name"],
                "inverter_power_kw": row["inverter_power_kw"],
                "availability_values": [],
            },
        )
        if row["availability_pct"] is not None:
            inverter["availability_values"].append(float(row["availability_pct"]))

    period_start = datetime.combine(from_date, datetime.min.time())
    period_end = datetime.combine(to_date + timedelta(days=1), datetime.min.time())
    sample_rows = conn.execute(
        """
        SELECT inverter_id, sample_time, active_power_kw
        FROM inverter_power_samples
        WHERE asset_id = ? AND provider = ? AND sample_time >= ? AND sample_time < ?
        ORDER BY inverter_id, sample_time
        """,
        (
            asset_id,
            INTEGRATION_PROVIDER_FUSIONSOLAR,
            period_start.isoformat(timespec="seconds"),
            period_end.isoformat(timespec="seconds"),
        ),
    ).fetchall()
    samples_by_inverter: dict[str, list[tuple[datetime, float]]] = {}
    for row in sample_rows:
        sample_time = parse_datetime_value(row["sample_time"])
        active_power = parse_float_value(row["active_power_kw"])
        if sample_time is None or active_power is None:
            continue
        samples_by_inverter.setdefault(str(row["inverter_id"]), []).append((sample_time, max(active_power, 0.0)))

    inverters = list(inverters_by_id.values())
    for inverter in inverters:
        percentages = inverter.pop("availability_values")
        inverter["average_pct"] = round(sum(percentages) / len(percentages), 2) if percentages else None
        inverter["chart"] = build_inverter_power_chart(
            samples_by_inverter.get(inverter["inverter_id"], []),
            period_start,
            period_end,
            parse_float_value(inverter["inverter_power_kw"]),
        )
    inverters.sort(key=lambda row: (row["average_pct"] is None, row["average_pct"] or 0, row["inverter_name"] or row["inverter_id"]))
    return {
        "asset_id": int(asset["id"]),
        "project_name": asset["project_name"],
        "inverters": inverters,
    }


def build_inverter_power_chart(
    samples: list[tuple[datetime, float]],
    period_start: datetime,
    period_end: datetime,
    rated_power_kw: float | None,
) -> dict[str, Any]:
    chart_width = 760.0
    chart_height = 260.0
    plot_left = 52.0
    plot_right = 16.0
    plot_top = 16.0
    plot_bottom = 36.0
    plot_width = chart_width - plot_left - plot_right
    plot_height = chart_height - plot_top - plot_bottom
    total_seconds = max((period_end - period_start).total_seconds(), 1.0)
    sorted_samples = sorted(samples, key=lambda item: item[0])
    if len(sorted_samples) > 600:
        step = max(len(sorted_samples) // 600, 1)
        reduced = sorted_samples[::step]
        if reduced[-1] != sorted_samples[-1]:
            reduced.append(sorted_samples[-1])
        sorted_samples = reduced
    observed_max = max((power for _, power in sorted_samples), default=0.0)
    y_max = max(observed_max, rated_power_kw or 0.0, 1.0)
    if y_max > 10:
        y_max = float((int(y_max + 4.999) // 5) * 5)
    else:
        y_max = float(int(y_max + 0.999))
    coordinates: list[tuple[float, float]] = []
    for sample_time, power in sorted_samples:
        elapsed = min(max((sample_time - period_start).total_seconds(), 0.0), total_seconds)
        x = plot_left + elapsed / total_seconds * plot_width
        y = plot_top + (1 - min(power / y_max, 1.0)) * plot_height
        coordinates.append((round(x, 2), round(y, 2)))
    line_path = " ".join(
        ("M" if index == 0 else "L") + f" {x} {y}"
        for index, (x, y) in enumerate(coordinates)
    )
    baseline = plot_top + plot_height
    area_path = ""
    if coordinates:
        area_path = (
            f"M {coordinates[0][0]} {baseline} "
            + " ".join(f"L {x} {y}" for x, y in coordinates)
            + f" L {coordinates[-1][0]} {baseline} Z"
        )
    x_ticks = []
    for index in range(5):
        ratio = index / 4
        tick_time = period_start + timedelta(seconds=total_seconds * ratio)
        x_ticks.append(
            {
                "x": round(plot_left + plot_width * ratio, 2),
                "label": tick_time.strftime("%H:%M") if (period_end - period_start).days <= 1 else tick_time.strftime("%d/%m"),
            }
        )
    y_ticks = [
        {
            "y": round(plot_top + plot_height * index / 4, 2),
            "label": round(y_max * (1 - index / 4), 1),
        }
        for index in range(5)
    ]
    return {
        "width": int(chart_width),
        "height": int(chart_height),
        "plot_left": plot_left,
        "plot_right": chart_width - plot_right,
        "plot_top": plot_top,
        "plot_bottom": baseline,
        "line_path": line_path,
        "area_path": area_path,
        "x_ticks": x_ticks,
        "y_ticks": y_ticks,
        "sample_count": len(samples),
        "max_power_kw": round(y_max, 1),
    }


def parse_datetime_value(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        timestamp = float(value)
        if timestamp > 10_000_000_000:
            timestamp /= 1000
        try:
            return datetime.fromtimestamp(timestamp)
        except (OSError, OverflowError, ValueError):
            return None
    raw = str(value).strip()
    try:
        timestamp = float(raw)
        if timestamp > 10_000_000_000:
            timestamp /= 1000
        return datetime.fromtimestamp(timestamp)
    except (OSError, OverflowError, ValueError):
        pass
    raw = raw.replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(raw)
        return parsed.replace(tzinfo=None) if parsed.tzinfo else parsed
    except ValueError:
        return None


def calculate_asset_availability(device_rows: list[dict[str, Any]]) -> dict[str, Any]:
    enabled_rows = [row for row in device_rows if int(row.get("enabled", 1) or 0) == 1]
    total = len(enabled_rows)
    available = sum(1 for row in enabled_rows if row.get("availability_status") == "available")
    unavailable = sum(
        1 for row in enabled_rows if row.get("availability_status") in {"unavailable", "no_communication"}
    )
    no_communication = sum(1 for row in enabled_rows if row.get("availability_status") == "no_communication")
    recent = sum(1 for row in enabled_rows if row.get("communication_status") == "recent")
    rated_values = [row.get("rated_power_kw") for row in enabled_rows]
    known_rated = [float(value) for value in rated_values if value is not None]
    affected_rows = [
        row for row in enabled_rows if row.get("availability_status") in {"unavailable", "no_communication"}
    ]
    affected_power_kw = (
        sum(float(row["rated_power_kw"]) for row in affected_rows)
        if affected_rows and all(row.get("rated_power_kw") is not None for row in affected_rows)
        else None
    )
    capacity_availability_pct = None
    if total and len(known_rated) == total and sum(known_rated) > 0:
        capacity_availability_pct = round(
            sum(float(row["rated_power_kw"]) for row in enabled_rows if row.get("availability_status") == "available")
            / sum(known_rated)
            * 100,
            2,
        )
    return {
        "inverter_availability_pct": round(available / total * 100, 2) if total else None,
        "capacity_availability_pct": capacity_availability_pct,
        "communication_availability_pct": round(recent / total * 100, 2) if total else None,
        "available_inverters": available,
        "total_inverters": total,
        "unavailable_inverters": unavailable,
        "no_communication_devices": no_communication,
        "affected_power_kw": affected_power_kw,
        "available_strings": sum(int(row.get("available_strings") or 0) for row in enabled_rows),
        "total_strings": sum(int(row.get("total_strings") or 0) for row in enabled_rows),
        "unavailable_strings": sum(int(row.get("unavailable_strings") or 0) for row in enabled_rows),
        "string_availability_pct": (
            round(
                sum(int(row.get("available_strings") or 0) for row in enabled_rows)
                / sum(int(row.get("total_strings") or 0) for row in enabled_rows)
                * 100,
                2,
            )
            if sum(int(row.get("total_strings") or 0) for row in enabled_rows)
            else None
        ),
    }


def format_fusionsolar_alarm_time(value: Any) -> str:
    if value in (None, ""):
        return ""
    raw = str(value).strip()
    if raw.isdigit():
        timestamp = int(raw)
        if timestamp > 10_000_000_000:
            timestamp = int(timestamp / 1000)
        try:
            return datetime.fromtimestamp(timestamp).isoformat(timespec="seconds")
        except (ValueError, OSError):
            return raw
    return raw


def normalize_fusionsolar_alarm(row: dict[str, Any]) -> dict[str, str]:
    alarm_name = first_non_empty(
        row,
        [
            "alarmName",
            "alarm_name",
            "name",
            "alarmType",
            "alarmTypeName",
            "faultName",
            "eventName",
            "cause",
        ],
    )
    device_name = first_non_empty(
        row,
        [
            "devName",
            "deviceName",
            "device_name",
            "equipmentName",
            "inverterName",
            "devAlias",
            "devTypeName",
            "devDn",
            "deviceDn",
        ],
    )
    return {
        "alarm_name": alarm_name or first_non_empty(row, ["alarmId", "alarm_id", "id"]) or "Alarme ativo",
        "device_name": device_name or "Aparelho nao identificado",
        "severity": first_non_empty(row, ["lev", "level", "severity", "alarmLevel"]),
        "raised_at": format_fusionsolar_alarm_time(first_non_empty(row, ["raiseTime", "startTime", "occurTime", "happenTime"])),
        "status": first_non_empty(row, ["status", "alarmStatus", "state"]),
    }


def summarize_fusionsolar_alarms(alarms: list[dict[str, Any]], limit: int = 3) -> dict[str, Any]:
    normalized = [normalize_fusionsolar_alarm(alarm) for alarm in alarms if isinstance(alarm, dict)]
    primary = normalized[0] if normalized else {}
    summary_parts = []
    for alarm in normalized[:limit]:
        label = alarm["alarm_name"]
        if alarm["device_name"]:
            label = f"{label} @ {alarm['device_name']}"
        if alarm["severity"]:
            label = f"{label} (sev. {alarm['severity']})"
        summary_parts.append(label)
    if len(normalized) > limit:
        summary_parts.append(f"+{len(normalized) - limit} alarmes")
    return {
        "primary_alarm_name": primary.get("alarm_name", ""),
        "primary_alarm_device": primary.get("device_name", ""),
        "primary_alarm_severity": primary.get("severity", ""),
        "primary_alarm_raised_at": primary.get("raised_at", ""),
        "alarm_summary": "; ".join(summary_parts),
        "normalized_alarms": normalized,
    }


def fusionsolar_alarm_severity_rank(value: Any) -> int | None:
    normalized = normalize_name(str(value or ""))
    if not normalized:
        return None
    if normalized in {"1", "1.0", "critical", "critica", "critico"}:
        return 1
    if normalized in {"2", "2.0", "major", "alta", "maior"}:
        return 2
    if normalized in {"3", "3.0", "minor", "menor", "baixa"}:
        return 3
    if normalized in {"4", "4.0", "warning", "aviso", "alerta"}:
        return 4
    return None


def derive_fusionsolar_monitoring_status(health_raw: Any, alarms: list[dict[str, Any]]) -> str:
    status = map_fusionsolar_status(health_raw)
    if status == "Desconectada":
        return status

    alarm_ranks = [
        rank
        for alarm in alarms
        if isinstance(alarm, dict)
        for rank in [fusionsolar_alarm_severity_rank(first_non_empty(alarm, ["lev", "level", "severity", "alarmLevel"]))]
        if rank is not None
    ]
    if any(rank <= 2 for rank in alarm_ranks):
        return "Erro"
    if alarm_ranks:
        return "Alerta"
    return status


def normalize_fusionsolar_plant_row(
    station_row: dict[str, Any],
    realtime_row: dict[str, Any] | None = None,
    alarms: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    external_id = str(station_row.get("plantCode") or station_row.get("stationCode") or "").strip()
    external_name = str(station_row.get("plantName") or station_row.get("stationName") or "").strip()
    if not external_name:
        raise ValueError("A resposta FusionSolar nao trouxe nome de central numa das linhas.")

    data_item_map = (realtime_row or {}).get("dataItemMap") or {}
    health_raw = data_item_map.get("real_health_state")
    raw_status = describe_fusionsolar_health_state(health_raw)

    active_alarms = alarms or []
    status = derive_fusionsolar_monitoring_status(health_raw, active_alarms)
    alarm_summary = summarize_fusionsolar_alarms(active_alarms)
    alarm_levels = sorted({str(item.get("lev")) for item in active_alarms if item.get("lev") is not None})
    notes_parts = [f"health_state={raw_status}"]
    if active_alarms:
        notes_parts.append(f"active_alarms={len(active_alarms)}")
        if alarm_levels:
            notes_parts.append(f"levels={','.join(alarm_levels)}")
        if alarm_summary["alarm_summary"]:
            notes_parts.append(f"alarm_details={alarm_summary['alarm_summary']}")

    return {
        "external_id": external_id,
        "external_name": external_name,
        "status": status,
        "raw_status": raw_status,
        "health_state": raw_status,
        "alarm_count": len(active_alarms),
        "alarm_levels": ",".join(alarm_levels),
        "primary_alarm_name": alarm_summary["primary_alarm_name"],
        "primary_alarm_device": alarm_summary["primary_alarm_device"],
        "primary_alarm_severity": alarm_summary["primary_alarm_severity"],
        "primary_alarm_raised_at": alarm_summary["primary_alarm_raised_at"],
        "alarm_summary": alarm_summary["alarm_summary"],
        "notes": "; ".join(notes_parts),
        "payload": {
            "station": station_row,
            "realtime": realtime_row or {},
            "alarms": active_alarms,
            "normalized_alarms": alarm_summary["normalized_alarms"],
        },
    }


def find_suggested_asset_id(conn: sqlite3.Connection, external_name: str) -> int | None:
    normalized_name = normalize_name(external_name)
    exact = find_asset_id(conn, external_name)
    if exact:
        return exact
    candidate = conn.execute(
        """
        SELECT id
        FROM assets
        WHERE REPLACE(LOWER(project_name), ' ', '') LIKE ?
        ORDER BY project_name COLLATE NOCASE
        LIMIT 1
        """,
        (f"%{normalized_name.replace(' ', '')}%",),
    ).fetchone()
    return int(candidate["id"]) if candidate else None


def parse_provider_payload_data(payload: dict[str, Any]) -> Any:
    data = payload.get("data")
    if isinstance(data, str):
        try:
            return json.loads(data)
        except json.JSONDecodeError:
            return data
    return data


def get_sigenergy_endpoint_config(config: sqlite3.Row | dict[str, Any]) -> dict[str, str]:
    config_map = dict(config)
    return {
        "base_url": str(config_map.get("base_url") or DEFAULT_SIGENERGY_BASE_URL).strip() or DEFAULT_SIGENERGY_BASE_URL,
        "login_endpoint": str(config_map.get("login_endpoint") or DEFAULT_SIGENERGY_AUTH_ENDPOINT).strip() or DEFAULT_SIGENERGY_AUTH_ENDPOINT,
        "systems_endpoint": str(config_map.get("plants_endpoint") or DEFAULT_SIGENERGY_SYSTEMS_ENDPOINT).strip() or DEFAULT_SIGENERGY_SYSTEMS_ENDPOINT,
        "real_time_endpoint": str(config_map.get("real_time_endpoint") or DEFAULT_SIGENERGY_REALTIME_ENDPOINT).strip() or DEFAULT_SIGENERGY_REALTIME_ENDPOINT,
        "energy_flow_endpoint": str(config_map.get("alarms_endpoint") or DEFAULT_SIGENERGY_ENERGY_FLOW_ENDPOINT).strip() or DEFAULT_SIGENERGY_ENERGY_FLOW_ENDPOINT,
        "region": str(config_map.get("region") or DEFAULT_SIGENERGY_REGION).strip() or DEFAULT_SIGENERGY_REGION,
    }


def sigenergy_configured_system_ids(config: sqlite3.Row | dict[str, Any]) -> list[str]:
    raw_value = str(dict(config).get("system_ids") or "").strip()
    return [item.strip() for item in re.split(r"[,;\s]+", raw_value) if item.strip()]


def get_sigenergy_token(config: sqlite3.Row | dict[str, Any], *, force_login: bool = False) -> str:
    endpoints = get_sigenergy_endpoint_config(config)
    app_key = str(config["username"] or "").strip()
    app_secret = str(config["password"] or "").strip()
    if not app_key or not app_secret:
        raise ValueError("Preenche App Key e App Secret da Sigenergy.")
    if not endpoints["base_url"]:
        raise ValueError("Falta a Base URL da Sigenergy.")

    cache_key = f"{endpoints['base_url']}|{app_key}|{endpoints['region']}"
    now = datetime.now()
    with SIGENERGY_TOKEN_LOCK:
        cached = SIGENERGY_TOKEN_CACHE.get(cache_key)
        if cached and not force_login and cached["expires_at"] > now:
            return str(cached["access_token"])

        auth_key = base64.b64encode(f"{app_key}:{app_secret}".encode("utf-8")).decode("ascii")
        response = requests.post(
            build_provider_url(endpoints["base_url"], endpoints["login_endpoint"]),
            json={"key": auth_key},
            headers={"Content-Type": "application/json", "Accept": "application/json"},
            timeout=30,
        )
        response.raise_for_status()
        payload = response.json()
        if int(payload.get("code") or 0) != 0:
            raise ValueError(f"{payload.get('msg') or 'Login Sigenergy falhou.'} (code={payload.get('code')})")
        token_data = parse_provider_payload_data(payload)
        if not isinstance(token_data, dict):
            raise ValueError("A resposta Sigenergy de login nao trouxe data JSON valido.")
        access_token = str(token_data.get("accessToken") or token_data.get("access_token") or "").strip()
        if not access_token:
            raise ValueError("A resposta Sigenergy de login nao trouxe accessToken.")
        expires_in = int(float(str(token_data.get("expiresIn") or token_data.get("expires_in") or 43199)))
        SIGENERGY_TOKEN_CACHE[cache_key] = {
            "access_token": access_token,
            "expires_at": now + timedelta(seconds=max(expires_in - 300, 300)),
        }
        return access_token


def sigenergy_headers(token: str, region: str) -> dict[str, str]:
    return {
        "Authorization": f"Bearer {token}",
        "sigen-region": region,
        "Accept": "application/json",
        "Content-Type": "application/json",
    }


def fetch_sigenergy_json(
    method: str,
    *,
    base_url: str,
    endpoint: str,
    token: str,
    region: str,
    json_payload: dict[str, Any] | None = None,
) -> dict[str, Any]:
    response = requests.request(
        method,
        build_provider_url(base_url, endpoint),
        headers=sigenergy_headers(token, region),
        json=json_payload,
        timeout=30,
    )
    response.raise_for_status()
    payload = response.json()
    code = payload.get("code")
    if code not in (None, 0, "0"):
        raise ValueError(f"{payload.get('msg') or 'Pedido Sigenergy falhou.'} (code={code})")
    return payload


def normalize_sigenergy_system_rows(data: Any) -> list[dict[str, Any]]:
    if isinstance(data, list):
        return [row for row in data if isinstance(row, dict)]
    if isinstance(data, dict):
        for key in ("list", "records", "systems", "systemList", "rows"):
            rows = data.get(key)
            if isinstance(rows, list):
                return [row for row in rows if isinstance(row, dict)]
        if any(key in data for key in ("systemId", "id", "systemName", "name")):
            return [data]
    return []


def fetch_sigenergy_systems(config: sqlite3.Row | dict[str, Any], token: str) -> list[dict[str, Any]]:
    configured_ids = sigenergy_configured_system_ids(config)
    if configured_ids:
        return [{"systemId": system_id, "systemName": system_id} for system_id in configured_ids]
    endpoints = get_sigenergy_endpoint_config(config)
    try:
        payload = fetch_sigenergy_json(
            "GET",
            base_url=endpoints["base_url"],
            endpoint=endpoints["systems_endpoint"],
            token=token,
            region=endpoints["region"],
        )
    except requests.HTTPError as exc:
        if exc.response is not None and exc.response.status_code == 404:
            raise ValueError(
                "A Sigenergy autenticou, mas nao disponibilizou a lista automatica de sistemas. "
                "Preenche SIGENERGY_SYSTEM_IDS no .env com os systemId da app mySigen."
            ) from exc
        raise
    rows = normalize_sigenergy_system_rows(parse_provider_payload_data(payload))
    if not rows:
        raise ValueError(
            "A API Sigenergy respondeu com sucesso, mas sem sistemas. "
            "Confirma no developer portal se a App Key tem sistemas onboarded/autorizados."
        )
    return rows


def fetch_sigenergy_realtime_data(config: sqlite3.Row | dict[str, Any], token: str, system_id: str) -> dict[str, Any]:
    endpoints = get_sigenergy_endpoint_config(config)
    payload = fetch_sigenergy_json(
        "POST",
        base_url=endpoints["base_url"],
        endpoint=endpoints["real_time_endpoint"],
        token=token,
        region=endpoints["region"],
        json_payload={"systemId": system_id},
    )
    data = parse_provider_payload_data(payload)
    return data if isinstance(data, dict) else {"raw_data": data}


def fetch_sigenergy_energy_flow(config: sqlite3.Row | dict[str, Any], token: str, system_id: str) -> dict[str, Any]:
    endpoints = get_sigenergy_endpoint_config(config)
    endpoint = endpoints["energy_flow_endpoint"].replace("{system_id}", system_id).replace("{systemId}", system_id)
    payload = fetch_sigenergy_json(
        "GET",
        base_url=endpoints["base_url"],
        endpoint=endpoint,
        token=token,
        region=endpoints["region"],
    )
    data = parse_provider_payload_data(payload)
    return data if isinstance(data, dict) else {"raw_data": data}


def map_sigenergy_status(raw_status: Any, energy_flow: dict[str, Any] | None = None) -> str:
    normalized = normalize_name(str(raw_status or ""))
    if normalized in {"fault", "error", "alarm", "shutdown"}:
        return "Erro"
    if normalized in {"disconnected", "offline", "communication lost"}:
        return "Desconectada"
    if normalized in {"running", "normal", "online", "standby", "grid connected", "grid-connected"}:
        return "Operacional"
    if energy_flow and any(energy_flow.get(key) not in (None, "") for key in ("pvPower", "batterySoc", "loadPower", "gridPower")):
        return "Operacional"
    return normalize_status(str(raw_status or "Operacional"))


def normalize_sigenergy_system_row(
    system_row: dict[str, Any],
    realtime_row: dict[str, Any] | None = None,
    energy_flow: dict[str, Any] | None = None,
) -> dict[str, Any]:
    external_id = first_non_empty(system_row, ["systemId", "id", "stationId", "plantId"])
    if not external_id:
        raise ValueError("A resposta Sigenergy nao trouxe systemId numa das linhas.")
    external_name = first_non_empty(system_row, ["systemName", "name", "stationName", "plantName"]) or external_id
    realtime = realtime_row or {}
    flow = energy_flow or {}
    raw_status = first_non_empty(realtime, ["systemStatus", "status", "runningStatus", "state"]) or first_non_empty(
        system_row,
        ["systemStatus", "status", "runningStatus", "state"],
    )
    status = map_sigenergy_status(raw_status, flow)
    notes_parts = [f"system_status={raw_status or 'unknown'}"]
    for key in ("pvPower", "gridPower", "batteryPower", "batterySoc", "loadPower"):
        if key in flow and flow[key] not in (None, ""):
            notes_parts.append(f"{key}={flow[key]}")
    return {
        "external_id": external_id,
        "external_name": external_name,
        "status": status,
        "raw_status": raw_status or "unknown",
        "notes": "; ".join(notes_parts),
        "payload": {
            "system": system_row,
            "realtime": realtime,
            "energy_flow": flow,
        },
    }


def run_sigenergy_check(conn: sqlite3.Connection, provider: str, dry_run: bool = False) -> dict[str, Any]:
    config = get_integration_config(conn, provider)
    if config is None:
        raise ValueError("Configuracao Sigenergy nao encontrada.")
    endpoints = get_sigenergy_endpoint_config(config)
    token = get_sigenergy_token(config)
    systems = fetch_sigenergy_systems(config, token)
    normalized_rows: list[dict[str, Any]] = []
    realtime_count = 0
    energy_flow_count = 0
    energy_flow_errors: list[str] = []
    for system_row in systems:
        system_id = first_non_empty(system_row, ["systemId", "id", "stationId", "plantId"])
        if not system_id:
            continue
        realtime = fetch_sigenergy_realtime_data(config, token, system_id)
        realtime_count += 1
        energy_flow: dict[str, Any] = {}
        try:
            energy_flow = fetch_sigenergy_energy_flow(config, token, system_id)
            energy_flow_count += 1
        except Exception as exc:
            energy_flow_errors.append(f"{system_id}: {exc}")
        normalized_rows.append(normalize_sigenergy_system_row(system_row, realtime, energy_flow))
        time.sleep(0.2)

    if not dry_run:
        conn.execute(
            """
            UPDATE integration_configs
            SET last_sync_status = ?, last_error = ?, updated_at = ?
            WHERE provider = ?
            """,
            ("success", "", datetime.now().isoformat(timespec="seconds"), provider),
        )
        conn.commit()
    return {
        "rows": normalized_rows,
        "station_count": len(systems),
        "realtime_count": realtime_count,
        "alarm_count": 0,
        "alarm_error": "; ".join(energy_flow_errors),
        "energy_flow_count": energy_flow_count,
        "base_url": endpoints["base_url"],
    }


def run_provider_check(conn: sqlite3.Connection, provider: str, dry_run: bool = False) -> dict[str, Any]:
    if provider == INTEGRATION_PROVIDER_SIGENERGY:
        return run_sigenergy_check(conn, provider, dry_run=dry_run)
    return run_fusionsolar_check(conn, provider, dry_run=dry_run)


def run_fusionsolar_check(conn: sqlite3.Connection, provider: str, dry_run: bool = False) -> dict[str, Any]:
    config = get_integration_config(conn, provider)
    if config is None:
        raise ValueError("Configuracao FusionSolar nao encontrada.")
    endpoints = get_fusionsolar_endpoint_config(config)
    if not endpoints["base_url"]:
        raise ValueError("Falta a Base URL do FusionSolar.")

    last_error: Exception | None = None
    for attempt in range(2):
        try:
            session, _ = get_fusionsolar_session(config, force_login=attempt == 1)
            stations = fetch_fusionsolar_stations(
                session,
                base_url=endpoints["base_url"],
                endpoint=endpoints["plants_endpoint"],
            )
            if not stations:
                raise ValueError("A API FusionSolar nao devolveu centrais para esta conta.")

            station_codes = [
                str(row.get("plantCode") or row.get("stationCode") or "").strip()
                for row in stations
                if str(row.get("plantCode") or row.get("stationCode") or "").strip()
            ]
            realtime_map = fetch_fusionsolar_realtime_map(
                session,
                base_url=endpoints["base_url"],
                endpoint=endpoints["real_time_endpoint"],
                station_codes=station_codes,
            )
            alarm_map: dict[str, list[dict[str, Any]]] = {}
            alarm_error = ""
            try:
                alarm_map = fetch_fusionsolar_alarm_map(
                    session,
                    base_url=endpoints["base_url"],
                    endpoint=endpoints["alarms_endpoint"],
                    station_codes=station_codes,
                )
            except Exception as exc:
                alarm_error = str(exc)
            normalized_rows = [
                normalize_fusionsolar_plant_row(
                    station_row,
                    realtime_map.get(str(station_row.get("plantCode") or station_row.get("stationCode") or "").strip()),
                    alarm_map.get(str(station_row.get("plantCode") or station_row.get("stationCode") or "").strip(), []),
                )
                for station_row in stations
            ]
            break
        except Exception as exc:
            invalidate_fusionsolar_session(config)
            last_error = exc
            if attempt == 1:
                raise
    else:
        raise last_error or ValueError("Falha desconhecida no FusionSolar.")

    if not dry_run:
        conn.execute(
            """
            UPDATE integration_configs
            SET last_sync_status = ?, last_error = ?, updated_at = ?
            WHERE provider = ?
            """,
            ("success", "", datetime.now().isoformat(timespec="seconds"), provider),
        )
        conn.commit()
    return {
        "rows": normalized_rows,
        "station_count": len(stations),
        "realtime_count": len(realtime_map),
        "alarm_count": sum(len(items) for items in alarm_map.values()),
        "alarm_error": alarm_error,
    }


def run_fusionsolar_production_sync(
    conn: sqlite3.Connection,
    provider: str = "FusionSolar",
    target_date: date | None = None,
    period_type: str = "day",
) -> dict[str, Any]:
    if period_type not in {"day", "month"}:
        raise ValueError("Periodo invalido para performance.")
    if target_date is None:
        target_date = date.today() - timedelta(days=1)
    if period_type == "month":
        target_date = target_date.replace(day=1)

    with FUSIONSOLAR_SYNC_LOCK:
        config = get_integration_config(conn, provider)
        if config is None:
            raise ValueError("Configuracao FusionSolar nao encontrada.")
        if not config["enabled"]:
            raise ValueError("A integracao FusionSolar esta desativada.")
        endpoints = get_fusionsolar_endpoint_config(config)
        if not endpoints["base_url"]:
            raise ValueError("Falta a Base URL do FusionSolar.")

        mapped_assets = query_all(
            conn,
            """
            SELECT
                a.id AS asset_id,
                a.project_name,
                a.kwp,
                ai.external_id,
                COALESCE(ps.enabled, 1) AS performance_enabled,
                ps.warning_deviation_pct,
                ps.alert_deviation_pct,
                ps.critical_deviation_pct,
                ps.baseline_years,
                ps.min_baseline_points,
                ps.monthly_budget_json
            FROM asset_integrations ai
            JOIN assets a ON a.id = ai.asset_id
            LEFT JOIN performance_settings ps ON ps.asset_id = a.id
            WHERE ai.provider = ?
              AND ai.enabled = 1
              AND COALESCE(ai.external_id, '') != ''
              AND COALESCE(a.monitoring_status, 'active') != 'disabled'
              AND COALESCE(ps.enabled, 1) = 1
            ORDER BY a.project_name COLLATE NOCASE
            """,
            (provider,),
        )
        if not mapped_assets:
            return {"processed": 0, "missing_data": 0, "no_reference": 0, "period_date": target_date.isoformat()}

        station_codes = [str(row["external_id"]).strip() for row in mapped_assets if str(row["external_id"] or "").strip()]
        session_obj, _ = get_fusionsolar_session(config)
        try:
            if period_type == "month":
                endpoint_used = endpoints["month_kpi_endpoint"]
                kpi_map = fetch_fusionsolar_kpi_month_map(
                    session_obj,
                    endpoints["base_url"],
                    endpoint_used,
                    station_codes,
                    target_date,
                )
            else:
                endpoint_used = endpoints["day_kpi_endpoint"]
                kpi_map = fetch_fusionsolar_kpi_day_map(
                    session_obj,
                    endpoints["base_url"],
                    endpoint_used,
                    station_codes,
                    target_date,
                )
        except Exception as exc:
            if is_fusionsolar_rate_limit_error(exc):
                reason = mark_fusionsolar_performance_rate_limited(conn)
                conn.commit()
                return {
                    "processed": 0,
                    "missing_data": 0,
                    "no_reference": 0,
                    "period_date": target_date.isoformat(),
                    "stopped_reason": reason,
                }
            raise

        processed = 0
        missing_data = 0
        no_reference = 0
        with_production = 0
        for row in mapped_assets:
            asset_id = int(row["asset_id"])
            external_id = str(row["external_id"] or "").strip()
            kpi_row = kpi_map.get(external_id, {})
            data_item_map = kpi_row.get("dataItemMap") if isinstance(kpi_row, dict) else {}
            if not isinstance(data_item_map, dict):
                data_item_map = {}
            production_kwh, selected_key, selected_raw_value = select_production_value(data_item_map)
            if production_kwh is not None:
                with_production += 1
            kwp = parse_kwp_value(row["kwp"])
            specific_yield = calculate_specific_yield(production_kwh, kwp)
            settings = get_performance_settings(conn, asset_id)
            settings.update({key: row[key] for key in row.keys() if key in settings and row[key] is not None})

            expected_kwh, expected_specific_yield, expected_source, baseline_quality = calculate_expected_production(
                conn,
                asset_id=asset_id,
                provider=provider,
                period_type=period_type,
                period_date=target_date,
                kwp=kwp,
                settings=settings,
            )
            performance_status, data_quality, deviation_pct = classify_performance_status(
                production_kwh,
                kwp,
                expected_kwh,
                warning_deviation_pct=float(settings.get("warning_deviation_pct") or -10),
                alert_deviation_pct=float(settings.get("alert_deviation_pct") or -20),
                critical_deviation_pct=float(settings.get("critical_deviation_pct") or -30),
            )
            if data_quality == "ok" and baseline_quality == "partial_history" and expected_source == "none":
                data_quality = "partial_history"
            if performance_status == "Sem dados":
                missing_data += 1
            if performance_status == "Sem referência":
                no_reference += 1

            notes_parts = []
            if production_kwh is None:
                notes_parts.append(
                    build_missing_production_note(
                        data_item_map,
                        station_code=external_id,
                        period_type=period_type,
                        period_date=target_date,
                    )
                )
            if kwp is None:
                notes_parts.append("kWp local em falta ou invalido.")
            if expected_source == "none":
                notes_parts.append("Sem histórico ou orçamento mensal para referência.")
            upsert_production_record(
                conn,
                asset_id=asset_id,
                provider=provider,
                external_id=external_id,
                period_type=period_type,
                period_date=target_date,
                production_kwh=production_kwh,
                specific_yield=specific_yield,
                expected_kwh=expected_kwh,
                expected_specific_yield=expected_specific_yield,
                deviation_pct=deviation_pct,
                performance_status=performance_status,
                expected_source=expected_source,
                data_quality=data_quality,
                notes=" ".join(notes_parts),
                payload_json=str(kpi_row.get("payload_json") or json.dumps(kpi_row, ensure_ascii=True)),
                selected_production_key=selected_key,
                selected_production_raw_value=selected_raw_value,
            )
            processed += 1

        recalculate_performance_references(
            conn,
            period_type=period_type,
            period_date=target_date,
            provider=provider,
        )
        logger = current_app.logger if has_app_context() else logging.getLogger(__name__)
        logger.info(
            "FusionSolar performance sync: requested_station_codes=%s endpoint=%s period_type=%s target_date=%s api_rows=%s with_production=%s missing_production=%s",
            len(station_codes),
            endpoint_used,
            period_type,
            target_date.isoformat(),
            len(kpi_map),
            with_production,
            missing_data,
        )
        conn.commit()
        return {
            "processed": processed,
            "missing_data": missing_data,
            "no_reference": no_reference,
            "period_date": target_date.isoformat(),
        }


def iter_daily_backfill_dates(from_year: int, to_year: int, *, today_value: date | None = None) -> list[date]:
    today_value = today_value or date.today()
    cursor = date(from_year, 1, 1)
    end_date = min(date(to_year, 12, 31), today_value - timedelta(days=1))
    days: list[date] = []
    while cursor <= end_date:
        days.append(cursor)
        cursor += timedelta(days=1)
    return days


def iter_daily_backfill_months(
    from_year: int,
    to_year: int,
    *,
    today_value: date | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
) -> list[date]:
    today_value = today_value or date.today()
    start_date = date_from or date(from_year, 1, 1)
    end_date = date_to or date(to_year, 12, 31)
    end_date = min(end_date, today_value - timedelta(days=1))
    if start_date > end_date:
        return []
    cursor = start_date.replace(day=1)
    end_month = end_date.replace(day=1)
    months: list[date] = []
    while cursor <= end_month:
        months.append(cursor)
        year = cursor.year + (1 if cursor.month == 12 else 0)
        month = 1 if cursor.month == 12 else cursor.month + 1
        cursor = date(year, month, 1)
    return months


def date_in_backfill_window(
    candidate: date,
    *,
    from_year: int,
    to_year: int,
    today_value: date,
    date_from: date | None = None,
    date_to: date | None = None,
) -> bool:
    start_date = date_from or date(from_year, 1, 1)
    end_date = date_to or date(to_year, 12, 31)
    end_date = min(end_date, today_value - timedelta(days=1))
    return start_date <= candidate <= end_date


def iter_monthly_backfill_dates(from_year: int, to_year: int, *, today_value: date | None = None) -> list[date]:
    today_value = today_value or date.today()
    current_month = today_value.replace(day=1)
    months: list[date] = []
    for year in range(from_year, to_year + 1):
        for month in range(1, 13):
            period_date = date(year, month, 1)
            if period_date >= current_month:
                continue
            months.append(period_date)
    return months


def get_fusionsolar_performance_assets(
    conn: sqlite3.Connection,
    provider: str,
    asset_id: int | None = None,
    asset_ids: list[int] | None = None,
) -> list[sqlite3.Row]:
    conditions = [
        "ai.provider = ?",
        "ai.enabled = 1",
        "COALESCE(ai.external_id, '') != ''",
        "COALESCE(a.monitoring_status, 'active') != 'disabled'",
        "COALESCE(ps.enabled, 1) = 1",
    ]
    params: list[Any] = [provider]
    if asset_id:
        conditions.append("a.id = ?")
        params.append(asset_id)
    if asset_ids:
        placeholders = ", ".join("?" for _ in asset_ids)
        conditions.append(f"a.id IN ({placeholders})")
        params.extend(asset_ids)
    return query_all(
        conn,
        f"""
        SELECT
            a.id AS asset_id,
            a.project_name,
            a.kwp,
            ai.external_id,
            COALESCE(ps.enabled, 1) AS performance_enabled,
            ps.warning_deviation_pct,
            ps.alert_deviation_pct,
            ps.critical_deviation_pct,
            ps.baseline_years,
            ps.min_baseline_points,
            ps.monthly_budget_json
        FROM asset_integrations ai
        JOIN assets a ON a.id = ai.asset_id
        LEFT JOIN performance_settings ps ON ps.asset_id = a.id
        WHERE {" AND ".join(conditions)}
        ORDER BY a.project_name COLLATE NOCASE
        """,
        params,
    )


def recalculate_production_expectations(
    conn: sqlite3.Connection,
    *,
    provider: str,
    asset_ids: list[int],
    period_type: str | None = None,
) -> int:
    if not asset_ids:
        return 0
    total = 0
    for asset_id in asset_ids:
        summary = recalculate_performance_references(
            conn,
            period_type=period_type,
            asset_id=asset_id,
            provider=provider,
        )
        total += summary["records_processed"]
    return total


def recalculate_performance_references(
    conn: sqlite3.Connection,
    period_type: str | None = None,
    period_date: date | str | None = None,
    asset_id: int | None = None,
    provider: str = "FusionSolar",
    today_value: date | None = None,
) -> dict[str, int]:
    conditions = ["pr.provider = ?"]
    params: list[Any] = [provider]
    if period_type:
        conditions.append("pr.period_type = ?")
        params.append(period_type)
    if period_date:
        normalized_date = period_date.isoformat() if isinstance(period_date, date) else str(period_date)
        conditions.append("pr.period_date = ?")
        params.append(normalized_date)
    if asset_id:
        conditions.append("pr.asset_id = ?")
        params.append(asset_id)
    rows = query_all(
        conn,
        f"""
        SELECT
            pr.*,
            a.project_name,
            a.kwp,
            ps.warning_deviation_pct,
            ps.alert_deviation_pct,
            ps.critical_deviation_pct,
            ps.baseline_years,
            ps.min_baseline_points,
            ps.monthly_budget_json
        FROM production_records pr
        JOIN assets a ON a.id = pr.asset_id
        LEFT JOIN performance_settings ps ON ps.asset_id = pr.asset_id
        WHERE {" AND ".join(conditions)}
        ORDER BY pr.period_date ASC, pr.id ASC
        """,
        params,
    )
    summary = {
        "records_processed": 0,
        "references_created": 0,
        "still_without_reference": 0,
        "missing_kwp": 0,
        "missing_production": 0,
    }
    for row in rows:
        target_date = parse_date_value(row["period_date"])
        if target_date is None:
            continue
        kwp = parse_kwp_value(row["kwp"])
        settings = get_performance_settings(conn, int(row["asset_id"]))
        settings.update({key: row[key] for key in row.keys() if key in settings and row[key] is not None})
        reference_result = calculate_expected_production_with_diagnostic(
            conn,
            asset_id=int(row["asset_id"]),
            provider=provider,
            period_type=row["period_type"],
            period_date=target_date,
            kwp=kwp,
            settings=settings,
            asset_name=row["project_name"],
            today_value=today_value,
        )
        expected_kwh = reference_result["expected_kwh"]
        expected_specific_yield = reference_result["expected_specific_yield"]
        expected_source = reference_result["expected_source"]
        performance_status, data_quality, deviation_pct = classify_performance_status(
            row["production_kwh"],
            kwp,
            expected_kwh,
            warning_deviation_pct=float(settings.get("warning_deviation_pct") or -10),
            alert_deviation_pct=float(settings.get("alert_deviation_pct") or -20),
            critical_deviation_pct=float(settings.get("critical_deviation_pct") or -30),
        )
        if data_quality == "ok" and reference_result["quality"] == "partial_history" and expected_source == "none":
            data_quality = "partial_history"
        notes = row["notes"] or ""
        reason = reference_result["diagnostic"].get("no_reference_reason") or ""
        if reason and reason not in notes:
            notes = f"{notes} {reason}".strip()
        conn.execute(
            """
            UPDATE production_records
            SET expected_kwh = ?, expected_specific_yield = ?, deviation_pct = ?,
                performance_status = ?, expected_source = ?, data_quality = ?,
                reference_diagnostic_json = ?, notes = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                expected_kwh,
                expected_specific_yield,
                deviation_pct,
                performance_status,
                expected_source,
                data_quality,
                json.dumps(reference_result["diagnostic"], ensure_ascii=True),
                notes,
                datetime.now().isoformat(timespec="seconds"),
                row["id"],
            ),
        )
        summary["records_processed"] += 1
        if expected_kwh is not None and expected_specific_yield is not None:
            summary["references_created"] += 1
        else:
            summary["still_without_reference"] += 1
        if kwp is None:
            summary["missing_kwp"] += 1
        if row["production_kwh"] is None:
            summary["missing_production"] += 1
    conn.commit()
    return summary


def _run_fusionsolar_production_backfill_legacy(
    conn: sqlite3.Connection,
    *,
    provider: str = "FusionSolar",
    period_type: str = "day",
    from_year: int,
    to_year: int,
    asset_id: int | None = None,
    today_value: date | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    max_days: int | None = None,
) -> dict[str, Any]:
    if period_type not in {"day", "month"}:
        raise ValueError("Tipo de período inválido.")
    if from_year > to_year:
        raise ValueError("Ano inicial nao pode ser superior ao ano final.")
    if (to_year - from_year + 1) > 3:
        raise ValueError("Intervalo superior a 3 anos. Reduz o período para executar o backfill.")

    today_value = today_value or date.today()
    config = get_integration_config(conn, provider)
    if config is None:
        raise ValueError("Configuracao FusionSolar nao encontrada.")
    if not config["enabled"]:
        raise ValueError("A integracao FusionSolar esta desativada.")
    endpoints = get_fusionsolar_endpoint_config(config)
    assets = get_fusionsolar_performance_assets(conn, provider, asset_id)
    dates = (
        iter_daily_backfill_dates(from_year, to_year, today_value=today_value)
        if period_type == "day"
        else iter_monthly_backfill_dates(from_year, to_year, today_value=today_value)
    )
    if period_type == "day":
        if date_from is not None:
            dates = [candidate for candidate in dates if candidate >= date_from]
        if date_to is not None:
            dates = [candidate for candidate in dates if candidate <= date_to]
        if max_days is not None and max_days > 0:
            dates = dates[:max_days]

    summary = {
        "assets_processed": 0,
        "records_updated": 0,
        "missing_production": 0,
        "api_errors": 0,
        "mtd_records_updated": 0,
        "baselines_recalculated": 0,
        "references_created": 0,
        "still_without_reference": 0,
    }
    logger = current_app.logger if has_app_context() else logging.getLogger(__name__)

    for asset in assets:
        summary["assets_processed"] += 1
        external_id = str(asset["external_id"] or "").strip()
        for period_date in dates:
            try:
                if period_type == "month":
                    kpi_map = fetch_fusionsolar_kpi_month_map(
                        session_obj,
                        endpoints["base_url"],
                        endpoints["month_kpi_endpoint"],
                        [external_id],
                        period_date,
                    )
                else:
                    kpi_map = fetch_fusionsolar_kpi_day_map(
                        session_obj,
                        endpoints["base_url"],
                        endpoints["day_kpi_endpoint"],
                        [external_id],
                        period_date,
                    )
                result = store_production_kpi_record(
                    conn,
                    asset_row=asset,
                    provider=provider,
                    external_id=external_id,
                    period_type=period_type,
                    period_date=period_date,
                    kpi_row=kpi_map.get(external_id, {}),
                    notes_prefix="Backfill histórico.",
                )
                summary["records_updated"] += 1
                if result["production_kwh"] is None:
                    summary["missing_production"] += 1
            except Exception as exc:
                summary["api_errors"] += 1
                logger.warning(
                    "FusionSolar performance backfill failed: asset_id=%s stationCode=%s period_type=%s period_date=%s error=%s",
                    asset["asset_id"],
                    external_id,
                    period_type,
                    period_date.isoformat(),
                    exc,
                )
                continue

    selected_asset_ids = [int(asset["asset_id"]) for asset in assets]
    recalc_targets = dates if period_type == "month" else ([max(dates)] if dates else [])
    if not summary["stopped_reason"]:
        for target in recalc_targets:
            recalc = recalculate_performance_references(
                conn,
                period_type=period_type,
                period_date=target,
                asset_id=asset_id,
                provider=provider,
                today_value=today_value,
            )
            summary["baselines_recalculated"] += recalc["records_processed"]
            summary["references_created"] += recalc["references_created"]
            summary["still_without_reference"] += recalc["still_without_reference"]

    current_month = today_value.replace(day=1)
    if selected_asset_ids:
        for asset in assets:
            external_id = str(asset["external_id"] or "").strip()
            try:
                kpi_map = fetch_fusionsolar_kpi_month_map(
                    session_obj,
                    endpoints["base_url"],
                    endpoints["month_kpi_endpoint"],
                    [external_id],
                    current_month,
                )
                store_production_kpi_record(
                    conn,
                    asset_row=asset,
                    provider=provider,
                    external_id=external_id,
                    period_type="mtd",
                    period_date=current_month,
                    kpi_row=kpi_map.get(external_id, {}),
                    notes_prefix="MTD recalculado após backfill.",
                )
                summary["mtd_records_updated"] += 1
            except Exception as exc:
                summary["api_errors"] += 1
                logger.warning(
                    "FusionSolar MTD recalculation failed after backfill: asset_id=%s stationCode=%s error=%s",
                    asset["asset_id"],
                    external_id,
                    exc,
                )
        recalc = recalculate_performance_references(
            conn,
            period_type="mtd",
            period_date=current_month,
            asset_id=asset_id,
            provider=provider,
            today_value=today_value,
        )
        summary["baselines_recalculated"] += recalc["records_processed"]
        summary["references_created"] += recalc["references_created"]
        summary["still_without_reference"] += recalc["still_without_reference"]
    conn.commit()
    return summary


def run_fusionsolar_production_backfill(
    conn: sqlite3.Connection,
    *,
    provider: str = "FusionSolar",
    period_type: str = "day",
    from_year: int,
    to_year: int,
    asset_id: int | None = None,
    today_value: date | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    max_days: int | None = None,
    max_api_calls: int | None = None,
    kpi_call_delay_seconds: float | None = None,
    sleeper: Any | None = None,
    max_wait_cycles: int = 24,
) -> dict[str, Any]:
    if period_type not in {"day", "month"}:
        raise ValueError("Tipo de periodo invalido.")
    if from_year > to_year:
        raise ValueError("Ano inicial nao pode ser superior ao ano final.")
    if (to_year - from_year + 1) > 3:
        raise ValueError("Intervalo superior a 3 anos. Reduz o periodo para executar o backfill.")

    today_value = today_value or date.today()
    config = get_integration_config(conn, provider)
    if config is None:
        raise ValueError("Configuracao FusionSolar nao encontrada.")
    if not config["enabled"]:
        raise ValueError("A integracao FusionSolar esta desativada.")
    endpoints = get_fusionsolar_endpoint_config(config)
    assets = get_fusionsolar_performance_assets(conn, provider, asset_id)
    dates = iter_monthly_backfill_dates(from_year, to_year, today_value=today_value)
    if period_type == "day":
        dates = iter_daily_backfill_months(
            from_year,
            to_year,
            today_value=today_value,
            date_from=date_from,
            date_to=date_to,
        )

    summary = {
        "assets_processed": 0,
        "records_updated": 0,
        "missing_production": 0,
        "api_errors": 0,
        "api_calls_used": 0,
        "months_processed": 0,
        "chunks_processed": 0,
        "mtd_records_updated": 0,
        "baselines_recalculated": 0,
        "references_created": 0,
        "still_without_reference": 0,
        "stopped_reason": "",
        "resume_hint": "",
        "wait_cycles": 0,
    }
    logger = current_app.logger if has_app_context() else logging.getLogger(__name__)
    station_codes = [str(asset["external_id"] or "").strip() for asset in assets if str(asset["external_id"] or "").strip()]
    assets_by_external_id = {
        str(asset["external_id"] or "").strip(): asset
        for asset in assets
        if str(asset["external_id"] or "").strip()
    }
    summary["assets_processed"] = len(assets_by_external_id)
    max_api_calls = max_api_calls if max_api_calls is not None else max_days
    max_api_calls = max_api_calls if max_api_calls is not None else FUSIONSOLAR_PERFORMANCE_MAX_API_CALLS
    kpi_call_delay_seconds = (
        FUSIONSOLAR_PERFORMANCE_KPI_DELAY_SECONDS
        if kpi_call_delay_seconds is None
        else kpi_call_delay_seconds
    )
    sleep_func = sleeper or time.sleep
    processed_dates: list[date] = []

    def wait_after_rate_limit(reason: str, resume_hint: date | None = None) -> bool:
        summary["wait_cycles"] += 1
        summary["resume_hint"] = resume_hint.isoformat() if resume_hint else summary["resume_hint"]
        if summary["wait_cycles"] > max_wait_cycles:
            summary["stopped_reason"] = f"Limite FusionSolar repetido demasiadas vezes. Ultimo estado: {reason}"
            logger.warning(
                "FusionSolar performance backfill stopped after repeated cooldowns: period_type=%s station_count=%s wait_cycles=%s reason=%s",
                period_type,
                len(station_codes),
                summary["wait_cycles"],
                reason,
            )
            return False
        conn.commit()
        seconds = fusionsolar_cooldown_sleep_seconds(conn)
        logger.warning(
            "FusionSolar performance backfill waiting for API cooldown: period_type=%s station_count=%s seconds=%s wait_cycle=%s",
            period_type,
            len(station_codes),
            seconds,
            summary["wait_cycles"],
        )
        sleep_func(seconds)
        return True

    cooldown_reason = get_fusionsolar_performance_cooldown_reason(conn)
    if cooldown_reason and not wait_after_rate_limit(cooldown_reason):
        summary["api_errors"] = 1
        return summary

    session_obj, _ = get_fusionsolar_session(config)

    def refresh_session_after_expiry(exc: Exception, context: str) -> None:
        nonlocal session_obj
        invalidate_fusionsolar_session(config)
        session_obj, _ = get_fusionsolar_session(config, force_login=True)
        logger.warning("FusionSolar session refreshed after expired login: context=%s error=%s", context, exc)

    def store_kpi_map(period_date_value: date, kpi_map_value: dict[str, dict[str, Any]], record_type: str, notes_prefix: str) -> None:
        for external_id_value, asset in assets_by_external_id.items():
            result = store_production_kpi_record(
                conn,
                asset_row=asset,
                provider=provider,
                external_id=external_id_value,
                period_type=record_type,
                period_date=period_date_value,
                kpi_row=kpi_map_value.get(external_id_value, {}),
                notes_prefix=notes_prefix,
            )
            if record_type == "mtd":
                summary["mtd_records_updated"] += 1
            else:
                summary["records_updated"] += 1
            if result["production_kwh"] is None:
                summary["missing_production"] += 1

    def wait_before_next_call() -> None:
        if kpi_call_delay_seconds and kpi_call_delay_seconds > 0:
            sleep_func(kpi_call_delay_seconds)

    if period_type == "day":
        for month_value in dates:
            month_had_records = False
            station_chunks = chunked(station_codes, 100)
            for chunk_index, station_group in enumerate(station_chunks, start=1):
                session_retry_used = False
                while True:
                    if summary["api_calls_used"] >= max_api_calls:
                        summary["stopped_reason"] = (
                            f"Limite local de {max_api_calls} chamadas API atingido. "
                            f"Retoma a partir de {month_value.isoformat()}."
                        )
                        summary["resume_hint"] = month_value.isoformat()
                        logger.info(
                            "FusionSolar performance backfill stopped by max calls: period_type=%s month=%s api_calls_used=%s max_api_calls=%s",
                            period_type,
                            month_value.isoformat(),
                            summary["api_calls_used"],
                            max_api_calls,
                        )
                        break
                    if summary["api_calls_used"] > 0:
                        wait_before_next_call()
                    try:
                        logger.info(
                            "FusionSolar daily performance backfill request: period_type=%s month=%s station_count=%s chunk_index=%s api_calls_used=%s",
                            period_type,
                            month_value.isoformat(),
                            len(station_group),
                            chunk_index,
                            summary["api_calls_used"],
                        )
                        rows = fetch_fusionsolar_kpi_day_rows(
                            session_obj,
                            endpoints["base_url"],
                            endpoints["day_kpi_endpoint"],
                            station_group,
                            month_value,
                        )
                        summary["api_calls_used"] += 1
                        summary["chunks_processed"] += 1
                        for row in rows:
                            external_id_value = str(row.get("stationCode") or row.get("plantCode") or "").strip()
                            asset = assets_by_external_id.get(external_id_value)
                            if asset is None:
                                continue
                            row_date = parse_fusionsolar_collect_date(row, month_value)
                            if row_date is None or row_date.replace(day=1) != month_value:
                                continue
                            if not date_in_backfill_window(
                                row_date,
                                from_year=from_year,
                                to_year=to_year,
                                today_value=today_value,
                                date_from=date_from,
                                date_to=date_to,
                            ):
                                continue
                            result = store_production_kpi_record(
                                conn,
                                asset_row=asset,
                                provider=provider,
                                external_id=external_id_value,
                                period_type="day",
                                period_date=row_date,
                                kpi_row=row,
                                notes_prefix="Backfill historico mensal diario.",
                            )
                            if result["upsert_status"] != "skipped_existing_valid":
                                summary["records_updated"] += 1
                                processed_dates.append(row_date)
                                month_had_records = True
                            if result["production_kwh"] is None:
                                summary["missing_production"] += 1
                        logger.info(
                            "FusionSolar daily performance backfill response: month=%s chunk_index=%s rows=%s records_updated=%s api_calls_used=%s",
                            month_value.isoformat(),
                            chunk_index,
                            len(rows),
                            summary["records_updated"],
                            summary["api_calls_used"],
                        )
                        break
                    except Exception as exc:
                        summary["api_errors"] += 1
                        if is_fusionsolar_rate_limit_error(exc):
                            reason = mark_fusionsolar_performance_rate_limited(conn)
                            logger.warning(
                                "FusionSolar performance backfill rate limited: period_type=%s month=%s station_count=%s chunk_index=%s api_calls_used=%s error=%s",
                                period_type,
                                month_value.isoformat(),
                                len(station_group),
                                chunk_index,
                                summary["api_calls_used"],
                                exc,
                            )
                            if wait_after_rate_limit(reason, month_value):
                                continue
                            break
                        if is_fusionsolar_session_expired_error(exc) and not session_retry_used:
                            session_retry_used = True
                            refresh_session_after_expiry(exc, f"daily:{month_value.isoformat()}:chunk:{chunk_index}")
                            continue
                        logger.warning(
                            "FusionSolar daily performance backfill chunk failed: period_type=%s month=%s station_count=%s chunk_index=%s error=%s",
                            period_type,
                            month_value.isoformat(),
                            len(station_group),
                            chunk_index,
                            exc,
                        )
                        break
                if summary["stopped_reason"]:
                    break
            if month_had_records:
                summary["months_processed"] += 1
            if summary["stopped_reason"]:
                break

        recalc_targets = sorted(set(processed_dates))
        for target in recalc_targets:
            recalc = recalculate_performance_references(
                conn,
                period_type="day",
                period_date=target,
                asset_id=asset_id,
                provider=provider,
                today_value=today_value,
            )
            summary["baselines_recalculated"] += recalc["records_processed"]
            summary["references_created"] += recalc["references_created"]
            summary["still_without_reference"] += recalc["still_without_reference"]
        selected_asset_ids = [int(asset["asset_id"]) for asset in assets]
        current_month = today_value.replace(day=1)
        if selected_asset_ids and not summary["stopped_reason"]:
            session_retry_used = False
            while True:
                try:
                    if summary["api_calls_used"] > 0:
                        wait_before_next_call()
                    kpi_map = fetch_fusionsolar_kpi_month_map(
                        session_obj,
                        endpoints["base_url"],
                        endpoints["month_kpi_endpoint"],
                        station_codes,
                        current_month,
                    )
                    summary["api_calls_used"] += max(1, len(chunked(station_codes, 100)))
                    store_kpi_map(current_month, kpi_map, "mtd", "MTD recalculado apos backfill.")
                    break
                except Exception as exc:
                    summary["api_errors"] += 1
                    if is_fusionsolar_rate_limit_error(exc):
                        reason = mark_fusionsolar_performance_rate_limited(conn)
                        if wait_after_rate_limit(reason, current_month):
                            continue
                    if is_fusionsolar_session_expired_error(exc) and not session_retry_used:
                        session_retry_used = True
                        refresh_session_after_expiry(exc, f"daily-mtd:{current_month.isoformat()}")
                        continue
                    logger.warning(
                        "FusionSolar MTD recalculation failed after backfill: station_count=%s error=%s",
                        len(station_codes),
                        exc,
                    )
                    break
            recalc = recalculate_performance_references(
                conn,
                period_type="mtd",
                period_date=current_month,
                asset_id=asset_id,
                provider=provider,
                today_value=today_value,
            )
            summary["baselines_recalculated"] += recalc["records_processed"]
            summary["references_created"] += recalc["references_created"]
            summary["still_without_reference"] += recalc["still_without_reference"]
        conn.commit()
        return summary

    for period_date_value in dates:
        stored_current_date = False
        session_retry_used = False
        while True:
            try:
                if summary["api_calls_used"] >= max_api_calls:
                    summary["stopped_reason"] = (
                        f"Limite local de {max_api_calls} chamadas API atingido. "
                        f"Retoma a partir de {period_date_value.isoformat()}."
                    )
                    summary["resume_hint"] = period_date_value.isoformat()
                    break
                if summary["api_calls_used"] > 0:
                    wait_before_next_call()
                kpi_map = fetch_fusionsolar_kpi_month_map(
                    session_obj,
                    endpoints["base_url"],
                    endpoints["month_kpi_endpoint"],
                    station_codes,
                    period_date_value,
                )
                summary["api_calls_used"] += max(1, len(chunked(station_codes, 100)))
                store_kpi_map(period_date_value, kpi_map, period_type, "Backfill historico.")
                stored_current_date = True
                break
            except Exception as exc:
                if is_fusionsolar_rate_limit_error(exc):
                    summary["api_errors"] += 1
                    reason = mark_fusionsolar_performance_rate_limited(conn)
                    logger.warning(
                        "FusionSolar performance backfill rate limited: period_type=%s period_date=%s station_count=%s error=%s",
                        period_type,
                        period_date_value.isoformat(),
                        len(station_codes),
                        exc,
                    )
                    if wait_after_rate_limit(reason, period_date_value):
                        continue
                    break
                if is_fusionsolar_session_expired_error(exc) and not session_retry_used:
                    session_retry_used = True
                    summary["api_errors"] += 1
                    refresh_session_after_expiry(exc, f"month-group:{period_date_value.isoformat()}")
                    continue
                logger.warning(
                    "FusionSolar grouped performance backfill failed, retrying per asset: period_type=%s period_date=%s station_count=%s error=%s",
                    period_type,
                    period_date_value.isoformat(),
                    len(station_codes),
                    exc,
                )
                break
        if summary["stopped_reason"]:
            break
        if stored_current_date:
            processed_dates.append(period_date_value)
            continue

        stored_any_for_date = False
        for external_id_value, asset in assets_by_external_id.items():
            session_retry_used = False
            while True:
                try:
                    if summary["api_calls_used"] >= max_api_calls:
                        summary["stopped_reason"] = (
                            f"Limite local de {max_api_calls} chamadas API atingido. "
                            f"Retoma a partir de {period_date_value.isoformat()}."
                        )
                        summary["resume_hint"] = period_date_value.isoformat()
                        break
                    if summary["api_calls_used"] > 0:
                        wait_before_next_call()
                    single_map = fetch_fusionsolar_kpi_month_map(
                        session_obj,
                        endpoints["base_url"],
                        endpoints["month_kpi_endpoint"],
                        [external_id_value],
                        period_date_value,
                    )
                    summary["api_calls_used"] += 1
                    store_kpi_map(period_date_value, single_map, period_type, "Backfill historico.")
                    stored_any_for_date = True
                    break
                except Exception as asset_exc:
                    summary["api_errors"] += 1
                    if is_fusionsolar_rate_limit_error(asset_exc):
                        reason = mark_fusionsolar_performance_rate_limited(conn)
                        logger.warning(
                            "FusionSolar performance backfill rate limited: asset_id=%s stationCode=%s period_type=%s period_date=%s error=%s",
                            asset["asset_id"],
                            external_id_value,
                            period_type,
                            period_date_value.isoformat(),
                            asset_exc,
                        )
                        if wait_after_rate_limit(reason, period_date_value):
                            continue
                        break
                    if is_fusionsolar_session_expired_error(asset_exc) and not session_retry_used:
                        session_retry_used = True
                        refresh_session_after_expiry(asset_exc, f"month-asset:{period_date_value.isoformat()}:{external_id_value}")
                        continue
                    logger.warning(
                        "FusionSolar performance backfill failed: asset_id=%s stationCode=%s period_type=%s period_date=%s error=%s",
                        asset["asset_id"],
                        external_id_value,
                        period_type,
                        period_date_value.isoformat(),
                        asset_exc,
                    )
                    break
            if summary["stopped_reason"]:
                break
        if summary["stopped_reason"]:
            break
        if stored_any_for_date:
            processed_dates.append(period_date_value)

    selected_asset_ids = [int(asset["asset_id"]) for asset in assets]
    recalc_targets = processed_dates
    for target in recalc_targets:
        recalc = recalculate_performance_references(
            conn,
            period_type=period_type,
            period_date=target,
            asset_id=asset_id,
            provider=provider,
            today_value=today_value,
        )
        summary["baselines_recalculated"] += recalc["records_processed"]
        summary["references_created"] += recalc["references_created"]
        summary["still_without_reference"] += recalc["still_without_reference"]

    current_month = today_value.replace(day=1)
    if selected_asset_ids and not summary["stopped_reason"]:
        session_retry_used = False
        while True:
            try:
                if summary["api_calls_used"] > 0:
                    wait_before_next_call()
                kpi_map = fetch_fusionsolar_kpi_month_map(
                    session_obj,
                    endpoints["base_url"],
                    endpoints["month_kpi_endpoint"],
                    station_codes,
                    current_month,
                )
                summary["api_calls_used"] += max(1, len(chunked(station_codes, 100)))
                store_kpi_map(current_month, kpi_map, "mtd", "MTD recalculado apos backfill.")
                break
            except Exception as exc:
                summary["api_errors"] += 1
                if is_fusionsolar_rate_limit_error(exc):
                    reason = mark_fusionsolar_performance_rate_limited(conn)
                    if wait_after_rate_limit(reason, current_month):
                        continue
                if is_fusionsolar_session_expired_error(exc) and not session_retry_used:
                    session_retry_used = True
                    refresh_session_after_expiry(exc, f"month-mtd:{current_month.isoformat()}")
                    continue
                logger.warning(
                    "FusionSolar MTD recalculation failed after backfill: station_count=%s error=%s",
                    len(station_codes),
                    exc,
                )
                break
        recalc = recalculate_performance_references(
            conn,
            period_type="mtd",
            period_date=current_month,
            asset_id=asset_id,
            provider=provider,
            today_value=today_value,
        )
        summary["baselines_recalculated"] += recalc["records_processed"]
        summary["references_created"] += recalc["references_created"]
        summary["still_without_reference"] += recalc["still_without_reference"]
    conn.commit()
    return summary


def run_fusionsolar_month_cycle(
    conn: sqlite3.Connection,
    *,
    provider: str = "FusionSolar",
    report_month: str,
    asset_ids: list[int],
    kpi_call_delay_seconds: float | None = None,
    sleeper: Any | None = None,
    max_wait_cycles: int = 24,
) -> dict[str, Any]:
    if not asset_ids:
        raise ValueError("Escolhe pelo menos uma instalacao para o ciclo.")
    month_start = datetime.strptime(normalize_report_month(report_month), "%Y-%m").date()
    _, month_last_day = calendar.monthrange(month_start.year, month_start.month)
    month_end = month_start.replace(day=month_last_day)
    today_value = date.today()
    date_to = min(month_end, today_value - timedelta(days=1))
    if date_to < month_start:
        raise ValueError("O mes escolhido ainda nao tem dias fechados para importar.")

    config = get_integration_config(conn, provider)
    if config is None:
        raise ValueError("Configuracao FusionSolar nao encontrada.")
    if not config["enabled"]:
        raise ValueError("A integracao FusionSolar esta desativada.")
    endpoints = get_fusionsolar_endpoint_config(config)
    assets = get_fusionsolar_performance_assets(conn, provider, asset_ids=asset_ids)
    if not assets:
        raise ValueError("Nenhuma das instalacoes escolhidas tem mapeamento FusionSolar ativo.")

    station_codes = [str(asset["external_id"] or "").strip() for asset in assets if str(asset["external_id"] or "").strip()]
    assets_by_external_id = {str(asset["external_id"] or "").strip(): asset for asset in assets if str(asset["external_id"] or "").strip()}
    station_chunks = chunked(station_codes, 100)
    session_obj, _ = get_fusionsolar_session(config)
    sleep_func = sleeper or time.sleep
    kpi_call_delay_seconds = FUSIONSOLAR_PERFORMANCE_KPI_DELAY_SECONDS if kpi_call_delay_seconds is None else kpi_call_delay_seconds
    logger = current_app.logger if has_app_context() else logging.getLogger(__name__)
    summary = {
        "assets_selected": len(asset_ids),
        "assets_mapped": len(assets_by_external_id),
        "month": month_start.strftime("%Y-%m"),
        "daily_chunks_total": len(station_chunks),
        "daily_chunks_completed": 0,
        "monthly_chunks_completed": 0,
        "records_updated": 0,
        "monthly_records_updated": 0,
        "missing_production": 0,
        "api_calls_used": 0,
        "api_errors": 0,
        "wait_cycles": 0,
        "status": "running",
    }
    processed_dates: set[date] = set()

    def wait_after_rate_limit(reason: str) -> None:
        summary["wait_cycles"] += 1
        if summary["wait_cycles"] > max_wait_cycles:
            raise ValueError(f"Limite FusionSolar repetido demasiadas vezes. Ultimo estado: {reason}")
        conn.commit()
        seconds = fusionsolar_cooldown_sleep_seconds(conn)
        logger.warning(
            "FusionSolar month cycle waiting for API cooldown: month=%s seconds=%s wait_cycle=%s",
            month_start.isoformat(),
            seconds,
            summary["wait_cycles"],
        )
        sleep_func(seconds)

    def wait_between_calls() -> None:
        if summary["api_calls_used"] > 0 and kpi_call_delay_seconds and kpi_call_delay_seconds > 0:
            sleep_func(kpi_call_delay_seconds)

    def refresh_session_after_expiry(exc: Exception, context: str) -> None:
        nonlocal session_obj
        invalidate_fusionsolar_session(config)
        session_obj, _ = get_fusionsolar_session(config, force_login=True)
        logger.warning("FusionSolar session refreshed after expired login: context=%s error=%s", context, exc)

    for chunk_index, station_group in enumerate(station_chunks, start=1):
        session_retry_used = False
        while True:
            try:
                wait_between_calls()
                rows = fetch_fusionsolar_kpi_day_rows(
                    session_obj,
                    endpoints["base_url"],
                    endpoints["day_kpi_endpoint"],
                    station_group,
                    month_start,
                )
                summary["api_calls_used"] += 1
                for row in rows:
                    external_id_value = str(row.get("stationCode") or row.get("plantCode") or "").strip()
                    asset = assets_by_external_id.get(external_id_value)
                    if asset is None:
                        continue
                    row_date = parse_fusionsolar_collect_date(row, month_start)
                    if row_date is None or row_date < month_start or row_date > date_to:
                        continue
                    result = store_production_kpi_record(
                        conn,
                        asset_row=asset,
                        provider=provider,
                        external_id=external_id_value,
                        period_type="day",
                        period_date=row_date,
                        kpi_row=row,
                        notes_prefix="Ciclo mensal automatico.",
                    )
                    if result["upsert_status"] != "skipped_existing_valid":
                        summary["records_updated"] += 1
                        processed_dates.add(row_date)
                    if result["production_kwh"] is None:
                        summary["missing_production"] += 1
                summary["daily_chunks_completed"] = chunk_index
                conn.commit()
                break
            except Exception as exc:
                summary["api_errors"] += 1
                if is_fusionsolar_rate_limit_error(exc):
                    wait_after_rate_limit(mark_fusionsolar_performance_rate_limited(conn))
                    continue
                if is_fusionsolar_session_expired_error(exc) and not session_retry_used:
                    session_retry_used = True
                    refresh_session_after_expiry(exc, f"month-cycle-day:{month_start.isoformat()}:chunk:{chunk_index}")
                    continue
                raise

    for chunk_index, station_group in enumerate(station_chunks, start=1):
        session_retry_used = False
        while True:
            try:
                wait_between_calls()
                kpi_map = fetch_fusionsolar_kpi_month_map(
                    session_obj,
                    endpoints["base_url"],
                    endpoints["month_kpi_endpoint"],
                    station_group,
                    month_start,
                )
                summary["api_calls_used"] += 1
                for external_id_value in station_group:
                    asset = assets_by_external_id.get(external_id_value)
                    if asset is None:
                        continue
                    result = store_production_kpi_record(
                        conn,
                        asset_row=asset,
                        provider=provider,
                        external_id=external_id_value,
                        period_type="month",
                        period_date=month_start,
                        kpi_row=kpi_map.get(external_id_value, {}),
                        notes_prefix="Ciclo mensal automatico.",
                    )
                    if result["upsert_status"] != "skipped_existing_valid":
                        summary["monthly_records_updated"] += 1
                    if result["production_kwh"] is None:
                        summary["missing_production"] += 1
                summary["monthly_chunks_completed"] = chunk_index
                conn.commit()
                break
            except Exception as exc:
                summary["api_errors"] += 1
                if is_fusionsolar_rate_limit_error(exc):
                    wait_after_rate_limit(mark_fusionsolar_performance_rate_limited(conn))
                    continue
                if is_fusionsolar_session_expired_error(exc) and not session_retry_used:
                    session_retry_used = True
                    refresh_session_after_expiry(exc, f"month-cycle-month:{month_start.isoformat()}:chunk:{chunk_index}")
                    continue
                raise

    for target in sorted(processed_dates):
        for selected_asset_id in asset_ids:
            recalculate_performance_references(
                conn,
                period_type="day",
                period_date=target,
                asset_id=selected_asset_id,
                provider=provider,
                today_value=today_value,
            )
    for selected_asset_id in asset_ids:
        recalculate_performance_references(
            conn,
            period_type="month",
            period_date=month_start,
            asset_id=selected_asset_id,
            provider=provider,
            today_value=today_value,
        )
    conn.commit()
    summary["status"] = "completed"
    return summary


def create_integration_run(conn: sqlite3.Connection, provider: str, trigger_type: str) -> int:
    cursor = conn.execute(
        """
        INSERT INTO integration_sync_runs (provider, started_at, trigger_type, status)
        VALUES (?, ?, ?, ?)
        """,
        (provider, datetime.now().isoformat(timespec="seconds"), trigger_type, "running"),
    )
    return int(cursor.lastrowid)


def finalize_integration_run(
    conn: sqlite3.Connection,
    run_id: int,
    *,
    status: str,
    matched_count: int,
    unresolved_count: int,
    auto_resolved_count: int,
    error_message: str = "",
    summary_json: dict[str, Any] | None = None,
) -> None:
    conn.execute(
        """
        UPDATE integration_sync_runs
        SET finished_at = ?, status = ?, matched_count = ?, unresolved_count = ?,
            auto_resolved_count = ?, error_message = ?, summary_json = ?
        WHERE id = ?
        """,
        (
            datetime.now().isoformat(timespec="seconds"),
            status,
            matched_count,
            unresolved_count,
            auto_resolved_count,
            error_message,
            json.dumps(summary_json or {}, ensure_ascii=True),
            run_id,
        ),
    )


def create_or_update_asset_integration(
    conn: sqlite3.Connection,
    asset_id: int,
    provider: str,
    external_id: str,
    external_name: str,
    status: str,
) -> None:
    existing = None
    if external_id:
        existing = conn.execute(
            "SELECT id FROM asset_integrations WHERE provider = ? AND external_id = ?",
            (provider, external_id),
        ).fetchone()
    if existing:
        conn.execute(
            """
            UPDATE asset_integrations
            SET asset_id = ?, external_name = ?, enabled = 1, last_sync_at = ?, last_status = ?, last_error = ''
            WHERE id = ?
            """,
            (asset_id, external_name, datetime.now().isoformat(timespec="seconds"), status, existing["id"]),
        )
        return

    candidate = conn.execute(
        """
        SELECT id
        FROM asset_integrations
        WHERE provider = ? AND asset_id = ?
        LIMIT 1
        """,
        (provider, asset_id),
    ).fetchone()
    if candidate:
        conn.execute(
            """
            UPDATE asset_integrations
            SET external_id = ?, external_name = ?, enabled = 1, last_sync_at = ?, last_status = ?, last_error = ''
            WHERE id = ?
            """,
            (external_id or None, external_name, datetime.now().isoformat(timespec="seconds"), status, candidate["id"]),
        )
        return

    conn.execute(
        """
        INSERT INTO asset_integrations (asset_id, provider, external_id, external_name, enabled, last_sync_at, last_status, last_error)
        VALUES (?, ?, ?, ?, 1, ?, ?, '')
        """,
        (
            asset_id,
            provider,
            external_id or None,
            external_name,
            datetime.now().isoformat(timespec="seconds"),
            status,
        ),
    )


def upsert_integration_unresolved(
    conn: sqlite3.Connection,
    *,
    provider: str,
    run_id: int,
    external_id: str,
    external_name: str,
    status: str,
    payload: dict[str, Any],
) -> None:
    normalized_name = normalize_name(external_name)
    suggested_asset_id = find_suggested_asset_id(conn, external_name)
    existing = conn.execute(
        """
        SELECT id
        FROM integration_unresolved
        WHERE provider = ? AND normalized_name = ? AND resolution_status = 'pending'
        LIMIT 1
        """,
        (provider, normalized_name),
    ).fetchone()
    if existing:
        conn.execute(
            """
            UPDATE integration_unresolved
            SET sync_run_id = ?, external_id = ?, external_status = ?, payload_json = ?, suggested_asset_id = ?, created_at = ?
            WHERE id = ?
            """,
            (
                run_id,
                external_id or None,
                status,
                json.dumps(payload, ensure_ascii=True),
                suggested_asset_id,
                datetime.now().isoformat(timespec="seconds"),
                existing["id"],
            ),
        )
        return
    conn.execute(
        """
        INSERT INTO integration_unresolved (
            provider, sync_run_id, external_id, external_name, normalized_name, external_status,
            payload_json, suggested_asset_id, resolution_status, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'pending', ?)
        """,
        (
            provider,
            run_id,
            external_id or None,
            external_name,
            normalized_name,
            status,
            json.dumps(payload, ensure_ascii=True),
            suggested_asset_id,
            datetime.now().isoformat(timespec="seconds"),
        ),
    )


def get_latest_availability_by_asset(conn: sqlite3.Connection, asset_id: int) -> dict[str, Any] | None:
    row = conn.execute(
        """
        SELECT *
        FROM availability_daily
        WHERE asset_id = ?
        ORDER BY period_date DESC, id DESC
        LIMIT 1
        """,
        (asset_id,),
    ).fetchone()
    return dict(row) if row else None


def get_dashboard_availability_summary(conn: sqlite3.Connection) -> dict[str, Any]:
    rows = query_all(
        conn,
        """
        SELECT ad.*
        FROM availability_daily ad
        JOIN (
            SELECT asset_id, MAX(period_date || 'T' || printf('%09d', id)) AS marker
            FROM availability_daily
            GROUP BY asset_id
        ) latest
          ON latest.asset_id = ad.asset_id
         AND latest.marker = ad.period_date || 'T' || printf('%09d', ad.id)
        """
    )
    pct_values = [row["inverter_availability_pct"] for row in rows if row["inverter_availability_pct"] is not None]
    affected_values = [row["affected_power_kw"] for row in rows if row["affected_power_kw"] is not None]
    return {
        "average_inverter_availability_pct": round(sum(pct_values) / len(pct_values), 2) if pct_values else None,
        "unavailable_inverters": sum(int(row["unavailable_inverters"] or 0) for row in rows),
        "no_communication_devices": sum(int(row["no_communication_devices"] or 0) for row in rows),
        "affected_power_kw": round(sum(float(value) for value in affected_values), 2) if affected_values else None,
        "string_availability_pct": (
            round(
                sum(int(row["available_strings"] or 0) for row in rows)
                / sum(int(row["total_strings"] or 0) for row in rows)
                * 100,
                2,
            )
            if sum(int(row["total_strings"] or 0) for row in rows)
            else None
        ),
        "unavailable_strings": sum(int(row["unavailable_strings"] or 0) for row in rows),
    }


def get_latest_device_rows_for_asset(conn: sqlite3.Connection, asset_id: int) -> list[dict[str, Any]]:
    return query_all(
        conn,
        """
        SELECT
            pd.device_name,
            pd.rated_power_kw,
            pd.last_seen_at,
            drs.*
        FROM provider_devices pd
        LEFT JOIN device_realtime_snapshots drs
          ON drs.id = (
              SELECT latest.id
              FROM device_realtime_snapshots latest
              WHERE latest.provider_device_id = pd.id
              ORDER BY latest.collected_at DESC, latest.id DESC
              LIMIT 1
          )
        WHERE pd.asset_id = ? AND pd.enabled = 1
        ORDER BY pd.device_name COLLATE NOCASE, pd.id
        """,
        (asset_id,),
    )


def upsert_provider_device(conn: sqlite3.Connection, asset_id: int, provider: str, row: dict[str, Any]) -> int:
    now = datetime.now().isoformat(timespec="seconds")
    enabled = int(row.get("enabled", 1))
    existing = conn.execute(
        "SELECT id FROM provider_devices WHERE provider = ? AND external_device_id = ?",
        (provider, row["external_device_id"]),
    ).fetchone()
    payload_json = json.dumps(row["payload"], ensure_ascii=True)
    if existing:
        conn.execute(
            """
            UPDATE provider_devices
            SET asset_id = ?, station_code = ?, dev_dn = ?, sn = ?, device_name = ?, dev_type_id = ?,
                model = ?, rated_power_kw = ?, enabled = ?, payload_json = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                asset_id,
                row["station_code"],
                row["dev_dn"],
                row["sn"],
                row["device_name"],
                row["dev_type_id"],
                row["model"],
                row["rated_power_kw"],
                enabled,
                payload_json,
                now,
                existing["id"],
            ),
        )
        return int(existing["id"])
    cursor = conn.execute(
        """
        INSERT INTO provider_devices (
            asset_id, provider, station_code, external_device_id, dev_dn, sn, device_name, dev_type_id,
            model, rated_power_kw, enabled, payload_json, created_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            asset_id,
            provider,
            row["station_code"],
            row["external_device_id"],
            row["dev_dn"],
            row["sn"],
            row["device_name"],
            row["dev_type_id"],
            row["model"],
            row["rated_power_kw"],
            enabled,
            payload_json,
            now,
            now,
        ),
    )
    return int(cursor.lastrowid)


def prepare_fusionsolar_inverter_history_context(
    conn: sqlite3.Connection,
    *,
    force_login: bool = False,
    history_call_delay_seconds: float = 0,
    sleeper: Any = time.sleep,
) -> dict[str, Any]:
    provider = INTEGRATION_PROVIDER_FUSIONSOLAR
    config = get_integration_config(conn, provider)
    if config is None or not config["enabled"]:
        raise ValueError("Configuracao FusionSolar indisponivel.")
    endpoints = get_fusionsolar_endpoint_config(config)
    mappings = query_all(
        conn,
        """
        SELECT asset_id, external_id
        FROM asset_integrations
        WHERE provider = ? AND enabled = 1 AND COALESCE(external_id, '') != ''
        """,
        (provider,),
    )
    station_to_asset = {str(row["external_id"]): int(row["asset_id"]) for row in mappings}
    if not station_to_asset:
        return {"provider": provider, "config": config, "endpoints": endpoints, "session": None, "devices": []}
    session, _ = get_fusionsolar_session(config, force_login=force_login)
    raw_devices = fetch_fusionsolar_device_list(
        session,
        base_url=endpoints["base_url"],
        endpoint=endpoints["device_list_endpoint"],
        station_codes=sorted(station_to_asset),
    )
    devices: list[dict[str, Any]] = []
    for raw_row in raw_devices:
        device = normalize_fusionsolar_device_identity(raw_row)
        if device["dev_type_id"] not in FUSIONSOLAR_INVERTER_DEVICE_TYPE_IDS:
            continue
        asset_id = station_to_asset.get(str(device["station_code"] or ""))
        if not asset_id or not device["external_device_id"]:
            continue
        device["asset_id"] = asset_id
        device["enabled"] = 0 if is_removed_inverter_name(device["device_name"]) else 1
        device["payload"] = raw_row
        device["provider_device_id"] = upsert_provider_device(conn, asset_id, provider, device)
        if not device["enabled"]:
            logging.info(
                "FusionSolar inverter excluded because it is marked as removed: asset_id=%s inverter_id=%s name=%s",
                asset_id,
                device["external_device_id"],
                device["device_name"],
            )
            continue
        devices.append(device)
    conn.commit()
    return {
        "provider": provider,
        "config": config,
        "endpoints": endpoints,
        "session": session,
        "devices": devices,
        "history_call_delay_seconds": history_call_delay_seconds,
        "sleeper": sleeper,
    }


def sync_fusionsolar_inverter_availability_for_date(
    conn: sqlite3.Connection,
    target_date: date,
    *,
    context: dict[str, Any] | None = None,
) -> dict[str, Any]:
    if target_date >= date.today():
        raise ValueError("A disponibilidade temporal requer um dia fechado.")
    sync_context = context or prepare_fusionsolar_inverter_history_context(conn)
    devices = sync_context["devices"]
    if not devices:
        return {"date": target_date.isoformat(), "samples": 0, "plants": 0, "inverters": 0}
    try:
        history_rows = fetch_fusionsolar_device_history(
            sync_context["session"],
            base_url=sync_context["endpoints"]["base_url"],
            endpoint=sync_context["endpoints"]["device_history_endpoint"],
            devices=devices,
            target_date=target_date,
            call_delay_seconds=float(sync_context.get("history_call_delay_seconds") or 0),
            sleeper=sync_context.get("sleeper") or time.sleep,
        )
    except Exception:
        logging.exception(
            "FusionSolar device history request failed: target_date=%s endpoint=%s inverter_count=%s",
            target_date,
            sync_context["endpoints"]["device_history_endpoint"],
            len(devices),
        )
        raise

    now = datetime.now().isoformat(timespec="seconds")
    for sample in history_rows:
        conn.execute(
            """
            INSERT INTO inverter_power_samples (
                asset_id, provider, external_station_id, inverter_id, inverter_name, inverter_power_kw,
                sample_time, active_power_kw, raw_payload, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(provider, inverter_id, sample_time) DO UPDATE SET
                asset_id = excluded.asset_id,
                external_station_id = excluded.external_station_id,
                inverter_name = excluded.inverter_name,
                inverter_power_kw = excluded.inverter_power_kw,
                active_power_kw = excluded.active_power_kw,
                raw_payload = excluded.raw_payload
            """,
            (
                sample["asset_id"],
                sync_context["provider"],
                sample["station_code"],
                sample["external_device_id"],
                sample["device_name"],
                sample["rated_power_kw"],
                sample["sample_time"].isoformat(timespec="seconds"),
                sample["active_power_kw"],
                json.dumps(sample["raw_payload"], ensure_ascii=True),
                now,
            ),
        )

    plants_written = 0
    devices_by_asset: dict[int, list[dict[str, Any]]] = {}
    for device in devices:
        devices_by_asset.setdefault(int(device["asset_id"]), []).append(device)
    for asset_id, plant_devices in devices_by_asset.items():
        stored_samples = query_all(
            conn,
            """
            SELECT inverter_id, sample_time, active_power_kw
            FROM inverter_power_samples
            WHERE asset_id = ? AND provider = ? AND sample_time >= ? AND sample_time < ?
            """,
            (
                asset_id,
                sync_context["provider"],
                datetime.combine(target_date, datetime.min.time()).isoformat(timespec="seconds"),
                datetime.combine(target_date + timedelta(days=1), datetime.min.time()).isoformat(timespec="seconds"),
            ),
        )
        samples_by_inverter: dict[str, list[dict[str, Any]]] = {}
        valid_slots: set[datetime] = set()
        for row in stored_samples:
            sample_time = parse_datetime_value(row["sample_time"])
            if sample_time is None:
                continue
            sample = {"sample_time": sample_time, "active_power_kw": row["active_power_kw"]}
            samples_by_inverter.setdefault(str(row["inverter_id"]), []).append(sample)
            if is_inverter_available(row["active_power_kw"]):
                valid_slots.add(inverter_availability_slot(sample_time))

        conn.execute(
            "DELETE FROM inverter_availability_daily WHERE asset_id = ? AND provider = ? AND availability_date = ?",
            (asset_id, sync_context["provider"], target_date.isoformat()),
        )
        inverter_results: list[dict[str, Any]] = []
        for device in plant_devices:
            device_samples = samples_by_inverter.get(str(device["external_device_id"]), [])
            if not device_samples:
                logging.info(
                    "FusionSolar inverter excluded because it has no samples: target_date=%s asset_id=%s inverter_id=%s name=%s",
                    target_date,
                    asset_id,
                    device["external_device_id"],
                    device["device_name"],
                )
                continue
            result = calculate_inverter_daily_availability(
                device_samples,
                valid_slots,
            )
            result.update(
                {
                    "asset_id": asset_id,
                    "inverter_id": device["external_device_id"],
                    "inverter_name": device["device_name"],
                    "inverter_power_kw": device["rated_power_kw"],
                }
            )
            inverter_results.append(result)
            conn.execute(
                """
                INSERT INTO inverter_availability_daily (
                    asset_id, provider, availability_date, inverter_id, inverter_name, inverter_power_kw,
                    valid_slots, available_slots, unavailable_slots, availability_pct, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(provider, inverter_id, availability_date) DO UPDATE SET
                    asset_id = excluded.asset_id,
                    inverter_name = excluded.inverter_name,
                    inverter_power_kw = excluded.inverter_power_kw,
                    valid_slots = excluded.valid_slots,
                    available_slots = excluded.available_slots,
                    unavailable_slots = excluded.unavailable_slots,
                    availability_pct = excluded.availability_pct,
                    updated_at = excluded.updated_at
                """,
                (
                    asset_id,
                    sync_context["provider"],
                    target_date.isoformat(),
                    device["external_device_id"],
                    device["device_name"],
                    device["rated_power_kw"],
                    result["valid_slots"],
                    result["available_slots"],
                    result["unavailable_slots"],
                    result["availability_pct"],
                    now,
                    now,
                ),
            )
        weighted_pct = calculate_weighted_plant_availability(inverter_results)
        tolerated_valid_slots = apply_inverter_edge_tolerance(valid_slots)
        conn.execute(
            """
            INSERT INTO plant_availability_daily (
                asset_id, provider, availability_date, valid_slots, weighted_availability_pct,
                inverter_count, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(provider, asset_id, availability_date) DO UPDATE SET
                valid_slots = excluded.valid_slots,
                weighted_availability_pct = excluded.weighted_availability_pct,
                inverter_count = excluded.inverter_count,
                updated_at = excluded.updated_at
            """,
            (
                asset_id,
                sync_context["provider"],
                target_date.isoformat(),
                len(tolerated_valid_slots),
                weighted_pct,
                len(inverter_results),
                now,
                now,
            ),
        )
        plants_written += 1
        logging.info(
            "FusionSolar inverter availability calculated: target_date=%s asset_id=%s inverters=%s valid_slots=%s availability_pct=%s",
            target_date,
            asset_id,
            len(inverter_results),
            len(tolerated_valid_slots),
            weighted_pct,
        )
    conn.commit()
    return {
        "date": target_date.isoformat(),
        "samples": len(history_rows),
        "plants": plants_written,
        "inverters": len(devices),
    }


def sync_fusionsolar_inverter_availability_range(
    conn: sqlite3.Connection,
    from_date: date,
    to_date: date,
) -> dict[str, Any]:
    context = prepare_fusionsolar_inverter_history_context(conn)
    totals = {"days": 0, "samples": 0, "plants": 0, "inverters": len(context["devices"])}
    current = from_date
    while current <= to_date:
        result = sync_fusionsolar_inverter_availability_for_date(conn, current, context=context)
        totals["days"] += 1
        totals["samples"] += int(result["samples"])
        totals["plants"] += int(result["plants"])
        current += timedelta(days=1)
    return totals


def run_fusionsolar_inverter_availability_backfill(
    conn: sqlite3.Connection,
    *,
    from_date: date,
    to_date: date,
    sleeper: Any = time.sleep,
    history_call_delay_seconds: float = FUSIONSOLAR_PERFORMANCE_KPI_DELAY_SECONDS,
    max_wait_cycles: int = 24,
) -> dict[str, Any]:
    if from_date > to_date or to_date >= date.today():
        raise ValueError("O backfill WAT requer um intervalo valido de dias fechados.")
    summary: dict[str, Any] = {
        "days": 0,
        "samples": 0,
        "plants": 0,
        "inverters": 0,
        "api_errors": 0,
        "wait_cycles": 0,
        "resume_hint": from_date.isoformat(),
        "stopped_reason": "",
    }

    def wait_after_rate_limit(reason: str, resume_date: date) -> bool:
        summary["wait_cycles"] += 1
        summary["resume_hint"] = resume_date.isoformat()
        if summary["wait_cycles"] > max_wait_cycles:
            summary["stopped_reason"] = f"Limite FusionSolar repetido demasiadas vezes. Ultimo estado: {reason}"
            logging.warning(
                "FusionSolar WAT backfill stopped after repeated cooldowns: resume_date=%s wait_cycles=%s reason=%s",
                resume_date,
                summary["wait_cycles"],
                reason,
            )
            return False
        conn.commit()
        seconds = fusionsolar_cooldown_sleep_seconds(conn)
        logging.warning(
            "FusionSolar WAT backfill waiting for API cooldown: resume_date=%s seconds=%s wait_cycle=%s",
            resume_date,
            seconds,
            summary["wait_cycles"],
        )
        sleeper(seconds)
        return True

    def prepare_context(force_login: bool = False) -> dict[str, Any] | None:
        while True:
            try:
                return prepare_fusionsolar_inverter_history_context(
                    conn,
                    force_login=force_login,
                    history_call_delay_seconds=history_call_delay_seconds,
                    sleeper=sleeper,
                )
            except Exception as exc:
                summary["api_errors"] += 1
                if is_fusionsolar_rate_limit_error(exc):
                    reason = mark_fusionsolar_performance_rate_limited(conn)
                    if not wait_after_rate_limit(reason, from_date):
                        return None
                    force_login = True
                    continue
                if is_fusionsolar_session_expired_error(exc):
                    force_login = True
                    continue
                raise

    cooldown_reason = get_fusionsolar_performance_cooldown_reason(conn)
    if cooldown_reason and not wait_after_rate_limit(cooldown_reason, from_date):
        return summary
    context = prepare_context()
    if context is None:
        return summary
    summary["inverters"] = len(context["devices"])

    current = from_date
    while current <= to_date:
        session_retry_used = False
        while True:
            try:
                logging.info("FusionSolar WAT backfill day started: target_date=%s", current)
                result = sync_fusionsolar_inverter_availability_for_date(conn, current, context=context)
                stored_counts = conn.execute(
                    """
                    SELECT
                        (SELECT COUNT(*) FROM inverter_power_samples
                         WHERE provider = ? AND sample_time >= ? AND sample_time < ?) AS power_samples,
                        (SELECT COUNT(*) FROM inverter_availability_daily
                         WHERE provider = ? AND availability_date = ?) AS inverter_daily_rows,
                        (SELECT COUNT(*) FROM plant_availability_daily
                         WHERE provider = ? AND availability_date = ?) AS plant_daily_rows
                    """,
                    (
                        INTEGRATION_PROVIDER_FUSIONSOLAR,
                        datetime.combine(current, datetime.min.time()).isoformat(timespec="seconds"),
                        datetime.combine(current + timedelta(days=1), datetime.min.time()).isoformat(timespec="seconds"),
                        INTEGRATION_PROVIDER_FUSIONSOLAR,
                        current.isoformat(),
                        INTEGRATION_PROVIDER_FUSIONSOLAR,
                        current.isoformat(),
                    ),
                ).fetchone()
                summary["days"] += 1
                summary["samples"] += int(result["samples"])
                summary["plants"] += int(result["plants"])
                summary["resume_hint"] = (current + timedelta(days=1)).isoformat()
                conn.commit()
                logging.info(
                    "FusionSolar WAT backfill day completed: target_date=%s fetched_samples=%s "
                    "stored_power_samples=%s stored_inverter_daily=%s stored_plant_daily=%s plants=%s inverters=%s",
                    current,
                    result["samples"],
                    stored_counts["power_samples"],
                    stored_counts["inverter_daily_rows"],
                    stored_counts["plant_daily_rows"],
                    result["plants"],
                    result["inverters"],
                )
                break
            except Exception as exc:
                summary["api_errors"] += 1
                if is_fusionsolar_rate_limit_error(exc):
                    reason = mark_fusionsolar_performance_rate_limited(conn)
                    if not wait_after_rate_limit(reason, current):
                        return summary
                    refreshed = prepare_context(force_login=True)
                    if refreshed is None:
                        return summary
                    context = refreshed
                    continue
                if is_fusionsolar_session_expired_error(exc) and not session_retry_used:
                    session_retry_used = True
                    invalidate_fusionsolar_session(context["config"])
                    refreshed = prepare_context(force_login=True)
                    if refreshed is None:
                        return summary
                    context = refreshed
                    continue
                raise
        current += timedelta(days=1)
    summary["resume_hint"] = ""
    return summary


def recalculate_stored_inverter_availability(
    conn: sqlite3.Connection,
    from_date: date,
    to_date: date,
    *,
    asset_id: int | None = None,
) -> dict[str, int]:
    provider = INTEGRATION_PROVIDER_FUSIONSOLAR
    conditions = ["provider = ?", "sample_time >= ?", "sample_time < ?"]
    params: list[Any] = [
        provider,
        datetime.combine(from_date, datetime.min.time()).isoformat(timespec="seconds"),
        datetime.combine(to_date + timedelta(days=1), datetime.min.time()).isoformat(timespec="seconds"),
    ]
    if asset_id is not None:
        conditions.append("asset_id = ?")
        params.append(asset_id)
    sample_dates = conn.execute(
        f"""
        SELECT DISTINCT asset_id, substr(sample_time, 1, 10) AS sample_date
        FROM inverter_power_samples
        WHERE {' AND '.join(conditions)}
        ORDER BY sample_date, asset_id
        """,
        params,
    ).fetchall()
    now = datetime.now().isoformat(timespec="seconds")
    totals = {"days": 0, "plants": 0, "inverters": 0}
    for date_row in sample_dates:
        current_asset_id = int(date_row["asset_id"])
        target_date = parse_date_value(date_row["sample_date"])
        if target_date is None:
            continue
        devices = query_all(
            conn,
            """
            SELECT external_device_id, device_name, rated_power_kw, model
            FROM provider_devices
            WHERE asset_id = ? AND provider = ? AND enabled = 1 AND dev_type_id IN (1, 38)
            ORDER BY device_name COLLATE NOCASE, external_device_id
            """,
            (current_asset_id, provider),
        )
        samples = query_all(
            conn,
            """
            SELECT inverter_id, sample_time, active_power_kw
            FROM inverter_power_samples
            WHERE asset_id = ? AND provider = ? AND sample_time >= ? AND sample_time < ?
            ORDER BY sample_time
            """,
            (
                current_asset_id,
                provider,
                datetime.combine(target_date, datetime.min.time()).isoformat(timespec="seconds"),
                datetime.combine(target_date + timedelta(days=1), datetime.min.time()).isoformat(timespec="seconds"),
            ),
        )
        samples_by_inverter: dict[str, list[dict[str, Any]]] = {}
        valid_slots: set[datetime] = set()
        for sample_row in samples:
            sample_time = parse_datetime_value(sample_row["sample_time"])
            if sample_time is None:
                continue
            sample = {"sample_time": sample_time, "active_power_kw": sample_row["active_power_kw"]}
            samples_by_inverter.setdefault(str(sample_row["inverter_id"]), []).append(sample)
            if is_inverter_available(sample_row["active_power_kw"]):
                valid_slots.add(inverter_availability_slot(sample_time))

        conn.execute(
            "DELETE FROM inverter_availability_daily WHERE asset_id = ? AND provider = ? AND availability_date = ?",
            (current_asset_id, provider, target_date.isoformat()),
        )
        inverter_results: list[dict[str, Any]] = []
        for device in devices:
            if is_removed_inverter_name(device["device_name"]):
                continue
            inverter_id = str(device["external_device_id"] or "")
            device_samples = samples_by_inverter.get(inverter_id, [])
            if not inverter_id or not device_samples:
                continue
            inverter_power_kw = (
                parse_float_value(device["rated_power_kw"])
                or infer_inverter_power_from_model(device["model"])
            )
            result = calculate_inverter_daily_availability(device_samples, valid_slots)
            result["inverter_power_kw"] = inverter_power_kw
            inverter_results.append(result)
            conn.execute(
                """
                INSERT INTO inverter_availability_daily (
                    asset_id, provider, availability_date, inverter_id, inverter_name, inverter_power_kw,
                    valid_slots, available_slots, unavailable_slots, availability_pct, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    current_asset_id,
                    provider,
                    target_date.isoformat(),
                    inverter_id,
                    device["device_name"],
                    inverter_power_kw,
                    result["valid_slots"],
                    result["available_slots"],
                    result["unavailable_slots"],
                    result["availability_pct"],
                    now,
                    now,
                ),
            )
        weighted_pct = calculate_weighted_plant_availability(inverter_results)
        tolerated_valid_slots = apply_inverter_edge_tolerance(valid_slots)
        conn.execute(
            """
            INSERT INTO plant_availability_daily (
                asset_id, provider, availability_date, valid_slots, weighted_availability_pct,
                inverter_count, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(provider, asset_id, availability_date) DO UPDATE SET
                valid_slots = excluded.valid_slots,
                weighted_availability_pct = excluded.weighted_availability_pct,
                inverter_count = excluded.inverter_count,
                updated_at = excluded.updated_at
            """,
            (
                current_asset_id,
                provider,
                target_date.isoformat(),
                len(tolerated_valid_slots),
                weighted_pct,
                len(inverter_results),
                now,
                now,
            ),
        )
        totals["plants"] += 1
        totals["inverters"] += len(inverter_results)
        totals["days"] += 1
    conn.commit()
    return totals


def run_fusionsolar_device_availability_sync(
    conn: sqlite3.Connection,
    provider: str,
    trigger_type: str = "manual",
) -> dict[str, Any]:
    config = get_integration_config(conn, provider)
    if config is None or not config["enabled"]:
        raise ValueError(f"Configuracao {provider} indisponivel.")
    endpoints = get_fusionsolar_endpoint_config(config)
    mappings = query_all(
        conn,
        """
        SELECT asset_id, external_id
        FROM asset_integrations
        WHERE provider = ? AND enabled = 1 AND COALESCE(external_id, '') != ''
        """,
        (provider,),
    )
    station_to_asset = {str(row["external_id"]): int(row["asset_id"]) for row in mappings}
    station_codes = sorted(station_to_asset)
    if not station_codes:
        return {"devices": 0, "snapshots": 0, "assets": 0}

    last_error: Exception | None = None
    for attempt in range(2):
        try:
            session, _ = get_fusionsolar_session(config, force_login=attempt == 1)
            device_rows = fetch_fusionsolar_device_list(
                session,
                base_url=endpoints["base_url"],
                endpoint=endpoints["device_list_endpoint"],
                station_codes=station_codes,
            )
            break
        except Exception as exc:
            invalidate_fusionsolar_session(config)
            last_error = exc
            if attempt == 1:
                raise
    else:
        raise last_error or ValueError("Falha desconhecida no FusionSolar.")

    tracked: list[dict[str, Any]] = []
    for raw_row in device_rows:
        normalized = normalize_fusionsolar_device_identity(raw_row)
        if normalized["dev_type_id"] not in FUSIONSOLAR_INVERTER_DEVICE_TYPE_IDS:
            continue
        if not normalized["station_code"] or not normalized["external_device_id"]:
            continue
        asset_id = station_to_asset.get(normalized["station_code"])
        if not asset_id:
            continue
        normalized["payload"] = raw_row
        normalized["enabled"] = 0 if is_removed_inverter_name(normalized["device_name"]) else 1
        normalized["provider_device_id"] = upsert_provider_device(conn, asset_id, provider, normalized)
        if not normalized["enabled"]:
            continue
        normalized["asset_id"] = asset_id
        tracked.append(normalized)

    realtime_map = fetch_fusionsolar_device_realtime_map(
        session,
        base_url=endpoints["base_url"],
        endpoint=endpoints["device_real_time_endpoint"],
        devices=tracked,
    )
    collected_at_dt = datetime.now()
    collected_at = collected_at_dt.isoformat(timespec="seconds")
    snapshots_by_asset: dict[int, list[dict[str, Any]]] = {}
    for device in tracked:
        realtime = next(
            (realtime_map[key] for key in (device["external_device_id"], device["dev_dn"], device["sn"]) if key and key in realtime_map),
            {},
        )
        data_map = realtime.get("dataItemMap") if isinstance(realtime.get("dataItemMap"), dict) else realtime
        seen_dt = parse_datetime_value(first_non_empty(realtime, ["collectTime", "collectedAt", "lastSeen", "last_seen_at"]))
        has_recent_data = True if realtime and seen_dt is None else bool(
            seen_dt and collected_at_dt - seen_dt <= timedelta(minutes=DEFAULT_DEVICE_COMMUNICATION_THRESHOLD_MINUTES)
        )
        availability_status = classify_fusionsolar_inverter_availability(
            {"inverter_state": first_non_empty(data_map, ["inverter_state", "inverterState"])},
            has_recent_data=has_recent_data,
        )
        communication_status = "recent" if has_recent_data else "stale"
        currents, voltages = parse_fusionsolar_pv_inputs(realtime)
        expected_string_indexes = learn_expected_strings_from_voltage(
            conn,
            device["provider_device_id"],
            voltages,
            collected_at,
        )
        pv_health = calculate_pv_input_health(
            currents,
            voltages,
            expected_string_indexes=expected_string_indexes,
        )
        snapshot = {
            **device,
            "availability_status": availability_status,
            "communication_status": communication_status,
            "rated_power_kw": device["rated_power_kw"],
            **pv_health,
        }
        conn.execute(
            """
            INSERT INTO device_realtime_snapshots (
                provider_device_id, asset_id, provider, station_code, collected_at, inverter_state,
                active_power_kw, day_energy_kwh, availability_status, communication_status,
                string_available_count, string_total_count, pv_current_json, pv_voltage_json, payload_json, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                device["provider_device_id"],
                device["asset_id"],
                provider,
                device["station_code"],
                collected_at,
                parse_int_value(first_non_empty(data_map, ["inverter_state", "inverterState"])),
                normalize_power_to_kw(first_non_empty(data_map, ["active_power", "activePower"])),
                parse_float_value(first_non_empty(data_map, ["day_cap", "dayEnergy", "day_energy"])),
                availability_status,
                communication_status,
                pv_health["available_strings"],
                pv_health["total_strings"],
                json.dumps(currents, ensure_ascii=True) if currents else None,
                json.dumps(voltages, ensure_ascii=True) if voltages else None,
                json.dumps(realtime, ensure_ascii=True),
                collected_at,
            ),
        )
        conn.execute(
            "UPDATE provider_devices SET last_seen_at = COALESCE(?, last_seen_at), updated_at = ? WHERE id = ?",
            (seen_dt.isoformat(timespec="seconds") if seen_dt else collected_at if realtime else None, collected_at, device["provider_device_id"]),
        )
        snapshots_by_asset.setdefault(device["asset_id"], []).append(snapshot)

    for asset_id, rows in snapshots_by_asset.items():
        summary = calculate_asset_availability(rows)
        now = datetime.now().isoformat(timespec="seconds")
        conn.execute(
            """
            INSERT INTO availability_daily (
                asset_id, provider, period_date, inverter_availability_pct, capacity_availability_pct,
                communication_availability_pct, string_availability_pct, available_inverters, total_inverters, unavailable_inverters,
                no_communication_devices, available_strings, total_strings, unavailable_strings, affected_power_kw,
                payload_json, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(asset_id, provider, period_date) DO UPDATE SET
                inverter_availability_pct = excluded.inverter_availability_pct,
                capacity_availability_pct = excluded.capacity_availability_pct,
                communication_availability_pct = excluded.communication_availability_pct,
                string_availability_pct = excluded.string_availability_pct,
                available_inverters = excluded.available_inverters,
                total_inverters = excluded.total_inverters,
                unavailable_inverters = excluded.unavailable_inverters,
                no_communication_devices = excluded.no_communication_devices,
                available_strings = excluded.available_strings,
                total_strings = excluded.total_strings,
                unavailable_strings = excluded.unavailable_strings,
                affected_power_kw = excluded.affected_power_kw,
                payload_json = excluded.payload_json,
                updated_at = excluded.updated_at
            """,
            (
                asset_id,
                provider,
                date.today().isoformat(),
                summary["inverter_availability_pct"],
                summary["capacity_availability_pct"],
                summary["communication_availability_pct"],
                summary["string_availability_pct"],
                summary["available_inverters"],
                summary["total_inverters"],
                summary["unavailable_inverters"],
                summary["no_communication_devices"],
                summary["available_strings"],
                summary["total_strings"],
                summary["unavailable_strings"],
                summary["affected_power_kw"],
                json.dumps(summary, ensure_ascii=True),
                now,
                now,
            ),
        )
    conn.commit()
    return {"devices": len(tracked), "snapshots": sum(len(rows) for rows in snapshots_by_asset.values()), "assets": len(snapshots_by_asset)}


def run_integration_sync(conn: sqlite3.Connection, provider: str, trigger_type: str = "manual") -> dict[str, Any]:
    return run_fusionsolar_sync(conn, provider, trigger_type=trigger_type)


def run_fusionsolar_sync(conn: sqlite3.Connection, provider: str, trigger_type: str = "manual") -> dict[str, Any]:
    with FUSIONSOLAR_SYNC_LOCK:
        config = get_integration_config(conn, provider)
        if config is None:
            raise ValueError(f"Configuracao {provider} nao encontrada.")
        if not config["enabled"]:
            raise ValueError(f"A integracao {provider} esta desativada.")

        run_id = create_integration_run(conn, provider, trigger_type)
        batch_id = create_monitoring_batch(
            conn,
            record_date=date.today().isoformat(),
            default_notes=f"Sync {provider} ({trigger_type})",
            raw_input="",
            source=provider,
        )

        try:
            result = run_provider_check(conn, provider, dry_run=True)
            rows = result["rows"]
            matched = 0
            unresolved = 0
            auto_resolved = 0
            synced_asset_ids: set[int] = set()
            alert_events: list[dict[str, Any]] = []
            now = datetime.now()

            for row in rows:
                external_id = row["external_id"]
                external_name = row["external_name"]
                status = row["status"]

                mapped_asset = None
                if external_id:
                    mapped_asset = conn.execute(
                        """
                        SELECT ai.asset_id
                        FROM asset_integrations ai
                        WHERE ai.provider = ? AND ai.external_id = ? AND ai.enabled = 1
                        LIMIT 1
                        """,
                        (provider, external_id),
                    ).fetchone()
                if mapped_asset is None:
                    mapped_asset = conn.execute(
                        """
                        SELECT ai.asset_id
                        FROM asset_integrations ai
                        WHERE ai.provider = ? AND ai.external_name = ? AND ai.enabled = 1
                        LIMIT 1
                        """,
                        (provider, external_name),
                    ).fetchone()
                asset_id = int(mapped_asset["asset_id"]) if mapped_asset else (find_asset_id(conn, external_name) or 0)

                if asset_id:
                    synced_asset_ids.add(asset_id)
                    previous = get_latest_monitoring_row(conn, asset_id)
                    duplicate_latest = (
                        previous is not None
                        and previous["status"] == status
                        and previous["record_date"] == date.today().isoformat()
                        and previous["source"] == provider
                    )
                    if not duplicate_latest:
                        conn.execute(
                            """
                            INSERT INTO monitoring_records (asset_id, status, record_date, notes, source, batch_id)
                            VALUES (?, ?, ?, ?, ?, ?)
                            """,
                            (
                                asset_id,
                                status,
                                date.today().isoformat(),
                                f"Sync {provider}: {row['notes']}",
                                provider,
                                batch_id,
                            ),
                        )
                        event = build_monitoring_alert_event(
                            conn,
                            asset_id=asset_id,
                            previous_status=previous["status"] if previous else "",
                            current_status=status,
                            happened_at=now.isoformat(timespec="seconds"),
                            alarm_context=row,
                        )
                        if event:
                            alert_events.append(event)
                    create_or_update_asset_integration(conn, asset_id, provider, external_id, external_name, status)
                    matched += 1
                else:
                    upsert_integration_unresolved(
                        conn,
                        provider=provider,
                        run_id=run_id,
                        external_id=external_id,
                        external_name=external_name,
                        status=status,
                        payload=row["payload"],
                    )
                    unresolved += 1

            mapped_assets = query_all(
                conn,
                """
                SELECT ai.asset_id
                FROM asset_integrations ai
                JOIN latest_monitoring_view lm ON lm.asset_id = ai.asset_id
                WHERE ai.provider = ? AND ai.enabled = 1 AND lm.status IN ('Erro', 'Desconectada')
                """,
                (provider,),
            )
            for row in mapped_assets:
                asset_id = int(row["asset_id"])
                if asset_id in synced_asset_ids:
                    continue
                existing_today = conn.execute(
                    """
                    SELECT 1
                    FROM monitoring_records
                    WHERE asset_id = ? AND record_date = ? AND source = ?
                    LIMIT 1
                    """,
                    (asset_id, date.today().isoformat(), provider),
                ).fetchone()
                if existing_today:
                    continue
                previous = get_latest_monitoring_row(conn, asset_id)
                conn.execute(
                    """
                    INSERT INTO monitoring_records (asset_id, status, record_date, notes, source, batch_id)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        asset_id,
                        "Resolvido",
                        date.today().isoformat(),
                        f"Resolvido automaticamente por ausencia no sync {provider}.",
                        provider,
                        batch_id,
                    ),
                )
                auto_resolved += 1
                event = build_monitoring_alert_event(
                    conn,
                    asset_id=asset_id,
                    previous_status=previous["status"] if previous else "",
                    current_status="Resolvido",
                    happened_at=now.isoformat(timespec="seconds"),
                )
                if event:
                    alert_events.append(event)

            conn.execute(
                """
                UPDATE integration_configs
                SET last_sync_at = ?, last_sync_status = 'success', last_error = '', updated_at = ?
                WHERE provider = ?
                """,
                (
                    datetime.now().isoformat(timespec="seconds"),
                    datetime.now().isoformat(timespec="seconds"),
                    provider,
                ),
            )
            conn.execute(
                """
                UPDATE monitoring_import_batches
                SET imported_count = ?, matched_count = ?, unmatched_count = ?, auto_resolved_count = ?
                WHERE id = ?
                """,
                (matched + unresolved, matched, unresolved, auto_resolved, batch_id),
            )
            finalize_integration_run(
                conn,
                run_id,
                status="success",
                matched_count=matched,
                unresolved_count=unresolved,
                auto_resolved_count=auto_resolved,
                summary_json={
                    "provider_rows": len(rows),
                    "alarm_rows": result.get("alarm_count", 0),
                    "alarm_error": result.get("alarm_error", ""),
                    "station_rows": result.get("station_count", len(rows)),
                    "realtime_rows": result.get("realtime_count", len(rows)),
                },
            )
            process_monitoring_alerts(conn, alert_events, batch_id, now)
            conn.commit()
            device_availability: dict[str, Any] | None = None
            if provider == INTEGRATION_PROVIDER_FUSIONSOLAR:
                try:
                    device_availability = run_fusionsolar_device_availability_sync(conn, provider, trigger_type=trigger_type)
                except Exception as exc:
                    logging.getLogger(__name__).warning("FusionSolar device availability sync failed: %s", exc)
            return {
                "matched": matched,
                "unresolved": unresolved,
                "auto_resolved": auto_resolved,
                "device_availability": device_availability,
            }
        except Exception as exc:
            conn.execute(
                """
                UPDATE integration_configs
                SET last_sync_status = 'error', last_error = ?, updated_at = ?
                WHERE provider = ?
                """,
                (str(exc), datetime.now().isoformat(timespec="seconds"), provider),
            )
            finalize_integration_run(
                conn,
                run_id,
                status="error",
                matched_count=0,
                unresolved_count=0,
                auto_resolved_count=0,
                error_message=str(exc),
            )
            conn.commit()
            raise


def run_all_integration_syncs(conn: sqlite3.Connection, trigger_type: str = "manual") -> dict[str, Any]:
    results: dict[str, Any] = {}
    errors: dict[str, str] = {}
    for provider in INTEGRATION_PROVIDER_OPTIONS:
        config = get_integration_config(conn, provider)
        if config is None or not config["enabled"]:
            continue
        try:
            results[provider] = run_integration_sync(conn, provider, trigger_type=trigger_type)
        except Exception as exc:
            errors[provider] = str(exc)
    if not results and errors:
        raise ValueError("; ".join(f"{provider}: {message}" for provider, message in errors.items()))
    return {"results": results, "errors": errors}


def resolve_fusionsolar_unresolved(conn: sqlite3.Connection, unresolved_id: int, asset_id: int) -> None:
    row = conn.execute(
        "SELECT * FROM integration_unresolved WHERE id = ? AND resolution_status = 'pending'",
        (unresolved_id,),
    ).fetchone()
    if row is None:
        raise ValueError("Entrada FusionSolar por resolver nao encontrada.")
    payload = json.loads(row["payload_json"] or "{}")
    create_or_update_asset_integration(
        conn,
        asset_id,
        row["provider"],
        str(row["external_id"] or ""),
        row["external_name"],
        row["external_status"] or "Operacional",
    )
    conn.execute(
        """
        UPDATE integration_unresolved
        SET resolution_status = 'resolved', resolved_at = ?, resolution_notes = ?
        WHERE id = ?
        """,
        (
            datetime.now().isoformat(timespec="seconds"),
            f"Associado ao asset {asset_id}",
            unresolved_id,
        ),
    )
    conn.commit()


def create_asset_from_unresolved(conn: sqlite3.Connection, unresolved_id: int) -> int:
    row = conn.execute(
        "SELECT * FROM integration_unresolved WHERE id = ? AND resolution_status = 'pending'",
        (unresolved_id,),
    ).fetchone()
    if row is None:
        raise ValueError("Entrada FusionSolar por resolver nao encontrada.")

    project_name = row["external_name"]
    installation_group = infer_installation_group(project_name)
    cursor = conn.execute(
        """
        INSERT INTO assets (project_name, installation_group, active_contract, notes, alias_blob)
        VALUES (?, ?, 'no', ?, ?)
        """,
        (
            project_name,
            installation_group,
            f"Criado a partir do provider {row['provider']}.",
            project_name,
        ),
    )
    asset_id = int(cursor.lastrowid)
    normalized_name = normalize_name(project_name)
    if normalized_name:
        conn.execute(
            "INSERT OR IGNORE INTO asset_aliases (asset_id, alias_name, normalized_alias, source) VALUES (?, ?, ?, ?)",
            (asset_id, project_name, normalized_name, "integration-create"),
        )
    resolve_fusionsolar_unresolved(conn, unresolved_id, asset_id)
    rebuild_asset_alias_blob(conn, asset_id)
    return asset_id


def ignore_fusionsolar_unresolved(conn: sqlite3.Connection, unresolved_id: int) -> None:
    conn.execute(
        """
        UPDATE integration_unresolved
        SET resolution_status = 'ignored', resolved_at = ?, resolution_notes = 'Ignorado manualmente'
        WHERE id = ?
        """,
        (datetime.now().isoformat(timespec="seconds"), unresolved_id),
    )
    conn.commit()


app = create_app()


def parse_cli_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Monitoring Board local app")
    parser.add_argument("--host", default="127.0.0.1", help="Host/IP onde a app vai escutar")
    parser.add_argument("--port", type=int, default=5000, help="Porta onde a app vai arrancar")
    parser.add_argument("--debug", action="store_true", help="Ativa debug do Flask")
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_cli_args()
    if DEFAULT_EXCEL_PATH and not DB_PATH.exists():
        with closing(get_db(str(DB_PATH))) as conn:
            import_excel_data(conn, DEFAULT_EXCEL_PATH)
    elif DEFAULT_EXCEL_PATH:
        with closing(get_db(str(DB_PATH))) as conn:
            if query_scalar(conn, "SELECT COUNT(*) FROM assets") == 0:
                import_excel_data(conn, DEFAULT_EXCEL_PATH)
    app.run(host=args.host, port=args.port, debug=args.debug)
