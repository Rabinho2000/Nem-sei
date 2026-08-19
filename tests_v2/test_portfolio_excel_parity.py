"""The portfolio workbook must be the document V1 already ships.

V2 has no portfolios as a domain concept yet, but the renderer is a pure
function of a payload, so its contract can be pinned against a real V1 output
now and cannot drift while the domain catches up.
"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from nemsei.reporting.portfolio_excel import (
    INSTALLATION_COLUMNS,
    METADATA_FIELDS,
    QUALITY_SOURCES,
    SHEET_NAMES,
    SUMMARY_METRICS,
    build_portfolio_report_workbook,
)


V1_REPORTS = Path("/opt/server/apps/Nem-sei/data/uploads/generated_reports")


def reference() -> Path | None:
    if not V1_REPORTS.is_dir():
        return None
    found = sorted(V1_REPORTS.glob("*/Output_Portfolio_*.xlsx"))
    return found[0] if found else None


REFERENCE = reference()
requires_reference = pytest.mark.skipif(REFERENCE is None, reason="no real V1 portfolio report is available here")


def payload(**overrides) -> dict:
    base = {
        "portfolio_name": "Output Portfolio 1",
        "period_label": "Fevereiro 2026",
        "profile": "Completo",
        "profile_version": 1,
        "engine_version": "portfolio-reporting-v1",
        "coverage_pct": 14.29,
        "totals": {"Potencia kWp": 10},
        "installations": [{"Instalacao": "Output Solar", "Nome local": "Output Solar", "NIF": "501123123", "Potencia kWp": 10}],
        "coverage": {"mapping": 100, "production": 0},
        "metadata": {"engine_version": "portfolio-reporting-v1", "period_start": "2026-02-01"},
    }
    base.update(overrides)
    return base


def rows_of(sheet):
    return [tuple(row) for row in sheet.iter_rows(values_only=True)]


@requires_reference
def test_sheets_match_v1_in_name_and_order() -> None:
    assert build_portfolio_report_workbook(payload()).sheetnames == load_workbook(REFERENCE).sheetnames == list(SHEET_NAMES)


@requires_reference
def test_all_forty_six_installation_columns_match_v1_exactly() -> None:
    expected = rows_of(load_workbook(REFERENCE, data_only=True)["Instalacoes"])[0]
    actual = rows_of(build_portfolio_report_workbook(payload())["Instalacoes"])[0]
    assert actual == expected
    assert len(actual) == len(INSTALLATION_COLUMNS) == 46


@requires_reference
def test_summary_metric_rows_match_v1() -> None:
    expected = [row[0] for row in rows_of(load_workbook(REFERENCE, data_only=True)["Resumo"])]
    actual = [row[0] for row in rows_of(build_portfolio_report_workbook(payload())["Resumo"])]
    # V1 appends three diagnostic keys after the metrics; the metric block itself must match.
    assert actual == expected[: len(actual)]
    assert [metric for metric in expected if metric in SUMMARY_METRICS] == list(SUMMARY_METRICS)


@requires_reference
def test_quality_and_metadata_rows_match_v1() -> None:
    reference = load_workbook(REFERENCE, data_only=True)
    workbook = build_portfolio_report_workbook(payload())
    assert [row[0] for row in rows_of(workbook["Qualidade dos dados"])] == ["Fonte", *QUALITY_SOURCES]
    expected_quality = [row[0] for row in rows_of(reference["Qualidade dos dados"])][: 1 + len(QUALITY_SOURCES)]
    assert [row[0] for row in rows_of(workbook["Qualidade dos dados"])] == expected_quality
    assert [row[0] for row in rows_of(workbook["Metadados"])] == list(METADATA_FIELDS)
    assert [row[0] for row in rows_of(reference["Metadados"])] == list(METADATA_FIELDS)


@requires_reference
def test_header_styling_and_column_widths_match_v1() -> None:
    reference = load_workbook(REFERENCE)
    workbook = build_portfolio_report_workbook(payload())
    for sheet_name in ("Instalacoes", "Qualidade dos dados", "Metadados"):
        mine, theirs = workbook[sheet_name]["A1"], reference[sheet_name]["A1"]
        assert bool(mine.font.b) == bool(theirs.font.b), sheet_name
        assert mine.fill.fgColor.rgb == theirs.fill.fgColor.rgb, sheet_name
        expected = {key: value.width for key, value in reference[sheet_name].column_dimensions.items() if value.width}
        actual = {key: value.width for key, value in workbook[sheet_name].column_dimensions.items() if value.width}
        assert actual == expected, sheet_name


@requires_reference
def test_v1_uses_no_merged_cells_in_the_portfolio_document() -> None:
    reference = load_workbook(REFERENCE)
    workbook = build_portfolio_report_workbook(payload())
    for sheet_name in SHEET_NAMES:
        assert list(reference[sheet_name].merged_cells.ranges) == []
        assert list(workbook[sheet_name].merged_cells.ranges) == []


def test_an_unmeasured_metric_is_blank_and_a_real_zero_is_kept() -> None:
    workbook = build_portfolio_report_workbook(payload(totals={"Potencia kWp": 10, "Producao real": 0}))
    summary = {row[0]: row[1] for row in rows_of(workbook["Resumo"]) if row[0]}
    assert summary["Producao real"] == "0.00"
    assert summary["Disponibilidade"] is None
    assert get_column_letter(46) == "AT"
