from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import pytest
from alembic import command
from alembic.config import Config
from openpyxl import Workbook
from sqlalchemy import create_engine, func, select
from sqlalchemy.exc import IntegrityError

from nemsei.assets.service import create_asset
from nemsei.db.session import build_session_factory
from nemsei.reporting.financial_workbook import FinancialModelParseError, parse_financial_model_workbook
from nemsei.reporting.models import FinancialModel, FinancialModelMonth, ReportSourceFile
from nemsei.reporting.service import import_financial_model, register_source_file


def upgrade(settings, monkeypatch) -> None:
    monkeypatch.setenv("NEMSEI_V2_ENV", "test")
    monkeypatch.setenv("NEMSEI_V2_DATABASE_URL", settings.database_url)
    command.upgrade(Config("alembic.ini"), "head")


def as_sold_workbook(path: Path, *, missing_month: int | None = None) -> Path:
    """A minimal workbook of the one family with a real file behind it."""
    workbook = Workbook()
    project = workbook.active
    project.title = "Projeto"
    project["C5"] = "Projeto Exemplo"
    project["H8"] = 17.5
    project["H14"] = 1234.5
    project["H22"] = 0.02
    project["H23"] = 0.005
    project["D26"], project["E26"], project["D28"], project["E28"] = 500, 8750, 700, 12250
    project["J5"], project["K5"], project["L5"], project["M5"] = "Month", "Monthly Production [kWh]", "AC [kWh]", "% AC"
    for month in range(1, 13):
        row = month + 5
        project.cell(row, 10, month)
        project.cell(row, 11, None if month == missing_month else month * 100)
        project.cell(row, 12, month * 60)
        project.cell(row, 13, 0.6)
    project["P5"], project["P6"], project["P7"] = 11700, 7800, 4680
    project["P8"], project["P9"], project["P10"] = 3120, 0.6, 0.4
    project["L32"], project["L33"], project["F46"] = 0.15, 0.086, 0.05
    project["G39"] = "2026/1"
    for row, label, energy, network in ((41, "SV", 0.08, 0.02), (42, "Vazio", 0.09, 0.03), (43, "Cheia", 0.10, 0.04), (44, "Ponta", 0.11, 0.05)):
        project.cell(row, 5, label)
        project.cell(row, 6, energy)
        project.cell(row, 7, network)
        project.cell(row, 8, energy + network)

    savings = workbook.create_sheet("Savings Yr1")
    savings.append([])
    savings.append([])
    savings.append([None, "Month", "Cons. [kWh]", "Faturas €", "AC [kWh]", "Save AC €", "Exced [kWh]"])
    for month in range(1, 13):
        savings.append([None, month, month * 150, month * 10, month * 60, month * 4, month * 40])
    savings["B16"], savings["C16"], savings["D16"] = "TOTAL", 11700, 780
    savings["E16"], savings["F16"], savings["G16"] = 4680, 312, 3120
    workbook.save(path)
    return path


@pytest.fixture
def asset_and_workbook(settings, monkeypatch, tmp_path):
    upgrade(settings, monkeypatch)
    factory = build_session_factory(create_engine(settings.database_url))
    with factory() as session, session.begin():
        asset = create_asset(session, canonical_name="Alpha Solar")
        asset_id = asset.id
    return factory, asset_id, as_sold_workbook(tmp_path / "model.xlsx")


def register(session, asset_id: int, path: Path, **kwargs):
    return register_source_file(
        session,
        asset_id=asset_id,
        path=path,
        original_filename=path.name,
        stored_path=f"uploads/financial_models/{asset_id}/{path.name}",
        uploaded_by="operator",
        **kwargs,
    )


def test_a_source_file_is_identified_by_its_hash_and_never_duplicated(asset_and_workbook) -> None:
    factory, asset_id, workbook = asset_and_workbook
    with factory() as session, session.begin():
        first = register(session, asset_id, workbook)
        assert len(first.sha256) == 64 and first.size_bytes == workbook.stat().st_size
        again = register(session, asset_id, workbook)
        assert again.id == first.id
    with factory() as session:
        assert session.scalar(select(func.count()).select_from(ReportSourceFile)) == 1


def test_the_same_file_cannot_be_claimed_by_two_assets(asset_and_workbook) -> None:
    factory, asset_id, workbook = asset_and_workbook
    with factory() as session, session.begin():
        register(session, asset_id, workbook)
        other = create_asset(session, canonical_name="Bravo Solar")
        with pytest.raises(ValueError, match="another asset"):
            register(session, other.id, workbook)


def test_import_persists_every_month_with_its_provenance(asset_and_workbook) -> None:
    factory, asset_id, workbook = asset_and_workbook
    parsed = parse_financial_model_workbook(workbook)
    with factory() as session, session.begin():
        source = register(session, asset_id, workbook)
        model = import_financial_model(session, source_file=source, workbook_path=workbook, operator="operator", confirm=True)
        model_id = model.id

    with factory() as session:
        model = session.get(FinancialModel, model_id)
        assert model.status == "confirmed" and model.version == 1
        assert model.workbook_format == "financial_automatic_as_sold"
        assert model.source_file_sha256 == session.get(ReportSourceFile, model.source_file_id).sha256
        assert model.parser_name == parsed.parser_name and model.parser_version == parsed.parser_version
        assert model.warnings_json == list(parsed.warnings)
        assert model.details_json["format"] == "financial_automatic_as_sold"

        months = session.scalars(
            select(FinancialModelMonth).where(FinancialModelMonth.financial_model_id == model_id).order_by(FinancialModelMonth.month)
        ).all()
        assert [month.month for month in months] == list(range(1, 13))
        for month, entry in zip(months, parsed.monthly, strict=True):
            assert month.expected_production_kwh == Decimal(str(entry["expected_production_kwh"]))
            # The cell behind the number travels with the number.
            assert month.source_fields_json == entry["source_fields"]
            assert month.source_fields_json["expected_production_kwh"]["cell"].startswith("Projeto!K")


def test_an_incomplete_workbook_is_refused_rather_than_filled_in(settings, monkeypatch, tmp_path) -> None:
    """V1's rule, kept deliberately: a gap invalidates the model, it is not a zero.

    A financial model drives expected production for a whole year, so a month
    the workbook does not state cannot be guessed, interpolated or defaulted.
    The parser refuses the file and nothing reaches the database.
    """
    upgrade(settings, monkeypatch)
    factory = build_session_factory(create_engine(settings.database_url))
    workbook = as_sold_workbook(tmp_path / "gap.xlsx", missing_month=7)
    with factory() as session, session.begin():
        asset = create_asset(session, canonical_name="Alpha Solar")
        source = register(session, asset.id, workbook)
        with pytest.raises(FinancialModelParseError, match="financial_model_missing_month"):
            import_financial_model(session, source_file=source, workbook_path=workbook, operator="operator")
    with factory() as session:
        assert session.scalar(select(func.count()).select_from(FinancialModel)) == 0
        assert session.scalar(select(func.count()).select_from(FinancialModelMonth)) == 0


def test_base_year_records_whether_the_workbook_or_an_operator_chose_it(asset_and_workbook) -> None:
    factory, asset_id, workbook = asset_and_workbook
    with factory() as session, session.begin():
        source = register(session, asset_id, workbook)
        from_workbook = import_financial_model(session, source_file=source, workbook_path=workbook, operator="operator")
        assert from_workbook.base_year == 2026
        assert from_workbook.base_year_source == "workbook"
        assert from_workbook.base_year_cell == "Projeto!G39"

        overridden = import_financial_model(
            session, source_file=source, workbook_path=workbook, operator="sergio", base_year_override=2025
        )
        assert overridden.base_year == 2025
        assert overridden.base_year_source == "operator"
        assert overridden.base_year_cell is None
        assert overridden.confirmed_by == "sergio"


def test_an_operator_base_year_must_name_its_actor(asset_and_workbook) -> None:
    factory, asset_id, workbook = asset_and_workbook
    with factory() as session:
        with pytest.raises(IntegrityError):
            session.execute(
                FinancialModel.__table__.insert().values(
                    source_file_id=None, asset_id=asset_id, version=99, status="draft",
                    base_year=2025, base_year_source="operator", workbook_format="unknown",
                    parser_name="x", parser_version="1", source_file_sha256="0" * 64,
                    warnings_json=[], details_json={}, source_cells_json={},
                    confirmed_by=None, created_at=func.now(), updated_at=func.now(),
                )
            )
        session.rollback()


def test_confirming_a_new_version_supersedes_the_previous_one(asset_and_workbook) -> None:
    factory, asset_id, workbook = asset_and_workbook
    with factory() as session, session.begin():
        source = register(session, asset_id, workbook)
        first = import_financial_model(session, source_file=source, workbook_path=workbook, operator="operator", confirm=True)
        second = import_financial_model(session, source_file=source, workbook_path=workbook, operator="operator", confirm=True)
        first_id, second_id = first.id, second.id
    with factory() as session:
        assert session.get(FinancialModel, first_id).status == "superseded"
        second = session.get(FinancialModel, second_id)
        assert second.status == "confirmed" and second.version == 2
        assert second.supersedes_model_id == first_id
