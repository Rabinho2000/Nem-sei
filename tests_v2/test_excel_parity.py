"""The V2 workbook must be the document V1 already ships, not a lookalike.

The reference is a real V1 output kept on the server, so the comparison is
against something a customer actually received rather than against a fixture
written from the same assumptions as the code under test.
"""
from __future__ import annotations

import ast
from pathlib import Path

import pytest
from openpyxl import load_workbook

from nemsei.reporting.excel import (
    ENERGY_METRICS,
    FINANCIAL_METRICS,
    METADATA_FIELDS,
    SHEET_NAMES,
    build_asset_report_workbook,
    format_value,
)


V1_REPORTS = Path("/opt/server/apps/Nem-sei/data/uploads/generated_reports")


def v1_reference() -> Path | None:
    if not V1_REPORTS.is_dir():
        return None
    candidates = sorted(V1_REPORTS.glob("*/Central_congelada_*.xlsx"))
    return candidates[0] if candidates else None


REFERENCE = v1_reference()
requires_reference = pytest.mark.skipif(REFERENCE is None, reason="no real V1 report is available here")


def sample_payload(**overrides) -> dict:
    payload = {
        "title": "Central congelada - 2026-06",
        "asset_name": "Central congelada",
        "energy": {"production_kwh": 321.0},
        "financial": {},
        "quality": {"production_state": "complete"},
        "metadata": {"period_start": "2026-06-01", "period_end": "2026-06-30"},
    }
    payload.update(overrides)
    return payload


def rows_of(sheet) -> list[tuple]:
    return [tuple(row) for row in sheet.iter_rows(values_only=True)]


def pairs(rows) -> dict:
    """Label/value pairs, ignoring the trailing layout columns."""
    return {row[0]: row[1] for row in rows if row and row[0] is not None}


@requires_reference
def test_the_workbook_has_v1s_sheets_in_v1s_order(tmp_path: Path) -> None:
    reference = load_workbook(REFERENCE, data_only=True)
    workbook = build_asset_report_workbook(sample_payload())
    assert workbook.sheetnames == reference.sheetnames == list(SHEET_NAMES)


@requires_reference
def test_metric_rows_match_v1_row_for_row() -> None:
    reference = load_workbook(REFERENCE, data_only=True)
    workbook = build_asset_report_workbook(sample_payload())
    for sheet_name, keys in (("Energia", ENERGY_METRICS), ("Financeiro", FINANCIAL_METRICS)):
        expected = [row[0] for row in rows_of(reference[sheet_name])]
        actual = [row[0] for row in rows_of(workbook[sheet_name])]
        assert actual == expected, sheet_name
        assert [key for key in expected if key in keys] == list(keys)


@requires_reference
def test_metadata_fields_match_v1() -> None:
    reference = load_workbook(REFERENCE, data_only=True)
    expected = [row[0] for row in rows_of(reference["Metadados"])][1:]
    workbook = build_asset_report_workbook(sample_payload())
    actual = [row[0] for row in rows_of(workbook["Metadados"])][1:]
    assert actual == expected == list(METADATA_FIELDS)


@requires_reference
def test_the_summary_sheet_keeps_v1s_two_column_layout() -> None:
    reference = load_workbook(REFERENCE, data_only=True)
    workbook = build_asset_report_workbook(sample_payload())
    expected_labels = [(row[0], row[3]) for row in rows_of(reference["Resumo"])[5:]]
    actual_labels = [(row[0], row[3]) for row in rows_of(workbook["Resumo"])[5:]]
    assert actual_labels == expected_labels


@requires_reference
def test_a_measured_value_is_written_exactly_as_v1_writes_it() -> None:
    reference = load_workbook(REFERENCE, data_only=True)
    reference_production = pairs(rows_of(reference["Energia"])[2:])["production_kwh"]
    workbook = build_asset_report_workbook(sample_payload())
    assert pairs(rows_of(workbook["Energia"])[2:])["production_kwh"] == reference_production == "321.00"


def test_a_missing_value_is_blank_and_never_a_zero() -> None:
    """The rule that protects the customer relationship, pinned on its own."""
    workbook = build_asset_report_workbook(sample_payload())
    energy = pairs(rows_of(workbook["Energia"])[2:])
    assert energy["self_use_kwh"] is None
    assert energy["production_kwh"] == "321.00"
    financial = pairs(rows_of(workbook["Financeiro"])[2:])
    assert set(financial.values()) == {None}


def test_zero_is_written_as_zero_and_stays_distinct_from_missing() -> None:
    workbook = build_asset_report_workbook(sample_payload(energy={"production_kwh": 0}))
    energy = pairs(rows_of(workbook["Energia"])[2:])
    assert energy["production_kwh"] == "0.00"
    assert energy["export_kwh"] is None


def test_formatting_matches_v1s_two_decimal_text() -> None:
    assert format_value(321) == "321.00"
    assert format_value(0.005) == f"{0.005:.2f}"  # V1 formats floats, not Decimals
    assert format_value(None) is None
    assert format_value("") is None
    assert format_value("complete") == "complete"


def test_rendering_never_reaches_a_database_or_a_provider() -> None:
    import inspect

    from nemsei.reporting import excel

    tree = ast.parse(inspect.getsource(excel))
    imported = {
        name.name.split(".")[0]
        for node in ast.walk(tree)
        if isinstance(node, ast.Import)
        for name in node.names
    } | {
        (node.module or "").split(".")[0]
        for node in ast.walk(tree)
        if isinstance(node, ast.ImportFrom)
    }
    assert imported <= {"__future__", "decimal", "typing", "openpyxl"}, imported


def style_of(cell) -> dict:
    """The visual facts a customer would notice, and nothing else."""
    return {
        "bold": bool(cell.font.b),
        "italic": bool(cell.font.i),
        "size": cell.font.sz,
        "colour": getattr(cell.font.color, "rgb", None) if cell.font.color else None,
        "fill": cell.fill.fgColor.rgb if cell.fill and cell.fill.fill_type else None,
    }


@requires_reference
def test_the_banner_rows_are_painted_exactly_as_v1_paints_them() -> None:
    reference = load_workbook(REFERENCE)
    workbook = build_asset_report_workbook(sample_payload())
    for sheet_name, cells in (("Resumo", ("A1", "A2", "A4")), ("Energia", ("A1", "A2", "B2"))):
        for ref in cells:
            assert style_of(workbook[sheet_name][ref]) == style_of(reference[sheet_name][ref]), f"{sheet_name}!{ref}"


@requires_reference
def test_merged_ranges_and_column_widths_match_v1() -> None:
    reference = load_workbook(REFERENCE)
    workbook = build_asset_report_workbook(sample_payload())
    for sheet_name in SHEET_NAMES:
        assert sorted(str(r) for r in workbook[sheet_name].merged_cells.ranges) == sorted(
            str(r) for r in reference[sheet_name].merged_cells.ranges
        ), sheet_name
        expected = {key: value.width for key, value in reference[sheet_name].column_dimensions.items() if value.width}
        actual = {key: value.width for key, value in workbook[sheet_name].column_dimensions.items() if value.width}
        assert actual == expected, sheet_name


@requires_reference
def test_the_title_row_keeps_v1s_height() -> None:
    reference = load_workbook(REFERENCE)
    workbook = build_asset_report_workbook(sample_payload())
    assert workbook["Resumo"].row_dimensions[1].height == reference["Resumo"].row_dimensions[1].height == 30.0
