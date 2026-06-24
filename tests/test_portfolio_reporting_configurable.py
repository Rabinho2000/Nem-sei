from __future__ import annotations

import sqlite3
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

import app as app_module
import monitoring_board.app_factory as app_factory_module
import monitoring_board.services.portfolio_reporting as portfolio_reporting_service
from app import ensure_database
from monitoring_board.portfolio_report_repository import (
    archive_profile,
    duplicate_profile,
    get_default_profile,
    get_snapshot_result,
    latest_profile_version,
    list_profiles,
    save_profile,
    set_default_profile,
    snapshot_portfolio_result,
)
from monitoring_board.portfolio_repository import create_portfolio
from monitoring_board.reporting.periods import build_period
from monitoring_board.reporting.portfolio import (
    PortfolioReportColumn,
    PortfolioReportProfile,
    PortfolioReportRow,
    aggregate_rows,
    comparison_values,
    data_coverage,
    default_profile,
)
from monitoring_board.services.portfolio_reporting import export_portfolio_result_workbook, prepare_portfolio_report


def connect(tmp_path: Path) -> sqlite3.Connection:
    db_path = tmp_path / "portfolio-reporting.db"
    ensure_database(str(db_path))
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def add_asset(conn: sqlite3.Connection, name: str = "Solar One") -> int:
    cursor = conn.execute(
        "INSERT INTO assets (project_name, nif, active_contract, kwp, mounting_date, start_contract) VALUES (?, '123456789', 'yes', '10', '2024-01-01', '2024-01-01')",
        (name,),
    )
    return int(cursor.lastrowid)


def add_portfolio_member(conn: sqlite3.Connection, asset_id: int) -> int:
    portfolio_id = create_portfolio(conn, name=f"Portfolio Test {asset_id}")
    conn.execute(
        """
        INSERT INTO portfolio_assets (portfolio_id, asset_id, external_name, active, mapping_status, mapping_confidence, display_order)
        VALUES (?, ?, 'Solar One', 1, 'manual', 1, 10)
        """,
        (portfolio_id, asset_id),
    )
    conn.commit()
    return int(portfolio_id)


def test_domain_recalculates_totals_and_comparison_without_averaging_percentages() -> None:
    columns = (
        PortfolioReportColumn("actual_production_kwh", "Producao real"),
        PortfolioReportColumn("adjusted_expected_kwh", "Esperada"),
        PortfolioReportColumn("installed_power_kwp", "kWp"),
        PortfolioReportColumn("availability_pct", "Disponibilidade"),
        PortfolioReportColumn("deviation_pct", "Desvio"),
    )
    rows = (
        PortfolioReportRow(1, {"actual_production_kwh": Decimal("100"), "adjusted_expected_kwh": Decimal("100"), "installed_power_kwp": Decimal("1"), "availability_pct": Decimal("100")}),
        PortfolioReportRow(2, {"actual_production_kwh": Decimal("50"), "adjusted_expected_kwh": Decimal("200"), "installed_power_kwp": Decimal("9"), "availability_pct": Decimal("80")}),
    )

    summary = aggregate_rows(rows, columns)
    previous = aggregate_rows((PortfolioReportRow(3, {"actual_production_kwh": Decimal("0"), "adjusted_expected_kwh": Decimal("0")}),), columns)
    comparison = comparison_values(summary, previous, "previous_period")

    assert summary.values["actual_production_kwh"] == Decimal("150.00")
    assert summary.values["availability_pct"] == Decimal("82.00")
    assert summary.values["deviation_pct"] == Decimal("-50.00")
    assert comparison.values["actual_production_kwh"]["delta_pct"] is None


def test_profiles_versions_archive_and_duplicate_are_persisted(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    profiles = list_profiles(conn)
    assert {row["name"] for row in profiles} >= {"Resumo operacional", "Performance", "Financeiro", "Qualidade dos dados", "Completo"}

    profile = PortfolioReportProfile(
        id=None,
        name="Custom",
        description="Custom profile",
        portfolio_id=None,
        period_type="quarterly",
        columns=(PortfolioReportColumn("installation", "Instalacao"), PortfolioReportColumn("actual_production_kwh", "kWh")),
        filters={},
    )
    profile_id = save_profile(conn, profile)
    duplicate_id = duplicate_profile(conn, profile_id, "Custom copy")
    archive_profile(conn, duplicate_id)

    assert latest_profile_version(conn, profile_id) == 1
    assert get_default_profile(conn).name == "Completo"
    assert all(row["id"] != duplicate_id for row in list_profiles(conn))


def test_service_periods_snapshots_and_excel_use_same_result(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    asset_id = add_asset(conn)
    portfolio_id = add_portfolio_member(conn, asset_id)
    profile = default_profile("Performance")
    profile_id = save_profile(conn, profile)
    profile = get_default_profile(conn)

    result = prepare_portfolio_report(
        conn,
        portfolio_id=portfolio_id,
        portfolio_name="Solcorelios I",
        profile=profile,
        period_type="quarterly",
        year=2026,
        quarter=1,
        profile_version=latest_profile_version(conn, profile_id),
    )
    snapshot_id = snapshot_portfolio_result(conn, result, "snap")
    conn.execute("UPDATE assets SET project_name = 'Changed' WHERE id = ?", (asset_id,))
    snapshot = get_snapshot_result(conn, snapshot_id)
    assert snapshot is not None
    workbook = export_portfolio_result_workbook(snapshot)
    output = tmp_path / "portfolio.xlsx"
    workbook.save(output)
    loaded = load_workbook(output)

    assert result.period.label == "T1 2026"
    assert snapshot.rows[0].values["installation"] == "Solar One"
    assert {"Resumo", "Instalacoes", "Qualidade dos dados", "Metadados"} <= set(loaded.sheetnames)


def test_period_coverage_marks_missing_multi_month_result() -> None:
    period = build_period("semiannual", year=2026, semester=1)
    coverage = data_coverage((), period.included_months)

    assert coverage.global_pct == Decimal("0.00")
    assert coverage.missing_months == tuple(month.isoformat() for month in period.included_months)


def test_service_keeps_simple_self_use_and_real_financial_values(monkeypatch) -> None:
    monthly = [
        {
            "asset_id": 1,
            "installation": "A",
            "installed_power_kwp": 10,
            "actual_production_kwh": 100,
            "adjusted_expected_kwh": 90,
            "self_use_kwh": 70,
            "self_use_simple_kwh": 70,
            "export_kwh": 30,
            "consumption_kwh": 120,
            "grid_import_kwh": 50,
            "estimated_value_eur": 14,
            "export_revenue_eur": 3,
            "esco_payment_eur": 7,
            "fixed_fee_eur": 0,
            "net_benefit_eur": 10,
            "availability_pct": 100,
            "warnings": [],
        }
    ]
    monkeypatch.setattr(portfolio_reporting_service, "build_portfolio_report_rows", lambda conn, portfolio_id, report_month: monthly)

    result = prepare_portfolio_report(
        None,
        portfolio_id=1,
        portfolio_name="P",
        profile=default_profile("Financeiro"),
        period_type="monthly",
        report_month="2026-01",
    )

    row = result.rows[0].values
    assert row["self_use_kwh"] == Decimal("70.00")
    assert row["export_kwh"] == Decimal("30.00")
    assert row["grid_import_kwh"] == Decimal("50.00")
    assert row["net_benefit_eur"] == Decimal("10.00")
    assert result.summary.values["net_benefit_eur"] == Decimal("10.00")


def test_service_does_not_convert_missing_financial_metrics_to_zero(monkeypatch) -> None:
    monthly = [{"asset_id": 1, "installation": "A", "installed_power_kwp": 10, "warnings": ["missing_self_use", "missing_billing"]}]
    monkeypatch.setattr(portfolio_reporting_service, "build_portfolio_report_rows", lambda conn, portfolio_id, report_month: monthly)

    result = prepare_portfolio_report(
        None,
        portfolio_id=1,
        portfolio_name="P",
        profile=default_profile("Financeiro"),
        period_type="monthly",
        report_month="2026-01",
    )

    assert result.rows[0].values.get("self_use_kwh") is None
    assert result.rows[0].values.get("net_benefit_eur") is None
    assert result.summary.values.get("net_benefit_eur") is None


def test_partial_coverage_uses_expected_installation_month_slots(monkeypatch) -> None:
    rows_by_month = {
        "2026-01": [{"asset_id": 1, "installation": "A", "installed_power_kwp": 10, "actual_production_kwh": 10, "warnings": []}],
        "2026-02": [{"asset_id": 1, "installation": "A", "installed_power_kwp": 10, "warnings": ["missing_monthly_production", "inferred_hourly_self_use"]}],
        "2026-03": [{"asset_id": 1, "installation": "A", "installed_power_kwp": 10, "actual_production_kwh": 20, "warnings": []}],
    }
    monkeypatch.setattr(portfolio_reporting_service, "build_portfolio_report_rows", lambda conn, portfolio_id, report_month: rows_by_month[report_month])

    result = prepare_portfolio_report(
        None,
        portfolio_id=1,
        portfolio_name="P",
        profile=default_profile("Qualidade dos dados"),
        period_type="quarterly",
        year=2026,
        quarter=1,
    )

    assert result.coverage.by_source["production"] == Decimal("66.67")
    assert result.coverage.by_source["self_use"] == Decimal("100.00")


def test_snapshot_preserves_comparison(tmp_path: Path, monkeypatch) -> None:
    conn = connect(tmp_path)
    portfolio_id = add_portfolio_member(conn, add_asset(conn))
    rows_by_month = {
        "2026-01": [{"asset_id": 1, "installation": "A", "actual_production_kwh": 100, "adjusted_expected_kwh": 100, "warnings": []}],
        "2025-12": [{"asset_id": 1, "installation": "A", "actual_production_kwh": 50, "adjusted_expected_kwh": 100, "warnings": []}],
    }
    monkeypatch.setattr(portfolio_reporting_service, "build_portfolio_report_rows", lambda conn, portfolio_id, report_month: rows_by_month[report_month])
    result = prepare_portfolio_report(
        conn,
        portfolio_id=portfolio_id,
        portfolio_name="P",
        profile=default_profile("Performance"),
        period_type="monthly",
        report_month="2026-01",
        comparison="previous_period",
    )
    snapshot_id = snapshot_portfolio_result(conn, result)
    snapshot = get_snapshot_result(conn, snapshot_id)

    assert snapshot is not None
    assert snapshot.comparison is not None
    assert snapshot.comparison.values["actual_production_kwh"]["delta"] == "50.00"
    workbook = export_portfolio_result_workbook(snapshot)
    assert "Comparacao" in [cell.value for row in workbook["Resumo"].iter_rows() for cell in row]


def test_profile_default_uniqueness_and_invalid_config(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    first = save_profile(conn, default_profile("Performance"), is_default=1)
    second = save_profile(conn, default_profile("Financeiro"), is_default=1)

    assert get_default_profile(conn).id == second
    assert conn.execute("SELECT COUNT(*) FROM portfolio_report_profiles WHERE portfolio_id IS NULL AND is_default = 1").fetchone()[0] == 1
    set_default_profile(conn, first)
    assert get_default_profile(conn).id == first
    invalid = PortfolioReportProfile(
        id=None,
        name="Invalid",
        description="",
        portfolio_id=None,
        period_type="weekly",
        columns=(PortfolioReportColumn("actual_production_kwh", "kWh"),),
        filters={},
    )
    try:
        save_profile(conn, invalid)
    except ValueError as exc:
        assert str(exc) == "invalid_profile_period_type"
    else:
        raise AssertionError("expected invalid period type")


def test_portfolio_reports_route_does_not_recalculate_legacy_when_canonical_succeeds(tmp_path: Path, monkeypatch) -> None:
    db_path = tmp_path / "no-legacy.db"
    ensure_database(str(db_path))
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    portfolio_id = add_portfolio_member(conn, add_asset(conn))
    conn.close()
    calls = {"legacy": 0}

    def legacy(*args, **kwargs):
        calls["legacy"] += 1
        return []

    monkeypatch.setattr(app_factory_module, "build_portfolio_report_rows", legacy)
    flask_app = app_module.app
    previous_db = flask_app.config["DATABASE"]
    previous_testing = flask_app.config.get("TESTING")
    flask_app.config["DATABASE"] = str(db_path)
    flask_app.config["TESTING"] = True
    client = flask_app.test_client()
    with client.session_transaction() as session:
        session["authenticated"] = True
        session["csrf_token"] = "token"
    try:
        response = client.get(f"/portfolio-reports?portfolio_id={portfolio_id}&report_month=2026-01")
        assert response.status_code == 200
        assert calls["legacy"] == 0
    finally:
        flask_app.config["DATABASE"] = previous_db
        flask_app.config["TESTING"] = previous_testing


def test_portfolio_report_routes_preview_snapshot_and_export(tmp_path: Path) -> None:
    db_path = tmp_path / "routes.db"
    conn = sqlite3.connect(db_path)
    conn.close()
    ensure_database(str(db_path))
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    asset_id = add_asset(conn)
    portfolio_id = add_portfolio_member(conn, asset_id)
    conn.close()

    flask_app = app_module.app
    previous_db = flask_app.config["DATABASE"]
    previous_testing = flask_app.config.get("TESTING")
    flask_app.config["DATABASE"] = str(db_path)
    flask_app.config["TESTING"] = True
    client = flask_app.test_client()
    with client.session_transaction() as session:
        session["authenticated"] = True
        session["csrf_token"] = "token"
    try:
        preview = client.get(f"/portfolio-reports?portfolio_id={portfolio_id}&period_type=annual&report_year=2026")
        assert preview.status_code == 200
        assert b"Portfolio configuravel" in preview.data
        assert b"Configurador de perfil" in preview.data

        post = client.post(
            "/portfolio-reports/generate",
            data={"csrf_token": "token", "portfolio_id": portfolio_id, "period_type": "annual", "report_year": "2026", "report_month": "2026-01"},
        )
        assert post.status_code in {302, 303}
        export = client.get(f"/portfolio-reports/export?portfolio_id={portfolio_id}&period_type=annual&report_year=2026&report_month=2026-01")
        assert export.status_code == 200
        assert export.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    finally:
        flask_app.config["DATABASE"] = previous_db
        flask_app.config["TESTING"] = previous_testing
