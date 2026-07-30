from __future__ import annotations

import sqlite3
import struct
import hashlib
import io
import json
from contextlib import contextmanager
from dataclasses import replace
from datetime import date
from pathlib import Path
from urllib.parse import parse_qs, urlparse

import app as app_module
import monitoring_board.app_factory as app_factory_module
from openpyxl import load_workbook
from pypdf import PdfReader
from app import ensure_database
from monitoring_board.customer_reports import prepare_customer_report
from monitoring_board.portfolio_report_repository import get_default_profile
from monitoring_board.portfolio_repository import create_portfolio
from monitoring_board.report_template_repository import (
    archive_template,
    duplicate_template,
    ensure_report_template_schema,
    get_default_template,
    get_template,
    latest_template_version,
    list_templates,
    save_template,
    set_default_template,
)
from monitoring_board.reporting.templates import (
    RENDERER_GENERIC_INDIVIDUAL,
    RENDERER_GENERIC_PORTFOLIO,
    RENDERER_SOLCOR_INDIVIDUAL,
    default_template,
    template_to_config,
)
from monitoring_board.reporting.templates import validate_template_scope
from monitoring_board.reporting_storage import reconcile_generated_reports
from monitoring_board.services.portfolio_reporting import prepare_portfolio_report
from monitoring_board.reporting.templates import TemplateSection
from monitoring_board.services.report_rendering import render_individual_excel, render_individual_pdf, render_portfolio_excel, render_portfolio_html, render_portfolio_pdf, render_zip, safe_filename


def connect(tmp_path: Path) -> sqlite3.Connection:
    db_path = tmp_path / "outputs.db"
    ensure_database(str(db_path))
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def add_asset(conn: sqlite3.Connection, name: str = "Output Solar") -> int:
    cursor = conn.execute(
        "INSERT INTO assets (project_name, nif, active_contract, kwp, mounting_date, start_contract) VALUES (?, '501123123', 'yes', '10', '2024-01-01', '2024-01-01')",
        (name,),
    )
    return int(cursor.lastrowid)


def add_fusionsolar_integration(
    conn: sqlite3.Connection,
    asset_id: int,
    name: str,
) -> None:
    conn.execute(
        """
        INSERT INTO asset_integrations
            (asset_id, provider, external_id, external_name, enabled)
        VALUES (?, 'FusionSolar', ?, ?, 1)
        """,
        (asset_id, f"FS-{asset_id}", name),
    )


def add_portfolio(conn: sqlite3.Connection, asset_id: int) -> int:
    portfolio_id = create_portfolio(conn, name=f"Output Portfolio {asset_id}")
    conn.execute(
        "INSERT INTO portfolio_assets (portfolio_id, asset_id, external_name, active, mapping_status, mapping_confidence, display_order) VALUES (?, ?, 'Output Solar', 1, 'manual', 1, 10)",
        (portfolio_id, asset_id),
    )
    conn.commit()
    return portfolio_id


@contextmanager
def report_test_client(db_path: Path):
    flask_app = app_module.app
    previous_db = flask_app.config["DATABASE"]
    previous_testing = flask_app.config.get("TESTING")
    flask_app.config["DATABASE"] = str(db_path)
    flask_app.config["TESTING"] = True
    client = flask_app.test_client()
    with client.session_transaction() as session:
        session["authenticated"] = True
        session["csrf_token"] = "token"
    try:
        yield client
    finally:
        flask_app.config["DATABASE"] = previous_db
        flask_app.config["TESTING"] = previous_testing


def install_custom_individual_template(conn: sqlite3.Connection) -> int:
    conn.execute("DELETE FROM report_template_versions")
    conn.execute("DELETE FROM report_templates")
    base = default_template("Individual padrao")
    custom = replace(
        base,
        name="EPC Cliente Personalizado",
        description="Configuração individual preservada",
        title="Relatório Solar {asset}",
        subtitle="Desempenho energético",
        sections=(
            TemplateSection("cover", "Capa personalizada", True, 10),
            TemplateSection("production", "Produção personalizada", True, 20),
            TemplateSection("warnings", "Avisos de qualidade", True, 30),
        ),
        branding=replace(
            base.branding,
            company_name="Solcor Cliente",
            client_name="Cliente AGA",
            logo_path="static/solcor-logo.png",
            primary_color="#123456",
            secondary_color="#65A430",
            footer="Rodapé preservado",
            contacts="reporting@example.invalid",
            disclaimer="Configuração de teste.",
        ),
        filename_pattern="EPC_Custom_{asset}_{period}",
    )
    return save_template(conn, custom, is_default=1)


def test_template_crud_default_version_and_invalid_config(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    templates = list_templates(conn)
    assert {"Individual padrao", "Portfolio executivo"} <= {row["name"] for row in templates}

    template_id = save_template(conn, default_template("Portfolio operacional"), is_default=1)
    duplicate_id = duplicate_template(conn, template_id, "Portfolio operacional copia")
    archive_template(conn, duplicate_id)
    set_default_template(conn, template_id)

    assert latest_template_version(conn, template_id) == 1
    assert get_default_template(conn, "portfolio").id == template_id
    invalid = default_template("Portfolio executivo")
    invalid = invalid.__class__(**{**invalid.__dict__, "name": "", "report_type": "portfolio"})
    try:
        save_template(conn, invalid)
    except ValueError as exc:
        assert str(exc) == "template_name_required"
    else:
        raise AssertionError("expected invalid template")


def test_custom_template_survives_schema_and_startup_byte_for_byte(
    tmp_path: Path,
) -> None:
    db_path = tmp_path / "custom-template-startup.db"
    ensure_database(str(db_path))
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        template_id = install_custom_individual_template(conn)
        custom_config = json.dumps(
            template_to_config(get_default_template(conn, "individual")),
            ensure_ascii=False,
            indent=2,
        )
        conn.execute(
            "UPDATE report_templates SET config_json = ? WHERE id = ?",
            (custom_config, template_id),
        )
        conn.execute(
            "UPDATE report_template_versions SET config_json = ? WHERE template_id = ? AND version = 1",
            (custom_config, template_id),
        )
        conn.commit()
        template_before = dict(
            conn.execute(
                "SELECT * FROM report_templates WHERE id = ?",
                (template_id,),
            ).fetchone()
        )
        versions_before = [
            dict(row)
            for row in conn.execute(
                "SELECT * FROM report_template_versions WHERE template_id = ? ORDER BY id",
                (template_id,),
            )
        ]

        ensure_report_template_schema(conn)
        conn.commit()

    ensure_database(str(db_path))
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        templates_after = [
            dict(row)
            for row in conn.execute("SELECT * FROM report_templates ORDER BY id")
        ]
        versions_after = [
            dict(row)
            for row in conn.execute(
                "SELECT * FROM report_template_versions ORDER BY id"
            )
        ]

    assert templates_after == [template_before]
    assert versions_after == versions_before


def test_persisted_custom_default_is_used_without_generic_reseed(
    tmp_path: Path,
) -> None:
    db_path = tmp_path / "custom-default.db"
    ensure_database(str(db_path))
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        template_id = install_custom_individual_template(conn)
        conn.commit()

    ensure_database(str(db_path))
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        selected = get_default_template(conn, "individual")
        rows = conn.execute(
            "SELECT id, name, is_default FROM report_templates ORDER BY id"
        ).fetchall()

    assert selected.id == template_id
    assert selected.name == "EPC Cliente Personalizado"
    assert selected.branding.primary_color == "#123456"
    assert [(row["id"], row["name"], row["is_default"]) for row in rows] == [
        (template_id, "EPC Cliente Personalizado", 1)
    ]


def test_legacy_standard_template_infers_renderer_without_rewriting_config(
    tmp_path: Path,
) -> None:
    db_path = tmp_path / "legacy-renderer-inference.db"
    ensure_database(str(db_path))
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            "SELECT id, config_json FROM report_templates WHERE name = 'Individual padrao'"
        ).fetchone()
        config = json.loads(row["config_json"])
        config.pop("renderer")
        legacy_json = json.dumps(config, ensure_ascii=False, indent=2)
        conn.execute(
            "UPDATE report_templates SET config_json = ? WHERE id = ?",
            (legacy_json, row["id"]),
        )
        conn.execute(
            "UPDATE report_template_versions SET config_json = ? WHERE template_id = ?",
            (legacy_json, row["id"]),
        )
        conn.commit()

        ensure_report_template_schema(conn)
        inferred = get_template(conn, row["id"])
        stored_template = conn.execute(
            "SELECT config_json FROM report_templates WHERE id = ?",
            (row["id"],),
        ).fetchone()[0]
        stored_version = conn.execute(
            "SELECT config_json FROM report_template_versions WHERE template_id = ?",
            (row["id"],),
        ).fetchone()[0]

    assert inferred.renderer == RENDERER_SOLCOR_INDIVIDUAL
    assert stored_template == legacy_json
    assert stored_version == legacy_json


def test_existing_template_version_remains_linked_to_historical_run(
    tmp_path: Path,
) -> None:
    db_path = tmp_path / "template-history-link.db"
    ensure_database(str(db_path))
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        template_id = install_custom_individual_template(conn)
        version = conn.execute(
            "SELECT id, version, config_json FROM report_template_versions WHERE template_id = ?",
            (template_id,),
        ).fetchone()
        conn.execute(
            """
            INSERT INTO report_generation_runs (
                template_id, template_version, report_type, asset_id,
                period_type, period_start, period_end, status,
                requested_count, completed_count, failed_count, created_at
            ) VALUES (?, ?, 'individual', 1956, 'monthly', '2026-06-01',
                      '2026-06-30', 'completed', 1, 1, 0, '2026-07-01T10:00:00')
            """,
            (template_id, version["version"]),
        )
        run_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        conn.commit()

    ensure_database(str(db_path))
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        linked = conn.execute(
            """
            SELECT r.id AS run_id, r.template_id, r.template_version,
                   v.id AS version_id, v.config_json
            FROM report_generation_runs r
            JOIN report_template_versions v
              ON v.template_id = r.template_id
             AND v.version = r.template_version
            WHERE r.id = ?
            """,
            (run_id,),
        ).fetchone()

    assert linked["template_id"] == template_id
    assert linked["template_version"] == version["version"]
    assert linked["version_id"] == version["id"]
    assert linked["config_json"] == version["config_json"]


def test_safe_filename_blocks_traversal_and_reserved_names() -> None:
    assert safe_filename("Cliente Instalação 2026 01", extension="pdf") == "Cliente_Instalacao_2026_01.pdf"
    assert safe_filename("CON", extension="pdf") == "_CON.pdf"
    try:
        safe_filename("../bad", extension="pdf")
    except ValueError as exc:
        assert str(exc) == "unsafe_filename"
    else:
        raise AssertionError("expected unsafe filename")


def test_portfolio_renderers_use_canonical_result(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    portfolio_id = add_portfolio(conn, add_asset(conn))
    profile = get_default_profile(conn, portfolio_id)
    result = prepare_portfolio_report(conn, portfolio_id=portfolio_id, portfolio_name="Output Portfolio", profile=profile, report_month="2026-01")
    template = get_default_template(conn, "portfolio", portfolio_id)

    html = render_portfolio_html(result, template)
    pdf = render_portfolio_pdf(result, template)
    excel = render_portfolio_excel(result, template)
    zipped = render_zip([pdf, excel])

    assert "Output Portfolio" in html
    assert pdf.content.startswith(b"%PDF-")
    assert excel.content.startswith(b"PK")
    assert zipped.content.startswith(b"PK")


def test_portfolio_defaults_have_explicit_generic_renderer() -> None:
    for name in (
        "Portfolio executivo",
        "Portfolio operacional",
        "Portfolio financeiro",
    ):
        template = default_template(name)
        assert template.renderer == RENDERER_GENERIC_PORTFOLIO
        assert template_to_config(template)["renderer"] == RENDERER_GENERIC_PORTFOLIO


def test_report_generation_routes_create_files_and_download(tmp_path: Path) -> None:
    db_path = tmp_path / "routes.db"
    ensure_database(str(db_path))
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    asset_id = add_asset(conn)
    portfolio_id = add_portfolio(conn, asset_id)
    template_id = next(row["id"] for row in list_templates(conn, "portfolio") if row["name"] == "Portfolio executivo")
    conn.close()

    flask_app = app_module.app
    previous_db = flask_app.config["DATABASE"]
    previous_testing = flask_app.config.get("TESTING")
    flask_app.config["DATABASE"] = str(db_path)
    flask_app.config["TESTING"] = True
    client = flask_app.test_client()
    with client.session_transaction() as session:
        session["authenticated"] = True
        session["csrf_token"] = "token"
    try:
        templates_page = client.get("/report-templates")
        assert templates_page.status_code == 200
        assert b"Templates" in templates_page.data
        preview = client.get(f"/report-generation/preview?portfolio_id={portfolio_id}&template_id={template_id}&report_month=2026-01")
        assert preview.status_code == 200
        assert b"Output Portfolio" in preview.data
        response = client.post(
            "/report-generation",
            data={
                "csrf_token": "token",
                "report_type": "portfolio",
                "template_id": str(template_id),
                "portfolio_id": str(portfolio_id),
                "report_month": "2026-01",
                "period_type": "monthly",
                "formats": ["pdf", "excel", "zip"],
            },
        )
        assert response.status_code in {302, 303}
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        file_row = conn.execute("SELECT * FROM report_generated_files WHERE format = 'zip' ORDER BY id DESC LIMIT 1").fetchone()
        conn.close()
        assert file_row is not None
        download = client.get(f"/report-generation/files/{file_row['id']}")
        assert download.status_code == 200
        assert download.mimetype == "application/zip"
    finally:
        flask_app.config["DATABASE"] = previous_db
        flask_app.config["TESTING"] = previous_testing


def test_individual_report_form_hides_portfolio_controls_and_requires_asset(
    tmp_path: Path,
) -> None:
    db_path = tmp_path / "individual-form.db"
    ensure_database(str(db_path))
    flask_app = app_module.app
    previous_db = flask_app.config["DATABASE"]
    previous_testing = flask_app.config.get("TESTING")
    flask_app.config["DATABASE"] = str(db_path)
    flask_app.config["TESTING"] = True
    client = flask_app.test_client()
    with client.session_transaction() as session:
        session["authenticated"] = True
        session["csrf_token"] = "token"
    try:
        page = client.get("/exports?tab=generate")
        stylesheet = client.get("/static/styles.css")
    finally:
        flask_app.config["DATABASE"] = previous_db
        flask_app.config["TESTING"] = previous_testing

    assert page.status_code == 200
    assert b'id="report-asset" required' in page.data
    assert b"Gerar, guardar e descarregar" in page.data
    assert b".reports-card [hidden]" in stylesheet.data
    assert b"display: none !important" in stylesheet.data


def test_exports_individual_asset_query_renders_selected_option(tmp_path: Path) -> None:
    db_path = tmp_path / "selected-asset.db"
    ensure_database(str(db_path))
    with sqlite3.connect(db_path) as conn:
        asset_id = add_asset(conn, "Selected Solar")
        add_fusionsolar_integration(conn, asset_id, "Selected Solar")
        conn.commit()

    with report_test_client(db_path) as client:
        page = client.get(
            "/exports",
            query_string={
                "tab": "generate",
                "asset_id": asset_id,
            },
        )

    assert page.status_code == 200
    html = page.get_data(as_text=True)
    option_start = html.index(f'<option value="{asset_id}" data-report-type=')
    assert " selected>" in html[option_start : option_start + 180]
    assert 'name="asset_id" id="report-asset" required' in html


def test_individual_generation_uses_persisted_custom_default(
    tmp_path: Path,
    monkeypatch,
) -> None:
    db_path = tmp_path / "custom-generation-template.db"
    ensure_database(str(db_path))
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        asset_id = add_asset(conn, "Custom Template Solar")
        template_id = install_custom_individual_template(conn)
        conn.commit()

    selected_templates = []

    def fake_report(conn, asset_id, period_job, **kwargs):
        return {
            "asset": {"project_name": "Custom Template Solar"},
            "period_label": "Junho 2026",
        }

    def fake_renderer(report, template):
        selected_templates.append(template)
        return app_factory_module.RenderedFile(
            filename="custom-template.pdf",
            content=b"%PDF-1.4 custom",
            mimetype="application/pdf",
            fmt="pdf",
        )

    monkeypatch.setattr(app_factory_module, "UPLOAD_DIR", tmp_path / "uploads")
    monkeypatch.setattr(
        app_factory_module,
        "store_runtime_relative_path",
        lambda path: str(path),
    )
    monkeypatch.setattr(
        app_factory_module,
        "build_individual_generation_report",
        fake_report,
    )
    monkeypatch.setattr(app_factory_module, "render_individual_pdf", fake_renderer)

    with report_test_client(db_path) as client:
        response = client.post(
            "/report-generation",
            data={
                "csrf_token": "token",
                "report_type": "individual",
                "asset_id": str(asset_id),
                "period_type": "monthly",
                "report_month": "2026-06",
                "formats": ["pdf"],
            },
        )

    assert response.status_code in {302, 303}
    assert len(selected_templates) == 1
    assert selected_templates[0].id == template_id
    assert selected_templates[0].name == "EPC Cliente Personalizado"
    assert selected_templates[0].title == "Relatório Solar {asset}"
    with sqlite3.connect(db_path) as conn:
        run = conn.execute(
            "SELECT template_id, template_version, status FROM report_generation_runs"
        ).fetchone()
    assert run == (template_id, 1, "completed")


def test_seeded_individual_epc_uses_recovered_solcor_layout() -> None:
    notice = "Rascunho - produção incompleta: 20/31 dias disponíveis"
    report = prepare_customer_report(
        {
            "asset": {
                "id": 1956,
                "project_name": "AGA (EPC)",
                "contract_type": "EPC (O&M)",
            },
            "period_type": "monthly",
            "period_start": "2026-07-01",
            "period_end": "2026-07-31",
            "period_label": "Julho 2026",
            "production_kwh": None,
            "self_use_kwh": None,
            "export_kwh": None,
            "consumption_kwh": None,
            "production_is_final": False,
            "daily_rows": [],
            "report_notes": [notice],
            "electricity_price": 0.2,
            "sell_price": 0.04,
            "coverage_pct": 0,
        }
    )

    rendered = render_individual_pdf(report, default_template("Individual padrao"))
    reader = PdfReader(io.BytesIO(rendered.content))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    page = reader.pages[0]
    xobjects = page["/Resources"].get("/XObject", {})

    assert len(reader.pages) == 1
    assert float(page.mediabox.width) > float(page.mediabox.height)
    assert "Relatório Mensal - Energia Solar" in text
    assert "Modelo EPC" in text
    assert "Produção Total" in text
    assert notice in text
    assert "Cover" not in text
    assert "Identification" not in text
    assert "Executive Summary" not in text
    assert len(xobjects) >= 1


def test_seeded_individual_esco_uses_recovered_solcor_layout_and_kpis() -> None:
    report = prepare_customer_report(
        {
            "asset": {
                "id": 2035,
                "project_name": "Sicobrita",
                "contract_type": "ESCO",
            },
            "period_type": "monthly",
            "period_start": "2026-06-01",
            "period_end": "2026-06-30",
            "period_label": "Junho 2026",
            "production_kwh": 1000,
            "self_use_kwh": 800,
            "export_kwh": 200,
            "consumption_kwh": 1200,
            "production_is_final": True,
            "daily_rows": [],
            "electricity_price": 0.2,
            "sell_price": 0.04,
        },
        solcor_price_per_kwh=0.08,
    )

    rendered = render_individual_pdf(report, default_template("Individual padrao"))
    reader = PdfReader(io.BytesIO(rendered.content))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)

    assert len(reader.pages) == 1
    assert float(reader.pages[0].mediabox.width) > float(reader.pages[0].mediabox.height)
    assert "Modelo ESCO" in text
    assert "Instalação operada em modelo ESCO" in text
    assert "Pagamento à Solcor" in text
    assert "Benefício Líquido" in text
    assert "Cover" not in text


def test_customised_individual_keeps_solcor_renderer_branding_and_sections(
    tmp_path: Path,
) -> None:
    conn = connect(tmp_path)
    template_id = install_custom_individual_template(conn)
    conn.commit()
    reloaded = get_template(conn, template_id)
    raw_config = json.loads(
        conn.execute(
            "SELECT config_json FROM report_templates WHERE id = ?",
            (template_id,),
        ).fetchone()[0]
    )
    conn.close()

    report = prepare_customer_report(
        {
            "asset": {
                "id": 1956,
                "project_name": "AGA Personalizada",
                "contract_type": "EPC",
            },
            "period_type": "monthly",
            "period_start": "2026-07-01",
            "period_end": "2026-07-31",
            "period_label": "Julho 2026",
            "production_kwh": 120,
            "self_use_kwh": 100,
            "export_kwh": 20,
            "consumption_kwh": 180,
            "production_is_final": True,
            "daily_rows": [],
            "electricity_price": 0.2,
            "sell_price": 0.04,
        }
    )
    rendered = render_individual_pdf(report, reloaded)
    reader = PdfReader(io.BytesIO(rendered.content))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    content_stream = reader.pages[0].get_contents().get_data()

    assert reloaded.renderer == RENDERER_SOLCOR_INDIVIDUAL
    assert raw_config["renderer"] == RENDERER_SOLCOR_INDIVIDUAL
    assert "Relatório Solar AGA Personalizada" in text
    assert "Desempenho energético" in text
    assert "Rodapé preservado" in text
    assert "Produção Diária de Eletricidade" in text
    assert "Destaques do Periodo" not in text
    assert b".070588 .203922 .337255 rg" in content_stream
    assert b".396078 .643137 .188235 rg" in content_stream


def test_individual_compact_remains_explicit_generic_visual_family() -> None:
    template = default_template("Individual compacto")
    report = {
        "asset": {"id": 1, "project_name": "Compact Test"},
        "report_type": "epc",
        "period_label": "Julho 2026",
        "period_type": "monthly",
        "period_start": "2026-07-01",
        "period_end": "2026-07-31",
        "production_kwh": 100,
        "self_use_kwh": 80,
        "export_kwh": 20,
        "consumption_kwh": 120,
        "net_benefit_eur": 12,
    }

    rendered = render_individual_pdf(report, template)
    reader = PdfReader(io.BytesIO(rendered.content))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)

    assert template.renderer == RENDERER_GENERIC_INDIVIDUAL
    assert float(reader.pages[0].mediabox.width) < float(reader.pages[0].mediabox.height)
    assert "Cover" in text
    assert "Identification" in text
    assert "Modelo EPC" not in text


def test_individual_generation_persists_scope_and_fallback_state(
    tmp_path: Path,
    monkeypatch,
) -> None:
    db_path = tmp_path / "individual-state.db"
    ensure_database(str(db_path))
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        asset_id = add_asset(conn, "Stateful Solar")
        add_fusionsolar_integration(conn, asset_id, "Stateful Solar")
        template_id = next(
            row["id"]
            for row in list_templates(conn, "individual")
            if row["name"] == "Individual padrao"
        )
        conn.commit()

    def fake_report(conn, asset_id, period_job, **kwargs):
        return {
            "asset": {"project_name": "Stateful Solar"},
            "period_label": "Abril 2026",
        }

    def fake_pdf(report, template):
        return app_factory_module.RenderedFile(
            filename="stateful.pdf",
            content=b"%PDF-1.4 fake",
            mimetype="application/pdf",
            fmt="pdf",
        )

    def fake_excel(report, template):
        return app_factory_module.RenderedFile(
            filename="stateful.xlsx",
            content=b"PK fake",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            fmt="xlsx",
        )

    monkeypatch.setattr(app_factory_module, "UPLOAD_DIR", tmp_path / "uploads")
    monkeypatch.setattr(app_factory_module, "store_runtime_relative_path", lambda path: str(path))
    monkeypatch.setattr(app_factory_module, "build_individual_generation_report", fake_report)
    monkeypatch.setattr(app_factory_module, "render_individual_pdf", fake_pdf)
    monkeypatch.setattr(app_factory_module, "render_individual_excel", fake_excel)

    with report_test_client(db_path) as client:
        response = client.post(
            "/report-generation",
            data={
                "csrf_token": "token",
                "report_type": "individual",
                "asset_id": str(asset_id),
                "portfolio_id": "999",
                "profile_id": "999",
                "template_id": str(template_id),
                "period_type": "monthly",
                "report_month": "2026-04",
                "report_year": "2026",
                "report_quarter": "2",
                "report_semester": "1",
                "include_availability": "on",
                "formats": ["pdf", "excel"],
            },
        )
        assert response.status_code in {302, 303}
        location = response.headers["Location"]
        parsed = urlparse(location)
        query = parse_qs(parsed.query)
        assert parsed.path == "/exports"
        assert parsed.fragment == "reports-history"
        assert query["tab"] == ["history"]
        assert query["report_type"] == ["individual"]
        assert query["asset_id"] == [str(asset_id)]
        assert query["template_id"] == [str(template_id)]
        assert query["period_type"] == ["monthly"]
        assert query["report_month"] == ["2026-04"]
        assert query["report_year"] == ["2026"]
        assert query["report_quarter"] == ["2"]
        assert query["report_semester"] == ["1"]
        assert query["include_availability"] == ["on"]
        assert "portfolio_id" not in query
        assert "profile_id" not in query
        assert query["run_id"][0].isdigit()

        history = client.get(location)

    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        run = conn.execute(
            "SELECT * FROM report_generation_runs ORDER BY id DESC LIMIT 1"
        ).fetchone()
        files = conn.execute(
            "SELECT * FROM report_generated_files WHERE run_id = ? ORDER BY id",
            (run["id"],),
        ).fetchall()

    assert run["status"] == "completed"
    assert run["asset_id"] == asset_id
    assert run["portfolio_id"] is None
    assert run["completed_count"] == 2
    assert len(files) == 2
    assert all(row["asset_id"] == asset_id for row in files)
    assert all(row["portfolio_id"] is None for row in files)
    assert any(row["format"] == "pdf" and row["status"] == "completed" for row in files)

    history_html = history.get_data(as_text=True)
    option_start = history_html.index(f'<option value="{asset_id}" data-report-type=')
    assert " selected>" in history_html[option_start : option_start + 180]
    template_option_start = history_html.index(
        f'<option value="{template_id}" data-template-type="individual"'
    )
    assert " selected>" in history_html[
        template_option_start : template_option_start + 260
    ]
    assert "Último relatório compatível" in history_html
    assert "stateful.pdf" in history_html
    assert f"#{run['id']}" in history_html
    assert "Ainda não existem gerações." not in history_html


def test_missing_individual_asset_is_rejected_without_persistence(tmp_path: Path) -> None:
    db_path = tmp_path / "missing-individual-asset.db"
    ensure_database(str(db_path))
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        template_id = next(
            row["id"]
            for row in list_templates(conn, "individual")
            if row["name"] == "Individual padrao"
        )

    with report_test_client(db_path) as client:
        response = client.post(
            "/report-generation",
            data={
                "csrf_token": "token",
                "report_type": "individual",
                "template_id": str(template_id),
                "period_type": "monthly",
                "report_month": "2026-04",
                "formats": ["pdf"],
            },
        )

    assert response.status_code in {302, 303}
    query = parse_qs(urlparse(response.headers["Location"]).query)
    assert query["report_type"] == ["individual"]
    assert "asset_id" not in query
    with sqlite3.connect(db_path) as conn:
        assert conn.execute("SELECT COUNT(*) FROM report_generation_runs").fetchone()[0] == 0
        assert conn.execute("SELECT COUNT(*) FROM report_generated_files").fetchone()[0] == 0


def test_portfolio_generation_preserves_portfolio_scope_without_asset_leak(
    tmp_path: Path,
    monkeypatch,
) -> None:
    db_path = tmp_path / "portfolio-state.db"
    ensure_database(str(db_path))
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        asset_id = add_asset(conn, "Portfolio Member")
        portfolio_id = add_portfolio(conn, asset_id)
        profile = get_default_profile(conn, portfolio_id)
        template_id = next(
            row["id"]
            for row in list_templates(conn, "portfolio")
            if row["name"] == "Portfolio executivo"
        )

    def fake_portfolio_pdf(result, template):
        return app_factory_module.RenderedFile(
            filename="portfolio.pdf",
            content=b"%PDF-1.4 fake",
            mimetype="application/pdf",
            fmt="pdf",
            asset_id=asset_id,
        )

    def fake_portfolio_excel(result, template):
        return app_factory_module.RenderedFile(
            filename="portfolio.xlsx",
            content=b"PK fake",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            fmt="xlsx",
            asset_id=asset_id,
        )

    monkeypatch.setattr(app_factory_module, "UPLOAD_DIR", tmp_path / "uploads")
    monkeypatch.setattr(app_factory_module, "store_runtime_relative_path", lambda path: str(path))
    monkeypatch.setattr(app_factory_module, "render_portfolio_pdf", fake_portfolio_pdf)
    monkeypatch.setattr(app_factory_module, "render_portfolio_excel", fake_portfolio_excel)

    with report_test_client(db_path) as client:
        response = client.post(
            "/report-generation",
            data={
                "csrf_token": "token",
                "report_type": "portfolio",
                "asset_id": str(asset_id),
                "portfolio_id": str(portfolio_id),
                "profile_id": str(profile.id),
                "template_id": str(template_id),
                "period_type": "monthly",
                "report_month": "2026-01",
                "formats": ["pdf", "excel"],
            },
        )
        location = response.headers["Location"]
        page = client.get(location)

    query = parse_qs(urlparse(location).query)
    assert query["report_type"] == ["portfolio"]
    assert query["portfolio_id"] == [str(portfolio_id)]
    assert query["profile_id"] == [str(profile.id)]
    assert "asset_id" not in query
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        run = conn.execute(
            "SELECT * FROM report_generation_runs ORDER BY id DESC LIMIT 1"
        ).fetchone()
        files = conn.execute(
            "SELECT * FROM report_generated_files WHERE run_id = ? ORDER BY id",
            (run["id"],),
        ).fetchall()
    assert run["status"] == "completed"
    assert run["portfolio_id"] == portfolio_id
    assert run["asset_id"] is None
    assert files
    assert all(row["portfolio_id"] == portfolio_id for row in files)
    assert all(row["asset_id"] is None for row in files)
    page_html = page.get_data(as_text=True)
    assert 'value="portfolio" checked' in page_html
    assert f'<option value="{portfolio_id}" selected>' in page_html


def test_report_scope_changes_only_replace_browser_url(tmp_path: Path) -> None:
    db_path = tmp_path / "report-scope-js.db"
    ensure_database(str(db_path))
    with report_test_client(db_path) as client:
        script = client.get("/static/reports.js").get_data(as_text=True)
        page = client.get("/exports?tab=generate").get_data(as_text=True)

    assert "window.history.replaceState" in script
    assert "select[name='asset_id']" in script
    assert "select[name='portfolio_id']" in script
    assert "fetch(" not in script
    assert 'name="action" value="save_billing_config"' in page


def test_individual_preview_is_rejected_without_internal_error(
    tmp_path: Path,
) -> None:
    db_path = tmp_path / "individual-preview.db"
    ensure_database(str(db_path))
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        template_id = next(
            row["id"]
            for row in list_templates(conn, "individual")
            if row["name"] == "Individual padrao"
        )
    flask_app = app_module.app
    previous_db = flask_app.config["DATABASE"]
    previous_testing = flask_app.config.get("TESTING")
    flask_app.config["DATABASE"] = str(db_path)
    flask_app.config["TESTING"] = True
    client = flask_app.test_client()
    with client.session_transaction() as session:
        session["authenticated"] = True
        session["csrf_token"] = "token"
    try:
        response = client.get(
            "/report-generation/preview",
            query_string={
                "report_type": "individual",
                "template_id": template_id,
                "report_month": "2026-07",
            },
            follow_redirects=True,
        )
    finally:
        flask_app.config["DATABASE"] = previous_db
        flask_app.config["TESTING"] = previous_testing

    assert response.status_code == 200
    assert b"Erro interno" not in response.data
    assert "apenas para relatórios de portefólio" in response.get_data(
        as_text=True
    )


def test_single_manual_pdf_is_saved_and_downloaded(
    tmp_path: Path,
    monkeypatch,
) -> None:
    db_path = tmp_path / "single-download.db"
    ensure_database(str(db_path))
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        asset_id = add_asset(conn, "Download Solar")
        template_id = next(
            row["id"]
            for row in list_templates(conn, "individual")
            if row["name"] == "Individual padrao"
        )
        conn.commit()

    def fake_report(conn, asset_id, period_job, **kwargs):
        return {
            "asset": {"id": asset_id, "project_name": "Download Solar"},
            "period_label": "Junho 2026",
            "period_type": "monthly",
            "period_start": "2026-06-01",
            "period_end": "2026-06-30",
            "production_kwh": 100,
            "self_use_kwh": 70,
            "export_kwh": 30,
            "consumption_kwh": 90,
            "grid_import_kwh": 20,
            "net_benefit_eur": 10,
        }

    output_root = tmp_path / "uploads"
    monkeypatch.setattr(app_factory_module, "UPLOAD_DIR", output_root)
    monkeypatch.setattr(
        app_factory_module,
        "store_runtime_relative_path",
        lambda path: str(path),
    )
    monkeypatch.setattr(
        app_factory_module,
        "build_individual_generation_report",
        fake_report,
    )
    flask_app = app_module.app
    previous_db = flask_app.config["DATABASE"]
    previous_testing = flask_app.config.get("TESTING")
    flask_app.config["DATABASE"] = str(db_path)
    flask_app.config["TESTING"] = True
    client = flask_app.test_client()
    with client.session_transaction() as session:
        session["authenticated"] = True
        session["csrf_token"] = "token"
    try:
        response = client.post(
            "/report-generation",
            data={
                "csrf_token": "token",
                "report_type": "individual",
                "template_id": str(template_id),
                "asset_id": str(asset_id),
                "report_month": "2026-06",
                "period_type": "monthly",
                "formats": ["pdf"],
            },
        )
        assert response.status_code in {302, 303}
        assert "/report-generation/files/" in response.headers["Location"]
        download = client.get(response.headers["Location"])
    finally:
        flask_app.config["DATABASE"] = previous_db
        flask_app.config["TESTING"] = previous_testing

    assert download.status_code == 200
    assert download.mimetype == "application/pdf"
    assert download.headers["Content-Disposition"].startswith("attachment;")
    with sqlite3.connect(db_path) as conn:
        run = conn.execute(
            "SELECT status, completed_count, failed_count "
            "FROM report_generation_runs ORDER BY id DESC LIMIT 1"
        ).fetchone()
    assert run == ("completed", 1, 0)


def test_batch_partial_counts_metadata_and_auxiliary_zip(tmp_path: Path, monkeypatch) -> None:
    db_path = tmp_path / "batch.db"
    ensure_database(str(db_path))
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    asset_ids = [add_asset(conn, f"Batch {index}") for index in range(3)]
    template_id = next(row["id"] for row in list_templates(conn, "individual") if row["name"] == "Individual padrao")
    conn.close()

    def fake_report(conn, *, asset_id, report_month, electricity_price, sell_price, billing_config, period=None, **kwargs):
        if asset_id == asset_ids[1]:
            return None
        return {
            "asset_id": asset_id,
            "asset": {"id": asset_id, "project_name": f"Batch {asset_id}"},
            "period_label": report_month,
            "period_type": "monthly",
            "period_start": f"{report_month}-01",
            "period_end": f"{report_month}-28",
            "production_kwh": 10,
            "self_use_kwh": 7,
            "export_kwh": 3,
            "consumption_kwh": 9,
            "grid_import_kwh": 2,
            "net_benefit_eur": 1,
        }

    monkeypatch.setattr(app_factory_module, "build_local_customer_production_report", fake_report)
    flask_app = app_module.app
    previous_db = flask_app.config["DATABASE"]
    previous_testing = flask_app.config.get("TESTING")
    flask_app.config["DATABASE"] = str(db_path)
    flask_app.config["TESTING"] = True
    client = flask_app.test_client()
    with client.session_transaction() as session:
        session["authenticated"] = True
        session["csrf_token"] = "token"
    try:
        response = client.post(
            "/report-generation",
            data={
                "csrf_token": "token",
                "report_type": "individual",
                "template_id": str(template_id),
                "asset_ids": [str(item) for item in asset_ids],
                "report_month": "2026-01",
                "formats": ["pdf", "zip"],
            },
        )
        assert response.status_code in {302, 303}
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        run = conn.execute("SELECT * FROM report_generation_runs ORDER BY id DESC LIMIT 1").fetchone()
        files = conn.execute("SELECT * FROM report_generated_files WHERE run_id = ? ORDER BY id", (run["id"],)).fetchall()
        conn.close()
        assert run["status"] == "partial"
        assert run["requested_count"] == 3
        assert run["completed_count"] == 2
        assert run["failed_count"] == 1
        assert {row["asset_id"] for row in files if row["status"] == "completed" and row["is_auxiliary"] == 0} == {asset_ids[0], asset_ids[2]}
        assert sum(row["is_auxiliary"] for row in files) == 1
    finally:
        flask_app.config["DATABASE"] = previous_db
        flask_app.config["TESTING"] = previous_testing


def test_portfolio_two_periods_pdf_excel_zip_counts_zip_auxiliary(tmp_path: Path) -> None:
    db_path = tmp_path / "portfolio-batch.db"
    ensure_database(str(db_path))
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    portfolio_id = add_portfolio(conn, add_asset(conn))
    template_id = next(row["id"] for row in list_templates(conn, "portfolio") if row["name"] == "Portfolio executivo")
    conn.close()
    flask_app = app_module.app
    previous_db = flask_app.config["DATABASE"]
    previous_testing = flask_app.config.get("TESTING")
    flask_app.config["DATABASE"] = str(db_path)
    flask_app.config["TESTING"] = True
    client = flask_app.test_client()
    with client.session_transaction() as session:
        session["authenticated"] = True
        session["csrf_token"] = "token"
    try:
        response = client.post(
            "/report-generation",
            data={
                "csrf_token": "token",
                "report_type": "portfolio",
                "template_id": str(template_id),
                "portfolio_id": str(portfolio_id),
                "report_months": "2026-01,2026-02",
                "formats": ["pdf", "excel", "zip"],
            },
        )
        assert response.status_code in {302, 303}
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        run = conn.execute("SELECT * FROM report_generation_runs ORDER BY id DESC LIMIT 1").fetchone()
        files = conn.execute("SELECT * FROM report_generated_files WHERE run_id = ?", (run["id"],)).fetchall()
        conn.close()
        assert run["requested_count"] == 4
        assert run["completed_count"] == 4
        assert run["failed_count"] == 0
        assert sum(row["is_auxiliary"] for row in files) == 1
    finally:
        flask_app.config["DATABASE"] = previous_db
        flask_app.config["TESTING"] = previous_testing


def test_invalid_formats_and_limits_are_rejected(tmp_path: Path) -> None:
    db_path = tmp_path / "invalid.db"
    ensure_database(str(db_path))
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    template_id = next(row["id"] for row in list_templates(conn, "portfolio") if row["name"] == "Portfolio executivo")
    conn.close()
    flask_app = app_module.app
    previous_db = flask_app.config["DATABASE"]
    previous_testing = flask_app.config.get("TESTING")
    flask_app.config["DATABASE"] = str(db_path)
    flask_app.config["TESTING"] = True
    client = flask_app.test_client()
    with client.session_transaction() as session:
        session["authenticated"] = True
        session["csrf_token"] = "token"
    try:
        response = client.post("/report-generation", data={"csrf_token": "token", "report_type": "portfolio", "template_id": str(template_id), "formats": ["zip"]})
        assert response.status_code in {302, 303}
        conn = sqlite3.connect(db_path)
        assert conn.execute("SELECT COUNT(*) FROM report_generation_runs").fetchone()[0] == 0
        conn.close()
    finally:
        flask_app.config["DATABASE"] = previous_db
        flask_app.config["TESTING"] = previous_testing


def test_individual_excel_reconciles_with_canonical_report(tmp_path: Path) -> None:
    template = default_template("Individual padrao")
    report = {
        "asset": {"id": 1, "project_name": "Excel Individual"},
        "period_label": "Janeiro 2026",
        "period_type": "monthly",
        "period_start": "2026-01-01",
        "period_end": "2026-01-31",
        "production_kwh": 100,
        "self_use_kwh": 70,
        "export_kwh": 30,
        "consumption_kwh": 90,
        "grid_import_kwh": 20,
        "solcor_payment_eur": 7,
        "fixed_monthly_fee_eur": 0,
        "net_benefit_eur": 10,
    }
    rendered = render_individual_excel(report, template)
    path = tmp_path / rendered.filename
    path.write_bytes(rendered.content)
    workbook = load_workbook(path)
    values = {row[0].value: row[1].value for row in workbook["Energia"].iter_rows(min_row=2)}

    assert workbook.sheetnames == ["Resumo", "Energia", "Financeiro", "Qualidade dos dados", "Metadados"]
    assert values["production_kwh"] == report["production_kwh"]
    assert values["self_use_kwh"] == report["self_use_kwh"]


def test_individual_pdf_and_excel_include_readable_incomplete_production_warning(tmp_path: Path) -> None:
    template = default_template("Individual padrao")
    notice = "Rascunho — produção incompleta: 23/30 dias disponíveis"
    report = {
        "asset": {"id": 1, "project_name": "Relatorio Parcial"},
        "period_label": "Abril 2026",
        "period_type": "monthly",
        "period_start": "2026-04-01",
        "period_end": "2026-04-30",
        "production_kwh": None,
        "raw_daily_total_kwh": 230,
        "production_quality_status": "partial",
        "coverage_pct": 0,
        "report_notes": [notice],
        "warnings": ["partial_monthly_production", "production_not_final"],
    }

    excel = render_individual_excel(report, template)
    excel_path = tmp_path / excel.filename
    excel_path.write_bytes(excel.content)
    workbook = load_workbook(excel_path)
    quality_values = [cell.value for row in workbook["Qualidade dos dados"].iter_rows() for cell in row]

    pdf = render_individual_pdf(report, template)
    pdf_path = tmp_path / pdf.filename
    pdf_path.write_bytes(pdf.content)
    pdf_text = "\n".join(page.extract_text() or "" for page in PdfReader(str(pdf_path)).pages)

    assert notice in quality_values
    assert "Rascunho" in pdf_text
    assert "23/30 dias" in pdf_text


def test_section_order_is_shared_by_preview_and_pdf(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    portfolio_id = add_portfolio(conn, add_asset(conn))
    result = prepare_portfolio_report(conn, portfolio_id=portfolio_id, portfolio_name="Order Portfolio", profile=get_default_profile(conn, portfolio_id), report_month="2026-01")
    template = default_template("Portfolio executivo")
    template = template.__class__(
        **{
            **template.__dict__,
            "sections": (
                TemplateSection("warnings", "Warnings First", True, 10),
                TemplateSection("kpis", "KPIs Second", True, 20),
                TemplateSection("installations_table", "Table Third", True, 30),
            ),
        }
    )
    html = render_portfolio_html(result, template)
    pdf = render_portfolio_pdf(result, template)
    pdf_path = tmp_path / "order.pdf"
    pdf_path.write_bytes(pdf.content)
    text = "\n".join(page.extract_text() or "" for page in PdfReader(str(pdf_path)).pages)

    assert html.index("Warnings First") < html.index("KPIs Second") < html.index("Table Third")
    assert text.index("Warnings First") < text.index("KPIs Second") < text.index("Table Third")


def test_portfolio_draft_hides_financials_in_html_pdf_and_excel(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    portfolio_id = add_portfolio(conn, add_asset(conn, "Draft Portfolio Asset"))
    result = prepare_portfolio_report(
        conn,
        portfolio_id=portfolio_id,
        portfolio_name="Draft Portfolio",
        profile=get_default_profile(conn, portfolio_id),
        report_month="2026-01",
        reference_date=date(2026, 2, 1),
    )
    template = default_template("Portfolio executivo")

    html = render_portfolio_html(result, template)
    assert "Indisponível — rascunho" in html
    assert result.summary.values["estimated_value_eur"] is None

    excel = render_portfolio_excel(result, template)
    excel_path = tmp_path / excel.filename
    excel_path.write_bytes(excel.content)
    workbook = load_workbook(excel_path)
    summary_values = {
        row[0].value: row[1].value
        for row in workbook["Resumo"].iter_rows()
        if len(row) > 1 and row[0].value
    }
    assert summary_values["Valor autoconsumo"] is None
    quality_text = [cell.value for row in workbook["Qualidade dos dados"].iter_rows() for cell in row]
    assert "production_financials_not_final" in quality_text

    pdf = render_portfolio_pdf(result, template)
    pdf_path = tmp_path / pdf.filename
    pdf_path.write_bytes(pdf.content)
    pdf_text = "\n".join(page.extract_text() or "" for page in PdfReader(str(pdf_path)).pages)
    assert "Indisponível" in pdf_text
    assert "rascunho" in pdf_text


def test_portfolio_pdf_includes_more_than_ten_columns(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    portfolio_id = add_portfolio(conn, add_asset(conn))
    result = prepare_portfolio_report(conn, portfolio_id=portfolio_id, portfolio_name="Wide Portfolio", profile=get_default_profile(conn, portfolio_id), report_month="2026-01")
    template = default_template("Portfolio operacional")
    pdf = render_portfolio_pdf(result, template)
    pdf_path = tmp_path / "wide.pdf"
    pdf_path.write_bytes(pdf.content)
    text = "\n".join(page.extract_text() or "" for page in PdfReader(str(pdf_path)).pages)

    assert len(result.columns) > 10
    assert "Beneficio liquido" in text
    assert "Warnings" in text


def test_template_client_scope_is_strict() -> None:
    template = default_template("Portfolio executivo")
    client_template = template.__class__(**{**template.__dict__, "client_key": "cliente-a"})

    validate_template_scope(template, "portfolio", portfolio_id=None, client_key=None)
    validate_template_scope(client_template, "portfolio", portfolio_id=None, client_key="cliente-a")
    for client_key in (None, "", "cliente-b"):
        try:
            validate_template_scope(client_template, "portfolio", portfolio_id=None, client_key=client_key)
        except ValueError as exc:
            assert str(exc) == "template_client_mismatch"
        else:
            raise AssertionError("expected client mismatch")


def test_snapshot_rejects_multiple_periods(tmp_path: Path) -> None:
    db_path = tmp_path / "snapshot-period.db"
    ensure_database(str(db_path))
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    portfolio_id = add_portfolio(conn, add_asset(conn))
    profile = get_default_profile(conn, portfolio_id)
    from monitoring_board.portfolio_report_repository import snapshot_portfolio_result

    result = prepare_portfolio_report(conn, portfolio_id=portfolio_id, portfolio_name="Snapshot P", profile=profile, report_month="2026-01")
    snapshot_id = snapshot_portfolio_result(conn, result)
    template_id = next(row["id"] for row in list_templates(conn, "portfolio") if row["name"] == "Portfolio executivo")
    conn.commit()
    conn.close()
    flask_app = app_module.app
    previous_db = flask_app.config["DATABASE"]
    previous_testing = flask_app.config.get("TESTING")
    flask_app.config["DATABASE"] = str(db_path)
    flask_app.config["TESTING"] = True
    client = flask_app.test_client()
    with client.session_transaction() as session:
        session["authenticated"] = True
        session["csrf_token"] = "token"
    try:
        client.post(
            "/report-generation",
            data={
                "csrf_token": "token",
                "report_type": "portfolio",
                "template_id": str(template_id),
                "portfolio_id": str(portfolio_id),
                "snapshot_id": str(snapshot_id),
                "report_months": "2026-01,2026-02",
                "formats": ["pdf", "excel"],
            },
        )
        conn = sqlite3.connect(db_path)
        assert conn.execute("SELECT COUNT(*) FROM report_generation_runs").fetchone()[0] == 0
        conn.close()
    finally:
        flask_app.config["DATABASE"] = previous_db
        flask_app.config["TESTING"] = previous_testing


def test_logo_upload_validation_and_storage() -> None:
    payload = b"\x89PNG\r\n\x1a\n" + b"\x00" * 8 + struct.pack(">II", 120, 40) + b"\x08\x02\x00\x00\x00" + b"x"

    class File:
        filename = "logo.png"

        def read(self):
            return payload

    path = app_factory_module.store_report_logo(File())
    assert path.endswith(".png")
    bad = File()
    bad.filename = "logo.svg"
    try:
        app_factory_module.store_report_logo(bad)
    except ValueError as exc:
        assert str(exc) == "invalid_logo_extension"
    else:
        raise AssertionError("expected invalid extension")


def test_storage_reconciliation_detects_missing_orphan_and_hash(tmp_path: Path) -> None:
    db_path = tmp_path / "storage.db"
    ensure_database(str(db_path))
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    root = tmp_path / "generated"
    root.mkdir()
    ok = root / "ok.pdf"
    ok.write_bytes(b"abc")
    orphan = root / "orphan.pdf"
    orphan.write_bytes(b"orphan")
    conn.execute("INSERT INTO report_generation_runs (report_type, status, requested_count, created_at) VALUES ('portfolio', 'completed', 1, '2026-01-01')")
    run_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.execute(
        """
        INSERT INTO report_generated_files
            (run_id, format, filename, relative_path, sha256, size_bytes, status, created_at)
        VALUES (?, 'pdf', 'ok.pdf', ?, 'bad-hash', 3, 'completed', '2026-01-01')
        """,
        (run_id, str(ok)),
    )

    statuses = {finding.status for finding in reconcile_generated_reports(conn, root=root)}
    assert {"hash_mismatch", "orphan_file"} <= statuses


def test_storage_reconciliation_ignores_failed_rows_without_file(tmp_path: Path) -> None:
    db_path = tmp_path / "storage-failed.db"
    ensure_database(str(db_path))
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    root = tmp_path / "generated"
    root.mkdir()
    conn.execute("INSERT INTO report_generation_runs (report_type, status, requested_count, created_at) VALUES ('portfolio', 'partial', 1, '2026-01-01')")
    run_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.execute(
        """
        INSERT INTO report_generated_files
            (run_id, format, filename, relative_path, sha256, size_bytes, status, error_message, created_at)
        VALUES (?, 'pdf', 'failed', '', '', 0, 'failed', 'Sem dados', '2026-01-01')
        """,
        (run_id,),
    )

    statuses = {finding.status for finding in reconcile_generated_reports(conn, root=root)}
    assert "invalid_path" not in statuses


def test_storage_reconciliation_resolves_runtime_relative_paths_from_explicit_root(tmp_path: Path) -> None:
    db_path = tmp_path / "storage-root.db"
    ensure_database(str(db_path))
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    root = tmp_path / "uploads" / "generated_reports"
    root.mkdir(parents=True)
    report = root / "1" / "ok.pdf"
    report.parent.mkdir()
    report.write_bytes(b"abc")
    conn.execute("INSERT INTO report_generation_runs (report_type, status, requested_count, created_at) VALUES ('portfolio', 'completed', 1, '2026-01-01')")
    run_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.execute(
        """
        INSERT INTO report_generated_files
            (run_id, format, filename, relative_path, sha256, size_bytes, status, created_at)
        VALUES (?, 'pdf', 'ok.pdf', 'uploads/generated_reports/1/ok.pdf', ?, 3, 'completed', '2026-01-01')
        """,
        (run_id, hashlib.sha256(b"abc").hexdigest()),
    )

    statuses = {finding.status for finding in reconcile_generated_reports(conn, root=root)}
    assert "invalid_path" not in statuses
    assert "ok" in statuses


def test_ensure_database_is_idempotent_for_reporting_outputs(tmp_path: Path) -> None:
    db_path = tmp_path / "idempotent.db"
    ensure_database(str(db_path))
    ensure_database(str(db_path))
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        templates = conn.execute(
            "SELECT name, report_type, active, is_default, config_json FROM report_templates ORDER BY id"
        ).fetchall()

    assert [row["name"] for row in templates] == [
        "Individual padrao",
        "Individual compacto",
        "Portfolio executivo",
        "Portfolio operacional",
        "Portfolio financeiro",
    ]
    assert all(row["active"] == 1 and json.loads(row["config_json"]) for row in templates)
    renderers = {
        row["name"]: json.loads(row["config_json"])["renderer"]
        for row in templates
    }
    assert renderers == {
        "Individual padrao": RENDERER_SOLCOR_INDIVIDUAL,
        "Individual compacto": RENDERER_GENERIC_INDIVIDUAL,
        "Portfolio executivo": RENDERER_GENERIC_PORTFOLIO,
        "Portfolio operacional": RENDERER_GENERIC_PORTFOLIO,
        "Portfolio financeiro": RENDERER_GENERIC_PORTFOLIO,
    }
    assert {
        (row["report_type"], row["name"])
        for row in templates
        if row["is_default"]
    } == {
        ("individual", "Individual padrao"),
        ("portfolio", "Portfolio executivo"),
    }


def test_reporting_health_route(tmp_path: Path) -> None:
    db_path = tmp_path / "health.db"
    ensure_database(str(db_path))
    flask_app = app_module.app
    previous_db = flask_app.config["DATABASE"]
    previous_testing = flask_app.config.get("TESTING")
    flask_app.config["DATABASE"] = str(db_path)
    flask_app.config["TESTING"] = True
    client = flask_app.test_client()
    with client.session_transaction() as session:
        session["authenticated"] = True
        session["csrf_token"] = "token"
    try:
        response = client.get("/reporting-health")
        assert response.status_code == 200
        assert response.json["database"] == "ok"
    finally:
        flask_app.config["DATABASE"] = previous_db
        flask_app.config["TESTING"] = previous_testing
