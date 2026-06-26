from __future__ import annotations

import io
import sqlite3
from pathlib import Path

from openpyxl import load_workbook

import app as app_module
from app import ensure_database
from monitoring_board import portfolio_management as domain
from monitoring_board import portfolio_repository as repo


def connect(tmp_path: Path) -> sqlite3.Connection:
    db_path = tmp_path / "portfolio-management.db"
    ensure_database(str(db_path))
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def add_asset(conn: sqlite3.Connection, name: str, nif: str = "") -> int:
    cursor = conn.execute(
        "INSERT INTO assets (project_name, nif, active_contract, kwp, alias_blob) VALUES (?, ?, 'yes', '10', '')",
        (name, nif),
    )
    asset_id = int(cursor.lastrowid)
    conn.commit()
    return asset_id


def csrf_client(db_path: Path):
    flask_app = app_module.app
    previous_db = flask_app.config["DATABASE"]
    previous_testing = flask_app.config.get("TESTING")
    flask_app.config["DATABASE"] = str(db_path)
    flask_app.config["TESTING"] = True
    client = flask_app.test_client()
    with client.session_transaction() as session:
        session["authenticated"] = True
        session["csrf_token"] = "token"
    return flask_app, previous_db, previous_testing, client


def test_portfolio_name_normalization_is_conservative() -> None:
    assert domain.normalize_name("Marmores Granja, Lda.") == domain.normalize_name("MARMORES GRANJA LDA")
    assert domain.normalize_name("Marmores Granja") == domain.normalize_name("Marmores Granja, Lda.")
    assert domain.normalize_name("A & B - Energia S.A.") == "a e b energia"
    assert domain.normalize_name("Grupo Lda Norte") == "grupo lda norte"
    assert domain.normalize_name("Marmores Granja Norte") != domain.normalize_name("Marmores Granja Sul")
    assert domain.compare_names("Marmores Granja Norte", "Marmores Granja Sul") < domain.STRONG_SUGGESTION_THRESHOLD


def test_alias_crud_conflicts_and_learning(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    asset_a = add_asset(conn, "Marmores Granja Norte", "501111111")
    asset_b = add_asset(conn, "Marmores Granja Sul", "502222222")

    alias_id = repo.upsert_alias(conn, asset_id=asset_a, alias_name="Marmores Granja", source="manual")
    assert alias_id is not None
    assert repo.upsert_alias(conn, asset_id=asset_a, alias_name="Marmores Granja", source="manual") is None
    try:
        repo.upsert_alias(conn, asset_id=asset_b, alias_name="Marmores Granja", source="manual")
    except ValueError as exc:
        assert str(exc) == "alias_conflict"
    else:
        raise AssertionError("expected alias conflict")

    repo.toggle_alias(conn, asset_id=asset_a, alias_id=alias_id, active=False)
    assert repo.list_aliases(conn, asset_id=asset_a)[0]["active"] == 0
    repo.toggle_alias(conn, asset_id=asset_a, alias_id=alias_id, active=True)
    repo.update_alias(conn, asset_id=asset_a, alias_id=alias_id, alias_name="Marmores Granja Norte Alias")
    assert repo.list_aliases(conn, asset_id=asset_a)[0]["source"] == "manual"

    portfolio_id = repo.create_portfolio(conn, name="Teste Alias")
    member_id = repo.add_member(conn, portfolio_id=portfolio_id, asset_id=None, external_name="Cliente Aprendido")
    repo.confirm_mapping(conn, member_id=member_id, portfolio_id=portfolio_id, asset_id=asset_b, create_alias=True)
    learned = conn.execute("SELECT * FROM asset_aliases WHERE asset_id = ? AND normalized_alias = ?", (asset_b, domain.normalize_name("Cliente Aprendido"))).fetchone()
    assert learned["source"] == "mapping_confirmed"


def test_mapping_decisions_exact_fuzzy_conflict_and_manual(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    asset_a = add_asset(conn, "Marmores Granja Norte", "501111111")
    asset_b = add_asset(conn, "Marmores Granja Sul", "502222222")
    repo.upsert_alias(conn, asset_id=asset_a, alias_name="Pedra Norte", source="manual")

    nif = repo.suggest_mapping(conn, external_name="Outro", nif="501111111")
    alias = repo.suggest_mapping(conn, external_name="Pedra Norte", nif="")
    fuzzy = repo.suggest_mapping(conn, external_name="Marmores Granja Nort", nif="")
    close = repo.suggest_mapping(conn, external_name="Marmores Granja", nif="")
    discordant = repo.suggest_mapping(conn, external_name="Pedra Norte", nif="502222222")

    assert nif.auto_mappable is True and nif.asset_id == asset_a and nif.method == "nif_exact"
    assert alias.auto_mappable is True and alias.asset_id == asset_a
    assert fuzzy.status in {"mapping_suggested", "mapped"} and fuzzy.score >= domain.POSSIBLE_SUGGESTION_THRESHOLD
    assert close.status in {"mapping_pending", "mapping_conflict"}
    assert discordant.status == "mapping_conflict"
    assert {candidate.asset_id for candidate in discordant.candidates} == {asset_a, asset_b}

    portfolio_id = repo.create_portfolio(conn, name="Mappings")
    member_id = repo.add_member(conn, portfolio_id=portfolio_id, asset_id=asset_a)
    repo.unmap_member(conn, member_id=member_id, portfolio_id=portfolio_id)
    assert conn.execute("SELECT asset_id FROM portfolio_assets WHERE id = ?", (member_id,)).fetchone()[0] is None


def test_portfolio_crud_members_order_copy_move_and_delete_rules(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    asset_a = add_asset(conn, "Central A")
    asset_b = add_asset(conn, "Central B")
    source = repo.create_portfolio(conn, name="Origem")
    target = repo.create_portfolio(conn, name="Destino")
    member_a = repo.add_member(conn, portfolio_id=source, asset_id=asset_a, sub_account="001")
    member_b = repo.add_member(conn, portfolio_id=source, asset_id=asset_b, sub_account="002")

    repo.update_portfolio(conn, portfolio_id=source, name="Origem Editada", description="D", notes="N")
    duplicate = repo.duplicate_portfolio(conn, portfolio_id=source, new_name="Copia")
    assert len(repo.list_portfolio_members(conn, duplicate)) == 2
    repo.reorder_members(conn, portfolio_id=source, ordered_ids=[member_b, member_a])
    assert [row["id"] for row in repo.list_portfolio_members(conn, source)] == [member_b, member_a]
    repo.copy_members(conn, source_portfolio_id=source, target_portfolio_id=target, member_ids=[member_a], move=False)
    assert len(repo.list_portfolio_members(conn, target)) == 1
    repo.copy_members(conn, source_portfolio_id=source, target_portfolio_id=target, member_ids=[member_b], move=True)
    assert {row["asset_id"] for row in repo.list_portfolio_members(conn, target)} == {asset_a, asset_b}
    repo.archive_portfolio(conn, source)
    repo.reactivate_portfolio(conn, source)
    repo.delete_portfolio(conn, source, confirm_name="Origem Editada")
    assert repo.get_portfolio(conn, source) is None

    blocked = repo.create_portfolio(conn, name="Com historico")
    conn.execute("INSERT INTO portfolio_report_runs (portfolio_id, report_month, created_at) VALUES (?, '2026-01', '2026-01-01')", (blocked,))
    try:
        repo.delete_portfolio(conn, blocked, confirm_name="Com historico")
    except ValueError as exc:
        assert str(exc) == "portfolio_has_report_history"
    else:
        raise AssertionError("expected delete block")


def test_available_asset_filters_and_backend_rejects_manipulated_ids(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    asset_a = add_asset(conn, "Filtro A")
    asset_b = add_asset(conn, "Filtro B")
    source = repo.create_portfolio(conn, name="Filtro Origem")
    target = repo.create_portfolio(conn, name="Filtro Destino")
    member_a = repo.add_member(conn, portfolio_id=source, asset_id=asset_a)
    repo.add_member(conn, portfolio_id=target, asset_id=asset_b)

    available = repo.list_available_assets(conn, portfolio_id=source, asset_filter="available")
    in_current = repo.list_available_assets(conn, portfolio_id=source, asset_filter="in_current")
    other = repo.list_available_assets(conn, portfolio_id=source, asset_filter="other_portfolio")

    assert asset_a not in {row["id"] for row in available}
    assert {row["id"] for row in in_current} == {asset_a}
    assert asset_b in {row["id"] for row in other}

    try:
        repo.remove_members(conn, portfolio_id=target, member_ids=[member_a])
    except ValueError as exc:
        assert str(exc) == "member_not_found"
    else:
        raise AssertionError("expected cross-portfolio member rejection")

    try:
        repo.reorder_members(conn, portfolio_id=source, ordered_ids=[member_a, member_a])
    except ValueError as exc:
        assert str(exc) == "duplicate_ids"
    else:
        raise AssertionError("expected duplicate order rejection")


def test_add_member_reactivates_inactive_member_and_prevents_active_duplicates(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    asset_id = add_asset(conn, "Reativar A")
    portfolio_id = repo.create_portfolio(conn, name="Reativar")
    member_id = repo.add_member(conn, portfolio_id=portfolio_id, asset_id=asset_id, external_name="Nome antigo")
    repo.update_member(
        conn,
        member_id=member_id,
        portfolio_id=portfolio_id,
        asset_id=asset_id,
        external_name="Nome antigo",
        nif="501111111",
        sub_account="001",
        notes="inativo",
        active=False,
    )

    reactivated_id = repo.add_member(
        conn,
        portfolio_id=portfolio_id,
        asset_id=asset_id,
        external_name="Nome novo",
        nif="502222222",
        sub_account="002",
        notes="reativado",
    )

    assert reactivated_id == member_id
    row = conn.execute("SELECT * FROM portfolio_assets WHERE id = ?", (member_id,)).fetchone()
    assert row["active"] == 1
    assert row["external_name"] == "Nome novo"
    assert row["sub_account"] == "002"
    assert conn.execute("SELECT COUNT(*) FROM portfolio_assets WHERE portfolio_id = ? AND asset_id = ?", (portfolio_id, asset_id)).fetchone()[0] == 1

    try:
        repo.add_member(conn, portfolio_id=portfolio_id, asset_id=asset_id)
    except ValueError as exc:
        assert str(exc) == "member_already_exists"
    else:
        raise AssertionError("expected active duplicate rejection")


def test_update_pending_member_merges_into_existing_asset_member(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    asset_id = add_asset(conn, "A Colmeia do Minho", "500001022")
    portfolio_id = repo.create_portfolio(conn, name="Merge Colmeia")
    existing_id = repo.add_member(conn, portfolio_id=portfolio_id, asset_id=asset_id, external_name="A Colmeia do Minho")
    pending_id = repo.add_member(
        conn,
        portfolio_id=portfolio_id,
        asset_id=None,
        external_name="A COLMEIA DO MINHO SA",
        nif="500001022",
        sub_account="001",
    )

    repo.update_member(
        conn,
        member_id=pending_id,
        portfolio_id=portfolio_id,
        asset_id=asset_id,
        external_name="A COLMEIA DO MINHO SA",
        nif="500001022",
        sub_account="001",
        notes="",
        active=True,
    )

    members = repo.list_portfolio_members(conn, portfolio_id)
    assert len(members) == 1
    assert members[0]["id"] == existing_id
    assert members[0]["asset_id"] == asset_id
    assert members[0]["external_name"] == "A COLMEIA DO MINHO SA"
    assert members[0]["nif"] == "500001022"
    assert members[0]["sub_account"] == "001"
    assert members[0]["mapping_status"] == "manual"
    assert conn.execute("SELECT 1 FROM portfolio_assets WHERE id = ?", (pending_id,)).fetchone() is None


def test_sync_portfolio_asset_members_replaces_asset_selection_and_keeps_pending_rows(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    asset_a = add_asset(conn, "Sync A")
    asset_b = add_asset(conn, "Sync B")
    asset_c = add_asset(conn, "Sync C")
    portfolio_id = repo.create_portfolio(conn, name="Sync Portfolio")
    repo.add_member(conn, portfolio_id=portfolio_id, asset_id=asset_a)
    repo.add_member(conn, portfolio_id=portfolio_id, asset_id=asset_b)
    pending = repo.add_member(conn, portfolio_id=portfolio_id, asset_id=None, external_name="Pendente")

    result = repo.sync_portfolio_asset_members(
        conn,
        portfolio_id=portfolio_id,
        asset_ids=[asset_b, asset_c, asset_c],
        asset_names={asset_b: "Nome B no Portfolio", asset_c: "Nome C no Portfolio"},
    )

    members = repo.list_portfolio_members(conn, portfolio_id)
    assert result == {"selected": 2, "added": 1, "removed": 1}
    assert {row["asset_id"] for row in members if row["asset_id"] is not None} == {asset_b, asset_c}
    assert {
        row["asset_id"]: row["external_name"]
        for row in members
        if row["asset_id"] is not None
    } == {asset_b: "Nome B no Portfolio", asset_c: "Nome C no Portfolio"}
    assert any(row["id"] == pending and row["asset_id"] is None for row in members)


def test_import_preview_apply_and_export_roundtrip(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    asset_id = add_asset(conn, "Central Importada", "509999999")
    csv_data = b"portfolio,sub_account,external_name,nif,asset_id,alias,notes,active\nImportado,001,Central Importada,509999999,%d,Alias Importado,Nota,1\n" % asset_id
    import_id = repo.create_import_preview(conn, portfolio_id=None, original_filename="config.csv", data=csv_data)
    run = repo.get_import_run(conn, import_id)
    assert run["rows_total"] == 1
    repo.apply_import_run(conn, import_id)
    portfolio = conn.execute("SELECT id FROM portfolio_groups WHERE name = 'Importado'").fetchone()
    assert portfolio is not None
    assert conn.execute("SELECT COUNT(*) FROM portfolio_assets WHERE portfolio_id = ?", (portfolio["id"],)).fetchone()[0] == 1
    assert conn.execute("SELECT source FROM asset_aliases WHERE alias_name = 'Alias Importado'").fetchone()[0] == "portfolio_import"

    workbook = repo.export_configuration_workbook(conn)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    loaded = load_workbook(output, read_only=True)
    assert {"Portfolios", "Membros", "Aliases", "Pendentes", "Conflitos"}.issubset(set(loaded.sheetnames))
    loaded.close()


def test_import_preview_validates_and_applies_selected_rows(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    asset_id = add_asset(conn, "Central Selectiva", "509999991")
    other_asset = add_asset(conn, "Outra Central", "508888888")
    portfolio_id = repo.create_portfolio(conn, name="Selecionado")
    repo.add_member(conn, portfolio_id=portfolio_id, asset_id=other_asset, sub_account="001")
    csv_data = (
        "portfolio,sub_account,external_name,nif,asset_id,alias,active\n"
        f"Selecionado,002,Central Selectiva,509999991,{asset_id},Alias Selectivo,1\n"
        "Selecionado,001,Subconta Duplicada,,999999,,1\n"
    ).encode()
    import_id = repo.create_import_preview(conn, portfolio_id=portfolio_id, original_filename="preview.csv", data=csv_data)
    preview = repo.import_preview_from_json(repo.get_import_run(conn, import_id)["preview_json"])

    assert preview.rows_total == 2
    assert preview.rows[1].errors == ("asset_not_found",)
    repo.apply_import_run(conn, import_id, selected_rows=[preview.rows[0].row_number])
    assert conn.execute("SELECT COUNT(*) FROM portfolio_assets WHERE portfolio_id = ?", (portfolio_id,)).fetchone()[0] == 2
    assert conn.execute("SELECT source FROM asset_aliases WHERE alias_name = 'Alias Selectivo'").fetchone()[0] == "portfolio_import"


def test_xlsx_multisheet_export_import_roundtrip(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    asset_id = add_asset(conn, "Central Roundtrip", "507777777")
    portfolio_id = repo.create_portfolio(conn, name="Roundtrip")
    repo.add_member(conn, portfolio_id=portfolio_id, asset_id=asset_id, external_name="Central Roundtrip", sub_account="001")
    repo.upsert_alias(conn, asset_id=asset_id, alias_name="Round Alias", source="manual")

    workbook = repo.export_configuration_workbook(conn)
    output = io.BytesIO()
    workbook.save(output)
    import_id = repo.create_import_preview(conn, portfolio_id=None, original_filename="roundtrip.xlsx", data=output.getvalue())
    preview = repo.import_preview_from_json(repo.get_import_run(conn, import_id)["preview_json"])

    assert any(row.action == "portfolio" for row in preview.rows)
    assert any(row.action == "new_member" for row in preview.rows)
    assert any(row.action == "alias" for row in preview.rows)


def test_portfolio_manager_routes_and_existing_portfolios_page(tmp_path: Path) -> None:
    db_path = tmp_path / "routes.db"
    ensure_database(str(db_path))
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    asset_a = add_asset(conn, "Central Route A", "501111111")
    asset_b = add_asset(conn, "Central Route B", "502222222")
    conn.close()
    flask_app, previous_db, previous_testing, client = csrf_client(db_path)
    try:
        create = client.post("/portfolio-manager/create", data={"csrf_token": "token", "name": "Rotas", "description": "D", "notes": "N"})
        assert create.status_code in {302, 303}
        with sqlite3.connect(db_path) as check:
            portfolio_id = check.execute("SELECT id FROM portfolio_groups WHERE name = 'Rotas'").fetchone()[0]

        assert client.get(f"/portfolio-manager?portfolio_id={portfolio_id}").status_code == 200
        assert client.post("/portfolio-manager/update", data={"csrf_token": "token", "portfolio_id": portfolio_id, "name": "Rotas Edit", "description": "D2", "notes": "N2"}).status_code in {302, 303}
        assert client.post("/portfolio-manager/members/add", data={"csrf_token": "token", "portfolio_id": portfolio_id, "asset_id": asset_a, "external_name": "Central Route A"}).status_code in {302, 303}
        assert client.post("/assets/%d/aliases/add" % asset_a, data={"csrf_token": "token", "alias_name": "Alias Route", "next": f"/portfolio-manager?portfolio_id={portfolio_id}"}).status_code in {302, 303}
        assert client.post("/portfolio-manager/mappings/suggest", data={"csrf_token": "token", "portfolio_id": portfolio_id}).status_code in {302, 303}
        with sqlite3.connect(db_path) as check:
            member_id = check.execute("SELECT id FROM portfolio_assets WHERE portfolio_id = ?", (portfolio_id,)).fetchone()[0]

        assert client.post("/portfolio-manager/members/reorder", data={"csrf_token": "token", "portfolio_id": portfolio_id, "ordered_ids": str(member_id)}).status_code in {302, 303}
        assert client.post("/portfolio-manager/duplicate", data={"csrf_token": "token", "portfolio_id": portfolio_id, "new_name": "Rotas Copia"}).status_code in {302, 303}
        with sqlite3.connect(db_path) as check:
            target_id = check.execute("SELECT id FROM portfolio_groups WHERE name = 'Rotas Copia'").fetchone()[0]

        assert client.post("/portfolio-manager/members/copy", data={"csrf_token": "token", "portfolio_id": portfolio_id, "target_portfolio_id": target_id, "member_ids": str(member_id)}).status_code in {302, 303}
        assert client.post("/portfolio-manager/members/move", data={"csrf_token": "token", "portfolio_id": portfolio_id, "target_portfolio_id": target_id, "member_ids": str(member_id)}).status_code in {302, 303}
        with sqlite3.connect(db_path) as check:
            moved_member_id = check.execute("SELECT id FROM portfolio_assets WHERE portfolio_id = ? AND asset_id = ?", (target_id, asset_a)).fetchone()[0]
        assert client.post("/portfolio-manager/mappings/confirm", data={"csrf_token": "token", "portfolio_id": target_id, "member_id": moved_member_id, "asset_id": asset_b, "create_alias": "on"}).status_code in {302, 303}
        assert client.post("/portfolio-manager/archive", data={"csrf_token": "token", "portfolio_id": target_id}).status_code in {302, 303}

        csv_data = b"portfolio,external_name,asset_id\nRotas Import,Central Route A,%d\n" % asset_a
        import_response = client.post(
            "/portfolio-manager/import",
            data={"csrf_token": "token", "portfolio_id": "", "file": (io.BytesIO(csv_data), "config.csv")},
            content_type="multipart/form-data",
        )
        assert import_response.status_code in {302, 303}
        with sqlite3.connect(db_path) as check:
            import_id = check.execute("SELECT id FROM portfolio_import_runs ORDER BY id DESC LIMIT 1").fetchone()[0]
        preview_html = client.get(f"/portfolio-manager?portfolio_id={portfolio_id}&import_id={import_id}")
        assert b"Aplicar linhas selecionadas" in preview_html.data
        assert b"Rotas Import" in preview_html.data
        assert client.post(f"/portfolio-manager/import/{import_id}/apply", data={"csrf_token": "token"}).status_code in {302, 303}
        export = client.get("/portfolio-manager/export")
        assert export.status_code == 200
        assert client.get(f"/portfolios?portfolio_id={target_id}").status_code == 200
    finally:
        flask_app.config["DATABASE"] = previous_db
        flask_app.config["TESTING"] = previous_testing


def test_portfolio_manager_add_modal_buttons_friendly_errors_and_state(tmp_path: Path) -> None:
    db_path = tmp_path / "portfolio-ui.db"
    ensure_database(str(db_path))
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    asset_a = add_asset(conn, "UI A", "501111111")
    asset_b = add_asset(conn, "UI B", "502222222")
    asset_c = add_asset(conn, "UI C", "503333333")
    portfolio_id = repo.create_portfolio(conn, name="UI Portfolio")
    repo.add_member(conn, portfolio_id=portfolio_id, asset_id=asset_a)
    inactive_id = repo.add_member(conn, portfolio_id=portfolio_id, asset_id=asset_b, external_name="Inativo")
    repo.update_member(
        conn,
        member_id=inactive_id,
        portfolio_id=portfolio_id,
        asset_id=asset_b,
        external_name="Inativo",
        nif="",
        sub_account="",
        notes="",
        active=False,
    )
    conn.commit()
    conn.close()
    flask_app, previous_db, previous_testing, client = csrf_client(db_path)
    try:
        page = client.get(f"/portfolio-manager?portfolio_id={portfolio_id}&tab=installations&search=UI&asset_filter=available")
        assert page.status_code == 200
        assert b"Adicionar instalacoes" in page.data
        assert b"Editar" in page.data
        assert b"Retirar associacao" in page.data
        assert b"Remover" in page.data

        add_many = client.post(
            "/portfolio-manager/members/add",
            data={
                "csrf_token": "token",
                "portfolio_id": portfolio_id,
                "tab": "installations",
                "search": "UI",
                "asset_filter": "available",
                "asset_ids": [str(asset_b), str(asset_c)],
            },
        )
        assert add_many.status_code in {302, 303}
        assert "tab=installations" in add_many.headers["Location"]
        assert "search=UI" in add_many.headers["Location"]
        with sqlite3.connect(db_path) as check:
            rows = {
                row[0]: row[1]
                for row in check.execute(
                    "SELECT asset_id, active FROM portfolio_assets WHERE portfolio_id = ? AND asset_id IS NOT NULL",
                    (portfolio_id,),
                )
            }
        assert rows == {asset_a: 1, asset_b: 1, asset_c: 1}

        duplicate = client.post(
            "/portfolio-manager/members/add",
            data={
                "csrf_token": "token",
                "portfolio_id": portfolio_id,
                "tab": "installations",
                "asset_id": asset_a,
            },
            follow_redirects=True,
        )
        assert "Esta instalacao ja pertence a este portfolio.".encode() in duplicate.data

        with sqlite3.connect(db_path) as check:
            member_id = check.execute("SELECT id FROM portfolio_assets WHERE portfolio_id = ? AND asset_id = ?", (portfolio_id, asset_a)).fetchone()[0]
        update = client.post(
            "/portfolio-manager/members/update",
            data={
                "csrf_token": "token",
                "portfolio_id": portfolio_id,
                "member_id": member_id,
                "asset_id": asset_c,
                "external_name": "UI A externo",
                "tab": "installations",
            },
            follow_redirects=True,
        )
        assert "Esta instalacao ja pertence a este portfolio.".encode() in update.data
    finally:
        flask_app.config["DATABASE"] = previous_db
        flask_app.config["TESTING"] = previous_testing
