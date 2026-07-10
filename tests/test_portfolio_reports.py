from __future__ import annotations

import sqlite3
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook

from app import ensure_database
from monitoring_board.portfolio_reports import (
    aggregate_portfolio_total,
    auto_map_portfolio_assets,
    build_portfolio_report_rows,
    calculate_degradation_factor,
    calculate_tariff_value,
    classify_tariff_period,
    import_financial_model_file,
    map_external_portfolio_entity,
    parse_financial_model_file,
    parse_helioscope_monthly_expected,
)


def connect(tmp_path: Path) -> sqlite3.Connection:
    db_path = tmp_path / "test.db"
    ensure_database(str(db_path))
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def add_asset(conn: sqlite3.Connection, name: str = "Solar One", nif: str = "123456789") -> int:
    cursor = conn.execute(
        "INSERT INTO assets (project_name, nif, kwp, mounting_date, start_contract) VALUES (?, ?, '10', '2024-01-01', '2024-01-01')",
        (name, nif),
    )
    asset_id = int(cursor.lastrowid)
    conn.execute(
        "INSERT INTO asset_aliases (asset_id, alias_name, normalized_alias, source) VALUES (?, ?, ?, 'test')",
        (asset_id, name, name.lower()),
    )
    return asset_id


def test_degradation_factor_uses_helioscope_phase_2_rule() -> None:
    factor = calculate_degradation_factor(date(2024, 1, 1), date(2025, 7, 1))
    assert round(factor, 5) == 0.97225


def test_helioscope_monthly_parser_reads_month_columns(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Metric", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"])
    sheet.append(["Expected kWh", 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12])
    path = tmp_path / "helioscope.xlsx"
    workbook.save(path)

    parsed = parse_helioscope_monthly_expected(path)

    assert parsed[1] == 1
    assert parsed[12] == 12


def build_financial_workbook(path: Path) -> None:
    workbook = Workbook()
    projeto = workbook.active
    projeto.title = "Projeto"
    projeto["C5"] = "Residencia Teste"
    projeto["H8"] = 6.5
    for month in range(1, 13):
        row = 5 + month
        projeto.cell(row=row, column=10, value=month)
        projeto.cell(row=row, column=11, value=month * 100)
    projeto["C41"] = datetime(2026, 1, 1)
    projeto["C42"] = datetime(2026, 1, 31)
    projeto["C44"] = "Simples"
    projeto["C45"] = "Diario"
    projeto["F41"] = 0.08
    projeto["G41"] = 0.0607
    projeto["H41"] = 0.1407
    projeto["F46"] = 0.045
    projeto["L33"] = 0.0555
    projeto["M33"] = 809
    projeto["P28"] = 20520.93
    helio = workbook.create_sheet("Helio&Cons")
    helio["L1"] = "New Production\n(8760)"
    helio["N1"] = "Dates of Year"
    helio["L2"] = 250
    helio["N2"] = datetime(2026, 1, 1, 0, 0)
    helio["L3"] = 500
    helio["N3"] = datetime(2026, 1, 1, 0, 15)
    workbook.save(path)


def test_financial_model_parser_reads_monthly_interval_and_tariff(tmp_path: Path) -> None:
    path = tmp_path / "financial.xlsx"
    build_financial_workbook(path)

    parsed = parse_financial_model_file(path)

    assert parsed.project_name == "Residencia Teste"
    assert parsed.monthly_expected[1] == 100
    assert parsed.monthly_expected[12] == 1200
    assert len(parsed.interval_expected) == 2
    assert parsed.interval_expected[0].expected_kwh == 0.25
    assert parsed.tariff is not None
    assert parsed.tariff.tariff_type == "simple"
    assert parsed.tariff.simple_price_eur_kwh == 0.1407


def test_financial_model_parser_reads_upac_prod_month_format(tmp_path: Path) -> None:
    path = tmp_path / "financial_upac.xlsm"
    workbook = Workbook()
    upac = workbook.active
    upac.title = "UPAC"
    upac["A4"] = "Usinage"
    upac["D4"] = 138.6
    upac["D6"] = 84546
    upac["H12"] = "Semanal"
    upac["H13"] = 0.064138
    upac["H14"] = 0.06806
    upac["H15"] = 0.099243
    upac["H16"] = 0.116027
    upac["H19"] = 0.045
    upac["K5"] = 199413.42
    upac["K7"] = 186000.02
    upac["K8"] = 13413.4
    upac["K14"] = 0.1142
    upac["K15"] = 0.1099
    upac["N4"] = 4500.33
    prod_month = workbook.create_sheet("Prod month")
    prod_month.append(["Row Labels", "Consumption", "PV"])
    for month in range(1, 13):
        prod_month.append([month, month * 1000, month * 100])
    workbook.save(path)

    parsed = parse_financial_model_file(path)

    assert parsed.project_name == "Usinage"
    assert parsed.installed_power_kwp == 138.6
    assert parsed.monthly_expected[1] == 100
    assert parsed.monthly_expected[12] == 1200
    assert parsed.tariff is not None
    assert parsed.tariff.tariff_type == "tetra-hourly"
    assert parsed.tariff.super_vazio_price_eur_kwh == 0.064138
    assert parsed.tariff.solcor_price_eur_kwh == 0.1099
    assert parsed.financial_outputs["annual_pv_production_kwh"] == 199413.42


class UploadStub:
    def __init__(self, source: Path, filename: str = "financial.xlsx") -> None:
        self.source = source
        self.filename = filename

    def save(self, target: Path) -> None:
        target.write_bytes(self.source.read_bytes())


def test_financial_model_import_replaces_helioscope_and_applies_tariff(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    asset_id = add_asset(conn)
    portfolio_id = conn.execute("SELECT id FROM portfolio_groups WHERE name = 'Solcorelios I'").fetchone()["id"]
    path = tmp_path / "financial.xlsx"
    build_financial_workbook(path)

    result = import_financial_model_file(
        conn,
        upload_dir=tmp_path / "uploads",
        file_storage=UploadStub(path),
        asset_id=asset_id,
        portfolio_id=portfolio_id,
        base_year=2026,
    )
    conn.commit()

    assert result["months"] == 12
    assert result["intervals_15m"] == 2
    assert conn.execute("SELECT COUNT(*) FROM helioscope_expected_production WHERE asset_id = ?", (asset_id,)).fetchone()[0] == 12
    assert conn.execute("SELECT COUNT(*) FROM helioscope_expected_interval_production WHERE asset_id = ?", (asset_id,)).fetchone()[0] == 2
    tariff = conn.execute("SELECT * FROM asset_tariffs WHERE asset_id = ?", (asset_id,)).fetchone()
    assert tariff["tariff_type"] == "simple"
    assert tariff["simple_price_eur_kwh"] == 0.1407
    billing = conn.execute("SELECT * FROM asset_billing_configs WHERE asset_id = ?", (asset_id,)).fetchone()
    assert billing["default_electricity_price"] == "0.1407"


def test_tariff_period_classification_handles_overnight_rule() -> None:
    rules = [{"weekday_type": "all", "start_time": "22:00", "end_time": "07:00", "period_name": "vazio"}]

    assert classify_tariff_period(datetime(2026, 1, 5, 23, 0), rules) == "vazio"
    assert classify_tariff_period(datetime(2026, 1, 6, 6, 0), rules) == "vazio"
    assert classify_tariff_period(datetime(2026, 1, 6, 12, 0), rules) is None


def test_simple_tariff_calculation() -> None:
    result = calculate_tariff_value(
        {"tariff_type": "simple", "simple_price_eur_kwh": 0.2},
        monthly_kwh=100,
        hourly_records=[],
        rules=[],
    )

    assert result["estimated_value_eur"] == 20
    assert result["warnings"] == []


def test_tri_hourly_calculation_uses_hourly_records() -> None:
    tariff = {
        "tariff_type": "tri-hourly",
        "ponta_price_eur_kwh": 0.3,
        "cheia_price_eur_kwh": 0.2,
        "vazio_price_eur_kwh": 0.1,
        "super_vazio_price_eur_kwh": None,
    }
    rules = [
        {"weekday_type": "all", "start_time": "08:00", "end_time": "10:00", "period_name": "ponta"},
        {"weekday_type": "all", "start_time": "10:00", "end_time": "20:00", "period_name": "cheia"},
        {"weekday_type": "all", "start_time": "20:00", "end_time": "08:00", "period_name": "vazio"},
    ]
    hourly = [
        {"period_start": "2026-01-05T08:00:00", "production_kwh": 10},
        {"period_start": "2026-01-05T11:00:00", "production_kwh": 10},
        {"period_start": "2026-01-05T21:00:00", "production_kwh": 10},
    ]

    result = calculate_tariff_value(tariff, monthly_kwh=999, hourly_records=hourly, rules=rules)

    assert result["period_kwh"]["ponta"] == 10
    assert result["period_kwh"]["cheia"] == 10
    assert result["period_kwh"]["vazio"] == 10
    assert result["estimated_value_eur"] is None
    assert "missing_hourly_self_use" in result["warnings"]


def test_report_flags_missing_data(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    asset_id = add_asset(conn)
    portfolio_id = conn.execute("SELECT id FROM portfolio_groups WHERE name = 'Solcorelios I'").fetchone()["id"]
    conn.execute(
        "INSERT INTO portfolio_assets (portfolio_id, asset_id, active, mapping_status, mapping_confidence) VALUES (?, ?, 1, 'manual', 1)",
        (portfolio_id, asset_id),
    )
    conn.commit()

    row = build_portfolio_report_rows(conn, portfolio_id, "2026-01")[0]

    assert row["data_status"] == "missing_data"
    assert "missing_monthly_production" in row["warnings"]
    assert "missing_helioscope_expected" in row["warnings"]
    assert "missing_tariff" in row["warnings"]


def test_report_uses_daily_production_when_monthly_row_is_missing(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    asset_id = add_asset(conn)
    portfolio_id = conn.execute("SELECT id FROM portfolio_groups WHERE name = 'Solcorelios I'").fetchone()["id"]
    conn.execute(
        "INSERT INTO portfolio_assets (portfolio_id, asset_id, active, mapping_status, mapping_confidence) VALUES (?, ?, 1, 'manual', 1)",
        (portfolio_id, asset_id),
    )
    conn.execute(
        """
        INSERT INTO production_records (
            asset_id, provider, external_id, period_type, period_date,
            production_kwh, data_quality, created_at, updated_at
        ) VALUES (?, 'FusionSolar', 'D1', 'day', '2026-01-01', 10, 'ok', '2026-01-02T00:00:00', '2026-01-02T00:00:00')
        """,
        (asset_id,),
    )
    conn.execute(
        """
        INSERT INTO production_records (
            asset_id, provider, external_id, period_type, period_date,
            production_kwh, data_quality, created_at, updated_at
        ) VALUES (?, 'FusionSolar', 'D2', 'day', '2026-01-02', 15, 'ok', '2026-01-03T00:00:00', '2026-01-03T00:00:00')
        """,
        (asset_id,),
    )
    conn.commit()

    row = next(item for item in build_portfolio_report_rows(conn, portfolio_id, "2026-01") if item["asset_id"] == asset_id)

    assert row["actual_production_kwh"] == 25
    assert "missing_monthly_production" not in row["warnings"]


def test_portfolio_total_aggregates_and_weights_availability() -> None:
    rows = [
        {"actual_production_kwh": 100, "adjusted_expected_kwh": 100, "installed_power_kwp": 100, "availability_pct": 90, "estimated_value_eur": 10, "production_ponta_kwh": 1, "production_cheia_kwh": 2, "production_vazio_kwh": 3, "production_super_vazio_kwh": 0, "helioscope_expected_kwh": 110, "warnings": []},
        {"actual_production_kwh": 50, "adjusted_expected_kwh": 300, "installed_power_kwp": 900, "availability_pct": 80, "estimated_value_eur": 5, "production_ponta_kwh": 4, "production_cheia_kwh": 5, "production_vazio_kwh": 6, "production_super_vazio_kwh": 0, "helioscope_expected_kwh": 330, "warnings": ["missing_hourly_production"]},
        {"actual_production_kwh": 25, "adjusted_expected_kwh": 100, "installed_power_kwp": None, "availability_pct": 10, "estimated_value_eur": 2, "production_ponta_kwh": 0, "production_cheia_kwh": 0, "production_vazio_kwh": 0, "production_super_vazio_kwh": 0, "helioscope_expected_kwh": 100, "warnings": []},
    ]

    total = aggregate_portfolio_total(rows)

    assert total["actual_production_kwh"] == 175
    assert total["installed_power_kwp"] == 1000
    assert total["availability_pct"] == 81.0
    assert total["deviation_pct"] == -65.0
    assert total["warnings"] == ["missing_hourly_production", "missing_installed_power"]


def test_mapping_by_nif_then_name(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    asset_id = add_asset(conn, name="Solar Alias", nif="501 222 333")

    assert map_external_portfolio_entity(conn, nif="501222333", external_name="Other")["asset_id"] == asset_id
    assert map_external_portfolio_entity(conn, nif="", external_name="Solar Alias")["asset_id"] == asset_id


def test_seed_external_portfolio_rows_includes_missing_subaccounts(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    solcorelios_ii = conn.execute("SELECT id FROM portfolio_groups WHERE name = 'Solcorelios II'").fetchone()["id"]

    rows = conn.execute(
        "SELECT sub_account, nif, mapping_status FROM portfolio_assets WHERE portfolio_id = ? AND sub_account <= '005' ORDER BY sub_account",
        (solcorelios_ii,),
    ).fetchall()

    assert [row["sub_account"] for row in rows] == ["001", "002", "003", "004", "005"]
    assert all(row["nif"] == "" for row in rows)
    assert all(row["mapping_status"] == "missing_source" for row in rows)


def test_repeated_nif_across_portfolios_is_not_false_conflict(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    asset_id = add_asset(conn, name="JETESETECAR EQUIPAMENTOS AUTO LDA", nif="503194387")

    result = auto_map_portfolio_assets(conn)
    rows = conn.execute(
        "SELECT asset_id, mapping_status FROM portfolio_assets WHERE nif = '503194387' ORDER BY portfolio_id"
    ).fetchall()

    assert result["conflicts"] == 0
    assert {row["mapping_status"] for row in rows} == {"mapped"}
    assert {row["asset_id"] for row in rows} == {asset_id}
