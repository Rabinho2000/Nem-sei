from __future__ import annotations

import sqlite3
from pathlib import Path

import pytest
from openpyxl import Workbook

from app import app, ensure_database
from monitoring_board.db import get_db
from monitoring_board.financial_model_repository import list_model_monthly
from monitoring_board.portfolio_reports import build_portfolio_report_rows
from monitoring_board.reporting.financial_models import FinancialModelParseError, parse_financial_model_workbook
from monitoring_board.services.financial_models import (
    FinancialModelError,
    confirm_financial_model_import,
    create_financial_model_preview,
)


class UploadStub:
    def __init__(self, path: Path, filename: str | None = None) -> None:
        self.path = path
        self.filename = filename or path.name

    def save(self, target: Path) -> None:
        target.write_bytes(self.path.read_bytes())


def connect(tmp_path: Path) -> sqlite3.Connection:
    db_path = tmp_path / "financial_models.db"
    ensure_database(str(db_path))
    conn = get_db(str(db_path))
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def add_asset(conn: sqlite3.Connection, *, name: str = "Usinage", nif: str = "505435748", kwp: str = "138.6") -> int:
    cursor = conn.execute(
        "INSERT INTO assets (project_name, nif, kwp, mounting_date, start_contract) VALUES (?, ?, ?, '2024-01-01', '2024-01-01')",
        (name, nif, kwp),
    )
    return int(cursor.lastrowid)


def financial_workbook(path: Path, *, sheet_name: str = "Prod month", unit_label: str = "PV Production (kWh)", missing_month: bool = False) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    if missing_month:
        months = months[:-1]
    sheet.append(["Metric", *months])
    sheet.append([unit_label, *[month * 100 for month in range(1, len(months) + 1)]])
    sheet.append(["Consumption (kWh)", *[month * 120 for month in range(1, len(months) + 1)]])
    sheet.append(["Self consumption (kWh)", *[month * 80 for month in range(1, len(months) + 1)]])
    meta = workbook.create_sheet("UPAC")
    meta["A4"] = "Usinage"
    meta["D4"] = 138.6
    workbook.save(path)


def usinage_style_financial_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Prod month"
    sheet.append([])
    sheet.append([])
    sheet.append(["Row Labels", "Soma de Consumption (kWh)", "Soma de PV (kWh)", "Soma de SC (kWh)", "% AC"])
    for month in range(1, 13):
        production = month * 1000.0
        self_use = month * 900.0
        sheet.append([month, month * 1200.0, production, self_use, self_use / production])
    sheet.append(["Grand Total", 93600.0, 78000.0, 70200.0, 0.9])

    upac = workbook.create_sheet("UPAC")
    upac["A4"] = "Usinage"
    upac["D4"] = 138.6
    upac["D6"] = 84546
    upac["H12"] = "Semanal"
    upac["G13"] = "SV"
    upac["H13"] = 0.064138
    upac["G14"] = "Vazio"
    upac["H14"] = 0.06806
    upac["G15"] = "Cheia"
    upac["H15"] = 0.099243
    upac["G16"] = "Ponta"
    upac["H16"] = 0.116027
    upac["G19"] = "OMIE"
    upac["H19"] = 0.045
    upac["K4"] = 93600
    upac["K5"] = 78000
    upac["K7"] = 70200
    upac["K8"] = 7800
    upac["K14"] = 0.114224
    upac["K15"] = 0.1099
    upac["N4"] = 4500.33
    upac["Q11"] = 37.46
    upac["C24"] = "Year"
    upac["D24"] = 2019

    proposal = workbook.create_sheet("Data PV Proposal")
    proposal.append([])
    proposal.append([None, None, "Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez", None, "Annual"])
    proposal.append([None, "Autoconsumo (kWh)", *[month * 900.0 for month in range(1, 13)], None, 70200.0])
    proposal.append([None, "Excedente (kWh)", *[month * 100.0 for month in range(1, 13)], None, 7800.0])
    proposal.append([None, "Total Benefit (EUR)", *[month * 10.0 for month in range(1, 13)], None, 780.0])

    invoice = workbook.create_sheet("Detalhes da fatura")
    invoice.append([])
    invoice.append([None, "Row Labels", "Sum of SC (kWh)", "Sum of Savings Energy"])
    invoice.append([None, "Cheia", 1000, 100, 0.5, "Ponta", -500, 0.1, 0.05, None, -50, -25])
    invoice.append([None, "Ponta", 500, 60, 0.25, "Cheia", -1000, 0.09, 0.04, None, -90, -40])
    invoice.append([None, "Grand Total", 1500, 160])
    invoice["F10"] = "Excedente"
    invoice["G10"] = 7800
    invoice["H10"] = 0.045
    invoice["I10"] = 351
    invoice["F12"] = "TOT"
    invoice["G12"] = 511
    invoice["F13"] = "EUR/kWh"
    invoice["G13"] = 0.1095
    workbook.save(path)


def financial_automatic_as_sold_workbook(path: Path, *, battery_charge: bool = False) -> None:
    workbook = Workbook()
    project = workbook.active
    project.title = "Projeto"
    project["C5"] = "Projeto Exemplo"
    project["H8"] = 17.5
    project["H14"] = 1234.5
    project["H22"] = 0.02
    project["H23"] = 0.005
    project["D26"] = 500
    project["E26"] = 8750
    project["D28"] = 700
    project["E28"] = 12250
    project["J5"] = "Month"
    project["K5"] = "Monthly Production [kWh]"
    project["L5"] = "AC [kWh]"
    project["M5"] = "% AC"
    for month in range(1, 13):
        row = month + 5
        project.cell(row, 10, month)
        project.cell(row, 11, month * 100)
        project.cell(row, 12, month * 60)
        project.cell(row, 13, 0.6)
    project["K18"] = 7800
    project["L18"] = 4680
    project["P5"] = 11700
    project["P6"] = 7800
    project["P7"] = 4680
    project["P8"] = 3120
    project["P9"] = 0.6
    project["P10"] = 0.4
    project["L32"] = 0.15
    project["G39"] = "2026/1"
    project["C44"] = "Tetra-horário"
    project["C45"] = "Semanal"
    project["F46"] = 0.05
    for row, label, energy, network in (
        (41, "SV", 0.08, 0.02),
        (42, "Vazio", 0.09, 0.03),
        (43, "Cheia", 0.10, 0.04),
        (44, "Ponta", 0.11, 0.05),
    ):
        project.cell(row, 5, label)
        project.cell(row, 6, energy)
        project.cell(row, 7, network)
        project.cell(row, 8, energy + network)

    savings = workbook.create_sheet("Savings Yr1")
    savings.append([])
    savings.append([])
    savings.append([None, "Month", "Cons. [kWh]", "Faturas €", "AC [kWh]", "Save AC €", "Exced [kWh]"])
    if battery_charge:
        savings["I3"] = "Exc. ESS [kWh]"
    for month in range(1, 13):
        savings.append([None, month, month * 150, month * 10, month * 60, month * 4, month * 40])
        if battery_charge:
            savings.cell(month + 3, 9, month * 10)
    savings["B16"] = "TOTAL"
    savings["C16"] = 11700
    savings["D16"] = 780
    savings["E16"] = 4680
    savings["F16"] = 312
    savings["G16"] = 3120
    for row, label, consumption, production, self_use, grid_import, export in (
        (41, "Ponta", 3000, 2000, 1200, 1800, 800),
        (42, "Cheia", 4000, 3000, 1800, 2200, 1200),
        (43, "Vazio", 3500, 2500, 1500, 2000, 1000),
        (44, "SV", 1200, 300, 180, 1020, 120),
    ):
        savings.cell(row, 2, label)
        savings.cell(row, 3, consumption)
        savings.cell(row, 4, production)
        savings.cell(row, 5, self_use)
        savings.cell(row, 6, grid_import)
        savings.cell(row, 7, export)
    savings["E45"] = 4680
    savings["F45"] = 7020
    savings["D52"] = 500
    for row, label, savings_value in (
        (48, "Ponta", 120),
        (49, "Cheia", 180),
        (50, "Vazio", 150),
        (51, "SV", 18),
    ):
        savings.cell(row, 2, label)
        savings.cell(row, 5, savings_value)

    noise = workbook.create_sheet("Outro resumo")
    noise.append(["Metric", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"])
    noise.append(["PV Production", *range(1, 13)])
    workbook.save(path)


def financial_automatic_shifted_workbook(path: Path) -> None:
    workbook = Workbook()
    project = workbook.active
    project.title = "Projeto"
    project["B5"] = "Nome do Projeto:"
    project["C5"] = "Projeto Sanitizado"
    project["G8"] = "Size Project"
    project["H8"] = 25
    project["J5"] = "Month"
    project["K5"] = "Monthly Production [kWh]"
    project["L5"] = "AC [kWh]"
    for month in range(1, 13):
        row = month + 5
        project.cell(row, 10, month)
        project.cell(row, 11, month * 100)
        project.cell(row, 12, month * 70)
    project.append([])
    project["E33"] = "Período"
    project["F33"] = "Energia"
    project["G33"] = "Redes"
    project["H33"] = "Total"
    for row, label in enumerate(("SV", "Vazio", "Cheia", "Ponta"), start=34):
        project.cell(row, 5, label)
        project.cell(row, 6, 0.08)
        project.cell(row, 7, 0.02)
        project.cell(row, 8, 0.10)

    savings = workbook.create_sheet("Savings Yr1")
    savings.append([])
    savings.append([])
    savings.append([None, "Month", "Cons. [kWh]", "Faturas EUR", None, "AC [kWh]", "Save AC EUR", None, "Exced [kWh]"])
    for month in range(1, 13):
        savings.append([None, month, month * 150, month * 10, None, month * 70, month * 5, None, month * 30])
    savings["B16"] = "TOTAL"
    savings["C16"] = 11700
    savings["D16"] = 780
    savings["F16"] = 5460
    savings["G16"] = 390
    savings["I16"] = 2340
    workbook.save(path)


def monthly_production_legacy_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Monthly Production"
    sheet.append([])
    sheet.append([])
    sheet.append([])
    sheet.append([None, None, "Row Labels", "Sum of Consumption (kWh)", "Sum of PV (kWh)", "Sum of SC (kWh)"])
    for month in range(1, 13):
        sheet.append([None, None, month, month * 150, month * 100, month * 70])
    upac = workbook.create_sheet("UPAC")
    upac["A4"] = "Projeto Sanitizado"
    upac["D4"] = 25
    upac["K5"] = 9999
    workbook.save(path)


def test_financial_model_schema_new_and_existing_db(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    try:
        assert conn.execute("SELECT name FROM sqlite_master WHERE name = 'financial_models'").fetchone()
        assert conn.execute("SELECT name FROM sqlite_master WHERE name = 'financial_model_monthly'").fetchone()
        assert "details_json" in {row["name"] for row in conn.execute("PRAGMA table_info(financial_models)").fetchall()}
        ensure_database(str(tmp_path / "financial_models.db"))
    finally:
        conn.close()


def test_parse_financial_model_workbook_xlsx_and_mwh_conversion(tmp_path: Path) -> None:
    path = tmp_path / "model.xlsx"
    financial_workbook(path, unit_label="PV Production (MWh)")

    parsed = parse_financial_model_workbook(path)

    assert parsed.detected_name == "Usinage"
    assert parsed.detected_kwp == 138.6
    assert len(parsed.monthly) == 12
    assert parsed.monthly[0]["expected_production_kwh"] == 100000
    assert parsed.monthly[0]["source_fields"]["expected_production_kwh"]["conversion"] == "mwh_to_kwh"
    assert parsed.monthly[0]["expected_export_kwh"] == 99920
    assert "financial_model_calculated_export" in parsed.monthly[0]["warnings"]


def test_parse_financial_model_workbook_xlsm(tmp_path: Path) -> None:
    path = tmp_path / "model.xlsm"
    financial_workbook(path)

    parsed = parse_financial_model_workbook(path)

    assert len(parsed.monthly) == 12
    assert parsed.monthly[11]["expected_production_kwh"] == 1200


def test_parse_usinage_style_financial_model_details(tmp_path: Path) -> None:
    path = tmp_path / "usinage.xlsm"
    usinage_style_financial_workbook(path)

    parsed = parse_financial_model_workbook(path)

    assert parsed.base_year == 2019
    assert parsed.monthly[0]["expected_production_kwh"] == 1000
    assert parsed.monthly[0]["expected_self_use_kwh"] == 900
    assert parsed.monthly[0]["expected_export_kwh"] == 100
    assert parsed.details["upac_summary"]
    assert any(item["key"] == "selling_price_total_eur" and item["value"] == 84546 for item in parsed.details["upac_summary"])
    assert len(parsed.details["tariff_periods"]) >= 4
    assert any(row["label"] == "Total Benefit (EUR)" for row in parsed.details["proposal_rows"])
    assert parsed.details["invoice_periods"][0]["period"] == "Cheia"


def test_parse_financial_automatic_as_sold_layout(tmp_path: Path) -> None:
    path = tmp_path / "financial_automatic_as_sold.xlsm"
    financial_automatic_as_sold_workbook(path)

    parsed = parse_financial_model_workbook(path)

    assert parsed.parser_version == "3"
    assert parsed.sheet_name == "Projeto"
    assert parsed.detected_name == "Projeto Exemplo"
    assert parsed.detected_kwp == 17.5
    assert parsed.base_year == 2026
    assert parsed.source_cells["base_year"] == "Projeto!G39"
    assert parsed.monthly[0]["expected_production_kwh"] == 100
    assert parsed.monthly[0]["expected_consumption_kwh"] == 150
    assert parsed.monthly[0]["expected_self_use_kwh"] == 60
    assert parsed.monthly[0]["expected_export_kwh"] == 40
    assert parsed.monthly[0]["expected_grid_import_kwh"] == 90
    assert parsed.monthly[0]["source_fields"]["expected_production_kwh"]["cell"] == "Projeto!K6"
    assert parsed.monthly[0]["source_fields"]["expected_consumption_kwh"]["cell"] == "Savings Yr1!C4"
    assert "financial_model_calculated_grid_import" in parsed.warnings
    assert parsed.details["format"] == "financial_automatic_as_sold"
    assert any(item["key"] == "installation_cost_total_eur" and item["value"] == 8750 for item in parsed.details["upac_summary"])
    assert any(item["key"] == "annual_grid_import_kwh" and item["value"] == 7020 for item in parsed.details["upac_summary"])
    assert parsed.details["tariff_periods"][0]["label"] == "SV"
    assert parsed.details["electricity_costs"][0]["energy_eur_kwh"] == 0.08
    assert parsed.details["invoice_periods"][0]["period"] == "Ponta"
    assert parsed.details["invoice_totals"][0]["value"] == 780
    assert next(row for row in parsed.details["proposal_rows"] if row["label"] == "Buy from grid (kWh)")["annual"] == 7020


def test_parse_financial_automatic_shifted_energy_columns(tmp_path: Path) -> None:
    path = tmp_path / "financial_automatic_shifted.xlsm"
    financial_automatic_shifted_workbook(path)

    parsed = parse_financial_model_workbook(path)

    assert parsed.monthly[0]["expected_production_kwh"] == 100
    assert parsed.monthly[0]["expected_consumption_kwh"] == 150
    assert parsed.monthly[0]["expected_self_use_kwh"] == 70
    assert parsed.monthly[0]["expected_export_kwh"] == 30
    assert parsed.monthly[0]["source_fields"]["expected_self_use_kwh"]["cell"] == "Savings Yr1!F4"
    assert parsed.monthly[0]["source_fields"]["expected_export_kwh"]["cell"] == "Savings Yr1!I4"
    assert parsed.details["tariff_periods"][0]["source_cell"] == "Projeto!H34"
    assert parsed.details["invoice_totals"][1]["source_cell"] == "Savings Yr1!G16"


def test_parse_financial_automatic_adjusts_export_stored_in_battery(tmp_path: Path) -> None:
    path = tmp_path / "financial_automatic_battery.xlsm"
    financial_automatic_as_sold_workbook(path, battery_charge=True)

    parsed = parse_financial_model_workbook(path)

    assert parsed.monthly[0]["expected_export_kwh"] == 30
    assert parsed.monthly[0]["calculated_fields"]["expected_export_kwh"] == "export_minus_battery_charge"
    assert parsed.monthly[0]["source_fields"]["expected_export_kwh"]["adjustment_cell"] == "Savings Yr1!I4"
    assert "financial_model_battery_export_adjusted" in parsed.warnings


def test_parse_legacy_monthly_production_with_offset_month_column(tmp_path: Path) -> None:
    path = tmp_path / "financial_legacy_monthly_production.xlsm"
    monthly_production_legacy_workbook(path)

    parsed = parse_financial_model_workbook(path)

    assert parsed.sheet_name == "Monthly Production"
    assert parsed.monthly[0]["expected_production_kwh"] == 100
    assert parsed.monthly[0]["expected_consumption_kwh"] == 150
    assert parsed.monthly[0]["expected_self_use_kwh"] == 70
    assert "financial_model_annual_total_mismatch_production" in parsed.warnings


def test_parse_financial_model_rejects_missing_month_and_ambiguous_sheets(tmp_path: Path) -> None:
    missing = tmp_path / "missing.xlsx"
    financial_workbook(missing, missing_month=True)
    with pytest.raises(FinancialModelParseError):
        parse_financial_model_workbook(missing)

    ambiguous = tmp_path / "ambiguous.xlsx"
    financial_workbook(ambiguous)
    workbook = Workbook()
    first = workbook.active
    first.title = "One"
    first.append(["Metric", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"])
    first.append(["PV Production", *range(1, 13)])
    second = workbook.create_sheet("Two")
    second.append(["Metric", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"])
    second.append(["PV Production", *range(1, 13)])
    workbook.save(ambiguous)
    with pytest.raises(FinancialModelParseError):
        parse_financial_model_workbook(ambiguous)


def test_create_preview_confirm_version_and_duplicate_hash(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    try:
        asset_id = add_asset(conn)
        path = tmp_path / "model.xlsx"
        financial_workbook(path)
        model_id = create_financial_model_preview(conn, upload_dir=tmp_path / "uploads", file_storage=UploadStub(path), asset_id=asset_id, base_year=2026)
        assert len(list_model_monthly(conn, model_id=model_id)) == 12
        version = confirm_financial_model_import(conn, model_id=model_id, asset_id=asset_id)
        conn.commit()
        assert version == 1
        model = conn.execute("SELECT * FROM financial_models WHERE id = ?", (model_id,)).fetchone()
        assert model["status"] == "confirmed"
        assert model["active"] == 1
        assert model["details_json"]
        with pytest.raises(FinancialModelError):
            create_financial_model_preview(conn, upload_dir=tmp_path / "uploads", file_storage=UploadStub(path), asset_id=asset_id, base_year=2026)
    finally:
        conn.close()


def test_financial_model_precedence_over_helioscope_and_year_isolated(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    try:
        asset_id = add_asset(conn)
        portfolio_id = conn.execute("SELECT id FROM portfolio_groups WHERE name = 'Solcorelios I'").fetchone()["id"]
        conn.execute("INSERT INTO portfolio_assets (portfolio_id, asset_id, active, mapping_status, mapping_confidence) VALUES (?, ?, 1, 'manual', 1)", (portfolio_id, asset_id))
        source_id = conn.execute(
            "INSERT INTO source_files (asset_id, file_type, original_filename, stored_path, uploaded_at) VALUES (?, 'helioscope', 'h.xlsx', 'h.xlsx', '2026-01-01')",
            (asset_id,),
        ).lastrowid
        conn.execute(
            "INSERT INTO helioscope_expected_production (asset_id, source_file_id, base_year, month, expected_kwh, imported_at) VALUES (?, ?, 2026, 1, 50, '2026-01-01')",
            (asset_id, source_id),
        )
        path = tmp_path / "model.xlsx"
        financial_workbook(path)
        model_id = create_financial_model_preview(conn, upload_dir=tmp_path / "uploads", file_storage=UploadStub(path), asset_id=asset_id, base_year=2026)
        confirm_financial_model_import(conn, model_id=model_id, asset_id=asset_id)
        conn.commit()

        row_2026 = next(row for row in build_portfolio_report_rows(conn, portfolio_id, "2026-01") if row["asset_id"] == asset_id)
        row_2027 = next(row for row in build_portfolio_report_rows(conn, portfolio_id, "2027-01") if row["asset_id"] == asset_id)

        assert row_2026["expected_production_kwh"] == 100
        assert row_2026["expected_production_source"] == "financial_model"
        assert row_2027["expected_production_source"] == "helioscope"
        assert row_2027["expected_production_kwh"] == 50
    finally:
        conn.close()


def test_preview_not_used_in_reports_and_real_values_stay_none(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    try:
        asset_id = add_asset(conn)
        portfolio_id = conn.execute("SELECT id FROM portfolio_groups WHERE name = 'Solcorelios I'").fetchone()["id"]
        conn.execute("INSERT INTO portfolio_assets (portfolio_id, asset_id, active, mapping_status, mapping_confidence) VALUES (?, ?, 1, 'manual', 1)", (portfolio_id, asset_id))
        path = tmp_path / "model.xlsx"
        financial_workbook(path)
        create_financial_model_preview(conn, upload_dir=tmp_path / "uploads", file_storage=UploadStub(path), asset_id=asset_id, base_year=2026)
        conn.commit()

        row = next(row for row in build_portfolio_report_rows(conn, portfolio_id, "2026-01") if row["asset_id"] == asset_id)

        assert row["expected_production_source"] == "none"
        assert row["actual_production_kwh"] is None
        assert row["self_use_kwh"] is None
        assert row["export_kwh"] is None
        assert row["consumption_kwh"] is None
    finally:
        conn.close()


def test_financial_model_routes_block_idor(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    try:
        asset_a = add_asset(conn, name="A")
        asset_b = add_asset(conn, name="B")
        path = tmp_path / "model.xlsx"
        financial_workbook(path)
        model_id = create_financial_model_preview(conn, upload_dir=tmp_path / "uploads", file_storage=UploadStub(path), asset_id=asset_a, base_year=2026)
        conn.commit()
    finally:
        conn.close()
    previous_db = app.config["DATABASE"]
    app.config["DATABASE"] = str(tmp_path / "financial_models.db")
    try:
        client = app.test_client()
        with client.session_transaction() as sess:
            sess["authenticated"] = True
            sess["username"] = "admin"
        assert client.get(f"/asset/{asset_a}/financial-model/{model_id}/preview").status_code == 200
        assert client.get(f"/asset/{asset_b}/financial-model/{model_id}/preview").status_code == 404
    finally:
        app.config["DATABASE"] = previous_db


def test_asset_detail_renders_active_financial_model_details(tmp_path: Path) -> None:
    conn = connect(tmp_path)
    try:
        asset_id = add_asset(conn)
        path = tmp_path / "usinage.xlsm"
        usinage_style_financial_workbook(path)
        model_id = create_financial_model_preview(conn, upload_dir=tmp_path / "uploads", file_storage=UploadStub(path), asset_id=asset_id)
        confirm_financial_model_import(conn, model_id=model_id, asset_id=asset_id)
        conn.commit()
    finally:
        conn.close()
    previous_db = app.config["DATABASE"]
    app.config["DATABASE"] = str(tmp_path / "financial_models.db")
    try:
        client = app.test_client()
        with client.session_transaction() as sess:
            sess["authenticated"] = True
            sess["username"] = "admin"
        response = client.get(f"/asset/{asset_id}")
        assert response.status_code == 200
        assert "Resumo do modelo".encode() in response.data
        assert b"Total Benefit" in response.data
    finally:
        app.config["DATABASE"] = previous_db
