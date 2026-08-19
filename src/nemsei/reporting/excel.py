"""Asset report workbook, matching the layout V1 already ships to customers.

The structure here is not a design choice: it is copied from a real V1 output on
the server, sheet by sheet and row by row, so that a customer receiving a V2
report sees the same document. Redesign comes later; parity comes first.

Two V1 behaviours are reproduced deliberately because they carry meaning:

- a value that is missing is written as an empty cell, never as a zero, so a
  month nobody measured never reads as a month that produced nothing;
- numbers are written with two decimals as text, exactly as V1 writes them.

This module takes a payload and returns a workbook. It touches no database and
no provider, so a report can be regenerated from a snapshot alone.
"""
from __future__ import annotations

from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


# V1's palette, read from a real report on the server rather than chosen here.
PRIMARY = "0B2D52"
SECONDARY = "4BA52E"
# Column widths differ per sheet in V1; these are the ones it actually writes.
SHEET_WIDTHS = {
    "Resumo": {"A": 27.0, "B": 16.0, "C": 4.0, "D": 27.0, "E": 16.0},
    "Energia": {"A": 30.0, "B": 18.0},
    "Financeiro": {"A": 32.0, "B": 18.0},
    "Qualidade dos dados": {"A": 30.0, "B": 90.0},
    "Metadados": {"A": 28.0, "B": 36.0},
}


ENERGY_METRICS = ("production_kwh", "self_use_kwh", "export_kwh", "consumption_kwh", "grid_import_kwh")
FINANCIAL_METRICS = ("savings_eur", "export_revenue_eur", "solcor_payment_eur", "fixed_monthly_fee_eur", "net_benefit_eur")
SUMMARY_ENERGY_LABELS = (
    ("Produção total (kWh)", "production_kwh"),
    ("Autoconsumo (kWh)", "self_use_kwh"),
    ("Excedente (kWh)", "export_kwh"),
    ("Consumo (kWh)", "consumption_kwh"),
    ("Importação rede (kWh)", "grid_import_kwh"),
)
SUMMARY_FINANCIAL_LABELS = (
    ("Poupança (EUR)", "savings_eur"),
    ("Receita excedente (EUR)", "export_revenue_eur"),
    ("Pagamento Solcor (EUR)", "solcor_payment_eur"),
    ("Benefício líquido (EUR)", "net_benefit_eur"),
    ("Cobertura (%)", "coverage_pct"),
)
METADATA_FIELDS = (
    "period_type",
    "period_start",
    "period_end",
    "months_count",
    "tariff_type",
    "billing_mode",
    "billing_energy_base",
)
QUALITY_FIELDS = (
    ("Estado da producao", "production_state"),
    ("Cobertura", "coverage_pct"),
    ("Total diario bruto kWh", "daily_total_kwh"),
)
SHEET_NAMES = ("Resumo", "Energia", "Financeiro", "Qualidade dos dados", "Metadados")


def format_value(value: Any) -> str | None:
    """Two decimals as text, and nothing at all when there is nothing to say."""
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float, Decimal)):
        # Float formatting, not Decimal: V1 formats floats, and Decimal would
        # round 0.005 to 0.00 where V1 writes 0.01.
        return f"{float(value):.2f}"
    return str(value)


def _apply_widths(sheet: Worksheet) -> None:
    for column, width in SHEET_WIDTHS.get(sheet.title, {}).items():
        sheet.column_dimensions[column].width = width


def _banner(sheet: Worksheet, span: str, *, bold: bool = True, italic: bool = False, size: float | None = None, fill: str | None = None) -> None:
    """V1 merges its banner rows and paints them; an unpainted banner reads wrong."""
    sheet.merge_cells(span)
    cell = sheet[span.split(":")[0]]
    cell.font = Font(bold=bold, italic=italic, size=size, color="FFFFFF" if fill else PRIMARY)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="left")


def _header_row(sheet: Worksheet, row: int, columns: int) -> None:
    for index in range(1, columns + 1):
        cell = sheet.cell(row, index)
        if cell.value is None:
            continue
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=SECONDARY)
        cell.alignment = Alignment(horizontal="left")


def _metric(payload: dict[str, Any], section: str, key: str) -> Any:
    return (payload.get(section) or {}).get(key)


def _write_metric_sheet(sheet: Worksheet, title: str, payload: dict[str, Any], section: str, keys: tuple[str, ...]) -> None:
    sheet.append([title])
    sheet.append(["Metrica", "Valor"])
    for key in keys:
        sheet.append([key, format_value(_metric(payload, section, key))])
    _banner(sheet, "A1:F1", size=14.0, fill=PRIMARY)
    _header_row(sheet, 2, 2)
    _apply_widths(sheet)


def build_asset_report_workbook(payload: dict[str, Any]) -> Workbook:
    """Render one asset report. The payload is the only input."""
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumo"

    # V1's layout, spacer rows included: title, subtitle, blank, the
    # installation/period line in A4, blank, then the header row in A6.
    summary.append([payload.get("title") or ""])
    summary.append([payload.get("subtitle") or "Relatorio de performance"])
    summary.append([])
    period = _metric(payload, "metadata", "period_label") or ""
    summary.append([f"Instalação: {payload.get('asset_name') or '-'}  |  Período: {period or '-'}"])
    summary.append([])
    summary.append(["Indicadores principais", "Valor", None, "Indicadores financeiros", "Valor"])
    for (energy_label, energy_key), (financial_label, financial_key) in zip(SUMMARY_ENERGY_LABELS, SUMMARY_FINANCIAL_LABELS, strict=True):
        section = "financial" if financial_key != "coverage_pct" else "quality"
        summary.append(
            [
                energy_label,
                format_value(_metric(payload, "energy", energy_key)),
                None,
                financial_label,
                format_value(_metric(payload, section, financial_key)),
            ]
        )

    _banner(summary, "A1:E1", size=18.0, fill=PRIMARY)
    _banner(summary, "A2:E2", bold=False, italic=True, size=11.0, fill=SECONDARY)
    _banner(summary, "A4:E4", bold=True)
    summary.row_dimensions[1].height = 30
    _apply_widths(summary)

    _write_metric_sheet(workbook.create_sheet("Energia"), "Energia e produção", payload, "energy", ENERGY_METRICS)
    _write_metric_sheet(workbook.create_sheet("Financeiro"), "Resumo financeiro", payload, "financial", FINANCIAL_METRICS)

    quality = workbook.create_sheet("Qualidade dos dados")
    quality.append(["Qualidade dos dados"])
    quality.append(["Campo", "Valor"])
    for label, key in QUALITY_FIELDS:
        quality.append([label, format_value(_metric(payload, "quality", key))])
    _banner(quality, "A1:F1", size=14.0, fill=PRIMARY)
    _header_row(quality, 2, 2)
    _apply_widths(quality)

    metadata = workbook.create_sheet("Metadados")
    metadata.append(["Metadados do relatório"])
    for field in METADATA_FIELDS:
        metadata.append([field, format_value(_metric(payload, "metadata", field))])
    _banner(metadata, "A1:F1", size=14.0, fill=PRIMARY)
    _header_row(metadata, 2, 2)
    _apply_widths(metadata)
    return workbook


def payload_from_snapshot(snapshot_payload: dict[str, Any]) -> dict[str, Any]:
    """A snapshot payload already is the report payload; kept as the seam."""
    return snapshot_payload
