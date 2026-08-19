"""Portfolio report workbook, matching the layout V1 already ships.

Like the asset report, the structure is taken from a real V1 output on the
server rather than designed here: four sheets, forty-six installation columns in
V1's order, V1's header fill and its column widths.

The renderer is a pure function of a payload, so it does not need V2 to have
portfolios as a domain concept yet. When portfolio membership arrives, it feeds
this; until then the document contract is pinned and cannot drift.

Missing stays missing: a metric nobody measured is an empty cell, never a zero.
"""
from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from nemsei.reporting.excel import format_value


HEADER_FILL = "D9EAF7"

INSTALLATION_COLUMNS = (
    "Instalacao", "Nome local", "NIF", "Subconta", "Potencia kWp", "Confianca mapping",
    "Producao real", "Estado producao", "Cobertura diaria", "Producao diaria bruta", "Helioscope", "Producao prevista",
    "Consumo previsto", "Autoconsumo previsto", "Excedente previsto", "Importacao prevista", "Taxa AC prevista", "Taxa AS prevista",
    "Specific yield previsto", "Origem prevista", "Modelo financeiro", "Versao do modelo", "Data efetiva do modelo", "Esperada ajustada",
    "Desvio", "Desvio %", "Performance vs esperado", "Specific yield", "Disponibilidade", "Autoconsumo",
    "Excedente", "Consumo", "Importacao rede", "Taxa autoconsumo", "Taxa autossuficiencia", "Valor autoconsumo",
    "Receita excedente", "Pagamento ESCO", "Mensalidade fixa", "Beneficio liquido", "Fatura", "Tarifa",
    "Estado", "Cobertura", "Warnings", "Avisos",
)

SUMMARY_METRICS = (
    "Potencia kWp", "Producao real", "Helioscope", "Producao prevista", "Consumo previsto",
    "Autoconsumo previsto", "Excedente previsto", "Importacao prevista", "Taxa AC prevista", "Taxa AS prevista",
    "Specific yield previsto", "Esperada ajustada", "Desvio", "Desvio %", "Performance vs esperado",
    "Specific yield", "Disponibilidade", "Autoconsumo", "Excedente", "Consumo", "Importacao rede",
    "Taxa autoconsumo", "Taxa autossuficiencia", "Valor autoconsumo", "Receita excedente",
    "Pagamento ESCO", "Mensalidade fixa", "Beneficio liquido", "Cobertura",
)

QUALITY_SOURCES = ("production", "financial_model", "availability", "tariff", "self_use", "invoice", "mapping")

METADATA_FIELDS = (
    "engine_version", "generated_at", "profile", "profile_version",
    "period_type", "period_start", "period_end", "months", "sources", "columns",
)

SHEET_WIDTHS = {
    "Resumo": {"A": 35.0, "B": 48.0},
    "Qualidade dos dados": {"A": 17.0, "B": 17.0, "C": 20.0, "D": 27.0, "E": 33.0, "F": 12.0, "G": 17.0, "H": 41.0},
    "Metadados": {"A": 17.0, "B": 48.0},
}
INSTALLATION_WIDTHS = (
    14.0, 14.0, 12.0, 12.0, 14.0, 19.0, 15.0, 17.0, 18.0, 23.0, 12.0, 19.0, 18.0, 22.0, 20.0, 21.0,
    18.0, 18.0, 25.0, 17.0, 19.0, 18.0, 24.0, 19.0, 12.0, 12.0, 25.0, 16.0, 17.0, 13.0, 12.0, 12.0,
    17.0, 18.0, 23.0, 19.0, 19.0, 16.0, 18.0, 19.0, 17.0, 12.0, 14.0, 12.0, 12.0, 48.0,
)
SHEET_NAMES = ("Resumo", "Instalacoes", "Qualidade dos dados", "Metadados")


def _header(sheet: Worksheet, columns: int) -> None:
    """V1 paints its header row a pale blue and bolds it; no merge, no banner."""
    for index in range(1, columns + 1):
        cell = sheet.cell(1, index)
        if cell.value is None:
            continue
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)


def _widths(sheet: Worksheet, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def build_portfolio_report_workbook(payload: dict[str, Any]) -> Workbook:
    """Render one portfolio report from a payload; no database, no provider."""
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumo"
    totals = payload.get("totals") or {}
    summary.append([payload.get("brand") or "Solcoraction"])
    for label, key in (
        ("Portfolio", "portfolio_name"),
        ("Periodo", "period_label"),
        ("Perfil", "profile"),
        ("Versao do perfil", "profile_version"),
        ("Engine", "engine_version"),
        ("Cobertura global", "coverage_pct"),
    ):
        summary.append([label, payload.get(key)])
    summary.append([])  # V1 leaves a blank row before the metric table
    summary.append(["Metrica", "Valor"])
    for metric in SUMMARY_METRICS:
        summary.append([metric, format_value(totals.get(metric))])
    _widths(summary, SHEET_WIDTHS["Resumo"])

    installations = workbook.create_sheet("Instalacoes")
    installations.append(list(INSTALLATION_COLUMNS))
    for row in payload.get("installations") or []:
        installations.append([format_value(row.get(column)) for column in INSTALLATION_COLUMNS])
    installations.append(["TOTAL"] + [format_value(totals.get(column)) for column in INSTALLATION_COLUMNS[1:]])
    _header(installations, len(INSTALLATION_COLUMNS))
    _widths(installations, {get_column_letter(index + 1): width for index, width in enumerate(INSTALLATION_WIDTHS)})

    quality = workbook.create_sheet("Qualidade dos dados")
    quality.append(["Fonte", "Cobertura"])
    coverage = payload.get("coverage") or {}
    for source in QUALITY_SOURCES:
        quality.append([source, coverage.get(source)])
    _header(quality, 2)
    _widths(quality, SHEET_WIDTHS["Qualidade dos dados"])

    metadata = workbook.create_sheet("Metadados")
    metadata_values = payload.get("metadata") or {}
    for index, field in enumerate(METADATA_FIELDS):
        metadata.append([field, metadata_values.get(field)])
    _header(metadata, 2)
    _widths(metadata, SHEET_WIDTHS["Metadados"])
    return workbook
